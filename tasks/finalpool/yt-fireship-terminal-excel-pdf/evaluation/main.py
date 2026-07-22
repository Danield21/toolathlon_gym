"""
Evaluation for yt-fireship-terminal-excel-pdf task.
"""
import csv
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, abs_tol=1.0, rel_tol=0.0):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_gt_quarterly_stats(groundtruth_workspace):
    path = os.path.join(groundtruth_workspace, "quarterly_stats.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return None


def load_gt_xlsx(groundtruth_workspace):
    path = os.path.join(groundtruth_workspace, "Fireship_Analytics.xlsx")
    if not os.path.exists(path):
        return None
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None


def check_csv(agent_workspace, groundtruth_workspace):
    print("\n=== Check 1: fireship_raw.csv ===")
    csv_path = os.path.join(agent_workspace, "fireship_raw.csv")
    if not os.path.exists(csv_path):
        record("fireship_raw.csv exists", False, f"Not found at {csv_path}")
        return
    record("fireship_raw.csv exists", True)

    try:
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            headers = reader.fieldnames or []
    except Exception as e:
        record("CSV readable", False, str(e))
        return
    record("CSV readable", True)

    required_cols = ["video_id", "title", "published_date", "duration_sec",
                     "view_count", "like_count"]
    headers_lower = [h.lower().strip() for h in headers]
    has_required = all(col in headers_lower for col in required_cols)
    record("CSV has all required columns", has_required, f"Found: {headers}")
    record("CSV has correct column ORDER",
           [h.lower().strip() for h in headers[:6]] == required_cols,
           f"Got: {headers[:6]}")

    # Compare to GT row count (104)
    gt_path = os.path.join(groundtruth_workspace, "fireship_raw.csv")
    expected_n = None
    if os.path.exists(gt_path):
        with open(gt_path, newline="", encoding="utf-8") as f:
            expected_n = sum(1 for _ in csv.DictReader(f))
    if expected_n is not None:
        record(f"CSV row count == {expected_n}", len(rows) == expected_n,
               f"Got {len(rows)}")
    else:
        record("CSV has at least 100 rows", len(rows) >= 100,
               f"Got {len(rows)}")

    # YYYY-MM-DD date format
    if rows:
        bad_dates = sum(1 for r in rows if not (
            len(r.get("published_date", "")) == 10 and
            r.get("published_date", "")[4] == "-" and
            r.get("published_date", "")[7] == "-"))
        record("All published_date values are YYYY-MM-DD",
               bad_dates == 0, f"{bad_dates} bad dates out of {len(rows)}")

        # duration_sec is integer
        bad_dur = 0
        for r in rows:
            try:
                int(r.get("duration_sec", ""))
            except Exception:
                bad_dur += 1
        record("All duration_sec values are integers",
               bad_dur == 0, f"{bad_dur} bad duration values")


def check_json(agent_workspace, groundtruth_workspace):
    print("\n=== Check 2: quarterly_stats.json ===")
    json_path = os.path.join(agent_workspace, "quarterly_stats.json")
    if not os.path.exists(json_path):
        record("quarterly_stats.json exists", False, f"Not found at {json_path}")
        return
    record("quarterly_stats.json exists", True)

    try:
        with open(json_path) as f:
            data = json.load(f)
    except Exception as e:
        record("quarterly_stats.json is valid JSON", False, str(e))
        return
    record("quarterly_stats.json is valid JSON", True)

    gt = load_gt_quarterly_stats(groundtruth_workspace) or []
    expected_n = len(gt) if gt else 4
    record(f"Has {expected_n} quarters", len(data) == expected_n,
           f"Found {len(data)} entries")

    if not data:
        return

    first = data[0]
    has_keys = all(k in first for k in ["Quarter", "Video_Count",
                                         "Total_Views", "Avg_Engagement_Rate"])
    record("Each entry has Quarter/Video_Count/Total_Views/Avg_Engagement_Rate",
           has_keys, f"Keys found: {list(first.keys())}")

    quarters_actual = [d.get("Quarter", "") for d in data]
    record("Quarters sorted chronologically",
           quarters_actual == sorted(quarters_actual),
           f"Got: {quarters_actual}")

    # Compare to GT for each quarter
    gt_by_q = {d["Quarter"]: d for d in gt}
    a_by_q = {d.get("Quarter"): d for d in data}
    for q, gt_d in gt_by_q.items():
        a_d = a_by_q.get(q)
        if a_d is None:
            record(f"Quarter '{q}' present", False)
            continue
        record(f"Quarter '{q}' present", True)
        record(f"'{q}' Video_Count == {gt_d['Video_Count']}",
               num_close(a_d.get("Video_Count"), gt_d["Video_Count"], abs_tol=0),
               f"got {a_d.get('Video_Count')}")
        # Total_Views: tight tol (1% relative or abs 100)
        record(f"'{q}' Total_Views ~ {gt_d['Total_Views']}",
               num_close(a_d.get("Total_Views"), gt_d["Total_Views"],
                         abs_tol=100, rel_tol=0.01),
               f"got {a_d.get('Total_Views')}")
        # Engagement: 0.01% tol
        record(f"'{q}' Avg_Engagement_Rate ~ {gt_d['Avg_Engagement_Rate']}",
               num_close(a_d.get("Avg_Engagement_Rate"),
                         gt_d["Avg_Engagement_Rate"], abs_tol=0.01),
               f"got {a_d.get('Avg_Engagement_Rate')}")


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Check 3: Fireship_Analytics.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Fireship_Analytics.xlsx")
    if not os.path.exists(xlsx_path):
        record("Fireship_Analytics.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Fireship_Analytics.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Raw_Data sheet", "raw_data" in sheet_names_lower)
    record("Has Quarterly_Stats sheet", "quarterly_stats" in sheet_names_lower)
    record("Has Summary sheet", "summary" in sheet_names_lower)

    gt_wb = load_gt_xlsx(groundtruth_workspace)

    # Raw_Data: rows match GT count
    if "raw_data" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("raw_data")]]
        a_rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        a_data = a_rows[1:]  # skip header
        if gt_wb and "Raw_Data" in gt_wb.sheetnames:
            gt_rows = [r for r in gt_wb["Raw_Data"].iter_rows(values_only=True)
                       if any(c is not None for c in r)]
            gt_data = gt_rows[1:]
            record(f"Raw_Data row count == {len(gt_data)}",
                   len(a_data) == len(gt_data),
                   f"Got {len(a_data)}")
        else:
            record("Raw_Data has at least 100 rows", len(a_data) >= 100)

        # Header order
        if a_rows:
            req_hdrs = ["video_id", "title", "published_date",
                        "duration_sec", "view_count", "like_count"]
            actual_hdrs = [str(c).strip().lower() if c else "" for c in a_rows[0][:6]]
            record("Raw_Data header order", actual_hdrs == req_hdrs,
                   f"Got {a_rows[0][:6]}")

    # Quarterly_Stats
    if "quarterly_stats" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("quarterly_stats")]]
        a_rows = [r for r in ws.iter_rows(values_only=True)
                  if any(c is not None for c in r)]
        if a_rows:
            req_hdrs = ["quarter", "video_count", "total_views", "avg_engagement_rate"]
            actual_hdrs = [str(c).strip().lower() if c else "" for c in a_rows[0][:4]]
            record("Quarterly_Stats header order matches",
                   actual_hdrs == req_hdrs, f"Got {a_rows[0][:4]}")

        if gt_wb and "Quarterly_Stats" in gt_wb.sheetnames:
            gt_rows = [r for r in gt_wb["Quarterly_Stats"].iter_rows(values_only=True)
                       if any(c is not None for c in r)]
            gt_data = gt_rows[1:]
            a_data = a_rows[1:]
            record(f"Quarterly_Stats row count == {len(gt_data)}",
                   len(a_data) == len(gt_data),
                   f"Got {len(a_data)}")

            # Sort: chronological
            a_quarters = [str(r[0]).strip() for r in a_data if r and r[0]]
            record("Quarterly_Stats sorted chronologically",
                   a_quarters == sorted(a_quarters), f"Got {a_quarters}")

            # Check each row matches GT
            gt_by_q = {str(r[0]).strip(): r for r in gt_data if r and r[0]}
            a_by_q = {str(r[0]).strip(): r for r in a_data if r and r[0]}
            for q, gt_r in gt_by_q.items():
                ar = a_by_q.get(q)
                if ar is None:
                    record(f"Quarterly_Stats '{q}' present", False)
                    continue
                # Video_Count exact
                record(f"Q-Stats '{q}' Video_Count exact",
                       num_close(ar[1], gt_r[1], abs_tol=0))
                # Total_Views with 1% tol
                record(f"Q-Stats '{q}' Total_Views (1% tol)",
                       num_close(ar[2], gt_r[2], abs_tol=100, rel_tol=0.01))
                # Engagement rate with 0.01 tol
                record(f"Q-Stats '{q}' Avg_Engagement_Rate (0.01)",
                       num_close(ar[3], gt_r[3], abs_tol=0.01))

    # Summary sheet
    if "summary" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows = [r for r in ws.iter_rows(values_only=True)
                if any(c is not None for c in r)]
        if rows:
            hdrs = [str(c).strip().lower() if c else "" for c in rows[0][:2]]
            record("Summary headers Metric/Value",
                   hdrs == ["metric", "value"],
                   f"Got {rows[0][:2]}")
        # Build lookup
        lookup = {}
        for r in rows[1:]:
            if r and r[0] is not None:
                lookup[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        if gt_wb and "Summary" in gt_wb.sheetnames:
            gt_rows = [r for r in gt_wb["Summary"].iter_rows(values_only=True)
                       if any(c is not None for c in r)]
            gt_lookup = {}
            for r in gt_rows[1:]:
                if r and r[0] is not None:
                    gt_lookup[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

            # Required keys (per task.md): Total_Videos, Total_Views, Best_Quarter,
            # Worst_Quarter, Overall_Avg_Engagement
            for key in ["total_videos", "total_views", "best_quarter",
                        "worst_quarter", "overall_avg_engagement"]:
                a_val = lookup.get(key)
                gt_val = gt_lookup.get(key)
                record(f"Summary has '{key}'", a_val is not None,
                       f"Available keys: {list(lookup.keys())}")
                if gt_val is None:
                    continue
                if isinstance(gt_val, (int, float)):
                    if key == "overall_avg_engagement":
                        record(f"Summary '{key}' value (0.01 tol)",
                               num_close(a_val, gt_val, abs_tol=0.01),
                               f"got {a_val}, expected {gt_val}")
                    elif key == "total_videos":
                        record(f"Summary '{key}' exact",
                               num_close(a_val, gt_val, abs_tol=0),
                               f"got {a_val}, expected {gt_val}")
                    elif key == "total_views":
                        record(f"Summary '{key}' (1% tol)",
                               num_close(a_val, gt_val, abs_tol=100, rel_tol=0.01),
                               f"got {a_val}, expected {gt_val}")
                else:
                    record(f"Summary '{key}' string match",
                           str_match(a_val, gt_val),
                           f"got {a_val}, expected {gt_val}")


def check_word(agent_workspace):
    print("\n=== Check 4: Fireship_Analytics_Report.docx ===")
    docx_path = os.path.join(agent_workspace, "Fireship_Analytics_Report.docx")
    if not os.path.exists(docx_path):
        record("Fireship_Analytics_Report.docx exists", False, f"Not found at {docx_path}")
        return
    record("Fireship_Analytics_Report.docx exists", True)

    try:
        import docx
        doc = docx.Document(docx_path)
        full_text = " ".join(p.text for p in doc.paragraphs).lower()
    except Exception as e:
        record("Word document readable", False, str(e))
        return
    record("Word document readable", True)

    record("Has Executive Summary section", "executive summary" in full_text)
    record("Has Quarterly Performance section", "quarterly performance" in full_text)
    record("Has Key Observations section", "key observations" in full_text)
    record("Has Appendix section", "appendix" in full_text)
    record("Mentions Fireship channel", "fireship" in full_text)
    # Should mention total videos count (104) — require 104 to co-occur with 'video' wording
    # (avoid matching '104' that appears in dates like '2024-10-04')
    import re as _re
    # Accept: "104 videos", "104 in total", "total ... 104", etc.
    total_videos_104 = (
        bool(_re.search(r"\b104\s+videos?\b", full_text)) or
        bool(_re.search(r"\b104\s+in\s+total\b", full_text)) or
        bool(_re.search(r"total[^.\n]{0,60}\b104\b", full_text)) or
        bool(_re.search(r"\b104\b[^.\n]{0,30}\s+total", full_text))
    )
    record("Mentions total videos count 104 (co-located with 'video'/'total')", total_videos_104,
           "Word doc should mention total video count (104) in context of videos/total")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_csv(args.agent_workspace, args.groundtruth_workspace)
    check_json(args.agent_workspace, args.groundtruth_workspace)
    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
