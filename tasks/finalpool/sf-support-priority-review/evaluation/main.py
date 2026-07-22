"""Evaluation for sf-support-priority-review."""
import argparse
import os
import sys
import openpyxl
import psycopg2


DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Support_Priority_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Support_Priority_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Priority Analysis
    print(f"  Checking Priority Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Priority Analysis")
    g_rows = load_sheet_rows(gt_wb, "Priority Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Priority Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Priority Analysis' not found in groundtruth")
    else:
        sheet_name = "Priority Analysis"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{key}.Ticket_Count: {a_row[1]} vs {g_row[1]} (tol=5)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.5):
                    errors.append(f"{key}.Avg_Response_Hrs: {a_row[2]} vs {g_row[2]} (tol=0.5)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.1):
                    errors.append(f"{key}.SLA_Target_Hrs: {a_row[3]} vs {g_row[3]} (tol=0.1)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not str_match(a_row[4], g_row[4]):
                    errors.append(f"{key}.SLA_Met: '{a_row[4]}' vs '{g_row[4]}'")

            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.1):
                    errors.append(f"{key}.Avg_CSAT: {a_row[5]} vs {g_row[5]} (tol=0.1)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # Tighter tolerance for small-integer counts (Priorities_Meeting_SLA, Priorities_Missing_SLA)
                metric_key = key
                if "priorities" in metric_key:
                    # Exact integer match for small-int counts
                    tol = 0
                elif "csat" in metric_key:
                    tol = 0.1
                elif "total_tickets" in metric_key or "total" in metric_key:
                    tol = 10.0
                else:
                    tol = 1.0
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check email (required by task.md)
    print("  Checking email...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE LOWER(to_addr::text) LIKE '%support-manager@company.com%'
              AND LOWER(subject) LIKE '%support%priority%analysis%report%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
        if not emails:
            all_errors.append("Email not found: required to=support-manager@company.com with subject 'Support Priority Analysis Report'")
        else:
            # Body should mention priorities not meeting SLA (3 priorities all No)
            body = (emails[0][2] or "").lower()
            # Per task.md: 'flagging any priorities that are not meeting SLA' -
            # body should mention priority terms
            mentioned_pri = sum(1 for p in ["high", "medium", "low"] if p in body)
            if mentioned_pri < 1:
                all_errors.append("Email body missing reference to priorities not meeting SLA (high/medium/low)")
            print("    PASS" if not all_errors or all_errors[-1] != "" else "")
    except Exception as e:
        all_errors.append(f"Email DB check error: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
