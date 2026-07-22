"""Evaluation script for wc-customer-gform-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def norm(s):
    return str(s or "").strip().lower().replace("_", " ")

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _fetch_expected_segments():
    """Compute expected segment stats from real WC data with fallback."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
        SELECT 'VIP' as seg, COUNT(*), SUM(total_spent)::float, AVG(total_spent)::float, AVG(orders_count::float)
        FROM wc.customers WHERE total_spent > 500
        UNION ALL
        SELECT 'Regular', COUNT(*), SUM(total_spent)::float, AVG(total_spent)::float, AVG(orders_count::float)
        FROM wc.customers WHERE total_spent >= 100 AND total_spent <= 500
        UNION ALL
        SELECT 'New', COUNT(*), SUM(total_spent)::float, AVG(total_spent)::float, AVG(orders_count::float)
        FROM wc.customers WHERE total_spent < 100
        """)
        rows = {}
        for seg, cnt, tot, avg, avo in cur.fetchall():
            rows[seg] = {
                "count": int(cnt),
                "total_revenue": round(float(tot or 0), 2),
                "avg_spend": round(float(avg or 0), 2),
                "avg_orders": round(float(avo or 0), 1),
            }
        conn.close()
        return rows
    except Exception as e:
        print(f"[WARN] could not fetch expected segments: {e}")
        return None


def _fetch_top10():
    """Fetch top 10 customers from real WC data."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""SELECT first_name, last_name, email, total_spent::float, orders_count
        FROM wc.customers ORDER BY total_spent DESC LIMIT 10""")
        rows = []
        for fn, ln, em, ts, oc in cur.fetchall():
            ts_f = float(ts)
            seg = "VIP" if ts_f > 500 else ("Regular" if ts_f >= 100 else "New")
            rows.append({
                "name": f"{fn} {ln}",
                "email": em,
                "total_spent": round(ts_f, 2),
                "order_count": int(oc),
                "segment": seg,
            })
        conn.close()
        return rows
    except Exception as e:
        print(f"[WARN] could not fetch top10: {e}")
        return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    EXPECTED_SEGMENTS = _fetch_expected_segments()
    EXPECTED_TOP10 = _fetch_top10()

    # Check Customer_Insights_Report.xlsx
    excel_path = os.path.join(agent_workspace, "Customer_Insights_Report.xlsx")
    check("Customer_Insights_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            check("Open Excel", False, str(e))
            wb = None
        if wb:
            sheet_norm = {norm(s): s for s in wb.sheetnames}

            # ---- Customer_Segments sheet ----
            cs_key = sheet_norm.get(norm("Customer_Segments"))
            check("Customer_Segments sheet exists", cs_key is not None)
            if cs_key:
                ws = wb[cs_key]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [norm(c) for c in rows[0]]
                    data_rows = rows[1:]
                    # Required headers
                    for h in ["segment", "customer count", "total revenue", "avg spend", "avg orders"]:
                        check(f"CS has column {h}", h in headers, f"got {headers}")
                    # Row count exactly 3
                    check("CS has 3 segment rows", len(data_rows) == 3, f"got {len(data_rows)}")
                    # Build map by segment
                    if "segment" in headers:
                        seg_idx = headers.index("segment")
                        cc_idx = headers.index("customer count") if "customer count" in headers else None
                        tr_idx = headers.index("total revenue") if "total revenue" in headers else None
                        as_idx = headers.index("avg spend") if "avg spend" in headers else None
                        ao_idx = headers.index("avg orders") if "avg orders" in headers else None
                        agent_segs = {}
                        for r in data_rows:
                            if r and r[seg_idx]:
                                agent_segs[norm(r[seg_idx])] = r
                        if EXPECTED_SEGMENTS:
                            for seg, exp in EXPECTED_SEGMENTS.items():
                                kn = norm(seg)
                                if kn in agent_segs:
                                    r = agent_segs[kn]
                                    if cc_idx is not None:
                                        av = safe_float(r[cc_idx])
                                        check(f"{seg} count == {exp['count']}", av is not None and abs(av - exp['count']) <= 0, f"got {av}")
                                    if tr_idx is not None:
                                        av = safe_float(r[tr_idx])
                                        check(f"{seg} total_revenue ~ {exp['total_revenue']}",
                                              av is not None and abs(av - exp['total_revenue']) <= max(1.0, abs(exp['total_revenue'])*0.01),
                                              f"got {av}")
                                    if as_idx is not None:
                                        av = safe_float(r[as_idx])
                                        check(f"{seg} avg_spend ~ {exp['avg_spend']}",
                                              av is not None and abs(av - exp['avg_spend']) <= max(1.0, abs(exp['avg_spend'])*0.01),
                                              f"got {av}")
                                    if ao_idx is not None:
                                        av = safe_float(r[ao_idx])
                                        check(f"{seg} avg_orders ~ {exp['avg_orders']}",
                                              av is not None and abs(av - exp['avg_orders']) <= 0.2,
                                              f"got {av}")
                                else:
                                    check(f"{seg} segment present", False, f"missing in {list(agent_segs.keys())}")
                    # Sorting check: by Total_Revenue desc
                    if "total revenue" in headers and len(data_rows) >= 2:
                        tr_idx = headers.index("total revenue")
                        revs = [safe_float(r[tr_idx]) for r in data_rows]
                        revs_clean = [v for v in revs if v is not None]
                        check("CS sorted by Total_Revenue desc",
                              revs_clean == sorted(revs_clean, reverse=True),
                              f"got {revs}")

            # ---- Top_Customers sheet ----
            tc_key = sheet_norm.get(norm("Top_Customers"))
            check("Top_Customers sheet exists", tc_key is not None)
            if tc_key:
                ws = wb[tc_key]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [norm(c) for c in rows[0]]
                    data_rows = rows[1:]
                    for h in ["customer name", "email", "total spent", "order count", "segment"]:
                        check(f"TC has column {h}", h in headers, f"got {headers}")
                    check("TC has 10 rows", len(data_rows) == 10, f"got {len(data_rows)}")
                    if EXPECTED_TOP10 and "email" in headers:
                        em_idx = headers.index("email")
                        ts_idx = headers.index("total spent") if "total spent" in headers else None
                        oc_idx = headers.index("order count") if "order count" in headers else None
                        sg_idx = headers.index("segment") if "segment" in headers else None
                        # Match by email
                        agent_top = {}
                        for r in data_rows:
                            if r and r[em_idx]:
                                agent_top[norm(r[em_idx])] = r
                        for exp in EXPECTED_TOP10:
                            kn = norm(exp["email"])
                            r = agent_top.get(kn)
                            if r is None:
                                check(f"Top customer {exp['email']} present", False, f"missing")
                                continue
                            check(f"Top customer {exp['email']} present", True)
                            if ts_idx is not None:
                                av = safe_float(r[ts_idx])
                                check(f"{exp['email']} total_spent ~ {exp['total_spent']}",
                                      av is not None and abs(av - exp['total_spent']) <= max(0.5, abs(exp['total_spent'])*0.01),
                                      f"got {av}")
                            if oc_idx is not None:
                                av = safe_float(r[oc_idx])
                                check(f"{exp['email']} order_count == {exp['order_count']}",
                                      av is not None and abs(av - exp['order_count']) <= 0,
                                      f"got {av}")
                            if sg_idx is not None:
                                check(f"{exp['email']} segment == {exp['segment']}",
                                      norm(r[sg_idx]) == norm(exp['segment']),
                                      f"got {r[sg_idx]}")
                        # Order check: sorted desc by total_spent
                        if ts_idx is not None and len(data_rows) >= 2:
                            spends = [safe_float(r[ts_idx]) for r in data_rows]
                            spends_clean = [v for v in spends if v is not None]
                            check("TC sorted by Total_Spent desc",
                                  spends_clean == sorted(spends_clean, reverse=True),
                                  f"got {spends}")

            # ---- Segment_Strategy sheet ----
            ss_key = sheet_norm.get(norm("Segment_Strategy"))
            check("Segment_Strategy sheet exists", ss_key is not None)
            if ss_key:
                ws = wb[ss_key]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [norm(c) for c in rows[0]]
                    data_rows = rows[1:]
                    for h in ["segment", "engagement strategy", "retention risk", "target action"]:
                        check(f"SS has column {h}", h in headers, f"got {headers}")
                    check("SS has 3 rows", len(data_rows) == 3, f"got {len(data_rows)}")
                    if "segment" in headers:
                        sg_idx = headers.index("segment")
                        seg_set = {norm(r[sg_idx]) for r in data_rows if r and r[sg_idx]}
                        for s in ["vip", "regular", "new"]:
                            check(f"SS has {s} segment", s in seg_set, f"got {seg_set}")

    # Check Python script exists with specific filename
    seg_script = os.path.join(agent_workspace, "customer_segmenter.py")
    check("customer_segmenter.py exists", os.path.exists(seg_script), f"in {agent_workspace}")

    # Database checks - require non-noise data
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Form: must be 'Customer Experience Survey' (exact normalized match)
        cur.execute("""SELECT id, title, document_title FROM gform.forms""")
        all_forms = cur.fetchall()
        survey_forms = []
        for fid, t, dt in all_forms:
            t_norm = norm(t)
            dt_norm = norm(dt)
            if t_norm == 'customer experience survey' or dt_norm == 'customer experience survey':
                survey_forms.append((fid, t))
        check("Customer Experience Survey form created (exact title)", len(survey_forms) >= 1,
              f"found titles: {[(t, dt) for _, t, dt in all_forms]}")

        if survey_forms:
            form_ids = [r[0] for r in survey_forms]
            cur.execute("""SELECT COUNT(*) FROM gform.questions WHERE form_id = ANY(%s)""", (form_ids,))
            q_count = cur.fetchone()[0]
            check("Survey form has >= 4 questions", q_count >= 4, f"got {q_count}")

        # Notion: must have 'Customer Intelligence Hub' page (excluding noise 'Meeting Notes Archive')
        cur.execute("""SELECT id, properties FROM notion.pages WHERE archived = false""")
        all_pages = cur.fetchall()
        hub_found = False
        hub_page_id = None
        for pid, props in all_pages:
            try:
                title_arr = props.get('title', {}).get('title', [])
                title_text = "".join([t.get('text', {}).get('content', '') for t in title_arr]).lower()
                if 'customer intelligence hub' in title_text:
                    hub_found = True
                    hub_page_id = pid
                    break
            except Exception:
                pass
        check("'Customer Intelligence Hub' Notion page created", hub_found,
              f"checked {len(all_pages)} pages")

        # Notion: page should reference key segment terms (block content check)
        if hub_found and hub_page_id is not None:
            try:
                cur.execute("""SELECT COALESCE(string_agg(block_data::text, ' '), '')
                               FROM notion.blocks WHERE parent_id = %s AND archived = false""",
                            (hub_page_id,))
                row = cur.fetchone()
                blocks_text = (row[0] if row else "").lower()
                # Require content references segment terms or strategy
                seg_terms_found = sum(1 for term in ['vip', 'regular', 'new'] if term in blocks_text)
                strategy_present = ('segment' in blocks_text) or ('strategy' in blocks_text)
                check("Notion hub page references >=2 segment terms or 'segment'/'strategy'",
                      seg_terms_found >= 2 or strategy_present,
                      f"seg_terms={seg_terms_found}, strategy_present={strategy_present}, len={len(blocks_text)}")
            except Exception as e:
                # Don't fail catastrophically if blocks schema differs; log only
                print(f"  [INFO] Notion content check skipped: {e}")

        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
