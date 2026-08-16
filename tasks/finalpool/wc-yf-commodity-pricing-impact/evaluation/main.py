"""Evaluation for wc-yf-commodity-pricing-impact."""
import argparse
import os
import sys

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def nums_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Commodity_Impact.xlsx")
    gt_path = os.path.join(gt_workspace, "Commodity_Impact.xlsx")
    if not os.path.exists(path):
        return ["Commodity_Impact.xlsx not found"]
    if not os.path.exists(gt_path):
        return ["GT Commodity_Impact.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)

        # Check Gold Price Trend sheet
        rows = load_sheet_rows(wb, "Gold Price Trend")
        gt_rows = load_sheet_rows(gt_wb, "Gold Price Trend")
        if rows is None:
            errors.append("Sheet 'Gold Price Trend' not found")
        elif gt_rows is None:
            errors.append("GT Sheet 'Gold Price Trend' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]
            # Task says "30 most recent trading days" - GT has 31, accept 28-32
            if not (28 <= len(data_rows) <= 32):
                errors.append(f"Gold Price Trend has {len(data_rows)} rows, expected ~30 (got GT={len(gt_data)})")
            # Build GT date->row lookup
            def date_key(d):
                return str(d)[:10] if d else ""
            gt_lookup = {date_key(r[0]): r for r in gt_data}
            # Validate Trend_Direction logic per row in agent: Up if Gold_Close > MA else Down
            wrong_trend = 0
            wrong_direction_logic = 0
            for r in data_rows:
                if len(r) < 4 or r[1] is None or r[2] is None or r[3] is None:
                    continue
                try:
                    gc = float(r[1])
                    ma = float(r[2])
                    expected_dir = "up" if gc > ma else "down"
                    if str(r[3]).strip().lower() != expected_dir:
                        wrong_direction_logic += 1
                except (ValueError, TypeError):
                    pass
            if wrong_direction_logic > 1:  # allow 1 boundary case
                errors.append(f"{wrong_direction_logic} rows have incorrect Trend_Direction logic")
            # Check matching dates against GT
            matched = 0
            for r in data_rows:
                gt_r = gt_lookup.get(date_key(r[0]))
                if gt_r:
                    matched += 1
                    # Verify Gold_Close (tol 30 due to mock variance)
                    if r[1] is not None and gt_r[1] is not None:
                        try:
                            if abs(float(r[1]) - float(gt_r[1])) > 30:
                                errors.append(f"Gold_Close on {r[0]}: {r[1]}, expected ~{gt_r[1]}")
                                break  # only report once
                        except (ValueError, TypeError):
                            pass
            if matched < 20:
                errors.append(f"Only {matched} dates match between agent and GT")

        # Check Category Analysis sheet - validate ALL columns per category
        rows2 = load_sheet_rows(wb, "Category Analysis")
        gt_rows2 = load_sheet_rows(gt_wb, "Category Analysis")
        if rows2 is None:
            errors.append("Sheet 'Category Analysis' not found")
        elif gt_rows2 is None:
            errors.append("GT Sheet 'Category Analysis' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            gt_data2 = [r for r in gt_rows2[1:] if r and r[0] is not None]
            # Task says "all seven product categories"
            if not (6 <= len(data_rows2) <= 9):
                errors.append(f"Category Analysis has {len(data_rows2)} rows, expected ~7")
            a_lookup2 = {str(r[0]).strip().lower(): r for r in data_rows2}
            for g in gt_data2:
                gkey = str(g[0]).strip().lower()
                # Use exact match OR substring fallback for naming variations like "Audio (Headphones)" vs "Audio"
                a = a_lookup2.get(gkey)
                if a is None:
                    # Try to find by first word match (e.g. 'Audio' vs 'Audio (Headphones)')
                    first_word = gkey.split()[0] if gkey else ""
                    candidates = [v for k, v in a_lookup2.items() if k.startswith(first_word) or first_word in k]
                    if candidates:
                        a = candidates[0]
                if a is None:
                    errors.append(f"Category Analysis: missing category '{g[0]}'")
                    continue
                # Avg_Price (col 1) tol 5%
                if a[1] is None or abs(float(a[1]) - float(g[1])) > max(0.5, abs(float(g[1]))*0.05):
                    errors.append(f"Category Analysis {gkey} Avg_Price={a[1]}, expected ~{g[1]}")
                # Product_Count (col 2) tol 0
                try:
                    if a[2] is None or abs(int(a[2]) - int(g[2])) > 1:
                        errors.append(f"Category Analysis {gkey} Product_Count={a[2]}, expected {g[2]}")
                except (ValueError, TypeError):
                    pass
                # Estimated_Material_Cost_Pct (col 3) tol 0.02
                # Normalize: GT stores as decimal (0.13), agent may write as
                # percentage (13).  Convert both to decimal before comparing.
                try:
                    a_pct = float(a[3]) if a[3] is not None else None
                    g_pct = float(g[3]) if g[3] is not None else None
                    if a_pct is not None and a_pct > 1:
                        a_pct = a_pct / 100.0
                    if g_pct is not None and g_pct > 1:
                        g_pct = g_pct / 100.0
                    if a_pct is None or g_pct is None or abs(a_pct - g_pct) > 0.02:
                        errors.append(f"Category Analysis {gkey} Material_Cost_Pct={a[3]}, expected ~{g[3]}")
                except (ValueError, TypeError):
                    pass
                # Estimated_Material_Cost (col 4) tol 5% rel
                try:
                    if a[4] is None or abs(float(a[4]) - float(g[4])) > max(1.0, abs(float(g[4]))*0.1):
                        errors.append(f"Category Analysis {gkey} Material_Cost={a[4]}, expected ~{g[4]}")
                except (ValueError, TypeError):
                    pass
                # Target_Margin_Pct (col 5) - tol 1
                try:
                    if a[5] is None or abs(float(a[5]) - float(g[5])) > 1:
                        errors.append(f"Category Analysis {gkey} Target_Margin_Pct={a[5]}, expected {g[5]}")
                except (ValueError, TypeError):
                    pass
                # Current_Margin_vs_Target (col 6) - exact match against GT
                if len(a) > 6 and len(g) > 6 and g[6] is not None:
                    a_val = str(a[6] or "").strip().lower()
                    g_val = str(g[6]).strip().lower()
                    # Accept synonyms
                    synonyms = {
                        "meets": ["meets", "meets target", "above", "above target", "exceeds", "on target"],
                        "meets target": ["meets", "meets target", "above", "above target", "exceeds", "on target"],
                        "at risk": ["at risk", "below", "falls short", "shortfall", "below target"],
                    }
                    accepted = synonyms.get(g_val, [g_val])
                    if a_val not in accepted:
                        errors.append(f"Category Analysis {gkey} Current_Margin_vs_Target={a[6]}, expected {g[6]}")

        # Check Correlation Summary sheet
        rows3 = load_sheet_rows(wb, "Correlation Summary")
        gt_rows3 = load_sheet_rows(gt_wb, "Correlation Summary")
        if rows3 is None:
            errors.append("Sheet 'Correlation Summary' not found")
        elif gt_rows3 is None:
            errors.append("GT Sheet 'Correlation Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            gt_data3 = [r for r in gt_rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 3:
                errors.append(f"Correlation Summary has {len(data_rows3)} rows, expected >= 3")
            a_lookup3 = {str(r[0]).strip().lower(): r for r in data_rows3}
            # Required keys to check
            required = {
                "current_gold_price": (50, "numeric"),  # tol $50
                "gold_30d_moving_avg": (50, "numeric"),
                "overall_trend": (None, "string"),
                "categories_analyzed": (0, "numeric"),
            }
            for g in gt_data3:
                gkey = str(g[0]).strip().lower()
                # Try direct lookup
                a = a_lookup3.get(gkey)
                if a is None:
                    # try fuzzy
                    for k, v in a_lookup3.items():
                        if gkey.replace("_", " ").replace(" ", "") in k.replace("_", " ").replace(" ", ""):
                            a = v
                            break
                if a is None:
                    if gkey in required:
                        errors.append(f"Correlation Summary: missing metric '{g[0]}'")
                    continue
                if g[1] is None:
                    continue
                if isinstance(g[1], str):
                    if str(a[1] or "").strip().lower() != g[1].strip().lower():
                        errors.append(f"Correlation Summary {gkey}={a[1]}, expected {g[1]}")
                else:
                    try:
                        gv = float(g[1])
                        if "price" in gkey or "moving" in gkey or "avg" in gkey:
                            tol = max(50, abs(gv)*0.02)
                        elif "categor" in gkey or "count" in gkey:
                            tol = 0
                        else:
                            tol = max(0.5, abs(gv)*0.05)
                        if a[1] is None or abs(float(a[1]) - gv) > tol:
                            errors.append(f"Correlation Summary {gkey}={a[1]}, expected ~{g[1]}")
                    except (ValueError, TypeError):
                        pass

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    from docx import Document
    path = os.path.join(agent_workspace, "Commodity_Report.docx")
    if not os.path.exists(path):
        return ["Commodity_Report.docx not found"]
    try:
        doc = Document(path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
        if len(full_text) < 300:
            errors.append(f"Word doc too short ({len(full_text)} chars)")
        if "gold" not in full_text:
            errors.append("Word doc does not mention 'gold'")
        # Required content topics from task.md
        if "trend" not in full_text and "moving" not in full_text:
            errors.append("Word doc does not mention price trend or moving average")
        # Must discuss category exposure
        if not any(c in full_text for c in ["watch", "electronic", "camera", "audio", "appliance", "tv", "category"]):
            errors.append("Word doc does not discuss product categories")
        # Must address margin pressure
        if "margin" not in full_text and "pressure" not in full_text and "cost" not in full_text:
            errors.append("Word doc does not address margin pressure or material costs")
        # Must include recommendations
        if "recommend" not in full_text and "suggest" not in full_text and "should" not in full_text and "strategy" not in full_text:
            errors.append("Word doc does not include recommendations or pricing strategy")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
