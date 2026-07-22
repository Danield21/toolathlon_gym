"""Evaluation for sf-sales-product-category."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def is_numeric(val):
    if val is None: return False
    try:
        float(str(val).replace(",", "").replace("$", "").strip())
        return True
    except (ValueError, TypeError):
        return False


def check_gsheet(errors):
    """Verify Google Sheet 'Product Category Report' exists with substantive data."""
    try:
        conn = psycopg2.connect(**DB); cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gsheet.spreadsheets
            WHERE LOWER(title) LIKE '%product category report%'
               OR LOWER(title) = LOWER('Product Category Report')
        """)
        sheets = cur.fetchall()
        if not sheets:
            errors.append("Google Sheet 'Product Category Report' not found")
            cur.close(); conn.close(); return
        sid = sheets[0][0]
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
        cell_count = cur.fetchone()[0]
        if cell_count < 10:
            errors.append(f"Google Sheet has only {cell_count} cells (expected >=10)")
        cur.close(); conn.close()
    except Exception as e:
        errors.append(f"Google Sheet check error: {e}")


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Product_Categories.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Product_Categories.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Product Categories
    print(f"  Checking Product Categories...")
    a_rows = load_sheet_rows(agent_wb, "Product Categories")
    g_rows = load_sheet_rows(gt_wb, "Product Categories")
    if a_rows is None:
        all_errors.append("Sheet 'Product Categories' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Product Categories' not found in groundtruth")
    else:
        sheet_name = "Product Categories"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):  # exact product count
                    errors.append(f"{key}.Product_Count: {a_row[1]} vs {g_row[1]} (exact)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.5):
                    errors.append(f"{key}.Avg_Price: {a_row[2]} vs {g_row[2]} (tol=0.5)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.5):
                    errors.append(f"{key}.Avg_Cost: {a_row[3]} vs {g_row[3]} (tol=0.5)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Avg_Margin: {a_row[4]} vs {g_row[4]} (tol=0.5)")

        # Sort order check: by Avg_Margin desc
        if a_data and len(a_data) >= 2:
            a_margins = []
            for r in a_data:
                if r and len(r) >= 5 and r[4] is not None:
                    try:
                        a_margins.append(float(r[4]))
                    except Exception:
                        pass
            if len(a_margins) >= 2:
                if a_margins != sorted(a_margins, reverse=True):
                    errors.append(f"Product Categories not sorted by Avg_Margin desc: {a_margins[:5]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if is_numeric(g_row[1]):
                    # Counts exact, averages ±0.05
                    tol = 0 if "count" in key or "total" in key else 0.05
                    if not num_close(a_row[1], g_row[1], tol):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
                else:
                    if str(a_row[1]).strip().lower() != str(g_row[1]).strip().lower():
                        errors.append(f"{key}.Value: '{a_row[1]}' vs '{g_row[1]}'")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    # Google Sheet check
    print(f"  Checking Google Sheet...")
    gs_errors = []
    check_gsheet(gs_errors)
    all_errors.extend(gs_errors)
    if gs_errors:
        for e in gs_errors:
            print(f"    {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
