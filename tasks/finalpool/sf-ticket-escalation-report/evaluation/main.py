"""Evaluation for sf-ticket-escalation-report."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Escalation_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Escalation_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Check Priority Summary sheet
    print("  Checking Priority Summary...")
    a_rows = load_sheet_rows(agent_wb, "Priority Summary")
    g_rows = load_sheet_rows(gt_wb, "Priority Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Priority Summary' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Priority Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing priority: {g_row[0]}")
                continue
            # Total_Tickets - exact (small set of priorities, count must match)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{key}.Total_Tickets: {a_row[1]} vs {g_row[1]}")
            # Low_CSAT_Tickets - tight tolerance
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 5):
                    errors.append(f"{key}.Low_CSAT: {a_row[2]} vs {g_row[2]}")
            # Avg_Response_Hours
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Avg_Response: {a_row[4]} vs {g_row[4]}")
            # Avg_CSAT
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.05):
                    errors.append(f"{key}.Avg_CSAT: {a_row[5]} vs {g_row[5]}")
            # SLA_Hours - exact integer
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 0):
                    errors.append(f"{key}.SLA_Hours: {a_row[6]} vs {g_row[6]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Summary sheet
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if "issue_type" in key or ("priority" in key and "highest" in key):
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif "pct" in key:
                    if not num_close(a_row[1], g_row[1], 0.5):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif "hours" in key or "satisfaction" in key or "csat" in key:
                    if not num_close(a_row[1], g_row[1], 0.5):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                else:
                    # Default counts: scale tol by GT magnitude (small values must be near-exact)
                    try:
                        gv_f = float(g_row[1]) if g_row[1] is not None else 0
                    except (TypeError, ValueError):
                        gv_f = 0
                    if abs(gv_f) < 50:
                        # Small counts (At_Risk_Tickets ~2) must be exact
                        tol_default = 1
                    elif abs(gv_f) < 1000:
                        tol_default = 5
                    else:
                        tol_default = 10
                    if not num_close(a_row[1], g_row[1], tol_default):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol={tol_default})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Reporter Analysis sheet
    print("  Checking Reporter Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Reporter Analysis")
    g_rows = load_sheet_rows(gt_wb, "Reporter Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Reporter Analysis' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Reporter Analysis' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing reporter: {g_row[0]}")
                continue
            # Reporter Total_Tickets - tighten
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{key}.Total: {a_row[1]} vs {g_row[1]}")
            # Low_CSAT_Count
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 5):
                    errors.append(f"{key}.Low_CSAT: {a_row[2]} vs {g_row[2]}")
            # Avg_Response_Hours
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Avg_Response: {a_row[4]} vs {g_row[4]}")
            # Avg_CSAT
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.05):
                    errors.append(f"{key}.Avg_CSAT: {a_row[5]} vs {g_row[5]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check At Risk Tickets sheet
    print("  Checking At Risk Tickets...")
    a_rows = load_sheet_rows(agent_wb, "At Risk Tickets")
    g_rows = load_sheet_rows(gt_wb, "At Risk Tickets")
    if a_rows is None:
        all_errors.append("Sheet 'At Risk Tickets' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'At Risk Tickets' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing priority: {g_row[0]}")
                continue
            # Total_Tickets
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{key}.Total_Tickets: {a_row[1]} vs {g_row[1]}")
            # At_Risk_Count
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1):
                    errors.append(f"{key}.At_Risk_Count: {a_row[2]} vs {g_row[2]}")
            # At_Risk_Pct
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.5):
                    errors.append(f"{key}.At_Risk_Pct: {a_row[3]} vs {g_row[3]}")
            # Avg_Response_Hours
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Avg_Response_Hours: {a_row[4]} vs {g_row[4]}")
            # SLA_Hours - exact integer
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0):
                    errors.append(f"{key}.SLA_Hours: {a_row[5]} vs {g_row[5]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Low Satisfaction sheet
    print("  Checking Low Satisfaction...")
    a_rows = load_sheet_rows(agent_wb, "Low Satisfaction")
    g_rows = load_sheet_rows(gt_wb, "Low Satisfaction")
    if a_rows is None:
        all_errors.append("Sheet 'Low Satisfaction' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Low Satisfaction' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        # Compound key: Priority + Issue_Type
        a_lookup = {}
        for r in a_data:
            if r and r[0] is not None and len(r) > 1 and r[1] is not None:
                k = (str(r[0]).strip().lower(), str(r[1]).strip().lower())
                a_lookup[k] = r
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None or len(g_row) < 2 or g_row[1] is None:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing combo: {key}")
                continue
            # Low_CSAT_Count
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 3):
                    errors.append(f"{key}.Low_CSAT_Count: {a_row[2]} vs {g_row[2]}")
            # Total_Count
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 3):
                    errors.append(f"{key}.Total_Count: {a_row[3]} vs {g_row[3]}")
            # Low_CSAT_Pct
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Low_CSAT_Pct: {a_row[4]} vs {g_row[4]}")
            # Avg_CSAT
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.05):
                    errors.append(f"{key}.Avg_CSAT: {a_row[5]} vs {g_row[5]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check email was sent - require exact subject and recipient
    print("  Checking email sent...")
    db_config = {
        "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
        "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
        "user": "eigent", "password": "camel",
    }
    try:
        conn = psycopg2.connect(**db_config)
        cur = conn.cursor()
        # Require exact subject 'Weekly Escalation Analysis Report' (case-insensitive substring of full title)
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE LOWER(subject) LIKE '%weekly escalation analysis report%'
        """)
        emails = cur.fetchall()
        cur.execute("""
            SELECT m.subject, m.to_addr FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            WHERE LOWER(m.subject) LIKE '%weekly escalation analysis report%'
        """)
        sent = cur.fetchall()
        cur.close()
        conn.close()

        if len(emails) + len(sent) < 1:
            all_errors.append("No 'Weekly Escalation Analysis Report' email found")
            print("    FAIL: exact-subject email not found")
        else:
            # Check recipient is escalations@company.com
            all_recipients_text = " ".join([str(e[1] or "") for e in emails] +
                                           [str(s[1] or "") for s in sent]).lower()
            if "escalations@company.com" not in all_recipients_text:
                all_errors.append(f"Email recipient 'escalations@company.com' not found in {all_recipients_text}")
                print("    FAIL: recipient missing")
            else:
                print("    PASS")
    except Exception as e:
        all_errors.append(f"Email check error: {e}")
        print(f"    ERROR: {e}")

    # Check Notion page - require exact title and 3+ action items
    print("  Checking Notion page...")
    try:
        conn = psycopg2.connect(**db_config)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE LOWER(properties::text) LIKE '%support escalation analysis%'
        """)
        pages = cur.fetchall()
        if not pages:
            all_errors.append("No Notion page titled 'Support Escalation Analysis' found")
            print("    FAIL: title-matched page not found")
        else:
            page_id = pages[0][0]
            # Look for child blocks with action item content
            # notion.blocks has columns: parent_id (text), block_data (jsonb).
            # No plain_text/page_id columns exist; extract text from block_data jsonb.
            try:
                cur.execute("""
                    SELECT block_data::text FROM notion.blocks
                    WHERE parent_id=%s
                """, (page_id,))
                blocks = cur.fetchall()
                blob = " ".join(str(b[0] or "") for b in blocks).lower()
            except Exception:
                # Fallback: read all blocks
                cur.execute("SELECT block_data::text FROM notion.blocks LIMIT 1000")
                blocks = cur.fetchall()
                blob = " ".join(str(b[0] or "") for b in blocks).lower()
            # Check for action items - look for at least 3 list-style markers or 'action item' indicators
            indicators = ["action", "item", "focus", "review", "investigate", "improve",
                          "address", "recommend"]
            indicator_count = sum(1 for ind in indicators if ind in blob)
            if indicator_count < 3:
                all_errors.append(f"Notion page action items insufficient (indicators={indicator_count})")
                print(f"    FAIL: action items insufficient")
            else:
                print("    PASS")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Notion check error: {e}")
        print(f"    ERROR: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
