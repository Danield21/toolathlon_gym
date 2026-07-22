#!/usr/bin/env python3
"""Evaluation script for research-grant-proposal-coordination.

Verifies the 6 phases of the task:
  - Researcher database (CSV/JSON) listing 3+ researchers with credentials
  - Notion workspace with proposal sections (overview, methodology, team, timeline, budget)
  - Main proposal Word doc covering required sections
  - Compliance checklist
  - Calendar event invitation
  - Email submissions/communications
"""

from argparse import ArgumentParser
import json
import os
import sys
from pathlib import Path
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

EXPECTED_RESEARCHERS = ["lee", "wang", "kumar"]  # from literature_sources.csv
EXPECTED_SECTIONS = ["project overview", "research methodology", "team", "timeline", "budget"]

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, cond, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if cond:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _read_text_any(path):
    p = Path(path)
    if not p.exists():
        return ""
    try:
        ext = p.suffix.lower()
        if ext == ".docx":
            from docx import Document
            return "\n".join(par.text for par in Document(str(p)).paragraphs)
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(str(p), data_only=True)
            txts = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            txts.append(str(cell))
            return " ".join(txts)
        if ext in (".txt", ".md", ".csv", ".json", ".tsv"):
            return p.read_text(errors="ignore")
    except Exception:
        return ""
    return ""


def check_researcher_db(workspace):
    """Find a researcher database file with at least 3 researchers."""
    base = Path(workspace)
    candidates = []
    for pat in ["*researcher*", "*team*", "*roster*", "*cv*", "*contributors*"]:
        for ext in ["csv", "xlsx", "json", "tsv", "md"]:
            candidates.extend(base.rglob(f"{pat}.{ext}"))
    check("Researcher database file exists",
          len(candidates) >= 1, f"found {len(candidates)} candidates")
    if not candidates:
        return
    text = "\n".join(_read_text_any(str(c)) for c in candidates).lower()
    found = [r for r in EXPECTED_RESEARCHERS if r in text]
    check("Researcher database contains at least 3 expected researchers",
          len(found) >= 3, f"found: {found}")


def check_proposal_doc(workspace):
    """Verify main proposal docx contains the required sections."""
    base = Path(workspace)
    candidates = []
    for pat in ["*proposal*", "*main_proposal*", "*grant*"]:
        candidates.extend(base.rglob(f"{pat}.docx"))
    # Filter out the template file from initial_workspace
    candidates = [c for c in candidates if "template" not in c.name.lower()]
    check("Main proposal Word document exists",
          len(candidates) >= 1, f"found {len(candidates)} files (excl. template)")
    if not candidates:
        return
    text = "\n".join(_read_text_any(str(c)) for c in candidates).lower()
    check("Proposal doc has substantive content (>=600 chars)",
          len(text) >= 600, f"length={len(text)}")
    missing = []
    for sect in EXPECTED_SECTIONS:
        if sect not in text:
            # try alt phrasings
            alt = {
                "project overview": ["overview", "project summary"],
                "research methodology": ["methodology", "methods"],
                "team": ["team composition", "researchers", "personnel"],
                "timeline": ["schedule", "milestones"],
                "budget": ["funding", "cost"],
            }.get(sect, [])
            if not any(a in text for a in alt):
                missing.append(sect)
    check("Proposal doc covers all required sections",
          len(missing) == 0, f"missing sections: {missing}")


def check_compliance_checklist(workspace):
    """Find a compliance checklist with funder-required topics."""
    base = Path(workspace)
    candidates = []
    for pat in ["*compliance*", "*checklist*", "*requirements*", "*submission_check*"]:
        for ext in ["csv", "xlsx", "json", "txt", "md", "docx"]:
            candidates.extend(base.rglob(f"{pat}.{ext}"))
    check("Compliance checklist file exists",
          len(candidates) >= 1, f"found {len(candidates)}")
    if not candidates:
        return
    text = "\n".join(_read_text_any(str(c)) for c in candidates).lower()
    must = ["page", "format", "citation", "deadline"]
    missing = [m for m in must if m not in text]
    check("Compliance checklist mentions page/format/citation/deadline topics",
          len(missing) <= 1, f"missing: {missing}")


def check_notion():
    """Verify Notion workspace has proposal sections."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        check("Notion DB reachable", False, str(e))
        return
    titles = []
    for _id, props in rows:
        if not props:
            continue
        try:
            t = props.get("title", {}) if isinstance(props, dict) else {}
            arr = t.get("title", [])
            for el in arr:
                content = el.get("text", {}).get("content", "")
                if content:
                    titles.append(content.lower())
        except Exception:
            pass
    # All proposal section pages present
    text = " | ".join(titles)
    section_kws = {
        "overview": ["overview", "project overview"],
        "methodology": ["methodology", "method"],
        "team": ["team"],
        "timeline": ["timeline", "schedule"],
        "budget": ["budget"],
    }
    found_sections = [k for k, kws in section_kws.items() if any(kw in t for kw in kws for t in titles)]
    check("Notion workspace has at least 3 proposal section pages",
          len(found_sections) >= 3,
          f"found sections in titles: {found_sections}; titles: {titles[:8]}")


def check_calendar(launch_time):
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, summary, COALESCE(description, ''), attendees
            FROM gcal.events
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        check("Calendar DB reachable", False, str(e))
        return
    matched = None
    for _id, summary, desc, attendees in rows:
        text = ((summary or "") + " " + (desc or "")).lower()
        if any(k in text for k in ["proposal", "grant", "review draft", "draft review"]):
            matched = (summary, desc, attendees)
            break
    check("Calendar event scheduled for grant proposal review meeting",
          matched is not None, f"checked {len(rows)} events; none matched")


def check_emails():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, from_addr, to_addr, COALESCE(body_text, body_html, '')
            FROM email.messages
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        check("Email DB reachable", False, str(e))
        return
    # Researcher submission emails OR coordination email
    matched = None
    for subj, frm, to, body in rows:
        text = ((subj or "") + " " + (body or "")).lower()
        if any(k in text for k in ["proposal", "grant", "cv", "submission"]):
            matched = (subj, frm, to)
            break
    check("Grant proposal email communication exists",
          matched is not None, f"checked {len(rows)} emails")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0; FAIL_COUNT = 0
    if not agent_workspace or not Path(agent_workspace).exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    check_researcher_db(agent_workspace)
    check_proposal_doc(agent_workspace)
    check_compliance_checklist(agent_workspace)
    check_notion()
    check_calendar(launch_time)
    check_emails()

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
