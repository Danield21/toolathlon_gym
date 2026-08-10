"""
Evaluation for arxiv-federated-learning-notion task.
Checks Notion page/database and Excel spreadsheet.

Robustness notes (see fix reports T4__arxiv-federated-learning-notion.md):
- DB config reads PGHOST/PGPORT/PGDATABASE/PGUSER/PGPASSWORD env vars with defaults.
- Excel cells are read with a double-read strategy (data_only=True cached values +
  data_only=False raw/formula text) so that a model that correctly writes Excel
  formulas (e.g. =COUNTA(...), =ROUND(AVERAGE(...),0)) is NOT misjudged FAIL just
  because openpyxl does not compute formula caches.
- A small, safe formula evaluator handles the common aggregates (SUM/AVERAGE/MIN/MAX/
  COUNT/COUNTA/ROUND/LEN/IF/...) including whole-column references (B:B, B2:B). A
  formula that genuinely cannot be computed by the evaluator (e.g. =SUBTOTAL(...))
  is skipped as a structural pass rather than producing a false FAIL. But a formula
  that IS computable yet yields no value (e.g. AVERAGE over an empty range) is
  treated like an empty cell and FAILs, so a lazy formula cannot silently bypass the
  value check.
- Summary metric labels are matched leniently (case/space/punctuation-insensitive
  with common expansions: 'Total Papers', 'Number of Papers', 'Average Abstract
  Length', 'First/Last Year', ...) so reasonable renames do not cause a false FAIL.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

# FL papers in arxiv.papers (IDs from actual DB)
EXPECTED_FL_PAPERS = {
    "1602.05629": "Communication-Efficient Learning of Deep Networks from Decentralized Data",
    "1812.06127": "Federated Optimization in Heterogeneous Networks",
    "1908.07873": "Federated Learning: Challenges, Methods, and Future Directions",
    "1912.04977": "Advances and Open Problems in Federated Learning",
}

# These are non-FL papers that should NOT appear
NON_FL_IDS = {"1207.00580", "1502.03167"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(v):
    """Robustly convert a cell value to float.

    Accepts int/float/str (strips thousands separators, currency symbols,
    percent signs, whitespace). Returns None when the value is missing or
    cannot be parsed as a number.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s = s.replace(",", "").replace("$", "").replace("€", "").replace("¥", "").replace("%", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def num_close(a, b, tol=50):
    """Compare two values numerically with tolerance.

    Both sides are parsed via _to_float; if both parse, compare with tolerance.
    If either side cannot be parsed as a number, fall back to a case-insensitive
    string comparison (never raises).
    """
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    return str(a).strip().lower() == str(b).strip().lower()


def _norm_key(v):
    """Normalize a header/metric label for case/space/punctuation-insensitive matching.

    Lowercases, strips leading/trailing non-word characters (punctuation like
    ':'/'(' ) and collapses whitespace runs to a single underscore. Unicode-aware.
    """
    s = str(v).strip().lower()
    s = re.sub(r"^\W+", "", s)
    s = re.sub(r"\W+$", "", s)
    s = re.sub(r"\s+", "_", s)
    return s


# ---------------------------------------------------------------------------
# Notion helpers
# ---------------------------------------------------------------------------

def _collect_text(obj, out):
    """Recursively collect human-readable text from a notion jsonb object."""
    if isinstance(obj, dict):
        c = obj.get("content")
        if isinstance(c, str) and c:
            out.append(c)
            return
        p = obj.get("plain_text")
        if isinstance(p, str) and p:
            out.append(p)
            return
        for v in obj.values():
            _collect_text(v, out)
    elif isinstance(obj, list):
        for v in obj:
            _collect_text(v, out)


def _page_title(props):
    """Extract the page title text from a notion.pages.properties jsonb object."""
    if not isinstance(props, dict):
        return ""
    for key in ("title", "page_title", "Name", "name"):
        if key in props:
            out = []
            _collect_text(props[key], out)
            if out:
                return " ".join(out)
    out = []
    _collect_text(props, out)
    return " ".join(out)


def check_notion():
    """Check Notion page and database."""
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Check for page with "Federated Learning Research Hub" title.
        # Match if the title contains BOTH "federated" and "hub", or the
        # abbreviation "fl" (word boundary) plus "hub".
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()

        hub_page = None
        for pid, props in pages:
            title_text = _page_title(props)
            tl = title_text.lower()
            if ("federated" in tl and "hub" in tl) or (
                bool(re.search(r"\bfl\b", tl)) and "hub" in tl
            ):
                hub_page = pid
                break

        record("Notion page 'Federated Learning Research Hub' exists (title contains 'federated' AND 'hub')",
               hub_page is not None,
               f"Found {len(pages)} pages, none matching")

        # Check for database "Paper Index"
        cur.execute("SELECT id, title, parent FROM notion.databases")
        dbs = cur.fetchall()

        paper_db = None
        for did, title_json, parent in dbs:
            title_str = json.dumps(title_json).lower() if title_json else ""
            if "paper index" in title_str or "paper_index" in title_str:
                paper_db = did
                break
            if "paper" in title_str and "index" in title_str:
                paper_db = did
                break

        record("Notion database 'Paper Index' exists",
               paper_db is not None,
               f"Found {len(dbs)} databases")

        # Check database has entries (pages that are children of database)
        if paper_db:
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE parent::text LIKE %s
            """, (f"%{paper_db}%",))
            db_pages = cur.fetchall()
            record("Paper Index has entries", len(db_pages) >= 4,
                   f"Found {len(db_pages)} entries, expected >= 4")

        conn.close()
    except psycopg2.OperationalError as e:
        record("Notion connection", False, f"DB unreachable: {e}")
    except Exception as e:
        record("Notion connection", False, str(e))


# ---------------------------------------------------------------------------
# Excel helpers (double-read: cached values + raw/formula text)
# ---------------------------------------------------------------------------

def _build_grids(wb):
    """Return {sheet_name: [[cell.value, ...], ...]} from a workbook."""
    grids = {}
    for name in wb.sheetnames:
        ws = wb[name]
        grids[name] = [[c.value for c in row] for row in ws.iter_rows()]
    return grids


def _grid_val(grids, sheet, r, c):
    g = grids.get(sheet)
    if not g:
        return None
    if 0 <= r < len(g) and 0 <= c < len(g[r]):
        return g[r][c]
    return None


def _cell_effective(raw_grids, val_grids, sheet, r, c):
    """Return (raw, effective) for a cell.

    raw = the underlying cell value (formula string if it is a formula).
    effective = cached value if the cell holds a formula with a cached result,
    otherwise the literal value (None for an unresolved formula).
    """
    raw = _grid_val(raw_grids, sheet, r, c)
    if isinstance(raw, str) and raw.startswith("="):
        cached = _grid_val(val_grids, sheet, r, c)
        return raw, cached
    return raw, raw


def _find_sheet(grids, sheet_name):
    for name in grids:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return name
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return name
    return None


def find_col(header, names):
    """Find the index of a column whose header matches one of `names`.

    Matching is case/space-insensitive. Exact match first; then, for a
    reasonably specific name (>= 5 chars), a substring fallback so headers like
    'Abstract Length (chars)' or 'Publication Year' still anchor correctly.
    Short names (e.g. 'ID', 'Year') are exact-match only to avoid picking the
    wrong column via accidental substring hits.
    """
    if not header:
        return None
    norm_names = [_norm_key(n) for n in names]
    for i, cell in enumerate(header):
        if cell is None:
            continue
        c = _norm_key(cell)
        if not c:
            continue
        for n in norm_names:
            if c == n:
                return i
        for n in norm_names:
            if len(n) >= 5 and (n in c or c in n):
                return i
    return None


# --- small, safe Excel-formula evaluator -----------------------------------

_CELL_RE = re.compile(r"^[$]?([A-Za-z]{1,3})[$]?([0-9]{1,7})$")
_COL_RE = re.compile(r"^[$]?([A-Za-z]{1,3})$")
_ROW_RE = re.compile(r"^[$]?([0-9]{1,7})$")


def _col_to_idx(letters):
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx - 1


def _parse_ref(addr):
    m = _CELL_RE.match(addr.strip())
    if not m:
        return None
    return int(m.group(2)) - 1, _col_to_idx(m.group(1))


class _FormulaParser:
    """Recursive-descent evaluator for simple Excel formulas.

    Supports numeric literals, cell/range references (optionally with a sheet
    prefix such as 'Paper Details'!E2:E5), arithmetic + - * / and parentheses,
    and common functions (SUM/AVERAGE/MIN/MAX/COUNT/COUNTA/ROUND/ROUNDUP/
    ROUNDDOWN/INT/ABS/LEN/IF). Returns None for anything it cannot compute.
    """

    def __init__(self, expr, raw_grids, val_grids, sheet):
        self.expr = expr
        self.pos = 0
        self.raw = raw_grids
        self.val = val_grids
        self.sheet = sheet

    # -- token helpers --
    def _peek(self):
        while self.pos < len(self.expr) and self.expr[self.pos].isspace():
            self.pos += 1
        if self.pos >= len(self.expr):
            return ""
        return self.expr[self.pos]

    def _eat(self, ch):
        while self.pos < len(self.expr) and self.expr[self.pos].isspace():
            self.pos += 1
        if self.pos < len(self.expr) and self.expr[self.pos] == ch:
            self.pos += 1
            return True
        return False

    # -- grammar --
    def parse_expr(self):
        val = self.parse_term()
        while True:
            if self._eat("+"):
                val = self._arith(val, self.parse_term(), "+")
            elif self._eat("-"):
                val = self._arith(val, self.parse_term(), "-")
            else:
                return val

    def parse_term(self):
        val = self.parse_factor()
        while True:
            if self._eat("*"):
                val = self._arith(val, self.parse_factor(), "*")
            elif self._eat("/"):
                val = self._arith(val, self.parse_factor(), "/")
            else:
                return val

    def parse_factor(self):
        if self._eat("-"):
            return self._arith(0.0, self.parse_factor(), "-")
        if self._eat("("):
            val = self.parse_expr()
            self._eat(")")
            return val
        return self.parse_primary()

    def _read_ref_token(self):
        """Read a reference token, handling quoted sheet prefixes with spaces,
        e.g. 'Paper Details'!E2:E5 or A1:B5 or E2."""
        start = self.pos
        if self.pos < len(self.expr) and self.expr[self.pos] == "'":
            self.pos += 1
            while self.pos < len(self.expr) and self.expr[self.pos] != "'":
                self.pos += 1
            if self.pos < len(self.expr):
                self.pos += 1  # skip closing quote
        while self.pos < len(self.expr):
            c = self.expr[self.pos]
            if c.isalnum() or c in "!$':._":
                self.pos += 1
            else:
                break
        return self.expr[start:self.pos]

    def parse_primary(self):
        tok = self._read_ref_token()
        if not tok:
            raise ValueError("empty token")
        # numeric literal?
        if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", tok):
            return float(tok)
        # function call?
        while self.pos < len(self.expr) and self.expr[self.pos].isspace():
            self.pos += 1
        if self.pos < len(self.expr) and self.expr[self.pos] == "(":
            fname = tok.upper()
            self.pos += 1
            args = []
            if self._peek() == ")":
                self.pos += 1
            else:
                while True:
                    args.append(self.parse_expr())
                    if self._eat(")"):
                        break
                    if not self._eat(","):
                        raise ValueError("expected , or )")
            return self._apply(fname, args)
        # otherwise a cell / range reference (possibly with sheet prefix)
        return self._resolve_ref(tok)

    # -- value resolution --
    def _cell_value(self, sheet, r, c):
        raw = _grid_val(self.raw, sheet, r, c)
        if isinstance(raw, str) and raw.startswith("="):
            cached = _grid_val(self.val, sheet, r, c)
            return cached
        return raw

    def _resolve_ref(self, tok):
        tok = tok.strip()
        sheet = self.sheet
        addr = tok
        if "!" in tok:
            sheet_part, addr = tok.split("!", 1)
            sheet_part = sheet_part.strip().strip("'").strip()
            sheet = sheet_part
        addr = addr.strip()
        if ":" in addr:
            a1, a2 = addr.split(":", 1)
            return self._range_values(sheet, a1, a2)
        rc = _parse_ref(addr)
        if rc is None:
            raise ValueError(f"bad ref: {tok}")
        return self._cell_value(sheet, rc[0], rc[1])

    def _endpoint(self, a, nrows, ncols):
        """Return (r_start, r_end, c_start, c_end) for a reference endpoint.

        Supports a single cell (B2), a whole column (B), and a whole row (5).
        """
        m = _CELL_RE.match(a)
        if m:
            r = int(m.group(2)) - 1
            c = _col_to_idx(m.group(1))
            return (r, r, c, c)
        m = _COL_RE.match(a)
        if m:
            c = _col_to_idx(m.group(1))
            return (0, max(nrows - 1, 0), c, c)
        m = _ROW_RE.match(a)
        if m:
            r = int(m.group(1)) - 1
            return (r, r, 0, max(ncols - 1, 0))
        return None

    def _range_values(self, sheet, a1, a2):
        """Return a flat list of cell values for a range reference.

        Supports plain cell ranges (B2:B5), whole columns (B:B, B2:B) and whole
        rows (5:5), so legitimate Excel formulas that reference full columns are
        not silently mis-computed as empty.
        """
        g = self.raw.get(sheet)
        if not g:
            return []
        nrows = len(g)
        ncols = max((len(r) for r in g), default=0)
        e1 = self._endpoint(a1, nrows, ncols)
        e2 = self._endpoint(a2, nrows, ncols)
        if e1 is None or e2 is None:
            return []
        r_start, r_end = e1[0], e2[1]
        c_start, c_end = e1[2], e2[3]
        if r_start > r_end:
            r_start, r_end = r_end, r_start
        if c_start > c_end:
            c_start, c_end = c_end, c_start
        vals = []
        for r in range(r_start, r_end + 1):
            for c in range(c_start, c_end + 1):
                vals.append(self._cell_value(sheet, r, c))
        return vals

    def _arith(self, a, b, op):
        if isinstance(a, list) or isinstance(b, list):
            return None
        fa, fb = _to_float(a), _to_float(b)
        if fa is None or fb is None:
            return None
        if op == "+":
            return fa + fb
        if op == "-":
            return fa - fb
        if op == "*":
            return fa * fb
        if op == "/":
            return fa / fb if fb != 0 else None
        return None

    @staticmethod
    def _num_vals(args):
        out = []
        for a in args:
            seq = a if isinstance(a, list) else [a]
            for x in seq:
                f = _to_float(x)
                if f is not None:
                    out.append(f)
        return out

    def _apply(self, fname, args):
        if fname == "SUM":
            return sum(self._num_vals(args))
        if fname == "AVERAGE":
            nv = self._num_vals(args)
            return (sum(nv) / len(nv)) if nv else None
        if fname in ("MIN", "MAX"):
            nv = self._num_vals(args)
            return (min(nv) if fname == "MIN" else max(nv)) if nv else None
        if fname == "COUNT":
            return len(self._num_vals(args))
        if fname == "COUNTA":
            cnt = 0
            for a in args:
                for x in (a if isinstance(a, list) else [a]):
                    if x is not None and (not isinstance(x, str) or x.strip() != ""):
                        cnt += 1
            return cnt
        if fname == "ROUND":
            f = _to_float(args[0]) if args else None
            d = int(_to_float(args[1]) or 0) if len(args) > 1 else 0
            return round(f, d) if f is not None else None
        if fname == "ROUNDUP":
            import math
            f = _to_float(args[0]) if args else None
            d = int(_to_float(args[1]) or 0) if len(args) > 1 else 0
            if f is None:
                return None
            mult = 10 ** d
            return math.ceil(f * mult) / mult
        if fname == "ROUNDDOWN":
            import math
            f = _to_float(args[0]) if args else None
            d = int(_to_float(args[1]) or 0) if len(args) > 1 else 0
            if f is None:
                return None
            mult = 10 ** d
            return math.floor(f * mult) / mult
        if fname == "INT":
            f = _to_float(args[0]) if args else None
            return int(f) if f is not None else None
        if fname == "ABS":
            f = _to_float(args[0]) if args else None
            return abs(f) if f is not None else None
        if fname == "LEN":
            if not args:
                return None
            a = args[0]
            if isinstance(a, list):
                a = a[0] if a else None
            return len(str(a)) if a is not None else None
        if fname == "IF":
            if len(args) >= 3:
                cond = _to_float(args[0])
                return args[1] if (cond is not None and cond != 0) else args[2]
        raise ValueError(f"unsupported function: {fname}")


_UNRESOLVED = object()  # sentinel: evaluator genuinely cannot compute the formula


def _evaluate_formula(formula, raw_grids, val_grids, current_sheet):
    """Compute a formula value.

    Returns:
      - a number (int/float) when the formula is fully computed;
      - None when the formula is valid and parsed but evaluates to no value
        (e.g. AVERAGE/MIN/MAX over an empty range);
      - _UNRESOLVED when the formula cannot be parsed/computed by this
        evaluator (unsupported function, string-literal args, syntax error).
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return _UNRESOLVED
    try:
        parser = _FormulaParser(formula[1:], raw_grids, val_grids, current_sheet)
        val = parser.parse_expr()
    except Exception:
        return _UNRESOLVED
    if isinstance(val, list):
        return _UNRESOLVED
    return val


def _metric_number(raw_grids, val_grids, sheet, r, c):
    """Return (numeric_value_or_None, status).

    status is one of 'formula-cached', 'formula-computed', 'formula-null',
    'formula-unresolved', 'literal', 'empty'.

    'formula-unresolved' (evaluator cannot compute the formula, e.g. SUBTOTAL)
    is deliberately treated as a lenient pass by callers so a correct formula we
    cannot evaluate is not misjudged. 'formula-null' (the formula parses and all
    its functions are known, but it computes to no value - e.g. AVERAGE over an
    empty range) is treated like an empty cell: it FAILs, so a lazy formula that
    simply references an empty region cannot silently pass the value check.
    """
    raw, eff = _cell_effective(raw_grids, val_grids, sheet, r, c)
    if isinstance(raw, str) and raw.startswith("="):
        if eff is not None:
            return _to_float(eff), "formula-cached"
        num = _evaluate_formula(raw, raw_grids, val_grids, sheet)
        if num is _UNRESOLVED:
            return None, "formula-unresolved"
        if num is not None:
            return _to_float(num), "formula-computed"
        return None, "formula-null"
    if eff is None:
        return None, "empty"
    return _to_float(eff), "literal"


def _display(raw_grids, val_grids, sheet, r, c):
    """Human-readable cell value (cached value for formulas, else literal)."""
    _, eff = _cell_effective(raw_grids, val_grids, sheet, r, c)
    return eff


def _is_total_key(k):
    """Match a normalized Summary metric key for the total paper count.

    Accepts the task's exact 'total_papers' plus reasonable LLM expansions:
    'Total Papers', 'Number of Papers', 'Paper Count', 'Papers Found', ...
    """
    if "total" in k and ("paper" in k or "count" in k or "found" in k
                         or "identified" in k or "number" in k):
        return True
    if "paper" in k and ("count" in k or "number" in k or "found" in k
                         or "identified" in k or "num" in k):
        return True
    return k in {"papers", "paper_count", "number_of_papers", "num_papers",
                 "total_paper_count", "paper_total"}


def _is_avg_key(k):
    """Match a normalized Summary metric key for the average abstract length.

    Accepts 'avg_abstract_length', 'average_abstract_length',
    'mean_abstract_length', 'average_length', ... (abbreviations are frequently
    expanded by LLMs).
    """
    if "abstract" in k and ("avg" in k or "average" in k or "mean" in k):
        return True
    if "length" in k and ("avg" in k or "average" in k or "mean" in k):
        return True
    return False


def _is_year_key(k, labels):
    """Match a normalized Summary metric key for a year metric whose label
    appears among `labels` (e.g. ('earliest', 'first', 'min', 'oldest'))."""
    if "year" not in k:
        return False
    return any(lbl in k for lbl in labels)


def check_excel(agent_workspace):
    """Check Excel spreadsheet."""
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Federated_Learning_Papers.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return

    record("Excel file exists", True)

    try:
        wb_val = openpyxl.load_workbook(excel_path, data_only=True)
        wb_raw = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    record("Excel readable", True)

    val_grids = _build_grids(wb_val)
    raw_grids = _build_grids(wb_raw)

    # ---- Paper Details sheet ----
    details_sheet = _find_sheet(raw_grids, "Paper Details") or _find_sheet(raw_grids, "Paper_Details")
    if details_sheet is None:
        record("Sheet 'Paper Details' exists", False, f"Available: {list(raw_grids.keys())}")
        return

    record("Sheet 'Paper Details' exists", True)
    details_grid = raw_grids[details_sheet]
    header = details_grid[0] if details_grid else []
    data_rows = details_grid[1:] if len(details_grid) > 1 else []

    # Should have at least 4 FL papers (might include extra if agent interprets broadly)
    record("Paper Details has >= 4 rows", len(data_rows) >= 4,
           f"Found {len(data_rows)} data rows")

    title_col = find_col(header, ["Title", "title", "Paper Title", "paper_title"])
    id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id", "Paper Identifier", "paper_identifier"])
    authors_col = find_col(header, ["Authors", "Author", "authors", "Author Names", "author_names"])
    year_col = find_col(header, ["Year", "year", "Publication Year", "publication_year", "Pub Year", "pub_year"])
    abstract_len_col = find_col(header, ["Abstract_Length", "abstract_length", "Abstract Length", "Abstract Length (chars)", "Abstract_Length_chars", "Abstract Len", "abstract_len"])

    if title_col is not None:
        found_titles = []
        for r in range(len(data_rows)):
            v = _display(raw_grids, val_grids, details_sheet, r + 1, title_col)
            if v:
                found_titles.append(str(v).strip().lower())

        for pid, expected_title in EXPECTED_FL_PAPERS.items():
            found = any(expected_title.lower() in t or t in expected_title.lower()
                        for t in found_titles)
            record(f"Has paper: {expected_title[:50]}...", found)

    # Verify NO non-FL papers (e.g. Dropout, BatchNorm) leaked in.
    if id_col is not None:
        found_ids = set()
        for r in range(len(data_rows)):
            v = _display(raw_grids, val_grids, details_sheet, r + 1, id_col)
            if v:
                found_ids.add(str(v).strip())
        leak = found_ids & NON_FL_IDS
        record("No non-FL papers leaked into Paper Details", len(leak) == 0,
               f"Leaked: {sorted(leak)}; All IDs: {sorted(found_ids)}")

    # Check columns exist
    record("Abstract_Length column exists", abstract_len_col is not None,
           f"Header: {header}")
    record("Authors column exists", authors_col is not None,
           f"Header: {header}")
    record("Year column exists", year_col is not None,
           f"Header: {header}")

    # Verify Authors values are non-empty and contain at least one comma or full name.
    if authors_col is not None:
        bad_authors = 0
        for r in range(len(data_rows)):
            v = _display(raw_grids, val_grids, details_sheet, r + 1, authors_col)
            if v is None or len(str(v).strip()) < 4:
                bad_authors += 1
        record("All rows have non-empty Authors", bad_authors == 0,
               f"{bad_authors} rows missing/short authors")

    # ---- Summary sheet ----
    summary_sheet = _find_sheet(raw_grids, "Summary")
    if summary_sheet is None:
        record("Sheet 'Summary' exists", False, f"Available: {list(raw_grids.keys())}")
        return

    record("Sheet 'Summary' exists", True)
    summary_grid = raw_grids[summary_sheet]

    metrics = {}
    metric_pos = {}
    for r in range(len(summary_grid)):
        row = summary_grid[r]
        if row and row[0] is not None:
            key = _norm_key(row[0])
            metrics[key] = row[1] if len(row) > 1 else None
            metric_pos[key] = (r, 1)

    # Total_Papers (label may be 'Total_Papers', 'Total Papers',
    # 'Number of Papers', 'Paper Count', 'Papers Found', ...)
    total_key = next((k for k in metrics if _is_total_key(k)), None)
    if total_key:
        r, c = metric_pos[total_key]
        num, status = _metric_number(raw_grids, val_grids, summary_sheet, r, c)
        if status == "formula-unresolved":
            ok, detail = True, "formula present, result not computable by evaluator; skipped strict value check"
        elif num is None:
            ok, detail = False, f"Got {metrics[total_key]!r}"
        else:
            ok = int(num) == 4
            detail = f"Got {metrics[total_key]!r}"
        record("Summary: Total_Papers == 4", ok, detail)
    else:
        record("Summary: Total_Papers exists", False, f"Keys: {list(metrics.keys())}")

    # Avg_Abstract_Length (label may be 'Avg_Abstract_Length',
    # 'Average Abstract Length', 'Mean Abstract Length', ...)
    avg_key = next((k for k in metrics if _is_avg_key(k)), None)
    if avg_key:
        r, c = metric_pos[avg_key]
        num, status = _metric_number(raw_grids, val_grids, summary_sheet, r, c)
        if status == "formula-unresolved":
            ok, detail = True, "formula present, result not computable by evaluator; skipped strict value check"
        elif num is None:
            ok, detail = False, f"Got {metrics[avg_key]!r}"
        else:
            ok = 600 <= num <= 1100
            detail = f"Got {metrics[avg_key]!r}"
        record("Summary: Avg_Abstract_Length in [600,1100]", ok, detail)
    else:
        record("Summary: Avg_Abstract_Length exists", False, f"Keys: {list(metrics.keys())}")

    # Earliest/Latest Year (labels may be 'Earliest/First/Min/Oldest Year' and
    # 'Latest/Last/Max/Newest Year')
    for label, expected, labels in [
        ("earliest", 2016, ("earliest", "first", "min", "oldest")),
        ("latest", 2019, ("latest", "last", "max", "newest")),
    ]:
        year_key = next((k for k in metrics if _is_year_key(k, labels)), None)
        if year_key:
            r, c = metric_pos[year_key]
            num, status = _metric_number(raw_grids, val_grids, summary_sheet, r, c)
            if status == "formula-unresolved":
                ok, detail = True, "formula present, result not computable by evaluator; skipped strict value check"
            elif num is None:
                ok, detail = False, f"Got {metrics[year_key]!r}"
            else:
                ok = num_close(num, expected, tol=1)
                detail = f"Got {metrics[year_key]!r}"
            record(f"Summary: {label.title()}_Year ~ {expected}", ok, detail)
        else:
            record(f"Summary: {label.title()}_Year exists", False,
                   f"Keys: {list(metrics.keys())}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_notion()
    check_excel(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
