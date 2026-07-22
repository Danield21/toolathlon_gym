"""Evaluation script for terminal-sf-wc-support-email-gcal-excel."""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_expected_ticket_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT "PRIORITY", COUNT(*) as cnt, ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY" ORDER BY cnt DESC
    """)
    rows = cur.fetchall()
    cur.execute('SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return rows, total


def get_expected_order_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT status, COUNT(*) FROM wc.orders GROUP BY status ORDER BY COUNT(*) DESC")
    rows = cur.fetchall()
    cur.execute("SELECT COUNT(*) FROM wc.orders")
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return rows, total


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Get expected data
    ticket_data, total_tickets = get_expected_ticket_data()
    order_data, total_orders = get_expected_order_data()

    # Check Excel
    excel_path = os.path.join(agent_workspace, "Product_Quality_Report.xlsx")
    check("Product_Quality_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # Ticket_by_Priority
        check("Ticket_by_Priority sheet exists", "Ticket_by_Priority" in wb.sheetnames)
        if "Ticket_by_Priority" in wb.sheetnames:
            ws = wb["Ticket_by_Priority"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Ticket_by_Priority has 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Priority', 'Ticket_Count', 'Avg_Response_Hours', 'Pct_of_Total']:
                check(f"Ticket_by_Priority has {col}", col.lower() in headers, f"headers: {headers[:5]}")

            # Verify ticket counts and avg response and pct match DB
            if data_rows:
                row_dict = {str(r[0]).strip(): r for r in data_rows if r[0]}
                for priority, cnt, avg_resp in ticket_data:
                    if priority in row_dict:
                        agent_cnt = safe_float(row_dict[priority][1])
                        if agent_cnt is not None:
                            check(f"Ticket count for {priority}", abs(agent_cnt - cnt) < 5,
                                  f"expected ~{cnt}, got {agent_cnt}")
                        agent_avg = safe_float(row_dict[priority][2])
                        if agent_avg is not None and avg_resp is not None:
                            check(f"Avg_Response_Hours for {priority}",
                                  abs(agent_avg - float(avg_resp)) <= 0.2,
                                  f"expected ~{avg_resp}, got {agent_avg}")
                        agent_pct = safe_float(row_dict[priority][3])
                        expected_pct = round(100.0 * cnt / total_tickets, 1) if total_tickets else 0
                        if agent_pct is not None:
                            check(f"Pct_of_Total for {priority}",
                                  abs(agent_pct - expected_pct) <= 0.5,
                                  f"expected ~{expected_pct}, got {agent_pct}")
                # Sort by ticket count descending
                counts_in_order = [safe_float(r[1]) for r in data_rows if r[0]]
                counts_in_order = [c for c in counts_in_order if c is not None]
                if counts_in_order:
                    check("Ticket_by_Priority sorted descending by Ticket_Count",
                          counts_in_order == sorted(counts_in_order, reverse=True),
                          f"got {counts_in_order}")

        # Order_Status_Breakdown — validate actual counts vs wc.orders
        check("Order_Status_Breakdown sheet exists", "Order_Status_Breakdown" in wb.sheetnames)
        if "Order_Status_Breakdown" in wb.sheetnames:
            ws = wb["Order_Status_Breakdown"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Order_Status_Breakdown has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Status', 'Order_Count', 'Quality_Flag']:
                check(f"Order_Status_Breakdown has {col}", col.lower() in headers,
                      f"headers: {headers[:5]}")

            # Build agent's status->count lookup
            status_idx = headers.index("status") if "status" in headers else 0
            count_idx = headers.index("order_count") if "order_count" in headers else 1
            agent_counts = {}
            for row in data_rows:
                if not row or row[status_idx] is None:
                    continue
                k = str(row[status_idx]).strip().lower()
                try:
                    agent_counts[k] = int(row[count_idx])
                except (TypeError, ValueError):
                    pass

            # Pull DB ground-truth
            try:
                conn_d = get_conn()
                cd = conn_d.cursor()
                cd.execute("SELECT status, COUNT(*) FROM wc.orders GROUP BY status")
                db_counts = {str(s).strip().lower(): int(c) for s, c in cd.fetchall()}
                cd.close()
                conn_d.close()
            except Exception:
                db_counts = {}

            for st, db_n in db_counts.items():
                a_n = agent_counts.get(st)
                check(f"Order_Status '{st}' count matches DB ({db_n})",
                      a_n is not None and a_n == db_n,
                      f"agent={a_n}, db={db_n}")

            # Check Review flags for refunded/cancelled
            for row in data_rows:
                status = str(row[0]).lower() if row[0] else ""
                flag = str(row[-1]).lower() if row[-1] else ""
                if status in ("refunded", "cancelled"):
                    check(f"Quality_Flag 'Review' for {status}", "review" in flag, f"got '{flag}'")

        # Quality_Action_Plan — content checks per row
        check("Quality_Action_Plan sheet exists", "Quality_Action_Plan" in wb.sheetnames)
        if "Quality_Action_Plan" in wb.sheetnames:
            ws = wb["Quality_Action_Plan"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Quality_Action_Plan has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Issue', 'Severity', 'Owner', 'Deadline', 'Action_Item']:
                check(f"Quality_Action_Plan has {col}", col.lower() in headers,
                      f"headers: {headers[:6]}")
            # Per-row content checks
            issue_i = headers.index("issue") if "issue" in headers else None
            sev_i = headers.index("severity") if "severity" in headers else None
            owner_i = headers.index("owner") if "owner" in headers else None
            action_i = headers.index("action_item") if "action_item" in headers else None
            valid_severities = {"high", "medium", "low"}
            sev_violations = 0
            owner_violations = 0
            action_violations = 0
            issue_violations = 0
            for row in data_rows:
                if not row:
                    continue
                if sev_i is not None and sev_i < len(row):
                    s = str(row[sev_i] or "").strip().lower()
                    if s not in valid_severities:
                        sev_violations += 1
                if owner_i is not None and owner_i < len(row):
                    o = str(row[owner_i] or "").strip()
                    if not o:
                        owner_violations += 1
                if action_i is not None and action_i < len(row):
                    a = str(row[action_i] or "").strip()
                    if not a:
                        action_violations += 1
                if issue_i is not None and issue_i < len(row):
                    i_v = str(row[issue_i] or "").strip()
                    if not i_v:
                        issue_violations += 1
            check("Quality_Action_Plan all Severity in {High,Medium,Low}",
                  sev_violations == 0, f"{sev_violations} bad severities")
            check("Quality_Action_Plan all Owner non-empty",
                  owner_violations == 0, f"{owner_violations} empty owners")
            check("Quality_Action_Plan all Action_Item non-empty",
                  action_violations == 0, f"{action_violations} empty actions")
            check("Quality_Action_Plan all Issue non-empty",
                  issue_violations == 0, f"{issue_violations} empty issues")

    # Check terminal script and outputs
    script_path = os.path.join(agent_workspace, "defect_correlation.py")
    check("defect_correlation.py exists", os.path.exists(script_path))
    if os.path.exists(script_path):
        with open(script_path) as f:
            script = f.read()
        # Script should reference both data sources and emit results
        check("Script references support_tickets.json", "support_tickets" in script,
              "missing reference")
        check("Script references order_data.json", "order_data" in script,
              "missing reference")
        check("Script outputs quality_analysis_results.json",
              "quality_analysis_results" in script, "missing output reference")
    results_path = os.path.join(agent_workspace, "quality_analysis_results.json")
    check("quality_analysis_results.json exists", os.path.exists(results_path))
    if os.path.exists(results_path):
        try:
            with open(results_path) as f:
                results = json.load(f)
            check("quality_analysis_results.json is valid JSON object",
                  isinstance(results, (dict, list)), f"type={type(results).__name__}")
        except Exception as e:
            check("quality_analysis_results.json parses", False, str(e))

    # Check emails — require exact subjects per task spec
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Email 1: exact "Product Quality Analysis Report Ready" -> quality-team
        cur.execute(
            "SELECT subject, to_addr FROM email.messages "
            "WHERE TRIM(LOWER(subject)) = %s",
            ('product quality analysis report ready',),
        )
        quality_emails = cur.fetchall()
        check("Quality analysis email with exact subject 'Product Quality Analysis Report Ready'",
              len(quality_emails) >= 1, f"found {len(quality_emails)}")
        if quality_emails:
            check("Quality analysis email to quality-team@company.com",
                  "quality-team@company.com" in str(quality_emails[0][1]).lower(),
                  f"to: {quality_emails[0][1]}")

        # Email 2: exact "Quality Review Meeting Required" -> operations
        cur.execute(
            "SELECT subject, to_addr FROM email.messages "
            "WHERE TRIM(LOWER(subject)) = %s",
            ('quality review meeting required',),
        )
        review_emails = cur.fetchall()
        check("Quality review meeting email with exact subject 'Quality Review Meeting Required'",
              len(review_emails) >= 1, f"found {len(review_emails)}")
        if review_emails:
            check("Review meeting email to operations@company.com",
                  "operations@company.com" in str(review_emails[0][1]).lower(),
                  f"to: {review_emails[0][1]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email checks", False, str(e))

    # Check calendar - exact dates and times
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events WHERE summary ILIKE %s", ('%quality review meeting%',))
        review_events = cur.fetchall()
        check("Quality Review Meeting event exists", len(review_events) >= 1, f"found {len(review_events)}")
        if review_events:
            ev = review_events[0]
            start_str = str(ev[2])
            check("Review Meeting on 2026-03-19", "2026-03-19" in start_str,
                  f"got {start_str}")
            try:
                if ev[2] and ev[3]:
                    check("Review Meeting starts at 10:00",
                          ev[2].hour == 10,
                          f"got {ev[2].hour}:00")
                    duration_min = (ev[3] - ev[2]).total_seconds() / 60
                    check("Review Meeting 90 minutes (10:00-11:30)",
                          85 <= duration_min <= 95,
                          f"duration {duration_min} min")
            except (AttributeError, TypeError):
                pass

        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events WHERE summary ILIKE %s", ('%quality improvement planning%',))
        plan_events = cur.fetchall()
        check("Quality Improvement Planning event exists", len(plan_events) >= 1, f"found {len(plan_events)}")
        if plan_events:
            ev = plan_events[0]
            start_str = str(ev[2])
            check("Planning event on 2026-03-21", "2026-03-21" in start_str,
                  f"got {start_str}")
            try:
                if ev[2] and ev[3]:
                    check("Planning event starts at 14:00",
                          ev[2].hour == 14,
                          f"got {ev[2].hour}:00")
                    duration_min = (ev[3] - ev[2]).total_seconds() / 60
                    check("Planning event 120 minutes (14:00-16:00)",
                          115 <= duration_min <= 125,
                          f"duration {duration_min} min")
            except (AttributeError, TypeError):
                pass

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar checks", False, str(e))

    check_reverse_validation(agent_workspace)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Product_Quality_Report.xlsx")
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        # No unexpected sheets beyond the 3 required ones
        expected_sheets = {"ticket_by_priority", "order_status_breakdown", "quality_action_plan"}
        unexpected = [s for s in wb.sheetnames if s.lower().replace(" ", "_") not in expected_sheets]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected sheets: {unexpected}")

        # Ticket counts should not be negative
        if "Ticket_by_Priority" in wb.sheetnames:
            ws = wb["Ticket_by_Priority"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[1] is not None:
                    val = safe_float(row[1])
                    if val is not None and val < 0:
                        check("No negative ticket counts", False, f"Found negative: {val}")
                        break
            else:
                check("No negative ticket counts", True)

    # Calendar: no events scheduled before today
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE summary ILIKE '%%quality%%'
              AND start_datetime < '2026-03-01'
        """)
        old_events = cur.fetchone()[0]
        check("No quality events before March 2026", old_events == 0,
              f"Found {old_events} old events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
