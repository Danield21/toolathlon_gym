"""Evaluation for yf-market-news-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected values from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM yf.news")
    total_articles = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT symbol) FROM yf.news")
    unique_tickers = cur.fetchone()[0]
    cur.execute("""
        SELECT data->'content'->'provider'->>'displayName' as publisher, COUNT(*) as cnt
        FROM yf.news GROUP BY 1 ORDER BY 2 DESC LIMIT 1
    """)
    row = cur.fetchone()
    top_publisher = row[0] if row else ""
    cur.close()
    conn.close()
    return total_articles, unique_tickers, top_publisher


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Market_News_Digest.xlsx."""
    print("\n=== Checking Market_News_Digest.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Market_News_Digest.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True
    total_articles, unique_tickers, top_publisher = get_expected_data()

    # Check News Articles sheet
    news_sheet = None
    for name in wb.sheetnames:
        if "news" in name.lower() and "article" in name.lower():
            news_sheet = wb[name]
            break
    if news_sheet is None:
        record("Sheet 'News Articles' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'News Articles' exists", True)
        rows = [r for r in news_sheet.iter_rows(min_row=2, values_only=True)
                if any(c is not None for c in r)]
        # Tighten: row count must be exact (count is integer from DB)
        ok_count = len(rows) == total_articles
        record("News Articles row count matches DB", ok_count,
               f"Expected exactly {total_articles}, got {len(rows)}")
        if not ok_count:
            all_ok = False

    # Check Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        summary = {}
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

        for key, val in summary.items():
            if "total" in key and "article" in key:
                # integer count - exact match
                ok = num_close(val, total_articles, 0)
                record("Summary Total_Articles", ok,
                       f"Expected {total_articles}, got {val}")
                if not ok:
                    all_ok = False
            elif "unique" in key and "ticker" in key:
                # integer count - exact match
                ok = num_close(val, unique_tickers, 0)
                record("Summary Unique_Tickers", ok,
                       f"Expected {unique_tickers}, got {val}")
                if not ok:
                    all_ok = False
            elif "top" in key and "publisher" in key:
                # Use exact normalised equality (case-insensitive trim) to prevent
                # 'Yahoo Reuters' from matching 'Reuters' via substring.
                expected_l = (top_publisher or "").strip().lower()
                actual_l = str(val or "").strip().lower()
                ok = (expected_l == actual_l) and bool(expected_l)
                record("Summary Top_Publisher", ok,
                       f"Expected '{top_publisher}', got '{val}'")
                if not ok:
                    all_ok = False

    return all_ok


EXPECTED_TICKERS = ["GOOGL", "AMZN", "JPM", "JNJ", "XOM", "GC=F", "^DJI"]


def check_notion():
    """Check Notion page titled exactly 'Market Intelligence Hub'."""
    print("\n=== Checking Notion Page ===")
    try:
        conn = psycopg2.connect(**DB)
    except psycopg2.OperationalError as e:
        record("Notion page check (DB connect)", False, str(e))
        return False
    cur = conn.cursor()

    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()

    found_page_id = None
    for page in pages:
        props = page[1] if isinstance(page[1], dict) else json.loads(page[1]) if page[1] else {}
        title_text = ""
        if "title" in props:
            t = props["title"]
            if isinstance(t, dict) and "title" in t:
                for item in t["title"]:
                    if isinstance(item, dict):
                        title_text += item.get("plain_text", item.get("text", {}).get("content", ""))
        if "market intelligence hub" in title_text.lower():
            found_page_id = page[0]
            break

    ok = found_page_id is not None
    record("Notion page 'Market Intelligence Hub' exists", ok,
           "No page with title 'Market Intelligence Hub'")

    notion_ok = ok
    # Check that the Notion page (children blocks) mentions all 7 tickers.
    # This is REQUIRED by task.md (per-ticker heading + paragraph), so failures here are
    # blocking. DB schema variations are handled by widening the query.
    if found_page_id:
        try:
            # Notion schema uses block_data (jsonb), with type column for heading detection
            cur.execute(
                "SELECT block_data, type FROM notion.blocks WHERE parent_id=%s",
                (found_page_id,)
            )
            blocks = cur.fetchall()
        except Exception as e:
            blocks = []
            print(f"  [WARN] Notion blocks query failed: {e}")
        block_text = ""
        # Heading-typed blocks (any of heading_1, heading_2, heading_3) are tracked separately
        heading_text = ""
        for b in blocks:
            try:
                d = b[0] if isinstance(b[0], dict) else json.loads(b[0]) if b[0] else {}
            except Exception:
                d = {}
            block_type = b[1] if len(b) > 1 else ""
            block_type = str(block_type or "").lower()
            if not isinstance(d, dict):
                continue
            # Notion block_data may be either {<type>: {rich_text: [...]}} or
            # {rich_text: [...]} directly. Walk both forms.
            def _collect_rich_text(node):
                texts = []
                if isinstance(node, dict):
                    rt = node.get("rich_text")
                    if isinstance(rt, list):
                        for it in rt:
                            if isinstance(it, dict):
                                texts.append(it.get("plain_text", "") or "")
                    for v in node.values():
                        if isinstance(v, dict):
                            texts.extend(_collect_rich_text(v))
                return texts
            collected = _collect_rich_text(d)
            for text in collected:
                block_text += " " + text
                if block_type.startswith("heading"):
                    heading_text += " " + text
        missing_tk = [t for t in EXPECTED_TICKERS if t.lower() not in block_text.lower()]
        tk_ok = len(missing_tk) == 0
        record("Notion page mentions all 7 tickers", tk_ok,
               f"Missing tickers: {missing_tk}")
        if not tk_ok:
            notion_ok = False
        # Per task.md: 'a heading with the ticker symbol' — at least 5 of 7 tickers
        # should appear in heading-typed blocks (some agents may use different heading
        # encoding for special tickers like ^DJI / GC=F).
        heading_hits = sum(1 for t in EXPECTED_TICKERS if t.lower() in heading_text.lower())
        # Only enforce when we actually saw heading blocks (avoid penalising schema variants)
        if heading_text.strip():
            ok_hd = heading_hits >= 5
            record("Notion page has ticker headings (>=5 of 7)", ok_hd,
                   f"Found {heading_hits}/7 ticker headings")
            if not ok_hd:
                notion_ok = False

    cur.close()
    conn.close()
    return notion_ok


def check_email():
    """Check email exists with subject 'Weekly Market News Digest' to portfolio-team@investco.com."""
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB)
    except psycopg2.OperationalError as e:
        record("Email check (DB connect)", False, str(e))
        return False
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%weekly%market%news%digest%'
           OR subject ILIKE '%market%news%digest%'
        LIMIT 10
    """)
    rows = cur.fetchall()

    matching = []
    for subj, to_addr, body in rows:
        if "portfolio-team@investco.com" in str(to_addr).lower():
            matching.append((subj, to_addr, body))

    has_email = len(matching) > 0
    record("Email 'Weekly Market News Digest' to portfolio-team@investco.com", has_email,
           f"Candidates examined: {len(rows)}")

    email_ok = has_email
    if has_email:
        # Validate body mentions tickers and has total numbers
        subj, to_addr, body = matching[0]
        body_l = (body or "").lower()
        # task says 'tickers covered' — require at least 6 of 7 to be more strict
        ticker_hits = sum(1 for t in EXPECTED_TICKERS if t.lower() in body_l)
        body_ok = ticker_hits >= 6
        record("Email body mentions at least 6 of 7 tickers", body_ok,
               f"Found {ticker_hits}/7 tickers in body")
        if not body_ok:
            email_ok = False
        # Also require a numeric total (any digit) and a non-trivial body length
        has_digit = any(c.isdigit() for c in body_l)
        record("Email body contains a numeric total", has_digit,
               f"body excerpt: {body_l[:100]}")
        if not has_digit:
            email_ok = False
        non_trivial = len(body_l.strip()) >= 60
        record("Email body length >= 60 chars", non_trivial,
               f"length={len(body_l.strip())}")
        if not non_trivial:
            email_ok = False

    cur.close()
    conn.close()
    return email_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)

    notion_ok = check_notion()
    email_ok = check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    overall = excel_ok and notion_ok and email_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
