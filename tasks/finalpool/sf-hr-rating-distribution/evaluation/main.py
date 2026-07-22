"""Evaluation for sf-hr-rating-distribution."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Rating_Distribution.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Rating_Distribution.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Rating Distribution
    print(f"  Checking Rating Distribution...")
    a_rows = load_sheet_rows(agent_wb, "Rating Distribution")
    g_rows = load_sheet_rows(gt_wb, "Rating Distribution")
    if a_rows is None:
        all_errors.append("Sheet 'Rating Distribution' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Rating Distribution' not found in groundtruth")
    else:
        sheet_name = "Rating Distribution"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # Tighter tolerance: max(2, 5% of GT) to handle very small buckets accurately
                gv = g_row[1]
                try:
                    tol = max(2, abs(float(gv)) * 0.02) if gv else 2
                except (TypeError, ValueError):
                    tol = 2
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}.Count: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # Scaled tolerance per metric type
                if "pct" in key or "percent" in key:
                    tol = 0.5
                elif "total_employees" in key:
                    tol = 1
                elif "count" in key:
                    tol = 2
                else:
                    tol = 1.0
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check email (BLOCKING)
    print("  Checking email (blocking)...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%%hr-analytics@company.com%%'
              AND subject ILIKE '%%performance rating distribution report%%'
        """)
        rows = cur.fetchall()
        if not rows:
            all_errors.append("Email to hr-analytics@company.com with subject 'Performance Rating Distribution Report' not found")
        else:
            body_concat = " ".join((r[2] or "").lower() for r in rows)
            # Body must contain BOTH at least one specific count AND at least one specific percentage/identifier
            # to verify the agent actually computed/inserted the distribution numbers (not just generic words).
            specific_counts = ["50000", "50,000", "5008", "5,008", "12464", "12,464", "2553", "2,553"]
            specific_pcts = ["34.9", "34.90", "34.9%", "high_performers_pct", "high performers"]
            has_count = any(kw in body_concat for kw in specific_counts)
            has_pct_or_label = any(kw in body_concat for kw in specific_pcts)
            if not (has_count and has_pct_or_label):
                all_errors.append(
                    f"Email body must include at least one specific count "
                    f"(50000/5008/12464/2553) AND at least one specific pct or label "
                    f"(34.9 or 'high performers'); got body excerpt: {body_concat[:200]}"
                )
            print(f"    PASS (found {len(rows)} matching emails)")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Email check failed: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
