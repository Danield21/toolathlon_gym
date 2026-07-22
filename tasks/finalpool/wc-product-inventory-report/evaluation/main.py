"""Evaluation for wc-product-inventory-report."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Inventory_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Inventory_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # ===== Inventory Report sheet =====
    print(f"  Checking Inventory Report...")
    a_rows = load_sheet_rows(agent_wb, "Inventory Report")
    g_rows = load_sheet_rows(gt_wb, "Inventory Report")
    if a_rows is None:
        all_errors.append("Sheet 'Inventory Report' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Inventory Report' not found in groundtruth")
    else:
        errors = []
        a_data = [r for r in a_rows[1:] if r and r[0] is not None and str(r[0]).strip()]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None and str(r[0]).strip()]

        # Row count: must equal expected
        if len(a_data) != len(g_data):
            errors.append(f"Row count mismatch: agent {len(a_data)} vs gt {len(g_data)}")

        # Build lookups keyed by name first, then fall back to (name, regular_price) tuple
        # to disambiguate any duplicate product names (rare but defensive).
        def _key(row):
            name = str(row[0]).strip().lower() if row[0] is not None else ""
            try:
                rp = float(row[2]) if len(row) > 2 and row[2] is not None else None
            except (TypeError, ValueError):
                rp = None
            return name, rp

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                k = _key(row)
                # Primary key by name; secondary by (name, price) tuple
                a_lookup[k[0]] = row
                a_lookup[k] = row
        g_lookup = {}
        for row in g_data:
            if row and row[0] is not None:
                k = _key(row)
                g_lookup[k[0]] = row
                g_lookup[k] = row
        # Iterate only on string-keyed entries (avoid double-counting tuple keys)
        g_lookup_iter = {k: v for k, v in g_lookup.items() if isinstance(k, str)}
        a_lookup = {k: v for k, v in a_lookup.items() if isinstance(k, str)}
        g_lookup = g_lookup_iter

        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Pad short rows with None
            def safe(r, i):
                return r[i] if i < len(r) else None

            # Type (col 1) -- string
            g_type, a_type = safe(g_row, 1), safe(a_row, 1)
            if not str_match(a_type, g_type):
                errors.append(f"{key[:60]}.Type: {a_type!r} vs {g_type!r}")

            # Regular_Price (col 2) -- tight 0.01 tolerance for dollar amounts
            g_rp, a_rp = safe(g_row, 2), safe(a_row, 2)
            if not num_close(a_rp, g_rp, 0.01):
                errors.append(f"{key[:60]}.Regular_Price: {a_rp} vs {g_rp} (tol=0.01)")

            # Sale_Price (col 3) -- 0.01 tolerance, allow 0 if not on sale
            g_sp, a_sp = safe(g_row, 3), safe(a_row, 3)
            if not num_close(a_sp, g_sp, 0.01):
                errors.append(f"{key[:60]}.Sale_Price: {a_sp} vs {g_sp} (tol=0.01)")

            # Stock_Qty (col 4) -- exact integer
            g_sq, a_sq = safe(g_row, 4), safe(a_row, 4)
            if not num_close(a_sq, g_sq, 0):
                errors.append(f"{key[:60]}.Stock_Qty: {a_sq} vs {g_sq} (exact)")

            # Stock_Status (col 5) -- string
            g_ss, a_ss = safe(g_row, 5), safe(a_row, 5)
            if not str_match(a_ss, g_ss):
                errors.append(f"{key[:60]}.Stock_Status: {a_ss!r} vs {g_ss!r}")

        # Detect extra rows in agent
        for key in a_lookup:
            if key not in g_lookup:
                errors.append(f"Extra row not in GT: {a_lookup[key][0]}")

        # Sort order: Regular_Price descending
        a_prices = []
        for row in a_data:
            try:
                a_prices.append(float(row[2]) if row[2] is not None else None)
            except (TypeError, ValueError):
                a_prices.append(None)
        # sorted desc check (skip Nones at end)
        non_none = [p for p in a_prices if p is not None]
        if non_none != sorted(non_none, reverse=True):
            errors.append(f"Inventory Report not sorted by Regular_Price descending (first 5: {a_prices[:5]})")

        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:8]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # ===== Summary sheet =====
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        # Support BOTH key-value (vertical) and single-row (horizontal) formats
        EXPECTED_KEYS = {
            "total_products": ("count", 0),
            "avg_price": ("price", 0.01),
            "on_sale_count": ("count", 0),
            "out_of_stock": ("count", 0),
        }

        # Build mapping from GT
        g_kv = {}
        for row in g_rows:
            if row and row[0] is not None and len(row) > 1 and row[1] is not None:
                key = str(row[0]).strip().lower().replace(" ", "_")
                g_kv[key] = row[1]
        # If GT also has horizontal layout: header row + value row
        if not all(k in g_kv for k in EXPECTED_KEYS):
            # try horizontal
            if len(g_rows) >= 2:
                hdr = [str(c).strip().lower().replace(" ", "_") if c else "" for c in g_rows[0]]
                vals = g_rows[1] if len(g_rows) > 1 else []
                for i, h in enumerate(hdr):
                    if h in EXPECTED_KEYS and i < len(vals):
                        g_kv.setdefault(h, vals[i])

        # Build mapping from agent (try both layouts)
        a_kv = {}
        for row in a_rows:
            if row and row[0] is not None and len(row) > 1 and row[1] is not None:
                key = str(row[0]).strip().lower().replace(" ", "_")
                if key in EXPECTED_KEYS:
                    a_kv[key] = row[1]
        if not all(k in a_kv for k in EXPECTED_KEYS):
            # try horizontal
            if len(a_rows) >= 2:
                hdr = [str(c).strip().lower().replace(" ", "_") if c else "" for c in a_rows[0]]
                vals = a_rows[1] if len(a_rows) > 1 else []
                for i, h in enumerate(hdr):
                    if h in EXPECTED_KEYS and i < len(vals):
                        a_kv.setdefault(h, vals[i])

        for k, (kind, tol) in EXPECTED_KEYS.items():
            if k not in g_kv:
                errors.append(f"Summary GT missing {k}")
                continue
            if k not in a_kv:
                errors.append(f"Summary agent missing {k}")
                continue
            if not num_close(a_kv[k], g_kv[k], tol):
                errors.append(f"Summary {k}: agent={a_kv[k]} gt={g_kv[k]} (tol={tol})")

        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:8]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:15]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
