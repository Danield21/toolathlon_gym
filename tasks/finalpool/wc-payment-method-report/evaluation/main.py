"""Evaluation for wc-payment-method-report."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Payment_Method_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Payment_Method_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # --- Check Payment Methods sheet ---
    print("  Checking Payment Methods sheet...")
    a_rows = load_sheet_rows(agent_wb, "Payment Methods")
    g_rows = load_sheet_rows(gt_wb, "Payment Methods")
    if a_rows is None:
        all_errors.append("Sheet 'Payment Methods' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Payment Methods' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row: {g_row[0]}")
                continue

            # Order_Count (col 1) - exact integer
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    all_errors.append(f"{key}.Order_Count: {a_row[1]} vs {g_row[1]}")

            # Total_Revenue (col 2) - tighten to tol=1.0 (was 50.0)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1.0):
                    all_errors.append(f"{key}.Total_Revenue: {a_row[2]} vs {g_row[2]} (tol=1.0)")

            # Avg_Order_Value (col 3) - tighten to tol=0.05 (was 5.0)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.05):
                    all_errors.append(f"{key}.Avg_Order_Value: {a_row[3]} vs {g_row[3]} (tol=0.05)")

        # Validate column headers
        if a_rows:
            header = [str(c or "").strip().lower() for c in a_rows[0]]
            for col in ["payment_method", "order_count", "total_revenue", "avg_order_value"]:
                if not any(col in h or col.replace("_", " ") in h for h in header):
                    all_errors.append(f"Payment Methods missing column header: {col}")

        # Validate alphabetical sort by payment method name
        a_method_names = [str(r[0] or "").strip() for r in a_data if r and r[0] is not None]
        if a_method_names:
            sorted_names = sorted(a_method_names, key=lambda s: s.lower())
            if a_method_names != sorted_names:
                all_errors.append(
                    f"Payment Methods not sorted alphabetically. Got {a_method_names}; "
                    f"expected {sorted_names}"
                )

        if not [e for e in all_errors if "Payment Methods" in e or "Missing row" in e]:
            print("    PASS")

    # --- Check Summary sheet ---
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Summary: {g_row[0]}")
                continue

            g_val = g_row[1]
            a_val = a_row[1]

            try:
                float(a_val); float(g_val)
                # Numeric tol depends on metric: counts are integer, totals to ~1.0
                tol = 0
                if "revenue" in key or "total_revenue" in key:
                    tol = 1.0
                if not num_close(a_val, g_val, tol):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val} (tol={tol})")
            except (TypeError, ValueError):
                if not str_match(a_val, g_val):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val}")

        if not [e for e in all_errors if "Summary" in e]:
            print("    PASS")

    # --- Check email sent ---
    print("  Checking email...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT m.subject, m.from_addr, m.to_addr, m.body_text FROM email.messages m")
        all_msgs = cur.fetchall()

        # Filter to recipient finance-lead@company.com
        target_emails = []
        for subj, from_addr, to_addr, body in all_msgs:
            to_str = str(to_addr or "").lower()
            if "finance-lead@company.com" in to_str:
                target_emails.append((subj, from_addr, to_addr, body))

        if not target_emails:
            all_errors.append("No email sent to finance-lead@company.com")
        else:
            # At least one must have exact subject 'Payment Method Analysis' (case-insensitive)
            ok_subj = False
            ok_from = False
            ok_body = False
            for subj, from_addr, to_addr, body in target_emails:
                sl = (subj or "").strip().lower()
                if sl == "payment method analysis":
                    ok_subj = True
                if "analytics@company.com" in (from_addr or "").lower():
                    ok_from = True
                bl = (body or "").lower()
                # Body must mention the most-used method 'Credit Card (Stripe)' AND total revenue '61712'
                if ("credit card" in bl or "stripe" in bl) and "61712" in bl:
                    ok_body = True
            if not ok_subj:
                all_errors.append(
                    f"No email with subject 'Payment Method Analysis'. "
                    f"Subjects: {[s for s, _, _, _ in target_emails]}"
                )
            if not ok_from:
                all_errors.append(
                    f"No email from analytics@company.com. "
                    f"Senders: {[f for _, f, _, _ in target_emails]}"
                )
            if not ok_body:
                all_errors.append(
                    f"No email body mentions most-used method (Credit Card/Stripe) AND total revenue 61712. "
                    f"Bodies (first 200): {[(b or '')[:200] for _, _, _, b in target_emails][:1]}"
                )
            if ok_subj and ok_from and ok_body:
                print("    PASS")

        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Email check error: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
