"""Evaluation for sf-hr-education-notion."""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

EXPECTED_PAGE_TITLE = "Workforce Education Analysis"

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _normalize_title(s):
    """Lowercase, replace punctuation with spaces, collapse whitespace."""
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(s).lower()).split())


def _levenshtein(a, b):
    """Small-scale DP edit distance between two strings."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def titles_match(expected, actual):
    """Exact Notion title match (task.md prescribes the exact title). After
    normalization, the titles must be identical — no fuzzy/edit-distance
    tolerance, so the model must follow the title given in task.md."""
    e = _normalize_title(expected)
    a = _normalize_title(actual)
    if not e or not a:
        return False
    return e == a


def get_sheet(wb, target):
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Education_Analysis.xlsx against groundtruth."""
    print("\n=== Checking Education_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Education_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Education_Analysis.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet: By Department ---
    agent_ws = get_sheet(agent_wb, "By Department")
    gt_ws = get_sheet(gt_wb, "By Department")

    if agent_ws is None:
        record("Sheet 'By Department' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'By Department' exists", True)
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        # Drop empty rows
        agent_rows = [r for r in agent_rows if r and r[0]]
        gt_rows = [r for r in gt_rows if r and r[0]]

        record("By Department row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")
        if len(agent_rows) != len(gt_rows):
            all_ok = False

        # Verify alphabetical sort
        agent_dept_order = [str(r[0]).strip() for r in agent_rows]
        sort_ok = agent_dept_order == sorted(agent_dept_order, key=lambda s: s.lower())
        record("By Department sorted alphabetically", sort_ok,
               f"order: {agent_dept_order}")
        if not sort_ok:
            all_ok = False

        agent_lookup = {}
        for r in agent_rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        # Tight integer tolerance (data is deterministic per DB snapshot)
        TOL_COUNT = 0
        TOL_PCT = 0.2

        for gt_row in gt_rows:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Department '{gt_row[0]}' present", False, "Missing")
                all_ok = False
                continue

            checks = [
                (1, "Total_Employees", TOL_COUNT),
                (2, "Bachelors_Count", TOL_COUNT),
                (3, "Masters_Count", TOL_COUNT),
                (4, "PhD_Count", TOL_COUNT),
                (5, "Bachelors_Pct", TOL_PCT),
                (6, "Masters_Pct", TOL_PCT),
                (7, "PhD_Pct", TOL_PCT),
            ]
            for col, name, tol in checks:
                exp = gt_row[col] if col < len(gt_row) else None
                act = a_row[col] if col < len(a_row) else None
                ok = num_close(act, exp, tol)
                record(f"'{gt_row[0]}' {name}", ok,
                       f"Expected {exp}, got {act}")
                if not ok:
                    all_ok = False

    # --- Sheet: Summary ---
    agent_ws2 = get_sheet(agent_wb, "Summary")
    gt_ws2 = get_sheet(gt_wb, "Summary")

    if agent_ws2 is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)

        agent_summary = {}
        for row in agent_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_summary[str(row[0]).strip().lower()] = row[1]

        gt_summary = {}
        for row in gt_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_summary[str(row[0]).strip().lower()] = row[1]

        # Tighter tolerances per metric kind
        SUMMARY_TOLS = {
            "total_employees": 0,           # exact integer
            "avg_salary_phd": 5,            # whole-dollar drift
            "avg_salary_bachelors": 5,
        }
        for metric, expected in gt_summary.items():
            actual = agent_summary.get(metric)
            if actual is None:
                record(f"Summary '{metric}' present", False, "Missing")
                all_ok = False
                continue
            if isinstance(expected, (int, float)):
                tol = SUMMARY_TOLS.get(metric, max(abs(expected) * 0.005, 1))
                ok = num_close(actual, expected, tol)
            else:
                # string match: allow case-insensitive or substring match for things like "Bachelor's"
                a = str(actual).strip().lower().replace("'", "").replace("’", "")
                e = str(expected).strip().lower().replace("'", "").replace("’", "")
                ok = a == e or e in a or a in e
            record(f"Summary '{metric}'", ok,
                   f"Expected {expected}, got {actual}")
            if not ok:
                all_ok = False

    return all_ok


def check_notion():
    """Check Notion page titled 'Workforce Education Analysis'."""
    print("\n=== Checking Notion Page ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("Notion check connectable", False, f"DB error: {e}")
        return False

    cur.execute("SELECT id, properties FROM notion.pages")
    pages = cur.fetchall()

    found_page = False
    page_id = None
    for page in pages:
        props = page[1] if isinstance(page[1], dict) else json.loads(page[1]) if page[1] else {}
        title_text = ""
        if "title" in props:
            t = props["title"]
            if isinstance(t, dict) and "title" in t:
                for item in t["title"]:
                    if isinstance(item, dict) and "text" in item:
                        title_text += item["text"].get("content", "")
                    elif isinstance(item, dict) and "plain_text" in item:
                        title_text += item["plain_text"]
        # Exact title match (task.md prescribes the exact page title)
        if titles_match(EXPECTED_PAGE_TITLE, title_text):
            found_page = True
            page_id = page[0]
            break

    record("Notion page 'Workforce Education Analysis' found", found_page,
           f"No page with exact title '{EXPECTED_PAGE_TITLE}'")

    notion_ok = found_page
    if found_page and page_id:
        # Check that the page has some blocks (content)
        cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s", (str(page_id),))
        block_count = cur.fetchone()[0]
        ok = block_count >= 3
        record("Notion page has content blocks", ok,
               f"Found {block_count} blocks")
        if not ok:
            notion_ok = False

        # Get all block content and types (notion.blocks stores JSON in block_data)
        cur.execute("SELECT type, block_data FROM notion.blocks WHERE parent_id = %s", (str(page_id),))
        blocks = cur.fetchall()
        all_text = ""
        heading_texts = []
        for (btype, block_data) in blocks:
            if block_data:
                text = json.dumps(block_data) if isinstance(block_data, dict) else str(block_data)
                all_text += text.lower()
                if btype and "heading" in str(btype).lower():
                    heading_texts.append(text.lower())

        # Per task: heading for EACH department
        EXPECTED_DEPTS = ["Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"]
        # Try to fetch live list from DB so updated data still works
        try:
            conn2 = psycopg2.connect(**DB_CONFIG)
            cur2 = conn2.cursor()
            cur2.execute(
                'SELECT DISTINCT "DEPARTMENT" FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" '
                'WHERE "DEPARTMENT" IS NOT NULL ORDER BY "DEPARTMENT"'
            )
            dept_rows = cur2.fetchall()
            if dept_rows:
                EXPECTED_DEPTS = [r[0] for r in dept_rows if r[0]]
            cur2.close()
            conn2.close()
        except Exception:
            pass

        for dept in EXPECTED_DEPTS:
            ok_dept = dept.lower() in all_text
            record(f"Notion page mentions department '{dept}'", ok_dept,
                   f"Not in page text")
            if not ok_dept:
                notion_ok = False

        # PhD percentage / total headcount mentioned
        record("Notion page mentions PhD", "phd" in all_text,
               "PhD not mentioned")
        if "phd" not in all_text:
            notion_ok = False

        # Highest PhD% department: query DB to find the actual top dept and require it
        # appears within ~120 chars of 'highest' (case-insensitive) in page content.
        top_dept = None
        try:
            conn3 = psycopg2.connect(**DB_CONFIG)
            cur3 = conn3.cursor()
            cur3.execute(
                """
                SELECT "DEPARTMENT" FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
                GROUP BY "DEPARTMENT"
                ORDER BY 100.0 * COUNT(*) FILTER (WHERE "EDUCATION_LEVEL" = 'PhD') / NULLIF(COUNT(*), 0) DESC
                LIMIT 1
                """
            )
            r0 = cur3.fetchone()
            if r0:
                top_dept = r0[0]
            cur3.close()
            conn3.close()
        except Exception:
            top_dept = None

        import re
        if top_dept:
            # Look for "highest ... <dept>" or "<dept> ... highest" within 120 chars of each other
            esc_dept = re.escape(top_dept.lower())
            pat1 = r"highest[^\n]{0,120}" + esc_dept
            pat2 = esc_dept + r"[^\n]{0,120}highest"
            ok_highest = (re.search(pat1, all_text, re.IGNORECASE) is not None
                          or re.search(pat2, all_text, re.IGNORECASE) is not None)
            record(f"Notion page references highest PhD% dept '{top_dept}' near 'highest'",
                   ok_highest,
                   f"Could not find {top_dept} within 120 chars of 'highest'")
        else:
            ok_highest = "highest" in all_text and "phd" in all_text
            record("Notion page references highest PhD% department", ok_highest,
                   "No 'highest' + 'phd' phrasing (DB lookup unavailable)")
        if not ok_highest:
            notion_ok = False

    cur.close()
    conn.close()
    return notion_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    notion_ok = check_notion()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and notion_ok
    print(f"  Excel: {'PASS' if excel_ok else 'FAIL'}, Notion: {'PASS' if notion_ok else 'FAIL'}")
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
