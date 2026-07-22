"""Evaluation for sf-support-resolution-analysis."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Support_Resolution_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Support_Resolution_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Issue Analysis
    print(f"  Checking Issue Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Issue Analysis")
    g_rows = load_sheet_rows(gt_wb, "Issue Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Issue Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Issue Analysis' not found in groundtruth")
    else:
        sheet_name = "Issue Analysis"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # Tightened: Snowflake is read-only, so tickets count is deterministic; tol=1
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Tickets: {a_row[1]} vs {g_row[1]} (tol=1)")

            if len(a_row) > 2 and len(g_row) > 2:
                # Tightened: 2-decimal task spec; allow up to 0.5 hour variance
                if not num_close(a_row[2], g_row[2], 0.5):
                    errors.append(f"{key}.Avg_Response_Hrs: {a_row[2]} vs {g_row[2]} (tol=0.5)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.1):
                    errors.append(f"{key}.Avg_CSAT: {a_row[3]} vs {g_row[3]} (tol=0.1)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        # Per-metric tolerance: small ints exact, large totals tol=50, strings exact
        TOL_MAP = {
            "total_issue_types": 0,
            "total_tickets": 50,
            "most_common_issue": None,  # string exact
            "best_csat_issue": None,    # string exact
        }
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                tol = TOL_MAP.get(key, 1.0)
                if tol is None:
                    # String exact match
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}.Value (str): {a_row[1]} vs {g_row[1]}")
                else:
                    if not num_close(a_row[1], g_row[1], tol):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check sheet: Issue Analysis sort order (Tickets descending)
    print(f"  Checking sort order (Tickets desc)...")
    a_rows = load_sheet_rows(agent_wb, "Issue Analysis")
    if a_rows and len(a_rows) > 2:
        tickets = []
        for row in a_rows[1:]:
            if row and len(row) > 1 and row[1] is not None:
                try:
                    tickets.append(float(row[1]))
                except (TypeError, ValueError):
                    pass
        is_sorted = all(tickets[i] >= tickets[i+1] for i in range(len(tickets)-1))
        if not is_sorted:
            all_errors.append(f"Issue Analysis not sorted by Tickets desc: {tickets}")
            print(f"    FAIL")
        else:
            print(f"    PASS")

    # Check Google Sheet: "Support Issue Dashboard"
    print(f"  Checking Google Sheet 'Support Issue Dashboard'...")
    try:
        import psycopg2
        conn = psycopg2.connect(host='localhost', port=5432, database='toolathlon_gym',
                                user='eigent', password='camel', connect_timeout=5)
        cur = conn.cursor()
        cur.execute(
            "SELECT id FROM gsheet.spreadsheets WHERE LOWER(title) LIKE %s",
            ('%support issue dashboard%',)
        )
        rows = cur.fetchall()
        if not rows:
            all_errors.append("Google Sheet 'Support Issue Dashboard' not found")
            print(f"    FAIL: spreadsheet not found")
        else:
            ss_id = rows[0][0]
            # Look for an "Issues" sheet
            cur.execute(
                "SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id=%s AND LOWER(title) LIKE %s",
                (ss_id, '%issue%')
            )
            sheets = cur.fetchall()
            if not sheets:
                all_errors.append("Google Sheet 'Issues' sheet not found in 'Support Issue Dashboard'")
                print(f"    FAIL: 'Issues' sheet missing")
            else:
                sheet_id = sheets[0][0]
                # Verify cells contain issue types AND their per-row Tickets values from GT
                cur.execute(
                    "SELECT value FROM gsheet.cells WHERE spreadsheet_id=%s AND sheet_id=%s",
                    (ss_id, sheet_id)
                )
                cell_vals = [str(r[0]).strip().lower() for r in cur.fetchall() if r[0] is not None]
                cell_blob = " | ".join(cell_vals)
                gt_issue_rows = load_sheet_rows(gt_wb, "Issue Analysis")
                expected_issue_types = []
                expected_tickets = []
                for r in (gt_issue_rows[1:] if gt_issue_rows else []):
                    if r and r[0] is not None:
                        expected_issue_types.append(str(r[0]).strip().lower())
                        if len(r) > 1 and r[1] is not None:
                            expected_tickets.append(int(float(r[1])))
                # Issue type name presence
                missing_types = [it for it in expected_issue_types if it not in cell_blob]
                if missing_types:
                    all_errors.append(f"Google Sheet 'Issues' missing issue types: {missing_types}")
                    print(f"    FAIL: missing types {missing_types}")
                # Numeric Tickets-count presence (each GT row's count must appear in cells)
                missing_counts = []
                for tc in expected_tickets:
                    candidates = [str(tc), f"{tc:,}", f"{tc}.0"]
                    if not any(c in cell_blob for c in candidates):
                        missing_counts.append(tc)
                if missing_counts:
                    all_errors.append(f"Google Sheet 'Issues' missing Tickets values: {missing_counts}")
                    print(f"    FAIL: missing tickets values {missing_counts}")
                if not missing_types and not missing_counts:
                    print(f"    PASS")
        cur.close()
        conn.close()
    except Exception as e:
        # Defense: if DB unreachable, fail loudly so verifier doesn't silently
        # accept the run; a real eval environment must have gsheet schema available.
        all_errors.append(f"Google Sheet check could not run (DB error): {e}")
        print(f"    ERROR: gsheet DB check failed: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
