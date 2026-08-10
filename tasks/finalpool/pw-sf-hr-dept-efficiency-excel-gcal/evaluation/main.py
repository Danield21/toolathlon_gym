"""Evaluation script for pw-sf-hr-dept-efficiency-excel-gcal."""
import argparse
import json
import os
import re
import sys

import openpyxl

TASK_NAME = "pw-sf-hr-dept-efficiency-excel-gcal"

# All DB connection params read from env with safe defaults so the evaluator
# connects to the same per-task PG the preprocess used (PGPORT/PGHOST/PGDATABASE
# are injected by the run harness in parallel-isolation mode).
DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _norm(s):
    """Normalize a header string: lowercase, non-alphanumerics -> spaces."""
    if s is None:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower()).strip()


# Synonym map used by the soft column matcher. The upstream runtime repo ships
# NO utils.verify_v2 semantic module, so the evaluator must not rely on it.
# This built-in fallback accepts reasonable phrasings the task prompt leaves
# open ("Dept" vs "Department", "Action Item" vs "Action") while still
# rejecting clearly-wrong columns.
_HEADER_SYNONYMS = {
    "department": ["department", "dept", "department name", "division", "team",
                   "business unit", "primary dimension", "dimension"],
    "metric": ["metric", "metric name", "key metric", "measure", "indicator",
               "kpi"],
    "value": ["value", "metric value", "amount", "number", "result", "score"],
    "priority": ["priority", "priority level", "priority order", "rank",
                 "level"],
    "action": ["action", "action item", "action items", "recommendation",
               "recommended action", "recommendations", "action plan",
               "next step", "initiative"],
}


def _header_matches(expected, header):
    """Soft-match an expected column name against an actual header.

    Accepts exact/normalized equality or a shared word token (including any
    synonym token). Replaces the (unavailable) utils.verify_v2 smart matcher.
    """
    e = _norm(expected)
    h = _norm(header)
    if not e or not h:
        return False
    if e == h:
        return True
    e_tokens = set(e.split())
    for syn in _HEADER_SYNONYMS.get(e, []):
        e_tokens.update(_norm(syn).split())
    h_tokens = set(h.split())
    return bool(e_tokens & h_tokens)


def _sheet_headers_and_rows(ws):
    """Return (headers, data_rows) after locating the header row.

    Scans the top rows for the first non-empty row and treats it as the header
    row (tolerates a leading title row or blank row). Data rows are all
    non-empty rows after it, so a multi-agent repeat-write cannot break the
    min-row count via stray blank rows.
    """
    max_scan = min(ws.max_row, 6)
    header_row = 1
    for r in range(1, max_scan + 1):
        vals = [c.value for c in ws[r]]
        non_empty = [v for v in vals if v is not None and str(v).strip() != ""]
        # A real header row has >= 2 populated cells; a lone title row (1 cell)
        # is skipped so we land on the actual column headers.
        if len(non_empty) >= 2:
            header_row = r
            break
    else:
        # No multi-cell row in the top rows: fall back to the first non-empty row.
        for r in range(1, max_scan + 1):
            vals = [c.value for c in ws[r]]
            if any(v is not None and str(v).strip() != "" for v in vals):
                header_row = r
                break
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[header_row]]
    data_rows = [
        row
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True)
        if any(v is not None and str(v).strip() != "" for v in row)
    ]
    return headers, data_rows


# --- GT-driven value checks --------------------------------------------------
# The original eval verified only structure (sheet/column/row presence), so a
# submission with fabricated numbers passed. We close that gap against the GT
# workbook while respecting the task's metric ambiguity: task.md describes the
# comparison column only as "our internal metric values" vs "external benchmark",
# and the warehouse exposes salary / work-life-balance / satisfaction while the
# mock page exposes a benchmark for each. A valid solution may therefore pick a
# non-salary metric. Accordingly we:
#   * always enforce gap self-consistency (gap ~= internal - benchmark), which
#     is metric-independent and catches fabricated numbers;
#   * compare the metric-independent counts (Total_Departments, Total_Employees)
#     to the GT;
#   * compare internal & benchmark values to the GT ONLY when the agent's own
#     column header shows it chose a salary metric, so a WLB / satisfaction
#     solution is never falsely failed.
# The GT self-test always passes because agent_workspace == groundtruth_workspace.

_HR_INTERNAL_KW = ("our", "internal", "actual", "current", "salary", "avg",
                   "average", "value", "score", "metric", "mean")
_HR_BENCH_KW = ("benchmark", "industry", "external", "target", "reference")
_HR_GAP_KW = ("gap", "diff", "delta", "variance", "shortfall", "deviation",
              "discrepancy", "overage", "spread")
_HR_SALARY_KW = ("salary", "compensation", "comp", "pay", "wage", "income")


def _norm_hr(s):
    return re.sub(r"[^a-z0-9]+", "", str(s or "").strip().lower())


def _to_float_hr(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return None
    s = s.replace(",", "").replace("$", "").replace("¥", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else None


def _classify_hr_columns(headers):
    """Return (dept_i, internal_i, benchmark_i, gap_i, internal_is_salary).

    Indices are 0-based against `headers`. Any role that cannot be confidently
    placed is None; callers skip the corresponding comparison so an unusual but
    valid layout is not penalized."""
    norm = [_norm_hr(h) for h in headers]
    dept_i = None
    for i, h in enumerate(norm):
        if "department" in h or "dept" in h:
            dept_i = i
            break
    if dept_i is None:
        dept_i = 0  # task.md: the first column is named exactly "Department"
    assigned = {dept_i}

    def _first(kws):
        for i, h in enumerate(norm):
            if i in assigned:
                continue
            if any(k in h for k in kws):
                assigned.add(i)
                return i
        return None

    # Claim the most specific roles first so a "Salary_Gap" header is taken by
    # the gap role rather than by the salary/internal role.
    gap_i = _first(_HR_GAP_KW)
    bench_i = _first(_HR_BENCH_KW)
    internal_i = _first(_HR_INTERNAL_KW)
    internal_is_salary = bool(internal_i is not None and
                              any(k in norm[internal_i] for k in _HR_SALARY_KW))
    return dept_i, internal_i, bench_i, gap_i, internal_is_salary


def _check_hr_data_values(agent_ws, gt_ws):
    """Compare Data_Analysis numbers to the GT (gap consistency always; salary
    values only when the agent used a salary metric)."""
    a_hdrs = [str(c.value).strip() if c.value is not None else ""
              for c in agent_ws[1]]
    g_hdrs = [str(c.value).strip() if c.value is not None else ""
              for c in gt_ws[1]]
    a_dept, a_int, a_bench, a_gap, a_is_salary = _classify_hr_columns(a_hdrs)
    g_dept, g_int, g_bench, g_gap, _ = _classify_hr_columns(g_hdrs)
    if a_int is None or a_bench is None or a_gap is None:
        return  # layout too unusual to verify -> leave to structure checks

    a_rows = {}
    for row in agent_ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        key = _norm_hr(row[a_dept]) if a_dept < len(row) and row[a_dept] else ""
        if key:
            a_rows[key] = row

    g_map = {}
    for row in gt_ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        key = _norm_hr(row[g_dept]) if g_dept < len(row) and row[g_dept] else ""
        if not key:
            continue
        g_map[key] = (
            row[g_int] if g_int is not None and g_int < len(row) else None,
            row[g_bench] if g_bench is not None and g_bench < len(row) else None,
            row[g_gap] if g_gap is not None and g_gap < len(row) else None,
        )

    for key, arow in a_rows.items():
        gval = g_map.get(key)
        if gval is None:
            gkey = next((gk for gk in g_map if key in gk or gk in key), None)
            if gkey is not None:
                gval = g_map[gkey]
        dept_label = arow[a_dept] if a_dept < len(arow) else key
        a_int_v = _to_float_hr(arow[a_int]) if a_int < len(arow) else None
        a_bench_v = _to_float_hr(arow[a_bench]) if a_bench < len(arow) else None
        a_gap_v = _to_float_hr(arow[a_gap]) if a_gap < len(arow) else None
        # gap self-consistency (metric-independent)
        if a_int_v is not None and a_bench_v is not None and a_gap_v is not None:
            expected_gap = a_int_v - a_bench_v
            ok = abs(a_gap_v - expected_gap) <= max(2.0, abs(expected_gap) * 0.01)
            check(f"Data_Analysis '{dept_label}' gap equals internal - benchmark",
                  ok, f"gap={a_gap_v}, internal-benchmark={expected_gap:.2f}")
        # salary value comparison only when the agent column is salary-typed
        if a_is_salary and gval is not None:
            g_int_v = _to_float_hr(gval[0])
            g_bench_v = _to_float_hr(gval[1])
            if g_int_v is not None and a_int_v is not None:
                check(f"Data_Analysis '{dept_label}' internal salary matches GT",
                      abs(a_int_v - g_int_v) <= max(2.0, abs(g_int_v) * 0.005),
                      f"agent={a_int_v}, gt={g_int_v}")
            if g_bench_v is not None and a_bench_v is not None:
                check(f"Data_Analysis '{dept_label}' benchmark matches GT",
                      abs(a_bench_v - g_bench_v) <= max(2.0, abs(g_bench_v) * 0.005),
                      f"agent={a_bench_v}, gt={g_bench_v}")


def _check_hr_metrics_values(agent_ws, gt_ws):
    """Compare the metric-independent summary counts (Total_Departments,
    Total_Employees) to the GT. Metric-dependent rows (Avg_Salary_Gap,
    Departments_Above / Departments_Below) are intentionally not compared."""
    def _map(ws):
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            out[_norm_hr(row[0])] = row[1] if len(row) > 1 else None
        return out
    a = _map(agent_ws)
    g = _map(gt_ws)
    for key in ("totaldepartments", "totalemployees"):
        if key in a and key in g:
            av = _to_float_hr(a[key])
            gv = _to_float_hr(g[key])
            if av is not None and gv is not None:
                check(f"Metrics {key} matches GT",
                      abs(av - gv) <= max(1.0, abs(gv) * 0.001),
                      f"agent={a[key]}, gt={g[key]}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Detect GT self-test (V1 parity test).
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    # GT workbook for value comparison (None if absent -> value checks skip).
    gt_xlsx_path = (os.path.join(groundtruth_workspace, "Hr_Dept_Efficiency_Report.xlsx")
                    if groundtruth_workspace else "")
    gt_wb = None
    if gt_xlsx_path and os.path.exists(gt_xlsx_path):
        try:
            gt_wb = openpyxl.load_workbook(gt_xlsx_path, data_only=True)
        except Exception:
            gt_wb = None

    excel_path = os.path.join(agent_workspace, "Hr_Dept_Efficiency_Report.xlsx")
    check("Hr_Dept_Efficiency_Report.xlsx exists", os.path.exists(excel_path))
    wb = None
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            check("Hr_Dept_Efficiency_Report.xlsx opens", False, str(e))

    if wb is not None:
        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows non-empty data rows, and
            contains the required columns (soft, synonym-tolerant match)."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _headers, _data_rows = _sheet_headers_and_rows(wb[sheet_name])
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            for _exp in expected_cols:
                _ok = any(_header_matches(_exp, h) for h in _headers)
                check(f"{sheet_name} has {_exp} column",
                      _ok, f"headers: {[h for h in _headers if h][:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            headers, _ = _sheet_headers_and_rows(wb["Data_Analysis"])
            check_columns('Data_Analysis', ['Department'], 5)

            # task.md says the sheet includes the primary dimension, our
            # internal metric values, the external benchmark values, and the
            # gap/difference between them. Accept any reasonable phrasing for
            # the internal / benchmark / gap columns (salary or WLB/satisfaction
            # path). Only the dimension column is pinned to Department.
            header_concat = " ".join(headers).lower()
            # Internal column names are left open by the prompt; the internal
            # data source (HR_ANALYTICS) exposes PERFORMANCE_RATING /
            # JOB_SATISFACTION / WORK_LIFE_BALANCE and the mock page exposes
            # Benchmark_WLB_Score, so a correct model may plausibly label the
            # internal column "Our Score", "Performance Score", "Current", or
            # "Actual" (gap-analysis style "Current vs Target"). Accept any of
            # those phrasings as well as the salary/satisfaction/wlb forms.
            has_internal = any(k in header_concat for k in
                               ['our_', 'ours', 'internal', 'salary', 'avg',
                                'average', 'value', 'satisfaction', 'wlb',
                                'engagement', 'rating', 'metric', 'mean',
                                'score', 'current', 'actual'])
            check("Data_Analysis has internal-metric column", has_internal,
                  f"headers: {[h for h in headers if h][:8]}")
            # "Target" is the exact column name the provided Report_Template.xlsx
            # Thresholds sheet uses for the external benchmark target, so a
            # model that reuses template terminology for the benchmark column
            # must not be penalized.
            has_benchmark = any(k in header_concat for k in
                                ['benchmark', 'industry', 'national',
                                 'external', 'reference', 'target'])
            check("Data_Analysis has benchmark column", has_benchmark,
                  f"headers: {[h for h in headers if h][:8]}")
            has_gap = any(k in header_concat for k in
                          ['gap', 'diff', 'delta', 'variance', 'spread',
                           'shortfall', 'deviation', 'discrepancy', 'overage'])
            check("Data_Analysis has gap/difference column", has_gap,
                  f"headers: {[h for h in headers if h][:8]}")

            # Value checks against the GT workbook (gap consistency always;
            # salary comparison gated on a salary-typed internal column).
            if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames:
                _check_hr_data_values(wb["Data_Analysis"], gt_wb["Data_Analysis"])

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            # The prompt says Metrics "summarizes total counts, averages, and
            # key statistics" without pinning a row count; a minimal-but-correct
            # summary (e.g. Total_Departments + Total_Employees) may have only
            # two data rows. Require >= 2 so an empty/near-empty sheet still
            # fails while a lean correct summary passes.
            check_columns('Metrics', ['Metric', 'Value'], 2)
            # Metric-independent count checks (Total_Departments, Total_Employees).
            if gt_wb is not None and "Metrics" in gt_wb.sheetnames:
                _check_hr_metrics_values(wb["Metrics"], gt_wb["Metrics"])

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            check_columns('Recommendations', ['Priority', 'Action'], 2)

    # Calendar and script checks run regardless of Excel state.
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Per task.md: title 'Analysis Review' on 2026-03-14 14:00-15:00 UTC.
        cur.execute(
            "SELECT summary, start_datetime, end_datetime FROM gcal.events "
            "WHERE summary ILIKE %s",
            ('%analysis%review%',)
        )
        events = cur.fetchall()
        if is_gt_self_test and len(events) == 0:
            print("  [WARN] 'Analysis Review' calendar event exists: 0 found (GT self-test, non-blocking)")
        else:
            check("'Analysis Review' calendar event exists",
                  len(events) >= 1,
                  f"found {len(events)} events with 'Analysis Review' in summary")
            if events:
                from datetime import datetime, timezone
                target_start = datetime(2026, 3, 14, 14, 0, tzinfo=timezone.utc)
                ok_date = False
                for sum_, start_dt, end_dt in events:
                    if start_dt is None:
                        continue
                    # Normalize to UTC for comparison; if naive assume UTC.
                    if hasattr(start_dt, 'tzinfo') and start_dt.tzinfo is not None:
                        start_utc = start_dt.astimezone(timezone.utc)
                        local_view = start_dt.tzinfo
                    else:
                        start_utc = (start_dt.replace(tzinfo=timezone.utc)
                                     if hasattr(start_dt, 'replace') else start_dt)
                        local_view = timezone.utc
                    # The canonical target is 14:00 UTC, but a model that passes
                    # a naive '2026-03-14T14:00:00' (no Z / offset) to the gcal
                    # MCP has PostgreSQL store it in the session/server timezone
                    # rather than UTC. Accept BOTH the UTC instant AND the
                    # "14:00 on 2026-03-14 in the stored local timezone" reading
                    # so the naive path is not penalized by PG timezone config.
                    try:
                        target_local = datetime(2026, 3, 14, 14, 0,
                                                tzinfo=local_view).astimezone(timezone.utc)
                    except Exception:
                        target_local = target_start
                    if (abs((start_utc - target_start).total_seconds()) < 3600 or
                            abs((start_utc - target_local).total_seconds()) < 3600):
                        ok_date = True
                        break
                check("'Analysis Review' starts 2026-03-14 14:00 (UTC or local)",
                      ok_date,
                      f"event starts: {[e[1] for e in events]}")
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))

    proc_path = os.path.join(agent_workspace, "sf_efficiency_processor.py")
    proc_exists = os.path.exists(proc_path)
    check("sf_efficiency_processor.py exists", proc_exists)
    if proc_exists:
        try:
            proc_size = os.path.getsize(proc_path)
            check("sf_efficiency_processor.py is non-trivial (>=200 bytes)",
                  proc_size >= 200, f"size={proc_size} bytes")
        except OSError as e:
            check("sf_efficiency_processor.py size check", False, str(e))

    # task.md requires the processor to output sf_efficiency_results.json.
    results_path = os.path.join(agent_workspace, "sf_efficiency_results.json")
    if is_gt_self_test and not os.path.exists(results_path):
        print("  [WARN] sf_efficiency_results.json exists: not found (GT self-test, non-blocking)")
    elif os.path.exists(results_path):
        try:
            with open(results_path, "r", encoding="utf-8") as _f:
                json.load(_f)
            check("sf_efficiency_results.json is valid JSON", True)
        except Exception as e:
            check("sf_efficiency_results.json is valid JSON", False, str(e))
    else:
        check("sf_efficiency_results.json exists", False)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
