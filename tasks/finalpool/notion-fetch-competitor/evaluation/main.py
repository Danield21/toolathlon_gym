"""
Evaluation script for notion-fetch-competitor task.

Checks per task.md (TaskFlow Pro vs 4 competitors with 3 sheets including SWOT):
1. Excel file (Competitor_Analysis.xlsx) - 3 sheets: Product Comparison (4 rows),
   Feature Matrix (>=6 features), SWOT Summary (>=4 SWOT rows covering all 4 categories)
2. Notion page titled like "Competitive Analysis - Q1 2026" with summary content
3. Memory file (memory/memory.json) has been updated with notes
"""

import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Per task.md: middle pricing tier per competitor
# ProjectHub Inc / pro / $25, AgileBoard Co / standard / $15,
# WorkStream Ltd / growth / $20, TeamSync Corp / business / $28
EXPECTED_COMPETITORS = [
    {"company_keywords": ["projecthub"], "pricing_tier": "pro",      "monthly_price": 25,  "rating": 4.3},
    {"company_keywords": ["agileboard"], "pricing_tier": "standard", "monthly_price": 15,  "rating": 4.6},
    {"company_keywords": ["workstream"], "pricing_tier": "growth",   "monthly_price": 20,  "rating": 4.1},
    {"company_keywords": ["teamsync"],   "pricing_tier": "business", "monthly_price": 28,  "rating": 4.4},
]

SWOT_CATEGORIES = ["strength", "weakness", "opportunity", "threat"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def find_sheet(wb, *keyword_groups):
    """Find a sheet whose lowercased name contains all keywords from any group."""
    for sn in wb.sheetnames:
        low = sn.lower()
        for group in keyword_groups:
            if all(k in low for k in group):
                return sn
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Competitor_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # ---- Sheet 1: Product Comparison ----
    pc_name = find_sheet(wb, ["product", "comparison"], ["comparison"])
    if not pc_name:
        record("Sheet 'Product Comparison' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Product Comparison' exists", True)
        ws = wb[pc_name]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else "" for c in rows[0]] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        # Required columns (case-insensitive)
        req_cols = ["company", "product_name", "pricing_tier", "monthly_price",
                    "key_features", "target_market", "rating"]
        header_low = [h.lower() for h in header]
        col_idx = {}
        for col in req_cols:
            for i, h in enumerate(header_low):
                if col in h or col.replace("_", " ") in h or col.replace("_", "") in h.replace("_", "").replace(" ", ""):
                    col_idx[col] = i
                    break
        missing_cols = [c for c in req_cols if c not in col_idx]
        if missing_cols:
            record("Product Comparison has required columns",
                   False, f"Missing: {missing_cols}; got header={header}")
            all_ok = False
        else:
            record("Product Comparison has required columns", True)

        if len(data_rows) != 4:
            record("Product Comparison has exactly 4 rows", False, f"Found {len(data_rows)}")
            all_ok = False
        else:
            record("Product Comparison has exactly 4 rows", True)

            # Check each competitor row matches one expected competitor
            matched_indices = set()
            for r_idx, row in enumerate(data_rows):
                if not row or row[0] is None:
                    continue
                row_text = " ".join(str(c).lower() for c in row if c is not None)
                # Match competitor by company keyword
                expected = None
                exp_idx = None
                for i, e in enumerate(EXPECTED_COMPETITORS):
                    if i in matched_indices:
                        continue
                    if any(kw in row_text for kw in e["company_keywords"]):
                        expected = e
                        exp_idx = i
                        break
                if expected is None:
                    record(f"Row {r_idx+1} matches a known competitor", False,
                           f"Row text: {row_text[:120]}")
                    all_ok = False
                    continue
                matched_indices.add(exp_idx)
                kw = expected["company_keywords"][0]
                # Pricing tier check
                if "pricing_tier" in col_idx:
                    tier_val = row[col_idx["pricing_tier"]]
                    tier_str = str(tier_val).strip().lower() if tier_val is not None else ""
                    if expected["pricing_tier"] not in tier_str:
                        record(f"{kw}: pricing tier matches '{expected['pricing_tier']}'",
                               False, f"Got '{tier_str}'")
                        all_ok = False
                    else:
                        record(f"{kw}: pricing tier matches '{expected['pricing_tier']}'", True)
                # Monthly price check
                if "monthly_price" in col_idx:
                    mp_val = row[col_idx["monthly_price"]]
                    if not num_close(mp_val, expected["monthly_price"], tol=1.0):
                        record(f"{kw}: monthly_price ~={expected['monthly_price']}",
                               False, f"Got {mp_val}")
                        all_ok = False
                    else:
                        record(f"{kw}: monthly_price ~={expected['monthly_price']}", True)
                # Rating check
                if "rating" in col_idx:
                    rt_val = row[col_idx["rating"]]
                    if not num_close(rt_val, expected["rating"], tol=0.2):
                        record(f"{kw}: rating ~={expected['rating']}",
                               False, f"Got {rt_val}")
                        all_ok = False
                    else:
                        record(f"{kw}: rating ~={expected['rating']}", True)

    # ---- Sheet 2: Feature Matrix ----
    fm_name = find_sheet(wb, ["feature", "matrix"])
    if not fm_name:
        record("Sheet 'Feature Matrix' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Feature Matrix' exists", True)
        ws = wb[fm_name]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else "" for c in rows[0]] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        # Required columns
        req = ["feature", "our", "competitor"]
        header_low = " ".join(h.lower() for h in header)
        if not all(k in header_low for k in req):
            record("Feature Matrix header has Feature/Our_Product/Competitor*", False,
                   f"Got {header}")
            all_ok = False
        else:
            record("Feature Matrix header has Feature/Our_Product/Competitor*", True)
        # at least 6 features
        if len(data_rows) < 6:
            record("Feature Matrix has at least 6 features", False,
                   f"Found {len(data_rows)} rows")
            all_ok = False
        else:
            record("Feature Matrix has at least 6 features", True)

        # Values must be Yes/No/Partial across competitor columns
        valid_vals = {"yes", "no", "partial"}
        bad_rows = 0
        for row in data_rows:
            for cell in row[1:]:
                if cell is None:
                    bad_rows += 1
                    break
                if str(cell).strip().lower() not in valid_vals:
                    bad_rows += 1
                    break
        if bad_rows > 0:
            record("Feature Matrix values are Yes/No/Partial", False,
                   f"{bad_rows} rows have invalid values")
            all_ok = False
        else:
            record("Feature Matrix values are Yes/No/Partial", True)

    # ---- Sheet 3: SWOT Summary ----
    swot_name = find_sheet(wb, ["swot"])
    if not swot_name:
        record("Sheet 'SWOT Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'SWOT Summary' exists", True)
        ws = wb[swot_name]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else "" for c in rows[0]] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        req = ["category", "description", "priority"]
        header_low = " ".join(h.lower() for h in header)
        if not all(k in header_low for k in req):
            record("SWOT header has Category/Description/Priority", False, f"Got {header}")
            all_ok = False
        else:
            record("SWOT header has Category/Description/Priority", True)

        if len(data_rows) < 4:
            record("SWOT has at least 4 rows", False, f"Found {len(data_rows)}")
            all_ok = False
        else:
            record("SWOT has at least 4 rows", True)

        # All 4 SWOT categories must appear
        categories_present = set()
        for row in data_rows:
            if row and row[0] is not None:
                cat = str(row[0]).strip().lower()
                for swot_cat in SWOT_CATEGORIES:
                    if swot_cat in cat:
                        categories_present.add(swot_cat)
        for swot_cat in SWOT_CATEGORIES:
            if swot_cat in categories_present:
                record(f"SWOT contains category '{swot_cat}'", True)
            else:
                record(f"SWOT contains category '{swot_cat}'", False, "Not found")
                all_ok = False

        # Priority values
        valid_priorities = {"high", "medium", "low"}
        bad_pri = 0
        for row in data_rows:
            if len(row) < 3 or row[2] is None:
                bad_pri += 1
                continue
            if str(row[2]).strip().lower() not in valid_priorities:
                bad_pri += 1
        if bad_pri > 0:
            record("SWOT priorities are High/Medium/Low", False,
                   f"{bad_pri} rows invalid")
            all_ok = False
        else:
            record("SWOT priorities are High/Medium/Low", True)

    wb.close()
    return all_ok


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false AND in_trash = false")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Notion DB accessible", False, str(e))
        return False

    all_ok = True
    target_page_id = None
    target_title = ""

    for page_id, props_raw in pages:
        if isinstance(props_raw, str):
            props = json.loads(props_raw)
        else:
            props = props_raw

        title_parts = []
        for key in ["title", "Name"]:
            prop = props.get(key, {})
            if isinstance(prop, dict):
                tp = prop.get("title", [])
                if tp:
                    title_parts = tp
                    break

        page_title = "".join(p.get("plain_text", "") for p in title_parts)
        low = page_title.lower()
        if "competitive analysis" in low:
            target_page_id = page_id
            target_title = page_title
            break

    if not target_page_id:
        record("Notion page 'Competitive Analysis - Q1 2026' exists", False,
               f"Found {len(pages)} pages but none match 'competitive analysis'")
        return False
    record("Notion page 'Competitive Analysis - Q1 2026' exists", True)

    # Title should mention Q1 2026 (per task.md)
    title_low = target_title.lower()
    has_q1 = "q1" in title_low and "2026" in title_low
    if not has_q1:
        record("Notion page title mentions 'Q1 2026'", False, f"Got '{target_title}'")
        all_ok = False
    else:
        record("Notion page title mentions 'Q1 2026'", True)

    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute("SELECT type, block_data FROM notion.blocks WHERE parent_id = %s",
                     (target_page_id,))
        blocks = cur2.fetchall()
        cur2.close()
        conn2.close()
    except Exception as e:
        record("Notion blocks readable", False, str(e))
        return False

    all_text = " ".join(str(b[1]) for b in blocks if b[1]).lower()

    if len(all_text) < 50:
        record("Notion page has substantive content (>=50 chars)", False,
               f"Got {len(all_text)} chars")
        all_ok = False
    else:
        record("Notion page has substantive content (>=50 chars)", True)

    # Per task.md content requirements: summary of competitors, key
    # differentiators, pricing recommendations, action items
    expected_topics = {
        "competitor summary": ["projecthub", "agileboard", "workstream", "teamsync"],
        "differentiator/recommendation": ["differentiat", "recommend", "strength"],
        "pricing": ["pricing", "price", "tier"],
        "action items": ["action", "next step", "todo", "to-do", "to do"],
    }
    # competitor summary: at least 2 of the 4 competitor names appear
    comp_hits = sum(1 for kw in expected_topics["competitor summary"] if kw in all_text)
    if comp_hits >= 2:
        record("Notion page mentions at least 2 competitor names", True)
    else:
        record("Notion page mentions at least 2 competitor names", False,
               f"Hits: {comp_hits}")
        all_ok = False

    for topic in ["differentiator/recommendation", "pricing", "action items"]:
        if any(kw in all_text for kw in expected_topics[topic]):
            record(f"Notion page covers {topic}", True)
        else:
            record(f"Notion page covers {topic}", False, "No matching keyword")
            all_ok = False

    return all_ok


def check_memory(agent_workspace):
    print("\n=== Checking Memory ===")
    memory_path = os.path.join(agent_workspace, "memory", "memory.json")
    if not os.path.isfile(memory_path):
        record("Memory file exists", False, f"Not found: {memory_path}")
        return False
    record("Memory file exists", True)

    try:
        with open(memory_path, "r") as f:
            data = json.load(f)
    except Exception as e:
        record("Memory file is valid JSON", False, str(e))
        return False
    record("Memory file is valid JSON", True)

    content = json.dumps(data).lower()
    # Should have notes; default empty {"notes": []} has length 14 — must be longer
    has_notes = len(content) > 50
    record("Memory file has been updated with notes (>50 chars)", has_notes)
    has_relevant = ("competitor" in content or "competitive" in content or
                    "taskflow" in content or "differentiator" in content)
    record("Memory mentions competitor/taskflow/differentiator", has_relevant)
    return has_notes and has_relevant


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    global FAIL_COUNT

    # Detect GT self-test (V1 parity test). When agent==gt, the Notion DB
    # cannot have a competitive-analysis page (no agent has run yet), so
    # tolerate Notion absence as a warning rather than a hard FAIL.
    try:
        gt_canon = os.path.realpath(args.groundtruth_workspace) if args.groundtruth_workspace else ""
        ag_canon = os.path.realpath(args.agent_workspace) if args.agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    excel_ok = check_excel(args.agent_workspace)

    fail_before_notion = FAIL_COUNT
    notion_ok = check_notion()
    notion_failures = FAIL_COUNT - fail_before_notion

    memory_ok = check_memory(args.agent_workspace)

    if is_gt_self_test and not notion_ok:
        print("\n  [WARN] Notion check failed during GT self-test; tolerated (no agent run yet).")
        # Roll back the Notion failures during self-test only.
        FAIL_COUNT -= notion_failures
        notion_ok = True

    overall = excel_ok and notion_ok and memory_ok and FAIL_COUNT == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Notion:   {'PASS' if notion_ok else 'FAIL'}")
    print(f"  Memory:   {'PASS' if memory_ok else 'FAIL'}")
    print(f"  Overall:  {'PASS' if overall else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
