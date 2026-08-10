"""Evaluation for sf-department-budget-analysis."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Department_Budget_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Department_Budget_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Budget Analysis
    print(f"  Checking Budget Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Budget Analysis")
    g_rows = load_sheet_rows(gt_wb, "Budget Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Budget Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Budget Analysis' not found in groundtruth")
    else:
        sheet_name = "Budget Analysis"
        errors = []
        # Header check (order-independent): all expected columns must be present.
        expected_headers = [
            "Department", "Budget", "Planned_Headcount", "Actual_Headcount",
            "Avg_Salary", "Total_Salary_Cost", "Budget_Utilization_Pct",
        ]
        actual_headers = [str(c or "").strip() for c in (a_rows[0] if a_rows else [])]
        header_lower = [h.lower() for h in actual_headers]
        missing_headers = [h for h in expected_headers if h.lower() not in header_lower]
        if missing_headers:
            errors.append(f"Budget Analysis missing headers: {missing_headers}; found {actual_headers}")
        # Map expected columns to the agent's column indices by header name so the
        # column order in the output does not matter.
        col_of = {h.lower(): i for i, h in enumerate(actual_headers)}
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Filter empty rows
        a_data = [r for r in a_data if r and r[0] is not None]
        g_data = [r for r in g_data if r and r[0] is not None]

        # Row count must match GT (no extras)
        if len(a_data) != len(g_data):
            errors.append(f"Budget Analysis row count {len(a_data)} vs GT {len(g_data)}")

        # Sort order: alphabetical by Department
        a_dept_names = [str(r[0]).strip() for r in a_data]
        sorted_dept_names = sorted(a_dept_names, key=lambda s: s.lower())
        if a_dept_names != sorted_dept_names:
            errors.append(f"Budget Analysis not sorted alphabetically: {a_dept_names}")

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Column-wise comparison using header-mapped indices, so the agent's
            # column order does not matter. GT columns are fixed (1..6).
            COLS = [
                ("budget", "Budget", 100.0, 1),
                ("planned_headcount", "Planned_Headcount", 1, 2),
                ("actual_headcount", "Actual_Headcount", 0, 3),
                ("avg_salary", "Avg_Salary", 10.0, 4),
                ("total_salary_cost", "Total_Salary_Cost", 1000.0, 5),
                ("budget_utilization_pct", "Budget_Utilization_Pct", 0.1, 6),
            ]
            for hkey, label, tol, gcol in COLS:
                ai = col_of.get(hkey)
                if ai is None:
                    continue  # missing header already reported above
                if ai < len(a_row) and gcol < len(g_row):
                    if not num_close(a_row[ai], g_row[gcol], tol):
                        errors.append(f"{key}.{label}: {a_row[ai]} vs {g_row[gcol]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary (per-metric tolerance)
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        # Header check
        expected_headers = ["Metric", "Value"]
        actual_headers = [str(c or "").strip() for c in (a_rows[0] if a_rows else [])]
        for i, h in enumerate(expected_headers):
            if i >= len(actual_headers) or actual_headers[i].lower() != h.lower():
                errors.append(f"Summary header[{i}]: '{actual_headers[i] if i < len(actual_headers) else None}' vs '{h}'")
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Per-metric tolerance map (rounded to 1 decimal => tol=0.1; integer count => tol=0).
        # avg_budget_utilization gets a wider margin (0.5): the task asks for an average
        # of per-department utilization, but agents may legitimately compute the aggregate
        # ratio (total salary / total budget) instead, which can differ by a few tenths.
        TOL_BY_METRIC = {
            "total_budget": 100.0,
            "total_salary_cost": 1000.0,
            "avg_budget_utilization": 0.5,
            "over_budget_depts": 0,
        }
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        # Required rows
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            tol = TOL_BY_METRIC.get(key, 1.0)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
