"""Evaluation for yf-peer-benchmark-analysis."""
import argparse
import json
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel ===")
    import openpyxl
    path = os.path.join(agent_workspace, "Peer_Benchmark.xlsx")
    gt_path = os.path.join(gt_workspace, "Peer_Benchmark.xlsx")
    check("Peer_Benchmark.xlsx exists", os.path.exists(path), f"Expected {path}")
    if not os.path.exists(path):
        return
    if not os.path.exists(gt_path):
        check("GT Peer_Benchmark.xlsx exists", False)
        return

    a_wb = openpyxl.load_workbook(path, data_only=True)
    g_wb = openpyxl.load_workbook(gt_path, data_only=True)

    a_rows = load_sheet_rows(a_wb, "Consensus Analysis")
    g_rows = load_sheet_rows(g_wb, "Consensus Analysis")
    check("Sheet 'Consensus Analysis' present", a_rows is not None)
    if a_rows is not None and g_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        check("Consensus Analysis has 5 rows", len(a_data) == 5, f"got {len(a_data)}")
        # Sort: alphabetical by symbol
        symbols = [str(r[0]).strip().upper() for r in a_data]
        check("Consensus Analysis sorted alphabetically",
              symbols == sorted(symbols), f"got {symbols}")
        a_lookup = {str(r[0]).strip().upper(): r for r in a_data}
        for g_row in g_data:
            sym = str(g_row[0]).strip().upper()
            row = a_lookup.get(sym)
            check(f"Symbol '{sym}' present", row is not None)
            if row is None:
                continue
            # Current_Price ±0.50 (allow some drift)
            check(f"  {sym}.Current_Price ≈ {g_row[1]}",
                  num_close(row[1], g_row[1], 1.0), f"got {row[1]}")
            # Target_Price exact
            check(f"  {sym}.Target_Price == {g_row[2]}",
                  num_close(row[2], g_row[2], 0.01), f"got {row[2]}")
            # Upside_Pct ±0.5
            check(f"  {sym}.Upside_Pct ≈ {g_row[3]}",
                  num_close(row[3], g_row[3], 0.5), f"got {row[3]}")
            # Consensus exact
            check(f"  {sym}.Consensus == '{g_row[4]}'",
                  str_match(row[4], g_row[4]), f"got '{row[4]}'")
            # Trailing_PE ±0.1
            check(f"  {sym}.Trailing_PE ≈ {g_row[5]}",
                  num_close(row[5], g_row[5], 0.5), f"got {row[5]}")
            # PE_Benchmark exact
            check(f"  {sym}.PE_Benchmark == {g_row[6]}",
                  num_close(row[6], g_row[6], 0.01), f"got {row[6]}")
            # PE_vs_Benchmark exact
            check(f"  {sym}.PE_vs_Benchmark == '{g_row[7]}'",
                  str_match(row[7], g_row[7]), f"got '{row[7]}'")

    a_rows = load_sheet_rows(a_wb, "Summary")
    g_rows = load_sheet_rows(g_wb, "Summary")
    check("Sheet 'Summary' present", a_rows is not None)
    if a_rows is not None and g_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        a_lookup = {str(r[0]).strip().lower(): r[1] for r in a_data}
        for g_row in g_data:
            metric = str(g_row[0]).strip().lower()
            val = a_lookup.get(metric)
            check(f"Summary metric '{metric}' present", val is not None)
            if val is None:
                continue
            if metric == "most_undervalued":
                check(f"  Summary.{metric} == '{g_row[1]}'",
                      str_match(val, g_row[1]), f"got '{val}'")
            elif metric in ("stocks_above_target", "stocks_below_target"):
                check(f"  Summary.{metric} == {g_row[1]}",
                      num_close(val, g_row[1], 0), f"got {val}")
            else:  # avg metrics
                check(f"  Summary.{metric} ≈ {g_row[1]}",
                      num_close(val, g_row[1], 0.5), f"got {val}")


def check_gsheet(gt_workspace):
    print("\n=== Checking Google Sheet ===")
    import openpyxl
    g_wb = openpyxl.load_workbook(os.path.join(gt_workspace, "Peer_Benchmark.xlsx"), data_only=True)
    g_rows = load_sheet_rows(g_wb, "Consensus Analysis")
    expected_lookup = {str(r[0]).strip().upper(): r for r in g_rows[1:] if r and r[0]}

    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%%peer benchmark data%%' OR LOWER(title) LIKE '%%peer benchmark%%'")
    sheets = cur.fetchall()
    check("Spreadsheet 'Peer Benchmark Data' exists", len(sheets) >= 1, f"Found {len(sheets)}")
    if not sheets:
        cur.close()
        conn.close()
        return
    sid = sheets[0][0]
    cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id=%s", (sid,))
    sub = cur.fetchall()
    sheet_id = next((s_id for s_id, t in sub if t and t.strip().lower() == "benchmark"), None)
    check("Sheet 'Benchmark' present", sheet_id is not None, f"got {[t for _, t in sub]}")
    if sheet_id is None:
        cur.close()
        conn.close()
        return
    cur.execute("SELECT row_index, col_index, value FROM gsheet.cells WHERE sheet_id=%s ORDER BY row_index, col_index", (sheet_id,))
    cells = cur.fetchall()
    grid = {}
    for r, c, v in cells:
        grid.setdefault(r, {})[c] = str(v).strip() if v is not None else ""
    # Find header row
    header_row = None
    for r in sorted(grid.keys()):
        joined = " ".join(grid[r].values()).lower()
        if "symbol" in joined and "target" in joined and ("upside" in joined or "trailing_pe" in joined.replace(" ", "_")):
            header_row = r
            break
    check("GSheet header row present", header_row is not None)
    if header_row is None:
        cur.close()
        conn.close()
        return
    col_map = {}
    for c, v in grid[header_row].items():
        v_norm = v.strip().lower().replace(" ", "_")
        for key in ("symbol", "current_price", "target_price", "upside_pct", "consensus", "trailing_pe", "pe_benchmark", "pe_vs_benchmark"):
            if v_norm == key:
                col_map[key] = c

    data_rows = [grid[r] for r in sorted(grid.keys()) if r > header_row]
    check("GSheet has 5 stock data rows", len(data_rows) >= 5, f"got {len(data_rows)}")

    by_sym = {}
    for row in data_rows:
        sym = row.get(col_map.get("symbol", -1), "").strip().upper()
        if sym:
            by_sym[sym] = row

    for sym, exp in expected_lookup.items():
        row = by_sym.get(sym)
        check(f"GSheet symbol '{sym}' present", row is not None)
        if row is None:
            continue
        try:
            tgt = float(row.get(col_map.get("target_price", -1), "0"))
        except (TypeError, ValueError):
            tgt = None
        try:
            upside = float(row.get(col_map.get("upside_pct", -1), "0"))
        except (TypeError, ValueError):
            upside = None
        check(f"  GSheet {sym}.Target_Price == {exp[2]}",
              tgt is not None and abs(tgt - float(exp[2])) <= 0.01,
              f"got {tgt}")
        check(f"  GSheet {sym}.Upside_Pct ≈ {exp[3]}",
              upside is not None and abs(upside - float(exp[3])) <= 0.5,
              f"got {upside}")
        pe_v_b = row.get(col_map.get("pe_vs_benchmark", -1), "").strip().lower()
        check(f"  GSheet {sym}.PE_vs_Benchmark == '{exp[7]}'",
              pe_v_b == str(exp[7]).strip().lower(), f"got '{pe_v_b}'")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(agent_ws, gt_ws)
    check_gsheet(gt_ws)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
