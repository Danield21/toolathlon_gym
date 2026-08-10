"""
Evaluation for canvas-assignment-deadline-gcal.

Checks:
1. Assignment_Tracker.xlsx matches groundtruth (sheets: Assignments, Summary)
2. Google Calendar events for each assignment deadline
   - GCal checks are runtime_only (won't block when agent didn't populate)
   - BUT if agent populated GCal events at all, missing/incorrect events ARE blocking
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
RUNTIME_ONLY_FAIL = 0


def record(name, passed, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, RUNTIME_ONLY_FAIL
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if runtime_only:
            RUNTIME_ONLY_FAIL += 1
        msg = f": {detail[:300]}" if detail else ""
        suffix = " (runtime-only)" if runtime_only else ""
        print(f"  [FAIL] {name}{suffix}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _date_match(a, b):
    """Compare two date-ish values by their YYYY-MM-DD portion.

    A datetime cell ("2014-02-26 00:00:00") written by openpyxl must still match
    the groundtruth's plain "2014-02-26" string. Non-date strings fall back to a
    case-insensitive exact comparison.
    """
    def norm(v):
        if v is None:
            return None
        s = str(v).strip()
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return s[:10]
        return s.lower()
    return norm(a) == norm(b)


def get_sheet(wb, target):
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


# ============================================================================
# Check 1: Assignment_Tracker.xlsx
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    """Compare Assignment_Tracker.xlsx against groundtruth."""
    print("\n=== Checking Assignment_Tracker.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Assignment_Tracker.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Assignment_Tracker.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet: Assignments ---
    agent_ws = get_sheet(agent_wb, "Assignments")
    gt_ws = get_sheet(gt_wb, "Assignments")

    if agent_ws is None:
        record("Sheet 'Assignments' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Assignments' exists", True)

        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("Assignments row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        # Build lookup by assignment name
        agent_lookup = {}
        for r in agent_rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for gt_row in gt_rows:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Assignment '{gt_row[0]}' present", False, "Missing")
                all_ok = False
                continue

            # Check Due_Date (col 1)
            ok_date = _date_match(a_row[1], gt_row[1])
            record(f"'{gt_row[0]}' Due_Date", ok_date,
                   f"Expected {gt_row[1]}, got {a_row[1]}")
            if not ok_date:
                all_ok = False

            # Check Points_Possible (col 2)
            ok_pts = num_close(a_row[2], gt_row[2], 0.5)
            record(f"'{gt_row[0]}' Points_Possible", ok_pts,
                   f"Expected {gt_row[2]}, got {a_row[2]}")
            if not ok_pts:
                all_ok = False

            # Check Assignment_Group (col 3)
            ok_grp = str_match(a_row[3], gt_row[3])
            record(f"'{gt_row[0]}' Assignment_Group", ok_grp,
                   f"Expected {gt_row[3]}, got {a_row[3]}")
            if not ok_grp:
                all_ok = False

    # --- Sheet: Summary ---
    agent_ws2 = get_sheet(agent_wb, "Summary")
    gt_ws2 = get_sheet(gt_wb, "Summary")

    if agent_ws2 is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)

        agent_summary = {}
        for row in agent_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_summary[str(row[0]).strip().lower()] = row[1]

        gt_summary = {}
        for row in gt_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_summary[str(row[0]).strip().lower()] = row[1]

        for metric, expected in gt_summary.items():
            actual = agent_summary.get(metric)
            if actual is None:
                record(f"Summary '{metric}' present", False, "Missing")
                all_ok = False
            else:
                ok = num_close(actual, expected, 1.0) if isinstance(expected, (int, float)) else _date_match(actual, expected)
                record(f"Summary '{metric}'", ok,
                       f"Expected {expected}, got {actual}")
                if not ok:
                    all_ok = False

    return all_ok


# ============================================================================
# Check 2: Google Calendar events
# ============================================================================

def _get_expected_assignments(gt_workspace):
    """Load expected assignment names from groundtruth Excel."""
    gt_file = os.path.join(gt_workspace, "Assignment_Tracker.xlsx")
    names = []
    if not os.path.isfile(gt_file):
        return names
    try:
        wb = openpyxl.load_workbook(gt_file, data_only=True)
        ws = get_sheet(wb, "Assignments")
        if ws is None:
            return names
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                names.append(str(row[0]).strip())
    except Exception:
        pass
    return names


def check_gcal(gt_workspace=None):
    """Check calendar events for assignment deadlines.

    Pattern: GCal checks are runtime_only by default. BUT if the agent has
    populated any 'Due:' / TMA-related events in the calendar, we treat
    missing/incorrect details as BLOCKING (because the agent partially
    completed the deliverable).
    """
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY summary")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")

    # Compute expected assignment list from GT file (fallback to hardcoded list if GT missing)
    expected_assignments = []
    if gt_workspace:
        expected_assignments = _get_expected_assignments(gt_workspace)
    if not expected_assignments:
        expected_assignments = [
            "TMA 25355",
            "TMA 25356",
            "TMA 25357",
            "TMA 25358",
            "TMA 25359",
            "TMA 25360",
            "Final Exam 25361",
        ]
    expected_count = len(expected_assignments)

    # Detect whether agent populated any relevant events at all.
    # A relevant event has either 'Due:' prefix, contains an expected assignment name,
    # or matches keywords ('TMA' / 'Final Exam').
    expected_keywords = [name.lower() for name in expected_assignments]
    relevant_events = []
    for e in events:
        summary_lower = (e[0] or "").lower()
        if "due:" in summary_lower:
            relevant_events.append(e)
            continue
        if any(kw in summary_lower for kw in expected_keywords):
            relevant_events.append(e)
            continue
        if "tma" in summary_lower or "final exam" in summary_lower:
            relevant_events.append(e)
    agent_populated = len(relevant_events) > 0

    # If agent populated events, missing/wrong events are real failures (blocking).
    is_runtime_only = not agent_populated

    # Should have one event per assignment
    record(f"At least {expected_count} calendar events created",
           len(events) >= expected_count,
           f"Found {len(events)}",
           runtime_only=is_runtime_only)

    # Check that events have "Due:" prefix
    due_events = [e for e in events if "due:" in (e[0] or "").lower()]
    record("Events have 'Due:' prefix in summary",
           len(due_events) >= expected_count,
           f"Found {len(due_events)} events with 'Due:' prefix",
           runtime_only=is_runtime_only)

    all_ok = True
    for name in expected_assignments:
        found = any(name.lower() in (e[0] or "").lower() for e in events)
        record(f"Calendar event for '{name}'", found, runtime_only=is_runtime_only)
        if not found:
            all_ok = False

    # Check descriptions mention points explicitly (not just any digit)
    import re
    events_with_points = [
        e for e in events
        if e[1] and (
            "point" in e[1].lower()
            or re.search(r'\b\d+(?:\.\d+)?\s*(?:pts?|points?)\b', e[1].lower())
        )
    ]
    record("Events have points in description",
           len(events_with_points) >= 5,
           f"Found {len(events_with_points)} events with points info",
           runtime_only=is_runtime_only)

    # Per-event date/time check: if an event matches an expected assignment, its
    # start_datetime should fall on the GT due date. This is BLOCKING when agent
    # has populated events (catches "wrong-date" attacks).
    if agent_populated and gt_workspace:
        gt_file = os.path.join(gt_workspace, "Assignment_Tracker.xlsx")
        if os.path.isfile(gt_file):
            try:
                wb = openpyxl.load_workbook(gt_file, data_only=True)
                ws = get_sheet(wb, "Assignments")
                gt_due = {}
                if ws is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0] and row[1]:
                            gt_due[str(row[0]).strip().lower()] = str(row[1]).strip()
                for e in relevant_events:
                    summary_lower = (e[0] or "").lower()
                    matched_name = None
                    for name in expected_assignments:
                        if name.lower() in summary_lower:
                            matched_name = name
                            break
                    if not matched_name:
                        continue
                    expected_date = gt_due.get(matched_name.lower())
                    if expected_date and e[2] is not None:
                        actual_date = str(e[2])[:10]
                        date_ok = expected_date[:10] == actual_date
                        # When agent populated, this is BLOCKING (not runtime_only).
                        record(f"Event for '{matched_name}' on correct date",
                               date_ok,
                               f"Expected {expected_date[:10]}, got {actual_date}",
                               runtime_only=False)
                        if not date_ok:
                            all_ok = False
            except Exception as e:
                print(f"  [WARN] Could not validate per-event dates: {e}")

    return all_ok and len(events) >= expected_count


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    gcal_ok = check_gcal(gt_dir)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT} (runtime-only fails: {RUNTIME_ONLY_FAIL})")

    blocking_fail = FAIL_COUNT - RUNTIME_ONLY_FAIL
    overall = blocking_fail == 0
    print(f"  Blocking failures: {blocking_fail}")
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
