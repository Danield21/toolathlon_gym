"""Evaluation for sf-hr-job-satisfaction-gform-excel."""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_satisfaction():
    """Query actual satisfaction data from DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               COUNT(*) as emp_count,
               AVG("JOB_SATISFACTION") as avg_js,
               AVG("WORK_LIFE_BALANCE") as avg_wlb
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    rows = cur.fetchall()
    cur.execute('SELECT COUNT(*), AVG("JOB_SATISFACTION"), AVG("WORK_LIFE_BALANCE") FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    totals = cur.fetchone()
    cur.close()
    conn.close()
    return rows, totals


def check_excel(agent_workspace):
    errors = []
    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl not installed")
        return errors

    agent_file = os.path.join(agent_workspace, "Employee_Satisfaction_Analysis.xlsx")
    if not os.path.exists(agent_file):
        errors.append("Employee_Satisfaction_Analysis.xlsx not found in agent workspace")
        return errors

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

    try:
        dept_rows, totals = get_expected_satisfaction()
    except Exception as e:
        errors.append(f"Cannot query expected data: {e}")
        return errors

    total_emp = int(totals[0])
    overall_avg_js = round(float(totals[1]), 2)
    overall_avg_wlb = round(float(totals[2]), 2)

    dept_data = {}
    for r in dept_rows:
        dept = str(r[0]).strip().lower()
        dept_data[dept] = {
            "count": int(r[1]),
            "avg_js": round(float(r[2]), 2),
            "avg_wlb": round(float(r[3]), 2),
            "combined": round((float(r[2]) + float(r[3])) / 2, 2),
        }

    # Find highest/lowest by combined score
    sorted_depts = sorted(dept_data.items(), key=lambda x: x[1]["combined"], reverse=True)
    highest_dept = sorted_depts[0][0]
    lowest_dept = sorted_depts[-1][0]

    # Check Department Scores sheet
    a_rows = load_sheet_rows(agent_wb, "Department Scores")
    if a_rows is None:
        errors.append("Sheet 'Department Scores' not found in agent output")
    else:
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        # Tighten: must equal expected dept count (no extras allowed)
        expected_dept_count = len(dept_data)
        if len(a_data) != expected_dept_count:
            errors.append(f"Department Scores: expected {expected_dept_count} data rows, got {len(a_data)}")
        if len(a_data) >= 1:
            a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r[0]}
            for dept_key, expected in dept_data.items():
                a_row = a_lookup.get(dept_key)
                if a_row is None:
                    errors.append(f"Missing department: {dept_key}")
                    continue
                # Employee_Count col 1 - tighten tol from 10 to 2
                if len(a_row) > 1 and not num_close(a_row[1], expected["count"], 2):
                    errors.append(f"{dept_key} Employee_Count: got {a_row[1]}, expected {expected['count']} (tol=2)")
                # Avg_Job_Satisfaction col 2
                if len(a_row) > 2 and not num_close(a_row[2], expected["avg_js"], 0.05):
                    errors.append(f"{dept_key} Avg_Job_Satisfaction: got {a_row[2]}, expected {expected['avg_js']} (tol=0.05)")
                # Avg_Work_Life_Balance col 3
                if len(a_row) > 3 and not num_close(a_row[3], expected["avg_wlb"], 0.05):
                    errors.append(f"{dept_key} Avg_Work_Life_Balance: got {a_row[3]}, expected {expected['avg_wlb']} (tol=0.05)")
                # Combined_Score col 4
                if len(a_row) > 4 and not num_close(a_row[4], expected["combined"], 0.05):
                    errors.append(f"{dept_key} Combined_Score: got {a_row[4]}, expected {expected['combined']} (tol=0.05)")

            # Verify sort order (Combined_Score descending)
            try:
                combined_vals = [float(r[4]) for r in a_data if r[4] is not None]
                if combined_vals and combined_vals != sorted(combined_vals, reverse=True):
                    errors.append(
                        f"Department Scores not sorted by Combined_Score descending: {combined_vals}"
                    )
            except (TypeError, ValueError):
                errors.append("Department Scores Combined_Score column has non-numeric values; cannot verify sort")

    # Check Summary sheet
    a_sum = load_sheet_rows(agent_wb, "Summary")
    if a_sum is None:
        errors.append("Sheet 'Summary' not found in agent output")
    else:
        a_sum_data = {str(r[0]).strip().lower(): r[1] for r in (a_sum[1:] if len(a_sum) > 1 else []) if r and r[0]}

        # Total_Employees - tighten tol from 100 to 5 (was very wide for a fixed-DB count)
        te = a_sum_data.get("total_employees")
        if te is None:
            errors.append("Summary missing Total_Employees")
        elif not num_close(te, total_emp, 5):
            errors.append(f"Total_Employees: got {te}, expected {total_emp} (tol=5)")

        # Highest_Satisfaction_Dept
        hsd = a_sum_data.get("highest_satisfaction_dept")
        if hsd is None:
            errors.append("Summary missing Highest_Satisfaction_Dept")
        elif str(hsd).strip().lower() != highest_dept:
            errors.append(f"Highest_Satisfaction_Dept: got '{hsd}', expected '{highest_dept}'")

        # Lowest_Satisfaction_Dept
        lsd = a_sum_data.get("lowest_satisfaction_dept")
        if lsd is None:
            errors.append("Summary missing Lowest_Satisfaction_Dept")
        elif str(lsd).strip().lower() != lowest_dept:
            errors.append(f"Lowest_Satisfaction_Dept: got '{lsd}', expected '{lowest_dept}'")

        # Overall_Avg_Satisfaction
        oas = a_sum_data.get("overall_avg_satisfaction")
        if oas is None:
            errors.append("Summary missing Overall_Avg_Satisfaction")
        elif not num_close(oas, overall_avg_js, 0.05):
            errors.append(f"Overall_Avg_Satisfaction: got {oas}, expected {overall_avg_js} (tol=0.05)")

        # Overall_Avg_WLB - newly added per task spec
        oawlb = a_sum_data.get("overall_avg_wlb")
        if oawlb is None:
            errors.append("Summary missing Overall_Avg_WLB")
        elif not num_close(oawlb, overall_avg_wlb, 0.05):
            errors.append(f"Overall_Avg_WLB: got {oawlb}, expected {overall_avg_wlb} (tol=0.05)")

    return errors


def check_gform():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Title must be Employee Wellbeing Survey 2026 (case-insensitive, allow extra whitespace)
        cur.execute("""
            SELECT f.id, f.title
            FROM gform.forms f
            WHERE LOWER(f.title) LIKE '%employee%wellbeing%survey%2026%'
        """)
        forms = cur.fetchall()
        if not forms:
            errors.append("No Google Form titled 'Employee Wellbeing Survey 2026' found")
            cur.close()
            conn.close()
            return errors
        form_id = forms[0][0]
        cur.execute("""
            SELECT id, title, question_type, required, config
            FROM gform.questions
            WHERE form_id = %s ORDER BY position
        """, (form_id,))
        qs = cur.fetchall()
        if len(qs) < 4:
            errors.append(f"Form has {len(qs)} questions, expected at least 4")
            cur.close()
            conn.close()
            return errors

        titles = [(q[1] or "").lower() for q in qs]
        types = [(q[2] or "").upper() for q in qs]
        required = [bool(q[3]) for q in qs]

        # Q1: required text question about department
        q1_ok = any(
            "department" in t and types[i] in ("TEXT", "SHORT_ANSWER", "PARAGRAPH") and required[i]
            for i, t in enumerate(titles)
        )
        if not q1_ok:
            errors.append(
                f"Missing required text question for department. "
                f"Questions: {list(zip(titles, types, required))}"
            )

        # Q2 + Q3: required 1-5 rating-style for role satisfaction and work-life balance
        # Look for SCALE / LINEAR_SCALE / RATING type or MULTIPLE_CHOICE with 1..5 choices
        def _is_rating(q):
            qtype = (q[2] or "").upper()
            cfg = q[4] or {}
            if isinstance(cfg, str):
                try:
                    cfg = json.loads(cfg)
                except (TypeError, ValueError):
                    cfg = {}
            choices = []
            if isinstance(cfg, dict):
                choices = cfg.get("choices") or cfg.get("options") or []
                # Linear scale config typically has min/max
                lo = cfg.get("low") or cfg.get("min")
                hi = cfg.get("high") or cfg.get("max")
                if (lo is not None and hi is not None
                        and int(lo) == 1 and int(hi) == 5):
                    return True
            if qtype in ("SCALE", "LINEAR_SCALE", "RATING", "STAR_RATING"):
                return True
            if qtype in ("RADIO", "MULTIPLE_CHOICE", "CHOICE"):
                # Need 1..5 choices
                vals = [str(c).strip() for c in choices]
                if len(vals) >= 5 and all(str(i) in vals for i in range(1, 6)):
                    return True
            return False

        rating_qs = [(i, q) for i, q in enumerate(qs) if _is_rating(q)]
        rating_topics_found = {"role": False, "wlb": False}
        for i, q in rating_qs:
            t = titles[i]
            if not required[i]:
                continue
            # role satisfaction (also accept 'position' as a synonym)
            if "satisf" in t and ("role" in t or "current role" in t or "job" in t or "position" in t):
                rating_topics_found["role"] = True
            elif ("work-life" in t or "work life" in t or "work/life" in t or "wlb" in t):
                rating_topics_found["wlb"] = True
        if not rating_topics_found["role"]:
            errors.append(
                f"Missing required 1-5 rating question about role satisfaction. "
                f"Rating questions: {[(titles[i], types[i], required[i]) for i, _ in rating_qs]}"
            )
        if not rating_topics_found["wlb"]:
            errors.append(
                f"Missing required 1-5 rating question about work-life balance. "
                f"Rating questions: {[(titles[i], types[i], required[i]) for i, _ in rating_qs]}"
            )

        # Q4: optional text question about improvements
        q4_ok = any(
            ("improv" in t or "suggest" in t)
            and types[i] in ("TEXT", "SHORT_ANSWER", "PARAGRAPH")
            for i, t in enumerate(titles)
        )
        if not q4_ok:
            errors.append(
                f"Missing text question about workplace improvements/suggestions. "
                f"Questions: {list(zip(titles, types, required))}"
            )

        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"GForm DB check error: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE LOWER(to_addr::text) LIKE '%hr.team@company.com%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
        if not emails:
            errors.append("No email sent to hr.team@company.com")
            return errors
        # Among hr.team emails, at least one must reference Q1 2026 + satisfaction
        found_subj = False
        found_body = False
        for subj, to_addr, body in emails:
            sl = (subj or "").lower()
            bl = (body or "").lower()
            if ("q1 2026" in sl or "q1, 2026" in sl or "q1-2026" in sl) and \
               ("satisfaction" in sl or "wellbeing" in sl):
                found_subj = True
            # Body should mention the Excel report or its filename
            if ("excel" in bl or "employee_satisfaction_analysis" in bl
                    or "spreadsheet" in bl or "report" in bl):
                found_body = True
        if not found_subj:
            errors.append(
                f"No email subject references Q1 2026 employee satisfaction analysis. "
                f"Subjects: {[s for s, _, _ in emails]}"
            )
        if not found_body:
            errors.append(
                f"No email body mentions the Excel report. "
                f"Bodies (first 200): {[(b or '')[:200] for _, _, b in emails]}"
            )
    except Exception as e:
        errors.append(f"Email DB check error: {e}")
    return errors


def check_notion():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Single-substring exact phrase to prevent loose 3-substring co-occurrence
        # ('Q1 2025 (and 2026) Employee Satisfaction Notes' would have falsely passed).
        cur.execute("""
            SELECT id, properties
            FROM notion.pages
            WHERE LOWER(properties::text) LIKE '%employee satisfaction q1 2026%'
        """)
        pages = cur.fetchall()
        if not pages:
            errors.append("No Notion page titled 'Employee Satisfaction Q1 2026' found")
            cur.close()
            conn.close()
            return errors

        # Verify page has block content with key findings (mention dept names or numbers)
        page_id = pages[0][0]
        cur.execute("""
            SELECT block_data::text FROM notion.blocks
            WHERE LOWER(block_data::text) ILIKE %s
               OR parent_id::text = %s
        """, (f"%{page_id}%", str(page_id)))
        blocks = cur.fetchall()
        block_text = " ".join(b[0].lower() for b in blocks if b[0])
        # Key findings: require >=2 of {highest, lowest, avg, score} co-occurring
        # with 'department' or an actual department name keyword.
        analysis_kws = ["highest", "lowest", "avg", "average", "score"]
        analysis_hits = sum(1 for k in analysis_kws if k in block_text)
        # Department-related context: keyword 'department' OR common dept names.
        dept_kws = ["department", "engineering", "sales", "marketing", "finance",
                    "operations", "support", "hr", "human resources"]
        has_dept = any(k in block_text for k in dept_kws)
        has_findings = analysis_hits >= 2 and has_dept
        if not has_findings:
            errors.append(
                f"Notion page 'Employee Satisfaction Q1 2026' lacks key-findings content "
                f"(analysis_hits={analysis_hits}/{len(analysis_kws)}, has_dept={has_dept}). "
                f"Sample block text (first 300): {block_text[:300]}"
            )
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Notion DB check error: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    print("\n=== Checking Google Form ===")
    gform_errors = check_gform()
    if gform_errors:
        for e in gform_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(gform_errors)
    else:
        print("  [PASS] GForm check passed")

    print("\n=== Checking Email ===")
    email_errors = check_email()
    if email_errors:
        for e in email_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(email_errors)
    else:
        print("  [PASS] Email check passed")

    print("\n=== Checking Notion Page ===")
    notion_errors = check_notion()
    if notion_errors:
        for e in notion_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(notion_errors)
    else:
        print("  [PASS] Notion check passed")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"errors": all_errors, "success": len(all_errors) == 0}, f, indent=2)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
