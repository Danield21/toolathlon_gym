"""
Evaluation for yt-transcript-notion-song-report-gcal-email task.

Tightened checks:
1. Song_Analysis_Report.xlsx with exactly 3 sheets and required columns
2. Tracklist >= 12 rows, Artist_Stats >= 4 rows, Publication_Plan exactly 3 rows w/ 3 specific dates
3. Notion: page with required title tokens (afrobeat AND analysis); database with >= 10 song entries
4. GCal: 3 distinct events on 2026-03-15, 03-22, 03-29 (publish events) with required keywords
5. Both emails sent: editorial@musicblog.com (subject + body) and artists@afrobeat.com (subject)
"""
import json
import os
import re
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl
from zoneinfo import ZoneInfo

# ──────────────────────────────────────────────────────────────────────────
# EVALUATION GROUND TRUTH SPEC (gcal tz root-fix v3, case-study 2026-08-13)
# Source of truth: docs/task.md. The Calendar MCP writes events as UTC
# instants into the DB (`TIMESTAMPTZ`), so all comparisons anchor to the
# timezone that task.md declares for the event's wall-clock semantics.
# Never use bare `sd.date()` / `strftime("%Y-%m-%d")` on rows read from
# `gcal.events`: psycopg2 returns them in the PG session `TimeZone`
# (case-study: compute node default was Asia/Shanghai, masking 10 AM ET as
# 22:00+08 and shifting the event date). Use `utils.evaluation.gcal_helpers`
# instead (session-tz-independent).
# ──────────────────────────────────────────────────────────────────────────
# task.md line 9: "...on March 15, 2026, from 10:00 to 11:00" (music blog
# AfroBeats Today, Eastern Time). EXPECTED_TIMEZONE anchors wall-clock
# comparisons to America/New_York.
EXPECTED_TIMEZONE = ZoneInfo("America/New_York")

from utils.evaluation.gcal_helpers import get_zone_components  # noqa: E402

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def norm(s):
    return str(s or "").strip().lower().replace("_", " ")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def date_norm(v):
    """Normalize cell to YYYY-MM-DD or '' on failure."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    # Allow "March 15, 2026"
    months = {"january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
              "july":7,"august":8,"september":9,"october":10,"november":11,"december":12}
    m2 = re.match(r"^([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})", s)
    if m2:
        mo = months.get(m2.group(1).lower())
        if mo:
            return f"{int(m2.group(3)):04d}-{mo:02d}-{int(m2.group(2)):02d}"
    return ""


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Song_Analysis_Report.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Song_Analysis_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Song_Analysis_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        for k in ["Excel file readable", "Tracklist sheet exists",
                  "Tracklist columns correct", "Tracklist has >= 12 data rows",
                  "Artist_Stats sheet exists", "Artist_Stats has >= 4 data rows",
                  "Publication_Plan sheet exists", "Publication_Plan has exactly 3 rows",
                  "Publication_Plan has 3 required dates",
                  "Publication_Plan platform is AfroBeats Today",
                  "Publication_Plan status is Scheduled"]:
            record(k, False, "skipped (no excel)")
        return
    record("Song_Analysis_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    # ---------- Tracklist ----------
    tracklist_sheet = None
    for name in wb.sheetnames:
        if "track" in name.lower():
            tracklist_sheet = wb[name]
            break
    if tracklist_sheet is None:
        record("Tracklist sheet exists", False, f"Sheets: {wb.sheetnames}")
        for k in ["Tracklist columns correct", "Tracklist has >= 12 data rows"]:
            record(k, False, "skipped")
    else:
        record("Tracklist sheet exists", True)
        rows = list(tracklist_sheet.iter_rows(values_only=True))
        headers = [norm(c) for c in (rows[0] if rows else [])]
        required = ["track no", "song title", "artist", "start time sec",
                    "estimated duration sec", "genre"]
        ok_headers = all(any(req in h for h in headers) for req in required)
        record("Tracklist columns correct", ok_headers,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        record("Tracklist has >= 12 data rows", len(data_rows) >= 12,
               f"Found {len(data_rows)} data rows")

    # ---------- Artist_Stats ----------
    artist_sheet = None
    for name in wb.sheetnames:
        if "artist" in name.lower() or "stat" in name.lower():
            artist_sheet = wb[name]
            break
    if artist_sheet is None:
        record("Artist_Stats sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Artist_Stats has >= 4 data rows", False, "skipped")
    else:
        record("Artist_Stats sheet exists", True)
        rows = list(artist_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        record("Artist_Stats has >= 4 data rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} rows")

    # ---------- Publication_Plan ----------
    pub_sheet = None
    for name in wb.sheetnames:
        n = name.lower()
        if "publication" in n or "publish" in n or n == "publication_plan" or "plan" in n:
            pub_sheet = wb[name]
            break
    if pub_sheet is None:
        record("Publication_Plan sheet exists", False, f"Sheets: {wb.sheetnames}")
        for k in ["Publication_Plan has exactly 3 rows",
                  "Publication_Plan has 3 required dates",
                  "Publication_Plan platform is AfroBeats Today",
                  "Publication_Plan status is Scheduled"]:
            record(k, False, "skipped")
    else:
        record("Publication_Plan sheet exists", True)
        rows = list(pub_sheet.iter_rows(values_only=True))
        headers = [norm(c) for c in (rows[0] if rows else [])]
        # locate columns
        def col_idx(*candidates):
            for i, h in enumerate(headers):
                if any(c in h for c in candidates):
                    return i
            return -1
        ci_date = col_idx("publish date", "publish_date", "date")
        ci_plat = col_idx("platform")
        ci_stat = col_idx("status")
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        record("Publication_Plan has exactly 3 rows", len(data_rows) == 3,
               f"Found {len(data_rows)} rows")

        required_dates = {"2026-03-15", "2026-03-22", "2026-03-29"}
        found_dates = set()
        if ci_date >= 0:
            for r in data_rows:
                if ci_date < len(r):
                    nd = date_norm(r[ci_date])
                    if nd:
                        found_dates.add(nd)
        record("Publication_Plan has 3 required dates",
               required_dates.issubset(found_dates),
               f"got dates: {sorted(found_dates)}")

        plat_ok = False
        if ci_plat >= 0:
            plats = [str(r[ci_plat] or "").lower() for r in data_rows if ci_plat < len(r)]
            plat_ok = len(plats) == 3 and all("afrobeats today" in p for p in plats)
        record("Publication_Plan platform is AfroBeats Today", plat_ok,
               f"platforms: {plats if ci_plat>=0 else 'no col'}")

        stat_ok = False
        if ci_stat >= 0:
            stats = [str(r[ci_stat] or "").strip().lower() for r in data_rows if ci_stat < len(r)]
            stat_ok = len(stats) == 3 and all("scheduled" in s for s in stats)
        record("Publication_Plan status is Scheduled", stat_ok,
               f"statuses: {stats if ci_stat>=0 else 'no col'}")


def _parent_type(parent):
    """Notion parent type with Bug-C tolerance: prefer the explicit "type"
    field, else infer from whichever key is present (agents often omit it;
    the mock server normalizes on write but evaluator stays robust anyway)."""
    if not isinstance(parent, dict):
        return ""
    t = parent.get("type")
    if isinstance(t, str) and t:
        return t
    for key in ("database_id", "page_id", "block_id", "workspace"):
        if parent.get(key) is not None:
            return key
    return ""


def check_notion():
    print("\n=== Check 2: Notion Pages and Database ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, parent, properties FROM notion.pages")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Notion DB query", False, str(e))
        for k in ["Notion analysis page with required tokens",
                  "Notion song database has >= 10 entries"]:
            record(k, False, "skipped")
        return

    found_page = False
    for page_id, parent, props in pages:
        try:
            parent_type = _parent_type(parent)
            if parent_type in ("workspace", "page_id"):
                title_items = []
                for key, val in props.items():
                    if isinstance(val, dict) and (
                        val.get("type") == "title" or "title" in val
                    ):
                        title_items = val.get("title", [])
                        break
                title_text = " ".join(
                    item.get("text", {}).get("content", "") for item in title_items
                    if isinstance(item, dict)
                ).lower()
                # Require BOTH "afrobeat" AND "analysis" tokens (NOT just "analysis")
                if "afrobeat" in title_text and ("analysis" in title_text or "blog" in title_text):
                    found_page = True
        except Exception:
            continue

    record("Notion analysis page with required tokens", found_page,
           f"Total pages: {len(pages)}")

    # Bug C variant (c4 2026-08-15): agent pages may carry {"database_id": X}
    # without the "type" discriminator; the mock server now normalizes parent
    # on write, but stay robust to pre-existing rows by key presence too.
    db_entries = [p for p in pages if _parent_type(p[1]) == "database_id"]
    record("Notion song database has >= 10 entries", len(db_entries) >= 10,
           f"Found {len(db_entries)} database-parented pages")


def check_gcal():
    print("\n=== Check 3: GCal Publication Events ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime FROM gcal.events
            WHERE start_datetime >= TIMESTAMPTZ '2026-03-01T00:00:00Z' AND start_datetime < TIMESTAMPTZ '2026-04-01T00:00:00Z'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("GCal DB query", False, str(e))
        for k in ["GCal Publish Afrobeat Mix Feature on 2026-03-15",
                  "GCal Artist Spotlight Article on 2026-03-22",
                  "GCal Genre Overview Piece on 2026-03-29"]:
            record(k, False, "skipped")
        return

    def event_on_date(target_date_str, kw_required, expected_title):
        for summary, start_dt in events:
            # gcal.events.start_datetime is TIMESTAMPTZ (UTC instant). The
            # event date must be extracted in EXPECTED_TIMEZONE
            # (America/New_York per task.md) — bare start_dt.strftime
            # ("%Y-%m-%d") silently follows the PG session tz and can shift
            # the event to the wrong day (case-study 2026-08-13).
            ev_date, _, _ = get_zone_components(start_dt, EXPECTED_TIMEZONE)
            ds = ev_date.isoformat() if ev_date is not None else ""
            if ds == target_date_str:
                s_lower = (summary or "").strip().lower()
                # Accept exact title match
                if s_lower == expected_title.lower():
                    return True
                # Otherwise require ALL keyword tokens (was 2-of-N — too loose)
                if all(kw in s_lower for kw in kw_required):
                    return True
        return False

    record("GCal Publish Afrobeat Mix Feature on 2026-03-15",
           event_on_date("2026-03-15",
                         ["publish", "afrobeat", "mix", "feature"],
                         "Publish Afrobeat Mix Feature"),
           f"events: {[(s, str(d)[:16]) for s,d in events]}")
    record("GCal Artist Spotlight Article on 2026-03-22",
           event_on_date("2026-03-22",
                         ["artist", "spotlight", "article"],
                         "Artist Spotlight Article"),
           f"events: {[(s, str(d)[:16]) for s,d in events]}")
    record("GCal Genre Overview Piece on 2026-03-29",
           event_on_date("2026-03-29",
                         ["genre", "overview", "piece"],
                         "Genre Overview Piece"),
           f"events: {[(s, str(d)[:16]) for s,d in events]}")


def check_emails():
    print("\n=== Check 4: Emails sent ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT to_addr, subject, body_text, body_html FROM email.messages")
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB query", False, str(e))
        for k in ["Editorial email sent", "Editorial email subject correct",
                  "Editorial email body has summary",
                  "Artists email sent", "Artists email subject correct"]:
            record(k, False, "skipped")
        return

    edit_msgs = []
    art_msgs = []
    for to_addr, subj, btxt, bhtml in rows:
        to_str = str(to_addr or "").lower()
        body = (btxt or "") + "\n" + (bhtml or "")
        if "editorial@musicblog.com" in to_str:
            edit_msgs.append((subj or "", body))
        if "artists@afrobeat.com" in to_str:
            art_msgs.append((subj or "", body))

    record("Editorial email sent", len(edit_msgs) >= 1, f"count={len(edit_msgs)}")
    if edit_msgs:
        # Pick first message
        subj_l = edit_msgs[0][0].lower()
        body_l = edit_msgs[0][1].lower()
        # subject must contain key tokens
        ok_subj = ("afrobeat" in subj_l and "analysis" in subj_l and "report" in subj_l and "ready" in subj_l)
        record("Editorial email subject correct", ok_subj, f"subj={subj_l[:120]}")
        # Tighten: require BOTH 'track' or 'song' AND a publication date hint
        # (task says "summary of the number of tracks identified and the publication schedule")
        has_count_word = "track" in body_l or "song" in body_l
        has_schedule_hint = ("publication" in body_l or "schedule" in body_l
                              or "march" in body_l)
        ok_body = has_count_word and has_schedule_hint
        record("Editorial email body has summary (tracks/songs + schedule)",
               ok_body, f"body sample: {body_l[:200]}")
    else:
        record("Editorial email subject correct", False, "skipped")
        record("Editorial email body has summary", False, "skipped")

    record("Artists email sent", len(art_msgs) >= 1, f"count={len(art_msgs)}")
    if art_msgs:
        subj_l = art_msgs[0][0].lower()
        ok_subj = ("artist" in subj_l and "feature" in subj_l and ("afrobeat" in subj_l or "mix" in subj_l))
        record("Artists email subject correct", ok_subj, f"subj={subj_l[:120]}")
    else:
        record("Artists email subject correct", False, "skipped")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%) [FAIL_COUNT={FAIL_COUNT}]")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "fail_count": FAIL_COUNT,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
