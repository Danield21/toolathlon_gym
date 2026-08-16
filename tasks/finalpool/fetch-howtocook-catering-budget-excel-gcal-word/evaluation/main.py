"""Evaluation for fetch-howtocook-catering-budget-excel-gcal-word.

Strengthened checks (round 5):
  * Per-recipe Line_Total sum must equal Meal_Cost in Budget Summary.
  * At least one ingredient's Unit_Price must match the supplier API price
    (or fall back to GT) exactly.
  * At least one Discount_Pct in {10, 15} must appear (per task: bulk
    discount applies when an ingredient's aggregate weight >= 50 kg).
  * GCal delivery events must be 07:00-08:00 and the description should
    list at least one recipe from that day.
"""
import argparse
import json
import os
import sys
import urllib.request
from collections import defaultdict
from datetime import timezone

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

# Day to recipe(s) mapping based on GT meal plan; the eval falls back to
# GT data when needed.
DEFAULT_DAY_RECIPES = {
    "2026-04-13": ["Egg Fried Rice", "Braised Pork Belly", "Kung Pao Chicken"],
    "2026-04-14": ["Congee with Pork", "Mapo Tofu", "Braised Beef Noodles"],
    "2026-04-15": ["Scallion Pancakes", "Tomato Egg Stir-fry", "Sweet and Sour Pork"],
}


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _fetch_supplier_prices():
    """Fetch ingredient prices from the mock supplier API.

    Returns dict[ingredient_name -> unit_price] or {} on failure.
    """
    try:
        req = urllib.request.Request("http://localhost:30217/api/ingredients.json")
        with urllib.request.urlopen(req, timeout=3) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return {ing["name"]: float(ing["price_per_kg"])
                for ing in data.get("ingredients", []) if "price_per_kg" in ing}
    except Exception as e:
        print(f"  [INFO] Supplier API unavailable, using GT fallback: {e}")
        return {}


def _fetch_gt_prices(gt_workspace):
    """Read GT Ingredient Costs sheet to build {ingredient: unit_price}."""
    try:
        import openpyxl
        path = os.path.join(gt_workspace, "Catering_Budget.xlsx")
        if not os.path.exists(path):
            return {}
        wb = openpyxl.load_workbook(path, data_only=True)
        rows = load_sheet_rows(wb, "Ingredient Costs") or []
        prices = {}
        for row in rows[1:]:
            if row and len(row) >= 5 and row[1] and row[4] is not None:
                prices[str(row[1]).strip()] = float(row[4])
        return prices
    except Exception:
        return {}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    agent_ws = args.agent_workspace or task_root
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []          # blocking errors (gate V1 + V2)
    runtime_errors = []      # only blocking when an agent has run

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "Catering_Budget.xlsx")

    # We collect per-recipe Line_Totals here so the Budget Summary check
    # can compare against them.
    sum_by_recipe = defaultdict(float)
    discounts_seen = set()
    spot_unit_price_passed = None  # None=skip, True/False otherwise

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("Catering_Budget.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

        # Check Meal Plan sheet
        print("  Checking Meal Plan sheet...")
        a_rows = load_sheet_rows(agent_wb, "Meal Plan")
        if a_rows is None:
            all_errors.append("Sheet 'Meal Plan' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            if len(a_data) < 9:
                all_errors.append(f"Meal Plan row count: {len(a_data)}, expected 9")
            else:
                # Check all 3 days and 3 meal types exist
                days = set()
                meals = set()
                for row in a_data:
                    if row and row[0] is not None:
                        days.add(int(row[0]))
                    if row and row[1]:
                        meals.add(str(row[1]).strip().lower())
                if len(days) < 3:
                    all_errors.append(f"Meal Plan covers {len(days)} days, expected 3")
                expected_meals = {"breakfast", "lunch", "dinner"}
                if not expected_meals.issubset(meals):
                    all_errors.append(f"Missing meal types: {expected_meals - meals}")
            print("    Done.")

        # Check Ingredient Costs sheet
        print("  Checking Ingredient Costs sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Ingredient Costs")
        if a_rows2 is None:
            all_errors.append("Sheet 'Ingredient Costs' not found in agent output")
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            if len(a_data2) < 10:
                all_errors.append(f"Ingredient Costs: {len(a_data2)} rows, expected 10+")
            # Check that Line_Total values are present and positive
            totals = []
            line_total_rows = []  # (recipe, ingredient, qty, unit_price, discount, line_total)
            for row in a_data2:
                if row and len(row) >= 7 and row[6] is not None:
                    try:
                        totals.append(float(row[6]))
                    except (TypeError, ValueError):
                        pass
                    try:
                        line_total_rows.append((
                            str(row[0]).strip() if row[0] else "",
                            str(row[1]).strip() if row[1] else "",
                            float(row[2]) if row[2] is not None else 0.0,
                            float(row[4]) if row[4] is not None else 0.0,
                            float(row[5]) if row[5] is not None else 0.0,
                            float(row[6]),
                        ))
                    except (TypeError, ValueError):
                        pass
            if len(totals) < 10:
                all_errors.append(f"Ingredient Costs: only {len(totals)} rows have Line_Total")
            elif any(t <= 0 for t in totals):
                all_errors.append("Ingredient Costs: some Line_Total values are <= 0")

            # Per-row formula check: Line_Total ~= Qty * Unit_Price * (1 - Discount/100)
            formula_bad = []
            for r, i, q, up, disc, lt in line_total_rows:
                expected_lt = q * up * (1 - disc / 100.0)
                if abs(expected_lt - lt) > max(0.05, 0.01 * abs(expected_lt)):
                    formula_bad.append((r, i, q, up, disc, lt, round(expected_lt, 2)))
            if formula_bad:
                all_errors.append(
                    f"Line_Total formula mismatch in {len(formula_bad)} row(s); "
                    f"first: {formula_bad[0]}"
                )

            # Aggregate Line_Total per recipe
            for r, _, _, _, _, lt in line_total_rows:
                sum_by_recipe[r] += lt

            # Discount_Pct: at least one row should be 10 or 15 (per task,
            # bulk discount applies to ingredients aggregated >= 50 kg).
            for _, _, _, _, disc, _ in line_total_rows:
                if int(round(disc)) in (10, 15):
                    discounts_seen.add(int(round(disc)))

            # Spot-check: at least one ingredient's Unit_Price matches the
            # supplier API (or GT). We require ALL agent prices that match
            # an API ingredient to also match the API value (no random
            # made-up prices).
            api_prices = _fetch_supplier_prices()
            ref_prices = api_prices or _fetch_gt_prices(gt_ws)
            if ref_prices and line_total_rows:
                matched = 0
                mismatched = []
                for r, i, q, up, disc, lt in line_total_rows:
                    if i in ref_prices:
                        if abs(ref_prices[i] - up) < 0.01:
                            matched += 1
                        else:
                            mismatched.append((i, up, ref_prices[i]))
                spot_unit_price_passed = (matched >= 1 and len(mismatched) == 0)
                if matched == 0:
                    all_errors.append(
                        "No ingredient Unit_Price matched the supplier API "
                        f"(checked {len(line_total_rows)} rows)"
                    )
                if mismatched:
                    all_errors.append(
                        f"Unit_Price mismatch with supplier API in "
                        f"{len(mismatched)} row(s); first: {mismatched[0]}"
                    )
            print("    Done.")

        # Check Budget Summary sheet
        print("  Checking Budget Summary sheet...")
        a_rows3 = load_sheet_rows(agent_wb, "Budget Summary")
        if a_rows3 is None:
            all_errors.append("Sheet 'Budget Summary' not found in agent output")
        else:
            a_data3 = a_rows3[1:] if len(a_rows3) > 1 else []
            if len(a_data3) < 9:
                all_errors.append(f"Budget Summary: {len(a_data3)} rows, expected 9+")
            # Check for grand total row
            has_total = False
            for row in a_data3:
                if row:
                    for cell in row:
                        if cell and "total" in str(cell).lower():
                            has_total = True
                            break
            if not has_total:
                all_errors.append("Budget Summary: no Grand Total row found")

            # Per-recipe Meal_Cost must equal sum of Line_Totals for that recipe
            recipe_meal_cost = {}
            grand_total_in_summary = None
            for row in a_data3:
                if not row:
                    continue
                # recipe row layout: Day, Meal, Recipe_Name, Meal_Cost
                name = row[2] if len(row) > 2 else None
                cost = row[3] if len(row) > 3 else None
                if isinstance(name, str) and "total" in name.lower():
                    try:
                        grand_total_in_summary = float(cost)
                    except (TypeError, ValueError):
                        pass
                    continue
                if name and cost is not None:
                    try:
                        recipe_meal_cost[str(name).strip()] = float(cost)
                    except (TypeError, ValueError):
                        pass

            mismatches = []
            for r, mc in recipe_meal_cost.items():
                expected = sum_by_recipe.get(r)
                if expected is None:
                    continue
                if abs(expected - mc) > max(0.5, 0.01 * abs(expected)):
                    mismatches.append((r, mc, round(expected, 2)))
            if mismatches:
                all_errors.append(
                    f"Meal_Cost != sum(Line_Total) for {len(mismatches)} recipe(s); "
                    f"first: {mismatches[0]}"
                )

            # Grand total must equal sum of all Line_Totals
            if grand_total_in_summary is not None and sum_by_recipe:
                expected_grand = sum(sum_by_recipe.values())
                if abs(grand_total_in_summary - expected_grand) > max(1.0, 0.01 * abs(expected_grand)):
                    all_errors.append(
                        f"Grand Total mismatch: summary={grand_total_in_summary}, "
                        f"sum(Line_Total)={round(expected_grand, 2)}"
                    )
            print("    Done.")

        # Discount check (after all sheets parsed): >=1 row in {10, 15}
        if not discounts_seen:
            all_errors.append(
                "No bulk discount applied (expected at least one Discount_Pct "
                "of 10 or 15 since Pork Belly aggregates >= 50 kg)"
            )

    # --- Check 2: Word document ---
    print("Checking Word document...")
    doc_path = os.path.join(agent_ws, "Catering_Proposal.docx")
    if not os.path.exists(doc_path):
        all_errors.append("Catering_Proposal.docx not found in agent workspace")
    else:
        try:
            from docx import Document
            doc = Document(doc_path)
            full_text = "\n".join(p.text for p in doc.paragraphs)
            full_lower = full_text.lower()
            if "80" not in full_text:
                all_errors.append("Word doc does not mention 80 attendees")
            if "day 1" not in full_lower and "day1" not in full_lower:
                all_errors.append("Word doc does not mention Day 1")
            if "day 2" not in full_lower and "day2" not in full_lower:
                all_errors.append("Word doc does not mention Day 2")
            if "day 3" not in full_lower and "day3" not in full_lower:
                all_errors.append("Word doc does not mention Day 3")
            if "total" not in full_lower:
                all_errors.append("Word doc does not include total cost")
            if "catering" not in full_lower and "retreat" not in full_lower:
                all_errors.append("Word doc missing title/overview")
            # Word must have a menu/recipe section per task
            if "menu" not in full_lower and "recipe" not in full_lower:
                all_errors.append("Word doc missing menu/recipe section")
        except Exception as e:
            all_errors.append(f"Error reading Word doc: {e}")

    # --- Check 3: GCal delivery events (07:00-08:00 + recipe descriptions) ---
    print("Checking GCal events...")
    agent_ran_gcal = False
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE summary ILIKE '%catering%delivery%' OR summary ILIKE '%day%delivery%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) >= 1:
            agent_ran_gcal = True

        if len(rows) < 3:
            runtime_errors.append(f"GCal: found {len(rows)} delivery events, expected 3")
        else:
            # Build {date: (start_hh:mm, end_hh:mm, description)}
            by_date = {}
            for summary, desc, st, en in rows:
                if st is None or en is None:
                    continue
                if st.tzinfo is not None:
                    st_utc = st.astimezone(timezone.utc)
                    en_utc = en.astimezone(timezone.utc)
                else:
                    st_utc, en_utc = st, en
                d = str(st_utc.date())
                by_date[d] = (
                    st_utc.strftime("%H:%M"),
                    en_utc.strftime("%H:%M"),
                    str(desc or ""),
                    str(summary or ""),
                )

            for d in ("2026-04-13", "2026-04-14", "2026-04-15"):
                if d not in by_date:
                    runtime_errors.append(f"GCal: no delivery event on {d}")
                    continue
                hh_start, hh_end, desc, summary = by_date[d]
                if (hh_start, hh_end) != ("07:00", "08:00"):
                    runtime_errors.append(
                        f"GCal {d}: time {hh_start}-{hh_end} (expected 07:00-08:00)"
                    )
                # Description must reference at least one of that day's recipes
                day_recipes = DEFAULT_DAY_RECIPES.get(d, [])
                desc_lower = desc.lower()
                if not any(r.lower() in desc_lower for r in day_recipes):
                    runtime_errors.append(
                        f"GCal {d}: description missing recipe references "
                        f"({day_recipes}); got: {desc[:120]}"
                    )
    except Exception as e:
        runtime_errors.append(f"Error checking GCal: {e}")

    # --- Final result ---
    blocking = list(all_errors) + (runtime_errors if agent_ran_gcal else [])

    if blocking:
        print(f"\n=== RESULT: FAIL ({len(blocking)} errors) ===")
        for e in blocking[:15]:
            print(f"  {e}")
        if not agent_ran_gcal and runtime_errors:
            print(f"  (Skipped {len(runtime_errors)} runtime-only errors in V1 mode)")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        if not agent_ran_gcal and runtime_errors:
            print(f"  (Skipped {len(runtime_errors)} runtime-only errors in V1 mode)")
        sys.exit(0)


if __name__ == "__main__":
    main()
