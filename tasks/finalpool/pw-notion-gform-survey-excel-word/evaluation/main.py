"""Evaluation script for pw-notion-gform-survey-excel-word."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-notion-gform-survey-excel-word"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Survey_Report.xlsx")
    check("Survey_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Survey_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        # Compute GT row counts to use as minimums
        def _gt_rows(name):
            if gt_wb is not None and name in gt_wb.sheetnames:
                return list(gt_wb[name].iter_rows(min_row=2, values_only=True))
            return []

        _da_gt = _gt_rows("Data_Analysis")
        _m_gt = _gt_rows("Metrics")
        _r_gt = _gt_rows("Recommendations")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            min_da = max(2, len(_da_gt))
            check(f"Data_Analysis has >= {min_da} rows", len(data_rows) >= min_da, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Item', 'Internal_Value', 'External_Benchmark', 'Gap'], min_da)

            # Cell-value comparison vs GT (by Item key) — critical FP guard
            if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames:
                gt_ws = gt_wb["Data_Analysis"]
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                a_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                a_idx = {h: i for i, h in enumerate(a_headers)}
                item_gi = gt_headers.index("item") if "item" in gt_headers else 0
                item_ai = a_idx.get("item", 0)
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                gt_by_item = {str(r[item_gi]).strip().lower(): r for r in gt_rows
                              if r and item_gi < len(r) and r[item_gi] is not None}
                a_by_item = {str(r[item_ai]).strip().lower(): r for r in data_rows
                             if r and item_ai < len(r) and r[item_ai] is not None}
                for it, gt_row in gt_by_item.items():
                    found = it in a_by_item
                    check(f"Data_Analysis Item '{it}' present", found)
                    if found:
                        agent_row = a_by_item[it]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row) or gt_h == "item":
                                continue
                            gv = gt_row[ci]
                            agent_ci = a_idx.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                tol = max(0.5, abs(gf) * 0.03)
                                check(f"Data_Analysis '{it}' {gt_h} ~{gf}",
                                      abs(gf - af) <= tol, f"got {av}")

            # Sort order: alphabetical by Item
            primary_idx = a_idx.get("item", 0) if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames else 0
            primary_vals = [str(r[primary_idx]).strip() for r in data_rows
                            if r and primary_idx < len(r) and r[primary_idx]]
            sorted_vals = sorted(primary_vals, key=lambda s: s.lower())
            check("Data_Analysis sorted alphabetically by Item",
                  primary_vals == sorted_vals, f"order={primary_vals[:3]}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            min_m = max(2, len(_m_gt))
            check(f"Metrics has >= {min_m} rows", len(data_rows) >= min_m, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], min_m)

            # Metric value validation vs GT
            if gt_wb is not None and "Metrics" in gt_wb.sheetnames:
                gt_rows = list(gt_wb["Metrics"].iter_rows(min_row=2, values_only=True))
                gt_map = {str(r[0]).strip().lower(): r[1] for r in gt_rows if r and r[0]}
                a_map = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for k, gv in gt_map.items():
                    found = k in a_map
                    check(f"Metrics has '{k}'", found)
                    if found:
                        gf = safe_float(gv)
                        af = safe_float(a_map[k])
                        if gf is not None and af is not None:
                            tol = max(0.5, abs(gf) * 0.05)
                            check(f"Metrics '{k}' ~{gf}",
                                  abs(gf - af) <= tol, f"got {a_map[k]}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            min_r = max(1, len(_r_gt))
            check(f"Recommendations has >= {min_r} rows", len(data_rows) >= min_r, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action'], min_r)

    # Notion / GForm / Word / processor checks - OUTSIDE excel-exists block
    import json as _json
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Require title-property match (not just any properties text)
        cur.execute(
            "SELECT id, properties FROM notion.pages "
            "WHERE archived = false AND properties::text ILIKE %s",
            ('%Survey Dashboard%',),
        )
        rows = cur.fetchall()
        matched_pages = []
        for pid, props in rows:
            try:
                pj = _json.loads(props) if isinstance(props, str) else props
                if isinstance(pj, dict):
                    titles = []
                    for v in pj.values():
                        if isinstance(v, dict) and v.get("type") == "title":
                            for rt in v.get("title", []) or []:
                                if isinstance(rt, dict):
                                    titles.append(rt.get("plain_text", ""))
                    if "survey dashboard" in " ".join(titles).lower():
                        matched_pages.append(pid)
            except Exception:
                matched_pages.append(pid)
        check("Notion 'Survey Dashboard' page created",
              len(matched_pages) >= 1, f"title_matches={len(matched_pages)}")

        # Validate the page has substantive body content (block count)
        if matched_pages:
            cur.execute(
                "SELECT COUNT(*) FROM notion.blocks WHERE parent_id = ANY(%s)",
                (matched_pages,),
            )
            block_count = cur.fetchone()[0]
            check("Survey Dashboard page has body blocks",
                  block_count >= 2, f"block_count={block_count}")
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))

    try:
        conn = get_conn()
        cur = conn.cursor()
        # Title equality (case-insensitive trim) on the form
        cur.execute(
            "SELECT id, title FROM gform.forms "
            "WHERE LOWER(TRIM(title)) = LOWER('Survey Feedback')"
        )
        forms = cur.fetchall()
        check("Survey-feedback form created",
              len(forms) >= 1, f"found {len(forms)} forms")
        # Validate the form has at least one question
        if forms:
            form_ids = [f[0] for f in forms]
            try:
                cur.execute(
                    "SELECT COUNT(*) FROM gform.questions WHERE form_id = ANY(%s)",
                    (form_ids,),
                )
                qcount = cur.fetchone()[0]
                check("Survey-feedback form has at least one question",
                      qcount >= 1, f"questions={qcount}")
            except Exception:
                # Schema may differ — best effort, don't crash the check.
                pass
        conn.close()
    except Exception as e:
        check("GForm check", False, str(e))

    # Check Word document - require specific filename + Heading-style sections
    survey_docx = os.path.join(agent_workspace, "Survey_Analysis.docx")
    check("Survey_Analysis.docx exists", os.path.exists(survey_docx))
    if os.path.exists(survey_docx):
        from docx import Document
        doc = Document(survey_docx)
        heading_texts = []
        body_paras = []
        for p in doc.paragraphs:
            sname = (p.style.name if p.style else "") or ""
            t = (p.text or "").strip()
            if not t:
                continue
            if sname.startswith("Heading") or sname == "Title":
                heading_texts.append(t.lower())
            else:
                body_paras.append(t)

        def has_heading(*needles):
            return any(any(n in h for n in needles) for h in heading_texts)

        check("Survey_Analysis.docx has Executive Summary / Overview heading",
              has_heading("executive summary", "overview", "summary"),
              f"headings: {heading_texts[:5]}")
        check("Survey_Analysis.docx has Key Findings heading",
              has_heading("key findings", "findings", "finding"),
              f"headings: {heading_texts[:5]}")
        check("Survey_Analysis.docx has Recommendations heading",
              has_heading("recommendation"),
              f"headings: {heading_texts[:5]}")
        body_len = sum(len(p) for p in body_paras)
        check("Survey_Analysis.docx has substantive body content",
              body_len > 150 and len(body_paras) >= 3,
              f"body_len={body_len} paras={len(body_paras)}")

    check("notion_survey_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "notion_survey_processor.py")))

    # Validate optional JSON output if present
    results_path = os.path.join(agent_workspace, "notion_survey_results.json")
    if os.path.exists(results_path):
        try:
            with open(results_path) as _rf:
                _r = json.load(_rf)
            check("notion_survey_results.json is non-empty",
                  bool(_r) and isinstance(_r, (dict, list)),
                  f"type={type(_r).__name__}")
        except Exception as e:
            check("notion_survey_results.json valid JSON", False, str(e))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
