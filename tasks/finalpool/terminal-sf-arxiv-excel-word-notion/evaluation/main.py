"""Evaluation for terminal-sf-arxiv-excel-word-notion.
Checks:
1. Retention_Strategy.xlsx with 3 sheets (Department_Analysis, Research_Summary, Action_Plan)
2. Retention_Strategy_Report.docx
3. Notion database "Retention Action Items" with 7 department pages
4. flight_risk_analysis.py and synthesis.py scripts exist
5. flight_risk_analysis.json and synthesis.json outputs exist
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

DEPARTMENTS = ["Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"]

# Hardcoded fallback flight risk data (sat<=4 AND perf>=4)
# Priority rules from task.md: pct>8.3 High, 7.9<=pct<=8.3 Medium, pct<7.9 Low
_FALLBACK_EXPECTED_DATA = {
    "Engineering": {"headcount": 7096, "flight_risk": 566, "pct": 7.98, "priority": "Medium"},
    "Finance":     {"headcount": 7148, "flight_risk": 598, "pct": 8.37, "priority": "High"},
    "HR":          {"headcount": 7077, "flight_risk": 594, "pct": 8.39, "priority": "High"},
    "Operations":  {"headcount": 7120, "flight_risk": 564, "pct": 7.92, "priority": "Medium"},
    "R&D":         {"headcount": 7083, "flight_risk": 576, "pct": 8.13, "priority": "Medium"},
    "Sales":       {"headcount": 7232, "flight_risk": 596, "pct": 8.24, "priority": "Medium"},
    "Support":     {"headcount": 7244, "flight_risk": 537, "pct": 7.41, "priority": "Low"},
}


def _get_expected_data_from_db():
    """Query sf_data schema to compute department headcounts and flight risk counts dynamically."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        try:
            # Get headcount per department
            cur.execute("""
                SELECT "DEPARTMENT", COUNT(*) as headcount
                FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
                GROUP BY "DEPARTMENT"
            """)
            headcounts = {r[0]: r[1] for r in cur.fetchall()}

            # Get flight risk count per department (satisfaction<=4 AND performance>=4)
            cur.execute("""
                SELECT "DEPARTMENT", COUNT(*) as flight_risk
                FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
                WHERE "JOB_SATISFACTION" <= 4 AND "PERFORMANCE_RATING" >= 4
                GROUP BY "DEPARTMENT"
            """)
            flight_risks = {r[0]: r[1] for r in cur.fetchall()}

            result = {}
            # task.md: pct > 8.3 -> High; 7.9 <= pct <= 8.3 -> Medium; pct < 7.9 -> Low
            for dept in DEPARTMENTS:
                hc = headcounts.get(dept, 0)
                fr = flight_risks.get(dept, 0)
                pct = round(fr / hc * 100, 2) if hc > 0 else 0
                if pct > 8.3:
                    priority = "High"
                elif pct < 7.9:
                    priority = "Low"
                else:
                    priority = "Medium"
                result[dept] = {"headcount": hc, "flight_risk": fr, "pct": pct, "priority": priority}
            return result
        finally:
            cur.close()
            conn.close()
    except Exception:
        return _FALLBACK_EXPECTED_DATA


EXPECTED_DATA = _get_expected_data_from_db()
EXPECTED_DATA_LOWER = {k.lower(): v for k, v in EXPECTED_DATA.items()}


def _canon_dept(name):
    """Canonical department key that maps common spellings to the expected key.

    Tolerates "R & D" / "R/D" / "RD" / "RND" / "Research & Development" -> "r&d"
    (task.md lists the department as "R&D"); other departments match on their
    lowercased, non-alphanumeric-stripped form.
    """
    if name is None:
        return None
    s = str(name).lower().strip()
    # '&' -> 'and' so "R & D" / "Research & Development" canonicalize like the spelled-out forms
    s = s.replace("&", " and ")
    t = re.sub(r"[^a-z0-9]", "", s)
    if t in ("rd", "rnd", "randd", "researchanddevelopment"):
        return "r&d"
    if t in ("humanresources",):
        return "hr"
    return t


CANON_DEPT = {_canon_dept(k): k for k in EXPECTED_DATA_LOWER}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _to_float(v):
    """Coerce str/int/float/None to float, tolerating %, currency symbols, thousands-sep, spaces."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace(",", "").replace("$", "").replace("€", "").replace("¥", "").replace("£", "")
    s = s.replace("%", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _cell_value(raw, cached=None):
    """Robust numeric value of a cell; formula cells fall back to their cached value."""
    if raw is not None and isinstance(raw, str) and raw.strip().startswith("="):
        raw = cached  # formula -> use cached (data_only) value; None if never recalculated
    return _to_float(raw)


def num_close(a, b, tol=2.0):
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # One side unparseable: fall back to case-insensitive string equality.
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def _find_header_idx(rows, keywords):
    """Index of the row that best matches the expected column keywords (default 0).

    Tolerates a title row above the real header row (R10).
    """
    best_idx = 0
    best_score = -1
    for i, row in enumerate(rows):
        cells = [str(c).lower() if c is not None else "" for c in row]
        score = sum(1 for kw in keywords if any(kw in c for c in cells))
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _norm_col_name(c):
    return str(c).lower().strip() if c is not None else ""


def _load_workbook_pair(path):
    """Return (raw_workbook, cached_workbook) or (None, None) on failure."""
    try:
        wb_raw = openpyxl.load_workbook(path, data_only=False)
    except Exception:
        return None, None
    try:
        wb_cached = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        wb_cached = None
    return wb_raw, wb_cached


def check_excel(workspace):
    print("\n=== Check 1: Retention_Strategy.xlsx ===")
    path = os.path.join(workspace, "Retention_Strategy.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb_raw, wb_cached = _load_workbook_pair(path)
    if wb_raw is None:
        check("Excel file readable", False, f"Failed to open {path}")
        return
    sheets = wb_raw.sheetnames
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Sheet 1: Department_Analysis
    da_idx = next((i for i, s in enumerate(sheets_lower) if "department" in s and "analysis" in s), 0)
    ws1 = wb_raw[sheets[da_idx]]
    ws1c = wb_cached[sheets[da_idx]] if wb_cached else None
    rows1 = list(ws1.iter_rows(values_only=True))
    rows1c = list(ws1c.iter_rows(values_only=True)) if ws1c else None
    hdr1 = _find_header_idx(rows1, ["department", "headcount", "flight", "pct", "satisfaction"])
    data1 = [r for r in rows1[hdr1 + 1:] if any(c for c in r)]
    check("Department_Analysis has 7 rows", len(data1) >= 7, f"Found {len(data1)}")

    headers = [_norm_col_name(c) for c in rows1[hdr1]]

    # Tolerate common spellings of the flight-risk percentage column: the task.md
    # column name is flight_risk_pct, but "Flight Risk Percentage"/"Flight Risk
    # Ratio"/"Risk %" etc. are all semantically identical (R2 review).
    def _is_pct_col(h):
        pct_kw = ("pct", "%", "percent", "percentage", "rate", "ratio")
        return ("flight" in h and any(k in h for k in pct_kw)) or \
               ("risk" in h and any(k in h for k in pct_kw))

    check("Has flight_risk_pct column",
          any(_is_pct_col(h) for h in headers),
          f"Headers: {rows1[hdr1]}")
    # "avg_satisfaction", "avg_sat", "Avg Job Sat", "satisfaction score" all match
    check("Has avg_satisfaction column",
          any("sat" in h for h in headers),
          f"Headers: {rows1[hdr1]}")

    dept_col = next((i for i, h in enumerate(headers) if "department" in h or "dept" in h), 0)
    hc_col = next((i for i, h in enumerate(headers) if "headcount" in h or "head_count" in h), 1)
    fr_col = next((i for i, h in enumerate(headers) if "flight_risk_count" in h or ("flight" in h and "count" in h)), 2)

    found_depts = 0
    for offset, raw_row in enumerate(data1):
        cached_row = rows1c[hdr1 + 1 + offset] if rows1c else None
        dept_key = _canon_dept(raw_row[dept_col] if len(raw_row) > dept_col else None)
        if dept_key is None or dept_key not in CANON_DEPT:
            continue
        dept_name = CANON_DEPT[dept_key]
        found_depts += 1
        exp = EXPECTED_DATA_LOWER[dept_name]
        hc_raw = raw_row[hc_col] if len(raw_row) > hc_col else None
        hc_cached = cached_row[hc_col] if cached_row is not None and len(cached_row) > hc_col else None
        hc_val = _cell_value(hc_raw, hc_cached)
        if hc_val is not None:
            check(f"{dept_name} headcount correct",
                  num_close(hc_val, exp["headcount"], 50),
                  f"Got {hc_raw}, expected ~{exp['headcount']}")
        fr_raw = raw_row[fr_col] if len(raw_row) > fr_col else None
        fr_cached = cached_row[fr_col] if cached_row is not None and len(cached_row) > fr_col else None
        fr_val = _cell_value(fr_raw, fr_cached)
        if fr_val is not None:
            check(f"{dept_name} flight_risk_count correct",
                  num_close(fr_val, exp["flight_risk"], 20),
                  f"Got {fr_raw}, expected ~{exp['flight_risk']}")
    check("All 7 departments found in Department_Analysis", found_depts >= 7, f"Found {found_depts}")

    # Sheet 2: Research_Summary
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s), 1)
    if rs_idx < len(sheets):
        ws2 = wb_raw[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        hdr2 = _find_header_idx(rows2, ["paper", "key", "applicability", "score"])
        data2 = [r for r in rows2[hdr2 + 1:] if any(c for c in r)]
        check("Research_Summary has 3 rows", len(data2) >= 3, f"Found {len(data2)}")
        headers2 = [_norm_col_name(c) for c in rows2[hdr2]]
        check("Has applicability_score column",
              any("applicability" in h or "score" in h for h in headers2),
              f"Headers: {rows2[hdr2]}")
        # Check that relevant papers are included (not the noise ones)
        all_text2 = " ".join(str(c) for r in data2 for c in r if c).lower()
        check("Contains retention-related paper",
              "retention" in all_text2 or "turnover" in all_text2 or "employee" in all_text2)
        check("Does NOT contain autonomous vehicle paper",
              not any(p in all_text2 for p in ["autonomous vehicle", "urban navigation", "carla"]),
              "Noise paper included in research summary")
        check("Does NOT contain quantum computing paper",
              not any(p in all_text2 for p in ["quantum computing", "protein folding"]),
              "Noise paper included in research summary")

    # Sheet 3: Action_Plan
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "action" in s or "plan" in s), 2)
    if ap_idx < len(sheets):
        ws3 = wb_raw[sheets[ap_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        hdr3 = _find_header_idx(rows3, ["department", "strategy", "priority", "cost"])
        data3 = [r for r in rows3[hdr3 + 1:] if any(c for c in r)]
        check("Action_Plan has 7 rows", len(data3) >= 7, f"Found {len(data3)}")

        headers3 = [_norm_col_name(c) for c in rows3[hdr3]]
        pri_col = next((i for i, h in enumerate(headers3) if "priority" in h), None)
        dept_col3 = next((i for i, h in enumerate(headers3) if "department" in h or "dept" in h), 0)

        if pri_col is not None:
            for row in data3:
                dept_key = _canon_dept(row[dept_col3] if len(row) > dept_col3 else None)
                if dept_key is not None and dept_key in CANON_DEPT:
                    dept_name = CANON_DEPT[dept_key]
                    exp_pri = EXPECTED_DATA_LOWER[dept_name]["priority"]
                    got_pri = _norm_col_name(row[pri_col] if len(row) > pri_col else None)
                    check(f"{dept_name} priority is {exp_pri}",
                          got_pri == exp_pri.lower(),
                          f"Got '{got_pri}', expected '{exp_pri}'")


def check_word(workspace):
    print("\n=== Check 2: Retention_Strategy_Report.docx ===")
    path = os.path.join(workspace, "Retention_Strategy_Report.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    check("Has title mentioning retention", "retention" in full_text and ("strategy" in full_text or "report" in full_text))
    check("Mentions flight risk", "flight risk" in full_text or "flight-risk" in full_text)
    check("Mentions executive summary", "executive summary" in full_text or "summary" in full_text)
    check("Mentions research findings", "research" in full_text and ("finding" in full_text or "paper" in full_text))
    check("Mentions recommendations", "recommend" in full_text)
    dept_found = sum(1 for d in DEPARTMENTS if re.search(rf"\b{re.escape(d.lower())}\b", full_text))
    check("Mentions specific departments", dept_found >= 5,
          f"Found {dept_found} departments")
    check("Has substantial content", len(full_text) > 500, f"Length: {len(full_text)}")
    check("Mentions priority levels", "high" in full_text and ("medium" in full_text or "low" in full_text))


def _prop_value(props, key_like):
    """Return the value object of the property whose key contains key_like (case-insensitive)."""
    if not isinstance(props, dict):
        return None
    for k, v in props.items():
        if key_like in k.lower():
            return v
    return None


def _select_name(value):
    """Extract an option name from a Notion select-style property value (robust to shapes)."""
    if isinstance(value, dict):
        sel = value.get("select")
        if isinstance(sel, dict) and "name" in sel:
            return sel["name"]
        if isinstance(sel, str):
            return sel
        if "name" in value:
            return value["name"]
    elif isinstance(value, str):
        return value
    return None


def _page_priority(props):
    """Structured priority extraction: properties->Priority->select->name, with substring fallback."""
    name = _select_name(_prop_value(props, "priority"))
    if name:
        return str(name).lower()
    text = json.dumps(props).lower()
    for kw in ("high", "medium", "low"):
        if f'"{kw}"' in text:
            return kw
    return None


def _page_status(props):
    """Structured status extraction: properties->Status->select->name, with substring fallback."""
    name = _select_name(_prop_value(props, "status"))
    if name:
        return str(name).lower()
    text = json.dumps(props).lower()
    if "not started" in text or "not_started" in text:
        return "not started"
    return None


def check_notion():
    print("\n=== Check 3: Notion Retention Action Items ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("Notion database connection", False, str(e))
        return
    cur = conn.cursor()
    try:
        # Check database exists - strict title match: must contain 'retention' + 'action' + 'items'
        # or exact 'retention action items'.
        cur.execute("""
            SELECT id, title, properties FROM notion.databases
            WHERE (title::text ILIKE '%retention%' AND title::text ILIKE '%action%' AND title::text ILIKE '%items%')
               OR title::text ILIKE '%retention action items%'
        """)
        dbs = cur.fetchall()
        if not dbs:
            cur.execute("SELECT id, title, properties FROM notion.databases")
            all_dbs = cur.fetchall()
            check("Notion database 'Retention Action Items' exists", False,
                  f"Found {len(all_dbs)} databases: {[str(d[1])[:50] for d in all_dbs]}")
            return
        check("Notion database 'Retention Action Items' exists", True)

        # Multi-agent runs may create duplicate same-name databases (some empty or
        # partial). Pick the matching database with the most pages so an incomplete
        # duplicate never sinks a correct model (R2 review).
        best_pages = []
        best_props = {}
        for db_row in dbs:
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE parent::text LIKE %s
            """, (f'%{db_row[0]}%',))
            pg = cur.fetchall()
            if len(pg) >= len(best_pages):
                best_pages = pg
                best_props = db_row[2] if db_row[2] else {}
        props = best_props

        # Check properties
        prop_names = [k.lower() for k in props.keys()] if isinstance(props, dict) else []
        check("Has Priority property", any("priority" in p for p in prop_names), f"Props: {prop_names}")
        check("Has Status property", any("status" in p for p in prop_names), f"Props: {prop_names}")
        check("Has Strategy property", any("strategy" in p for p in prop_names), f"Props: {prop_names}")

        # Check pages
        pages = best_pages
        check("Has 7 department pages", len(pages) >= 7, f"Found {len(pages)} pages")

        if pages:
            # Check that pages have correct priorities (structured parse, not bare substring).
            # Expected counts are derived from EXPECTED_DATA so they stay consistent with the
            # Action_Plan priority checks and with whatever the DB actually yields.
            exp_counts = {}
            for v in EXPECTED_DATA.values():
                exp_counts[v["priority"]] = exp_counts.get(v["priority"], 0) + 1
            high_count = sum(1 for p in pages if _page_priority(p[1] if p[1] else {}) == "high")
            medium_count = sum(1 for p in pages if _page_priority(p[1] if p[1] else {}) == "medium")
            low_count = sum(1 for p in pages if _page_priority(p[1] if p[1] else {}) == "low")

            check("Has High priority pages count matching data", high_count == exp_counts.get("High", 0),
                  f"Found {high_count} High priority pages, expected {exp_counts.get('High', 0)}")
            check("Has Medium priority pages count matching data", medium_count == exp_counts.get("Medium", 0),
                  f"Found {medium_count} Medium priority pages, expected {exp_counts.get('Medium', 0)}")
            check("Has Low priority pages count matching data", low_count == exp_counts.get("Low", 0),
                  f"Found {low_count} Low priority pages, expected {exp_counts.get('Low', 0)}")

            # Check Status is Not Started
            not_started_count = sum(1 for p in pages if _page_status(p[1] if p[1] else {}) == "not started")
            check("All pages have Status 'Not Started'", not_started_count >= 7,
                  f"Found {not_started_count} with 'Not Started'")

    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python Scripts ===")
    check("flight_risk_analysis.py exists",
          os.path.exists(os.path.join(workspace, "flight_risk_analysis.py")))
    check("synthesis.py exists",
          os.path.exists(os.path.join(workspace, "synthesis.py")))


def check_json_outputs(workspace):
    print("\n=== Check 5: JSON Outputs ===")
    fr_path = os.path.join(workspace, "flight_risk_analysis.json")
    if os.path.exists(fr_path):
        check("flight_risk_analysis.json exists", True)
        try:
            with open(fr_path) as f:
                fr_data = json.load(f)
            check("flight_risk_analysis.json is valid JSON", True)
            # Check it has department data
            fr_text = json.dumps(fr_data).lower()
            check("flight_risk_analysis.json mentions departments",
                  sum(1 for d in DEPARTMENTS if d.lower() in fr_text) >= 5)
        except (json.JSONDecodeError, Exception) as e:
            check("flight_risk_analysis.json is valid JSON", False, str(e))
    else:
        check("flight_risk_analysis.json exists", False)

    syn_path = os.path.join(workspace, "synthesis.json")
    if os.path.exists(syn_path):
        check("synthesis.json exists", True)
        try:
            with open(syn_path) as f:
                syn_data = json.load(f)
            check("synthesis.json is valid JSON", True)
        except (json.JSONDecodeError, Exception) as e:
            check("synthesis.json is valid JSON", False, str(e))
    else:
        check("synthesis.json exists", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_scripts(args.agent_workspace)
    check_json_outputs(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Tightened: require all checks to pass (previously >=70%).
    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
