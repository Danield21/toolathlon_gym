"""Evaluation for yf-market-cap-ranking."""
import argparse
import json
import os
import sys
import openpyxl

try:
    import psycopg2
    HAS_PG = True
except ImportError:
    HAS_PG = False

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")), "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_notion_page():
    """Check Notion page titled 'Market Cap Rankings' exists.
    Returns True if found, False if not found or DB unavailable.
    DB errors are treated as FAIL (not SKIP) per benchmark policy."""
    if not HAS_PG:
        print("    WARNING: psycopg2 not available, treating as FAIL")
        return False
    try:
        conn = psycopg2.connect(**DB)
    except psycopg2.OperationalError as e:
        print(f"    WARNING: Notion DB unavailable: {e}")
        return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        for page_id, props in cur.fetchall():
            p = props if isinstance(props, dict) else json.loads(props) if props else {}
            title_text = ""
            # Walk all properties looking for title-type field
            for prop_name, prop_val in (p.items() if isinstance(p, dict) else []):
                if isinstance(prop_val, dict):
                    title_arr = prop_val.get("title")
                    if isinstance(title_arr, list):
                        for item in title_arr:
                            if isinstance(item, dict):
                                title_text += item.get("plain_text", "") or item.get("text", {}).get("content", "") if isinstance(item.get("text"), dict) else item.get("plain_text", "")
            # Fallback: top-level title key
            if not title_text and "title" in p:
                t = p["title"]
                if isinstance(t, dict) and "title" in t:
                    for item in t["title"]:
                        if isinstance(item, dict):
                            title_text += item.get("plain_text", item.get("text", {}).get("content", "")) if isinstance(item.get("text"), dict) else item.get("plain_text", "")
            tn = title_text.strip().lower()
            # Require exact title equality or "market cap rankings" core (allow trailing words)
            if tn == "market cap rankings" or tn.startswith("market cap rankings"):
                cur.close()
                conn.close()
                return True
        cur.close()
        conn.close()
        return False
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        print(f"    WARNING: Notion check error: {e}")
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "YF_Market_Cap.xlsx")
    gt_file = os.path.join(gt_dir, "YF_Market_Cap.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Market Cap Ranking
    print(f"  Checking Market Cap Ranking...")
    a_rows = load_sheet_rows(agent_wb, "Market Cap Ranking")
    g_rows = load_sheet_rows(gt_wb, "Market Cap Ranking")
    if a_rows is None:
        all_errors.append("Sheet 'Market Cap Ranking' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Market Cap Ranking' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Name (col 1) - exact (case-insensitive)
            if len(a_row) > 1 and len(g_row) > 1 and g_row[1] is not None:
                if not str_match(a_row[1], g_row[1]):
                    errors.append(f"{key}.Name: {a_row[1]!r} vs {g_row[1]!r}")
            # Sector (col 2) - exact (case-insensitive)
            if len(a_row) > 2 and len(g_row) > 2 and g_row[2] is not None:
                if not str_match(a_row[2], g_row[2]):
                    errors.append(f"{key}.Sector: {a_row[2]!r} vs {g_row[2]!r}")
            # Market_Cap_B (col 3) - tighter tolerance: 1% of GT value or 1.0, whichever larger
            if len(a_row) > 3 and len(g_row) > 3:
                try:
                    gt_val = float(g_row[3])
                    tol_mc = max(1.0, abs(gt_val) * 0.01)
                except (TypeError, ValueError):
                    tol_mc = 1.0
                if not num_close(a_row[3], g_row[3], tol_mc):
                    errors.append(f"{key}.Market_Cap_B: {a_row[3]} vs {g_row[3]} (tol={tol_mc:.2f})")
            # Recommendation (col 4) - exact (case-insensitive)
            if len(a_row) > 4 and len(g_row) > 4 and g_row[4] is not None:
                if not str_match(a_row[4], g_row[4]):
                    errors.append(f"{key}.Recommendation: {a_row[4]!r} vs {g_row[4]!r}")

        # Verify descending sort by Market_Cap_B in agent file
        try:
            mc_vals = []
            for row in a_data:
                if row and row[0] is not None and len(row) > 3:
                    try:
                        mc_vals.append(float(row[3]))
                    except (TypeError, ValueError):
                        pass
            for i in range(1, len(mc_vals)):
                if mc_vals[i] > mc_vals[i-1] + 0.01:
                    errors.append(f"Market Cap Ranking not sorted descending: row {i+1} ({mc_vals[i]}) > row {i} ({mc_vals[i-1]})")
                    break
        except Exception:
            pass

        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                # Total_Companies is integer count - tolerance 0
                if "total_companies" in key or "companies" in key.replace("_", "") and "total" in key.replace("_", ""):
                    if not num_close(a_row[1], g_row[1], 0):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=0 integer)")
                # Largest/Smallest are company names (string)
                elif "largest" in key or "smallest" in key:
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}.Value: {a_row[1]!r} vs {g_row[1]!r}")
                else:
                    # Total_Market_Cap_B - tighter tol (1%)
                    try:
                        gt_val = float(g_row[1])
                        tol_s = max(1.0, abs(gt_val) * 0.01)
                    except (TypeError, ValueError):
                        tol_s = 1.0
                    if not num_close(a_row[1], g_row[1], tol_s):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol_s:.2f})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check Notion page (FAIL if not found or DB unavailable)
    print(f"  Checking Notion page...")
    notion_status = check_notion_page()
    if notion_status:
        print(f"    PASS")
    else:
        all_errors.append("Notion page titled 'Market Cap Rankings' not found")
        print(f"    FAIL: page not found")


    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
