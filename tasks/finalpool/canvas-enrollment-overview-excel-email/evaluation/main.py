"""Evaluation for canvas-enrollment-overview-excel-email."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Spring2014_Enrollment.xlsx")
    gt_file = os.path.join(gt_dir, "Spring2014_Enrollment.xlsx")

    file_errors = []
    db_errors = []

    if not os.path.exists(agent_file):
        file_errors.append(f"Agent output not found: {agent_file}")
    if not os.path.exists(gt_file):
        file_errors.append(f"Groundtruth not found: {gt_file}")

    if not file_errors:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Enrollment sheet
        print("  Checking Enrollment...")
        a_rows = load_sheet_rows(agent_wb, "Enrollment")
        g_rows = load_sheet_rows(gt_wb, "Enrollment")
        if a_rows is None:
            file_errors.append("Sheet 'Enrollment' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Enrollment' not found in groundtruth")
        else:
            # Validate header (case/underscore-insensitive)
            expected_headers = ["course_name", "course_code", "student_count", "teacher_count", "ta_count", "student_teacher_ratio"]
            if not a_rows or not a_rows[0]:
                file_errors.append("Enrollment header row missing")
            else:
                a_headers = [str(h or "").strip().lower().replace(" ", "_") for h in a_rows[0]]
                for i, expected in enumerate(expected_headers):
                    if i >= len(a_headers) or a_headers[i] != expected:
                        file_errors.append(f"Enrollment header[{i}]: expected '{expected}' got '{a_headers[i] if i < len(a_headers) else 'MISSING'}'")

            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            # Validate row count matches GT exactly
            a_data = [r for r in a_data if r and r[0] is not None]
            g_data = [r for r in g_data if r and r[0] is not None]
            if len(a_data) != len(g_data):
                file_errors.append(f"Enrollment row count: {len(a_data)} vs expected {len(g_data)}")

            a_lookup = {}
            for row in a_data:
                a_lookup[str(row[0]).strip().lower()] = row
            for g_row in g_data:
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing row: {g_row[0]}")
                    continue

                # Course_Code (col 1)
                if len(a_row) > 1 and len(g_row) > 1:
                    if not str_match(a_row[1], g_row[1]):
                        file_errors.append(f"{key}.Course_Code: {a_row[1]} vs {g_row[1]}")
                # Student_Count (col 2) — Canvas data is deterministic, tol=0
                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 0):
                        file_errors.append(f"{key}.Student_Count: {a_row[2]} vs {g_row[2]}")
                # Teacher_Count (col 3)
                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 0):
                        file_errors.append(f"{key}.Teacher_Count: {a_row[3]} vs {g_row[3]}")
                # TA_Count (col 4)
                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 0):
                        file_errors.append(f"{key}.TA_Count: {a_row[4]} vs {g_row[4]}")
                # Student_Teacher_Ratio (col 5)
                if len(a_row) > 5 and len(g_row) > 5:
                    if not num_close(a_row[5], g_row[5], 0.2):
                        file_errors.append(f"{key}.Student_Teacher_Ratio: {a_row[5]} vs {g_row[5]}")

            # Validate sort order: ascending by Student_Count
            try:
                counts = [float(r[2]) for r in a_data if r and len(r) > 2 and r[2] is not None]
                if counts != sorted(counts):
                    file_errors.append(f"Enrollment not sorted by Student_Count ascending: {counts}")
            except (TypeError, ValueError) as e:
                file_errors.append(f"Sort validation error: {e}")

        # Check Summary sheet
        print("  Checking Summary...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            file_errors.append("Sheet 'Summary' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing summary row: {g_row[0]}")
                    continue
                if len(a_row) > 1 and len(g_row) > 1:
                    if isinstance(g_row[1], (int, float)):
                        # Counts and totals: tight tolerance (Canvas data deterministic)
                        # Use absolute tolerance of 1 for integer fields
                        if not num_close(a_row[1], g_row[1], 1):
                            file_errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
                    else:
                        if not str_match(a_row[1], g_row[1]):
                            file_errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")

    # Check email (DB check) — now BLOCKING
    print("  Checking email...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Sent folder: lookup by name (avoid hardcoding folder_id)
        cur.execute("SELECT id FROM email.folders WHERE LOWER(name) IN ('sent', 'sent items')")
        sent_ids = [r[0] for r in cur.fetchall()]
        if not sent_ids:
            # Surface preprocess corruption rather than silently using folder id 2.
            db_errors.append("No sent folder found (preprocess malfunction)")
            rows = []
        else:
            cur.execute("""
                SELECT subject, COALESCE(body_text, body_html, ''), to_addr
                FROM email.messages
                WHERE folder_id = ANY(%s)
            """, (sent_ids,))
            rows = cur.fetchall()

        target_subj = "spring 2014 enrollment summary"
        target_to = "registrar@university.edu"
        match = None
        for subj, body, to_addr in rows:
            subj_l = (subj or "").strip().lower()
            # to_addr is jsonb (list of strings)
            tos = to_addr if isinstance(to_addr, list) else [str(to_addr or "")]
            to_str = " ".join(str(t).lower() for t in tos)
            if target_subj in subj_l and target_to in to_str:
                match = (subj, body, to_addr)
                break
        if sent_ids and not match:
            db_errors.append(f"No email with subject '{target_subj}' to {target_to} found in sent folder")
        elif match:
            body_lower = (match[1] or "").lower()
            # GT: total_courses=6, total_students=7804, smallest=Environmental Economics, largest=Creative Computing
            if "7804" not in body_lower and "7,804" not in body_lower:
                db_errors.append("Email body missing total student count (7804)")
            if "creative computing" not in body_lower:
                db_errors.append("Email body missing largest course (Creative Computing)")
            if "environmental economics" not in body_lower:
                db_errors.append("Email body missing smallest course (Environmental Economics)")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Email check: {e}")

    # Check calendar event (DB check) — now BLOCKING + validates date/time
    print("  Checking calendar event...")
    try:
        from datetime import datetime, timezone, timedelta
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Force session to UTC so timestamps come back as UTC
        cur.execute("SET TIME ZONE 'UTC'")
        cur.execute("""
            SELECT summary,
                   start_datetime AT TIME ZONE 'UTC',
                   end_datetime AT TIME ZONE 'UTC'
            FROM gcal.events
            WHERE LOWER(summary) LIKE '%spring 2014 enrollment review%'
        """)
        events = cur.fetchall()
        if not events:
            db_errors.append("No calendar event 'Spring 2014 Enrollment Review' found")
        else:
            target_start = datetime(2026, 3, 14, 14, 0, 0)
            target_end = datetime(2026, 3, 14, 15, 0, 0)
            ok = False
            examples = []
            for summ, st, et in events:
                # Strip tz to compare UTC components
                st_naive = st.replace(tzinfo=None) if hasattr(st, 'replace') and st else None
                et_naive = et.replace(tzinfo=None) if hasattr(et, 'replace') and et else None
                examples.append((summ, str(st_naive), str(et_naive)))
                if st_naive == target_start and et_naive == target_end:
                    ok = True
                    break
            if not ok:
                db_errors.append(f"No event with start={target_start} UTC, end={target_end} UTC. Found: {examples[:3]}")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Calendar check: {e}")

    # Final result — db_errors are now BLOCKING
    print(f"\n=== SUMMARY ===")
    print(f"  File errors: {len(file_errors)}")
    print(f"  DB errors:   {len(db_errors)}")
    if db_errors:
        for e in db_errors[:10]:
            print(f"    [DB] {e}")
    if file_errors:
        for e in file_errors[:10]:
            print(f"    [FILE] {e}")
    total = len(file_errors) + len(db_errors)
    if total > 0:
        print(f"  Overall: FAIL")
        sys.exit(1)
    else:
        print(f"  Overall: PASS")
        sys.exit(0)


if __name__ == "__main__":
    main()
