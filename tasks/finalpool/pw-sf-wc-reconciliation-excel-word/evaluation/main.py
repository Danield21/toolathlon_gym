"""Evaluation script for pw-sf-wc-reconciliation-excel-word."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-sf-wc-reconciliation-excel-word"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Detect GT self-test (V1 parity test).
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    excel_path = os.path.join(agent_workspace, "Wc_Reconciliation_Report.xlsx")
    check("Wc_Reconciliation_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Wc_Reconciliation_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            # task.md says "primary dimension (such as department, product, region, or topic)"
            # Mock has Product_Category; Snowflake/WC have Region. Accept either.
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            header_concat = " ".join(headers)
            has_dim = any(k in header_concat for k in
                          ['region', 'category', 'product', 'department',
                           'topic', 'dimension'])
            check("Data_Analysis has a primary dimension column", has_dim,
                  f"headers: {headers[:8]}")
            # Tightened (per QA): drop overly-broad 'value' which would match
            # benchmark-side 'Industry_Avg_Value'. Keep tokens that hint at
            # the agent's own organization (our_/internal/...) or canonical
            # business metrics computed from the warehouse.
            has_internal = any(k in header_concat for k in
                               ['our_', 'internal', 'order_count', 'revenue',
                                'orders', 'sales', 'count', 'company'])
            check("Data_Analysis has internal-metric column", has_internal,
                  f"headers: {headers[:8]}")
            has_benchmark = any(k in header_concat for k in
                                ['benchmark', 'industry', 'market', 'avg_',
                                 'external', 'reference'])
            check("Data_Analysis has benchmark column", has_benchmark,
                  f"headers: {headers[:8]}")
            has_gap = any(k in header_concat for k in
                          ['gap', 'diff', 'delta', 'variance', 'spread'])
            check("Data_Analysis has gap/difference column", has_gap,
                  f"headers: {headers[:8]}")
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 4)
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            # task.md hints "primary dimension (such as department, product,
            # region, or topic)" — Recommendations should reference whichever
            # primary dimension was used in Data_Analysis. Accept any of them
            # rather than hardcoding 'Region' (per QA finding).
            check_columns('Recommendations', ['Priority', 'Action'], 2)
            r_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            r_concat = " ".join(r_headers)
            has_dim = any(k in r_concat for k in
                           ['region', 'category', 'product', 'department',
                            'topic', 'dimension', 'segment', 'area'])
            check("Recommendations has primary-dimension column",
                  has_dim, f"headers: {r_headers[:8]}")
    # Word/script checks should run regardless of Excel state.
    # Per task.md: Wc_Reconciliation_Analysis.docx with executive summary,
    # key findings, and recommendations sections.
    word_path = os.path.join(agent_workspace, "Wc_Reconciliation_Analysis.docx")
    if not os.path.exists(word_path):
        import glob as globmod
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        if word_files:
            word_path = word_files[0]
    check("Wc_Reconciliation_Analysis.docx (or any *.docx) exists",
          os.path.exists(word_path), f"checked {word_path}")
    if os.path.exists(word_path):
        from docx import Document
        doc = Document(word_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Word doc has substantive content (>=200 chars)",
              len(text) >= 200, f"text length: {len(text)}")
        check("Word doc has executive summary",
              "executive" in text and "summary" in text,
              "missing 'executive summary'")
        check("Word doc has key findings",
              "finding" in text, "missing 'findings'")
        check("Word doc has recommendations",
              "recommend" in text, "missing 'recommend'")

    proc_path = os.path.join(agent_workspace, "sf_wc_reconcile_processor.py")
    proc_exists = os.path.exists(proc_path)
    check("sf_wc_reconcile_processor.py exists", proc_exists)
    if proc_exists:
        try:
            proc_size = os.path.getsize(proc_path)
            check("sf_wc_reconcile_processor.py is non-trivial (>=200 bytes)",
                  proc_size >= 200, f"size={proc_size} bytes")
        except OSError as e:
            check("sf_wc_reconcile_processor.py size check", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
