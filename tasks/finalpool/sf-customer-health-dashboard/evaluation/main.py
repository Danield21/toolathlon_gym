"""
Evaluation script for sf-customer-health-dashboard task.

Checks:
1. Excel file Customer_Health.xlsx with 3 sheets
2. Notion page created
3. Calendar events for critical accounts with score < 15
"""
import argparse
import json
import os
import sys
from datetime import date

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

# The analysis "today" date. Health scores (and the resulting status buckets,
# below-15 set, and calendar follow-up schedule) shift with this date, so it
# must be pinned. 2026-04-25 reproduces groundtruth_workspace/Customer_Health.xlsx
# exactly (2000 customers: 1557 Critical / 443 At Risk / 0 Healthy; 39 below 15;
# per-segment totals and averages match). Follow-up calls start the Monday after.
AS_OF_DATE = "2026-04-25"

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _levenshtein(a, b):
    """Pure-Python Levenshtein distance for short strings."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def extract_page_title(props):
    """Extract a Notion page title from its properties JSON blob."""
    if isinstance(props, str):
        try:
            props = json.loads(props)
        except Exception:
            return ""
    if not isinstance(props, dict):
        return ""
    for prop in props.values():
        if isinstance(prop, dict) and isinstance(prop.get("title"), list) and prop["title"]:
            return "".join(
                (p.get("plain_text", "") or (p.get("text") or {}).get("content", ""))
                for p in prop["title"] if isinstance(p, dict)
            ).strip()
    return ""


def loose_title_match(actual, expected, keywords=None):
    """Exact title match: task.md says "Use this exact phrase as the page
    title". After normalization the expected title must appear intact in the
    actual title — no edit-distance / keyword fuzzy tolerance."""
    a = " ".join(str(actual or "").lower().split())
    e = " ".join(str(expected or "").lower().split())
    if not a or not e:
        return False
    return e in a


def extract_block_text(block_data):
    """Extract plain text from a Notion block_data JSON blob.

    block_data is the stored body of a block (e.g. {'paragraph': {'rich_text':
    [{'plain_text': '...'}]}}). Walk the JSON and collect every rich_text
    plain_text value so content checks work regardless of block type.
    """
    if isinstance(block_data, str):
        try:
            block_data = json.loads(block_data)
        except Exception:
            return ""

    texts = []

    def walk(node):
        if isinstance(node, dict):
            if "plain_text" in node and isinstance(node["plain_text"], str):
                texts.append(node["plain_text"])
            if isinstance(node.get("rich_text"), list):
                for rt in node["rich_text"]:
                    walk(rt)
            else:
                for v in node.values():
                    walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    walk(block_data)
    return " ".join(t for t in texts if t)


def compute_expected_values():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        WITH order_stats AS (
            SELECT "CUSTOMER_ID",
                COUNT(*) as order_count,
                MAX("ORDER_DATE") as last_order_date
            FROM sf_data."SALES_DW__PUBLIC__ORDERS"
            GROUP BY "CUSTOMER_ID"
        ),
        health AS (
            SELECT c."CUSTOMER_ID", c."CUSTOMER_NAME", c."SEGMENT", c."REGION",
                ROUND(GREATEST(0, 100 - (%(asof)s::date - COALESCE(os.last_order_date, c."SIGNUP_DATE")))::numeric, 2) as recency_score,
                ROUND(LEAST(100, COALESCE(os.order_count, 0) * 10)::numeric, 2) as frequency_score,
                ROUND(LEAST(100, c."LIFETIME_VALUE" / 50.0)::numeric, 2) as monetary_score,
                ROUND((0.4 * GREATEST(0, 100 - (%(asof)s::date - COALESCE(os.last_order_date, c."SIGNUP_DATE"))) +
                0.3 * LEAST(100, COALESCE(os.order_count, 0) * 10) +
                0.3 * LEAST(100, c."LIFETIME_VALUE" / 50.0))::numeric, 2) as health_score
            FROM sf_data."SALES_DW__PUBLIC__CUSTOMERS" c
            LEFT JOIN order_stats os ON c."CUSTOMER_ID" = os."CUSTOMER_ID"
        )
        SELECT "CUSTOMER_NAME", "SEGMENT", "REGION",
            recency_score, frequency_score, monetary_score, health_score,
            CASE WHEN health_score >= 80 THEN 'Healthy'
                 WHEN health_score >= 50 THEN 'At Risk'
                 ELSE 'Critical' END as status
        FROM health
        ORDER BY health_score ASC
    """, {"asof": AS_OF_DATE})
    all_rows = cur.fetchall()

    # Summary by segment
    segments = {}
    for r in all_rows:
        seg = r[1]
        if seg not in segments:
            segments[seg] = {"total": 0, "healthy": 0, "at_risk": 0, "critical": 0, "scores": []}
        segments[seg]["total"] += 1
        segments[seg]["scores"].append(float(r[6]))
        if r[7] == "Healthy":
            segments[seg]["healthy"] += 1
        elif r[7] == "At Risk":
            segments[seg]["at_risk"] += 1
        else:
            segments[seg]["critical"] += 1

    segment_summary = []
    for seg in sorted(segments.keys()):
        s = segments[seg]
        avg = round(sum(s["scores"]) / len(s["scores"]), 2)
        segment_summary.append({
            "Segment": seg,
            "Total_Customers": s["total"],
            "Healthy_Count": s["healthy"],
            "At_Risk_Count": s["at_risk"],
            "Critical_Count": s["critical"],
            "Avg_Health_Score": avg,
        })

    critical_rows = [r for r in all_rows if r[7] == "Critical"]
    below_15 = [r for r in all_rows if float(r[6]) < 15]

    cur.close()
    conn.close()

    return {
        "total_customers": len(all_rows),
        "critical_count": len(critical_rows),
        "at_risk_count": sum(1 for r in all_rows if r[7] == "At Risk"),
        "healthy_count": sum(1 for r in all_rows if r[7] == "Healthy"),
        "segment_summary": segment_summary,
        "below_15": below_15,
        "top5_lowest": all_rows[:5],
    }


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def header_map(ws):
    """Map lowercased header text -> 0-based column index, read from row 1.

    Lets the checks locate columns by name rather than by position, so an
    agent that inserts an extra column or reorders columns does not fail
    purely on column-order grounds. Callers fall back to a positional
    default (matching the column order documented in task.md) if a header
    is absent.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}
    return {str(h).strip().lower(): i for i, h in enumerate(header_row) if h is not None}


def cell(row, idx):
    """Safe cell access; returns None when idx is out of range (robust to
    rows shorter than the widest header)."""
    if row is None or idx is None:
        return None
    return row[idx] if idx < len(row) else None


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Customer_Health.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return
    check("Excel file readable", True)

    for sn in ["Health Scores", "Critical Accounts", "Summary by Segment"]:
        found = any(str_match(s, sn) for s in wb.sheetnames)
        check(f"Sheet '{sn}' exists", found, f"Found: {wb.sheetnames}")

    # --- Health Scores ---
    print("\n--- Health Scores ---")
    ws = get_sheet(wb, "Health Scores")
    if ws:
        col = header_map(ws)
        name_idx = col.get("customer_name", 0)
        score_idx = col.get("health_score", 6)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Strict: total customers exact (this is fixed by DB)
        check("Health Scores total rows",
              len(rows) == expected["total_customers"],
              f"Expected {expected['total_customers']}, got {len(rows)}")

        # Check that first 5 rows are the actual lowest scores (sorted ascending)
        if len(rows) >= 5:
            for i, exp_row in enumerate(expected["top5_lowest"][:5]):
                agent_row = rows[i]
                agent_score = cell(agent_row, score_idx)
                if agent_score is not None:
                    check(f"Row {i+1} health score == {exp_row[6]} (tol 0.5)",
                          num_close(agent_score, exp_row[6], abs_tol=0.5, rel_tol=0.0),
                          f"Expected ~{exp_row[6]}, got {agent_score}")
                # Per-rank name match with tie-tolerance: agent's rank-i name should be
                # in the expected top-5 set (allows ties at the same score to swap order).
                agent_name_val = cell(agent_row, name_idx)
                if agent_name_val is not None:
                    agent_name = str(agent_name_val).strip().lower()
                    expected_top5_names = {str(r[0] or "").strip().lower() for r in expected["top5_lowest"][:5]}
                    check(f"Row {i+1} customer name in expected top-5 (tie-tolerance)",
                          agent_name in expected_top5_names,
                          f"agent rank-{i+1}={agent_name}, expected_top5={expected_top5_names}")
            # Set-overlap requirement tightened to >=4/5 (was >=3/5).
            agent_top5_names = {str(cell(r, name_idx) or "").strip().lower() for r in rows[:5]}
            expected_top5_names = {str(r[0] or "").strip().lower() for r in expected["top5_lowest"][:5]}
            overlap = len(agent_top5_names & expected_top5_names)
            check(f"Top-5 lowest customer names overlap (>=4/5)",
                  overlap >= 4,
                  f"agent: {agent_top5_names}, expected: {expected_top5_names}")

            # Verify ascending sort (by health_score)
            scores = [float(r[score_idx]) for r in rows if cell(r, score_idx) is not None]
            check("Health Scores sorted ascending by Health_Score",
                  scores == sorted(scores),
                  f"first 5: {scores[:5]}")

    # --- Critical Accounts ---
    print("\n--- Critical Accounts ---")
    ws = get_sheet(wb, "Critical Accounts")
    if ws:
        col = header_map(ws)
        status_idx = col.get("status", 7)
        score_idx = col.get("health_score", 6)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Critical count is deterministic from DB - require exact equality.
        check("Critical Accounts count",
              len(rows) == expected["critical_count"],
              f"Expected {expected['critical_count']}, got {len(rows)}")
        # Verify all entries have Status == Critical
        if rows:
            non_critical = [r for r in rows if cell(r, status_idx) and not str_match(cell(r, status_idx), "Critical")]
            check("All Critical Accounts rows have Status='Critical'",
                  len(non_critical) == 0,
                  f"Found {len(non_critical)} non-Critical rows")
            # Sort ascending
            scores = [float(r[score_idx]) for r in rows if cell(r, score_idx) is not None]
            check("Critical Accounts sorted ascending",
                  scores == sorted(scores),
                  f"first 5: {scores[:5]}")

    # --- Summary by Segment ---
    print("\n--- Summary by Segment ---")
    ws = get_sheet(wb, "Summary by Segment")
    if ws:
        col = header_map(ws)
        seg_idx = col.get("segment", 0)
        tot_idx = col.get("total_customers", 1)
        healthy_idx = col.get("healthy_count", 2)
        atrisk_idx = col.get("at_risk_count", 3)
        crit_idx = col.get("critical_count", 4)
        avg_idx = col.get("avg_health_score", 5)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Summary rows", len(rows) == 4,
              f"Expected 4, got {len(rows)}")

        # Alphabetical sort
        seg_names = [str(cell(r, seg_idx)) for r in rows if cell(r, seg_idx)]
        check("Summary sorted alphabetically by Segment",
              seg_names == sorted(seg_names),
              f"order: {seg_names}")

        for exp_seg in expected["segment_summary"]:
            seg = exp_seg["Segment"]
            matched = None
            for r in rows:
                if str_match(cell(r, seg_idx), seg):
                    matched = r
                    break
            if matched:
                tot = cell(matched, tot_idx)
                hlth = cell(matched, healthy_idx)
                ar = cell(matched, atrisk_idx)
                cr = cell(matched, crit_idx)
                avg = cell(matched, avg_idx)
                # Exact integer counts (was tol=5/10)
                check(f"{seg} Total_Customers == {exp_seg['Total_Customers']}",
                      num_close(tot, exp_seg["Total_Customers"], abs_tol=1, rel_tol=0.0),
                      f"Expected {exp_seg['Total_Customers']}, got {tot}")
                check(f"{seg} Healthy_Count == {exp_seg['Healthy_Count']}",
                      num_close(hlth, exp_seg["Healthy_Count"], abs_tol=2, rel_tol=0.0),
                      f"Expected {exp_seg['Healthy_Count']}, got {hlth}")
                check(f"{seg} At_Risk_Count == {exp_seg['At_Risk_Count']}",
                      num_close(ar, exp_seg["At_Risk_Count"], abs_tol=2, rel_tol=0.0),
                      f"Expected {exp_seg['At_Risk_Count']}, got {ar}")
                check(f"{seg} Critical_Count == {exp_seg['Critical_Count']}",
                      num_close(cr, exp_seg["Critical_Count"], abs_tol=2, rel_tol=0.0),
                      f"Expected {exp_seg['Critical_Count']}, got {cr}")
                # Tighten avg health score (was tol=2.0); use 0.5 (2-decimal precision)
                check(f"{seg} Avg_Health_Score ~= {exp_seg['Avg_Health_Score']}",
                      num_close(avg, exp_seg["Avg_Health_Score"], abs_tol=0.5, rel_tol=0.0),
                      f"Expected {exp_seg['Avg_Health_Score']}, got {avg}")
            else:
                check(f"Segment '{seg}' found", False, "Not in output")


def check_notion(expected):
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Lenient title match for "Customer Health Dashboard": normalized
    # containment, Levenshtein <= 3, or keyword coverage (was ILIKE exact).
    cur.execute("""
        SELECT id, properties FROM notion.pages
        ORDER BY id
    """)
    all_pages = cur.fetchall()
    expected_title = "Customer Health Dashboard"
    titles_seen = []
    page_id = None
    for pid, props in all_pages:
        title = extract_page_title(props)
        if title:
            titles_seen.append(title)
        haystack = title if title else (
            json.dumps(props) if isinstance(props, dict) else str(props))
        if page_id is None and loose_title_match(
                haystack, expected_title,
                keywords=["customer", "health", "dashboard"]):
            page_id = pid
    check("Notion page titled 'Customer Health Dashboard' exists (exact title)",
          page_id is not None,
          f"Titles seen: {titles_seen[:10]}")

    if page_id is not None:
        # notion.blocks schema: (id, object, parent_type, parent_id, ...,
        # block_data jsonb, position). Content blocks are children of the page
        # (parent_id = page id); text lives inside block_data rich_text entries.
        cur.execute("""
            SELECT block_data FROM notion.blocks
            WHERE parent_id = %s
            ORDER BY position
        """, (page_id,))
        blocks = cur.fetchall()
        check("Notion page has content blocks", len(blocks) > 0,
              f"Found {len(blocks)} blocks")

        # Validate page content includes total counts and at least 5 lowest customers
        all_text = " ".join(extract_block_text(b[0]) for b in blocks).lower()
        # Counts (strip thousands separators so "2,000" matches "2000")
        check(f"Notion mentions total customers ({expected['total_customers']})",
              str(expected["total_customers"]) in all_text.replace(",", ""),
              f"Sample: {all_text[:300]}")
        # Mention of categories
        check("Notion mentions 'Healthy' / 'At Risk' / 'Critical' status terms",
              "healthy" in all_text and ("at risk" in all_text or "at-risk" in all_text) and "critical" in all_text,
              f"Sample: {all_text[:300]}")
        # 5 lowest customers' names
        names_found = sum(1 for r in expected["top5_lowest"][:5]
                          if str(r[0] or "").lower() in all_text)
        check("Notion lists at least 4 of the 5 lowest-score customer names",
              names_found >= 4,
              f"Found {names_found}/5 names in content")

    cur.close()
    conn.close()


def check_calendar(expected):
    print("\n=== Checking Calendar Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%follow%up%' OR LOWER(summary) LIKE '%follow-up%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()

    below_15_count = len(expected["below_15"])
    # Tighten to exact count
    check(f"Calendar events for critical accounts (score < 15) == {below_15_count}",
          len(events) == below_15_count,
          f"Expected {below_15_count}, got {len(events)}")

    # Check ALL events match lowest-score customers in order
    if events and expected["below_15"]:
        match_count = 0
        for i in range(min(len(events), len(expected["below_15"]))):
            evt_summary = events[i][0] or ""
            exp_name = expected["below_15"][i][0]
            if exp_name.lower() in evt_summary.lower():
                match_count += 1
        n_check = min(len(events), len(expected["below_15"]))
        check(f"All events match lowest-score order ({match_count}/{n_check})",
              match_count >= n_check - 1,
              f"Order mismatch")

    # Check events start on the Monday after the analysis date (2026-04-27)
    # and one per day at 10:00.
    from datetime import time
    if events:
        first_dt = events[0][2]
        if first_dt:
            first_date = first_dt.date() if hasattr(first_dt, 'date') else first_dt
            check("First event on 2026-04-27",
                  str(first_date) == "2026-04-27",
                  f"First event date: {first_date}")
            # 10:00 AM
            try:
                start_time = first_dt.time() if hasattr(first_dt, 'time') else None
                check("First event at 10:00 AM",
                      start_time is not None and start_time.hour == 10 and start_time.minute == 0,
                      f"Got time: {start_time}")
            except AttributeError:
                pass
        # 30-minute duration on each event
        if len(events) > 0:
            ok_30min = 0
            for ev in events:
                if ev[2] and ev[3]:
                    delta = (ev[3] - ev[2]).total_seconds() / 60
                    if 25 <= delta <= 35:
                        ok_30min += 1
            check(f"Events are 30 minutes long ({ok_30min}/{len(events)})",
                  ok_30min >= len(events) - 1,
                  f"{ok_30min}/{len(events)} have 30-min duration")
        # Description mentions segment, region, score
        desc_ok = 0
        for ev in events:
            d = str(ev[1] or "").lower()
            if any(s in d for s in ["segment", "region", "score"]):
                desc_ok += 1
        check(f"Event descriptions reference segment/region/score ({desc_ok}/{len(events)})",
              desc_ok >= len(events) - 1,
              f"{desc_ok}/{len(events)} have substantive description")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Computing Expected Values ===")
    try:
        expected = compute_expected_values()
        print(f"  Total customers: {expected['total_customers']}")
        print(f"  Critical: {expected['critical_count']}")
        print(f"  Customers with score < 15: {len(expected['below_15'])}")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    check_excel(args.agent_workspace, expected)
    check_notion(expected)
    check_calendar(expected)

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    # Require all checks pass (FAIL_COUNT == 0)
    success = FAIL_COUNT == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
