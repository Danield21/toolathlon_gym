"""Evaluation for sf-sla-breach-audit."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "SLA_Breach_Audit.xlsx")
    gt_file = os.path.join(gt_dir, "SLA_Breach_Audit.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # --- Check Breach Analysis sheet ---
    print("  Checking Breach Analysis sheet...")
    a_rows = load_sheet_rows(agent_wb, "Breach Analysis")
    g_rows = load_sheet_rows(gt_wb, "Breach Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Breach Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Breach Analysis' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Breach Analysis: {g_row[0]}")
                continue

            # Total_Tickets (col 1) - exact match (counts must be precise)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    all_errors.append(f"{key}.Total_Tickets: {a_row[1]} vs {g_row[1]}")

            # Avg_Response_Hours (col 2) - tighten ±0.05
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.05):
                    all_errors.append(f"{key}.Avg_Response_Hours: {a_row[2]} vs {g_row[2]}")

            # SLA_Target_Hours (col 3) - exact (policy values)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0):
                    all_errors.append(f"{key}.SLA_Target_Hours: {a_row[3]} vs {g_row[3]}")

            # Breached_Count (col 4) - exact match (counts)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0):
                    all_errors.append(f"{key}.Breached_Count: {a_row[4]} vs {g_row[4]}")

            # Breach_Rate (col 5) - tighten ±0.1 percentage points
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.1):
                    all_errors.append(f"{key}.Breach_Rate: {a_row[5]} vs {g_row[5]}")

        if not [e for e in all_errors if "Breach Analysis" in e]:
            print("    PASS")
        else:
            print(f"    ERRORS found")

    # --- Check Summary sheet ---
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Summary: {g_row[0]}")
                continue

            g_val = g_row[1]
            a_val = a_row[1]

            try:
                float(a_val); float(g_val)
                # Tighter tolerance for percentages and counts in summary
                tol = 0.1 if "rate" in key else 0
                if not num_close(a_val, g_val, tol):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val} (tol={tol})")
            except (TypeError, ValueError):
                if not str_match(a_val, g_val):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val}")

        if not [e for e in all_errors if "Summary" in e]:
            print("    PASS")

    # --- Check email sent ---
    print("  Checking email...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT m.subject, m.to_addr, m.body_text, m.from_addr
            FROM email.messages m
        """)
        all_msgs = cur.fetchall()

        # Tightened: to ops-director, subject 'SLA Breach Audit Results',
        # body contains overall breach rate value and worst priority name
        # (sender address is fixed by the email environment and not checked)
        # Determine expected values from GT Summary sheet
        gt_summary = load_sheet_rows(gt_wb, "Summary")
        expected_overall_rate = None
        expected_worst = None
        for row in gt_summary[1:] if gt_summary else []:
            if not row or row[0] is None: continue
            k = str(row[0]).strip().lower()
            if k == "overall_breach_rate" and row[1] is not None:
                try:
                    expected_overall_rate = float(row[1])
                except (TypeError, ValueError):
                    pass
            elif k == "worst_priority" and row[1] is not None:
                expected_worst = str(row[1]).strip()

        found_email = None
        for subj, to_addr, body, from_addr in all_msgs:
            subj_str = str(subj or "").lower()
            to_str = str(to_addr or "").lower()
            if ("sla breach audit results" in subj_str or
                ("sla" in subj_str and "breach" in subj_str and "audit" in subj_str)):
                if "ops-director@company.com" in to_str:
                    found_email = (subj, to_addr, body, from_addr)
                    break

        if not found_email:
            all_errors.append("Email not found: to=ops-director@company.com, subject contains 'SLA Breach Audit Results'")
        else:
            print("    Email envelope PASS")
            body_str = str(found_email[2] or "")
            # Body must contain overall breach rate value (as a number) and worst priority name
            if expected_overall_rate is not None:
                rate_str = f"{expected_overall_rate:g}"
                rate_alt = f"{expected_overall_rate:.2f}"
                if rate_str not in body_str and rate_alt not in body_str:
                    all_errors.append(f"Email body missing overall breach rate {rate_str} or {rate_alt}; body: {body_str[:200]}")
                else:
                    print(f"    Body contains overall_breach_rate={rate_str}")
            if expected_worst is not None:
                if expected_worst.lower() not in body_str.lower():
                    all_errors.append(f"Email body missing worst priority '{expected_worst}'; body: {body_str[:200]}")
                else:
                    print(f"    Body mentions worst_priority={expected_worst}")

        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Email check error: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
