"""
Evaluation for 12306-team-qufu-conference-excel-gcal-email.

Checks:
1. Conference_Travel_Plan.xlsx exists
2. "Outbound" sheet has at least 4 data rows
3. "Outbound" sheet has Train_No column containing G235 or G168
4. "Return" sheet has at least 4 data rows
5. "Coordination_Notes" sheet has at least 3 rows
6. GCal has at least 2 new travel events (beyond the conference itself)
7. Email sent to beijing_team@uni.edu
8. Email sent to shanghai_team@uni.edu
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Conference_Travel_Plan.xlsx ===")
    import glob

    pattern = os.path.join(agent_workspace, "*.xlsx")
    all_xlsx = glob.glob(pattern)
    conf_files = [f for f in all_xlsx if any(
        kw in os.path.basename(f).lower()
        for kw in ["conference", "travel", "qufu", "plan"]
    )]

    if not conf_files:
        record("Conference travel xlsx exists", False,
               f"No matching xlsx in {agent_workspace}")
        return
    record("Conference travel xlsx exists", True)

    xlsx_path = conf_files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Outbound sheet
    if "outbound" not in sheet_names_lower:
        record("Outbound sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Outbound sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("outbound")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        # Task says 5 team members (3 Beijing + 2 Shanghai) on outbound.
        record("Outbound has exactly 5 data rows", len(data_rows) == 5,
               f"Found {len(data_rows)} data rows")

        import re
        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        has_g235 = bool(re.search(r"\bg235\b", all_text))
        has_g168 = bool(re.search(r"\bg168\b", all_text))
        # Task requires BOTH trains (one per origin city).
        record("Outbound contains both G235 and G168 train numbers",
               has_g235 and has_g168,
               f"G235:{has_g235} G168:{has_g168}")

    # Check Return sheet
    if "return" not in sheet_names_lower:
        record("Return sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Return sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("return")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Return has exactly 5 data rows", len(data_rows) == 5,
               f"Found {len(data_rows)} data rows")

    # Check Coordination_Notes sheet
    coord_match = [s for s in sheet_names_lower if "coord" in s or "note" in s]
    if not coord_match:
        record("Coordination_Notes sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Coordination_Notes sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(coord_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        # Task says at least four coordination checkpoints.
        record("Coordination_Notes has at least 4 rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} data rows")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Conference_Travel_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        # Coordination_Notes can be ordered differently; allow >= GT rows for that sheet.
        sn_lower = gt_sheet_name.strip().lower()
        if "coord" in sn_lower or "note" in sn_lower:
            record(f"GT '{gt_sheet_name}' row count >= GT", len(agent_rows) >= len(gt_rows),
                   f"Expected >= {len(gt_rows)}, got {len(agent_rows)}")
            # Skip per-row value match for coordination notes (free-form text).
            continue

        record(f"GT '{gt_sheet_name}' row count exact", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        # Build agent rows index keyed by Team_Member (column 0) so that
        # row order between agent and GT can differ — the agent might list
        # Beijing or Shanghai team first, or sort alphabetically.
        agent_by_member = {}
        for ar in agent_rows:
            if ar and ar[0] is not None:
                key = str(ar[0]).strip().lower()
                agent_by_member[key] = ar

        # Iterate ALL GT rows; lookup by Team_Member name (col 0); ints tol=0, floats 2%.
        for idx, gt_row in enumerate(gt_rows):
            gt_member = gt_row[0]
            member_key = str(gt_member).strip().lower() if gt_member is not None else ""
            a_row = agent_by_member.get(member_key)
            if a_row is None:
                record(f"GT '{gt_sheet_name}' member '{gt_member}' present", False,
                       f"Agent rows have members: {list(agent_by_member.keys())}")
                continue
            row_ok = True
            fail_detail = None
            for col_idx in range(min(len(gt_row), len(a_row))):
                gt_val = gt_row[col_idx]
                a_val = a_row[col_idx]
                if gt_val is None:
                    continue
                if isinstance(gt_val, int):
                    ok = num_close(a_val, gt_val, 0)
                elif isinstance(gt_val, float):
                    ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.02, 1.0))
                else:
                    ok = str_match(a_val, gt_val)
                if not ok:
                    fail_detail = (col_idx, gt_val, a_val)
                    row_ok = False
                    break
            if row_ok:
                record(f"GT '{gt_sheet_name}' member '{gt_member}' values match", True)
            else:
                record(f"GT '{gt_sheet_name}' member '{gt_member}' col {fail_detail[0]+1}",
                       False, f"Expected {fail_detail[1]}, got {fail_detail[2]}")
    gt_wb.close()


def check_gcal():
    print("\n=== Check 2: Calendar travel events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE (start_datetime >= '2026-03-12' AND start_datetime < '2026-03-13')
          AND summary NOT ILIKE '%confucian studies conference%'
        ORDER BY start_datetime
    """)
    mar12_events = cur.fetchall()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE (start_datetime >= '2026-03-15' AND start_datetime < '2026-03-16')
          AND summary NOT ILIKE '%confucian studies conference%'
        ORDER BY start_datetime
    """)
    mar15_events = cur.fetchall()
    cur.close()
    conn.close()
    # Task: 2 outbound events on Mar 12 (Beijing team + Shanghai team) + 1 return event on Mar 15.
    bj = any("beijing" in (e[0] or "").lower() or "g235" in (e[0] or "").lower() for e in mar12_events)
    sh = any("shanghai" in (e[0] or "").lower() or "g168" in (e[0] or "").lower() for e in mar12_events)
    record(
        "Mar 12 has Beijing team travel event",
        bj,
        f"Mar12 events: {[e[0] for e in mar12_events]}",
    )
    record(
        "Mar 12 has Shanghai team travel event",
        sh,
        f"Mar12 events: {[e[0] for e in mar12_events]}",
    )
    record(
        "Mar 15 has return travel event",
        len(mar15_events) >= 1,
        f"Mar15 events: {[e[0] for e in mar15_events]}",
    )


def _query_email_count(addr_pattern):
    """Query email count using a fresh connection to avoid transaction aborts."""
    cnt = 0
    sent = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s"
            " AND from_addr NOT ILIKE %s",
            (f"%{addr_pattern}%", f"%{addr_pattern}%"),
        )
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
    except Exception:
        pass
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COUNT(*) FROM email.sent_log WHERE to_addr::text ILIKE %s",
            (f"%{addr_pattern}%",),
        )
        sent = cur2.fetchone()[0]
        cur2.close()
        conn2.close()
    except Exception:
        pass
    return cnt, sent


def check_notion():
    print("\n=== Check 3: Notion 'Qufu Conference Travel Coordination' page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Notion page query", False, str(e))
        return

    found = False
    titles = []
    for page_id, props in pages:
        title = ""
        if isinstance(props, dict):
            title_obj = props.get("title", {})
            if isinstance(title_obj, dict):
                title_list = title_obj.get("title", [])
                if isinstance(title_list, list):
                    for t in title_list:
                        if isinstance(t, dict):
                            title += t.get("text", {}).get("content", "")
        titles.append(title)
        tl = title.lower()
        if "qufu" in tl and ("conference" in tl or "travel" in tl) and "coordin" in tl:
            found = True
            break
    record(
        "Notion 'Qufu Conference Travel Coordination' page exists",
        found,
        f"Notion titles found: {titles[:6]}",
    )


def check_emails():
    print("\n=== Check 4: Emails to teams ===")

    beijing_cnt, beijing_sent = _query_email_count("beijing_team@uni.edu")
    record("Email sent to beijing_team@uni.edu", beijing_cnt >= 1 or beijing_sent >= 1,
           f"messages: {beijing_cnt}, sent_log: {beijing_sent}")

    shanghai_cnt, shanghai_sent = _query_email_count("shanghai_team@uni.edu")
    record("Email sent to shanghai_team@uni.edu", shanghai_cnt >= 1 or shanghai_sent >= 1,
           f"messages: {shanghai_cnt}, sent_log: {shanghai_sent}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_notion()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0 and PASS_COUNT > 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
