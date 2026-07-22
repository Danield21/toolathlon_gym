"""Evaluation script for fetch-yf-dividend-strategy-gsheet-email.

Uses verify_v2 smart column matching: accepts semantically-equivalent agent
column names (e.g. "Primary Dimension" → "Symbol") when the underlying data
matches the ground truth. Column existence is judged by LLM (deterministic,
cached); value correctness is compared programmatically.
"""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-yf-dividend-strategy-gsheet-email"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Dividend_Strategy_Report.xlsx")
    check("Dividend_Strategy_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Dividend_Strategy_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, has >= min_rows, and contains required columns
            (via exact match or LLM semantic mapping when verify_v2 is available)."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(data_rows) >= min_rows, f"got {len(data_rows)}")

            if _HAS_VERIFY_V2 and gt_wb is not None:
                raw_headers, agent_dict_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for expected_col in expected_cols:
                    gt_vals = get_gt_column_values(gt_wb, sheet_name, expected_col)
                    ok, matched_col, reason = smart_column_exists(
                        expected_col=expected_col,
                        agent_headers=raw_headers,
                        gt_samples=gt_vals[:3],
                        agent_rows=agent_dict_rows,
                        task_name=TASK_NAME,
                    )
                    detail = reason
                    if ok and matched_col and matched_col.lower() != expected_col.lower():
                        detail = f"LLM-mapped to {matched_col!r}"
                    check(f"{sheet_name} has {expected_col} column", ok, detail)
            else:
                # Fallback to legacy strict header check if verify_v2 unavailable
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                for expected_col in expected_cols:
                    check(f"{sheet_name} has {expected_col} column",
                          expected_col.lower() in headers, f"headers: {headers[:8]}")

        # Tightened row counts == GT
        gt_data_rows = []
        gt_metrics_rows = []
        if gt_wb is not None:
            if "Data_Analysis" in gt_wb.sheetnames:
                gt_data_rows = list(gt_wb["Data_Analysis"].iter_rows(min_row=2, values_only=True))
            if "Metrics" in gt_wb.sheetnames:
                gt_metrics_rows = list(gt_wb["Metrics"].iter_rows(min_row=2, values_only=True))

        check_columns("Data_Analysis",
                      ['Symbol', 'Current_Price', 'Target_Price', 'Upside'],
                      min_rows=len(gt_data_rows) if gt_data_rows else 5)
        check_columns("Metrics", ['Metric', 'Value'],
                      min_rows=len(gt_metrics_rows) if gt_metrics_rows else 3)
        check_columns("Recommendations", ['Priority', 'Action'], min_rows=2)

        # Per-Symbol value comparison vs GT
        if gt_wb is not None and "Data_Analysis" in wb.sheetnames and "Data_Analysis" in gt_wb.sheetnames:
            ws = wb["Data_Analysis"]
            gt_ws = gt_wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            header_map = {h: i for i, h in enumerate(headers)}
            sym_idx_gt = gt_headers.index("symbol") if "symbol" in gt_headers else 0
            sym_idx_a = header_map.get("symbol", 0)
            gt_by_sym = {str(r[sym_idx_gt]).strip().upper(): r for r in gt_data_rows if r and r[sym_idx_gt]}
            agent_by_sym = {str(r[sym_idx_a]).strip().upper(): r for r in data_rows if r and len(r) > sym_idx_a and r[sym_idx_a]}
            for sym, gt_row in gt_by_sym.items():
                found = sym in agent_by_sym
                check(f"Data_Analysis symbol '{sym}' present", found)
                if found:
                    agent_row = agent_by_sym[sym]
                    for ci, gt_h in enumerate(gt_headers):
                        if not gt_h or ci >= len(gt_row) or gt_h == "symbol":
                            continue
                        gv = gt_row[ci]
                        agent_ci = header_map.get(gt_h)
                        if agent_ci is None or agent_ci >= len(agent_row):
                            continue
                        av = agent_row[agent_ci]
                        gf = safe_float(gv)
                        af = safe_float(av)
                        if gf is not None and af is not None:
                            # Tighter price tolerance: 5% rel
                            tol = max(0.5, abs(gf) * 0.05)
                            check(f"Data_Analysis '{sym}' {gt_h} ~{gf}",
                                  abs(gf - af) <= tol, f"got {av}")

        # Email: subject = 'Analysis Report Complete' AND recipient team-lead@company.com
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Tighten subject to equality after lower/strip
            cur.execute(
                """SELECT subject FROM email.messages
                   WHERE to_addr::text ILIKE %s
                     AND LOWER(TRIM(subject)) = LOWER(%s)""",
                ('%team-lead@company.com%', 'Analysis Report Complete'),
            )
            emails = cur.fetchall()
            check("Email to team-lead@company.com with subject 'Analysis Report Complete' sent",
                  len(emails) >= 1, f"found {len(emails)} matching emails")
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        # GSheet: title equals 'Portfolio Dividend Tracker' (case-insensitive)
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Title equality after lower/strip; also accept the legacy 'Yf Dividend Tracker' name
            cur.execute("""SELECT title FROM gsheet.spreadsheets
                           WHERE LOWER(TRIM(title)) IN (LOWER('Portfolio Dividend Tracker'), LOWER('Yf Dividend Tracker'))""")
            sheets = cur.fetchall()
            check("Cloud spreadsheet titled 'Portfolio Dividend Tracker' created",
                  len(sheets) >= 1, f"found {len(sheets)} matching sheets")
            conn.close()
        except Exception as e:
            check("GSheet check", False, str(e))

    check("yf_dividend_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "yf_dividend_processor.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
