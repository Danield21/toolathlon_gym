"""Evaluation for terminal-sf-fetch-excel-notion-gcal.
Checks:
1. SLA_Compliance_Report.xlsx with 4 sheets and correct data
2. Notion database "SLA Compliance Dashboard" with entries
3. Google Calendar monthly SLA review events
4. sla_analyzer.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _find_sheet(sheets_lower, sheets, *kws):
    """Find first sheet whose lowercase name contains all given keywords (any of)."""
    for i, s in enumerate(sheets_lower):
        if any(k in s for k in kws):
            return i
    return None


def _compute_compliance_from_db():
    """Compute per-priority compliance rate dynamically from Snowflake.

    Formula per SLA_Framework.pdf:
        compliance_rate = min(100, (SLA_target / avg_response_time) * 100)
    Returns dict {priority_lower: rate_percent}.
    """
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            'SELECT t."PRIORITY", AVG(t."RESPONSE_TIME_HOURS"), s."RESPONSE_TARGET_HOURS" '
            'FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t '
            'JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" s '
            '   ON t."PRIORITY" = s."PRIORITY" '
            'GROUP BY t."PRIORITY", s."RESPONSE_TARGET_HOURS"'
        )
        out = {}
        for pri, avg_rt, target in cur.fetchall():
            if pri is None or avg_rt is None or target is None:
                continue
            rate = min(100.0, float(target) / float(avg_rt) * 100.0)
            out[str(pri).strip().lower()] = round(rate, 1)
        cur.close()
        conn.close()
        return out
    except Exception:
        # Fallback to hardcoded GT values if DB unavailable.
        return {"high": 64.2, "medium": 65.1, "low": 93.2}


def check_excel(workspace):
    print("\n=== Check 1: SLA_Compliance_Report.xlsx ===")
    path = os.path.join(workspace, "SLA_Compliance_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    # Ticket_Summary sheet - must match priorities High/Medium/Low with compliance rates
    ts_idx = _find_sheet(sheets_lower, sheets, "ticket", "summary")
    if ts_idx is None:
        check("Ticket_Summary sheet exists", False, f"sheets: {sheets}")
    else:
        ws = wb[sheets[ts_idx]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        check("Ticket_Summary has 3 priority rows",
              len(data_rows) == 3, f"Found {len(data_rows)}")
        # Build row lookup by priority (col 0)
        by_pri = {}
        for r in data_rows:
            if r and r[0]:
                by_pri[str(r[0]).strip().lower()] = r
        for pri in ["high", "medium", "low"]:
            check(f"Ticket_Summary has '{pri}' priority row",
                  pri in by_pri, f"got priorities: {list(by_pri)}")
        # Validate compliance rates dynamically from Snowflake (formula in
        # SLA_Framework.pdf: min(100, target/avg_rt * 100)). Tol 2.0% absolute
        # to allow rounding differences.
        expected_compliance = _compute_compliance_from_db()
        # Locate the compliance_rate column from header
        header_row_ts = list(rows[0]) if rows else []
        compliance_col = None
        for ci, h in enumerate(header_row_ts):
            if h and "complian" in str(h).lower():
                compliance_col = ci
                break
        for pri, exp_pct in expected_compliance.items():
            row = by_pri.get(pri)
            if not row:
                continue
            if compliance_col is not None and compliance_col < len(row):
                v = row[compliance_col]
                if isinstance(v, (int, float)):
                    ok = abs(float(v) - exp_pct) <= 2.0
                    check(f"Ticket_Summary {pri} compliance rate ~{exp_pct}",
                          ok, f"got {v} (col {compliance_col})")
                    continue
            # Fallback: any numeric value in the row that matches.
            nums = [c for c in row if isinstance(c, (int, float))]
            ok = any(abs(float(c) - exp_pct) <= 2.0 for c in nums)
            check(f"Ticket_Summary {pri} compliance rate ~{exp_pct}",
                  ok, f"got nums {nums}")

    # Benchmark_Comparison sheet - must contain 3 metrics + Meets/Below status
    bc_idx = _find_sheet(sheets_lower, sheets, "benchmark", "comparison")
    if bc_idx is None:
        check("Benchmark_Comparison sheet exists", False, f"sheets: {sheets}")
    else:
        ws2 = wb[sheets[bc_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Benchmark_Comparison has 3 rows",
              len(data_rows2) == 3, f"Found {len(data_rows2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Contains response time metric", "response" in all_text2, f"Text: {all_text2[:120]}")
        check("Contains satisfaction metric", "satisfaction" in all_text2, f"Text: {all_text2[:120]}")
        check("Contains compliance metric", "compliance" in all_text2, f"Text: {all_text2[:120]}")
        # Status must be Meets or Below for each of the 3 rows.
        # Per SLA_Framework.pdf: response time better-if-lower, satisfaction
        # and compliance_rate better-if-higher. Internal: avg_resp=15.03,
        # csat=3.26, compliance=73.2 (overall); benchmark from mock dashboard
        # is response=8, csat=3.5, compliance=85. All 3 rows should be
        # "Below". (We only enforce "no row has clearly invalid status".)
        status_vals = []
        for r in data_rows2:
            for c in r:
                if isinstance(c, str) and c.strip().lower() in ("meets", "below"):
                    status_vals.append(c.strip().lower())
        check("Benchmark_Comparison has 3 valid status values",
              len(status_vals) >= 3, f"got {status_vals}")
        # Validate per-row Status against gap sign: each data row must have
        # a status field, and Status must be 'Below' iff the internal value is
        # worse than benchmark for that metric (response_time: internal>bench
        # means Below; csat/compliance: internal<bench means Below).
        # Locate columns
        bc_headers = [str(c).strip().lower() if c else "" for c in (rows2[0] if rows2 else [])]
        def col_idx(*kws):
            for i, h in enumerate(bc_headers):
                if all(k in h for k in kws):
                    return i
            return None
        metric_col = col_idx("metric")
        internal_col = col_idx("internal")
        bench_col = col_idx("benchmark") or col_idx("industry")
        status_col = col_idx("status")
        if all(x is not None for x in (metric_col, internal_col, bench_col, status_col)):
            for r in data_rows2:
                if metric_col >= len(r) or status_col >= len(r):
                    continue
                metric_name = str(r[metric_col] or "").lower()
                try:
                    iv = float(r[internal_col]) if r[internal_col] is not None else None
                    bv = float(r[bench_col]) if r[bench_col] is not None else None
                except (TypeError, ValueError):
                    continue
                if iv is None or bv is None:
                    continue
                status = str(r[status_col] or "").strip().lower()
                # Determine expected Below: response_time better-if-lower
                if "response" in metric_name:
                    expected_below = iv > bv
                else:
                    # satisfaction & compliance rate: better-if-higher
                    expected_below = iv < bv
                if expected_below:
                    check(f"Benchmark_Comparison '{metric_name[:30]}' status=Below",
                          status == "below", f"got '{status}'")
                else:
                    check(f"Benchmark_Comparison '{metric_name[:30]}' status=Meets",
                          status == "meets", f"got '{status}'")

    # Agent_Performance sheet
    ap_idx = _find_sheet(sheets_lower, sheets, "agent", "performance")
    if ap_idx is None:
        check("Agent_Performance sheet exists", False, f"sheets: {sheets}")
    else:
        ws3 = wb[sheets[ap_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Agent_Performance has at least 5 rows",
              len(data_rows3) >= 5, f"Found {len(data_rows3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        # Yes/No status indicating SLA met (must use word-bounded values, not
        # 'no' as substring of 'novelty').
        import re as _re
        status_tokens_ap = []
        for r in data_rows3:
            for c in r:
                if isinstance(c, str) and c.strip().lower() in ("yes", "no"):
                    status_tokens_ap.append(c.strip().lower())
        check("Agent_Performance has yes/no SLA flag column",
              len(status_tokens_ap) >= len(data_rows3),
              f"got status tokens: {status_tokens_ap}")
        # Validate at least one known agent name appears.
        try:
            conn_ap = psycopg2.connect(**DB_CONFIG)
            cur_ap = conn_ap.cursor()
            cur_ap.execute(
                'SELECT DISTINCT "AGENT_NAME" '
                'FROM sf_data."SUPPORT_CENTER__PUBLIC__AGENTS" '
                'WHERE "AGENT_NAME" IS NOT NULL LIMIT 20'
            )
            db_agents = [str(r[0]).lower() for r in cur_ap.fetchall() if r[0]]
            cur_ap.close()
            conn_ap.close()
            agents_present = sum(1 for a in db_agents if a in all_text3)
            check("Agent_Performance includes real agent names from DB",
                  agents_present >= 3,
                  f"only {agents_present}/{len(db_agents)} known agents found")
        except Exception:
            pass

    # Monthly_Trend sheet
    mt_idx = _find_sheet(sheets_lower, sheets, "monthly", "trend")
    if mt_idx is None:
        check("Monthly_Trend sheet exists", False, f"sheets: {sheets}")
    else:
        ws4 = wb[sheets[mt_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Monthly_Trend has at least 3 rows",
              len(data_rows4) >= 3, f"Found {len(data_rows4)}")
        # Validate numeric content: must have a positive total_tickets value
        # and a compliance_pct value in [0, 100] for at least 3 rows.
        # Avoids fabricating obviously wrong data.
        mt_headers = [str(c).strip().lower() if c else "" for c in (rows4[0] if rows4 else [])]
        tickets_col = None
        pct_col = None
        for ci, h in enumerate(mt_headers):
            if "ticket" in h and tickets_col is None:
                tickets_col = ci
            if ("complian" in h or "pct" in h or "%" in h) and pct_col is None:
                pct_col = ci
        valid_rows = 0
        for r in data_rows4:
            t_ok = False
            p_ok = False
            if tickets_col is not None and tickets_col < len(r):
                v = r[tickets_col]
                if isinstance(v, (int, float)) and v > 0:
                    t_ok = True
            else:
                # fallback: any positive integer-like value
                t_ok = any(isinstance(c, (int, float)) and c > 100 for c in r)
            if pct_col is not None and pct_col < len(r):
                v = r[pct_col]
                if isinstance(v, (int, float)) and 0.0 <= float(v) <= 100.0:
                    p_ok = True
            else:
                p_ok = any(isinstance(c, (int, float)) and 0.0 <= float(c) <= 100.0 for c in r)
            if t_ok and p_ok:
                valid_rows += 1
        check("Monthly_Trend rows have positive ticket counts and compliance 0-100",
              valid_rows >= 3, f"only {valid_rows} rows have valid numeric content")


def check_notion():
    print("\n=== Check 2: Notion SLA Compliance Dashboard ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM notion.databases")
    dbs = cur.fetchall()
    sla_db = None
    for db_id, title in dbs:
        title_str = ""
        if isinstance(title, list):
            title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
        elif isinstance(title, str):
            try:
                parsed = json.loads(title)
                if isinstance(parsed, list):
                    title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                else:
                    title_str = str(title)
            except Exception:
                title_str = str(title)
        else:
            title_str = str(title) if title else ""
        if "sla" in title_str.lower() and ("compliance" in title_str.lower() or "dashboard" in title_str.lower()):
            sla_db = (db_id, title_str)
            break

    check("SLA Compliance Dashboard exists", sla_db is not None,
          f"Databases: {[d[1] for d in dbs]}")

    if sla_db:
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE parent->>'database_id' = %s
        """, (sla_db[0],))
        pages = cur.fetchall()
        check("Dashboard has metric entries", len(pages) >= 3,
              f"Found {len(pages)} pages")
        # Validate at least 3 pages have meaningful Value AND
        # Benchmark numeric properties + Status (Meets|Below).
        valid_pages = 0
        for _pid, props in pages:
            if not props:
                continue
            if isinstance(props, str):
                try:
                    props = json.loads(props)
                except Exception:
                    continue
            if not isinstance(props, dict):
                continue
            has_value = False
            has_bench = False
            has_status = False
            for key, prop in props.items():
                if not isinstance(prop, dict):
                    continue
                key_l = key.lower()
                ptype = prop.get("type", "")
                # Number properties: Value, Benchmark
                if ptype == "number":
                    num = prop.get("number")
                    if num is not None and isinstance(num, (int, float)):
                        if "value" in key_l and "bench" not in key_l:
                            has_value = True
                        if "bench" in key_l:
                            has_bench = True
                # Select properties: Status
                if ptype == "select":
                    sel = prop.get("select") or {}
                    sel_name = (sel.get("name") or "").lower() if isinstance(sel, dict) else ""
                    if "status" in key_l and sel_name in ("meets", "below"):
                        has_status = True
                # Rich text fallback for status
                if ptype == "rich_text" and "status" in key_l:
                    rts = prop.get("rich_text") or []
                    if isinstance(rts, list):
                        joined = " ".join(rt.get("plain_text", "") for rt in rts if isinstance(rt, dict)).lower()
                        if "meets" in joined or "below" in joined:
                            has_status = True
            if has_value and has_bench and has_status:
                valid_pages += 1
        check("Dashboard pages have Value, Benchmark, Status (Meets/Below) properties",
              valid_pages >= 3,
              f"only {valid_pages} pages have full property set")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Check 3: Monthly SLA Review Events ===")
    # Required: 3 events on first Monday of April, May, June 2026 at 10 AM, 90min
    # First Monday of April 2026 = 2026-04-06
    # First Monday of May 2026 = 2026-05-04
    # First Monday of June 2026 = 2026-06-01
    expected_dates = ["2026-04-06", "2026-05-04", "2026-06-01"]

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute(
        "SELECT summary, start_datetime AT TIME ZONE 'UTC' as st_utc, "
        "end_datetime AT TIME ZONE 'UTC' as ed_utc FROM gcal.events "
        "WHERE lower(summary) LIKE '%%sla%%review%%' "
        "   OR lower(summary) LIKE '%%monthly%%sla%%' "
        "ORDER BY start_datetime")
    events = cur.fetchall()
    check("At least 3 SLA review events",
          len(events) >= 3, f"Found {len(events)} events")

    matched_dates = set()
    duration_ok_count = 0
    for s, sd, ed in events:
        if not sd:
            continue
        date_str = sd.date().isoformat() if hasattr(sd, 'date') else None
        if date_str not in expected_dates:
            continue
        # 10:00 UTC start
        if sd.hour != 10:
            continue
        matched_dates.add(date_str)
        # 90-minute duration
        if ed and (ed - sd).total_seconds() == 5400:
            duration_ok_count += 1
    check("Events on first Mondays of Apr/May/Jun 2026 at 10:00 UTC",
          matched_dates == set(expected_dates),
          f"matched dates: {sorted(matched_dates)}, expected {expected_dates}")
    check("All 3 events have 90-minute duration",
          duration_ok_count >= 3,
          f"duration-ok events: {duration_ok_count}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: sla_analyzer.py ===")
    path = os.path.join(workspace, "sla_analyzer.py")
    exists = os.path.exists(path)
    check("sla_analyzer.py exists", exists)
    if exists:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
        except Exception:
            content = ""
        text_l = content.lower()
        # Script must have non-trivial size and reference compliance computation logic.
        check("sla_analyzer.py is non-empty",
              len(content.strip()) >= 80,
              f"only {len(content.strip())} bytes")
        # Look for any of the typical compliance / SLA computation hints.
        keywords = ("complian", "sla", "/avg", "response_time", "target", "ratio", "percent", "/100", "* 100")
        found = sum(1 for k in keywords if k in text_l)
        check("sla_analyzer.py contains compliance computation logic",
              found >= 2,
              f"only matched {found} of {len(keywords)} hint keywords")


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: compliance rate values must be 0-100 (gap can be negative; counts/times must be >=0)
    path = os.path.join(workspace, "SLA_Compliance_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
        bad_cell = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            # Identify columns where negative is suspicious (counts, rates 0-100)
            no_neg_cols = []
            for ci, h in enumerate(headers):
                if any(kw in h for kw in ("count", "ticket_count", "compliance_rate",
                                          "compliance_pct", "tickets_handled",
                                          "avg_response", "satisfaction",
                                          "target_hours", "total_tickets")):
                    no_neg_cols.append(ci)
            for row in ws.iter_rows(min_row=2, values_only=True):
                for ci in no_neg_cols:
                    if ci < len(row) and isinstance(row[ci], (int, float)) and row[ci] < 0:
                        bad_cell = (sheet_name, headers[ci], row[ci])
                        break
                if bad_cell:
                    break
            if bad_cell:
                break
        check("No negative compliance/count values in SLA report",
              bad_cell is None,
              f"found negative value: {bad_cell}")

    # Notion: no duplicate SLA dashboard databases
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE LOWER(title::text) LIKE '%%sla%%'
        """)
        db_count = cur.fetchone()[0]
        check("No duplicate SLA databases in Notion", db_count <= 1,
              f"Found {db_count} SLA databases")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_gcal()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Strict: all checks must pass (no 70% threshold)
    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
