"""Evaluation for yf-earnings-calendar-report."""
import argparse
import os
import sys
import psycopg2
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def num_close_pct(a, b, pct=0.05):
    """Allow percentage-based tolerance for large numbers."""
    try:
        a, b = float(a), float(b)
        if b == 0:
            return abs(a) < 1.0
        return abs(a - b) / abs(b) <= pct
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    all_errors = []

    # ---- Check Excel ----
    agent_excel = os.path.join(args.agent_workspace, "Stock_Financial_Summary.xlsx")
    gt_excel = os.path.join(gt_dir, "Stock_Financial_Summary.xlsx")

    if not os.path.exists(agent_excel):
        all_errors.append("Agent output Stock_Financial_Summary.xlsx not found")
    elif not os.path.exists(gt_excel):
        all_errors.append("Groundtruth Stock_Financial_Summary.xlsx not found")
    else:
        agent_wb = openpyxl.load_workbook(agent_excel, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_excel, data_only=True)

        # Check Stock Overview sheet
        print("  Checking Stock Overview...")
        a_rows = load_sheet_rows(agent_wb, "Stock Overview")
        g_rows = load_sheet_rows(gt_wb, "Stock Overview")
        if a_rows is None:
            all_errors.append("Sheet 'Stock Overview' not found in agent output")
        elif g_rows is None:
            all_errors.append("Sheet 'Stock Overview' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []
            if len(a_data) != len(g_data):
                all_errors.append(f"Stock Overview row count: agent={len(a_data)}, expected={len(g_data)}")
            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().upper()] = row
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().upper()
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing ticker: {key}")
                    continue
                # Company_Name (col 1) - tolerant string match (substring)
                if len(a_row) > 1 and len(g_row) > 1 and g_row[1]:
                    a_name = str(a_row[1] or "").strip().lower()
                    g_name = str(g_row[1] or "").strip().lower()
                    # Accept if shared first word matches or g_name in a_name (or vice versa)
                    g_first = g_name.split()[0] if g_name else ""
                    if g_first not in a_name and a_name not in g_name and g_name not in a_name:
                        all_errors.append(f"{key}.Company_Name: '{a_row[1]}' vs '{g_row[1]}'")
                # Sector (col 2)
                if len(a_row) > 2 and len(g_row) > 2 and g_row[2]:
                    if not str_match(a_row[2], g_row[2]):
                        all_errors.append(f"{key}.Sector: '{a_row[2]}' vs '{g_row[2]}'")
                # Market Cap (col 3) - use percentage tolerance
                if len(a_row) > 3 and len(g_row) > 3 and g_row[3] is not None:
                    if not num_close_pct(a_row[3], g_row[3], 0.05):
                        all_errors.append(f"{key}.Market_Cap: {a_row[3]} vs {g_row[3]}")
                # Trailing PE (col 4)
                if len(a_row) > 4 and len(g_row) > 4 and g_row[4] is not None:
                    if not num_close(a_row[4], g_row[4], 0.5):
                        all_errors.append(f"{key}.Trailing_PE: {a_row[4]} vs {g_row[4]}")
                # Forward PE (col 5)
                if len(a_row) > 5 and len(g_row) > 5 and g_row[5] is not None:
                    if not num_close(a_row[5], g_row[5], 0.5):
                        all_errors.append(f"{key}.Forward_PE: {a_row[5]} vs {g_row[5]}")
                # Dividend_Yield (col 6) - decimal form per task.md (e.g., 0.02 for 2%).
                # GT stores decimal form. Tolerance 0.005 = 0.5 percentage points.
                if len(a_row) > 6 and len(g_row) > 6 and g_row[6] is not None:
                    a_dy = a_row[6]
                    g_dy = float(g_row[6])
                    if a_dy is not None:
                        try:
                            a_val = float(a_dy)
                            if not num_close(a_val, g_dy, 0.005):
                                all_errors.append(f"{key}.Dividend_Yield: {a_dy} vs {g_dy} (decimal form)")
                        except (TypeError, ValueError):
                            all_errors.append(f"{key}.Dividend_Yield not numeric: {a_dy}")
                # Trailing EPS (col 7)
                if len(a_row) > 7 and len(g_row) > 7 and g_row[7] is not None:
                    if not num_close(a_row[7], g_row[7], 0.5):
                        all_errors.append(f"{key}.Trailing_EPS: {a_row[7]} vs {g_row[7]}")
            # Sort by ticker alphabetical
            agent_tickers = [str(r[0]).strip().upper() for r in a_data if r and r[0]]
            if agent_tickers != sorted(agent_tickers):
                all_errors.append(f"Stock Overview not sorted alphabetically by Ticker: {agent_tickers}")
            if not any("Stock Overview" in e or "ticker" in e.lower() for e in all_errors):
                print("    PASS")

        # Check Profitability Metrics sheet
        print("  Checking Profitability Metrics...")
        a_rows = load_sheet_rows(agent_wb, "Profitability Metrics")
        g_rows = load_sheet_rows(gt_wb, "Profitability Metrics")
        if a_rows is None:
            all_errors.append("Sheet 'Profitability Metrics' not found in agent output")
        elif g_rows is None:
            all_errors.append("Sheet 'Profitability Metrics' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []
            if len(a_data) != len(g_data):
                all_errors.append(f"Profitability Metrics row count: agent={len(a_data)}, expected={len(g_data)}")
            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().upper()] = row
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().upper()
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing ticker in Profitability: {key}")
                    continue
                # Total Revenue (col 1) - pct tolerance
                if len(a_row) > 1 and len(g_row) > 1 and g_row[1] is not None:
                    if not num_close_pct(a_row[1], g_row[1], 0.05):
                        all_errors.append(f"{key}.Total_Revenue: {a_row[1]} vs {g_row[1]}")
                # Revenue_Growth (col 2) - must match GT (decimal form per task.md / GT)
                if len(a_row) > 2 and len(g_row) > 2 and g_row[2] is not None:
                    a_rg = a_row[2]
                    g_rg = float(g_row[2])
                    if a_rg is not None:
                        try:
                            a_val = float(a_rg)
                            if not num_close(a_val, g_rg, 0.005):
                                all_errors.append(f"{key}.Revenue_Growth: {a_rg} vs {g_rg} (decimal form)")
                        except (TypeError, ValueError):
                            all_errors.append(f"{key}.Revenue_Growth not numeric: {a_rg}")
                # Profit Margins (col 3)
                if len(a_row) > 3 and len(g_row) > 3 and g_row[3] is not None:
                    if not num_close(a_row[3], g_row[3], 0.02):
                        all_errors.append(f"{key}.Profit_Margins: {a_row[3]} vs {g_row[3]}")
                # Return_On_Equity (col 4) - decimal form per GT
                if len(a_row) > 4 and len(g_row) > 4 and g_row[4] is not None:
                    a_re = a_row[4]
                    g_re = float(g_row[4])
                    if a_re is not None:
                        try:
                            a_val = float(a_re)
                            if not num_close(a_val, g_re, 0.005):
                                all_errors.append(f"{key}.Return_On_Equity: {a_re} vs {g_re} (decimal form)")
                        except (TypeError, ValueError):
                            all_errors.append(f"{key}.Return_On_Equity not numeric: {a_re}")
                # Forward EPS (col 5)
                if len(a_row) > 5 and len(g_row) > 5 and g_row[5] is not None:
                    if not num_close(a_row[5], g_row[5], 0.5):
                        all_errors.append(f"{key}.Forward_EPS: {a_row[5]} vs {g_row[5]}")
            if not any("Profitability" in e for e in all_errors):
                print("    PASS")

    # ---- Check Google Sheet ----
    print("  Checking Google Sheet...")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    spreadsheets = cur.fetchall()
    found_sheet = False
    target_ss_id = None
    target_sheet_id = None
    for ss in spreadsheets:
        if ss[1] and "portfolio" in str(ss[1]).lower() and "watch" in str(ss[1]).lower():
            found_sheet = True
            target_ss_id = ss[0]
            # Check for Summary sheet
            cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id=%s", (target_ss_id,))
            sheets = cur.fetchall()
            summary_found = False
            for s in sheets:
                if s[1] and "summary" in str(s[1]).lower():
                    summary_found = True
                    target_sheet_id = s[0]
                    # Verify Summary has data: at least header row + 5 data rows
                    cur.execute("SELECT MAX(row_index) FROM gsheet.cells WHERE spreadsheet_id=%s AND sheet_id=%s",
                                (target_ss_id, target_sheet_id))
                    max_row = cur.fetchone()[0] or 0
                    if max_row < 5:
                        all_errors.append(f"Google Sheet Summary only has {max_row} rows, expected >= 5 data rows")
                    break
            if not summary_found:
                all_errors.append("Google Sheet missing 'Summary' sheet")
            break
    if not found_sheet:
        all_errors.append("Google Sheet 'Portfolio Watch List' not found")
    else:
        # Validate cells: each ticker AND its company-name should appear (per-row content check).
        if target_ss_id and target_sheet_id:
            cur.execute("SELECT value FROM gsheet.cells WHERE spreadsheet_id=%s AND sheet_id=%s",
                        (target_ss_id, target_sheet_id))
            cell_values_raw = [str(c[0] or "") for c in cur.fetchall()]
            cell_values = " ".join(cell_values_raw).upper()
            cell_values_lower = " ".join(cell_values_raw).lower()
            ticker_company = {
                "AMZN": "amazon",
                "GOOGL": "alphabet",
                "JNJ": "johnson",
                "JPM": "jp morgan",
                "XOM": "exxon",
            }
            for ticker, comp_kw in ticker_company.items():
                if ticker not in cell_values:
                    all_errors.append(f"Google Sheet missing ticker '{ticker}' in Summary cells")
                # Also check company-name keyword appears (avoids ticker-only stub rows).
                if comp_kw not in cell_values_lower and ticker.lower() != comp_kw:
                    all_errors.append(f"Google Sheet Summary missing company name keyword '{comp_kw}' for {ticker}")
        if not any("Google Sheet" in e for e in all_errors):
            print("    PASS")

    cur.close()
    conn.close()

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:15]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
