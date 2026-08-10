"""Evaluation for sf-sales-customer-loyalty.

Expected values are derived from the data warehouse at evaluation time (see
``_derive_expected``) so the checks are immune to LIFETIME_VALUE data drift.
The agent is expected to compute the same aggregates from the Snowflake/
warehouse data the task points it at:

  * Customers        = number of customers in that (segment, region) group
  * Avg_LTV          = average of CUSTOMERS.LIFETIME_VALUE per group
  * Total_Orders     = number of ORDER rows for those customers
  * Overall_Avg_LTV  = customer-count-weighted average of LIFETIME_VALUE

A fallback derived from the frozen groundtruth is used only if the warehouse
is unreachable, so an offline run still produces a sensible result.
"""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

# Fallback values (frozen groundtruth) — used only when the warehouse is
# unreachable. Keyed by lowercased "segment - region".
_FALLBACK_LOYALTY = {
    "consumer - latin america": (98, 1820.62, 1008),
    "consumer - europe": (110, 1737.73, 1132),
    "consumer - north america": (113, 1645.34, 1134),
    "consumer - asia pacific": (105, 1588.73, 1068),
    "consumer - middle east": (106, 1565.52, 1081),
    "enterprise - asia pacific": (107, 1782.56, 1098),
    "enterprise - europe": (104, 1697.17, 1018),
    "enterprise - latin america": (93, 1692.71, 932),
    "enterprise - middle east": (92, 1498.54, 913),
    "enterprise - north america": (117, 1411.42, 1097),
    "government - middle east": (94, 1625.88, 902),
    "government - north america": (95, 1602.48, 975),
    "government - latin america": (85, 1580.38, 836),
    "government - europe": (97, 1530.44, 943),
    "government - asia pacific": (103, 1468.67, 1023),
    "smb - middle east": (101, 1803.66, 976),
    "smb - europe": (99, 1706.02, 1007),
    "smb - asia pacific": (101, 1566.13, 1009),
    "smb - north america": (88, 1421.89, 927),
    "smb - latin america": (92, 1255.97, 921),
}
_FALLBACK_SUMMARY = {
    "total_combinations": 20,
    "highest_ltv_group": "Consumer - Latin America",
    "overall_avg_ltv": 1602.48,
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def is_numeric(val):
    if val is None:
        return False
    try:
        float(str(val).replace(",", "").replace("$", "").strip())
        return True
    except (ValueError, TypeError):
        return False


def _derive_expected():
    """Derive expected loyalty aggregates from the warehouse at runtime.

    Returns (loyalty_lookup, summary) where:
      loyalty_lookup maps "segment - region" (lowercased) -> (customers, avg_ltv, total_orders)
      summary maps "total_combinations"/"highest_ltv_group"/"overall_avg_ltv" -> value
    Falls back to the frozen groundtruth if the warehouse is unreachable.
    """
    CUST = 'sf_data."SALES_DW__PUBLIC__CUSTOMERS"'
    ORD = 'sf_data."SALES_DW__PUBLIC__ORDERS"'
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Per (segment, region): customer count, avg lifetime value, total orders.
        # Total orders = count of ORDERS rows joined by CUSTOMER_ID.
        cur.execute(f'''
            WITH oc AS (
                SELECT "CUSTOMER_ID", COUNT(*) AS order_count
                FROM {ORD}
                GROUP BY "CUSTOMER_ID"
            )
            SELECT
                c."SEGMENT",
                c."REGION",
                COUNT(*)                              AS customers,
                AVG(c."LIFETIME_VALUE")               AS avg_ltv,
                COALESCE(SUM(oc.order_count), 0)      AS total_orders
            FROM {CUST} c
            LEFT JOIN oc ON oc."CUSTOMER_ID" = c."CUSTOMER_ID"
            GROUP BY c."SEGMENT", c."REGION"
        ''')
        rows = cur.fetchall()

        lookup = {}
        weighted_ltv_sum = 0.0
        weighted_cust = 0
        highest_key = None
        highest_avg = None
        for seg, reg, customers, avg_ltv, total_orders in rows:
            key = f"{seg} - {reg}".strip().lower()
            avg_ltv = round(float(avg_ltv), 2) if avg_ltv is not None else None
            lookup[key] = (int(customers), avg_ltv, int(total_orders))
            if avg_ltv is not None:
                weighted_ltv_sum += avg_ltv * int(customers)
                weighted_cust += int(customers)
                if highest_avg is None or avg_ltv > highest_avg:
                    highest_avg = avg_ltv
                    highest_key = f"{seg} - {reg}"
        overall = round(weighted_ltv_sum / weighted_cust, 2) if weighted_cust else None
        cur.close()
        conn.close()
        summary = {
            "total_combinations": len(lookup),
            "highest_ltv_group": highest_key,
            "overall_avg_ltv": overall,
        }
        if not lookup or highest_key is None or overall is None:
            raise RuntimeError("empty warehouse result")
        print("[derive] Expected values derived from warehouse")
        return lookup, summary
    except Exception as e:
        print(f"[WARN] Could not derive expected from warehouse ({e}); using fallback")
        return dict(_FALLBACK_LOYALTY), dict(_FALLBACK_SUMMARY)


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_email(errors):
    """Verify email to marketing@company.com with subject 'Customer Loyalty Analysis'."""
    try:
        conn = psycopg2.connect(**DB); cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, COALESCE(body_text, body_html, '')
            FROM email.messages
        """)
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        errors.append(f"Email check error: {e}"); return
    target_subj = "customer loyalty analysis"
    target_to = "marketing@company.com"
    matched = None
    for subj, to_addr, body in rows:
        subj_l = (subj or "").strip().lower()
        to_str = str(to_addr or "").lower()
        if target_subj in subj_l and target_to in to_str:
            matched = (subj, to_addr, body)
            break
    if not matched:
        errors.append(f"Email '{target_subj}' to {target_to} not found (checked {len(rows)} emails)")
        return
    body_l = (matched[2] or "").lower()
    # Body should highlight highest-value segment-region(s)
    must = ["segment", "region"]
    missing = [m for m in must if m not in body_l]
    if missing:
        errors.append(f"Email body missing terms: {missing}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_loyalty, expected_summary = _derive_expected()

    agent_file = os.path.join(args.agent_workspace, "Customer_Loyalty_Report.xlsx")
    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    all_errors = []

    # Check sheet: Customer Loyalty
    print(f"  Checking Customer Loyalty...")
    a_rows = load_sheet_rows(agent_wb, "Customer Loyalty")
    if a_rows is None:
        all_errors.append("Sheet 'Customer Loyalty' not found in agent output")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []

        if len(a_data) != len(expected_loyalty):
            errors.append(f"Customer Loyalty row count: agent={len(a_data)}, "
                          f"expected={len(expected_loyalty)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for key, (exp_cust, exp_ltv, exp_orders) in expected_loyalty.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {key}")
                continue
            # Customers (exact)
            if not num_close(a_row[1], exp_cust, 0):
                errors.append(f"{key}.Customers: {a_row[1]} vs {exp_cust} (exact)")
            # Avg_LTV (tol=0.5)
            if len(a_row) > 2 and not num_close(a_row[2], exp_ltv, 0.5):
                errors.append(f"{key}.Avg_LTV: {a_row[2]} vs {exp_ltv} (tol=0.5)")
            # Total_Orders (exact)
            if len(a_row) > 3 and not num_close(a_row[3], exp_orders, 0):
                errors.append(f"{key}.Total_Orders: {a_row[3]} vs {exp_orders} (exact)")

        # Sort order: alphabetical by Segment, then Avg_LTV desc within segment.
        # Build the expected key order from derived data.
        def seg_of(label):
            return label.split(" - ")[0].strip().lower()
        expected_order = sorted(
            expected_loyalty.keys(),
            key=lambda k: (seg_of(k), -(expected_loyalty[k][1] or 0))
        )
        a_keys = [str(r[0]).strip().lower() for r in a_data if r and r[0]]
        if len(a_keys) >= 2:
            # Compare within each segment block (segment boundary robust to
            # cross-segment ties at the tail). Group agent keys by segment.
            for seg, exp_block in _group_by_segment(expected_order):
                got_block = [k for k in a_keys if seg_of(k) == seg]
                if got_block != exp_block:
                    errors.append(f"Segment '{seg}' not in expected order "
                                  f"(Avg_LTV desc); expected {exp_block[:3]}..., got {got_block[:3]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        metric_expected = {
            "total_combinations": expected_summary["total_combinations"],
            "highest_ltv_group": expected_summary["highest_ltv_group"],
            "overall_avg_ltv": expected_summary["overall_avg_ltv"],
        }
        for key, exp_val in metric_expected.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {key}")
                continue
            a_val = a_row[1] if len(a_row) > 1 else None
            if is_numeric(exp_val):
                tol = 1.0 if "avg" in key or "ltv" in key else 0
                if not num_close(a_val, exp_val, tol):
                    errors.append(f"{key}.Value: {a_val} vs {exp_val} (tol={tol})")
            else:
                if not str_match(a_val, exp_val):
                    errors.append(f"{key}.Value: '{a_val}' vs '{exp_val}' (string)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Email check
    print(f"  Checking email...")
    email_errors = []
    check_email(email_errors)
    all_errors.extend(email_errors)
    if email_errors:
        for e in email_errors:
            print(f"    {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


def _group_by_segment(ordered_keys):
    """Yield (segment, [keys]) preserving the derived order within each segment."""
    last_seg = None
    block = []
    for k in ordered_keys:
        s = k.split(" - ")[0].strip().lower()
        if s != last_seg:
            if block:
                yield last_seg, block
            last_seg = s
            block = [k]
        else:
            block.append(k)
    if block:
        yield last_seg, block


if __name__ == "__main__":
    main()
