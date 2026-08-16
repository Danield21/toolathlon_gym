"""Evaluation for sf-deal-pipeline-review."""
import argparse
import os
import sys

import openpyxl
import psycopg2

# ──────────────────────────────────────────────────────────────────────────
# EVALUATION GROUND TRUTH SPEC (gcal tz root-fix v3, case-study 2026-08-13)
# gcal.events.start_datetime is TIMESTAMPTZ; bare ev[2].hour silently compares
# wrong in non-UTC PG sessions. Use gcal_helpers.
# ──────────────────────────────────────────────────────────────────────────
# task.md line 7: "9 AM UTC on consecutive business days" → UTC target
from utils.evaluation.gcal_helpers import get_utc_components  # noqa: E402


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Pipeline_Review.xlsx")
    gt_file = os.path.join(gt_dir, "Pipeline_Review.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Check Regional Performance
    print("  Checking Regional Performance...")
    a_rows = load_sheet_rows(agent_wb, "Regional Performance")
    g_rows = load_sheet_rows(gt_wb, "Regional Performance")
    if a_rows is None:
        all_errors.append("Sheet 'Regional Performance' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Regional Performance' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing region: {g_row[0]}")
                continue
            # Target (exact - from mock dashboard)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Target: {a_row[1]} vs {g_row[1]}")
            # Actual (tighter; values around 60k-90k, tol 100)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 100):
                    errors.append(f"{key}.Actual: {a_row[2]} vs {g_row[2]}")
            # Gap (=Actual-Target; same tol)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 100):
                    errors.append(f"{key}.Gap: {a_row[3]} vs {g_row[3]}")
            # Gap_Pct (1-decimal percentage)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Gap_Pct: {a_row[4]} vs {g_row[4]}")
            # On_Track (Yes/No)
            if len(a_row) > 5 and len(g_row) > 5:
                if not str_match(a_row[5], g_row[5]):
                    errors.append(f"{key}.On_Track: {a_row[5]} vs {g_row[5]}")
        # Verify alphabetical sort
        agent_regions = [str(r[0]).strip() for r in a_data if r and r[0] is not None]
        if agent_regions != sorted(agent_regions):
            errors.append(f"Regional Performance not sorted alphabetically: {agent_regions}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Segment Breakdown
    print("  Checking Segment Breakdown...")
    a_rows = load_sheet_rows(agent_wb, "Segment Breakdown")
    g_rows = load_sheet_rows(gt_wb, "Segment Breakdown")
    if a_rows is None:
        all_errors.append("Sheet 'Segment Breakdown' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Segment Breakdown' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for r in a_data:
            if r and r[0] is not None and r[1] is not None:
                k = f"{str(r[0]).strip().lower()}|{str(r[1]).strip().lower()}"
                a_lookup[k] = r
        errors = []
        # Row count check
        if len(a_data) != len(g_data):
            errors.append(f"Segment Breakdown row count: {len(a_data)} vs {len(g_data)}")
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}|{g_row[1]}")
                continue
            # Target
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1):
                    errors.append(f"{key}.Target: {a_row[2]} vs {g_row[2]}")
            # Actual
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 100):
                    errors.append(f"{key}.Actual: {a_row[3]} vs {g_row[3]}")
            # Gap_Pct (tighter)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.5):
                    errors.append(f"{key}.Gap_Pct: {a_row[5]} vs {g_row[5]}")
            # On_Track
            if len(a_row) > 6 and len(g_row) > 6:
                if not str_match(a_row[6], g_row[6]):
                    errors.append(f"{key}.On_Track: {a_row[6]} vs {g_row[6]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Summary
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if key in ("worst_region", "best_region"):
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif key in ("regions_on_track", "regions_behind"):
                    if not num_close(a_row[1], g_row[1], 0):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif key.endswith("_pct"):
                    # percentage: tol 0.5
                    if not num_close(a_row[1], g_row[1], 0.5):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                else:
                    # money: tol 200 (vs base ~480k)
                    if not num_close(a_row[1], g_row[1], 200):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Check Google Calendar events
    print("  Checking Google Calendar events...")
    try:
        db_config = {
            "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
            "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
            "user": "eigent", "password": "camel",
        }
        conn = psycopg2.connect(**db_config)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE summary ILIKE '%Pipeline%' AND summary ILIKE '%Review%'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()
        cur.close()
        conn.close()

        # Determine expected event count by reading GT (regions with Gap_Pct < -15)
        gt_data = load_sheet_rows(gt_wb, "Regional Performance") or []
        behind_regions = []
        for r in (gt_data[1:] if len(gt_data) > 1 else []):
            if r and len(r) > 4 and r[4] is not None:
                try:
                    if float(r[4]) < -15:
                        behind_regions.append(str(r[0]))
                except (TypeError, ValueError):
                    pass
        expected_n = len(behind_regions)
        if len(events) != expected_n:
            all_errors.append(f"Expected {expected_n} review meetings, found {len(events)}")
            print(f"    FAIL: expected {expected_n} meetings, found {len(events)}")
        else:
            # Region names appear in summary
            for region in behind_regions:
                if not any(region.lower() in (e[0] or "").lower() for e in events):
                    all_errors.append(f"GCal: missing meeting for region '{region}'")
            # Each event 1-hour — require ALL events to conform
            ok_1h = 0
            for ev in events:
                if ev[2] and ev[3]:
                    delta = (ev[3] - ev[2]).total_seconds() / 3600
                    if 0.9 <= delta <= 1.1:
                        ok_1h += 1
            if events and ok_1h != len(events):
                all_errors.append(f"GCal: {ok_1h}/{len(events)} events are exactly 1 hour")
            # Start at 9 AM UTC — check ALL events, not just first violation
            # gcal.events.start_datetime is TIMESTAMPTZ; use helper for UTC hour.
            non_9am = []
            for ev in events:
                if ev[2]:
                    _utc_date, ev_h, _utc_m = get_utc_components(ev[2])
                    if ev_h != 9:
                        non_9am.append((ev[0], ev_h))
            if non_9am:
                all_errors.append(f"GCal: {len(non_9am)} events not at 9AM UTC: {non_9am[:3]}")
            # First event on 2026-03-10 (Tuesday) or after consecutive business days
            if events:
                first_dt = events[0][2]
                if first_dt:
                    first_date, _h, _m = get_utc_components(first_dt)
                    if first_date is None or str(first_date) != "2026-03-10":
                        all_errors.append(f"GCal: first event not on 2026-03-10 (got {first_date})")
            # Description mentions gap percentage
            desc_ok = 0
            for ev in events:
                d = str(ev[1] or "").lower()
                if "%" in d or "gap" in d:
                    desc_ok += 1
            if desc_ok < len(events) - 1:
                all_errors.append(f"GCal: {desc_ok}/{len(events)} descriptions mention gap%")
            if not [e for e in all_errors if e.startswith("GCal:")]:
                print("    PASS")
    except Exception as e:
        all_errors.append(f"GCal check error: {e}")
        print(f"    ERROR: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
