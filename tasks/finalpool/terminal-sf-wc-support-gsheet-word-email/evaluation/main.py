"""Evaluation for terminal-sf-wc-support-gsheet-word-email.

Checks:
1. Support_Quality_Audit.xlsx with 3 sheets
2. Support_Audit_Report.docx
3. Google Sheet "Support Quality Audit"
4. Email sent to cs_leadership@company.com
"""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_dynamic_ticket_data():
    """Dynamically query ticket priority data from the DB."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT "PRIORITY", COUNT(*) as cnt,
                   ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp,
                   ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) as avg_sat
            FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
            GROUP BY "PRIORITY" ORDER BY "PRIORITY"
        """)
        rows = cur.fetchall()
        cur.execute('SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
        total = cur.fetchone()[0]
        cur.execute("""
            SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
            WHERE "PRIORITY" = 'High'
        """)
        high_count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return {r[0].strip().lower(): {"count": r[1], "avg_resp": float(r[2]), "avg_sat": float(r[3])} for r in rows}, total, high_count
    except Exception:
        return None, None, None


def get_dynamic_order_data():
    """Dynamically query WC order data from the DB.

    The wc schema stores order line items as JSONB on wc.orders.line_items,
    and product categories as JSONB on wc.products.categories. Matches the
    semantics used to build the GT (count-by-line-item, not by distinct
    order-category pair).
    """
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM wc.orders")
        total_orders = cur.fetchone()[0]
        cur.execute("""
            WITH order_categories AS (
                SELECT o.id as order_id, o.status,
                       cat->>'name' as category
                FROM wc.orders o,
                     jsonb_array_elements(o.line_items) li,
                     wc.products p,
                     jsonb_array_elements(p.categories) cat
                WHERE (li->>'product_id')::int = p.id
            )
            SELECT category, COUNT(*) as total,
                   SUM(CASE WHEN status IN ('refunded','cancelled','failed')
                            THEN 1 ELSE 0 END) as problems
            FROM order_categories
            GROUP BY category
            ORDER BY category
        """)
        cat_rows = cur.fetchall()
        cur.close()
        conn.close()
        return total_orders, {r[0].strip().lower(): {"total": r[1], "problems": r[2]} for r in cat_rows}
    except Exception:
        return None, None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Support_Quality_Audit.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Support_Quality_Audit.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Support_Quality_Audit.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Get dynamic expected data
    ticket_data, total_tickets, high_count = get_dynamic_ticket_data()
    total_orders, cat_data = get_dynamic_order_data()

    # Fallback: use groundtruth file if DB queries fail
    gt_wb = None
    try:
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception:
        pass

    # Sheet 1: Ticket_Priority_Summary
    print("  Checking Ticket_Priority_Summary...")
    a_sheet = get_sheet(agent_wb, "Ticket_Priority_Summary")
    check("Sheet 'Ticket_Priority_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        check("Ticket_Priority_Summary has 3 rows", len(a_rows) == 3, f"Got {len(a_rows)}")

        a_lookup = {str(r[0]).strip().lower(): r for r in a_rows if r and r[0]}

        if ticket_data:
            # Use dynamic DB data
            for priority, vals in ticket_data.items():
                a_row = a_lookup.get(priority)
                if a_row is None:
                    check(f"Priority '{priority}' present", False, "Missing")
                    continue
                if len(a_row) > 1:
                    check(f"'{priority}' Ticket_Count",
                          num_close(a_row[1], vals["count"], 1),
                          f"Expected {vals['count']}, got {a_row[1]}")
                if len(a_row) > 2:
                    check(f"'{priority}' Avg_Response_Hours",
                          num_close(a_row[2], vals["avg_resp"], 0.15),
                          f"Expected {vals['avg_resp']}, got {a_row[2]}")
                if len(a_row) > 3:
                    check(f"'{priority}' Avg_Satisfaction",
                          num_close(a_row[3], vals["avg_sat"], 0.1),
                          f"Expected {vals['avg_sat']}, got {a_row[3]}")
        elif gt_wb:
            # Fallback to groundtruth
            g_sheet = get_sheet(gt_wb, "Ticket_Priority_Summary")
            if g_sheet:
                g_rows = list(g_sheet.iter_rows(min_row=2, values_only=True))
                for g_row in g_rows:
                    if not g_row or not g_row[0]:
                        continue
                    key = str(g_row[0]).strip().lower()
                    a_row = a_lookup.get(key)
                    if a_row is None:
                        check(f"Priority '{g_row[0]}' present", False, "Missing")
                        continue
                    if len(a_row) > 1 and len(g_row) > 1:
                        check(f"'{key}' Ticket_Count",
                              num_close(a_row[1], g_row[1], 1),
                              f"Expected {g_row[1]}, got {a_row[1]}")

    # Sheet 2: Product_Problem_Rates
    print("  Checking Product_Problem_Rates...")
    a_sheet = get_sheet(agent_wb, "Product_Problem_Rates")
    check("Sheet 'Product_Problem_Rates' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        a_lookup = {}
        for r in a_rows:
            if r and r[0]:
                a_lookup[str(r[0]).strip().lower()] = r

        if cat_data:
            expected_cats = len(cat_data)
            check(f"Product_Problem_Rates has {expected_cats} rows",
                  len(a_rows) == expected_cats, f"Got {len(a_rows)}")
            for cat_name, vals in cat_data.items():
                a_row = a_lookup.get(cat_name)
                if a_row is None:
                    for ak, av in a_lookup.items():
                        if cat_name.split()[0] in ak:
                            a_row = av
                            break
                if a_row is None:
                    check(f"Category '{cat_name}' present", False,
                          f"Missing from {list(a_lookup.keys())}")
                    continue
                if len(a_row) > 1:
                    check(f"'{cat_name}' Total_Orders",
                          num_close(a_row[1], vals["total"], 1),
                          f"Expected {vals['total']}, got {a_row[1]}")
                if len(a_row) > 2:
                    check(f"'{cat_name}' Problem_Orders",
                          num_close(a_row[2], vals["problems"], 1),
                          f"Expected {vals['problems']}, got {a_row[2]}")
                if len(a_row) > 3 and vals["total"] > 0:
                    expected_rate = round(vals["problems"] / vals["total"] * 100, 1)
                    check(f"'{cat_name}' Problem_Rate_Pct",
                          num_close(a_row[3], expected_rate, 2.0),
                          f"Expected {expected_rate}, got {a_row[3]}")
        else:
            check("Product_Problem_Rates has 8 rows", len(a_rows) == 8, f"Got {len(a_rows)}")

    # Sheet 3: Audit_Summary
    print("  Checking Audit_Summary...")
    a_sheet = get_sheet(agent_wb, "Audit_Summary")
    check("Sheet 'Audit_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_data = {}
        for row in a_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_data[str(row[0]).strip().lower()] = row[1]

        if total_tickets is not None and total_orders is not None:
            check("Total_Tickets",
                  num_close(a_data.get("total_tickets"), total_tickets, 1),
                  f"Expected {total_tickets}, got {a_data.get('total_tickets')}")
            check("Total_WC_Orders",
                  num_close(a_data.get("total_wc_orders"), total_orders, 1),
                  f"Expected {total_orders}, got {a_data.get('total_wc_orders')}")
            if total_orders > 0 and cat_data:
                total_problems = sum(v["problems"] for v in cat_data.values())
                expected_rate = round(total_problems / sum(v["total"] for v in cat_data.values()) * 100, 1)
                check("Overall_Problem_Rate_Pct",
                      num_close(a_data.get("overall_problem_rate_pct"), expected_rate, 3.0),
                      f"Expected {expected_rate}, got {a_data.get('overall_problem_rate_pct')}")
            if high_count is not None and total_tickets > 0:
                expected_hp = round(high_count / total_tickets * 100, 1)
                check("High_Priority_Pct",
                      num_close(a_data.get("high_priority_pct"), expected_hp, 2.0),
                      f"Expected {expected_hp}, got {a_data.get('high_priority_pct')}")
            # Validate Highest_Problem_Category (string match)
            if cat_data:
                try:
                    worst_cat = max(
                        cat_data.items(),
                        key=lambda kv: (kv[1]["problems"] / max(kv[1]["total"], 1))
                    )[0]
                except Exception:
                    worst_cat = None
                if worst_cat is not None:
                    actual_cat = a_data.get("highest_problem_category")
                    actual_str = str(actual_cat).strip().lower() if actual_cat is not None else ""
                    check("Highest_Problem_Category matches",
                          actual_str == worst_cat.strip().lower(),
                          f"Expected '{worst_cat}', got '{actual_cat}'")
        elif gt_wb:
            g_sheet = get_sheet(gt_wb, "Audit_Summary")
            if g_sheet:
                g_data = {}
                for row in g_sheet.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        g_data[str(row[0]).strip().lower()] = row[1]
                check("Total_Tickets",
                      num_close(a_data.get("total_tickets"), g_data.get("total_tickets"), 100),
                      f"Expected {g_data.get('total_tickets')}, got {a_data.get('total_tickets')}")
                check("Total_WC_Orders",
                      num_close(a_data.get("total_wc_orders"), g_data.get("total_wc_orders"), 5),
                      f"Expected {g_data.get('total_wc_orders')}, got {a_data.get('total_wc_orders')}")


def check_word(agent_workspace):
    print("\n=== Checking Support_Audit_Report.docx ===")
    docx_path = os.path.join(agent_workspace, "Support_Audit_Report.docx")
    check("Support_Audit_Report.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        # Require substantive content matching task spec sections
        check("Document has substantial content", len(text) >= 400, f"Length: {len(text)}")
        check("Contains executive summary",
              "executive" in text and "summary" in text,
              "Missing executive summary section")
        check("Contains support/ticket reference",
              "support" in text and "ticket" in text)
        check("Contains problem/complaint reference",
              "problem" in text or "complaint" in text or "refund" in text)
        check("Contains recommendation",
              "recommend" in text or "suggestion" in text or "action" in text)

        # Dynamically derive the highest-problem-rate category from DB
        _, cat_data = get_dynamic_order_data()
        if cat_data:
            # category with highest problem rate (problems/total)
            try:
                worst_cat = max(
                    cat_data.items(),
                    key=lambda kv: (kv[1]["problems"] / max(kv[1]["total"], 1))
                )[0]
                # Accept singular or plural variants (e.g. 'camera'/'cameras')
                worst_lc = worst_cat.lower()
                singular = worst_lc.rstrip("s")
                check(f"Contains highest-problem category '{worst_cat}'",
                      worst_lc in text or singular in text,
                      f"Missing highest-problem category '{worst_cat}'")
            except Exception as e:
                check("Contains highest-problem category", False, str(e))
        else:
            # DB unavailable - skip rather than guess
            check("DB-derived highest-problem category check",
                  False, "DB unavailable; cannot validate highest-problem category")
    except ImportError:
        check("python-docx available", False, "Cannot verify Word content")
    except Exception as e:
        check("Word document readable", False, str(e))


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gsheet.spreadsheets
            WHERE lower(title) LIKE '%%support%%audit%%'
        """)
        rows = cur.fetchall()
        check("Google Sheet 'Support Quality Audit' exists", len(rows) > 0,
              f"Found {len(rows)} matching sheets")

        if rows:
            ss_id = rows[0][0]
            cur.execute("""
                SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s
            """, (ss_id,))
            sheets_meta = cur.fetchall()
            sheet_names = [s[1].lower() for s in sheets_meta]
            check("GSheet has Ticket_Priority_Summary",
                  any("ticket" in s and "priority" in s for s in sheet_names),
                  f"Sheets: {sheet_names}")
            check("GSheet has Product_Problem_Rates",
                  any("product" in s and "problem" in s for s in sheet_names),
                  f"Sheets: {sheet_names}")
            check("GSheet has Audit_Summary",
                  any("audit" in s and "summary" in s for s in sheet_names),
                  f"Sheets: {sheet_names}")
            # Validate cells: each named sheet must contain at least 2 data rows
            for sheet_id, sheet_title in sheets_meta:
                cur.execute(
                    "SELECT COUNT(DISTINCT row_index) FROM gsheet.cells "
                    "WHERE spreadsheet_id = %s AND sheet_id = %s "
                    "AND row_index > 0 AND value IS NOT NULL AND value != ''",
                    (ss_id, sheet_id),
                )
                row_count = cur.fetchone()[0] or 0
                check(f"GSheet '{sheet_title}' has data rows (>=2)",
                      row_count >= 2, f"got {row_count}")
            # Validate Audit_Summary content: at least Total_Tickets and
            # Total_WC_Orders metric names + matching numeric values, so
            # GSheet content is not just empty placeholder rows.
            ticket_data, total_tickets, _ = get_dynamic_ticket_data()
            total_orders, _ = get_dynamic_order_data()
            audit_sheet_id = None
            for sid, stitle in sheets_meta:
                if "audit" in stitle.lower() and "summary" in stitle.lower():
                    audit_sheet_id = sid
                    break
            if audit_sheet_id is not None:
                cur.execute(
                    "SELECT row_index, col_index, value FROM gsheet.cells "
                    "WHERE spreadsheet_id = %s AND sheet_id = %s "
                    "AND value IS NOT NULL AND value != ''",
                    (ss_id, audit_sheet_id),
                )
                cells = cur.fetchall()
                # Build {row: {col: value}}
                grid = {}
                for ri, ci, val in cells:
                    grid.setdefault(ri, {})[ci] = val
                # Find rows with metric labels
                gs_metrics = {}
                for ri, cols in grid.items():
                    label = str(cols.get(0, "") or "").strip().lower()
                    val_str = str(cols.get(1, "") or "").strip()
                    if label and val_str:
                        gs_metrics[label] = val_str
                check("GSheet Audit_Summary has Total_Tickets metric",
                      "total_tickets" in gs_metrics,
                      f"keys: {list(gs_metrics.keys())[:8]}")
                check("GSheet Audit_Summary has Total_WC_Orders metric",
                      "total_wc_orders" in gs_metrics or "total_orders" in gs_metrics,
                      f"keys: {list(gs_metrics.keys())[:8]}")
                if total_tickets is not None and "total_tickets" in gs_metrics:
                    try:
                        v = float(gs_metrics["total_tickets"])
                        check("GSheet Audit_Summary Total_Tickets matches DB",
                              num_close(v, total_tickets, 1),
                              f"GSheet={v}, DB={total_tickets}")
                    except ValueError:
                        check("GSheet Audit_Summary Total_Tickets numeric",
                              False, gs_metrics.get("total_tickets"))
                if total_orders is not None:
                    key = "total_wc_orders" if "total_wc_orders" in gs_metrics else "total_orders"
                    if key in gs_metrics:
                        try:
                            v = float(gs_metrics[key])
                            check("GSheet Audit_Summary Total_WC_Orders matches DB",
                                  num_close(v, total_orders, 1),
                                  f"GSheet={v}, DB={total_orders}")
                        except ValueError:
                            check("GSheet Audit_Summary Total_WC_Orders numeric",
                                  False, gs_metrics.get(key))

        cur.close()
        conn.close()
    except Exception as e:
        check("Google Sheet check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Check sent_log joined with messages
        cur.execute("""
            SELECT m.subject, m.to_addr, m.body_text
            FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            WHERE lower(m.subject) LIKE '%%support%%quality%%audit%%'
               OR lower(m.subject) LIKE '%%q1%%support%%'
        """)
        rows = cur.fetchall()
        if not rows:
            # Also check messages directly for sent emails
            cur.execute("""
                SELECT subject, to_addr, body_text FROM email.messages
                WHERE lower(subject) LIKE '%%support%%quality%%audit%%'
                   OR lower(subject) LIKE '%%q1%%support%%'
            """)
            rows = cur.fetchall()
        check("Audit email sent", len(rows) > 0, f"Found {len(rows)} matching emails")
        if rows:
            subj, to_addr, body = rows[0]
            to_str = str(to_addr).lower() if to_addr else ""
            check("Email to cs_leadership",
                  "cs_leadership" in to_str,
                  f"To: {to_addr}")
            # Body content checks
            body_lower = (body or "").lower()
            check("Email body has substantive content (>=100 chars)",
                  len(body_lower) >= 100,
                  f"Body length: {len(body_lower)}")
            check("Email body mentions ticket or support",
                  "ticket" in body_lower or "support" in body_lower,
                  "Expected ticket/support reference")
            check("Email body mentions priority or category",
                  "priorit" in body_lower or "categor" in body_lower or "product" in body_lower,
                  "Expected priority/category reference")
            # Email body should reference at least one quantitative finding
            # (e.g. a percentage, ticket count, problem rate). This catches
            # fluff bodies that lack any factual content.
            import re as _re
            has_number = _re.search(r"\d+", body_lower) is not None
            check("Email body contains at least one numeric finding",
                  has_number,
                  "Body lacks any digits (no count/percentage)")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    agent_file = os.path.join(workspace, "Support_Quality_Audit.xlsx")
    if os.path.exists(agent_file):
        wb = openpyxl.load_workbook(agent_file, data_only=True)
        # No unexpected sheets
        expected_keywords = {"ticket", "priority", "product", "problem", "audit", "summary"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        # Problem_Rate_Pct should not exceed 100
        ppr = get_sheet(wb, "Product_Problem_Rates")
        if ppr:
            for row in ppr.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 3 and row[3] is not None:
                    try:
                        rate = float(row[3])
                        if rate > 100:
                            check("No problem rate > 100%", False, f"Found {rate} for {row[0]}")
                            break
                    except (ValueError, TypeError):
                        pass
            else:
                check("No problem rate > 100%", True)

    # Email: no emails to wrong recipients
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE lower(subject) LIKE '%%support%%quality%%audit%%'
              AND to_addr::text NOT ILIKE '%%cs_leadership%%'
        """)
        wrong_emails = cur.fetchone()[0]
        check("No audit emails to wrong recipients", wrong_emails == 0,
              f"Found {wrong_emails} misrouted emails")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gsheet()
    check_email()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
