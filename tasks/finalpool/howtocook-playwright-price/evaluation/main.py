"""
Evaluation script for howtocook-playwright-price task.

Checks:
1. Excel file (Recipe_Cost_Analysis.xlsx) exists with 2 sheets
2. Dish Costs sheet has 5 data rows
3. Ingredient Prices sheet has data
4. Google Sheet exists with recipe/cost in title
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
_AGENT_DISH_NAMES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    """Check the Excel output file."""
    print("\n=== Check 1: Excel File ===")

    excel_path = os.path.join(agent_workspace, "Recipe_Cost_Analysis.xlsx")
    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return

    record("Excel file readable", True)

    # Check that there are at least 2 sheets
    record("Excel has >= 2 sheets", len(wb.sheetnames) >= 2,
           f"Found sheets: {wb.sheetnames}")

    # Find Dish Costs sheet
    dc_sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "dish" in name_lower or ("cost" in name_lower and "ingredient" not in name_lower):
            dc_sheet = name
            break
    if not dc_sheet:
        # fallback: any sheet with "cost" in name
        for name in wb.sheetnames:
            if "cost" in name.lower():
                dc_sheet = name
                break

    agent_dish_names_lower = set()
    if not dc_sheet:
        record("Dish Costs sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Dish Costs sheet exists", True)
        ws = wb[dc_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)] if len(rows) > 1 else []
        record("Dish Costs has 5 data rows", len(data_rows) >= 5,
               f"Found {len(data_rows)} data rows")

        # NOTE: Task.md allows agent to "select 5 dishes from different categories",
        # so we do NOT hardcode specific dish names. Validate distinct dish names.
        dish_names = set()
        for r in data_rows:
            if r and r[0] is not None:
                dish_names.add(str(r[0]).strip())
                agent_dish_names_lower.add(str(r[0]).strip().lower())
        record("Dish Costs has >=5 distinct dish names",
               len(dish_names) >= 5,
               f"Found {len(dish_names)} distinct names")

        # Determine column indices via headers
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            total_idx = next((i for i, h in enumerate(headers) if "total" in h or ("estimated" in h and "cost" in h and "serving" not in h)), None)
            cps_idx = next((i for i, h in enumerate(headers) if "per_serving" in h or "per serving" in h), None)

            # Verify Cost_Per_Serving numeric in >=4 rows
            if cps_idx is not None:
                cps_count = 0
                for r in data_rows:
                    if r and len(r) > cps_idx:
                        try:
                            float(r[cps_idx])
                            cps_count += 1
                        except (TypeError, ValueError):
                            pass
                record("Dish Costs has numeric Cost_Per_Serving in >=4 rows",
                       cps_count >= 4, f"numeric count={cps_count}")

            # Self-consistency: Cost_Per_Serving == round(Total/4, 2) within tol
            if total_idx is not None and cps_idx is not None:
                consistent_rows = 0
                for r in data_rows:
                    if r and len(r) > max(total_idx, cps_idx):
                        try:
                            tot = float(r[total_idx])
                            cps = float(r[cps_idx])
                            if abs(round(tot / 4, 2) - cps) <= 0.05:
                                consistent_rows += 1
                        except (TypeError, ValueError):
                            pass
                record("Dish Costs self-consistency Cost_Per_Serving == Total/4 in >=4 rows",
                       consistent_rows >= 4,
                       f"consistent_rows={consistent_rows}")

    # Find Ingredient Prices sheet
    ip_sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "ingredient" in name_lower or "price" in name_lower:
            ip_sheet = name
            break

    if not ip_sheet:
        record("Ingredient Prices sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Ingredient Prices sheet exists", True)
        ws = wb[ip_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)] if len(rows) > 1 else []
        record("Ingredient Prices has >= 10 rows", len(data_rows) >= 10,
               f"Found {len(data_rows)} data rows")

    # --- Budget Summary self-consistency ---
    bs_sheet = None
    for name in wb.sheetnames:
        if "budget" in name.lower() or "summary" in name.lower():
            bs_sheet = name
            break
    if bs_sheet:
        ws = wb[bs_sheet]
        bs_rows = list(ws.iter_rows(values_only=True))
        bs_data = [r for r in bs_rows[1:] if r and any(c is not None for c in r)]
        bs_lookup = {}
        total_cost_val = None
        for r in bs_data:
            if r and r[0]:
                key = str(r[0]).strip().lower()
                bs_lookup[key] = r[1] if len(r) > 1 else None
        # Required metrics
        for req in ["total_cost", "average_cost_per_dish", "cheapest_dish", "most_expensive_dish"]:
            present = req in bs_lookup or req.replace("_", " ") in bs_lookup
            # also accept variants without underscore
            if not present:
                for k in bs_lookup:
                    if k.replace("_", " ").replace("  ", " ") == req.replace("_", " "):
                        present = True
                        break
            record(f"Budget Summary contains '{req}'", present,
                   f"keys: {list(bs_lookup)}")
        # Self-consistency: Average_Cost_Per_Dish == Total_Cost / 5 (within 0.05)
        try:
            tot = float(bs_lookup.get("total_cost"))
            avg = float(bs_lookup.get("average_cost_per_dish"))
            if abs(round(tot / 5, 2) - avg) <= 0.05:
                record("Budget Summary Average_Cost_Per_Dish == Total_Cost/5", True)
            else:
                record("Budget Summary Average_Cost_Per_Dish == Total_Cost/5", False,
                       f"total={tot}, avg={avg}, expected={round(tot/5, 2)}")
        except (TypeError, ValueError):
            record("Budget Summary Average_Cost_Per_Dish numeric and consistent",
                   False, "Could not parse Total_Cost or Average_Cost_Per_Dish")
        # Cheapest != Most expensive
        cheap = str(bs_lookup.get("cheapest_dish") or "").strip()
        expensive = str(bs_lookup.get("most_expensive_dish") or "").strip()
        if cheap and expensive:
            record("Cheapest_Dish != Most_Expensive_Dish",
                   cheap.lower() != expensive.lower(),
                   f"cheap={cheap}, expensive={expensive}")
            # Both should be in the agent's Dish Costs dish names
            if agent_dish_names_lower:
                record("Cheapest_Dish appears in Dish Costs",
                       cheap.lower() in agent_dish_names_lower,
                       f"cheap={cheap}")
                record("Most_Expensive_Dish appears in Dish Costs",
                       expensive.lower() in agent_dish_names_lower,
                       f"expensive={expensive}")

    # GT existence-only check (free dish choice, so per-row comparison disabled)
    gt_path = os.path.join(groundtruth_workspace, "Recipe_Cost_Analysis.xlsx")
    if os.path.isfile(gt_path):
        record("GT file present (existence only - free dish choice)", True)

    # Stash agent dish names for GSheet check
    global _AGENT_DISH_NAMES
    _AGENT_DISH_NAMES = list(agent_dish_names_lower)
    wb.close()


def check_gsheet():
    """Check that a Google Sheet was created with recipe/cost data."""
    print("\n=== Check 2: Google Sheet ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()

        if not spreadsheets:
            record("Google Sheet exists", False, "No spreadsheets found")
            cur.close()
            conn.close()
            return

        record("Google Sheet exists", True)

        # Find sheet with title 'Cafeteria Menu Cost Analysis' (per task.md)
        target_ss = None
        for ss_id, ss_title in spreadsheets:
            title_lower = (ss_title or "").lower()
            # Tightened: require ('cafeteria' AND 'cost') OR ('menu' AND 'cost' AND 'analysis')
            if ("cafeteria" in title_lower and "cost" in title_lower) \
                    or ("menu" in title_lower and "cost" in title_lower
                        and "analysis" in title_lower):
                target_ss = (ss_id, ss_title)
                break

        record("'Cafeteria Menu Cost Analysis' Google Sheet exists",
               target_ss is not None,
               f"All titles: {[s[1] for s in spreadsheets]}")
        if target_ss is None:
            cur.close()
            conn.close()
            return

        # Check cells exist
        cur.execute("""
            SELECT c.row_index, c.col_index, c.value
            FROM gsheet.cells c
            JOIN gsheet.sheets s ON c.sheet_id = s.id
            WHERE s.spreadsheet_id = %s
            ORDER BY c.row_index, c.col_index
        """, (target_ss[0],))
        cells = cur.fetchall()

        record("Google Sheet has data", len(cells) >= 5,
               f"Found {len(cells)} cells")

        # Tightened: GSheet should contain at least 4 of the agent's 5 dish names
        cell_text = " | ".join(str(v or "").strip().lower() for _, _, v in cells)
        if _AGENT_DISH_NAMES:
            mentioned = [d for d in _AGENT_DISH_NAMES if d and d in cell_text]
            record("GSheet contains >=4 of agent's 5 dish names",
                   len(mentioned) >= 4,
                   f"mentioned={len(mentioned)} of {len(_AGENT_DISH_NAMES)}: {mentioned[:5]}")

        cur.close()
        conn.close()

    except Exception as e:
        record("Google Sheet check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gsheet()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
