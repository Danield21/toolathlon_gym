"""Evaluation for canvas-assessment-quality-audit."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(v):
    """Robustly coerce a cell value to float (or None if not parseable).

    Handles int/float directly; strips thousands separators, currency
    symbols, '%' and spaces from strings before parsing."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "").replace("$", "").replace("¥", "").replace("€", "")\
        .replace("%", "").replace(" ", "")
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=2.0):
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= tol


_FORMULA_UNRESOLVED = object()


def _effective_value(wb_val, wb_formula, sheet, r, c):
    """Return a cell's effective value.

    Literal cells -> their literal value. Formula cells -> the cached value
    (data_only workbook); if the cache is None the cell cannot be resolved
    without a spreadsheet engine, so return the _FORMULA_UNRESOLVED sentinel.
    """
    if isinstance(wb_formula[sheet][r + 1][c].value, str) and \
            wb_formula[sheet][r + 1][c].value.startswith("="):
        cached = wb_val[sheet][r + 1][c].value
        return _FORMULA_UNRESOLVED if cached is None else cached
    return wb_val[sheet][r + 1][c].value


def load_sheet_rows(wb_val, wb_formula, sheet_name):
    for name in wb_val.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            rows = []
            for r, row in enumerate(wb_val[name].iter_rows()):
                rows.append([_effective_value(wb_val, wb_formula, name, r, c)
                             for c in range(len(row))])
            return rows
    return None


def get_expected_quiz_data():
    """Get expected quiz data from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name, q.title,
               (SELECT COUNT(*) FROM canvas.quiz_questions qq WHERE qq.quiz_id = q.id) as qcount,
               ROUND(AVG(qs.score), 1) as avg_score,
               q.points_possible
        FROM canvas.quizzes q
        JOIN canvas.courses c ON c.id = q.course_id
        LEFT JOIN canvas.quiz_submissions qs ON qs.quiz_id = q.id
        GROUP BY c.name, q.id, q.title, q.points_possible
        ORDER BY c.name, q.title
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    quiz_data = []
    for course, quiz, qcount, avg_score, pts in rows:
        # Skip quizzes with no submissions (avg_score = None) - difficulty undefined
        if avg_score is None or qcount == 0 or pts is None or float(pts) <= 0:
            continue
        difficulty = round(float(avg_score) / (float(pts) * qcount), 3)
        flagged = difficulty < 0.3 or difficulty > 0.8
        quiz_data.append({
            "course": course, "quiz": quiz, "qcount": qcount,
            "avg_score": float(avg_score),
            "difficulty": difficulty, "flagged": flagged,
        })
    return quiz_data


def _find_col(header, target):
    """Find column index in header (case-insensitive, underscore/space tolerant).
    Uses exact-match only (after normalization) to avoid substring collisions
    e.g. 'avg_score' vs 'org_score'."""
    target_norm = str(target).lower().replace(" ", "_").replace("-", "_")
    for i, h in enumerate(header):
        h_norm = str(h or "").lower().replace(" ", "_").replace("-", "_")
        if h_norm == target_norm:
            return i
    return -1


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Assessment_Quality.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Assessment_Quality.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Assessment_Quality.xlsx exists", True)

    try:
        wb_val = openpyxl.load_workbook(xlsx_path, data_only=True)
        wb_formula = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_quiz_data()
    total_quizzes = len(expected)
    flagged_count = sum(1 for q in expected if q["flagged"])
    total_questions = sum(q["qcount"] for q in expected)
    expected_by_key = {(q["course"], q["quiz"]): q for q in expected}
    flag_rate_pct = round(100.0 * flagged_count / total_quizzes, 1) if total_quizzes else 0.0

    # Quiz Overview sheet
    qo_rows = load_sheet_rows(wb_val, wb_formula, "Quiz Overview")
    if qo_rows is None:
        check("Sheet 'Quiz Overview' exists", False, f"Available: {wb_val.sheetnames}")
    else:
        check("Sheet 'Quiz Overview' exists", True)
        data_rows = [r for r in (qo_rows[1:] if len(qo_rows) > 1 else []) if r and r[0]]
        check(f"Quiz Overview row count == {total_quizzes}",
              len(data_rows) == total_quizzes,
              f"Found {len(data_rows)}, expected {total_quizzes}")

        header = qo_rows[0] if qo_rows else []
        # Exact normalized match (consistent with _find_col) to avoid
        # substring collisions such as 'avg_score' vs 'org_score'.
        for col in ["course", "quiz", "question_count", "avg_score", "avg_difficulty"]:
            check(f"Column '{col}' present", _find_col(header, col) >= 0,
                  f"Header: {header}")

        # Per-row value validation
        idx_course = _find_col(header, "course")
        idx_quiz = _find_col(header, "quiz")
        idx_qcount = _find_col(header, "question_count")
        idx_avg = _find_col(header, "avg_score")
        idx_diff = _find_col(header, "avg_difficulty")

        if all(i >= 0 for i in [idx_course, idx_quiz, idx_qcount, idx_avg, idx_diff]):
            row_match = 0
            qcount_resolvable = 0
            qcount_ok = 0
            avg_resolvable = 0
            avg_ok = 0
            diff_resolvable = 0
            diff_ok = 0
            for r in data_rows:
                key = (str(r[idx_course] or "").strip(), str(r[idx_quiz] or "").strip())
                if key in expected_by_key:
                    row_match += 1
                    exp = expected_by_key[key]
                    qc = r[idx_qcount]
                    if qc is not _FORMULA_UNRESOLVED:
                        qcount_resolvable += 1
                        # Parse via _to_float so integer-like representations
                        # (18, 18.0, "18", "18.0") all count as correct, while a
                        # genuinely wrong count (17, 17.5, "N/A") still fails.
                        qc_parsed = _to_float(qc)
                        if qc_parsed is not None and abs(qc_parsed - exp["qcount"]) <= 0.001:
                            qcount_ok += 1
                    av = r[idx_avg]
                    if av is not _FORMULA_UNRESOLVED:
                        avg_resolvable += 1
                        if num_close(av, exp["avg_score"], tol=0.5):
                            avg_ok += 1
                    df = r[idx_diff]
                    if df is not _FORMULA_UNRESOLVED:
                        diff_resolvable += 1
                        if num_close(df, exp["difficulty"], tol=0.05):
                            diff_ok += 1
            check("Quiz Overview rows match expected (course, quiz)",
                  row_match == total_quizzes,
                  f"Matched {row_match}/{total_quizzes}")
            # A value check passes when every resolvable row is correct; rows
            # whose value is an unresolved formula are skipped (task requires
            # literal values, so a fully literal workbook validates everything).
            check("Quiz Overview Question_Count values correct",
                  qcount_resolvable == 0 or qcount_ok == qcount_resolvable,
                  f"{qcount_ok}/{qcount_resolvable} correct")
            check("Quiz Overview Avg_Score values correct (all rows)",
                  avg_resolvable == 0 or avg_ok == avg_resolvable,
                  f"{avg_ok}/{avg_resolvable} within tol 0.5")
            check("Quiz Overview Avg_Difficulty values correct (all rows)",
                  diff_resolvable == 0 or diff_ok == diff_resolvable,
                  f"{diff_ok}/{diff_resolvable} within tol 0.05")

            # Sort order validation (course then quiz)
            actual_keys = [(str(r[idx_course] or "").strip(), str(r[idx_quiz] or "").strip())
                           for r in data_rows]
            check("Quiz Overview sorted by course then quiz",
                  actual_keys == sorted(actual_keys),
                  f"First 3: {actual_keys[:3]}")

    # Flagged Items sheet
    fi_rows = load_sheet_rows(wb_val, wb_formula, "Flagged Items")
    if fi_rows is None:
        check("Sheet 'Flagged Items' exists", False, f"Available: {wb_val.sheetnames}")
    else:
        check("Sheet 'Flagged Items' exists", True)
        data_rows = [r for r in (fi_rows[1:] if len(fi_rows) > 1 else []) if r and r[0]]
        check(f"Flagged Items row count == {flagged_count}",
              len(data_rows) == flagged_count,
              f"Found {len(data_rows)}, expected {flagged_count}")

        # Validate the flagged set matches and Issue values are correct
        header = fi_rows[0] if fi_rows else []
        idx_course = _find_col(header, "course")
        idx_quiz = _find_col(header, "quiz")
        idx_issue = _find_col(header, "issue")

        if data_rows and all(i >= 0 for i in [idx_course, idx_quiz, idx_issue]):
            expected_flagged = {(q["course"], q["quiz"]): q for q in expected if q["flagged"]}
            match_count = 0
            issue_correct = 0
            for r in data_rows:
                key = (str(r[idx_course] or "").strip(), str(r[idx_quiz] or "").strip())
                if key in expected_flagged:
                    match_count += 1
                    exp = expected_flagged[key]
                    issue_text = str(r[idx_issue] or "").strip().lower()
                    if exp["difficulty"] > 0.8 and "easy" in issue_text:
                        issue_correct += 1
                    elif exp["difficulty"] < 0.3 and "hard" in issue_text:
                        issue_correct += 1
            check("Flagged items set matches expected",
                  match_count == flagged_count,
                  f"{match_count}/{flagged_count} match")
            check("Flagged items have correct Issue label (all rows)",
                  issue_correct == flagged_count,
                  f"{issue_correct}/{flagged_count} correct labels")

    # Summary sheet
    sum_rows = load_sheet_rows(wb_val, wb_formula, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb_val.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        check(f"Total_Quizzes = {total_quizzes}",
              lookup.get("total_quizzes") is _FORMULA_UNRESOLVED or
              num_close(lookup.get("total_quizzes"), total_quizzes, tol=0),
              f"Got {lookup.get('total_quizzes')}")
        check(f"Total_Questions = {total_questions} (exact)",
              lookup.get("total_questions") is _FORMULA_UNRESOLVED or
              num_close(lookup.get("total_questions"), total_questions, tol=0),
              f"Got {lookup.get('total_questions')}")
        check(f"Flagged_Quizzes = {flagged_count} (exact)",
              lookup.get("flagged_quizzes") is _FORMULA_UNRESOLVED or
              num_close(lookup.get("flagged_quizzes"), flagged_count, tol=0),
              f"Got {lookup.get('flagged_quizzes')}")
        # Validate Flag_Rate (parse percentage)
        fr_raw = lookup.get("flag_rate")
        fr_val = None
        if fr_raw is not None and fr_raw is not _FORMULA_UNRESOLVED:
            try:
                fr_str = str(fr_raw).replace("%", "").strip()
                fr_val = float(fr_str)
                # Could be expressed 0-1 or 0-100
                if fr_val <= 1.0:
                    fr_val *= 100.0
            except (ValueError, TypeError):
                pass
        check(f"Flag_Rate ~= {flag_rate_pct}% (tol 0.5pp)",
              fr_raw is _FORMULA_UNRESOLVED or
              (fr_val is not None and abs(fr_val - flag_rate_pct) <= 0.5),
              f"Got {fr_raw}")


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Assessment_Report.docx")
    if not os.path.isfile(docx_path):
        check("Assessment_Report.docx exists", False, f"Not found: {docx_path}")
        return
    check("Assessment_Report.docx exists", True)
    check("Word doc has content (> 1KB)", os.path.getsize(docx_path) > 1000,
          f"Size: {os.path.getsize(docx_path)}")

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Report mentions difficulty", "difficult" in all_text, f"Sample: {all_text[:200]}")
        check("Report mentions flagged or quality", "flag" in all_text or "quality" in all_text,
              f"Sample: {all_text[:200]}")
        check("Report has recommendations", "recommend" in all_text,
              f"Sample: {all_text[:200]}")
    except ImportError:
        check("python-docx available", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
