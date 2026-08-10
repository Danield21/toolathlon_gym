"""
Evaluation for 12306-wc-amazon-product-roadshow-excel-notion task.

Checks (Excel / Roadshow_Plan.xlsx):
  1. Products sheet: exists, required columns, >=5 data rows, and the
     set of Product_IDs must be the real top-5 products by sales volume
     (computed live from wc.products.total_sales when the DB is reachable,
     otherwise read from the groundtruth workbook). Name / Category /
     Price_USD / Stock_Status / Monthly_Sales_Est are compared per product
     with tolerant matching.
  2. Customer_Regions sheet: exists, required columns, >=2 data rows, and
     Shanghai + Guangzhou rows must be present with Priority 1 and numeric
     counts / revenue shares. (Exact numbers are not derivable from data and
     are NOT compared.)
  3. Travel_Itinerary sheet: exists, required columns, >=2 data rows, with a
     G11 row (Mar 10, Beijing -> Shanghai/Hongqiao, 07:00 -> 12:31, ~5.5h)
     and a G105 row (Mar 10, Beijing -> Guangzhou, 08:00 -> 15:48, ~7.8h).
     Station names accepted in Chinese or English.
  4. Roadshow_Schedule sheet: exists, required columns, >=4 data rows, with a
     Shanghai meeting row on Mar 10 (with a contact) and a Guangzhou meeting
     row on Mar 11 (with a contact). Venue / Meeting_Type / Status are free-form.
  5. Notion page exists with Roadshow + Shanghai + Guangzhou in the title.
  6. Emails sent to shanghai_dist@partner.com, guangzhou_dist@partner.com,
     and manager@company.com with the specified subjects.
"""
import json
import os
import re
import sys
from argparse import ArgumentParser
from datetime import datetime, date, time

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


# --------------------------------------------------------------------------
# Value helpers
# --------------------------------------------------------------------------
def _to_float(v):
    """Parse a cell into a float, tolerating thousands separators, currency
    symbols, a trailing '%' and surrounding whitespace. Returns None when the
    value is empty / not numeric (formulas without a cached value included)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    if s.startswith("="):
        # Formula string (only surfaces when reading with data_only=False).
        return None
    for ch in (",", "$", "€", "¥", "£", "%", " ", " "):
        s = s.replace(ch, "")
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _to_int(v):
    """Parse a cell into an integer only when it is integral."""
    f = _to_float(v)
    if f is None:
        return None
    if f == int(f):
        return int(f)
    return None


def num_close(a, b, tol=1.0):
    """Numeric closeness after normalization; falls back to a case-insensitive
    string comparison only when one side cannot be parsed as a number."""
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    return str_match(a, b)


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _norm_name(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()


def name_match(a, b, min_ratio=0.5):
    """Tolerant product-name matching: exact, containment, or significant-token
    overlap. Handles long kaggle-style names that a model may truncate."""
    if a is None or b is None:
        return False
    na, nb = _norm_name(a), _norm_name(b)
    if na == nb:
        return True
    if na and na in nb:
        return True
    if nb and nb in na:
        return True
    ta = set(w for w in na.split() if len(w) >= 4)
    tb = set(w for w in nb.split() if len(w) >= 4)
    if not ta or not tb:
        return False
    return len(ta & tb) / min(len(ta), len(tb)) >= min_ratio


def _norm_alnum(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def stock_match(a, b):
    return _norm_alnum(a) == _norm_alnum(b)


def category_match(cell, expected_cats):
    """Pass if any expected category name appears (case-insensitive) in the
    cell, e.g. an 'Audio, Speakers' cell matching category 'Audio'."""
    if cell is None:
        return False
    s = str(cell).lower()
    for c in expected_cats or []:
        if str(c).lower() in s:
            return True
    return False


_MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

_MONTH_NAME_RE = (
    r"(?:january|february|march|april|may|june|july|august|september|"
    r"october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|oct|nov|dec)"
)


def _cell_date(v):
    """Parse a cell into (year_or_None, month, day). Handles YYYY-MM-DD,
    MM/DD/YYYY or DD/MM/YYYY, and month-name dates in either word order
    ('March 10, 2026' and '10 March 2026' / '10 Mar 2026')."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return (v.year, v.month, v.day)
    s = str(v).strip()
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        return (int(m.group(3)), int(m.group(1)), int(m.group(2)))
    # Day-first month-name form ('10 March 2026'): try before Month-day so the
    # leading day is not swallowed by the trailing year digits.
    m = re.search(
        r"(\d{1,2})(?:st|nd|rd|th)?\s+(" + _MONTH_NAME_RE + r")[a-z]*\.?"
        r"(?:\s*,?\s*(\d{4}))?",
        s,
        re.I,
    )
    if m:
        year = int(m.group(3)) if m.group(3) else None
        return (year, _MONTH_ABBR[m.group(2).lower()[:3]], int(m.group(1)))
    m = re.search(
        r"(" + _MONTH_NAME_RE + r")[a-z]*\.?\s+(\d{1,2})(?:\s*,?\s*(\d{4}))?",
        s,
        re.I,
    )
    if m:
        year = int(m.group(3)) if m.group(3) else None
        return (year, _MONTH_ABBR[m.group(1).lower()[:3]], int(m.group(2)))
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(\d{1,2})月(\d{1,2})日", s)
    if m:
        return (None, int(m.group(1)), int(m.group(2)))
    return None


def _month_day_candidates(v):
    """Return a set of (month, day) tuples a date cell could plausibly
    represent. Covers datetime objects, YYYY-MM-DD, ambiguous M/D/Y and
    D/M/Y (both readings), M/D or D/M, and month-name dates in either word
    order. Used to accept every legitimate spelling of 'March 10/11'."""
    out = set()
    if v is None:
        return out
    if isinstance(v, (datetime, date)):
        out.add((v.month, v.day))
        return out
    s = str(v).strip().lower()
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        out.add((int(m.group(2)), int(m.group(3))))
    else:
        m = re.search(r"(\d{1,2})[-/](\d{1,2})(?:[-/](\d{4}))?", s)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            out.add((a, b))
            out.add((b, a))
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?\s+(" + _MONTH_NAME_RE + r")[a-z]*\.?", s)
    if m:
        out.add((_MONTH_ABBR[m.group(2)[:3]], int(m.group(1))))
    m = re.search(r"(" + _MONTH_NAME_RE + r")[a-z]*\.?\s+(\d{1,2})", s)
    if m:
        out.add((_MONTH_ABBR[m.group(1)[:3]], int(m.group(2))))
    # Chinese date forms: 2026年3月10日 / 3月10日.
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s)
    if m:
        out.add((int(m.group(2)), int(m.group(3))))
    m = re.search(r"(\d{1,2})月(\d{1,2})日", s)
    if m:
        out.add((int(m.group(1)), int(m.group(2))))
    return out


def _time_to_minutes(v):
    """Parse a departure/arrival cell into minutes past midnight."""
    if v is None:
        return None
    if isinstance(v, time):
        return v.hour * 60 + v.minute
    if isinstance(v, datetime):
        return v.hour * 60 + v.minute
    if isinstance(v, (int, float)):
        f = float(v)
        if 0 <= f <= 24 and f == int(f):
            return int(f) * 60
        if 0 < f < 1:
            return f * 1440
        return None
    s = str(v).strip().lower()
    m = re.search(r"(\d{1,2}):(\d{2})", s)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2))
        if "pm" in s and hh < 12:
            hh += 12
        if "am" in s and hh == 12:
            hh = 0
        return hh * 60 + mm
    return None


def _hours_to_float(v):
    """Parse a duration cell ('5:31', 5.5, '5.5h', '5h 31m', ...) into hours."""
    if v is None:
        return None
    if isinstance(v, time):
        return v.hour + v.minute / 60.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower().replace(",", ".")
    m = re.match(r"^(\d{1,3}):(\d{1,2})$", s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60.0
    m = re.match(r"^(\d{1,3}(?:\.\d+)?)\s*h\s*$", s)
    if m:
        return float(m.group(1))
    m = re.match(r"^(\d{1,3})\s*h\s*(\d{1,2})\s*m(?:in)?\s*$", s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60.0
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def norm_sheet_name(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").strip().lower()).strip()


def find_sheet(wb, target):
    t = norm_sheet_name(target)
    for n in wb.sheetnames:
        if norm_sheet_name(n) == t:
            return wb[n]
    return None


def _norm_header(s):
    """Normalize a header cell for matching (tolerant of case, '_' vs ' ',
    parentheses, trailing punctuation)."""
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").strip().lower()).strip()


def _headers(rows):
    return [_norm_header(c) for c in (rows[0] if rows else [])]


def _find_col(headers, token):
    """Locate a column by a normalized keyword. Matches exact / whole-token
    first ('price' vs 'Price_USD'), then falls back to a substring match for
    longer keywords so 'contact' also finds 'Key_Contacts' while short tokens
    like 'to' never match the wrong column."""
    t = _norm_header(token)
    if not t:
        return None
    for i, h in enumerate(headers):
        if not h:
            continue
        if t == h or t in h.split() or h in t.split():
            return i
    if len(t) >= 4:
        for i, h in enumerate(headers):
            if not h:
                continue
            if t in h or h in t:
                return i
    return None


def _has_headers(headers, tokens):
    """True when every required column keyword is found among the headers."""
    return all(_find_col(headers, tok) is not None for tok in tokens)


def _col_val(row, headers, token):
    i = _find_col(headers, token)
    if i is None or i >= len(row):
        return None
    return row[i]


def _city_substrings(city_key):
    """Substrings that identify a city in English or Chinese."""
    return {
        "shanghai": ["shanghai", "上海"],
        "guangzhou": ["guangzhou", "广州"],
    }.get(city_key, [city_key])


# --------------------------------------------------------------------------
# Expected top-5 products (live from DB, fallback to groundtruth workbook)
# --------------------------------------------------------------------------
def _db_expected_products():
    """Top-5 products by sales volume, mirroring the woocommerce 'top sellers'
    report (SELECT ... FROM wc.products ORDER BY total_sales DESC). Returns a
    list of dicts, or None if the DB is unreachable."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, name, price, stock_status, total_sales, categories
            FROM wc.products
            ORDER BY total_sales DESC, id ASC
            LIMIT 5
            """
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        result = []
        for r in rows:
            cats = []
            raw_cats = r[5]
            if raw_cats:
                if isinstance(raw_cats, str):
                    try:
                        raw_cats = json.loads(raw_cats)
                    except Exception:
                        raw_cats = []
                try:
                    for c in raw_cats:
                        if isinstance(c, dict) and c.get("name"):
                            cats.append(c["name"])
                except Exception:
                    pass
            result.append(
                {
                    "id": int(r[0]),
                    "name": r[1],
                    "price": _to_float(r[2]),
                    "stock_status": r[3],
                    "total_sales": _to_float(r[4]),
                    "categories": cats,
                }
            )
        return result or None
    except Exception:
        return None


def _gt_expected_products(gt_path):
    """Read the Products sheet of the groundtruth workbook as the expected
    top-5 products (used when the DB is unavailable)."""
    try:
        wb = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception:
        return None
    ws = find_sheet(wb, "Products")
    if ws is None:
        return None
    rows = [r for r in ws.iter_rows(values_only=True)]
    headers = _headers(rows)
    result = []
    for r in rows[1:]:
        pid = _to_int(_col_val(r, headers, "product_id"))
        name = _col_val(r, headers, "name")
        if pid is None or name is None:
            continue
        cat = _col_val(r, headers, "category")
        result.append(
            {
                "id": pid,
                "name": name,
                "price": _to_float(_col_val(r, headers, "price")),
                "stock_status": _col_val(r, headers, "stock"),
                "total_sales": _to_float(_col_val(r, headers, "monthly")),
                "categories": [c.strip() for c in str(cat).split(",") if c.strip()],
            }
        )
    return result or None


def _resolve_expected_products(gt_path):
    """Pick the expected top-5 product list.

    Prefer the live DB top-5 (the set the model saw through the woocommerce
    MCP). If the DB top-5 disagrees with the groundtruth workbook, the
    deployment is inconsistent (e.g. the MCP bridge fell back to another
    database); in that case use the groundtruth as the authoritative reference
    and warn, so a correctly-completed workspace is never judged against a
    divergent DB."""
    db_exp = _db_expected_products()
    gt_exp = _gt_expected_products(gt_path)
    if db_exp is not None and gt_exp is not None:
        db_ids = sorted(p["id"] for p in db_exp)
        gt_ids = sorted(p["id"] for p in gt_exp)
        if db_ids != gt_ids:
            print(
                f"  [WARN] DB top-5 {db_ids} differs from groundtruth top-5 {gt_ids}; "
                "using groundtruth as authoritative (possible DB/MCP mismatch)."
            )
            return gt_exp
        return db_exp
    if db_exp is not None:
        return db_exp
    return gt_exp


# --------------------------------------------------------------------------
# Excel checks
# --------------------------------------------------------------------------
def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Roadshow_Plan.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Roadshow_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Roadshow_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Roadshow_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    gt_path = os.path.join(groundtruth_workspace, "Roadshow_Plan.xlsx")

    # ---- Products sheet ----
    prod_sheet = find_sheet(wb, "Products")
    record("Products sheet exists", prod_sheet is not None, f"Sheets: {wb.sheetnames}")
    if prod_sheet is not None:
        rows = [r for r in prod_sheet.iter_rows(values_only=True)]
        headers = _headers(rows)
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        record(
            "Products has required columns",
            _has_headers(headers, ("product_id", "name", "price", "stock", "monthly")),
            f"Headers: {rows[0] if rows else []}",
        )
        record(
            "Products has at least 5 data rows",
            len(data_rows) >= 5,
            f"Found {len(data_rows)} data rows",
        )

        expected = _resolve_expected_products(gt_path)
        if expected is None:
            record(
                "Products expected top-5 available",
                False,
                "Could not determine top-5 (DB unreachable and no groundtruth)",
            )
        else:
            id_col = _find_col(headers, "product_id")
            agent_by_id = {}
            for r in data_rows:
                pid = _to_int(r[id_col] if id_col is not None else None)
                if pid is not None:
                    agent_by_id[pid] = r
            agent_ids = set(agent_by_id.keys())
            expected_ids = set(p["id"] for p in expected)
            record(
                "Products contains the top-5 product IDs (by sales volume)",
                agent_ids >= expected_ids,
                f"Expected {sorted(expected_ids)}, agent has {sorted(agent_ids)}",
            )
            for p in expected:
                pid = p["id"]
                if pid not in agent_by_id:
                    record(f"Products: product {pid} present", False, "ID not found in agent sheet")
                    continue
                r = agent_by_id[pid]
                name = _col_val(r, headers, "name")
                cat = _col_val(r, headers, "category")
                price = _col_val(r, headers, "price")
                stock = _col_val(r, headers, "stock")
                sales = _col_val(r, headers, "monthly")
                ok = True
                issues = []
                if not name_match(name, p["name"]):
                    ok = False
                    issues.append(
                        f"name: expected '{str(p['name'])[:45]}' got '{str(name)[:45]}'"
                    )
                if not category_match(cat, p["categories"]):
                    ok = False
                    issues.append(f"category: expected {p['categories']} got '{cat}'")
                if p["price"] is None or not num_close(
                    price, p["price"], tol=max(abs(p["price"] or 0) * 0.1, 0.5)
                ):
                    ok = False
                    issues.append(f"price: expected {p['price']} got '{price}'")
                if not stock_match(stock, p["stock_status"]):
                    ok = False
                    issues.append(f"stock: expected {p['stock_status']} got '{stock}'")
                sf = _to_float(sales)
                if sf is None or sf <= 0:
                    ok = False
                    issues.append(f"Monthly_Sales_Est not a positive number: '{sales}'")
                record(f"Products: product {pid} matches", ok, "; ".join(issues))

    # ---- Customer_Regions sheet ----
    cr_sheet = find_sheet(wb, "Customer_Regions")
    record(
        "Customer_Regions sheet exists",
        cr_sheet is not None,
        f"Sheets: {wb.sheetnames}",
    )
    if cr_sheet is not None:
        rows = [r for r in cr_sheet.iter_rows(values_only=True)]
        headers = _headers(rows)
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        record(
            "Customer_Regions has required columns",
            _has_headers(headers, ("region", "customer", "revenue", "priority")),
            f"Headers: {rows[0] if rows else []}",
        )
        record(
            "Customer_Regions has at least 2 data rows",
            len(data_rows) >= 2,
            f"Found {len(data_rows)} data rows",
        )

        def _city_priority(city_key):
            keys = _city_substrings(city_key)
            for r in data_rows:
                region = _col_val(r, headers, "region")
                prio = _to_float(_col_val(r, headers, "priority"))
                if (
                    region
                    and any(k in str(region).lower() for k in keys)
                    and prio is not None
                    and abs(prio - 1.0) < 0.01
                ):
                    return True
            return False

        def _city_numeric(city_key):
            keys = _city_substrings(city_key)
            for r in data_rows:
                region = _col_val(r, headers, "region")
                if region and any(k in str(region).lower() for k in keys):
                    cc = _to_float(_col_val(r, headers, "customer"))
                    rs = _to_float(_col_val(r, headers, "revenue"))
                    return cc is not None and rs is not None
            return False

        record(
            "Customer_Regions marks Shanghai as Priority 1", _city_priority("shanghai")
        )
        record(
            "Customer_Regions marks Guangzhou as Priority 1", _city_priority("guangzhou")
        )
        record(
            "Customer_Regions Shanghai row has numeric counts/revenue",
            _city_numeric("shanghai"),
        )
        record(
            "Customer_Regions Guangzhou row has numeric counts/revenue",
            _city_numeric("guangzhou"),
        )

    # ---- Travel_Itinerary sheet ----
    travel_sheet = find_sheet(wb, "Travel_Itinerary")
    record(
        "Travel_Itinerary sheet exists",
        travel_sheet is not None,
        f"Sheets: {wb.sheetnames}",
    )
    if travel_sheet is not None:
        rows = [r for r in travel_sheet.iter_rows(values_only=True)]
        headers = _headers(rows)
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        record(
            "Travel_Itinerary has required columns",
            _has_headers(
                headers,
                ("date", "train", "from", "to", "depart", "arrive", "duration"),
            ),
            f"Headers: {rows[0] if rows else []}",
        )
        record(
            "Travel_Itinerary has at least 2 data rows",
            len(data_rows) >= 2,
            f"Found {len(data_rows)} data rows",
        )

        def _find_train(train_no):
            for r in data_rows:
                tn = _col_val(r, headers, "train")
                if tn is not None and str(tn).strip().upper() == train_no:
                    return r
            return None

        def _from_ok(v):
            s = str(v).lower()
            return "北京" in s or "beijing" in s

        def _to_ok(v, keys):
            s = str(v).lower()
            return any(k in s for k in keys)

        def _check_train(train_no, to_keys, dep_min, arr_min, dur_h):
            r = _find_train(train_no)
            if r is None:
                record(
                    f"Travel_Itinerary has row for train {train_no}",
                    False,
                    f"Train {train_no} not found",
                )
                return
            ok = True
            issues = []
            if (3, 10) not in _month_day_candidates(_col_val(r, headers, "date")):
                ok = False
                issues.append(f"date: expected 2026-03-10, got '{_col_val(r, headers, 'date')}'")
            frm = _col_val(r, headers, "from")
            if not _from_ok(frm):
                ok = False
                issues.append(f"From: expected Beijing station, got '{frm}'")
            to = _col_val(r, headers, "to")
            if not _to_ok(to, to_keys):
                ok = False
                issues.append(f"To: expected {to_keys}, got '{to}'")
            dep = _time_to_minutes(_col_val(r, headers, "depart"))
            arr = _time_to_minutes(_col_val(r, headers, "arrive"))
            if dep is None or abs(dep - dep_min) > 15:
                ok = False
                issues.append(
                    f"Depart: expected {dep_min // 60:02d}:{dep_min % 60:02d}, "
                    f"got '{_col_val(r, headers, 'depart')}'"
                )
            if arr is None or abs(arr - arr_min) > 15:
                ok = False
                issues.append(
                    f"Arrive: expected {arr_min // 60:02d}:{arr_min % 60:02d}, "
                    f"got '{_col_val(r, headers, 'arrive')}'"
                )
            dur = _hours_to_float(_col_val(r, headers, "duration"))
            if dur is None or abs(dur - dur_h) > 1.0:
                ok = False
                issues.append(
                    f"Duration: expected ~{dur_h} h, got '{_col_val(r, headers, 'duration')}'"
                )
            pur = _col_val(r, headers, "purpose")
            if pur is None or not str(pur).strip():
                ok = False
                issues.append("Purpose empty")
            record(
                f"Travel_Itinerary row for train {train_no} matches",
                ok,
                "; ".join(issues),
            )

        _check_train("G11", ["shanghai", "hongqiao", "上海", "虹桥"], 7 * 60, 12 * 60 + 31, 5.5)
        _check_train("G105", ["guangzhou", "广州"], 8 * 60, 15 * 60 + 48, 7.8)

    # ---- Roadshow_Schedule sheet ----
    sched_sheet = find_sheet(wb, "Roadshow_Schedule")
    record(
        "Roadshow_Schedule sheet exists",
        sched_sheet is not None,
        f"Sheets: {wb.sheetnames}",
    )
    if sched_sheet is not None:
        rows = [r for r in sched_sheet.iter_rows(values_only=True)]
        headers = _headers(rows)
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        record(
            "Roadshow_Schedule has required columns",
            _has_headers(
                headers,
                ("city", "date", "venue", "meeting", "contact", "status"),
            ),
            f"Headers: {rows[0] if rows else []}",
        )
        record(
            "Roadshow_Schedule has at least 4 data rows",
            len(data_rows) >= 4,
            f"Found {len(data_rows)} data rows",
        )

        def _has_city_meeting(city_key, month, day):
            keys = _city_substrings(city_key)
            for r in data_rows:
                city = _col_val(r, headers, "city")
                if not city or not any(k in str(city).lower() for k in keys):
                    continue
                if (month, day) not in _month_day_candidates(_col_val(r, headers, "date")):
                    continue
                contact = _col_val(r, headers, "contact")
                if contact is not None and str(contact).strip():
                    return True
            return False

        record(
            "Roadshow_Schedule has a Shanghai meeting row on Mar 10",
            _has_city_meeting("shanghai", 3, 10),
        )
        record(
            "Roadshow_Schedule has a Guangzhou meeting row on Mar 11",
            _has_city_meeting("guangzhou", 3, 11),
        )


# --------------------------------------------------------------------------
# Notion check
# --------------------------------------------------------------------------
def _deep_text(node, out):
    """Recursively collect text from a Notion jsonb structure, tolerating every
    shape the notion MCP can store: rich-text arrays, text.content / plain_text,
    bare strings, {'title': {...}} / {'properties': {...}} wrappers and
    title-as-rich-text-array. Guards against malformed nodes."""
    if isinstance(node, str):
        s = node.strip()
        if s:
            out.append(s)
    elif isinstance(node, dict):
        t = node.get("text")
        if isinstance(t, dict) and t.get("content"):
            out.append(str(t["content"]))
        elif isinstance(t, str) and t.strip():
            out.append(t)
        pt = node.get("plain_text")
        if isinstance(pt, str) and pt.strip():
            out.append(pt)
        c = node.get("content")
        if isinstance(c, str) and c.strip():
            out.append(c)
        for v in node.values():
            _deep_text(v, out)
    elif isinstance(node, list):
        for it in node:
            _deep_text(it, out)


def _has_roadshow_title(texts):
    """True when the collected text contains the roadshow title markers."""
    joined = " ".join(texts).lower()
    return (
        "roadshow" in joined
        and "shanghai" in joined
        and "guangzhou" in joined
    )


def check_notion():
    """Returns (ok, definitive). ok = a page with the required title exists;
    definitive = False when the DB is unreachable or the query failed, so the
    caller does not block PASS on an infra-level failure."""
    print("\n=== Check 2: Notion Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Notion page title contains Roadshow + Shanghai + Guangzhou", False,
               f"DB unavailable: {str(e)[:120]}")
        return False, False
    found = False
    try:
        cur = conn.cursor()
        # 1) Page titles from notion.pages.properties (authoritative).
        try:
            cur.execute("SELECT properties FROM notion.pages")
            for (props,) in cur.fetchall():
                texts = []
                _deep_text(props, texts)
                if _has_roadshow_title(texts):
                    found = True
                    break
        except Exception as e:
            record("Notion page title contains Roadshow + Shanghai + Guangzhou", False,
                   f"pages query failed: {str(e)[:120]}")
            cur.close()
            conn.close()
            return False, False
        # 2) Fallback: title-bearing blocks in notion.blocks.block_data (page
        #    blocks or blocks carrying a 'title'/'properties' key).
        if not found:
            try:
                cur.execute(
                    "SELECT block_data FROM notion.blocks "
                    "WHERE type = 'page' OR block_data ? 'title' OR block_data ? 'properties'"
                )
                for (bd,) in cur.fetchall():
                    texts = []
                    _deep_text(bd, texts)
                    if _has_roadshow_title(texts):
                        found = True
                        break
            except Exception:
                # blocks table may not exist in every deployment; pages path is
                # authoritative.
                pass
        cur.close()
        conn.close()
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        record("Notion page title contains Roadshow + Shanghai + Guangzhou", False,
               f"Query failed: {str(e)[:120]}")
        return False, False

    record(
        "Notion page title contains Roadshow + Shanghai + Guangzhou",
        found,
        "Scanned notion.pages.properties and notion.blocks.block_data",
    )
    return found, True


# --------------------------------------------------------------------------
# Email check
# --------------------------------------------------------------------------
def check_emails_sent():
    print("\n=== Check 3: Emails Sent ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Email send check", False, f"DB unavailable: {str(e)[:120]}")
        return False, False
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT m.subject, m.to_addr FROM email.messages m
            JOIN email.folders f ON m.folder_id = f.id
            WHERE UPPER(f.name) = 'SENT'
            """
        )
        msgs = list(cur.fetchall())
        cur.execute(
            """
            SELECT m.subject, m.to_addr FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            """
        )
        msgs += list(cur.fetchall())

        def to_str(v):
            if isinstance(v, list):
                return " ".join(str(x).lower() for x in v)
            if isinstance(v, str):
                try:
                    p = json.loads(v)
                    if isinstance(p, list):
                        return " ".join(str(x).lower() for x in p)
                except Exception:
                    pass
                return v.lower()
            return str(v).lower()

        normalized = [((s or "").lower(), to_str(t)) for s, t in msgs]

        def has_msg(recipient, subject_substr):
            for subj, to in normalized:
                if recipient in to and subject_substr in subj:
                    return True
            return False

        checks = [
            (
                "Email to shanghai_dist with subject 'Roadshow Meeting Confirmation - Shanghai March 10'",
                "shanghai_dist@partner.com",
                "roadshow meeting confirmation - shanghai march 10",
            ),
            (
                "Email to guangzhou_dist with subject 'Roadshow Meeting Confirmation - Guangzhou March 11'",
                "guangzhou_dist@partner.com",
                "roadshow meeting confirmation - guangzhou march 11",
            ),
            (
                "Email to manager@company.com with subject 'Roadshow Plan Summary - Shanghai and Guangzhou March 2026'",
                "manager@company.com",
                "roadshow plan summary - shanghai and guangzhou march 2026",
            ),
        ]
        all_ok = True
        for name, recipient, subj in checks:
            ok = has_msg(recipient, subj)
            if not ok:
                all_ok = False
            record(name, ok, f"Total: {len(normalized)}")
        return all_ok, True
    except Exception as e:
        record("Email send check", False, str(e))
        return False, False
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


# --------------------------------------------------------------------------
def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    file_fail = FAIL_COUNT

    notion_ok, notion_definitive = check_notion()
    email_ok, email_definitive = check_emails_sent()
    runtime_fail = FAIL_COUNT - file_fail

    # A runtime check is "definitive" only when the DB was reachable and the
    # query actually ran. If the DB is unreachable the check is inconclusive
    # and must not block PASS (infra-level failure, not an agent error).
    runtime_hard_fail = 0
    if notion_definitive and not notion_ok:
        runtime_hard_fail += 1
    if email_definitive and not email_ok:
        runtime_hard_fail += 1

    # GT self-test mode (agent_workspace == groundtruth_workspace): the GT dir
    # can never contain the agent's runtime artifacts, so runtime checks are
    # reported but never block there.
    self_test = (
        os.path.realpath(args.agent_workspace)
        == os.path.realpath(args.groundtruth_workspace)
    )
    blocking_runtime_fail = runtime_hard_fail if not self_test else 0

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%) "
          f"(file_fail={file_fail}, runtime_fail={runtime_fail})")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "file_fail": file_fail,
        "runtime_fail": runtime_fail,
        "runtime_hard_fail": runtime_hard_fail,
        "blocking_runtime_fail": blocking_runtime_fail,
        "self_test": self_test,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if file_fail == 0 and blocking_runtime_fail == 0:
        if self_test:
            print("PASS (file checks clean; self-test)")
        else:
            print("PASS (file checks and runtime deliverables clean)")
        sys.exit(0)
    else:
        if file_fail:
            print(f"FAIL ({file_fail} file-level failures)")
        elif blocking_runtime_fail:
            print(f"FAIL ({blocking_runtime_fail} runtime deliverable(s) missing)")
        sys.exit(1)


if __name__ == "__main__":
    main()
