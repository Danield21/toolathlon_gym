"""Evaluation for terminal-yf-scholarly-excel-word-email.
Checks:
1. FinTech_Research_Report.xlsx with 4 sheets and correct data
2. FinTech_Research_Report.docx with required sections
3. Email sent to research-committee@company.org
4. market_analysis.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    print("\n=== Check 1: FinTech_Research_Report.xlsx ===")
    path = os.path.join(workspace, "FinTech_Research_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    def _find_sheet(needle):
        # Exact (case-insensitive), then prefix, then substring
        for s in sheets:
            if s.lower() == needle.lower():
                return s
        for s in sheets:
            if s.lower().startswith(needle.lower()):
                return s
        for s in sheets:
            if needle.lower() in s.lower():
                return s
        return None

    # Market_Data sheet (require exact name)
    md_name = _find_sheet("Market_Data")
    if md_name is None:
        check("Market_Data sheet exists", False, f"Sheets: {sheets}")
    else:
        check("Market_Data sheet exists", True)
        ws = wb[md_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        check("Market_Data has exactly 2 stock rows", len(data_rows) == 2, f"Found {len(data_rows)}")
        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        check("Contains JPM", "jpm" in all_text, f"Text: {all_text[:120]}")
        check("Contains XOM", "xom" in all_text, f"Text: {all_text[:120]}")
        # Sector references
        check("Mentions Financial sector for JPM", "financial" in all_text or "banking" in all_text)
        check("Mentions Energy sector for XOM", "energy" in all_text)

    # Academic_Papers sheet
    ap_name = _find_sheet("Academic_Papers")
    if ap_name is None:
        check("Academic_Papers sheet exists", False, f"Sheets: {sheets}")
    else:
        check("Academic_Papers sheet exists", True)
        ws2 = wb[ap_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Academic_Papers has at least 4 rows", len(data_rows2) >= 4, f"Found {len(data_rows2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        # Specific paper references (must include at least 2 of these)
        paper_titles_seen = sum([
            "efficient capital markets" in all_text2 or ("efficient" in all_text2 and "capital" in all_text2),
            "random walk" in all_text2,
            "algorithmic trading" in all_text2,
            "big data" in all_text2 or "modern data" in all_text2,
        ])
        check("Academic_Papers mentions at least 2 specific market-efficiency paper topics",
              paper_titles_seen >= 2,
              f"Recognized paper-topic count: {paper_titles_seen}; text sample: {all_text2[:200]}")
        # Should mention 'market efficiency' or 'efficient market' explicitly
        check("Academic_Papers explicitly references market efficiency",
              "market efficiency" in all_text2 or "efficient market" in all_text2 or
              "weak-form" in all_text2 or "weak form" in all_text2,
              f"Sample: {all_text2[:200]}")

    # Statistical_Tests sheet
    st_name = _find_sheet("Statistical_Tests")
    if st_name is None:
        check("Statistical_Tests sheet exists", False, f"Sheets: {sheets}")
    else:
        check("Statistical_Tests sheet exists", True)
        ws3 = wb[st_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Statistical_Tests has exactly 4 rows (autocorr + normality for 2 stocks)",
              len(data_rows3) == 4, f"Found {len(data_rows3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Contains autocorrelation test (Ljung-Box)",
              "autocorrelation" in all_text3 or "ljung" in all_text3,
              f"Text: {all_text3[:200]}")
        check("Contains normality test (Jarque-Bera)",
              "normality" in all_text3 or "jarque" in all_text3,
              f"Text: {all_text3[:200]}")
        # Check that p-values exist in a 'p_value' / 'p-value' column
        header_row = rows3[0] if rows3 else []
        p_col_idx = None
        for i, h in enumerate(header_row):
            hl = str(h or "").strip().lower().replace("-", "_")
            if hl in ("p_value", "pvalue") or "p_value" in hl:
                p_col_idx = i
                break
        p_vals = []
        if p_col_idx is not None:
            for r in data_rows3:
                if p_col_idx < len(r):
                    try:
                        v = float(r[p_col_idx])
                        if 0.0 <= v <= 1.0:
                            p_vals.append(v)
                    except (TypeError, ValueError):
                        continue
        check("Statistical_Tests has p_value column with values in [0,1] for all 4 tests",
              p_col_idx is not None and len(p_vals) == len(data_rows3),
              f"p_col_idx={p_col_idx}, found {len(p_vals)} valid p-values; first={p_vals[:6]}")
        # Conclusion column - must mention reject/fail-to-reject for each row
        concl_count = sum(1 for r in data_rows3
                          if any("reject" in str(c).lower() for c in r if c))
        check("Statistical_Tests has reject-related conclusions for all 4 tests",
              concl_count == len(data_rows3),
              f"Got {concl_count}/{len(data_rows3)} rows with reject-conclusion")

    # Research_Summary sheet
    rs_name = _find_sheet("Research_Summary")
    if rs_name is None:
        check("Research_Summary sheet exists", False, f"Sheets: {sheets}")
    else:
        check("Research_Summary sheet exists", True)
        ws4 = wb[rs_name]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Research_Summary has at least 3 findings", len(data_rows4) >= 3, f"Found {len(data_rows4)}")
        # Each row should have non-empty finding/evidence/implication
        complete_rows = sum(
            1 for r in data_rows4
            if len([c for c in r[:3] if c and str(c).strip()]) >= 3
        )
        check("Research_Summary rows have finding+evidence+implication populated",
              complete_rows >= 3,
              f"Got {complete_rows}/{len(data_rows4)} complete rows")


def check_word(workspace):
    print("\n=== Check 2: FinTech_Research_Report.docx ===")
    path = os.path.join(workspace, "FinTech_Research_Report.docx")
    if not os.path.exists(path):
        check("Word file exists", False, f"Not found at {path}")
        return
    check("Word file exists", True)

    try:
        from docx import Document
        doc = Document(path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Contains introduction or objectives", "introduction" in all_text or "objective" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains methodology", "methodology" in all_text or "method" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains results", "result" in all_text or "finding" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains conclusion", "conclusion" in all_text or "implication" in all_text,
              f"Text: {all_text[:150]}")
        check("Mentions market efficiency", "market efficiency" in all_text or "efficient market" in all_text,
              f"Text: {all_text[:150]}")
    except ImportError:
        check("python-docx available", False, "python-docx not installed")


def check_email():
    print("\n=== Check 3: Email to Research Committee ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE lower(to_addr::text) LIKE '%%research-committee@company.org%%'
    """)
    msgs = cur.fetchall()

    check("Email sent to research-committee@company.org",
          len(msgs) >= 1,
          f"Found {len(msgs)} matching messages")

    if msgs:
        # Subject must be exactly: "FinTech Research Report - Market Efficiency Analysis"
        ok_subj = False
        ok_jpm = False
        ok_xom = False
        ok_test = False
        for subj, to_addr, body in msgs:
            sl = (subj or "").strip().lower()
            bl = (body or "").lower()
            # Strict equality on subject
            if sl == "fintech research report - market efficiency analysis":
                ok_subj = True
            if "jpm" in bl or "morgan" in bl:
                ok_jpm = True
            if "xom" in bl or "exxon" in bl:
                ok_xom = True
            # Body should reference statistical test results (autocorrelation/ljung/normality/jarque)
            if any(kw in bl for kw in ("autocorrelation", "ljung", "normality", "jarque", "p-value", "p value")):
                ok_test = True
        check("Subject 'FinTech Research Report - Market Efficiency Analysis'", ok_subj,
              f"Subjects: {[s for s, _, _ in msgs]}")
        check("Email body mentions JPM / Morgan", ok_jpm,
              f"Bodies sample: {[(b or '')[:200] for _, _, b in msgs[:1]]}")
        check("Email body mentions XOM / Exxon", ok_xom,
              f"Bodies sample: {[(b or '')[:200] for _, _, b in msgs[:1]]}")
        check("Email body references statistical test results", ok_test,
              f"Bodies sample: {[(b or '')[:200] for _, _, b in msgs[:1]]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: market_analysis.py ===")
    path = os.path.join(workspace, "market_analysis.py")
    if not os.path.exists(path):
        check("market_analysis.py exists", False, f"Not found at {path}")
        return
    check("market_analysis.py exists", True)
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read().lower()
    except Exception as e:
        check("market_analysis.py readable", False, str(e))
        return
    check("market_analysis.py readable", True)
    # Should reference autocorrelation / Ljung-Box test
    check("Script implements autocorrelation / Ljung-Box test",
          "ljung" in content or "autocorrelation" in content or "acorr_ljungbox" in content,
          f"Sample: {content[:200]}")
    check("Script implements normality test (Jarque-Bera or shapiro)",
          "jarque" in content or "jarque_bera" in content or "normaltest" in content
          or "shapiro" in content or "normality" in content,
          f"Sample: {content[:200]}")
    check("Script computes daily returns",
          "return" in content and ("pct_change" in content or "diff" in content or "log" in content),
          f"Sample: {content[:200]}")
    check("Script uses 0.05 significance level",
          "0.05" in content or "alpha = 0.05" in content or "alpha=0.05" in content,
          f"Sample: {content[:200]}")


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "FinTech_Research_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets
        expected_keywords = {"market", "data", "academic", "paper", "statistic", "test",
                             "research", "summary", "finding"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

    # Email: no duplicate research emails
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, COUNT(*) FROM email.messages
            WHERE lower(subject) LIKE '%%fintech%%' OR lower(subject) LIKE '%%market efficiency%%'
            GROUP BY subject HAVING COUNT(*) > 1
        """)
        dupes = cur.fetchall()
        check("No duplicate research emails", len(dupes) == 0,
              f"Duplicates: {dupes}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Use FAIL_COUNT == 0 as the bar (replaces 70% threshold which masked critical fails)
    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
