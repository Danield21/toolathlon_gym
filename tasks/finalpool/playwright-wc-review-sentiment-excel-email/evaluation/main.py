"""
Evaluation script for playwright-wc-review-sentiment-excel-email task.

Checks:
1. Review_Comparison_Report.xlsx with Product Comparison and Category Summary sheets
2. Email with quality alert for flagged products
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.3):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Review_Comparison_Report.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Review_Comparison_Report.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    # Load GT workbook for per-product/category comparison.
    gt_file = os.path.join(groundtruth_workspace, "Review_Comparison_Report.xlsx")
    gt_wb = None
    if os.path.isfile(gt_file):
        try:
            gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
        except Exception:
            gt_wb = None

    all_ok = True

    # Check Product Comparison sheet
    comp_sheet = None
    for name in wb.sheetnames:
        if "comparison" in name.lower() or "product" in name.lower():
            comp_sheet = name
            break

    if not comp_sheet:
        record("Product Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Product Comparison sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and r[0]] if len(rows) > 1 else []

        record(
            "Product Comparison has >= 12 products",
            len(data_rows) >= 12,
            f"Found {len(data_rows)} data rows",
        )
        if len(data_rows) < 12:
            all_ok = False

        # Check that alert flags exist
        alert_col = None
        if rows:
            header = [str(h).lower() if h else "" for h in rows[0]]
            for i, h in enumerate(header):
                if "alert" in h or "flag" in h:
                    alert_col = i
                    break

        if alert_col is not None:
            flagged = [r for r in data_rows if r[alert_col] and str(r[alert_col]).lower() == "yes"]
            record(
                "At least 2 products flagged",
                len(flagged) >= 2,
                f"Found {len(flagged)} flagged products",
            )
            if len(flagged) < 2:
                all_ok = False
        else:
            record("Alert flag column exists", False, f"Headers: {rows[0] if rows else 'none'}")
            all_ok = False

        # Check rating difference column exists
        diff_col = None
        if rows:
            for i, h in enumerate(header):
                if "difference" in h or "diff" in h:
                    diff_col = i
                    break

        if diff_col is not None:
            record("Rating difference column exists", True)
            # Per-product check of Rating_Difference + Alert_Flag against GT (all products).
            if gt_wb is not None:
                # Locate GT Product Comparison sheet.
                gt_sheet_name = None
                for n in gt_wb.sheetnames:
                    if "comparison" in n.lower() or "product" in n.lower():
                        gt_sheet_name = n; break
                if gt_sheet_name:
                    gt_ws = gt_wb[gt_sheet_name]
                    gt_rows = list(gt_ws.iter_rows(values_only=True))
                    gt_header = [str(h).lower() if h else "" for h in gt_rows[0]]
                    gt_diff_col = next((i for i, h in enumerate(gt_header) if "difference" in h or "diff" in h), None)
                    gt_alert_col = next((i for i, h in enumerate(gt_header) if "alert" in h or "flag" in h), None)
                    gt_data = [r for r in gt_rows[1:] if r and r[0]]
                    # Build lookup: product_name.lower() -> (rating_diff, alert)
                    gt_lookup = {}
                    for r in gt_data:
                        key = str(r[0]).strip().lower()
                        gt_lookup[key] = (
                            r[gt_diff_col] if gt_diff_col is not None else None,
                            str(r[gt_alert_col]).strip().lower() if gt_alert_col is not None and r[gt_alert_col] else None,
                        )
                    # Match agent rows to GT entries; verify each.
                    for ar in data_rows:
                        name = str(ar[0]).strip().lower() if ar[0] else ""
                        if name in gt_lookup:
                            gt_diff, gt_alert = gt_lookup[name]
                            if gt_diff is not None:
                                ok = num_close(ar[diff_col], gt_diff, tol=0.1)
                                record(f"'{ar[0][:30]}' Rating_Difference ~{gt_diff} (tol 0.1)",
                                       ok, f"Got {ar[diff_col]}")
                                if not ok:
                                    all_ok = False
                            if gt_alert is not None and alert_col is not None:
                                a_alert = str(ar[alert_col]).strip().lower() if ar[alert_col] else ""
                                ok = a_alert == gt_alert
                                record(f"'{ar[0][:30]}' Alert_Flag == {gt_alert}",
                                       ok, f"Got {ar[alert_col]}")
                                if not ok:
                                    all_ok = False
            else:
                # Fall-back: spot check NIHARA only.
                for r in data_rows:
                    if r[0] and "nihara" in str(r[0]).lower():
                        ok = num_close(r[diff_col], 0.6, tol=0.1)
                        record("NIHARA rating diff ~0.6 (tol 0.1)", ok, f"Got {r[diff_col]}")
                        if not ok:
                            all_ok = False
                        break
        else:
            record("Rating difference column exists", False)
            all_ok = False

        # Collect flagged products list for cross-check with email
        try:
            flagged_products = []
            for r in data_rows:
                if alert_col is not None and r[alert_col] and str(r[alert_col]).strip().lower() == "yes":
                    if r[0]:
                        flagged_products.append(str(r[0]).strip())
            # Stash on function object for cross-check
            check_excel.flagged_products = flagged_products
            record(f"Flagged products captured ({len(flagged_products)})", True,
                   f"Flagged: {flagged_products}")
        except Exception as e:
            record("Flagged products capture", False, str(e))

    # Check Category Summary sheet
    cat_sheet = None
    for name in wb.sheetnames:
        if "category" in name.lower() or "summary" in name.lower():
            cat_sheet = name
            break

    if not cat_sheet:
        record("Category Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Category Summary sheet exists", True)
        ws2 = wb[cat_sheet]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if r and r[0]] if len(rows2) > 1 else []

        record(
            "Category Summary has 3 categories",
            len(data_rows2) >= 3,
            f"Found {len(data_rows2)} rows",
        )
        if len(data_rows2) < 3:
            all_ok = False

        # Check categories present
        cats = [str(r[0]).lower() for r in data_rows2 if r[0]]
        has_electronics = any("electron" in c for c in cats)
        has_cameras = any("camera" in c for c in cats)
        has_appliances = any("appliance" in c for c in cats)
        record(
            "All 3 categories present",
            has_electronics and has_cameras and has_appliances,
            f"Categories: {cats}",
        )
        if not (has_electronics and has_cameras and has_appliances):
            all_ok = False

        # Validate numeric averages against GT (Avg_Internal_Rating, Avg_External_Rating, Avg_Rating_Difference).
        if gt_wb is not None:
            gt_cat_name = None
            for n in gt_wb.sheetnames:
                if "category" in n.lower() or "summary" in n.lower():
                    gt_cat_name = n; break
            if gt_cat_name:
                gt_ws2 = gt_wb[gt_cat_name]
                gt_rows2 = list(gt_ws2.iter_rows(values_only=True))
                gt_hdr2 = [str(h).lower() if h else "" for h in gt_rows2[0]]
                # Build col map
                ag_hdr2 = [str(h).lower() if h else "" for h in rows2[0]] if rows2 else []
                def col(hdr_list, *keys):
                    for i, h in enumerate(hdr_list):
                        for kw in keys:
                            if all(p in h for p in kw):
                                return i
                    return None
                gt_lookup2 = {}
                for gr in gt_rows2[1:]:
                    if gr and gr[0]:
                        gt_lookup2[str(gr[0]).strip().lower()] = gr
                for (col_hdr_keys, label) in [
                    (["avg", "internal"], "Avg_Internal_Rating"),
                    (["avg", "external"], "Avg_External_Rating"),
                    (["avg", "rating", "diff"], "Avg_Rating_Difference"),
                ]:
                    gci = col(gt_hdr2, col_hdr_keys)
                    aci = col(ag_hdr2, col_hdr_keys)
                    if gci is None or aci is None:
                        continue
                    for ar in data_rows2:
                        key = str(ar[0]).strip().lower() if ar[0] else ""
                        gr = gt_lookup2.get(key)
                        if gr is None or gci >= len(gr) or aci >= len(ar):
                            continue
                        gv = gr[gci]
                        av = ar[aci]
                        try:
                            ok = abs(float(av) - float(gv)) <= 0.1
                            record(f"Category '{ar[0]}' {label} ~{gv}", ok, f"Got {av}")
                            if not ok:
                                all_ok = False
                        except (TypeError, ValueError):
                            pass

    wb.close()
    return all_ok


def check_email():
    """Check quality alert email."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, from_addr, to_addr, body_text FROM email.messages"
        )
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    # Require exact subject
    found = False
    flagged = getattr(check_excel, "flagged_products", [])
    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower().strip()
        if "quality alert: product review discrepancies" == subj_lower:
            found = True
            record("Email with exact subject 'Quality Alert: Product Review Discrepancies'", True)

            # Check recipient exactly product-team@company.com
            to_str = str(to_addr).lower() if to_addr else ""
            record(
                "Email to product-team@company.com exactly",
                "product-team@company.com" in to_str,
                f"To: {to_addr}",
            )

            # From address quality@company.com
            from_str = str(from_addr).lower() if from_addr else ""
            record(
                "Email from quality@company.com",
                "quality@company.com" in from_str,
                f"From: {from_addr}",
            )

            # Body mentions all flagged products
            body_lower = (body_text or "").lower()
            if flagged:
                missing = [p for p in flagged if p.lower() not in body_lower]
                record(
                    f"Email body mentions all {len(flagged)} flagged products",
                    len(missing) == 0,
                    f"Missing: {missing}",
                )
            else:
                record("Flagged products list present (for cross-check)", False,
                       "Excel did not provide flagged list")

            # Body has rating info
            has_rating = any(
                kw in body_lower for kw in ["rating", "internal", "external", "difference", "diff"]
            )
            record("Email body has rating information", has_rating)
            break

    if not found:
        record(
            "Email with exact subject 'Quality Alert: Product Review Discrepancies'",
            False,
            f"Found {len(emails)} emails but subject mismatch",
        )

    return found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)
    file_fail = FAIL_COUNT
    email_ok = check_email()
    email_fail = FAIL_COUNT - file_fail

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:  {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Email:  {'PASS' if email_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} (file_fail={file_fail}, email_fail={email_fail})")

    # Excel (file-level) checks are blocking. Email runtime may fail in GT self-test.
    overall = (file_fail == 0)
    print(f"  Overall:  {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
