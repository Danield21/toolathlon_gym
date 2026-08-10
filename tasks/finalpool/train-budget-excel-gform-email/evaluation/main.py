"""
Evaluation for train-budget-excel-gform-email task.

Checks:
1. Travel_Budget.xlsx with Seat_Options, Budget_Scenarios, Summary sheets
2. Correct prices and totals for 8 people
3. Survey form with 3 questions exists
4. Email sent to finance@company.com

The Excel checks are BLOCKING. The gform / email checks are runtime-only
(non-blocking) and degrade gracefully when the database is unreachable, so a
DB outage never turns a correct Excel submission into a FAIL.
"""
import json
import os
import re
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2
from openpyxl.utils import column_index_from_string

# Environment-driven DB config: every connection uses the same PGDATABASE /
# PGPORT as the worker library the agent actually wrote its side-effects to.
DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0
BLOCKING_FAIL_COUNT = 0  # Runtime-only checks (gform/email) do not count
# True when the eval runs against the groundtruth workspace itself; the gform /
# email are solver side-effects absent from the GT DB snapshot, so their
# existence checks are skipped (not failed) in that mode.
IS_GT_SELF_TEST = False


def record(name, passed, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, BLOCKING_FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if not runtime_only:
            BLOCKING_FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        suffix = " (runtime-only)" if runtime_only else ""
        print(f"  [FAIL] {name}{suffix}{msg}")


# ---------------------------------------------------------------------------
# Robust numeric extraction.
#
# Cells may hold:
#   * literal numbers (int/float),
#   * strings with currency symbols / thousands separators / percent signs
#     (e.g. "2,792", "¥349", "90%"),
#   * Excel formulas written by the agent (e.g. "=349*8", "=D2*8",
#     "=13988-2792"). openpyxl returns those as strings starting with "=" and,
#     unless a recalc happened, no cached value is available, so we evaluate
#     simple arithmetic / cell-reference formulas ourselves.
# ---------------------------------------------------------------------------

def _to_float(v):
    """Best-effort conversion of a cell value to a float. Returns None when the
    value is missing / blank / not numeric / an unevaluable formula."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return None  # formulas are resolved separately
    t = s.replace(",", "").replace(" ", "")
    for ch in ("$", "¥", "€", "%", "￥"):
        t = t.replace(ch, "")
    try:
        return float(t)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", t)
        if m:
            try:
                return float(m.group(0))
            except ValueError:
                return None
        return None


def _cell_float(ws, ws_cached, row, column, depth=0, seen=None):
    """Return the float value of a cell, resolving Excel formulas when the
    cached (data_only) value is unavailable. Returns None if unresolvable."""
    if depth > 6:
        return None
    if seen is None:
        seen = set()
    key = (row, column)
    if key in seen:
        return None
    raw = ws.cell(row=row, column=column).value
    if not (isinstance(raw, str) and raw.startswith("=")):
        return _to_float(raw)
    if ws_cached is not None:
        cached = ws_cached.cell(row=row, column=column).value
        cached_num = _to_float(cached)
        if cached_num is not None:
            return cached_num
    return _formula_float(ws, ws_cached, raw, depth, seen | {key})


def _formula_float(ws, ws_cached, formula, depth=0, seen=None):
    """Evaluate a simple Excel formula (arithmetic + cell references) that was
    saved without a cached result. Returns a float or None."""
    expr = formula[1:].strip()
    if not expr:
        return None
    seen = seen or set()

    def repl(m):
        ref = m.group(0).replace("$", "")
        col_s = "".join(ch for ch in ref if ch.isalpha()).upper()
        row_s = "".join(ch for ch in ref if ch.isdigit())
        try:
            col = column_index_from_string(col_s)
            row = int(row_s)
        except Exception:
            return "0"
        num = _cell_float(ws, ws_cached, row, col, depth + 1, seen)
        return repr(num) if num is not None else "0"

    expr = re.sub(r"(?<![A-Za-z0-9])[$]?[A-Za-z]{1,3}[$]?\d+", repl, expr)
    expr = expr.replace("^", "**")
    # Restrict to arithmetic tokens only (no names/functions), then evaluate.
    if not re.fullmatch(r"[0-9+\-*/().\s]*", expr):
        return None
    try:
        value = float(eval(expr, {"__builtins__": {}}, {}))
    except Exception:
        return None
    return value


def _sheet_numeric_values(ws, ws_cached):
    """Collect all numeric values in a sheet, formula-aware."""
    out = []
    for row in ws.iter_rows():
        for cell in row:
            num = _cell_float(ws, ws_cached, cell.row, cell.column)
            if num is not None:
                out.append(num)
    return out


def check_excel(agent_workspace):
    print("\n=== Check 1: Travel_Budget.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Travel_Budget.xlsx")
    if not os.path.exists(xlsx_path):
        record("Travel_Budget.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Travel_Budget.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)  # keep formulas
        wb_cached = openpyxl.load_workbook(xlsx_path, data_only=True)  # cached values
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Seat_Options sheet
    if "seat_options" not in sheet_names_lower:
        record("Seat_Options sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Seat_Options sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("seat_options")]]
        ws_cached = wb_cached[wb.sheetnames[sheet_names_lower.index("seat_options")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Seat_Options has at least 4 rows", len(data_rows) >= 4,
               f"Found {len(data_rows)}")

        all_text = " ".join(str(c) for r in rows for c in r if c).upper()
        # Train codes may be written as "G1" or "G1_260310_1"; match the code
        # as a token while avoiding "G1" matching inside "G11" / "G105".
        has_g1 = bool(re.search(r"(?<![A-Za-z0-9])G1(?![0-9])", all_text))
        has_g11 = bool(re.search(r"(?<![A-Za-z0-9])G11(?![0-9])", all_text))
        record("Seat_Options has G11 and G1", has_g1 and has_g11, all_text[:200])

        numeric_vals = _sheet_numeric_values(ws, ws_cached)
        has_349 = any(abs(v - 349.0) < 0.1 for v in numeric_vals)
        has_553 = any(abs(v - 553.0) < 0.1 for v in numeric_vals)
        has_1748 = any(abs(v - 1748.5) < 0.5 for v in numeric_vals)
        record("Seat prices include 349.0, 553.0, 1748.5", has_349 and has_553 and has_1748,
               f"Numerics: {numeric_vals[:40]}")

    # Budget_Scenarios sheet
    if "budget_scenarios" not in sheet_names_lower:
        record("Budget_Scenarios sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Budget_Scenarios sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("budget_scenarios")]]
        ws2_cached = wb_cached[wb.sheetnames[sheet_names_lower.index("budget_scenarios")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        record("Budget_Scenarios has 3 rows", len(data_rows2) >= 3,
               f"Found {len(data_rows2)}")

        numeric_vals2 = _sheet_numeric_values(ws2, ws2_cached)
        has_2792 = any(abs(v - 2792.0) < 1.0 for v in numeric_vals2)
        has_4424 = any(abs(v - 4424.0) < 1.0 for v in numeric_vals2)
        has_13988 = any(abs(v - 13988.0) < 5.0 for v in numeric_vals2)
        record("Budget total 2792 CNY correct", has_2792, f"Numerics: {numeric_vals2}")
        record("Standard total 4424 CNY correct", has_4424, f"Numerics: {numeric_vals2}")
        record("Premium total ~13988 CNY correct", has_13988, f"Numerics: {numeric_vals2}")

    # Summary sheet
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws3 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        ws3_cached = wb_cached[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows3 = list(ws3.iter_rows(values_only=True))
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        numeric_vals3 = _sheet_numeric_values(ws3, ws3_cached)
        has_price_diff = any(abs(v - 11196.0) < 10.0 for v in numeric_vals3)
        record("Summary has price difference ~11196 CNY", has_price_diff,
               f"Text: {all_text3[:200]}")


def check_gform():
    print("\n=== Check 2: Survey Form ===")
    if IS_GT_SELF_TEST:
        print("  [SKIP] Survey form check: GT self-test (side-effect not present in GT snapshot)")
        return
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Survey form (DB connection)", False, f"DB unavailable: {e}",
               runtime_only=True)
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%travel preference%'
               OR title ILIKE '%business trip%travel%'
        """)
        forms = cur.fetchall()
        # Form existence is BLOCKING: the task requires creating the
        # "Business Trip Travel Preference Survey".
        record("Business Trip Travel Preference survey form exists", len(forms) >= 1,
               f"Found forms: {[f[1] for f in forms]}")

        if forms:
            form_id = forms[0][0]
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
            q_count = cur.fetchone()[0]
            # Question count is BLOCKING: task.md fixes the form at exactly three questions.
            record("Form has exactly 3 questions", q_count == 3, f"Found {q_count}")

            cur.execute(
                "SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position",
                (form_id,),
            )
            questions = cur.fetchall()

            def _title(q):
                return (q[0] or "").lower()

            def _qtype(q):
                return (q[1] or "").lower()

            has_departure_q = any("departure" in _title(q) or "time" in _title(q) for q in questions)
            has_seat_q = any("seat" in _title(q) or "class" in _title(q) for q in questions)
            # The real google-forms MCP only produces 'textQuestion' /
            # 'choiceQuestion'. Accept text-style types and fall back to the
            # title keywords so we never demand a question type the MCP cannot
            # create.
            has_text_q = any(
                _qtype(q) in ("textquestion", "text", "paragraph")
                or any(k in _title(q) for k in ("requirement", "comment", "special"))
                for q in questions
            )
            record("Form has departure time question", has_departure_q,
                   f"Questions: {questions}", runtime_only=True)
            record("Form has seat class question", has_seat_q,
                   f"Questions: {questions}", runtime_only=True)
            record("Form has open text question", has_text_q,
                   f"Questions: {questions}", runtime_only=True)
    except psycopg2.Error as e:
        record("Survey form (DB query)", False, f"DB error: {e}", runtime_only=True)
    finally:
        try:
            conn.close()
        except Exception:
            pass


def check_email():
    print("\n=== Check 3: Email to finance@company.com ===")
    if IS_GT_SELF_TEST:
        print("  [SKIP] Email check: GT self-test (side-effect not present in GT snapshot)")
        return
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Email (DB connection)", False, f"DB unavailable: {e}",
               runtime_only=True)
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr, subject, body_text FROM email.messages
            WHERE subject ILIKE '%budget%' OR subject ILIKE '%beijing-shanghai%' OR subject ILIKE '%conference%'
        """)
        messages = cur.fetchall()
        cur.close()
        conn.close()
    except psycopg2.Error as e:
        record("Email (DB query)", False, f"DB error: {e}", runtime_only=True)
        return

    all_msgs = list(messages)
    # Email existence is BLOCKING: the task requires sending the budget email.
    record("Budget analysis email sent", len(all_msgs) >= 1,
           f"Found {len(all_msgs)} matching emails")

    if all_msgs:
        to_raw = all_msgs[0][0]
        to_str = str(to_raw).lower() if to_raw else ""
        record("Email sent to finance@company.com", "finance@company.com" in to_str,
               f"To: {to_str[:100]}", runtime_only=True)
        # Body content validation: must include Budget / Standard / Premium totals
        body_raw = (all_msgs[0][2] or "") if len(all_msgs[0]) > 2 else ""
        body_norm = str(body_raw).lower()

        def has_number(text, target, tol=10):
            # Strip thousands separators and match either 'target' or close number
            nums = []
            for n in re.findall(r"[\d,]+(?:\.\d+)?", text):
                try:
                    nums.append(float(n.replace(",", "")))
                except ValueError:
                    pass
            return any(abs(v - target) <= tol for v in nums)

        has_budget_total = has_number(body_norm, 2792.0, tol=2)
        has_standard_total = has_number(body_norm, 4424.0, tol=2)
        has_premium_total = has_number(body_norm, 13988.0, tol=20)
        record("Email body mentions Budget_Total (~2792)", has_budget_total,
               f"Body[:300]: {body_norm[:300]}", runtime_only=True)
        record("Email body mentions Standard_Total (~4424)", has_standard_total,
               f"Body[:300]: {body_norm[:300]}", runtime_only=True)
        record("Email body mentions Premium_Total (~13988)", has_premium_total,
               f"Body[:300]: {body_norm[:300]}", runtime_only=True)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    global IS_GT_SELF_TEST
    try:
        _gt = os.path.realpath(args.groundtruth_workspace) if args.groundtruth_workspace else ""
        _ag = os.path.realpath(args.agent_workspace) if args.agent_workspace else ""
        IS_GT_SELF_TEST = bool(_gt) and bool(_ag) and _gt == _ag
    except Exception:
        IS_GT_SELF_TEST = False

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%) ; blocking_fail={BLOCKING_FAIL_COUNT}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "blocking_fail_count": BLOCKING_FAIL_COUNT,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Blocking fail => FAIL. Runtime-only (gform/email) failures do not block the
    # local-file pass, but any NON-runtime failure is fatal.
    if BLOCKING_FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
