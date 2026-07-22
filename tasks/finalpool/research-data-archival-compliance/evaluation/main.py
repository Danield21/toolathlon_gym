#!/usr/bin/env python3
"""Evaluation script for research-data-archival-compliance.

Verifies the 6 phases of the task:
  - Archival registry lists 3 datasets with required metadata
  - Compliance certificate (Word) covers funders + retention rules
  - Compliance email sent to institutional leadership
  - Calendar event scheduled for the review meeting
"""

from argparse import ArgumentParser
import csv
import json
import os
import sys
from pathlib import Path
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

EXPECTED_DATASETS = ["NSF-2021-CLIM-001", "NIH-2022-HLTH-002", "DOE-2023-ENRG-003"]
EXPECTED_FUNDERS = {"NSF", "NIH", "DOE"}
LEADERSHIP_EMAIL = "research_leadership@institution.org"

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, cond, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if cond:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _read_text_any(path):
    """Best-effort read from .docx, .txt, .csv, .xlsx, .json — return concatenated text."""
    p = Path(path)
    if not p.exists():
        return ""
    try:
        ext = p.suffix.lower()
        if ext == ".docx":
            try:
                from docx import Document
                doc = Document(str(p))
                return "\n".join(par.text for par in doc.paragraphs)
            except Exception:
                return ""
        if ext == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(p), data_only=True)
                txts = []
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell is not None:
                                txts.append(str(cell))
                return " ".join(txts)
            except Exception:
                return ""
        if ext in (".txt", ".md", ".csv", ".json", ".log"):
            return p.read_text(errors="ignore")
    except Exception:
        return ""
    return ""


def _glob_text(workspace, patterns):
    """Aggregate text across multiple file patterns under workspace (recursive)."""
    text = []
    base = Path(workspace)
    for pat in patterns:
        for hit in base.rglob(pat):
            text.append(_read_text_any(str(hit)))
    return "\n".join(text)


def find_registry_file(workspace):
    """Find an archival registry (csv/xlsx/json/txt) by name."""
    base = Path(workspace)
    candidates = []
    for pat in ["*registry*", "*inventory*", "*archival*", "*dataset*", "*compliance_registry*"]:
        for ext in ["csv", "xlsx", "json", "tsv", "md"]:
            candidates.extend(base.rglob(f"{pat}.{ext}"))
    # de-duplicate, prefer top-level
    seen = []
    for c in candidates:
        if c not in seen:
            seen.append(c)
    return seen


def check_registry(workspace):
    """Verify archival registry exists and contains 3 datasets with metadata."""
    files = find_registry_file(workspace)
    check("Archival registry file exists",
          len(files) >= 1,
          f"searched: registry/inventory/archival/dataset matching csv/xlsx/json — found {len(files)} files")
    if not files:
        return
    text_combined = "\n".join(_read_text_any(str(f)) for f in files)
    text_lower = text_combined.lower()
    # All 3 datasets present
    missing_datasets = [d for d in EXPECTED_DATASETS if d.lower() not in text_lower]
    check("Registry lists all 3 expected datasets",
          len(missing_datasets) == 0,
          f"missing: {missing_datasets}")
    # All 3 funders mentioned
    missing_funders = [f for f in EXPECTED_FUNDERS if f.lower() not in text_lower]
    check("Registry mentions all 3 funders (NSF, NIH, DOE)",
          len(missing_funders) == 0,
          f"missing: {missing_funders}")
    # Required metadata field tokens
    md_tokens = ["retention", "format", "size"]
    missing_md = [t for t in md_tokens if t not in text_lower]
    check("Registry includes retention/format/size metadata terms",
          len(missing_md) == 0,
          f"missing: {missing_md}")


def check_compliance_doc(workspace):
    """Verify compliance certificate / plan (Word doc preferred) covers funders + retention."""
    candidates = []
    base = Path(workspace)
    for pat in ["*compliance*", "*certif*", "*archival_plan*"]:
        for ext in ["docx", "txt", "md", "pdf"]:
            candidates.extend(base.rglob(f"{pat}.{ext}"))
    check("Compliance document exists (compliance/certificate/archival_plan)",
          len(candidates) >= 1,
          f"found {len(candidates)} files")
    if not candidates:
        return
    text = "\n".join(_read_text_any(str(c)) for c in candidates).lower()
    check("Compliance doc has substantive content (>=400 chars)",
          len(text) >= 400,
          f"length={len(text)}")
    # Mentions all funders
    missing_funders = [f for f in EXPECTED_FUNDERS if f.lower() not in text]
    check("Compliance doc mentions all funders (NSF/NIH/DOE)",
          len(missing_funders) == 0,
          f"missing: {missing_funders}")
    # Mentions retention + sensitive/PHI handling + preservation
    must_have = ["retention", "preservation", "sensitive"]
    missing = [k for k in must_have if k not in text]
    check("Compliance doc covers retention/preservation/sensitive topics",
          len(missing) == 0,
          f"missing: {missing}")


def check_email():
    """Verify compliance verification email sent to leadership."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, COALESCE(body_text, body_html, '')
            FROM email.messages
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        check("Email DB reachable", False, str(e))
        return
    check("At least one email message exists",
          len(rows) >= 1, f"got {len(rows)}")
    if not rows:
        return

    def to_list(to_addr):
        if to_addr is None:
            return []
        if isinstance(to_addr, list):
            return [str(t).lower() for t in to_addr]
        if isinstance(to_addr, str):
            try:
                p = json.loads(to_addr)
                if isinstance(p, list):
                    return [str(t).lower() for t in p]
            except Exception:
                pass
            return [to_addr.lower()]
        return [str(to_addr).lower()]

    matched = None
    for subj, to_addr, body in rows:
        tos = to_list(to_addr)
        subj_l = (subj or "").lower()
        body_l = (body or "").lower()
        # leadership recipient and compliance topic
        if any("leadership" in t or "compliance" in t for t in tos) or any("compliance" in s for s in [subj_l]):
            if any(k in subj_l + body_l for k in ["compliance", "archival", "archive"]):
                matched = (subj, to_addr, body)
                break
    check("Compliance verification email sent (subject/body mentions compliance or archival)",
          matched is not None,
          "no email with compliance/archival in subject")
    if matched:
        body_l = (matched[2] or "").lower()
        funders_in_body = sum(1 for f in EXPECTED_FUNDERS if f.lower() in body_l)
        check("Compliance email body mentions >=2 funder agencies",
              funders_in_body >= 2,
              f"found {funders_in_body} funders in body")


def check_calendar(launch_time):
    """Verify a meeting scheduled with research teams about data organization/compliance."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, summary, COALESCE(description, ''), start_datetime, attendees
            FROM gcal.events
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        check("Calendar DB reachable", False, str(e))
        return
    check("At least one calendar event exists",
          len(rows) >= 1, f"got {len(rows)}")
    if not rows:
        return
    matched = None
    for _id, summary, desc, _start, attendees in rows:
        text = ((summary or "") + " " + (desc or "")).lower()
        if any(k in text for k in ["compliance", "archival", "data review", "research data"]):
            matched = (summary, desc, attendees)
            break
    check("Calendar event scheduled for compliance/archival/data review meeting",
          matched is not None,
          f"checked {len(rows)} events, none matched compliance/archival keywords")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0
    if not agent_workspace or not Path(agent_workspace).exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    # GT self-test toleration: when running with agent==groundtruth (V1 GT-only
    # smoke test), the email/gcal checks would fail because the agent never
    # actually sent the email or scheduled the meeting. Skip them.
    is_gt_self_test = (
        agent_workspace and groundtruth_workspace
        and os.path.abspath(agent_workspace) == os.path.abspath(groundtruth_workspace)
    )

    check_registry(agent_workspace)
    check_compliance_doc(agent_workspace)
    if not is_gt_self_test:
        check_email()
        check_calendar(launch_time)
    else:
        print("  [SKIP] email and calendar checks (GT self-test mode)")

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
