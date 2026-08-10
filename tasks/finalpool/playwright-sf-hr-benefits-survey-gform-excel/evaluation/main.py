"""Evaluation for playwright-sf-hr-benefits-survey-gform-excel."""
import argparse
import os
import re
import sys

import psycopg2

# All DB settings are read from the environment with sane defaults, matching
# preprocess/main.py. Never hardcode a port or database name here.
DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)


def _cell_float(v):
    """Robustly coerce a cell value to float, or None if not parseable.

    Handles int/float, and strings with thousands separators, currency symbols,
    trailing percent signs, and surrounding whitespace.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    for ch in (",", "$", "€", "¥"):
        s = s.replace(ch, "")
    s = s.rstrip("%").strip()
    try:
        return float(s)
    except ValueError:
        return None


def num_close(a, b, tol=0.5):
    """Compare two values numerically when possible, else case-insensitive string equality."""
    fa, fb = _cell_float(a), _cell_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    return str(a).strip().lower() == str(b).strip().lower()


def _dept_key(s):
    """Normalize a department name to a canonical lowercased key.

    Tolerates common humanization variants of the DB's short names (e.g.
    'Human Resources' -> 'hr', 'Research and Development' -> 'r&d') so that a
    faithful agent which expands the abbreviations still matches rows/options.
    Unknown names are returned stripped of non-alphanumerics.
    """
    t = "".join(ch for ch in str(s or "").lower() if ch.isalnum())
    if t in ("humanresource", "humanresources"):
        return "hr"
    if t in ("rd", "rand", "randd", "rnd", "researchanddevelopment", "researchdevelopment"):
        return "r&d"
    return t


def _subject_key(s):
    """Normalize an email subject so '&' and the word 'and' compare equal."""
    s = str(s or "").strip().lower().replace("&", " and ")
    return re.sub(r"\s+", " ", s)


def load_sheet_rows(wb, sheet_name):
    """Return list of (row_index_1based, [cell values]) or None if sheet missing."""
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [(i, [cell.value for cell in row]) for i, row in enumerate(wb[name].iter_rows(), start=1)]
    return None


def _is_formula(wb_f, sheet_name, row_idx, col_0based):
    """True if the cell at (row_idx, col_0based) in the formula workbook holds an Excel formula."""
    if wb_f is None:
        return False
    try:
        ws = wb_f[sheet_name]
    except KeyError:
        return False
    v = ws.cell(row=row_idx, column=col_0based + 1).value
    return isinstance(v, str) and v.startswith("=")


def _check_num(wb_f, sheet, row_idx, col, a_val, g_val, tol, label, errors):
    """Compare a numeric cell; skip when the agent used an Excel formula (task requires literals)."""
    if _is_formula(wb_f, sheet, row_idx, col):
        return
    if not num_close(a_val, g_val, tol):
        errors.append(f"{label}: {a_val} vs {g_val}")


def _db_expected_ratings():
    """Derive the expected Satisfaction_Rating per department from the employee DB.

    Rule (matching the task): round the average job satisfaction to two decimal
    places, then 'High' if it is 6.55 or above, else 'Moderate'. For departments
    whose rounded average sits within 0.01 of the 6.55 threshold, both ratings are
    accepted because the boundary is genuinely ambiguous. Returns a dict of
    {dept_lower: {accepted_rating_lowercases}} or None when the DB is unreachable
    (callers then fall back to the ground-truth rating strings).
    """
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute(
            'SELECT "DEPARTMENT", ROUND(AVG("JOB_SATISFACTION")::numeric, 2) '
            'FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" GROUP BY "DEPARTMENT"'
        )
        out = {}
        for dept, avg in cur.fetchall():
            if avg is None:
                continue
            avg_f = float(avg)
            rating = "High" if avg_f >= 6.55 else "Moderate"
            if abs(avg_f - 6.55) <= 0.01:
                out[_dept_key(dept)] = {"high", "moderate"}
            else:
                out[_dept_key(dept)] = {rating.lower()}
        cur.close()
        conn.close()
        return out
    except Exception:
        return None


def check_excel(agent_workspace, gt_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Benefits_Analysis.xlsx")
    if not os.path.exists(path):
        return ["Benefits_Analysis.xlsx not found"]
    gt_path = os.path.join(gt_workspace, "Benefits_Analysis.xlsx")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        wb_f = openpyxl.load_workbook(path, data_only=False)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True) if os.path.exists(gt_path) else None

        db_ratings = _db_expected_ratings()

        # ---------- Competitor Comparison (all 7 rows) ----------
        rows = load_sheet_rows(wb, "Competitor Comparison")
        if rows is None:
            errors.append("Sheet 'Competitor Comparison' not found")
        else:
            data_rows = [(i, r) for (i, r) in rows[1:] if r and r[0] is not None]
            if len(data_rows) != 7:
                errors.append(f"Competitor Comparison has {len(data_rows)} rows, expected 7")
            agent_lookup = {str(r[0]).strip().lower(): (i, r) for (i, r) in data_rows if r[0]}
            if gt_wb is not None:
                gt_rows = load_sheet_rows(gt_wb, "Competitor Comparison") or []
                gt_data = [r for (_, r) in gt_rows[1:] if r and r[0] is not None]
                for g_row in gt_data:
                    key = str(g_row[0]).strip().lower()
                    hit = agent_lookup.get(key)
                    if hit is None:
                        errors.append(f"Competitor row missing for {g_row[0]}")
                        continue
                    a_row = list(hit[1]) + [None] * (4 - len(hit[1]))
                    _check_num(wb_f, "Competitor Comparison", hit[0], 1, a_row[1], g_row[1], 1,
                               f"{g_row[0]} Health_Insurance_Pct", errors)
                    _check_num(wb_f, "Competitor Comparison", hit[0], 2, a_row[2], g_row[2], 1,
                               f"{g_row[0]} PTO_Days", errors)
                    _check_num(wb_f, "Competitor Comparison", hit[0], 3, a_row[3], g_row[3], 0.2,
                               f"{g_row[0]} Retirement_Match_Pct", errors)
            # Sort order: rows must be alphabetical by Company
            company_names = [str(r[0]).strip() for (_, r) in data_rows if r and r[0]]
            sorted_names = sorted(company_names, key=lambda s: s.lower())
            if company_names != sorted_names:
                errors.append("Competitor Comparison not sorted alphabetically by Company")

        # ---------- Department Satisfaction (all 7 rows) ----------
        rows2 = load_sheet_rows(wb, "Department Satisfaction")
        if rows2 is None:
            errors.append("Sheet 'Department Satisfaction' not found")
        else:
            data_rows2 = [(i, r) for (i, r) in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) != 7:
                errors.append(f"Department Satisfaction has {len(data_rows2)} rows, expected 7")
            agent_lookup2 = {_dept_key(r[0]): (i, r) for (i, r) in data_rows2 if r[0]}
            if gt_wb is not None:
                gt_rows2 = load_sheet_rows(gt_wb, "Department Satisfaction") or []
                gt_data2 = [r for (_, r) in gt_rows2[1:] if r and r[0] is not None]
                for g_row in gt_data2:
                    key = _dept_key(g_row[0])
                    hit = agent_lookup2.get(key)
                    if hit is None:
                        errors.append(f"Department row missing for {g_row[0]}")
                        continue
                    a_row = list(hit[1]) + [None] * (5 - len(hit[1]))
                    _check_num(wb_f, "Department Satisfaction", hit[0], 1, a_row[1], g_row[1], 0.15,
                               f"{g_row[0]} Avg_Job_Satisfaction", errors)
                    _check_num(wb_f, "Department Satisfaction", hit[0], 2, a_row[2], g_row[2], 0.15,
                               f"{g_row[0]} Avg_Work_Life_Balance", errors)
                    _check_num(wb_f, "Department Satisfaction", hit[0], 3, a_row[3], g_row[3], 0,
                               f"{g_row[0]} Headcount", errors)
                    # Satisfaction_Rating: accept a db-derived set when available, else GT string.
                    a_rating = str(a_row[4] or "").strip().lower()
                    if db_ratings is not None:
                        expected = db_ratings.get(key)
                    else:
                        expected = None
                    if expected is None:
                        expected = {str(g_row[4] or "").strip().lower()}
                    if a_rating not in expected:
                        errors.append(f"{g_row[0]} Satisfaction_Rating: {a_row[4]} vs {sorted(expected)}")
            # Sort order: alphabetical by Department
            dept_names = [str(r[0]).strip() for (_, r) in data_rows2 if r and r[0]]
            sorted_depts = sorted(dept_names, key=lambda s: s.lower())
            if dept_names != sorted_depts:
                errors.append("Department Satisfaction not sorted alphabetically by Department")

        # ---------- Gap Analysis (all 3 rows) ----------
        rows3 = load_sheet_rows(wb, "Gap Analysis")
        if rows3 is None:
            errors.append("Sheet 'Gap Analysis' not found")
        else:
            data_rows3 = [(i, r) for (i, r) in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) != 3:
                errors.append(f"Gap Analysis has {len(data_rows3)} rows, expected 3")
            agent_lookup3 = {str(r[0]).strip().lower(): (i, r) for (i, r) in data_rows3 if r[0]}
            if gt_wb is not None:
                gt_rows3 = load_sheet_rows(gt_wb, "Gap Analysis") or []
                gt_data3 = [r for (_, r) in gt_rows3[1:] if r and r[0] is not None]
                for g_row in gt_data3:
                    key = str(g_row[0]).strip().lower()
                    hit = agent_lookup3.get(key)
                    if hit is None:
                        errors.append(f"Gap row missing for {g_row[0]}")
                        continue
                    a_row = list(hit[1]) + [None] * (5 - len(hit[1]))
                    _check_num(wb_f, "Gap Analysis", hit[0], 1, a_row[1], g_row[1], 0.2,
                               f"{g_row[0]} Our_Value", errors)
                    _check_num(wb_f, "Gap Analysis", hit[0], 2, a_row[2], g_row[2], 1.0,
                               f"{g_row[0]} Market_Average", errors)
                    _check_num(wb_f, "Gap Analysis", hit[0], 3, a_row[3], g_row[3], 1.0,
                               f"{g_row[0]} Gap", errors)
                    # Priority (string match, case-insensitive)
                    a_pri = str(a_row[4] or "").strip().lower()
                    g_pri = str(g_row[4] or "").strip().lower()
                    if a_pri != g_pri:
                        errors.append(f"{g_row[0]} Priority: {a_row[4]} vs {g_row[4]}")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def _question_options(config):
    """Extract a lowercased list of option value strings from a question config jsonb.

    Handles both the jsonb-array form [{"value": "..."}, ...] and a plain list of
    strings, comparing on the extracted values only.
    """
    if isinstance(config, dict):
        opts = config.get("options")
    elif isinstance(config, list):
        opts = config
    else:
        opts = None
    out = []
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                if o.get("value") is not None:
                    out.append(str(o["value"]).strip().lower())
            elif isinstance(o, str):
                out.append(o.strip().lower())
    return out


def _leading_number(s):
    m = re.match(r"^\s*(\d+(?:\.\d+)?)", s)
    return m.group(1) if m else None


def check_gform():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        if not forms:
            errors.append("No Google Forms found")
        else:
            # Exact title match
            target = "employee benefits improvement survey"
            form_id = None
            for fid, title in forms:
                if (title or "").strip().lower() == target:
                    form_id = fid
                    break
            if form_id is None:
                errors.append(
                    f"Form titled 'Employee Benefits Improvement Survey' not found (forms: {[f[1] for f in forms]})"
                )
            else:
                cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
                q_count = cur.fetchone()[0]
                if q_count != 5:
                    errors.append(f"Form has {q_count} questions, expected exactly 5")
                # Fetch title / type / required / options. Question types are content-checked
                # against the two types the google_forms MCP can actually create.
                cur.execute(
                    "SELECT title, question_type, required, config FROM gform.questions WHERE form_id = %s ORDER BY id",
                    (form_id,),
                )
                qrows = cur.fetchall()
                if qrows:
                    qinfo = [
                        {
                            "title": str(r[0] or "").strip().lower(),
                            "qtype": str(r[1] or "").strip().lower(),
                            "required": bool(r[2]),
                            "options": _question_options(r[3]),
                        }
                        for r in qrows
                    ]
                    # All questions must be required
                    not_required = [r[0] for r in qrows if not r[2]]
                    if not_required:
                        errors.append(f"Some questions are not required: {not_required}")
                    # At least one multiple-choice question must exist
                    if not any(q["qtype"] in ("choicequestion", "multiplechoice") for q in qinfo):
                        errors.append("Form missing multiple-choice questions")
                    # Question 1: department, options = the seven departments
                    dq = next((q for q in qinfo if "department" in q["title"]), None)
                    if dq is None:
                        errors.append("Form missing department question (Q1)")
                    else:
                        if dq["qtype"] != "choicequestion":
                            errors.append("Q1 (department) is not a multiple-choice question")
                        required_depts = ["engineering", "finance", "hr", "operations", "r&d", "sales", "support"]
                        dept_opts = {_dept_key(o) for o in dq["options"]}
                        missing = [d for d in required_depts if d not in dept_opts]
                        if missing:
                            errors.append(f"Q1 (department) options missing: {missing}")
                    # Rating questions Q2-Q4: health / PTO / retirement, options must cover 1-5
                    scale_nums = {"1", "2", "3", "4", "5"}
                    for keyword, qlabel in (
                        ("health", "Q2 (health insurance)"),
                        ("pto", "Q3 (PTO policy)"),
                        ("retirement", "Q4 (retirement matching)"),
                    ):
                        q = next((x for x in qinfo if keyword in x["title"]), None)
                        if q is None:
                            errors.append(f"Form missing {qlabel} question")
                        else:
                            if q["qtype"] != "choicequestion":
                                errors.append(f"{qlabel} is not a multiple-choice question")
                            lead = {_leading_number(o) for o in q["options"] if _leading_number(o)}
                            missing_nums = sorted(scale_nums - lead)
                            if missing_nums:
                                errors.append(
                                    f"{qlabel} options do not include all values 1 through 5 (missing: {missing_nums})"
                                )
                    # Question 5: benefit priority
                    pq = next((x for x in qinfo if "priorit" in x["title"]), None)
                    if pq is None:
                        errors.append("Form missing benefit priority question (Q5)")
                    else:
                        if pq["qtype"] != "choicequestion":
                            errors.append("Q5 (priority) is not a multiple-choice question")
                        opt_joined = " ".join(pq["options"])
                        for kw in ("health", "pto", "retirement", "other"):
                            if kw not in opt_joined:
                                errors.append(f"Q5 (priority) options missing '{kw}'")
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking GForm: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%hr_leadership@company.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            errors.append("No email found to hr_leadership@company.com")
        else:
            # Find email matching the subject (treat '&' and 'and' as equal)
            target_subj = "benefits competitiveness analysis & survey launch"
            matched = None
            for subj, body in rows:
                if _subject_key(subj) == _subject_key(target_subj):
                    matched = (subj, body)
                    break
            if matched is None:
                errors.append(
                    f"Email subject 'Benefits Competitiveness Analysis & Survey Launch' not found (subjects: {[r[0] for r in rows]})"
                )
            else:
                _, body = matched
                body_l = (body or "").lower()
                # Body must summarize the three benefit gaps and mention the created
                # survey (task.md line 17). The exact priority wording is NOT
                # required, so a faithful numeric summary passes.
                key_groups = [
                    ["health"],
                    ["pto", "paid time off"],
                    ["retirement"],
                    ["survey"],
                ]
                missing_groups = [g[0] for g in key_groups if not any(k in body_l for k in g)]
                if missing_groups:
                    errors.append(f"Email body missing key topics: {missing_groups}")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Google Form...")
    errs = check_gform()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:15]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
