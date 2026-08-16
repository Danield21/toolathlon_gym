"""Evaluation script for wc-howtocook-meal-kit-excel-word-email."""
import os
import argparse, json, sys
import openpyxl
import psycopg2


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:300] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def num_close(a, b, abs_tol=0.01, rel_tol=0.0):
    a = safe_float(a)
    b = safe_float(b)
    if a is None or b is None:
        return False
    return abs(a - b) <= max(abs_tol, abs(b) * rel_tol)


def get_sheet(wb, name):
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            return wb[sn]
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    kit_names_in_proposals = []
    kit_proposals_count = 0

    # === Excel ===
    excel_path = os.path.join(agent_workspace, "Meal_Kit_Report.xlsx")
    check("Meal_Kit_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            check("Excel readable", False, str(e))
            wb = None

        if wb is not None:
            # Required sheet names
            for sn in ["Product_Catalog", "Recipe_Collection", "Kit_Proposals", "Summary"]:
                check(f"Sheet '{sn}' exists", get_sheet(wb, sn) is not None,
                      f"sheets: {wb.sheetnames}")

            # Per-sheet column requirements
            sheet_columns = {
                "Product_Catalog": ["product_name", "price", "stock", "category"],
                "Recipe_Collection": ["recipe_name", "category", "difficulty", "ingredient_count"],
                "Kit_Proposals": ["kit_name", "recipe_name", "estimated_cost",
                                  "margin_pct", "recommended_price"],
                "Summary": ["metric", "value"],
            }
            for sn, cols in sheet_columns.items():
                ws = get_sheet(wb, sn)
                if ws is None:
                    continue
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    check(f"{sn} has rows", False)
                    continue
                header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
                for c in cols:
                    check(f"{sn} has column '{c}'",
                          c in header, f"headers: {header}")

            # Sort by Product_Name
            ws = get_sheet(wb, "Product_Catalog")
            if ws is not None:
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                rows = [r for r in rows if r and r[0]]
                names = [str(r[0]).strip().lower() for r in rows]
                check("Product_Catalog sorted by Product_Name",
                      names == sorted(names), f"first few: {names[:5]}")

            # Sort by Recipe_Name
            ws = get_sheet(wb, "Recipe_Collection")
            if ws is not None:
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                rows = [r for r in rows if r and r[0]]
                names = [str(r[0]).strip().lower() for r in rows]
                check("Recipe_Collection sorted by Recipe_Name",
                      names == sorted(names), f"first few: {names[:5]}")

            # Kit_Proposals min 5 rows
            ws = get_sheet(wb, "Kit_Proposals")
            if ws is not None:
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                data = [r for r in rows if r and r[0]]
                kit_proposals_count = len(data)
                check("Kit_Proposals has at least 5 rows", len(data) >= 5,
                      f"got {len(data)}")
                # Each row must have all fields (not None) AND numeric sanity
                for r in data:
                    name = str(r[0]) if r and r[0] else "?"
                    kit_names_in_proposals.append(name)
                    if len(r) >= 5:
                        check(f"Kit '{name}' has Estimated_Cost, Margin_Pct, Recommended_Price",
                              all(v is not None for v in [r[2], r[3], r[4]]),
                              f"row: {r}")
                        # Numeric sanity
                        cost = safe_float(r[2])
                        margin = safe_float(r[3])
                        price = safe_float(r[4])
                        check(f"Kit '{name}' cost > 0",
                              cost is not None and cost > 0,
                              f"cost={r[2]}")
                        check(f"Kit '{name}' 0 < margin_pct < 100",
                              margin is not None and 0 < margin < 100,
                              f"margin={r[3]}")
                        check(f"Kit '{name}' recommended_price > cost",
                              price is not None and cost is not None and price > cost,
                              f"price={r[4]}, cost={r[2]}")

            # Summary required metrics with sane values
            ws = get_sheet(wb, "Summary")
            if ws is not None:
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                data = [r for r in rows if r and r[0]]
                lookup = {str(r[0]).strip().lower(): r[1] if len(r) > 1 else None
                          for r in data}
                for m in ["total_products", "total_recipes", "proposed_kits",
                          "avg_kit_cost", "avg_margin_pct"]:
                    check(f"Summary has '{m}'", m in lookup, f"keys: {list(lookup.keys())}")
                # Value sanity: Proposed_Kits should equal len(Kit_Proposals data rows)
                if "proposed_kits" in lookup:
                    val = safe_float(lookup["proposed_kits"])
                    check(
                        f"Summary Proposed_Kits == Kit_Proposals row count ({kit_proposals_count})",
                        val is not None and int(val) == kit_proposals_count and val >= 5,
                        f"got {lookup['proposed_kits']}, expected {kit_proposals_count}",
                    )

    # === Word doc ===
    docx_path = os.path.join(agent_workspace, "Meal_Kit_Proposal.docx")
    check("Meal_Kit_Proposal.docx exists", os.path.exists(docx_path))
    if os.path.exists(docx_path):
        try:
            from docx import Document
            doc = Document(docx_path)
            text = "\n".join(p.text for p in doc.paragraphs)
            text_lower = text.lower()
            check("Document has substantial content", len(text) > 100,
                  f"length: {len(text)}")
            check("Document title 'Meal Kit Service Launch Proposal'",
                  "meal kit service launch proposal" in text_lower,
                  "")
            for sec in ["Product Analysis", "Recipe Selection", "Kit Design",
                        "Financial Projections"]:
                check(f"Document section '{sec}'",
                      sec.lower() in text_lower, "")
        except Exception as e:
            check("Word doc readable", False, str(e))

    # === Python script: meal_kit_designer.py exact filename ===
    py_path = os.path.join(agent_workspace, "meal_kit_designer.py")
    check("meal_kit_designer.py exists", os.path.exists(py_path),
          f"workspace: {os.listdir(agent_workspace)[:20]}")

    # === JSON files ===
    for jf in ["store_products.json", "recipes.json", "meal_kit_plans.json"]:
        jpath = os.path.join(agent_workspace, jf)
        check(f"{jf} exists", os.path.exists(jpath), jpath)
        if os.path.exists(jpath):
            try:
                with open(jpath) as f:
                    json.load(f)
                check(f"{jf} is valid JSON", True)
            except json.JSONDecodeError as e:
                check(f"{jf} valid JSON", False, str(e))

    # === Email checks ===
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Look for email with exact subject AND recipient
        cur.execute("SELECT subject, from_addr, to_addr, body_text, folder_id FROM email.messages")
        all_emails = cur.fetchall()

        target = None
        for subj, from_addr, to_addr, body, folder_id in all_emails:
            if (subj or "").strip().lower() == "meal kit service proposal ready":
                to_str = json.dumps(to_addr).lower() if isinstance(to_addr, (list, dict)) else str(to_addr or "").lower()
                if "product-team@company.com" in to_str:
                    target = (subj, from_addr, to_addr, body)
                    break
        check("Email with exact subject + recipient sent",
              target is not None,
              f"Subjects: {[(e[0], e[2]) for e in all_emails[:5]]}")
        if target:
            import re
            body = (target[3] or "").lower()
            # Require 'top 3' word-boundary phrase
            top3_match = re.search(r"\btop\s+3\b", body) is not None
            check("Email body mentions 'top 3' (word boundary) and 'kit'",
                  "kit" in body and top3_match,
                  f"body head: {body[:200]}")
            # Both cost AND margin (task says 'estimated costs and margins')
            check("Email body mentions both cost AND margin",
                  "cost" in body and "margin" in body,
                  f"body head: {body[:200]}")

        # Reverse: noise newsletter not in Sent
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE folder_id IN (
                SELECT id FROM email.folders WHERE LOWER(name) LIKE '%sent%'
            )
            AND subject ILIKE '%newsletter%'
        """)
        noise_sent = cur.fetchone()[0]
        check("No noise/newsletter emails in Sent folder", noise_sent == 0,
              f"found {noise_sent}")

        conn.close()
    except Exception as e:
        check("Email DB check error", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": success}, f)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
