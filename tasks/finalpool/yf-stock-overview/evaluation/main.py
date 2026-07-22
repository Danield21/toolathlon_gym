"""Evaluation for yf-stock-overview."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "YF_Stock_Overview.xlsx")
    gt_file = os.path.join(gt_dir, "YF_Stock_Overview.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Check sheet: Stock Overview
    print(f"  Checking Stock Overview...")
    a_rows = load_sheet_rows(agent_wb, "Stock Overview")
    g_rows = load_sheet_rows(gt_wb, "Stock Overview")
    if a_rows is None:
        all_errors.append("Sheet 'Stock Overview' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Stock Overview' not found in groundtruth")
    else:
        sheet_name = "Stock Overview"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Verify row count matches GT exactly (excludes indices/commodities)
        # Filter out empty trailing rows
        a_data_nonempty = [r for r in a_data if r and r[0] is not None]
        g_data_nonempty = [r for r in g_data if r and r[0] is not None]
        if len(a_data_nonempty) != len(g_data_nonempty):
            errors.append(
                f"Stock Overview row count: {len(a_data_nonempty)} vs expected {len(g_data_nonempty)} "
                f"(check: only stocks with sector, exclude indices/commodities)"
            )

        a_lookup = {}
        for row in a_data_nonempty:
            a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data_nonempty:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Name (col 1) - exact match
            if len(a_row) > 1 and len(g_row) > 1:
                if not str_match(a_row[1], g_row[1]):
                    errors.append(f"{key}.Name: '{a_row[1]}' vs '{g_row[1]}'")
            # Sector (col 2) - exact match
            if len(a_row) > 2 and len(g_row) > 2:
                if not str_match(a_row[2], g_row[2]):
                    errors.append(f"{key}.Sector: '{a_row[2]}' vs '{g_row[2]}'")
            # Market_Cap (col 3) - relative tolerance 5% (allows minor live updates)
            if len(a_row) > 3 and len(g_row) > 3:
                gv = g_row[3]
                try:
                    tol_mc = max(1e6, abs(float(gv)) * 0.05)
                except (TypeError, ValueError):
                    tol_mc = 1e6
                if not num_close(a_row[3], gv, tol_mc):
                    errors.append(f"{key}.Market_Cap: {a_row[3]} vs {gv} (tol={tol_mc:.0f})")
            # PE_Ratio (col 4) - relative tolerance 10%
            if len(a_row) > 4 and len(g_row) > 4:
                gv = g_row[4]
                try:
                    tol_pe = max(0.5, abs(float(gv)) * 0.1)
                except (TypeError, ValueError):
                    tol_pe = 0.5
                if not num_close(a_row[4], gv, tol_pe):
                    errors.append(f"{key}.PE_Ratio: {a_row[4]} vs {gv} (tol={tol_pe:.2f})")

        # Check sort order: Market_Cap descending
        try:
            mcaps = []
            for row in a_data_nonempty:
                if len(row) > 3 and row[3] is not None:
                    try:
                        mcaps.append(float(row[3]))
                    except (TypeError, ValueError):
                        pass
            if mcaps != sorted(mcaps, reverse=True):
                errors.append(
                    f"Stock Overview not sorted by Market_Cap descending: {mcaps}"
                )
        except Exception:
            pass

        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:10]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Per-metric tolerances
            if len(a_row) > 1 and len(g_row) > 1:
                gv = g_row[1]
                if key == "total_stocks":
                    # exact integer count
                    if not num_close(a_row[1], gv, 0):
                        errors.append(f"{key}.Value: {a_row[1]} vs {gv} (tol=0)")
                elif key in ("total_market_cap", "largest_market_cap"):
                    try:
                        tol = max(1e6, abs(float(gv)) * 0.05)
                    except (TypeError, ValueError):
                        tol = 1e6
                    if not num_close(a_row[1], gv, tol):
                        errors.append(f"{key}.Value: {a_row[1]} vs {gv} (tol={tol:.0f})")
                elif key == "largest_company":
                    if not str_match(a_row[1], gv):
                        errors.append(f"{key}.Value: '{a_row[1]}' vs '{gv}'")
                else:
                    if not num_close(a_row[1], gv, 0.5):
                        errors.append(f"{key}.Value: {a_row[1]} vs {gv}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check cloud spreadsheet "Stock Watch Dashboard" with "Stocks" tab
    print(f"  Checking cloud spreadsheet 'Stock Watch Dashboard'...")
    try:
        import psycopg2
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=int(os.environ.get("PGPORT", "5432")),
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user=os.environ.get("PGUSER", "eigent"),
            password=os.environ.get("PGPASSWORD", "camel"),
        )
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, title FROM gsheet.spreadsheets
            WHERE LOWER(title) = LOWER('Stock Watch Dashboard')
            """
        )
        ss_rows = cur.fetchall()
        if not ss_rows:
            all_errors.append("Cloud spreadsheet 'Stock Watch Dashboard' not found")
        else:
            ss_id = ss_rows[0][0]
            cur.execute(
                """
                SELECT id, title FROM gsheet.sheets
                WHERE spreadsheet_id = %s AND LOWER(title) = LOWER('Stocks')
                """,
                (ss_id,),
            )
            sheet_rows = cur.fetchall()
            if not sheet_rows:
                all_errors.append(
                    "Cloud spreadsheet 'Stock Watch Dashboard' missing 'Stocks' tab"
                )
            else:
                sheet_id = sheet_rows[0][0]
                # Verify that at least the GT count of data rows are present
                cur.execute(
                    """
                    SELECT COUNT(DISTINCT row_index) FROM gsheet.cells
                    WHERE spreadsheet_id = %s AND sheet_id = %s
                    """,
                    (ss_id, sheet_id),
                )
                row_count = cur.fetchone()[0]
                # Header + 5 stocks = 6 rows
                if row_count < 6:
                    all_errors.append(
                        f"Cloud spreadsheet 'Stocks' tab has {row_count} rows, expected at least 6"
                    )
                else:
                    # Stronger content check: ensure all 5 expected ticker symbols appear in column A
                    expected_tickers = {"GOOGL", "AMZN", "JPM", "XOM", "JNJ"}
                    cur.execute(
                        """
                        SELECT value FROM gsheet.cells
                        WHERE spreadsheet_id = %s AND sheet_id = %s
                          AND col_index = 0
                        """,
                        (ss_id, sheet_id),
                    )
                    col_a_values = {
                        str(r[0]).strip().upper() for r in cur.fetchall()
                        if r[0] is not None
                    }
                    missing = expected_tickers - col_a_values
                    if missing:
                        all_errors.append(
                            f"Cloud spreadsheet 'Stocks' tab missing tickers in column A: {sorted(missing)}"
                        )
                    else:
                        print("    PASS")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking cloud spreadsheet: {e}")

    

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
