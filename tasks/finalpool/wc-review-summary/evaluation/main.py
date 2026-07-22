"""Evaluation for wc-review-summary."""
import argparse
import os
import sys
import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_notion():
    """Check Notion page titled exactly 'Product Review Dashboard' exists
    AND its content body mentions top-rated and lowest-rated product names.
    Returns (ok: bool, detail: str).
    """
    import json as _json
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties::text ILIKE %s
        """, ('%product review dashboard%',))
        rows = cur.fetchall()

        # Find candidate pages whose title field equals 'product review dashboard'.
        candidate_ids = []
        for row in rows:
            props = row[1]
            if isinstance(props, str):
                try:
                    props = _json.loads(props)
                except Exception:
                    continue
            if not isinstance(props, dict):
                continue
            for prop_name, prop_val in props.items():
                if not isinstance(prop_val, dict):
                    continue
                title_arr = prop_val.get("title") if prop_val.get("type") == "title" else prop_val.get("title")
                if title_arr is None:
                    title_arr = prop_val.get("Name") or prop_val.get("name")
                if not isinstance(title_arr, list):
                    continue
                title_text = ""
                for item in title_arr:
                    if isinstance(item, dict):
                        plain = item.get("plain_text")
                        if plain:
                            title_text += plain
                        elif "text" in item and isinstance(item["text"], dict):
                            title_text += item["text"].get("content", "")
                if title_text.strip().lower() == "product review dashboard":
                    candidate_ids.append(row[0])
                    break

        if not candidate_ids:
            cur.close()
            conn.close()
            return False, "no page with exact title 'Product Review Dashboard'"

        # Pull top-rated and lowest-rated product names dynamically.
        cur.execute("""
            SELECT p.name FROM wc.product_reviews r
            JOIN wc.products p ON r.product_id = p.id
            GROUP BY p.name
            HAVING COUNT(r.id) >= 3
            ORDER BY AVG(r.rating) DESC, LOWER(p.name) ASC
            LIMIT 1
        """)
        top_row = cur.fetchone()
        cur.execute("""
            SELECT p.name FROM wc.product_reviews r
            JOIN wc.products p ON r.product_id = p.id
            GROUP BY p.name
            HAVING COUNT(r.id) >= 3
            ORDER BY AVG(r.rating) ASC, LOWER(p.name) ASC
            LIMIT 1
        """)
        bot_row = cur.fetchone()
        # Fallback names with case-insensitive variant
        top_names = []
        bot_names = []
        if top_row and top_row[0]:
            top_names.append(top_row[0])
        if bot_row and bot_row[0]:
            bot_names.append(bot_row[0])

        # Pull alternative top with ASCII sort to also accept that variant
        cur.execute("""
            SELECT p.name FROM wc.product_reviews r
            JOIN wc.products p ON r.product_id = p.id
            GROUP BY p.name
            HAVING COUNT(r.id) >= 3
            ORDER BY AVG(r.rating) DESC, p.name ASC
            LIMIT 1
        """)
        top_row2 = cur.fetchone()
        if top_row2 and top_row2[0]:
            top_names.append(top_row2[0])

        # For each candidate page, gather content text and confirm at least one
        # top + one lowest product name (40-char prefix matching for truncation).
        for page_id in candidate_ids:
            cur.execute(
                "SELECT block_data FROM notion.blocks WHERE parent_id = %s",
                (page_id,),
            )
            blocks = cur.fetchall()
            content_text = ""
            for b in blocks:
                bd = b[0]
                if isinstance(bd, dict):
                    content_text += " " + _json.dumps(bd)
                elif isinstance(bd, str):
                    content_text += " " + bd
            cl = content_text.lower()

            def _name_in(name):
                n = (name or "").strip().lower()
                if not n:
                    return False
                # match full name OR 40-char prefix to handle truncation
                if n in cl:
                    return True
                if len(n) > 40 and n[:40] in cl:
                    return True
                return False

            top_match = any(_name_in(n) for n in top_names)
            bot_match = any(_name_in(n) for n in bot_names)
            if top_match and bot_match:
                cur.close()
                conn.close()
                return True, "ok"
        cur.close()
        conn.close()
        return False, (f"page found but content lacks top-rated/lowest-rated names; "
                       f"top={top_names}, bot={bot_names}")
    except Exception as e:
        print(f"  WARNING: Notion check error: {e}")
        return False, f"exception: {e}"


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Review_Summary.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Review_Summary.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Review Summary
    print(f"  Checking Review Summary...")
    a_rows = load_sheet_rows(agent_wb, "Review Summary")
    g_rows = load_sheet_rows(gt_wb, "Review Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Review Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Review Summary' not found in groundtruth")
    else:
        sheet_name = "Review Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Build full-name -> row maps. GT product names in this sheet are sometimes
        # truncated to ~50 chars + "..."; agent likely produces full names from the DB.
        # Use prefix-overlap (>=40 chars) matching to be robust to either form.
        def _norm_name(s):
            s = str(s or "").strip().lower()
            # strip trailing ellipsis-like marks for matching
            if s.endswith("..."):
                s = s[:-3].rstrip()
            return s

        a_norm_rows = []
        for row in a_data:
            if row and row[0] is not None:
                a_norm_rows.append((_norm_name(row[0]), row))

        def _find_agent_row(gt_key):
            # exact normalised match
            for k, r in a_norm_rows:
                if k == gt_key:
                    return r, k
            # prefix overlap (>=40 chars, both directions)
            for k, r in a_norm_rows:
                if not k or not gt_key:
                    continue
                short = min(len(k), len(gt_key))
                if short >= 40 and k[:40] == gt_key[:40]:
                    return r, k
            return None, None

        matched_agent_keys = set()
        gt_count = 0
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            gt_count += 1
            key = _norm_name(g_row[0])
            a_row, matched_key = _find_agent_row(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            matched_agent_keys.add(matched_key)

            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):  # exact integer for Reviews count
                    errors.append(f"{key[:40]}.Reviews: {a_row[1]} vs {g_row[1]} (tol=0)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.05):
                    errors.append(f"{key[:40]}.Avg_Rating: {a_row[2]} vs {g_row[2]} (tol=0.05)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.05):
                    errors.append(f"{key[:40]}.Min_Rating: {a_row[3]} vs {g_row[3]} (tol=0.05)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.05):
                    errors.append(f"{key[:40]}.Max_Rating: {a_row[4]} vs {g_row[4]} (tol=0.05)")
        # Reject extra rows in agent not in GT
        for k, r in a_norm_rows:
            if k not in matched_agent_keys:
                errors.append(f"Extra row not in GT: {str(r[0])[:60]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        # Per-metric tolerance map
        tol_map = {
            "products_with_reviews": 0,
            "total_reviews": 0,
            "overall_avg_rating": 0.05,
            "top_rated_product": 0,  # will fall through to str_match
        }
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                tol = tol_map.get(key, 0.05)
                # for product names compare strings
                if key == "top_rated_product":
                    a_val = str(a_row[1] or "").strip().lower()
                    g_val = str(g_row[1] or "").strip().lower()
                    # GT names are sometimes truncated mid-word with "...". Accept the agent value
                    # if either is a meaningful prefix of the other (>=40 chars overlap),
                    # OR if the agent value matches a different valid 5-star tied product
                    # (case-insensitive sort vs ASCII sort may yield different first names).
                    valid_top_rated = set()
                    try:
                        with psycopg2.connect(
                            host=os.environ.get("PGHOST", "localhost"), port=5432,
                            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
                            user="eigent", password="camel",
                        ) as _c:
                            _cur = _c.cursor()
                            _cur.execute("""
                                SELECT p.name FROM wc.product_reviews r
                                JOIN wc.products p ON r.product_id = p.id
                                GROUP BY p.name
                                HAVING COUNT(r.id) >= 3
                                ORDER BY AVG(r.rating) DESC, LOWER(p.name) ASC
                                LIMIT 1
                            """)
                            row1 = _cur.fetchone()
                            if row1:
                                valid_top_rated.add(row1[0].strip().lower())
                            _cur.execute("""
                                SELECT p.name FROM wc.product_reviews r
                                JOIN wc.products p ON r.product_id = p.id
                                GROUP BY p.name
                                HAVING COUNT(r.id) >= 3
                                ORDER BY AVG(r.rating) DESC, p.name ASC
                                LIMIT 1
                            """)
                            row2 = _cur.fetchone()
                            if row2:
                                valid_top_rated.add(row2[0].strip().lower())
                    except Exception as _exc:
                        print(f"  WARNING: Top_Rated_Product DB lookup error: {_exc}")
                    if g_val:
                        valid_top_rated.add(g_val)

                    def _matches(av, candidates):
                        for cv in candidates:
                            if not cv or not av:
                                continue
                            short = min(len(cv), len(av))
                            if short < 40:
                                # Both must equal exactly when short
                                if cv == av:
                                    return True
                            else:
                                # Prefix-match: 40-char prefixes must agree
                                if cv[:40] == av[:40]:
                                    return True
                                if cv == av:
                                    return True
                        return False
                    if not _matches(a_val, valid_top_rated):
                        errors.append(f"{key}.Value: '{a_row[1]}' vs valid {sorted(valid_top_rated)}")
                else:
                    if not num_close(a_row[1], g_row[1], tol):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    # Notion check
    print("  Checking Notion page 'Product Review Dashboard'...")
    notion_ok, notion_detail = check_notion()
    if not notion_ok:
        all_errors.append(f"Notion page 'Product Review Dashboard' check: {notion_detail}")
        print(f"    FAIL: {notion_detail}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
