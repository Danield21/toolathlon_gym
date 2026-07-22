"""Evaluation for terminal-canvas-gsheet-word-notion-gcal."""
import argparse
import json
import os
import sys
from datetime import datetime, date

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
COURSE_IDS = (20, 21)


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def get_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
    SELECT a.course_id, COUNT(*) submissions, AVG(s.score)::numeric(10,2) avg,
           PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY s.score)::numeric(10,2) median,
           (COUNT(CASE WHEN s.score >= 60 THEN 1 END)*100.0/COUNT(*))::numeric(10,2) pass_rate
    FROM canvas.submissions s
    JOIN canvas.assignments a ON a.id = s.assignment_id
    WHERE a.course_id IN %s AND s.score IS NOT NULL
    GROUP BY a.course_id
    ORDER BY a.course_id
    """, (COURSE_IDS,))
    courses = {r[0]: {"submissions": int(r[1]), "avg": float(r[2]), "median": float(r[3]), "pass_rate": float(r[4])} for r in cur.fetchall()}
    cur.execute("""
    SELECT
      CASE WHEN s.score < 60 THEN 'F' WHEN s.score < 70 THEN 'D' WHEN s.score < 80 THEN 'C' WHEN s.score < 90 THEN 'B' ELSE 'A' END AS grade,
      COUNT(*)
    FROM canvas.submissions s
    JOIN canvas.assignments a ON a.id = s.assignment_id
    WHERE a.course_id IN %s AND s.score IS NOT NULL
    GROUP BY 1
    """, (COURSE_IDS,))
    grades = {r[0]: int(r[1]) for r in cur.fetchall()}
    cur.close()
    conn.close()
    return {"courses": courses, "grades": grades}


def check_excel(workspace, expected):
    print("\n=== Check 1: Academic_Advising_Report.xlsx ===")
    path = os.path.join(workspace, "Academic_Advising_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheets_lower = [s.lower() for s in sheets]
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    # Course_Summary
    cs_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s and "summary" in s), None)
    if cs_idx is None:
        cs_idx = next((i for i, s in enumerate(sheets_lower) if "summary" in s), 0)
    ws = wb[sheets[cs_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
    check(f"Course_Summary has 2 course rows", len(data_rows) == 2, f"Found {len(data_rows)}")

    # Verify per-course numeric values
    for row in data_rows:
        if not row or row[0] is None:
            continue
        try:
            cid = int(row[0])
        except (TypeError, ValueError):
            continue
        if cid not in expected["courses"]:
            check(f"Course {cid} in expected list", False, f"got cid={cid}")
            continue
        info = expected["courses"][cid]
        # row layout: course_id, course_name, enrollment, avg, median, pass_rate
        if len(row) >= 6:
            try:
                avg = float(row[3]) if row[3] is not None else None
                median = float(row[4]) if row[4] is not None else None
                pass_rate = float(row[5]) if row[5] is not None else None
            except (TypeError, ValueError):
                avg = median = pass_rate = None
            check(f"Course {cid}.avg ≈ {info['avg']}",
                  avg is not None and abs(avg - info["avg"]) <= 0.5,
                  f"got {avg}")
            check(f"Course {cid}.median ≈ {info['median']}",
                  median is not None and abs(median - info["median"]) <= 0.5,
                  f"got {median}")
            check(f"Course {cid}.pass_rate ≈ {info['pass_rate']}",
                  pass_rate is not None and abs(pass_rate - info["pass_rate"]) <= 0.5,
                  f"got {pass_rate}")

    all_text = " ".join(str(c) for r in rows for c in r if c is not None).lower()
    check("Course_Summary mentions Global Governance",
          "global governance" in all_text or "governance" in all_text,
          all_text[:120])

    # Grade_Distribution
    gd_idx = next((i for i, s in enumerate(sheets_lower) if "grade" in s and "distribution" in s), None)
    if gd_idx is None:
        gd_idx = next((i for i, s in enumerate(sheets_lower) if "grade" in s or "distribution" in s), 1)
    ws2 = wb[sheets[gd_idx]]
    rows2 = list(ws2.iter_rows(values_only=True))
    data_rows2 = [r for r in rows2[1:] if any(c is not None for c in r)]
    check("Grade_Distribution has 5 range rows", len(data_rows2) == 5, f"Found {len(data_rows2)}")
    # Verify per-range counts (match on first non-space character of the label)
    found_grades = {}
    for row in data_rows2:
        label = str(row[0] or "").strip()
        if not label:
            continue
        try:
            cnt = int(float(row[1])) if row[1] is not None else None
        except (TypeError, ValueError):
            cnt = None
        first_char = label[0].upper()
        if first_char in ("A", "B", "C", "D", "F"):
            found_grades[first_char] = cnt
        elif "fail" in label.lower() or "below" in label.lower():
            found_grades["F"] = cnt
    for grade, count in expected["grades"].items():
        check(f"Grade range {grade} count == {count}",
              found_grades.get(grade) == count,
              f"got {found_grades.get(grade)}")

    # Advising_Needs
    an_idx = next((i for i, s in enumerate(sheets_lower) if "advising" in s and "need" in s), None)
    if an_idx is None:
        an_idx = next((i for i, s in enumerate(sheets_lower) if "advising" in s or "need" in s), 2)
    ws3 = wb[sheets[an_idx]]
    rows3 = list(ws3.iter_rows(values_only=True))
    data_rows3 = [r for r in rows3[1:] if any(c is not None for c in r)]
    check("Advising_Needs has 4 category rows", len(data_rows3) == 4, f"Found {len(data_rows3)}")
    all_text3 = " ".join(str(c) for r in rows3 for c in r if c is not None).lower()
    for cat in ("urgent intervention", "academic support", "progress monitoring", "no action"):
        check(f"Advising_Needs contains '{cat}'", cat in all_text3, f"text: {all_text3[:200]}")

    # Appointment_Schedule
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "appointment" in s or "schedule" in s), 3)
    ws4 = wb[sheets[ap_idx]]
    rows4 = list(ws4.iter_rows(values_only=True))
    data_rows4 = [r for r in rows4[1:] if any(c is not None for c in r)]
    check("Appointment_Schedule has 5 weekday rows", len(data_rows4) == 5, f"Found {len(data_rows4)}")
    expected_dates = ["2026-03-16", "2026-03-17", "2026-03-18", "2026-03-19", "2026-03-20"]
    found_dates = []
    for r in data_rows4:
        if r[0] is None:
            continue
        d_str = r[0].isoformat() if hasattr(r[0], "isoformat") else str(r[0])
        found_dates.append(d_str.strip())
    for d in expected_dates:
        check(f"Appointment_Schedule has date {d}", any(d in fd for fd in found_dates), f"got dates {found_dates}")


def check_gsheet():
    print("\n=== Check 2: Google Sheet Analytics ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return

    cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%%academic advising%%'")
    sheets = cur.fetchall()
    check("Spreadsheet 'Academic Advising Analytics' exists", len(sheets) >= 1,
          f"Found: {[s[1] for s in sheets]}")

    if sheets:
        ss_id = sheets[0][0]
        cur.execute("SELECT row_index, col_index, value FROM gsheet.cells WHERE spreadsheet_id = %s", (ss_id,))
        cells = cur.fetchall()
        all_text = " ".join(str(v).lower() for _, _, v in cells if v is not None)
        check("GSheet contains course names (governance/geopolitics)",
              "governance" in all_text or "geopolitics" in all_text,
              f"sample: {all_text[:200]}")
        # Check >=2 numeric cells > 0 (enrollment + avg scores at minimum)
        numeric_count = 0
        for _, _, v in cells:
            try:
                if float(v) > 0:
                    numeric_count += 1
            except (TypeError, ValueError):
                pass
        check("GSheet has at least 6 numeric cells (course summary data)",
              numeric_count >= 6, f"got {numeric_count}")

    cur.close()
    conn.close()


def check_word(workspace):
    print("\n=== Check 3: Advising_Recommendations.docx ===")
    path = os.path.join(workspace, "Advising_Recommendations.docx")
    if not os.path.exists(path):
        check("Word file exists", False, f"Not found at {path}")
        return
    check("Word file exists", True)
    try:
        from docx import Document
        doc = Document(path)
    except Exception as e:
        check("Word readable", False, str(e))
        return
    all_text = " ".join(p.text for p in doc.paragraphs)
    text_lower = all_text.lower()
    check("Word contains performance overview", "overview" in text_lower or "performance" in text_lower)
    check("Word contains at-risk identification", "risk" in text_lower or "intervention" in text_lower)
    check("Word contains recommendations", "recommend" in text_lower)
    check("Word mentions Global Governance/Geopolitics",
          "governance" in text_lower or "geopolitics" in text_lower)
    # Reference to specific advising categories
    for cat in ("urgent intervention", "academic support", "progress monitoring"):
        check(f"Word mentions advising category '{cat}'", cat in text_lower)


def check_notion():
    print("\n=== Check 4: Notion Student Advising Tracker ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("SELECT id, title FROM notion.databases")
    dbs = cur.fetchall()
    advising_db = None
    for db_id, title in dbs:
        title_str = ""
        if isinstance(title, list):
            title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
        elif isinstance(title, str):
            try:
                parsed = json.loads(title)
                if isinstance(parsed, list):
                    title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                else:
                    title_str = str(title)
            except Exception:
                title_str = str(title)
        else:
            title_str = str(title) if title else ""
        if "advising" in title_str.lower() and "tracker" in title_str.lower():
            advising_db = (db_id, title_str)
            break

    check("Notion 'Student Advising Tracker' database exists", advising_db is not None,
          f"DBs: {[d[1] for d in dbs]}")
    if not advising_db:
        cur.close()
        conn.close()
        return

    cur.execute("SELECT id, properties FROM notion.pages WHERE parent->>'database_id' = %s", (advising_db[0],))
    pages = cur.fetchall()
    check("Tracker has at least 4 advising category entries", len(pages) >= 4, f"Found {len(pages)} pages")

    # Verify entries cover all 4 categories
    categories_found = set()
    for pid, props in pages:
        ptext = json.dumps(props or {}).lower()
        for cat in ("urgent intervention", "academic support", "progress monitoring", "no action"):
            if cat in ptext:
                categories_found.add(cat)
    for cat in ("urgent intervention", "academic support", "progress monitoring", "no action"):
        check(f"Notion has page for '{cat}'", cat in categories_found,
              f"found cats {categories_found}")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Check 5: Calendar Advising Events ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return

    cur.execute("""
        SELECT summary, description, start_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%%advising%%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("Exactly 5 advising events created", len(events) == 5, f"Found {len(events)} events")

    expected_dates = {date(2026, 3, 16), date(2026, 3, 17), date(2026, 3, 18), date(2026, 3, 19), date(2026, 3, 20)}
    found_dates = set()
    for sm, desc, start in events:
        if start is not None:
            found_dates.add(start.date() if hasattr(start, "date") else start)
    for d in expected_dates:
        check(f"Calendar has event on {d}", d in found_dates, f"got {found_dates}")

    # All event summaries contain 'Academic Advising'
    if events:
        all_with_phrase = all("academic advising" in (s or "").lower() for s, _, _ in events)
        check("All events summary contains 'Academic Advising'", all_with_phrase,
              f"summaries: {[s for s, _, _ in events]}")
        # All descriptions mention Global Governance program
        descs = " ".join(d or "" for _, d, _ in events).lower()
        check("Event descriptions mention Global Governance program",
              ("global governance" in descs) or ("governance program" in descs),
              f"sample: {descs[:200]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 6: advising_analyzer.py ===")
    path = os.path.join(workspace, "advising_analyzer.py")
    check("advising_analyzer.py exists", os.path.exists(path))
    if os.path.exists(path):
        with open(path) as f:
            content = f.read()
        check("Script references median calculation",
              "median" in content.lower(), f"sample: {content[:200]}")
        check("Script references pass rate / 60 threshold",
              "60" in content and ("pass" in content.lower() or "rate" in content.lower()),
              "missing 60 threshold")


def check_reverse_validation():
    print("\n=== Reverse Validation ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("""
        SELECT summary, start_datetime, EXTRACT(DOW FROM start_datetime) AS dow
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%%advising%%'
    """)
    events = cur.fetchall()
    weekend_events = [e for e in events if e[2] in (0, 6)]
    check("No advising events on weekends", len(weekend_events) == 0,
          f"weekend events: {weekend_events}")
    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = get_expected()
    check_excel(args.agent_workspace, expected)
    check_gsheet()
    check_word(args.agent_workspace)
    check_notion()
    check_gcal()
    check_script(args.agent_workspace)
    check_reverse_validation()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_ok = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": all_ok}, f)
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
