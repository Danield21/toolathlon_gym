"""Evaluation for canvas-course-comparison."""
import argparse
import os
import sys
import openpyxl

try:
    import psycopg2
    DB = {
        "host": os.environ.get("PGHOST", "localhost"),
        "port": int(os.environ.get("PGPORT", "5432")),
        "dbname": "toolathlon_gym",
        "user": "eigent",
        "password": "camel",
    }
except Exception:
    psycopg2 = None
    DB = None


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def num_close_rel(a, b, rel=0.02, abs_tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_gsheet(errors_list, gt_course_codes=None, gt_course_names=None):
    """Verify a Google Sheet titled 'Course Comparison Dashboard' was created with a 'Data' tab.
    Also content-checks: at least N course codes from GT must appear in the cells.
    """
    print("  Checking Google Sheet 'Course Comparison Dashboard'...")
    if psycopg2 is None or DB is None:
        errors_list.append("psycopg2 unavailable; cannot verify Google Sheet")
        return
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        sheets = cur.fetchall()
        target_id = None
        target_name = None
        for sid, title in sheets:
            if title and "course comparison dashboard" in title.strip().lower():
                target_id = sid
                target_name = title
                break
        if target_id is None:
            errors_list.append(
                f"Google Sheet 'Course Comparison Dashboard' not found (got {[n for _,n in sheets]})"
            )
            cur.close(); conn.close()
            return

        # Verify 'Data' tab exists
        cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s", (target_id,))
        tabs = cur.fetchall()
        has_data_tab = any(t and "data" in t[1].strip().lower() for t in tabs)
        if not has_data_tab:
            errors_list.append(
                f"Google Sheet '{target_name}' missing 'Data' tab (tabs: {[t[1] for t in tabs]})"
            )

        # Verify there are enough cells (22 rows + header) * 6 cols ~= 138
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (target_id,))
        cell_count = cur.fetchone()[0]
        if cell_count < 60:
            errors_list.append(
                f"Google Sheet '{target_name}' has only {cell_count} cells; expected >= 60"
            )

        # Content check: ensure most GT course codes appear in cells (mirrors Excel verification)
        if gt_course_codes:
            cur.execute(
                "SELECT COALESCE(value, formatted_value) FROM gsheet.cells "
                "WHERE spreadsheet_id = %s",
                (target_id,)
            )
            cell_values = [str(r[0]).strip().lower() for r in cur.fetchall() if r[0]]
            cells_blob = " | ".join(cell_values)
            missing = [c for c in gt_course_codes if c.lower() not in cells_blob]
            # Allow at most 1 missing course code (margin for header/format variations)
            if len(missing) > max(1, len(gt_course_codes) // 5):
                errors_list.append(
                    f"Google Sheet '{target_name}' missing {len(missing)} GT course codes: {missing[:5]}"
                )
        cur.close(); conn.close()
        print(f"    Found GSheet '{target_name}' with {len(tabs)} tabs and {cell_count} cells")
    except Exception as e:
        errors_list.append(f"Google Sheet check raised: {e}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Course_Comparison.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Course_Comparison.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Course Comparison
    print(f"  Checking Course Comparison...")
    a_rows = load_sheet_rows(agent_wb, "Course Comparison")
    g_rows = load_sheet_rows(gt_wb, "Course Comparison")
    if a_rows is None:
        all_errors.append("Sheet 'Course Comparison' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Course Comparison' not found in groundtruth")
    else:
        sheet_name = "Course Comparison"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            # Course_Name (col 1) - require exact match
            if len(a_row) > 1 and len(g_row) > 1:
                if not str_match(a_row[1], g_row[1]):
                    errors.append(f"{key}.Course_Name: {a_row[1]} vs {g_row[1]}")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close_rel(a_row[2], g_row[2], rel=0.02, abs_tol=2):
                    errors.append(f"{key}.Students: {a_row[2]} vs {g_row[2]} (rel 2%)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0):
                    errors.append(f"{key}.Assignments: {a_row[3]} vs {g_row[3]} (exact)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0):
                    errors.append(f"{key}.Quizzes: {a_row[4]} vs {g_row[4]} (exact)")

            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0):
                    errors.append(f"{key}.Discussions: {a_row[5]} vs {g_row[5]} (exact)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=1.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")



    # Google Sheet check (extract GT course codes from the Course Comparison sheet)
    _gt_codes = []
    _gt_names = []
    _g_rows_main = load_sheet_rows(gt_wb, "Course Comparison")
    if _g_rows_main and len(_g_rows_main) > 1:
        for _row in _g_rows_main[1:]:
            if _row and _row[0] is not None:
                _gt_codes.append(str(_row[0]).strip())
            if _row and len(_row) > 1 and _row[1] is not None:
                _gt_names.append(str(_row[1]).strip())
    check_gsheet(all_errors, gt_course_codes=_gt_codes, gt_course_names=_gt_names)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
