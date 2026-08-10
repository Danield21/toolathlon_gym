"""Evaluation for canvas-enrollment-forecast-excel-gform-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def norm_qtype(t):
    # Normalise a question type: lowercase, drop spaces/hyphens/underscores.
    return "".join(ch for ch in str(t or "").lower() if ch.isalnum())


# Question type families: members of the same family are treated as equivalent.
TEXT_TYPES = {norm_qtype(t) for t in ("text", "short_answer", "paragraph", "long_answer")}
CHOICE_TYPES = {norm_qtype(t) for t in ("choice", "multiple_choice", "radio", "dropdown", "checkbox")}


def check_excel(agent_workspace, gt_dir):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Enrollment_Analysis.xlsx")
    gt_path = os.path.join(gt_dir, "Enrollment_Analysis.xlsx")

    if not os.path.isfile(xlsx_path):
        check("Enrollment_Analysis.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Enrollment_Analysis.xlsx exists", True)

    if not os.path.isfile(gt_path):
        check("Groundtruth Excel exists", False, f"Not found: {gt_path}")
        return

    try:
        agent_wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # Sheet 1: Enrollment Trends
    print("  --- Enrollment Trends ---")
    a_rows = load_sheet_rows(agent_wb, "Enrollment Trends")
    g_rows = load_sheet_rows(gt_wb, "Enrollment Trends")
    if a_rows is None:
        check("Sheet 'Enrollment Trends' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Enrollment Trends' exists in GT", False)
    else:
        check("Sheet 'Enrollment Trends' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Enrollment Trends has 22 data rows", len(a_data) == 22, f"Found {len(a_data)}")

        # Build lookup: (base_name_lower, semester_lower, year) -> row
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower(), int(row[2]) if row[2] else 0)
                a_lookup[key] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower(), int(g_row[2]) if g_row[2] else 0)
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]} {g_row[1]} {g_row[2]}")
                continue
            # Student_Count (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1):
                    errors.append(f"{key}: Student_Count {a_row[3]} vs {g_row[3]}")
        if errors:
            for e in errors[:5]:
                check(f"Enrollment Trends data", False, e)
        else:
            check("Enrollment Trends data matches", True)

    # Sheet 2: Course Capacity
    print("  --- Course Capacity ---")
    a_rows = load_sheet_rows(agent_wb, "Course Capacity")
    g_rows = load_sheet_rows(gt_wb, "Course Capacity")
    if a_rows is None:
        check("Sheet 'Course Capacity' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Course Capacity' exists in GT", False)
    else:
        check("Sheet 'Course Capacity' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Course Capacity has 7 data rows", len(a_data) == 7, f"Found {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            # Latest_Enrollment (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}: Latest_Enrollment {a_row[1]} vs {g_row[1]}")
            # Growth_Trend (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not str_match(a_row[2], g_row[2]):
                    errors.append(f"{key}: Growth_Trend '{a_row[2]}' vs '{g_row[2]}'")
            # Needs_Split (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not str_match(a_row[3], g_row[3]):
                    errors.append(f"{key}: Needs_Split '{a_row[3]}' vs '{g_row[3]}'")
            # Consider_Consolidation (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not str_match(a_row[4], g_row[4]):
                    errors.append(f"{key}: Consider_Consolidation '{a_row[4]}' vs '{g_row[4]}'")
            # Projected_Next (col 5) - whole number from linear regression; allow tol=2
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 2):
                    errors.append(f"{key}: Projected_Next {a_row[5]} vs {g_row[5]}")
            # Faculty_Needed (col 6)
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 1):
                    errors.append(f"{key}: Faculty_Needed {a_row[6]} vs {g_row[6]}")
        if errors:
            for e in errors[:5]:
                check(f"Course Capacity data", False, e)
        else:
            check("Course Capacity data matches", True)

    # Sheet 3: Department Summary
    print("  --- Department Summary ---")
    a_rows = load_sheet_rows(agent_wb, "Department Summary")
    g_rows = load_sheet_rows(gt_wb, "Department Summary")
    if a_rows is None:
        check("Sheet 'Department Summary' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Department Summary' exists in GT", False)
    else:
        check("Sheet 'Department Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing metric: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                # Total_Projected_Enrollment may sum 7 per-course rounding errors;
                # tighten tol from 14 to 7 (1 per course max).
                tol = 7 if "projected" in key else 1
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            for e in errors[:5]:
                check(f"Department Summary data", False, e)
        else:
            check("Department Summary data matches", True)


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Match exact title 'Course Preference Survey' (case-insensitive)
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE LOWER(TRIM(title)) = 'course preference survey'
        """)
        forms = cur.fetchall()
        check("Course Preference Survey form exists (exact title)", len(forms) >= 1,
              f"Found {len(forms)} forms with exact title")

        if forms:
            form_id = forms[0][0]
            cur.execute("""
                SELECT id, title, question_type, config FROM gform.questions
                WHERE form_id = %s ORDER BY position
            """, (form_id,))
            questions = cur.fetchall()
            check("Form has 4 questions", len(questions) == 4,
                  f"Found {len(questions)} questions: {[q[1] for q in questions]}")

            if len(questions) >= 4:
                # Check question types (normalised; compared by family)
                types = [norm_qtype(q[2]) for q in questions]
                titles = [q[1].lower() if q[1] else "" for q in questions]

                has_name_text = any("name" in t and types[i] in TEXT_TYPES
                                    for i, t in enumerate(titles))
                check("Has student name text question", has_name_text,
                      f"Questions: {list(zip(titles, types))}")

                choice_qs = [q for i, q in enumerate(questions) if types[i] in CHOICE_TYPES]
                check("Has checkbox/choice question for courses", len(choice_qs) >= 1,
                      f"Types: {types}")
                check("Has radio/choice question for schedule", len(choice_qs) >= 2,
                      f"Types: {types}")

                # Build a list of all choices across questions for content checks
                def _choices(cfg):
                    if not cfg:
                        return []
                    if isinstance(cfg, str):
                        try:
                            cfg = json.loads(cfg)
                        except (TypeError, ValueError):
                            return []
                    if isinstance(cfg, dict):
                        ch = cfg.get("choices") or cfg.get("options") or []
                        return [str(c).strip().lower() for c in ch]
                    return []

                # All 7 base course names should appear as options in some checkbox/MC question
                base_courses = [
                    "applied analytics & algorithms",
                    "biochemistry & bioinformatics",
                    "creative computing & culture",
                    "data-driven design",
                    "environmental economics & ethics",
                    "foundations of finance",
                    "global governance & geopolitics",
                ]
                # Match by normalised substring/inclusion to handle minor whitespace/punct variation
                def _norm(s):
                    return "".join(ch for ch in s.lower() if ch.isalnum())

                def _covers_all(required, choices):
                    norm_choices = [_norm(c) for c in choices]
                    return all(any(_norm(r) in nc or nc in _norm(r)
                                   for nc in norm_choices if nc)
                               for r in required)

                # Some choice-family question must list all 7 base course names
                course_q = None
                for q in choice_qs:
                    if _covers_all(base_courses, _choices(q[3])):
                        course_q = q
                        break
                check("Checkbox question lists all 7 base course names",
                      course_q is not None,
                      f"Got choices: {[(q[1], _choices(q[3])) for q in choice_qs]}")

                # Some choice-family question must have Morning/Afternoon/Evening options
                sched_choices = []
                for q in choice_qs:
                    if q is course_q:
                        continue
                    cand = _choices(q[3])
                    cand_norm = [c.strip().lower() for c in cand]
                    if (any(c == "morning" or c.startswith("morning") for c in cand_norm)
                            and any(c == "afternoon" or c.startswith("afternoon") for c in cand_norm)
                            and any(c == "evening" or c.startswith("evening") for c in cand_norm)):
                        sched_choices = cand
                        break
                sched_norm = [c.strip().lower() for c in sched_choices]
                has_morning = any("morning" == c or c.startswith("morning") for c in sched_norm)
                has_afternoon = any("afternoon" == c or c.startswith("afternoon") for c in sched_norm)
                has_evening = any("evening" == c or c.startswith("evening") for c in sched_norm)
                check("Schedule radio has Morning option", has_morning,
                      f"Choices: {sched_choices}")
                check("Schedule radio has Afternoon option", has_afternoon,
                      f"Choices: {sched_choices}")
                check("Schedule radio has Evening option", has_evening,
                      f"Choices: {sched_choices}")

                # Fourth question: accessibility / accommodations (text)
                has_accessibility = any(
                    ("access" in t or "accommod" in t or "special need" in t)
                    and types[i] in TEXT_TYPES
                    for i, t in enumerate(titles)
                )
                check("Has accessibility/accommodations text question",
                      has_accessibility,
                      f"Questions: {list(zip(titles, types))}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Google Form check", False, str(e))


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Each chair has a department keyword that should appear in subject and body
        chair_emails = {
            "analytics_chair@university.edu": ["analytics"],
            "biochem_chair@university.edu": ["biochem", "biochemistry"],
            "computing_chair@university.edu": ["computing", "creative"],
            "design_chair@university.edu": ["design"],
            "economics_chair@university.edu": ["economics"],
            "finance_chair@university.edu": ["finance"],
            "governance_chair@university.edu": ["governance"],
        }

        cur.execute("""
            SELECT id, subject, to_addr, body_text FROM email.messages
            WHERE subject ILIKE '%%enrollment%%'
               OR subject ILIKE '%%forecast%%'
               OR subject ILIKE '%%projection%%'
        """)
        emails = cur.fetchall()
        check("Exactly 7 enrollment-related emails sent (one per chair)", len(emails) == 7,
              f"Found {len(emails)} matching emails")

        # Map chair -> emails sent to that chair
        per_chair = {chair: [] for chair in chair_emails}
        for email_row in emails:
            _, subject, to_addr, body = email_row
            if isinstance(to_addr, str):
                try:
                    to_addr_parsed = json.loads(to_addr)
                except json.JSONDecodeError:
                    to_addr_parsed = [to_addr]
            else:
                to_addr_parsed = to_addr
            if not isinstance(to_addr_parsed, list):
                to_addr_parsed = [to_addr_parsed]
            for addr in to_addr_parsed:
                addr_lower = str(addr).lower().strip()
                for chair in chair_emails:
                    if chair in addr_lower:
                        per_chair[chair].append((subject or "", body or ""))

        for chair, dept_keys in chair_emails.items():
            mails = per_chair[chair]
            check(f"Email sent to {chair}", len(mails) >= 1,
                  f"Found {len(mails)} email(s)")
            if not mails:
                continue
            # Each chair email subject should contain "Enrollment Forecast" + dept keyword
            ok_subj = False
            ok_dept_in_subj = False
            ok_body_content = False
            for subj, body in mails:
                sl = subj.lower()
                bl = body.lower()
                if "enrollment forecast" in sl or ("enrollment" in sl and "forecast" in sl):
                    ok_subj = True
                if any(k in sl for k in dept_keys):
                    ok_dept_in_subj = True
                # Body should mention projection (number) AND faculty / staff need.
                # Require a digit within +/-100 chars of a 'project' occurrence.
                has_proj = "project" in bl
                has_fac = "facult" in bl or "staff" in bl or "instructor" in bl or "teacher" in bl
                # Scoped digit-near-projection check
                import re as _re_pj
                digit_near_proj = False
                for m in _re_pj.finditer(r"project", bl):
                    window = bl[max(0, m.start() - 100): m.end() + 100]
                    if any(ch.isdigit() for ch in window):
                        digit_near_proj = True
                        break
                if has_proj and has_fac and digit_near_proj:
                    ok_body_content = True
            check(f"  Subject 'Enrollment Forecast' for {chair}", ok_subj,
                  f"Subjects: {[m[0] for m in mails]}")
            check(f"  Subject mentions department for {chair}", ok_dept_in_subj,
                  f"Subjects: {[m[0] for m in mails]}; expected dept keys: {dept_keys}")
            check(f"  Body has projection & faculty info for {chair}", ok_body_content,
                  f"Bodies (first 200 chars): {[m[1][:200] for m in mails]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
