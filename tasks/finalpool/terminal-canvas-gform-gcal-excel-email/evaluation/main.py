"""Evaluation for terminal-canvas-gform-gcal-excel-email."""
import argparse
import datetime as _dt
import json
import os
import re
import sys

import openpyxl
import psycopg2


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


# R1: all DB settings read from env with defaults consistent with preprocess/main.py.
DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)

PASS_COUNT = 0
FAIL_COUNT = 0

# Quizzes below 75 avg from courses 7 and 11 — derived from DB at module load.
# Falls back to hardcoded snapshot if DB unreachable so eval still runs.
def _fetch_below_75_quizzes():
    fallback = [
        "CMA 24295", "CMA 24298",
        "CMA 25341", "CMA 25343", "CMA 25344",
        "CMA 25345", "CMA 25346", "CMA 25347",
    ]
    try:
        import psycopg2 as _ps
        conn = _ps.connect(**DB)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT q.title FROM canvas.quizzes q
            JOIN canvas.quiz_submissions s ON s.quiz_id = q.id
            WHERE q.course_id IN (7, 11) AND s.score IS NOT NULL
            GROUP BY q.title
            HAVING AVG(s.score) < 75
            ORDER BY q.title
            """
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if rows:
            return [r[0] for r in rows]
        return fallback
    except Exception:
        return fallback


BELOW_75_QUIZZES = _fetch_below_75_quizzes()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def _skip(name, detail=""):
    """Informational skip that does not affect PASS/FAIL counts (used when a
    cell is an Excel formula whose cached value we cannot read, per R2)."""
    d = f": {str(detail)[:200]}" if detail else ""
    print(f"  [SKIP] {name}{d}")


def _cell_value(val):
    """Robust numeric extraction (R2).

    - int/float -> float
    - str      -> strip % / currency / thousands separators, then float
    - formula (starts with '=') -> None (cannot evaluate a cached value here)
    - None / unparseable -> None
    """
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("="):
            return None
        s2 = (s.replace(",", "")
               .replace("%", "")
               .replace("$", "")
               .replace("¥", "")
               .replace("€", "")
               .replace("£", "")
               .strip())
        try:
            return float(s2)
        except ValueError:
            return None
    return None


def _expected_sender():
    """Derive the From address the emails MCP will use from the task dir's
    email_config.json (R7). Falls back to the task narrative address."""
    candidates = [
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "email_config.json"),
        os.path.join(os.getcwd(), "email_config.json"),
    ]
    for cfg_path in candidates:
        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            em = (cfg.get("email") or "").strip().lower()
            if em:
                return em
        except Exception:
            continue
    return "coordinator@assessment.example.com"


def _is_text_type(qtype):
    """MCP google-forms only ever writes 'textQuestion' / 'choiceQuestion' (R6)."""
    q = (qtype or "").upper()
    return ("TEXT" in q) or ("PARAGRAPH" in q) or ("SHORT_ANSWER" in q)


def _is_choice_type(qtype):
    q = (qtype or "").upper()
    return ("CHOICE" in q) or ("RADIO" in q)


def _get_choices(cfg):
    """Extract the actual option value strings from a question config (R5).

    The MCP stores options as a jsonb array like [{"value":"A"},{"value":"B"}]
    under config.options (or config.choices). Compare value strings, not the
    object or its str() repr.
    """
    if cfg is None:
        return []
    if isinstance(cfg, str):
        try:
            cfg = json.loads(cfg)
        except (TypeError, ValueError):
            return []
    if isinstance(cfg, dict):
        opts = cfg.get("choices") or cfg.get("options") or []
        out = []
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    out.append(str(v).strip().lower())
            elif o is not None:
                out.append(str(o).strip().lower())
        return out
    if isinstance(cfg, list):
        out = []
        for o in cfg:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    out.append(str(v).strip().lower())
            elif o is not None:
                out.append(str(o).strip().lower())
        return out
    return []


def _numbers_in_choices(choices):
    """Collect every integer digit-run found inside the choice values."""
    nums = set()
    for c in choices:
        m = re.search(r"\d+", c or "")
        if m:
            nums.add(int(m.group()))
    return nums


def _date_iso_from(val):
    """Normalize a date cell / timestamptz to an ISO 'YYYY-MM-DD' string.

    For datetimes the UTC date is used so the result does not depend on the DB
    session timezone (R9). For strings a handful of common formats is tried.
    """
    if isinstance(val, _dt.datetime):
        try:
            if val.tzinfo is not None:
                val = val.astimezone(_dt.timezone.utc)
            return val.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(val or "").strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%d/%m/%Y", "%d.%m.%Y", "%Y.%m.%d"):
        try:
            return _dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    m = re.search(r"\d{4}-\d{2}-\d{2}", s)
    if m:
        return m.group()
    return ""


def _hour_from_time_string(s):
    """Parse a time cell like '14:00', '2:00 PM', '2pm', '14:00:00', or a full
    datetime like '2026-03-16 14:00:00' -> int hour. The clock-time token is
    searched anywhere in the cell so a leading date cannot be mistaken for the
    hour (a bare '14:00' in a datetime string must not read the year)."""
    s = str(s or "").strip().lower().replace(".", "")
    # HH:MM[:SS] with optional am/pm -- the canonical clock-time token.
    m = re.search(r"(\d{1,2}):(\d{2})(?::\d{2})?\s*(am|pm)?", s)
    if m:
        hour = int(m.group(1))
        if m.group(3) == "pm":
            if hour != 12:
                hour += 12
        elif m.group(3) == "am":
            if hour == 12:
                hour = 0
        return hour
    # Bare hour with am/pm: '2 PM', '9am'
    m = re.search(r"\b(\d{1,2})\s*(am|pm)\b", s)
    if m:
        hour = int(m.group(1))
        if m.group(2) == "pm":
            if hour != 12:
                hour += 12
        elif m.group(2) == "am":
            if hour == 12:
                hour = 0
        return hour
    # Bare 24h hour token (0..23).
    m = re.search(r"\b(\d{1,2})\b", s)
    if m:
        h = int(m.group(1))
        if h <= 23:
            return h
    return None


def _utc_hour(sd):
    """Extract the UTC hour of an event start time, or None if unparseable."""
    try:
        if isinstance(sd, _dt.datetime):
            utc = sd.astimezone(_dt.timezone.utc) if sd.tzinfo is not None else sd
            return utc.hour
        if isinstance(sd, str):
            m = re.search(r"(\d{1,2}):(\d{2})", str(sd))
            if m:
                return int(m.group(1)) % 24
    except Exception:
        pass
    return None


def _is_valid_start_time(sd):
    """A review event counts as the '2:00 PM' session regardless of timezone (R9).

    task.md specifies 2:00 PM but no timezone, so the *hour* of the stored instant
    cannot be required: a wall-clock 14:00 local maps to any UTC hour 0..23
    depending on the timezone the agent used (e.g. +09 -> 05:00 UTC, +10 -> 04:00,
    -09 -> 23:00, -10 -> 00:00 the next day). The timezone-independent properties
    of a '2:00 PM' event that we can require here are:
      - it is on the hour (minute == 0), and
      - every session shares the same time-of-day ('each subsequent session one
        day later at the same time') -- enforced via _utc_hour in check_gcal().

    The literal '14:00' requirement is still enforced on the Excel
    Remediation_Schedule Start_Time column, which is an explicit wall-clock time
    written by the agent with no timezone ambiguity.
    """
    s = str(sd or "")
    if "14:00" in s:
        return True
    try:
        if isinstance(sd, _dt.datetime):
            utc = sd.astimezone(_dt.timezone.utc) if sd.tzinfo is not None else sd
            return utc.minute == 0
        if isinstance(sd, str):
            m = re.search(r"(\d{1,2}):(\d{2})", s)
            if m:
                return int(m.group(2)) == 0
    except Exception:
        pass
    return False


def check_excel(ws_path):
    """Check Quiz_Performance_Report.xlsx."""
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "Quiz_Performance_Report.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        return
    check("Excel file exists", True)

    try:
        # R2: read formulas (data_only=False) so a formula cell is detected and
        # handled gracefully instead of silently reading None.
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Quiz_Performance
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
    qp_name = None
    for candidate in ["quiz_performance", "quiz performance"]:
        if candidate in sheet_names_lower:
            qp_name = sheet_names_lower[candidate]
            break
    if qp_name is None:
        check("Quiz_Performance sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Quiz_Performance sheet exists", True)
        ws = wb[qp_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        # Query dynamic quiz count from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.quizzes WHERE course_id IN (7, 11)")
            expected_quiz_count = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_quiz_count = 11
        # At least all quizzes must be present; a correct agent may also add a
        # totals row, so use >= (the per-quiz checks below still require every
        # quiz to be present and numerically correct).
        check(f"Quiz_Performance has at least {expected_quiz_count} rows",
              len(data_rows) >= expected_quiz_count,
              f"Found {len(data_rows)} data rows")

        # Query dynamic quiz avg scores + pass rates from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("""
                SELECT q.id,
                       q.title,
                       AVG(qs.score),
                       100.0 * SUM(CASE WHEN qs.score >= 70 THEN 1 ELSE 0 END) / COUNT(*)
                FROM canvas.quizzes q
                JOIN canvas.quiz_submissions qs ON q.id = qs.quiz_id
                WHERE q.course_id IN (7, 11)
                GROUP BY q.id, q.title
            """)
            quiz_stats = {}
            for qid, title, avg, pr in cur.fetchall():
                quiz_stats[str(qid)] = {
                    "title": title,
                    "avg": float(avg),
                    "pass_rate": float(pr),
                }
            cur.close(); conn.close()
        except Exception as e:
            print(f"  [INFO] Could not fetch quiz stats from DB: {e}")
            quiz_stats = {}

        # Check ALL quizzes from DB against agent's data.
        # Match by quiz id, with a case-insensitive title fallback (R13) so the
        # row is still found if Quiz_ID/Quiz_Title columns are swapped.
        a_by_qid = {}
        a_by_title = {}
        for r in data_rows:
            try:
                qid = int(r[2]) if r[2] is not None else None
            except (TypeError, ValueError):
                qid = None
            if qid is not None:
                a_by_qid[str(qid)] = r
            for col in (r[2], r[3]):
                if isinstance(col, str) and col.strip():
                    a_by_title[col.strip().lower()] = r

        for qid, stat in quiz_stats.items():
            r = a_by_qid.get(qid) or a_by_title.get(str(stat["title"]).strip().lower())
            if r is None:
                check(f"Quiz {qid} ({stat['title']}) present", False,
                      f"Not found in Excel")
                continue
            avg = _cell_value(r[4])
            pr = _cell_value(r[5])
            needs = str(r[6]).strip().lower() if len(r) > 6 and r[6] else ""
            if avg is None:
                _skip(f"Quiz {qid} Avg_Score numeric", "formula/non-literal value, skipped")
            else:
                check(f"Quiz {qid} ({stat['title']}) Avg_Score ~{stat['avg']:.2f}",
                      abs(avg - stat["avg"]) < 1.0,
                      f"Got {avg}")
            if pr is None:
                _skip(f"Quiz {qid} Pass_Rate numeric", "formula/non-literal value, skipped")
            else:
                # R8-ish: allow pass rate as either a percentage (62.91) or a
                # decimal fraction (0.6291); normalize the fraction upward.
                expected_pr = stat["pass_rate"]
                pr_n = pr
                if expected_pr > 1.0 and pr_n < 1.0:
                    pr_n *= 100.0
                check(f"Quiz {qid} Pass_Rate ~{stat['pass_rate']:.2f}",
                      abs(pr_n - expected_pr) < 2.0,
                      f"Got {pr}")
            expected_needs = "yes" if stat["avg"] < 75 else "no"
            # Tighten: exact match (case-insensitive trim) instead of 'in' substring
            check(f"Quiz {qid} Needs_Review = {expected_needs}",
                  needs == expected_needs,
                  f"Got '{r[6]}' for avg {stat['avg']:.2f}")

    # Sheet 2: Feedback_Summary
    fb_name = None
    for candidate in ["feedback_summary", "feedback summary"]:
        if candidate in sheet_names_lower:
            fb_name = sheet_names_lower[candidate]
            break
    if fb_name is None:
        check("Feedback_Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Feedback_Summary sheet exists", True)
        ws2 = wb[fb_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Feedback_Summary has 5 rows", len(data_rows2) == 5,
              f"Found {len(data_rows2)} rows")
        # Validate question types per task spec (R6: only text / multiple-choice
        # are creatable by the MCP; Q4 is a single-choice 1..5 scale).
        expected_types = {
            1: ["text", "short_answer", "paragraph", "free_text"],
            2: ["multiple_choice", "radio", "choice"],
            3: ["multiple_choice", "radio", "choice"],
            4: ["multiple_choice", "radio", "choice", "scale", "rating"],
            5: ["text", "short_answer", "paragraph", "free_text"],
        }
        for r in data_rows2:
            try:
                qn = int(r[0])
            except (TypeError, ValueError):
                continue
            qtype = str(r[2] or "").strip().lower().replace(" ", "_")
            if qn in expected_types:
                ok = any(t in qtype for t in expected_types[qn])
                check(f"Feedback_Summary Q{qn} type matches expected",
                      ok,
                      f"Got '{r[2]}', expected one of {expected_types[qn]}")
        # Q1 should mention 'challenging' or 'topic'
        q1_text = next((str(r[1] or "").lower() for r in data_rows2 if r[0] == 1), "")
        check("Q1 mentions challenging course topics",
              "challeng" in q1_text or "topic" in q1_text or "difficult" in q1_text,
              f"Q1 text: {q1_text}")

    # Sheet 3: Remediation_Schedule
    rs_name = None
    for candidate in ["remediation_schedule", "remediation schedule"]:
        if candidate in sheet_names_lower:
            rs_name = sheet_names_lower[candidate]
            break
    if rs_name is None:
        check("Remediation_Schedule sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Remediation_Schedule sheet exists", True)
        ws3 = wb[rs_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("Remediation_Schedule has 8 rows", len(data_rows3) == 8,
              f"Found {len(data_rows3)} rows")
        # Expected dates: 2026-03-16 .. 2026-03-23 (one per day) at 14:00, 1 hour
        expected_dates = [f"2026-03-{16+i:02d}" for i in range(8)]
        seen_dates = set()
        for r in data_rows3:
            d = _date_iso_from(r[1])
            t = str(r[2] or "")
            dur = _cell_value(r[3])
            if d in expected_dates:
                seen_dates.add(d)
            hour = _hour_from_time_string(t)
            check(f"Remediation row '{r[0]}' Start_Time = 14:00",
                  hour == 14,
                  f"Got '{r[2]}'")
            if dur is None:
                _skip(f"Remediation row '{r[0]}' Duration_Hours numeric",
                      "formula/non-literal value, skipped")
            else:
                check(f"Remediation row '{r[0]}' Duration_Hours = 1",
                      abs(dur - 1.0) < 0.05,
                      f"Got {r[3]}")
        check("All 8 expected session dates (Mar 16-23) present",
              len(seen_dates) == 8,
              f"Got dates {sorted(seen_dates)}, expected {expected_dates}")

    wb.close()


def _gform_score(questions):
    """Score a form's questions against the Q1..Q5 task requirements.

    Returns a dict with the question count and booleans for each of Q1..Q5, plus a
    total score. Used to select the most complete form when several similarly-named
    forms exist (issue4) and to drive the individual PASS/FAIL checks.
    """
    types = [(q[2] or "").upper() for q in questions]
    titles = [(q[1] or "").lower() for q in questions]

    q1 = any(
        ("challeng" in t or "topic" in t or "difficult" in t) and _is_text_type(types[i])
        for i, t in enumerate(titles)
    )
    q2 = False
    for i, q in enumerate(questions):
        t = titles[i]
        if "hour" in t and "stud" in t and _is_choice_type(types[i]):
            choices = _get_choices(q[3])
            expected_choices = ["less than 3", "3 to 5", "5 to 8", "more than 8"]
            if sum(1 for ec in expected_choices if any(ec in c for c in choices)) >= 3:
                q2 = True
                break
    q3 = False
    for i, q in enumerate(questions):
        t = titles[i]
        if ("learn" in t or "format" in t) and _is_choice_type(types[i]):
            choices = _get_choices(q[3])
            expected_choices = ["lectures", "hands-on labs", "group projects", "self-paced online"]
            if sum(1 for ec in expected_choices if any(ec in c or c in ec for c in choices)) >= 3:
                q3 = True
                break
    q4 = False
    for i, q in enumerate(questions):
        t = titles[i]
        if "difficult" in t or "rate" in t or "scale" in t:
            if _is_choice_type(types[i]):
                choices = _get_choices(q[3])
                nums = _numbers_in_choices(choices)
                if 1 in nums and 5 in nums and len(nums) >= 3:
                    q4 = True
                    break
            elif _is_text_type(types[i]):
                q4 = True
                break
    q5 = any(
        ("suggest" in t or "improv" in t) and _is_text_type(types[i])
        for i, t in enumerate(titles)
    )
    return {
        "count": len(questions),
        "q1": q1, "q2": q2, "q3": q3, "q4": q4, "q5": q5,
        "score": int(q1) + int(q2) + int(q3) + int(q4) + int(q5),
    }


def check_gform():
    """Check Google Form creation."""
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    def _title_match(t):
        return ("academic" in t and "performance" in t and ("self" in t or "assess" in t)) \
            or ("performance" in t and "assess" in t)

    matching = [(fid, title) for fid, title in forms if _title_match((title or "").lower())]

    check("Assessment feedback form 'Academic Performance Self-Assessment' created",
          len(matching) > 0,
          f"Forms: {[f[1] for f in forms]}")

    form_id = None
    questions = []
    score = {"count": 0, "q1": False, "q2": False, "q3": False, "q4": False, "q5": False, "score": 0}
    if matching:
        # Load questions for every candidate and select the most complete form. A
        # swarm / retry may have left multiple similarly-named forms (some partial);
        # the complete form wins instead of the first incomplete one (issue4).
        candidates = []
        for fid, _title in matching:
            cur.execute("""
                SELECT id, title, question_type, config FROM gform.questions
                WHERE form_id = %s ORDER BY position
            """, (fid,))
            qs = cur.fetchall()
            candidates.append((fid, qs, _gform_score(qs)))
        # Prefer a form with exactly 5 questions; among those the one passing the
        # most Q1..Q5 checks; otherwise fall back to the most complete attempt.
        with5 = [c for c in candidates if c[2]["count"] == 5]
        if with5:
            form_id, questions, score = max(with5, key=lambda c: (c[2]["score"], c[2]["count"]))
        else:
            form_id, questions, score = max(candidates, key=lambda c: (c[2]["count"], c[2]["score"]))

    if form_id:
        check("Form has exactly 5 questions", score["count"] == 5,
              f"Found {score['count']}")

        if score["count"] >= 5:
            types = [(q[2] or "").upper() for q in questions]
            titles = [(q[1] or "").lower() for q in questions]
            check("Q1: free text question about challenging topics", score["q1"],
                  f"Questions: {list(zip(titles, types))}")
            check("Q2: multiple choice for study hours with 4 specific options", score["q2"],
                  f"Questions+choices: {[(t, types[i], _get_choices(questions[i][3])) for i, t in enumerate(titles)]}")
            check("Q3: multiple choice for learning format with 4 specific options", score["q3"],
                  f"Questions: {list(zip(titles, types))}")
            check("Q4: 1-5 rating for course difficulty (single choice or text)", score["q4"],
                  f"Questions: {list(zip(titles, types))}")
            check("Q5: free text question about suggestions", score["q5"],
                  f"Questions: {list(zip(titles, types))}")

    conn.close()


def check_gcal():
    """Check calendar events."""
    print("\n=== Checking Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT summary, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()

    # Filter to "Quiz Review:" events
    review_events = [e for e in events
                     if "quiz review" in (e[0] or "").lower()]
    check("Exactly 8 'Quiz Review:' events created",
          len(review_events) == 8,
          f"Found {len(review_events)} Quiz Review events: {[e[0] for e in review_events]}")

    # Reference all 8 below-75 quiz titles
    found_quizzes = []
    for quiz_title in BELOW_75_QUIZZES:
        quiz_num = quiz_title.split()[-1]
        if any(quiz_num in (e[0] or "") for e in review_events):
            found_quizzes.append(quiz_title)
    check(f"Review events reference all {len(BELOW_75_QUIZZES)} below-75 quizzes",
          len(found_quizzes) == len(BELOW_75_QUIZZES),
          f"Referenced quizzes: {found_quizzes}; missing: {set(BELOW_75_QUIZZES) - set(found_quizzes)}")

    # Validate dates (UTC-normalized so the DB session timezone cannot shift the
    # calendar date, R9). task.md does not state a timezone, so a correct agent in
    # UTC-10..UTC-12 creates '14:00 local on Mar 16' as 00:00..02:00 UTC on Mar 17,
    # shifting the whole block to Mar 17..Mar 24. Accept both the base block and
    # that single legitimate +1-day shift; any other date block still FAILs.
    expected_dates = [f"2026-03-{16+i:02d}" for i in range(len(BELOW_75_QUIZZES))]
    expected_dates_shifted = [f"2026-03-{17+i:02d}" for i in range(len(BELOW_75_QUIZZES))]
    seen_dates = set()
    for _, sd in review_events:
        d = _date_iso_from(sd)
        if d:
            seen_dates.add(d)
    dates_ok = (seen_dates == set(expected_dates)) or (seen_dates == set(expected_dates_shifted))
    check("All 8 expected calendar dates (Mar 16-23, or Mar 17-24 for UTC-10..-12) present",
          dates_ok,
          f"Got {sorted(seen_dates)}, expected {expected_dates} or {expected_dates_shifted}")

    # Check 2 PM start time (timezone tolerant, R9): every event on the hour and
    # all events at the same time-of-day ('...one day later at the same time').
    on_the_hour = bool(review_events) and all(_is_valid_start_time(sd) for _, sd in review_events)
    utc_hours = {_utc_hour(sd) for _, sd in review_events if _utc_hour(sd) is not None}
    same_time = len(utc_hours) == 1
    check("All 8 review events start on the hour at the same time each day (2:00 PM equivalent)",
          on_the_hour and same_time and len(review_events) == 8,
          f"on_the_hour={on_the_hour}, distinct UTC hours={sorted(utc_hours)}, count={len(review_events)}")

    conn.close()


def _email_recipients(to_addr):
    """Normalize the jsonb/list to_addr column into a list of lowercase strings."""
    if isinstance(to_addr, list):
        return [str(r).strip().lower() for r in to_addr if r is not None]
    if isinstance(to_addr, str):
        try:
            parsed = json.loads(to_addr)
            if isinstance(parsed, list):
                return [str(r).strip().lower() for r in parsed if r is not None]
        except (json.JSONDecodeError, TypeError):
            pass
        return [str(to_addr).strip().lower()]
    return []


def check_email():
    """Check email sent."""
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    expected_from = _expected_sender()

    # A swarm / multi-agent retry may leave earlier test or partial emails also
    # addressed to faculty. Consider every such email and evaluate the most
    # complete one, so an incomplete first email cannot shadow the correct summary
    # email (issue3).
    candidates = [e for e in all_emails
                  if "faculty@assessment.example.com" in _email_recipients(e[2])]

    def _score_email(e):
        subj, from_addr, to_addr, body = e
        score = 0
        sl = re.sub(r"\s+", " ", (subj or "").lower()).strip()
        if all(p in sl for p in ["quiz performance analysis", "cross-course review"]):
            score += 1
        if expected_from in (from_addr or "").lower():
            score += 1
        b = (body or "").lower()
        if "remediation" in b or "review" in b or "flagged" in b:
            score += 1
        if "survey" in b or "feedback" in b:
            score += 1
        if re.search(r"\b(11|eleven)\b", body or "", re.IGNORECASE):
            score += 1
        if re.search(r"\b(8|eight)\b", body or "", re.IGNORECASE):
            score += 1
        if re.findall(r'\b(7[0-9]\.\d{1,2}|7[0-9])\b', body or ""):
            score += 1
        return score

    check("Email sent to faculty@assessment.example.com", len(candidates) > 0,
          f"Total emails: {len(all_emails)}")

    if candidates:
        target_email = max(candidates, key=_score_email)
        subj, from_addr, to_addr, body = target_email
        sl = (subj or "").lower()
        # Subject: "Quiz Performance Analysis - Cross-Course Review"
        # Normalized, then require the two key phrases (relaxed vs exact match
        # so minor punctuation/dash differences don't cause a false FAIL).
        _norm_subj = re.sub(r"\s+", " ", sl).strip()
        _expected_parts = ["quiz performance analysis", "cross-course review"]
        check("Email subject 'Quiz Performance Analysis - Cross-Course Review'",
              all(p in _norm_subj for p in _expected_parts),
              f"Subject: '{subj}' (normalized: '{_norm_subj}')")
        # R7: expected From derived from the task's email_config.json.
        check(f"Email from {expected_from}",
              expected_from in (from_addr or "").lower(),
              f"From: {from_addr}")
        body_lower = (body or "").lower()
        check("Email body mentions remediation or review",
              "remediation" in body_lower or "review" in body_lower or "flagged" in body_lower,
              "Expected remediation/review content in body")
        check("Email body mentions survey or feedback",
              "survey" in body_lower or "feedback" in body_lower,
              "Expected survey/feedback mention in body")
        # Body should include total quizzes count (11) and flagged count (8).
        # Accept both the numerals and their spelled-out forms (R10).
        check("Email body mentions total quizzes (11 or eleven) word-boundary",
              re.search(r"\b(11|eleven)\b", body or "", re.IGNORECASE) is not None,
              f"Body sample: {body_lower[:300]}")
        check("Email body mentions flagged count (8 or eight) word-boundary",
              re.search(r"\b(8|eight)\b", body or "", re.IGNORECASE) is not None,
              f"Body sample: {body_lower[:300]}")
        # Body should include the program average (around 73-74)
        avg_nums = re.findall(r'\b(7[0-9]\.\d{1,2}|7[0-9])\b', body or "")
        check("Email body includes program average score (in 70-80 range)",
              len(avg_nums) >= 1,
              f"Numbers found: {avg_nums}; body sample: {(body or '')[:300]}")

    conn.close()


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check that noise emails are not in agent output (using assessment.example.com domain noise)
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        expected_from = _expected_sender()
        # The stored from_addr is the formatted header ("Name <email>"), so use
        # a LIKE filter instead of an exact equality (problem 7).
        cur.execute(
            "SELECT to_addr FROM email.messages WHERE from_addr LIKE %s",
            (f"%{expected_from}%",)
        )
        sent_emails = cur.fetchall()
        noise_recipients = [
            # Mismatched domains
            "all-staff@university.edu", "faculty@university.edu",
            "all@university.edu", "researchers@university.edu",
            # Same domain but wrong recipient
            "admin@assessment.example.com", "it@assessment.example.com",
            "facilities@assessment.example.com", "grants@assessment.example.com",
        ]
        for email_row in sent_emails:
            to_str = str(email_row[0]).lower()
            for noise in noise_recipients:
                if noise in to_str:
                    check("No email sent to noise recipients", False,
                          f"Sent to noise recipient: {noise}")
                    cur.close(); conn.close()
                    return
        check("No email sent to noise recipients", True)
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-CANVAS-GFORM-GCAL-EXCEL-EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_gform()
    check_gcal()
    check_email()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
