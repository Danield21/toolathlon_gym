"""
Evaluation for yt-canvas-tech-curriculum-excel-notion task.

Checks:
1. Curriculum_Video_Map.xlsx exists with Course_Videos and Summary sheets, exact columns + values from GT
2. Summary numeric values validated against GT
3. Notion page 'Tech Course Video Resources' exists with course content
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Curriculum_Video_Map.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Curriculum_Video_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Curriculum_Video_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Curriculum_Video_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Course_Videos sheet
    cv_idx = next((i for i, s in enumerate(sheet_names_lower) if "course_video" in s or "course video" in s), None)
    if cv_idx is None:
        record("Course_Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Course_Videos sheet exists", True)

    ws = wb[wb.sheetnames[cv_idx]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        record("Course_Videos has data rows", False, "Sheet is empty")
        return

    headers = [str(c).strip() if c else "" for c in rows[0]]
    REQUIRED_COLS = ["Course_Name", "Course_Code", "Video_ID", "Video_Title",
                     "View_Count", "Duration_Min", "Relevance_Match"]
    for col in REQUIRED_COLS:
        record(f"Course_Videos column '{col}'", col in headers, f"got {headers}")

    h_idx = {h: i for i, h in enumerate(headers)}
    data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)]
    # 2-3 videos per course
    course_counts = {}
    for r in data_rows:
        if r and "Course_Name" in h_idx:
            cn = str(r[h_idx["Course_Name"]] or "").strip()
            course_counts[cn] = course_counts.get(cn, 0) + 1
    record(
        "Each matched course has 2 or 3 video rows",
        all(2 <= c <= 3 for c in course_counts.values()) and len(course_counts) > 0,
        f"course counts: {course_counts}",
    )

    # Sort order: Course_Name asc, then View_Count desc within each course
    if data_rows and "Course_Name" in h_idx and "View_Count" in h_idx:
        cn_idx = h_idx["Course_Name"]
        vc_idx = h_idx["View_Count"]
        # Group by course; within each group, view_counts must be descending
        sort_ok = True
        for cn, group in __import__("itertools").groupby(data_rows, key=lambda r: str(r[cn_idx] or "")):
            views = [r[vc_idx] for r in group]
            views_dec = all(views[i] >= views[i + 1] for i in range(len(views) - 1))
            if not views_dec:
                sort_ok = False
                break
        # Also: course names must be ascending across rows
        course_names_seen = []
        seen_names = set()
        for r in data_rows:
            cn = str(r[cn_idx] or "")
            if cn not in seen_names:
                course_names_seen.append(cn)
                seen_names.add(cn)
        names_asc = course_names_seen == sorted(course_names_seen)
        record("Course_Videos sorted by Course_Name asc, View_Count desc", sort_ok and names_asc,
               f"sort_ok={sort_ok} names_asc={names_asc}")

    # ---- GT comparison ----
    gt_path = os.path.join(groundtruth_workspace, "Curriculum_Video_Map.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        # Course_Videos GT comparison
        gt_ws = None
        for sn in gt_wb.sheetnames:
            if sn.lower().startswith("course_video"):
                gt_ws = gt_wb[sn]; break
        if gt_ws is not None:
            # Read header row first
            gt_all_rows = list(gt_ws.iter_rows(values_only=True))
            gt_headers = [str(c).strip() if c else "" for c in (gt_all_rows[0] if gt_all_rows else [])]
            gt_h_idx = {h: i for i, h in enumerate(gt_headers)}
            gt_rows = [r for r in gt_all_rows[1:] if r and any(c is not None for c in r)]
            record("GT Course_Videos exact row count", len(data_rows) == len(gt_rows),
                   f"agent={len(data_rows)} gt={len(gt_rows)}")
            # Build lookup by (Course_Name, Video_ID)
            gt_lookup = {}
            for gr in gt_rows:
                key = (str(gr[gt_h_idx["Course_Name"]] or "").strip(),
                       str(gr[gt_h_idx["Video_ID"]] or "").strip())
                gt_lookup[key] = gr
            for ar in data_rows:
                key = (str(ar[h_idx["Course_Name"]] or "").strip(),
                       str(ar[h_idx["Video_ID"]] or "").strip())
                gt_row = gt_lookup.get(key)
                if gt_row is None:
                    record(f"Row {key} present in GT", False)
                    continue
                # Validate cells
                for col in REQUIRED_COLS:
                    if col not in h_idx or col not in gt_h_idx:
                        continue
                    av = ar[h_idx[col]]
                    gv = gt_row[gt_h_idx[col]]
                    if isinstance(gv, (int, float)) and not isinstance(gv, bool):
                        if not num_close(av, gv, max(abs(gv) * 0.05, 0.5)):
                            record(f"Row {key} {col}", False, f"agent={av} gt={gv}")
                    else:
                        if not str_match(av, gv):
                            record(f"Row {key} {col}", False, f"agent={av} gt={gv}")

    # ---- Summary sheet ----
    sum_idx = next((i for i, s in enumerate(sheet_names_lower) if "summary" in s), None)
    if sum_idx is None:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Summary sheet exists", True)

    ws_sum = wb[wb.sheetnames[sum_idx]]
    sum_rows = list(ws_sum.iter_rows(values_only=True))
    sum_data = {}
    for r in sum_rows:
        if r and r[0]:
            key = str(r[0]).strip()
            sum_data[key] = r[1] if len(r) > 1 else None

    REQUIRED_SUMMARY = ["Total_Courses_Mapped", "Total_Videos_Recommended", "Avg_Views_Recommended"]
    for k in REQUIRED_SUMMARY:
        record(f"Summary has '{k}' label", k in sum_data, f"got {list(sum_data.keys())}")

    # Compare against GT summary numerics
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        gt_sum_ws = None
        for sn in gt_wb.sheetnames:
            if "summary" in sn.lower():
                gt_sum_ws = gt_wb[sn]; break
        if gt_sum_ws is not None:
            gt_sum = {str(r[0]).strip(): r[1] for r in gt_sum_ws.iter_rows(values_only=True) if r and r[0]}
            for k in REQUIRED_SUMMARY:
                if k in sum_data and k in gt_sum:
                    actual = sum_data[k]
                    expected = gt_sum[k]
                    if isinstance(expected, (int, float)):
                        # Avg_Views_Recommended tol broader
                        tol = 0 if k != "Avg_Views_Recommended" else max(1, expected * 0.01)
                        record(f"Summary {k} value", num_close(actual, expected, tol),
                               f"agent={actual} gt={expected}")
                    else:
                        record(f"Summary {k} value", str_match(actual, expected),
                               f"agent={actual} gt={expected}")


def check_notion(agent_workspace="."):
    print("\n=== Check 2: Notion page 'Tech Course Video Resources' ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    # Strict title check: page properties.title.title[*].plain_text exactly matches
    cur.execute("""
        SELECT id, properties FROM notion.pages
        ORDER BY created_time DESC
    """)
    rows = cur.fetchall()
    matched = []
    for pid, props in rows:
        title_text = ""
        if isinstance(props, dict):
            for k, v in props.items():
                if isinstance(v, dict):
                    if v.get("type") == "title" or "title" in v:
                        title_arr = v.get("title")
                        if isinstance(title_arr, list):
                            for piece in title_arr:
                                if isinstance(piece, dict):
                                    title_text += piece.get("plain_text", "")
                                    inner = piece.get("text") or {}
                                    if isinstance(inner, dict):
                                        title_text += inner.get("content", "")
        if title_text.strip().lower() == "tech course video resources":
            matched.append(pid)
    record("Notion page 'Tech Course Video Resources' exists (exact title)",
           len(matched) > 0,
           f"No matching page found; checked {len(rows)} pages")

    # Determine expected number of courses from agent's Course_Videos sheet (or GT)
    expected_course_count = 0
    try:
        xlsx_path = os.path.join(agent_workspace, "Curriculum_Video_Map.xlsx")
        if os.path.isfile(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            for sn in wb.sheetnames:
                if sn.lower().startswith("course_video"):
                    ws = wb[sn]
                    rs = list(ws.iter_rows(values_only=True))
                    if rs:
                        hdrs = [str(c).strip() if c else "" for c in rs[0]]
                        if "Course_Name" in hdrs:
                            cn_i = hdrs.index("Course_Name")
                            names = set()
                            for r in rs[1:]:
                                if r and len(r) > cn_i and r[cn_i]:
                                    names.add(str(r[cn_i]).strip())
                            expected_course_count = len(names)
                    break
    except Exception:
        pass
    if expected_course_count <= 0:
        expected_course_count = 6  # fallback to default

    if matched:
        page_id = matched[0]
        cur.execute("""
            SELECT block_data FROM notion.blocks
            WHERE parent_id = %s
            ORDER BY position
        """, (page_id,))
        blocks = cur.fetchall()
        all_content = " ".join(str(b[0] or "") for b in blocks).lower()

        record("Notion page mentions 'fireship' (video source)", "fireship" in all_content,
               f"snippet: {all_content[:200]}")
        record("Notion page mentions 'data' OR 'computing' course type",
               "data" in all_content or "computing" in all_content,
               f"snippet: {all_content[:200]}")
        # Heading-per-course requirement: heading count >= number of courses
        cur.execute(
            "SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s "
            "AND block_data::text ILIKE '%heading_%'",
            (page_id,),
        )
        heading_count = cur.fetchone()[0]
        record(f"Notion page has at least {expected_course_count} headings (one per course)",
               heading_count >= expected_course_count,
               f"heading_count={heading_count} expected>={expected_course_count}")

    cur.close()
    conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    print(f"\nOverall: {PASS_COUNT}/{total} checks passed")

    result = {"total_passed": PASS_COUNT, "total_checks": total}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Strict: all must pass (FAIL_COUNT == 0)
    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL ({FAIL_COUNT} failures)")
        sys.exit(1)


if __name__ == "__main__":
    main()
