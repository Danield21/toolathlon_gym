"""Evaluation for yf-sector-comparison-gform-excel-email.

Checks:
1. Sector_Comparison.xlsx with Metrics sheet (5 rows) and Sector_Summary sheet (5 rows)
2. GForm "Investment Preference Survey" with 4 questions (titles + options)
3. Email to investors@fund.example.com with "Sector Comparison" in subject
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0

SYMBOLS = ["GOOGL", "AMZN", "JPM", "JNJ", "XOM"]
SECTORS_BY_SYM = {
    "GOOGL": "Communication Services",
    "AMZN": "Consumer Cyclical",
    "JPM": "Financial Services",
    "JNJ": "Healthcare",
    "XOM": "Energy",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, abs_tol=None, tol_pct=None):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs_tol is not None:
        return abs(a - b) <= abs_tol
    if abs(b) < 1e-6:
        return abs(a) < 0.01
    return abs(a - b) / abs(b) * 100 <= (tol_pct or 1.0)


def get_expected_data():
    """Query DB for expected market caps + prices."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    syms = tuple(SYMBOLS)
    cur.execute(
        "SELECT symbol, (data->>'marketCap')::float, (data->>'trailingPE')::float, "
        "(data->>'fiftyTwoWeekHigh')::float, (data->>'fiftyTwoWeekLow')::float "
        "FROM yf.stock_info WHERE symbol IN %s",
        (syms,),
    )
    info_rows = cur.fetchall()
    cur.execute(
        "SELECT sp.symbol, sp.close FROM yf.stock_prices sp WHERE sp.symbol IN %s "
        "AND sp.date = (SELECT MAX(date) FROM yf.stock_prices WHERE symbol = sp.symbol)",
        (syms,),
    )
    prices = {r[0]: float(r[1]) for r in cur.fetchall()}
    conn.close()
    info = {r[0]: {"mktcap_b": float(r[1]) / 1_000_000_000 if r[1] else None,
                   "pe": float(r[2]) if r[2] else None,
                   "high": float(r[3]) if r[3] else None,
                   "low": float(r[4]) if r[4] else None}
            for r in info_rows}
    avg_price = sum(prices.values()) / len(prices) if prices else 0
    return prices, info, avg_price


def check_excel(agent_ws, prices, info, avg_price):
    print("\n=== Check 1: Sector_Comparison.xlsx ===")
    path = os.path.join(agent_ws, "Sector_Comparison.xlsx")
    check("File Sector_Comparison.xlsx exists", os.path.isfile(path))
    if not os.path.isfile(path):
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel is readable", False, str(e))
        return

    # ---------- Metrics ----------
    metrics_ws = None
    for sname in wb.sheetnames:
        if sname.lower() == "metrics":
            metrics_ws = wb[sname]
            break
    if metrics_ws is None:
        for sname in wb.sheetnames:
            if "metric" in sname.lower():
                metrics_ws = wb[sname]
                break
    check("Sheet 'Metrics' exists", metrics_ws is not None, f"Sheets: {wb.sheetnames}")

    if metrics_ws is not None:
        all_rows = list(metrics_ws.iter_rows(values_only=True))
        header = [str(c or "").strip() for c in (all_rows[0] if all_rows else [])]
        REQUIRED = ["Symbol", "Company", "Sector", "Market_Cap_B", "Current_Price",
                    "PE_Ratio", "High_52w", "Low_52w"]
        for col in REQUIRED:
            check(f"Metrics column '{col}'", col in header, f"got {header}")
        h_idx = {h: i for i, h in enumerate(header)}

        rows = [r for r in all_rows[1:] if r and any(c is not None for c in r)]
        check("Metrics sheet has exactly 5 rows", len(rows) == 5, f"Got {len(rows)}")

        sym_idx = h_idx.get("Symbol", 0)
        agent_lookup = {str(r[sym_idx]).strip(): r for r in rows}
        for sym in SYMBOLS:
            row = agent_lookup.get(sym)
            check(f"Metrics has row for {sym}", row is not None)
            if row is None:
                continue
            # Sector
            if "Sector" in h_idx:
                actual = str(row[h_idx["Sector"]] or "").strip()
                expected = SECTORS_BY_SYM[sym]
                check(f"{sym} Sector", actual == expected, f"actual={actual} expected={expected}")
            # Market_Cap_B (tol 1.0 billion absolute)
            if "Market_Cap_B" in h_idx:
                actual = row[h_idx["Market_Cap_B"]]
                expected = info.get(sym, {}).get("mktcap_b")
                if expected is not None:
                    check(f"{sym} Market_Cap_B", num_close(actual, expected, abs_tol=1.0),
                          f"actual={actual} expected={expected:.2f}")
            # Current_Price (tol 0.05)
            if "Current_Price" in h_idx:
                actual = row[h_idx["Current_Price"]]
                expected = prices.get(sym)
                check(f"{sym} Current_Price", num_close(actual, expected, abs_tol=0.05),
                      f"actual={actual} expected={expected}")
            # PE_Ratio (tol 0.5)
            if "PE_Ratio" in h_idx:
                actual = row[h_idx["PE_Ratio"]]
                expected = info.get(sym, {}).get("pe")
                if expected is None:
                    # Accept N/A
                    check(f"{sym} PE_Ratio is N/A", str(actual or "").upper() == "N/A" or actual is None,
                          f"actual={actual}")
                else:
                    check(f"{sym} PE_Ratio", num_close(actual, expected, abs_tol=0.5),
                          f"actual={actual} expected={expected:.2f}")
            # High_52w (tol 1.0)
            if "High_52w" in h_idx:
                actual = row[h_idx["High_52w"]]
                expected = info.get(sym, {}).get("high")
                if expected is not None:
                    check(f"{sym} High_52w", num_close(actual, expected, abs_tol=1.0),
                          f"actual={actual} expected={expected}")
            # Low_52w (tol 1.0)
            if "Low_52w" in h_idx:
                actual = row[h_idx["Low_52w"]]
                expected = info.get(sym, {}).get("low")
                if expected is not None:
                    check(f"{sym} Low_52w", num_close(actual, expected, abs_tol=1.0),
                          f"actual={actual} expected={expected}")

    # ---------- Sector_Summary ----------
    summary_ws = None
    for sname in wb.sheetnames:
        if "summary" in sname.lower():
            summary_ws = wb[sname]
            break
    check("Sheet 'Sector_Summary' exists", summary_ws is not None, f"Sheets: {wb.sheetnames}")

    if summary_ws is not None:
        rows = list(summary_ws.iter_rows(values_only=True))
        header = [str(c or "").strip() for c in (rows[0] if rows else [])]
        REQUIRED2 = ["Sector", "Symbol", "Market_Cap_B", "Price_Assessment"]
        for col in REQUIRED2:
            check(f"Sector_Summary column '{col}'", col in header, f"got {header}")
        h_idx2 = {h: i for i, h in enumerate(header)}

        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)]
        check("Sector_Summary sheet has exactly 5 rows", len(data_rows) == 5, f"Got {len(data_rows)}")

        # Per-row Price_Assessment validation
        if "Symbol" in h_idx2 and "Price_Assessment" in h_idx2:
            sy_idx = h_idx2["Symbol"]
            pa_idx = h_idx2["Price_Assessment"]
            for r in data_rows:
                sym = str(r[sy_idx] or "").strip()
                actual_pa = str(r[pa_idx] or "").strip()
                if sym in prices:
                    expected_pa = "Above_Avg" if prices[sym] > avg_price else "Below_Avg"
                    check(
                        f"{sym} Price_Assessment",
                        actual_pa == expected_pa,
                        f"actual='{actual_pa}' expected='{expected_pa}' (price={prices[sym]} vs avg={avg_price:.2f})",
                    )


def check_gform():
    print("\n=== Check 2: Google Form ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    check("At least one Google Form exists", len(forms) > 0,
          "No forms found in gform.forms")

    # Match exact title
    target_title = "investment preference survey"
    found_form_id = None
    for form_id, title in forms:
        if (title or "").strip().lower() == target_title:
            found_form_id = form_id
            break
    check("Form titled 'Investment Preference Survey'", found_form_id is not None,
          f"Forms: {[r[1] for r in forms]}")

    if found_form_id:
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (found_form_id,))
        q_count = cur.fetchone()[0]
        check("Form has exactly 4 questions", q_count == 4, f"Got {q_count} questions")

        cur.execute(
            "SELECT title, config FROM gform.questions WHERE form_id = %s ORDER BY position",
            (found_form_id,),
        )
        questions = cur.fetchall()
        q_titles = [q[0] for q in questions]
        q_text = " ".join(q_titles).lower()

        EXPECTED_QUESTIONS = [
            ("Q1: sector preference", "sector"),
            ("Q2: investment horizon", "horizon"),
            ("Q3: risk tolerance", "risk"),
            ("Q4: financial metric", "metric"),
        ]
        for label, kw in EXPECTED_QUESTIONS:
            check(f"Form has {label}", kw in q_text, f"Questions: {q_titles}")

        # Verify options for Q1: sector list
        EXPECTED_OPTIONS = {
            "sector": ["communication services", "consumer cyclical", "financial services", "healthcare", "energy"],
            "horizon": ["short-term", "medium-term", "long-term"],
            "risk": ["low", "medium", "high"],
            "metric": ["dividend yield", "p/e ratio", "revenue growth rate", "market capitalization"],
        }
        # Parse config as jsonb dict, extract options[].value list, normalize
        import json as _json
        for q_title, q_config in questions:
            q_low = (q_title or "").lower()
            # q_config is jsonb -> psycopg2 returns a dict already
            if isinstance(q_config, dict):
                cfg = q_config
            else:
                try:
                    cfg = _json.loads(q_config) if q_config else {}
                except Exception:
                    cfg = {}
            opt_values = []
            for o in cfg.get("options") or []:
                if isinstance(o, dict):
                    opt_values.append(str(o.get("value") or "").strip().lower())
                else:
                    opt_values.append(str(o).strip().lower())
            opt_text = " | ".join(opt_values)
            for keyword, opts in EXPECTED_OPTIONS.items():
                if keyword in q_low:
                    # Each expected option must appear as a substring of one of the option values
                    missing_opts = [
                        o for o in opts
                        if not any(o in v for v in opt_values)
                    ]
                    check(
                        f"Q '{q_title[:40]}' includes all required options",
                        not missing_opts and len(opt_values) >= len(opts),
                        f"missing: {missing_opts}; got_options: {opt_text}",
                    )

    cur.close()
    conn.close()


def check_email(prices, info):
    print("\n=== Check 3: Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Match recipient + subject contains 'Sector Comparison'
    cur.execute(
        "SELECT subject, from_addr, to_addr, body_text FROM email.messages "
        "WHERE to_addr::text ILIKE '%investors@fund.example.com%'"
    )
    rows = cur.fetchall()
    check("Email sent to investors@fund.example.com", len(rows) >= 1, "no matching recipient")

    # Find email with subject containing 'Sector Comparison'
    matched = None
    for subj, fr, ta, body in rows:
        if "sector comparison" in (subj or "").lower():
            matched = (subj, fr, ta, body)
            break
    check("Email subject contains 'Sector Comparison'", matched is not None,
          f"subjects: {[r[0] for r in rows]}")
    if matched is not None:
        subj, fr, ta, body = matched
        check("Email from research@fund.example.com",
              "research@fund.example.com" in str(fr or "").lower(), f"from={fr}")
        body_l = (body or "").lower()
        # Largest market cap = GOOGL
        largest_mktcap_sym = max(info.keys(), key=lambda s: info[s].get("mktcap_b") or 0)
        check(f"Email body mentions largest market cap {largest_mktcap_sym}",
              largest_mktcap_sym.lower() in body_l, f"body[:300]: {body_l[:300]}")
        # Highest PE
        highest_pe_sym = max(info.keys(), key=lambda s: info[s].get("pe") or 0)
        check(f"Email body mentions highest P/E {highest_pe_sym}",
              highest_pe_sym.lower() in body_l, f"body[:300]: {body_l[:300]}")
        # 52-week range mentioned for >=2 stocks: require the SYMBOL to appear within
        # ~200 chars of both the high and low values, to avoid integers matching
        # unrelated numbers elsewhere in the body.
        import re as _re
        mentions = 0
        for sym in SYMBOLS:
            high = info.get(sym, {}).get("high")
            low = info.get(sym, {}).get("low")
            if not (high and low):
                continue
            sym_low = sym.lower()
            # Find every occurrence of the symbol; check a +/- 200-char window for both values
            sym_positions = [m.start() for m in _re.finditer(_re.escape(sym_low), body_l)]
            high_strs = [str(int(high)), f"{float(high):.2f}"]
            low_strs = [str(int(low)), f"{float(low):.2f}"]
            ok = False
            for pos in sym_positions:
                window = body_l[max(0, pos - 200): pos + 200]
                if any(h in window for h in high_strs) and any(l in window for l in low_strs):
                    ok = True
                    break
            if ok:
                mentions += 1
        check("Email body mentions 52-week range for >=2 stocks (symbol-near-value)", mentions >= 2,
              f"only {mentions} stocks have a symbol-near-range pattern")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Evaluation: yf-sector-comparison-gform-excel-email ===")

    try:
        prices, info, avg_price = get_expected_data()
    except Exception as e:
        print(f"DB query failed: {e}")
        # Fall back to hardcoded values
        prices = {"GOOGL": 300.88, "AMZN": 218.94, "JPM": 293.55, "JNJ": 239.63, "XOM": 150.76}
        info = {
            "GOOGL": {"mktcap_b": 3639.75, "pe": 27.81, "high": 349, "low": 129.87},
            "AMZN": {"mktcap_b": 2350.30, "pe": 30.49, "high": 258.6, "low": 151.61},
            "JPM": {"mktcap_b": 791.71, "pe": 14.66, "high": 335.87, "low": 172.68},
            "JNJ": {"mktcap_b": 577.48, "pe": 21.71, "high": 251.71, "low": 134.92},
            "XOM": {"mktcap_b": 628.18, "pe": 22.50, "high": 159.61, "low": 94.55},
        }
        avg_price = sum(prices.values()) / len(prices)

    check_excel(args.agent_workspace, prices, info, avg_price)
    check_gform()
    check_email(prices, info)

    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"pass": PASS_COUNT, "fail": FAIL_COUNT}, f)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
