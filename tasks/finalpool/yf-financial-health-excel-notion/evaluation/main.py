"""Evaluation for yf-financial-health-excel-notion."""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

SYMBOLS = ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]


def num_close(a, b, rel_tol=0.02, abs_tol=10):
    try:
        af = float(a); bf = float(b)
        if bf == 0:
            return abs(af) <= abs_tol
        return abs(af - bf) <= max(abs_tol, abs(bf) * rel_tol)
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def fetch_expected():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    income = {}
    balance = {}
    for sym in SYMBOLS:
        cur.execute("""SELECT (data->>'Total Revenue')::float, (data->>'Net Income')::float
            FROM yf.financial_statements
            WHERE symbol=%s AND stmt_type='income_stmt' AND freq='annual'
            ORDER BY period_end DESC LIMIT 1""", (sym,))
        r = cur.fetchone()
        if r:
            rev, ni = r
            income[sym] = {"revenue": float(rev), "net_income": float(ni),
                           "margin": round(float(ni) / float(rev) * 100, 1) if rev else 0}
        cur.execute("""SELECT (data->>'Total Assets')::float,
                              (data->>'Total Liabilities Net Minority Interest')::float,
                              (data->>'Stockholders Equity')::float
            FROM yf.financial_statements
            WHERE symbol=%s AND stmt_type='balance_sheet' AND freq='annual'
            ORDER BY period_end DESC LIMIT 1""", (sym,))
        r = cur.fetchone()
        if r:
            a, l, e = r
            balance[sym] = {"assets": float(a), "liabilities": float(l), "equity": float(e)}
    cur.close()
    conn.close()
    return income, balance


def check_excel(agent_workspace, income, balance):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Financial_Health_Report.xlsx")
    if not os.path.exists(path):
        return ["Financial_Health_Report.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return [f"Error reading Excel: {e}"]

    # === Income Statement ===
    rows = load_sheet_rows(wb, "Income Statement")
    if rows is None:
        errors.append("Sheet 'Income Statement' not found")
    else:
        # Header
        if rows and rows[0]:
            ah = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
            expected = ["symbol", "total_revenue", "net_income", "profit_margin_pct"]
            for i, eh in enumerate(expected):
                got = ah[i] if i < len(ah) else "MISSING"
                if got != eh:
                    errors.append(f"IncomeStmt header[{i}]: expected '{eh}' got '{got}'")

        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        if len(data_rows) != 5:
            errors.append(f"Income Statement has {len(data_rows)} rows, expected 5")

        # Sort: alphabetical
        symbols_in_order = [str(r[0]).strip().upper() for r in data_rows]
        if symbols_in_order != sorted(SYMBOLS):
            errors.append(f"Income Statement not alphabetically sorted: {symbols_in_order}")

        # Per-symbol numeric checks
        a_lookup = {str(r[0]).strip().upper(): r for r in data_rows}
        for sym in SYMBOLS:
            r = a_lookup.get(sym)
            if r is None:
                errors.append(f"Symbol {sym} missing from Income Statement")
                continue
            if sym not in income:
                continue
            exp = income[sym]
            if not num_close(r[1], exp["revenue"]):
                errors.append(f"{sym} Total_Revenue: {r[1]} vs expected ~{exp['revenue']:.0f}")
            if not num_close(r[2], exp["net_income"]):
                errors.append(f"{sym} Net_Income: {r[2]} vs expected ~{exp['net_income']:.0f}")
            if not num_close(r[3], exp["margin"], rel_tol=0, abs_tol=0.2):
                errors.append(f"{sym} Profit_Margin_Pct: {r[3]} vs expected {exp['margin']}")

    # === Balance Sheet ===
    rows2 = load_sheet_rows(wb, "Balance Sheet")
    if rows2 is None:
        errors.append("Sheet 'Balance Sheet' not found")
    else:
        if rows2 and rows2[0]:
            ah = [str(h or "").strip().lower().replace(" ", "_") for h in rows2[0]]
            expected = ["symbol", "total_assets", "total_liabilities", "equity"]
            for i, eh in enumerate(expected):
                got = ah[i] if i < len(ah) else "MISSING"
                if got != eh:
                    errors.append(f"BalanceSheet header[{i}]: expected '{eh}' got '{got}'")

        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        if len(data_rows2) != 5:
            errors.append(f"Balance Sheet has {len(data_rows2)} rows, expected 5")

        symbols_in_order = [str(r[0]).strip().upper() for r in data_rows2]
        if symbols_in_order != sorted(SYMBOLS):
            errors.append(f"Balance Sheet not alphabetically sorted: {symbols_in_order}")

        b_lookup = {str(r[0]).strip().upper(): r for r in data_rows2}
        for sym in SYMBOLS:
            r = b_lookup.get(sym)
            if r is None:
                errors.append(f"Symbol {sym} missing from Balance Sheet")
                continue
            if sym not in balance:
                continue
            exp = balance[sym]
            if not num_close(r[1], exp["assets"]):
                errors.append(f"{sym} Total_Assets: {r[1]} vs expected ~{exp['assets']:.0f}")
            if not num_close(r[2], exp["liabilities"]):
                errors.append(f"{sym} Total_Liabilities: {r[2]} vs expected ~{exp['liabilities']:.0f}")
            if not num_close(r[3], exp["equity"]):
                errors.append(f"{sym} Equity: {r[3]} vs expected ~{exp['equity']:.0f}")

    return errors


def check_notion(income, balance):
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
        rows = cur.fetchall()

        # Find page titled 'Financial Health Dashboard'
        target = "financial health dashboard"
        target_page_id = None
        for pid, props in rows:
            if not isinstance(props, dict):
                continue
            for prop in props.values():
                if isinstance(prop, dict) and prop.get("type") == "title":
                    title_arr = prop.get("title", [])
                    title_text = "".join(t.get("plain_text", "") or t.get("text", {}).get("content", "") for t in title_arr)
                    if target in title_text.strip().lower():
                        target_page_id = pid
                        break
            if target_page_id:
                break

        if not target_page_id:
            errors.append(f"No Notion page titled '{target}' found")
            cur.close(); conn.close()
            return errors

        # Get page block content from notion.blocks.block_data
        body_text = ""
        try:
            cur.execute("SELECT block_data FROM notion.blocks WHERE parent_id = %s AND archived = false", (target_page_id,))
            body_text = " ".join(json.dumps(r[0]) if r[0] else "" for r in cur.fetchall())
        except Exception as ex:
            try:
                conn.rollback()
                cur.execute("SELECT * FROM notion.blocks WHERE parent_id = %s LIMIT 100", (target_page_id,))
                body_text = " ".join(str(r) for r in cur.fetchall())
            except Exception:
                pass
        body_lower = body_text.lower()

        # Compute who has highest revenue
        if income:
            top_rev_sym = max(income.keys(), key=lambda s: income[s]["revenue"])
            top_margin_sym = max(income.keys(), key=lambda s: income[s]["margin"])
            if top_rev_sym.lower() not in body_lower:
                errors.append(f"Notion page doesn't mention highest-revenue symbol {top_rev_sym}")
            if top_margin_sym.lower() not in body_lower:
                errors.append(f"Notion page doesn't mention best-margin symbol {top_margin_sym}")

        # Strongest balance sheet: highest equity (typically) — accept either highest equity
        # or highest assets-to-liabilities ratio as candidates (handle ties)
        if balance:
            equity_max = max(balance.values(), key=lambda b: b["equity"])["equity"]
            ratio_max = max(
                balance.values(),
                key=lambda b: (b["assets"] / b["liabilities"]) if b["liabilities"] else 0
            )
            ratio_max_val = (ratio_max["assets"] / ratio_max["liabilities"]) if ratio_max["liabilities"] else 0
            equity_candidates = {s.lower() for s, b in balance.items() if b["equity"] == equity_max}
            ratio_candidates = {
                s.lower() for s, b in balance.items()
                if b["liabilities"] and abs((b["assets"] / b["liabilities"]) - ratio_max_val) < 1e-6
            }
            candidates = equity_candidates | ratio_candidates
            if not any(c in body_lower for c in candidates):
                errors.append(
                    f"Notion page doesn't mention strongest-balance-sheet symbol (one of {sorted(candidates)})"
                )

        cur.close(); conn.close()
    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, COALESCE(body_text, body_html, '')
            FROM email.messages
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()

        target_subj = "financial health report"
        target_to = "finance_team@company.com"
        match = None
        for subj, to_addr, body in rows:
            sl = (subj or "").strip().lower()
            tos = []
            if isinstance(to_addr, list):
                tos = [str(t).lower() for t in to_addr]
            elif isinstance(to_addr, str):
                try:
                    p = json.loads(to_addr)
                    tos = [str(t).lower() for t in p] if isinstance(p, list) else [to_addr.lower()]
                except Exception:
                    tos = [to_addr.lower()]
            if target_subj in sl and any(target_to in t for t in tos):
                match = (subj, to_addr, body)
                break

        if not match:
            errors.append(f"No email subject '{target_subj}' to {target_to} found")
            return errors

        body_lower = (match[2] or "").lower()
        if "excel" not in body_lower and "financial_health_report" not in body_lower:
            errors.append("Email body should mention the Excel file")
        if "notion" not in body_lower and "dashboard" not in body_lower:
            errors.append("Email body should mention the Notion page/dashboard")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    income, balance = fetch_expected()

    all_errors = []
    print("  Checking Excel file...")
    errs = check_excel(agent_ws, income, balance)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion page...")
    errs = check_notion(income, balance)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
