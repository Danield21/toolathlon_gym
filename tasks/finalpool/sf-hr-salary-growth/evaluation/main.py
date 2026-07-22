"""Evaluation for sf-hr-salary-growth."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Salary_Growth.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Salary_Growth.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Salary Growth
    print(f"  Checking Salary Growth...")
    a_rows = load_sheet_rows(agent_wb, "Salary Growth")
    g_rows = load_sheet_rows(gt_wb, "Salary Growth")
    if a_rows is None:
        all_errors.append("Sheet 'Salary Growth' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Salary Growth' not found in groundtruth")
    else:
        sheet_name = "Salary Growth"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        gt_keys = set()
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            gt_keys.add(key)
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Counts are deterministic: require exact match (tol=0)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    errors.append(f"{key}.Employees_With_Changes: {a_row[1]} vs {g_row[1]} (exact)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0):
                    errors.append(f"{key}.Total_Changes: {a_row[2]} vs {g_row[2]} (exact)")

            # Avg_Increase: tighten tol from 5.0 to 1% relative or 1.0 absolute
            if len(a_row) > 3 and len(g_row) > 3:
                gv3 = g_row[3]
                try:
                    gf3 = float(gv3) if gv3 is not None else None
                except (TypeError, ValueError):
                    gf3 = None
                tol3 = max(1.0, abs(gf3) * 0.01) if gf3 is not None else 1.0
                if not num_close(a_row[3], g_row[3], tol3):
                    errors.append(f"{key}.Avg_Increase: {a_row[3]} vs {g_row[3]} (tol={tol3:.2f})")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.2):
                    errors.append(f"{key}.Avg_Pct_Change: {a_row[4]} vs {g_row[4]} (tol=0.2)")
        # Check for extra rows in agent output that should not exist
        extra_keys = [k for k in a_lookup.keys() if k not in gt_keys]
        if extra_keys:
            errors.append(f"Salary Growth has {len(extra_keys)} extra unexpected row(s): {extra_keys[:5]}")

        # Verify sort order: rows in a_data (after header) must be sorted by Avg_Pct_Change descending
        # (only validate against rows that match GT keys, in agent's order)
        ordered_pcts = []
        for row in a_data:
            if not row or row[0] is None:
                continue
            k = str(row[0]).strip().lower()
            if k not in gt_keys:
                continue
            try:
                p = float(row[4]) if len(row) > 4 and row[4] is not None else None
            except (TypeError, ValueError):
                p = None
            if p is not None:
                ordered_pcts.append((k, p))
        # Check that pcts are non-increasing
        for i in range(len(ordered_pcts) - 1):
            cur_k, cur_p = ordered_pcts[i]
            nxt_k, nxt_p = ordered_pcts[i + 1]
            # allow ties; require descending order
            if cur_p < nxt_p - 1e-9:
                errors.append(
                    f"Sort order violation: '{cur_k}' Avg_Pct_Change={cur_p} < '{nxt_k}' Avg_Pct_Change={nxt_p} (must be descending)"
                )
                break
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        gt_keys = set()
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            gt_keys.add(key)
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                gv = g_row[1]
                av = a_row[1]
                # Best_Dept (text) -> str_match; numeric metrics scaled tol
                try:
                    gv_f = float(gv) if gv is not None else None
                except (TypeError, ValueError):
                    gv_f = None
                if gv_f is not None:
                    if "total_salary_changes" in key:
                        tol = 0  # exact int count - deterministic aggregate
                    elif "overall_avg_increase" in key:
                        # tighten from 5.0 to 1% relative or 1.0 absolute
                        tol = max(1.0, abs(gv_f) * 0.01)
                    else:
                        tol = 1.0
                    if not num_close(av, gv, tol):
                        errors.append(f"{key}.Value: {av} vs {gv} (tol={tol})")
                else:
                    if not str_match(av, gv):
                        errors.append(f"{key}.Value: '{av}' vs '{gv}' (str_match)")
        # Check for extra rows in agent Summary
        extra_keys = [k for k in a_lookup.keys() if k not in gt_keys]
        if extra_keys:
            errors.append(f"Summary has {len(extra_keys)} extra unexpected row(s): {extra_keys[:5]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
