"""Evaluation script for pw-wc-review-analysis-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-wc-review-analysis-excel-notion"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def _norm(s):
    """Case-insensitive, space/underscore-insensitive key normalizer.
    Used so reasonable column/key variants (e.g. 'External Rating' vs
    'External_Rating') do not cause false FAILs."""
    return str(s or "").strip().lower().replace(" ", "").replace("_", "")

def _collect_text(value, out):
    """Recursively collect every string leaf (text.content / plain_text / any
    key or value) from a nested notion block/property structure.  Robust to all
    known shapes: dict, list, rich-text arrays, bare strings."""
    if isinstance(value, str):
        out.append(value)
    elif isinstance(value, (list, tuple)):
        for v in value:
            _collect_text(v, out)
    elif isinstance(value, dict):
        for v in value.values():
            _collect_text(v, out)

def _sentiment_bucket(v):
    """Collapse equivalent sentiment labels (case-insensitive) into buckets so
    that reasonable binning (e.g. 'Very Positive' -> 'Positive', or
    'Very Negative' -> 'Negative') is not treated as a wrong answer."""
    s = str(v or "").strip().lower()
    if s in ("very positive", "positive", "pos", "+"):
        return "positive"
    if s in ("very negative", "negative", "neg", "-"):
        return "negative"
    if s in ("mixed",):
        return "mixed"
    if s in ("neutral",):
        return "neutral"
    return s

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Review_Analysis_Report.xlsx")
    check("Review_Analysis_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Review_Analysis_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [_norm(c.value) for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _norm(_exp) in _headers, f"headers: {_headers[:8]}")

        check("External_Reviews sheet exists", "External_Reviews" in wb.sheetnames)
        if "External_Reviews" in wb.sheetnames:
            ws = wb["External_Reviews"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("External_Reviews has >= 8 rows", len(data_rows) >= 8, f"got {len(data_rows)}")

            check_columns('External_Reviews', ['Product', 'External_Rating', 'Review_Count', 'Sentiment', 'Common_Complaint'], 8)
        check("Review_Summary sheet exists", "Review_Summary" in wb.sheetnames)
        if "Review_Summary" in wb.sheetnames:
            ws = wb["Review_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Review_Summary has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            check_columns('Review_Summary', ['Metric', 'Value'], 4)
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Require exact title match for 'Product Review Tracker'
            cur.execute(
                "SELECT id, properties FROM notion.pages "
                "WHERE properties::text ILIKE %s AND archived = false",
                ('%Product Review Tracker%',))
            rows = cur.fetchall()
            check("Notion 'Product Review Tracker' page exists",
                  len(rows) >= 1, f"got {len(rows)} matching pages")
            # Validate page (or its descendant blocks) mentions products needing
            # attention and their common complaints.  The task only asks for a
            # summary of the attention products (Smart Watch, USB-C Hub) and
            # their complaints (Screen durability, Compatibility issues), so a
            # faithful paraphrase must pass.  Accept the product names (incl.
            # common spellings) or the complaint themes (incl. reasonable
            # synonyms such as connectivity).  Text is collected from ALL
            # descendant blocks (recursively, so summaries nested >1 level
            # deep are found) plus the page properties, robust to every known
            # block_data shape.
            tracker_ids = [r[0] for r in rows]
            content_ok = False
            if tracker_ids:
                collected = []
                frontier = list(tracker_ids)
                seen = set()
                for _depth in range(20):
                    if not frontier:
                        break
                    cur.execute(
                        "SELECT id, block_data FROM notion.blocks "
                        "WHERE parent_id = ANY(%s)", (frontier,))
                    batch = cur.fetchall()
                    nxt = []
                    for bid, bdata in batch:
                        if bid is None or bid in seen:
                            continue
                        seen.add(bid)
                        nxt.append(bid)
                        if bdata is not None:
                            _collect_text(bdata, collected)
                    frontier = nxt
                blob = (" ".join(collected) + " " +
                        " ".join(str(r[1] or "") for r in rows)).lower()
                attention_kws = [
                    "usb-c hub", "usb c hub", "usb hub", "usbchub", "usbhub",
                    "smart watch", "smartwatch", "watch", "hub",
                    "compatibility", "connectivity", "screen durability",
                    "durability", "needs attention",
                ]
                content_ok = any(kw in blob for kw in attention_kws)
            check("Notion tracker page mentions products needing attention",
                  content_ok, "no attention products / complaints found in page content")
            conn.close()
        except Exception as e:
            check("Notion check", False, str(e))
        # Validate Sentiment column values are from enumerated set
        if "External_Reviews" in wb.sheetnames:
            ws = wb["External_Reviews"]
            headers_raw = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            headers_norm = [_norm(c.value) for c in ws[1]]
            if _norm("sentiment") in headers_norm:
                s_idx = headers_norm.index(_norm("sentiment"))
                allowed = {"positive", "very positive", "negative", "very negative",
                           "mixed", "neutral"}
                vals = [str(r[s_idx]).strip().lower() for r in ws.iter_rows(min_row=2, values_only=True)
                        if r and r[s_idx] is not None]
                bad = [v for v in vals if v not in allowed]
                check("External_Reviews Sentiment values valid",
                      len(vals) >= 5 and len(bad) == 0,
                      f"bad values: {bad[:3]} of {len(vals)}")

            # NEW (R14): per-product value validation against GT.
            # Build lookup keyed by product name (case-insensitive).
            if gt_wb and "External_Reviews" in gt_wb.sheetnames:
                gt_ws = gt_wb["External_Reviews"]
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
                agent_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]

                def col_idx(headers, name):
                    nm = _norm(name)
                    hn = [_norm(h) for h in headers]
                    return hn.index(nm) if nm in hn else -1

                a_p_idx = col_idx(headers_raw, "product")
                a_r_idx = col_idx(headers_raw, "external_rating")
                a_s_idx = col_idx(headers_raw, "sentiment")
                a_c_idx = col_idx(headers_raw, "common_complaint")
                g_p_idx = col_idx(gt_headers, "product")
                g_r_idx = col_idx(gt_headers, "external_rating")
                g_s_idx = col_idx(gt_headers, "sentiment")
                g_c_idx = col_idx(gt_headers, "common_complaint")

                if a_p_idx >= 0 and g_p_idx >= 0:
                    a_lookup = {_norm(r[a_p_idx]): r for r in agent_rows
                                if a_p_idx < len(r) and r[a_p_idx]}
                    for g_row in gt_rows:
                        if g_p_idx >= len(g_row) or not g_row[g_p_idx]:
                            continue
                        prod_key = _norm(g_row[g_p_idx])
                        a_row = a_lookup.get(prod_key)
                        check(f"External_Reviews has product '{g_row[g_p_idx]}'",
                              a_row is not None,
                              f"missing product key {prod_key!r}")
                        if a_row is None:
                            continue
                        # Sentiment - must match GT semantically.  Compare
                        # sentiment buckets so 'Very Positive'/'Positive' and
                        # 'Very Negative'/'Negative' are treated as equivalent
                        # (a model that bins intensities is still correct).
                        if g_s_idx >= 0 and g_s_idx < len(g_row) and a_s_idx >= 0 and a_s_idx < len(a_row):
                            gv = str(g_row[g_s_idx]).strip().lower() if g_row[g_s_idx] is not None else ""
                            av = str(a_row[a_s_idx]).strip().lower() if a_row[a_s_idx] is not None else ""
                            check(f"External_Reviews '{g_row[g_p_idx]}' Sentiment matches GT",
                                  _sentiment_bucket(gv) == _sentiment_bucket(av),
                                  f"expected {gv!r}, got {av!r}")
                        # External_Rating - tolerant of integer rounding (max
                        # rounding error 0.5).  An empty/unparseable agent cell
                        # must FAIL (not silently skip), since this is a core
                        # value extracted from the source data.
                        if g_r_idx >= 0 and g_r_idx < len(g_row) and a_r_idx >= 0 and a_r_idx < len(a_row):
                            gf = safe_float(g_row[g_r_idx])
                            af = safe_float(a_row[a_r_idx])
                            if gf is not None:
                                check(f"External_Reviews '{g_row[g_p_idx]}' External_Rating ~ {gf}",
                                      af is not None and abs(gf - af) <= 0.5,
                                      f"got {af!r}")

        # Validate Review_Summary metrics by name and value (now compared to GT)
        if "Review_Summary" in wb.sheetnames:
            ws = wb["Review_Summary"]
            metrics = {}
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r and r[0]:
                    metrics[_norm(r[0])] = r[1] if len(r) > 1 else None
            required = [_norm(m) for m in (
                "total_products_reviewed", "avg_external_rating",
                "positive_products", "needs_attention")]
            missing = [m for m in required if m not in metrics]
            check("Review_Summary has required metric keys",
                  len(missing) == 0, f"missing: {missing}")

            # NEW (R14): metric value validation against GT.
            if gt_wb and "Review_Summary" in gt_wb.sheetnames:
                gt_ws = gt_wb["Review_Summary"]
                gt_metrics = {}
                for r in gt_ws.iter_rows(min_row=2, values_only=True):
                    if r and r[0]:
                        gt_metrics[_norm(r[0])] = r[1] if len(r) > 1 else None
                # Exact match for counts; tolerance for avg.  An empty or
                # unparseable agent value must FAIL (not silently skip) so a
                # blank Value cell cannot pass the numeric checks.
                for m in ("total_products_reviewed", "positive_products", "needs_attention"):
                    m = _norm(m)
                    if m in gt_metrics and m in metrics:
                        gf = safe_float(gt_metrics[m])
                        af = safe_float(metrics[m])
                        if gf is not None:
                            check(f"Review_Summary {m} == {gf}",
                                  af is not None and abs(gf - af) <= 1, f"got {af!r}")
                m = _norm("avg_external_rating")
                if m in gt_metrics and m in metrics:
                    gf = safe_float(gt_metrics[m])
                    af = safe_float(metrics[m])
                    if gf is not None:
                        check(f"Review_Summary avg_external_rating ~ {gf}",
                              af is not None and abs(gf - af) <= 0.25, f"got {af!r}")
        check("review_analyzer.py exists", os.path.exists(os.path.join(agent_workspace, "review_analyzer.py")))
        # Verify all 3 JSON files mentioned in task.md exist
        for _fname in ("external_reviews.json", "internal_reviews.json", "review_insights.json"):
            _fp = os.path.join(agent_workspace, _fname)
            check(f"{_fname} exists", os.path.exists(_fp))
            if os.path.exists(_fp):
                try:
                    with open(_fp, "r", encoding="utf-8") as _f:
                        _data = json.load(_f)
                    check(f"{_fname} is valid non-empty JSON",
                          (isinstance(_data, (list, dict)) and len(_data) > 0),
                          f"type={type(_data).__name__}, len={len(_data) if hasattr(_data, '__len__') else 0}")
                except Exception as _e:
                    check(f"{_fname} is valid non-empty JSON", False, str(_e)[:120])


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
