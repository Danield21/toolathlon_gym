"""Evaluation script for fetch-sf-sales-competitor-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-sf-sales-competitor-excel-notion"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
WARN_COUNT = 0
IS_GT_SELF_TEST = False

def check(name, condition, detail="", db_side=False, gt_optional=False):
    global PASS_COUNT, FAIL_COUNT, WARN_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        if IS_GT_SELF_TEST and (db_side or gt_optional):
            WARN_COUNT += 1
            detail_str = str(detail)[:200] if detail else ""
            print(f"  [WARN] {name} (GT self-test mode): {detail_str}")
        else:
            FAIL_COUNT += 1
            detail_str = str(detail)[:200] if detail else ""
            print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, WARN_COUNT, IS_GT_SELF_TEST
    PASS_COUNT = 0
    FAIL_COUNT = 0
    WARN_COUNT = 0

    # Detect GT self-test mode
    try:
        if groundtruth_workspace and os.path.exists(groundtruth_workspace):
            IS_GT_SELF_TEST = (
                os.path.realpath(agent_workspace) ==
                os.path.realpath(groundtruth_workspace)
            )
    except Exception:
        IS_GT_SELF_TEST = False

    
    excel_path = os.path.join(agent_workspace, "Sales_Competitor_Report.xlsx")
    check("Sales_Competitor_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Sales_Competitor_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        # Deterministic synonym table used as a fallback before invoking the
        # LLM mapper. The task.md describes the primary dimension generically
        # ("department, product, region, or topic"), so an agent may name the
        # column `Dimension`, `Category`, etc. — these still represent the
        # same conceptual key and should pass the column-presence check.
        COLUMN_SYNONYMS = {
            'Region':     ['region', 'area', 'territory', 'market', 'geography',
                           'dimension', 'primary_dimension', 'segment',
                           'category', 'department', 'product', 'topic', 'group'],
            'Order_Count': ['order_count', 'orders', 'order count', 'count',
                            'order_volume', 'volume', 'num_orders', 'order_qty'],
            'Revenue':    ['revenue', 'internal_revenue', 'sales', 'total_sales',
                           'total_revenue', 'internal', 'internal_metric',
                           'internal_value', 'actual'],
            'Market_Size_M':  ['market_size_m', 'market size', 'market_size',
                               'external_benchmark', 'external', 'benchmark',
                               'expected', 'target_market', 'addressable_market'],
            'Market_Penetration_Pct': ['market_penetration_pct', 'market_penetration',
                                       'penetration', 'penetration_pct', 'gap',
                                       'difference', 'delta', 'variance', 'pct_diff'],
            'Metric':     ['metric', 'name', 'kpi', 'key', 'measure'],
            'Value':      ['value', 'amount', 'number', 'val'],
            'Priority':   ['priority', 'rank', 'order', 'level'],
            'Action':     ['action', 'recommendation', 'recommended_action',
                           'next_step', 'item'],
        }

        def _header_match_synonym(expected, headers_lower):
            for syn in COLUMN_SYNONYMS.get(expected, []):
                if syn.lower() in headers_lower:
                    return syn
            for syn in COLUMN_SYNONYMS.get(expected, []):
                for h in headers_lower:
                    if syn.lower() in h or h in syn.lower():
                        return h
            return None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Lookup order: exact -> deterministic synonym -> LLM mapper."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            _raw_headers = [str(c.value).strip() if c.value else "" for c in _ws[1]]
            _headers_lower = [h.lower() for h in _raw_headers]
            for _exp in expected_cols:
                if _exp.lower() in _headers_lower:
                    check(f"{sheet_name} has {_exp} column", True, "exact match")
                    continue
                _syn = _header_match_synonym(_exp, _headers_lower)
                if _syn:
                    check(f"{sheet_name} has {_exp} column", True,
                          f"matched via synonym {_syn!r}")
                    continue
                if _HAS_VERIFY_V2 and gt_wb is not None:
                    _raw_h_v2, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_h_v2,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
                else:
                    check(f"{sheet_name} has {_exp} column",
                          False, f"headers: {_raw_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Region', 'Order_Count', 'Revenue', 'Market_Size_M', 'Market_Penetration_Pct'], 5)

            # Value-level: dimension key values overlap with GT (>=3 of 5).
            # Task.md offers "department, product, region, or topic" as the
            # primary dimension; agent may pick any. We accept overlap with
            # GT region values OR validate that the chosen dimension has a
            # plausible non-empty distinct set.
            if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames:
                try:
                    gt_ws = gt_wb["Data_Analysis"]
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    gt_regions = {str(r[0]).strip().lower() for r in gt_rows if r and r[0]}
                    agent_keys = {str(r[0]).strip().lower() for r in data_rows if r and r[0]}
                    overlap = gt_regions & agent_keys
                    if len(overlap) >= 3:
                        check("Data_Analysis dimension values match GT regions (>=3 of 5)",
                              True, f"overlap={overlap}")
                    else:
                        # Alternative dimension chosen — accept if it has 4-8 distinct
                        # non-empty entries (matching the API/data range)
                        n_distinct = len([k for k in agent_keys if k])
                        check("Data_Analysis primary-dimension has 4-8 distinct values",
                              4 <= n_distinct <= 8,
                              f"distinct_keys={n_distinct}, sample={sorted(agent_keys)[:6]}, "
                              f"gt_region_overlap={len(overlap)}")
                except Exception as _e:
                    check("Data_Analysis dimension value-check", False, str(_e))
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 4)

            # Value-level: a Total_X count exists with value in [4,8].
            # Agent might use Total_Regions / Total_Departments / Total_Products etc.
            # depending on which dimension they chose.
            metric_dict = {}
            for r in data_rows:
                if r and len(r) >= 2 and r[0]:
                    metric_dict[str(r[0]).strip().lower().replace(' ', '_')] = r[1]
            total_keys = ('total_regions', 'total_departments', 'total_products',
                          'total_topics', 'total_categories', 'total_dimensions',
                          'total_segments', 'total_items', 'total_count',
                          'total_groups', 'total')
            tr = None
            for tk in total_keys:
                if tk in metric_dict:
                    tr = safe_float(metric_dict.get(tk))
                    break
            check("Metrics has a Total_<dimension> count in [4,8]",
                  tr is not None and 4 <= tr <= 8,
                  f"got Total_<dim>={tr} (sampled keys: {list(metric_dict.keys())[:8]})")
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action', 'Region'], 2)
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Tightened: require BOTH 'competitor' AND ('dashboard' OR 'analysis')
            cur.execute(
                """SELECT id, properties FROM notion.pages
                   WHERE archived = false
                     AND properties::text ILIKE %s
                     AND (properties::text ILIKE %s OR properties::text ILIKE %s)""",
                ('%competitor%', '%dashboard%', '%analysis%'))
            pages = cur.fetchall()
            check("Notion 'Sf Competitor Dashboard' page created",
                  len(pages) >= 1, f"found {len(pages)} matching pages",
                  db_side=True)
            conn.close()
        except Exception as e:
            check("Notion check", False, str(e), db_side=True)

        check("sf_competitor_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "sf_competitor_processor.py")),
              gt_optional=True)


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
