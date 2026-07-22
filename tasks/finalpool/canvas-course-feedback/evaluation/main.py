"""
Evaluation script for canvas-course-feedback task.

Check 1: Google Form exists in gform schema with correct structure.
Check 2: Excel file has correct course statistics.
"""

import argparse
import os
import sys
import json
import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

# Ground truth data from Canvas (read-only)
EXPECTED_COURSES = [
    {
        "Course_Code": "AAA-2014J",
        "Course_Name": "Applied Analytics & Algorithms",
        "Enrollment": 365,
        "Avg_Score": 67.85,
        "Assignment_Count": 6,
        "Quiz_Count": 0,
        "Total_Assessments": 6,
    },
    {
        "Course_Code": "BBB-2014J",
        "Course_Name": "Biochemistry & Bioinformatics",
        "Enrollment": 2292,
        "Avg_Score": 64.31,
        "Assignment_Count": 6,
        "Quiz_Count": 0,
        "Total_Assessments": 6,
    },
    {
        "Course_Code": "CCC-2014J",
        "Course_Name": "Creative Computing & Culture",
        "Enrollment": 2498,
        "Avg_Score": 70.22,
        "Assignment_Count": 10,
        "Quiz_Count": 4,
        "Total_Assessments": 14,
    },
    {
        "Course_Code": "DDD-2014J",
        "Course_Name": "Data-Driven Design",
        "Enrollment": 1803,
        "Avg_Score": 69.99,
        "Assignment_Count": 7,
        "Quiz_Count": 0,
        "Total_Assessments": 7,
    },
    {
        "Course_Code": "EEE-2014J",
        "Course_Name": "Environmental Economics & Ethics",
        "Enrollment": 1188,
        "Avg_Score": 81.27,
        "Assignment_Count": 5,
        "Quiz_Count": 0,
        "Total_Assessments": 5,
    },
    {
        "Course_Code": "FFF-2014J",
        "Course_Name": "Foundations of Finance",
        "Enrollment": 2365,
        "Avg_Score": 76.51,
        "Assignment_Count": 13,
        "Quiz_Count": 7,
        "Total_Assessments": 20,
    },
    {
        "Course_Code": "GGG-2014J",
        "Course_Name": "Global Governance & Geopolitics",
        "Enrollment": 749,
        "Avg_Score": 76.60,
        "Assignment_Count": 10,
        "Quiz_Count": 6,
        "Total_Assessments": 16,
    },
]


def check_google_form():
    """Check that a Fall 2014 feedback form was created with proper questions."""
    print("[eval] Checking Google Form in gform schema...")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    errors = []

    try:
        # Check form exists with matching title
        cur.execute(
            "SELECT id, title FROM gform.forms WHERE LOWER(title) LIKE '%fall 2014%' AND LOWER(title) LIKE '%feedback%'"
        )
        forms = cur.fetchall()
        if not forms:
            errors.append("No form found with title containing 'Fall 2014' and 'Feedback'")
            return False, errors

        form_id = forms[0][0]
        form_title = forms[0][1]
        print(f"  Found form: '{form_title}' (id={form_id})")

        # Check questions
        cur.execute(
            "SELECT id, title, question_type, config FROM gform.questions WHERE form_id = %s ORDER BY position",
            (form_id,),
        )
        questions = cur.fetchall()

        if len(questions) != 5:
            errors.append(f"Expected 5 questions, found {len(questions)}")

        # Count types
        radio_count = sum(1 for q in questions if q[2] == "choiceQuestion")
        text_count = sum(1 for q in questions if q[2] == "textQuestion")

        if radio_count < 2:
            errors.append(f"Expected at least 2 RADIO/choice questions, found {radio_count}")
        if text_count < 2:
            errors.append(f"Expected at least 2 TEXT questions, found {text_count}")

        # Check the course selection question has 7 options
        course_q_found = False
        for q in questions:
            q_title = q[1].lower() if q[1] else ""
            if "course" in q_title and "feedback" in q_title:
                course_q_found = True
                config = q[3] if isinstance(q[3], dict) else json.loads(q[3]) if q[3] else {}
                options = config.get("options", [])
                if len(options) != 7:
                    errors.append(
                        f"Course selection question has {len(options)} options, expected 7"
                    )
                break

        if not course_q_found:
            # Try alternative: look for any choice question with 7 options
            for q in questions:
                if q[2] == "choiceQuestion":
                    config = q[3] if isinstance(q[3], dict) else json.loads(q[3]) if q[3] else {}
                    options = config.get("options", [])
                    if len(options) == 7:
                        course_q_found = True
                        print(f"  Found course selection question by 7 options: '{q[1]}'")
                        break

        if not course_q_found:
            errors.append("No course selection question found with 7 options")

        # Check for satisfaction question: exactly 5 options covering 1..5
        satisfaction_q_found = False
        for q in questions:
            q_title = (q[1] or "").lower()
            config = q[3] if isinstance(q[3], dict) else (json.loads(q[3]) if q[3] else {})
            if q[2] == "choiceQuestion" and ("satisf" in q_title or "rate" in q_title or "score" in q_title or "rating" in q_title or "1-5" in q_title or "1 to 5" in q_title):
                options = config.get("options", [])
                opt_texts = [str(o).strip() for o in options]
                # Require exactly 5 options and all of 1..5 present
                if len(options) == 5:
                    opts_combined = " ".join(opt_texts)
                    has_all_5 = all(str(v) in opts_combined for v in [1, 2, 3, 4, 5])
                    if has_all_5:
                        satisfaction_q_found = True
                        break
        if not satisfaction_q_found:
            errors.append("Satisfaction question must have exactly 5 options (1, 2, 3, 4, 5)")

        # Check for recommendation question: exactly 3 options Yes/No/Maybe
        recommend_q_found = False
        for q in questions:
            q_title = (q[1] or "").lower()
            config = q[3] if isinstance(q[3], dict) else (json.loads(q[3]) if q[3] else {})
            if q[2] == "choiceQuestion" and ("recommend" in q_title):
                options = config.get("options", [])
                opt_texts_lower = " ".join(str(o).lower() for o in options)
                if len(options) == 3 and ("yes" in opt_texts_lower and "no" in opt_texts_lower and "maybe" in opt_texts_lower):
                    recommend_q_found = True
                    break
        if not recommend_q_found:
            errors.append("Recommendation question must have exactly 3 options (Yes, No, Maybe)")

        if errors:
            return False, errors

        print("  Google Form check passed.")
        return True, []

    finally:
        cur.close()
        conn.close()


def load_expected_courses_from_db():
    """Query Canvas DB for Fall 2014 courses with enrollment/avg_score/assignment_count/quiz_count. Fallback if hardcoded EXPECTED_COURSES diverges."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT c.course_code, c.name,
                   (SELECT COUNT(*) FROM canvas.enrollments e WHERE e.course_id = c.id AND e.type = 'StudentEnrollment') as enrollment,
                   (SELECT ROUND(AVG(user_avg)::numeric, 2) FROM (
                       SELECT s.user_id, AVG(s.score) AS user_avg
                       FROM canvas.submissions s
                       JOIN canvas.assignments a ON s.assignment_id = a.id
                       WHERE a.course_id = c.id AND s.score IS NOT NULL
                       GROUP BY s.user_id
                   ) u) as avg_score,
                   (SELECT COUNT(*) FROM canvas.assignments a WHERE a.course_id = c.id) as assignment_count,
                   (SELECT COUNT(*) FROM canvas.quizzes q WHERE q.course_id = c.id) as quiz_count
            FROM canvas.courses c
            WHERE c.course_code LIKE '%%2014J%%'
            ORDER BY c.course_code
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        out = []
        for r in rows:
            code, name, enr, avg, ac, qc = r
            out.append({
                "Course_Code": code,
                "Course_Name": name,
                "Enrollment": int(enr) if enr else 0,
                "Avg_Score": float(avg) if avg else 0.0,
                "Assignment_Count": int(ac) if ac else 0,
                "Quiz_Count": int(qc) if qc else 0,
                "Total_Assessments": (int(ac) if ac else 0) + (int(qc) if qc else 0),
            })
        return out
    except Exception as e:
        print(f"[eval] Warning: could not load courses from DB: {e}")
        return None


def check_excel(agent_workspace):
    """Check the Excel report against expected course data."""
    print("[eval] Checking Excel file...")
    errors = []

    excel_path = os.path.join(agent_workspace, "Fall_2014_Course_Report.xlsx")
    if not os.path.exists(excel_path):
        errors.append(f"Excel file not found: {excel_path}")
        return False, errors

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        errors.append(f"Cannot open Excel file: {e}")
        return False, errors

    # Check sheet name
    if "Course Statistics" not in wb.sheetnames:
        errors.append(
            f"Sheet 'Course Statistics' not found. Available sheets: {wb.sheetnames}"
        )
        return False, errors

    ws = wb["Course Statistics"]

    # Read header row
    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "Course_Code",
        "Course_Name",
        "Enrollment",
        "Avg_Score",
        "Assignment_Count",
        "Quiz_Count",
        "Total_Assessments",
    ]

    # Map header indices
    col_map = {}
    for eh in expected_headers:
        found = False
        for idx, h in enumerate(headers):
            if h and eh.lower().replace("_", "") == str(h).lower().replace("_", "").replace(" ", ""):
                col_map[eh] = idx
                found = True
                break
        if not found:
            # Try partial match
            for idx, h in enumerate(headers):
                if h and eh.lower().replace("_", " ") in str(h).lower().replace("_", " "):
                    col_map[eh] = idx
                    found = True
                    break
        if not found:
            errors.append(f"Column '{eh}' not found in headers: {headers}")

    if errors:
        return False, errors

    # Read data rows
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_map["Course_Code"]] is not None:
            data_rows.append(row)

    if len(data_rows) != 7:
        errors.append(f"Expected 7 data rows, found {len(data_rows)}")
        return False, errors

    # Try DB fallback if available (resilient to data updates)
    db_courses = load_expected_courses_from_db()
    expected_list = EXPECTED_COURSES
    if db_courses and len(db_courses) == 7:
        # Only use DB data if we got all 7 courses
        expected_list = db_courses
        print(f"  Using DB-derived expected data for {len(db_courses)} courses")

    # Check each course
    for expected in expected_list:
        code = expected["Course_Code"]
        found = False
        for row in data_rows:
            row_code = str(row[col_map["Course_Code"]]).strip()
            if row_code == code:
                found = True

                # Check enrollment (tolerance 10)
                enrollment = row[col_map["Enrollment"]]
                if enrollment is not None:
                    if abs(int(enrollment) - expected["Enrollment"]) > 10:
                        errors.append(
                            f"{code}: Enrollment {enrollment} differs from expected {expected['Enrollment']} (tolerance 10)"
                        )

                # Check avg score (tolerance 2.0)
                avg_score = row[col_map["Avg_Score"]]
                if avg_score is not None:
                    if abs(float(avg_score) - expected["Avg_Score"]) > 2.0:
                        errors.append(
                            f"{code}: Avg_Score {avg_score} differs from expected {expected['Avg_Score']} (tolerance 2.0)"
                        )

                # Check assignment count (tolerance 1)
                assign_count = row[col_map["Assignment_Count"]]
                if assign_count is not None:
                    if abs(int(assign_count) - expected["Assignment_Count"]) > 1:
                        errors.append(
                            f"{code}: Assignment_Count {assign_count} differs from expected {expected['Assignment_Count']} (tolerance 1)"
                        )

                # Check quiz count (tolerance 1)
                quiz_count = row[col_map["Quiz_Count"]]
                if quiz_count is not None:
                    if abs(int(quiz_count) - expected["Quiz_Count"]) > 1:
                        errors.append(
                            f"{code}: Quiz_Count {quiz_count} differs from expected {expected['Quiz_Count']} (tolerance 1)"
                        )

                break

        if not found:
            errors.append(f"Course {code} not found in Excel data")

    if errors:
        return False, errors

    print("  Excel check passed.")
    return True, []


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, help="Launch time")
    args = parser.parse_args()

    all_passed = True

    # Check 1: Google Form
    form_pass, form_errors = check_google_form()
    if not form_pass:
        all_passed = False
        print("[FAIL] Google Form check failed:")
        for e in form_errors:
            print(f"  - {e}")
    else:
        print("[PASS] Google Form check passed.")

    # Check 2: Excel file
    excel_pass, excel_errors = check_excel(args.agent_workspace)
    if not excel_pass:
        all_passed = False
        print("[FAIL] Excel check failed:")
        for e in excel_errors:
            print(f"  - {e}")
    else:
        print("[PASS] Excel check passed.")

    if all_passed:
        print("\nAll checks passed!")
        sys.exit(0)
    else:
        print("\nSome checks failed.")
        sys.exit(1)


if __name__ == "__main__":
    main()
