"""Evaluation for canvas-grade-summary."""
import argparse
import os
import sys
import openpyxl

try:
    import psycopg2
    DB = {
        "host": os.environ.get("PGHOST", "localhost"),
        "port": int(os.environ.get("PGPORT", "5432")),
        "dbname": "toolathlon_gym",
        "user": "eigent",
        "password": "camel",
    }
except Exception:
    psycopg2 = None
    DB = None


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_gsheet(errors_list):
    """Verify a Google Sheet titled 'Grade Summary Report' was created with summary data."""
    print("  Checking Google Sheet 'Grade Summary Report'...")
    if psycopg2 is None or DB is None:
        errors_list.append("psycopg2 unavailable; cannot verify Google Sheet")
        return
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        sheets = cur.fetchall()
        target_id = None
        target_name = None
        for sid, title in sheets:
            if title and "grade summary report" in title.strip().lower():
                target_id = sid
                target_name = title
                break
        if target_id is None:
            errors_list.append(
                f"Google Sheet 'Grade Summary Report' not found (got {[n for _,n in sheets]})"
            )
            cur.close(); conn.close()
            return

        # Verify there is sufficient content and required metric labels
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (target_id,))
        cell_count = cur.fetchone()[0]
        if cell_count < 6:  # Total_Courses, Highest_Avg_Course, Overall_Avg_Score plus headers => >= 6
            errors_list.append(
                f"Google Sheet '{target_name}' has only {cell_count} cells; expected >= 6"
            )
        # Pull all cell values (lower-cased) and require Total_Courses, Highest_Avg_Course, Overall_Avg_Score labels
        cur.execute("SELECT value FROM gsheet.cells WHERE spreadsheet_id = %s", (target_id,))
        all_vals = [str(r[0] or "").strip().lower() for r in cur.fetchall()]
        joined = " | ".join(all_vals)
        for required in ["total_courses", "highest_avg_course", "overall_avg_score"]:
            # Allow either snake_case or space-separated label variants
            label_alt = required.replace("_", " ")
            if required not in joined and label_alt not in joined:
                errors_list.append(
                    f"Google Sheet '{target_name}' missing required label: {required}"
                )
        cur.close(); conn.close()
        print(f"    Found GSheet '{target_name}' with {cell_count} cells")
    except Exception as e:
        errors_list.append(f"Google Sheet check raised: {e}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Grade_Summary.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Grade_Summary.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Grade Summary
    print(f"  Checking Grade Summary...")
    a_rows = load_sheet_rows(agent_wb, "Grade Summary")
    g_rows = load_sheet_rows(gt_wb, "Grade Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Grade Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Grade Summary' not found in groundtruth")
    else:
        sheet_name = "Grade Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Students_Submitted: {a_row[1]} vs {g_row[1]} (tol=1)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.5):
                    errors.append(f"{key}.Avg_Score: {a_row[2]} vs {g_row[2]} (tol=0.5)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.5):
                    errors.append(f"{key}.Max_Score: {a_row[3]} vs {g_row[3]} (tol=0.5)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Min_Score: {a_row[4]} vs {g_row[4]} (tol=0.5)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                gt_v = g_row[1]
                if isinstance(gt_v, (int, float)):
                    if not num_close(a_row[1], gt_v, 1.0):
                        errors.append(f"{key}.Value: {a_row[1]} vs {gt_v} (tol=1.0)")
                else:
                    # String value: accept exact match OR (for course names) prefix
                    a_str = str(a_row[1] or "").strip()
                    g_str = str(gt_v or "").strip()
                    if a_str.lower() != g_str.lower():
                        # For course name fields, accept agent's value if it starts with GT prefix
                        # (which would happen if the GT was truncated) or vice versa
                        if not (a_str.lower().startswith(g_str.lower()) or g_str.lower().startswith(a_str.lower())):
                            errors.append(f"{key}.Value: '{a_row[1]}' vs '{gt_v}'")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")



    # Check Google Sheet 'Grade Summary Report'
    check_gsheet(all_errors)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
