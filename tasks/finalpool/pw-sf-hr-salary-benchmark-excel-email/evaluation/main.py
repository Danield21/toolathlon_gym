"""Evaluation script for pw-sf-hr-salary-benchmark-excel-email."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Salary_Benchmark_Report.xlsx")
    check("Salary_Benchmark_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Salary_Benchmark_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        # Helper: validate per-row data against groundtruth
        def _validate_rows_by_dept(sheet_name, gt_check_cols, tol_pct=0.05):
            """For each GT row keyed by Department, verify agent has matching row."""
            if sheet_name not in wb.sheetnames or not gt_wb or sheet_name not in gt_wb.sheetnames:
                return
            ws = wb[sheet_name]
            gt_ws = gt_wb[sheet_name]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
            dept_idx_a = headers.index("department") if "department" in headers else 0
            dept_idx_g = gt_headers.index("department") if "department" in gt_headers else 0
            agent_by = {str(r[dept_idx_a]).strip(): r for r in data_rows if r and r[dept_idx_a]}
            for gt_row in gt_rows:
                if not gt_row or not gt_row[dept_idx_g]:
                    continue
                dept = str(gt_row[dept_idx_g]).strip()
                arow = agent_by.get(dept)
                check(f"{sheet_name} has dept '{dept}'", arow is not None)
                if arow is None:
                    continue
                for col in gt_check_cols:
                    if col not in gt_headers or col not in headers:
                        continue
                    gv = gt_row[gt_headers.index(col)]
                    av = arow[headers.index(col)]
                    gf = safe_float(gv)
                    af = safe_float(av)
                    if gf is not None and af is not None:
                        tol = max(1.0, abs(gf) * tol_pct)
                        check(f"{sheet_name} '{dept}' {col} ~{gf}",
                              abs(gf - af) <= tol, f"got {av}")
                    elif gv is not None and av is not None:
                        check(f"{sheet_name} '{dept}' {col} text",
                              str(gv).strip().lower() == str(av).strip().lower(),
                              f"expected {gv}, got {av}")

        # Compute GT row counts
        def _gt_rows_count(name):
            if gt_wb and name in gt_wb.sheetnames:
                return len(list(gt_wb[name].iter_rows(min_row=2, values_only=True)))
            return 0

        check("Compensation_Comparison sheet exists", "Compensation_Comparison" in wb.sheetnames)
        if "Compensation_Comparison" in wb.sheetnames:
            ws = wb["Compensation_Comparison"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            min_cc = max(7, _gt_rows_count("Compensation_Comparison"))
            check(f"Compensation_Comparison has >= {min_cc} rows", len(data_rows) >= min_cc, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Employee_Count', 'Our_Avg_Salary', 'Industry_Benchmark', 'Difference', 'Status']:
                check(f"Compensation_Comparison has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            # Per-dept value validation (include difference_pct for full coverage)
            _validate_rows_by_dept("Compensation_Comparison",
                                   ['employee_count', 'our_avg_salary',
                                    'industry_benchmark', 'difference',
                                    'difference_pct', 'status'])

        check("Department_Details sheet exists", "Department_Details" in wb.sheetnames)
        if "Department_Details" in wb.sheetnames:
            ws = wb["Department_Details"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            min_dd = max(7, _gt_rows_count("Department_Details"))
            check(f"Department_Details has >= {min_dd} rows", len(data_rows) >= min_dd, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Avg_Experience', 'Avg_Performance']:
                check(f"Department_Details has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            _validate_rows_by_dept("Department_Details",
                                   ['avg_experience', 'avg_performance',
                                    'our_avg_salary', 'salary_per_year_exp'])

        check("Executive_Summary sheet exists", "Executive_Summary" in wb.sheetnames)
        if "Executive_Summary" in wb.sheetnames:
            ws = wb["Executive_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            min_es = max(5, _gt_rows_count("Executive_Summary"))
            check(f"Executive_Summary has >= {min_es} rows", len(data_rows) >= min_es, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Executive_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Validate metric values against GT
            if gt_wb and "Executive_Summary" in gt_wb.sheetnames:
                gt_ws = gt_wb["Executive_Summary"]
                gt_metrics = {str(r[0]).strip().lower(): r[1] for r in gt_ws.iter_rows(min_row=2, values_only=True) if r and r[0]}
                agent_metrics = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for gt_m, gt_v in gt_metrics.items():
                    found = gt_m in agent_metrics
                    check(f"Executive_Summary has '{gt_m}'", found)
                    if found:
                        av = agent_metrics[gt_m]
                        gf = safe_float(gt_v)
                        af = safe_float(av)
                        if gf is not None and af is not None:
                            tol = max(1.0, abs(gf) * 0.05)
                            check(f"Executive_Summary '{gt_m}' ~{gf}",
                                  abs(gf - af) <= tol, f"got {av}")
                        elif gt_v is not None and av is not None:
                            check(f"Executive_Summary '{gt_m}' text",
                                  str(gt_v).strip().lower() == str(av).strip().lower(),
                                  f"expected {gt_v}, got {av}")

        # Check email was sent - require EXACT subject and recipient
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "SELECT subject, to_addr, body_text FROM email.messages "
                "WHERE subject ILIKE %s AND to_addr::text ILIKE %s",
                ('%2026 Salary Benchmark Analysis Complete%', '%hr-director@company.com%'))
            emails = cur.fetchall()
            check("Email with exact subject + hr-director recipient sent",
                  len(emails) >= 1, f"found {len(emails)} matching emails")
            if emails:
                body = (emails[0][2] or "").lower()
                # Body should include key findings: above/below counts, biggest gap, overall status
                check("Email body mentions 'above' or 'below' benchmark",
                      "above" in body or "below" in body, body[:200])
                check("Email body mentions overall status",
                      "competitive" in body or "needs attention" in body or "status" in body,
                      body[:200])
            conn.close()
        except Exception as e:
            check("Email verification", False, str(e))

        # Check terminal artifacts (all three JSONs explicitly named in task.md)
        check("salary_processor.py exists", os.path.exists(os.path.join(agent_workspace, "salary_processor.py")))
        for jname in ("benchmark_raw.json", "internal_salaries.json", "salary_comparison.json"):
            jpath = os.path.join(agent_workspace, jname)
            check(f"{jname} exists", os.path.exists(jpath))
            if os.path.exists(jpath):
                try:
                    with open(jpath, "r") as _f:
                        _d = json.load(_f)
                    check(f"{jname} is valid JSON and non-empty", bool(_d), "empty/null")
                except Exception as _e:
                    check(f"{jname} is valid JSON and non-empty", False, str(_e))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
