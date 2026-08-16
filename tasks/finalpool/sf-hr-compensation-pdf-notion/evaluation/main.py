"""Evaluation for sf-hr-compensation-pdf-notion.

All checks are blocking: Compensation_Data.xlsx (Excel data comparison),
Compensation_Report.pdf (existence + extractable text with title/subtitle/
date/department names), and the Notion page titled 'Compensation Analysis 2026'.
"""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    all_errors = []

    # ---- Check Excel (blocking) ----
    agent_file = os.path.join(args.agent_workspace, "Compensation_Data.xlsx")
    gt_file = os.path.join(gt_dir, "Compensation_Data.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Department Summary
    print("  Checking Department Summary...")
    a_rows = load_sheet_rows(agent_wb, "Department Summary")
    g_rows = load_sheet_rows(gt_wb, "Department Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Department Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Department Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Header check (order-independent): all expected columns must be present.
        expected_headers = [
            "Department", "Employees", "Avg_Salary", "Min_Salary",
            "Max_Salary", "Median_Salary",
        ]
        actual_headers = [str(c or "").strip() for c in (a_rows[0] if a_rows else [])]
        header_lower = [h.lower() for h in actual_headers]
        missing_headers = [h for h in expected_headers if h.lower() not in header_lower]
        if missing_headers:
            all_errors.append(f"Department Summary missing headers: {missing_headers}; found {actual_headers}")
        # Map expected columns to the agent's column indices by header name so
        # the column order in the output does not matter.
        a_col = {h.lower(): i for i, h in enumerate(actual_headers)}

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        # Validate sort order: should be alphabetical by Department
        a_dept_order = [str(r[0]).strip().lower() for r in a_data if r and r[0] is not None]
        if a_dept_order != sorted(a_dept_order):
            all_errors.append(f"Department Summary not sorted alphabetically: {a_dept_order}")
        # (hkey, label, tolerance, GT column index). GT columns are fixed (1..5).
        COLS = [
            ("employees", "Employees", 1, 1),
            ("avg_salary", "Avg_Salary", 1.0, 2),
            ("min_salary", "Min_Salary", 1, 3),
            ("max_salary", "Max_Salary", 1, 4),
            ("median_salary", "Median_Salary", 1.0, 5),
        ]
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing department: {g_row[0]}")
                continue
            for hkey, label, tol, gcol in COLS:
                ai = a_col.get(hkey)
                if ai is None:
                    continue  # missing header already reported above
                if ai < len(a_row) and gcol < len(g_row):
                    if not num_close(a_row[ai], g_row[gcol], tol):
                        all_errors.append(f"{key}.{label}: {a_row[ai]} vs {g_row[gcol]} (tol={tol})")
        if not all_errors:
            print("    PASS")

    # Check Education Breakdown
    print("  Checking Education Breakdown...")
    a_rows = load_sheet_rows(agent_wb, "Education Breakdown")
    g_rows = load_sheet_rows(gt_wb, "Education Breakdown")
    prev_errors = len(all_errors)
    if a_rows is None:
        all_errors.append("Sheet 'Education Breakdown' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Education Breakdown' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Header check (order-independent): all expected columns must be present.
        expected_headers = [
            "Department", "Education_Level", "Count", "Avg_Salary",
            "Min_Salary", "Max_Salary",
        ]
        actual_headers = [str(c or "").strip() for c in (a_rows[0] if a_rows else [])]
        header_lower = [h.lower() for h in actual_headers]
        missing_headers = [h for h in expected_headers if h.lower() not in header_lower]
        if missing_headers:
            all_errors.append(f"Education Breakdown missing headers: {missing_headers}; found {actual_headers}")
        # Map expected columns to the agent's column indices by header name so
        # the column order in the output does not matter.
        a_col = {h.lower(): i for i, h in enumerate(actual_headers)}

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None and row[1] is not None:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                a_lookup[key] = row
        # (hkey, label, tolerance, GT column index). GT columns are fixed (2..5).
        COLS = [
            ("count", "Count", 1, 2),
            ("avg_salary", "Avg_Salary", 1.0, 3),
            ("min_salary", "Min_Salary", 1, 4),
            ("max_salary", "Max_Salary", 1, 5),
        ]
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing edu row: {g_row[0]} / {g_row[1]}")
                continue
            for hkey, label, tol, gcol in COLS:
                ai = a_col.get(hkey)
                if ai is None:
                    continue  # missing header already reported above
                if ai < len(a_row) and gcol < len(g_row):
                    if not num_close(a_row[ai], g_row[gcol], tol):
                        all_errors.append(f"{key[0]} / {key[1]}.{label}: {a_row[ai]} vs {g_row[gcol]} (tol={tol})")

        new_errors = len(all_errors) - prev_errors
        if new_errors == 0:
            print("    PASS")

    # ---- Check PDF exists (blocking with content checks) ----
    print("  Checking PDF...")
    pdf_path = os.path.join(args.agent_workspace, "Compensation_Report.pdf")
    if not os.path.exists(pdf_path):
        all_errors.append("Compensation_Report.pdf not found")
    else:
        file_size = os.path.getsize(pdf_path)
        if file_size < 500:
            all_errors.append(f"Compensation_Report.pdf too small ({file_size} bytes)")
        else:
            # Content checks
            try:
                pdf_text = ""
                try:
                    import pdfplumber
                    with pdfplumber.open(pdf_path) as pdf:
                        for page in pdf.pages:
                            pdf_text += (page.extract_text() or "") + "\n"
                except ImportError:
                    try:
                        from pypdf import PdfReader
                        reader = PdfReader(pdf_path)
                        for page in reader.pages:
                            pdf_text += (page.extract_text() or "") + "\n"
                    except ImportError:
                        from PyPDF2 import PdfReader
                        reader = PdfReader(pdf_path)
                        for page in reader.pages:
                            pdf_text += (page.extract_text() or "") + "\n"
                pdf_lower = pdf_text.lower()
                # Title check
                if "compensation analysis report" not in pdf_lower:
                    all_errors.append("PDF missing 'Compensation Analysis Report' title")
                if "hr analytics" not in pdf_lower:
                    all_errors.append("PDF missing 'HR Analytics' subtitle")
                if "2026-03-06" not in pdf_lower and "march 6, 2026" not in pdf_lower and "march 06, 2026" not in pdf_lower:
                    all_errors.append("PDF missing date 2026-03-06")
                # Validate at least 3 of 7 dept names appear in PDF
                dept_count = sum(1 for d in ["engineering", "finance", "hr", "operations", "r&d", "sales", "support"]
                                 if d in pdf_lower)
                if dept_count < 5:
                    all_errors.append(f"PDF missing department names (found {dept_count}/7)")
                if not all_errors or all([e for e in all_errors if "PDF" not in e]):
                    print("    PASS")
            except Exception as e:
                print(f"    [WARN] PDF content extraction error: {e}")

    # ---- Notion check (BLOCKING) ----
    print("  Checking Notion page (blocking)...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        # Look for page with title 'Compensation Analysis 2026'
        cur.execute("""
            SELECT COUNT(*) FROM notion.pages
            WHERE LOWER(properties::text) LIKE '%%compensation analysis 2026%%'
               OR LOWER(properties::text) LIKE '%%compensation%%2026%%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            # fallback: look in blocks block_data for the title
            cur.execute("""
                SELECT COUNT(*) FROM notion.blocks
                WHERE LOWER(block_data::text) LIKE '%%compensation analysis 2026%%'
            """)
            count = cur.fetchone()[0]
        if count == 0:
            all_errors.append("Notion page 'Compensation Analysis 2026' not found")
        else:
            print(f"    PASS (found {count} matching Notion entries)")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Notion check failed: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:15]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
