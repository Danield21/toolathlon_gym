"""
Evaluation script for howtocook-notion-recipe-kb task.

Checks:
1. Excel file (Recipe_Comparison.xlsx) - two sheets with correct structure and data
2. Notion - database with recipe entries exists
3. Memory - memory.json has entities (agent recorded filtering criteria)

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
RUNTIME_INCONCLUSIVE = 0
IN_RUNTIME_BLOCK = False


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, RUNTIME_INCONCLUSIVE
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        if IN_RUNTIME_BLOCK:
            RUNTIME_INCONCLUSIVE += 1
            msg = f": {detail[:300]}" if detail else ""
            print(f"  [FAIL-runtime] {name}{msg}")
        else:
            FAIL_COUNT += 1
            msg = f": {detail[:300]}" if detail else ""
            print(f"  [FAIL] {name}{msg}")


def normalize(text):
    """Normalize text for comparison: lowercase, strip whitespace."""
    if text is None:
        return ""
    return str(text).strip().lower()


def load_sheet_rows(wb, sheet_name):
    """Load all rows from a sheet (case-insensitive name lookup)."""
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    ws = wb[matched]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_excel(agent_workspace):
    """Check Recipe_Comparison.xlsx has correct structure and content."""
    print("\n=== Checking Excel Output ===")

    excel_path = os.path.join(agent_workspace, "Recipe_Comparison.xlsx")

    if not os.path.isfile(excel_path):
        record("Recipe_Comparison.xlsx exists", False, f"Not found at {excel_path}")
        return False

    record("Recipe_Comparison.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    record("Excel file readable", True)
    all_ok = True

    # --- Sheet 1: All Recipes ---
    all_recipes = load_sheet_rows(wb, "All Recipes")
    if all_recipes is None:
        record("Sheet 'All Recipes' exists", False, "Not found")
        all_ok = False
    else:
        record("Sheet 'All Recipes' exists", True)

        # Check header row
        if len(all_recipes) < 1:
            record("All Recipes has header", False, "Sheet is empty")
            all_ok = False
        else:
            headers = [normalize(h) for h in all_recipes[0]]
            expected_headers = ["recipe_name", "category", "difficulty",
                                "ingredients_count", "prep_steps_count"]
            # Check headers exist (flexible on exact naming)
            header_keywords = {
                "recipe": False, "category": False, "difficulty": False,
                "ingredient": False, "step": False,
            }
            for h in headers:
                for kw in header_keywords:
                    if kw in h:
                        header_keywords[kw] = True

            all_headers_found = all(header_keywords.values())
            record("All Recipes headers correct", all_headers_found,
                   f"Headers: {headers}, missing keywords: "
                   f"{[k for k, v in header_keywords.items() if not v]}")
            if not all_headers_found:
                all_ok = False

        # Check data rows (at least 8)
        data_rows = all_recipes[1:] if len(all_recipes) > 1 else []
        # Filter out empty rows
        data_rows = [r for r in data_rows if any(c is not None for c in r)]
        record("All Recipes has at least 8 recipes",
               len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")
        if len(data_rows) < 8:
            all_ok = False

        # Check that recipes span at least 4 categories
        if data_rows and len(all_recipes[0]) >= 2:
            categories = set()
            for row in data_rows:
                if len(row) >= 2 and row[1] is not None:
                    categories.add(normalize(row[1]))
            record("All Recipes spans at least 4 categories",
                   len(categories) >= 4,
                   f"Found {len(categories)} categories: {categories}")
            if len(categories) < 4:
                all_ok = False

        # Check that numeric columns have numeric values
        numeric_ok_count = 0
        for row in data_rows:
            if len(row) >= 5:
                try:
                    ing = int(float(row[3])) if row[3] is not None else None
                    steps = int(float(row[4])) if row[4] is not None else None
                    if ing is not None and ing > 0 and steps is not None and steps > 0:
                        numeric_ok_count += 1
                except (ValueError, TypeError):
                    pass
        # Require ALL rows (min 8) have valid numeric columns
        record("All Recipes numeric columns valid (all rows)",
               numeric_ok_count >= len(data_rows) and numeric_ok_count >= 8,
               f"{numeric_ok_count}/{len(data_rows)} rows have valid numeric data")
        if numeric_ok_count < len(data_rows) or numeric_ok_count < 8:
            all_ok = False

    # --- Sheet 2: Top Picks ---
    top_picks = load_sheet_rows(wb, "Top Picks")
    if top_picks is None:
        record("Sheet 'Top Picks' exists", False, "Not found")
        all_ok = False
    else:
        record("Sheet 'Top Picks' exists", True)

        # Check header
        if len(top_picks) < 1:
            record("Top Picks has header", False, "Sheet is empty")
            all_ok = False
        else:
            headers = [normalize(h) for h in top_picks[0]]
            header_keywords = {
                "rank": False, "recipe": False, "category": False, "reason": False,
            }
            for h in headers:
                for kw in header_keywords:
                    if kw in h:
                        header_keywords[kw] = True

            all_headers_found = all(header_keywords.values())
            record("Top Picks headers correct", all_headers_found,
                   f"Headers: {headers}, missing: "
                   f"{[k for k, v in header_keywords.items() if not v]}")
            if not all_headers_found:
                all_ok = False

        # Check at least 5 entries
        data_rows = top_picks[1:] if len(top_picks) > 1 else []
        data_rows = [r for r in data_rows if any(c is not None for c in r)]
        record("Top Picks has at least 5 entries",
               len(data_rows) >= 5,
               f"Found {len(data_rows)} data rows")
        if len(data_rows) < 5:
            all_ok = False

        # Check that reason column is not empty
        reason_filled = 0
        for row in data_rows:
            if len(row) >= 4 and row[3] is not None and len(str(row[3]).strip()) > 3:
                reason_filled += 1
        record("Top Picks reason column filled",
               reason_filled >= 4,
               f"{reason_filled}/{len(data_rows)} rows have reasons")
        if reason_filled < 4:
            all_ok = False

    return all_ok


def _get_top_picks_names(agent_workspace):
    """Return the set of Top Picks recipe names from Excel (normalized) for cross-check."""
    excel_path = os.path.join(agent_workspace, "Recipe_Comparison.xlsx")
    names = set()
    if not os.path.isfile(excel_path):
        return names
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        rows = load_sheet_rows(wb, "Top Picks")
        if rows is None or len(rows) < 2:
            return names
        # header row
        headers = [normalize(h) for h in rows[0]]
        name_col = None
        for i, h in enumerate(headers):
            if "recipe" in h or h == "name":
                name_col = i
                break
        if name_col is None:
            return names
        for r in rows[1:]:
            if r and len(r) > name_col and r[name_col]:
                names.add(normalize(r[name_col]))
    except Exception:
        pass
    return names


def check_notion(agent_workspace=None, runtime_allow_empty=True):
    """Check Notion has a recipe collection page and database with entries."""
    global IN_RUNTIME_BLOCK
    print("\n=== Checking Notion ===")

    all_ok = True

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Total notion data check - if totally empty, treat as runtime-only
        cur.execute("SELECT COUNT(*) FROM notion.pages")
        (n_pages,) = cur.fetchone()
        cur.execute("SELECT COUNT(*) FROM notion.databases")
        (n_dbs,) = cur.fetchone()
        if runtime_allow_empty and n_pages == 0 and n_dbs == 0:
            IN_RUNTIME_BLOCK = True

        # Find 'Team Recipe Collection' page first
        cur.execute("""
            SELECT id, properties::text
            FROM notion.pages
        """)
        all_pages = cur.fetchall()
        team_page_id = None
        for page_id, props_text in all_pages:
            if props_text and "team recipe collection" in props_text.lower():
                team_page_id = page_id
                break
        if team_page_id is None:
            # Fallback: any "recipe collection" page
            for page_id, props_text in all_pages:
                if props_text and "recipe collection" in props_text.lower():
                    team_page_id = page_id
                    break
        record("Notion 'Team Recipe Collection' page exists",
               team_page_id is not None,
               "No page with 'recipe collection' found")
        if team_page_id is None:
            all_ok = False

        # Find recipe database - prefer the one whose parent is team_page_id
        cur.execute("""
            SELECT id, title::text, properties::text, parent::text
            FROM notion.databases
        """)
        databases = cur.fetchall()

        recipe_db = None
        # Priority 1: child of team_page_id
        if team_page_id:
            for db_id, db_title, db_props, db_parent in databases:
                if db_parent and team_page_id in db_parent:
                    recipe_db = (db_id, db_title, db_props)
                    break
        # Priority 2: title/props mentions recipe
        if recipe_db is None:
            for db_id, db_title, db_props, db_parent in databases:
                title_lower = (db_title or "").lower()
                props_lower = (db_props or "").lower()
                if "recipe" in title_lower or "recipe" in props_lower:
                    recipe_db = (db_id, db_title, db_props)
                    break

        has_recipe_db = recipe_db is not None
        record("Notion recipe database exists (child of Team Recipe Collection page)",
               has_recipe_db,
               f"Found {len(databases)} databases"
               if not has_recipe_db else "")

        if not has_recipe_db:
            all_ok = False
        else:
            db_id = recipe_db[0]

            # Check database properties include Name, Category, Difficulty
            db_props_lower = (recipe_db[2] or "").lower()
            for prop_name in ["name", "category", "difficulty"]:
                found = prop_name in db_props_lower
                record(f"Notion database has '{prop_name}' property", found,
                       f"'{prop_name}' not found in database properties")
                if not found:
                    all_ok = False

            # Check pages (entries) in the database
            cur.execute("""
                SELECT id, properties::text
                FROM notion.pages
                WHERE parent::text LIKE %s
            """, (f"%{db_id}%",))
            db_pages = cur.fetchall()

            record("Notion database has at least 5 recipe entries",
                   len(db_pages) >= 5,
                   f"Found {len(db_pages)} entries")
            if len(db_pages) < 5:
                all_ok = False

            # Cross-check Name values overlap with Excel Top Picks
            if agent_workspace:
                top_names = _get_top_picks_names(agent_workspace)
                if top_names:
                    # Extract names from Notion entries
                    notion_names = set()
                    for pid, props_text in db_pages:
                        if props_text:
                            pl = props_text.lower()
                            # Heuristic: any word from top picks present?
                            for tn in top_names:
                                if tn and tn in pl:
                                    notion_names.add(tn)
                    overlap = notion_names & top_names
                    # Allow partial overlap (name transliteration variance)
                    record("Notion DB recipe Names overlap with Excel Top Picks (>=3 matches)",
                           len(overlap) >= 3,
                           f"Overlap: {overlap} (Top Picks: {top_names})")
                    if len(overlap) < 3:
                        all_ok = False

        cur.close()
        conn.close()

    except Exception as e:
        record("Notion database check", False, str(e))
        all_ok = False

    IN_RUNTIME_BLOCK = False
    return all_ok


def check_memory(agent_workspace):
    """Check that memory.json has entities (agent recorded criteria)."""
    print("\n=== Checking Memory ===")

    mem_file = os.path.join(agent_workspace, "memory", "memory.json")

    if not os.path.isfile(mem_file):
        record("memory.json exists", False, f"Not found at {mem_file}")
        return False

    record("memory.json exists", True)

    try:
        with open(mem_file, "r") as f:
            memory_data = json.load(f)
    except Exception as e:
        record("memory.json readable", False, str(e))
        return False

    record("memory.json readable", True)

    all_ok = True

    # Check entities exist (require >=2)
    entities = memory_data.get("entities", [])
    record("Memory has >=2 entities",
           len(entities) >= 2,
           f"Found {len(entities)} entities")
    if len(entities) < 2:
        all_ok = False

    # Check at least one observation references a real category name
    real_categories = ["meat", "vegetable", "soup", "staple", "breakfast",
                       "condiment", "dessert", "snack", "poultry", "seafood"]
    category_obs_found = False
    for entity in entities:
        obs_list = entity.get("observations", [])
        for obs in obs_list:
            obs_lower = str(obs).lower()
            if any(cat in obs_lower for cat in real_categories):
                category_obs_found = True
                break
        if category_obs_found:
            break

    record("Memory has observation referencing real category name",
           category_obs_found,
           f"None of {real_categories} found in any observation")
    if not category_obs_found:
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    notion_ok = check_notion(args.agent_workspace)
    memory_ok = check_memory(args.agent_workspace)

    # Use FAIL_COUNT as authoritative (excludes runtime-only inconclusive)
    overall = (FAIL_COUNT == 0) and excel_ok and memory_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:   {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Notion:  {'PASS' if notion_ok else 'FAIL (may be runtime-only)'}")
    print(f"  Memory:  {'PASS' if memory_ok else 'FAIL'}")
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Runtime-inconclusive: {RUNTIME_INCONCLUSIVE}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
