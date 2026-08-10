"""Evaluation for terminal-wc-sf-notion-excel-email.
Checks:
1. Support_Quality_Audit.xlsx with 4 sheets and correct data
2. Notion database with critical product entries
3. Two emails sent to correct recipients
4. Python scripts exist

Expected values are recomputed from the read-only DB at evaluation time, so the
ground truth never goes stale relative to the immutable db seed.
"""
import argparse
import json
import os
import re
import sys
from collections import defaultdict

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _to_float(val):
    """Robust numeric parser.

    Handles str/int/float/None. Strips thousand separators, currency symbols,
    percent signs and surrounding whitespace. "90%" parses as 90.0. Returns
    None when the value cannot be interpreted as a number (including Excel
    formula strings that carry no cached result).
    """
    if val is None:
        return None
    if isinstance(val, bool):
        return float(val)
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.startswith("="):
        return None
    for ch in (",", "$", "¥", "€", "%"):
        s = s.replace(ch, "")
    s = s.strip()
    try:
        return float(s)
    except ValueError:
        return None


def safe_float(val, default=None):
    f = _to_float(val)
    return f if f is not None else default


def _number_strings(f):
    """Several textual representations of a float, for formula embedding checks."""
    out = []
    if f == int(f):
        out.append(str(int(f)))
    out.append(f"{f:.2f}")
    out.append(repr(f))
    out.append(f"{f:g}")
    return list(dict.fromkeys(out))


def _formula_embeds_number(formula, num):
    """Return True if a formula string (with no cached result) embeds the
    expected number as a standalone token, e.g. '=31588', '=3.26',
    '=SUM(6466,9348,15774)'. A range reference like '=SUM(A1:A360)' does NOT
    count as embedding 36. This only fires for formula cells that carry no
    cached value; it never lets a wrong number pass (the token must match).
    """
    if not (isinstance(formula, str) and formula.startswith("=")):
        return False
    f = _to_float(num)
    if f is None:
        return False
    for rep in _number_strings(f):
        pattern = r"(?<![A-Za-z0-9_.])" + re.escape(rep) + r"(?![A-Za-z0-9_.])"
        if re.search(pattern, formula):
            return True
    return False


def num_close(a, b, tol=2.0):
    """Numeric closeness.

    R3 semantics: if both sides parse as numbers, compare numerically with
    tolerance. If exactly one side fails to parse because it is an Excel
    formula with no cached result, accept only when that formula embeds the
    expected number as a standalone literal token. If exactly one side fails
    to parse and the other did parse (empty cell / garbage text), fall back to
    a case-insensitive string comparison (which fails). If neither side parses,
    fail (do not silently pass).
    """
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if fa is None and fb is None:
        return False
    # One side is a non-numeric formula without a cached result: only accept
    # if the formula literally embeds the other side's number.
    if fb is not None and _formula_embeds_number(a, fb):
        return True
    if fa is not None and _formula_embeds_number(b, fa):
        return True
    return str(a).strip().lower() == str(b).strip().lower()


def get_groundtruth_from_db():
    """Compute expected values from read-only DB data."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Problem products
    cur.execute("""
        SELECT (unnest_item->>'product_id')::int as pid, COUNT(DISTINCT o.id)
        FROM wc.orders o, jsonb_array_elements(o.line_items) as unnest_item
        WHERE o.status IN ('refunded','failed')
        GROUP BY pid
    """)
    refund_products = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("""
        SELECT product_id, COUNT(*)
        FROM wc.product_reviews WHERE rating <= 2
        GROUP BY product_id
    """)
    low_review_products = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT id, name, categories FROM wc.products")
    products = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    all_pids = set(refund_products.keys()) | set(low_review_products.keys())
    problem_list = []
    for pid in all_pids:
        rc = refund_products.get(pid, 0)
        lrc = low_review_products.get(pid, 0)
        severity = rc * 30 + lrc * 40
        name = products.get(pid, ("Unknown", []))[0][:60]
        cats = products.get(pid, ("", []))[1]
        cat = cats[0]['name'] if cats else 'Unknown'
        problem_list.append((pid, name, cat, rc, lrc, severity))
    problem_list.sort(key=lambda x: (-x[5], x[0]))

    severities = sorted([p[5] for p in problem_list])
    p80_idx = int(len(severities) * 0.8)
    p80_val = severities[p80_idx] if p80_idx < len(severities) else severities[-1]
    critical = [p for p in problem_list if p[5] > p80_val]

    # Priority data
    cur.execute("""
        SELECT "PRIORITY", COUNT(*),
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY" ORDER BY "PRIORITY"
    """)
    priority_data = cur.fetchall()

    # Issue type data
    cur.execute("""
        SELECT "ISSUE_TYPE", COUNT(*),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE" ORDER BY COUNT(*) DESC
    """)
    issue_data = cur.fetchall()

    total_tickets = sum(r[1] for r in priority_data)
    cur.execute("""
        SELECT ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
    """)
    overall_sat = float(cur.fetchone()[0])

    cur.close()
    conn.close()

    return {
        "problem_list": problem_list,
        "critical": critical,
        "priority_data": priority_data,
        "issue_data": issue_data,
        "total_tickets": total_tickets,
        "overall_sat": overall_sat,
        "p80_val": p80_val,
    }


def _load_workbook(path):
    """Load a workbook twice.

    Returns (formulas_wb, values_wb). formulas_wb is loaded with data_only=False
    (default) so structure/sheet names are always visible; values_wb is loaded
    with data_only=True so cached results of formula cells are available.
    values_wb may be None if the values pass fails.
    """
    formulas_wb = openpyxl.load_workbook(path)
    try:
        values_wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        values_wb = None
    return formulas_wb, values_wb


def _cell_value(cell, values_ws):
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        if values_ws is not None:
            cached = values_ws.cell(row=cell.row, column=cell.column).value
            if cached is not None:
                return cached
    return v


def _read_rows(ws, values_ws=None, min_row=1):
    rows = []
    for row in ws.iter_rows(min_row=min_row):
        rows.append([_cell_value(c, values_ws) for c in row])
    return rows


def _dedup_rows(rows, key_idx=0):
    """Content-based dedup on the leading key column (case-insensitive)."""
    seen = set()
    out = []
    for r in rows:
        if r[key_idx] is None:
            continue
        key = str(r[key_idx]).strip().lower()
        if key and key not in seen:
            seen.add(key)
            out.append(r)
    return out


# Header hints used to (a) auto-detect an optional header row and (b) locate a
# sheet by content when its name does not match the expected keywords.
HEADER_HINTS = {
    "pp": ["product", "severity", "refund", "category", "low", "review", "id", "name", "score"],
    "sp": ["priority", "response", "ticket", "satisfaction"],
    "it": ["issue", "type", "ticket", "count", "satisfaction"],
    "es": ["metric", "value", "summary", "key", "result"],
}


def _norm(s):
    """Normalize a string for header matching: lowercase, spaces/dashes -> _."""
    s = str(s).strip().lower().replace(" ", "_").replace("-", "_")
    return re.sub(r"[^a-z0-9_]", "", s)


def _cell_matches_hint(cell, hint):
    """True when a cell looks like a header column matching `hint`, using
    token-aware matching so 'Avg_Response_Hours' matches 'response' and
    'Total_Tickets' matches 'ticket'."""
    s = _norm(cell)
    if not s:
        return False
    if s == hint:
        return True
    tokens = s.split("_")
    return any(t == hint or t.startswith(hint) or hint.startswith(t) for t in tokens)


def _header_score(row, hints):
    """Count how many cells in row look like a header column name."""
    matches = 0
    for c in row:
        if c is None:
            continue
        if any(_cell_matches_hint(c, h) for h in hints):
            matches += 1
    return matches


def _row_is_header(row, hints):
    """True when the row looks like a header (>=2 cells match known columns)."""
    return _header_score(row, hints) >= 2


def _filter_blank_rows(rows):
    """Drop rows that are entirely empty (trailing/format-inflated blank rows)."""
    return [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]


def _read_data_rows(ws, values_ws, hints):
    """Read a sheet's data rows, tolerating an optional header row.

    Skips leading blank rows; if the first non-blank row is a header it is
    skipped; all-blank rows (e.g. trailing rows inflated by max_row) are
    dropped. Returns only non-blank data rows.
    """
    rows = _read_rows(ws, values_ws, min_row=1)
    start = 0
    while start < len(rows):
        if any(c is not None and str(c).strip() != "" for c in rows[start]):
            break
        start += 1
    if start < len(rows) and _row_is_header(rows[start], hints):
        start += 1
    return _filter_blank_rows(rows[start:])


def _locate_sheet(wb, values_wb, canonical_name, keywords, hints, fallback_idx, used):
    """Locate a sheet by (1) exact canonical name, (2) keyword substring,
    (3) content header match, (4) positional fallback. `used` holds already
    claimed indices so no sheet is double-claimed by two checks."""
    sheets = wb.sheetnames
    # 1. exact canonical name (case-insensitive)
    target = _norm(canonical_name)
    for i, s in enumerate(sheets):
        if i in used:
            continue
        if _norm(s) == target:
            return i
    # 2. keyword substring
    for i, s in enumerate(sheets):
        if i in used:
            continue
        sl = s.lower()
        if any(k in sl for k in keywords):
            return i
    # 3. content header match
    best, best_score = None, 0
    for i, s in enumerate(sheets):
        if i in used:
            continue
        ws = wb[s]
        vs = values_wb[s] if values_wb is not None else None
        first = next((r for r in _read_rows(ws, vs, min_row=1)
                      if any(c is not None and str(c).strip() != "" for c in r)), None)
        if first is None:
            continue
        score = _header_score(first, hints)
        if score >= 2 and score > best_score:
            best, best_score = i, score
    if best is not None:
        return best
    # 4. positional fallback: never reuse an already-claimed index
    if fallback_idx not in used and fallback_idx < len(sheets):
        return fallback_idx
    for i in range(len(sheets)):
        if i not in used:
            return i
    return 0


def check_excel(workspace, gt):
    print("\n=== Check 1: Support_Quality_Audit.xlsx ===")
    path = os.path.join(workspace, "Support_Quality_Audit.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb, wb_values = _load_workbook(path)
    sheets = wb.sheetnames

    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")
    if len(sheets) == 0:
        return

    def values_sheet(name):
        if wb_values is None:
            return None
        try:
            return wb_values[name]
        except Exception:
            return None

    # Each sheet is located by exact name, keyword substring, content header
    # match, then positional fallback; indices already claimed by an earlier
    # check are never re-used.
    used = set()

    # Problem_Products sheet
    pp_idx = _locate_sheet(wb, wb_values, "Problem_Products", ["problem", "product"],
                           HEADER_HINTS["pp"], 0, used)
    used.add(pp_idx)
    ws_pp = wb[sheets[pp_idx]]
    rows_pp = _read_data_rows(ws_pp, values_sheet(sheets[pp_idx]), HEADER_HINTS["pp"])
    expected_count = len(gt["problem_list"])
    check(f"Problem_Products has ~{expected_count} rows",
          abs(len(rows_pp) - expected_count) <= 2,
          f"Found {len(rows_pp)} data rows, expected {expected_count}")

    # Check top product by severity
    if rows_pp:
        top_row = rows_pp[0]
        top_pid = safe_float(top_row[0])
        expected_top = gt["problem_list"][0]
        check("Top product ID correct",
              top_pid is not None and int(top_pid) == expected_top[0],
              f"Got pid={top_pid}, expected {expected_top[0]}")
        top_severity = top_row[5] if len(top_row) > 5 else top_row[-1]
        check("Top product severity correct",
              num_close(top_severity, expected_top[5], tol=5),
              f"Got {top_severity}, expected {expected_top[5]}")

    # Check a mid-range product exists
    if len(gt["problem_list"]) > 5:
        mid_product = gt["problem_list"][3]
        all_text = " ".join(str(c) for r in rows_pp for c in r if c).lower()
        check("Contains expected mid-range product",
              str(mid_product[0]) in all_text or mid_product[1][:15].lower() in all_text,
              f"Looking for pid={mid_product[0]} or name={mid_product[1][:15]}")

    # Support_By_Priority sheet
    sp_idx = _locate_sheet(wb, wb_values, "Support_By_Priority", ["priority", "support"],
                           HEADER_HINTS["sp"], 1, used)
    used.add(sp_idx)
    if sp_idx < len(sheets):
        ws_sp = wb[sheets[sp_idx]]
        rows_sp = _dedup_rows(_read_data_rows(ws_sp, values_sheet(sheets[sp_idx]), HEADER_HINTS["sp"]))
        check("Support_By_Priority has at least 3 rows", len(rows_sp) >= 3,
              f"Found {len(rows_sp)} rows")

        if rows_sp:
            all_text_sp = " ".join(str(c) for r in rows_sp for c in r if c).lower()
            check("Has High priority", "high" in all_text_sp)
            check("Has Medium priority", "medium" in all_text_sp)
            check("Has Low priority", "low" in all_text_sp)

            # Check ticket count + avg satisfaction per priority (content-matched)
            for expected in gt["priority_data"]:
                key = str(expected[0]).lower()
                row = next((r for r in rows_sp if r[0] and key in str(r[0]).lower()), None)
                if row is None:
                    continue
                count = row[1] if len(row) > 1 else None
                check(f"{expected[0]} priority count ~{expected[1]}",
                      num_close(count, expected[1], tol=50),
                      f"Got {count}, expected {expected[1]}")
                sat = row[3] if len(row) > 3 else row[-1]
                check(f"{expected[0]} priority avg satisfaction ~{expected[3]}",
                      num_close(sat, float(expected[3]), tol=0.1),
                      f"Got {sat}, expected {expected[3]}")

    # Issue_Type_Breakdown sheet
    it_idx = _locate_sheet(wb, wb_values, "Issue_Type_Breakdown", ["issue", "type"],
                           HEADER_HINTS["it"], 2, used)
    used.add(it_idx)
    if it_idx < len(sheets):
        ws_it = wb[sheets[it_idx]]
        rows_it = _dedup_rows(_read_data_rows(ws_it, values_sheet(sheets[it_idx]), HEADER_HINTS["it"]))
        expected_it_rows = len(gt["issue_data"])
        check(f"Issue_Type_Breakdown has at least {expected_it_rows} rows",
              len(rows_it) >= expected_it_rows,
              f"Found {len(rows_it)} rows, expected {expected_it_rows}")

        if rows_it:
            all_text_it = " ".join(str(c) for r in rows_it for c in r if c).lower()
            check("Has Bug issue type", "bug" in all_text_it)
            check("Has Performance Issue type", "performance" in all_text_it)

            # Content-matched per-issue-type ticket counts
            for expected in gt["issue_data"]:
                key = str(expected[0]).lower()
                row = next((r for r in rows_it if r[0] and key in str(r[0]).lower()), None)
                if row is None:
                    continue
                count = row[1] if len(row) > 1 else None
                check(f"{expected[0]} ticket count ~{expected[1]}",
                      num_close(count, expected[1], tol=50),
                      f"Got {count}, expected {expected[1]}")

    # Executive_Summary sheet
    es_idx = _locate_sheet(wb, wb_values, "Executive_Summary", ["executive", "summary"],
                           HEADER_HINTS["es"], 3, used)
    used.add(es_idx)
    if es_idx < len(sheets):
        ws_es = wb[sheets[es_idx]]
        rows_es = _read_data_rows(ws_es, values_sheet(sheets[es_idx]), HEADER_HINTS["es"])
        check("Executive_Summary has at least 5 rows", len(rows_es) >= 5,
              f"Found {len(rows_es)} rows")

        if rows_es:
            summary_dict = {}
            for r in rows_es:
                if r[0]:
                    summary_dict[str(r[0]).lower()] = r[1]

            # Total Problem Products
            tp_key = next((k for k in summary_dict if "total" in k and "problem" in k), None)
            if tp_key:
                check("Total Problem Products correct",
                      num_close(summary_dict[tp_key], expected_count, tol=2),
                      f"Got {summary_dict[tp_key]}, expected {expected_count}")

            # Critical Products
            cp_key = next((k for k in summary_dict if "critical" in k), None)
            if cp_key:
                check("Critical Products count correct",
                      num_close(summary_dict[cp_key], len(gt["critical"]), tol=1),
                      f"Got {summary_dict[cp_key]}, expected {len(gt['critical'])}")

            # Total Support Tickets
            tt_key = next((k for k in summary_dict if "ticket" in k), None)
            if tt_key:
                check("Total Support Tickets correct",
                      num_close(summary_dict[tt_key], gt["total_tickets"], tol=100),
                      f"Got {summary_dict[tt_key]}, expected {gt['total_tickets']}")

            # Overall Avg Satisfaction
            sat_key = next((k for k in summary_dict if "satisfaction" in k), None)
            if sat_key:
                check("Overall Avg Satisfaction correct",
                      num_close(summary_dict[sat_key], gt["overall_sat"], tol=0.1),
                      f"Got {summary_dict[sat_key]}, expected {gt['overall_sat']}")

            # Highest Risk Category - dynamically derive from problem_list
            cat_key = next((k for k in summary_dict if "category" in k or "risk" in k), None)
            if cat_key:
                cat_severity = defaultdict(int)
                for (pid, name, cat, rc, lrc, sev) in gt["problem_list"]:
                    cat_severity[cat] += sev
                expected_top_cat = max(cat_severity, key=cat_severity.get) if cat_severity else "Electronics"
                cell_text = str(summary_dict[cat_key]).strip().lower()
                exp_lower = expected_top_cat.lower()
                check(f"Highest Risk Category is {expected_top_cat}",
                      exp_lower in cell_text or cell_text in exp_lower,
                      f"Got {summary_dict[cat_key]}, expected {expected_top_cat}")


def _unique_page_props(pages):
    """Content-dedup pages by their full properties payload (case-insensitive)."""
    unique = set()
    for p in pages:
        props = p[1] if p[1] else {}
        unique.add(json.dumps(props, sort_keys=True).lower())
    return unique


def check_notion(gt):
    print("\n=== Check 2: Notion Database ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title, properties FROM notion.databases
        WHERE title::text ILIKE '%Support Quality%' OR title::text ILIKE '%support_quality%'
    """)
    dbs = cur.fetchall()
    check("Notion database 'Support Quality Tracker' exists", len(dbs) >= 1,
          f"Found {len(dbs)} matching databases")

    if dbs:
        db_id = str(dbs[0][0])
        props = dbs[0][2] if dbs[0][2] else {}
        if not isinstance(props, dict):
            props = {}

        # Check properties exist
        prop_names_lower = {k.lower(): k for k in props.keys()}
        check("Has Severity property", any("severity" in k for k in prop_names_lower))
        check("Has Status property", any("status" in k for k in prop_names_lower))
        check("Has Product property", any("product" in k for k in prop_names_lower))

        # Check pages
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE parent::text LIKE %s AND NOT archived
        """, (f'%{db_id}%',))
        pages = cur.fetchall()
        unique_props = _unique_page_props(pages)
        check(f"Has at least 6 distinct entries for critical products",
              len(unique_props) >= 6,
              f"Found {len(pages)} pages ({len(unique_props)} distinct)")

        if pages:
            # Check at least one page has Critical severity
            critical_found = False
            for page in pages:
                page_props = page[1] if page[1] else {}
                props_text = json.dumps(page_props).lower()
                if "critical" in props_text:
                    critical_found = True
                    break
            check("At least one entry has Critical severity", critical_found)

            # Check pages reference expected products
            all_page_text = " ".join(json.dumps(p[1]).lower() for p in pages if p[1])
            check("Pages reference AGARO or tripod product",
                  "agaro" in all_page_text or "tripod" in all_page_text,
                  f"Text snippet: {all_page_text[:200]}")

    cur.close()
    conn.close()


def check_emails(gt):
    print("\n=== Check 3: Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Check support team email
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%support quality%' AND subject ILIKE '%priority%'
        AND to_addr::text ILIKE '%support_team%'
    """)
    support_emails = cur.fetchall()
    check("Email to support_team sent", len(support_emails) >= 1,
          f"Found {len(support_emails)} matching emails")

    if support_emails:
        body = (support_emails[0][2] or "").lower()
        sat_val = gt["overall_sat"]
        sat_str_1 = f"{sat_val:.1f}"
        sat_str_2 = f"{sat_val:.2f}"
        has_sat_num = sat_str_1 in body or sat_str_2 in body
        check("Support email mentions satisfaction",
              "satisfaction" in body or has_sat_num,
              f"Body snippet: {body[:150]} (expected ~{sat_val})")

    # Check product team email
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%support quality%' AND subject ILIKE '%product%'
        AND to_addr::text ILIKE '%product_team%'
    """)
    product_emails = cur.fetchall()
    check("Email to product_team sent", len(product_emails) >= 1,
          f"Found {len(product_emails)} matching emails")

    if product_emails:
        body = (product_emails[0][2] or "").lower()
        check("Product email mentions critical products",
              "critical" in body or "severity" in body or "agaro" in body,
              f"Body snippet: {body[:150]}")
        check("Product email mentions total problem products",
              "36" in body or str(len(gt["problem_list"])) in body or "problem" in body,
              f"Body snippet: {body[:150]}")

    cur.close()
    conn.close()


def check_reverse_validation(gt):
    print("\n=== Reverse Validation ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # Check Notion does not include non-critical (below 80th percentile) products
        cur.execute("""
            SELECT id, title FROM notion.databases
            WHERE title::text ILIKE '%%Support Quality%%' OR title::text ILIKE '%%support_quality%%'
        """)
        dbs = cur.fetchall()
        if dbs:
            db_id = str(dbs[0][0])
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE parent::text LIKE %s AND NOT archived
            """, (f'%{db_id}%',))
            pages = cur.fetchall()
            unique_props = _unique_page_props(pages)
            # Should have no more than ~top 20% products; certainly not all problem products
            max_expected = len(gt["critical"]) + 2  # small tolerance
            check("Notion does not include non-critical products",
                  len(unique_props) <= max_expected,
                  f"Found {len(unique_props)} distinct pages, expected at most {max_expected} (critical={len(gt['critical'])})")

        # Check no emails sent to wrong recipients
        noise_recipients = [
            "all-staff@company.com",
            "hr@company.com",
            "newsletter@company.com",
            "finance@company.com",
        ]
        for addr in noise_recipients:
            cur.execute(
                "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s",
                (f"%{addr}%",),
            )
            cnt = cur.fetchone()[0]
            check(f"No email sent to noise recipient {addr}", cnt == 0,
                  f"Found {cnt} emails to {addr}")
    except Exception as e:
        check("Reverse validation", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python Scripts ===")
    for script in ["correlate_issues.py", "support_metrics.py"]:
        path = os.path.join(workspace, script)
        check(f"{script} exists", os.path.exists(path), f"Not found at {path}")

    # Check output JSON files
    for jf in ["problem_products.json", "support_analysis.json"]:
        path = os.path.join(workspace, jf)
        check(f"{jf} exists", os.path.exists(path), f"Not found at {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    gt = get_groundtruth_from_db()

    check_excel(args.agent_workspace, gt)
    check_notion(gt)
    check_emails(gt)
    check_scripts(args.agent_workspace)
    check_reverse_validation(gt)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Tightened from 70% to 85% to prevent bad-case bypass (which passed at 78%) while allowing runtime-only service checks to fail in GT test
    sys.exit(0 if accuracy >= 85 else 1)


if __name__ == "__main__":
    main()
