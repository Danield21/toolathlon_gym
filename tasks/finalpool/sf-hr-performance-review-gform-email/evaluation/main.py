"""
Evaluation for sf-hr-performance-review-gform-email task.

Checks:
1. Performance_Review_Setup.xlsx exists with "Performance Analysis" sheet
2. Sheet has 7 department rows
3. Engineering row has Employee_Count=7096, Avg_Performance~3.21
4. GForm "Annual Performance Review Form" exists with at least 4 questions
5. GForm has department and rating questions
6. Email sent to hr_director@company.com
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

EXPECTED_DEPARTMENTS = {"Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"}


def _fetch_dept_expected():
    """Dynamically query DB for per-department stats. Fallback to hardcoded on failure."""
    try:
        conn = psycopg2.connect(**DB_CONFIG, connect_timeout=5)
        cur = conn.cursor()
        cur.execute("""
            SELECT "DEPARTMENT", COUNT(*),
                   ROUND(AVG("PERFORMANCE_RATING")::numeric, 2),
                   SUM(CASE WHEN "PERFORMANCE_RATING" < 3 THEN 1 ELSE 0 END),
                   SUM(CASE WHEN "PERFORMANCE_RATING" >= 4 THEN 1 ELSE 0 END)
            FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
            GROUP BY "DEPARTMENT"
        """)
        result = {}
        for r in cur.fetchall():
            result[r[0]] = {
                "count": int(r[1]),
                "avg": float(r[2]),
                "low": int(r[3]),
                "high": int(r[4]),
            }
        cur.close(); conn.close()
        return result
    except Exception as e:
        print(f"[fallback] DB unreachable: {e}")
        return None


DEPT_EXPECTED = _fetch_dept_expected() or {
    "Engineering": {"count": 7096, "avg": 3.21, "low": 1381, "high": 2535},
    "Finance":     {"count": 7148, "avg": 3.21, "low": 1387, "high": 2510},
    "HR":          {"count": 7077, "avg": 3.20, "low": 1425, "high": 2485},
    "Operations":  {"count": 7120, "avg": 3.18, "low": 1431, "high": 2426},
    "R&D":         {"count": 7083, "avg": 3.20, "low": 1415, "high": 2478},
    "Sales":       {"count": 7232, "avg": 3.19, "low": 1494, "high": 2523},
    "Support":     {"count": 7244, "avg": 3.20, "low": 1479, "high": 2515},
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Excel Performance_Review_Setup.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Performance_Review_Setup.xlsx")
    if not os.path.exists(xlsx_path):
        record("Performance_Review_Setup.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Performance_Review_Setup.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    # Find Performance Analysis sheet
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    perf_idx = next((i for i, s in enumerate(sheet_names_lower) if "performance" in s or "analysis" in s), None)
    if perf_idx is None:
        perf_idx = 0  # use first sheet

    ws = wb[wb.sheetnames[perf_idx]]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        record("Sheet has data", False, "Sheet is empty")
        return

    data_rows = [r for r in rows[1:] if any(c for c in r)]
    record("Sheet has 7 department rows", len(data_rows) >= 7,
           f"Found {len(data_rows)} data rows")

    # Check headers
    headers = [str(c).lower().strip() if c else "" for c in rows[0]]
    has_dept = any("dept" in h or "department" in h for h in headers)
    has_count = any("count" in h or "employee" in h for h in headers)
    has_avg = any("avg" in h or "average" in h or "performance" in h for h in headers)

    record("Has Department column", has_dept, f"Headers: {rows[0]}")
    record("Has Employee_Count column", has_count, f"Headers: {rows[0]}")
    record("Has Avg_Performance column", has_avg, f"Headers: {rows[0]}")

    # Check for Engineering row with correct data
    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    has_engineering = "engineering" in all_text
    record("Contains Engineering department data", has_engineering, "No Engineering row found")

    # Check all expected departments present
    found_depts = set()
    for row in data_rows:
        for cell in row:
            if cell and str(cell).strip() in EXPECTED_DEPARTMENTS:
                found_depts.add(str(cell).strip())
    record("All 7 departments present", len(found_depts) == 7,
           f"Found departments: {found_depts}")

    # Verify Engineering count (7096) and avg (~3.21)
    engineering_row = None
    dept_col = next((i for i, h in enumerate(headers) if "dept" in h or "department" in h), 0)
    for row in data_rows:
        if row[dept_col] and str(row[dept_col]).strip() == "Engineering":
            engineering_row = row
            break

    # Identify column indices using normalized header matching
    dept_col  = next((i for i, h in enumerate(headers) if "dept" in h or "department" in h), 0)
    count_col = next((i for i, h in enumerate(headers) if "count" in h and "low" not in h and "high" not in h), 1)
    avg_col   = next((i for i, h in enumerate(headers) if "avg" in h or "average" in h), 2)
    low_col   = next((i for i, h in enumerate(headers) if "low" in h and "count" in h), None)
    high_col  = next((i for i, h in enumerate(headers) if "high" in h and "count" in h), None)

    record("Has Low_Performers_Count column", low_col is not None, f"Headers: {rows[0]}")
    record("Has High_Performers_Count column", high_col is not None, f"Headers: {rows[0]}")

    # Build dept -> row dict
    dept_to_row = {}
    for row in data_rows:
        name = str(row[dept_col]).strip() if dept_col < len(row) and row[dept_col] else ""
        if name:
            dept_to_row[name] = row

    # Validate every expected department
    for dept, exp in DEPT_EXPECTED.items():
        row = dept_to_row.get(dept)
        if row is None:
            record(f"{dept} row present", False, f"Not found; got {list(dept_to_row.keys())}")
            continue
        record(f"{dept} row present", True)
        try:
            cnt = int(row[count_col]) if count_col < len(row) and row[count_col] is not None else None
            record(f"{dept} Employee_Count={exp['count']}", cnt == exp["count"], f"Got {cnt}")
        except (TypeError, ValueError):
            record(f"{dept} Employee_Count={exp['count']}", False, f"Parse failed: {row[count_col]}")
        try:
            avg = float(row[avg_col]) if avg_col < len(row) and row[avg_col] is not None else None
            record(f"{dept} Avg_Performance={exp['avg']}", avg is not None and abs(avg - exp["avg"]) < 0.015, f"Got {avg}")
        except (TypeError, ValueError):
            record(f"{dept} Avg_Performance={exp['avg']}", False, f"Parse failed: {row[avg_col]}")
        if low_col is not None:
            try:
                lv = int(row[low_col]) if low_col < len(row) and row[low_col] is not None else None
                record(f"{dept} Low_Performers_Count={exp['low']}", lv == exp["low"], f"Got {lv}")
            except (TypeError, ValueError):
                record(f"{dept} Low_Performers_Count={exp['low']}", False, f"Parse failed: {row[low_col]}")
        if high_col is not None:
            try:
                hv = int(row[high_col]) if high_col < len(row) and row[high_col] is not None else None
                record(f"{dept} High_Performers_Count={exp['high']}", hv == exp["high"], f"Got {hv}")
            except (TypeError, ValueError):
                record(f"{dept} High_Performers_Count={exp['high']}", False, f"Parse failed: {row[high_col]}")


def check_gform():
    print("\n=== Check 2: GForm Annual Performance Review Form ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    # Exact (case-insensitive) match: "Annual Performance Review Form"
    review_form = None
    for form_id, title in forms:
        tlow = (title or "").strip().lower()
        if tlow == "annual performance review form":
            review_form = (form_id, title)
            break
    # Fallback to looser match only if exact not found
    if review_form is None:
        for form_id, title in forms:
            tlow = (title or "").lower()
            if "annual performance review" in tlow and "form" in tlow:
                review_form = (form_id, title)
                break

    record("Annual Performance Review Form exists (exact title)", review_form is not None,
           f"Forms found: {[f[1] for f in forms]}")

    if review_form:
        form_id, title = review_form
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        record("Form has exactly 6 questions", q_count == 6,
               f"Found {q_count} questions")

        # Check for department and rating questions
        cur.execute("SELECT title, question_type, required, config FROM gform.questions WHERE form_id = %s ORDER BY position", (form_id,))
        questions = cur.fetchall()
        q_titles_lower = [q[0].lower() for q in questions]

        has_emp_id_q = any("employee" in t and "id" in t for t in q_titles_lower)
        has_dept_q = any("department" in t for t in q_titles_lower)
        has_rating_q = any("rating" in t for t in q_titles_lower) or any("performance" in t and "rating" in t for t in q_titles_lower)
        has_achievement_q = any("achievement" in t for t in q_titles_lower)
        has_improvement_q = any("improvement" in t for t in q_titles_lower)
        has_goals_q = any("goal" in t for t in q_titles_lower)

        record("Form has Employee ID question", has_emp_id_q,
               f"Questions: {[q[0] for q in questions]}")
        record("Form has Department question", has_dept_q,
               f"Questions: {[q[0] for q in questions]}")
        record("Form has Performance Rating question", has_rating_q,
               f"Questions: {[q[0] for q in questions]}")
        record("Form has Key Achievements question", has_achievement_q,
               f"Questions: {[q[0] for q in questions]}")
        record("Form has Areas for Improvement question", has_improvement_q,
               f"Questions: {[q[0] for q in questions]}")
        record("Form has Goals for next year question", has_goals_q,
               f"Questions: {[q[0] for q in questions]}")

        # Question type / options checks
        dept_q = next((q for q, t in zip(questions, q_titles_lower) if "department" in t), None)
        rating_q = next((q for q, t in zip(questions, q_titles_lower) if "rating" in t), None)

        if dept_q is not None:
            record("Department question is RADIO",
                   (dept_q[1] or "").upper() == "RADIO",
                   f"Got type {dept_q[1]}")
            try:
                cfg = dept_q[3] if isinstance(dept_q[3], dict) else json.loads(dept_q[3]) if dept_q[3] else {}
                opts = set(str(o).strip() for o in (cfg.get("options") or []))
                expected_opts = {"Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"}
                record("Department question has all 7 dept options",
                       expected_opts.issubset(opts),
                       f"Got options: {opts}")
            except Exception as e:
                record("Department question has all 7 dept options", False, str(e))
        if rating_q is not None:
            record("Performance Rating question is RADIO",
                   (rating_q[1] or "").upper() == "RADIO",
                   f"Got type {rating_q[1]}")
            try:
                cfg = rating_q[3] if isinstance(rating_q[3], dict) else json.loads(rating_q[3]) if rating_q[3] else {}
                opts = cfg.get("options") or []
                record("Performance Rating has 5 rating options",
                       len(opts) == 5,
                       f"Got {len(opts)} options: {opts}")
            except Exception as e:
                record("Performance Rating has 5 rating options", False, str(e))

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email to hr_director@company.com ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "hr_director@company.com" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    record("Email sent to hr_director@company.com", matching is not None,
           f"Messages found: {len(messages)}")

    if matching:
        subject, _, _, body_text = matching
        subj_lower = (subject or "").lower()
        body_lower = (body_text or "").lower()
        record("Email subject is 'Annual Performance Review Process Setup Complete'",
               "annual performance review process setup complete" in subj_lower,
               f"Subject: {subject}")
        # Body should summarize department distribution
        dept_mentions = sum(1 for d in EXPECTED_DEPARTMENTS if d.lower() in body_lower)
        record("Email body summarizes department distribution (>=4 dept names)",
               dept_mentions >= 4,
               f"Mentioned {dept_mentions} depts in body")
        record("Email body mentions performance distribution",
               ("performance" in body_lower and "distribution" in body_lower) or
               ("rating" in body_lower),
               f"Body sample: {body_lower[:200]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL ({FAIL_COUNT} checks failed)")
        sys.exit(1)


if __name__ == "__main__":
    main()
