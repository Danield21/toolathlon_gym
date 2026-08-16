"""Evaluation for yf-sector-rotation-excel-notion-email.

Checks:
1. Excel file (Sector_Rotation.xlsx) with 3 sheets, correct structure and values
2. Notion database "Sector Research" with 5 stock entries
3. Two emails sent to investment_team@firm.com and trading_desk@firm.com
"""
import argparse
import os
import sys
import psycopg2
import openpyxl

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym", user="eigent", password="camel")

SYMBOLS = ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]

# Tolerances (tightened)
PRICE_TOL = 0.05  # 2-decimal price
RETURN_TOL = 0.2  # 2-decimal percentage rounding
RS_TOL = 0.01     # 3-decimal RS


def load_sheet_rows(wb, sheet_name):
    """Load rows from a sheet, case-insensitive match."""
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol):
    """Check if two numbers are close within tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison."""
    try:
        return str(a).strip().lower() == str(b).strip().lower()
    except (TypeError, AttributeError):
        return False


def check_excel(agent_ws, gt_ws):
    errors = []
    agent_path = os.path.join(agent_ws, "Sector_Rotation.xlsx")
    gt_path = os.path.join(gt_ws, "Sector_Rotation.xlsx")

    if not os.path.exists(agent_path):
        return ["Sector_Rotation.xlsx not found in agent workspace"]
    if not os.path.exists(gt_path):
        return ["Sector_Rotation.xlsx not found in groundtruth workspace"]

    try:
        wb_agent = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        return [f"Error loading Excel files: {e}"]

    # --- Sheet 1: Momentum Analysis ---
    agent_rows = load_sheet_rows(wb_agent, "Momentum Analysis")
    gt_rows = load_sheet_rows(wb_gt, "Momentum Analysis")

    if agent_rows is None:
        errors.append("Sheet 'Momentum Analysis' not found")
    elif gt_rows is None:
        errors.append("Groundtruth 'Momentum Analysis' sheet missing")
    else:
        agent_data = [r for r in agent_rows[1:] if r and r[0] is not None]
        gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]

        if len(agent_data) < 5:
            errors.append(f"Momentum Analysis has {len(agent_data)} rows, expected 5")
        else:
            agent_by_sym = {str(r[0]).strip().upper(): r for r in agent_data}
            gt_by_sym = {str(r[0]).strip().upper(): r for r in gt_data}

            for sym in SYMBOLS:
                if sym not in agent_by_sym:
                    errors.append(f"{sym} missing from Momentum Analysis")
                    continue
                a = agent_by_sym[sym]
                g = gt_by_sym.get(sym)
                if g is None:
                    continue

                # Company_Name (col 1)
                if not str_match(a[1], g[1]):
                    errors.append(f"{sym}: Company_Name '{a[1]}' != '{g[1]}'")
                # Sector (col 2)
                if not str_match(a[2], g[2]):
                    errors.append(f"{sym}: Sector '{a[2]}' != '{g[2]}'")
                # Latest_Price (col 3)
                if not num_close(a[3], g[3], PRICE_TOL):
                    errors.append(f"{sym}: Latest_Price {a[3]} != {g[3]} (tol {PRICE_TOL})")

                # returns (cols 4,5,6)
                for idx, name in [(4, "Return_1M_Pct"), (5, "Return_3M_Pct"), (6, "Return_6M_Pct")]:
                    if not num_close(a[idx], g[idx], RETURN_TOL):
                        errors.append(f"{sym}: {name} {a[idx]} != {g[idx]} (tol {RETURN_TOL})")

                # Composite_Momentum (col 7)
                if not num_close(a[7], g[7], RETURN_TOL):
                    errors.append(f"{sym}: Composite_Momentum {a[7]} != {g[7]}")

                # Benchmark_Momentum (col 8)
                if not num_close(a[8], g[8], RETURN_TOL):
                    errors.append(f"{sym}: Benchmark_Momentum {a[8]} != {g[8]}")

                # Signal (col 9)
                if not str_match(a[9], g[9]):
                    errors.append(f"{sym}: Signal '{a[9]}' != '{g[9]}'")

        # Sort: alphabetical
        if 'agent_data' in dir() and agent_data:
            symbols_in_order = [str(r[0]).strip().upper() for r in agent_data]
            if symbols_in_order != sorted(SYMBOLS):
                errors.append(f"Momentum Analysis not alphabetically sorted: {symbols_in_order}")

    # --- Sheet 2: Relative Strength ---
    agent_rs = load_sheet_rows(wb_agent, "Relative Strength")
    gt_rs = load_sheet_rows(wb_gt, "Relative Strength")

    if agent_rs is None:
        errors.append("Sheet 'Relative Strength' not found")
    elif gt_rs is None:
        errors.append("Groundtruth 'Relative Strength' sheet missing")
    else:
        agent_data = [r for r in agent_rs[1:] if r and r[0] is not None]
        gt_data = [r for r in gt_rs[1:] if r and r[0] is not None]

        if len(agent_data) < 5:
            errors.append(f"Relative Strength has {len(agent_data)} rows, expected 5")
        else:
            agent_by_sym = {str(r[0]).strip().upper(): r for r in agent_data}
            gt_by_sym = {str(r[0]).strip().upper(): r for r in gt_data}

            for sym in SYMBOLS:
                if sym not in agent_by_sym:
                    errors.append(f"{sym} missing from Relative Strength")
                    continue
                a = agent_by_sym[sym]
                g = gt_by_sym.get(sym)
                if g is None:
                    continue

                for idx, name in [(1, "RS_1M"), (2, "RS_3M"), (3, "RS_6M"), (4, "Avg_RS")]:
                    if not num_close(a[idx], g[idx], RS_TOL):
                        errors.append(f"{sym}: {name} {a[idx]} != {g[idx]} (tol {RS_TOL})")

                # RS_Rank
                if a[5] is not None and g[5] is not None:
                    if int(a[5]) != int(g[5]):
                        errors.append(f"{sym}: RS_Rank {a[5]} != {g[5]}")

    # --- Sheet 3: Strategy Summary ---
    agent_sum = load_sheet_rows(wb_agent, "Strategy Summary")
    gt_sum = load_sheet_rows(wb_gt, "Strategy Summary")

    if agent_sum is None:
        errors.append("Sheet 'Strategy Summary' not found")
    elif gt_sum is None:
        errors.append("Groundtruth 'Strategy Summary' sheet missing")
    else:
        def build_summary_dict(rows):
            d = {}
            for r in rows:
                if r and r[0] is not None:
                    d[str(r[0]).strip()] = r[1]
            return d

        a_sum = build_summary_dict(agent_sum[1:])
        g_sum = build_summary_dict(gt_sum[1:])

        for label in ["Overweight_Count", "Neutral_Count", "Underweight_Count"]:
            if label in g_sum:
                a_val = a_sum.get(label)
                g_val = g_sum[label]
                if a_val is None:
                    errors.append(f"Summary missing: {label}")
                elif int(a_val) != int(g_val):
                    errors.append(f"Summary {label}: {a_val} != {g_val}")

        for label in ["Top_Momentum_Stock", "Bottom_Momentum_Stock", "Portfolio_Signal"]:
            if label in g_sum:
                if not str_match(a_sum.get(label, ""), g_sum[label]):
                    errors.append(f"Summary {label}: '{a_sum.get(label)}' != '{g_sum[label]}'")

        if "Avg_Composite_Momentum" in g_sum:
            if not num_close(a_sum.get("Avg_Composite_Momentum", 0), g_sum["Avg_Composite_Momentum"], RETURN_TOL):
                errors.append(f"Summary Avg_Composite_Momentum: {a_sum.get('Avg_Composite_Momentum')} != {g_sum['Avg_Composite_Momentum']}")

        # Benchmark_Return_6M (numeric)
        if "Benchmark_Return_6M" in g_sum:
            if not num_close(a_sum.get("Benchmark_Return_6M", 0), g_sum["Benchmark_Return_6M"], RETURN_TOL):
                errors.append(f"Summary Benchmark_Return_6M: {a_sum.get('Benchmark_Return_6M')} != {g_sum['Benchmark_Return_6M']}")
        # Analysis_Date (string match)
        if "Analysis_Date" in g_sum:
            if not str_match(a_sum.get("Analysis_Date", ""), g_sum["Analysis_Date"]):
                errors.append(f"Summary Analysis_Date: '{a_sum.get('Analysis_Date')}' != '{g_sum['Analysis_Date']}'")

    return errors


def check_notion():
    """Check for Sector Research database with 5 stock entries."""
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Find database titled exactly 'Sector Research'
        cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
        all_dbs = cur.fetchall()
        target_db = None
        for did, title in all_dbs:
            title_text = ""
            if isinstance(title, list):
                for t in title:
                    if isinstance(t, dict):
                        title_text += t.get("plain_text", "") or t.get("text", {}).get("content", "")
            else:
                title_text = str(title or "")
            if title_text.strip().lower() == "sector research":
                target_db = did
                break
        if target_db is None:
            errors.append(f"No Notion database titled 'Sector Research'. Found: {[t for _, t in all_dbs]}")
            cur.close(); conn.close()
            return errors

        # 5 pages parented to target db
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND parent->>'database_id' = %s
        """, (target_db,))
        pages = cur.fetchall()
        if len(pages) != 5:
            errors.append(f"Sector Research database has {len(pages)} pages, expected 5")

        # Each page: validate Stock title, Sector select, Signal select, Momentum number, Last_Updated date
        valid_signals = {"overweight", "neutral", "underweight"}
        signals_seen = set()
        symbols_seen = set()
        for pid, props in pages:
            if not isinstance(props, dict):
                continue
            page_signal_present = False
            page_signal_valid = False
            for k, v in props.items():
                if not isinstance(v, dict):
                    continue
                t = v.get("type")
                kn = (k or "").lower()
                if t == "title" and ("stock" in kn or kn == "name"):
                    arr = v.get("title", [])
                    name = "".join(x.get("plain_text", "") for x in arr).strip().upper()
                    if name in [s.upper() for s in SYMBOLS]:
                        symbols_seen.add(name)
                if t == "select" and "signal" in kn:
                    page_signal_present = True
                    sel = v.get("select")
                    if isinstance(sel, dict):
                        sname = (sel.get("name") or "").strip().lower()
                        if sname in valid_signals:
                            page_signal_valid = True
                            signals_seen.add(sname)
                        else:
                            errors.append(f"Page {pid}: invalid signal '{sname}' (must be one of {sorted(valid_signals)})")
                if t == "number" and "momentum" in kn:
                    n = v.get("number")
                    if not isinstance(n, (int, float)):
                        errors.append(f"Page {pid}: Momentum is not a number")
                if t == "date" and ("last_updated" in kn.replace(" ", "_") or "updated" in kn):
                    d = v.get("date")
                    if not isinstance(d, dict) or not d.get("start"):
                        errors.append(f"Page {pid}: Last_Updated has no date")
            if not page_signal_present:
                errors.append(f"Page {pid}: missing Signal property")
            elif not page_signal_valid:
                # already added invalid signal error above; this catches None select
                pass

        missing_syms = set(SYMBOLS) - symbols_seen
        if missing_syms:
            errors.append(f"Notion DB missing symbols: {sorted(missing_syms)}")

        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def check_emails():
    """Check for 2 emails: investment_team@firm.com (Sector Rotation Analysis Update) and trading_desk@firm.com (Actionable Sector Signals)."""
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        cur.execute("""
            SELECT subject, to_addr, COALESCE(body_text, body_html, '')
            FROM email.messages
        """)
        all_emails = cur.fetchall()
        cur.close()
        conn.close()

        def find_email(target_to, target_subj):
            import json as _json
            for subj, to_addr, body in all_emails:
                tos = []
                if isinstance(to_addr, list):
                    tos = [str(t).lower() for t in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        p = _json.loads(to_addr)
                        tos = [str(t).lower() for t in p] if isinstance(p, list) else [to_addr.lower()]
                    except Exception:
                        tos = [to_addr.lower()]
                if (target_subj in (subj or "").strip().lower() and
                        any(target_to in t for t in tos)):
                    return subj, to_addr, body
            return None

        e1 = find_email("investment_team@firm.com", "sector rotation analysis update")
        if not e1:
            errors.append("No email subject 'Sector Rotation Analysis Update' to investment_team@firm.com")
        else:
            body_lower = (e1[2] or "").lower()
            if not any(k in body_lower for k in ["overweight", "underweight", "bullish", "bearish", "mixed"]):
                errors.append("investment_team email body lacks portfolio signal/strength keywords")

        e2 = find_email("trading_desk@firm.com", "actionable sector signals")
        if not e2:
            errors.append("No email subject 'Actionable Sector Signals' to trading_desk@firm.com")
        else:
            body_lower = (e2[2] or "").lower()
            symbols_in_body = sum(1 for s in SYMBOLS if s.lower() in body_lower)
            if symbols_in_body < 5:
                errors.append(f"trading_desk email body mentions only {symbols_in_body}/5 stock symbols")
    except Exception as e:
        errors.append(f"Error checking emails: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion database...")
    errs = check_notion()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking emails...")
    errs = check_emails()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
