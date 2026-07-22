"""Evaluation for sf-sales-order-status."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _db_order_status():
    """Compute expected order-status aggregates from live DB. Returns dict, or None if DB unavailable."""
    try:
        import psycopg2
        conn = psycopg2.connect(host='localhost', port=5432,
                                dbname='toolathlon_gym', user='eigent', password='camel')
        cur = conn.cursor()
        cur.execute('''
            SELECT "STATUS",
                   COUNT(*) AS orders,
                   SUM("TOTAL_AMOUNT") AS revenue,
                   AVG("QUANTITY")::double precision AS avg_qty,
                   AVG("DISCOUNT")::double precision AS avg_disc
            FROM sf_data."SALES_DW__PUBLIC__ORDERS"
            GROUP BY "STATUS"
        ''')
        rows = cur.fetchall()
        conn.close()
        by_status = {}
        total_rev = 0.0
        total_orders = 0
        for r in rows:
            status, orders, revenue, avg_qty, avg_disc = r
            orders = int(orders)
            revenue = float(revenue) if revenue is not None else 0.0
            avg_qty = float(avg_qty) if avg_qty is not None else 0.0
            avg_disc = float(avg_disc) if avg_disc is not None else 0.0
            by_status[status] = {
                "orders": orders,
                "revenue": revenue,
                "avg_qty": avg_qty,
                "avg_disc": avg_disc,
            }
            total_rev += revenue
            total_orders += orders
        # fill revenue_pct
        for k, v in by_status.items():
            v["revenue_pct"] = (v["revenue"] / total_rev * 100) if total_rev else 0.0
        summary = {
            "total_orders": total_orders,
            "total_revenue": round(total_rev, 2),
            "delivered_revenue": round(by_status.get("Delivered", {}).get("revenue", 0.0), 2),
            "cancel_rate_pct": round(by_status.get("Cancelled", {}).get("orders", 0) / total_orders * 100, 2) if total_orders else 0.0,
        }
        return {"status": by_status, "summary": summary}
    except Exception as e:
        print(f"  (DB query skipped: {e})")
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Order_Status.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Order_Status.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    db_data = _db_order_status()

    # Check sheet: Order Status
    print(f"  Checking Order Status...")
    a_rows = load_sheet_rows(agent_wb, "Order Status")
    g_rows = load_sheet_rows(gt_wb, "Order Status")
    if a_rows is None:
        all_errors.append("Sheet 'Order Status' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Order Status' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Sort check: Revenue descending (column index 2)
        a_revenues = []
        for r in a_data:
            if r and r[0] is not None and len(r) > 2 and r[2] is not None:
                try:
                    a_revenues.append(float(r[2]))
                except (TypeError, ValueError):
                    pass
        if a_revenues and a_revenues != sorted(a_revenues, reverse=True):
            errors.append(f"Order Status not sorted by Revenue descending; got {a_revenues}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        # Primary source: DB if available, else GT
        primary = []
        if db_data:
            for status, vals in db_data["status"].items():
                primary.append([status, vals["orders"], round(vals["revenue"], 2),
                                round(vals["avg_qty"], 1), round(vals["avg_disc"], 2),
                                round(vals["revenue_pct"], 1)])
        else:
            primary = g_data

        for g_row in primary:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{key}.Orders: {a_row[1]} vs {g_row[1]} (tol=5)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 100.0):
                    errors.append(f"{key}.Revenue: {a_row[2]} vs {g_row[2]} (tol=100.0)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.5):
                    errors.append(f"{key}.Avg_Quantity: {a_row[3]} vs {g_row[3]} (tol=0.5)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Avg_Discount: {a_row[4]} vs {g_row[4]} (tol=0.5)")

            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.5):
                    errors.append(f"{key}.Revenue_Pct: {a_row[5]} vs {g_row[5]} (tol=0.5)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        # Graded tolerances by metric (money wide, counts tight, rates tight)
        SUM_TOLS = {
            "total_orders": 1,
            "total_revenue": 100.0,
            "delivered_revenue": 100.0,
            "cancel_rate_pct": 0.5,
        }

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            expected = g_row[1] if len(g_row) > 1 else None
            if db_data and key in db_data["summary"]:
                expected = db_data["summary"][key]
            tol = SUM_TOLS.get(key, 100.0)

            if len(a_row) > 1:
                if not num_close(a_row[1], expected, tol):
                    errors.append(f"{key}.Value: {a_row[1]} vs {expected} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
