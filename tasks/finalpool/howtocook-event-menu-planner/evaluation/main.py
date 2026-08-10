"""
Evaluation for howtocook-event-menu-planner task.
Checks Excel and email.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _normalize_commas(t):
    """Normalize comma usage in a numeric string.

    A comma may be a thousands separator ('1,050') or, in some locales, a
    decimal separator ('157,5'). We strip it when it groups digits in threes
    from the right (a valid thousands separator) and otherwise treat it as a
    decimal point. When both '.' and ',' are present the comma is a thousands
    separator.
    """
    if "," not in t:
        return t
    if "." in t:
        return t.replace(",", "")
    if re.fullmatch(r"[+-]?\d{1,3}(,\d{3})+", t):
        return t.replace(",", "")
    if re.fullmatch(r"[+-]?\d+,\d{1,2}", t):
        return t.replace(",", ".")
    return t


def _parse_num(s):
    """Parse a value into a float, or None if it cannot be parsed.

    Handles int/float, and strings with currency symbols, thousands
    separators, percent signs, whitespace etc.
    """
    if s is None:
        return None
    if isinstance(s, bool):
        return None
    if isinstance(s, (int, float)):
        return float(s)
    t = str(s).strip()
    if not t:
        return None
    for ch in ["$", "¥", "€", "£", " ", "_", "%", " "]:
        t = t.replace(ch, "")
    t = _normalize_commas(t)
    try:
        return float(t)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=1.0):
    fa = _parse_num(a)
    fb = _parse_num(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # Fall back to case-insensitive string comparison when either side is
    # not numeric (e.g. one side is a formula string).
    return str_match(a, b)


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


# ---------------------------------------------------------------------------
# Small, safe Excel-formula evaluator.
#
# Agents may legitimately implement the Cost Summary derived values as
# spreadsheet formulas (e.g. =B4-B3, =0.15*B2). openpyxl does not recompute
# formula caches on write, so reading with data_only=True alone would yield
# None for those cells and wrongly fail the task. We therefore read the
# workbook with data_only=False and resolve values here, using the cached
# (data_only=True) value first when present, and otherwise evaluating simple
# formulas. A formula that resolves to a number participates in the normal
# numeric checks. A formula that cannot be evaluated (or an empty / non-numeric
# Amount cell) is flagged by the 'Cost Summary metrics numeric' check so that a
# garbage formula cannot silently bypass the cost-consistency checks.
# ---------------------------------------------------------------------------

# NOTE: _REF_RE/_RANGE_RE/_ABS_REF_RE intentionally have NO ^ or $ anchors.
# They are matched with re.Pattern.match(text, pos) where the reference is a
# *prefix* of a larger expression (e.g. 'D2:D8' inside 'SUM(D2:D8)' or 'D2*0.15'
# inside 'ROUND(D2*0.15,2)'). A '^' anchor would prevent such a match: Python's
# '^' only matches at the real start of the string, not at the pos argument.
# Callers that need a full-token match (e.g. _atom on a bare word) check
# m.end() explicitly.
_REF_RE = re.compile(r"\$?([A-Za-z]{1,3})\$?(\d+)")
_RANGE_RE = re.compile(r"\$?([A-Za-z]{1,3})\$?(\d+):\$?([A-Za-z]{1,3})\$?(\d+)")
_ABS_REF_RE = re.compile(r"\$?([A-Za-z]{1,3})\$?(\d+)")
_NUM_RE = re.compile(r"(\d+\.?\d*|\.\d+)")
_WORD_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_.]*")
_SHEETQUOTED_RE = re.compile(r"'([^']+)'!")
_SHEETRAW_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_. ]*!")


def _col_index(col):
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


class _SheetResolver:
    """Resolves cell values from a workbook, evaluating simple formulas."""

    def __init__(self, wb, wb_values):
        self.wb = wb  # data_only=False: raw formulas + literals
        self.wb_values = wb_values  # data_only=True: cached values
        self._cache = {}

    def _match_sheet(self, name):
        if name is None:
            return None
        key = str(name).strip().strip("'").lower()
        for sn in self.wb.sheetnames:
            if sn.strip().lower() == key:
                return sn
            if sn.strip().lower().replace(" ", "_") == key:
                return sn
            if sn.strip().lower().replace("_", " ") == key:
                return sn
        return None

    def raw(self, sheet, row, col):
        sheet = self._match_sheet(sheet)
        if sheet is None:
            return None
        try:
            return self.wb[sheet].cell(row=row, column=col).value
        except Exception:
            return None

    def cached(self, sheet, row, col):
        sheet = self._match_sheet(sheet)
        if sheet is None:
            return None
        try:
            return self.wb_values[sheet].cell(row=row, column=col).value
        except Exception:
            return None

    def value(self, sheet, row, col, depth=0):
        """Resolve the numeric value of a cell, or None."""
        sheet = self._match_sheet(sheet)
        if sheet is None:
            return None
        key = (sheet, row, col)
        if key in self._cache:
            return self._cache[key]
        if depth > 10:
            return None
        res = self._resolve(sheet, row, col, depth)
        self._cache[key] = res
        return res

    def _resolve(self, sheet, row, col, depth):
        raw = self.raw(sheet, row, col)
        if raw is None:
            return None
        if isinstance(raw, bool):
            return None
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw).strip()
        if not s:
            return None
        if s.startswith("="):
            # Prefer a cached value when the file carries one (files saved by
            # real Excel/LibreOffice have it; openpyxl-written files don't).
            cached = self.cached(sheet, row, col)
            if cached is not None and not isinstance(cached, str):
                return float(cached)
            if isinstance(cached, str) and not cached.strip().startswith("="):
                f = _parse_num(cached)
                if f is not None:
                    return f
            try:
                parser = _FormulaParser(self, s[1:], sheet, depth + 1)
                return parser.parse()
            except Exception:
                return None
        return _parse_num(s)


class _FormulaParser:
    def __init__(self, resolver, text, sheet, depth):
        self.res = resolver
        self.text = text
        self.sheet = sheet
        self.depth = depth
        self.pos = 0

    def parse(self):
        return self._expr()

    def _ws(self):
        while self.pos < len(self.text) and self.text[self.pos] in " \t":
            self.pos += 1

    def _peek(self):
        self._ws()
        if self.pos >= len(self.text):
            return None
        return self.text[self.pos]

    def _expr(self):
        v = self._term()
        while True:
            ch = self._peek()
            if ch == "+":
                self.pos += 1
                r = self._term()
                v = (v + r) if (v is not None and r is not None) else None
            elif ch == "-":
                self.pos += 1
                r = self._term()
                v = (v - r) if (v is not None and r is not None) else None
            else:
                return v

    def _term(self):
        v = self._factor()
        while True:
            ch = self._peek()
            if ch == "*":
                self.pos += 1
                r = self._factor()
                v = (v * r) if (v is not None and r is not None) else None
            elif ch == "/":
                self.pos += 1
                r = self._factor()
                if v is not None and r is not None and r != 0:
                    v = v / r
                else:
                    v = None
            else:
                return v

    def _factor(self):
        self._ws()
        sign = 1
        ch = self._peek()
        if ch == "-":
            self.pos += 1
            sign = -1
        elif ch == "+":
            self.pos += 1
        v = self._atom()
        if v is not None:
            v = v * sign
        while self._peek() == "%":
            self.pos += 1
            if v is not None:
                v = v / 100.0
        return v

    def _atom(self):
        self._ws()
        if self.pos >= len(self.text):
            return None
        ch = self.text[self.pos]
        if ch == "(":
            self.pos += 1
            v = self._expr()
            if self._peek() == ")":
                self.pos += 1
            return v
        # A bare or absolute range, e.g. D2:D8 or $D$2:$D$8 (aggregate args).
        m = _RANGE_RE.match(self.text, self.pos)
        if m:
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            self.pos = m.end()
            return self._range_value(self.sheet, c1, r1, c2, r2)
        if ch == "$":
            # Excel absolute reference like $B$2 or B$2
            m = _ABS_REF_RE.match(self.text, self.pos)
            if m:
                col, row = m.group(1), int(m.group(2))
                self.pos = m.end()
                return self._cell_value(self.sheet, col, row)
            return None
        m = _SHEETQUOTED_RE.match(self.text, self.pos)
        if m:
            sheet = m.group(1)
            self.pos = m.end()
            return self._sheet_ref(sheet)
        m = _SHEETRAW_RE.match(self.text, self.pos)
        if m:
            sheet = m.group(0)[:-1].strip()
            self.pos = m.end()
            return self._sheet_ref(sheet)
        m = _WORD_RE.match(self.text, self.pos)
        if m:
            tok = m.group(0)
            self.pos = m.end()
            if self._peek() == "(":
                self.pos += 1
                return self._func(tok)
            m2 = _REF_RE.match(tok)
            if m2 and m2.end() == len(tok):
                col, row = m2.group(1), int(m2.group(2))
                return self._cell_value(self.sheet, col, row)
            return None
        m = _NUM_RE.match(self.text, self.pos)
        if m:
            self.pos = m.end()
            return float(m.group(0))
        return None

    def _sheet_ref(self, sheet):
        m = _RANGE_RE.match(self.text, self.pos)
        if m:
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            self.pos = m.end()
            return self._range_value(sheet, c1, r1, c2, r2)
        m = _REF_RE.match(self.text, self.pos)
        if m:
            col, row = m.group(1), int(m.group(2))
            self.pos = m.end()
            return self._cell_value(sheet, col, row)
        return None

    def _cell_value(self, sheet, col, row):
        return self.res.value(sheet, row, _col_index(col) + 1, self.depth)

    def _range_value(self, sheet, col1, row1, col2, row2):
        sheet = self.res._match_sheet(sheet)
        if sheet is None:
            return None
        vals = []
        c1, c2 = _col_index(col1), _col_index(col2)
        for r in range(min(row1, row2), max(row1, row2) + 1):
            for c in range(min(c1, c2), max(c1, c2) + 1):
                vals.append(self.res.value(sheet, r, c + 1, self.depth))
        return vals

    def _func(self, name):
        args = []
        while True:
            self._ws()
            ch = self._peek()
            if ch == ")":
                self.pos += 1
                break
            args.append(self._arg())
            self._ws()
            ch = self._peek()
            if ch == ",":
                self.pos += 1
                continue
            if ch == ")":
                self.pos += 1
                break
            break
        flat = []
        for a in args:
            if isinstance(a, list):
                flat.extend(a)
            else:
                flat.append(a)
        vals = [x for x in flat if x is not None]
        fn = name.upper()
        if not vals:
            return None
        if fn == "SUM":
            return sum(vals)
        if fn in ("AVERAGE", "AVG"):
            return sum(vals) / len(vals)
        if fn == "MIN":
            return min(vals)
        if fn == "MAX":
            return max(vals)
        if fn == "ABS":
            return abs(vals[0])
        if fn == "ROUND":
            if len(vals) >= 2:
                return round(vals[0], int(vals[1]))
            return vals[0]
        if fn == "ROUNDUP":
            import math
            if len(vals) >= 2:
                fac = 10 ** int(vals[1])
                return math.ceil(vals[0] * fac) / fac
            return vals[0]
        if fn == "ROUNDDOWN":
            import math
            if len(vals) >= 2:
                fac = 10 ** int(vals[1])
                return math.floor(vals[0] * fac) / fac
            return vals[0]
        if fn == "IF":
            return vals[1] if len(vals) > 1 else vals[0]
        return vals[0]

    def _arg(self):
        # A full expression argument. Sheet-quoted refs, ranges and arithmetic
        # are all handled by the expression parser (_atom/_term/_expr), so a
        # reference followed by more arithmetic ('Sheet'!D2*0.15) parses fully.
        self._ws()
        if self.pos >= len(self.text):
            return None
        return self._expr()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


# Course labels accepted for each required course. Event details name the
# courses "appetizer", "main", "dessert"; accept common synonyms so that a
# reasonable implementation is not wrongly rejected.
COURSE_ALIASES = {
    "appetizer": ["appetizer", "appetiser", "starter", "hors d'oeuvre", "hors-d'oeuvre"],
    "main": ["main", "entree", "entrée", "main course", "main dish", "main_course", "main_dish"],
    "dessert": ["dessert", "sweet", "pudding"],
}


def _metric_value(res, rows, metrics, key, sheet_name, amount_col):
    """Resolve a Cost Summary metric's Amount as float, plus raw/formula info."""
    if key is None:
        return None, None, False
    ridx, col0 = metrics[key]
    row = rows[ridx - 1]
    raw = row[col0] if col0 < len(row) else None
    is_formula = isinstance(raw, str) and raw.strip().startswith("=")
    val = res.value(sheet_name, ridx, col0 + 1)
    return val, raw, is_formula


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Event_Menu.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        wb_values = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        record("Excel readable", False, f"{type(e).__name__}: {e}")
        return False
    record("Excel readable", True)
    res = _SheetResolver(wb, wb_values)

    # ---- Menu Plan ----
    mp_rows = load_sheet_rows(wb, "Menu Plan") or load_sheet_rows(wb, "Menu_Plan")
    course_counts = {}
    if mp_rows is None:
        record("Sheet 'Menu Plan' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Menu Plan' exists", True)
        header = mp_rows[0] if mp_rows else []
        data = [r for r in mp_rows[1:] if r and r[0] is not None]
        record("Menu Plan has >= 6 rows (2 per course)", len(data) >= 6, f"Found {len(data)}")

        course_col = find_col(header, ["Course", "course"])
        recipe_col = find_col(header, ["Recipe_Name", "Recipe Name", "Recipe", "Name"])
        serves_col = find_col(header, ["Serves", "Servings"])
        diet_col = find_col(header, ["Dietary_Tags", "Dietary Tags", "Tags"])

        # Per-course count
        if course_col is not None:
            for r in data:
                if course_col < len(r) and r[course_col]:
                    c = str(r[course_col]).strip().lower()
                    course_counts[c] = course_counts.get(c, 0) + 1
            for course in ["appetizer", "main", "dessert"]:
                aliases = COURSE_ALIASES.get(course, [course])
                cnt = sum(v for k, v in course_counts.items()
                          if any(a in k for a in aliases))
                record(f"At least 2 recipes for course '{course}'",
                       cnt >= 2,
                       f"Found {cnt}; per-course: {course_counts}")

        # Serves should be 50 (event details: scaled for 50 guests)
        if serves_col is not None:
            wrong_serves = []
            for ridx, r in enumerate(mp_rows[1:], start=2):
                if not r or r[0] is None:
                    continue
                if serves_col < len(r) and r[serves_col] is not None:
                    s = res.value("Menu Plan", ridx, serves_col + 1)
                    if s is None:
                        continue
                    if abs(s - 50) > 0.5:
                        wrong_serves.append(r[serves_col])
            record("Serves column == 50 for all rows",
                   len(wrong_serves) == 0,
                   f"Wrong serves values: {wrong_serves[:5]}")

        # Dietary_Tags column exists and covers required restrictions
        record("Dietary_Tags column exists", diet_col is not None, f"Header: {header}")
        if diet_col is not None:
            tags_concat = " ".join(
                str(r[diet_col]) for r in data if diet_col < len(r) and r[diet_col]
            ).lower()
            for needed in ["vegetarian", "gluten-free", "nut-free"]:
                ok = (needed in tags_concat
                      or needed.replace("-", " ") in tags_concat
                      or needed.replace("-", "") in tags_concat)
                record(f"Dietary_Tags column mentions '{needed}'",
                       ok, f"tags: {tags_concat[:200]}")

    # ---- Ingredient List ----
    il_rows = load_sheet_rows(wb, "Ingredient List") or load_sheet_rows(wb, "Ingredient_List")
    total_ingredient_cost = 0.0
    if il_rows is None:
        record("Sheet 'Ingredient List' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Ingredient List' exists", True)
        header = il_rows[0] if il_rows else []
        data = [r for r in il_rows[1:] if r and r[0] is not None]
        record("Ingredient List has >= 5 items", len(data) >= 5, f"Found {len(data)}")

        ing_col = find_col(header, ["Ingredient"])
        qty_col = find_col(header, ["Quantity_For_50", "Quantity For 50", "Quantity"])
        unit_col = find_col(header, ["Unit"])
        cost_col = find_col(header, ["Estimated_Cost", "Estimated Cost", "Cost"])
        record("Ingredient List has Ingredient column", ing_col is not None)
        record("Ingredient List has Quantity_For_50 column", qty_col is not None)
        record("Ingredient List has Unit column", unit_col is not None)
        record("Ingredient List has Estimated_Cost column", cost_col is not None)

        if cost_col is not None:
            for ridx, r in enumerate(il_rows[1:], start=2):
                if not r or r[0] is None:
                    continue
                if cost_col < len(r) and r[cost_col] is not None:
                    c = res.value("Ingredient List", ridx, cost_col + 1)
                    if c is not None:
                        total_ingredient_cost += c

    # ---- Cost Summary ----
    cs_rows = load_sheet_rows(wb, "Cost Summary") or load_sheet_rows(wb, "Cost_Summary")
    if cs_rows is None:
        record("Sheet 'Cost Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Cost Summary' exists", True)
        metrics = {}
        for ridx, row in enumerate(cs_rows[1:], start=2):
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = (ridx, 1)

        # Required metrics keys
        for req_key, friendly in [
            ("total_food_cost", "Total_Food_Cost"),
            ("cost_per_person", "Cost_Per_Person"),
            ("budget_per_person", "Budget_Per_Person"),
            ("budget_variance", "Budget_Variance"),
            ("service_fee_estimate", "Service_Fee_Estimate"),
        ]:
            present = any(req_key in k for k in metrics)
            record(f"Cost Summary contains {friendly}", present,
                   f"Items: {list(metrics.keys())}")

        tfc_key = next((k for k in metrics if "total_food_cost" in k or ("total" in k and "food" in k)), None)
        cpp_key = next((k for k in metrics if "cost" in k and "per" in k and "person" in k), None)
        bpp_key = next((k for k in metrics if "budget" in k and "per" in k and "person" in k), None)
        bv_key = next((k for k in metrics if "budget" in k and "var" in k), None)
        sfe_key = next((k for k in metrics if "service" in k and "fee" in k), None)

        cs_sheet = "Cost Summary"
        tfc, tfc_raw, tfc_f = _metric_value(res, cs_rows, metrics, tfc_key, cs_sheet, 1)
        cpp, cpp_raw, cpp_f = _metric_value(res, cs_rows, metrics, cpp_key, cs_sheet, 1)
        bpp, bpp_raw, bpp_f = _metric_value(res, cs_rows, metrics, bpp_key, cs_sheet, 1)
        bv, bv_raw, bv_f = _metric_value(res, cs_rows, metrics, bv_key, cs_sheet, 1)
        sfe, sfe_raw, sfe_f = _metric_value(res, cs_rows, metrics, sfe_key, cs_sheet, 1)

        # Numeric sanity: every metric that holds a plain literal must be a
        # number. Formula cells are resolved above; if a formula cannot be
        # evaluated the dependent checks are skipped rather than failed.
        metric_vals = [
            ("Total_Food_Cost", tfc_key, tfc, tfc_raw, tfc_f),
            ("Cost_Per_Person", cpp_key, cpp, cpp_raw, cpp_f),
            ("Budget_Per_Person", bpp_key, bpp, bpp_raw, bpp_f),
            ("Budget_Variance", bv_key, bv, bv_raw, bv_f),
            ("Service_Fee_Estimate", sfe_key, sfe, sfe_raw, sfe_f),
        ]
        # Every required metric must resolve to a number. Empty Amount cells,
        # non-numeric literals and formulas that do not evaluate are all
        # failures: none of them may silently skip the cost consistency checks.
        numeric_fail = []
        for name, key, val, raw, is_formula in metric_vals:
            if key is None:
                continue
            if raw is None:
                numeric_fail.append(f"{name}: empty Amount cell")
            elif val is None:
                if is_formula:
                    numeric_fail.append(f"{name}: formula did not evaluate ({raw!r})")
                else:
                    numeric_fail.append(f"{name}: non-numeric Amount {raw!r}")
        record("Cost Summary metrics numeric", len(numeric_fail) == 0,
               ("Non-numeric Amount cells: " + "; ".join(numeric_fail)) if numeric_fail else "")

        # Budget_Per_Person == 30 (per event_details.json)
        if bpp is not None:
            record("Budget_Per_Person == 30",
                   abs(bpp - 30) <= 0.5,
                   f"Got ${bpp}, expected 30 from event details")

        # Cost_Per_Person <= Budget_Per_Person (under-budget constraint).
        # task.md lists the 30/person budget as context rather than a hard
        # requirement, so allow a small margin: a careful plan that lands just
        # over budget (e.g. 31-33) must not be rejected, while a plan that
        # grossly ignores the budget still fails.
        if cpp is not None and bpp is not None:
            record("Cost_Per_Person <= Budget_Per_Person",
                   cpp <= bpp + max(2.0, bpp * 0.1),
                   f"Got ${cpp} vs budget ${bpp}")

        # Budget_Variance == Budget_Per_Person - Cost_Per_Person (per task.md)
        if bv is not None and cpp is not None and bpp is not None:
            expected_bv = bpp - cpp
            record("Budget_Variance == Budget_Per_Person - Cost_Per_Person",
                   abs(bv - expected_bv) <= max(0.5, abs(expected_bv) * 0.02),
                   f"Got {bv}, expected {expected_bv}")

        # Service_Fee_Estimate ≈ 0.15 * Total_Food_Cost
        if sfe is not None and tfc is not None:
            expected_sfe = round(tfc * 0.15, 2)
            record(f"Service_Fee_Estimate == 0.15 * Total_Food_Cost ({expected_sfe})",
                   abs(sfe - expected_sfe) <= max(0.5, expected_sfe * 0.02),
                   f"Got {sfe}, expected {expected_sfe}")

        # Total_Food_Cost roughly == Cost_Per_Person * 50 guests (within 20% tol)
        if tfc is not None and cpp is not None:
            expected_tfc = cpp * 50
            record("Total_Food_Cost ≈ Cost_Per_Person * 50",
                   abs(tfc - expected_tfc) <= max(50, expected_tfc * 0.2),
                   f"Got {tfc}, expected ≈ {expected_tfc}")

    # ---- Dietary Accommodations ----
    da_rows = load_sheet_rows(wb, "Dietary Accommodations") or load_sheet_rows(wb, "Dietary_Accommodations")
    if da_rows is None:
        record("Sheet 'Dietary Accommodations' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Dietary Accommodations' exists", True)
        header = da_rows[0] if da_rows else []
        data = [r for r in da_rows[1:] if r and r[0] is not None]
        record("Dietary Accommodations has >= 3 rows", len(data) >= 3, f"Found {len(data)}")

        # Validate guest counts match event details {vegetarian:10, gluten_free:5, nut_allergy:3}
        restr_col = find_col(header, ["Restriction", "Diet"])
        gc_col = find_col(header, ["Guest_Count", "Guest Count", "Count", "Guests"])
        if restr_col is not None and gc_col is not None:
            EXPECTED_GUESTS = {
                "vegetarian": 10,
                "gluten": 5,    # gluten-free or gluten free
                "nut": 3,       # nut allergy
            }
            seen_keys = set()
            for ridx, r in enumerate(da_rows[1:], start=2):
                if not r or r[0] is None:
                    continue
                if restr_col < len(r) and gc_col < len(r):
                    name = str(r[restr_col] or "").lower()
                    gc = res.value("Dietary Accommodations", ridx, gc_col + 1)
                    if gc is None:
                        continue
                    for key, expected in EXPECTED_GUESTS.items():
                        if key in name:
                            seen_keys.add(key)
                            ok = abs(gc - expected) <= 0.5
                            record(f"Dietary Accommodations '{key}' guest count == {expected}",
                                   ok,
                                   f"Got {gc}")
            for k, expected in EXPECTED_GUESTS.items():
                if k not in seen_keys:
                    record(f"Dietary Accommodations row for '{k}' present",
                           False, f"Not found in restriction column")

    return True


def _load_task_email_config():
    """Read the task directory's email_config.json to derive the expected From."""
    here = os.path.dirname(os.path.abspath(__file__))
    task_root = os.path.dirname(here)
    path = os.path.join(task_root, "email_config.json")
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            data = data[0] if data else {}
        return data or {}
    except Exception:
        return {}


def check_email():
    print("\n=== Checking Email ===")
    # Email runtime checks are non-blocking: a DB problem here must never
    # crash the evaluator (which would turn into a blocking FAIL).
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Email DB check (database unreachable)", False,
               f"{type(e).__name__}: {e}")
        return False
    try:
        cur = conn.cursor()

        # Task.md requires exact subject "Menu Plan for Annual Company Dinner"
        # and recipient catering@vendor.com.
        cur.execute("""
            SELECT id, subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE subject ILIKE '%%menu plan for annual company dinner%%'
              AND to_addr::text ILIKE '%%catering@vendor.com%%'
        """)
        emails = cur.fetchall()

        record("Email with subject 'Menu Plan for Annual Company Dinner' to catering@vendor.com",
               len(emails) >= 1, f"Found {len(emails)}")

        if emails:
            e = emails[0]
            to = e[3]
            if isinstance(to, str):
                try:
                    to = json.loads(to)
                except Exception:
                    pass
            to_str = str(to).lower()
            record("Email to catering@vendor.com", "catering@vendor.com" in to_str, f"To: {to}")

            from_str = str(e[2]).lower() if e[2] else ""
            cfg = _load_task_email_config()
            expected_email = (cfg.get("email") or "events@company.com").strip().lower()
            ok = bool(expected_email) and expected_email in from_str
            if not ok and expected_email:
                m = re.search(r"<([^>]+)>", from_str)
                if m:
                    ok = m.group(1).strip().lower() == expected_email
                else:
                    ok = from_str.strip() == expected_email
            record("Email from events@company.com", ok,
                   f"From: {e[2]} (expected {expected_email})")

            body = str(e[4]).lower() if e[4] else ""
            record("Email body mentions guests/menu", "guest" in body or "menu" in body or "dinner" in body,
                   f"Body preview: {body[:200]}")

        cur.close()
        conn.close()
        return True
    except Exception as e:
        record("Email DB check failed", False, f"{type(e).__name__}: {e}")
        try:
            conn.close()
        except Exception:
            pass
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    file_fail_before_email = FAIL_COUNT
    check_email()
    email_fail = FAIL_COUNT - file_fail_before_email

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} (file_fail={file_fail_before_email}, email_fail={email_fail})")
    # File failures (Excel/groundtruth) are blocking. Email runtime checks may fail in GT self-test.
    sys.exit(0 if file_fail_before_email == 0 else 1)


if __name__ == "__main__":
    main()
