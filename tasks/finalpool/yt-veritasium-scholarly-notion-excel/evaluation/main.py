"""
Evaluation for yt-veritasium-scholarly-notion-excel task.

Checks:
1. Science_Resource_Map.xlsx exists with Videos and Papers sheets
2. Videos sheet has 5 rows with top Veritasium videos and topics
3. Papers sheet has 15 rows (papers found per topic)
4. Notion page "Science Video-Paper Resource Map" exists with content
"""
import json
import os
import re
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def norm(s):
    return str(s or "").strip().lower().replace("_", " ")


def safe_float(v):
    try:
        if v is None: return None
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


# Top 5 Veritasium video metadata (READ-ONLY, fixed in DB)
EXPECTED_TOP5 = [
    {"rank": 1, "video_id": "SC2eSujzrUY", "title_kw": ["poisoned", "planet"], "view_count": 32716281},
    {"rank": 2, "video_id": "88bMVbx1dzM", "title_kw": ["zooming"], "view_count": 24771026},
    {"rank": 3, "video_id": "Q56PMJbCFXQ", "title_kw": ["skyscraper"], "view_count": 23724425},
    {"rank": 4, "video_id": "qJZ1Ez28C-A", "title_kw": ["quantum"], "view_count": 17545470},
    {"rank": 5, "video_id": "Q10_srZ-pbs", "title_kw": ["theory", "everything"], "view_count": 16347242},
]


def _fetch_top5_from_db():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT v.video_id, v.title, v.view_count
            FROM youtube.videos v
            JOIN youtube.channels c ON v.channel_id=c.channel_id
            WHERE c.title ILIKE '%veritasium%'
            ORDER BY v.view_count DESC LIMIT 5
        """)
        rows = []
        for vid, title, vc in cur.fetchall():
            rows.append({"video_id": vid, "title": title, "view_count": int(vc)})
        conn.close()
        return rows
    except Exception as e:
        print(f"[WARN] DB top5 fetch: {e}")
        return None


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Science_Resource_Map.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Science_Resource_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Science_Resource_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Science_Resource_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_norm = {norm(s): s for s in wb.sheetnames}
    db_top5 = _fetch_top5_from_db() or [
        {"video_id": e["video_id"], "title": "", "view_count": e["view_count"]}
        for e in EXPECTED_TOP5
    ]

    # Videos sheet
    v_key = sheet_norm.get(norm("Videos"))
    record("Videos sheet exists (exact)", v_key is not None, f"Sheets: {wb.sheetnames}")
    if v_key:
        ws = wb[v_key]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [norm(c) for c in rows[0]]
            for h in ["rank", "video id", "title", "view count", "main topic", "paper count"]:
                record(f"Videos has '{h}' column", h in headers, f"Headers: {rows[0]}")
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        record("Videos sheet has 5 rows", len(data_rows) == 5,
               f"Found {len(data_rows)} rows")

        if rows and data_rows and len(headers) >= 6:
            vid_idx = headers.index("video id") if "video id" in headers else None
            rank_idx = headers.index("rank") if "rank" in headers else None
            vc_idx = headers.index("view count") if "view count" in headers else None
            mt_idx = headers.index("main topic") if "main topic" in headers else None
            pc_idx = headers.index("paper count") if "paper count" in headers else None

            agent_vids = {}
            for r in data_rows:
                if vid_idx is not None and r[vid_idx]:
                    agent_vids[str(r[vid_idx]).strip()] = r

            # Each expected top5 video must be present
            for v in db_top5:
                vid = v["video_id"]
                r = agent_vids.get(vid)
                if r is None:
                    record(f"Video {vid} present", False,
                           f"agent vids: {list(agent_vids.keys())}")
                    continue
                record(f"Video {vid} present", True)
                if vc_idx is not None:
                    av = safe_float(r[vc_idx])
                    record(f"Video {vid} view_count ~ {v['view_count']}",
                           av is not None and abs(av - v['view_count']) <= max(1.0, v['view_count']*0.005),
                           f"got {av}")
                if mt_idx is not None:
                    record(f"Video {vid} has Main_Topic",
                           r[mt_idx] is not None and str(r[mt_idx]).strip() != "",
                           f"got {r[mt_idx]}")
                if pc_idx is not None:
                    av = safe_float(r[pc_idx])
                    # paper_count should be integer 1-3 (each topic should have papers found)
                    record(f"Video {vid} paper_count is integer 1-3",
                           av is not None and 1 <= av <= 3,
                           f"got {av}")

            # Sorting by Rank
            if rank_idx is not None:
                ranks = [safe_float(r[rank_idx]) for r in data_rows]
                ranks_clean = [v for v in ranks if v is not None]
                record("Videos sorted by Rank asc",
                       ranks_clean == sorted(ranks_clean),
                       f"got {ranks}")
                # Ranks should be exactly 1..5
                record("Ranks are 1..5",
                       sorted(ranks_clean) == [1.0, 2.0, 3.0, 4.0, 5.0],
                       f"got {ranks_clean}")

            # Total paper_count across the 5 videos
            if pc_idx is not None:
                total_pc = sum(safe_float(r[pc_idx]) or 0 for r in data_rows)
                # Up to 3 per topic, 5 topics = max 15; require >= 10 (most topics covered)
                record("Total Paper_Count between 10 and 15",
                       10 <= total_pc <= 15,
                       f"got {total_pc}")

    # Papers sheet
    p_key = sheet_norm.get(norm("Papers"))
    record("Papers sheet exists (exact)", p_key is not None, f"Sheets: {wb.sheetnames}")
    if p_key:
        ws2 = wb[p_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        if rows2:
            headers2 = [norm(c) for c in rows2[0]]
            for h in ["topic", "paper title", "authors", "year", "citations"]:
                record(f"Papers has '{h}' column", h in headers2, f"Headers: {rows2[0]}")
            # Up to 15 papers; should be at least 10 (most topics covered)
            record("Papers sheet has >= 10 rows",
                   len(data_rows2) >= 10, f"Found {len(data_rows2)} rows")
            # Validate each row has non-empty fields
            if all(h in headers2 for h in ["topic", "paper title", "authors", "year"]):
                t_idx = headers2.index("topic")
                pt_idx = headers2.index("paper title")
                au_idx = headers2.index("authors")
                yr_idx = headers2.index("year")
                bad = 0
                for r in data_rows2:
                    if not (r[t_idx] and r[pt_idx] and r[au_idx] and r[yr_idx]):
                        bad += 1
                record("All Papers rows have non-empty Topic/Title/Authors/Year",
                       bad == 0, f"{bad} bad rows")
                # Years should be 2023 (per preprocess data)
                bad_yr = 0
                for r in data_rows2:
                    yv = safe_float(r[yr_idx])
                    if yv is None or yv < 2010 or yv > 2030:
                        bad_yr += 1
                record("All Papers years in [2010, 2030]",
                       bad_yr == 0, f"{bad_yr} bad year rows")


def check_notion():
    print("\n=== Check 2: Notion page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Match exact title
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false
        """)
        all_pages = cur.fetchall()
        target = None
        for pid, props in all_pages:
            try:
                title_arr = props.get('title', {}).get('title', [])
                title_text = "".join([t.get('text', {}).get('content', '') for t in title_arr])
                if norm(title_text) == norm("Science Video-Paper Resource Map"):
                    target = pid
                    break
            except Exception:
                pass
        record("Notion page 'Science Video-Paper Resource Map' exists",
               target is not None, f"checked {len(all_pages)} pages")

        if target:
            # Check page content - look at blocks
            cur.execute("SELECT block_type, content FROM notion.blocks WHERE page_id=%s ORDER BY position", (target,))
            blocks = cur.fetchall()
            full_text = ""
            for btype, content in blocks:
                if isinstance(content, dict):
                    rt = content.get("rich_text", []) or content.get("text", [])
                    if isinstance(rt, list):
                        for piece in rt:
                            if isinstance(piece, dict):
                                # piece may have 'text' -> 'content' or 'plain_text'
                                full_text += " " + piece.get("plain_text", "") + " " + piece.get("text", {}).get("content", "")
                else:
                    full_text += " " + str(content)
            full_text_lower = full_text.lower()
            # Should mention all 5 video topic keywords or video ids
            kw_hits = 0
            for v in EXPECTED_TOP5:
                if any(kw in full_text_lower for kw in v["title_kw"]):
                    kw_hits += 1
            record("Notion page mentions >= 4 of 5 videos by keyword",
                   kw_hits >= 4, f"hits={kw_hits}, text[:300]={full_text_lower[:300]}")

        cur.close()
        conn.close()
    except Exception as e:
        record("Notion check", False, str(e))


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
