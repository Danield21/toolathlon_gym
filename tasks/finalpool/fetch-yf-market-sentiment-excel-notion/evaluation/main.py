"""Evaluation script for fetch-yf-market-sentiment-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-yf-market-sentiment-excel-notion"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Market_Sentiment_Report.xlsx")
    check("Market_Sentiment_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Market_Sentiment_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames and gt_wb is not None:
            ws = wb["Data_Analysis"]
            gt_ws = gt_wb["Data_Analysis"] if "Data_Analysis" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Data_Analysis has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Data_Analysis', ['Symbol', 'Current_Price', 'Target_Price', 'Upside'],
                          len(gt_rows) if gt_ws is not None else 5)
            if gt_ws is not None:
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                header_map = {h: i for i, h in enumerate(headers)}
                sym_idx_gt = gt_headers.index("symbol") if "symbol" in gt_headers else 0
                sym_idx_a = header_map.get("symbol", 0)
                gt_by_sym = {str(r[sym_idx_gt]).strip().upper(): r for r in gt_rows if r and r[sym_idx_gt]}
                agent_by_sym = {str(r[sym_idx_a]).strip().upper(): r for r in data_rows if r and len(r) > sym_idx_a and r[sym_idx_a]}
                for sym, gt_row in gt_by_sym.items():
                    found = sym in agent_by_sym
                    check(f"Data_Analysis symbol '{sym}' present", found)
                    if found:
                        agent_row = agent_by_sym[sym]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row) or gt_h == "symbol":
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                tol = max(0.25, abs(gf) * 0.05)
                                check(f"Data_Analysis '{sym}' {gt_h} ~{gf}",
                                      abs(gf - af) <= tol, f"got {av}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames and gt_wb is not None:
            ws = wb["Metrics"]
            gt_ws = gt_wb["Metrics"] if "Metrics" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Metrics has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Metrics', ['Metric', 'Value'], len(gt_rows) if gt_ws is not None else 3)
            if gt_ws is not None:
                gt_metric_map = {str(r[0]).strip().lower(): r[1] for r in gt_rows if r and r[0]}
                agent_metric_map = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for gt_m, gt_v in gt_metric_map.items():
                    found = gt_m in agent_metric_map
                    check(f"Metrics has '{gt_m}'", found)
                    if found:
                        av = agent_metric_map[gt_m]
                        gf = safe_float(gt_v)
                        af = safe_float(av)
                        if gf is not None and af is not None:
                            tol = max(0.25, abs(gf) * 0.05)
                            check(f"Metrics '{gt_m}' ~{gf}",
                                  abs(gf - af) <= tol, f"got {av}")
                        else:
                            # Non-numeric metric value: enforce exact (case-insensitive) match
                            if gt_v is not None and av is not None:
                                check(f"Metrics '{gt_m}' == '{gt_v}'",
                                      str(gt_v).strip().lower() == str(av).strip().lower(),
                                      f"got {av}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames and gt_wb is not None:
            ws = wb["Recommendations"]
            gt_ws = gt_wb["Recommendations"] if "Recommendations" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [r for r in data_rows if r and any(v is not None and str(v).strip() for v in r)]
            # Task spec: "at least two actionable items" — relax row count to >= 2
            check("Recommendations has >= 2 rows",
                  len(data_rows) >= 2, f"got {len(data_rows)}")
            check_columns('Recommendations', ['Priority', 'Action'], 2)
            # Validate each row has both Priority and Action populated; do not enforce
            # an exact (Priority, Action) tuple from GT because the task description only
            # mandates "at least two actionable items". We check that priorities are
            # distinct (so two genuinely different items) and that actions are non-empty
            # buy/hold/sell-style verbs (or at least non-empty strings).
            if data_rows:
                # Identify column indexes (Priority, Action)
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                p_idx = headers.index("priority") if "priority" in headers else 0
                a_idx = headers.index("action") if "action" in headers else 1
                priorities = [str(r[p_idx]).strip() for r in data_rows if len(r) > p_idx and r[p_idx] is not None and str(r[p_idx]).strip()]
                actions = [str(r[a_idx]).strip() for r in data_rows if len(r) > a_idx and r[a_idx] is not None and str(r[a_idx]).strip()]
                check("Recommendations priorities non-empty for all rows",
                      len(priorities) == len(data_rows),
                      f"priorities: {priorities}")
                check("Recommendations actions non-empty for all rows",
                      len(actions) == len(data_rows),
                      f"actions: {actions}")
                check("Recommendations has at least 2 distinct priorities",
                      len(set(p.lower() for p in priorities)) >= 2,
                      f"priorities: {priorities}")
                # Each Action should be a recognised verb: Buy, Hold, or Sell.
                allowed_actions = {"buy", "hold", "sell"}
                action_words = [a.lower() for a in actions]
                check("Recommendations actions in {Buy, Hold, Sell}",
                      all(any(w in a for w in allowed_actions) for a in action_words) and len(action_words) >= 2,
                      f"actions: {actions}")

        # Notion: page titled 'Market Sentiment Dashboard'
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("""SELECT id, properties FROM notion.pages
                           WHERE archived = false""")
            rows = cur.fetchall()
            target_lowers = {"market sentiment dashboard"}
            found_count = 0
            for pid, props in rows:
                if not props:
                    continue
                try:
                    if isinstance(props, str):
                        props_obj = json.loads(props)
                    else:
                        props_obj = props
                    titles = []
                    for k, v in (props_obj.items() if isinstance(props_obj, dict) else []):
                        if isinstance(v, dict) and v.get("type") == "title":
                            tlist = v.get("title", [])
                            for t in tlist:
                                if isinstance(t, dict):
                                    pt = t.get("plain_text", "") or (t.get("text", {}).get("content", "") if isinstance(t.get("text"), dict) else "")
                                    if pt:
                                        titles.append(pt)
                    title_text = " ".join(titles).lower().strip()
                    if any(t == title_text or t in title_text for t in target_lowers):
                        found_count += 1
                except Exception:
                    if any(t in str(props).lower() for t in target_lowers):
                        found_count += 1
            check("Knowledge base 'Market Sentiment Dashboard' page created",
                  found_count >= 1, f"found {found_count} matching pages")
            conn.close()
        except Exception as e:
            check("Notion check", False, str(e))

    check("yf_sentiment_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "yf_sentiment_processor.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
