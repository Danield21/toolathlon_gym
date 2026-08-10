"""Evaluation for yf-portfolio-stress-test-excel-word-gcal."""
import argparse
import os
import sys

import psycopg2


def _to_float(v):
    """Robust numeric conversion for str/int/float/None.

    Strips thousands separators, currency symbols and % so that formatted
    numbers (e.g. "633,555,641", "$260,000", "25%") compare equal to their
    plain numeric equivalents. Returns None when the value cannot be parsed.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "").replace("$", "").replace("¥", "").replace("€", "").replace("%", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


def num_close(a, b, tol=0.5):
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # Fallback only when either side is non-numeric: case-insensitive string equality.
    return str(a).strip().lower() == str(b).strip().lower()


def _sheet_index(wb, sheet_name):
    """Case-insensitive exact sheet name match, with a lenient keyword fallback."""
    key = sheet_name.strip().lower()
    for i, name in enumerate(wb.sheetnames):
        if name.strip().lower() == key:
            return i
    for i, name in enumerate(wb.sheetnames):
        if key.split()[0] in name.strip().lower():
            return i
    return None


def load_workbook_pair(path):
    """Load a workbook twice: cached values (data_only=True) and raw cells
    (data_only=False, which exposes formula strings).

    openpyxl stores only one of these per load, so both passes are required to
    tell a formula cell (with no cached result) apart from an empty cell.
    """
    import openpyxl
    wb_vals = openpyxl.load_workbook(path, data_only=True)
    wb_raw = openpyxl.load_workbook(path, data_only=False)
    return wb_vals, wb_raw


def _pair_rows(wb_vals, wb_raw, sheet_name):
    """Return (values_rows, raw_rows) for a sheet, or (None, None) if absent."""
    idx = _sheet_index(wb_vals, sheet_name)
    if idx is None:
        return None, None
    ws_v = wb_vals.worksheets[idx]
    ws_r = wb_raw.worksheets[idx]
    vals = [[cell.value for cell in row] for row in ws_v.iter_rows()]
    raws = [[cell.value for cell in row] for row in ws_r.iter_rows()]
    return vals, raws


def _is_formula(raw):
    return isinstance(raw, str) and raw.startswith("=")


def _cmp_cell(a_val, a_raw, gt_val, tol, errors, label):
    """Compare one numeric cell.

    Leniently skips formula cells whose cached value is None (Excel has not
    recalculated them). GT is a literal value, so there is no formula side to
    compare against; skipping avoids false FAILs for a formula-using agent.
    An empty (None) cell is still compared and will fail, as expected.
    """
    if _is_formula(a_raw) and a_val is None:
        return
    if not num_close(a_val, gt_val, tol):
        errors.append(f"{label}: {a_val} vs expected {gt_val}")


def check_excel(agent_workspace, groundtruth_workspace):
    errors = []

    agent_path = os.path.join(agent_workspace, "Stress_Test_Report.xlsx")
    gt_path = os.path.join(groundtruth_workspace, "Stress_Test_Report.xlsx")

    if not os.path.exists(agent_path):
        return ["Stress_Test_Report.xlsx not found in agent workspace"]
    if not os.path.exists(gt_path):
        return ["Stress_Test_Report.xlsx not found in groundtruth workspace"]

    try:
        wb_agent_v, wb_agent_r = load_workbook_pair(agent_path)
        wb_gt_v, _ = load_workbook_pair(gt_path)

        # --- Sheet: Portfolio Overview ---
        agent_rows_v, agent_rows_r = _pair_rows(wb_agent_v, wb_agent_r, "Portfolio Overview")
        gt_rows_v, _ = _pair_rows(wb_gt_v, wb_gt_v, "Portfolio Overview")
        if agent_rows_v is None:
            errors.append("Sheet 'Portfolio Overview' not found")
        elif gt_rows_v is None:
            errors.append("Groundtruth 'Portfolio Overview' missing")
        else:
            agent_data_v = [r for r in agent_rows_v[1:] if r and r[0] is not None]
            agent_data_r = [r for r in agent_rows_r[1:] if r and r[0] is not None]
            gt_data_v = [r for r in gt_rows_v[1:] if r and r[0] is not None]

            if len(agent_data_v) < len(gt_data_v):
                errors.append(f"Portfolio Overview: {len(agent_data_v)} rows, expected {len(gt_data_v)}")

            # Check each stock (lookup by symbol, order-independent)
            agent_lookup = {str(rv[0]).strip().upper(): (rv, rr) for rv, rr in zip(agent_data_v, agent_data_r)}
            gt_lookup = {str(r[0]).strip().upper(): r for r in gt_data_v}

            for sym, gt_row in gt_lookup.items():
                if sym not in agent_lookup:
                    errors.append(f"Portfolio Overview: {sym} missing")
                    continue
                a_v, a_r = agent_lookup[sym]
                _cmp_cell(a_v[1], a_r[1], gt_row[1], 0.5, errors, f"{sym} Allocation_Pct")
                _cmp_cell(a_v[3], a_r[3], gt_row[3], 1.0, errors, f"{sym} Current_Price")
                _cmp_cell(a_v[5], a_r[5], gt_row[5], 0.5, errors, f"{sym} Monthly_Volatility_Pct")
                _cmp_cell(a_v[6], a_r[6], gt_row[6], 0.5, errors, f"{sym} Worst_Monthly_Return_Pct")
                _cmp_cell(a_v[7], a_r[7], gt_row[7], 0.1, errors, f"{sym} Sharpe_Ratio")

        # --- Sheet: Stress Scenarios ---
        agent_rows2_v, agent_rows2_r = _pair_rows(wb_agent_v, wb_agent_r, "Stress Scenarios")
        gt_rows2_v, _ = _pair_rows(wb_gt_v, wb_gt_v, "Stress Scenarios")
        if agent_rows2_v is None:
            errors.append("Sheet 'Stress Scenarios' not found")
        elif gt_rows2_v is None:
            errors.append("Groundtruth 'Stress Scenarios' missing")
        else:
            agent_data2_v = [r for r in agent_rows2_v[1:] if r and r[0] is not None]
            agent_data2_r = [r for r in agent_rows2_r[1:] if r and r[0] is not None]
            gt_data2_v = [r for r in gt_rows2_v[1:] if r and r[0] is not None]

            if len(agent_data2_v) < len(gt_data2_v) - 4:  # Allow some tolerance for summary rows
                errors.append(f"Stress Scenarios: {len(agent_data2_v)} rows, expected ~{len(gt_data2_v)}")

            # Check portfolio totals for each scenario
            gt_totals = {}
            for r in gt_data2_v:
                if r[1] and str(r[1]).strip() == "Portfolio_Total":
                    gt_totals[str(r[0]).strip().lower()] = (r[4], r[5])  # Scenario_Value, Scenario_PnL

            agent_totals = {}
            for rv, rr in zip(agent_data2_v, agent_data2_r):
                if rv[1] and str(rv[1]).strip() == "Portfolio_Total":
                    agent_totals[str(rv[0]).strip().lower()] = (rv, rr)

            for sc_name, (gt_val, gt_pnl) in gt_totals.items():
                if sc_name not in agent_totals:
                    errors.append(f"Stress Scenarios: Portfolio_Total row missing for {sc_name}")
                else:
                    a_v, a_r = agent_totals[sc_name]
                    # Magnitude-based tolerance aligned with the per-stock rows: a
                    # correct agent may derive each per-stock Scenario_Value from the
                    # displayed 2-decimal Scenario_Return_Pct (GT's own value column
                    # is built from unrounded returns), so the summed total can differ
                    # by the rounding propagation (observed ~7-50 for this portfolio).
                    # 1% of magnitude is consistent with the per-stock leniency and
                    # still rejects grossly wrong totals / omitted stocks.
                    tol_total_val = max(100.0, abs(_to_float(gt_val) or 0.0) * 0.01)
                    tol_total_pnl = max(100.0, abs(_to_float(gt_pnl) or 0.0) * 0.01)
                    _cmp_cell(a_v[4], a_r[4], gt_val, tol_total_val, errors, f"{sc_name} total value")
                    _cmp_cell(a_v[5], a_r[5], gt_pnl, tol_total_pnl, errors, f"{sc_name} total PnL")

            # Per-stock Scenario_PnL checks for every (Scenario, Stock) combination.
            # Scenario names are normalized to lowercase on both sides, matching the
            # gcal / Risk Summary checks, so a correct agent writing e.g. 'market
            # crash' instead of 'Market Crash' is not falsely FAILed.
            gt_pnl_lookup = {}
            for r in gt_data2_v:
                if r[0] and r[1] and str(r[1]).strip() != "Portfolio_Total":
                    gt_pnl_lookup[(str(r[0]).strip().lower(), str(r[1]).strip().upper())] = (r[3], r[4], r[5])

            agent_pnl_lookup = {}
            for rv, rr in zip(agent_data2_v, agent_data2_r):
                if rv[0] and rv[1] and str(rv[1]).strip() != "Portfolio_Total":
                    agent_pnl_lookup[(str(rv[0]).strip().lower(), str(rv[1]).strip().upper())] = (rv, rr)

            for (sc_name, sym), (gt_ret, gt_val, gt_pnl) in gt_pnl_lookup.items():
                if (sc_name, sym) not in agent_pnl_lookup:
                    errors.append(f"Stress Scenarios: row missing for {sc_name} / {sym}")
                    continue
                a_v, a_r = agent_pnl_lookup[(sc_name, sym)]
                # tolerance: 1% of magnitude or 50 abs, whichever larger
                tol_pnl = max(50.0, abs(_to_float(gt_pnl) or 0.0) * 0.01)
                tol_val = max(50.0, abs(_to_float(gt_val) or 0.0) * 0.01)
                # keep the original leniency: only compare when the cell is filled
                if a_v[5] is not None or _is_formula(a_r[5]):
                    _cmp_cell(a_v[5], a_r[5], gt_pnl, tol_pnl, errors, f"{sc_name}/{sym} Scenario_PnL")
                if a_v[4] is not None or _is_formula(a_r[4]):
                    _cmp_cell(a_v[4], a_r[4], gt_val, tol_val, errors, f"{sc_name}/{sym} Scenario_Value")

        # --- Sheet: Risk Summary ---
        agent_rows3_v, agent_rows3_r = _pair_rows(wb_agent_v, wb_agent_r, "Risk Summary")
        gt_rows3_v, _ = _pair_rows(wb_gt_v, wb_gt_v, "Risk Summary")
        if agent_rows3_v is None:
            errors.append("Sheet 'Risk Summary' not found")
        elif gt_rows3_v is None:
            errors.append("Groundtruth 'Risk Summary' missing")
        else:
            agent_data3_v = [r for r in agent_rows3_v[1:] if r and r[0] is not None]
            agent_data3_r = [r for r in agent_rows3_r[1:] if r and r[0] is not None]
            gt_data3_v = [r for r in gt_rows3_v[1:] if r and r[0] is not None]

            agent_metrics = {}
            for rv, rr in zip(agent_data3_v, agent_data3_r):
                agent_metrics[str(rv[0]).strip().lower()] = (rv[1], rr[1])
            gt_metrics = {str(r[0]).strip().lower(): r[1] for r in gt_data3_v}

            for metric, gt_val in gt_metrics.items():
                if metric not in agent_metrics:
                    errors.append(f"Risk Summary: {metric} missing")
                    continue
                a_val, a_raw = agent_metrics[metric]
                if metric in ("worst_scenario", "best_scenario", "breach_threshold"):
                    if str(a_val).strip().lower() != str(gt_val).strip().lower():
                        errors.append(f"Risk Summary {metric}: '{a_val}' vs expected '{gt_val}'")
                elif metric in ("total_portfolio_value",):
                    _cmp_cell(a_val, a_raw, gt_val, 100, errors, f"Risk Summary {metric}")
                elif metric == "portfolio_var_95":
                    # '5th percentile' is ambiguous across interpolation conventions
                    # (numpy/pandas linear R-7 = GT -32096.75, nearest-rank 2nd-smallest
                    # ≈ -33994, PERCENTILE.EXC ≈ -37529, midpoint-of-bracket ≈ -27671,
                    # R-2..R-9 variants ≈ -35408..-38708). A flat 500 only matches the
                    # R-7 default and false-FAILs every other legitimate 5th-percentile
                    # rule. Use a wide absolute tolerance that covers the full spread of
                    # standard 5th-percentile methods (~[-38.7k, -27.7k]) while still
                    # rejecting a wrong percentile (e.g. 10th ≈ -21k) or a VaR computed
                    # from daily returns (~-10k).
                    _cmp_cell(a_val, a_raw, gt_val, 7000, errors, f"Risk Summary {metric}")
                elif metric in ("worst_scenario_loss", "best_scenario_pnl"):
                    _cmp_cell(a_val, a_raw, gt_val, 500, errors, f"Risk Summary {metric}")
                elif metric in ("max_historical_drawdown_pct", "worst_scenario_loss_pct"):
                    _cmp_cell(a_val, a_raw, gt_val, 1.0, errors, f"Risk Summary {metric}")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    try:
        from docx import Document

        path = os.path.join(agent_workspace, "Risk_Assessment.docx")
        if not os.path.exists(path):
            return ["Risk_Assessment.docx not found"]

        doc = Document(path)
        full_text = "\n".join([p.text for p in doc.paragraphs]).lower()

        required_sections = [
            "executive summary",
            "portfolio composition",
            "stress test results",
            "risk metrics",
            "scenario comparison",
            "recommendation",
        ]
        for section in required_sections:
            if section not in full_text:
                errors.append(f"Word doc missing section: {section}")

        # Check for key content
        if "var" not in full_text and "value at risk" not in full_text:
            errors.append("Word doc missing VaR discussion")
        if "drawdown" not in full_text:
            errors.append("Word doc missing drawdown discussion")

        # Ensure document cites at least one substantial numeric figure
        # (drawdown %, VaR amount, scenario loss). Look for >=2 percent values
        # and at least one large dollar amount.
        import re
        pct_matches = re.findall(r"-?\d+(?:\.\d+)?\s*%", full_text)
        # also look for explicit pct words
        pct_matches += re.findall(r"-?\d+(?:\.\d+)?\s*percent", full_text)
        if len(pct_matches) < 2:
            errors.append(
                f"Word doc has too few percent figures ({len(pct_matches)} < 2)"
            )

        dollar_matches = re.findall(r"\$\s*\d{1,3}(?:[,]\d{3})+(?:\.\d+)?", full_text)
        # also accept plain large numbers (e.g. 260000)
        big_num_matches = re.findall(r"\b\d{5,}\b", full_text)
        if len(dollar_matches) + len(big_num_matches) < 1:
            errors.append("Word doc cites no large dollar amount")

    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=int(os.environ.get("PGPORT", "5432")),
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user=os.environ.get("PGUSER", "eigent"),
            password=os.environ.get("PGPASSWORD", "camel"),
        )
        cur = conn.cursor()

        # Check for 4 stress test review meetings
        cur.execute("""
            SELECT summary, start_datetime::date, description
            FROM gcal.events
            WHERE LOWER(summary) LIKE '%stress test%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) < 4:
            errors.append(f"Expected 4 Stress Test Review meetings, found {len(rows)}")

        # Check scenario names in titles
        expected_scenarios = ["market crash", "sector rotation", "inflation shock", "historical replay"]
        found_scenarios = set()
        for summary, date, desc in rows:
            summary_lower = summary.lower() if summary else ""
            for sc in expected_scenarios:
                if sc in summary_lower:
                    found_scenarios.add(sc)

        for sc in expected_scenarios:
            if sc not in found_scenarios:
                errors.append(f"No calendar event found for scenario: {sc}")

        # Check each expected date appears at least once (set-based, so duplicate
        # events created by redundant sub-agents do not cause false failures).
        if rows:
            import datetime

            expected_dates = [
                datetime.date(2026, 3, 16),
                datetime.date(2026, 3, 23),
                datetime.date(2026, 3, 30),
                datetime.date(2026, 4, 6),
            ]
            actual_dates = set(r[1] for r in rows if r[1] is not None)
            for exp_date in expected_dates:
                if exp_date not in actual_dates:
                    errors.append(f"Missing calendar event on expected date {exp_date}")

    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )
    gt_ws = args.groundtruth_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
