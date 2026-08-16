"""Evaluation script for fetch-wc-inventory-forecast-excel-gcal-email."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-wc-inventory-forecast-excel-gcal-email"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Inventory_Forecast_Report.xlsx")
    check("Inventory_Forecast_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Inventory_Forecast_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Stock_Status sheet exists", "Stock_Status" in wb.sheetnames)
        if "Stock_Status" in wb.sheetnames and gt_wb is not None:
            ws = wb["Stock_Status"]
            gt_ws = gt_wb["Stock_Status"] if "Stock_Status" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Stock_Status has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Stock_Status', ['Product', 'Current_Stock', 'Total_Sales', 'Daily_Rate', 'Days_Remaining', 'Needs_Restock'],
                          len(gt_rows) if gt_ws is not None else 8)
            # Per-product row checks (match by 1st 30 chars of product name to handle truncation)
            if gt_ws is not None:
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                header_map = {h: i for i, h in enumerate(headers)}
                p_idx_gt = gt_headers.index("product") if "product" in gt_headers else 0
                p_idx_a = header_map.get("product", 0)
                gt_by_p = {str(r[p_idx_gt]).strip().lower()[:30]: r for r in gt_rows if r and r[p_idx_gt]}
                agent_by_p = {str(r[p_idx_a]).strip().lower()[:30]: r for r in data_rows if r and len(r) > p_idx_a and r[p_idx_a]}
                for p_key, gt_row in gt_by_p.items():
                    found = p_key in agent_by_p
                    check(f"Stock_Status product '{p_key[:25]}...' present", found)
                    if found:
                        agent_row = agent_by_p[p_key]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row) or gt_h == "product":
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                tol = max(0.5, abs(gf) * 0.10)
                                check(f"Stock_Status '{p_key[:20]}' {gt_h} ~{gf}",
                                      abs(gf - af) <= tol, f"got {av}")
                            elif gv is not None and av is not None:
                                check(f"Stock_Status '{p_key[:20]}' {gt_h} == {gv}",
                                      str(gv).strip().lower() == str(av).strip().lower(),
                                      f"got {av}")

        check("Supplier_Info sheet exists", "Supplier_Info" in wb.sheetnames)
        if "Supplier_Info" in wb.sheetnames and gt_wb is not None:
            ws = wb["Supplier_Info"]
            gt_ws = gt_wb["Supplier_Info"] if "Supplier_Info" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Supplier_Info has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Supplier_Info', ['Supplier', 'Lead_Time_Days', 'Min_Order_Qty', 'Reliability_Score'],
                          len(gt_rows) if gt_ws is not None else 4)
            if gt_ws is not None:
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                header_map = {h: i for i, h in enumerate(headers)}
                s_idx_gt = gt_headers.index("supplier") if "supplier" in gt_headers else 0
                s_idx_a = header_map.get("supplier", 0)
                gt_by_s = {str(r[s_idx_gt]).strip().lower(): r for r in gt_rows if r and r[s_idx_gt]}
                agent_by_s = {str(r[s_idx_a]).strip().lower(): r for r in data_rows if r and len(r) > s_idx_a and r[s_idx_a]}
                for s_key, gt_row in gt_by_s.items():
                    found = s_key in agent_by_s
                    check(f"Supplier_Info supplier '{s_key}' present", found)

        check("Restock_Summary sheet exists", "Restock_Summary" in wb.sheetnames)
        if "Restock_Summary" in wb.sheetnames and gt_wb is not None:
            ws = wb["Restock_Summary"]
            gt_ws = gt_wb["Restock_Summary"] if "Restock_Summary" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Restock_Summary has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Restock_Summary', ['Metric', 'Value'], len(gt_rows) if gt_ws is not None else 4)
            if gt_ws is not None:
                gt_metric_map = {str(r[0]).strip().lower(): r[1] for r in gt_rows if r and r[0]}
                agent_metric_map = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for gt_m, gt_v in gt_metric_map.items():
                    found = gt_m in agent_metric_map
                    check(f"Restock_Summary has '{gt_m}'", found)
                    if found:
                        av = agent_metric_map[gt_m]
                        gf = safe_float(gt_v)
                        af = safe_float(av)
                        if gf is not None and af is not None:
                            tol = max(0.5, abs(gf) * 0.10)
                            check(f"Restock_Summary '{gt_m}' ~{gf}",
                                  abs(gf - af) <= tol, f"got {av}")

        # Calendar: 'Inventory Review Meeting' on 2026-03-12 09:00-10:00 UTC, description lists products
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("""SELECT summary, description, start_datetime, end_datetime FROM gcal.events
                           WHERE summary ILIKE %s""",
                        ('%Inventory Review Meeting%',))
            events = cur.fetchall()
            check("'Inventory Review Meeting' event exists",
                  len(events) >= 1, f"found {len(events)} matching events")
            date_match = False
            start_match = False
            end_match = False
            desc_with_products = False
            for s, d, sd, ed in events:
                # node-pg/psycopg2 return timezone-aware datetimes rendered
                # in the SESSION TZ (Asia/Shanghai in the container): a
                # 09:00Z event prints as "17:00+08:00". Normalize to UTC
                # before matching wall-clock, else every correct event
                # fails the time check (case-study 2026-08-16 round-3).
                def utc_str(x):
                    if x is None:
                        return ""
                    if hasattr(x, "astimezone"):
                        import datetime as _dt
                        if x.tzinfo is None:
                            x = x.replace(tzinfo=_dt.timezone.utc)
                        x = x.astimezone(_dt.timezone.utc)
                    return str(x)
                sd_str = utc_str(sd)
                ed_str = utc_str(ed)
                if "2026-03-12" in sd_str:
                    date_match = True
                if "09:00" in sd_str:
                    start_match = True
                if "10:00" in ed_str:
                    end_match = True
                if d and len(d) > 30:
                    desc_with_products = True
            check("'Inventory Review Meeting' date 2026-03-12", date_match)
            check("'Inventory Review Meeting' start 09:00 UTC", start_match)
            check("'Inventory Review Meeting' end 10:00 UTC", end_match)
            check("'Inventory Review Meeting' description lists products", desc_with_products)
            conn.close()
        except Exception as e:
            check("Calendar DB checks", False, str(e))

        # Email: subject 'Urgent Restock Alert' AND recipient procurement@company.com
        # Use a SEPARATE try/except so a calendar failure doesn't mask the email check.
        # NOTE: email.messages.to_addr is a JSONB column; ILIKE requires ::text cast.
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("""SELECT subject, to_addr FROM email.messages
                           WHERE to_addr::text ILIKE %s AND subject ILIKE %s""",
                        ('%procurement@company.com%', '%Urgent Restock Alert%'))
            emails = cur.fetchall()
            check("Email to procurement@company.com with subject 'Urgent Restock Alert' sent",
                  len(emails) >= 1, f"found {len(emails)} matching emails")
            conn.close()
        except Exception as e:
            check("Email DB checks", False, str(e))

    check("inventory_forecaster.py exists",
          os.path.exists(os.path.join(agent_workspace, "inventory_forecaster.py")))
    # Required JSON intermediate / output files per task.md
    for json_name in ("supplier_data.json", "product_stock.json", "restock_plan.json"):
        jpath = os.path.join(agent_workspace, json_name)
        check(f"{json_name} exists", os.path.exists(jpath))
        if os.path.exists(jpath):
            try:
                with open(jpath, "r") as _f:
                    _data = json.load(_f)
                check(f"{json_name} is valid JSON and non-empty",
                      bool(_data),
                      f"empty or null content")
            except Exception as _e:
                check(f"{json_name} is valid JSON and non-empty", False, str(_e))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
