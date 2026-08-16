"""Evaluation script for sf-notion-project-tracker-excel-gcal-email."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Check Project_Portfolio.xlsx
    excel_path = os.path.join(agent_workspace, "Project_Portfolio.xlsx")
    check("Project_Portfolio.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Project_Portfolio.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        if gt_wb:
            for sheet_name in gt_wb.sheetnames:
                check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    gt_ws = gt_wb[sheet_name]
                    # Check headers
                    gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                    for h in gt_headers:
                        if h:
                            check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                    # Check row count
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                    # Proposed_Projects must have >= 5 per task spec; otherwise mirror GT row count
                    if sheet_name.strip().lower() == "proposed_projects":
                        min_rows = 5
                    else:
                        min_rows = max(1, len(gt_rows) - 2)
                    check(f"{sheet_name} has >= {min_rows} data rows", len(data_rows) >= min_rows, f"got {len(data_rows)}")

                    # Cell value comparison against groundtruth - iterate ALL rows
                    header_map = {h: i for i, h in enumerate(headers)}
                    # Build agent lookup keyed by first column (for name-based join)
                    agent_lookup = {}
                    if data_rows and data_rows[0]:
                        for r in data_rows:
                            if r and r[0] is not None:
                                agent_lookup[str(r[0]).strip().lower()] = r
                    # Columns where strict text equality is expected
                    strict_text_cols = {"department", "priority", "metric"}
                    # Columns where numeric comparison applies (tighter tol)
                    for gt_row in gt_rows:
                        if not gt_row or gt_row[0] is None:
                            continue
                        key = str(gt_row[0]).strip().lower()
                        agent_row = agent_lookup.get(key)
                        if agent_row is None:
                            # Row missing from agent output
                            check(f"{sheet_name} row '{gt_row[0]}' present", False,
                                  f"row not found in agent output")
                            continue
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row):
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                # Tightened: 1% rel (large numbers) + 0.5 abs (small numbers).
                                # Total_Available_Staff is intentionally a synthetic round
                                # aggregate and may diverge from a per-department sum, so
                                # accept any positive integer.
                                if gt_h == "total_available_staff":
                                    check(f"{sheet_name} '{key}' {gt_h} positive int",
                                          af >= 1, f"got {af}")
                                else:
                                    tol = max(0.5, abs(gf) * 0.01)
                                    check(f"{sheet_name} '{key}' {gt_h} ~{gf:.1f}",
                                          abs(gf - af) <= tol, f"got {af}")
                            elif gv is not None and av is not None:
                                gs = str(gv).strip().lower()
                                avs = str(av).strip().lower()
                                if gs:
                                    if gt_h in strict_text_cols:
                                        # Strict equality for canonical fields
                                        check(f"{sheet_name} '{key}' {gt_h} text strict",
                                              gs == avs,
                                              f"expected '{gs}', got '{avs}'")
                                    # else: skip subjective string fields (Top_Performer, project names)

    # Check Python script exists (terminal usage)
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")

    # Database checks
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Email: must be to all-managers@company.com with subject mentioning Q2 Portfolio
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages "
                    "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
                    "AND to_addr::text ILIKE '%all-managers@company.com%'")
        email_rows = cur.fetchall()
        # Match Q2 Portfolio subject
        good_email = None
        for subj, to_addr, body in email_rows:
            subj_l = (subj or "").lower()
            if "q2" in subj_l and ("portfolio" in subj_l or "project" in subj_l):
                good_email = (subj, to_addr, body)
                break
        check("Email to all-managers@company.com with Q2 Project Portfolio subject",
              good_email is not None,
              f"got {len(email_rows)} matching to_addr emails")
        if good_email:
            body_text = (good_email[2] or "").lower()
            # Body must reference projects, leads, AND kickoff dates per task spec
            # Support multiple date formats: "March 19", "Mar 19", "19 March", "19 Mar",
            # "03/19", "03-19", "3/19", "3-19", "2026-03-19", "2026/03/19".
            day_tokens_with_month = []
            for d in ("19", "20", "21"):
                day_tokens_with_month.extend([
                    f"march {d}", f"mar {d}",
                    f"{d} march", f"{d} mar",
                    f"03/{d}", f"03-{d}",
                    f"3/{d}", f"3-{d}",
                    f"2026-03-{d}", f"2026/03/{d}",
                ])
            has_kickoff_date = any(tok in body_text for tok in day_tokens_with_month)
            has_lead_or_project = (
                "lead" in body_text or "project" in body_text or "kickoff" in body_text
            )
            check("Email body lists projects, leads, and kickoff dates",
                  good_email[2] is not None and len(good_email[2].strip()) >= 80
                  and has_lead_or_project and has_kickoff_date,
                  f"body too short or missing required content: {(good_email[2] or '')[:80]}")
        # Calendar: 3 events on March 19/20/21 2026 from 10:00-11:00 UTC
        # Use timezone-aware comparison via SQL conversion to UTC
        cur.execute(
            "SELECT summary, start_datetime AT TIME ZONE 'UTC' as st_utc, "
            "end_datetime AT TIME ZONE 'UTC' as ed_utc FROM gcal.events "
            "WHERE (start_datetime AT TIME ZONE 'UTC')::date "
            "IN ('2026-03-19','2026-03-20','2026-03-21')")
        event_rows = cur.fetchall()
        # Filter for project-relevant events at the right time (10:00 UTC)
        seen_dates = set()
        for s, sd_utc, ed_utc in event_rows:
            if not sd_utc:
                continue
            d = sd_utc.date() if hasattr(sd_utc, 'date') else None
            hour = sd_utc.hour if hasattr(sd_utc, 'hour') else None
            if d is None or hour != 10:
                continue
            # Subject suggests project kickoff
            sl = (s or "").lower()
            if "project" in sl or "kickoff" in sl or "kick-off" in sl or "q2" in sl:
                seen_dates.add(d.isoformat())
        check("3 project kickoff calendar events on Mar 19/20/21 at 10:00 UTC",
              len(seen_dates) >= 3 and
              "2026-03-19" in seen_dates and
              "2026-03-20" in seen_dates and
              "2026-03-21" in seen_dates,
              f"got dates: {sorted(seen_dates)}")
        # Notion: 'Department Project Dashboard' page
        cur.execute("SELECT id, properties FROM notion.pages "
                    "WHERE properties::text ILIKE '%Department Project Dashboard%' "
                    "AND archived = false")
        notion_rows = cur.fetchall()
        check("Notion 'Department Project Dashboard' page exists",
              len(notion_rows) >= 1, f"got {len(notion_rows)} matching pages")
        if notion_rows:
            page_ids = [r[0] for r in notion_rows]
            cur.execute(
                "SELECT block_data FROM notion.blocks WHERE parent_id = ANY(%s)",
                (page_ids,))
            blob = " ".join(str(r[0] or "").lower() for r in cur.fetchall())
            # Page must reference Q2 2026 Project Portfolio AND have content
            # describing project-related themes from the task spec.
            check("Notion page mentions Q2 2026 Project Portfolio heading",
                  ("q2 2026" in blob or "q2" in blob)
                  and "portfolio" in blob and "project" in blob,
                  "missing Q2 2026 Project Portfolio content")
            # At least 3 proposed projects with leaders mentioned
            has_lead_words = ("lead" in blob or "leader" in blob)
            check("Notion page mentions project leaders / objectives",
                  has_lead_words
                  and ("objective" in blob or "timeline" in blob or "team" in blob),
                  "missing leaders / objectives / team / timeline content")
        # Reverse verification: noise emails should not be in Sent folder
        cur.execute("SELECT COUNT(*) FROM email.messages WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) AND subject ILIKE '%newsletter%'")
        noise_sent = cur.fetchone()[0]
        check("No noise emails in Sent folder", noise_sent == 0, f"found {noise_sent} noise emails in Sent")
        # Reverse verification: noise events should not be deleted
        cur.execute("SELECT COUNT(*) FROM gcal.events WHERE summary ILIKE '%standup%' OR summary ILIKE '%lunch%'")
        noise_events = cur.fetchone()[0]
        check("Noise events exist (not deleted by agent)", noise_events >= 1, f"noise events: {noise_events}")
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()