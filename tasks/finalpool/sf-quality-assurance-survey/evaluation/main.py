"""
Evaluation script for sf-quality-assurance-survey task.

Checks:
1. Excel file QA_Assessment.xlsx with 3 sheets
2. Google Form created with correct structure

Expected values are computed directly from the (immutable) PostgreSQL seed at
eval time, so the expected numbers are always derivable.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

# DB connection reads the same env vars as preprocess/main.py (R1).
DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0
BLOCKING_FAIL_COUNT = 0
# True when the harness runs the eval against the groundtruth workspace itself
# (agent_workspace == groundtruth_workspace). The Google Form is a solver
# side-effect that does not exist in the groundtruth DB snapshot, so the
# existence check is skipped in that mode instead of falsely failing the GT.
IS_GT_SELF_TEST = False


def check(name, condition, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, BLOCKING_FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if not runtime_only:
            BLOCKING_FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        suffix = " (runtime-only)" if runtime_only else ""
        print(f"  [FAIL] {name}{suffix}{detail_str}")


_NUM_TOKEN = re.compile(r'-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?(?:[eE][+-]?\d+)?')


def _to_float(v):
    """Robustly parse a cell value into a float (R2/R3 + round-2 hardening).

    Handles int/float/bool and strings containing thousands separators,
    currency symbols, percent signs, surrounding whitespace, and trailing
    unit words (e.g. '15.03 hours', '6,804 tickets', '92.4%'). Unparsable
    values (including raw formula strings such as "=B2/C2" and non-numeric
    text) return None.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return None
    m = _NUM_TOKEN.search(s)
    if m is None:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    """Numeric closeness.

    - Both sides parse to numbers: |a-b| <= max(abs_tol, |b|*rel_tol).
    - Both sides are genuinely blank: accept (GT also omits the value).
    - Agent side is blank/unresolvable while GT expects a number: FAIL. A
      blank cell (or an Excel formula with no cached value) must never
      silently pass a core numeric check -- the arithmetic has to be real.
    - Otherwise fall back to a case-insensitive string comparison.
    """
    af, bf = _to_float(a), _to_float(b)
    if af is not None and bf is not None:
        return abs(af - bf) <= max(abs_tol, abs(bf) * rel_tol)
    if a is None and b is None:
        return True
    if a is None:
        return False  # agent left the cell blank -> fail the check
    if b is None:
        return False
    # Agent wrote a non-numeric literal: only pass if it equals the target.
    return str(a).strip().lower() == str(b).strip().lower()


def _norm(s):
    if s is None:
        return ""
    return str(s).strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return _norm(a) == _norm(b)


def compute_expected_values():
    """Query PostgreSQL to compute expected values."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Agent Scorecard: distinct ticket reporters are the agents.
    # Targets: satisfaction >= 4.0, resolution >= 95%, response_time compared to 8 (medium target).
    cur.execute("""
        SELECT
            "REPORTER" as agent,
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp,
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) as avg_sat,
            ROUND(COUNT(CASE WHEN "STATUS"='Resolved' THEN 1 END)*100.0/COUNT(*)::numeric, 1) as resolution_rate
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "REPORTER"
        ORDER BY "REPORTER"
    """)
    agent_rows = cur.fetchall()
    agent_scorecard = []
    for row in agent_rows:
        agent_name, avg_resp, avg_sat, res_rate = row
        below = []
        if float(avg_resp) > 8.0:
            below.append("Response Time")
        if float(avg_sat) < 4.0:
            below.append("Satisfaction")
        if float(res_rate) < 95.0:
            below.append("Resolution Rate")
        agent_scorecard.append({
            "Agent": agent_name,
            "Avg_Response_Time": float(avg_resp),
            "Avg_Satisfaction": float(avg_sat),
            "Resolution_Rate": float(res_rate),
            "Below_Target_Areas": ", ".join(below) if below else "None",
        })

    # Issue Type Analysis
    cur.execute("""
        SELECT
            "ISSUE_TYPE",
            COUNT(*) as cnt,
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp,
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) as avg_sat
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE"
        ORDER BY "ISSUE_TYPE"
    """)
    issue_rows = cur.fetchall()
    issue_type_analysis = []
    for row in issue_rows:
        issue_type_analysis.append({
            "Issue_Type": row[0],
            "Ticket_Count": int(row[1]),
            "Avg_Response_Time": float(row[2]),
            "Avg_Satisfaction": float(row[3]),
        })

    # Summary
    cur.execute("""
        SELECT
            COUNT(*),
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2),
            ROUND(COUNT(CASE WHEN "STATUS"='Resolved' THEN 1 END)*100.0/COUNT(*)::numeric, 1)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
    """)
    s = cur.fetchone()
    total_tickets = int(s[0])
    overall_avg_resp = float(s[1])
    overall_avg_sat = float(s[2])
    overall_res_rate = float(s[3])

    agents_below_sat = sum(1 for a in agent_scorecard if a["Avg_Satisfaction"] < 4.0)
    agents_below_resp = sum(1 for a in agent_scorecard if a["Avg_Response_Time"] > 8.0)

    summary = {
        "Total_Tickets": total_tickets,
        "Overall_Avg_Response_Time": overall_avg_resp,
        "Overall_Avg_Satisfaction": overall_avg_sat,
        "Overall_Resolution_Rate": overall_res_rate,
        "Agents_Below_Satisfaction_Target": agents_below_sat,
        "Agents_Below_Response_Target": agents_below_resp,
    }

    cur.close()
    conn.close()

    return {
        "agent_scorecard": agent_scorecard,
        "issue_type_analysis": issue_type_analysis,
        "summary": summary,
    }


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


# --- Excel reading helpers (R2 / R10) ------------------------------------

def _resolve_cell(ws_struct, ws_vals, row, col):
    """Effective value of a cell.

    A formula cell (value starts with '=') is resolved to its cached value from
    the data_only=True workbook; None if no cached value was stored (the agent
    did not recalc before saving). Literal cells are returned unchanged.
    """
    c1 = ws_struct.cell(row=row, column=col)
    v = c1.value
    if isinstance(v, str) and v.startswith("="):
        return ws_vals.cell(row=row, column=col).value
    return v


def _normalize_header(v):
    return _norm(v)


# First-cell header values that mark the header row of each sheet (R10).
_HEADER_WORDS = {
    "agent": {"agent", "agentname", "name"},
    "issue": {"issuetype", "issue", "issuetypeid"},
}


def _find_header_row(ws, kind):
    """1-based row index of the header row whose first cell matches `kind`.

    Returns None when no header row is detected (caller then assumes row 1).
    """
    words = _HEADER_WORDS.get(kind, set())
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            continue
        if _normalize_header(v) in words:
            return r
    return None


def _read_data_rows(ws_struct, ws_vals, kind, pad_to=5):
    """Data rows (tuples) below the header row, formula-resolved.

    Fully blank rows are dropped, and every row is padded to `pad_to` columns
    so downstream index access (matched[1]..matched[4]) is safe even when the
    agent wrote fewer columns.
    """
    hdr = _find_header_row(ws_struct, kind)
    if hdr is None:
        hdr = 1  # assume row 1 is the header
    rows = []
    for r in range(hdr + 1, ws_struct.max_row + 1):
        row = tuple(_resolve_cell(ws_struct, ws_vals, r, c)
                    for c in range(1, ws_struct.max_column + 1))
        if any(v is not None and str(v).strip() != "" for v in row):
            if len(row) < pad_to:
                row = tuple(list(row) + [None] * (pad_to - len(row)))
            rows.append(row)
    return rows


def _find_col(ws_struct, hdr_row, keywords, default):
    """1-based column index whose normalized header contains any keyword.

    Falls back to `default` when no column matches (R10 header anchoring).
    """
    for c in range(1, ws_struct.max_column + 1):
        h = _norm(ws_struct.cell(row=hdr_row, column=c).value)
        if any(k in h for k in keywords):
            return c
    return default


def _cell(row, idx):
    return row[idx] if 0 <= idx < len(row) else None


def _sheet_pair(wb_struct, wb_vals, name):
    for i, sn in enumerate(wb_struct.sheetnames):
        if str_match(sn, name):
            return wb_struct[sn], wb_vals[wb_vals.sheetnames[i]]
    return None, None


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "QA_Assessment.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        # data_only=False keeps formulas so we can detect them; data_only=True
        # provides cached values for formula cells (R2).
        wb_struct = openpyxl.load_workbook(agent_file, data_only=False)
        wb_vals = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return
    check("Excel file readable", True)

    # Check sheets exist
    for sn in ["Agent Scorecard", "Issue Type Analysis", "Summary"]:
        found = any(str_match(s, sn) for s in wb_struct.sheetnames)
        check(f"Sheet '{sn}' exists", found, f"Found: {wb_struct.sheetnames}")

    # --- Agent Scorecard ---
    print("\n--- Agent Scorecard ---")
    ws, wsv = _sheet_pair(wb_struct, wb_vals, "Agent Scorecard")
    if ws:
        hdr_row = _find_header_row(ws, "agent") or 1
        c_name = _find_col(ws, hdr_row, ("agent", "name", "reporter"), 1)
        c_resp = _find_col(ws, hdr_row, ("response",), 2)
        c_sat = _find_col(ws, hdr_row, ("satisfaction", "csat"), 3)
        c_res = _find_col(ws, hdr_row, ("resolution",), 4)
        c_below = _find_col(ws, hdr_row, ("below",), 5)
        pad_to = max(6, c_name, c_resp, c_sat, c_res, c_below)
        by_name = {}
        for row in _read_data_rows(ws, wsv, "agent", pad_to=pad_to):
            nm = _cell(row, c_name - 1)
            if nm is not None and str(nm).strip() != "":
                by_name[str(nm).strip().lower()] = row
        exp = expected["agent_scorecard"]
        check("Agent Scorecard row count", len(by_name) >= len(exp),
              f"Expected at least {len(exp)}, got {len(by_name)}")

        for e_row in exp:
            agent = e_row["Agent"]
            matched = by_name.get(str(agent).strip().lower())
            if matched:
                check(f"{agent} Avg_Response_Time",
                      num_close(_cell(matched, c_resp - 1),
                                e_row["Avg_Response_Time"], 0.3),
                      f"Expected {e_row['Avg_Response_Time']}, got {_cell(matched, c_resp - 1)}")
                check(f"{agent} Avg_Satisfaction",
                      num_close(_cell(matched, c_sat - 1),
                                e_row["Avg_Satisfaction"], 0.1),
                      f"Expected {e_row['Avg_Satisfaction']}, got {_cell(matched, c_sat - 1)}")
                check(f"{agent} Resolution_Rate",
                      num_close(_cell(matched, c_res - 1),
                                e_row["Resolution_Rate"], 0.5),
                      f"Expected {e_row['Resolution_Rate']}, got {_cell(matched, c_res - 1)}")
                # Below_Target_Areas: every expected area must be present; a
                # blank cell is a FAIL, not a silent skip (review issue 5).
                expected_areas = [
                    x.strip().lower()
                    for x in e_row["Below_Target_Areas"].lower().split(",")
                    if x.strip() and x.strip().lower() != "none"
                ]
                agent_below = str(_cell(matched, c_below - 1) or "").lower()
                if expected_areas:
                    all_present = all(area in agent_below for area in expected_areas)
                    check(f"{agent} below areas '{e_row['Below_Target_Areas']}'",
                          all_present,
                          f"Expected '{e_row['Below_Target_Areas']}' in '{_cell(matched, c_below - 1)}'")
                else:
                    no_areas = (agent_below == "" or agent_below in
                                ("none", "n/a", "na", "no areas",
                                 "no below-target areas"))
                    check(f"{agent} below areas (expect none)",
                          no_areas,
                          f"Expected none, got '{_cell(matched, c_below - 1)}'")
            else:
                check(f"Agent '{agent}' found", False, "Not in agent output")

    # --- Issue Type Analysis ---
    print("\n--- Issue Type Analysis ---")
    ws, wsv = _sheet_pair(wb_struct, wb_vals, "Issue Type Analysis")
    if ws:
        hdr_row = _find_header_row(ws, "issue") or 1
        c_type = _find_col(ws, hdr_row, ("issue", "type"), 1)
        c_count = _find_col(ws, hdr_row, ("count", "ticket", "number"), 2)
        c_resp = _find_col(ws, hdr_row, ("response",), 3)
        c_sat = _find_col(ws, hdr_row, ("satisfaction", "csat"), 4)
        pad_to = max(5, c_type, c_count, c_resp, c_sat)
        by_type = {}
        for row in _read_data_rows(ws, wsv, "issue", pad_to=pad_to):
            t = _cell(row, c_type - 1)
            if t is not None and str(t).strip() != "":
                by_type[str(t).strip().lower()] = row
        exp = expected["issue_type_analysis"]
        check("Issue Type row count", len(by_type) >= len(exp),
              f"Expected at least {len(exp)}, got {len(by_type)}")

        for e_row in exp:
            it = e_row["Issue_Type"]
            matched = by_type.get(str(it).strip().lower())
            if matched:
                # Ticket_Count should be exact (integer)
                check(f"{it} Ticket_Count",
                      num_close(_cell(matched, c_count - 1),
                                e_row["Ticket_Count"], 0),
                      f"Expected {e_row['Ticket_Count']}, got {_cell(matched, c_count - 1)}")
                check(f"{it} Avg_Response_Time",
                      num_close(_cell(matched, c_resp - 1),
                                e_row["Avg_Response_Time"], 0.1),
                      f"Expected {e_row['Avg_Response_Time']}, got {_cell(matched, c_resp - 1)}")
                check(f"{it} Avg_Satisfaction",
                      num_close(_cell(matched, c_sat - 1),
                                e_row["Avg_Satisfaction"], 0.1),
                      f"Expected {e_row['Avg_Satisfaction']}, got {_cell(matched, c_sat - 1)}")
            else:
                check(f"Issue Type '{it}' found", False, "Not in output")

    # --- Summary ---
    print("\n--- Summary ---")
    ws, wsv = _sheet_pair(wb_struct, wb_vals, "Summary")
    if ws:
        data = {}
        for r in range(1, ws.max_row + 1):
            k = _resolve_cell(ws, wsv, r, 1)
            v = _resolve_cell(ws, wsv, r, 2)
            if k is not None and str(k).strip() != "":
                data[str(k).strip().lower().replace(" ", "_")] = v

        for key, gt_val in expected["summary"].items():
            key_lower = key.lower()
            agent_val = data.get(key_lower)
            if agent_val is None:
                for ak, av in data.items():
                    if key_lower.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                check(f"Summary '{key}'",
                      num_close(agent_val, gt_val, 1.0),
                      f"Expected {gt_val}, got {agent_val}")
            else:
                check(f"Summary '{key}'",
                      str_match(agent_val, gt_val),
                      f"Expected '{gt_val}', got '{agent_val}'")


# --- Google Form helpers (R5 / R6) ---------------------------------------

def _extract_options(cfg):
    """Extract option strings from a question config.

    Handles the MCP's stored format {"type": "RADIO", "options": [{"value": "A"}, ...]}
    as well as plain string lists and dicts keyed by label/text.
    """
    if not isinstance(cfg, dict):
        return []
    for key in ("choices", "options", "values", "items"):
        val = cfg.get(key)
        if isinstance(val, list):
            out = []
            for item in val:
                if isinstance(item, dict):
                    v = item.get("value")
                    if v is None:
                        v = item.get("label", item.get("text"))
                    if v is None:
                        for _k, _vv in item.items():
                            if isinstance(_vv, str) and _vv.strip():
                                v = _vv
                                break
                    if v is not None:
                        out.append(str(v).strip())
                else:
                    out.append(str(item).strip())
            return out
    return []


def check_form():
    print("\n=== Checking Google Form ===")
    # GT self-test: the form is a solver-created side-effect absent from the
    # groundtruth DB snapshot, so skip rather than falsely fail the GT run.
    if IS_GT_SELF_TEST:
        print("  [SKIP] Google Form check: GT self-test (side-effect not present in GT snapshot)")
        return
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        # Preserve graceful degradation: a DB outage must not turn a correct
        # Excel submission into a FAIL, so the form check is skipped, not failed.
        print(f"  [WARN] Google Form check skipped: DB unavailable ({e})")
        return
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE LOWER(title) LIKE '%qa%' OR LOWER(title) LIKE '%quality%' OR LOWER(title) LIKE '%self assess%'
        ORDER BY created_at DESC LIMIT 1
    """)
    form_row = cur.fetchone()
    # Form existence is BLOCKING: the task explicitly requires creating the
    # "QA Self Assessment Survey", so a solver that skips it must not pass.
    check("Google Form exists", form_row is not None,
          f"Found forms with QA/quality/self-assess title: {form_row}")
    if not form_row:
        cur.close()
        conn.close()
        return

    form_id = form_row[0]

    cur.execute("""
        SELECT title, question_type, required, config, position
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()
    # Question count is BLOCKING: task.md fixes the form at exactly five questions.
    check("Form has 5 questions", len(questions) == 5,
          f"Found {len(questions)}")

    # The MCP writes question_type as 'textQuestion' / 'choiceQuestion'
    # (schema may also store 'TEXT' / 'RADIO' / 'PARAGRAPH'). We accept the
    # real spellings and common variants; we never discriminate on dropdown /
    # scale / linear / paragraph / checkbox type names (R6).
    def _is_text_q(qt):
        ql = (qt or "").lower()
        return "text" in ql or "paragraph" in ql

    def _is_choice_q(qt):
        ql = (qt or "").lower()
        return "choice" in ql or "radio" in ql or "select" in ql or "checkbox" in ql

    if len(questions) >= 5:
        # Q1: name (text, required)
        check("Q1 is text type", _is_text_q(questions[0][1]),
              f"Got {questions[0][1]}", runtime_only=True)
        check("Q1 is required", questions[0][2] is True, runtime_only=True)

        # Q2: team (choice, required)
        check("Q2 is choice type", _is_choice_q(questions[1][1]),
              f"Got {questions[1][1]}", runtime_only=True)
        check("Q2 is required", questions[1][2] is True, runtime_only=True)

        # Q3: response time rating (choice, required)
        check("Q3 is choice type", _is_choice_q(questions[2][1]),
              f"Got {questions[2][1]}", runtime_only=True)
        check("Q3 is required", questions[2][2] is True, runtime_only=True)

        # Q4: challenge (text, required)
        check("Q4 is text type", _is_text_q(questions[3][1]),
              f"Got {questions[3][1]}", runtime_only=True)
        check("Q4 is required", questions[3][2] is True, runtime_only=True)

        # Q5: suggestions (text, not required)
        check("Q5 is text type", _is_text_q(questions[4][1]),
              f"Got {questions[4][1]}", runtime_only=True)
        check("Q5 is not required", questions[4][2] is False or questions[4][2] is None,
              f"Got required={questions[4][2]}", runtime_only=True)

        # Q3 (the 3rd question, position-independent) options per task.md:
        # must include "Exceeds Target", "Meets Target", "Below Target",
        # "Needs Improvement". Once the agent has created a form this is
        # BLOCKING (we hold them to the spec), but the option parsing is robust
        # to the stored JSONB format.
        q3_opts = [c.lower() for c in _extract_options(questions[2][3])]
        if q3_opts:
            required_q3_keywords = ["exceeds target", "meets target",
                                    "below target", "needs improvement"]
            joined = " | ".join(q3_opts)
            all_present = all(kw in joined for kw in required_q3_keywords)
            check("Q3 has the 4 required rating options "
                  "(Exceeds/Meets/Below Target, Needs Improvement)",
                  all_present,
                  f"Opts: {q3_opts}")

        q2_opts = [c.lower() for c in _extract_options(questions[1][3])]
        if q2_opts:
            check("Q2 has >= 2 team choices", len(q2_opts) >= 2,
                  f"Opts: {q2_opts}", runtime_only=True)

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    global IS_GT_SELF_TEST
    try:
        _gt = os.path.realpath(args.groundtruth_workspace) if args.groundtruth_workspace else ""
        _ag = os.path.realpath(args.agent_workspace) if args.agent_workspace else ""
        IS_GT_SELF_TEST = bool(_gt) and bool(_ag) and _gt == _ag
    except Exception:
        IS_GT_SELF_TEST = False

    print("=== Computing Expected Values ===")
    try:
        expected = compute_expected_values()
        print(f"  Agent scorecard: {len(expected['agent_scorecard'])} agents")
        print(f"  Issue types: {len(expected['issue_type_analysis'])} types")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    check_excel(args.agent_workspace, expected)
    check_form()

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    # Gate on BLOCKING_FAIL_COUNT
    success = BLOCKING_FAIL_COUNT == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT} (blocking_fail={BLOCKING_FAIL_COUNT})")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
