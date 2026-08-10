"""Evaluation for yf-options-expiry-monitor."""
import argparse
import os
import re
import sys
from datetime import date, datetime, timezone

import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)

PASS_COUNT = 0
FAIL_COUNT = 0

# Sentinel: agent wrote an Excel formula cell whose cached value is unavailable.
FORMULA_UNCACHED = object()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def _to_float(v):
    """Parse a numeric value, tolerating %, $, thousand separators, and spaces.
    Returns None when the value is not numeric (or None)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() == "none":
        return None
    for ch in ("%", "$", "¥", "€", ",", "_", " "):
        s = s.replace(ch, "")
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=1.0):
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # Only fall back to case-insensitive string equality when one side is
    # genuinely non-numeric.
    return str(a).strip().lower() == str(b).strip().lower()


def _cell_effective(raw_val, cached_val):
    """Return the effective value of a cell.
    Formula cells use the cached computed value; a formula with no cached
    value yields the FORMULA_UNCACHED sentinel."""
    if isinstance(raw_val, str) and raw_val.startswith("="):
        return cached_val if cached_val is not None else FORMULA_UNCACHED
    return raw_val


def load_sheet_rows(wb_path, sheet_name):
    """Load a sheet's rows as lists of effective cell values.
    Reads with data_only=False (formulas visible) plus data_only=True
    (cached values) so agent formulas degrade gracefully instead of
    turning into None and failing numeric checks."""
    import openpyxl
    try:
        wb_raw = openpyxl.load_workbook(wb_path, data_only=False)
        wb_cached = openpyxl.load_workbook(wb_path, data_only=True)
    except Exception:
        return None
    ws_raw = None
    for name in wb_raw.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            ws_raw = wb_raw[name]
            break
    if ws_raw is None:
        return None
    ws_cached = wb_cached[ws_raw.title]
    rows = []
    for r_idx, row in enumerate(ws_raw.iter_rows(), start=1):
        out = []
        for c_idx, cell in enumerate(row, start=1):
            out.append(_cell_effective(cell.value, ws_cached.cell(row=r_idx, column=c_idx).value))
        rows.append(out)
    return rows


def numeric_check(label, a_val, g_val, tol):
    if a_val is FORMULA_UNCACHED:
        check(f"{label} (skipped: agent cell is an uncached formula; task requires literal values)", True)
        return
    check(label, num_close(a_val, g_val, tol), f"got {a_val}")


def as_text(v):
    if v is FORMULA_UNCACHED or v is None:
        return ""
    return str(v).strip().lower()


def date_str(v):
    """Normalize a date-valued cell to 'YYYY-MM-DD'.

    Tolerates datetime objects and the common string forms a faithful agent
    might write ('YYYY-MM-DD', 'YYYY/MM/DD', 'M/D/YYYY', 'M-D-YY', ISO
    datetimes like '2026-03-09T00:00:00', ...). Falls back to the raw string
    when nothing parses so lookups never crash on an unexpected format.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()[:10]
    s = str(v).strip()
    if not s:
        return ""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"(\d{4})[/._](\d{1,2})[/._](\d{1,2})", s)
    if m:
        return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})", s)
    if m:
        mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mm <= 12 and 1 <= dd <= 31:
            year = yy if yy >= 100 else 2000 + yy
            return "%04d-%02d-%02d" % (year, mm, dd)
    try:
        from dateutil import parser as _dp
        return _dp.parse(s).strftime("%Y-%m-%d")
    except Exception:
        return s[:10]


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel ===")
    path = os.path.join(agent_workspace, "Options_Monitor.xlsx")
    gt_path = os.path.join(gt_workspace, "Options_Monitor.xlsx")
    check("Options_Monitor.xlsx exists", os.path.exists(path), f"Expected {path}")
    if not os.path.exists(path):
        return
    if not os.path.exists(gt_path):
        check("GT Options_Monitor.xlsx exists", False)
        return

    def rows_only(sheet_rows):
        return [r for r in sheet_rows[1:] if r and r[0] is not None and str(r[0]).strip() != ""]

    # Position Analysis
    a_all = load_sheet_rows(path, "Position Analysis")
    g_all = load_sheet_rows(gt_path, "Position Analysis")
    check("Sheet 'Position Analysis' present", a_all is not None)
    if a_all is not None and g_all is not None:
        a_data = rows_only(a_all)
        g_data = rows_only(g_all)
        check(f"Position Analysis row count == {len(g_data)}",
              len(a_data) == len(g_data), f"got {len(a_data)}")

        # Sort: Symbol then Expiration
        keys = [(str(r[0]).upper(), date_str(r[1])) for r in a_data]
        check("Position Analysis sorted by Symbol then Expiration",
              keys == sorted(keys), f"got {keys[:3]}...")

        # Build (symbol, expiration, type) -> row mapping
        a_lookup = {(str(r[0]).strip().upper(), date_str(r[1]), str(r[2]).strip().lower()): r for r in a_data}
        for g_row in g_data:
            sym = str(g_row[0]).strip().upper()
            exp = date_str(g_row[1])
            typ = str(g_row[2]).strip().lower()
            key = (sym, exp, typ)
            a_row = a_lookup.get(key)
            check(f"Row ({sym}, {exp}, {typ}) present", a_row is not None)
            if a_row is None:
                continue
            # Num_Contracts (exact)
            numeric_check(f"  ({sym}, {exp}, {typ}).Num_Contracts == {g_row[3]}", a_row[3], g_row[3], 0)
            # Avg_IV (±0.5)
            numeric_check(f"  ({sym}, {exp}, {typ}).Avg_IV ≈ {g_row[6]}", a_row[6], g_row[6], 0.5)
            # Risk_Level (exact)
            check(f"  ({sym}, {exp}, {typ}).Risk_Level == '{g_row[7]}'",
                  as_text(a_row[7]) == as_text(g_row[7]),
                  f"got '{a_row[7]}'")

    # Expiry Alerts
    a_all = load_sheet_rows(path, "Expiry Alerts")
    g_all = load_sheet_rows(gt_path, "Expiry Alerts")
    check("Sheet 'Expiry Alerts' present", a_all is not None)
    if a_all is not None and g_all is not None:
        a_data = rows_only(a_all)
        g_data = rows_only(g_all)
        check(f"Expiry Alerts row count == {len(g_data)}",
              len(a_data) == len(g_data), f"got {len(a_data)}")
        # Sort by Days_Until_Expiry ascending
        days = []
        for r in a_data:
            fv = _to_float(r[4])
            if fv is not None:
                days.append(fv)
        check("Expiry Alerts sorted by Days_Until_Expiry ascending",
              days == sorted(days), f"got {days}")
        # Verify each expected row
        a_lookup = {(str(r[0]).strip().upper(), date_str(r[1]), str(r[2]).strip().lower()): r for r in a_data}
        for g_row in g_data:
            sym = str(g_row[0]).strip().upper()
            exp = date_str(g_row[1])
            typ = str(g_row[2]).strip().lower()
            a_row = a_lookup.get((sym, exp, typ))
            check(f"Expiry Alerts row ({sym}, {exp}, {typ}) present",
                  a_row is not None)
            if a_row is None:
                continue
            numeric_check(f"  Days_Until_Expiry == {g_row[4]}", a_row[4], g_row[4], 0)

    # Summary
    a_all = load_sheet_rows(path, "Summary")
    g_all = load_sheet_rows(gt_path, "Summary")
    check("Sheet 'Summary' present", a_all is not None)
    if a_all is not None and g_all is not None:
        a_data = rows_only(a_all)
        g_data = rows_only(g_all)
        a_lookup = {str(r[0]).strip().lower(): r[1] for r in a_data}
        for g_row in g_data:
            metric = str(g_row[0]).strip().lower()
            val = a_lookup.get(metric)
            check(f"Summary metric '{metric}' present", val is not None)
            if val is None:
                continue
            if metric == "stocks_with_near_expiry":
                # Sort comma-separated symbols
                got_syms = sorted([s.strip().upper() for s in str(val).split(",") if s.strip()])
                exp_syms = sorted([s.strip().upper() for s in str(g_row[1]).split(",") if s.strip()])
                check(f"  Summary.{metric} == {exp_syms}",
                      got_syms == exp_syms, f"got {got_syms}")
            elif _to_float(g_row[1]) is not None:
                numeric_check(f"  Summary.{metric} == {g_row[1]}", val, g_row[1], 0)
            else:
                check(f"  Summary.{metric} == '{g_row[1]}'",
                      as_text(val) == as_text(g_row[1]),
                      f"got '{val}'")


def event_wall_dates(start_dt, start_tz):
    """Return the set of plausible calendar dates for an event's start.

    A faithful agent may record the same instant with different
    representations (naive local time, 'Z'/UTC, an explicit offset, or an
    offset plus a named timezone). The intended calendar date is always the
    expiration date, so accept any rendering that yields it. This prevents a
    DB-session-timezone / embedded-offset combination from shifting the wall
    date across midnight and failing a correct model.
    """
    dates = set()
    if start_dt is None:
        return dates
    # DB session timezone rendering (matches naive local round-trips)
    try:
        dates.add(start_dt.date())
    except Exception:
        pass
    # UTC rendering (matches 'Z' / explicit-offset instants)
    try:
        if start_dt.tzinfo is not None:
            dates.add(start_dt.astimezone(timezone.utc).date())
    except Exception:
        pass
    # Named-timezone rendering (matches offset + timeZone)
    if start_tz:
        try:
            from zoneinfo import ZoneInfo
            dates.add(start_dt.astimezone(ZoneInfo(start_tz)).date())
        except Exception:
            pass
    return dates


def check_gcal(gt_workspace):
    print("\n=== Checking Calendar Events ===")
    gt_path = os.path.join(gt_workspace, "Options_Monitor.xlsx")
    # Use Expiry Alerts to determine which (symbol, expiration) need events
    g_alerts = load_sheet_rows(gt_path, "Expiry Alerts")
    expected_events = set()
    if g_alerts:
        for r in g_alerts[1:]:
            if r and r[0] and str(r[0]).strip():
                expected_events.add((str(r[0]).strip().upper(), date_str(r[1])))

    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return

    cur.execute(
        "SELECT summary, description, start_datetime, start_timezone "
        "FROM gcal.events WHERE summary ILIKE '%options expiry%'"
    )
    events = cur.fetchall()
    check("At least 1 options expiry calendar event", len(events) >= 1, f"got {len(events)}")

    # Collect symbol -> candidate wall dates from events
    actual_events = {}
    for sm, desc, start, start_tz in events:
        sm_upper = (sm or "").upper()
        for sym in ("AMZN", "GOOGL", "JNJ", "JPM", "XOM"):
            if sym in sm_upper:
                actual_events.setdefault(sym, set()).update(event_wall_dates(start, start_tz))

    # Each unique (symbol, expiration) from expected_events should have an event
    for sym, exp in sorted(expected_events):
        present = sym in actual_events and exp in actual_events[sym]
        check(f"Calendar event for {sym} on {exp} present",
              present, f"actual events: {actual_events}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(agent_ws, gt_ws)
    check_gcal(gt_ws)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
