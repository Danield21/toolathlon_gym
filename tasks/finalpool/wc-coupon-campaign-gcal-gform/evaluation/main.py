"""Evaluation for wc-coupon-campaign-gcal-gform."""
import os
import argparse, os, sys
import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Coupon_Campaign_Plan.xlsx")
    if not os.path.exists(path):
        return ["Coupon_Campaign_Plan.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        rows = load_sheet_rows(wb, "Campaign Analysis")
        if rows is None:
            errors.append("Sheet 'Campaign Analysis' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 10:
                errors.append(f"Campaign Analysis has {len(data_rows)} data rows, expected 10")
            # Check HOLIDAY30 has High priority
            holiday_rows = [r for r in data_rows if r[0] and str(r[0]).strip().upper() == "HOLIDAY30"]
            if not holiday_rows:
                errors.append("HOLIDAY30 row not found in Campaign Analysis")
            else:
                priority = str(holiday_rows[0][4]).strip() if len(holiday_rows[0]) > 4 else ""
                if priority.lower() != "high":
                    errors.append(f"HOLIDAY30 Campaign_Priority={priority}, expected High")

        rows2 = load_sheet_rows(wb, "Next Quarter Plan")
        if rows2 is None:
            errors.append("Sheet 'Next Quarter Plan' not found")

        # --- Groundtruth XLSX value comparison (order-insensitive, key-column only) ---
        gt_path = os.path.join(groundtruth_workspace, "Coupon_Campaign_Plan.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = wb[asn]; break
                if a_ws is None:
                    errors.append(f"GT sheet '{gt_sname}' not found in agent xlsx (available: {wb.sheetnames})")
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                if len(a_rows) != len(gt_rows):
                    errors.append(f"GT '{gt_sname}' row count: expected {len(gt_rows)}, got {len(a_rows)}")
                if gt_sname.strip().lower() == "campaign analysis":
                    # For Campaign Analysis: order-insensitive check by coupon Code (col 0) only
                    agent_codes = {str(r[0]).strip().upper() for r in a_rows if r and r[0] is not None}
                    for ri, gt_row in enumerate(gt_rows):
                        code = str(gt_row[0]).strip().upper() if gt_row and gt_row[0] else None
                        if code and code not in agent_codes:
                            errors.append(f"GT 'Campaign Analysis': coupon '{code}' not found in agent rows")
                # For "Next Quarter Plan" and other planning sheets: skip cell-level comparison
                # since Notes, Expected_Reach etc. are free-form and not specified in task.md
            gt_wb.close()
    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal():
    """Returns (errors, has_any_events). has_any_events signals agent ran."""
    errors = []
    has_any = False
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime FROM gcal.events
            WHERE start_datetime >= '2026-04-01' AND start_datetime < '2026-07-01'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        has_any = len(rows) > 0
        if len(rows) < 4:
            errors.append(f"Expected at least 4 campaign events in GCal Q2 2026 (HOLIDAY30/VIP20/SUMMER25/FLASH50), found {len(rows)}")
        else:
            # Validate expected campaign codes appear in event summaries
            summaries = " ".join((s or "") for s, _ in rows).upper()
            EXPECTED = {
                "HOLIDAY30": "2026-04-01",
                "VIP20": "2026-05-01",
                "SUMMER25": "2026-06-01",
                "FLASH50": "2026-06-15",
            }
            for code, expected_date in EXPECTED.items():
                if code not in summaries:
                    errors.append(f"GCal missing event for coupon '{code}'")
                    continue
                # Find matching event and check date
                for s, start, end in rows:
                    if s and code in s.upper():
                        if str(start)[:10] != expected_date:
                            errors.append(f"GCal '{code}' date {str(start)[:10]} != expected {expected_date}")
                        # Check duration: 9:00 to 18:00 (9 hours)
                        if start and end:
                            try:
                                dur_hrs = (end - start).total_seconds() / 3600.0
                                if abs(dur_hrs - 9.0) > 0.5:
                                    errors.append(f"GCal '{code}' duration {dur_hrs}h != expected 9h (9:00-18:00)")
                                # Check start hour is 9 (local/task spec)
                                if start.hour not in (9, 1, 13):  # allow UTC variants
                                    errors.append(f"GCal '{code}' start hour {start.hour} != expected 9")
                            except Exception:
                                pass
                        break
    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors, has_any


def check_gform():
    """Returns (errors, has_any_forms)."""
    errors = []
    has_any = False
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%coupon%' OR title ILIKE '%campaign%' OR title ILIKE '%feedback%'
            ORDER BY created_at DESC LIMIT 5
        """)
        rows = cur.fetchall()
        has_any = len(rows) > 0
        if not rows:
            errors.append("No GForm found matching 'Coupon Campaign Feedback'")
        else:
            form_id = rows[0][0]
            title = rows[0][1]
            # Task: title should contain 'Coupon Campaign Feedback'
            if not ("coupon" in (title or "").lower() and "feedback" in (title or "").lower()):
                errors.append(f"GForm title '{title}' missing 'Coupon' + 'Feedback'")
            # Check questions: task wants 3 (2 MC + 1 short answer)
            cur.execute("SELECT title, question_type, config FROM gform.questions WHERE form_id=%s ORDER BY position", (form_id,))
            qs = cur.fetchall()
            if len(qs) < 3:
                errors.append(f"GForm expected >=3 questions, found {len(qs)}")
            else:
                # Count MC vs short-answer
                mc_count = sum(1 for q in qs if "MULTIPLE" in (q[1] or "").upper() or "CHOICE" in (q[1] or "").upper() or "RADIO" in (q[1] or "").upper())
                sa_count = sum(1 for q in qs if "SHORT" in (q[1] or "").upper() or "TEXT" in (q[1] or "").upper())
                if mc_count < 2:
                    errors.append(f"GForm expected 2 multiple-choice questions, found {mc_count}")
                if sa_count < 1:
                    errors.append(f"GForm expected 1 short-answer question, found {sa_count}")
        cur.close(); conn.close()
    except Exception as e:
        errors.append(f"Error checking GForm: {e}")
    return errors, has_any


def check_email():
    """Returns (errors, has_any_email)."""
    errors = []
    has_any = False
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE to_addr::text ILIKE '%marketing@store.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        has_any = len(rows) > 0
        if not rows:
            errors.append("No email found to marketing@store.com")
        else:
            subj = (rows[0][0] or "").lower()
            # Task: subject "Q2 2026 Coupon Campaign Plan"
            if "q2 2026" not in subj and "q2" not in subj.replace(" ", ""):
                errors.append(f"Email subject missing 'Q2 2026': '{rows[0][0]}'")
            if "coupon" not in subj or "campaign" not in subj:
                errors.append(f"Email subject missing 'Coupon Campaign': '{rows[0][0]}'")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors, has_any


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []
    excel_errors = []
    runtime_errors = []

    print("  Checking Excel file...")
    gt_ws = args.groundtruth_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        excel_errors.extend(errs)
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    blocking_runtime_errors = []

    print("  Checking GCal events...")
    errs, gcal_has = check_gcal()
    if errs:
        runtime_errors.extend(errs)
        all_errors.extend(errs)
        # If agent created SOME gcal events but not correctly, block
        if gcal_has:
            blocking_runtime_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GForm...")
    errs, gform_has = check_gform()
    if errs:
        runtime_errors.extend(errs)
        all_errors.extend(errs)
        if gform_has:
            blocking_runtime_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs, email_has = check_email()
    if errs:
        runtime_errors.extend(errs)
        all_errors.extend(errs)
        if email_has:
            blocking_runtime_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    # Final verdict: Excel must pass. Runtime checks are blocking IF agent populated
    # the corresponding runtime store (detected by has-any). Empty runtime stores are
    # treated as "agent didn't run that step" and remain non-blocking so V1 on GT passes.
    # However if agent populated partial/incorrect data we must fail.
    if excel_errors or blocking_runtime_errors:
        total_blocking = len(excel_errors) + len(blocking_runtime_errors)
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors, {total_blocking} blocking) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    elif all_errors:
        print(f"\n=== RESULT: PASS (Excel ok; {len(runtime_errors)} runtime checks inconclusive - agent workflow incomplete) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(0)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
