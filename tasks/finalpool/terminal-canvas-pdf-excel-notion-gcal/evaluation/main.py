"""Evaluation script for terminal-canvas-pdf-excel-notion-gcal."""
import os
import argparse, json, sys, re

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def get_expected_from_db():
    """Query canvas schema dynamically for expected average grades.

    The task defines avg_gpa as "averaging the per-course average grades", so the
    canonical expected value is the mean of per-course submission-score means
    (=75.40 on the seed).  We also compute the plain global average (=75.80) as a
    fallback reference so that agents who legitimately average all submission
    scores directly (or use enrollment current_score, ~73) still land in range.
    db_query_ok=False on failure.
    """
    expected = {
        "avg_gpa": None,          # mean of per-course submission-score means
        "avg_gpa_global": None,   # plain global AVG(score)
        "db_query_ok": False,
    }
    try:
        import psycopg2
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Mean of per-course submission-score means (matches the task definition).
        cur.execute("""
            SELECT AVG(course_mean) FROM (
                SELECT a.course_id, AVG(s.score) AS course_mean
                FROM canvas.submissions s
                JOIN canvas.assignments a ON a.id = s.assignment_id
                WHERE s.score IS NOT NULL AND a.course_id IS NOT NULL
                GROUP BY a.course_id
            ) t
        """)
        row = cur.fetchone()
        if row and row[0] is not None:
            expected["avg_gpa"] = float(row[0])
        # Plain global average (fallback reference).
        cur.execute("SELECT AVG(score) FROM canvas.submissions WHERE score IS NOT NULL")
        row = cur.fetchone()
        if row and row[0] is not None:
            expected["avg_gpa_global"] = float(row[0])
        expected["db_query_ok"] = expected["avg_gpa"] is not None
        cur.close()
        conn.close()
    except Exception as e:
        print(f"  [WARN] DB query for expected values failed: {e}")
    return expected


EXPECTED = get_expected_from_db()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def _to_float(x):
    """Robust numeric coercion. Handles int/float, strings with %/currency/
    thousands separators / whitespace, Excel-percent style values, and formulas
    that begin with '='. Returns None when the value cannot be parsed."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none", "null", "n/a", "na", "-", "--"):
        return None
    if s.startswith("="):
        s = s[1:].strip()  # a formula: try to read its cached/literal tail
    # strip currency symbols and thousands separators
    s = s.replace("$", "").replace("€", "").replace("¥", "").replace("£", "").replace(",", "")
    # percent sign removed; caller may scale if the task uses percent semantics
    s = s.replace("%", "").strip()
    m = re.match(r"^[-+]?\d*\.?\d+", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=2.0):
    """Tolerant numeric comparison. Both sides are coerced via _to_float; if both
    parse, compare |a-b| <= tol. If either side cannot be parsed, fall back to a
    case-insensitive string comparison."""
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def threshold_near(value, expected, tol=0.05):
    """Check an accreditation threshold cell against the expected numeric value.

    Tolerates: bare numbers (78), strings with '%' ('78.0%'), Excel percentage-
    formatted cells (stored as 0.78 meaning 78%), and values written in the wrong
    scale (65 instead of 0.65)."""
    a = _to_float(value)
    if a is None:
        return False
    if abs(a - expected) <= tol:
        return True
    # Excel-percent style: cell holds 0.78 but means 78
    if expected >= 1 and 0 < a < 1:
        return abs(a * 100 - expected) <= max(tol, 1.0)
    # value written on the wrong scale, e.g. 65 for 0.65
    if 0 < expected < 1 and a > 1:
        return abs(a / 100 - expected) <= max(tol, 0.05)
    return False


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _parse_launch_time(launch_time):
    """Parse the run's launch timestamp robustly.

    The runtime harness passes it as 'YYYY-MM-DD HH:MM:SS' plus a weekday
    suffix (e.g. '2026-08-07 12:34:56 Friday'); other callers may pass it
    without the weekday or with an ISO 'T' separator. Any trailing alpha
    weekday token is stripped before parsing, and several common formats are
    tried. Returns a naive datetime, or None when nothing parses."""
    from datetime import datetime
    if not launch_time:
        return None
    s = str(launch_time).strip()
    # Drop trailing weekday names (e.g. 'Friday'), including the '%A' token.
    for _ in range(2):
        parts = s.rsplit(" ", 1)
        if len(parts) == 2 and parts[1].isalpha() and not parts[1].isdigit():
            s = parts[0]
        else:
            break
    s = s.replace("T", " ", 1)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                "%Y/%m/%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel Workbook ===")
    import openpyxl

    path = os.path.join(agent_workspace, "Accreditation_Self_Study.xlsx")
    check("Accreditation_Self_Study.xlsx exists", os.path.exists(path))
    if not os.path.exists(path):
        return

    wb = openpyxl.load_workbook(path)
    sheet_names = [s.lower() for s in wb.sheetnames]

    # Check 4 sheets exist
    has_course = any("course" in s and "data" in s for s in sheet_names)
    has_matrix = any("accreditation" in s or "matrix" in s for s in sheet_names)
    has_gap = any("gap" in s for s in sheet_names)
    has_resource = any("resource" in s for s in sheet_names)
    check("Course_Data sheet exists", has_course, f"Sheets: {wb.sheetnames}")
    check("Accreditation_Matrix sheet exists", has_matrix, f"Sheets: {wb.sheetnames}")
    check("Gap_Analysis sheet exists", has_gap, f"Sheets: {wb.sheetnames}")
    check("Resource_Needs sheet exists", has_resource, f"Sheets: {wb.sheetnames}")

    # Check Course_Data content
    for sn in wb.sheetnames:
        if "course" in sn.lower() and "data" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Course_Data has 22 course rows", len(rows) >= 22, f"Got {len(rows)}")
            if rows:
                # Check some known course names
                names = [str(r[0]).lower() if r[0] else "" for r in rows]
                check("Course_Data includes Creative Computing", any("creative" in n for n in names))
                check("Course_Data includes Foundations of Finance", any("finance" in n for n in names))
            # Check headers
            headers = [str(c.value).lower() if c.value else "" for c in ws[1]]
            check("Course_Data has enrollment column", any("enroll" in h for h in headers), f"Headers: {headers}")
            check("Course_Data has grade column", any("grade" in h or "avg" in h for h in headers), f"Headers: {headers}")
            break

    # Check Accreditation_Matrix content
    for sn in wb.sheetnames:
        if "accreditation" in sn.lower() or "matrix" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Accreditation_Matrix has 8 criteria rows", len(rows) >= 8, f"Got {len(rows)}")
            if rows:
                all_text = " ".join(str(c).lower() for r in rows for c in r if c)
                check("Matrix has Compliant status", "compliant" in all_text)
                check("Matrix has Partial status", "partial" in all_text)
                # Check specific criteria
                check("Matrix mentions Student Learning Outcomes", "student learning" in all_text or "learning outcome" in all_text, f"Text sample: {all_text[:200]}")
                # Check threshold values (robust to '%' strings and Excel percent formats)
                has_78 = any(threshold_near(r[1], 78.0, tol=1.0) for r in rows if r[1] is not None)
                has_065 = any(threshold_near(r[1], 0.65, tol=0.05) for r in rows if r[1] is not None)
                check("Matrix has C1 threshold ~78", has_78, f"Thresholds: {[r[1] for r in rows]}")
                check("Matrix has C2 threshold ~0.65", has_065, f"Thresholds: {[r[1] for r in rows]}")
            break

    # Check Gap_Analysis content
    for sn in wb.sheetnames:
        if "gap" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            non_empty = [r for r in rows if any(c for c in r)]
            check("Gap_Analysis has rows for partial/non-compliant criteria", len(non_empty) >= 2, f"Got {len(non_empty)}")
            if non_empty:
                all_text = " ".join(str(c).lower() for r in non_empty for c in r if c)
                check("Gap_Analysis mentions remediation", "remed" in all_text or "action" in all_text or "improve" in all_text)
            break

    # Check Resource_Needs content
    for sn in wb.sheetnames:
        if "resource" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            non_empty = [r for r in rows if any(c for c in r)]
            check("Resource_Needs has at least 3 rows", len(non_empty) >= 3, f"Got {len(non_empty)}")
            if non_empty:
                headers = [str(c.value).lower() if c.value else "" for c in ws[1]]
                check("Resource_Needs has cost column", any("cost" in h or "estimated" in h for h in headers), f"Headers: {headers}")
            break


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Find accreditation database
        cur.execute("SELECT id, title, properties FROM notion.databases")
        dbs = cur.fetchall()
        accred_db = None
        for db_id, title_json, props_json in dbs:
            title_text = ""
            if title_json:
                titles = title_json if isinstance(title_json, list) else json.loads(title_json) if isinstance(title_json, str) else []
                for t in titles:
                    if isinstance(t, dict) and "text" in t:
                        title_text += t["text"].get("content", "")
            if "accreditation" in title_text.lower() and "action" in title_text.lower():
                accred_db = (db_id, props_json)
                break

        check("Accreditation Action Items database exists", accred_db is not None,
              f"Found DBs: {[str(d[1])[:80] for d in dbs]}")

        if accred_db:
            db_id, props = accred_db
            props_data = props if isinstance(props, dict) else json.loads(props) if isinstance(props, str) else {}

            # Check ALL required properties (Criterion title, Status, Owner, Due_Date, Evidence)
            prop_names = [k.lower() for k in props_data.keys()]
            check("DB has Criterion (title) property",
                  any("criterion" in p or "name" in p or "title" in p for p in prop_names),
                  f"Props: {list(props_data.keys())}")
            check("DB has Status property", any("status" in p for p in prop_names),
                  f"Props: {list(props_data.keys())}")
            check("DB has Owner property", any("owner" in p for p in prop_names),
                  f"Props: {list(props_data.keys())}")
            check("DB has Due_Date property",
                  any("due" in p and "date" in p for p in prop_names) or any(p == "date" for p in prop_names),
                  f"Props: {list(props_data.keys())}")
            check("DB has Evidence property", any("evidence" in p for p in prop_names),
                  f"Props: {list(props_data.keys())}")

            # Check pages
            cur.execute("SELECT id, properties FROM notion.pages WHERE parent::text LIKE %s AND archived = false",
                        (f'%{db_id}%',))
            pages = cur.fetchall()
            check("Notion has exactly 8 accreditation criterion pages", len(pages) == 8, f"Got {len(pages)}")

            if pages:
                # Check page content
                statuses = []
                for _, page_props in pages:
                    pp = page_props if isinstance(page_props, dict) else json.loads(page_props) if isinstance(page_props, str) else {}
                    for k, v in pp.items():
                        if "status" in k.lower() and isinstance(v, dict):
                            sel = v.get("select", {})
                            if sel and isinstance(sel, dict):
                                statuses.append(sel.get("name", ""))
                has_compliant = any("compliant" in s.lower() for s in statuses if s and "non" not in s.lower() and "partial" not in s.lower())
                has_partial = any("partial" in s.lower() for s in statuses if s)
                check("Pages include Compliant status", has_compliant, f"Statuses: {statuses}")
                check("Pages include Partial status", has_partial, f"Statuses: {statuses}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion accessible", False, str(e))


def check_gcal(launch_time):
    print("\n=== Checking Google Calendar Events ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        launch_dt = _parse_launch_time(launch_time)

        cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
        summaries = [str(e[0]).lower() for e in events]

        has_evidence = any("evidence" in s and "deadline" in s for s in summaries)
        has_draft = any("draft" in s and ("report" in s or "due" in s) for s in summaries)
        has_final = any("final" in s and "submission" in s for s in summaries)
        check("Evidence Collection Deadline event exists", has_evidence, f"Events: {summaries}")
        check("Draft Report Due event exists", has_draft, f"Events: {summaries}")
        check("Final Submission event exists", has_final, f"Events: {summaries}")

        # Check descriptions mention accreditation
        accred_events = [e for e in events if any(kw in str(e[0]).lower() for kw in ["evidence", "draft report", "final submission"])]
        if accred_events:
            descs = " ".join(str(e[1]).lower() for e in accred_events if e[1])
            check("Calendar events mention accreditation", "accreditation" in descs or "self-study" in descs or "self study" in descs,
                  f"Descriptions: {descs[:200]}")

        # Check timing (roughly 30/60/90 days from launch)
        if launch_dt and accred_events:
            from datetime import timedelta
            for e in events:
                s = str(e[0]).lower()
                if e[2] and "evidence" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Evidence deadline ~30 days from launch", 28 <= days_diff <= 32, f"Days: {days_diff}")
                elif e[2] and "draft" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Draft due ~60 days from launch", 58 <= days_diff <= 62, f"Days: {days_diff}")
                elif e[2] and "final" in s and "submission" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Final submission ~90 days from launch", 88 <= days_diff <= 92, f"Days: {days_diff}")

        cur.close()
        conn.close()
    except Exception as e:
        check("GCal accessible", False, str(e))


def check_scripts(agent_workspace):
    print("\n=== Checking Scripts and Outputs ===")
    check("compute_metrics.py exists", os.path.exists(os.path.join(agent_workspace, "compute_metrics.py")))
    check("evaluate_compliance.py exists", os.path.exists(os.path.join(agent_workspace, "evaluate_compliance.py")))
    check("generate_summary.py exists", os.path.exists(os.path.join(agent_workspace, "generate_summary.py")))

    # Check program_metrics.json
    metrics_path = os.path.join(agent_workspace, "program_metrics.json")
    check("program_metrics.json exists", os.path.exists(metrics_path))
    if os.path.exists(metrics_path):
        with open(metrics_path) as f:
            metrics = json.load(f)
        # Accept avg_gpa or any key that contains 'gpa' (e.g. average_gpa)
        gpa_key = None
        for k in metrics:
            if isinstance(k, str) and "gpa" in k.lower():
                gpa_key = k
                break
        check("Metrics has avg_gpa", gpa_key is not None, f"Keys: {list(metrics.keys())}")
        # Record DB query failure as a check failure (no silent fallback)
        check("DB expected-value query succeeded (avg_gpa)",
              EXPECTED.get("db_query_ok") is True,
              "DB query for expected avg_gpa failed; cannot validate metric")
        if gpa_key and EXPECTED.get("db_query_ok"):
            # Accept the per-course mean-of-means (=task definition) OR the plain
            # global average; tolerance is generous so agents computing from
            # enrollment current_score (~73) also pass.
            ok = False
            detail = f"Got {metrics[gpa_key]}"
            if EXPECTED.get("avg_gpa") is not None:
                ok = num_close(metrics[gpa_key], EXPECTED["avg_gpa"], tol=3.0)
                detail += f", expected ~{EXPECTED['avg_gpa']:.1f} (per-course avg)"
            if not ok and EXPECTED.get("avg_gpa_global") is not None:
                ok = num_close(metrics[gpa_key], EXPECTED["avg_gpa_global"], tol=3.0)
                detail += f" or ~{EXPECTED['avg_gpa_global']:.1f} (global avg)"
            check("avg_gpa roughly correct", ok, detail)

    # Check compliance_assessment.json
    compliance_path = os.path.join(agent_workspace, "compliance_assessment.json")
    check("compliance_assessment.json exists", os.path.exists(compliance_path))
    if os.path.exists(compliance_path):
        with open(compliance_path) as f:
            compliance = json.load(f)
        # Should be a list or dict with 8 criteria
        if isinstance(compliance, list):
            check("Compliance has >= 8 entries", len(compliance) >= 8, f"Got {len(compliance)}")
        elif isinstance(compliance, dict):
            # Common patterns: top-level is criteria_count or has 'criteria' key with list
            criteria_list = None
            for k in ["criteria", "items", "results", "assessments"]:
                if k in compliance and isinstance(compliance[k], list):
                    criteria_list = compliance[k]
                    break
            if criteria_list is not None:
                check("Compliance dict has 'criteria' list with >= 8 items",
                      len(criteria_list) >= 8, f"Got {len(criteria_list)}")
            else:
                # 8 criteria as keys
                check("Compliance has >= 8 keys/criteria", len(compliance) >= 8,
                      f"Keys: {list(compliance.keys())[:10]}")

    # Check accreditation_summary.txt
    summary_path = os.path.join(agent_workspace, "accreditation_summary.txt")
    check("accreditation_summary.txt exists", os.path.exists(summary_path))
    if os.path.exists(summary_path):
        with open(summary_path) as f:
            text = f.read().lower()
        check("Summary mentions compliant", "compliant" in text)
        check("Summary mentions partial", "partial" in text)
        # Compliance percentage: '%', the word 'percent', or an 'x/y' / 'x out of y'
        has_pct = ("%" in text or "percent" in text
                   or re.search(r"\d+\s*(/|out of|of)\s*\d+", text))
        check("Summary has compliance percentage", bool(has_pct), f"Length: {len(text)}")


def check_reverse_validation():
    """Check that noise Notion pages are NOT in the accreditation tracker."""
    print("\n=== Reverse Validation ===")
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Noise notion pages that should NOT appear in accreditation database
        noise_titles = ["q1 budget review", "server migration", "marketing campaign launch"]

        # Find the accreditation database
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        accred_db_id = None
        for db_id, title_json in dbs:
            title_text = ""
            if title_json:
                titles = title_json if isinstance(title_json, list) else json.loads(title_json) if isinstance(title_json, str) else []
                for t in titles:
                    if isinstance(t, dict) and "text" in t:
                        title_text += t["text"].get("content", "")
            if "accreditation" in title_text.lower() and "action" in title_text.lower():
                accred_db_id = db_id
                break

        if accred_db_id:
            cur.execute(
                "SELECT properties FROM notion.pages WHERE parent::text LIKE %s AND archived = false",
                (f'%{accred_db_id}%',))
            pages = cur.fetchall()
            page_titles = []
            for (props,) in pages:
                pp = props if isinstance(props, dict) else json.loads(props) if isinstance(props, str) else {}
                for k, v in pp.items():
                    if "title" in k.lower() or k.lower() == "name":
                        if isinstance(v, dict) and "title" in v:
                            for t in v["title"]:
                                if isinstance(t, dict) and "text" in t:
                                    page_titles.append(t["text"].get("content", "").lower())

            no_noise = not any(nt in " ".join(page_titles) for nt in noise_titles)
            check("No noise Notion pages in accreditation tracker (budget, migration, marketing)",
                  no_noise,
                  f"Page titles in accreditation DB: {page_titles}")
        else:
            check("No noise Notion pages in accreditation tracker", True,
                  "No accreditation DB found to check")

        # Check noise 'Project Milestones' database pages are not mixed in
        cur.execute("SELECT id, title FROM notion.databases")
        all_dbs = cur.fetchall()
        noise_db_titles = []
        for db_id, title_json in all_dbs:
            title_text = ""
            if title_json:
                titles = title_json if isinstance(title_json, list) else json.loads(title_json) if isinstance(title_json, str) else []
                for t in titles:
                    if isinstance(t, dict) and "text" in t:
                        title_text += t["text"].get("content", "")
            if "project milestones" in title_text.lower():
                noise_db_titles.append(title_text)
        # The noise DB existing is fine (it was injected), but accreditation DB should be separate
        check("Accreditation DB is distinct from noise Project Milestones DB",
              accred_db_id is not None,
              "Accreditation Action Items DB should exist separately")

        cur.close()
        conn.close()
    except Exception as e:
        check("Reverse validation (notion noise)", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_gcal(args.launch_time)
    check_scripts(args.agent_workspace)
    check_reverse_validation()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    # Strict gate: ALL checks must pass
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
