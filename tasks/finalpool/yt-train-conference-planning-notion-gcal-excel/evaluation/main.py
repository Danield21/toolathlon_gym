"""
Evaluation for yt-train-conference-planning-notion-gcal-excel task.

Checks:
1. Conference_Planning.xlsx exists with Pre_Conference_Videos, Speaker_Travel, Summary sheets (correct values).
2. Notion page 'Technology Innovation Forum 2026' exists with the three required sections.
3. Three GCal events with exact summaries, dates, and times.
4. Three confirmation emails (one per speaker) with required body content.
"""
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


SPEAKERS = [
    {"name": "Dr. Alex Kim", "email": "alex.kim@tech.edu", "city": "Beijing",
     "train": "G235", "depart": "17:30", "arrive": "19:16", "price": 109.5},
    {"name": "Dr. Maria Chen", "email": "maria.chen@research.org", "city": "Beijing",
     "train": "G235", "depart": "17:30", "arrive": "19:16", "price": 109.5},
    {"name": "Prof. James Wu", "email": "james.wu@university.edu", "city": "Shanghai",
     "train": "G168", "depart": "17:00", "arrive": "18:55", "price": 119.5},
]

CONF_NAME = "Technology Innovation Forum 2026"
TOTAL_BUDGET = 338.5  # 109.5 + 109.5 + 119.5


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace("_", " ") == name.strip().lower().replace("_", " "):
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Conference_Planning.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Conference_Planning.xlsx")
    if not os.path.exists(xlsx_path):
        record("Conference_Planning.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Conference_Planning.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    # === Pre_Conference_Videos ===
    ws = get_sheet(wb, "Pre_Conference_Videos")
    record("Sheet 'Pre_Conference_Videos' exists", ws is not None,
           f"sheets: {wb.sheetnames}")
    if ws is not None:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
            for col in ["rank", "video_id", "title", "view_count", "duration_sec"]:
                record(f"PCV column '{col}'", col in header, f"headers: {header}")
            data = [r for r in rows[1:] if any(c is not None for c in r)]
            record("PCV has exactly 6 rows", len(data) == 6, f"got {len(data)}")
            # Sorted by view_count descending
            try:
                vc_idx = header.index("view_count")
                vcs = [int(r[vc_idx]) for r in data]
                sorted_ok = all(vcs[i] >= vcs[i+1] for i in range(len(vcs)-1))
                record("PCV sorted by View_Count descending", sorted_ok,
                       f"View_Counts: {vcs}")
            except Exception:
                pass

            # Per-row Video_ID validation against expected top-N from DB.
            # Use a strict-but-lenient match: at least 4 of agent's top-6 video_ids
            # must be in the DB-derived top-8 fireship AI-related set.
            try:
                vid_idx = header.index("video_id")
                ttl_idx = header.index("title") if "title" in header else None
            except ValueError:
                vid_idx = None
            try:
                conn2 = psycopg2.connect(**DB_CONFIG)
                cur2 = conn2.cursor()
                cur2.execute(
                    """
                    SELECT video_id, title FROM youtube.videos
                    WHERE channel_title ILIKE '%fireship%'
                      AND (title ~* '\\mAI\\M'
                           OR title ILIKE '%DeepSeek%'
                           OR title ILIKE '%OpenAI%'
                           OR title ILIKE '%ChatGPT%'
                           OR title ILIKE '%Grok%'
                           OR title ILIKE '%Llama%')
                    ORDER BY view_count DESC LIMIT 12
                    """
                )
                top_ids = [r[0] for r in cur2.fetchall()]
                cur2.close()
                conn2.close()
                if vid_idx is not None and top_ids:
                    agent_ids = [str(r[vid_idx]).strip() for r in data if r[vid_idx]]
                    in_set = sum(1 for vid in agent_ids if vid in top_ids)
                    record(
                        "PCV: at least 4/6 video_ids match DB top fireship AI videos",
                        in_set >= 4,
                        f"match {in_set}/{len(agent_ids)}; agent={agent_ids[:6]}; db_top={top_ids[:6]}",
                    )
                    # Top row should be the absolute top by view_count
                    if agent_ids and top_ids:
                        record(
                            "PCV row 1 video_id is the top fireship AI video",
                            agent_ids[0] == top_ids[0],
                            f"got {agent_ids[0]}, expected {top_ids[0]}",
                        )
            except Exception as e:
                print(f"[WARN] PCV video_id check skipped: {e}")

    # === Speaker_Travel ===
    ws = get_sheet(wb, "Speaker_Travel")
    record("Sheet 'Speaker_Travel' exists", ws is not None,
           f"sheets: {wb.sheetnames}")
    if ws is not None:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
            for col in ["speaker_name", "email", "from_city", "train_code",
                         "depart_time", "arrive_time", "ticket_price_cny"]:
                record(f"ST column '{col}'", col in header, f"headers: {header}")

            data = [r for r in rows[1:] if any(c is not None for c in r)]
            record("ST has exactly 3 rows", len(data) == 3, f"got {len(data)}")

            # Build lookup by lowercase name (with/without prefix)
            lookup = {}
            for r in data:
                if r and r[0]:
                    name = str(r[0]).strip().lower()
                    lookup[name] = r

            idx = {h: i for i, h in enumerate(header)}
            for sp in SPEAKERS:
                # Try with full prefix and without
                key1 = sp["name"].lower()
                key2 = sp["name"].split(maxsplit=1)[-1].lower()
                a = lookup.get(key1) or lookup.get(key2)
                # Search by partial name
                if a is None:
                    last = sp["name"].split()[-1].lower()
                    for k, r in lookup.items():
                        if last in k:
                            a = r
                            break
                record(f"ST has {sp['name']}", a is not None,
                       f"keys: {list(lookup.keys())}")
                if a is None:
                    continue
                if "email" in idx:
                    record(f"{sp['name']} email correct",
                           str(a[idx['email']] or "").strip().lower() == sp["email"].lower(),
                           f"got {a[idx['email']]}")
                if "from_city" in idx:
                    record(f"{sp['name']} city correct",
                           str(a[idx['from_city']] or "").strip().lower() == sp["city"].lower(),
                           f"got {a[idx['from_city']]}")
                if "train_code" in idx:
                    record(f"{sp['name']} train code correct",
                           str(a[idx['train_code']] or "").strip().upper() == sp["train"],
                           f"got {a[idx['train_code']]}")
                if "depart_time" in idx:
                    record(f"{sp['name']} depart time correct",
                           sp["depart"] in str(a[idx['depart_time']] or ""),
                           f"got {a[idx['depart_time']]}")
                if "arrive_time" in idx:
                    record(f"{sp['name']} arrive time correct",
                           sp["arrive"] in str(a[idx['arrive_time']] or ""),
                           f"got {a[idx['arrive_time']]}")
                if "ticket_price_cny" in idx:
                    record(f"{sp['name']} ticket price correct",
                           num_close(a[idx['ticket_price_cny']], sp["price"], 1.0),
                           f"got {a[idx['ticket_price_cny']]}, expected {sp['price']}")

    # === Summary ===
    ws = get_sheet(wb, "Summary")
    record("Sheet 'Summary' exists", ws is not None, f"sheets: {wb.sheetnames}")
    if ws is not None:
        rows = list(ws.iter_rows(values_only=True))
        data = [r for r in rows[1:] if any(c is not None for c in r)]
        lookup = {str(r[0]).strip().lower(): r[1] if len(r) > 1 else None for r in data if r and r[0]}

        record("Summary Conference_Name == 'Technology Innovation Forum 2026'",
               str(lookup.get("conference_name", "")).strip().lower() == CONF_NAME.lower(),
               f"got '{lookup.get('conference_name')}'")
        record("Summary Video_Count == 6",
               num_close(lookup.get("video_count"), 6, 0),
               f"got {lookup.get('video_count')}")
        record("Summary Speaker_Count == 3",
               num_close(lookup.get("speaker_count"), 3, 0),
               f"got {lookup.get('speaker_count')}")
        record(f"Summary Total_Travel_Budget_CNY == {TOTAL_BUDGET}",
               num_close(lookup.get("total_travel_budget_cny"), TOTAL_BUDGET, 1.0),
               f"got {lookup.get('total_travel_budget_cny')}")


def check_notion():
    print("\n=== Check 2: Notion page 'Technology Innovation Forum 2026' ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
        pages = cur.fetchall()
        target = None
        for pid, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "technology innovation forum 2026" in props_str:
                target = (pid, props)
                break
        record("Notion page 'Technology Innovation Forum 2026' exists",
               target is not None,
               f"Found {len(pages)} pages")

        if target:
            page_id = target[0]
            cur.execute("""
                SELECT type, block_data FROM notion.blocks WHERE parent_id = %s AND archived = false
            """, (page_id,))
            blocks = cur.fetchall()
            all_text = " ".join(json.dumps(b[1]) for b in blocks if b[1]).lower()

            # Collect heading-block texts only for section validation.
            def _block_plaintext(bd):
                if not isinstance(bd, dict):
                    return ""
                # Look at common heading payload keys in flat order
                pieces = []
                for k in ("rich_text", "text"):
                    arr = bd.get(k)
                    if isinstance(arr, list):
                        for it in arr:
                            if isinstance(it, dict):
                                pieces.append(it.get("plain_text") or
                                              (it.get("text") or {}).get("content") or "")
                            elif isinstance(it, str):
                                pieces.append(it)
                # Some MCPs nest under a key matching the type
                for k in list(bd.keys()):
                    v = bd.get(k)
                    if isinstance(v, dict):
                        arr = v.get("rich_text") or v.get("text")
                        if isinstance(arr, list):
                            for it in arr:
                                if isinstance(it, dict):
                                    pieces.append(it.get("plain_text") or
                                                  (it.get("text") or {}).get("content") or "")
                                elif isinstance(it, str):
                                    pieces.append(it)
                return " ".join(pieces).strip()

            heading_texts = []
            for btype, bdata in blocks:
                if (btype or "").lower() in ("heading_1", "heading_2", "heading_3"):
                    heading_texts.append(_block_plaintext(bdata).lower())
            heading_blob = " | ".join(heading_texts)

            for sec in ["Pre-Conference Resources", "Speaker Travel Details",
                        "Conference Agenda"]:
                # Prefer heading block match; fall back to all_text only if no heading blocks were found.
                ok = (sec.lower() in heading_blob) or (
                    not heading_texts and sec.lower() in all_text
                )
                record(f"Notion page heading-block contains '{sec}'",
                       ok,
                       f"headings={heading_texts[:6]}; blocks_count={len(blocks)}")
            # Check Day1/Day2/Day3 placeholder text
            for d in ["Day 1", "Day 2", "Day 3"]:
                record(f"Notion page mentions '{d}'", d.lower() in all_text, "")
            # Check for at least one speaker name
            for sp in SPEAKERS:
                record(f"Notion page mentions speaker {sp['name'].split()[-1]}",
                       sp['name'].split()[-1].lower() in all_text, "")
    except Exception as e:
        record("Notion page check error", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 3: Google Calendar Conference Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE start_datetime >= '2026-03-12' AND start_datetime < '2026-03-15'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()

        expected = [
            ("Conference Day 1: Opening Ceremony", "2026-03-12", 9, 18),
            ("Conference Day 2: Main Sessions",   "2026-03-13", 9, 18),
            ("Conference Day 3: Closing",          "2026-03-14", 9, 15),
        ]
        for exp_title, exp_date, sh, eh in expected:
            match = None
            for s, sdt, edt in events:
                if (s or "").strip().lower() == exp_title.lower() and sdt and sdt.strftime("%Y-%m-%d") == exp_date:
                    match = (s, sdt, edt)
                    break
            record(f"GCal '{exp_title}' on {exp_date}", match is not None,
                   f"events: {[e[0] for e in events]}")
            if match:
                _, sdt, edt = match
                record(f"'{exp_title}' time {sh:02d}:00-{eh:02d}:00",
                       sdt.hour == sh and (edt is not None and edt.hour == eh),
                       f"start={sdt}, end={edt}")
    except Exception as e:
        record("GCal check error", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_emails():
    print("\n=== Check 4: Speaker Confirmation Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()

        expected_subject = "Conference Travel Confirmation"
        for sp in SPEAKERS:
            match = None
            for subj, to, body in emails:
                to_str = json.dumps(to).lower() if isinstance(to, (list, dict)) else str(to or "").lower()
                if (subj or "").strip().lower() == expected_subject.lower() and sp["email"].lower() in to_str:
                    match = (subj, to, body)
                    break
            record(f"Email to {sp['email']} with exact subject", match is not None,
                   f"emails: {[(e[0], e[1]) for e in emails[:5]]}")
            if match:
                body = (match[2] or "").lower()
                # Greeting includes name (use last name)
                last = sp["name"].split()[-1].lower()
                record(f"Email to {sp['email']} body greets {last}",
                       last in body, "")
                # Conference dates
                record(f"Email to {sp['email']} body mentions Mar 12 to 14",
                       ("march 12" in body and ("14" in body)) or "2026-03-12" in body,
                       "")
                # Train code
                record(f"Email to {sp['email']} body mentions train {sp['train']}",
                       sp["train"].lower() in body, "")
                # Departure/arrival times
                record(f"Email to {sp['email']} body mentions depart {sp['depart']}",
                       sp["depart"] in body, "")
                record(f"Email to {sp['email']} body mentions arrive {sp['arrive']}",
                       sp["arrive"] in body, "")
                # Price
                record(f"Email to {sp['email']} body mentions price {sp['price']}",
                       str(sp["price"]) in body or str(int(sp["price"])) in body,
                       "")
    except Exception as e:
        record("Email check error", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    all_passed = (FAIL_COUNT == 0)
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed")
    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "success": all_passed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
