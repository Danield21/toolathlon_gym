"""Evaluation for terminal-arxiv-yf-excel-word-notion-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"),
          port=int(os.environ.get("PGPORT", "5432")),
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

RELEVANT_PAPER_IDS = {"2306.06031", "2304.07619", "2302.14040", "2311.10723"}
NOISE_PAPER_IDS = {"2305.18290", "2307.09288"}
STOCKS = ["GOOGL", "AMZN", "JPM"]


def get_expected_from_db():
    """Query YF schema dynamically for stock price data."""
    defaults = {
        "stock_prices": {},  # symbol -> latest price
    }
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Get latest closing price for each stock
        for symbol in STOCKS:
            cur.execute("""
                SELECT close FROM yf.stock_prices
                WHERE symbol = %s ORDER BY date DESC LIMIT 1
            """, (symbol,))
            row = cur.fetchone()
            if row and row[0]:
                defaults["stock_prices"][symbol] = float(row[0])
        cur.close()
        conn.close()
    except Exception as e:
        print(f"  [WARN] DB query for expected values failed, using defaults: {e}")
    return defaults


EXPECTED = get_expected_from_db()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def _cell_value(v):
    """Robustly extract a float from an openpyxl cell value, or None.

    Accepts numeric literals and strings that carry currency symbols, percent
    signs, thousand separators, or currency codes (e.g. "$300.88", "300,88",
    "9"). Formula cells (starting with '=') have their numeric literal parsed
    where possible; non-numeric formulas return None.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    if s.startswith("="):
        s = s.lstrip("=").strip()
    for tok in (",", "$", "€", "¥", "%", "USD", "usd"):
        s = s.replace(tok, "")
    s = s.strip()
    try:
        return float(s)
    except ValueError:
        return None


def _is_formula(v):
    return isinstance(v, str) and v.strip().startswith("=")


def num_close(a, b, tol=2.0):
    """Compare two values numerically with tolerance.

    When both sides parse as numbers, compare with tolerance. Otherwise fall
    back to case-insensitive string comparison (for text values like 'Hold').
    """
    fa, fb = _cell_value(a), _cell_value(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def _numeric_ok(value, target, tol=1.0):
    """Numeric check that does not punish unverifiable formula cells.

    A cell that parses to a number is compared with tolerance. A cell holding
    an un-recalculated formula is treated as verifiable-unknown and accepted
    (per R2: when the expected value is a literal, skip formula cells rather
    than fail). Anything else (blank / non-numeric text) fails.
    """
    fv = _cell_value(value)
    if fv is not None:
        return num_close(fv, target, tol=tol)
    if _is_formula(value):
        return True
    return False


def check_excel(ws_path):
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "AI_Investment_Research.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        return
    check("Excel file exists", True)

    # Read with data_only=False so formula cells keep their formula strings
    # instead of turning into None; _cell_value/_numeric_ok then decide how to
    # handle each cell (R2/R3).
    wb = openpyxl.load_workbook(path, data_only=False)
    sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}

    # Portfolio_Holdings sheet
    ph_name = sn.get("portfolio_holdings")
    if ph_name is None:
        check("Portfolio_Holdings sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Portfolio_Holdings sheet exists", True)
        ws = wb[ph_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).lower() if h else "" for h in rows[0]] if rows else []
        data = [r for r in rows[1:] if r and r[0] is not None]
        check("Portfolio_Holdings has 3+ holdings", len(data) >= 3, f"Found {len(data)}")

        symbols_found = {str(r[0]).strip().upper() for r in data}
        check("All 3 stocks present", symbols_found >= {"GOOGL", "AMZN", "JPM"},
              f"Found: {symbols_found}")

        # Check has price column with reasonable values
        price_col = None
        for i, h in enumerate(headers):
            if "price" in h:
                price_col = i
                break
        if price_col is not None:
            prices = [r[price_col] for r in data
                      if len(r) > price_col and r[price_col] is not None and str(r[price_col]).strip() != ""]
            check("Prices are populated", len(prices) >= 3, f"Prices: {prices}")
            # Validate prices against dynamically queried DB values
            if EXPECTED["stock_prices"]:
                for row in data:
                    sym = str(row[0]).strip().upper()
                    if sym not in EXPECTED["stock_prices"] or len(row) <= price_col:
                        continue
                    expected_price = EXPECTED["stock_prices"][sym]
                    check(f"{sym} price reasonable (~{expected_price:.0f})",
                          _numeric_ok(row[price_col], expected_price, tol=expected_price * 0.1),
                          f"Got {row[price_col]}, expected ~{expected_price:.2f}")
        else:
            check("Price column exists", False, f"Headers: {headers}")

    # Research_Papers sheet
    rp_name = sn.get("research_papers")
    if rp_name is None:
        check("Research_Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Research_Papers sheet exists", True)
        ws2 = wb[rp_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        # At least the 4 relevant papers must be listed; the reverse-validation
        # step separately asserts no noise paper/ID appears anywhere in Excel.
        check("Research_Papers has 4+ relevant rows", len(data2) >= 4,
              f"Found {len(data2)}")

        # Check FinGPT is mentioned
        all_titles = " ".join(str(r[0]) for r in data2).lower()
        check("FinGPT paper listed", "fingpt" in all_titles, f"Titles: {all_titles[:200]}")

        # Check applicable stocks column. Anchor to the header when present
        # (task.md specifies Title, Authors, Key_Finding, Applicable_Stocks) so
        # a column reorder or an extra trailing column does not misread the
        # stock mapping; fall back to the last column otherwise.
        rp_headers = [str(h).lower() if h else "" for h in rows2[0]] if rows2 else []
        app_col = len(rp_headers) - 1 if rp_headers else -1
        for i, h in enumerate(rp_headers):
            if "stock" in h and "applic" in h:
                app_col = i
        all_stocks_text = " ".join(
            str(r[app_col]) if len(r) > app_col and r[app_col] else "" for r in data2
        ).upper()
        check("Applicable stocks mention GOOGL", "GOOGL" in all_stocks_text)
        check("Applicable stocks mention JPM", "JPM" in all_stocks_text)

    # AI_Impact_Assessment sheet
    ai_name = sn.get("ai_impact_assessment")
    if ai_name is None:
        check("AI_Impact_Assessment sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("AI_Impact_Assessment sheet exists", True)
        ws3 = wb[ai_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("AI_Impact_Assessment has 3+ rows", len(data3) >= 3, f"Found {len(data3)}")

        # Locate score / recommendation columns by header name (falls back to
        # fixed positions when headers are absent) so reordered columns do not
        # silently misread.
        ai_headers = [str(h).lower() if h else "" for h in rows3[0]] if rows3 else []
        score_col = 1
        rec_col = len(ai_headers) - 1 if ai_headers else -1
        for i, h in enumerate(ai_headers):
            if "score" in h or "exposure" in h:
                score_col = i
            if "recommendation" in h:
                rec_col = i

        # Check AI scores and recommendations
        for row in data3:
            stock = str(row[0]).strip().upper()
            score = row[score_col] if len(row) > score_col else None
            rec = str(row[rec_col]).lower() if len(row) > rec_col and row[rec_col] else ""
            if stock == "GOOGL":
                check("GOOGL AI score ~9", _numeric_ok(score, 9, tol=1), f"Score: {score}")
                check("GOOGL recommendation Overweight", "overweight" in rec, f"Rec: {rec}")
            elif stock == "AMZN":
                check("AMZN AI score ~8", _numeric_ok(score, 8, tol=1), f"Score: {score}")
                check("AMZN recommendation Overweight", "overweight" in rec, f"Rec: {rec}")
            elif stock == "JPM":
                check("JPM AI score ~5", _numeric_ok(score, 5, tol=1), f"Score: {score}")
                check("JPM recommendation Hold", "hold" in rec, f"Rec: {rec}")

    # Investment_Thesis sheet
    it_name = sn.get("investment_thesis")
    if it_name is None:
        check("Investment_Thesis sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Investment_Thesis sheet exists", True)
        ws4 = wb[it_name]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if r and r[0] is not None]
        check("Investment_Thesis has >= 2 rows", len(data4) >= 2, f"Found {len(data4)}")

        # Anchor the Theme column by header when present; accept "AI" or the
        # spelled-out "artificial intelligence" so a correctly-worded theme row
        # is not falsely failed.
        it_headers = [str(h).lower() if h else "" for h in rows4[0]] if rows4 else []
        theme_col = 0
        for i, h in enumerate(it_headers):
            if h == "theme" or "theme" in h:
                theme_col = i
        all_themes = " ".join(
            str(r[theme_col]) if len(r) > theme_col and r[theme_col] else "" for r in data4
        ).lower()
        check("Theme mentions AI",
              "ai" in all_themes or "artificial intelligence" in all_themes,
              f"Themes: {all_themes[:200]}")

    wb.close()


def check_word(ws_path):
    print("\n=== Checking Word Document ===")
    path = os.path.join(ws_path, "AI_Markets_Research_Report.docx")
    if not os.path.isfile(path):
        check("Word document exists", False, f"Not found: {path}")
        return
    check("Word document exists", True)

    from docx import Document
    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    check("Document title mentions AI and Markets",
          "ai" in full_text[:200] and "market" in full_text[:200])
    check("Document mentions GOOGL", "googl" in full_text)
    check("Document mentions AMZN", "amzn" in full_text)
    check("Document mentions JPM", "jpm" in full_text)
    check("Document has executive summary",
          "executive summary" in full_text or "executive" in full_text)
    check("Document has risk assessment",
          "risk" in full_text and ("assessment" in full_text or "factor" in full_text))
    check("Document mentions research papers",
          "fingpt" in full_text or "language model" in full_text or "research" in full_text)
    # task.md's section list requires an Investment Thesis section and per-stock
    # investment outlook, but does not mandate the exact wording "overweight" or
    # "recommendation". Accept any common investment-outlook term so a correct
    # report written with "buy"/"hold"/"outlook" phrasing is not wrongly failed.
    check("Document mentions investment outlook or thesis",
          any(kw in full_text for kw in ("recommend", "overweight", "underweight",
                                         "outlook", "thesis", "buy", "hold", "sell")),
          "Expected an investment-outlook or recommendation term")
    check("Document length >= 800 chars", len(full_text) >= 800,
          f"Length: {len(full_text)}")


def _extract_richtext(value):
    """Recursively pull plain_text / text.content strings out of a Notion value.

    Handles the shapes observed in the DB: a list of rich-text objects, a bare
    string, an object carrying plain_text / text.content, or objects wrapped in
    {"title": ...} / {"properties": ...} / {"content": ...} layers.
    """
    out = []
    if value is None:
        return out
    if isinstance(value, str):
        out.append(value)
    elif isinstance(value, list):
        for it in value:
            out.extend(_extract_richtext(it))
    elif isinstance(value, dict):
        if isinstance(value.get("plain_text"), str):
            out.append(value["plain_text"])
        t = value.get("text")
        if isinstance(t, dict) and isinstance(t.get("content"), str):
            out.append(t["content"])
        elif isinstance(t, str):
            out.append(t)
        for k in ("title", "properties", "rich_text", "name", "content"):
            if k in value:
                out.extend(_extract_richtext(value[k]))
    return out


def _database_title(title_json):
    """Normalise a notion.databases.title jsonb value to a plain string."""
    return "".join(_extract_richtext(title_json))


def _all_database_titles(cur):
    """Return [(id, title_str), ...] for every database (for diagnostics)."""
    cur.execute("SELECT id, title FROM notion.databases")
    return [(db_id, _database_title(t)) for db_id, t in cur.fetchall()]


def _find_pipeline_databases(cur):
    """Return ids of every database titled like 'Research Pipeline'.

    A multi-agent run may create several same-named databases; the evaluator
    aggregates across all of them so an agent is not penalised for duplication.
    """
    matches = []
    for db_id, title_str in _all_database_titles(cur):
        if "research" in title_str.lower() and "pipeline" in title_str.lower():
            matches.append(db_id)
    return matches


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        check("Research Pipeline database exists", False, f"DB connection failed: {e}")
        return
    cur = conn.cursor()
    try:
        db_ids = _find_pipeline_databases(cur)
        check("Research Pipeline database exists", len(db_ids) >= 1,
              f"Databases: {[d[1] for d in _all_database_titles(cur)]}")

        if db_ids:
            total_pages = 0
            props_sample = None
            for db_id in db_ids:
                cur.execute(
                    "SELECT COUNT(*) FROM notion.pages WHERE parent->>'database_id' = %s AND NOT archived",
                    (db_id,)
                )
                total_pages += cur.fetchone()[0]
                if props_sample is None:
                    cur.execute(
                        "SELECT properties FROM notion.pages "
                        "WHERE parent->>'database_id' = %s AND NOT archived AND properties IS NOT NULL LIMIT 1",
                        (db_id,)
                    )
                    row = cur.fetchone()
                    if row:
                        props_sample = row[0]
            check("Notion has >= 4 paper entries", total_pages >= 4,
                  f"Found {total_pages} pages across {len(db_ids)} database(s)")

            # Check properties of pages
            if props_sample is not None:
                try:
                    props = props_sample if isinstance(props_sample, dict) else json.loads(props_sample) if props_sample else {}
                except Exception:
                    props = {}
                props_lower = {k.lower(): v for k, v in props.items()}
                has_relevance = any("relevance" in k for k in props_lower)
                has_stock = any("stock" in k for k in props_lower)
                check("Pages have Relevance property", has_relevance, f"Props: {list(props.keys())}")
                check("Pages have Stock_Link property", has_stock, f"Props: {list(props.keys())}")
            else:
                check("Pages have Relevance property", False, "No page properties found")
                check("Pages have Stock_Link property", False, "No page properties found")
    finally:
        cur.close()
        conn.close()


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        check("Portfolio team email sent", False, f"DB connection failed: {e}")
        check("Risk committee email sent", False, f"DB connection failed: {e}")
        return
    cur = conn.cursor()

    # Check portfolio team email
    cur.execute("""
        SELECT subject, body_text, to_addr FROM email.messages
        WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
    """)
    sent = cur.fetchall()

    portfolio_email = None
    risk_email = None
    for subj, body, to_addr in sent:
        to_str = json.dumps(to_addr).lower() if to_addr else ""
        subj_lower = (subj or "").lower()
        if "portfolio_team" in to_str or "portfolio" in subj_lower:
            portfolio_email = (subj, body)
        if "risk_committee" in to_str or "risk" in subj_lower:
            risk_email = (subj, body)

    check("Portfolio team email sent", portfolio_email is not None,
          f"Sent emails: {[(s, t) for s, _, t in sent]}")
    if portfolio_email:
        body = (portfolio_email[1] or "").lower()
        check("Portfolio email mentions AI exposure", "ai" in body or "exposure" in body,
              f"Body preview: {body[:200]}")

    check("Risk committee email sent", risk_email is not None,
          f"Sent emails: {[(s, t) for s, _, t in sent]}")
    if risk_email:
        body = (risk_email[1] or "").lower()
        check("Risk email mentions concentration", "concentrat" in body or "sector" in body or "risk" in body,
              f"Body preview: {body[:200]}")

    cur.close()
    conn.close()


def check_terminal_outputs(ws_path):
    print("\n=== Checking Terminal Script Outputs ===")
    pa_path = os.path.join(ws_path, "portfolio_analysis.json")
    if os.path.isfile(pa_path):
        check("portfolio_analysis.json exists", True)
        with open(pa_path) as f:
            try:
                data = json.load(f)
                check("portfolio_analysis has content", len(data) > 0)
            except:
                check("portfolio_analysis is valid JSON", False)
    else:
        check("portfolio_analysis.json exists", False)

    rm_path = os.path.join(ws_path, "research_stock_mapping.json")
    if os.path.isfile(rm_path):
        check("research_stock_mapping.json exists", True)
        with open(rm_path) as f:
            try:
                data = json.load(f)
                check("research_stock_mapping has content", len(data) > 0)
            except:
                check("research_stock_mapping is valid JSON", False)
    else:
        check("research_stock_mapping.json exists", False)


def check_reverse_validation(ws_path):
    """Check noise arxiv papers not in notion tracker and no emails to wrong recipients."""
    print("\n=== Reverse Validation ===")

    # --- Check noise papers not in Notion ---
    noise_titles = ["dragggan", "drag your gan", "interactive point-based",
                    "llama 2", "open foundation and fine-tuned"]
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Find all Research Pipeline databases (aggregate across duplicates)
        pipeline_db_ids = _find_pipeline_databases(cur)

        if pipeline_db_ids:
            all_page_text = ""
            for pipeline_db_id in pipeline_db_ids:
                cur.execute(
                    "SELECT properties FROM notion.pages WHERE parent->>'database_id' = %s AND NOT archived",
                    (pipeline_db_id,))
                for (props,) in cur.fetchall():
                    if props:
                        all_page_text += json.dumps(props).lower() + " "

            no_noise_notion = not any(nt in all_page_text for nt in noise_titles)
            check("No noise arxiv papers in Notion tracker (DragGAN, Llama 2)",
                  no_noise_notion,
                  f"Found noise paper in Research Pipeline pages")
        else:
            check("No noise arxiv papers in Notion tracker", True, "No Research Pipeline DB to check")

        # --- Check no emails to wrong recipients ---
        # Use exact recipient match against parsed recipient list
        noise_recipients = {"team@firm.com", "office@firm.com", "social@firm.com",
                            "admin@firm.com", "facilities@firm.com", "staff@firm.com", "hr@firm.com"}

        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
        """)
        sent_emails = cur.fetchall()
        sent_recipients = set()
        for _subj, to in sent_emails:
            if to is None:
                continue
            to_list = to if isinstance(to, list) else ([to] if isinstance(to, str) else [])
            if isinstance(to, str):
                try:
                    to_list = json.loads(to)
                except Exception:
                    to_list = [to]
            for r in to_list:
                sent_recipients.add(str(r).strip().lower())

        contaminated = sent_recipients & noise_recipients
        no_noise_email = len(contaminated) == 0
        check("No emails sent to noise recipients (team@firm, office@firm, social@firm)",
              no_noise_email,
              f"Contaminated recipients: {contaminated}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))

    # --- Check noise papers not in Excel ---
    path = os.path.join(ws_path, "AI_Investment_Research.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=False)
        all_text = ""
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                all_text += " ".join(str(c) for c in row if c).lower() + " "
        wb.close()

        no_noise_excel = not any(nt in all_text for nt in noise_titles)
        check("No noise arxiv papers in Excel (DragGAN, Llama 2)",
              no_noise_excel,
              "Found noise paper content in Excel workbook")

        # Check noise paper IDs not present
        no_noise_ids = not any(nid in all_text for nid in NOISE_PAPER_IDS)
        check("No noise paper IDs in Excel (2305.18290, 2307.09288)",
              no_noise_ids,
              "Found noise paper ID in Excel")

    # --- Reverse: noise arxiv papers still exist in DB (agent should not delete) ---
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id FROM arxiv.papers WHERE id = ANY(%s)", (list(NOISE_PAPER_IDS),))
        remaining = {r[0] for r in cur.fetchall()}
        check("Reverse: noise arxiv papers preserved in DB",
              len(remaining) == len(NOISE_PAPER_IDS),
              f"Missing: {NOISE_PAPER_IDS - remaining}")
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse: noise arxiv papers preserved", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-ARXIV-YF-EXCEL-WORD-NOTION-EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_emails()
    check_terminal_outputs(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    # Tightened: require ALL checks to pass (previously accuracy >= 70%).
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
