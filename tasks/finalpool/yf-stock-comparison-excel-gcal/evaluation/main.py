"""
Evaluation script for yf-stock-comparison-excel-gcal task.

Checks:
1. Excel file Stock_Comparison_Report.xlsx - 2 sheets with correct structure and data
2. Google Calendar has investment review event
3. Email sent to analyst@investment.com
4. Word document Stock_Analysis_Report.docx exists
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_by_name(wb, name):
    for sname in wb.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            return [[cell.value for cell in row] for row in wb[sname].iter_rows()]
    return None


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Stock_Comparison_Report.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        return False

    agent_file = os.path.join(agent_workspace, "Stock_Comparison_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Stock_Comparison_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth file exists", False, f"Not found: {gt_file}")
        return False

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_ok = True

    # Check Price History sheet
    a_hist = load_sheet_by_name(agent_wb, "Price History")
    g_hist = load_sheet_by_name(gt_wb, "Price History")
    record("Sheet 'Price History' exists", a_hist is not None)

    if a_hist is not None:
        a_data = [r for r in a_hist[1:] if any(v is not None for v in r)]
        record("Price History has at least 18 rows (trading days)",
               len(a_data) >= 18,
               f"Found {len(a_data)} rows")
        if len(a_data) < 18:
            all_ok = False

        # Check headers
        if a_hist and len(a_hist) > 0:
            header = [str(v).strip().lower() if v else "" for v in a_hist[0]]
            record("Price History has Date column", any("date" in h for h in header))
            record("Price History has GOOGL column",
                   any("googl" in h for h in header))
            record("Price History has AMZN column",
                   any("amzn" in h for h in header))
            record("Price History has JPM column",
                   any("jpm" in h for h in header))

    # Check Performance Summary sheet
    a_summ = load_sheet_by_name(agent_wb, "Performance Summary")
    g_summ = load_sheet_by_name(gt_wb, "Performance Summary")
    record("Sheet 'Performance Summary' exists", a_summ is not None)

    if a_summ is not None and g_summ is not None:
        a_data = [r for r in a_summ[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_summ[1:] if any(v is not None for v in r)]
        record("Performance Summary has 3 rows (GOOGL, AMZN, JPM)",
               len(a_data) == 3,
               f"Found {len(a_data)} rows")

        # Build lookup by symbol
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().upper()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            sym = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(sym)
            if a_row is None:
                record(f"{sym} row in Performance Summary", False, "Not found")
                all_ok = False
                continue
            record(f"{sym} row exists", True)

            # Latest price (col 1) - tighten 5.0 -> 1.0
            if len(g_row) > 1 and len(a_row) > 1:
                ok = num_close(a_row[1], g_row[1], 1.0)
                record(f"{sym}: Latest_Price correct",
                       ok,
                       f"got {a_row[1]}, expected {g_row[1]} (tol=1.0)")
                if not ok:
                    all_ok = False
            # Month_Start_Price (col 2)
            if len(g_row) > 2 and len(a_row) > 2:
                ok = num_close(a_row[2], g_row[2], 1.0)
                record(f"{sym}: Month_Start_Price correct", ok,
                       f"got {a_row[2]}, expected {g_row[2]}")
                if not ok:
                    all_ok = False
            # Price_Change (col 3)
            if len(g_row) > 3 and len(a_row) > 3:
                ok = num_close(a_row[3], g_row[3], 1.0)
                record(f"{sym}: Price_Change correct", ok,
                       f"got {a_row[3]}, expected {g_row[3]}")
                if not ok:
                    all_ok = False
            # Return_Pct (col 4) - tighten 2.0 -> 0.5
            if len(g_row) > 4 and len(a_row) > 4:
                ok = num_close(a_row[4], g_row[4], 0.5)
                record(f"{sym}: Return_Pct correct",
                       ok,
                       f"got {a_row[4]}, expected {g_row[4]} (tol=0.5)")
                if not ok:
                    all_ok = False
            # Min_Price (col 5)
            if len(g_row) > 5 and len(a_row) > 5:
                ok = num_close(a_row[5], g_row[5], 1.0)
                record(f"{sym}: Min_Price correct", ok,
                       f"got {a_row[5]}, expected {g_row[5]}")
                if not ok:
                    all_ok = False
            # Max_Price (col 6)
            if len(g_row) > 6 and len(a_row) > 6:
                ok = num_close(a_row[6], g_row[6], 1.0)
                record(f"{sym}: Max_Price correct", ok,
                       f"got {a_row[6]}, expected {g_row[6]}")
                if not ok:
                    all_ok = False
            # Volatility (col 7) - REMOVED 'or True' bug; check value is positive numeric and matches GT
            if len(g_row) > 7 and len(a_row) > 7:
                ok = num_close(a_row[7], g_row[7], 1.0)
                record(f"{sym}: Volatility_Score correct", ok,
                       f"got {a_row[7]}, expected {g_row[7]} (tol=1.0)")
                if not ok:
                    all_ok = False

    return all_ok


# ============================================================================
# Check 2: Google Calendar
# ============================================================================

def check_gcal():
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")

    # Tighten: title must contain ALL: 'monthly', 'investment', 'portfolio', 'review' (case-insensitive)
    portfolio_events = [e for e in events
                        if e[0] and "monthly" in e[0].lower()
                        and "investment" in e[0].lower()
                        and "portfolio" in e[0].lower()
                        and "review" in e[0].lower()]
    ok_evt = len(portfolio_events) >= 1
    record("Event titled 'Monthly Investment Portfolio Review' exists",
           ok_evt,
           f"Events: {[e[0] for e in events[:5]]}")

    if not ok_evt:
        return False

    # Date 2026-03-18
    ev = portfolio_events[0]
    s = str(ev[1])
    ok_date = "2026-03-18" in s
    record("Event on 2026-03-18", ok_date, f"start: {s}")
    # Time 10am-11am: hour 10 in local or 14/15 in UTC
    ok_time = False
    try:
        if len(s) >= 13 and int(s[11:13]) in (10, 14, 15):
            ok_time = True
    except ValueError:
        pass
    record("Event starts at 10am (local) or UTC equivalent", ok_time, f"start: {s}")
    # Duration 60 min
    if ev[2] and ev[1]:
        try:
            dur = (ev[2] - ev[1]).total_seconds() / 60.0
            ok_dur = abs(dur - 60) <= 5
            record("Event duration is 60 minutes", ok_dur, f"got {dur} min")
        except Exception:
            pass
    return ok_evt and ok_date and ok_time


# ============================================================================
# Check 3: Email
# ============================================================================

def check_emails():
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Filter strictly by recipient analyst@investment.com (AND not OR)
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE to_addr::text ILIKE '%%analyst@investment.com%%'
    """)
    target_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(target_emails)} emails to analyst@investment.com.")
    found_email = len(target_emails) >= 1
    record("Email sent to analyst@investment.com", found_email,
           f"No email to analyst@investment.com found")

    if not found_email:
        return False

    ok_subj = False
    ok_body_sym = False
    ok_body_perf = False
    for subject, from_addr, to_addr, body_text in target_emails:
        sl = (subject or "").lower()
        bl = (body_text or "").lower()
        # Subject must be 'Stock Performance Report March 2026'
        if "stock performance report" in sl and "march 2026" in sl:
            ok_subj = True
        # Body should mention all 3 stock symbols (or names)
        sym_count = sum(1 for s in ["googl", "amzn", "jpm", "alphabet", "amazon", "morgan"] if s in bl)
        if sym_count >= 3:
            ok_body_sym = True
        # Body should mention best/worst with a return percentage near each
        # (regex-based proximity to avoid bare-substring FP)
        import re
        # Match e.g. "best ... 12.3%" or "highest return: 12%"
        best_pct = bool(re.search(r"(best|highest)[^\n]{0,120}\d+(?:\.\d+)?\s*%", bl))
        worst_pct = bool(re.search(r"(worst|lowest)[^\n]{0,120}\d+(?:\.\d+)?\s*%", bl))
        if best_pct and worst_pct:
            ok_body_perf = True
    record("Email subject 'Stock Performance Report March 2026'",
           ok_subj, f"Subjects: {[e[0] for e in target_emails]}")
    record("Email body mentions all 3 stock symbols", ok_body_sym)
    record("Email body mentions best/worst + return percentage", ok_body_perf)

    return ok_subj and ok_body_sym and ok_body_perf


# ============================================================================
# Check 4: Word document
# ============================================================================

def check_word(agent_workspace):
    print("\n=== Checking Stock_Analysis_Report.docx ===")

    docx_path = os.path.join(agent_workspace, "Stock_Analysis_Report.docx")
    if not os.path.isfile(docx_path):
        record("Word file exists", False, f"Not found: {docx_path}")
        return False
    record("Word file exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        record("Word doc has content", len(all_text.strip()) >= 100,
               f"Content length: {len(all_text.strip())}")
        # Heading 'Stock Comparison Analysis March 2026'
        record("Word doc has heading 'Stock Comparison Analysis March 2026'",
               "stock comparison analysis" in all_text and "march 2026" in all_text,
               f"Sample: {all_text[:200]}")
        # Should mention all 3 stocks
        ok_3 = sum(1 for s in ["googl", "amzn", "jpm", "alphabet", "amazon", "morgan"] if s in all_text) >= 3
        record("Word doc mentions all 3 stocks (GOOGL/AMZN/JPM)",
               ok_3,
               f"Sample: {all_text[:200]}")
        return (ok_3 and "stock comparison analysis" in all_text
                and "march 2026" in all_text)
    except ImportError:
        size = os.path.getsize(docx_path)
        record("Word file has content (>2KB)", size > 2000, f"Size: {size} bytes")
        return size > 2000
    except Exception as e:
        record("Word file readable", False, str(e))
        return False


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    gcal_ok = check_gcal()
    email_ok = check_emails()
    word_ok = check_word(args.agent_workspace)

    all_passed = excel_ok and gcal_ok and email_ok and word_ok

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
