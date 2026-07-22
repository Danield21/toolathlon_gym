"""Evaluation for sf-hr-department-budget."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace, gt_dir):
    """Check Budget_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Budget_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Budget_Analysis.xlsx")

    if not os.path.exists(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_ok = True

    # Sheet 1: Department Budget
    a_rows = load_sheet_rows(agent_wb, "Department Budget")
    g_rows = load_sheet_rows(gt_wb, "Department Budget")

    if a_rows is None:
        record("Sheet 'Department Budget' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Department Budget' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        record("Department Budget row count", len(a_data) == len(g_data),
               f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Department '{g_row[0]}' present", False, "Missing")
                all_ok = False
                continue

            errors = []
            # Budget (col 1) — exact dollar amount expected
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 100):
                    errors.append(f"Budget: {a_row[1]} vs {g_row[1]}")

            # Actual_Spend (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 100):
                    errors.append(f"Actual_Spend: {a_row[2]} vs {g_row[2]}")

            # Variance (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 100):
                    errors.append(f"Variance: {a_row[3]} vs {g_row[3]}")

            # Variance_Pct (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"Variance_Pct: {a_row[4]} vs {g_row[4]}")

            # Employee_Count (col 5) — exact integer
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0):
                    errors.append(f"Employee_Count: {a_row[5]} vs {g_row[5]}")

            # Avg_Salary (col 6)
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 10):
                    errors.append(f"Avg_Salary: {a_row[6]} vs {g_row[6]}")

            if errors:
                record(f"Department '{g_row[0]}' data", False, "; ".join(errors))
                all_ok = False
            else:
                record(f"Department '{g_row[0]}' data", True)

    # Sheet 2: Summary
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    if a_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary: {g_row[0]} present", False, "Missing")
                all_ok = False
                continue

            # Large values need slightly larger tolerance for cumulative rounding
            if "depts" in key or "over" in key or "under" in key:
                tol = 0  # exact count
            elif "budget" in key or "spend" in key or ("variance" in key and "pct" not in key):
                tol = 200  # cumulative rounding allowed
            else:
                tol = 0.5

            if len(a_row) > 1 and len(g_row) > 1:
                ok = num_close(a_row[1], g_row[1], tol)
                record(f"Summary: {g_row[0]}", ok,
                       f"Expected {g_row[1]}, got {a_row[1]} (tol={tol})")
                if not ok:
                    all_ok = False

    return all_ok


def check_gsheet():
    """Check Google Sheet was created with title 'Department Budget Overview'."""
    print("\n=== Checking Google Sheet ===")

    try:
        conn = psycopg2.connect(**{"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym",
                                   "user": "eigent", "password": "camel"})
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gsheet.spreadsheets
            WHERE LOWER(title) = LOWER('Department Budget Overview')
               OR LOWER(title) LIKE '%department budget overview%'
        """)
        sheets = cur.fetchall()
        record("Google Sheet titled 'Department Budget Overview' exists",
               len(sheets) >= 1, f"got {len(sheets)} matching sheets")
        if not sheets:
            cur.close(); conn.close()
            return False

        sid = sheets[0][0]
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
        cell_count = cur.fetchone()[0]
        record("Google Sheet has substantive data (>=20 cells)",
               cell_count >= 20, f"Cell count: {cell_count}")

        # Verify a 'Summary' sheet/tab exists with department data
        cur.execute("""
            SELECT s.title FROM gsheet.sheets s
            WHERE s.spreadsheet_id = %s
              AND LOWER(s.title) = 'summary'
        """, (sid,))
        has_summary = cur.fetchone() is not None
        record("Google Sheet has a 'Summary' tab",
               has_summary, "tab not found")

        cur.close(); conn.close()
        return cell_count >= 20
    except Exception as e:
        record("Google Sheet check", False, str(e))
        return False


def check_emails():
    """Check email was sent to CFO with required content."""
    print("\n=== Checking Emails ===")

    try:
        conn = psycopg2.connect(**{"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym",
                                   "user": "eigent", "password": "camel"})
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, COALESCE(body_text, body_html, '') FROM email.messages")
        all_emails = cur.fetchall()
        cur.close()
        conn.close()

        # Required exact subject per task: "Department Budget Analysis"
        target_subj = "department budget analysis"
        target_to = "cfo@company.example.com"
        matched = None
        for subject, from_addr, to_addr, body in all_emails:
            subj_lower = (subject or "").lower().strip()
            to_str = str(to_addr or "").lower()
            if target_subj in subj_lower and target_to in to_str:
                matched = (subject, to_addr, body)
                break
        record(f"Email with subject 'Department Budget Analysis' sent to {target_to}",
               matched is not None, f"checked {len(all_emails)} emails")
        if matched is None:
            return False
        body_lower = (matched[2] or "").lower()
        # Must mention required totals (total budget / total spend / overall variance)
        topic_terms = ["total budget", "total spend", "variance"]
        missing = [t for t in topic_terms if t not in body_lower]
        record("Email body includes total budget / total spend / variance",
               len(missing) == 0, f"missing: {missing}")
        # Must mention at least one over-budget department (numeric amount or dept name)
        record("Email body mentions over-budget department(s)",
               "over budget" in body_lower or "exceeded" in body_lower
               or "over-budget" in body_lower,
               "missing 'over budget'/'exceeded' term")
        return True
    except Exception as e:
        record("Email check", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    gsheet_ok = check_gsheet()
    emails_ok = check_emails()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
