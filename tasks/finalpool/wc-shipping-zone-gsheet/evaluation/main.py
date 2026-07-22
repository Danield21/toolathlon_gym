"""Evaluation for wc-shipping-zone-gsheet."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_zones():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
    SELECT
      CASE
          WHEN billing->>'state' = 'CA' AND billing->>'country' = 'US' THEN 'California'
          WHEN billing->>'country' = 'US' THEN 'Domestic US'
          ELSE 'International'
      END as zone,
      COUNT(*),
      ROUND(SUM(total::numeric), 2),
      ROUND(AVG(total::numeric), 2)
    FROM wc.orders
    WHERE status IN ('completed', 'processing')
    GROUP BY 1
    """)
    by_zone = {z: {"count": int(cnt), "rev": float(rev), "aov": float(aov)} for z, cnt, rev, aov in cur.fetchall()}
    cur.close()
    conn.close()
    # All 3 zones must appear (zero-filled if missing)
    for z in ("California", "Domestic US", "International"):
        if z not in by_zone:
            by_zone[z] = {"count": 0, "rev": 0.0, "aov": 0.0}
    return by_zone


def get_expected_methods():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
    SELECT z.name AS zone, m.method_id AS method, m.method_title AS title, m.settings
    FROM wc.shipping_zone_methods m
    JOIN wc.shipping_zones z ON z.id = m.zone_id
    """)
    methods = []
    for zone, mid, title, settings in cur.fetchall():
        cost = None
        if isinstance(settings, dict):
            cost = settings.get("cost") or settings.get("min_amount")
        methods.append({"zone": zone, "method": mid, "title": title, "cost": cost})
    cur.close()
    conn.close()
    return methods


def check_excel(args):
    print("\n=== Checking Excel Output ===")
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_file = os.path.join(args.agent_workspace, "Shipping_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Shipping_Analysis.xlsx")

    check("Agent Excel exists", os.path.exists(agent_file), f"Expected {agent_file}")
    if not os.path.exists(agent_file):
        return
    if not os.path.exists(gt_file):
        check("GT Excel exists", False, f"Expected {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    expected_zones = get_expected_zones()

    # Zone Summary
    a_rows = load_sheet_rows(agent_wb, "Zone Summary")
    g_rows = load_sheet_rows(gt_wb, "Zone Summary")
    check("Sheet 'Zone Summary' present", a_rows is not None)
    if a_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        check("Zone Summary has 3 rows (incl. zero-rows)", len(a_data) == 3, f"got {len(a_data)}")

        # Sort order: alphabetical by Zone_Name
        zones_seen = [str(r[0]).strip() for r in a_data]
        check("Zone Summary sorted alphabetically", zones_seen == sorted(zones_seen),
              f"got {zones_seen}")

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for zone, exp in expected_zones.items():
            row = a_lookup.get(zone.lower())
            check(f"Zone '{zone}' present", row is not None)
            if row is None:
                continue
            check(f"  {zone}.Order_Count == {exp['count']}",
                  num_close(row[1] if len(row) > 1 else None, exp["count"], 0),
                  f"got {row[1] if len(row) > 1 else None}")
            check(f"  {zone}.Total_Revenue ≈ {exp['rev']}",
                  num_close(row[2] if len(row) > 2 else None, exp["rev"], 1.0),
                  f"got {row[2] if len(row) > 2 else None}")
            check(f"  {zone}.Avg_Order_Value ≈ {exp['aov']}",
                  num_close(row[3] if len(row) > 3 else None, exp["aov"], 0.01),
                  f"got {row[3] if len(row) > 3 else None}")

    # Methods sheet - per-row content verification
    a_rows = load_sheet_rows(agent_wb, "Methods")
    check("Sheet 'Methods' present", a_rows is not None)
    if a_rows is not None:
        a_data = [r for r in a_rows[1:] if r and any(c is not None for c in r)]
        expected_methods = get_expected_methods()
        check(f"Methods has at least {len(expected_methods)} rows",
              len(a_data) >= len(expected_methods),
              f"got {len(a_data)}, expected {len(expected_methods)}")
        # Per-method check: each expected method must be present
        # by (zone, method, title) tuple
        for em in expected_methods:
            matched = False
            for ar in a_data:
                if not ar or len(ar) < 3:
                    continue
                z = str(ar[0] or "").strip().lower()
                m = str(ar[1] or "").strip().lower()
                d = str(ar[2] or "").strip().lower()
                if z == em["zone"].lower() and m == em["method"].lower() and em["title"].lower() in d:
                    matched = True
                    break
            check(f"Methods row for {em['zone']}/{em['method']}/{em['title']} present",
                  matched, "missing")


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%%shipping zone%%dashboard%%' OR LOWER(title) LIKE '%%shipping zone dashboard%%'")
    sheets = cur.fetchall()
    check("Spreadsheet 'Shipping Zone Dashboard' exists", len(sheets) >= 1, f"Found {len(sheets)}")
    if not sheets:
        cur.close()
        conn.close()
        return
    sid = sheets[0][0]
    cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id=%s", (sid,))
    sub = cur.fetchall()
    sheet_id = next((s_id for s_id, t in sub if t and t.strip().lower() == "summary"), None)
    check("Sheet 'Summary' present", sheet_id is not None, f"got {[t for _, t in sub]}")
    if sheet_id is None:
        cur.close()
        conn.close()
        return
    cur.execute("SELECT row_index, col_index, value FROM gsheet.cells WHERE sheet_id=%s ORDER BY row_index, col_index", (sheet_id,))
    cells = cur.fetchall()
    grid = {}
    for r, c, v in cells:
        grid.setdefault(r, {})[c] = str(v).strip() if v is not None else ""
    expected_zones = get_expected_zones()
    header_row = None
    for r in sorted(grid.keys()):
        joined = " ".join(grid[r].values()).lower()
        if "zone" in joined and ("count" in joined or "order" in joined) and "revenue" in joined:
            header_row = r
            break
    check("GSheet header row present", header_row is not None)
    if header_row is None:
        cur.close()
        conn.close()
        return

    col_map = {}
    for col, val in grid[header_row].items():
        v = val.strip().lower().replace(" ", "_")
        if v == "zone_name":
            col_map["zone"] = col
        elif v == "order_count":
            col_map["count"] = col
        elif v == "total_revenue":
            col_map["rev"] = col
        elif v == "avg_order_value":
            col_map["aov"] = col

    data_rows = [grid[r] for r in sorted(grid.keys()) if r > header_row]
    by_zone = {}
    for row in data_rows:
        z = row.get(col_map.get("zone", -1), "").strip()
        if z:
            by_zone[z.lower()] = row

    for zone, exp in expected_zones.items():
        row = by_zone.get(zone.lower())
        check(f"GSheet zone '{zone}' present", row is not None)
        if row is None:
            continue
        try:
            cnt = int(float(row.get(col_map.get("count", -1), "0")))
        except (TypeError, ValueError):
            cnt = None
        try:
            rev = float(row.get(col_map.get("rev", -1), "0"))
        except (TypeError, ValueError):
            rev = None
        try:
            aov = float(row.get(col_map.get("aov", -1), "0"))
        except (TypeError, ValueError):
            aov = None
        check(f"GSheet {zone}.Order_Count == {exp['count']}", cnt == exp["count"], f"got {cnt}")
        check(f"GSheet {zone}.Total_Revenue ≈ {exp['rev']}",
              rev is not None and abs(rev - exp["rev"]) <= 1.0, f"got {rev}")
        check(f"GSheet {zone}.Avg_Order_Value ≈ {exp['aov']}",
              aov is not None and abs(aov - exp["aov"]) <= 0.01, f"got {aov}")
    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args)
    check_gsheet()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
