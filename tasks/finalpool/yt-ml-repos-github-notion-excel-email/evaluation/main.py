"""
Evaluation for yt-ml-repos-github-notion-excel-email task.

Checks:
1. ML_Research_Tracker.xlsx exists
2. "Videos" sheet has >= 5 data rows, has Video_ID and Title columns
3. "Papers" sheet has >= 5 data rows, has ArXiv_ID and Title columns
4. "Summary" sheet has >= 5 rows
5. Notion page 'ML Tech Research Hub' exists + 'Research Items' database (Type/Status selects, >= 7 entries)
6. Email sent to research@lab.edu (subject references the catalog; body mentions videos / tracker / knowledge base)

GT xlsx comparison is key-based (order-independent) and tolerates a partial
subset of rows so that a fully-correct run still PASSes when a multi-agent
harness aggregates the same file from different writers.
"""
import os
import sys
import json
import re
from argparse import ArgumentParser
from datetime import datetime, date

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(v):
    """Robust numeric parsing: str/int/float/None; strip %, currency, thousands separators."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "").replace("$", "").replace("¥", "").replace("€", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def num_close(a, b, tol=1.0):
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # Fall back to case-insensitive string equality only when a side is unparseable.
    return str_match(a, b)


def _norm_text(v):
    """Normalize a text value for tolerant comparison: lowercase, collapse
    internal whitespace, strip, and drop trailing sentence punctuation. This
    keeps exact-string comparisons from failing on harmless presentation
    differences such as an accidental trailing period."""
    s = str(v).strip().lower()
    s = re.sub(r"\s+", " ", s)
    while s and s[-1] in ".,;:!?)]}'\"":
        s = s[:-1]
    return s


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return _norm_text(a) == _norm_text(b)


def _norm_alpha(s):
    """Lowercase alphanumeric-only view of a string (separator-insensitive)."""
    return re.sub(r"[^a-z0-9]+", "", _norm_text(s))


def _split_author_names(v):
    """Split an author string into a set of name tokens. Accepts comma,
    semicolon, ampersand, newline and 'and' separators so that 'A, B, and C',
    'A; B; C', 'A\\nB\\nC' all parse to the same set as 'A, B, C'."""
    if v is None:
        return set()
    parts = re.split(r"[,;&\n]|\s+and\s+", str(v), flags=re.IGNORECASE)
    names = set()
    for p in parts:
        p = p.strip().strip(".'\"").strip()
        if p:
            names.add(p.lower())
    return names


def _authors_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    na, nb = _split_author_names(a), _split_author_names(b)
    if na and nb:
        return na == nb
    return _norm_text(a) == _norm_text(b)


def _url_norm(v):
    if v is None:
        return None
    s = str(v).strip().strip("\"'`")
    while s and s[-1] in ".,;:!?)]}'\"":
        s = s[:-1]
    while s and s[-1] == "/":
        s = s[:-1]
    s = re.sub(r"^[a-z][a-z0-9+.-]*://", "", s)
    if s.startswith("www."):
        s = s[4:]
    return s.lower()


def _url_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    na, nb = _url_norm(a), _url_norm(b)
    if na is None or nb is None:
        return _norm_text(a) == _norm_text(b)
    if na == nb:
        return True
    # Tolerate case / separator differences (e.g. 'FlashAttention' vs
    # 'flash-attention') without accepting a different repository.
    return _norm_alpha(na) == _norm_alpha(nb)


# Topic field: the task's topic_mapping.json is the canonical list, but a
# correct model may write a title-derived variant (e.g. 'FlashAttention',
# 'Mixtral', 'Stable Diffusion') instead of the mapping string. Accept the
# canonical name plus its recognized variants, separator-insensitively.
_TOPIC_ALIASES = {
    "mixtral": "mixture of experts",
    "moe": "mixture of experts",
    "stable diffusion": "diffusion models",
    "stablediffusion": "diffusion models",
    "pgvector": "vector databases",
    "vectordb": "vector databases",
    "vector db": "vector databases",
}


def _topic_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    na, nb = _norm_alpha(a), _norm_alpha(b)
    if not na or not nb:
        return _norm_text(a) == _norm_text(b)
    if na == nb:
        return True
    for cand in (na, nb):
        alias = _TOPIC_ALIASES.get(cand)
        if alias and _norm_alpha(alias) in (na, nb):
            return True
    return False


def _cell_value(v):
    """Interpret a spreadsheet cell (literals expected). Formula cells fall back to
    their encoded literal so a stray '=285000' still compares as 285000."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s.startswith("="):
        parsed = _to_float(s[1:])
        return parsed if parsed is not None else s
    return v


def _date_norm(v):
    """Normalize a date-ish value to YYYY-MM-DD for comparison."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^\d{4}-\d{2}-\d{2}", s)
    return m.group(0) if m else s


def _row_match(a_row, g_row, headers_lower):
    """Compare one agent row against one GT row. GT values drive which checks run;
    extra columns on the agent side are ignored."""
    for ci in range(min(len(g_row), len(a_row))):
        gv, av = _cell_value(g_row[ci]), _cell_value(a_row[ci])
        if gv is None:
            continue
        h = (headers_lower[ci] if ci < len(headers_lower) else "").lower()
        if "date" in h or "published" in h:
            if _date_norm(av).lower() != _date_norm(gv).lower():
                return False
        elif "view_count" in h:
            # View counts come straight from the API; a small relative tolerance is fine.
            if not num_close(av, gv, max(abs(_to_float(gv) or 0.0) * 0.01, 1.0)):
                return False
        elif isinstance(gv, (int, float)) or "count" in h:
            # Count columns must match exactly.
            fa, fg = _to_float(av), _to_float(gv)
            if fa is None or fg is None or fa != fg:
                return False
        elif "author" in h:
            # Author lists: accept 'A, B, and C' / 'A; B; C' / 'A and B' spellings.
            if not _authors_match(av, gv):
                return False
        elif "topic" in h:
            # Topic column: tolerant to title-derived variants of the mapping.
            if not _topic_match(av, gv):
                return False
        elif "github" in h or "url" in h:
            # Repository URL: tolerate trailing punctuation, scheme, 'www.' and case.
            if not _url_match(av, gv):
                return False
        else:
            if not str_match(av, gv):
                return False
    return True


def _row_key(row):
    """Stable, order-independent key: normalized first column
    (Video_ID / ArXiv_ID / Topic)."""
    if not row or row[0] is None:
        return ""
    return re.sub(r"\s+", "", str(row[0]).strip().lower())


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-4: ML_Research_Tracker.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "ML_Research_Tracker.xlsx")
    if not os.path.exists(xlsx_path):
        record("ML_Research_Tracker.xlsx exists", False, f"Not found at {xlsx_path}")
        for msg in ["Videos sheet has >= 5 rows with Video_ID and Title",
                    "Papers sheet has >= 5 rows with ArXiv_ID and Title",
                    "Summary sheet has >= 5 rows"]:
            record(msg, False, "File missing")
        return
    record("ML_Research_Tracker.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Videos sheet — at least 5 data rows (7 when fully aggregated); precise headers
    videos_key = sheet_names_lower.get("videos") or next((sheet_names_lower[k] for k in sheet_names_lower if "video" in k), None)
    if not videos_key:
        record("Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        ws = wb[videos_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        headers = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]] if rows else []
        expected_h = ["video_id", "title", "topic", "github_url", "published_date", "view_count"]
        for i, eh in enumerate(expected_h):
            ah = headers[i] if i < len(headers) else "MISSING"
            record(f"Videos header[{i}] = '{eh}'", ah == eh, f"Got '{ah}'")
        record("Videos sheet has >= 5 data rows",
               len(data_rows) >= 5,
               f"Rows: {len(data_rows)}")

    # Papers sheet — at least 5 rows
    papers_key = sheet_names_lower.get("papers") or next((sheet_names_lower[k] for k in sheet_names_lower if "paper" in k), None)
    if not papers_key:
        record("Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        ws2 = wb[papers_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        headers2 = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows2[0]] if rows2 else []
        expected_h2 = ["arxiv_id", "title", "authors", "published", "topic", "related_video_id"]
        for i, eh in enumerate(expected_h2):
            ah = headers2[i] if i < len(headers2) else "MISSING"
            record(f"Papers header[{i}] = '{eh}'", ah == eh, f"Got '{ah}'")
        record("Papers sheet has >= 5 data rows",
               len(data_rows2) >= 5,
               f"Rows: {len(data_rows2)}")

    # Summary sheet — at least 5 rows (7 when fully aggregated)
    summary_key = sheet_names_lower.get("summary") or next((sheet_names_lower[k] for k in sheet_names_lower if "summar" in k), None)
    if not summary_key:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        ws3 = wb[summary_key]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)] if rows3 else []
        headers3 = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows3[0]] if rows3 else []
        expected_h3 = ["topic", "video_count", "paper_count", "github_repos_count"]
        for i, eh in enumerate(expected_h3):
            ah = headers3[i] if i < len(headers3) else "MISSING"
            record(f"Summary header[{i}] = '{eh}'", ah == eh, f"Got '{ah}'")
        record("Summary sheet has >= 5 data rows",
               len(data_rows3) >= 5,
               f"Rows: {len(data_rows3)}")

    # --- Groundtruth XLSX value comparison (key-based, order-independent) ---
    gt_path = os.path.join(groundtruth_workspace, "ML_Research_Tracker.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=False)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_headers = [str(c).strip().lower().replace(" ", "_") if c else ""
                         for c in next(a_ws.iter_rows(values_only=True))] if a_ws.max_row else []

            # Group GT rows by normalized key so agent rows can be matched
            # irrespective of the order they were written in.
            gt_by_key = {}
            for gi, g in enumerate(gt_rows):
                gt_by_key.setdefault(_row_key(g), []).append((gi, g))

            matched = 0
            unmatched = []
            used = set()  # GT row indices already claimed
            for a in a_rows:
                ka = _row_key(a)
                found = False
                # 1) exact normalized-key match (order-independent)
                for gi, g in gt_by_key.get(ka, []):
                    if gi in used:
                        continue
                    if _row_match(a, g, a_headers):
                        used.add(gi)
                        found = True
                        matched += 1
                        break
                # 2) tolerant fallback: e.g. Summary topic 'Mixtral' is a valid
                #    variant of the GT topic 'Mixture of Experts', whose normalized
                #    key differs. _row_match itself compares every column (incl. the
                #    key column), so this cannot accept a wrong identifier row.
                if not found:
                    for gi, g in enumerate(gt_rows):
                        if gi in used:
                            continue
                        if _row_match(a, g, a_headers):
                            used.add(gi)
                            found = True
                            matched += 1
                            break
                if not found:
                    unmatched.append(f"key={ka!r} values={a[:5]}")
            record(f"GT '{gt_sname}': all agent rows match ground truth (key-based)",
                   matched == len(a_rows) and len(a_rows) > 0,
                   f"{matched}/{len(a_rows)} matched" + (f"; e.g. {unmatched[0]}" if unmatched else ""))
        gt_wb.close()


def _richtext_tokens(v, out):
    """Recursively collect every text token (plain_text / content / bare string)
    from a Notion rich-text value, whatever shape it takes."""
    if isinstance(v, str):
        out.append(v)
    elif isinstance(v, list):
        for it in v:
            _richtext_tokens(it, out)
    elif isinstance(v, dict):
        if isinstance(v.get("plain_text"), str):
            out.append(v["plain_text"])
        if isinstance(v.get("content"), str):
            out.append(v["content"])
        for val in v.values():
            _richtext_tokens(val, out)


def _page_title_text(props):
    """Robustly pull the title text out of a page's properties JSON. Handles the
    standard Notion shape (a property object typed 'title'), a 'title'/'Name' key
    whose value is a rich-text array / object / string, a 'properties' wrapper,
    and finally a full scan of every text token in the properties."""
    if isinstance(props, dict) and isinstance(props.get("properties"), dict):
        props = props["properties"]
    if not isinstance(props, dict):
        return ""
    for v in props.values():
        if isinstance(v, dict) and v.get("type") == "title":
            out = []
            _richtext_tokens(v, out)
            t = "".join(out).strip()
            if t:
                return t
    for k, v in props.items():
        if str(k).strip().lower() in ("title", "name"):
            out = []
            _richtext_tokens(v, out)
            t = "".join(out).strip()
            if t:
                return t
    out = []
    _richtext_tokens(props, out)
    return "".join(out)


def check_notion():
    print("\n=== Check 5: Notion 'ML Tech Research Hub' page + 'Research Items' database ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
        pages = cur.fetchall()

        target_title = "ml tech research hub"
        page_id = None
        for pid, props in pages:
            name = _page_title_text(props).strip().lower()
            if target_title in name:
                page_id = pid
                break
        record("Notion page 'ML Tech Research Hub' exists", page_id is not None,
               "Not found")

        # Database 'Research Items'
        cur.execute("SELECT id, title, properties FROM notion.databases WHERE archived = false")
        dbs = cur.fetchall()
        target_db = None
        for did, title, dprops in dbs:
            out = []
            _richtext_tokens(title, out)
            title_text = "".join(out).strip().lower()
            if title_text == "research items":
                target_db = (did, dprops)
                break
        record("Notion database 'Research Items' exists",
               target_db is not None,
               "Not found")
        if target_db:
            db_id, dprops = target_db
            # Validate the database has a select property with Video/Paper options
            # (the 'Type' field) and one with New/In Review/Archived (the 'Status'
            # field). The property key name is not enforced so that a correct model
            # that names them slightly differently is not penalized.
            if isinstance(dprops, dict):
                has_type_select = False
                has_status_select = False
                for v in dprops.values():
                    if not isinstance(v, dict):
                        continue
                    if v.get("type") != "select":
                        continue
                    opts = v.get("select", {}).get("options", [])
                    names = {str(o.get("name", "")).lower() for o in opts if isinstance(o, dict)}
                    if "video" in names and "paper" in names:
                        has_type_select = True
                    if "new" in names and "in review" in names and "archived" in names:
                        has_status_select = True
                record("Research Items has 'Type' select with Video/Paper", has_type_select)
                record("Research Items has 'Status' select with New/In Review/Archived", has_status_select)

            # 7+ entries (videos) plus paper entries
            cur.execute("SELECT COUNT(*) FROM notion.pages WHERE archived = false AND parent->>'database_id' = %s", (db_id,))
            n = cur.fetchone()[0]
            record("Research Items database has >= 7 entries", n >= 7, f"Got {n}")
        cur.close()
        conn.close()
    except Exception as e:
        record("Notion check", False, str(e))


def _email_subject_ok(subj):
    """Does the subject reference the ML research catalog? Accepts 'catalog',
    'tracker', or an ML research/summary phrasing without demanding a literal 'ml'
    token next to 'catalog'/'tracker'."""
    s = (subj or "").strip().lower()
    return ("catalog" in s or "tracker" in s
            or ("ml" in s and ("research" in s or "summary" in s)))


def _score_candidate(subj, body, fname):
    bl = (body or "").lower()
    score = 0
    if "ml_research_tracker" in bl or "research_tracker" in bl or "tracker.xlsx" in bl:
        score += 2
    if "sent" in (fname or "").lower():
        score += 1
    return score


def _pick_deliverable(candidates):
    """Pick the deliverable email from (subj, body, folder) candidates. Prefer
    the higher score (body naming the tracker, folder Sent); on a tie keep the
    LATEST candidate (candidates arrive in ascending message-id order), because
    the final deliverable is sent after any earlier progress/test note."""
    match = None
    best_score = -1
    for subj, body, fname in candidates:
        score = _score_candidate(subj, body, fname)
        if score >= best_score:
            best_score = score
            match = (subj, body)
    return match


def check_email():
    print("\n=== Check 6: Email to research@lab.edu (subject + body) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT m.subject, m.to_addr, COALESCE(m.body_text, m.body_html, ''),
                   m.from_addr, f.name
            FROM email.messages m
            JOIN email.folders f ON f.id = m.folder_id
            ORDER BY m.id
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()

        target_to = "research@lab.edu"
        # Only consider mail the agent actually sent. The preprocess injects one
        # PI incoming message (pi@lab.edu) addressed to research@lab.edu — it must
        # never be selected as the agent's deliverable.
        candidates = []
        for subj, to_addr, body, from_addr, folder in rows:
            tos = []
            if isinstance(to_addr, list):
                tos = [str(t).lower() for t in to_addr]
            elif isinstance(to_addr, str):
                try:
                    p = json.loads(to_addr)
                    tos = [str(t).lower() for t in p] if isinstance(p, list) else [to_addr.lower()]
                except Exception:
                    tos = [to_addr.lower()]
            if not any(target_to in t for t in tos):
                continue
            from_l = (from_addr or "").lower()
            if "pi@lab.edu" in from_l:
                continue
            candidates.append((subj, body, (folder or "").lower()))

        # If several candidate messages exist (e.g. one per sub-agent, or a progress
        # note followed by the final delivery), prefer the deliverable email.
        match = _pick_deliverable(candidates)

        record("Email to research@lab.edu exists", match is not None, "")
        if match:
            body_lower = (match[1] or "").lower()
            # Subject must reference the ML research catalog (tolerant).
            record("Email subject references ML research catalog",
                   _email_subject_ok(match[0]),
                   f"Subject: {match[0]}")
            # Body must mention number of videos and papers
            record("Email body mentions video count (7)",
                   "7" in body_lower or "seven" in body_lower,
                   "")
            record("Email body mentions ML_Research_Tracker.xlsx file",
                   "ml_research_tracker" in body_lower or "research_tracker" in body_lower or "tracker.xlsx" in body_lower,
                   "")
            record("Email body mentions knowledge base / Notion",
                   "knowledge base" in body_lower or "notion" in body_lower or "research hub" in body_lower,
                   "")
    except Exception as e:
        record("Email check", False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print(f"Running evaluation for yt-ml-repos-github-notion-excel-email")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_notion()
    check_email()

    all_passed = FAIL_COUNT == 0
    summary = f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"
    print(f"\n{'='*40}")
    print(f"Result: {'PASS' if all_passed else 'FAIL'} - {summary}")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "all_passed": all_passed}, f)

    return all_passed, summary


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
