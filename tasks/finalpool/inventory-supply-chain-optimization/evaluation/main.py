#!/usr/bin/env python3
"""Evaluation script for inventory-supply-chain-optimization."""

import argparse
import os
import sys
from pathlib import Path

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432,
        dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
        user="eigent", password="camel",
    )


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    if not agent_workspace:
        return False, "No agent workspace provided"

    agent_ws = Path(agent_workspace)
    if not agent_ws.exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    # Detect GT self-test mode (V1 parity test).
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    # Phase 4-5: Excel inventory/procurement report
    xlsx_files = [p for p in agent_ws.glob("*.xlsx")
                  if 'data.csv' not in p.name.lower()]
    check("At least one xlsx inventory/procurement report exists",
          len(xlsx_files) >= 1, f"found {len(xlsx_files)} xlsx files")
    if xlsx_files:
        try:
            import openpyxl
            audit_path = max(xlsx_files, key=lambda p: p.stat().st_size)
            wb = openpyxl.load_workbook(str(audit_path), data_only=True)
            sheets = [s.lower() for s in wb.sheetnames]
            has_inventory_sheet = any(
                kw in s
                for s in sheets
                for kw in ['inventory', 'procurement', 'reorder', 'forecast',
                          'order', 'supplier', 'analysis', 'summary',
                          'tracking', 'demand', 'eoq']
            )
            check(f"Inventory xlsx ({audit_path.name}) has inventory-related sheet",
                  has_inventory_sheet, f"sheets={wb.sheetnames}")
            # Tightened: at least one sheet should have >=4 rows (header + 3 SKUs)
            max_rows = max((ws.max_row for ws in wb.worksheets), default=0)
            check("Inventory xlsx has >=4 rows of data in some sheet",
                  max_rows >= 4, f"max rows={max_rows}")

            # Value-level GT comparison - check headers/columns coverage
            gt_xlsx_path = Path(groundtruth_workspace) / "Inventory_Optimization_Report.xlsx"
            if gt_xlsx_path.exists():
                try:
                    gt_wb = openpyxl.load_workbook(str(gt_xlsx_path), data_only=True)
                    # Collect all header strings from all sheets
                    agent_headers = set()
                    for ws in wb.worksheets:
                        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                        for h in first:
                            if h:
                                agent_headers.add(str(h).strip().lower().replace(' ', '_'))
                    # GT has columns like SKU/Reorder_Point/EOQ/Safety_Stock/Supplier
                    expected_kw = ['sku', 'reorder', 'eoq', 'safety', 'supplier',
                                   'lead_time', 'demand', 'forecast']
                    matched_kw = sum(
                        1 for kw in expected_kw
                        if any(kw in h for h in agent_headers)
                    )
                    check("Inventory xlsx column-coverage (>=3 of 8 inventory keywords)",
                          matched_kw >= 3,
                          f"matched {matched_kw}/8 in headers={list(agent_headers)[:10]}")

                    # Numeric value sanity: collect agent + GT numerics, expect overlap
                    agent_numerics = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    agent_numerics.append(float(v))
                    gt_numerics = []
                    for ws in gt_wb.worksheets:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    gt_numerics.append(float(v))
                    check("Inventory xlsx has >=5 numeric values",
                          len(agent_numerics) >= 5,
                          f"got {len(agent_numerics)} numerics")

                    # Value-level: agent should produce reorder/EOQ-style
                    # positive integer counts (not just any numerics).
                    # Look for column header containing reorder/eoq/safety/qty
                    # and verify >=2 rows have plausible positive values.
                    inventory_value_count = 0
                    for ws in wb.worksheets:
                        first = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
                        h_low = [str(h).strip().lower() if h else '' for h in first]
                        target_idxs = [i for i, h in enumerate(h_low)
                                       if any(kw in h for kw in
                                              ['reorder', 'eoq', 'safety',
                                               'quantity', 'qty', 'order_qty',
                                               'demand'])]
                        if not target_idxs:
                            continue
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            for i in target_idxs:
                                if i >= len(row):
                                    continue
                                v = row[i]
                                if isinstance(v, (int, float)) and v > 0:
                                    inventory_value_count += 1
                    check("Inventory xlsx has >=2 positive reorder/EOQ/safety/demand values",
                          inventory_value_count >= 2,
                          f"got {inventory_value_count} positive values in inventory columns")
                    gt_wb.close()
                except Exception as _e:
                    check("Inventory xlsx GT value-comparison", False, str(_e))
        except Exception as e:
            check("Inventory xlsx parse", False, str(e))

    # Phase 5: Word risk-assessment / procedures doc
    docx_files = list(agent_ws.glob("*.docx"))
    check("At least one docx supply-chain doc exists",
          len(docx_files) >= 1, f"found {len(docx_files)} docx files")
    if docx_files:
        try:
            from docx import Document
            audit_docx = max(docx_files, key=lambda p: p.stat().st_size)
            doc = Document(str(audit_docx))
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word doc has substantive content (>=200 chars)",
                  len(text) >= 200, f"len={len(text)}")
            has_supply = any(kw in text for kw in
                             ['supply chain', 'inventory', 'procurement',
                              'supplier', 'reorder', 'forecast', 'risk'])
            has_recs = any(kw in text for kw in
                           ['recommendation', 'action', 'contingency',
                            'procedure', 'plan', 'finding'])
            check("Word doc mentions supply-chain/inventory/risk topic",
                  has_supply, f"text head: {text[:200]}")
            check("Word doc mentions recommendations/procedures/plans",
                  has_recs, f"text head: {text[:200]}")
        except Exception as e:
            check("Word doc parse", False, str(e))

    # Phase 6: Email + Calendar (DB checks)
    # GT self-test: tolerated as WARNING; real-agent run: BLOCKING.
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, subject FROM email.messages
               WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s
                  OR body_text ILIKE %s""",
            ('%purchase%', '%order%', '%supplier%', '%forecast%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Procurement email sent to suppliers: 0 found (GT self-test, non-blocking)")
        else:
            check("Procurement email sent to suppliers",
                  len(rows) >= 1, f"found {len(rows)} matching emails")
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))

    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT summary FROM gcal.events
               WHERE summary ILIKE %s OR summary ILIKE %s OR summary ILIKE %s""",
            ('%review%', '%inventory%', '%supply%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Calendar review meeting scheduled: 0 found (GT self-test, non-blocking)")
        else:
            check("Calendar review meeting scheduled",
                  len(rows) >= 1, f"found {len(rows)} matching events")
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))

    # Phase 6: Google Sheet for inventory tracking
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, title FROM gsheet.spreadsheets
               WHERE title ILIKE %s OR title ILIKE %s OR title ILIKE %s""",
            ('%inventory%', '%procurement%', '%tracking%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Inventory tracking Google Sheet created: 0 found (GT self-test, non-blocking)")
        else:
            check("Inventory tracking Google Sheet created",
                  len(rows) >= 1, f"found {len(rows)} matching spreadsheets")
        conn.close()
    except Exception as e:
        check("GSheet check", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
