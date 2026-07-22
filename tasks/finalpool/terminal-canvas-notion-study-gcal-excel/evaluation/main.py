"""Evaluation for terminal-canvas-notion-study-gcal-excel."""
import argparse
import json
import os
import sys
from datetime import date

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

COURSE_IDS = (3, 4, 7, 16, 17)

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').strip())
    except (TypeError, ValueError):
        return default


def get_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
    SELECT c.id, c.name,
           (SELECT COUNT(DISTINCT user_id) FROM canvas.enrollments WHERE course_id = c.id),
           (SELECT COUNT(*) FROM canvas.assignments WHERE course_id = c.id),
           (SELECT COUNT(*) FROM canvas.quizzes WHERE course_id = c.id)
    FROM canvas.courses c
    WHERE c.id IN %s
    """, (COURSE_IDS,))
    courses = []
    for cid, name, enroll, asgn, quiz in cur.fetchall():
        weekly = 3 + 0.5 * asgn + 1 * quiz
        if weekly > 8:
            priority = "High"
        elif weekly >= 5:
            priority = "Medium"
        else:
            priority = "Low"
        courses.append({
            "id": cid,
            "name": name,
            "enrollment": int(enroll),
            "assignments": int(asgn),
            "quizzes": int(quiz),
            "weekly_hours": weekly,
            "priority": priority,
        })
    cur.close()
    conn.close()
    return courses


def check_excel(workspace, expected):
    print("\n=== Check 1: Study_Plan_Report.xlsx ===")
    path = os.path.join(workspace, "Study_Plan_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheets_lower = [s.lower() for s in sheets]
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    # Course_Analysis
    ca_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s and "analysis" in s), None)
    if ca_idx is None:
        ca_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s), 0)
    ws = wb[sheets[ca_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
    check("Course_Analysis has exactly 5 rows", len(data_rows) == 5, f"Found {len(data_rows)}")

    # All 5 expected courses present with correct values
    a_lookup = {}
    for row in data_rows:
        if not row or row[0] is None:
            continue
        a_lookup[str(row[0]).strip().lower()] = row
    for exp in expected:
        row = a_lookup.get(exp["name"].lower())
        check(f"Course '{exp['name']}' row present", row is not None, f"got names: {list(a_lookup)}")
        if row is None:
            continue
        # Expected: Course_Name, Enrollment, Assignments, Quizzes, Weekly_Hours, Priority
        try:
            enroll = int(float(row[1])) if row[1] is not None else None
        except (TypeError, ValueError):
            enroll = None
        try:
            asgn = int(float(row[2])) if row[2] is not None else None
        except (TypeError, ValueError):
            asgn = None
        try:
            quiz = int(float(row[3])) if row[3] is not None else None
        except (TypeError, ValueError):
            quiz = None
        wh = safe_float(row[4])
        pri = str(row[5] or "").strip().lower()
        check(f"  '{exp['name']}'.Enrollment == {exp['enrollment']}",
              enroll == exp["enrollment"], f"got {enroll}")
        check(f"  '{exp['name']}'.Assignments == {exp['assignments']}",
              asgn == exp["assignments"], f"got {asgn}")
        check(f"  '{exp['name']}'.Quizzes == {exp['quizzes']}",
              quiz == exp["quizzes"], f"got {quiz}")
        check(f"  '{exp['name']}'.Weekly_Hours ≈ {exp['weekly_hours']}",
              wh is not None and abs(wh - exp["weekly_hours"]) <= 0.1,
              f"got {wh}")
        check(f"  '{exp['name']}'.Priority == {exp['priority']}",
              pri == exp["priority"].lower(), f"got {pri}")

    # Header check
    headers = [str(c).lower() if c else "" for c in rows[0]]
    for h in ("course_name", "enrollment", "assignments", "quizzes", "weekly_hours", "priority"):
        check(f"Header '{h}' present", any(h in hh for hh in headers), f"got {headers}")

    # Weekly_Schedule
    ws_idx = next((i for i, s in enumerate(sheets_lower) if "weekly" in s and "schedule" in s), None)
    if ws_idx is None:
        ws_idx = next((i for i, s in enumerate(sheets_lower) if "schedule" in s), 1)
    ws2 = wb[sheets[ws_idx]]
    rows2 = list(ws2.iter_rows(values_only=True))
    data_rows2 = [r for r in rows2[1:] if any(c is not None for c in r)]
    check("Weekly_Schedule has exactly 5 rows", len(data_rows2) == 5, f"Found {len(data_rows2)}")
    days_found = []
    for row in data_rows2:
        d = str(row[0] or "").strip().lower()
        days_found.append(d)
    for d in ("monday", "tuesday", "wednesday", "thursday", "friday"):
        check(f"Weekly_Schedule has {d}", d in days_found, f"got {days_found}")

    # Verify duration matches priority for each session
    for row in data_rows2:
        course_name = str(row[2] or "").strip()
        try:
            duration = float(row[3])
        except (TypeError, ValueError):
            duration = None
        # Find expected priority from course
        exp_course = next((c for c in expected if c["name"].lower() == course_name.lower()), None)
        if exp_course is None:
            continue
        if exp_course["priority"] == "High":
            expected_dur = 2.0
        elif exp_course["priority"] == "Medium":
            expected_dur = 1.5
        else:
            expected_dur = 1.0
        check(f"  '{course_name}' Duration_Hours == {expected_dur}",
              duration is not None and abs(duration - expected_dur) <= 0.01,
              f"got {duration}")

    # Priority_Matrix
    pm_idx = next((i for i, s in enumerate(sheets_lower) if "priority" in s and "matrix" in s), None)
    if pm_idx is None:
        pm_idx = next((i for i, s in enumerate(sheets_lower) if "priority" in s), 2)
    ws3 = wb[sheets[pm_idx]]
    rows3 = list(ws3.iter_rows(values_only=True))
    data_rows3 = [r for r in rows3[1:] if any(c is not None for c in r)]
    check("Priority_Matrix has 3 levels", len(data_rows3) == 3, f"Found {len(data_rows3)}")
    pm_lookup = {str(row[0] or "").strip().lower(): row for row in data_rows3}
    for level in ("High", "Medium", "Low"):
        row = pm_lookup.get(level.lower())
        check(f"Priority_Matrix has '{level}' row", row is not None)
        if row is not None:
            exp_count = sum(1 for c in expected if c["priority"] == level)
            exp_total = sum(c["weekly_hours"] for c in expected if c["priority"] == level)
            try:
                cnt = int(float(row[1])) if row[1] is not None else None
            except (TypeError, ValueError):
                cnt = None
            try:
                total = float(row[2]) if row[2] is not None else None
            except (TypeError, ValueError):
                total = None
            check(f"  Priority {level}.Course_Count == {exp_count}",
                  cnt == exp_count, f"got {cnt}")
            check(f"  Priority {level}.Total_Weekly_Hours == {exp_total}",
                  total is not None and abs(total - exp_total) <= 0.1,
                  f"got {total}")


def check_notion(expected):
    print("\n=== Check 2: Notion Study Planner Database ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return

    cur.execute("SELECT id, title FROM notion.databases")
    dbs = cur.fetchall()
    planner_db = None
    for db_id, title in dbs:
        title_str = ""
        if isinstance(title, list):
            title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
        elif isinstance(title, str):
            try:
                parsed = json.loads(title)
                if isinstance(parsed, list):
                    title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                else:
                    title_str = str(title)
            except Exception:
                title_str = str(title)
        else:
            title_str = str(title) if title else ""
        if "study" in title_str.lower() and "planner" in title_str.lower():
            planner_db = (db_id, title_str)
            break
    check("Notion 'Student Study Planner' database exists", planner_db is not None,
          f"Databases: {[d[1] for d in dbs]}")
    if not planner_db:
        cur.close()
        conn.close()
        return
    cur.execute("SELECT id, properties FROM notion.pages WHERE parent->>'database_id' = %s", (planner_db[0],))
    pages = cur.fetchall()
    check("Database has 5 course entries", len(pages) == 5, f"Found {len(pages)} pages")

    # Verify each course covered
    found_courses = set()
    for pid, props in pages:
        ptext = json.dumps(props or {}).lower()
        for exp in expected:
            if exp["name"].lower() in ptext:
                found_courses.add(exp["name"])
    for exp in expected:
        check(f"Notion has page for course '{exp['name']}'",
              exp["name"] in found_courses, f"got {found_courses}")

    cur.close()
    conn.close()


def check_gcal(expected):
    print("\n=== Check 3: Google Calendar Study Sessions ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return

    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    check("Exactly 5 study session events", len(events) == 5, f"Found {len(events)} events")

    # Date span: Mar 9-13 2026 (Mon-Fri)
    expected_dates = {date(2026, 3, 9), date(2026, 3, 10), date(2026, 3, 11), date(2026, 3, 12), date(2026, 3, 13)}
    found_dates = set()
    for sm, desc, start, end in events:
        if start is not None:
            found_dates.add(start.date() if hasattr(start, "date") else start)
    for d in expected_dates:
        check(f"Calendar has event on {d}", d in found_dates, f"got {found_dates}")

    # All summaries should contain 'Study Session'
    if events:
        summaries = [s or "" for s, _, _, _ in events]
        check("All events summary contains 'Study Session'",
              all("study session" in s.lower() for s in summaries),
              f"summaries: {summaries}")
        # Each summary contains a course name
        course_names = [c["name"].lower() for c in expected]
        per_summary_match = sum(1 for s in summaries if any(cn in s.lower() for cn in course_names))
        check("All 5 events reference one of the 5 courses",
              per_summary_match == 5, f"matched {per_summary_match}/5")

    # Duration matches priority
    for sm, desc, start, end in events:
        if start is None or end is None:
            continue
        duration_hours = (end - start).total_seconds() / 3600
        # Find course by summary
        match_course = None
        for c in expected:
            if c["name"].lower() in (sm or "").lower():
                match_course = c
                break
        if match_course is None:
            continue
        if match_course["priority"] == "High":
            expected_dur = 2.0
        elif match_course["priority"] == "Medium":
            expected_dur = 1.5
        else:
            expected_dur = 1.0
        check(f"Calendar event for '{match_course['name']}' duration ≈ {expected_dur}h",
              abs(duration_hours - expected_dur) <= 0.05,
              f"got {duration_hours:.2f}h")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: study_planner.py ===")
    path = os.path.join(workspace, "study_planner.py")
    check("study_planner.py exists", os.path.exists(path))
    if os.path.exists(path):
        with open(path) as f:
            content = f.read()
        # Verify formula keywords
        check("Script references base 3 hours", "3" in content,
              "should reference base 3 hours")
        check("Script references 0.5 multiplier for assignments",
              "0.5" in content or "0.5 hours" in content.lower(),
              "should reference 0.5 multiplier")
        check("Script references priority high/medium/low",
              all(kw in content.lower() for kw in ("high", "medium", "low")),
              "missing priority keywords")


def check_reverse_validation():
    print("\n=== Reverse Validation ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("""
        SELECT summary, start_datetime, EXTRACT(DOW FROM start_datetime) AS dow
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%%study%%'
    """)
    events = cur.fetchall()
    weekend_events = [e for e in events if e[2] in (0, 6)]
    check("No study session events on weekends", len(weekend_events) == 0,
          f"weekend events: {weekend_events}")
    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = get_expected()
    check_excel(args.agent_workspace, expected)
    check_notion(expected)
    check_gcal(expected)
    check_script(args.agent_workspace)
    check_reverse_validation()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_ok = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": all_ok}, f)
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
