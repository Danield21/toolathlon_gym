"""
Evaluation script for arxiv-research-landscape-report task.

Checks:
1. Excel file (Research_Landscape.xlsx) with 3 sheets
2. Paper Analysis has 5 target papers with correct IDs and citation counts
3. Conference Fit has 3 conferences
4. Summary has required metrics
5. Word document (Landscape_Report.docx) exists with substantive content
"""
import argparse
import json
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0

EXPECTED_PAPERS = {
    "2401.00001": {"title": "Efficient Transformers for NLP", "citation_count": 350},
    "2401.00002": {"title": "Deep RL with Human Feedback", "citation_count": 520},
    "2401.00003": {"title": "Generative Models for Code", "citation_count": 280},
    "2401.00004": {"title": "Knowledge Graph Embeddings", "citation_count": 190},
    "2401.00005": {"title": "Optimization in Deep Learning", "citation_count": 150},
}

NOISE_IDS = {"2401.00006", "2401.00007", "2401.00008"}

CONFERENCES = ["NeurIPS 2026", "ICML 2026", "AAAI 2026"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=50):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel Output ===")
    path = os.path.join(workspace, "Research_Landscape.xlsx")
    if not os.path.isfile(path):
        record("Excel file exists", False, f"Not found: {path}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False
    record("Excel readable", True)

    all_ok = True

    # Paper Analysis sheet
    pa_rows = load_sheet_rows(wb, "Paper Analysis") or load_sheet_rows(wb, "Paper_Analysis")
    if pa_rows is None:
        record("Sheet 'Paper Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Paper Analysis' exists", True)
        header = pa_rows[0] if pa_rows else []
        data = pa_rows[1:]

        id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        title_col = find_col(header, ["Title", "title"])
        cite_col = find_col(header, ["Citation_Count", "Citation Count", "Citations", "citation_count"])

        record("Paper Analysis has 5 data rows", len(data) == 5,
               f"Found {len(data)}")
        if len(data) != 5:
            all_ok = False

        if id_col is not None:
            found_ids = {str(r[id_col]).strip() for r in data if id_col < len(r) and r[id_col]}
            for eid in EXPECTED_PAPERS:
                present = eid in found_ids
                record(f"Paper {eid} present", present, f"Found: {found_ids}")
                if not present:
                    all_ok = False
            for nid in NOISE_IDS:
                absent = nid not in found_ids
                record(f"Noise {nid} absent", absent)
                if not absent:
                    all_ok = False
        else:
            record("Paper_ID column found", False, f"Header: {header}")
            all_ok = False

        if cite_col is not None and id_col is not None:
            for row in data:
                pid = str(row[id_col]).strip() if id_col < len(row) and row[id_col] else ""
                if pid in EXPECTED_PAPERS:
                    ok = num_close(row[cite_col] if cite_col < len(row) else None,
                                   EXPECTED_PAPERS[pid]["citation_count"], tol=50)
                    record(f"Citation count for {pid}", ok,
                           f"Got {row[cite_col] if cite_col < len(row) else None}, expected {EXPECTED_PAPERS[pid]['citation_count']}")
                    if not ok:
                        all_ok = False

            # NEW: Check Paper Analysis sorted by Citation_Count DESC
            try:
                cites = [float(r[cite_col]) for r in data if cite_col < len(r) and r[cite_col] is not None]
                sorted_ok = cites == sorted(cites, reverse=True)
                record("Paper Analysis sorted by Citation_Count DESC", sorted_ok,
                       f"Got order: {cites}")
                if not sorted_ok:
                    all_ok = False
            except Exception:
                pass

    # Conference Fit sheet
    cf_rows = load_sheet_rows(wb, "Conference Fit") or load_sheet_rows(wb, "Conference_Fit")
    if cf_rows is None:
        record("Sheet 'Conference Fit' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Conference Fit' exists", True)
        header = cf_rows[0] if cf_rows else []
        data = cf_rows[1:]
        record("Conference Fit has 3 rows", len(data) == 3, f"Found {len(data)}")
        if len(data) != 3:
            all_ok = False
        # Check conference names present
        conf_col = find_col(header, ["Conference", "Conference_Name", "Name"])
        if conf_col is not None:
            found_confs = " ".join(str(r[conf_col]).lower() for r in data if conf_col < len(r) and r[conf_col])
            for c in ["neurips", "icml", "aaai"]:
                ok = c in found_confs
                record(f"Conference Fit mentions {c.upper()}", ok, f"Found: {found_confs}")
                if not ok:
                    all_ok = False
        # Check Matching_Papers and Avg_Citations columns exist
        mp_col = find_col(header, ["Matching_Papers", "Matching Papers", "Paper_Count"])
        ac_col = find_col(header, ["Avg_Citations", "Average_Citations", "Avg Citations"])
        record("Conference Fit has Matching_Papers column", mp_col is not None, f"Header: {header}")
        record("Conference Fit has Avg_Citations column", ac_col is not None, f"Header: {header}")
        if mp_col is None or ac_col is None:
            all_ok = False

        # NEW: Matching_Papers numeric values. Given topics + papers mapping:
        # NeurIPS 2026 (deep learning, RL, NLP) matches 2 papers (00001, 00002)
        # ICML 2026 (optimization, generative models) matches 2 papers (00003, 00005)
        # AAAI 2026 (knowledge graphs, planning, NLP) matches 2 papers (00001, 00004)
        expected_mp = {"neurips": 2, "icml": 2, "aaai": 2}
        if mp_col is not None and conf_col is not None:
            for row in data:
                conf = str(row[conf_col]).lower() if conf_col < len(row) and row[conf_col] else ""
                mp_val = row[mp_col] if mp_col < len(row) else None
                for key, expected in expected_mp.items():
                    if key in conf:
                        ok = num_close(mp_val, expected, tol=1)
                        record(f"Conference Fit Matching_Papers for {key.upper()}", ok,
                               f"Got {mp_val}, expected ~{expected}")
                        if not ok:
                            all_ok = False

        # NEW: Avg_Citations numeric values
        # NeurIPS 2026: (520 + 350) / 2 = 435
        # ICML 2026: (280 + 150) / 2 = 215
        # AAAI 2026: (350 + 190) / 2 = 270
        expected_ac = {"neurips": 435, "icml": 215, "aaai": 270}
        if ac_col is not None and conf_col is not None:
            for row in data:
                conf = str(row[conf_col]).lower() if conf_col < len(row) and row[conf_col] else ""
                ac_val = row[ac_col] if ac_col < len(row) else None
                for key, expected in expected_ac.items():
                    if key in conf:
                        ok = num_close(ac_val, expected, tol=30)
                        record(f"Conference Fit Avg_Citations for {key.upper()}", ok,
                               f"Got {ac_val}, expected ~{expected}")
                        if not ok:
                            all_ok = False

    # Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        metrics = {}
        for row in sum_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        # Total Papers
        tp_key = next((k for k in metrics if "total" in k and "paper" in k), None)
        if tp_key:
            ok = num_close(metrics[tp_key], 5, tol=0)
            record("Summary: Total_Papers = 5", ok, f"Got {metrics[tp_key]}")
            if not ok:
                all_ok = False
        else:
            record("Summary: Total_Papers exists", False)
            all_ok = False

        # Average Citations
        avg_key = next((k for k in metrics if "avg" in k or "average" in k), None)
        expected_avg = (350 + 520 + 280 + 190 + 150) / 5  # 298.0
        if avg_key:
            ok = num_close(metrics[avg_key], expected_avg, tol=20)
            record("Summary: Avg_Citations", ok,
                   f"Got {metrics[avg_key]}, expected ~{expected_avg}")
            if not ok:
                all_ok = False
        else:
            record("Summary: Avg_Citations exists", False)
            all_ok = False

        # Highest Cited Paper - stricter: must match full phrase "deep rl with human feedback"
        hc_key = next((k for k in metrics if "highest" in k or "most" in k), None)
        if hc_key:
            val = str(metrics[hc_key]).lower() if metrics[hc_key] else ""
            # Require the full title "Deep RL with Human Feedback" (allowing minor whitespace variations)
            ok = "deep rl with human feedback" in val or (
                "deep rl" in val and "human feedback" in val and "deep rl with human feedback".replace(" ", "") in val.replace(" ", "")
            )
            record("Summary: Highest_Cited is Deep RL paper (exact phrase)", ok, f"Got: {metrics[hc_key]}")
            if not ok:
                all_ok = False

        # NEW: Best_Conference_Fit should be one of the top conferences (tied at 2)
        bcf_key = next((k for k in metrics if "best" in k and ("conference" in k or "fit" in k)), None)
        if bcf_key:
            val = str(metrics[bcf_key]).lower() if metrics[bcf_key] else ""
            ok = any(c in val for c in ["neurips", "icml", "aaai"])
            record("Summary: Best_Conference_Fit is a valid conference", ok, f"Got: {metrics[bcf_key]}")
            if not ok:
                all_ok = False

    return all_ok


def check_word(workspace):
    print("\n=== Checking Word Document ===")
    path = os.path.join(workspace, "Landscape_Report.docx")
    if not os.path.isfile(path):
        record("Word document exists", False, f"Not found: {path}")
        return False
    record("Word document exists", True)

    try:
        from docx import Document
        doc = Document(path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()

        # Require longer document (>= 4 sections of analysis)
        all_ok = True
        ok = len(full_text) > 400
        record("Document has substantial content (>400 chars)", ok, f"Only {len(full_text)} chars")
        if not ok:
            all_ok = False
        ok = sum(1 for c in ["neurips", "icml", "aaai"] if c in full_text) >= 3
        record("Mentions all three conferences (NeurIPS, ICML, AAAI)", ok, "Missing one or more conferences")
        if not ok:
            all_ok = False
        ok = "research" in full_text or "landscape" in full_text
        record("Mentions research/landscape", ok)
        if not ok:
            all_ok = False
        # Require at least 3 topic keywords (not just 1)
        topic_kws = ["transformer", "reinforcement", "optimization", "knowledge graph", "generative", "nlp"]
        topic_hits = sum(1 for kw in topic_kws if kw in full_text)
        ok = topic_hits >= 3
        record(f"Mentions >=3 topic keywords ({topic_hits}/6)", ok)
        if not ok:
            all_ok = False

        # Check 4 sections: overview, alignment, priority, gap analysis
        has_overview = any(kw in full_text for kw in ["overview", "summary", "analysis"])
        has_alignment = any(kw in full_text for kw in ["align", "mapping", "map to", "fit"])
        has_priority = any(kw in full_text for kw in ["priority", "recommend", "preferred", "prioritize"])
        has_gap = any(kw in full_text for kw in ["gap", "missing", "coverage", "uncovered", "topic"])
        record("Word has overview section", has_overview)
        record("Word has alignment section", has_alignment)
        record("Word has priority section", has_priority)
        record("Word has gap analysis section", has_gap)
        # Require at least 3 of 4 sections present (gap may be mentioned implicitly via 'topic')
        section_count = sum([has_overview, has_alignment, has_priority, has_gap])
        ok = section_count >= 3
        record(f"Word has >=3 of 4 required sections ({section_count}/4)", ok)
        if not ok:
            all_ok = False

        # Reference check: confirm the report draws on research_priorities.md /
        # publication_history.json by mentioning at least one of the group's
        # past venues (ACL/NeurIPS/ICML/AAAI/ICLR/CVPR) AND at least one
        # stated group expertise keyword (transformer/alignment/knowledge graph/optimization).
        past_venues = ["acl", "neurips", "icml", "aaai", "iclr", "cvpr"]
        venue_hits = sum(1 for v in past_venues if v in full_text)
        ok = venue_hits >= 3
        record(f"Report references >=3 past venues from publication_history ({venue_hits}/6)", ok)
        if not ok:
            all_ok = False

        expertise_kws = ["transformer", "alignment", "knowledge graph", "optimization",
                         "reinforcement learning", "generative", "nlp"]
        expertise_hits = sum(1 for kw in expertise_kws if kw in full_text)
        ok = expertise_hits >= 3
        record(f"Report references >=3 group expertise areas ({expertise_hits}/7)", ok)
        if not ok:
            all_ok = False

        return all_ok
    except Exception as e:
        record("Word document readable", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    word_ok = check_word(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Excel: {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Word:  {'PASS' if word_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
