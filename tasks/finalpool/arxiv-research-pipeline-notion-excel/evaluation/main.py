"""Evaluation script for arxiv-research-pipeline-notion-excel."""
import os
import argparse, json, os, sys
import re
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        s = str(val).replace(",", "").replace("%", "").replace("$", "")
        s = s.replace("¥", "").replace("€", "").replace("£", "").strip()
        if s == "": return default
        return float(s)
    except (ValueError, TypeError):
        return default

# arXiv IDs can legitimately appear as bare "2301.00001", as
# "http://arxiv.org/abs/2301.00001" (scholarly MCP's "Entry ID:" format),
# as "arxiv://2301.00001", or with a version suffix "2301.00001v2".
# Normalize to the bare arXiv id before comparing.
ARXIV_ID_RE = re.compile(r"(\d{4})\.(\d{4,5})")

def normalize_arxiv_id(v):
    if v is None:
        return None
    s = str(v).strip()
    m = ARXIV_ID_RE.search(s)
    if m:
        return m.group(0)
    return s.lower()

def norm_header(v):
    """Normalize an Excel header for case/punctuation-insensitive matching.

    'Citation_Count', 'Citation Count' and 'citation_count' all become
    'citationcount' so the evaluator does not penalize cosmetic header
    variants while the task still specifies the canonical column names.
    """
    if v is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(v).strip().lower())

# Notion topic coverage. task.md asks the page to cover four topics: research
# landscape overview, key papers summary, methodology comparison highlights,
# and research gaps with future-work recommendations. A faithful page may word
# any of these very differently from the task text, so each topic is matched
# against a broad set of synonym phrases (e.g. "challenges and directions" for
# research gaps, "how different methods compare" for methodology comparison).
TOPIC_KEYWORDS = {
    "research landscape": [
        "research landscape", "landscape",
        "research overview", "overview of the field", "field overview",
        "overview of the research", "research trends", "state of the art",
        "current research", "research areas", "big picture", "overview",
        "survey of the field",
    ],
    "key papers": [
        "key paper", "key papers", "notable papers", "main papers",
        "important papers", "selected papers", "most relevant papers",
        "seminal papers", "core papers", "highlighted papers",
        "representative papers", "major papers", "paper summary",
        "summary of papers", "papers reviewed", "key publications",
        "notable publications", "relevant publications", "most important papers",
    ],
    "methodology comparison": [
        "methodology", "method comparison", "methods comparison",
        "comparing methods", "comparison of methods", "how methods compare",
        "how different methods compare", "compare the methods", "compares methods",
        "compare methods", "comparing approaches", "approaches compared",
        "comparison", "compare", "contrast", "benchmark comparison",
        "method analysis", "approach comparison", "different approaches",
    ],
    "research gaps": [
        "research gap", "research gaps", "open problem", "open problems",
        "future work", "future direction", "future directions",
        "remaining challenge", "remaining challenges", "open question",
        "open questions", "challenges", "challenge", "unsolved problems",
        "limitations", "gap analysis", "gaps", "next steps",
        "recommendations for future work", "directions for future work",
        "future outlook",
    ],
}

def covered_topics(joined):
    """Return the task topics whose keyword phrases appear in the page text."""
    j = (joined or "").lower()
    return [t for t, kws in TOPIC_KEYWORDS.items() if any(k in j for k in kws)]

def _rich_text_content(obj):
    """Extract text from a Notion rich_text object."""
    if not isinstance(obj, dict):
        return ""
    txt = obj.get("text")
    if isinstance(txt, dict):
        c = txt.get("content")
        if c:
            return c
        p = txt.get("plain_text")
        if p:
            return p
    return obj.get("plain_text", "") or ""

def extract_page_title(props):
    """Robustly extract a page title from Notion page properties jsonb.

    Handles the standard Notion shape:
        {"title": {"title": [{"type": "text", "text": {"content": "..."}}]}}
    as well as the direct-list shape:
        {"title": [{"type": "text", "text": {"content": "..."}}]}
    and falls back to scanning the whole properties JSON for the hub name.
    """
    if not isinstance(props, dict):
        return ""
    title_obj = props.get("title")
    if isinstance(title_obj, dict):
        parts = [_rich_text_content(t) for t in (title_obj.get("title", []) or [])]
        return "".join(parts)
    if isinstance(title_obj, list):
        return "".join(_rich_text_content(t) for t in title_obj)
    return ""

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Check Research_Knowledge_Base.xlsx
    excel_path = os.path.join(agent_workspace, "Research_Knowledge_Base.xlsx")
    check("Research_Knowledge_Base.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path)
            gt_path = os.path.join(groundtruth_workspace, "Research_Knowledge_Base.xlsx")
            gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

            if gt_wb:
                # Match sheets case/punctuation-insensitively, so 'Paper_Catalog'
                # in GT also matches 'Paper Catalog' in the agent file.
                agent_sheets = {norm_header(sn): sn for sn in wb.sheetnames}
                for sheet_name in gt_wb.sheetnames:
                    agent_sheet = agent_sheets.get(norm_header(sheet_name))
                    check(f"{sheet_name} sheet exists", agent_sheet is not None,
                          f"sheets: {wb.sheetnames}")
                    if agent_sheet:
                        ws = wb[agent_sheet]
                        gt_ws = gt_wb[sheet_name]
                        # Check headers (case/punctuation-insensitive)
                        gt_headers = [norm_header(c.value) for c in gt_ws[1]]
                        headers = [norm_header(c.value) for c in ws[1]]
                        for h in gt_headers:
                            if h:
                                check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                        # Check row count
                        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                                   if any(c is not None for c in r)]
                        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                                     if any(c is not None for c in r)]

                        sheet_key = norm_header(agent_sheet)
                        if "researchgap" in sheet_key:
                            # Task: at least 4 identified gaps.
                            check(f"{sheet_name} has >= 4 data rows", len(data_rows) >= 4,
                                  f"got {len(data_rows)}")
                            # Validate Priority column values are in expected enum.
                            header_map = {h: i for i, h in enumerate(headers)}
                            prio_idx = header_map.get("priority", -1)
                            if prio_idx >= 0:
                                valid_priorities = {"critical", "important", "nice-to-have"}
                                invalid = []
                                for r in data_rows:
                                    if prio_idx < len(r) and r[prio_idx] is not None:
                                        val = str(r[prio_idx]).strip().lower()
                                        if val and val not in valid_priorities:
                                            invalid.append(val)
                                check(f"{sheet_name} Priority values are in {{Critical,Important,Nice-to-have}}",
                                      not invalid, f"invalid: {invalid[:5]}")
                            # Verify all rows have non-empty Gap_Area, Current_State,
                            # Opportunity, Priority.
                            non_empty_ok = True
                            for h in ["gaparea", "currentstate", "opportunity", "priority"]:
                                idx = header_map.get(h, -1)
                                if idx < 0:
                                    continue
                                for r in data_rows:
                                    if idx >= len(r) or r[idx] is None or str(r[idx]).strip() == "":
                                        non_empty_ok = False
                                        break
                                if not non_empty_ok:
                                    break
                            check(f"{sheet_name} all rows non-empty in core columns",
                                  non_empty_ok, "")
                        elif "papercatalog" in sheet_key:
                            # Task: at least 5 relevant papers. GT has only 3 rows
                            # (illustrative). Validate >= 5 rows here.
                            check(f"{sheet_name} has >= 5 data rows (task: at least 5 relevant papers)",
                                  len(data_rows) >= 5, f"got {len(data_rows)}")
                            # Validate columns Paper_ID, Title, Authors, Year are non-empty.
                            header_map = {h: i for i, h in enumerate(headers)}
                            non_empty_ok = True
                            for h in ["paperid", "title", "authors", "year"]:
                                idx = header_map.get(h, -1)
                                if idx < 0:
                                    continue
                                for r in data_rows:
                                    if idx >= len(r) or r[idx] is None or str(r[idx]).strip() == "":
                                        non_empty_ok = False
                                        break
                                if not non_empty_ok:
                                    break
                            check(f"{sheet_name} all rows non-empty in Paper_ID/Title/Authors/Year",
                                  non_empty_ok, "")
                            # Validate sort order: Citation_Count descending.
                            ci_idx = header_map.get("citationcount", -1)
                            if ci_idx >= 0 and len(data_rows) >= 2:
                                sort_ok = True
                                prev = None
                                for r in data_rows:
                                    if ci_idx >= len(r):
                                        continue
                                    v = safe_float(r[ci_idx])
                                    if v is None:
                                        continue
                                    if prev is not None and v > prev + 0.001:
                                        sort_ok = False
                                        break
                                    prev = v
                                check(f"{sheet_name} sorted by Citation_Count descending",
                                      sort_ok, "")
                            # Anti-fabrication: Paper_IDs must match real injected
                            # arxiv IDs (the 5 relevant ones). Reject the 2 noise
                            # papers (Quantum/Ocean) and any fabricated IDs.
                            # IDs are normalized to the bare arXiv id so both the
                            # arxiv_local JSON format ("2301.00001") and the
                            # scholarly text format
                            # ("http://arxiv.org/abs/2301.00001") are accepted.
                            pid_idx = header_map.get("paperid", -1)
                            if pid_idx >= 0:
                                relevant_ids = {
                                    "2301.00001",  # LLM Reasoning Survey
                                    "2302.00002",  # Prompt Engineering Guide
                                    "2303.00003",  # In-Context Learning Theory
                                    "2304.00004",  # Chain-of-Thought Reasoning
                                    "2305.00005",  # Survey of In-Context Learning
                                }
                                noise_ids = {"2304.99901", "2305.99902"}
                                seen_pids = []
                                for r in data_rows:
                                    if pid_idx < len(r) and r[pid_idx] is not None:
                                        nid = normalize_arxiv_id(r[pid_idx])
                                        if nid:
                                            seen_pids.append(nid)
                                included_relevant = relevant_ids & set(seen_pids)
                                included_noise = noise_ids & set(seen_pids)
                                check(
                                    f"{sheet_name} includes the 5 relevant arxiv IDs (no fabricated catalog)",
                                    len(included_relevant) == 5,
                                    f"included relevant: {included_relevant}, all seen: {seen_pids}",
                                )
                                check(
                                    f"{sheet_name} excludes noise papers (Quantum Computing, Ocean Modeling)",
                                    len(included_noise) == 0,
                                    f"noise leaked: {included_noise}",
                                )
                        elif "methodcomparison" in sheet_key:
                            # task.md does not state a minimum method count for
                            # this sheet (only Paper_Catalog >= 5 and Research_Gaps
                            # >= 4 are stated). Only 2 of the 5 papers have LaTeX
                            # source for deep methodology extraction
                            # (arxiv_latex.papers: 2301.00001, 2302.00002), so a
                            # thorough model that extracts 2 methods is fully
                            # correct. Require >= 2 rows (a comparison needs at
                            # least 2 methods) plus populated core columns; GT's 3
                            # illustrative rows are a valid superset.
                            check(f"{sheet_name} has >= 2 data rows", len(data_rows) >= 2,
                                  f"got {len(data_rows)}")
                            header_map = {h: i for i, h in enumerate(headers)}
                            non_empty_ok = True
                            for h in ["methodname", "papersource", "keyinnovation", "applicability"]:
                                idx = header_map.get(h, -1)
                                if idx < 0:
                                    continue
                                for r in data_rows:
                                    if idx >= len(r) or r[idx] is None or str(r[idx]).strip() == "":
                                        non_empty_ok = False
                                        break
                                if not non_empty_ok:
                                    break
                            check(f"{sheet_name} all rows non-empty in Method/Paper/Innovation/Applicability",
                                  non_empty_ok, "")
                            # Validate Applicability enum
                            ap_idx = header_map.get("applicability", -1)
                            if ap_idx >= 0:
                                valid_ap = {"high", "medium", "low"}
                                invalid_ap = []
                                for r in data_rows:
                                    if ap_idx < len(r) and r[ap_idx] is not None:
                                        val = str(r[ap_idx]).strip().lower()
                                        if val and val not in valid_ap:
                                            invalid_ap.append(val)
                                check(f"{sheet_name} Applicability values in {{High,Medium,Low}}",
                                      not invalid_ap, f"invalid: {invalid_ap[:5]}")
                        else:
                            # Default: >= GT row count.
                            check(f"{sheet_name} has >= {len(gt_rows)} data rows",
                                  len(data_rows) >= len(gt_rows), f"got {len(data_rows)}")
        except Exception as e:
            check("Research_Knowledge_Base.xlsx readable and valid", False, str(e))

    # Check the specifically-named Python script and JSON outputs exist.
    workspace_files = set(os.listdir(agent_workspace))
    check(
        "research_synthesizer.py exists",
        "research_synthesizer.py" in workspace_files,
        f"workspace: {sorted(workspace_files)[:25]}",
    )
    check(
        "papers_metadata.json exists",
        "papers_metadata.json" in workspace_files,
        f"workspace: {sorted(workspace_files)[:25]}",
    )
    check(
        "paper_contents.json exists",
        "paper_contents.json" in workspace_files,
        f"workspace: {sorted(workspace_files)[:25]}",
    )
    check(
        "research_synthesis.json exists",
        "research_synthesis.json" in workspace_files,
        f"workspace: {sorted(workspace_files)[:25]}",
    )

    # Database checks - validate Notion title and heading content.
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
        page_rows = cur.fetchall()

        page_count = len(page_rows)
        check("At least one non-archived Notion page exists", page_count >= 1,
              f"page count: {page_count}")

        # Find the page titled 'LLM Research Hub' (or similar).
        hub_page_id = None
        titles_seen = []
        for pid, props in page_rows:
            title = extract_page_title(props)
            if not title:
                # Last-resort fallback: page properties mention the hub name
                # somewhere (unusual storage shape).
                try:
                    s = json.dumps(props).lower()
                    if "llm research hub" in s:
                        title = "LLM Research Hub"
                except Exception:
                    pass
            titles_seen.append(title)
            tl = title.lower()
            if ("llm" in tl or "large language" in tl) and (
                "research hub" in tl or "research" in tl
            ):
                hub_page_id = pid
                break

        check(
            "Notion page 'LLM Research Hub' exists",
            hub_page_id is not None,
            f"Titles seen: {titles_seen[:10]}",
        )

        # Look for heading 'Large Language Model Research Dashboard' anywhere
        # in notion.blocks under that page (or any block).
        cur.execute(
            """
            SELECT parent_id, type, block_data FROM notion.blocks
            WHERE archived = false
            """
        )
        blocks = cur.fetchall()
        joined = " ".join(json.dumps(c) if c is not None else "" for _, _, c in blocks).lower()
        check(
            "Notion contains 'Large Language Model Research Dashboard' heading",
            "large language model research dashboard" in joined,
            f"sample: {joined[:200]}",
        )
        # Heuristic: page should mention each of 4 topics from task.md.
        # A faithful page may word any one of them in a fully novel way, so the
        # per-topic check passes if the topic's broad synonym set hits OR the
        # page demonstrably covers at least 3 of the 4 required topics. A stub /
        # off-topic page covering 2 or fewer topics still FAILs, so the check
        # keeps testing substantive multi-topic content.
        covered = covered_topics(joined)
        overall_ok = len(covered) >= 3
        for topic in TOPIC_KEYWORDS:
            check(
                f"Notion page mentions topic: {topic}",
                topic in covered or overall_ok,
                f"covered {len(covered)}/4 topics",
            )
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
