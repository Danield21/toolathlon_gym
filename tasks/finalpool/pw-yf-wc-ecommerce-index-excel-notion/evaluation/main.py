"""Evaluation script for pw-yf-wc-ecommerce-index-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-yf-wc-ecommerce-index-excel-notion"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Detect GT self-test (V1 parity test).
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    excel_path = os.path.join(agent_workspace, "Wc_Ecommerce_Index_Report.xlsx")
    check("Wc_Ecommerce_Index_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Wc_Ecommerce_Index_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Category', 'Our_Avg_Price', 'Market_Avg_Price', 'Price_Gap_Pct'], 6)
            # Verify Price_Gap_Pct contains numeric values (not just placeholder text)
            headers_lower = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            gap_idx = None
            for i, h in enumerate(headers_lower):
                if 'gap' in h and 'pct' in h:
                    gap_idx = i
                    break
                if 'gap' in h and '%' in h:
                    gap_idx = i
                    break
            if gap_idx is not None:
                gap_vals = []
                for row in data_rows:
                    val = safe_float(row[gap_idx])
                    if val is not None:
                        gap_vals.append(val)
                check("Price_Gap_Pct contains numeric values",
                      len(gap_vals) >= max(3, len(data_rows) // 2),
                      f"only {len(gap_vals)}/{len(data_rows)} numeric")
                # Plausible range: -100 to +100% typically, allow up to +/-500
                in_range = [v for v in gap_vals if -500 <= v <= 500]
                check("Price_Gap_Pct values in plausible range (-500 to +500)",
                      len(in_range) == len(gap_vals),
                      f"out of range: {[v for v in gap_vals if v not in in_range][:3]}")
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 3)
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action', 'Category'], 2)
    # Notion + processor.py checks moved out of Excel-existence block
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Find dashboard page (title-aware match)
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false
              AND (properties::text ILIKE %s OR properties::text ILIKE %s)
        """, ('%ecommerce dashboard%', '%yf wc ecommerce%'))
        pages = cur.fetchall()
        if is_gt_self_test and len(pages) == 0:
            print("  [WARN] Notion dashboard page created: 0 found (GT self-test, non-blocking)")
        else:
            check("Notion dashboard page created (title contains 'Ecommerce Dashboard' or 'Yf Wc Ecommerce')",
                  len(pages) >= 1, f"found {len(pages)} pages")
        # Verify dashboard page has substantive body content (>=3 blocks, >=1 text-bearing).
        if len(pages) >= 1:
            page_id = pages[0][0]
            cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s", (page_id,))
            block_count = cur.fetchone()[0]
            check("Dashboard page has >=3 body content blocks",
                  block_count >= 3, f"got {block_count} blocks")
            # Verify at least one block has text content (non-empty rich text / plain text)
            cur.execute(
                """SELECT type, block_data::text FROM notion.blocks
                   WHERE parent_id = %s""",
                (page_id,)
            )
            rows = cur.fetchall()
            text_blocks = 0
            for blk_type, content_str in rows:
                if not content_str:
                    continue
                # Plain text content presence: strip JSON quotes/whitespace and check length
                stripped = content_str.replace("null", "").strip()
                # Heuristic: alphanumeric content beyond JSON skeleton (>=20 chars of meaningful text)
                alnum_chars = sum(1 for ch in stripped if ch.isalnum())
                if alnum_chars >= 20:
                    text_blocks += 1
            check("Dashboard page has >=1 text-bearing block",
                  text_blocks >= 1, f"text_blocks={text_blocks}")
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))

    proc_path = os.path.join(agent_workspace, "yf_wc_ecommerce_processor.py")
    check("yf_wc_ecommerce_processor.py exists", os.path.exists(proc_path))
    if os.path.exists(proc_path):
        try:
            with open(proc_path, "r", encoding="utf-8", errors="ignore") as _pf:
                _proc_content = _pf.read()
            check("yf_wc_ecommerce_processor.py is non-trivial (>200 bytes + def + import)",
                  len(_proc_content) >= 200 and "def" in _proc_content and "import" in _proc_content,
                  f"size={len(_proc_content)}")
        except Exception as _e:
            check("yf_wc_ecommerce_processor.py readable", False, str(_e))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
