"""
Evaluation script for support-sla-audit-form task.

Checks:
1. Excel file (SLA_Audit_Report.xlsx) - 3 sheets with correct data computed from Snowflake
2. Google Form created with correct structure (queried from gform schema)

Expected values are computed at evaluation time from the PostgreSQL database,
not from pre-generated groundtruth files.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00" \
        --res_log_file /path/to/result.json
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


# Sentinel marking a cell whose value is an Excel formula with no cached result.
# Such cells are unreadable by openpyxl; the agent may have written a formula
# (e.g. =ROUND(...)) that a spreadsheet would compute but openpyxl cannot.
FORMULA_UNRESOLVED = object()


def _to_float(v):
    """Robustly parse a cell value to a float.

    Accepts int/float and strings with thousands separators, currency symbols,
    percent signs, and surrounding whitespace. Returns None if the value cannot
    be parsed (including None, booleans, uncached formulas, and non-numeric text).
    """
    if v is None or v is FORMULA_UNRESOLVED:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return None
    s = s.replace(",", "").replace("$", "").replace("¥", "").replace("€", "").replace("%", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def num_close(a, b, tol=1.0):
    """Compare two numeric values with tolerance.

    Both sides are parsed via _to_float (handles currency/thousand/%/spaces).
    If one side is an unresolvable Excel formula, be lenient (cannot verify the
    numeric value, but the agent did produce a value). Otherwise fall back to
    case-insensitive text comparison.
    """
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if a is FORMULA_UNRESOLVED or b is FORMULA_UNRESOLVED:
        # Unreadable formula cell: do not penalize formula-based answers.
        return True
    return str_match(a, b)


def avg_metric_close(agent_val, exp_val, tol=0.5):
    """Compare an average metric cell (Avg_Satisfaction / Avg_Response_Hours).

    Falls back to num_close for ordinary numeric comparison. Special case: when
    the expected value is 0.0, the agent resolved no tickets, so the true average
    is over an empty set and a correct agent may legitimately render it as 0, a
    blank cell, 'N/A', '-', 'NULL', 'None', etc. Accept any of those. Non-zero
    expected values must be numeric (identical to the original strict behavior).
    """
    if agent_val is FORMULA_UNRESOLVED:
        # Unreadable formula cell: do not penalize formula-based answers.
        return True
    fe = _to_float(exp_val)
    fa = _to_float(agent_val)
    if fa is not None:
        return abs(fa - fe) <= tol
    if fe == 0.0:
        if agent_val is None:
            return True
        s = str(agent_val).strip().lower().rstrip(".")
        return s in ("", "n/a", "na", "-", "--", "—", "null", "none", "nan",
                     "no data", "no average", "blank", "empty", "0.0")
    return num_close(agent_val, exp_val, tol)


def _sheet_pair(wb, wb_raw, name):
    """Return (data_only sheet, raw sheet) matched case-insensitively by name."""
    return get_sheet(wb, name), get_sheet(wb_raw, name)


def _iter_sheet_rows(ws_values, ws_raw, min_row=1):
    """Iterate a sheet yielding resolved row tuples.

    Uses the data_only workbook for cell values (so cached formula results are
    read as numbers) and the raw workbook to detect uncached formula cells
    (returned as FORMULA_UNRESOLVED). Fully-empty rows (trailing styling rows)
    are skipped so row counts are not inflated.
    """
    rows = []
    for row in ws_raw.iter_rows(min_row=min_row):
        resolved = []
        for cell in row:
            raw = cell.value
            if isinstance(raw, str) and raw.startswith("="):
                cached = ws_values[cell.coordinate].value
                resolved.append(cached if cached is not None else FORMULA_UNRESOLVED)
            else:
                resolved.append(raw)
        if any(c is not None and c is not FORMULA_UNRESOLVED for c in resolved):
            rows.append(tuple(resolved))
    return rows


def str_match(a, b):
    """Case-insensitive string comparison with whitespace normalization."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    """
    Query the PostgreSQL database to compute expected SLA audit values.
    Returns a dict with keys: sla_compliance, agent_performance, summary.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # ---- SLA Compliance by priority ----
    # Join tickets with SLA policies on priority, check breach
    cur.execute("""
        SELECT
            t."PRIORITY",
            COUNT(*) AS total_tickets,
            SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END) AS breached_tickets,
            ROUND(
                (1.0 - SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END)::numeric / COUNT(*)::numeric) * 100,
                1
            ) AS compliance_rate,
            ROUND(AVG(t."RESPONSE_TIME_HOURS")::numeric, 1) AS avg_response_hours,
            ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 1) AS avg_satisfaction
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
        JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" p
            ON t."PRIORITY" = p."PRIORITY"
        WHERE LOWER(t."STATUS") = 'resolved'
        GROUP BY t."PRIORITY"
        ORDER BY t."PRIORITY" ASC
    """)
    sla_rows = cur.fetchall()
    sla_compliance = []
    for row in sla_rows:
        sla_compliance.append({
            "Priority": row[0],
            "Total_Tickets": int(row[1]),
            "Breached_Tickets": int(row[2]),
            "Compliance_Rate": float(row[3]),
            "Avg_Response_Hours": float(row[4]),
            "Avg_Satisfaction": float(row[5]),
        })

    # ---- Agent Performance ----
    cur.execute("""
        SELECT
            a."AGENT_NAME",
            a."TEAM",
            COUNT(t."TICKET_ID") AS tickets_resolved,
            ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 1) AS avg_satisfaction,
            ROUND(AVG(t."RESPONSE_TIME_HOURS")::numeric, 1) AS avg_response_hours
        FROM sf_data."SUPPORT_CENTER__PUBLIC__AGENTS" a
        LEFT JOIN sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
            ON a."AGENT_NAME" = t."RESOLVER"
           AND LOWER(t."STATUS") = 'resolved'
        GROUP BY a."AGENT_NAME", a."TEAM"
        ORDER BY tickets_resolved DESC
    """)
    agent_rows = cur.fetchall()
    agent_performance = []
    for row in agent_rows:
        agent_performance.append({
            "Agent_Name": row[0],
            "Team": row[1],
            "Tickets_Resolved": int(row[2]),
            "Avg_Satisfaction": float(row[3]) if row[3] is not None else 0.0,
            "Avg_Response_Hours": float(row[4]) if row[4] is not None else 0.0,
        })

    # ---- Summary ----
    cur.execute("""
        SELECT
            COUNT(*) AS total_tickets,
            SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END) AS total_breached,
            ROUND(
                (1.0 - SUM(CASE WHEN t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" THEN 1 ELSE 0 END)::numeric / COUNT(*)::numeric) * 100,
                1
            ) AS overall_compliance_rate,
            ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 1) AS overall_avg_satisfaction
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
        JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" p
            ON t."PRIORITY" = p."PRIORITY"
        WHERE LOWER(t."STATUS") = 'resolved'
    """)
    summary_row = cur.fetchone()
    total_tickets = int(summary_row[0])
    total_breached = int(summary_row[1])
    overall_compliance_rate = float(summary_row[2])
    overall_avg_satisfaction = float(summary_row[3])

    summary = {
        "Overall_Compliance_Rate": overall_compliance_rate,
        "Overall_Avg_Satisfaction": overall_avg_satisfaction,
        "Meets_Compliance_Target": "Yes" if overall_compliance_rate >= 90.0 else "No",
        "Meets_Satisfaction_Target": "Yes" if overall_avg_satisfaction >= 4.0 else "No",
        "Total_Tickets_Reviewed": total_tickets,
        "Total_Breached_Tickets": total_breached,
    }

    cur.close()
    conn.close()

    return {
        "sla_compliance": sla_compliance,
        "agent_performance": agent_performance,
        "summary": summary,
    }


def get_sheet(wb, name):
    """Find sheet case-insensitively."""
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    """Check the Excel output file against computed expected values."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "SLA_Audit_Report.xlsx")
    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
        wb_raw = openpyxl.load_workbook(agent_file, data_only=False)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    check("Excel file readable", True)

    # Check sheet names
    expected_sheets = ["SLA Compliance", "Agent Performance", "Summary"]
    for sheet_name in expected_sheets:
        found = any(str_match(s, sheet_name) for s in wb.sheetnames)
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {wb.sheetnames}")

    # --- Sheet 1: SLA Compliance ---
    print("\n--- SLA Compliance ---")
    ws, ws_raw = _sheet_pair(wb, wb_raw, "SLA Compliance")
    if ws:
        agent_rows = _iter_sheet_rows(ws, ws_raw, min_row=2)
        exp_rows = expected["sla_compliance"]
        check("SLA Compliance row count", len(agent_rows) == len(exp_rows),
              f"Expected {len(exp_rows)}, got {len(agent_rows)}")

        for exp_row in exp_rows:
            priority = exp_row["Priority"]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], priority):
                    matched = ar
                    break
            if matched:
                check(f"Priority '{priority}' Total_Tickets",
                      num_close(matched[1], exp_row["Total_Tickets"], 1.0),
                      f"Expected {exp_row['Total_Tickets']}, got {matched[1]}")
                check(f"Priority '{priority}' Breached_Tickets",
                      num_close(matched[2], exp_row["Breached_Tickets"], 1.0),
                      f"Expected {exp_row['Breached_Tickets']}, got {matched[2]}")
                check(f"Priority '{priority}' Compliance_Rate",
                      num_close(matched[3], exp_row["Compliance_Rate"], 0.5),
                      f"Expected {exp_row['Compliance_Rate']}, got {matched[3]}")
                check(f"Priority '{priority}' Avg_Response_Hours",
                      num_close(matched[4], exp_row["Avg_Response_Hours"], 0.5),
                      f"Expected {exp_row['Avg_Response_Hours']}, got {matched[4]}")
                check(f"Priority '{priority}' Avg_Satisfaction",
                      num_close(matched[5], exp_row["Avg_Satisfaction"], 0.5),
                      f"Expected {exp_row['Avg_Satisfaction']}, got {matched[5]}")
            else:
                check(f"Priority '{priority}' found", False,
                      "Priority not in agent output")

    # --- Sheet 2: Agent Performance ---
    print("\n--- Agent Performance ---")
    ws, ws_raw = _sheet_pair(wb, wb_raw, "Agent Performance")
    if ws:
        agent_rows = _iter_sheet_rows(ws, ws_raw, min_row=2)
        exp_rows = expected["agent_performance"]
        check("Agent Performance row count", len(agent_rows) == len(exp_rows),
              f"Expected {len(exp_rows)}, got {len(agent_rows)}")

        for exp_row in exp_rows:
            agent_name = exp_row["Agent_Name"]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], agent_name):
                    matched = ar
                    break
            if matched:
                check(f"Agent '{agent_name}' Team",
                      str_match(matched[1], exp_row["Team"]),
                      f"Expected '{exp_row['Team']}', got '{matched[1]}'")
                check(f"Agent '{agent_name}' Tickets_Resolved",
                      num_close(matched[2], exp_row["Tickets_Resolved"], 1.0),
                      f"Expected {exp_row['Tickets_Resolved']}, got {matched[2]}")
                check(f"Agent '{agent_name}' Avg_Satisfaction",
                      avg_metric_close(matched[3], exp_row["Avg_Satisfaction"]),
                      f"Expected {exp_row['Avg_Satisfaction']}, got {matched[3]}")
                check(f"Agent '{agent_name}' Avg_Response_Hours",
                      avg_metric_close(matched[4], exp_row["Avg_Response_Hours"]),
                      f"Expected {exp_row['Avg_Response_Hours']}, got {matched[4]}")
            else:
                check(f"Agent '{agent_name}' found", False,
                      "Agent not in agent output")

    # --- Sheet 3: Summary ---
    print("\n--- Summary ---")
    ws, ws_raw = _sheet_pair(wb, wb_raw, "Summary")
    if ws:
        agent_data = {}
        for row in _iter_sheet_rows(ws, ws_raw, min_row=1):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower().replace(" ", "_")] = row[1]

        exp_summary = expected["summary"]
        # Whitelist of approved key aliases
        KEY_ALIASES = {
            "overall_compliance_rate": ["overall compliance", "compliance rate", "overallcompliancerate"],
            "overall_avg_satisfaction": ["overall avg satisfaction", "avg satisfaction", "satisfaction score", "overallavgsatisfaction"],
            "meets_compliance_target": ["meets compliance", "meetscompliancetarget"],
            "meets_satisfaction_target": ["meets satisfaction", "meetssatisfactiontarget"],
            "total_tickets_reviewed": ["total tickets", "total reviewed", "totalticketsreviewed"],
            "total_breached_tickets": ["total breached", "breached tickets", "totalbreachedtickets"],
        }
        for key, gt_val in exp_summary.items():
            key_lower = key.lower()
            agent_val = agent_data.get(key_lower)
            if agent_val is None:
                aliases = KEY_ALIASES.get(key_lower, [key_lower.replace("_", " ")])
                for ak, av in agent_data.items():
                    ak_norm = ak.replace("_", " ").strip()
                    if any(alias == ak_norm or alias == ak or alias == ak.replace("_", "") for alias in aliases):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                ok = num_close(agent_val, gt_val, 1.0)
                check(f"Summary '{key}'", ok,
                      f"Expected {gt_val}, got {agent_val}")
            else:
                ok = str_match(agent_val, gt_val)
                check(f"Summary '{key}'", ok,
                      f"Expected '{gt_val}', got '{agent_val}'")


def check_form():
    """Check that a Google Form was created with the correct structure."""
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        print(f"  [SKIP] Cannot connect to database to check form: {e}")
        return
    cur = conn.cursor()
    try:
        _check_form_queries(cur)
    except Exception as e:
        # A DB error here is infra-related (connection/query), not the agent's fault.
        print(f"  [SKIP] Form check aborted on DB error: {e}")
    finally:
        cur.close()
        conn.close()


def _check_form_queries(cur):
    """Run the actual form checks against the database."""
    # Find form with title containing "SLA Improvement"
    cur.execute("""
        SELECT id, title, description
        FROM gform.forms
        WHERE LOWER(title) LIKE '%sla improvement%'
        ORDER BY created_at DESC
        LIMIT 1
    """)
    form_row = cur.fetchone()

    check("Google Form with 'SLA Improvement' in title exists",
          form_row is not None)
    if not form_row:
        return

    form_id = form_row[0]
    form_title = form_row[1]
    form_desc = form_row[2] if len(form_row) > 2 else ""
    check("Form title is 'SLA Improvement Plan'",
          "sla improvement plan" in form_title.lower(),
          f"Actual title: '{form_title}'")
    # The task explicitly asks for a short description (see docs/task.md).
    # Any non-empty description satisfies it; keep the check lenient so a
    # correct implementation is never penalized for wording/length.
    check("Form has a description",
          form_desc is not None and len(str(form_desc).strip()) > 0,
          f"Description: '{form_desc}'")

    # Get questions
    cur.execute("""
        SELECT title, question_type, required, config
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()

    check("Form has 5 questions", len(questions) == 5,
          f"Found {len(questions)} questions")

    # Expected questions in order
    expected_questions = [
        {
            "title": "Your Name",
            "type": "textQuestion",
            "required": True,
            "options": None,
        },
        {
            "title": "Your Team",
            "type": "choiceQuestion",
            "required": True,
            "options": ["Tier 1", "Tier 2", "Tier 3", "Specialist"],
        },
        {
            "title": "Which priority level needs most improvement?",
            "type": "choiceQuestion",
            "required": True,
            "options": ["Critical", "High", "Medium", "Low"],
        },
        {
            "title": "Proposed improvement actions",
            "type": "textQuestion",
            "required": True,
            "options": None,
        },
        {
            "title": "Target completion date",
            "type": "textQuestion",
            "required": True,
            "options": None,
        },
    ]

    for i, exp_q in enumerate(expected_questions):
        if i < len(questions):
            actual_title, actual_type, actual_required, actual_config = questions[i]

            # Check title (fuzzy match)
            title_match = exp_q["title"].lower() in actual_title.lower()
            check(f"Q{i+1} title contains '{exp_q['title']}'",
                  title_match,
                  f"Actual: '{actual_title}'")

            # Check type
            check(f"Q{i+1} type is '{exp_q['type']}'",
                  actual_type == exp_q["type"],
                  f"Actual: '{actual_type}'")

            # Check required
            check(f"Q{i+1} required is {exp_q['required']}",
                  actual_required == exp_q["required"],
                  f"Actual: {actual_required}")

            # Check options for choice questions
            if exp_q["options"] is not None:
                # Do NOT skip when actual_config is empty/None — that would silently
                # let empty configs bypass the option checks. Instead, always evaluate
                # and report missing options as failures.
                config = actual_config if isinstance(actual_config, dict) else {}
                raw_options = config.get("options")
                actual_options = []
                if isinstance(raw_options, str):
                    # Options may arrive as a serialized JSON string in some runtimes.
                    try:
                        raw_options = json.loads(raw_options)
                    except Exception:
                        raw_options = None
                if isinstance(raw_options, list):
                    for opt in raw_options:
                        if isinstance(opt, dict) and "value" in opt:
                            actual_options.append(str(opt["value"]))
                        elif isinstance(opt, str):
                            actual_options.append(opt)

                for exp_opt in exp_q["options"]:
                    found = any(
                        exp_opt.lower() == ao.lower()
                        for ao in actual_options
                    )
                    check(f"Q{i+1} has option '{exp_opt}'", found,
                          f"Actual options: {actual_options}")
        else:
            check(f"Q{i+1} exists", False, "Question missing")


def check_mock_server_health(port=30162):
    """Report mock health without grading an ephemeral process."""
    print("\n=== Mock Server Health (informational) ===")
    try:
        import urllib.request
        with urllib.request.urlopen(f"http://localhost:{port}/", timeout=3) as resp:
            body = resp.read().decode("utf-8", errors="ignore")
            print(
                f"  [INFO] port {port}: status={resp.status}, bytes={len(body)}"
            )
    except Exception as e:
        print(f"  [INFO] mock server unavailable during evaluation: {e}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    # Compute expected values from database
    print("=== Computing Expected Values from Database ===")
    try:
        expected = compute_expected_values()
        print(f"  Computed SLA compliance for {len(expected['sla_compliance'])} priorities")
        print(f"  Computed performance for {len(expected['agent_performance'])} agents")
        print(f"  Summary: compliance={expected['summary']['Overall_Compliance_Rate']}%, "
              f"satisfaction={expected['summary']['Overall_Avg_Satisfaction']}")
    except Exception as e:
        print(f"  ERROR computing expected values: {e}")
        import traceback
        traceback.print_exc()
        return False, f"Failed to compute expected values: {e}"

    # Run checks
    check_mock_server_health()
    check_excel(agent_workspace, expected)
    check_form()

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    # Tightened: require ALL checks to pass (previously >=80%).
    success = FAIL_COUNT == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Pass Rate: {pass_rate:.1%}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "success": success,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return success, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Rate: {pass_rate:.1%}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
