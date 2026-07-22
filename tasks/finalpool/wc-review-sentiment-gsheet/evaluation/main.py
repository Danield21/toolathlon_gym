"""Evaluation for wc-review-sentiment-gsheet."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected data from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Total reviews and avg rating
    cur.execute("SELECT COUNT(*), ROUND(AVG(rating)::numeric, 2) FROM wc.product_reviews")
    total_reviews, overall_avg = cur.fetchone()

    # Rating distribution
    cur.execute("""
        SELECT rating, COUNT(*),
               ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM wc.product_reviews), 1)
        FROM wc.product_reviews
        GROUP BY rating ORDER BY rating
    """)
    rating_dist = cur.fetchall()

    # Top 20 most-reviewed products
    cur.execute("""
        SELECT p.name, COUNT(r.id) as review_count,
               ROUND(AVG(r.rating)::numeric, 2) as avg_rating,
               SUM(CASE WHEN r.rating=5 THEN 1 ELSE 0 END) as five_star,
               SUM(CASE WHEN r.rating=1 THEN 1 ELSE 0 END) as one_star
        FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        ORDER BY COUNT(r.id) DESC, p.name ASC
        LIMIT 20
    """)
    top_products = cur.fetchall()

    # Most reviewed product (alphabetically first if tied)
    cur.execute("""
        SELECT p.name FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        ORDER BY COUNT(r.id) DESC, p.name ASC
        LIMIT 1
    """)
    most_reviewed = cur.fetchone()[0]

    # Highest rated (min 3 reviews)
    cur.execute("""
        SELECT p.name FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        HAVING COUNT(r.id) >= 3
        ORDER BY AVG(r.rating) DESC, p.name ASC
        LIMIT 1
    """)
    highest_rated = cur.fetchone()[0]

    cur.close()
    conn.close()
    return {
        "total_reviews": total_reviews,
        "overall_avg": float(overall_avg),
        "rating_dist": rating_dist,
        "top_products": top_products,
        "most_reviewed": most_reviewed,
        "highest_rated": highest_rated,
    }


def check_excel(agent_workspace):
    """Check Review_Sentiment_Analysis.xlsx."""
    print("\n=== Checking Review_Sentiment_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Review_Sentiment_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True
    expected = get_expected_data()

    # Check Product Reviews sheet
    pr_sheet = None
    for name in wb.sheetnames:
        if "product" in name.lower() and "review" in name.lower():
            pr_sheet = wb[name]
            break
    if pr_sheet is None:
        record("Sheet 'Product Reviews' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Product Reviews' exists", True)
        rows = [r for r in pr_sheet.iter_rows(min_row=2, values_only=True) if r and any(c is not None for c in r)]
        rows_ok = len(rows) == 20
        record("Product Reviews has exactly 20 rows", rows_ok, f"Got {len(rows)}")
        if not rows_ok:
            all_ok = False

        # Validate every product in agent's top-20 is in expected top-20 (set match, order-independent for now)
        if rows and expected["top_products"]:
            expected_top_names = {p[0].strip().lower() for p in expected["top_products"]}
            agent_names = [str(r[0]).strip().lower() if r and r[0] is not None else "" for r in rows]
            non_matching = [n for n in agent_names if n not in expected_top_names]
            record("All agent product names are in expected top-20",
                   len(non_matching) == 0,
                   f"Non-matching: {non_matching[:5]}")
            if non_matching:
                all_ok = False

            # Validate sort order: Review_Count descending, then product name ascending
            sort_violations = []
            for i in range(len(rows) - 1):
                if not rows[i] or not rows[i+1]:
                    continue
                rc_a = rows[i][1] if len(rows[i]) > 1 else None
                rc_b = rows[i+1][1] if len(rows[i+1]) > 1 else None
                name_a = str(rows[i][0] or "").strip()
                name_b = str(rows[i+1][0] or "").strip()
                try:
                    rc_a_f = float(rc_a) if rc_a is not None else None
                    rc_b_f = float(rc_b) if rc_b is not None else None
                except (TypeError, ValueError):
                    rc_a_f = rc_b_f = None
                if rc_a_f is not None and rc_b_f is not None:
                    if rc_a_f < rc_b_f:
                        sort_violations.append(f"row {i}: rc {rc_a_f} < {rc_b_f}")
                    elif rc_a_f == rc_b_f and name_a > name_b:
                        sort_violations.append(f"row {i}: tie at rc={rc_a_f}, '{name_a}' > '{name_b}'")
            record("Product Reviews sort: Review_Count desc, name asc",
                   len(sort_violations) == 0,
                   f"Violations: {sort_violations[:3]}")
            if sort_violations:
                all_ok = False

            # Validate per-row review_count, avg_rating, five_star, one_star match expected
            expected_by_name = {p[0].strip().lower(): p for p in expected["top_products"]}
            mismatch_count = 0
            for r in rows:
                if not r or r[0] is None:
                    continue
                name_key = str(r[0]).strip().lower()
                exp = expected_by_name.get(name_key)
                if exp is None:
                    continue
                exp_rc, exp_avg, exp_five, exp_one = exp[1], float(exp[2]), exp[3], exp[4]
                if len(r) > 1 and r[1] is not None and not num_close(r[1], exp_rc, 0):
                    mismatch_count += 1
                if len(r) > 2 and r[2] is not None and not num_close(r[2], exp_avg, 0.05):
                    mismatch_count += 1
                if len(r) > 3 and r[3] is not None and not num_close(r[3], exp_five, 0):
                    mismatch_count += 1
                if len(r) > 4 and r[4] is not None and not num_close(r[4], exp_one, 0):
                    mismatch_count += 1
            record("Product Reviews per-row metrics match expected",
                   mismatch_count == 0,
                   f"{mismatch_count} cell mismatches")
            if mismatch_count:
                all_ok = False

    # Check Rating Distribution sheet
    rd_sheet = None
    for name in wb.sheetnames:
        if "rating" in name.lower() and "distribution" in name.lower():
            rd_sheet = wb[name]
            break
    if rd_sheet is None:
        record("Sheet 'Rating Distribution' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Rating Distribution' exists", True)
        rows = list(rd_sheet.iter_rows(min_row=2, values_only=True))
        record("Rating Distribution has 5 rows", len(rows) == 5, f"Got {len(rows)}")

        for exp_r in expected["rating_dist"]:
            rating_val = exp_r[0]
            found = False
            for r in rows:
                if r and r[0] is not None and int(r[0]) == rating_val:
                    found = True
                    ok_count = num_close(r[1], exp_r[1], 2)
                    record(f"Rating {rating_val} count", ok_count,
                           f"Expected {exp_r[1]}, got {r[1]}")
                    if not ok_count:
                        all_ok = False
                    break
            if not found:
                record(f"Rating {rating_val} found", False, "Missing")
                all_ok = False

    # Check Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        summary = {}
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

        # Validate all required metrics: Total_Reviews, Avg_Rating, Most_Reviewed_Product, Highest_Rated_Product
        total_reviews_found = False
        avg_rating_found = False
        most_reviewed_found = False
        highest_rated_found = False
        for key, val in summary.items():
            if "total" in key and "review" in key:
                total_reviews_found = True
                ok = num_close(val, expected["total_reviews"], 0)  # exact integer
                record("Summary Total_Reviews", ok,
                       f"Expected {expected['total_reviews']}, got {val}")
                if not ok:
                    all_ok = False
            elif key == "avg_rating" or ("avg" in key and "rating" in key):
                avg_rating_found = True
                ok = num_close(val, expected["overall_avg"], 0.05)
                record("Summary Avg_Rating", ok,
                       f"Expected {expected['overall_avg']}, got {val}")
                if not ok:
                    all_ok = False
            elif "most" in key and "reviewed" in key:
                most_reviewed_found = True
                ok = val is not None and str(val).strip().lower() == expected["most_reviewed"].strip().lower()
                record("Summary Most_Reviewed_Product", ok,
                       f"Expected '{expected['most_reviewed']}', got '{val}'")
                if not ok:
                    all_ok = False
            elif "highest" in key and "rated" in key:
                highest_rated_found = True
                ok = val is not None and str(val).strip().lower() == expected["highest_rated"].strip().lower()
                record("Summary Highest_Rated_Product", ok,
                       f"Expected '{expected['highest_rated']}', got '{val}'")
                if not ok:
                    all_ok = False
        if not total_reviews_found:
            record("Summary Total_Reviews row", False, "Missing")
            all_ok = False
        if not avg_rating_found:
            record("Summary Avg_Rating row", False, "Missing")
            all_ok = False
        if not most_reviewed_found:
            record("Summary Most_Reviewed_Product row", False, "Missing")
            all_ok = False
        if not highest_rated_found:
            record("Summary Highest_Rated_Product row", False, "Missing")
            all_ok = False

    return all_ok


def check_word(agent_workspace):
    """Check Review_Sentiment_Report.docx."""
    print("\n=== Checking Review_Sentiment_Report.docx ===")
    from docx import Document

    docx_file = os.path.join(agent_workspace, "Review_Sentiment_Report.docx")
    if not os.path.isfile(docx_file):
        record("Word file exists", False, f"Not found: {docx_file}")
        return False
    record("Word file exists", True)

    try:
        doc = Document(docx_file)
    except Exception as e:
        record("Word readable", False, str(e))
        return False

    all_text = " ".join(p.text.lower() for p in doc.paragraphs)
    word_ok = True
    title_ok = "product review sentiment analysis" in all_text
    record("Word title 'Product Review Sentiment Analysis'", title_ok, all_text[:200])
    if not title_ok:
        word_ok = False
    date_ok = "2026-03-06" in all_text
    record("Word contains date '2026-03-06'", date_ok, "Date not found")
    if not date_ok:
        word_ok = False
    sent_ok = "sentiment" in all_text
    record("Word mentions 'sentiment'", sent_ok, "No mention of 'sentiment'")
    if not sent_ok:
        word_ok = False
    rev_ok = "review" in all_text
    record("Word mentions 'review'", rev_ok, "No mention of 'review'")
    if not rev_ok:
        word_ok = False

    return word_ok


def check_gsheet():
    """Check Google Sheet titled 'Review Sentiment Dashboard' with 'Rating Distribution' tab."""
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except Exception as e:
        record("GSheet DB connection", False, str(e))
        return False

    # Title check: must contain 'review sentiment' AND 'dashboard'
    cur.execute("""
        SELECT id, title FROM gsheet.spreadsheets
        WHERE LOWER(title) LIKE '%review sentiment%'
           AND LOWER(title) LIKE '%dashboard%'
    """)
    rows = cur.fetchall()
    found = len(rows) > 0
    if not found:
        cur.execute("SELECT title FROM gsheet.spreadsheets")
        all_titles = [r[0] for r in cur.fetchall()]
        record("GSheet titled 'Review Sentiment Dashboard'", False, f"Got titles: {all_titles}")
        cur.close()
        conn.close()
        return False
    record("GSheet titled 'Review Sentiment Dashboard'", True)

    spreadsheet_id = rows[0][0]
    # Check for 'Rating Distribution' tab
    cur.execute("""
        SELECT id, title FROM gsheet.sheets
        WHERE spreadsheet_id = %s AND LOWER(title) LIKE '%%rating distribution%%'
    """, (spreadsheet_id,))
    sheets = cur.fetchall()
    sheet_ok = len(sheets) > 0
    record("GSheet has 'Rating Distribution' tab", sheet_ok, "Tab not found")

    cur.close()
    conn.close()
    return found and sheet_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    word_ok = check_word(args.agent_workspace)
    gsheet_ok = check_gsheet()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    overall = excel_ok and word_ok and gsheet_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
