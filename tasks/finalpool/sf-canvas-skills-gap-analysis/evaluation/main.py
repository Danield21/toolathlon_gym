"""Evaluation for sf-canvas-skills-gap-analysis."""
import argparse
import os
import sys

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def nums_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _exact_dept_match(dept_names, target):
    """Match a department name exactly (case-insensitive, ignoring '&'/'and' variants)."""
    target_norm = target.strip().lower().replace("&", "and").replace("  ", " ")
    for d in dept_names:
        dn = str(d).strip().lower().replace("&", "and").replace("  ", " ")
        if dn == target_norm:
            return d
    return None


def check_excel(agent_workspace, groundtruth_workspace=None):
    errors = []
    import openpyxl

    path = os.path.join(agent_workspace, "Skills_Gap.xlsx")
    if not os.path.exists(path):
        return ["Skills_Gap.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Try to load GT for cross-validation
        gt_wb = None
        if groundtruth_workspace:
            gt_path = os.path.join(groundtruth_workspace, "Skills_Gap.xlsx")
            if os.path.exists(gt_path):
                try:
                    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
                except Exception:
                    gt_wb = None

        # Sheet 1: Department Skills
        rows = load_sheet_rows(wb, "Department Skills")
        if rows is None:
            errors.append("Sheet 'Department Skills' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) != 7:
                errors.append(
                    f"Department Skills has {len(data_rows)} rows, expected 7"
                )
            dept_names = [str(r[0]).strip() for r in data_rows if r[0]]
            # Exact match (no substring) for all 7 departments
            for expected in ["Engineering", "Sales", "Finance", "Operations", "Support", "HR", "R&D"]:
                if _exact_dept_match(dept_names, expected) is None:
                    errors.append(f"Department '{expected}' not found (exact match)")

            # Validate gap_score for ALL 7 departments via GT
            if gt_wb is not None:
                gt_rows = load_sheet_rows(gt_wb, "Department Skills") or []
                gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]
                gt_gap = {str(r[0]).strip().lower().replace("&", "and"): r[5]
                          for r in gt_data if len(r) > 5}
                a_gap = {str(r[0]).strip().lower().replace("&", "and"): r[5]
                         for r in data_rows if len(r) > 5}
                for dept_norm, gv in gt_gap.items():
                    av = a_gap.get(dept_norm)
                    if av is None:
                        errors.append(f"Gap_Score missing for dept '{dept_norm}'")
                        continue
                    try:
                        gf = float(gv) if gv is not None else 0.0
                        af = float(av) if av is not None else 0.0
                        if not nums_close(af, gf, abs_tol=0.02, rel_tol=0.05):
                            errors.append(
                                f"Gap_Score for '{dept_norm}' expected {gf}, got {af}"
                            )
                    except (TypeError, ValueError):
                        errors.append(
                            f"Gap_Score for '{dept_norm}' non-numeric: got {av}"
                        )

        # Sheet 2: Training Mapping
        rows2 = load_sheet_rows(wb, "Training Mapping")
        if rows2 is None:
            errors.append("Sheet 'Training Mapping' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 10:
                errors.append(
                    f"Training Mapping has {len(data_rows2)} rows, expected >= 10"
                )
            # Exact-match check for each required skill area against GT
            if gt_wb is not None:
                gt_rows2 = load_sheet_rows(gt_wb, "Training Mapping") or []
                gt_data2 = [r for r in gt_rows2[1:] if r and r[0] is not None]
                gt_skills = {str(r[0]).strip().lower() for r in gt_data2 if r[0]}
                a_skills = {str(r[0]).strip().lower() for r in data_rows2 if r[0]}
                missing = gt_skills - a_skills
                if missing:
                    errors.append(
                        f"Training Mapping missing skill areas: {sorted(missing)[:5]}"
                    )

        # Sheet 3: Priority Actions
        rows3 = load_sheet_rows(wb, "Priority Actions")
        if rows3 is None:
            errors.append("Sheet 'Priority Actions' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) != 7:
                errors.append(
                    f"Priority Actions has {len(data_rows3)} rows, expected 7"
                )
            # Per-dept Priority validation against GT
            if gt_wb is not None:
                gt_rows3 = load_sheet_rows(gt_wb, "Priority Actions") or []
                gt_data3 = [r for r in gt_rows3[1:] if r and r[0] is not None]
                gt_pri = {str(r[0]).strip().lower().replace("&", "and"):
                          str(r[2]).strip().lower() if len(r) > 2 and r[2] else ""
                          for r in gt_data3}
                a_pri = {str(r[0]).strip().lower().replace("&", "and"):
                         str(r[2]).strip().lower() if len(r) > 2 and r[2] else ""
                         for r in data_rows3}
                for dept_norm, gp in gt_pri.items():
                    if dept_norm not in a_pri:
                        errors.append(f"Priority Actions missing dept '{dept_norm}'")
                        continue
                    if a_pri[dept_norm] != gp:
                        errors.append(
                            f"Priority for '{dept_norm}' expected '{gp}', got '{a_pri[dept_norm]}'"
                        )
                # Per-dept Contact_Email
                gt_email = {str(r[0]).strip().lower().replace("&", "and"):
                            str(r[4]).strip().lower() if len(r) > 4 and r[4] else ""
                            for r in gt_data3}
                a_email = {str(r[0]).strip().lower().replace("&", "and"):
                           str(r[4]).strip().lower() if len(r) > 4 and r[4] else ""
                           for r in data_rows3}
                for dept_norm, ge in gt_email.items():
                    if dept_norm in a_email and ge and a_email[dept_norm] != ge:
                        errors.append(
                            f"Contact_Email for '{dept_norm}' expected '{ge}', got '{a_email[dept_norm]}'"
                        )

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_emails():
    errors = []
    try:
        import psycopg2

        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user=os.environ.get("PGUSER", "eigent"),
            password=os.environ.get("PGPASSWORD", "camel"),
        )
        cur = conn.cursor()
        cur.execute(
            """
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE subject ILIKE '%%training recommendation%%'
               OR subject ILIKE '%%training recommendations%%'
            ORDER BY id DESC LIMIT 30
        """
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # Departments with non-zero gap (per groundtruth): Sales, HR, R&D, Operations, Support
        gap_dept_map = {
            "sales": "m.rodriguez",
            "hr": "d.thompson",
            "r&d": "l.wang",
            "operations": "r.kim",
            "support": "a.foster",
        }
        if len(rows) < 5:
            errors.append(
                f"Expected at least 5 training recommendation emails (one per non-zero-gap dept), found {len(rows)}"
            )

        # Per-department: check there's an email whose subject contains the dept
        # name AND 'Training Recommendation', whose to_addr contains the dept
        # head email, AND whose body has substantive content (>30 chars).
        for dept, head in gap_dept_map.items():
            matched = False
            body_ok = False
            for subj, to_addr, body in rows:
                subj_l = (subj or "").lower()
                to_l = str(to_addr or "").lower()
                if dept in subj_l and head in to_l:
                    matched = True
                    if body and len(str(body).strip()) > 30:
                        body_ok = True
                    break
            if not matched:
                errors.append(
                    f"No 'Training Recommendations' email for dept '{dept}' (expected to {head}@...)"
                )
            elif not body_ok:
                errors.append(
                    f"Email for dept '{dept}' has empty or trivially short body"
                )
    except Exception as e:
        errors.append(f"Error checking emails: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, args.groundtruth_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"))
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking emails...")
    errs = check_emails()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
