"""Evaluation for canvas-at-risk-intervention."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def norm(s):
    return str(s or "").strip().lower().replace("_", " ")


def safe_float(v):
    try:
        if v is None: return None
        return float(v)
    except (ValueError, TypeError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _fetch_expected_per_course():
    """Fetch real expected at-risk counts and avg scores per course."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT c.name,
              COUNT(*) FILTER (WHERE (e.grades->>'current_score')::numeric < 50 AND e.grades->>'current_score' IS NOT NULL)::int,
              AVG((e.grades->>'current_score')::numeric) FILTER (WHERE e.grades->>'current_score' IS NOT NULL)::float
            FROM canvas.enrollments e
            JOIN canvas.courses c ON e.course_id = c.id
            WHERE e.type = 'StudentEnrollment'
            GROUP BY c.name
        """)
        result = {}
        for name, cnt, avg in cur.fetchall():
            result[name] = {"at_risk": int(cnt), "avg_score": round(float(avg), 1) if avg is not None else None}
        cur.execute("""
            SELECT COUNT(*) FROM canvas.enrollments e
            WHERE e.type = 'StudentEnrollment'
              AND (e.grades->>'current_score')::numeric < 50
              AND e.grades->>'current_score' IS NOT NULL
        """)
        total = cur.fetchone()[0]
        conn.close()
        return result, int(total)
    except Exception as e:
        print(f"[WARN] expected fetch: {e}")
        return None, None


def _top_risk_keywords(n=5):
    """Return lowercase leading-subject keywords for the top-N courses by at-risk
    count, derived from the database. Used so the email-body high-risk-course
    check tracks the data instead of being hardcoded to specific course names."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT c.name,
              COUNT(*) FILTER (WHERE (e.grades->>'current_score')::numeric < 50
                                 AND e.grades->>'current_score' IS NOT NULL)::int AS atrisk
            FROM canvas.enrollments e
            JOIN canvas.courses c ON e.course_id = c.id
            WHERE e.type = 'StudentEnrollment'
            GROUP BY c.name
            ORDER BY atrisk DESC, c.name
            LIMIT %s
        """, (n,))
        seen = []
        for name, _ in cur.fetchall():
            # Drop the "(Fall 2014)" term suffix, then take the leading subject
            # phrase up to the first '&' (e.g. "Creative Computing" from
            # "Creative Computing & Culture (Fall 2014)").
            base = str(name).split("(")[0].strip()
            head = base.split("&")[0].strip().lower()
            if head and head not in seen:
                seen.append(head)
        conn.close()
        return seen
    except Exception as e:
        print(f"[WARN] top risk keywords: {e}")
        return []


def check_excel(agent_workspace, expected_per_course, expected_total):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "At_Risk_Report.xlsx")
    if not os.path.isfile(xlsx_path):
        check("At_Risk_Report.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("At_Risk_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # === At Risk Students ===
    ar_rows = load_sheet_rows(wb, "At Risk Students")
    if ar_rows is None:
        check("Sheet 'At Risk Students' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'At Risk Students' exists", True)
        data_rows = [r for r in ar_rows[1:] if any(c is not None for c in r)]
        # Exact total at risk count from DB (deterministic)
        if expected_total:
            check(f"At Risk Students has exactly {expected_total} rows",
                  len(data_rows) == expected_total,
                  f"Found {len(data_rows)}, expected {expected_total}")

        header = ar_rows[0] if ar_rows else []
        header_norm = [norm(h) for h in header]
        for col in ["student name", "course", "current score", "risk level", "recommended support"]:
            check(f"Column '{col}' present", col in header_norm,
                  f"Header: {header}")

        if data_rows and "current score" in header_norm and "risk level" in header_norm:
            cs_idx = header_norm.index("current score")
            rl_idx = header_norm.index("risk level")

            # All scores < 50
            scores = [safe_float(r[cs_idx]) for r in data_rows if r[cs_idx] is not None]
            check("All scores < 50",
                  len(scores) > 0 and all(s < 50 for s in scores),
                  f"sample: {scores[:5]}")

            # Risk level values per row consistent with score
            mismatches = 0
            for r in data_rows:
                s = safe_float(r[cs_idx])
                if s is None: continue
                rl = norm(r[rl_idx])
                if s < 40 and rl != "critical":
                    mismatches += 1
                elif 40 <= s < 50 and rl != "warning":
                    mismatches += 1
            check(f"Risk level consistent with score (Critical < 40, Warning 40-50)",
                  mismatches == 0, f"mismatches: {mismatches}")

            # Sort ascending
            check("At Risk Students sorted by Current_Score ascending",
                  scores == sorted(scores), f"first 5: {scores[:5]}")

            # Both Critical and Warning levels present
            risk_values = {norm(r[rl_idx]) for r in data_rows if r[rl_idx]}
            check("Has both Critical and Warning risk levels",
                  "critical" in risk_values and "warning" in risk_values,
                  f"got: {risk_values}")

    # === Course Summary ===
    cs_rows = load_sheet_rows(wb, "Course Summary")
    if cs_rows is None:
        check("Sheet 'Course Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Course Summary' exists", True)
        data_rows = [r for r in cs_rows[1:] if any(c is not None for c in r)]
        if expected_per_course:
            check(f"Course Summary has {len(expected_per_course)} rows",
                  len(data_rows) == len(expected_per_course),
                  f"Found {len(data_rows)}, expected {len(expected_per_course)}")
        elif data_rows:
            check("Course Summary has 22 rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        header = cs_rows[0] if cs_rows else []
        header_norm = [norm(h) for h in header]
        for col in ["course", "at risk count", "avg score"]:
            check(f"Column '{col}' present", col in header_norm,
                  f"Header: {header}")

        if data_rows and expected_per_course and len(header_norm) >= 3:
            c_idx = header_norm.index("course") if "course" in header_norm else 0
            ar_idx = header_norm.index("at risk count") if "at risk count" in header_norm else 1
            avg_idx = header_norm.index("avg score") if "avg score" in header_norm else 2
            agent_courses = {}
            for r in data_rows:
                if r[c_idx]:
                    agent_courses[norm(r[c_idx])] = r
            # Check each expected course
            ok_count = 0
            ok_avg = 0
            missing = []
            for name, exp in expected_per_course.items():
                key = norm(name)
                row = agent_courses.get(key)
                if row is None:
                    missing.append(name)
                    continue
                actual_count = safe_float(row[ar_idx])
                actual_avg = safe_float(row[avg_idx])
                # at_risk count is an integer derived deterministically from DB
                if actual_count is not None and int(actual_count) == int(exp["at_risk"]):
                    ok_count += 1
                # avg_score tightened to 0.5 (was 2.0)
                if actual_avg is not None and exp["avg_score"] is not None and abs(actual_avg - exp["avg_score"]) <= 0.5:
                    ok_avg += 1
            n_courses = len(expected_per_course)
            check(f"All {n_courses} course at-risk counts correct (deterministic)",
                  ok_count == n_courses, f"correct: {ok_count}/{n_courses}, missing: {missing[:3]}")
            check(f"All {n_courses} avg_scores within tolerance",
                  ok_avg == n_courses, f"correct avg: {ok_avg}/{n_courses}")

            # Alphabetical sort
            course_names = [str(r[c_idx]) for r in data_rows if r[c_idx]]
            check("Course Summary sorted alphabetically",
                  course_names == sorted(course_names), f"first 3: {course_names[:3]}")

    # === Action Plan ===
    ap_rows = load_sheet_rows(wb, "Action Plan")
    if ap_rows is None:
        check("Sheet 'Action Plan' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Action Plan' exists", True)
        data_rows = [r for r in ap_rows[1:] if any(c is not None for c in r)]
        check("Action Plan has at least 5 actions", len(data_rows) >= 5,
              f"Found {len(data_rows)}")
        header = ap_rows[0] if ap_rows else []
        header_norm = [norm(h) for h in header]
        for col in ["priority", "action", "target group", "timeline", "responsible"]:
            check(f"Action Plan has '{col}' column", col in header_norm,
                  f"Header: {header}")
        # Verify the 5 required action types are mentioned
        all_text = " ".join(str(c).lower() for r in data_rows for c in r if c)
        for kw, tag in [
            ("mandatory", "mandatory meetings"),
            ("tutor", "tutoring arrangements"),
            ("study group", "study group creation"),
            ("counsel", "counseling referrals"),
            ("monitor", "progress monitoring"),
        ]:
            check(f"Action Plan covers {tag}",
                  kw in all_text,
                  f"keyword '{kw}' missing")


def check_email(expected_total, risk_keywords):
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
        """)
        all_msgs = cur.fetchall()
        # Match exact subject + recipient
        target = None
        for mid, subj, to_addr, body in all_msgs:
            if not subj: continue
            ok_subj = norm(subj) == norm("At-Risk Student Intervention Report")
            ok_to = "academic-support@uni.edu" in str(to_addr or "").lower()
            if ok_subj and ok_to:
                target = (mid, subj, to_addr, body)
                break
        check("Email sent: exact subject + recipient", target is not None,
              f"all subjects: {[m[1] for m in all_msgs]}")
        if target:
            body = str(target[3] or "").lower()
            check("Email body has substantive content (>100 chars)",
                  len(body) > 100, f"len={len(body)}")
            # Body should contain exact at-risk count.
            if expected_total:
                # Tolerate thousand-separator variants: "2,297" / "2 297" / "2297"
                # all represent the same number. Normalize the body by stripping
                # commas and spaces before the literal count check (B.1.5 pattern:
                # evaluator assumption too strict; task.md does not mandate a
                # specific number format).
                body_normalized = body.replace(",", "").replace(" ", "")
                check(f"Email body mentions exact at-risk count {expected_total}",
                      str(expected_total) in body_normalized,
                      f"body sample: {body[:200]}")
            # Body mentions courses with high at-risk counts: require >= 2 of the
            # top-N at-risk course keywords derived from the database (so the check
            # tracks the data instead of being hardcoded to specific course names).
            keywords = risk_keywords or ["creative computing", "biochemistry", "data-driven"]
            hits = sum(1 for t in keywords if t in body)
            check("Email body lists high-risk courses (>=2 top-course keywords)",
                  hits >= 2,
                  f"hits={hits}/{len(keywords)} keywords={keywords}; body sample: {body[:200]}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_per_course, expected_total = _fetch_expected_per_course()
    risk_keywords = _top_risk_keywords(5)
    check_excel(args.agent_workspace, expected_per_course, expected_total)
    check_email(expected_total, risk_keywords)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
