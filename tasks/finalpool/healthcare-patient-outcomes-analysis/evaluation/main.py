#!/usr/bin/env python3
"""Evaluation script for healthcare-patient-outcomes-analysis."""

import argparse
import os
import sys
from pathlib import Path

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432,
        dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
        user="eigent", password="camel",
    )


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    if not agent_workspace:
        return False, "No agent workspace provided"

    agent_ws = Path(agent_workspace)
    if not agent_ws.exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    # Detect GT self-test mode (V1 parity test) — DB checks for agent-side
    # deliverables (sent emails, calendar events) cannot be staged by GT files,
    # so they are tolerated as warnings during self-test.
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    # Phase 5: Excel evidence/outcomes report
    xlsx_files = [p for p in agent_ws.glob("*.xlsx")
                  if 'data.csv' not in p.name.lower()]
    check("At least one xlsx outcomes/evidence report exists",
          len(xlsx_files) >= 1, f"found {len(xlsx_files)} xlsx files")
    if xlsx_files:
        try:
            import openpyxl
            audit_path = max(xlsx_files, key=lambda p: p.stat().st_size)
            wb = openpyxl.load_workbook(str(audit_path), data_only=True)
            sheets = [s.lower() for s in wb.sheetnames]
            has_outcome_sheet = any(
                kw in s
                for s in sheets
                for kw in ['outcome', 'effectiveness', 'safety', 'protocol',
                          'comparison', 'analysis', 'summary', 'evidence',
                          'metric', 'tracking']
            )
            check(f"Outcomes xlsx ({audit_path.name}) has outcome-related sheet",
                  has_outcome_sheet, f"sheets={wb.sheetnames}")
            # Tightened min-rows: at least one sheet should have >=4 rows
            # (header + 3 protocol/provider/metric entries) to show real comparison.
            max_rows = max((ws.max_row for ws in wb.worksheets), default=0)
            check("Outcomes xlsx has >=4 rows of data in some sheet",
                  max_rows >= 4, f"max rows={max_rows}")

            # Value-level GT comparison if GT file exists
            gt_xlsx_path = Path(groundtruth_workspace) / "Clinical_Outcomes_Report.xlsx"
            if gt_xlsx_path.exists():
                try:
                    gt_wb = openpyxl.load_workbook(str(gt_xlsx_path), data_only=True)
                    # Collect all numeric cells from agent vs GT to test "field/value coverage"
                    agent_numerics = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    agent_numerics.append(float(v))
                    gt_numerics = []
                    for ws in gt_wb.worksheets:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    gt_numerics.append(float(v))
                    # Agent should have at least 3 numerics (e.g. effectiveness %)
                    check("Outcomes xlsx has >=3 numeric metric values",
                          len(agent_numerics) >= 3,
                          f"got {len(agent_numerics)} numerics")
                    # Soft value overlap: at least 1 of GT numbers within 20% tolerance
                    if gt_numerics and agent_numerics:
                        matches = 0
                        for gv in gt_numerics:
                            tol = max(abs(gv) * 0.20, 0.5)
                            if any(abs(av - gv) <= tol for av in agent_numerics):
                                matches += 1
                        # Tightened (per QA): require >=3 value matches within 20% tolerance
                        # GT has 24 numerics so >=3 is reasonable - ensures agent
                        # picked the right patient cohort, not just random numbers.
                        check("Outcomes xlsx value-overlap with GT (>=3 within 20%)",
                              matches >= 3,
                              f"matched {matches}/{len(gt_numerics)} GT values")

                    # Tightened (per QA): at least one canonical clinical-metric
                    # name should appear in agent headers. Generic numeric
                    # cells alone don't prove the analysis is about clinical
                    # outcomes - the column header is the strongest signal.
                    agent_headers = set()
                    for ws in wb.worksheets:
                        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                        for h in first:
                            if h:
                                agent_headers.add(
                                    str(h).strip().lower().replace('_', ' ').replace('-', ' '))
                    canonical_metrics = [
                        'effectiveness', 'adverse', 'readmit', 'readmission',
                        'length of stay', 'los', 'mortality', 'satisfaction',
                        'recovery', 'complication', 'survival', 'response rate',
                        'cure rate', 'remission', 'outcome', 'patient count',
                        'patient', 'protocol',
                    ]
                    matched_metrics = [
                        m for m in canonical_metrics
                        if any(m in h for h in agent_headers)
                    ]
                    check(
                        "Outcomes xlsx headers reference clinical metrics (>=2)",
                        len(matched_metrics) >= 2,
                        f"matched={matched_metrics}, headers={list(agent_headers)[:8]}",
                    )
                    gt_wb.close()
                except Exception as _e:
                    check("Outcomes xlsx GT value-comparison", False, str(_e))
        except Exception as e:
            check("Outcomes xlsx parse", False, str(e))

    # Phase 5: Word clinical evidence report
    docx_files = list(agent_ws.glob("*.docx"))
    check("At least one docx clinical evidence report exists",
          len(docx_files) >= 1, f"found {len(docx_files)} docx files")
    if docx_files:
        try:
            from docx import Document
            audit_docx = max(docx_files, key=lambda p: p.stat().st_size)
            doc = Document(str(audit_docx))
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word doc has substantive content (>=200 chars)",
                  len(text) >= 200, f"len={len(text)}")
            has_outcome = any(kw in text for kw in
                              ['outcome', 'patient', 'protocol', 'treatment',
                               'effectiveness', 'clinical'])
            has_recommendation = any(kw in text for kw in
                                     ['recommendation', 'best practice',
                                      'finding', 'evidence'])
            check("Word doc mentions outcome/protocol/treatment",
                  has_outcome, f"text head: {text[:200]}")
            check("Word doc mentions recommendations/findings/evidence",
                  has_recommendation, f"text head: {text[:200]}")
        except Exception as e:
            check("Word doc parse", False, str(e))

    # Phase 6: Email and Calendar (DB checks)
    # GT self-test: tolerated as WARNING (cannot stage emails/events as files);
    # real-agent run: BLOCKING (agent must actually send + schedule).
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, subject FROM email.messages
               WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s
                  OR body_text ILIKE %s""",
            ('%clinical%', '%outcome%', '%treatment%', '%clinical%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Clinical evidence email sent: 0 found (GT self-test, non-blocking)")
        else:
            check("Clinical evidence email sent",
                  len(rows) >= 1, f"found {len(rows)} matching emails")
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))

    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT summary FROM gcal.events
               WHERE summary ILIKE %s OR summary ILIKE %s OR summary ILIKE %s""",
            ('%clinical%', '%review%', '%protocol%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Calendar clinical review meeting scheduled: 0 found (GT self-test, non-blocking)")
        else:
            check("Calendar clinical review meeting scheduled",
                  len(rows) >= 1, f"found {len(rows)} matching events")
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
