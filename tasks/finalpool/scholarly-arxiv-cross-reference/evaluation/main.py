"""
Evaluation for scholarly-arxiv-cross-reference task.
Checks Excel sheets and email.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# scholarly.arxiv_papers has 5 papers, arxiv.papers has 6
# Overlap IDs: 1602.05629, 1812.06127, 1908.07873 (3 papers)
# Scholarly only: 2001.08361, 2005.14165 (2 papers)
# Arxiv only: 1207.00580, 1502.03167, 1912.04977 (3 papers)
OVERLAP_IDS = {"1602.05629", "1812.06127", "1908.07873"}
SCHOLARLY_ONLY_IDS = {"2001.08361", "2005.14165"}
ARXIV_ONLY_IDS = {"1207.00580", "1502.03167", "1912.04977"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == c:
                return i
    return None


def check_excel(agent_workspace):
    """Check Excel file."""
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Citation_Cross_Reference.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    # Check Scholarly_Papers sheet
    scholarly_rows = load_sheet_rows(wb, "Scholarly_Papers")
    if scholarly_rows is None:
        scholarly_rows = load_sheet_rows(wb, "Scholarly Papers")
    if scholarly_rows is not None:
        record("Sheet 'Scholarly_Papers' exists", True)
        data_rows = scholarly_rows[1:] if len(scholarly_rows) > 1 else []
        # scholarly.arxiv_papers has 5 papers
        record("Scholarly_Papers has exactly 5 rows", len(data_rows) == 5,
               f"Found {len(data_rows)} rows")
        # Check expected paper IDs
        sch_header = scholarly_rows[0] if scholarly_rows else []
        id_col = find_col(sch_header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        if id_col is not None:
            ids_in_sheet = {str(r[id_col]).strip() for r in data_rows if id_col < len(r) and r[id_col] is not None}
            expected_sch = OVERLAP_IDS | SCHOLARLY_ONLY_IDS
            record("Scholarly_Papers contains expected paper IDs",
                   expected_sch.issubset(ids_in_sheet),
                   f"missing={expected_sch - ids_in_sheet}, extra={ids_in_sheet - expected_sch}")
    else:
        record("Sheet 'Scholarly_Papers' exists", False, f"Available: {wb.sheetnames}")

    # Check Arxiv_Papers sheet
    arxiv_rows = load_sheet_rows(wb, "Arxiv_Papers")
    if arxiv_rows is None:
        arxiv_rows = load_sheet_rows(wb, "Arxiv Papers")
    if arxiv_rows is not None:
        record("Sheet 'Arxiv_Papers' exists", True)
        data_rows = arxiv_rows[1:] if len(arxiv_rows) > 1 else []
        record("Arxiv_Papers has 6 rows", len(data_rows) == 6,
               f"Found {len(data_rows)} rows")
        ax_header = arxiv_rows[0] if arxiv_rows else []
        id_col_ax = find_col(ax_header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        if id_col_ax is not None:
            ids_ax = {str(r[id_col_ax]).strip() for r in data_rows if id_col_ax < len(r) and r[id_col_ax] is not None}
            expected_ax = OVERLAP_IDS | ARXIV_ONLY_IDS
            record("Arxiv_Papers contains expected paper IDs",
                   expected_ax.issubset(ids_ax),
                   f"missing={expected_ax - ids_ax}, extra={ids_ax - expected_ax}")
    else:
        record("Sheet 'Arxiv_Papers' exists", False, f"Available: {wb.sheetnames}")

    # Check Overlap_Analysis sheet
    overlap_rows = load_sheet_rows(wb, "Overlap_Analysis")
    if overlap_rows is None:
        overlap_rows = load_sheet_rows(wb, "Overlap Analysis")
    if overlap_rows is not None:
        record("Sheet 'Overlap_Analysis' exists", True)
        header = overlap_rows[0] if overlap_rows else []
        data_rows = overlap_rows[1:] if len(overlap_rows) > 1 else []

        id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        in_scholarly_col = find_col(header, ["In_Scholarly", "In Scholarly", "in_scholarly"])
        in_arxiv_col = find_col(header, ["In_Arxiv", "In Arxiv", "in_arxiv"])

        record("Overlap has Paper_ID column", id_col is not None, f"Header: {header}")
        record("Overlap has In_Scholarly column", in_scholarly_col is not None, f"Header: {header}")
        record("Overlap has In_Arxiv column", in_arxiv_col is not None, f"Header: {header}")

        # Check total rows (should be union of both databases = 8)
        record("Overlap_Analysis has exactly 8 rows", len(data_rows) == 8,
               f"Found {len(data_rows)} rows")

        # Check that each ID is correctly flagged
        if id_col is not None and in_scholarly_col is not None and in_arxiv_col is not None:
            overlap_found = 0
            scholarly_only_found = 0
            arxiv_only_found = 0
            for row in data_rows:
                if id_col >= len(row) or not row[id_col]:
                    continue
                pid = str(row[id_col]).strip()
                s_val = str(row[in_scholarly_col]).strip().lower() if in_scholarly_col < len(row) and row[in_scholarly_col] is not None else ""
                a_val = str(row[in_arxiv_col]).strip().lower() if in_arxiv_col < len(row) and row[in_arxiv_col] is not None else ""
                s_yes = "yes" in s_val
                a_yes = "yes" in a_val
                if pid in OVERLAP_IDS and s_yes and a_yes:
                    overlap_found += 1
                elif pid in SCHOLARLY_ONLY_IDS and s_yes and not a_yes:
                    scholarly_only_found += 1
                elif pid in ARXIV_ONLY_IDS and not s_yes and a_yes:
                    arxiv_only_found += 1

            record("Overlap papers marked Yes/Yes (3)", overlap_found == 3,
                   f"Found {overlap_found}/3 correctly marked overlap papers")
            record("Scholarly-only papers marked Yes/No (2)",
                   scholarly_only_found == 2,
                   f"Found {scholarly_only_found}/2")
            record("Arxiv-only papers marked No/Yes (3)",
                   arxiv_only_found == 3,
                   f"Found {arxiv_only_found}/3")
    else:
        record("Sheet 'Overlap_Analysis' exists", False, f"Available: {wb.sheetnames}")


def check_email():
    """Check email was sent."""
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find sent folder ID
        cur.execute("SELECT id FROM email.folders WHERE name='Sent'")
        sent_row = cur.fetchone()
        if not sent_row:
            cur.execute("SELECT id FROM email.folders WHERE name ILIKE '%%sent%%' LIMIT 1")
            sent_row = cur.fetchone()

        expected_subject = "Cross-Reference Analysis: Scholarly vs Arxiv Paper Databases"
        cur.execute("SELECT id, subject, from_addr, to_addr, body_text FROM email.messages")
        all_emails = cur.fetchall()

        target = None
        for em in all_emails:
            subj = (em[1] or "").strip().lower()
            if subj == expected_subject.lower():
                target = em
                break
        record("Email with exact subject 'Cross-Reference Analysis: Scholarly vs Arxiv Paper Databases'",
               target is not None,
               f"Subjects: {[e[1] for e in all_emails]}")

        if target is not None:
            _, subject, from_addr, to_addr, body_text = target
            from_str = str(from_addr or "").lower()
            record("Email from librarian@university.edu",
                   "librarian@university.edu" in from_str,
                   f"From: {from_addr}")
            try:
                if isinstance(to_addr, str):
                    to_parsed = json.loads(to_addr)
                else:
                    to_parsed = to_addr
            except Exception:
                to_parsed = to_addr
            to_str = json.dumps(to_parsed).lower() if to_parsed is not None else ""
            record("Email to research-lead@university.edu",
                   "research-lead@university.edu" in to_str,
                   f"To: {to_addr}")

            import re
            body = str(body_text).lower() if body_text else ""

            # Word-boundary digit checks contextualized to a count-word
            # Scholarly count 5 — match patterns like "5 papers ... scholarly", "scholarly ... 5",
            # accept also "5 in scholarly".
            sch5_re = (
                r"\b5\s*(?:papers|in\s+scholarly|scholarly)"
                r"|scholarly[^\n]{0,80}\b5\b"
                r"|\b5\b[^\n]{0,80}scholarly"
            )
            arxiv6_re = (
                r"\b6\s*(?:papers|in\s+arxiv|arxiv)"
                r"|arxiv[^\n]{0,80}\b6\b"
                r"|\b6\b[^\n]{0,80}arxiv"
            )
            overlap3_re = (
                r"(?:overlap|shared|common|both)[^\n]{0,80}\b3\b"
                r"|\b3\b[^\n]{0,80}(?:overlap|shared|common|both)"
            )

            record("Body mentions Scholarly count 5 (contextual)",
                   re.search(sch5_re, body) is not None,
                   f"Body preview: {body[:200]}")
            record("Body mentions Arxiv count 6 (contextual)",
                   re.search(arxiv6_re, body) is not None,
                   f"Body preview: {body[:200]}")
            record("Body mentions overlap count 3 (contextual)",
                   re.search(overlap3_re, body) is not None,
                   f"Body preview: {body[:200]}")
            record("Body mentions overlap/shared/common term",
                   "overlap" in body or "common" in body or "shared" in body,
                   f"Body preview: {body[:200]}")
            # Mention unique papers
            record("Body mentions unique papers",
                   "unique" in body or "only in" in body or "exclusive" in body,
                   f"Body preview: {body[:200]}")
            # Both database names mentioned
            record("Body mentions 'scholarly' database",
                   "scholarly" in body, f"Body preview: {body[:200]}")
            record("Body mentions 'arxiv' database",
                   "arxiv" in body, f"Body preview: {body[:200]}")

        conn.close()
    except Exception as e:
        record("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
