"""Evaluation for wc-order-refund-analysis."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_expected_refund_summary():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*), ROUND(SUM(total::numeric), 2), ROUND(AVG(total::numeric), 2) FROM wc.orders WHERE status = 'refunded'")
    cnt, total, avg = cur.fetchone()
    cur.close()
    conn.close()
    return {"count": int(cnt), "total": float(total) if total is not None else 0.0, "avg": float(avg) if avg is not None else 0.0}


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Refund_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Refund_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    print("--- Refund Details ---")
    agent_ws = get_sheet(agent_wb, "Refund Details")
    gt_ws = get_sheet(gt_wb, "Refund Details")
    check("Sheet 'Refund Details' exists", agent_ws is not None, f"Found: {agent_wb.sheetnames}")

    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = [r for r in agent_rows if any(c is not None for c in r)]
        check("Refund Details row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        # Sort order: Order_ID ascending
        oids = [r[0] for r in agent_rows if r and r[0] is not None]
        check("Sorted by Order_ID ascending", oids == sorted(oids), f"got {oids}")

        for gt_row in gt_rows:
            oid, cname, total, date_str = gt_row
            matched = None
            for ar in agent_rows:
                if ar and num_close(ar[0], oid, 0):
                    matched = ar
                    break
            if matched:
                check(f"Order {oid} Customer_Name == '{cname}'",
                      str_match(matched[1], cname),
                      f"got '{matched[1]}'")
                check(f"Order {oid} Total ≈ {total}",
                      num_close(matched[2], total, 0.01),
                      f"got {matched[2]}")
                # Date: accept date object or YYYY-MM-DD string
                date_val = matched[3]
                if hasattr(date_val, "isoformat"):
                    date_str_a = date_val.isoformat()[:10]
                else:
                    date_str_a = str(date_val).strip()[:10]
                check(f"Order {oid} Date == '{date_str}'",
                      date_str_a == date_str,
                      f"got '{date_str_a}'")
            else:
                check(f"Order {oid} found", False)

    print("--- Summary ---")
    agent_sum = get_sheet(agent_wb, "Summary")
    gt_sum = get_sheet(gt_wb, "Summary")
    check("Sheet 'Summary' exists", agent_sum is not None, f"Found: {agent_wb.sheetnames}")

    if agent_sum and gt_sum:
        gt_data = {}
        for row in gt_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_data[str(row[0]).strip().lower()] = row[1]
        agent_data = {}
        for row in agent_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]
        for key, gt_val in gt_data.items():
            agent_val = agent_data.get(key)
            if isinstance(gt_val, (int, float)):
                check(f"Summary '{key}' == {gt_val}",
                      num_close(agent_val, gt_val, 0.01),
                      f"got {agent_val}")
            else:
                check(f"Summary '{key}' == '{gt_val}'",
                      str_match(agent_val, gt_val),
                      f"got '{agent_val}'")

    return True


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return False
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    conn.close()

    expected = get_expected_refund_summary()

    def find_email_for_recipient(recipient):
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr is None:
                continue
            recipients = []
            if isinstance(to_addr, list):
                recipients = [str(r).strip().lower() for r in to_addr]
            elif isinstance(to_addr, str):
                try:
                    parsed = json.loads(to_addr)
                    if isinstance(parsed, list):
                        recipients = [str(r).strip().lower() for r in parsed]
                    else:
                        recipients = [str(to_addr).strip().lower()]
                except (json.JSONDecodeError, TypeError):
                    recipients = [str(to_addr).strip().lower()]
            if recipient.lower() in recipients:
                return subj, from_addr, to_addr, body
        return None

    result = find_email_for_recipient("manager@store.example.com")
    check("Summary email sent to manager@store.example.com", result is not None)
    if not result:
        return False
    subj, from_addr, to_addr, body = result
    check("Email subject == 'Refund Analysis Report'",
          (subj or "").strip().lower() == "refund analysis report",
          f"got: {subj}")
    check("Email from_addr == 'analytics@store.example.com'",
          str(from_addr or "").strip().lower() == "analytics@store.example.com",
          f"got: {from_addr}")
    body_text = body or ""
    body_lower = body_text.lower()
    check(f"Email body mentions total refund count {expected['count']}",
          str(expected["count"]) in body_text,
          f"missing {expected['count']}")
    total_2 = f"{expected['total']:.2f}"
    total_int = str(int(expected['total']))
    check(f"Email body mentions total amount {expected['total']}",
          total_2 in body_text or total_int in body_text or f"{int(expected['total']):,}" in body_text,
          f"missing total {expected['total']}")
    avg_2 = f"{expected['avg']:.2f}"
    check(f"Email body mentions average refund {expected['avg']}",
          avg_2 in body_text or str(int(expected['avg'])) in body_text,
          f"missing avg {expected['avg']}")
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("WC ORDER REFUND ANALYSIS - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, gt_dir)
    check_emails()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
