"""Evaluation script for pw-canvas-quiz-benchmark-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-canvas-quiz-benchmark-excel-notion"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Quiz_Benchmark_Report.xlsx")
    check("Quiz_Benchmark_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Quiz_Benchmark_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames and gt_wb is not None:
            ws = wb["Data_Analysis"]
            gt_ws = gt_wb["Data_Analysis"] if "Data_Analysis" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Data_Analysis has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Data_Analysis', ['Course', 'Code', 'Enrollment', 'Avg_Score', 'Pass_Rate'],
                          len(gt_rows) if gt_ws is not None else 6)
            if gt_ws is not None:
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                header_map = {h: i for i, h in enumerate(headers)}
                code_idx_gt = gt_headers.index("code") if "code" in gt_headers else 1
                code_idx_a = header_map.get("code", 1)
                gt_by_code = {str(r[code_idx_gt]).strip().upper(): r for r in gt_rows if r and r[code_idx_gt]}
                agent_by_code = {str(r[code_idx_a]).strip().upper(): r for r in data_rows if r and len(r) > code_idx_a and r[code_idx_a]}
                for code, gt_row in gt_by_code.items():
                    found = code in agent_by_code
                    check(f"Data_Analysis course code '{code}' present", found)
                    if found:
                        agent_row = agent_by_code[code]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row) or gt_h == "code":
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                tol = max(0.5, abs(gf) * 0.03)
                                check(f"Data_Analysis '{code}' {gt_h} ~{gf}",
                                      abs(gf - af) <= tol, f"got {av}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames and gt_wb is not None:
            ws = wb["Metrics"]
            gt_ws = gt_wb["Metrics"] if "Metrics" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Metrics has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Metrics', ['Metric', 'Value'], len(gt_rows) if gt_ws is not None else 4)
            if gt_ws is not None:
                gt_metric_map = {str(r[0]).strip().lower(): r[1] for r in gt_rows if r and r[0]}
                agent_metric_map = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for gt_m, gt_v in gt_metric_map.items():
                    found = gt_m in agent_metric_map
                    check(f"Metrics has '{gt_m}'", found)
                    if found:
                        av = agent_metric_map[gt_m]
                        gf = safe_float(gt_v)
                        af = safe_float(av)
                        if gf is not None and af is not None:
                            tol = max(0.5, abs(gf) * 0.03)
                            check(f"Metrics '{gt_m}' ~{gf}",
                                  abs(gf - af) <= tol, f"got {av}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check_columns('Recommendations', ['Priority', 'Action', 'Course'], 2)

        # Validate alphabetical sort by primary dimension (Course or first column)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            primary_idx = headers.index("course") if "course" in headers else 0
            primary_vals = [str(r[primary_idx]).strip() for r in data_rows
                            if r and primary_idx < len(r) and r[primary_idx]]
            sorted_vals = sorted(primary_vals, key=lambda s: s.lower())
            check("Data_Analysis sorted alphabetically by primary dimension",
                  primary_vals == sorted_vals,
                  f"order={primary_vals[:3]}")

    # Notion: page title 'Course Quiz Dashboard' — match the title field
    # specifically, not arbitrary properties text. Hoisted out of excel guard.
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Title in Notion is stored as the rich-text content under properties.title.
        # We require a properties.title plain_text that contains the marker phrase.
        cur.execute(
            """SELECT id, properties FROM notion.pages
               WHERE archived = false
                 AND properties::text ILIKE %s""",
            ('%Course Quiz Dashboard%',),
        )
        pages = cur.fetchall()
        # Filter to pages whose title text actually matches (rich_text plain_text).
        import json as _json
        matched = []
        for pid, props in pages:
            try:
                if isinstance(props, str):
                    pj = _json.loads(props)
                else:
                    pj = props
                # walk for any title-typed property's plain_text
                title_texts = []
                if isinstance(pj, dict):
                    for v in pj.values():
                        if isinstance(v, dict) and v.get("type") == "title":
                            for rt in v.get("title", []) or []:
                                if isinstance(rt, dict):
                                    title_texts.append(rt.get("plain_text", ""))
                blob = " ".join(title_texts).lower()
                if "course quiz dashboard" in blob:
                    matched.append(pid)
            except Exception:
                # Strict: if title parsing fails, do NOT accept substring match
                # (avoids false positives where the marker phrase appears in
                # arbitrary property content rather than the actual title).
                continue
        check("Notion 'Course Quiz Dashboard' page created",
              len(matched) >= 1,
              f"raw_substr_matches={len(pages)} title_matches={len(matched)}")
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))

    check("course_quiz_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "course_quiz_processor.py")))

    # Validate course_quiz_results.json content if produced
    results_path = os.path.join(agent_workspace, "course_quiz_results.json")
    if os.path.exists(results_path):
        try:
            with open(results_path) as _rf:
                _r = json.load(_rf)
            check("course_quiz_results.json is a non-empty object/array",
                  bool(_r) and isinstance(_r, (dict, list)),
                  f"type={type(_r).__name__}")
        except Exception as e:
            check("course_quiz_results.json valid JSON", False, str(e))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
