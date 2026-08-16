"""Evaluation script for fetch-arxiv-citation-network-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-arxiv-citation-network-excel-notion"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
WARN_COUNT = 0
IS_GT_SELF_TEST = False

def check(name, condition, detail="", db_side=False):
    global PASS_COUNT, FAIL_COUNT, WARN_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        if IS_GT_SELF_TEST and db_side:
            WARN_COUNT += 1
            detail_str = str(detail)[:200] if detail else ""
            print(f"  [WARN] {name} (GT self-test, DB-side): {detail_str}")
        else:
            FAIL_COUNT += 1
            detail_str = str(detail)[:200] if detail else ""
            print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, WARN_COUNT, IS_GT_SELF_TEST
    PASS_COUNT = 0
    FAIL_COUNT = 0
    WARN_COUNT = 0

    # Detect GT self-test mode (V1 parity check)
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        IS_GT_SELF_TEST = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        IS_GT_SELF_TEST = False

    excel_path = os.path.join(agent_workspace, "Citation_Network_Report.xlsx")
    check("Citation_Network_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Citation_Network_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Area', 'Citations', 'Relevance_Score'], 4)

            # Value-level validation: Area names must match GT areas (alphabetical sort allowed)
            if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames:
                try:
                    _agent_headers, _agent_rows = get_sheet_rows_as_dicts(wb, "Data_Analysis") if _HAS_VERIFY_V2 else (None, None)
                    gt_ws = gt_wb["Data_Analysis"]
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    gt_areas = {str(r[0]).strip().lower() for r in gt_rows if r and r[0]}
                    agent_first_col = {str(r[0]).strip().lower() for r in data_rows if r and r[0]}
                    overlap = gt_areas & agent_first_col
                    check("Data_Analysis Area values match GT (>=3 of 4)",
                          len(overlap) >= 3,
                          f"overlap={overlap}, gt={gt_areas}")
                except Exception as _e:
                    check("Data_Analysis Area value-check", False, str(_e))
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 4)

            # Value-level: must have Total_Papers metric set to 4
            metric_dict = {}
            for r in data_rows:
                if r and r[0] is not None and len(r) >= 2:
                    metric_dict[str(r[0]).strip().lower().replace(' ', '_')] = r[1]
            tp = metric_dict.get('total_papers')
            check("Metrics Total_Papers ~= 4 (3-5 acceptable)",
                  tp is not None and 3 <= safe_float(tp, 0) <= 5,
                  f"got Total_Papers={tp}")
            top = (str(metric_dict.get('top_area', '')) or '').lower()
            check("Metrics Top_Area is LLMs",
                  'llm' in top or 'llms' in top,
                  f"got Top_Area={top!r}")
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action'], 2)
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Tightened: must be a page whose title contains BOTH 'arxiv' AND 'citation' AND ('dashboard' or 'analysis')
            # This excludes the noise 'Old Project Notes' page and any unrelated 'dashboard' pages.
            cur.execute(
                """SELECT id, properties FROM notion.pages
                   WHERE archived = false
                     AND properties::text ILIKE %s
                     AND properties::text ILIKE %s
                     AND (properties::text ILIKE %s OR properties::text ILIKE %s)""",
                ('%arxiv%', '%citation%', '%dashboard%', '%analysis%'))
            pages = cur.fetchall()
            check("Notion 'Arxiv Citation Dashboard' page created",
                  len(pages) >= 1, f"found {len(pages)} matching pages",
                  db_side=True)
            conn.close()
        except Exception as e:
            check("Notion check", False, str(e), db_side=True)

        # Processor script: required at runtime; treated as runtime-side check
        # so V1 GT self-test (where the script isn't authored) doesn't block.
        proc_path = os.path.join(agent_workspace, "arxiv_citation_processor.py")
        check("arxiv_citation_processor.py exists",
              os.path.exists(proc_path), "", db_side=True)
        if os.path.exists(proc_path):
            try:
                with open(proc_path, "r", encoding="utf-8", errors="ignore") as _pf:
                    _src = _pf.read()
                check("arxiv_citation_processor.py is non-trivial (>=200 bytes)",
                      len(_src) >= 200, f"size={len(_src)} bytes")
                check("arxiv_citation_processor.py has Python imports",
                      ('import ' in _src) or ('from ' in _src and 'import' in _src),
                      "no import statement found")
            except OSError as e:
                check("arxiv_citation_processor.py readable", False, str(e))

        # Optional Gap column awareness: task.md describes a "gap or difference"
        # column. If the agent provides one, validate it equals Citations - Relevance_Score
        # (or close); if not provided, no penalty (GT also omits it).
        try:
            if "Data_Analysis" in wb.sheetnames:
                _ws = wb["Data_Analysis"]
                _hdrs = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                gap_col_idx = None
                for _i, _h in enumerate(_hdrs):
                    if _h in ("gap", "difference", "diff", "delta"):
                        gap_col_idx = _i; break
                if gap_col_idx is not None:
                    rows = list(_ws.iter_rows(min_row=2, values_only=True))
                    nonblank_gap = sum(1 for r in rows
                                       if r and len(r) > gap_col_idx and r[gap_col_idx] is not None)
                    check("Data_Analysis Gap column has values",
                          nonblank_gap >= 2,
                          f"gap col idx={gap_col_idx}, populated={nonblank_gap}")
        except Exception:
            pass


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
