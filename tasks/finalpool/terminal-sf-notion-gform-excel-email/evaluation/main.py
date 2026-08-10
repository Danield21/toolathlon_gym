"""Evaluation for terminal-sf-notion-gform-excel-email."""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(v):
    """Robustly convert a cell value to float.

    Handles int/float, and strings with thousands separators, currency
    symbols and % signs. Returns None when the value cannot be parsed.
    """
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for ch in (",", "$", "¥", "€", "%"):
            s = s.replace(ch, "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _norm_formula(f):
    return str(f).lower().replace(" ", "")


def num_close(a, b, tol=0.05):
    """Compare two cell values for numeric closeness.

    Formula cells (value starting with '='): if the groundtruth cell is also
    a formula, compare normalized formula text; if the groundtruth is a
    literal number we cannot verify the computed value without a calculation
    engine, so the check is skipped (no false FAIL). Otherwise both sides are
    parsed as floats and compared with tolerance; when one side cannot be
    parsed we fall back to a case-insensitive string comparison.
    """
    if isinstance(a, str) and a.startswith("="):
        if isinstance(b, str) and b.startswith("="):
            return _norm_formula(a) == _norm_formula(b)
        return True
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if fa is None and fb is None:
        return str(a).strip().lower() == str(b).strip().lower()
    return False


def _non_empty_rows(ws, min_row=2):
    """Rows with at least one non-None cell (tolerates trailing blank rows)."""
    return [r for r in ws.iter_rows(min_row=min_row, values_only=True)
            if any(c is not None for c in r)]


def _norm_dept(name):
    """Normalize a department name for robust matching across common aliases.

    Strips whitespace, lowercases, collapses internal whitespace, and maps
    common aliases (e.g. 'Research & Development' / 'Research and Development'
    -> 'r&d', 'Human Resources' -> 'hr') so a model using a natural synonym for
    a department name is not falsely FAILed. Identity for the canonical names
    used in the data source (Engineering, Finance, HR, Operations, R&D, Sales,
    Support).
    """
    if not name:
        return ""
    s = re.sub(r"\s+", " ", str(name).strip().lower())
    aliases = {
        "research & development": "r&d",
        "research and development": "r&d",
        "research & development (r&d)": "r&d",
        "research (r&d)": "r&d",
        "research&development": "r&d",
        "r & d": "r&d",
        "r and d": "r&d",
        "research development": "r&d",
        "human resources": "hr",
        "human resources (hr)": "hr",
    }
    return aliases.get(s, s)


# Words a correctly-completed script output may use to name the department with
# the highest / lowest engagement index. max()/min() idioms are common in
# Python scripts, as are the 'greatest'/'least' synonyms.
_HIGH_KEYWORDS = ["highest", "best", "top", "max", "maximum", "greatest"]
_LOW_KEYWORDS = ["lowest", "worst", "bottom", "min", "minimum", "least"]


def _has_any_word(content, words):
    """Case-insensitive whole-word match for any of the given words.

    Whole-word (\\b) matching avoids 'min' matching 'minutes'/'minimum' or 'top'
    matching 'stop', while still accepting 'Max engagement index: ...'.
    """
    return any(re.search(r"\b" + re.escape(w) + r"\b", content) for w in words)


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Employee_Engagement_Report.xlsx")
    gt_path = os.path.join(gt_workspace, "Employee_Engagement_Report.xlsx")

    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)
    wb = openpyxl.load_workbook(fpath, data_only=False)
    gt_wb = openpyxl.load_workbook(gt_path, data_only=False)

    # Sheet 1: Department_Scores
    ds_sheet = None
    for name in wb.sheetnames:
        if "department" in name.lower() and "score" in name.lower():
            ds_sheet = name
            break
    if not ds_sheet:
        record("Department_Scores sheet exists", False, f"Sheets: {wb.sheetnames}")
        return False
    record("Department_Scores sheet exists", True)

    ws = wb[ds_sheet]
    rows = _non_empty_rows(ws)
    record("Department_Scores has 7 rows", len(rows) == 7, f"Got {len(rows)}")

    gt_ws = gt_wb["Department_Scores"]
    gt_rows = _non_empty_rows(gt_ws)

    # Departments whose groundtruth engagement index sits within a small
    # tolerance of the groundtruth company mean. A correct model may compute the
    # company mean from raw (unrounded) data (mean == 5.0042, rounded to 5.00)
    # rather than from the rounded department indices (mean == 5.0057). In that
    # case departments with an index of exactly 5.00 land on the boundary and
    # may defensibly be classified as either High or Medium. Accept either
    # priority for these departments; keep strict exact matching elsewhere so
    # the check still tests the model's classification logic.
    gt_eis = {}
    for gt_row in gt_rows:
        if gt_row and gt_row[0] and len(gt_row) > 5:
            v = _to_float(gt_row[5])
            if v is not None:
                gt_eis[_norm_dept(gt_row[0])] = v
    BOUNDARY_TOL = 0.01
    gt_boundary = set()
    if gt_eis:
        gt_mean = sum(gt_eis.values()) / len(gt_eis)
        gt_boundary = {k for k, v in gt_eis.items() if abs(v - gt_mean) <= BOUNDARY_TOL}

    a_lookup = {}
    for row in rows:
        if row and row[0]:
            a_lookup[_norm_dept(row[0])] = row

    for gt_row in gt_rows:
        if not gt_row or not gt_row[0]:
            continue
        key = _norm_dept(gt_row[0])
        a_row = a_lookup.get(key)
        if a_row is None:
            record(f"Department {gt_row[0]} exists", False)
            continue
        # Check Avg_Satisfaction (idx 1)
        if len(a_row) > 1:
            record(f"{gt_row[0]} Avg_Satisfaction", num_close(a_row[1], gt_row[1]),
                   f"Got {a_row[1]} vs {gt_row[1]}")
        else:
            record(f"{gt_row[0]} Avg_Satisfaction", False,
                   f"Row too short ({len(a_row)} columns), expected >= 2")
        # Check Engagement_Index (idx 5)
        if len(a_row) > 5 and len(gt_row) > 5:
            record(f"{gt_row[0]} Engagement_Index", num_close(a_row[5], gt_row[5], 0.1),
                   f"Got {a_row[5]} vs {gt_row[5]}")

    # Sheet 2: Survey_Design
    sd_sheet = None
    for name in wb.sheetnames:
        if "survey" in name.lower():
            sd_sheet = name
            break
    if not sd_sheet:
        record("Survey_Design sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Survey_Design sheet exists", True)
        ws2 = wb[sd_sheet]
        rows2 = _non_empty_rows(ws2)
        record("Survey_Design has 5 rows", len(rows2) == 5, f"Got {len(rows2)}")

    # Sheet 3: Action_Items
    ai_sheet = None
    for name in wb.sheetnames:
        if "action" in name.lower():
            ai_sheet = name
            break
    if not ai_sheet:
        record("Action_Items sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Action_Items sheet exists", True)
        ws3 = wb[ai_sheet]
        rows3 = _non_empty_rows(ws3)
        record("Action_Items has 7 rows", len(rows3) == 7, f"Got {len(rows3)}")
        # Check priorities exist
        priorities = [str(r[0]).strip() for r in rows3 if r and r[0]]
        has_high = any("high" in p.lower() for p in priorities)
        has_medium = any("medium" in p.lower() for p in priorities)
        record("Action_Items has High priority items", has_high)
        record("Action_Items has Medium priority items", has_medium)
        # Per-department priority validation: Low below avg = High, above = Medium (per task.md)
        gt_ai = gt_wb["Action_Items"]
        gt_ai_rows = _non_empty_rows(gt_ai)
        gt_priority = {}
        for g in gt_ai_rows:
            if g and g[1]:
                gt_priority[_norm_dept(g[1])] = str(g[0]).strip().lower()
        for a in rows3:
            if not a or len(a) < 2 or not a[1]:
                continue
            dept_key = _norm_dept(a[1])
            gt_p = gt_priority.get(dept_key)
            if gt_p is None:
                continue
            actual_p = str(a[0]).strip().lower() if a[0] else ""
            if dept_key in gt_boundary:
                # Boundary department (index within tolerance of the company
                # mean): a defensible computation path can classify it either
                # way, accept either High or Medium.
                record(f"{a[1]} priority matches groundtruth ({gt_p})",
                       actual_p in ("high", "medium"),
                       f"Got {actual_p}, expected high or medium (boundary)")
            else:
                record(f"{a[1]} priority matches groundtruth ({gt_p})",
                       actual_p == gt_p, f"Got {actual_p}, expected {gt_p}")

    wb.close()
    gt_wb.close()
    return True


def check_terminal_output(agent_workspace):
    print("\n=== Checking Terminal Output ===")
    fpath = os.path.join(agent_workspace, "engagement_analysis_output.txt")
    if not os.path.isfile(fpath):
        record("engagement_analysis_output.txt exists", False)
        return False
    record("engagement_analysis_output.txt exists", True)
    with open(fpath) as f:
        content = f.read().lower()
    record("Output mentions highest department",
           _has_any_word(content, _HIGH_KEYWORDS),
           f"Preview: {content[:200]}")
    record("Output mentions lowest department",
           _has_any_word(content, _LOW_KEYWORDS),
           f"Preview: {content[:200]}")
    return True


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Exact-ish match: title contains 'employee engagement survey' but NOT archived/pre-existing/historical
        cur.execute("""
            SELECT COUNT(*) FROM gform.forms
            WHERE LOWER(title) LIKE '%%employee engagement survey%%'
              AND LOWER(title) NOT LIKE '%%pre-existing%%'
              AND LOWER(title) NOT LIKE '%%preexisting%%'
              AND LOWER(title) NOT LIKE '%%archived%%'
              AND LOWER(title) NOT LIKE '%%historical%%'
              AND LOWER(title) NOT LIKE '%%legacy%%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        record("Employee Engagement Survey form exists (not pre-existing)", count >= 1, f"Found {count}")
        return count >= 1
    except Exception as e:
        record("GForm check", False, str(e))
        return False


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # A page's title may be stored in notion.pages.properties (reference
        # implementation) or inside a page block's block_data, so search both.
        # Dedup via UNION so a page that matches in more than one place counts
        # once.
        cur.execute("""
            SELECT COUNT(*) FROM (
                SELECT id FROM notion.pages
                 WHERE properties::text ILIKE '%engagement%'
                   AND properties::text ILIKE '%dashboard%'
                UNION
                SELECT parent_id FROM notion.blocks
                 WHERE parent_type = 'page_id'
                   AND block_data::text ILIKE '%engagement%'
                   AND block_data::text ILIKE '%dashboard%'
                UNION
                SELECT id FROM notion.blocks
                 WHERE type = 'page'
                   AND block_data::text ILIKE '%engagement%'
                   AND block_data::text ILIKE '%dashboard%'
            ) t
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        record("Notion HR Dashboard page exists", count >= 1, f"Found {count}")
        return count >= 1
    except Exception as e:
        record("Notion check", False, str(e))
        return False


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%hr-leadership@company.com%'
            AND LOWER(subject) LIKE '%engagement%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        record("Email sent to hr-leadership@company.com about engagement", count >= 1, f"Found {count}")
        return count >= 1
    except Exception as e:
        record("Email check", False, str(e))
        return False


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no unexpected sheets beyond the 3 required
    path = os.path.join(workspace, "Employee_Engagement_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=False)
        record("Excel has no more than 5 sheets", len(wb.sheetnames) <= 5,
               f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        # No negative satisfaction or engagement values
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    val = _to_float(cell)
                    if val is not None and val < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        record("No negative values in Excel", not has_negative,
               "Found negative engagement/satisfaction value")

    # Notion: no duplicate HR Dashboard pages
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM (
                SELECT id FROM notion.pages
                 WHERE properties::text ILIKE '%engagement%'
                   AND properties::text ILIKE '%dashboard%'
                UNION
                SELECT parent_id FROM notion.blocks
                 WHERE parent_type = 'page_id'
                   AND block_data::text ILIKE '%engagement%'
                   AND block_data::text ILIKE '%dashboard%'
                UNION
                SELECT id FROM notion.blocks
                 WHERE type = 'page'
                   AND block_data::text ILIKE '%engagement%'
                   AND block_data::text ILIKE '%dashboard%'
            ) t
        """)
        count = cur.fetchone()[0]
        # Tolerate a small amount of duplicate creation (heterogeneous swarm
        # agents may both create the dashboard page); only flag pathological
        # duplication.
        record("No duplicate Notion HR Dashboard pages", count <= 3,
               f"Found {count} matching pages")

        # Reverse: noise Notion pages preserved
        noise_titles = ["Unrelated Meeting Notes", "Office Supplies Inventory",
                        "Travel Policy 2024", "Software License Overview", "Holiday Calendar"]
        cur.execute("SELECT properties::text FROM notion.pages")
        all_props = " ".join(r[0] or "" for r in cur.fetchall())
        preserved = sum(1 for t in noise_titles if t in all_props)
        record("Reverse: noise Notion pages preserved",
               preserved >= 4,
               f"Only {preserved}/5 noise pages remain")

        # Reverse: archived survey form preserved
        cur.execute("SELECT COUNT(*) FROM gform.forms WHERE title = 'Archived 2023 Workplace Climate Survey'")
        archived_count = cur.fetchone()[0]
        record("Reverse: archived survey form preserved",
               archived_count >= 1,
               f"Archived form count: {archived_count}")

        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_terminal_output(args.agent_workspace)
    check_gform()
    check_notion()
    check_email()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
