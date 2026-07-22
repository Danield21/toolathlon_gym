"""Evaluation for yf-financial-statements-overview."""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def fetch_expected():
    """DB-driven expectations."""
    out = {}
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT symbol, stmt_type, COUNT(*)
            FROM yf.financial_statements
            GROUP BY 1, 2 ORDER BY 1, 2
        """)
        rows = cur.fetchall()
        per_pair = {f"{sym} - {stype}": int(c) for sym, stype, c in rows}
        out["per_pair"] = per_pair
        cur.execute("SELECT COUNT(*) FROM yf.financial_statements")
        out["total_records"] = int(cur.fetchone()[0])
        cur.execute("SELECT COUNT(DISTINCT symbol) FROM yf.financial_statements")
        out["unique_symbols"] = int(cur.fetchone()[0])
        cur.execute("SELECT COUNT(DISTINCT stmt_type) FROM yf.financial_statements")
        out["statement_types"] = int(cur.fetchone()[0])
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[fallback] {e}")
    return out


def normalize_key(s):
    """Normalize Symbol_Type for robust comparison.

    Accept forms like 'AMZN - income_stmt', 'AMZN-income_stmt', 'AMZN_income_stmt',
    'amzn - income_stmt', etc.
    """
    return re.sub(r"[\s\-_]+", "", str(s or "").strip().lower())


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    agent_file = os.path.join(args.agent_workspace, "YF_Financial_Statements.xlsx")
    record("YF_Financial_Statements.xlsx exists", os.path.exists(agent_file), agent_file)
    if not os.path.exists(agent_file):
        sys.exit(1)

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        sys.exit(1)

    expected = fetch_expected()

    # === Financial Overview ===
    a_rows = load_sheet_rows(agent_wb, "Financial Overview")
    record("Sheet 'Financial Overview' exists", a_rows is not None,
           f"sheets: {agent_wb.sheetnames}")
    if a_rows is not None:
        header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in a_rows[0]]
        # Required columns: Symbol_Type, Periods (others allowed)
        record("FO has 'symbol_type' column",
               "symbol_type" in header, f"headers: {header}")
        record("FO has 'periods' column",
               "periods" in header, f"headers: {header}")

        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        per_pair = expected.get("per_pair", {})
        record(f"FO has exactly {len(per_pair)} rows",
               len(a_data) == len(per_pair),
               f"got {len(a_data)} expected {len(per_pair)}")

        # Build lookup using normalized key
        a_lookup = {}
        for r in a_data:
            a_lookup[normalize_key(r[0])] = r

        for key, exp_count in per_pair.items():
            r = a_lookup.get(normalize_key(key))
            record(f"FO row '{key}' present", r is not None,
                   f"keys: {list(a_lookup.keys())[:10]}")
            if r is None:
                continue
            # Find Periods value
            periods_idx = header.index("periods") if "periods" in header else 1
            if periods_idx < len(r):
                record(f"FO '{key}' Periods == {exp_count}",
                       num_close(r[periods_idx], exp_count, 0),
                       f"got {r[periods_idx]}, expected {exp_count}")

        # Sort alphabetically by Symbol_Type
        labels = [str(r[0]).strip() for r in a_data]
        record("FO sorted alphabetically by Symbol_Type",
               labels == sorted(labels), f"got: {labels}")

    # === Summary ===
    a_rows = load_sheet_rows(agent_wb, "Summary")
    record("Sheet 'Summary' exists", a_rows is not None, f"sheets: {agent_wb.sheetnames}")
    if a_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for metric, exp_val in [("total_records", expected.get("total_records")),
                                 ("unique_symbols", expected.get("unique_symbols")),
                                 ("statement_types", expected.get("statement_types"))]:
            row = a_lookup.get(metric)
            record(f"Summary metric '{metric}' present", row is not None,
                   f"keys: {list(a_lookup.keys())}")
            if row is None or exp_val is None:
                continue
            if len(row) > 1:
                record(f"Summary {metric} == {exp_val}",
                       num_close(row[1], exp_val, 0),
                       f"got {row[1]}, expected {exp_val}")

    all_passed = (FAIL_COUNT == 0)
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": all_passed}, f)
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
