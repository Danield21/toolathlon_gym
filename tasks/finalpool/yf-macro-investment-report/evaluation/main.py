"""Evaluation for yf-macro-investment-report."""
import argparse, os, sys


FAIL_COUNT = 0


def record(label, ok, detail=""):
    global FAIL_COUNT
    status = "PASS" if ok else "FAIL"
    extra = f" - {detail}" if detail and not ok else ""
    print(f"    [{status}] {label}{extra}")
    if not ok:
        FAIL_COUNT += 1


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


EXPECTED_SECTORS = {
    "AMZN": "Consumer Cyclical",
    "GOOGL": "Communication Services",
    "JNJ": "Healthcare",
    "JPM": "Financial Services",
    "XOM": "Energy",
}


def _compute_expected_returns():
    """Compute expected 90-day returns at runtime from yf.stock_prices.

    Falls back to hardcoded values if DB is unreachable.
    """
    fallback = {"AMZN": -4.61, "GOOGL": -6.29, "JNJ": 19.30, "JPM": -6.40, "XOM": 30.22}
    try:
        import psycopg2
        conn = psycopg2.connect(host='localhost', port=5432, dbname='toolathlon_gym',
                                user='eigent', password='camel')
        out = {}
        with conn.cursor() as cur:
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                cur.execute(
                    """
                    WITH latest AS (
                        SELECT date, close FROM yf.stock_prices WHERE symbol=%s ORDER BY date DESC LIMIT 1
                    ),
                    target_date AS (
                        SELECT (SELECT date FROM latest) - INTERVAL '90 days' as d
                    ),
                    old_price AS (
                        SELECT close FROM yf.stock_prices
                        WHERE symbol=%s AND date <= (SELECT d FROM target_date)
                        ORDER BY date DESC LIMIT 1
                    )
                    SELECT ((SELECT close FROM latest) - (SELECT close FROM old_price))
                           / (SELECT close FROM old_price) * 100
                    """,
                    (sym, sym),
                )
                row = cur.fetchone()
                if row and row[0] is not None:
                    out[sym] = float(row[0])
                else:
                    out[sym] = fallback[sym]
        conn.close()
        return out
    except Exception:
        return fallback


EXPECTED_RETURNS = _compute_expected_returns()
EXPECTED_MACRO = {
    "us_10y_yield": (4.25, "Headwind for growth stocks"),
    "fed_funds_rate": (5.25, "Restrictive"),
    "cpi_yoy": (3.1, "Inflationary pressure"),
    "unemployment": (3.7, "Tight labor market"),
    "gdp_growth_q4": (2.8, "Supportive"),
    "sp500_pe_ratio": (22.5, "Above historical average"),
    "vix": (18.3, "Low volatility"),
}


def check_excel(agent_workspace):
    import openpyxl
    path = os.path.join(agent_workspace, "Macro_Investment.xlsx")
    if not os.path.exists(path):
        record("Macro_Investment.xlsx exists", False, "file not found")
        # Increment FAIL_COUNT for each missed check
        for _ in range(20):
            record("Excel content (skipped due to missing file)", False)
        return
    record("Macro_Investment.xlsx exists", True)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # ===== Stock Performance sheet =====
        rows = load_sheet_rows(wb, "Stock Performance")
        record("Sheet 'Stock Performance' exists", rows is not None)
        if rows is None:
            for _ in range(10):
                record("Stock row check (skipped)", False)
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            record(f"Stock Performance has 5 rows", len(data_rows) == 5,
                   f"got {len(data_rows)}")
            by_sym = {}
            for r in data_rows:
                if r and r[0]:
                    by_sym[str(r[0]).strip().upper()] = r
            for sym in ["AMZN", "GOOGL", "JNJ", "JPM", "XOM"]:
                row = by_sym.get(sym)
                record(f"{sym} present in Stock Performance", row is not None)
                if row is not None:
                    # Sector check
                    if len(row) > 2 and row[2]:
                        record(f"{sym} sector",
                               str(row[2]).strip().lower() == EXPECTED_SECTORS[sym].lower(),
                               f"got '{row[2]}', expected '{EXPECTED_SECTORS[sym]}'")
                    else:
                        record(f"{sym} sector", False, "missing")
                    # Return value check (tighter abs_tol=0.1)
                    if len(row) > 5 and row[5] is not None:
                        record(f"{sym} Return_90d_Pct",
                               num_close(row[5], EXPECTED_RETURNS[sym], abs_tol=0.1, rel_tol=0.0),
                               f"got {row[5]}, expected ~{EXPECTED_RETURNS[sym]}")
                    else:
                        record(f"{sym} Return_90d_Pct", False, "missing")
            # Verify alphabetical sort by Symbol
            symbols_in_order = [str(r[0]).strip().upper() for r in data_rows if r and r[0]]
            record("Stock Performance sorted alphabetically",
                   symbols_in_order == sorted(symbols_in_order),
                   f"got {symbols_in_order}")

        # ===== Macro Context sheet =====
        rows2 = load_sheet_rows(wb, "Macro Context")
        record("Sheet 'Macro Context' exists", rows2 is not None)
        if rows2 is None:
            for _ in range(15):
                record("Macro row check (skipped)", False)
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            record("Macro Context has 7 rows", len(data_rows2) == 7,
                   f"got {len(data_rows2)}")
            lookup = {str(r[0]).strip().lower(): r for r in data_rows2 if r and r[0]}
            for ind_key, (exp_val, exp_label) in EXPECTED_MACRO.items():
                r = lookup.get(ind_key)
                record(f"Indicator '{ind_key}' present", r is not None)
                if r is not None:
                    val_ok = len(r) > 1 and r[1] is not None and num_close(
                        r[1], exp_val, abs_tol=0.1, rel_tol=0.0)
                    record(f"'{ind_key}' value", val_ok,
                           f"got {r[1] if len(r)>1 else None}, expected {exp_val}")
                    label_ok = len(r) > 2 and r[2] is not None and (
                        str(r[2]).strip().lower() == exp_label.lower())
                    record(f"'{ind_key}' Impact_Assessment", label_ok,
                           f"got '{r[2] if len(r)>2 else None}', expected '{exp_label}'")

        # ===== Portfolio Summary sheet =====
        rows3 = load_sheet_rows(wb, "Portfolio Summary")
        record("Sheet 'Portfolio Summary' exists", rows3 is not None)
        if rows3 is None:
            for _ in range(5):
                record("Portfolio summary row (skipped)", False)
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            lookup3 = {str(r[0]).strip().lower(): r[1] for r in data_rows3 if r and r[0]}
            # Average_90d_Return = avg of returns ~ 6.44
            v = lookup3.get("average_90d_return")
            record("Average_90d_Return ~ 6.44",
                   v is not None and num_close(v, 6.44, abs_tol=0.05, rel_tol=0.0),
                   f"got {v}")
            v = lookup3.get("best_performer")
            record("Best_Performer == XOM",
                   v is not None and str(v).strip().upper() == "XOM",
                   f"got {v}")
            v = lookup3.get("worst_performer")
            record("Worst_Performer == JPM",
                   v is not None and str(v).strip().upper() == "JPM",
                   f"got {v}")
            v = lookup3.get("avg_trailing_pe")
            record("Avg_Trailing_PE ~ 23.43",
                   v is not None and num_close(v, 23.43, abs_tol=0.1, rel_tol=0.0),
                   f"got {v}")
            v = lookup3.get("macro_risk_score")
            # Per task: 7 macro indicators with negative-sounding labels:
            # Headwind, Restrictive, Inflationary pressure, Above historical average
            # = 4 (us_10y_yield, fed_funds_rate, cpi_yoy, sp500_pe_ratio)
            # However task originally accepted 3 (excludes Above historical avg).
            # Accept 3 or 4 to be lenient given ambiguity but reject 0/1/5/6/7.
            try:
                vf = float(v) if v is not None else None
            except Exception:
                vf = None
            record("Macro_Risk_Score in {3,4}",
                   vf is not None and 3 <= vf <= 4,
                   f"got {v}")

    except Exception as e:
        record(f"Error reading Excel: {e}", False)


def check_word(agent_workspace):
    path = os.path.join(agent_workspace, "Investment_Report.docx")
    if not os.path.exists(path):
        record("Investment_Report.docx exists", False)
        for _ in range(8):
            record("Word section (skipped due to missing file)", False)
        return
    record("Investment_Report.docx exists", True)
    try:
        from docx import Document
        doc = Document(path)
        paras = [p.text for p in doc.paragraphs]
        full_text = "\n".join(paras)
        text_lower = full_text.lower()
        record("Investment_Report.docx >= 600 chars", len(full_text) >= 600,
               f"got {len(full_text)} chars")
        # Title (must appear in first 5 paragraphs at the top)
        first5_lower = "\n".join(paras[:5]).lower()
        record("Title 'Q1 2026 Macro Investment Report' at top (first 5 paras)",
               "q1 2026 macro investment report" in first5_lower)
        # Sections
        for section_kw in ["macroeconomic", "portfolio", "sector", "outlook"]:
            record(f"Word section keyword '{section_kw}'",
                   section_kw in text_lower)
        # Sector commentary should mention all 5 sectors and symbols
        sector_hits = sum(1 for sym in ["googl", "amzn", "jpm", "jnj", "xom"]
                          if sym in text_lower)
        record("Word references all 5 of 5 portfolio symbols",
               sector_hits == 5, f"got {sector_hits}")
        sector_word_hits = sum(1 for s in [
            "communication services", "consumer cyclical",
            "financial services", "healthcare", "energy"]
                               if s in text_lower)
        record("Word references all 5 of 5 sector names",
               sector_word_hits == 5, f"got {sector_word_hits}")
    except Exception as e:
        record(f"Error reading Word doc: {e}", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace")

    print("  Checking Excel file...")
    check_excel(agent_ws)

    print("  Checking Word document...")
    check_word(agent_ws)

    if FAIL_COUNT == 0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL ({FAIL_COUNT} errors) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
