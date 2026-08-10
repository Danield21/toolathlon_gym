"""
Evaluation script for canvas-quiz-analysis-email task.

Checks:
1. Excel Quiz_Performance.xlsx with "Quiz Analysis" sheet documenting each quiz's
   title, points possible, and question count.
2. Email sent to the course instructor.

The quiz metrics (Points_Possible, Question_Count) come straight from
canvas.quizzes — no per-submission data is required.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    # Quiz metrics come from canvas.quizzes directly (title, points_possible,
    # question_count). Ordered by question_count desc, title asc to match the
    # required worksheet sort order.
    cur.execute("""
        SELECT q.title, q.points_possible, q.question_count
        FROM canvas.quizzes q
        WHERE q.course_id = 3
        ORDER BY q.question_count DESC, q.title ASC
    """)
    quiz_rows = cur.fetchall()

    # Get instructor email (earliest-enrolled teacher for the course).
    cur.execute("""
        SELECT u.email FROM canvas.enrollments e
        JOIN canvas.users u ON e.user_id = u.id
        WHERE e.course_id = 3 AND e.type = 'TeacherEnrollment'
        ORDER BY e.id LIMIT 1
    """)
    row = cur.fetchone()
    instructor_email = row[0] if row else None

    conn.close()
    return {"quizzes": quiz_rows, "instructor_email": instructor_email}


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Quiz_Performance.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Sheet 'Quiz Analysis' exists", get_sheet(wb, "Quiz Analysis") is not None,
          f"Found: {wb.sheetnames}")

    ws = get_sheet(wb, "Quiz Analysis")
    if ws and expected:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["quizzes"]
        check("Quiz Analysis row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        # Build lookup by title (Quiz_Title is column 0)
        agent_by_title = {}
        for row in agent_rows:
            if row and row[0]:
                agent_by_title[str(row[0]).strip().lower()] = row

        for exp_row in exp:
            title = exp_row[0]
            agent_row = agent_by_title.get(title.strip().lower())
            if agent_row:
                # exp_row is (title, points_possible, question_count)
                check(f"Quiz '{title}' Points_Possible",
                      num_close(agent_row[1], float(exp_row[1]), 0.01),
                      f"Expected {exp_row[1]}, got {agent_row[1]}")
                check(f"Quiz '{title}' Question_Count",
                      num_close(agent_row[2], float(exp_row[2]), 0.0),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
            else:
                check(f"Quiz '{title}' found in output", False, "Not in agent output")

        # Check sort order (by Question_Count descending).
        if len(agent_rows) >= 2:
            counts = []
            for r in agent_rows:
                if r and r[2] is not None:
                    try:
                        counts.append(float(r[2]))
                    except (TypeError, ValueError):
                        pass
            check("Sorted by Question_Count descending",
                  all(counts[i] >= counts[i + 1] for i in range(len(counts) - 1)),
                  f"Question counts: {counts}")


def check_email(expected):
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for email check", False, str(e))
        return

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()

    conn.close()

    all_items = list(messages)
    check("At least one sent email message exists", len(all_items) > 0,
          f"Found {len(messages)} messages")

    found_email = False
    for item in all_items:
        subj = str(item[0] or "").lower()
        if "quiz" in subj and "performance" in subj:
            found_email = True
            check("Email subject contains 'Quiz Performance'", True)

            from_addr = item[1]
            from_str = json.dumps(from_addr).lower() if isinstance(from_addr, (list, dict)) else str(from_addr or "").lower()
            check("Email sent from coordinator@university.edu",
                  "coordinator@university.edu" in from_str,
                  f"From: {from_str}")

            to_addr = item[2]
            if expected and expected.get("instructor_email"):
                exp_email = expected["instructor_email"].lower()
                to_str = json.dumps(to_addr).lower() if isinstance(to_addr, (list, dict)) else str(to_addr or "").lower()
                check("Email sent to instructor",
                      exp_email in to_str,
                      f"Expected to contain '{exp_email}', got '{to_str}'")

            body = str(item[3] or "")
            check("Email body is not empty", len(body) > 20,
                  f"Body length: {len(body)}")
            # The quiz with the most questions should be mentioned in the body.
            # expected["quizzes"] is already ordered by question_count DESC, so the
            # first entry is the quiz with the most questions.
            if expected and expected.get("quizzes"):
                most_q_title = str(expected["quizzes"][0][0] or "").strip()
                if most_q_title:
                    check("Email body mentions the quiz with the most questions",
                          most_q_title.lower() in body.lower(),
                          f"Expected '{most_q_title}' in body")
            break

    if not found_email:
        check("Quiz Performance email found", False,
              f"Subjects: {[str(i[0]) for i in all_items]}")

    # Reverse validation: noise subject/body should not be present in the
    # Quiz Performance email that the agent sends.
    noise_phrases = [
        "weekly all-hands agenda",
        "it security reminder",
        "cafeteria menu update",
    ]
    for item in all_items:
        subj = str(item[0] or "").lower()
        body = str(item[3] or "").lower()
        if "quiz" in subj and "performance" in subj:
            for phrase in noise_phrases:
                check(
                    f"Quiz Performance email does not reuse noise phrase '{phrase}'",
                    phrase not in subj and phrase not in body,
                    f"Noise phrase leaked into output email",
                )


def check_excel_gt(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel (vs groundtruth) ===")
    agent_file = os.path.join(agent_workspace, "Quiz_Performance.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Quiz_Performance.xlsx")
    check("Excel file exists", os.path.isfile(agent_file))
    check("Groundtruth file exists", os.path.isfile(gt_file))
    if not os.path.isfile(agent_file) or not os.path.isfile(gt_file):
        return
    agent_wb = openpyxl.load_workbook(agent_file)
    gt_wb = openpyxl.load_workbook(gt_file)
    check("Sheet 'Quiz Analysis' exists", get_sheet(agent_wb, "Quiz Analysis") is not None)
    a_ws = get_sheet(agent_wb, "Quiz Analysis")
    g_ws = get_sheet(gt_wb, "Quiz Analysis")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Row count matches", len(a_rows) == len(g_rows),
              f"Expected {len(g_rows)}, got {len(a_rows)}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()
    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        check_excel(agent_workspace, expected)
    else:
        print("INFO: Falling back to groundtruth Excel")
        check_excel_gt(agent_workspace, groundtruth_workspace)

    check_email(expected)

    total_pass = PASS_COUNT
    total_fail = FAIL_COUNT
    all_ok = FAIL_COUNT == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": all_ok}
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return all_ok, f"Passed: {total_pass}, Failed: {total_fail}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
