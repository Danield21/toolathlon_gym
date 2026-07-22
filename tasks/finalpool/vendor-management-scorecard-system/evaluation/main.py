"""Evaluation for vendor-management-scorecard-system.

Verifies the agent produced concrete vendor scorecard deliverables across:
  - Excel scorecard / spreadsheet for ongoing tracking
  - Word vendor scorecard report with rankings
  - Google Sheet for ongoing performance tracking
  - Email feedback to underperforming vendors
  - Calendar vendor review meetings
"""
import argparse
import json
import os
import sys
from pathlib import Path

import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


# Required topical keywords
SCORECARD_KEYWORDS = ["delivery", "quality", "cost", "service"]
TIER_KEYWORDS = ["strategic", "preferred", "acceptable"]


def check_excel_scorecard(workspace):
    print("\n=== Check: Excel scorecard / tracking workbook ===")
    ws = Path(workspace)
    xlsx_files = list(ws.glob("*.xlsx"))
    check("At least one XLSX file produced", len(xlsx_files) >= 1, f"found {[p.name for p in xlsx_files]}")
    if not xlsx_files:
        return
    import openpyxl
    found_scorecard = False
    found_vendor_rows = 0
    all_text_combined = ""
    for p in xlsx_files:
        try:
            wb = openpyxl.load_workbook(str(p), data_only=True)
        except Exception as e:
            check(f"XLSX {p.name} readable", False, str(e))
            continue
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        all_text_combined += str(cell) + " "
            non_empty = [r for r in sheet.iter_rows(values_only=True) if any(c is not None for c in r)]
            if len(non_empty) >= 3:
                found_vendor_rows += len(non_empty) - 1
        wb.close()
    check("XLSX has at least 3 vendor data rows total", found_vendor_rows >= 3, f"got {found_vendor_rows}")
    text_lower = all_text_combined.lower()
    for kw in SCORECARD_KEYWORDS:
        check(f"XLSX content references '{kw}' dimension", kw in text_lower, f"missing '{kw}'")
    check("XLSX content references vendor scoring/ranking",
          ("score" in text_lower or "rank" in text_lower or "tier" in text_lower),
          "no scoring/ranking keywords")

    # Value-level: at least one numeric score-style cell in plausible 0-100 range
    score_in_range = False
    for p in xlsx_files:
        try:
            wb = openpyxl.load_workbook(str(p), data_only=True)
        except Exception:
            continue
        for sheet in wb.worksheets:
            headers = [str(c.value).lower() if c.value else '' for c in sheet[1]]
            score_cols = [i for i, h in enumerate(headers)
                          if any(k in h for k in ('score', 'delivery', 'quality', 'cost', 'service', 'composite'))]
            if not score_cols:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                for ci in score_cols:
                    if ci < len(row) and row[ci] is not None:
                        try:
                            fv = float(row[ci])
                            if 0 <= fv <= 100:
                                score_in_range = True
                                break
                        except (ValueError, TypeError):
                            pass
                if score_in_range:
                    break
            if score_in_range:
                break
        wb.close()
        if score_in_range:
            break
    check("XLSX has at least one numeric score in 0-100 range",
          score_in_range, "no plausible numeric score found")


def check_docx_report(workspace):
    print("\n=== Check: Word vendor scorecard report ===")
    ws = Path(workspace)
    docx_files = list(ws.glob("*.docx")) + list(ws.glob("**/*.docx"))
    check("At least one DOCX file produced", len(docx_files) >= 1, f"found {[p.name for p in docx_files][:5]}")
    if not docx_files:
        return
    largest = max(docx_files, key=lambda p: p.stat().st_size)
    try:
        from docx import Document
        doc = Document(str(largest))
    except Exception as e:
        check(f"DOCX {largest.name} readable", False, str(e))
        return
    text = "\n".join(p.text for p in doc.paragraphs)
    text_lower = text.lower()
    check(f"DOCX '{largest.name}' has >= 800 chars of content",
          len(text) >= 800, f"got {len(text)} chars")
    for kw in SCORECARD_KEYWORDS:
        check(f"DOCX mentions '{kw}' dimension", kw in text_lower, "missing")
    check("DOCX mentions vendor scoring or ranking",
          ("score" in text_lower or "rank" in text_lower),
          "missing")
    check("DOCX mentions tiering (strategic/preferred/acceptable)",
          any(t in text_lower for t in TIER_KEYWORDS),
          "missing tier keywords")
    check("DOCX includes recommendation section",
          "recommend" in text_lower, "missing recommendation")


def check_gsheet(is_gt_self_test=False):
    print("\n=== Check: Google Sheet vendor tracking ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    sheets = cur.fetchall()
    matched = 0
    for sid, t in sheets:
        if t and any(kw in t.lower() for kw in ("vendor", "scorecard", "performance", "supplier")):
            matched += 1
    if is_gt_self_test and matched == 0:
        print("  [WARN] GSheet vendor tracking checks skipped (GT self-test, non-blocking)")
        cur.close()
        conn.close()
        return
    check("At least 1 Google Sheet created", len(sheets) >= 1, f"found {len(sheets)}")
    if not sheets:
        cur.close()
        conn.close()
        return
    check("At least 1 GSheet titled vendor/scorecard/performance/supplier",
          matched >= 1, f"titles: {[t for _, t in sheets]}")
    # At least 10 cells of data in any sheet
    cur.execute("SELECT COUNT(*) FROM gsheet.cells")
    cell_count = cur.fetchone()[0]
    check("GSheet has at least 10 cells of data", cell_count >= 10, f"got {cell_count}")
    cur.close()
    conn.close()


def check_email(is_gt_self_test=False):
    print("\n=== Check: Vendor feedback emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("SELECT id FROM email.folders WHERE LOWER(name) IN ('sent','drafts')")
    folder_ids = [r[0] for r in cur.fetchall()]
    if not folder_ids:
        if is_gt_self_test:
            print("  [WARN] Email folders not configured (GT self-test, non-blocking)")
            cur.close()
            conn.close()
            return
        check("Email sent/drafts folders exist", False, "")
        cur.close()
        conn.close()
        return
    placeholders = ",".join(["%s"] * len(folder_ids))
    cur.execute(f"SELECT subject, body_text FROM email.messages WHERE folder_id IN ({placeholders})", folder_ids)
    sent = cur.fetchall()
    if is_gt_self_test and len(sent) == 0:
        print("  [WARN] Vendor email checks skipped (GT self-test, non-blocking)")
        cur.close()
        conn.close()
        return
    check("At least 1 email composed/sent by agent", len(sent) >= 1, f"found {len(sent)}")
    matched = 0
    for subj, body in sent:
        text = f"{subj or ''} {body or ''}".lower()
        if any(kw in text for kw in ("vendor", "supplier", "scorecard", "performance")):
            matched += 1
    check("At least 1 sent email is vendor/scorecard topical",
          matched >= 1, f"matched {matched}/{len(sent)}")
    cur.close()
    conn.close()


def check_gcal(is_gt_self_test=False):
    print("\n=== Check: Vendor review meetings on calendar ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("SELECT id, summary, description FROM gcal.events")
    events = cur.fetchall()
    if is_gt_self_test and len(events) == 0:
        print("  [WARN] Vendor calendar checks skipped (GT self-test, non-blocking)")
        cur.close()
        conn.close()
        return
    check("At least 1 calendar event created", len(events) >= 1, f"found {len(events)}")
    matched = 0
    for eid, summary, desc in events:
        text = f"{summary or ''} {desc or ''}".lower()
        if any(kw in text for kw in ("vendor", "supplier", "review", "scorecard")):
            matched += 1
    check("At least 1 event mentions vendor/supplier/review/scorecard",
          matched >= 1, f"matched {matched}/{len(events)}")
    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # Detect GT self-test (V1 parity test).
    try:
        gt_canon = os.path.realpath(args.groundtruth_workspace) if args.groundtruth_workspace else ""
        ag_canon = os.path.realpath(args.agent_workspace) if args.agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    check_excel_scorecard(args.agent_workspace)
    check_docx_report(args.agent_workspace)
    check_gsheet(is_gt_self_test)
    check_email(is_gt_self_test)
    check_gcal(is_gt_self_test)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_ok = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": all_ok}, f)
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
