"""Evaluation for sf-hr-manager-report."""
import argparse
import os
import re
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _normalize_title(s):
    """Lowercase, replace punctuation with spaces, collapse whitespace."""
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(s).lower()).split())


def _levenshtein(a, b):
    """Small-scale DP edit distance between two strings."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def titles_match(expected, actual):
    """Exact Notion title match (task.md prescribes the exact title). After
    normalization the titles must be identical — no fuzzy/edit-distance
    tolerance, so the model must follow the title given in task.md."""
    e = _normalize_title(expected)
    a = _normalize_title(actual)
    if not e or not a:
        return False
    return e == a


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_header(h):
    return " ".join(str(h or "").strip().lower().replace("-", " ").replace("_", " ").split())


def _header_map(rows):
    """Map normalized header text -> 0-based column index from row 1 of a sheet.
    Lets checks locate columns by name regardless of column order."""
    if not rows or not rows[0]:
        return {}
    return {_norm_header(h): i for i, h in enumerate(rows[0]) if _norm_header(h)}


def _cell(row, idx):
    """Safe cell access (None when idx out of range)."""
    if row is None or idx is None:
        return None
    try:
        return row[idx] if idx < len(row) else None
    except (TypeError, IndexError):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Manager_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Manager_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Manager Report
    print(f"  Checking Manager Report...")
    a_rows = load_sheet_rows(agent_wb, "Manager Report")
    g_rows = load_sheet_rows(gt_wb, "Manager Report")
    if a_rows is None:
        all_errors.append("Sheet 'Manager Report' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Manager Report' not found in groundtruth")
    else:
        sheet_name = "Manager Report"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        # Sort order check: by Department alphabetically
        a_dept_order = [str(r[0]).strip().lower() for r in a_data if r and r[0] is not None]
        if a_dept_order != sorted(a_dept_order):
            errors.append(f"Manager Report not sorted alphabetically: {a_dept_order}")
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Total - exact (deterministic)
            av = _cell(a_row, a_col.get("total"))
            gv = _cell(g_row, g_col.get("total"))
            if av is not None and gv is not None:
                if not num_close(av, gv, 1):
                    errors.append(f"{key}.Total: {av} vs {gv} (tol=1)")

            # High_Performers - tight tolerance for deterministic count
            av = _cell(a_row, a_col.get("high performers"))
            gv = _cell(g_row, g_col.get("high performers"))
            if av is not None and gv is not None:
                if not num_close(av, gv, 1):
                    errors.append(f"{key}.High_Performers: {av} vs {gv} (tol=1)")

            # Low_Performers - tight tolerance
            av = _cell(a_row, a_col.get("low performers"))
            gv = _cell(g_row, g_col.get("low performers"))
            if av is not None and gv is not None:
                if not num_close(av, gv, 1):
                    errors.append(f"{key}.Low_Performers: {av} vs {gv} (tol=1)")

            # Avg_Salary rounded to integer - tol=1
            av = _cell(a_row, a_col.get("avg salary"))
            gv = _cell(g_row, g_col.get("avg salary"))
            if av is not None and gv is not None:
                if not num_close(av, gv, 1.0):
                    errors.append(f"{key}.Avg_Salary: {av} vs {gv} (tol=1.0)")

            av = _cell(a_row, a_col.get("avg experience"))
            gv = _cell(g_row, g_col.get("avg experience"))
            if av is not None and gv is not None:
                if not num_close(av, gv, 0.2):
                    errors.append(f"{key}.Avg_Experience: {av} vs {gv} (tol=0.2)")

            av = _cell(a_row, a_col.get("high perf pct"))
            gv = _cell(g_row, g_col.get("high perf pct"))
            if av is not None and gv is not None:
                if not num_close(av, gv, 0.2):
                    errors.append(f"{key}.High_Perf_Pct: {av} vs {gv} (tol=0.2)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            av = _cell(a_row, a_col.get("value"))
            gv = _cell(g_row, g_col.get("value"))
            if av is not None and gv is not None:
                # Best_Dept is non-numeric -> exact str match; numeric metrics tighter tol
                try:
                    gv_f = float(gv) if gv is not None else None
                except (TypeError, ValueError):
                    gv_f = None
                if gv_f is not None:
                    # Choose tolerance per metric type
                    if "pct" in key or "percent" in key:
                        tol = 0.2
                    elif "total" in key or "count" in key or "performers" in key:
                        tol = 1.0
                    else:
                        tol = 1.0
                    if not num_close(av, gv, tol):
                        errors.append(f"{key}.Value: {av} vs {gv} (tol={tol})")
                else:
                    # String comparison - exact
                    if not str_match(av, gv):
                        errors.append(f"{key}.Value: '{av}' vs '{gv}' (str_match)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check Notion page (BLOCKING) - extract page title from properties JSON and
    # require a fuzzy match against 'HR Department Performance Report'
    # (containment / Levenshtein <= 3 / keyword overlap, case-insensitive).
    print("  Checking Notion page (blocking)...")
    try:
        import json as _json
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
        pages = cur.fetchall()
        target_phrase = "hr department performance report"
        found_count = 0
        for pid, props in pages:
            if not props:
                continue
            try:
                if isinstance(props, str):
                    props_obj = _json.loads(props)
                else:
                    props_obj = props
                titles = []
                for k, v in (props_obj.items() if isinstance(props_obj, dict) else []):
                    if isinstance(v, dict) and v.get("type") == "title":
                        for t in v.get("title", []):
                            if isinstance(t, dict):
                                pt = t.get("plain_text", "") or (
                                    t.get("text", {}).get("content", "")
                                    if isinstance(t.get("text"), dict) else ""
                                )
                                if pt:
                                    titles.append(pt)
                title_text = " ".join(titles).strip().lower()
                if titles_match(target_phrase, title_text):
                    found_count += 1
            except Exception:
                continue
        if found_count == 0:
            all_errors.append(
                "Notion page titled 'HR Department Performance Report' not found "
                "(fuzzy title match via title-property field)"
            )
        else:
            print(f"    PASS (found {found_count} matching Notion page titles)")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Notion check failed: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
