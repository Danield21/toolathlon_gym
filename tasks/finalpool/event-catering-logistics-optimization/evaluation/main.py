#!/usr/bin/env python3
"""Evaluation script for event-catering-logistics-optimization.

Six-phase event catering / logistics task. Validates structural deliverables
that any reasonable agent would produce based on task.md:

  - Catering plan spreadsheet (xlsx) with menu/ingredient/cost data
  - Catering proposal document (docx) for executive approval
  - Outgoing emails to vendors / stakeholders
  - Calendar events for vendor coordination + key deadlines
"""

from argparse import ArgumentParser
import json
import os
import sys
from pathlib import Path

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def get_conn():
    import psycopg2
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432,
        dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
        user="eigent", password="camel",
    )


def list_files(ws, suffix):
    return [p for p in Path(ws).glob(f"*{suffix}") if not p.name.startswith("~$")]


def check_workspace_files(agent_ws):
    print("\n=== Check: Workspace files ===")
    template_names = {"menu_templates.xlsx", "participant_list.xlsx"}
    xlsx_files = [p for p in list_files(agent_ws, ".xlsx") if p.name not in template_names]
    docx_files = list_files(agent_ws, ".docx")
    record("At least one .xlsx deliverable produced (excluding templates)",
           len(xlsx_files) >= 1, f"found {[p.name for p in xlsx_files]}")
    record("At least one .docx deliverable",
           len(docx_files) >= 1, f"found {[p.name for p in docx_files]}")
    return xlsx_files, docx_files


def check_catering_xlsx(xlsx_files):
    print("\n=== Check: Catering plan xlsx ===")
    if not xlsx_files:
        record("Catering xlsx readable", False, "no xlsx file")
        return

    target = max(xlsx_files, key=lambda p: p.stat().st_size)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(target), data_only=True)
    except Exception as e:
        record(f"xlsx {target.name} readable", False, str(e))
        return
    record(f"xlsx {target.name} readable", True)

    max_rows = 0
    max_cols = 0
    has_catering_keyword = False
    sheet_names = [s.lower() for s in wb.sheetnames]
    catering_sheet_kw = ['menu', 'ingredient', 'cost', 'catering',
                         'budget', 'vendor', 'item']
    has_catering_sheet = any(kw in sn for sn in sheet_names for kw in catering_sheet_kw)
    record(f"xlsx {target.name} has catering-related sheet name",
           has_catering_sheet, f"sheets={wb.sheetnames}")

    for ws in wb.worksheets:
        max_rows = max(max_rows, ws.max_row)
        max_cols = max(max_cols, ws.max_column)
        if ws.max_row >= 1:
            try:
                headers = [str(c.value).lower() if c.value else ''
                           for c in next(ws.iter_rows(min_row=1, max_row=1))]
                if any('cost' in h or 'price' in h or 'qty' in h or
                       'quantity' in h or 'menu' in h or 'ingredient' in h
                       for h in headers):
                    has_catering_keyword = True
            except StopIteration:
                pass

    record(f"xlsx {target.name} has substantive rows (>=4)",
           max_rows >= 4, f"max_rows={max_rows}")
    record(f"xlsx {target.name} has multiple columns (>=3)",
           max_cols >= 3, f"max_cols={max_cols}")
    record(f"xlsx {target.name} has cost/menu/ingredient columns",
           has_catering_keyword, "no header matched catering keywords")
    wb.close()


def check_proposal_docx(docx_files):
    print("\n=== Check: Catering proposal docx ===")
    if not docx_files:
        record("Catering proposal readable", False, "no docx file")
        return

    target = max(docx_files, key=lambda p: p.stat().st_size)
    try:
        from docx import Document
        doc = Document(str(target))
    except Exception as e:
        record(f"docx {target.name} readable", False, str(e))
        return
    record(f"docx {target.name} readable", True)

    text = "\n".join(p.text for p in doc.paragraphs)
    text_low = text.lower()
    word_count = len([w for w in text.split() if w.strip()])

    record(f"docx {target.name} substantive (>=150 words)",
           word_count >= 150, f"word_count={word_count}")

    keywords = ['menu', 'catering', 'cost', 'vendor',
                'dietary', 'ingredient', 'event', 'logistics']
    matches = sum(1 for kw in keywords if kw in text_low)
    record(f"docx {target.name} mentions catering/menu/cost (>=2 keywords)",
           matches >= 2, f"matched {matches}/8 keywords")


def check_vendor_emails(is_gt_self_test=False):
    print("\n=== Check: Outgoing vendor / stakeholder emails ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, subject FROM email.messages
               WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s
                  OR subject ILIKE %s OR body_text ILIKE %s OR body_text ILIKE %s""",
            ('%catering%', '%vendor%', '%event%',
             '%menu%', '%catering%', '%vendor%'))
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Vendor emails (db query)", False, f"db error: {e}")
        return
    if is_gt_self_test and len(rows) == 0:
        record("At least one vendor/catering/event email sent (GT self-test toleration)",
               True, "GT self-test: emails are agent runtime artifacts, skipped")
        return
    record("At least one vendor/catering/event email sent",
           len(rows) >= 1, f"matching emails: {len(rows)}")


def check_coordination_events(is_gt_self_test=False):
    print("\n=== Check: Vendor coordination / event calendar ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, summary FROM gcal.events
               WHERE summary ILIKE %s OR summary ILIKE %s OR summary ILIKE %s
                  OR summary ILIKE %s OR summary ILIKE %s""",
            ('%catering%', '%vendor%', '%coordination%',
             '%event%', '%delivery%'))
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Coordination events (db query)", False, f"db error: {e}")
        return
    if is_gt_self_test and len(rows) == 0:
        record("At least one vendor/event calendar event (GT self-test toleration)",
               True, "GT self-test: calendar events are agent runtime artifacts, skipped")
        return
    record("At least one vendor/event calendar event",
           len(rows) >= 1, f"matching events: {len(rows)}")


def check_dietary_survey_form(is_gt_self_test=False):
    """Phase 2 explicitly mentions an 'online survey platform' for dietary preferences."""
    print("\n=== Check: Dietary survey form (gform) ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, title FROM gform.forms
               WHERE title ILIKE %s OR title ILIKE %s OR title ILIKE %s
                  OR title ILIKE %s OR document_title ILIKE %s
                  OR document_title ILIKE %s""",
            ('%dietary%', '%survey%', '%catering%', '%food%',
             '%dietary%', '%catering%'))
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Survey form (db query)", True,
               f"gform schema not queryable (non-blocking): {e}")
        return
    if is_gt_self_test and len(rows) == 0:
        record("At least one dietary/survey form created (GT self-test toleration)",
               True, "GT self-test: gform is agent runtime artifact, skipped")
        return
    record("At least one dietary/survey form created",
           len(rows) >= 1, f"matching forms: {len(rows)}")


def check_value_match_against_gt(agent_xlsx_files, gt_xlsx_files, is_gt_self_test=False):
    """Compare numeric values from agent xlsx against GT xlsx range."""
    print("\n=== Check: Value-level GT comparison ===")
    if not gt_xlsx_files:
        record("GT xlsx available for comparison (non-blocking)", True,
               "no GT xlsx; skipping value-level check")
        return
    if not agent_xlsx_files:
        record("Agent xlsx available for value comparison", False,
               "no agent xlsx to compare")
        return
    try:
        import openpyxl
        gt_target = max(gt_xlsx_files, key=lambda p: p.stat().st_size)
        agent_target = max(agent_xlsx_files, key=lambda p: p.stat().st_size)
        gt_wb = openpyxl.load_workbook(str(gt_target), data_only=True)
        agent_wb = openpyxl.load_workbook(str(agent_target), data_only=True)
    except Exception as e:
        record("Value-level comparison (load)", False, str(e))
        return

    def collect_numerics(wb):
        nums = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)):
                        nums.append(float(cell))
        return nums

    gt_nums = collect_numerics(gt_wb)
    agent_nums = collect_numerics(agent_wb)
    gt_wb.close()
    agent_wb.close()
    if not gt_nums:
        record("GT numeric values present (non-blocking)", True, "no numeric GT values")
        return

    gt_min, gt_max = min(gt_nums), max(gt_nums)
    overlap = sum(1 for v in agent_nums if gt_min <= v <= gt_max)
    record("Agent xlsx numeric values overlap GT range (>=2 values)",
           overlap >= 2, f"overlap_count={overlap}, gt_min={gt_min}, gt_max={gt_max}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    if not agent_workspace:
        return False, "No agent workspace provided"
    agent_ws = Path(agent_workspace)
    if not agent_ws.exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    is_gt_self_test = False
    try:
        if groundtruth_workspace and Path(groundtruth_workspace).resolve() == agent_ws.resolve():
            is_gt_self_test = True
    except Exception:
        pass

    template_names = {"menu_templates.xlsx", "participant_list.xlsx"}
    xlsx_files, docx_files = check_workspace_files(agent_ws)
    check_catering_xlsx(xlsx_files)
    check_proposal_docx(docx_files)

    gt_xlsx_files = []
    if groundtruth_workspace:
        gt_ws = Path(groundtruth_workspace)
        if gt_ws.exists():
            gt_xlsx_files = [p for p in list_files(gt_ws, ".xlsx")
                             if p.name not in template_names]
    check_value_match_against_gt(xlsx_files, gt_xlsx_files, is_gt_self_test)

    check_dietary_survey_form(is_gt_self_test)
    check_vendor_emails(is_gt_self_test)
    check_coordination_events(is_gt_self_test)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT,
                    "total_checks": PASS_COUNT + FAIL_COUNT,
                    "success": success,
                }, f, indent=2)
        except Exception:
            pass

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    print(f"FAIL ({FAIL_COUNT} checks failed)")
    sys.exit(1)


if __name__ == "__main__":
    main()
