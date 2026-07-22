"""Evaluation for sf-sales-regional-summary."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    print("=== Checking Excel ===")
    agent_file = os.path.join(agent_workspace, "Regional_Sales.xlsx")
    gt_file = os.path.join(gt_dir, "Regional_Sales.xlsx")

    if not os.path.exists(agent_file):
        check("Agent Excel exists", False, agent_file)
        return
    check("Agent Excel exists", True)
    if not os.path.exists(gt_file):
        check("Groundtruth Excel exists", False, gt_file)
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Regional Breakdown sheet
    a_rows = load_sheet_rows(agent_wb, "Regional Breakdown")
    g_rows = load_sheet_rows(gt_wb, "Regional Breakdown")
    check("Sheet 'Regional Breakdown' present", a_rows is not None)
    if a_rows is not None and g_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]

        # Sort order verification (alphabetical by Region)
        agent_regions = [str(r[0]).strip() for r in a_data]
        check(
            "Regional Breakdown sorted alphabetically by Region",
            agent_regions == sorted(agent_regions),
            f"Got {agent_regions}",
        )

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            check(f"Region row '{g_row[0]}' present", a_row is not None)
            if a_row is None:
                continue
            # Orders: exact (tol 0)
            check(
                f"{g_row[0]}.Orders == {g_row[1]}",
                num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], 0),
                f"got {a_row[1] if len(a_row) > 1 else None}",
            )
            # Customers: exact (tol 0)
            check(
                f"{g_row[0]}.Customers == {g_row[2]}",
                num_close(a_row[2] if len(a_row) > 2 else None, g_row[2], 0),
                f"got {a_row[2] if len(a_row) > 2 else None}",
            )
            # Revenue: ±1.0
            check(
                f"{g_row[0]}.Revenue ≈ {g_row[3]}",
                num_close(a_row[3] if len(a_row) > 3 else None, g_row[3], 1.0),
                f"got {a_row[3] if len(a_row) > 3 else None}",
            )
            # AOV: ±0.01
            check(
                f"{g_row[0]}.Avg_Order_Value ≈ {g_row[4]}",
                num_close(a_row[4] if len(a_row) > 4 else None, g_row[4], 0.01),
                f"got {a_row[4] if len(a_row) > 4 else None}",
            )
            # Pct: ±0.1
            check(
                f"{g_row[0]}.Revenue_Share_Pct ≈ {g_row[5]}",
                num_close(a_row[5] if len(a_row) > 5 else None, g_row[5], 0.1),
                f"got {a_row[5] if len(a_row) > 5 else None}",
            )

    # Summary sheet
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    check("Sheet 'Summary' present", a_rows is not None)
    if a_rows is not None and g_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            check(f"Summary metric '{g_row[0]}' present", a_row is not None)
            if a_row is None:
                continue
            metric = str(g_row[0]).strip().lower()
            # Top_Region is a string name
            if metric == "top_region":
                check(
                    f"Summary.Top_Region == '{g_row[1]}'",
                    str_match(a_row[1] if len(a_row) > 1 else None, g_row[1]),
                    f"got {a_row[1] if len(a_row) > 1 else None}",
                )
            else:
                check(
                    f"Summary.{g_row[0]} ≈ {g_row[1]}",
                    num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], 1.0),
                    f"got {a_row[1] if len(a_row) > 1 else None}",
                )


def get_expected_regional():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
    SELECT c."REGION",
           COUNT(DISTINCT o."ORDER_ID") AS orders,
           COUNT(DISTINCT c."CUSTOMER_ID") AS customers,
           ROUND(SUM(o."TOTAL_AMOUNT")::numeric, 2) AS revenue
    FROM sf_data."SALES_DW__PUBLIC__ORDERS" o
    JOIN sf_data."SALES_DW__PUBLIC__CUSTOMERS" c ON o."CUSTOMER_ID" = c."CUSTOMER_ID"
    WHERE o."STATUS" = 'Delivered'
    GROUP BY c."REGION"
    ORDER BY c."REGION"
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    expected = {}
    for region, orders, customers, revenue in rows:
        expected[region.lower()] = {
            "Region": region,
            "Orders": int(orders),
            "Customers": int(customers),
            "Revenue": float(revenue),
        }
    return expected


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("Connect to gsheet DB", False, str(e))
        return
    cur.execute("""
    SELECT id, title FROM gsheet.spreadsheets
    WHERE LOWER(title) LIKE '%%regional sales%%' OR LOWER(title) LIKE '%%regional%%dashboard%%'
    """)
    sheets = cur.fetchall()
    check(
        "Spreadsheet titled 'Regional Sales Dashboard' exists",
        len(sheets) >= 1,
        f"Found {len(sheets)} matching spreadsheets",
    )
    if not sheets:
        cur.close()
        conn.close()
        return
    ss_id = sheets[0][0]

    # Verify "Overview" sheet exists
    cur.execute(
        "SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s",
        (ss_id,),
    )
    sheet_rows = cur.fetchall()
    sheet_titles = [r[1] for r in sheet_rows]
    check(
        "Sheet titled 'Overview' exists in spreadsheet",
        any(t and t.strip().lower() == "overview" for t in sheet_titles),
        f"Got sheets: {sheet_titles}",
    )

    # Get cells
    cur.execute(
        """
        SELECT c.row_index, c.col_index, c.value
        FROM gsheet.cells c
        WHERE c.spreadsheet_id = %s
        ORDER BY c.row_index, c.col_index
        """,
        (ss_id,),
    )
    cells = cur.fetchall()
    grid = {}
    for r, col, v in cells:
        grid.setdefault(r, {})[col] = str(v) if v is not None else ""

    expected = get_expected_regional()

    # Find header row containing 'region' and 'orders' and 'revenue'
    header_row = None
    for r in sorted(grid.keys()):
        joined = " ".join(grid[r].values()).lower()
        if "region" in joined and "orders" in joined and "revenue" in joined:
            header_row = r
            break
    check("GSheet header row with Region/Orders/Revenue found", header_row is not None)

    if header_row is not None:
        header_cells = grid[header_row]
        col_map = {}
        for col, val in header_cells.items():
            v = val.strip().lower()
            if v == "region":
                col_map["region"] = col
            elif v == "orders":
                col_map["orders"] = col
            elif v == "customers":
                col_map["customers"] = col
            elif v == "revenue":
                col_map["revenue"] = col
            elif v in ("avg_order_value", "avg order value"):
                col_map["aov"] = col
            elif "share" in v or "pct" in v:
                col_map["pct"] = col
        # Build region rows
        data_rows = [grid[r] for r in sorted(grid.keys()) if r > header_row]
        gsheet_lookup = {}
        for row in data_rows:
            region_val = row.get(col_map.get("region", -1), "").strip()
            if not region_val:
                continue
            gsheet_lookup[region_val.lower()] = row
        for region_key, exp in expected.items():
            row = gsheet_lookup.get(region_key)
            check(f"GSheet row for region '{exp['Region']}' present", row is not None)
            if row is None:
                continue
            try:
                orders = int(float(row.get(col_map.get("orders", -1), "0")))
            except (ValueError, TypeError):
                orders = None
            try:
                customers = int(float(row.get(col_map.get("customers", -1), "0")))
            except (ValueError, TypeError):
                customers = None
            try:
                revenue = float(row.get(col_map.get("revenue", -1), "0"))
            except (ValueError, TypeError):
                revenue = None
            check(
                f"GSheet '{exp['Region']}'.Orders == {exp['Orders']}",
                orders == exp["Orders"],
                f"got {orders}",
            )
            check(
                f"GSheet '{exp['Region']}'.Customers == {exp['Customers']}",
                customers == exp["Customers"],
                f"got {customers}",
            )
            check(
                f"GSheet '{exp['Region']}'.Revenue ≈ {exp['Revenue']}",
                revenue is not None and abs(revenue - exp["Revenue"]) <= 1.0,
                f"got {revenue}",
            )

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gsheet()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_passed = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
