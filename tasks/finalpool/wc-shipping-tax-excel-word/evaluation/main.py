"""Evaluation for wc-shipping-tax-excel-word."""
import os
import argparse, os, sys
import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Shipping_Tax_Analysis.xlsx")
    if not os.path.exists(path):
        return ["Shipping_Tax_Analysis.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        rows1 = load_sheet_rows(wb, "Shipping Zones")
        if rows1 is None:
            errors.append("Sheet 'Shipping Zones' not found")
        else:
            data_rows = [r for r in rows1[1:] if r and r[0] is not None]
            if len(data_rows) < 3:
                errors.append(f"Shipping Zones has {len(data_rows)} rows, expected at least 3")
            zone_names = {str(r[0]).strip().lower() for r in data_rows}
            for expected in ["domestic us", "california", "international"]:
                if not any(expected in z for z in zone_names):
                    errors.append(f"Expected zone '{expected}' not found in Shipping Zones")

        rows2 = load_sheet_rows(wb, "Tax Rates")
        if rows2 is None:
            errors.append("Sheet 'Tax Rates' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 6:
                errors.append(f"Tax Rates has {len(data_rows2)} rows, expected at least 6")

        # --- Groundtruth XLSX value comparison ---
        gt_path = os.path.join(groundtruth_workspace, "Shipping_Tax_Analysis.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = wb[asn]; break
                if a_ws is None:
                    errors.append(f"GT sheet '{gt_sname}' not found in agent xlsx (available: {wb.sheetnames})")
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                if len(a_rows) != len(gt_rows):
                    errors.append(f"GT '{gt_sname}' row count: expected {len(gt_rows)}, got {len(a_rows)}")
                # For each GT row, find a matching agent row (by all non-numeric/bool fields and rate)
                a_unmatched = list(a_rows)
                for ri, gtr in enumerate(gt_rows):
                    found_idx = None
                    for ai, ar in enumerate(a_unmatched):
                        if ar is None:
                            continue
                        all_match = True
                        for ci in range(min(len(gtr), len(ar))):
                            gv, av = gtr[ci], ar[ci]
                            if gv is None and av is None:
                                continue
                            if gv is None or av is None:
                                # Treat None vs empty string as match
                                if (gv is None and (av == "" or av == 0)) or (av is None and (gv == "" or gv == 0)):
                                    continue
                                all_match = False
                                break
                            if isinstance(gv, bool) or isinstance(av, bool):
                                if bool(gv) != bool(av):
                                    all_match = False
                                    break
                            elif isinstance(gv, (int, float)):
                                tol = max(abs(float(gv)) * 0.02, 0.05)
                                if not num_close(av, gv, tol):
                                    all_match = False
                                    break
                            else:
                                if not str_match(av, gv):
                                    all_match = False
                                    break
                        if all_match:
                            found_idx = ai
                            break
                    if found_idx is not None:
                        a_unmatched[found_idx] = None  # consume
                    else:
                        errors.append(f"GT '{gt_sname}' row {ri+1} {gtr} not matched in agent")
            gt_wb.close()
    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word_doc(agent_workspace):
    errors = []
    doc_path = os.path.join(agent_workspace, "Operations_Report.docx")
    if not os.path.exists(doc_path):
        return ["Operations_Report.docx not found"]
    try:
        from docx import Document
        doc = Document(doc_path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
        if "shipping" not in full_text:
            errors.append("Word doc does not contain 'shipping' keyword")
        if "tax" not in full_text:
            errors.append("Word doc does not contain 'tax' keyword")
        # Tightened length: 400 chars (substantive content)
        if len(full_text.strip()) < 400:
            errors.append(f"Word doc content too short ({len(full_text.strip())} chars, < 400)")
        # Recommendations section required
        if "recommendation" not in full_text:
            errors.append("Word doc missing 'recommendations' section")
        # Should mention zone count summary
        if "zone" not in full_text:
            errors.append("Word doc does not mention zones")
        # Should mention specific zones (at least 2 of GT zones)
        zone_mentions = sum(1 for z in ["california", "domestic", "international"] if z in full_text)
        if zone_mentions < 2:
            errors.append(f"Word doc mentions only {zone_mentions} zones (need >=2 of California/Domestic/International)")
        # Should mention BOTH tax class names (Standard AND Reduced)
        if "standard" not in full_text:
            errors.append("Word doc missing tax class name 'Standard'")
        if "reduced" not in full_text:
            errors.append("Word doc missing tax class name 'Reduced'")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            ORDER BY id DESC
        """)
        all_rows = cur.fetchall()
        cur.close(); conn.close()

        # Strict recipient match: parse to_addr (jsonb list or scalar) and require
        # exactly the target email to appear in the parsed list.
        import json as _json
        import re as _re_em
        target_email = "operations@store.com"
        rows = []
        for subj, to_addr, body in all_rows:
            recipients = []
            if isinstance(to_addr, list):
                recipients = [str(x).lower().strip() for x in to_addr]
            elif isinstance(to_addr, str):
                t = to_addr.strip()
                try:
                    parsed = _json.loads(t)
                    if isinstance(parsed, list):
                        recipients = [str(x).lower().strip() for x in parsed]
                    else:
                        recipients = [str(parsed).lower().strip()]
                except (TypeError, ValueError, _json.JSONDecodeError):
                    # Fallback: extract email-shaped tokens.
                    recipients = [m.lower() for m in _re_em.findall(r"[\w.+-]+@[\w.-]+", t)]
            else:
                continue
            # Require the target email to appear as one of the parsed recipients.
            if target_email in recipients:
                rows.append((subj, to_addr, body))

        if not rows:
            errors.append(f"No email found to exact recipient {target_email}")
        else:
            # Subject must contain 'Shipping and Tax Configuration Report'
            target_subj = "Shipping and Tax Configuration Report"
            target_match = None
            for subj, to, body in rows:
                if subj and target_subj.lower() in subj.lower():
                    target_match = (subj, to, body)
                    break
            if target_match is None:
                errors.append(f"No email with subject '{target_subj}' to {target_email} (got: {[r[0] for r in rows]})")
            else:
                body_lower = (target_match[2] or "").lower()
                # Body must mention shipping and tax
                if "shipping" not in body_lower:
                    errors.append("Email body does not mention shipping")
                if "tax" not in body_lower:
                    errors.append("Email body does not mention tax")
                # Tighten min length and require at least one specific zone name
                if len(body_lower.strip()) < 200:
                    errors.append(f"Email body too short ({len(body_lower)} chars; need >= 200)")
                zone_kw = ["california", "domestic", "international"]
                if not any(z in body_lower for z in zone_kw):
                    errors.append("Email body must mention at least one specific zone (California/Domestic/International)")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    gt_ws = args.groundtruth_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word_doc(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
