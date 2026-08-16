"""Evaluation for wc-customer-spending-report."""
import argparse
import os
import sys
import openpyxl

try:
    import psycopg2
    HAS_PG = True
except ImportError:
    HAS_PG = False

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_google_sheet():
    """Verify Google Sheet 'Customer Analysis' with sheet 'Spending' exists and has rows."""
    errors = []
    if not HAS_PG:
        return ["psycopg2 not available; cannot verify Google Sheet"]
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except psycopg2.OperationalError as e:
        return [f"Cannot connect to DB to verify Google Sheet: {e}"]
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gsheet.spreadsheets
            WHERE LOWER(title) LIKE %s
        """, ('%customer analysis%',))
        sheets = cur.fetchall()
        if not sheets:
            errors.append("Google Sheet titled 'Customer Analysis' not found")
            cur.close()
            conn.close()
            return errors
        ss_id = sheets[0][0]
        # Verify a child sheet named 'Spending' exists
        cur.execute("""
            SELECT id, title FROM gsheet.sheets
            WHERE spreadsheet_id = %s AND LOWER(title) LIKE %s
        """, (ss_id, '%spending%'))
        child_sheets = cur.fetchall()
        if not child_sheets:
            errors.append("Sub-sheet 'Spending' not found in 'Customer Analysis'")
            cur.close()
            conn.close()
            return errors
        sub_id = child_sheets[0][0]
        # Verify the 'Spending' sheet has at least 2 distinct rows (header + 1 data) and at least 6 columns
        cur.execute("""
            SELECT COUNT(DISTINCT row_index) FROM gsheet.cells
            WHERE spreadsheet_id = %s AND sheet_id = %s
        """, (ss_id, sub_id))
        row_count = cur.fetchone()[0]
        cur.execute("""
            SELECT COUNT(DISTINCT col_index) FROM gsheet.cells
            WHERE spreadsheet_id = %s AND sheet_id = %s AND row_index = 0
        """, (ss_id, sub_id))
        col_count = cur.fetchone()[0]
        # Header row + at least one data row; gt has 43 active customers
        if row_count < 2:
            errors.append(f"GSheet 'Spending' has too few rows: {row_count}")
        if col_count < 6:
            errors.append(f"GSheet 'Spending' has too few columns in header: {col_count}")
        # Check headers contain Customer_ID, Total_Spent
        cur.execute("""
            SELECT value FROM gsheet.cells
            WHERE spreadsheet_id = %s AND sheet_id = %s AND row_index = 0
            ORDER BY col_index
        """, (ss_id, sub_id))
        headers = [str(r[0] or "").strip().lower().replace(" ", "_") for r in cur.fetchall()]
        for required_h in ("customer_id", "total_spent"):
            if not any(required_h in h for h in headers):
                errors.append(f"GSheet 'Spending' missing required column: {required_h} (headers={headers})")
        cur.close()
        conn.close()
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        errors.append(f"Error verifying Google Sheet: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Customer_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Customer_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Customer Spending
    print(f"  Checking Customer Spending...")
    a_rows = load_sheet_rows(agent_wb, "Customer Spending")
    g_rows = load_sheet_rows(gt_wb, "Customer Spending")
    if a_rows is None:
        all_errors.append("Sheet 'Customer Spending' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Customer Spending' not found in groundtruth")
    else:
        sheet_name = "Customer Spending"
        errors = []
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if any(c is not None for c in r)]
        g_data = [r for r in (g_rows[1:] if len(g_rows) > 1 else []) if any(c is not None for c in r)]

        # Exact row count
        if len(a_data) != len(g_data):
            errors.append(f"Customer Spending row count: agent={len(a_data)} vs gt={len(g_data)}")

        # Total_Spent must be sorted descending (task.md requirement)
        totals_in_order = []
        for row in a_data:
            if row and len(row) > 4 and row[4] is not None:
                try:
                    totals_in_order.append(float(row[4]))
                except (TypeError, ValueError):
                    pass
        if totals_in_order and totals_in_order != sorted(totals_in_order, reverse=True):
            errors.append("Customer Spending not sorted by Total_Spent descending")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Validate Name (col 1) and Email (col 2) exact match
            if len(a_row) > 1 and len(g_row) > 1 and g_row[1] is not None:
                if not str_match(a_row[1], g_row[1]):
                    errors.append(f"{key}.Name: '{a_row[1]}' vs '{g_row[1]}'")
            if len(a_row) > 2 and len(g_row) > 2 and g_row[2] is not None:
                if not str_match(a_row[2], g_row[2]):
                    errors.append(f"{key}.Email: '{a_row[2]}' vs '{g_row[2]}'")

            # Orders is a count - exact match
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0):
                    errors.append(f"{key}.Orders: {a_row[3]} vs {g_row[3]} (exact)")

            if len(a_row) > 4 and len(g_row) > 4:
                # Tighter tol: rounded to 2 decimals → 0.05
                if not num_close(a_row[4], g_row[4], 0.05):
                    errors.append(f"{key}.Total_Spent: {a_row[4]} vs {g_row[4]} (tol=0.05)")

            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.05):
                    errors.append(f"{key}.Avg_Order_Value: {a_row[5]} vs {g_row[5]} (tol=0.05)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # Per-metric tolerance: counts exact, totals/averages within 1.0
                gv = g_row[1]
                tol = 0 if key in ("total_active_customers",) else 1.0
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check Google Sheet 'Customer Analysis'
    print(f"  Checking Google Sheet 'Customer Analysis'...")
    gs_errors = check_google_sheet()
    if gs_errors:
        all_errors.extend(gs_errors)
        for e in gs_errors[:5]:
            print(f"      {e}")
    else:
        print(f"    PASS")


    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
