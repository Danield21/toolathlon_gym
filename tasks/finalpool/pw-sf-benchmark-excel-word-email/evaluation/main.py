"""Evaluation script for pw-sf-benchmark-excel-word-email."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        s = str(val).replace(',', '').replace('%', '').replace('$', '').strip()
        if s.upper() == 'N/A':
            return default
        return float(s)
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def values_close(a, b, tolerance=0.03):
    """Check if two numeric values are within tolerance (relative)."""
    if a is None or b is None:
        return False
    if b == 0:
        return abs(a) < 1.0
    return abs(a - b) / max(abs(b), 1e-6) <= tolerance


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # --- Excel Evaluation ---
    excel_path = os.path.join(agent_workspace, "Benchmark_Analysis.xlsx")
    check("Benchmark_Analysis.xlsx exists", os.path.exists(excel_path))

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Benchmark_Analysis.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        # Sheet 1: Internal Metrics
        check("Internal Metrics sheet exists", "Internal Metrics" in wb.sheetnames,
              f"sheets: {wb.sheetnames}")
        if "Internal Metrics" in wb.sheetnames:
            ws = wb["Internal Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            # GT has 6 metric rows; require exact count
            gt_im_rows = []
            if gt_wb and "Internal Metrics" in gt_wb.sheetnames:
                gt_im_rows = list(gt_wb["Internal Metrics"].iter_rows(min_row=2, values_only=True))
            min_im = max(6, len(gt_im_rows))
            # Strict equality on row count when GT is known
            if gt_im_rows:
                check(f"Internal Metrics has == {min_im} rows", len(data_rows) == min_im, f"got {len(data_rows)}")
            else:
                check(f"Internal Metrics has >= {min_im} rows", len(data_rows) >= min_im, f"got {len(data_rows)}")

            # Sort: alphabetical by Metric
            metric_vals = [str(r[0]).strip() for r in data_rows if r and r[0]]
            check("Internal Metrics sorted alphabetically",
                  metric_vals == sorted(metric_vals, key=lambda s: s.lower()),
                  f"order: {metric_vals[:3]}")

            for expected_col in ['metric', 'internal_value', 'source_note']:
                check(f"Internal Metrics has '{expected_col}' column",
                      expected_col in headers, f"headers: {headers}")

            # Verify key metric values against groundtruth
            if gt_wb and "Internal Metrics" in gt_wb.sheetnames:
                gt_ws = gt_wb["Internal Metrics"]
                gt_data = {str(r[0]).strip(): r[1] for r in gt_ws.iter_rows(min_row=2, values_only=True) if r[0]}
                agent_data = {str(r[0]).strip(): r[1] for r in data_rows if r and r[0]}

                for metric_name in ['Avg_Salary', 'Avg_Order_Value', 'Avg_Satisfaction']:
                    gt_val = safe_float(gt_data.get(metric_name))
                    # Try variations of metric name
                    agent_val = None
                    for k, v in agent_data.items():
                        if metric_name.lower().replace('_', '') in k.lower().replace('_', '').replace(' ', ''):
                            agent_val = safe_float(v)
                            break
                    if agent_val is None:
                        agent_val = safe_float(agent_data.get(metric_name))
                    check(f"Internal Metrics: {metric_name} value matches",
                          values_close(agent_val, gt_val),
                          f"agent={agent_val}, gt={gt_val}")

        # Sheet 2: Gap Analysis
        check("Gap Analysis sheet exists", "Gap Analysis" in wb.sheetnames,
              f"sheets: {wb.sheetnames}")
        if "Gap Analysis" in wb.sheetnames:
            ws = wb["Gap Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            # GT Gap Analysis has 4 calculable rows (N/A metrics excluded)
            gt_ga_rows = []
            if gt_wb and "Gap Analysis" in gt_wb.sheetnames:
                gt_ga_rows = list(gt_wb["Gap Analysis"].iter_rows(min_row=2, values_only=True))
            min_ga = max(4, len(gt_ga_rows))
            if gt_ga_rows:
                check(f"Gap Analysis has == {min_ga} rows", len(data_rows) == min_ga, f"got {len(data_rows)}")
            else:
                check(f"Gap Analysis has >= {min_ga} rows", len(data_rows) >= min_ga, f"got {len(data_rows)}")

            # Sort: alphabetical by Metric (per task.md)
            metric_vals = [str(r[0]).strip() for r in data_rows if r and r[0]]
            check("Gap Analysis sorted alphabetically by metric",
                  metric_vals == sorted(metric_vals, key=lambda s: s.lower()),
                  f"order: {metric_vals[:3]}")

            for expected_col in ['metric', 'internal_value', 'industry_avg', 'classification', 'priority']:
                check(f"Gap Analysis has '{expected_col}' column",
                      expected_col in headers, f"headers: {headers}")

            # Check classifications
            if gt_wb and "Gap Analysis" in gt_wb.sheetnames:
                gt_ws = gt_wb["Gap Analysis"]
                gt_class = {}
                for r in gt_ws.iter_rows(min_row=2, values_only=True):
                    if r and r[0]:
                        gt_class[str(r[0]).strip()] = str(r[7]).strip() if r[7] else ""

                agent_class = {}
                class_col_idx = None
                for i, h in enumerate(headers):
                    if 'classification' in h:
                        class_col_idx = i
                        break
                if class_col_idx is not None:
                    for r in data_rows:
                        if r and r[0]:
                            agent_class[str(r[0]).strip()] = str(r[class_col_idx]).strip() if r[class_col_idx] else ""

                for metric_name in ['Avg_Order_Value', 'Avg_Salary', 'Avg_Satisfaction', 'Revenue_Per_Employee']:
                    gt_c = gt_class.get(metric_name, "")
                    agent_c = None
                    for k, v in agent_class.items():
                        if metric_name.lower().replace('_', '') in k.lower().replace('_', '').replace(' ', ''):
                            agent_c = v
                            break
                    if agent_c is None:
                        agent_c = agent_class.get(metric_name, "")
                    check(f"Gap Analysis: {metric_name} classification matches",
                          gt_c.lower() == str(agent_c).lower(),
                          f"agent='{agent_c}', gt='{gt_c}'")

        # Sheet 3: Action Plan
        check("Action Plan sheet exists", "Action Plan" in wb.sheetnames,
              f"sheets: {wb.sheetnames}")
        if "Action Plan" in wb.sheetnames:
            ws = wb["Action Plan"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_ap_rows = []
            if gt_wb and "Action Plan" in gt_wb.sheetnames:
                gt_ap_rows = list(gt_wb["Action Plan"].iter_rows(min_row=2, values_only=True))
            min_ap = max(2, len(gt_ap_rows))
            if gt_ap_rows:
                check(f"Action Plan has == {min_ap} rows", len(data_rows) == min_ap, f"got {len(data_rows)}")
            else:
                check(f"Action Plan has >= {min_ap} rows", len(data_rows) >= min_ap, f"got {len(data_rows)}")

            for expected_col in ['metric', 'classification', 'recommended_action',
                                 'estimated_impact', 'timeline']:
                check(f"Action Plan has '{expected_col}' column",
                      expected_col in headers, f"headers: {headers}")

            # Sort: by priority ascending. Critical (priority 1) before Moderate (priority 2)
            class_idx = headers.index('classification') if 'classification' in headers else None
            if class_idx is not None and data_rows:
                priority_order = []
                for r in data_rows:
                    if r and class_idx < len(r) and r[class_idx]:
                        c = str(r[class_idx]).lower()
                        if 'critical' in c:
                            priority_order.append(1)
                        elif 'moderate' in c:
                            priority_order.append(2)
                        elif 'on track' in c:
                            priority_order.append(3)
                        elif 'leading' in c:
                            priority_order.append(4)
                        else:
                            priority_order.append(99)
                check("Action Plan sorted by priority ascending",
                      priority_order == sorted(priority_order),
                      f"priorities={priority_order}")

            # Action Plan content vs GT: each metric has non-empty action/impact/timeline
            if gt_ap_rows:
                ra_idx = headers.index('recommended_action') if 'recommended_action' in headers else None
                ei_idx = headers.index('estimated_impact') if 'estimated_impact' in headers else None
                tl_idx = headers.index('timeline') if 'timeline' in headers else None
                gt_metrics = {str(r[0]).strip() for r in gt_ap_rows if r and r[0]}
                a_by_metric = {str(r[0]).strip(): r for r in data_rows if r and r[0]}
                for m in gt_metrics:
                    if m in a_by_metric:
                        r = a_by_metric[m]
                        for label, idx in [("recommended_action", ra_idx),
                                           ("estimated_impact", ei_idx),
                                           ("timeline", tl_idx)]:
                            if idx is None or idx >= len(r):
                                continue
                            v = (str(r[idx]).strip() if r[idx] is not None else "")
                            check(f"Action Plan '{m}' {label} non-empty",
                                  len(v) >= 5, f"got {v!r}")

    # --- Word Document Evaluation ---
    word_path = os.path.join(agent_workspace, "Benchmark_Report.docx")
    check("Benchmark_Report.docx exists", os.path.exists(word_path))

    if os.path.exists(word_path):
        from docx import Document
        doc = Document(word_path)
        full_text = "\n".join([p.text for p in doc.paragraphs])
        full_lower = full_text.lower()

        required_sections = [
            "Executive Summary", "Methodology", "Internal Performance Overview",
            "Benchmark Comparison", "Gap Analysis", "Strategic Recommendations",
            "Implementation Timeline"
        ]
        for section in required_sections:
            check(f"Word doc has '{section}' section",
                  section.lower() in full_lower,
                  "section not found")

        # Derive expected reference values from GT instead of hardcoding.
        gt_salary = None
        gt_order_value = None
        if gt_wb and "Internal Metrics" in gt_wb.sheetnames:
            for r in gt_wb["Internal Metrics"].iter_rows(min_row=2, values_only=True):
                if r and r[0]:
                    name = str(r[0]).strip().lower()
                    if name == "avg_salary":
                        gt_salary = safe_float(r[1])
                    elif name == "avg_order_value":
                        gt_order_value = safe_float(r[1])
        # Build acceptable salary tokens (handles 58396, 58,396, 58396.14)
        def _value_tokens(v):
            if v is None:
                return []
            s_int = str(int(round(v)))
            s_int_comma = f"{int(round(v)):,}"
            s_full = f"{v}"
            s_one = f"{v:.1f}"
            return list({s_int, s_int_comma, s_full, s_one})
        salary_tokens = _value_tokens(gt_salary or 58396)
        order_tokens = _value_tokens(gt_order_value or 152.45)
        check("Word doc mentions avg salary value",
              any(t in full_text for t in salary_tokens),
              f"expected one of {salary_tokens[:3]}")
        check("Word doc mentions avg order value",
              any(t in full_text for t in order_tokens),
              f"expected one of {order_tokens[:3]}")
        # Scope: classification labels must appear inside the Gap Analysis
        # section of the report (not just any random word elsewhere).
        ga_idx = full_lower.find("gap analysis")
        if ga_idx >= 0:
            ga_section = full_lower[ga_idx:ga_idx + 4000]
        else:
            ga_section = ""
        check("Word doc Gap Analysis section has 'Critical' classification",
              "critical" in ga_section,
              "label not found within 4000 chars after 'Gap Analysis' heading")
        check("Word doc Gap Analysis section has 'Moderate' classification",
              "moderate" in ga_section,
              "label not found within 4000 chars after 'Gap Analysis' heading")

    # --- Email Evaluation ---
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Operations head email - exact subject match
        cur.execute("""SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s
            AND subject ILIKE %s""",
            ('%operations_head%', '%Annual Benchmark Analysis - Complete Report%'))
        ops_emails = cur.fetchall()
        check("Email to operations_head@company.com with correct subject",
              len(ops_emails) >= 1, f"found {len(ops_emails)}")
        # Body should mention BOTH a numeric metric value AND a classification term
        if ops_emails:
            body = (ops_emails[0][2] or "").lower()
            has_number = any(t.lower() in body for t in (salary_tokens + order_tokens))
            has_class = any(t in body for t in ["critical", "moderate", "on track", "leading"])
            check("Operations email body mentions a metric value",
                  has_number, body[:200])
            check("Operations email body mentions a classification label",
                  has_class, body[:200])

        # HR director email - exact subject
        cur.execute("""SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s
            AND subject ILIKE %s""",
            ('%hr_director%', '%Benchmark Review - HR Metrics%'))
        hr_emails = cur.fetchall()
        check("Email to hr_director@company.com with correct subject",
              len(hr_emails) >= 1, f"found {len(hr_emails)}")
        # HR email must cover BOTH salary AND satisfaction (both are HR-domain metrics).
        if hr_emails:
            body = (hr_emails[0][2] or "").lower()
            has_salary_topic = ("salary" in body) or any(t.lower() in body for t in salary_tokens)
            has_satisfaction_topic = ("satisfaction" in body) or ("6.55" in body) or ("6.5" in body)
            check("HR email body mentions salary topic",
                  has_salary_topic, body[:200])
            check("HR email body mentions satisfaction topic",
                  has_satisfaction_topic, body[:200])

        # Sales VP email - exact subject
        cur.execute("""SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s
            AND subject ILIKE %s""",
            ('%sales_vp%', '%Benchmark Review - Sales Metrics%'))
        sales_emails = cur.fetchall()
        check("Email to sales_vp@company.com with correct subject",
              len(sales_emails) >= 1, f"found {len(sales_emails)}")
        # Sales email must cover BOTH order value AND revenue topics.
        if sales_emails:
            body = (sales_emails[0][2] or "").lower()
            has_order_topic = ("order value" in body) or ("order_value" in body) or any(t.lower() in body for t in order_tokens)
            has_revenue_topic = ("revenue" in body) or ("60.98" in body) or ("60.9" in body)
            check("Sales email body mentions order-value topic",
                  has_order_topic, body[:200])
            check("Sales email body mentions revenue topic",
                  has_revenue_topic, body[:200])

        conn.close()
    except Exception as e:
        check("Email verification", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
