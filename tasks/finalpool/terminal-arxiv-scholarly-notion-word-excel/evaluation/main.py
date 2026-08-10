"""Evaluation for terminal-arxiv-scholarly-notion-word-excel."""
import argparse
import os
import re
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"),
          port=int(os.environ.get("PGPORT", "5432")),
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user=os.environ.get("PGUSER", "eigent"),
          password=os.environ.get("PGPASSWORD", "camel"))

PASS_COUNT = 0
FAIL_COUNT = 0
RUNTIME_ONLY_FAIL = 0

TRANSFORMER_IDS = {"1706.03762", "1810.04805", "2005.14165",
                   "1409.0473", "1910.10683", "2009.06732"}
NOISE_IDS = {"1207.00580", "1502.03167", "1312.06199"}
CATEGORIES = ["Architecture Design", "Training Methods", "Applications", "Survey"]

_ARXIV_ID_RE = re.compile(r"(\d{4}\.\d{4,5})")
_WS_RE = re.compile(r"\s+")
# Default empty sheet-tab names produced by Excel/openpyxl/LibreOffice when a
# workbook is created ("Sheet", "Sheet1", "Sheet2", ...).  A faithful model
# that fills in the three required sheets but leaves the default blank tab is
# not doing anything wrong, so these must not count as "unexpected".
_DEFAULT_SHEET_RE = re.compile(r"^sheet\d*$", re.IGNORECASE)

# Case-/whitespace-insensitive set of valid methodology categories.  Category
# matching must tolerate a model normalizing to lowercase ("architecture
# design"), which is common LLM formatting behavior.
CATEGORY_KEYS = {_WS_RE.sub(" ", c).strip().lower() for c in CATEGORIES}


def _norm(v):
    """Normalize a cell value for text comparison (strip whitespace and a
    leading '=' that openpyxl returns for formula cells when reading with
    data_only=False)."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v).strip()
    return str(v).strip().lstrip("=")


def _paper_id(v):
    """Extract a canonical arXiv-style paper ID from any cell value.

    Handles bare IDs ("1706.03762"), URLs ("http://arxiv.org/abs/1706.03762"),
    and versioned IDs ("1706.03762v2"). Falls back to the lowercased raw
    string when no arXiv-style ID pattern is present.
    """
    s = _norm(v)
    if not s:
        return None
    m = _ARXIV_ID_RE.search(s)
    return m.group(1) if m else s.lower()


def _find_sheet(wb, *keywords):
    """Locate a sheet by keyword match on a normalized name, falling back to
    the exact normalized lookup. Handles 'Paper_Catalog', 'Paper Catalog',
    'PaperCatalog', 'citation matrix', etc."""
    sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}
    for s in wb.sheetnames:
        low = s.lower().replace(" ", "_")
        if all(k in low for k in keywords):
            return s
    if len(keywords) == 1:
        return sn.get(keywords[0])
    return None


def _norm_key(v):
    """Normalize a value for keyword/category matching: strip, collapse
    internal whitespace, lowercase.  Keeps category checks immune to case and
    spacing variants ("Architecture Design" vs "architecture design")."""
    return _WS_RE.sub(" ", _norm(v)).strip().lower()


def _cell(r, idx):
    """Safely index into a worksheet row; None when out of range or missing."""
    if r is None or idx is None:
        return None
    try:
        return r[idx] if 0 <= idx < len(r) else None
    except TypeError:
        return None


def _col(header, keywords, default):
    """Locate a data column by header keyword (case- and whitespace-insensitive),
    falling back to the spec position.  Anchoring by header keyword keeps the
    checks correct when a model writes the same columns in a different order or
    with slightly different header text."""
    if header:
        for i, h in enumerate(header):
            if h is None:
                continue
            norm = _WS_RE.sub("", str(h)).lower()
            if any(k in norm for k in keywords):
                return i
    return default


def _sheet_is_blank(wb, name):
    """True when a sheet contains no non-empty cells (e.g. the leftover default
    empty tab that a new workbook is born with)."""
    ws = wb[name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                return False
    return True


def _is_yes(v):
    """Interpret a cell as a boolean 'yes' (Yes / Y / true / 1)."""
    s = _norm(v).strip().lower()
    return s in ("y", "yes", "true", "1")


def _rt_text(v):
    """Recursively extract plain text from a Notion rich-text shaped JSON value.
    Handles a list of rich-text objects, a bare string, a dict carrying
    plain_text / text.content, or nested property wrappers."""
    out = []

    def rec(x):
        if isinstance(x, str):
            out.append(x)
        elif isinstance(x, dict):
            if x.get("plain_text"):
                out.append(x["plain_text"])
            txt = x.get("text")
            if isinstance(txt, dict) and txt.get("content"):
                out.append(txt["content"])
            elif isinstance(txt, str):
                out.append(txt)
            for v2 in x.values():
                rec(v2)
        elif isinstance(x, (list, tuple)):
            for v2 in x:
                rec(v2)

    rec(v)
    return " ".join(str(t) for t in out if t and str(t).strip())


def check(name, condition, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, RUNTIME_ONLY_FAIL
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if runtime_only:
            RUNTIME_ONLY_FAIL += 1
        d = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def check_excel(ws_path):
    """Check Research_Paper_Analysis.xlsx."""
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "Research_Paper_Analysis.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        return
    check("Excel file exists", True)

    # Read with data_only=False so formula cells are not silently lost (R2).
    wb = openpyxl.load_workbook(path, data_only=False)

    # Paper_Catalog
    pc_name = _find_sheet(wb, "paper", "catalog")
    if pc_name is None:
        check("Paper_Catalog sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Paper_Catalog sheet exists", True)
        ws = wb[pc_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0] if rows else None
        id_col = _col(header, ("paperid", "paper_id"), 0)
        cat_col = _col(header, ("category",), 5)
        data = [r for r in rows[1:] if r and _norm(_cell(r, id_col))]
        # All 9 papers are reachable through the task's search keywords, so a
        # complete submission has all 9.  Use a lower bound (all 6 transformer
        # papers at minimum) so a faithful submission is never failed over row
        # count alone.
        check("Paper_Catalog has at least 6 rows", len(data) >= 6,
              f"Found {len(data)}")

        # Check that transformer papers are present
        ids_found = {pid for r in data if (pid := _paper_id(_cell(r, id_col)))}
        transformer_found = len(TRANSFORMER_IDS & ids_found)
        check("All 6 transformer papers listed", transformer_found == 6,
              f"Found {transformer_found}/6")

        # Check categories assigned (case- and whitespace-insensitive)
        cats_found = {_norm_key(_cell(r, cat_col)) for r in data}
        cats_found.discard("")
        valid_cats = sum(1 for c in cats_found if c in CATEGORY_KEYS)
        check("Valid categories assigned", valid_cats >= 2,
              f"Categories found: {cats_found}")

    # Method_Comparison
    mc_name = _find_sheet(wb, "method", "comparison")
    if mc_name is None:
        check("Method_Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Method_Comparison sheet exists", True)
        ws2 = wb[mc_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if r and _norm(r[0])]
        check("Method_Comparison has at least 3 category rows", len(data2) >= 3,
              f"Found {len(data2)}")

    # Citation_Matrix
    cm_name = _find_sheet(wb, "citation", "matrix")
    if cm_name is None:
        check("Citation_Matrix sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Citation_Matrix sheet exists", True)
        ws3 = wb[cm_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        header3 = rows3[0] if rows3 else None
        id3 = _col(header3, ("paperid", "paper_id"), 0)
        overlap_col = _col(header3, ("overlap",), 4)
        data3 = [r for r in rows3[1:] if r and _norm(_cell(r, id3))]
        check("Citation_Matrix has at least 6 rows", len(data3) >= 6,
              f"Found {len(data3)}")

        # Check overlap marking (tolerant of Yes / Y / true / 1)
        overlap_count = 0
        for r in data3:
            pid = _paper_id(_cell(r, id3))
            if pid in TRANSFORMER_IDS:
                if _is_yes(_cell(r, overlap_col)):
                    overlap_count += 1
        check("All 6 transformer papers marked as overlap", overlap_count == 6,
              f"Found {overlap_count} (expected 6)")

    wb.close()


def check_word(ws_path):
    """Check Transformer_Literature_Review.docx."""
    print("\n=== Checking Word Document ===")
    path = os.path.join(ws_path, "Transformer_Literature_Review.docx")
    if not os.path.isfile(path):
        check("Word document exists", False, f"Not found: {path}")
        return
    check("Word document exists", True)

    from docx import Document
    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    check("Document mentions transformer", "transformer" in full_text)
    check("Document mentions attention", "attention" in full_text)
    check("Document mentions BERT or pre-training",
          "bert" in full_text or "pre-train" in full_text or "pretrain" in full_text)
    check("Document has conclusion section",
          "conclusion" in full_text or "summary" in full_text)
    check("Document length >= 500 chars", len(full_text) >= 500,
          f"Length: {len(full_text)}")


def check_notion():
    """Check Notion database creation.

    Notion checks are runtime-only soft indicators.  Any DB connection or
    query failure must degrade to a soft FAIL instead of crashing the whole
    evaluator.
    """
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        check("Notion DB reachable", False, str(e), runtime_only=True)
        return

    try:
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases")
        databases = cur.fetchall()

        found_db = None
        for db_id, title_json in databases:
            title_str = _rt_text(title_json)
            if "research" in title_str.lower() and "paper" in title_str.lower():
                found_db = db_id
                break
            if "tracker" in title_str.lower():
                found_db = db_id
                break

        check("Research Paper Tracker database exists", found_db is not None,
              f"Databases: {[d[1] for d in databases]}",
              runtime_only=True)

        if found_db:
            cur.execute(
                "SELECT COUNT(*) FROM notion.pages WHERE parent->>'database_id' = %s",
                (found_db,)
            )
            page_count = cur.fetchone()[0]
            check("Notion database has >= 6 paper entries", page_count >= 6,
                  f"Found {page_count} pages",
                  runtime_only=True)
        conn.close()
    except Exception as e:
        check("Notion checks completed", False, str(e), runtime_only=True)
        try:
            conn.close()
        except Exception:
            pass


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Research_Paper_Analysis.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            # Check no unexpected sheets beyond the 3 required
            valid_keywords = ["paper", "catalog", "method", "comparison", "citation", "matrix"]
            unexpected = [
                s for s in wb.sheetnames
                if not any(k in s.lower() for k in valid_keywords)
                and not _DEFAULT_SHEET_RE.match(s)
                and not _sheet_is_blank(wb, s)
            ]
            check("No unexpected sheets in Excel", len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")
            # Check no duplicate paper IDs
            pc_name = _find_sheet(wb, "paper", "catalog")
            if pc_name:
                ws = wb[pc_name]
                rows = list(ws.iter_rows(values_only=True))
                header = rows[0] if rows else None
                id_col = _col(header, ("paperid", "paper_id"), 0)
                ids = [pid for r in rows[1:] if r and (pid := _paper_id(_cell(r, id_col)))]
                check("No duplicate paper IDs in Paper_Catalog",
                      len(ids) == len(set(ids)),
                      f"Found {len(ids)} IDs but {len(set(ids))} unique")
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-ARXIV-SCHOLARLY-NOTION-WORD-EXCEL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    non_runtime_fail = FAIL_COUNT - RUNTIME_ONLY_FAIL
    overall = non_runtime_fail == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (runtime-only fails: {RUNTIME_ONLY_FAIL})")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
