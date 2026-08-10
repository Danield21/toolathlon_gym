"""Evaluation for terminal-canvas-excel-word-notion-email.
Checks:
1. Student_Risk_Analysis.xlsx with 4 sheets and correct data
2. Intervention_Plan.docx with required sections
3. Notion database "Student Risk Tracker" with 2 entries
4. Email sent to academic_advisors@university.edu
5. risk_scorer.py script exists

Hardening notes (from the audit):
- R1: DB connection fully reads env vars (PGHOST/PGPORT/PGDATABASE/PGUSER/PGPASSWORD) with
  defaults aligned to preprocess/main.py.
- R2: workbooks are loaded both with data_only=False (to detect formulas) and data_only=True
  (cached values). A formula cell whose cached value is None is SKIPPED (not FAILed); literal
  values are parsed robustly (strips commas, currency, %).
- R3: num_close parses both sides to float; falls back to case-insensitive string compare only
  when one side is unparseable.
- P5: Intervention_Plan timeline keywords accept both numeric ("1 week") and spelled-out
  ("one week") forms.
- P6: course_name comparison is substring/token tolerant (handles parens/dashes/year variants).
- Multi-agent robustness: Notion entries are aggregated across every matching "Student Risk
  Tracker" database so a duplicated/partial creation does not fail a correct solution.
- R4: Notion entries are validated by their aggregated numeric data (avg score / pass rate /
  student count per course), not by literal '2013'/'2014' substrings, so entries named by
  course id still pass. risk_level cells accept exact/substring/separator-collapsed variants.
- R5: Word doc title / section checks are keyword-tolerant (accepts longer titles, 'Risk
  Breakdown', summary synonyms) and course coverage is checked via years OR course ids.
- R6: email body high-risk count uses a newline-tolerant proximity regex.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def _to_float(v):
    """Robustly convert a value to float; None if not parseable.

    Handles int/float/str (strips thousand-separators, currency symbols, '%', spaces).
    A string starting with '=' is treated as an unparseable formula placeholder.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    if s.startswith("="):
        return None
    for ch in [",", "$", "¥", "€", "%", " ", " "]:
        s = s.replace(ch, "")
    try:
        return float(s)
    except ValueError:
        return None


def num_close(a, b, tol=2.0):
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # one side unparseable -> fall back to case-insensitive string comparison
    return str(a or "").strip().lower() == str(b or "").strip().lower()


def _cell_is_formula(raw_val):
    return raw_val is not None and isinstance(raw_val, str) and raw_val.startswith("=")


def _cell_float(raw_val, cache_val):
    """Float for a raw cell value; uses cached value for formula cells.

    Returns None if the cell is an unevaluated formula or empty/unparseable.
    """
    if _cell_is_formula(raw_val):
        return _to_float(cache_val)
    return _to_float(raw_val)


def check_num(name, raw_val, cache_val, g_val, tol, detail=""):
    """Numeric comparison that tolerates formula cells without cached values (SKIP)."""
    if _cell_is_formula(raw_val) and _to_float(cache_val) is None:
        print(f"  [SKIP] {name}: formula cell with no cached value")
        return
    check(name, num_close(_cell_float(raw_val, cache_val), g_val, tol), detail)


def _norm(s):
    return str(s or "").strip().lower().replace(" ", "_")


def _norm_id(v):
    s = str(v or "").strip().lower()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _idx(headers, h):
    return headers.index(h) if h in headers else -1


def _name_matches(a, g, cid=None):
    a = str(a or "").lower()
    g = str(g or "").lower()
    if not a or not g:
        return False
    if a == g:
        return True
    if g in a or a in g:
        return True
    # tolerant fallback: same subject + same year
    if "foundations of finance" in a and "foundations of finance" in g:
        ya = re.search(r"(20\d\d)", a)
        yg = re.search(r"(20\d\d)", g)
        if ya and yg and ya.group(1) == yg.group(1):
            return True
    # a correct model may write the course-id form ('Course 16') instead of the full name
    if cid is not None:
        an = re.sub(r"[^a-z0-9]", "", a)
        if an in ("course" + _norm_id(cid), "courseid" + _norm_id(cid)):
            return True
    return False


def _risk_matches(a, g):
    """Case/separator-insensitive match of a risk-level cell against a reference key.

    Accepts exact ('high'), substring ('High Risk' contains 'high'), and
    separator-collapsed ('High-risk' / 'Highrisk' -> 'highrisk') variants so a model that
    writes 'High Risk' / 'High-risk' in the risk_level column is not failed.
    """
    a = str(a or "").strip().lower()
    g = str(g or "").strip().lower()
    if not a or not g:
        return False
    if a == g:
        return True
    if g in a or a in g:
        return True
    aa = re.sub(r"[^a-z0-9]", "", a)
    gg = re.sub(r"[^a-z0-9]", "", g)
    if aa == gg or gg in aa or aa in gg:
        return True
    return False


def _load_wb(path):
    """Return (raw_wb, cache_wb). raw preserves formula strings; cache holds computed values."""
    return (openpyxl.load_workbook(path, data_only=False),
            openpyxl.load_workbook(path, data_only=True))


def _data_pairs(rows_raw, rows_cache):
    """Return aligned [(raw_row, cache_row), ...] for data rows after the header.

    Filtering is driven by the raw rows so an all-None cache row (formulas without a
    cached value) does not desync indices.
    """
    pairs = []
    for i in range(1, len(rows_raw)):
        rr = rows_raw[i]
        if rr and any(c is not None for c in rr):
            cr = rows_cache[i] if i < len(rows_cache) else None
            pairs.append((rr, cr))
    return pairs


def _gt_sheet_data(gt_wb_raw, sheet):
    """Return (headers, data_rows) for a GT sheet loaded with data_only=False."""
    if gt_wb_raw is None or sheet not in gt_wb_raw.sheetnames:
        return None, None
    ws = gt_wb_raw[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = [_norm(c) for c in rows[0]] if rows else []
    data = [r for r in rows[1:] if r and any(c is not None for c in r)]
    return headers, data


def check_excel(workspace, gt_dir):
    print("\n=== Check 1: Student_Risk_Analysis.xlsx ===")
    path = os.path.join(workspace, "Student_Risk_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    try:
        wb_raw, wb_cache = _load_wb(path)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return
    check("Excel file readable", True)

    gt_path = os.path.join(gt_dir, "Student_Risk_Analysis.xlsx")
    gt_wb_raw = None
    gt_wb_cache = None
    if os.path.exists(gt_path):
        try:
            gt_wb_raw, gt_wb_cache = _load_wb(gt_path)
        except Exception as e:
            print(f"  [WARN] Could not load GT workbook: {e}")

    sheets = wb_raw.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # ------- Course_Overview -------
    co_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s and "overview" in s), None)
    if co_idx is None:
        check("Course_Overview sheet exists", False, f"Sheets: {sheets}")
    else:
        check("Course_Overview sheet exists", True)
        ws1 = wb_raw[sheets[co_idx]]
        ws1c = wb_cache[sheets[co_idx]]
        rows1 = list(ws1.iter_rows(values_only=True))
        rows1c = list(ws1c.iter_rows(values_only=True))
        pairs1 = _data_pairs(rows1, rows1c)
        check("Course_Overview has at least 2 course rows", len(pairs1) >= 2, f"Found {len(pairs1)}")
        if rows1:
            headers = [_norm(c) for c in rows1[0]]
            for h in ["course_id", "course_name", "enrollment_count", "avg_score", "pass_rate"]:
                check(f"Course_Overview has '{h}' header", h in headers, f"got {headers}")
            gt_headers, gt_data = _gt_sheet_data(gt_wb_raw, "Course_Overview")
            if gt_headers is not None and gt_data is not None:
                cid_idx = _idx(headers, "course_id")
                a_lookup = {}
                for i, (rr, _cr) in enumerate(pairs1):
                    if cid_idx >= 0 and cid_idx < len(rr):
                        a_lookup[_norm_id(rr[cid_idx])] = i
                for g_row in gt_data:
                    gcid = _norm_id(g_row[gt_headers.index("course_id")])
                    idx = a_lookup.get(gcid)
                    if idx is None:
                        check(f"Course_Overview has course {gcid}", False,
                              f"Not in {list(a_lookup.keys())}")
                        continue
                    rr, cr = pairs1[idx]
                    # course_name (substring / token tolerant)
                    cn_idx = _idx(headers, "course_name")
                    if cn_idx >= 0 and cn_idx < len(rr):
                        a_name = str(rr[cn_idx] or "")
                        g_name = str(g_row[gt_headers.index("course_name")] or "")
                        check(f"course {gcid} name matches GT", _name_matches(a_name, g_name, gcid),
                              f"a='{a_name}' g='{g_name}'")
                    # enrollment_count (tol 5)
                    ec_idx = _idx(headers, "enrollment_count")
                    if ec_idx >= 0 and ec_idx < len(rr):
                        cache = cr[ec_idx] if cr is not None and ec_idx < len(cr) else None
                        check_num(f"course {gcid} enrollment_count", rr[ec_idx], cache,
                                  g_row[gt_headers.index("enrollment_count")], 5,
                                  f"a={rr[ec_idx]} g={g_row[gt_headers.index('enrollment_count')]}")
                    # avg_score (tol 1.0)
                    as_idx = _idx(headers, "avg_score")
                    if as_idx >= 0 and as_idx < len(rr):
                        cache = cr[as_idx] if cr is not None and as_idx < len(cr) else None
                        check_num(f"course {gcid} avg_score", rr[as_idx], cache,
                                  g_row[gt_headers.index("avg_score")], 1.0,
                                  f"a={rr[as_idx]} g={g_row[gt_headers.index('avg_score')]}")
                    # pass_rate (tol 1.0)
                    pr_idx = _idx(headers, "pass_rate")
                    if pr_idx >= 0 and pr_idx < len(rr):
                        cache = cr[pr_idx] if cr is not None and pr_idx < len(cr) else None
                        check_num(f"course {gcid} pass_rate", rr[pr_idx], cache,
                                  g_row[gt_headers.index("pass_rate")], 1.0,
                                  f"a={rr[pr_idx]} g={g_row[gt_headers.index('pass_rate')]}")

    # ------- Risk_Distribution -------
    rd_idx = next((i for i, s in enumerate(sheets_lower) if "risk" in s and "dist" in s), None)
    if rd_idx is None:
        check("Risk_Distribution sheet exists", False)
    else:
        check("Risk_Distribution sheet exists", True)
        ws2 = wb_raw[sheets[rd_idx]]
        ws2c = wb_cache[sheets[rd_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        rows2c = list(ws2c.iter_rows(values_only=True))
        pairs2 = _data_pairs(rows2, rows2c)
        check("Risk_Distribution has at least 3 rows", len(pairs2) >= 3, f"Found {len(pairs2)}")
        if rows2:
            headers2 = [_norm(c) for c in rows2[0]]
            for h in ["risk_level", "student_count", "pct"]:
                check(f"Risk_Distribution has '{h}' header", h in headers2, f"got {headers2}")
            gt_headers2, gt_data2 = _gt_sheet_data(gt_wb_raw, "Risk_Distribution")
            if gt_headers2 is not None and gt_data2 is not None:
                rl_idx = _idx(headers2, "risk_level")
                for g_row in gt_data2:
                    grl = str(g_row[gt_headers2.index("risk_level")]).strip().lower()
                    idx = next(
                        (i for i, (rr, _cr) in enumerate(pairs2)
                         if rl_idx >= 0 and rl_idx < len(rr) and _risk_matches(rr[rl_idx], grl)),
                        None)
                    if idx is None:
                        check(f"Risk_Distribution has '{grl}' row", False,
                              f"risk_levels: {[str(r[rl_idx]) if rl_idx < len(r) else '' for r, _ in pairs2]}")
                        continue
                    rr, cr = pairs2[idx]
                    sc_idx = _idx(headers2, "student_count")
                    pct_idx = _idx(headers2, "pct")
                    if sc_idx >= 0 and sc_idx < len(rr):
                        cache = cr[sc_idx] if cr is not None and sc_idx < len(cr) else None
                        check_num(f"Risk_Distribution {grl} student_count", rr[sc_idx], cache,
                                  g_row[gt_headers2.index("student_count")], 5,
                                  f"a={rr[sc_idx]} g={g_row[gt_headers2.index('student_count')]}")
                    if pct_idx >= 0 and pct_idx < len(rr):
                        cache = cr[pct_idx] if cr is not None and pct_idx < len(cr) else None
                        check_num(f"Risk_Distribution {grl} pct", rr[pct_idx], cache,
                                  g_row[gt_headers2.index("pct")], 1.0,
                                  f"a={rr[pct_idx]} g={g_row[gt_headers2.index('pct')]}")

    # ------- At_Risk_Students -------
    ar_idx = next((i for i, s in enumerate(sheets_lower) if "at_risk" in s or "risk_student" in s), None)
    if ar_idx is None:
        check("At_Risk_Students sheet exists", False)
    else:
        check("At_Risk_Students sheet exists", True)
        ws3 = wb_raw[sheets[ar_idx]]
        ws3c = wb_cache[sheets[ar_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        rows3c = list(ws3c.iter_rows(values_only=True))
        pairs3 = _data_pairs(rows3, rows3c)
        check("At_Risk_Students has at least 2 course rows", len(pairs3) >= 2, f"Found {len(pairs3)}")
        if rows3:
            headers3 = [_norm(c) for c in rows3[0]]
            for h in ["course_name", "high_risk_count", "medium_risk_count", "low_risk_count"]:
                check(f"At_Risk_Students has '{h}' header", h in headers3, f"got {headers3}")
            gt_headers3, gt_data3 = _gt_sheet_data(gt_wb_raw, "At_Risk_Students")
            if gt_headers3 is not None and gt_data3 is not None:
                cn3_idx = _idx(headers3, "course_name")

                def _ar_keys(cell_text):
                    """Return the lookup keys (year and/or course id) implied by a course_name cell."""
                    t = str(cell_text or "")
                    keys = []
                    ym = re.search(r"(20\d\d)", t)
                    if ym:
                        keys.append(("year", ym.group(1)))
                    cm = re.search(r"course\s*(\d+)", t, re.I)
                    if cm:
                        keys.append(("cid", _norm_id(cm.group(1))))
                    return keys

                # map GT year -> course id from the GT Course_Overview sheet, so a GT row keyed by
                # year can be matched to an agent row keyed by course id (e.g. 'Course 16').
                gt_co_h, gt_co_d = _gt_sheet_data(gt_wb_raw, "Course_Overview")
                year_to_cid = {"2013": "16", "2014": "17"}
                if (gt_co_h is not None and gt_co_d is not None
                        and "course_name" in gt_co_h and "course_id" in gt_co_h):
                    for gr in gt_co_d:
                        ym = re.search(r"(20\d\d)", str(gr[gt_co_h.index("course_name")] or ""))
                        if ym:
                            year_to_cid[ym.group(1)] = _norm_id(gr[gt_co_h.index("course_id")])

                a_keys = {}
                for i, (rr, _cr) in enumerate(pairs3):
                    if cn3_idx >= 0 and cn3_idx < len(rr):
                        for k in _ar_keys(rr[cn3_idx]):
                            a_keys.setdefault(k, i)
                for g_row in gt_data3:
                    gm = re.search(r"(20\d\d)", str(g_row[gt_headers3.index("course_name")] or ""))
                    gyear = gm.group(1) if gm else "?"
                    wanted = [("year", gyear)]
                    if gyear in year_to_cid:
                        wanted.append(("cid", year_to_cid[gyear]))
                    idx = next((a_keys[k] for k in wanted if k in a_keys), None)
                    if idx is None:
                        check(f"At_Risk_Students has row for {gyear}", False,
                              f"course_names: {[str(r[cn3_idx]) if cn3_idx < len(r) else '' for r, _ in pairs3]}")
                        continue
                    rr, cr = pairs3[idx]
                    for col_h, tol in [("high_risk_count", 0), ("medium_risk_count", 0), ("low_risk_count", 0)]:
                        c_idx = _idx(headers3, col_h)
                        if c_idx >= 0 and c_idx < len(rr):
                            cache = cr[c_idx] if cr is not None and c_idx < len(cr) else None
                            check_num(f"At_Risk_Students {gyear} {col_h}", rr[c_idx], cache,
                                      g_row[gt_headers3.index(col_h)], tol,
                                      f"a={rr[c_idx]} g={g_row[gt_headers3.index(col_h)]}")

    # ------- Intervention_Plan -------
    ip_idx = next((i for i, s in enumerate(sheets_lower) if "intervention" in s), None)
    if ip_idx is None:
        check("Intervention_Plan sheet exists", False)
    else:
        check("Intervention_Plan sheet exists", True)
        ws4 = wb_raw[sheets[ip_idx]]
        ws4c = wb_cache[sheets[ip_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        rows4c = list(ws4c.iter_rows(values_only=True))
        pairs4 = _data_pairs(rows4, rows4c)
        check("Intervention_Plan has at least 3 risk-level rows", len(pairs4) >= 3, f"Found {len(pairs4)}")
        if rows4:
            headers4 = [_norm(c) for c in rows4[0]]
            for h in ["risk_level", "action", "timeline", "responsible"]:
                check(f"Intervention_Plan has '{h}' header", h in headers4, f"got {headers4}")
            rl_idx = _idx(headers4, "risk_level")
            # keywords accept both numeric and spelled-out timeline forms (task/PDF use
            # "within one week", scoring_model.json uses "Within 1 week")
            EXPECTED = {
                "high": (["advis"], ["1 week", "one week"], ["academic advisor", "academic adviser"]),
                "medium": (["tutor"], ["2 week", "two week"], ["tutor"]),
                "low": (["self-paced", "self paced", "selfpaced"],
                        ["1 month", "one month"], ["success"]),
            }
            for level, (kw_action, kw_time, kw_resp) in EXPECTED.items():
                idx = next(
                    (i for i, (rr, _cr) in enumerate(pairs4)
                     if rl_idx >= 0 and rl_idx < len(rr) and _risk_matches(rr[rl_idx], level)),
                    None)
                if idx is None:
                    check(f"Intervention_Plan has '{level}' row", False,
                          f"risk_levels: {[str(r[rl_idx]) if rl_idx < len(r) else '' for r, _ in pairs4]}")
                    continue
                rr, _cr = pairs4[idx]
                row_text = " ".join(str(c or "").lower() for c in rr)
                check(f"Intervention_Plan {level} action mentions '{kw_action[0]}'",
                      any(k in row_text for k in kw_action), f"row: {row_text}")
                check(f"Intervention_Plan {level} timeline mentions '1 week/one week...'",
                      any(k in row_text for k in kw_time), f"row: {row_text}")
                check(f"Intervention_Plan {level} responsible mentions '{kw_resp[0]}'",
                      any(k in row_text for k in kw_resp), f"row: {row_text}")


def _doc_mentions_both_courses(text):
    """True if the doc clearly references both courses (2013 & 2014, or Course 16 & 17).

    Accepts year-based ('Fall 2013'), course-id-based ('Course 16'), or mixed references so a
    correct model that names the courses by id rather than year is not failed.
    """
    has_2013 = "2013" in text
    has_2014 = "2014" in text
    if has_2013 and has_2014:
        return True
    c16 = re.search(r"course\s*16", text) is not None
    c17 = re.search(r"course\s*17", text) is not None
    if c16 and c17:
        return True
    return (has_2013 and c17) or (has_2014 and c16)


def check_word(workspace):
    print("\n=== Check 2: Intervention_Plan.docx ===")
    path = os.path.join(workspace, "Intervention_Plan.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)
    try:
        doc = Document(path)
    except Exception as e:
        check("Word document readable", False, str(e))
        return
    check("Word document readable", True)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    # Title tolerant: the task-required phrase may be wrapped in a longer title, e.g.
    # "Student Retention Risk Analysis and Intervention Plan".
    check("Document title 'Student Retention Intervention Plan' present",
          "student retention" in full_text and "intervention plan" in full_text,
          f"Text begin: {full_text[:200]}")
    check("Document has executive summary section",
          any(k in full_text for k in
              ["executive summary", "summary", "key findings", "overview", "introduction", "findings"]))
    check("Document mentions risk distribution",
          any(k in full_text for k in
              ["risk distribution", "risk breakdown", "risk overview", "risk profile",
               "distribution of risk", "risk composition"]))
    check("Document mentions interventions", "intervention" in full_text)
    for lvl in ["high", "medium", "low"]:
        check(f"Document mentions '{lvl}' risk level", lvl in full_text)
    check("Document mentions both courses", _doc_mentions_both_courses(full_text),
          "No year (2013/2014) or course-id (16/17) reference found")
    check("Document has substantial content (>=300 chars)", len(full_text) > 300, f"Length: {len(full_text)}")


def _extract_db_title(title):
    """Extract text from a notion.database title jsonb column (list or JSON stringified)."""
    if title is None:
        return ""
    if isinstance(title, list):
        return " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
    if isinstance(title, str):
        try:
            parsed = json.loads(title)
        except Exception:
            return title
        if isinstance(parsed, list):
            return " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
        if isinstance(parsed, dict):
            return str(parsed.get("content", "") or parsed.get("plain_text", "") or "")
        return str(parsed)
    return str(title) if title else ""


def _walk_json(o, path, out):
    """Flatten a JSON object into (path, value) leaves; path is the list of keys."""
    if isinstance(o, dict):
        for k, v in o.items():
            _walk_json(v, path + [str(k)], out)
    elif isinstance(o, list):
        for v in o:
            _walk_json(v, path, out)
    else:
        out.append((path, o))


def _notion_metrics(props_text):
    """Extract (avg_scores, pass_rates, student_counts) numeric lists from a Notion page's
    properties JSON.

    Tolerates every shape a db-backed Notion mock may store: {"Average Score": {"number":
    77.14}}, {"average_score": 77.14}, {"Average Score": "77.14"}, nested rich-text, etc.
    Property-key matching is substring based, so name variants ("Avg Score", "Passing Rate",
    "student_count") still classify correctly. The title rich-text (which may contain years)
    is never classified as a metric because its key path has no avg/pass/count keyword.
    """
    avg, pr, cnt = [], [], []
    try:
        obj = json.loads(props_text) if isinstance(props_text, str) else props_text
    except Exception:
        return avg, pr, cnt
    pairs = []
    _walk_json(obj, [], pairs)
    for path, val in pairs:
        f = _to_float(val)
        if f is None:
            continue
        p = " ".join(path).lower()
        if ("averag" in p or "avg" in p) and ("score" in p or "point" in p):
            avg.append(f)
        elif "pass" in p and "rate" in p:
            pr.append(f)
        elif "count" in p or "enrollment" in p:
            cnt.append(f)
    return avg, pr, cnt


def _all_match(a_vals, g_vals, tol):
    """True if every GT value has at least one agent value within tol (order-insensitive)."""
    return all(any(abs(a - g) <= tol for a in a_vals) for g in g_vals)


def _gt_notion_expected(gt_dir):
    """Expected (avg_scores, pass_rates, counts) read from the GT Course_Overview sheet.

    Falls back to the seed-verified constants if the GT workbook is unavailable, so the
    check still works when run without a groundtruth xlsx.
    """
    fallback = ([77.14, 78.72], [88.8, 91.6], [2285, 2369])
    gt_path = os.path.join(gt_dir, "Student_Risk_Analysis.xlsx")
    if not os.path.exists(gt_path):
        return fallback
    try:
        wb_raw, _ = _load_wb(gt_path)
    except Exception:
        return fallback
    if wb_raw is None or "Course_Overview" not in wb_raw.sheetnames:
        return fallback
    ws = wb_raw["Course_Overview"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [_norm(c) for c in rows[0]] if rows else []
    ai = _idx(headers, "avg_score")
    pi = _idx(headers, "pass_rate")
    ci = _idx(headers, "enrollment_count")
    avg, pr, cnt = [], [], []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        if 0 <= ai < len(r):
            f = _to_float(r[ai])
            if f is not None:
                avg.append(f)
        if 0 <= pi < len(r):
            f = _to_float(r[pi])
            if f is not None:
                pr.append(f)
        if 0 <= ci < len(r):
            f = _to_float(r[ci])
            if f is not None:
                cnt.append(f)
    return (avg or fallback[0], pr or fallback[1], cnt or fallback[2])


def check_notion(gt_dir=""):
    print("\n=== Check 3: Notion Student Risk Tracker ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("Student Risk Tracker database exists", False, f"DB connection failed: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        tracker_db_ids = []
        for db_id, title in dbs:
            title_str = _extract_db_title(title)
            if ("student" in title_str.lower() and "risk" in title_str.lower()
                    and "tracker" in title_str.lower()):
                tracker_db_ids.append(db_id)
        check("Student Risk Tracker database exists", len(tracker_db_ids) >= 1,
              f"Databases: {[_extract_db_title(d[1]) for d in dbs]}")

        if tracker_db_ids:
            # Aggregate entries across every matching database so that a swarm that
            # split creation across sub-agents (or duplicated the db) still passes.
            cur.execute("""
                SELECT properties::text FROM notion.pages
                WHERE parent->>'database_id' = ANY(%s) AND archived = false
            """, (tracker_db_ids,))
            rows = cur.fetchall()
            check("Tracker has exactly 2 course entries", len(rows) == 2, f"Found {len(rows)}")
            # Verify the aggregated data for both courses is present, rather than requiring the
            # years '2013'/'2014' as literal substrings. A model may name entries by course id
            # (e.g. 'Course 16'/'Course 17'); what matters is that each course's aggregated
            # average score / pass rate / student count appear in the entries.
            exp_avg, exp_pr, exp_cnt = _gt_notion_expected(gt_dir)
            all_avg, all_pr, all_cnt = [], [], []
            for r in rows:
                a, p, c = _notion_metrics(r[0])
                all_avg += a
                all_pr += p
                all_cnt += c
            check("Notion entries carry both courses' average scores",
                  _all_match(all_avg, exp_avg, 1.0), f"avg values found: {sorted(all_avg)}")
            check("Notion entries carry both courses' pass rates",
                  _all_match(all_pr, exp_pr, 1.0), f"pass rate values found: {sorted(all_pr)}")
            check("Notion entries carry both courses' student counts",
                  _all_match(all_cnt, exp_cnt, 5), f"student count values found: {sorted(all_cnt)}")
    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Check 4: Email to Academic Advisors ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("Email sent to academic_advisors@university.edu", False, f"DB connection failed: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%academic_advisors@university.edu%'
        """)
        rows = cur.fetchall()
        check("Email sent to academic_advisors@university.edu", len(rows) >= 1, "No matching email")
        if rows:
            # Collect every row whose subject matches (exact OR the task-required pieces).
            # Body checks evaluate ALL matching emails so a correct final email is not hidden
            # behind an earlier partial/draft email to the same recipient.
            matching = []
            for subj, to_addr, body in rows:
                subj_l = (subj or "").strip().lower()
                if (subj_l == "student retention risk analysis - action required"
                        or ("student retention risk analysis" in subj_l and "action required" in subj_l)):
                    matching.append((subj, to_addr, body))
            check(
                "Email subject 'Student Retention Risk Analysis - Action Required'",
                len(matching) >= 1,
                f"subjects: {[r[0] for r in rows]}",
            )
            if matching:
                # Allow newlines between 'high' and the total count: a bulleted / multi-line
                # body may separate them. Window widened to 120 chars.
                high_count_re = (
                    r"high[\s\S]{0,120}\b\d+\b"
                    r"|\b\d+\b[\s\S]{0,120}high"
                )
                bodies = [(body or "").lower() for (_, _, body) in matching]
                check("Email body mentions high-risk count (with numeric adjacency)",
                      any(re.search(high_count_re, b) for b in bodies),
                      f"bodies: {[b[:120] for b in bodies]}")
                check("Email body mentions next steps",
                      any(("step" in b or "intervention" in b) for b in bodies),
                      f"bodies: {[b[:120] for b in bodies]}")
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: risk_scorer.py ===")
    path = os.path.join(workspace, "risk_scorer.py")
    check("risk_scorer.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("No risk emails sent to wrong recipients", False, f"DB connection failed: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT to_addr FROM email.messages
            WHERE subject ILIKE '%risk%' OR subject ILIKE '%retention%'
        """)
        emails = cur.fetchall()
        noise_recipients = ["all-staff@university.edu", "it@university.edu",
                            "facilities@university.edu"]
        for email_row in emails:
            to_str = str(email_row[0]).lower()
            for noise in noise_recipients:
                if noise in to_str:
                    check("No risk emails sent to wrong recipients", False,
                          f"Sent to noise recipient: {noise}")
                    return
        check("No risk emails sent to wrong recipients", True)
    except Exception as e:
        check("Reverse validation", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_word(args.agent_workspace)
    check_notion(args.groundtruth_workspace)
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed")

    result = {"total_passed": PASS_COUNT, "total_checks": total}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Strict: all checks must pass (FAIL_COUNT == 0)
    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL ({FAIL_COUNT} failures)")
        sys.exit(1)


if __name__ == "__main__":
    main()
