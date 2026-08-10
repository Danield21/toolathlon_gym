"""Evaluation script for fetch-howtocook-catering-excel-gcal-email."""
import os
import argparse
import sys
from datetime import datetime, timezone
import openpyxl


def _to_float(val):
    """Robustly convert a cell value to a float.

    Handles numeric types, numeric strings with thousands separators / currency
    symbols / trailing '%' / whitespace. Excel formula strings (start with '=')
    and unparseable values return None.
    """
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.startswith("="):
        return None  # Excel formula string -- cannot evaluate here
    s = s.replace(",", "").replace("$", "").replace("¥", "").replace("€", "")
    s = s.replace("%", "").replace(" ", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    fa, fb = _to_float(a), _to_float(b)
    if fa is None or fb is None:
        return str(a).strip().lower() == str(b).strip().lower()
    return abs(fa - fb) <= max(abs_tol, abs(fb) * rel_tol)


# R1: read DB config from environment with defaults (mirrors preprocess/main.py).
# The MCP servers the agent writes through (emails-mcp, calendar) read the
# underscore names PG_DATABASE/PG_USER/PG_PASSWORD; the swarm runtime bridges
# PGDATABASE/PGUSER/PGPASSWORD onto them (tool_servers.py _pg_bridge) and some
# task-level MCP configs set the underscore names directly. Prefer the
# underscore forms first so the evaluator always queries the exact
# database/role the MCP servers write to, falling back to the standard names
# and then to the deployed defaults.
DB_CONFIG = {
    "host": os.environ.get("PG_HOST") or os.environ.get("PGHOST") or "localhost",
    "port": int(os.environ.get("PG_PORT") or os.environ.get("PGPORT") or "5432"),
    "dbname": os.environ.get("PG_DATABASE") or os.environ.get("PGDATABASE") or "toolathlon_gym",
    "user": os.environ.get("PG_USER") or os.environ.get("PGUSER") or "eigent",
    "password": os.environ.get("PG_PASSWORD") or os.environ.get("PGPASSWORD") or "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _norm(s):
    """Normalize a string for case/whitespace/separator-insensitive comparison."""
    if s is None:
        return ""
    return str(s).strip().lower().replace("_", " ")


def _find_sheet(wb, *keywords):
    """Locate a sheet by exact (case-insensitive) name, else keyword substring."""
    exact = {_norm(n): n for n in wb.sheetnames}
    target = _norm(keywords[0])
    if target in exact:
        return wb[exact[target]]
    for name in wb.sheetnames:
        nl = _norm(name)
        if all(_norm(k) in nl for k in keywords):
            return wb[name]
    return None


def _headers(ws):
    return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]


def _find_col(headers, name):
    hl = [_norm(h) for h in headers]
    target = _norm(name)
    return hl.index(target) if target in hl else -1


def _data_rows(ws):
    """Data rows excluding fully-empty rows (robust to trailing/blank rows)."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return [r for r in rows if any(v is not None and str(v).strip() != "" for v in r)]


def _dedup(rows):
    """Drop exact-duplicate rows (robust to multi-agent duplicate writes)."""
    seen = set()
    out = []
    for r in rows:
        key = tuple(r)
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


def _parse_dt(val):
    """Parse a cell/DB datetime into a datetime, tolerating ISO variants."""
    if isinstance(val, datetime):
        return val
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    if "T" not in s and " " in s:
        s = s.replace(" ", "T", 1)
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


def _hour_min_ok(dt, hour, minute):
    """True if dt represents the given hour:minute on its own (session) wall
    clock OR in UTC. psycopg2 returns timestamptz values rendered in the DB
    session timezone, so a naive timestamp written by the agent (per the task
    instructions) reads back with the same wall clock, while an explicit
    '...Z' UTC timestamp reads back shifted by the session offset. Accepting
    either the session-local or the UTC reading means both a literal naive
    input and a correct UTC (Z) input pass regardless of the DB session
    timezone. Naive datetimes are interpreted as UTC.
    """
    if dt is None:
        return False
    local_ok = dt.hour == hour and dt.minute == minute
    utc = dt if dt.tzinfo is None else dt.astimezone(timezone.utc)
    utc_ok = utc.hour == hour and utc.minute == minute
    return local_ok or utc_ok


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # --- Check Excel file ---
    excel_path = os.path.join(agent_workspace, "Wellness_Menu_Plan.xlsx")
    check("Wellness_Menu_Plan.xlsx exists", os.path.exists(excel_path))

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # --- Sheet 1: Daily Menu ---
        ws = _find_sheet(wb, "Daily Menu")
        check("Daily Menu sheet exists", ws is not None)
        if ws is not None:
            headers = _headers(ws)
            data_rows = _dedup(_data_rows(ws))

            check("Daily Menu has 10 rows", len(data_rows) == 10,
                  f"got {len(data_rows)}")

            for col in ["Day", "Meal_Type", "Recipe_Name", "Servings", "Ingredient_Count",
                        "Estimated_Cost", "Is_Vegetarian"]:
                check(f"Daily Menu has {col} column",
                      _find_col(headers, col) >= 0, f"headers: {headers}")

            # Verify all 5 weekdays appear, each exactly twice, in Mon-Fri order
            # (two meals per day). Accepts day-major or meal-major row layout.
            day_col = _find_col(headers, "Day")
            if day_col >= 0 and len(data_rows) >= 10:
                days = [_norm(r[day_col]) for r in data_rows if r[day_col]]
                first_seen = []
                for d in days:
                    if d not in first_seen:
                        first_seen.append(d)
                expected_first = ["monday", "tuesday", "wednesday", "thursday", "friday"]
                from collections import Counter
                check("Days sorted Monday-Friday (2 meals per day)",
                      first_seen == expected_first and
                      Counter(days) == Counter(["monday"] * 2 + ["tuesday"] * 2 +
                                               ["wednesday"] * 2 + ["thursday"] * 2 +
                                               ["friday"] * 2),
                      f"got {days}")

            # Verify servings are all 50
            serv_col = _find_col(headers, "Servings")
            if serv_col >= 0:
                servings = [_to_float(r[serv_col]) for r in data_rows]
                check("All servings are 50",
                      all(s == 50 for s in servings if s is not None),
                      f"servings: {servings}")

            # Verify Is_Vegetarian has at least 2 Yes values
            veg_col = _find_col(headers, "Is_Vegetarian")
            if veg_col >= 0:
                veg_vals = [str(r[veg_col]).strip() for r in data_rows if r[veg_col]]
                yes_count = sum(1 for v in veg_vals if _norm(v) == "yes")
                check("At least 2 vegetarian options", yes_count >= 2,
                      f"found {yes_count} vegetarian")

            # Verify cost formula: Ingredient_Count * 0.50 * 50
            ic_col = _find_col(headers, "Ingredient_Count")
            ec_col = _find_col(headers, "Estimated_Cost")
            if ic_col >= 0 and ec_col >= 0:
                cost_ok = True
                for row in data_rows:
                    ic = _to_float(row[ic_col])
                    ec = _to_float(row[ec_col])
                    if ic is not None and ec is not None:
                        expected = round(ic * 0.50 * 50, 2)
                        if abs(ec - expected) > 0.01:
                            cost_ok = False
                            break
                check("Estimated_Cost = Ingredient_Count * 0.50 * 50", cost_ok)

        # --- Sheet 2: Ingredient Summary ---
        ws = _find_sheet(wb, "Ingredient Summary")
        check("Ingredient Summary sheet exists", ws is not None)
        if ws is not None:
            headers = _headers(ws)
            data_rows = _dedup(_data_rows(ws))

            check("Ingredient Summary has >= 20 rows", len(data_rows) >= 20,
                  f"got {len(data_rows)}")

            for col in ["Ingredient_Name", "Total_Quantity", "Unit", "Times_Used", "Total_Cost"]:
                check(f"Ingredient Summary has {col} column",
                      _find_col(headers, col) >= 0, f"headers: {headers}")

            # Verify sorted alphabetically. Ingredient names are Chinese
            # (HowToCook recipes), so 'alphabetical' can legitimately mean
            # either code-point order (Python sorted()) or pinyin order.
            # Accept code-point / casefold order, and pinyin order when
            # pypinyin happens to be available; otherwise fall back to the
            # code-point order the ground truth itself uses.
            name_col = _find_col(headers, "Ingredient_Name")
            if name_col >= 0:
                names = [str(r[name_col]).strip() for r in data_rows
                         if r[name_col] and str(r[name_col]).strip()]
                def _is_sorted(lst, key=None):
                    return lst == sorted(lst, key=key)
                unicode_ok = _is_sorted(names) or _is_sorted(names, key=str.casefold)
                pinyin_ok = False
                try:
                    from pypinyin import lazy_pinyin
                    pinyin_ok = _is_sorted(names, key=lazy_pinyin)
                except Exception:
                    pass
                check("Ingredients sorted alphabetically",
                      unicode_ok or pinyin_ok, f"first few: {names[:5]}")

            # Verify cost formula: Times_Used * 0.50 * 50
            tu_col = _find_col(headers, "Times_Used")
            tc_col = _find_col(headers, "Total_Cost")
            if tu_col >= 0 and tc_col >= 0:
                cost_ok = True
                for row in data_rows:
                    tu = _to_float(row[tu_col])
                    tc = _to_float(row[tc_col])
                    if tu is not None and tc is not None:
                        expected = round(tu * 0.50 * 50, 2)
                        if abs(tc - expected) > 0.01:
                            cost_ok = False
                            break
                check("Ingredient Total_Cost = Times_Used * 0.50 * 50", cost_ok)

        # --- Sheet 3: Budget Overview ---
        ws = _find_sheet(wb, "Budget Overview")
        check("Budget Overview sheet exists", ws is not None)
        if ws is not None:
            headers = _headers(ws)
            data_rows = _dedup(_data_rows(ws))

            check("Budget Overview has 6 rows", len(data_rows) == 6,
                  f"got {len(data_rows)}")

            for col in ["Label", "Value"]:
                check(f"Budget Overview has {col} column",
                      _find_col(headers, col) >= 0, f"headers: {headers}")

            # Build label->value map (normalized labels)
            label_col = _find_col(headers, "Label")
            value_col = _find_col(headers, "Value")
            budget_map = {}
            if label_col >= 0 and value_col >= 0:
                for row in data_rows:
                    if row[label_col]:
                        budget_map[_norm(row[label_col])] = _to_float(row[value_col])

            check("Has Total_Budget label", "total budget" in budget_map)
            tb = budget_map.get("total budget")
            if tb is not None:
                check("Total_Budget is 2000", tb == 2000.0, f"got {tb}")

            # Verify budget consistency
            tc = budget_map.get("total estimated cost")
            br = budget_map.get("budget remaining")
            if tc is not None and br is not None:
                check("Budget_Remaining = 2000 - Total_Estimated_Cost",
                      abs(br - (2000.0 - tc)) < 0.01,
                      f"remaining={br}, expected={2000.0 - tc}")

            acd = budget_map.get("avg cost per day")
            if tc is not None and acd is not None:
                check("Avg_Cost_Per_Day = Total / 5",
                      abs(acd - tc / 5) < 0.01,
                      f"avg={acd}, expected={tc / 5}")

            acm = budget_map.get("avg cost per meal")
            if tc is not None and acm is not None:
                check("Avg_Cost_Per_Meal = Total / 10",
                      abs(acm - tc / 10) < 0.01,
                      f"avg={acm}, expected={tc / 10}")

            # Verify Days_Over_Budget threshold logic ($400) -- only when the
            # daily Estimated_Cost cells are parseable (skip formulas/None)
            dob = budget_map.get("days over budget")
            daily_ws = _find_sheet(wb, "Daily Menu")
            if dob is not None and daily_ws is not None:
                daily_headers = _headers(daily_ws)
                day_col = _find_col(daily_headers, "Day")
                ec_col = _find_col(daily_headers, "Estimated_Cost")
                day_totals = {}
                if day_col >= 0 and ec_col >= 0:
                    for r in _dedup(_data_rows(daily_ws)):
                        d = str(r[day_col]).strip() if r[day_col] else None
                        ec = _to_float(r[ec_col])
                        if d and ec is not None:
                            day_totals[d] = day_totals.get(d, 0) + ec
                if day_totals:
                    expected_dob = sum(1 for v in day_totals.values() if v > 400)
                    check("Days_Over_Budget matches $400 threshold",
                          abs(dob - expected_dob) < 0.01,
                          f"got {dob}, expected {expected_dob}")

    # --- Check Calendar Events ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE summary ILIKE '%meal prep%'
            ORDER BY start_datetime
        """)
        cal_rows = cur.fetchall()
        cur.close()
        conn.close()

        check("At least 5 meal prep calendar events", len(cal_rows) >= 5,
              f"found {len(cal_rows)}")

        # Dates must cover exactly March 16-20 (set-equality tolerates duplicate
        # events created by mis-delegating sub-agents)
        dates = []
        for r in cal_rows:
            dt = _parse_dt(r[1])
            dates.append(dt.date().isoformat() if dt else str(r[1])[:10])
        expected_dates = ["2026-03-16", "2026-03-17", "2026-03-18",
                          "2026-03-19", "2026-03-20"]
        check("Calendar events on March 16-20",
              set(dates) == set(expected_dates), f"dates: {dates}")

        # Times: each event starts 07:00 and ends 08:00. Accept the session
        # wall-clock OR the UTC reading so both literal naive inputs (task's
        # stated format) and explicit Z/offset UTC inputs pass in any DB
        # session timezone (see _hour_min_ok).
        for row in cal_rows:
            st = _parse_dt(row[1])
            en = _parse_dt(row[2])
            if st is not None:
                check(f"Event {row[1]} starts at 07:00",
                      _hour_min_ok(st, 7, 0), f"got {st.isoformat()}")
            if en is not None:
                check(f"Event {row[1]} ends at 08:00",
                      _hour_min_ok(en, 8, 0), f"got {en.isoformat()}")
    except Exception as e:
        check("Calendar events query", False, str(e))

    # --- Check Emails ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr, subject, body_text
            FROM email.messages
            WHERE subject ILIKE '%wellness%'
            ORDER BY subject
        """)
        email_rows = cur.fetchall()
        cur.close()
        conn.close()

        check("At least 2 wellness emails sent", len(email_rows) >= 2,
              f"found {len(email_rows)}")

        # to_addr is jsonb, convert to string for matching
        recipients = [str(r[0]) for r in email_rows]
        subjects = [r[1] for r in email_rows]

        # Check vendor email
        vendor_found = any("catering_vendor" in str(r) for r in recipients)
        check("Email sent to catering_vendor@company.com", vendor_found,
              f"recipients: {recipients}")

        # Check committee email
        committee_found = any("wellness_committee" in str(r) for r in recipients)
        check("Email sent to wellness_committee@company.com", committee_found,
              f"recipients: {recipients}")

        # Check subjects
        ingredient_subj = any("ingredient" in _norm(s) for s in subjects)
        check("Vendor email has ingredient-related subject", ingredient_subj,
              f"subjects: {subjects}")

        menu_subj = any("menu" in _norm(s) for s in subjects)
        check("Committee email has menu-related subject", menu_subj,
              f"subjects: {subjects}")

    except Exception as e:
        check("Email query", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
