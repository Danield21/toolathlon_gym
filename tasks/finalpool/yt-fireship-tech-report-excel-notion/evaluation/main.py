"""
Evaluation for yt-fireship-tech-report-excel-notion task.

Checks:
1. Tech_Trend_Report.xlsx exists with "Videos" sheet having >= 8 data rows
2. "Videos" sheet has Video_ID, Title, View_Count columns (case-insensitive)
3. "Topic_Summary" sheet exists with >= 3 topic rows
4. Topic_Summary has Topic and Video_Count columns
5. Notion page exists with title containing "Fireship" or "Tech Trends"
6. GCal has a new event in March 2026 with "Tech Review" in summary (not the noise Team Sync)
7. Email was sent to techteam@company.com
"""
import os
import sys
import json
from argparse import ArgumentParser

import psycopg2
import openpyxl
from zoneinfo import ZoneInfo

# ──────────────────────────────────────────────────────────────────────────
# EVALUATION GROUND TRUTH SPEC (gcal tz root-fix v3, case-study 2026-08-13)
# Source of truth: docs/task.md. The Calendar MCP writes events as UTC
# instants into the DB (`TIMESTAMPTZ`), so all comparisons anchor to the
# timezone that task.md declares for the event's wall-clock semantics.
# Never use bare `sd.hour` / `sd.date()` on rows read from `gcal.events`:
# psycopg2 returns them in the PG session `TimeZone` (case-study: compute
# node default was Asia/Shanghai, masking 14:00 UTC as 22:00+08). Use
# `utils.evaluation.gcal_helpers` instead (session-tz-independent).
# ──────────────────────────────────────────────────────────────────────────
# task.md line 11: "Monday March 9 2026 from 14:00 to 15:00. Use the UTC
# timezone."
EXPECTED_TIMEZONE = ZoneInfo("UTC")
EXPECTED_MEETING_HOUR = 14
EXPECTED_MEETING_MINUTE = 0
EXPECTED_MEETING_DURATION_MIN = 60

from utils.evaluation.gcal_helpers import get_utc_components  # noqa: E402

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-4: Tech_Trend_Report.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Tech_Trend_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Trend_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Videos sheet has >= 8 data rows", False, "File missing")
        record("Videos sheet has required columns", False, "File missing")
        record("Topic_Summary sheet exists with >= 3 rows", False, "File missing")
        record("Topic_Summary has Topic and Video_Count columns", False, "File missing")
        return
    record("Tech_Trend_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Check Videos sheet
    videos_key = None
    for k in sheet_names_lower:
        if "video" in k and "summary" not in k and "topic" not in k:
            videos_key = sheet_names_lower[k]
            break
    if not videos_key:
        record("Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Videos sheet has 10 data rows", False, "Sheet missing")
        record("Videos sheet has required columns", False, "Sheet missing")
    else:
        ws = wb[videos_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        # Strict 10 rows (was >= 8)
        record("Videos sheet has 10 data rows", len(data_rows) == 10,
               f"Found {len(data_rows)} data rows")
        if rows:
            # Stricter: check exact header presence
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            for required in ["video_id", "title", "published_date", "duration_seconds",
                             "view_count", "like_count", "primary_topic"]:
                # Tolerate underscore/space
                norm_required = required.replace("_", "")
                hit = any(h.replace("_", "").replace(" ", "") == norm_required for h in headers)
                record(f"Videos sheet has '{required}' column", hit,
                       f"Headers: {headers}")
        else:
            record("Videos sheet has required columns", False, "Sheet is empty")

    # Check Topic_Summary sheet
    topic_key = None
    for k in sheet_names_lower:
        if "topic" in k:
            topic_key = sheet_names_lower[k]
            break
    if not topic_key:
        record("Topic_Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Topic_Summary has required columns", False, "Sheet missing")
    else:
        ws2 = wb[topic_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        record("Topic_Summary sheet has at least 3 rows", len(data_rows2) >= 3,
               f"Found {len(data_rows2)} data rows")
        if rows2:
            headers2 = [str(c).strip().lower() if c else "" for c in rows2[0]]
            for required in ["topic", "video_count", "total_views", "avg_duration"]:
                norm_required = required.replace("_", "")
                hit = any(h.replace("_", "").replace(" ", "") == norm_required for h in headers2)
                record(f"Topic_Summary has '{required}' column", hit,
                       f"Headers: {headers2}")
        else:
            record("Topic_Summary has required columns", False, "Sheet is empty")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Tech_Trend_Report.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            # For Videos: validate by Video_ID match (all 10 should match GT video IDs)
            if gt_sname.lower() == "videos":
                # Build GT video ID set
                gt_vids = {str(r[0]).strip() for r in gt_rows if r and r[0]}
                a_vids = {str(r[0]).strip() for r in a_rows if r and r[0]}
                overlap = len(gt_vids & a_vids)
                record(f"Videos sheet has all {len(gt_vids)} expected video IDs",
                       overlap == len(gt_vids), f"matched {overlap}/{len(gt_vids)}")
                # For each matched video ID, validate view_count
                gt_lookup = {str(r[0]).strip(): r for r in gt_rows if r and r[0]}
                a_lookup = {str(r[0]).strip(): r for r in a_rows if r and r[0]}
                vc_match = 0
                for vid in (gt_vids & a_vids):
                    gtr = gt_lookup[vid]
                    ar = a_lookup[vid]
                    if len(gtr) > 4 and len(ar) > 4:
                        try:
                            if int(gtr[4]) == int(ar[4]):
                                vc_match += 1
                        except (TypeError, ValueError):
                            pass
                record(f"View counts match for all matched videos",
                       vc_match == overlap and overlap > 0, f"{vc_match}/{overlap}")
            elif gt_sname.lower() == "topic_summary":
                # Topic labeling is a free-form clustering: task.md says the
                # Primary_Topic is "derived from the video tags or title
                # keywords", and youtube.videos carries no authoritative
                # topic column. Locking GT's human-chosen labels ("AI/ML",
                # "Big Tech") over-penalizes equally-defensible taxonomies
                # ("Artificial Intelligence", "Quantum Computing" for the
                # same 2329914-view video). Instead check SELF-CONSISTENCY:
                # the summary must partition the agent's own Videos sheet —
                # every video covered exactly once, and per-topic counts
                # matching the Videos rows (case-study 2026-08-16 round-3).
                vids_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == "videos":
                        vids_ws = wb[asn]
                        break
                if vids_ws is None:
                    record("Topic_Summary self-consistent with Videos", False,
                           "Videos sheet missing")
                else:
                    v_rows = [r for r in vids_ws.iter_rows(min_row=2, values_only=True)
                              if any(c is not None for c in r)]
                    try:
                        v_hdrs = [str(c).strip().lower() if c else "" for c in
                                  next(vids_ws.iter_rows(max_row=1, values_only=True))]
                    except StopIteration:
                        v_hdrs = []
                    topic_col = next((i for i, h in enumerate(v_hdrs)
                                      if h.replace("_", "").replace(" ", "") == "primarytopic"), 6)
                    vid_topic_pairs = [(str(r[0]).strip().lower(),
                                        str(r[topic_col]).strip().lower() if topic_col < len(r) and r[topic_col] else "")
                                       for r in v_rows if r and r[0]]
                    from collections import Counter
                    expect_counts = Counter(t for _, t in vid_topic_pairs if t)
                    a_topics = {str(r[0]).strip().lower(): r for r in a_rows if r and r[0]}
                    counts_ok = all(
                        a_topics.get(t) and len(a_topics[t]) > 1 and
                        int(a_topics[t][1]) == expect_counts.get(t, -1)
                        for t in expect_counts
                    )
                    covered = sum(expect_counts.values())
                    record("Topic_Summary covers all agent topics with correct counts",
                           counts_ok and covered == len(vid_topic_pairs),
                           f"expected {dict(expect_counts)}; agent topics: {list(a_topics.keys())}; "
                           f"videos covered {covered}/{len(vid_topic_pairs)}")
        gt_wb.close()


def check_notion():
    print("\n=== Check 5: Notion page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE properties::text ILIKE '%Fireship Tech Trends Q1 2026%'
            """)
            rows = cur.fetchall()
            target_id = rows[0][0] if rows else None
        record("Notion page titled 'Fireship Tech Trends Q1 2026' exists",
               len(rows) > 0,
               f"Found {len(rows)} matching pages")
        if target_id:
            with conn.cursor() as cur:
                # notion.blocks stores the payload in block_data (JSONB); there
                # is no `content` column (c4 case-study 2026-08-15: the old
                # query failed with "column content does not exist").
                cur.execute(
                    "SELECT block_data FROM notion.blocks WHERE parent_id = %s",
                    (target_id,))
                blocks = cur.fetchall()
            # Extract readable text from the block JSON: rich_text arrays in
            # paragraphs / headings / table cells.
            def _block_text(bd):
                if not isinstance(bd, dict):
                    return ""
                parts = []

                def _extract(value):
                    if isinstance(value, list):
                        for item in value:
                            if isinstance(item, dict):
                                if "text" in item:
                                    t = item["text"]
                                    parts.append(t.get("content", "") if isinstance(t, dict) else str(t))
                                elif "plain_text" in item:
                                    parts.append(item.get("plain_text", ""))
                                else:
                                    _extract(item)
                            elif isinstance(item, str):
                                parts.append(item)
                    elif isinstance(value, dict):
                        # Notion block_data nests: {heading_2: {rich_text: [...]}}
                        for sub in value.values():
                            _extract(sub)

                for v in bd.values():
                    _extract(v)
                return " ".join(p for p in parts if p)
            all_text = " ".join(_block_text(b[0]) for b in blocks).lower()
            # Heading + summary paragraph + top 5 topics table
            record("Notion page has substantive content (>200 chars blocks)",
                   len(all_text) > 200,
                   f"Got {len(all_text)} chars across {len(blocks)} blocks")
        conn.close()
    except Exception as e:
        record("Notion page check", False, str(e))


def check_gcal():
    print("\n=== Check 6: GCal Tech Review event ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, summary, start_datetime, end_datetime, recurrence
                FROM gcal.events
                WHERE summary ILIKE '%Tech Review Meeting%'
                  AND start_datetime >= TIMESTAMPTZ '2026-03-01T00:00:00Z'
                  AND start_datetime <  TIMESTAMPTZ '2026-04-01T00:00:00Z'
            """)
            rows = cur.fetchall()
        conn.close()
        record("GCal has 'Tech Review Meeting' event in March 2026",
               len(rows) > 0,
               f"Found {len(rows)} matching events")
        if rows:
            ev_id, summary, start_dt, end_dt, recurrence = rows[0]
            # Date 2026-03-09. Extract via UTC components — task.md declares
            # the event in UTC, so bare start_dt.date()/str(start_dt)[:10]
            # would silently shift a day in a non-UTC PG session
            # (case-study 2026-08-13).
            ev_date, ev_hour, ev_minute = get_utc_components(start_dt)
            if ev_date is not None:
                date_str = ev_date.isoformat()
                record("Tech Review Meeting on 2026-03-09",
                       date_str == "2026-03-09",
                       f"got {date_str}")
                # 14:00 UTC start (task.md: "from 14:00 to 15:00. Use the UTC timezone").
                record("Tech Review Meeting starts at 14:00",
                       ev_hour == EXPECTED_MEETING_HOUR and ev_minute == EXPECTED_MEETING_MINUTE,
                       f"got {ev_hour:02d}:{ev_minute:02d}" if ev_hour is not None else "unparseable")
            else:
                record("Tech Review Meeting on 2026-03-09", False, "start_datetime missing")
                record("Tech Review Meeting starts at 14:00", False, "start_datetime missing")
            # Duration 1 hour (timedelta on UTC instants is tz-safe).
            if end_dt and start_dt:
                duration = (end_dt - start_dt).total_seconds() / 60
                record("Tech Review Meeting is 1 hour",
                       EXPECTED_MEETING_DURATION_MIN - 5 <= duration <= EXPECTED_MEETING_DURATION_MIN + 5,
                       f"duration {duration:.0f} min")
            # Recurring (weekly) - check recurrence field exists
            rec_str = str(recurrence or "").upper()
            record("Tech Review Meeting is recurring (weekly)",
                   "WEEKLY" in rec_str or "RRULE" in rec_str,
                   f"recurrence: {recurrence}")
    except Exception as e:
        record("GCal check", False, str(e))


def check_email():
    print("\n=== Check 7: Email sent to techteam@company.com ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, subject, to_addr, body_text FROM email.messages
                WHERE to_addr::text ILIKE '%techteam@company.com%'
            """)
            rows = cur.fetchall()
        conn.close()
        record("Email sent to techteam@company.com",
               len(rows) > 0,
               f"Found {len(rows)} matching emails")
        if rows:
            # Find email with subject referencing tech trend report
            target = None
            for r in rows:
                subj = (r[1] or "").lower()
                if "tech trend" in subj or "tech report" in subj or "fireship" in subj:
                    target = r
                    break
            if target is None:
                target = rows[0]
            subject = (target[1] or "")
            body = (target[3] or "").lower()
            record("Email subject references tech trend/report",
                   "tech" in subject.lower() and ("trend" in subject.lower() or "report" in subject.lower()),
                   f"Subject: {subject}")
            # Body must mention topics, meeting, and Tech_Trend_Report.xlsx
            record("Email body mentions topics",
                   "topic" in body or any(t in body for t in ["javascript", "python", "ai/ml", "ai", "ml"]),
                   f"body sample: {body[:200]}")
            record("Email body mentions meeting",
                   "meeting" in body,
                   f"body sample: {body[:200]}")
            record("Email body mentions Tech_Trend_Report.xlsx",
                   "tech_trend_report.xlsx" in body or "tech_trend_report" in body,
                   f"body sample: {body[:200]}")
    except Exception as e:
        record("Email check", False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print(f"Running evaluation for yt-fireship-tech-report-excel-notion")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_notion()
    check_gcal()
    check_email()

    all_passed = FAIL_COUNT == 0
    summary = f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"
    print(f"\n{'='*40}")
    print(f"Result: {'PASS' if all_passed else 'FAIL'} - {summary}")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "all_passed": all_passed}, f)

    return all_passed, summary


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
