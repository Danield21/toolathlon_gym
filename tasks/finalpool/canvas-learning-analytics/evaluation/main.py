"""Evaluation for canvas-learning-analytics."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_kpis():
    """Get expected KPIs from DB."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except psycopg2.OperationalError:
        return {}
    cur = conn.cursor()
    results = {}
    cur.execute("SELECT id, name FROM canvas.courses ORDER BY name")
    courses = cur.fetchall()

    for cid, cname in courses:
        # Completion rate
        cur.execute("""
            SELECT COUNT(*) FILTER (WHERE s.workflow_state = 'graded'),
                   COUNT(*)
            FROM canvas.submissions s
            JOIN canvas.assignments a ON s.assignment_id = a.id
            WHERE a.course_id = %s
        """, (cid,))
        graded, total = cur.fetchone()
        comp_rate = round(graded / total * 100, 1) if total > 0 else 0

        # Avg grade
        cur.execute("""
            SELECT ROUND(AVG((grades->>'current_score')::numeric), 1)
            FROM canvas.enrollments
            WHERE course_id = %s AND type = 'StudentEnrollment'
              AND grades->>'current_score' IS NOT NULL
        """, (cid,))
        avg_grade = float(cur.fetchone()[0] or 0)

        # Late rate
        cur.execute("""
            SELECT SUM(CASE WHEN s.late THEN 1 ELSE 0 END)::numeric / NULLIF(COUNT(s.id), 0) * 100
            FROM canvas.submissions s
            JOIN canvas.assignments a ON s.assignment_id = a.id
            WHERE a.course_id = %s
        """, (cid,))
        late_rate = round(float(cur.fetchone()[0] or 0), 1)

        # Discussion count
        cur.execute("""
            SELECT COUNT(*) FROM canvas.discussion_topics
            WHERE course_id = %s
        """, (cid,))
        try:
            disc_count = int(cur.fetchone()[0] or 0)
        except Exception:
            disc_count = 0

        meets = comp_rate >= 75 and avg_grade >= 70 and late_rate <= 15
        results[cname.lower()] = {
            "completion_rate": comp_rate,
            "avg_grade": avg_grade,
            "late_rate": late_rate,
            "discussion_count": disc_count,
            "meets": meets,
        }

    cur.close()
    conn.close()
    return results


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Learning_Analytics.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Learning_Analytics.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Learning_Analytics.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_kpis()
    n_courses = len(expected) if expected else 22
    meets_count = sum(1 for v in expected.values() if v["meets"])

    # Course KPIs sheet
    kpi_rows = load_sheet_rows(wb, "Course KPIs")
    if kpi_rows is None:
        check("Sheet 'Course KPIs' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Course KPIs' exists", True)
        data_rows = kpi_rows[1:] if len(kpi_rows) > 1 else []
        # Filter empty rows
        data_rows = [r for r in data_rows if r and r[0] is not None]
        check(f"Course KPIs has {n_courses} data rows",
              len(data_rows) == n_courses,
              f"Found {len(data_rows)}, expected {n_courses}")

        header = kpi_rows[0] if kpi_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        # Locate column indices
        def _col_idx(*aliases):
            for i, h in enumerate(header_lower):
                for a in aliases:
                    if a in h:
                        return i
            return None
        course_col = _col_idx("course") if "course" in (header_lower[0] or "") else 0
        # Use first column for course name regardless
        course_col = 0
        comp_col = _col_idx("completion_rate", "completion")
        grade_col = _col_idx("avg_grade", "avg")
        late_col = _col_idx("late_rate", "late")
        disc_col = _col_idx("discussion")
        meets_col = _col_idx("meets")

        for col, idx in [("course", course_col), ("completion_rate", comp_col),
                         ("avg_grade", grade_col), ("late_rate", late_col),
                         ("discussion_count", disc_col),
                         ("meets_benchmark", meets_col)]:
            check(f"Column '{col}' present", idx is not None, f"Header: {header}")

        # Validate all courses' KPI values
        if expected:
            row_by_name = {}
            for row in data_rows:
                if row and row[course_col]:
                    row_by_name[str(row[course_col]).strip().lower()] = row
            mismatches = 0
            for cname_lower, exp in expected.items():
                row = row_by_name.get(cname_lower)
                if row is None:
                    check(f"Course KPIs row for '{cname_lower}'", False,
                          "course missing in sheet")
                    mismatches += 1
                    continue
                if comp_col is not None:
                    if not num_close(row[comp_col], exp["completion_rate"], 1.0):
                        mismatches += 1
                if grade_col is not None:
                    if not num_close(row[grade_col], exp["avg_grade"], 1.0):
                        mismatches += 1
                if late_col is not None:
                    if not num_close(row[late_col], exp["late_rate"], 1.0):
                        mismatches += 1
                if meets_col is not None:
                    val = str(row[meets_col] or "").strip().lower()
                    expected_meets = "yes" if exp["meets"] else "no"
                    if val not in (expected_meets, expected_meets[0],
                                   "true" if exp["meets"] else "false"):
                        mismatches += 1
            check("All courses' KPI values match DB (tol 1.0)",
                  mismatches == 0,
                  f"{mismatches} value mismatches across {n_courses} courses")

    # Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        check(f"Total_Courses = {n_courses}",
              num_close(lookup.get("total_courses"), n_courses, 0),
              f"Got {lookup.get('total_courses')}")
        check(f"Meets_All_Benchmarks = {meets_count} (exact)",
              num_close(lookup.get("meets_all_benchmarks"), meets_count, 0),
              f"Got {lookup.get('meets_all_benchmarks')}, expected {meets_count}")
        check(f"Below_Benchmark = {n_courses - meets_count} (exact)",
              num_close(lookup.get("below_benchmark"), n_courses - meets_count, 0),
              f"Got {lookup.get('below_benchmark')}")
        if n_courses > 0:
            expected_rate = round(meets_count / n_courses * 100, 1)
            check(f"Benchmark_Rate ~ {expected_rate}%",
                  num_close(lookup.get("benchmark_rate"), expected_rate, 1.0),
                  f"Got {lookup.get('benchmark_rate')}")

    # Trends sheet
    tr_rows = load_sheet_rows(wb, "Trends")
    if tr_rows is None:
        check("Sheet 'Trends' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Trends' exists", True)
        data_rows = tr_rows[1:] if len(tr_rows) > 1 else []
        data_rows = [r for r in data_rows if r and r[0] is not None]
        # Should have N courses * 3 categories = 3N rows (exact)
        check(f"Trends has {n_courses * 3} rows (exact)",
              len(data_rows) == n_courses * 3,
              f"Found {len(data_rows)}, expected {n_courses * 3}")

        # Validate Performance label per row matches DB-computed label
        if expected and len(data_rows) > 0:
            t_header = tr_rows[0] if tr_rows else []
            t_header_lower = [str(h).lower() if h else "" for h in t_header]
            cat_col = next((i for i, h in enumerate(t_header_lower)
                            if "category" in h), None)
            perf_col = next((i for i, h in enumerate(t_header_lower)
                             if "performance" in h), None)
            if cat_col is not None and perf_col is not None:
                trend_mismatches = 0
                for row in data_rows:
                    cname = str(row[0] or "").strip().lower()
                    cat = str(row[cat_col] or "").strip().lower()
                    perf = str(row[perf_col] or "").strip().lower()
                    exp = expected.get(cname)
                    if exp is None:
                        continue
                    if "completion" in cat:
                        gt = "above" if exp["completion_rate"] >= 75 else "below"
                        if gt not in perf:
                            trend_mismatches += 1
                    elif "grade" in cat:
                        gt = "above" if exp["avg_grade"] >= 70 else "below"
                        if gt not in perf:
                            trend_mismatches += 1
                    elif "late" in cat:
                        gt = "acceptable" if exp["late_rate"] <= 15 else "high"
                        if gt not in perf:
                            trend_mismatches += 1
                check("Trends Performance labels match DB-computed values",
                      trend_mismatches == 0,
                      f"{trend_mismatches} mismatches across {len(data_rows)} rows")


def check_notion():
    print("\n=== Checking Notion Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()
        found_page = None
        for page_id, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "learning analytics dashboard" in props_str:
                found_page = page_id
                break
        check("Notion page titled 'Learning Analytics Dashboard' exists",
              found_page is not None,
              f"Found {len(pages)} pages")
        if found_page:
            cur.execute(
                "SELECT content FROM notion.blocks WHERE parent_id = %s",
                (found_page,))
            blocks = cur.fetchall()
            full_text = " ".join(json.dumps(b[0]).lower() if b[0] else "" for b in blocks)
            check("Notion page has at least 4 blocks",
                  len(blocks) >= 4,
                  f"Found {len(blocks)} blocks")
            # Each required section
            sections = [
                ("benchmark", "overall benchmark performance"),
                ("meet", "courses meeting all targets"),
                ("attention", "courses needing attention"),
                ("recommend", "recommendations"),
            ]
            for keyword, label in sections:
                check(f"Notion page mentions {label}",
                      keyword in full_text,
                      f"Missing keyword '{keyword}'")
        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
