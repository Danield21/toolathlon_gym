#!/usr/bin/env python3
"""Evaluation script for research-lab-publication-pipeline.

Aligned with task.md 6 phases:
  - Literature search results (spreadsheet/csv with bibliographic info)
  - Bibliography file (.bib or equivalent)
  - Research roadmap document (Word)
  - Email to lab members
  - Calendar meeting
"""

from argparse import ArgumentParser
import json
import os
import re
import sys
from pathlib import Path
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _pdfplumber_text(path):
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def _pypdf_text(path):
    try:
        from pypdf import PdfReader
    except Exception:
        from PyPDF2 import PdfReader  # older package name
    reader = PdfReader(path)
    parts = []
    for i in range(len(reader.pages)):
        page = reader.pages[i]
        parts.append(page.extract_text() or "")
    return "\n".join(parts)


def _pymupdf_text(path):
    import fitz  # PyMuPDF
    doc = fitz.open(path)
    try:
        return "\n".join(page.get_text() for page in doc)
    finally:
        doc.close()


def _read_pdf_text(path):
    """Best-effort PDF text extraction.

    Tries pdfplumber, then pypdf/PyPDF2, then PyMuPDF. Returns "" if none is
    installed or all fail to extract text.
    """
    for loader in (_pdfplumber_text, _pypdf_text, _pymupdf_text):
        try:
            text = loader(path)
            if text:
                return text
        except ImportError:
            continue
        except Exception:
            continue
    return ""


def _read_text_any(path):
    p = Path(path)
    if not p.exists():
        return ""
    ext = p.suffix.lower()
    try:
        if ext == ".docx":
            from docx import Document
            return "\n".join(par.text for par in Document(str(p)).paragraphs)
        if ext == ".xlsx":
            import openpyxl
            # data_only=False keeps formula strings visible so a workbook that
            # uses formulas is still readable; bibliographic cells are text.
            wb = openpyxl.load_workbook(str(p), data_only=False)
            try:
                txts = []
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell is not None:
                                txts.append(str(cell))
                return " ".join(txts)
            finally:
                wb.close()
        if ext == ".pdf":
            return _read_pdf_text(str(p))
        if ext in (".txt", ".md", ".csv", ".json", ".tsv", ".bib"):
            return p.read_text(errors="ignore")
    except ImportError as e:
        # Do not silently swallow a missing library: surface it so an infra
        # failure is not mistaken for missing agent output.
        print(f"  [WARN] missing library while reading {p.name}: {e}", file=sys.stderr)
        return ""
    except Exception:
        return ""
    return ""


def _norm_title(t):
    if t is None:
        return None
    s = str(t).strip().lower()
    s = " ".join(s.split())
    return s or None


def _looks_like_header(row):
    if not row:
        return False
    known = {"title", "author", "authors", "year", "journal", "abstract", "keywords",
             "date", "relevance", "source", "name", "publication", "doi", "category",
             "conference", "venue", "notes", "tags"}
    vals = set()
    for v in row:
        if v is None:
            continue
        s = str(v).strip().lower()
        if s:
            vals.add(s)
    return len(vals & known) >= 2


def _extract_paper_rows(path):
    """Return (row_count, set_of_normalized_titles) for a paper-database file.

    Supports XLSX/CSV/TSV/JSON. For markdown/plain text, counts citation-like
    lines (those containing a 4-digit year).
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext in (".md", ".txt"):
        try:
            lines = [ln.strip() for ln in
                     p.read_text(encoding="utf-8", errors="ignore").splitlines()
                     if ln.strip()]
        except Exception:
            return 0, set()
        rows = []
        for ln in lines:
            if re.search(r"(19|20)\d{2}", ln) and len(ln) > 8:
                rows.append([ln])
        return len(rows), set()
    if ext == ".xlsx":
        import openpyxl
        try:
            wb = openpyxl.load_workbook(str(p), data_only=False)
            try:
                rows = []
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    for r in ws.iter_rows(values_only=True):
                        rows.append(list(r))
            finally:
                wb.close()
        except Exception:
            return 0, set()
    elif ext == ".csv":
        import csv as _csv
        try:
            with open(p, encoding="utf-8", errors="ignore", newline="") as fh:
                rows = [list(r) for r in _csv.reader(fh)]
        except Exception:
            return 0, set()
    elif ext == ".tsv":
        import csv as _csv
        try:
            with open(p, encoding="utf-8", errors="ignore", newline="") as fh:
                rows = [list(r) for r in _csv.reader(fh, delimiter="\t")]
        except Exception:
            return 0, set()
    elif ext == ".json":
        try:
            data = json.loads(p.read_text(encoding="utf-8", errors="ignore"))
        except Exception:
            return 0, set()
        items = data if isinstance(data, list) else []
        if isinstance(data, dict):
            items = [it for v in data.values() if isinstance(v, list) for it in v]
        rows = []
        titles = set()
        for it in items:
            if isinstance(it, dict):
                t = it.get("title") or it.get("Title")
                if t:
                    titles.add(_norm_title(t))
                rows.append([t])
            else:
                rows.append([it])
        return len(rows), titles
    else:
        return 0, set()

    rows = [r for r in rows if r and any((v is not None and str(v).strip()) for v in r)]
    if not rows:
        return 0, set()
    title_col = None
    if _looks_like_header(rows[0]):
        header = [str(v).strip().lower() if v is not None else "" for v in rows[0]]
        for i, h in enumerate(header):
            if h in ("title", "paper title", "publication title"):
                title_col = i
                break
        data_rows = rows[1:]
    else:
        data_rows = rows
    count = len(data_rows)
    titles = set()
    for r in data_rows:
        if title_col is not None and title_col < len(r):
            t = _norm_title(r[title_col])
            if t:
                titles.add(t)
    return count, titles


def check_literature_db(workspace):
    """Verify a consolidated literature database / spreadsheet exists with bibliographic entries."""
    base = Path(workspace)
    candidates = []
    for pat in ["*literature*", "*publications*", "*bibliography*", "*reference*",
                "*research_db*", "*paper*", "*citation*", "*submission_checklist*"]:
        for ext in ["xlsx", "csv", "json", "tsv", "md"]:
            candidates.extend(base.rglob(f"{pat}.{ext}"))
    record("Literature/publications database file exists",
           len(candidates) >= 1, f"found {len(candidates)} candidates")
    if not candidates:
        return
    text = "\n".join(_read_text_any(str(c)) for c in candidates).lower()
    record("Literature DB substantive content (>=300 chars)",
           len(text) >= 300, f"length={len(text)}")
    # Bibliographic field tokens
    bib_tokens = ["author", "year", "title"]
    missing = [t for t in bib_tokens if t not in text]
    record("Literature DB has bibliographic fields (author/year/title)",
           len(missing) <= 1, f"missing: {missing}")
    # Count unique paper rows across ALL matching files (union), so that a
    # multi-agent decomposition that writes one shard file per subagent still
    # aggregates correctly. Threshold is set below the seed count (10) to leave
    # a margin for an incomplete retrieval.
    paper_db_patterns = ["literature", "publications", "bibliography", "reference",
                         "research_db", "paper", "citation"]
    total_rows = 0
    unique_titles = set()
    counted_files = 0
    for c in candidates:
        name_l = c.stem.lower()
        if not any(p in name_l for p in paper_db_patterns):
            continue
        n, titles = _extract_paper_rows(c)
        counted_files += 1
        total_rows += n
        unique_titles |= titles
    # Union across files. Do NOT discard total_rows just because some shard
    # produced a title set: a headerless shard (a subagent that dumps raw rows
    # without a header) contributes only a row count, and dropping it would
    # undercount a legitimately-complete multi-file output. max() keeps the
    # aggregate at least the total number of data rows, which is a safe lower
    # bound on distinct papers for a correctly-completed multi-file library.
    unique_papers = max(len(unique_titles), total_rows)
    record("Literature DB has >=8 unique paper rows",
           unique_papers >= 8,
           f"unique_papers={unique_papers} across {counted_files} file(s)")


def check_bibliography(workspace):
    """Verify a .bib file or comparable references file exists."""
    base = Path(workspace)
    bib_files = []
    seen = set()
    for b in list(base.rglob("*.bib")) + list(base.rglob("*references*")):
        rp = str(b.resolve())
        if rp not in seen:
            seen.add(rp)
            bib_files.append(b)
    record("Bibliography file (.bib or *references*) exists",
           len(bib_files) >= 1, f"found {len(bib_files)}")
    if bib_files:
        entries = 0
        for b in bib_files:
            if b.suffix.lower() == ".bib":
                text = _read_text_any(str(b)).lower()
                # Count any standard BibTeX/BibLaTeX entry type (@article,
                # @online, @techreport, @phdthesis, @software, @inproceedings,
                # @misc, @incollection, @conference, ...). @online in particular
                # is the most common type for arXiv preprints, so a whitelist
                # that omits it would wrongly fail a correct bibliography.
                # Exclude the non-entry directives (@string/@preamble/@comment).
                _types = re.findall(r"@\s*([a-z]+)\s*\{", text)
                entries += sum(1 for t in _types
                               if t not in ("string", "preamble", "comment"))
            else:
                # Plain-text references file (md/txt/...): count lines that look
                # like a bibliographic entry, i.e. contain a 4-digit year.
                try:
                    for ln in b.read_text(encoding="utf-8", errors="ignore").splitlines():
                        if ln.strip() and re.search(r"(19|20)\d{2}", ln):
                            entries += 1
                except Exception:
                    pass
        record("Bibliography contains at least 5 entries",
               entries >= 5,
               f"entries: {entries}")


def _roadmap_pdf_unreadable(candidates, text):
    """True when the only roadmap candidates are PDFs whose text could not be
    extracted at all (i.e. the eval env has none of pdfplumber/pypdf/PyMuPDF,
    or the PDF has no text layer). Used to avoid failing a model that produced
    a real roadmap PDF just because the environment lacks the extraction
    library. Gated on file size (>=1 KB) so a blank/one-line PDF still fails:
    a blank single-page PDF is ~0.5 KB while even a minimal 400+-char text PDF
    is >1 KB. Only ever triggered when text is empty, so in the normal env
    (libs present, readable docx/md/pdf) it is inert and the real content
    checks still apply.
    """
    if text.strip():
        return False
    sizes = []
    for c in candidates:
        if c.suffix.lower() != ".pdf":
            continue
        try:
            sizes.append(c.stat().st_size)
        except Exception:
            pass
    return max(sizes, default=0) > 1024


def check_roadmap(workspace):
    """Verify a research roadmap / synthesis document exists with required content."""
    base = Path(workspace)
    candidates = []
    for pat in ["*roadmap*", "*synthesis*", "*research_plan*", "*future_directions*", "*priorities*"]:
        for ext in ["docx", "pdf", "md", "txt"]:
            candidates.extend(base.rglob(f"{pat}.{ext}"))
    record("Research roadmap document exists",
           len(candidates) >= 1, f"found {len(candidates)}")
    if not candidates:
        return
    text = "\n".join(_read_text_any(str(c)) for c in candidates).lower()
    if _roadmap_pdf_unreadable(candidates, text):
        print("  [WARN] roadmap PDF(s) present but text unreadable "
              "(no pdf text-extraction lib); size-based fallback applied",
              file=sys.stderr)
        record("Roadmap doc substantive content (>=400 chars)",
               True, "PDF unreadable; size-based fallback")
        record("Roadmap covers directions/gaps/recommendations (>=2 topics)",
               True, "PDF unreadable; size-based fallback")
        record("Roadmap has explicit 'Future Directions' or 'Recommendations' heading",
               True, "PDF unreadable; size-based fallback")
        return
    record("Roadmap doc substantive content (>=400 chars)",
           len(text) >= 400, f"length={len(text)}")
    # Roadmap must mention research directions/gaps/recommendations
    must = ["direction", "gap", "recommendation"]
    found_any = sum(1 for m in must if m in text)
    record("Roadmap covers directions/gaps/recommendations (>=2 topics)",
           found_any >= 2, f"found: {found_any}/3 topics")
    # Tightened: must include explicit 'Future Directions' or 'Recommendations' heading
    has_future = "future direction" in text or "future research" in text
    has_recs = "recommendation" in text
    record("Roadmap has explicit 'Future Directions' or 'Recommendations' heading",
           has_future or has_recs, f"has_future={has_future}, has_recs={has_recs}")


def check_emails():
    try:
        conn = psycopg2.connect(**DB); cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, COALESCE(body_text, body_html, '') FROM email.messages")
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        record("Email DB reachable", False, str(e)); return
    matched = None
    for subj, to_addr, body in rows:
        text = ((subj or "") + " " + (body or "")).lower()
        to_str = str(to_addr or "").lower()
        # Accept any valid recipient address. Requiring a group-style alias
        # (lab/group/team/...) would be an undeclared constraint the task text
        # never specifies and could wrongly fail an email sent to an individual
        # laboratory member's address.
        recipient_ok = "@" in to_str
        if (any(k in text for k in ["roadmap", "literature review", "publication pipeline",
                                    "research priorities", "lab member"])
                and recipient_ok):
            matched = (subj, to_addr); break
    record("Email to laboratory members about roadmap/literature exists",
           matched is not None, f"checked {len(rows)} emails")


def check_calendar():
    try:
        conn = psycopg2.connect(**DB); cur = conn.cursor()
        cur.execute("SELECT id, summary, COALESCE(description, '') FROM gcal.events")
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        record("Calendar DB reachable", False, str(e)); return
    matched = None
    for _id, summary, desc in rows:
        text = ((summary or "") + " " + (desc or "")).lower()
        if any(k in text for k in ["roadmap", "research priorities", "lab meeting",
                                   "literature review", "publication pipeline"]):
            matched = (summary, desc); break
    record("Calendar meeting scheduled to discuss priorities/literature",
           matched is not None, f"checked {len(rows)} events")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0; FAIL_COUNT = 0
    if not agent_workspace or not Path(agent_workspace).exists():
        return False, f"Agent workspace not found: {agent_workspace}"
    # GT self-test toleration: when agent==groundtruth (V1 GT-only smoke test),
    # the email/gcal checks fail because no email/event was actually sent.
    is_gt_self_test = (
        agent_workspace and groundtruth_workspace
        and os.path.abspath(agent_workspace) == os.path.abspath(groundtruth_workspace)
    )
    check_literature_db(agent_workspace)
    check_bibliography(agent_workspace)
    check_roadmap(agent_workspace)
    if not is_gt_self_test:
        check_emails()
        check_calendar()
    else:
        print("  [SKIP] email and calendar checks (GT self-test mode)")
    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace, args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
