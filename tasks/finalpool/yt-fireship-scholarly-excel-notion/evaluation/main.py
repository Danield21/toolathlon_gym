"""
Evaluation for yt-fireship-scholarly-excel-notion task.

Strict checks (no accuracy threshold; FAIL_COUNT must be 0):
1. Tech_Research_Map.xlsx exists with Video_Paper_Mapping and All_Papers sheets.
2. Video_Paper_Mapping rank order matches DB top-8 by view count, with required topics.
3. All_Papers contains expected paper titles per topic and is sorted by Tech_Topic, Citations DESC.
4. Notion page 'Tech Content Research Bridge' exists with summary + 8 video references.
"""
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

# Authoritative video ranking from yt DB (Fireship channel, top 8 by view_count desc)
EXPECTED_VIDEOS = [
    (1, "Big Tech in panic mode... Did DeepSeek R1 just pop the AI bubble?", 3878491, "AI and Large Language Models"),
    (2, "This free Chinese AI just crushed OpenAI's $200 o1 model...", 3115193, "AI and Large Language Models"),
    (3, "100+ Linux Things you Need to Know", 2891861, "Linux and Systems"),
    (4, "25 crazy software bugs explained", 2496715, "Software Engineering"),
    (5, "Some bad code just broke a billion Windows machines", 2467547, "Software Engineering"),
    (6, "DeepSeek stole our tech... says OpenAI", 2436569, "AI and Large Language Models"),
    (7, "Microsoft's new chip looks like science fiction…", 2329914, "Hardware and Computing"),
    (8, "Hackers are destroying the Internet's history book right now", 1927279, "Security"),
]

# Top paper title per topic (highest citation_count among injected papers)
TOP_PAPER_BY_TOPIC = {
    "AI and Large Language Models": ("DeepSeek R1: Incentivizing Reasoning Capability in LLMs via Reinforcement Learning", 1204),
    "Linux and Systems": ("Performance Analysis of Linux Kernel Scheduler Algorithms in Cloud Environments", 245),
    "Software Engineering": ("Automated Detection and Classification of Software Bugs Using Deep Learning", 312),
    "Hardware and Computing": ("Next-Generation Quantum-Classical Hybrid Computing Architectures", 423),
    "Security": ("Supply Chain Security Vulnerabilities in Open Source Software: A Systematic Review", 334),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def norm(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def get_int(v):
    try:
        if v is None:
            return None
        return int(float(str(v).replace(",", "").strip()))
    except (TypeError, ValueError):
        return None


def find_sheet(wb, *terms):
    """Find sheet matching all lowercase terms in name."""
    for n in wb.sheetnames:
        ln = n.lower()
        if all(t in ln for t in terms):
            return wb[n]
    return None


def check_excel(agent_workspace):
    print("\n=== Check 1: Tech_Research_Map.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Tech_Research_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Research_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Tech_Research_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    # ----- Video_Paper_Mapping sheet -----
    mapping = find_sheet(wb, "video", "paper") or find_sheet(wb, "mapping") or find_sheet(wb, "video_paper")
    record("Has Video_Paper_Mapping sheet", mapping is not None,
           f"Sheets: {wb.sheetnames}")

    if mapping is not None:
        rows = [r for r in mapping.iter_rows(values_only=True) if any(c is not None for c in r)]
        if rows:
            headers = [norm(c) for c in rows[0]]
            data_rows = rows[1:]
            record("Video_Paper_Mapping has 8 data rows", len(data_rows) == 8,
                   f"Got {len(data_rows)}")
            req_headers = ["video_rank", "video_title", "video_views", "tech_topic",
                           "paper_count", "top_paper_title", "top_paper_year",
                           "top_paper_citations"]
            for h in req_headers:
                record(f"Video_Paper_Mapping has '{h}' column", h in headers,
                       f"Got headers: {headers}")
            hcol = {h: i for i, h in enumerate(headers)}

            # Per-rank check
            for rank, exp_title, exp_views, exp_topic in EXPECTED_VIDEOS:
                if rank > len(data_rows):
                    record(f"VPM row rank={rank} present", False, f"only {len(data_rows)} rows")
                    continue
                row = data_rows[rank - 1]
                rank_idx = hcol.get("video_rank")
                title_idx = hcol.get("video_title")
                views_idx = hcol.get("video_views")
                topic_idx = hcol.get("tech_topic")
                top_paper_idx = hcol.get("top_paper_title")

                act_rank = get_int(row[rank_idx]) if rank_idx is not None else None
                act_title = norm(row[title_idx]) if title_idx is not None else ""
                act_views = get_int(row[views_idx]) if views_idx is not None else None
                act_topic = norm(row[topic_idx]) if topic_idx is not None else ""
                act_top_paper = norm(row[top_paper_idx]) if top_paper_idx is not None else ""

                record(f"VPM rank {rank} == {rank}", act_rank == rank,
                       f"got {act_rank}")
                # Title comparison: trim ellipsis variants
                norm_exp = norm(exp_title).replace("…", "...")
                norm_act = act_title.replace("…", "...")
                title_match = (norm_act == norm_exp) or (norm_exp[:80] in norm_act) or (norm_act[:80] in norm_exp)
                record(f"VPM rank {rank} title", title_match,
                       f"expected '{exp_title[:60]}...', got '{row[title_idx] if title_idx is not None else ''}'"[:200])
                record(f"VPM rank {rank} views == {exp_views}",
                       act_views == exp_views, f"got {act_views}")
                record(f"VPM rank {rank} topic == '{exp_topic}'",
                       act_topic == norm(exp_topic), f"got '{row[topic_idx] if topic_idx is not None else ''}'")
                # Top paper consistency
                exp_top = TOP_PAPER_BY_TOPIC.get(exp_topic)
                if exp_top:
                    record(f"VPM rank {rank} Top_Paper_Title matches topic top",
                           exp_top[0].lower() in act_top_paper or act_top_paper in exp_top[0].lower(),
                           f"expected '{exp_top[0][:60]}', got '{row[top_paper_idx] if top_paper_idx is not None else ''}'"[:200])

    # ----- All_Papers sheet -----
    papers = find_sheet(wb, "all", "paper") or find_sheet(wb, "all_papers") or find_sheet(wb, "papers")
    record("Has All_Papers sheet", papers is not None, f"Sheets: {wb.sheetnames}")

    if papers is not None:
        rows = [r for r in papers.iter_rows(values_only=True) if any(c is not None for c in r)]
        if rows:
            headers = [norm(c) for c in rows[0]]
            data_rows = rows[1:]
            req_headers = ["tech_topic", "paper_title", "first_author", "year",
                           "citations", "abstract_snippet"]
            for h in req_headers:
                record(f"All_Papers has '{h}' column", h in headers, f"Got: {headers}")
            record("All_Papers has >= 10 rows (5 topics x 2 papers)",
                   len(data_rows) >= 10, f"Got {len(data_rows)} rows")

            hcol = {h: i for i, h in enumerate(headers)}
            topic_idx = hcol.get("tech_topic")
            citations_idx = hcol.get("citations")
            title_idx = hcol.get("paper_title")

            # Sort check: rows sorted by Tech_Topic asc, then Citations desc within topic
            if topic_idx is not None and citations_idx is not None:
                last_topic = ""
                last_citation_per_topic = {}
                sort_ok = True
                for r in data_rows:
                    t = norm(r[topic_idx]) if topic_idx < len(r) else ""
                    c = get_int(r[citations_idx]) if citations_idx < len(r) else 0
                    if t < last_topic:
                        sort_ok = False
                        break
                    if t == last_topic:
                        prev = last_citation_per_topic.get(t)
                        if prev is not None and c is not None and c > prev:
                            sort_ok = False
                            break
                    last_topic = t
                    last_citation_per_topic[t] = c
                record("All_Papers sorted by Tech_Topic asc, Citations desc within topic",
                       sort_ok, "Sort order violated")

            # Verify the top paper for each topic appears with correct citations
            if title_idx is not None and citations_idx is not None and topic_idx is not None:
                title_to_citations_by_topic = {}
                for r in data_rows:
                    t = norm(r[topic_idx]) if topic_idx < len(r) else ""
                    title_to_citations_by_topic.setdefault(t, {})[norm(r[title_idx])] = get_int(r[citations_idx])

                for topic, (top_title, top_cit) in TOP_PAPER_BY_TOPIC.items():
                    bucket = title_to_citations_by_topic.get(norm(topic), {})
                    if not bucket:
                        record(f"All_Papers: topic '{topic}' present", False,
                               f"got topics: {list(title_to_citations_by_topic.keys())}")
                        continue
                    record(f"All_Papers: topic '{topic}' present", True)
                    # Find best match
                    found = None
                    for k, v in bucket.items():
                        if top_title.lower() in k or k in top_title.lower():
                            found = v
                            break
                    record(f"All_Papers top paper for '{topic}' has correct citations ({top_cit})",
                           found == top_cit, f"got {found} (have: {bucket})")


def check_notion():
    print("\n=== Check 2: Notion page 'Tech Content Research Bridge' ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()
        matching = []
        for pid, props in pages:
            props_str = json.dumps(props).lower() if props else ""
            if "tech content research bridge" in props_str:
                matching.append((pid, props))
        record("Notion page 'Tech Content Research Bridge' exists", len(matching) >= 1,
               f"Found {len(pages)} total pages, {len(matching)} matching")

        if matching:
            page_id = matching[0][0]
            cur.execute("""
                SELECT block_data FROM notion.blocks WHERE parent_id = %s AND archived = false
            """, (page_id,))
            blocks = cur.fetchall()
            all_text = " ".join(json.dumps(b[0]) for b in blocks if b[0]).lower()
            record("Notion page has substantial content (>=500 chars)",
                   len(all_text) >= 500, f"Got {len(all_text)} chars")

            # Verify all 8 video titles or distinctive snippets appear in Notion content
            for rank, title, _, _ in EXPECTED_VIDEOS:
                # Use the first ~30 chars of the title as snippet (avoid encoding issues)
                snippet = title.lower()[:30]
                # Strip common punctuation
                snippet_clean = snippet.replace("…", "")
                record(f"Notion page mentions video #{rank} ('{snippet_clean[:30]}')",
                       snippet_clean in all_text or snippet[:20] in all_text,
                       "Video reference not found")
    except Exception as e:
        record("Notion page check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    print(f"\nOverall: {PASS_COUNT}/{total} checks passed")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "total_passed": PASS_COUNT,
                "total_checks": total,
                "failed": FAIL_COUNT,
            }, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL ({FAIL_COUNT} failures)")
        sys.exit(1)


if __name__ == "__main__":
    main()
