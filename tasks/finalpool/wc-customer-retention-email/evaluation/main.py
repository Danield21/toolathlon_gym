"""Evaluation for wc-customer-retention-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check the Excel output against groundtruth."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "VIP_Customer_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "VIP_Customer_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # Find VIP Customers sheet (case-insensitive)
    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    agent_ws = get_sheet(agent_wb, "VIP Customers")
    gt_ws = get_sheet(gt_wb, "VIP Customers")

    check("Sheet 'VIP Customers' exists", agent_ws is not None,
          f"Found sheets: {agent_wb.sheetnames}")
    if not agent_ws or not gt_ws:
        return False

    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
    agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

    check("Row count matches (10 VIP customers)", len(agent_rows) == 10,
          f"Expected 10, got {len(agent_rows)}")

    all_ok = True
    for gt_row in gt_rows:
        rank, name, email, orders, total = gt_row
        # Find matching row by email
        matched = None
        for ar in agent_rows:
            if ar and len(ar) >= 5 and str_match(ar[2], email):
                matched = ar
                break
        if matched:
            check(f"Customer {email} rank", num_close(matched[0], rank, 0),
                  f"Expected {rank}, got {matched[0]}")
            check(f"Customer {email} name == '{name}'",
                  str_match(matched[1], name),
                  f"Expected '{name}', got '{matched[1]}'")
            check(f"Customer {email} orders_count",
                  num_close(matched[3], orders, 0),
                  f"Expected {orders}, got {matched[3]}")
            check(f"Customer {email} total_spent",
                  num_close(matched[4], total, 0.01),
                  f"Expected {total}, got {matched[4]}")
        else:
            check(f"Customer {email} found", False, "Not in agent output")
            all_ok = False

    return all_ok


def check_emails():
    """Check that VIP emails were sent to the correct addresses."""
    print("\n=== Checking Emails ===")

    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return False

    # Get expected top 10 customers
    cur.execute("""
        SELECT email, first_name, ROUND(total_spent::numeric, 2)
        FROM wc.customers
        ORDER BY total_spent::numeric DESC
        LIMIT 10
    """)
    expected_customers = cur.fetchall()

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    conn.close()

    def find_email_for_recipient(recipient):
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    check("At least 10 emails total", len(all_emails) >= 10,
          f"Found {len(all_emails)} emails")

    for email_addr, fname, total in expected_customers:
        result = find_email_for_recipient(email_addr)
        check(f"Email sent to {email_addr}", result is not None)
        if not result:
            continue
        subj, from_addr, to_addr, body = result
        check(f"  Email to {email_addr}: subject == 'Thank You, VIP Customer!'",
              (subj or "").strip().lower() == "thank you, vip customer!",
              f"got: {subj}")
        check(f"  Email to {email_addr}: from_addr == 'vip-program@store.example.com'",
              str(from_addr or "").strip().lower() == "vip-program@store.example.com",
              f"got: {from_addr}")
        body_text = body or ""
        body_lower = body_text.lower()
        check(f"  Email to {email_addr}: body addresses by first name '{fname}'",
              fname.lower() in body_lower, "missing first name")
        # body must contain total_spent value (formatted with 2 decimals or as integer)
        total_str_2 = f"{float(total):.2f}"
        total_str_int = str(int(float(total)))
        check(f"  Email to {email_addr}: body mentions total spent {total}",
              total_str_2 in body_text or total_str_int in body_text,
              f"missing {total} in body")

    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("WC CUSTOMER RETENTION EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, gt_dir)
    check_emails()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
