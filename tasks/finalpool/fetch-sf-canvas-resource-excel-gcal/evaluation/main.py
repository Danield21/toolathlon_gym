"""Evaluation script for fetch-sf-canvas-resource-excel-gcal."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-sf-canvas-resource-excel-gcal"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Canvas_Resource_Report.xlsx")
    check("Canvas_Resource_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Canvas_Resource_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames and gt_wb is not None:
            ws = wb["Data_Analysis"]
            gt_ws = gt_wb["Data_Analysis"] if "Data_Analysis" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Data_Analysis has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}, gt={len(gt_rows)}")
            check_columns('Data_Analysis', ['Course', 'Code', 'Enrollment', 'Avg_Score', 'Pass_Rate'],
                          len(gt_rows) if gt_ws is not None else 6)
            # Cell value comparison vs GT (all rows by Code key, exact)
            if gt_ws is not None:
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                header_map = {h: i for i, h in enumerate(headers)}
                code_idx_gt = gt_headers.index("code") if "code" in gt_headers else 1
                code_idx_agent = header_map.get("code", 1)
                gt_by_code = {str(r[code_idx_gt]).strip().upper(): r for r in gt_rows if r and r[code_idx_gt]}
                agent_by_code = {str(r[code_idx_agent]).strip().upper(): r for r in data_rows if r and len(r) > code_idx_agent and r[code_idx_agent]}
                for code, gt_row in gt_by_code.items():
                    found = code in agent_by_code
                    check(f"Data_Analysis course code '{code}' present", found)
                    if found:
                        agent_row = agent_by_code[code]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row):
                                continue
                            if gt_h == "code":
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                # Tighter tolerance for percent metrics (Pass_Rate, Avg_Score range 0-100)
                                if gt_h in ("pass_rate", "avg_score"):
                                    tol = max(1.0, abs(gf) * 0.03)
                                else:
                                    tol = max(1.0, abs(gf) * 0.05)
                                check(f"Data_Analysis '{code}' {gt_h} ~{gf}",
                                      abs(gf - af) <= tol, f"got {av}")
                            elif gv is not None and av is not None:
                                # Allow Course name to be substring/extended (e.g. with subtitle variations)
                                gs = str(gv).strip().lower()
                                avs = str(av).strip().lower()
                                check(f"Data_Analysis '{code}' {gt_h}",
                                      gs == avs or gs in avs or avs in gs,
                                      f"expected {gs[:60]}, got {avs[:60]}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames and gt_wb is not None:
            ws = wb["Metrics"]
            gt_ws = gt_wb["Metrics"] if "Metrics" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Metrics has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Metrics', ['Metric', 'Value'], len(gt_rows) if gt_ws is not None else 4)
            if gt_ws is not None:
                gt_metric_map = {str(r[0]).strip().lower(): r[1] for r in gt_rows if r and r[0]}
                agent_metric_map = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for gt_m, gt_v in gt_metric_map.items():
                    found = gt_m in agent_metric_map
                    check(f"Metrics has '{gt_m}'", found)
                    if found:
                        av = agent_metric_map[gt_m]
                        gf = safe_float(gt_v)
                        af = safe_float(av)
                        if gf is not None and af is not None:
                            # Tighter tolerance for rate metrics
                            if "rate" in gt_m or "score" in gt_m:
                                tol = max(1.0, abs(gf) * 0.03)
                            else:
                                tol = max(1.0, abs(gf) * 0.05)
                            check(f"Metrics '{gt_m}' ~{gf}",
                                  abs(gf - af) <= tol, f"got {av}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check_columns('Recommendations', ['Priority', 'Action', 'Course'], 2)

        # Calendar check: title equals 'Analysis Review' AND date 2026-03-14 14:00-15:00 UTC
        try:
            from datetime import datetime, timezone, timedelta
            conn = get_conn()
            cur = conn.cursor()
            # Tighten to title equality (case-insensitive trim)
            cur.execute("""SELECT summary, start_datetime, end_datetime FROM gcal.events
                           WHERE LOWER(TRIM(summary)) = LOWER('Analysis Review')""")
            events = cur.fetchall()
            check("'Analysis Review' event exists",
                  len(events) >= 1, f"found {len(events)} events with exact title")
            target_start_utc = datetime(2026, 3, 14, 14, 0, 0, tzinfo=timezone.utc)
            target_end_utc = datetime(2026, 3, 14, 15, 0, 0, tzinfo=timezone.utc)
            date_match = False
            start_match = False
            end_match = False
            for s, sd, ed in events:
                # Convert to UTC if TZ-aware; if naive, assume UTC
                def to_utc(dt):
                    if dt is None:
                        return None
                    if hasattr(dt, "tzinfo") and dt.tzinfo is not None:
                        return dt.astimezone(timezone.utc)
                    return dt.replace(tzinfo=timezone.utc) if hasattr(dt, "replace") else None
                sd_utc = to_utc(sd)
                ed_utc = to_utc(ed)
                if sd_utc and sd_utc.date() == target_start_utc.date():
                    date_match = True
                if sd_utc and abs((sd_utc - target_start_utc).total_seconds()) < 60:
                    start_match = True
                if ed_utc and abs((ed_utc - target_end_utc).total_seconds()) < 60:
                    end_match = True
            check("'Analysis Review' on 2026-03-14 (UTC)", date_match)
            check("'Analysis Review' starts at 14:00 UTC", start_match)
            check("'Analysis Review' ends at 15:00 UTC", end_match)
            conn.close()
        except Exception as e:
            check("Calendar check", False, str(e))

    # Move outside excel-exists guard
    check("sf_canvas_resource_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "sf_canvas_resource_processor.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
