"""Evaluation for sf-sales-monthly-trends."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_month(v):
    """Normalize month cell to YYYY-MM-DD string."""
    if v is None:
        return None
    try:
        return v.strftime("%Y-%m-%d")
    except Exception:
        s = str(v).strip()
        # Keep just first 10 chars if long date
        if len(s) >= 10 and s[4] == '-':
            return s[:10]
        return s


def _db_monthly_fetch():
    """Fetch live monthly revenue data from DB as cross-reference. Returns None if DB unavailable."""
    try:
        import psycopg2
        conn = psycopg2.connect(
            host=os.environ.get('PGHOST', 'localhost'),
            port=int(os.environ.get('PGPORT', '5432')),
            dbname=os.environ.get('PGDATABASE', 'toolathlon_gym'),
            user=os.environ.get('PGUSER', 'eigent'),
            password=os.environ.get('PGPASSWORD', 'camel'))
        cur = conn.cursor()
        cur.execute('SELECT * FROM sf_data."SALES_DW__ANALYTICS__MONTHLY_REVENUE" ORDER BY "MONTH_KEY"')
        rows = cur.fetchall()
        conn.close()
        results = []
        for r in rows:
            month_key, total_rev, order_cnt, uniq_cust, avg_ov = r[0], r[1], r[2], r[3], r[4]
            results.append({
                "month": month_key.strftime("%Y-%m-%d") if month_key else None,
                "revenue": float(total_rev) if total_rev is not None else None,
                "order_count": int(order_cnt) if order_cnt is not None else None,
                "unique_customers": int(uniq_cust) if uniq_cust is not None else None,
                "avg_order_value": float(avg_ov) if avg_ov is not None else None,
            })
        return results
    except Exception as e:
        print(f"  (DB query skipped: {e})")
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Monthly_Trends.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Monthly_Trends.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # Cross-reference with live DB (fallback to GT if unavailable)
    db_rows = _db_monthly_fetch()

    # Check sheet: Monthly Trends
    print(f"  Checking Monthly Trends...")
    a_rows = load_sheet_rows(agent_wb, "Monthly Trends")
    g_rows = load_sheet_rows(gt_wb, "Monthly Trends")
    if a_rows is None:
        all_errors.append("Sheet 'Monthly Trends' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Monthly Trends' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # --- Chronological order check on agent data ---
        a_months = [_norm_month(r[0]) for r in a_data if r and r[0] is not None]
        a_months_sorted = sorted(a_months)
        if a_months != a_months_sorted:
            errors.append(f"Monthly Trends rows not sorted chronologically by Month (got {a_months[:3]}... expected sorted)")

        # --- Row count check: must cover all expected months ---
        expected_month_set = set()
        if db_rows:
            expected_month_set = {r["month"] for r in db_rows if r["month"]}
        else:
            for g_row in g_data:
                if g_row and g_row[0] is not None:
                    expected_month_set.add(_norm_month(g_row[0]))
        a_month_set = {m for m in a_months if m}
        missing_months = expected_month_set - a_month_set
        if missing_months:
            errors.append(f"Missing months in Monthly Trends: {sorted(missing_months)[:5]}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[_norm_month(row[0]).lower()] = row

        # --- Cross reference primary source: db_rows if available, else gt ---
        if db_rows:
            for d_row in db_rows:
                key = d_row["month"].lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing row: {d_row['month']}")
                    continue
                if len(a_row) > 1 and not num_close(a_row[1], d_row["revenue"], 10.0):
                    errors.append(f"{key}.Revenue: {a_row[1]} vs {d_row['revenue']} (tol=10.0)")
                if len(a_row) > 2 and not num_close(a_row[2], d_row["order_count"], 2):
                    errors.append(f"{key}.Order_Count: {a_row[2]} vs {d_row['order_count']} (tol=2)")
                if len(a_row) > 3 and not num_close(a_row[3], d_row["unique_customers"], 2):
                    errors.append(f"{key}.Unique_Customers: {a_row[3]} vs {d_row['unique_customers']} (tol=2)")
                if len(a_row) > 4 and not num_close(a_row[4], d_row["avg_order_value"], 1.0):
                    errors.append(f"{key}.Avg_Order_Value: {a_row[4]} vs {d_row['avg_order_value']} (tol=1.0)")
        else:
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = _norm_month(g_row[0]).lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing row: {g_row[0]}")
                    continue

                if len(a_row) > 1 and len(g_row) > 1:
                    if not num_close(a_row[1], g_row[1], 10.0):
                        errors.append(f"{key}.Revenue: {a_row[1]} vs {g_row[1]} (tol=10.0)")

                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 2):
                        errors.append(f"{key}.Order_Count: {a_row[2]} vs {g_row[2]} (tol=2)")

                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 2):
                        errors.append(f"{key}.Unique_Customers: {a_row[3]} vs {g_row[3]} (tol=2)")

                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 1.0):
                        errors.append(f"{key}.Avg_Order_Value: {a_row[4]} vs {g_row[4]} (tol=1.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Derive DB-authoritative summary if available
        db_summary = None
        if db_rows:
            total_rev = sum(r["revenue"] for r in db_rows)
            total_months = len(db_rows)
            avg_rev = total_rev / total_months if total_months else 0.0
            peak = max(db_rows, key=lambda r: r["revenue"])
            db_summary = {
                "total_months": total_months,
                "total_revenue": round(total_rev, 2),
                "avg_monthly_revenue": round(avg_rev, 2),
                "peak_month": peak["month"],
                "peak_revenue": round(peak["revenue"], 2),
            }

        # Per-field tolerances: strings use exact match, integers tight, revenues relative
        SUMMARY_TOLS = {
            "total_months": 0,
            "total_revenue": 100,
            "peak_revenue": None,  # relative 0.5%
            "avg_monthly_revenue": 10.0,
            "peak_month": "STRING",
            "best_month": "STRING",
            "worst_month": "STRING",
        }

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                tol_spec = SUMMARY_TOLS.get(key, 100)
                # Prefer DB-derived expected if we have it
                expected = g_row[1]
                if db_summary and key in db_summary:
                    expected = db_summary[key]
                actual = a_row[1]
                if tol_spec == "STRING":
                    # Normalize month strings - allow "YYYY-MM-DD" or "YYYY-MM" or datetime
                    a_norm = _norm_month(actual) or str(actual).strip().lower() if actual else None
                    e_norm = _norm_month(expected) or str(expected).strip().lower() if expected else None
                    if a_norm and e_norm:
                        # Accept either full date or YYYY-MM prefix
                        if not (a_norm == e_norm or a_norm[:7] == e_norm[:7]):
                            errors.append(f"{key}.Value: '{actual}' vs '{expected}' (string)")
                    elif not str_match(actual, expected):
                        errors.append(f"{key}.Value: '{actual}' vs '{expected}' (string)")
                elif tol_spec is None:
                    # relative 0.5% for peak revenue
                    try:
                        g_val = float(expected)
                        tol = max(100.0, abs(g_val) * 0.005)
                    except (TypeError, ValueError):
                        tol = 100.0
                    if not num_close(actual, expected, tol):
                        errors.append(f"{key}.Value: {actual} vs {expected} (tol={tol:.2f})")
                else:
                    # try numeric first; fall back to string for non-numeric values
                    try:
                        float(expected)
                        if not num_close(actual, expected, float(tol_spec)):
                            errors.append(f"{key}.Value: {actual} vs {expected} (tol={tol_spec})")
                    except (TypeError, ValueError):
                        if not str_match(actual, expected):
                            errors.append(f"{key}.Value: '{actual}' vs '{expected}' (string)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    docx_path = os.path.join(args.agent_workspace, "Trends_Summary.docx")
    if not os.path.exists(docx_path):
        all_errors.append("Trends_Summary.docx not found")
    else:
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs).lower()
            _headings = " ".join(p.text for p in _doc.paragraphs if p.style.name.startswith("Heading")).lower()
            if len(_text.strip()) < 50:
                all_errors.append("Trends_Summary.docx has too little text content (< 50 chars)")
            _kws = ["trends", "summary"]
            _missing = [k for k in _kws if k not in _text and k not in _headings]
            if len(_missing) == len(_kws):
                all_errors.append(f"Trends_Summary.docx missing expected keywords: {_missing}")
        except ImportError:
            if os.path.getsize(docx_path) < 100:
                all_errors.append("Trends_Summary.docx too small")
        except Exception as _e:
            all_errors.append(f"Error reading Trends_Summary.docx: {_e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
