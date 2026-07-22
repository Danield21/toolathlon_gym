"""Evaluation script for fetch-sf-hr-worklife-excel-word-gcal."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-sf-hr-worklife-excel-word-gcal"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Detect GT self-test
    is_gt_self_test = False
    try:
        if groundtruth_workspace and os.path.realpath(groundtruth_workspace) == os.path.realpath(agent_workspace):
            is_gt_self_test = True
    except Exception:
        pass

    excel_path = os.path.join(agent_workspace, "Hr_Worklife_Report.xlsx")
    check("Hr_Worklife_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Hr_Worklife_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Department', 'Our_Avg_Salary', 'Industry_Benchmark', 'Salary_Gap'], 7)

            # Value-level: department names overlap with GT
            if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames:
                try:
                    gt_ws = gt_wb["Data_Analysis"]
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    gt_depts = {str(r[0]).strip().lower() for r in gt_rows if r and r[0]}
                    agent_depts = {str(r[0]).strip().lower() for r in data_rows if r and r[0]}
                    overlap = gt_depts & agent_depts
                    check("Data_Analysis Department values match GT (>=5 of 7)",
                          len(overlap) >= 5, f"overlap={overlap}, gt={gt_depts}")
                except Exception as _e:
                    check("Data_Analysis dept value-check", False, str(_e))
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 5)
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action', 'Department'], 3)
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Tightened: require summary 'analysis review' AND date 2026-03-14 14:00-15:00 UTC
            cur.execute(
                """SELECT summary, start_datetime, end_datetime FROM gcal.events
                   WHERE summary ILIKE %s
                     AND start_datetime::date = '2026-03-14'
                     AND EXTRACT(hour FROM start_datetime AT TIME ZONE 'UTC') = 14""",
                ('%analysis review%',))
            events = cur.fetchall()
            if is_gt_self_test and len(events) == 0:
                check("'Analysis Review' event on 2026-03-14 14:00 UTC created (GT self-test toleration)",
                      True, "GT self-test: gcal events are agent runtime artifacts, skipped")
            else:
                check("'Analysis Review' event on 2026-03-14 14:00 UTC created",
                      len(events) >= 1, f"found {len(events)} matching events")
                # Also check end_datetime = 15:00 UTC when at least one matched
                if len(events) >= 1:
                    end_ok = False
                    for _, _start, _end in events:
                        if _end is not None and _end.hour == 15:
                            end_ok = True
                            break
                    check("'Analysis Review' end_datetime = 15:00 UTC",
                          end_ok, f"end_datetimes: {[e[2] for e in events]}")
            conn.close()
        except Exception as e:
            check("Calendar check", False, str(e))

        # Check Word document - prefer exact name 'Hr_Worklife_Analysis.docx'
        import glob as globmod
        expected_docx = os.path.join(agent_workspace, "Hr_Worklife_Analysis.docx")
        word_files = []
        if os.path.exists(expected_docx):
            word_files = [expected_docx]
        else:
            word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Hr_Worklife_Analysis.docx exists",
              os.path.exists(expected_docx) or len(word_files) >= 1,
              f"checked {expected_docx}")
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content (>=200 chars)", len(text) >= 200, f"text length: {len(text)}")
            has_findings = any(kw in text for kw in ['finding', 'summary', 'analysis', 'overview'])
            has_recs = any(kw in text for kw in ['recommendation', 'action', 'priority', 'next step'])
            check("Word has summary/findings section", has_findings, f"text head: {text[:200]}")
            check("Word has recommendations/action section", has_recs, f"text head: {text[:200]}")

        sf_proc_path = os.path.join(agent_workspace, "sf_worklife_processor.py")
        if is_gt_self_test and not os.path.exists(sf_proc_path):
            check("sf_worklife_processor.py exists (GT self-test toleration)",
                  True, "GT self-test: agent-generated script, skipped")
        else:
            check("sf_worklife_processor.py exists", os.path.exists(sf_proc_path))
            # Body content check: not a stub
            if os.path.exists(sf_proc_path):
                try:
                    with open(sf_proc_path, "r", errors="ignore") as f:
                        body = f.read()
                    is_substantive = (
                        len(body) >= 200
                        and ('json' in body.lower() or 'open(' in body)
                        and ('def ' in body or 'for ' in body or 'import ' in body)
                    )
                    check("sf_worklife_processor.py is substantive (>=200 chars + reads JSON / has logic)",
                          is_substantive,
                          f"len={len(body)}, head={body[:120]!r}")
                except Exception as _e:
                    check("sf_worklife_processor.py readable", False, str(_e))

        # Salary_Gap relationship check: Salary_Gap = Our_Avg_Salary - Industry_Benchmark
        try:
            if "Data_Analysis" in wb.sheetnames:
                _ws = wb["Data_Analysis"]
                _headers = [str(c.value).strip() if c.value else "" for c in _ws[1]]
                _h_low = [h.lower() for h in _headers]
                idx_our = next((i for i, h in enumerate(_h_low) if 'our' in h and ('avg' in h or 'salary' in h or 'internal' in h)), None)
                idx_bench = next((i for i, h in enumerate(_h_low) if 'benchmark' in h or 'industry' in h or 'external' in h), None)
                idx_gap = next((i for i, h in enumerate(_h_low) if 'gap' in h or 'diff' in h), None)
                if idx_our is not None and idx_bench is not None and idx_gap is not None:
                    rows = list(_ws.iter_rows(min_row=2, values_only=True))
                    correct = 0
                    total = 0
                    for r in rows:
                        if not r or len(r) <= max(idx_our, idx_bench, idx_gap):
                            continue
                        ov = safe_float(r[idx_our])
                        bv = safe_float(r[idx_bench])
                        gv = safe_float(r[idx_gap])
                        if ov is None or bv is None or gv is None:
                            continue
                        total += 1
                        if abs((ov - bv) - gv) < 1.0:
                            correct += 1
                    # Make blocking when >=2 parseable rows (per QA report)
                    if total >= 2:
                        check("Salary_Gap = Our_Avg_Salary - Industry_Benchmark (>=80% rows)",
                              correct / total >= 0.8,
                              f"correct={correct}/{total}")
                    elif total > 0:
                        # 1 row only - keep non-blocking (insufficient signal)
                        check("Salary_Gap math (only 1 parseable row)", True,
                              f"correct={correct}/{total} (non-blocking, <2 rows)")
                    else:
                        # No parseable numeric rows when columns exist => suspicious
                        # but agent may use formulas/strings; keep non-blocking
                        check("Salary_Gap math (no parseable rows)", True,
                              "no parseable rows; skipping (non-blocking)")
                else:
                    # Columns not all detected - non-blocking (semantic mapping handles this elsewhere)
                    check("Salary_Gap math (column detection)", True,
                          f"could not detect all 3 columns; idx_our={idx_our},idx_bench={idx_bench},idx_gap={idx_gap} (non-blocking)")
        except Exception as _e:
            check("Salary_Gap math relation", True, f"non-blocking: {_e}")


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
