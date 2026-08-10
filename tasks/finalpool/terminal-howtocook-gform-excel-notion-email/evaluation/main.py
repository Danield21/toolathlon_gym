"""Evaluation for terminal-howtocook-gform-excel-notion-email.
Checks:
1. Meal_Program_Plan.xlsx with 4 sheets
2. Google Form survey with 5 questions
3. Notion "Recipe Knowledge Base" with menu entries
4. Email sent to all_staff
5. menu_planner.py script exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0
LOCAL_FAIL_COUNT = 0
CURRENT_CATEGORY = "runtime"


def _to_float(v):
    """Parse a numeric value from int/float/str (handles currency symbols, thousand
    separators, percent signs, trailing spaces, and units). Returns None when unparseable."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        low = s.lower()
        if low in ("n/a", "na", "none", "-", "nil", "null"):
            return None
        s2 = s.replace(",", "").replace("$", "").replace("€", "").replace("¥", "")
        s2 = s2.replace("%", "").replace(" ", "")
        for word in ("usd", "cny", "rmb", "yuan", "dollar", "dollars"):
            if s2.lower().endswith(word):
                s2 = s2[:-len(word)].strip()
        if not s2:
            return None
        try:
            return float(s2)
        except ValueError:
            return None
    return None


def _cell_value(cell, wb_cached=None):
    """Numeric value of an openpyxl cell. Formula cells are resolved through the
    cached (data_only=True) workbook; if no cached value exists, returns None so the
    numeric comparison is skipped instead of crashing."""
    if cell is None:
        return None
    v = cell.value
    if isinstance(v, str) and v.strip().startswith("="):
        if wb_cached is not None:
            try:
                cached = wb_cached[cell.parent.title][cell.coordinate].value
            except Exception:
                cached = None
            if cached is not None:
                return _to_float(cached)
        return None
    return _to_float(v)


def _is_total_row(cells):
    """True if a row looks like a summary/total row rather than a data row."""
    for c in cells:
        if c is None or c.value is None:
            continue
        v = str(c.value).strip().lower()
        if any(k in v for k in ("total", "sum", "avg", "average", "subtotal", "合计", "小计")):
            return True
    return False


def _contains_any(text, tokens):
    """True if any token (case-sensitive; pass pre-lowered text) is a substring."""
    return any(t in text for t in tokens)


# Cuisine / dietary question anchors. Broad enough to accept reasonable
# paraphrases (e.g. "Which style of food do you like best?" / "Any allergies
# or dislikes to report?") while still failing surveys that cover neither topic.
CUISINE_TOKENS = ("cuisine", "prefer", "favorite", "favourite",
                  "style of food", "type of food", "of food")
DIETARY_TOKENS = ("dietary", "restriction", "allerg", "dislike", "intolerance",
                  "cannot eat", "can't eat", "avoid", "no pork", "no seafood",
                  "vegetarian", "gluten", "vegan")

# Weekday anchors: English full names / abbreviations plus common Chinese forms.
MONDAY_TOKENS = ("monday", "mon", "周一", "星期一", "礼拜一")
FRIDAY_TOKENS = ("friday", "fri", "周五", "星期五", "礼拜五")

# Cost per serving by recipe category. Exact canonical labels (the Chinese
# categories returned by the HowToCook db and their English translations) match
# first; the substring fallback is ordered most-specific-first so that a
# compound label such as "Vegetable Soup" resolves to the soup cost (4) rather
# than the vegetable cost (5).
CATEGORY_COST_EXACT = {
    "荤菜": 8, "水产": 8, "素菜": 5, "主食": 6, "汤": 4,
    "meat": 8, "meat dish": 8, "meat dishes": 8, "meats": 8,
    "seafood": 8, "fish": 8, "fishes": 8, "aquatic": 8,
    "vegetable": 5, "vegetable dish": 5, "vegetable dishes": 5,
    "vegetables": 5, "veg": 5, "veggie": 5,
    "staple": 6, "staple food": 6, "staple foods": 6, "staples": 6,
    "soup": 4, "soups": 4,
}
CATEGORY_COST_SUBSTR = (
    ("荤菜", 8), ("水产", 8), ("seafood", 8), ("fish", 8), ("meat", 8),
    ("soup", 4), ("汤", 4),
    ("staple", 6), ("主食", 6),
    ("vegetable", 5), ("素菜", 5),
)


def _category_cost(cat):
    """Expected cost per serving for a recipe category, or None if unknown."""
    c = (cat or "").strip().lower()
    if not c:
        return None
    if c in CATEGORY_COST_EXACT:
        return CATEGORY_COST_EXACT[c]
    for k, v in CATEGORY_COST_SUBSTR:
        if k in c:
            return v
    return None


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, LOCAL_FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if CURRENT_CATEGORY == "local":
            LOCAL_FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def check_excel(workspace):
    global CURRENT_CATEGORY
    CURRENT_CATEGORY = "local"
    print("\n=== Check 1: Meal_Program_Plan.xlsx ===")
    path = os.path.join(workspace, "Meal_Program_Plan.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=False)   # keep formulas
    wb_cached = openpyxl.load_workbook(path, data_only=True)  # cached values for formula cells
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Survey_Questions
    sq_idx = next((i for i, s in enumerate(sheets_lower) if "survey" in s or "question" in s), 0)
    ws1 = wb[sheets[sq_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Survey_Questions has 5 rows", len(data1) >= 5, f"Found {len(data1)}")
    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Has cuisine preference question", _contains_any(all_text1, CUISINE_TOKENS),
          f"Text: {all_text1[:100]}")
    check("Has dietary restriction question", _contains_any(all_text1, DIETARY_TOKENS),
          f"Text: {all_text1[:100]}")

    # Recipe_Selection
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "recipe" in s and "select" in s), 1)
    if rs_idx < len(sheets):
        ws2 = wb[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Recipe_Selection has 7+ recipes", len(data2) >= 7, f"Found {len(data2)}")
        if rows2:
            headers = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has difficulty column", any("difficult" in h for h in headers),
                  f"Headers: {rows2[0]}")

    # Weekly_Menu
    wm_idx = next((i for i, s in enumerate(sheets_lower) if "weekly" in s or "menu" in s), 2)
    recipe_to_cat = {}
    recipe_to_diff = {}
    if rs_idx < len(sheets):
        ws2 = wb[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        if rows2:
            for r in rows2[1:]:
                if r and r[0]:
                    try:
                        recipe_to_cat[str(r[0]).strip().lower()] = str(r[1]).strip().lower() if r[1] else ""
                        recipe_to_diff[str(r[0]).strip().lower()] = int(r[2]) if r[2] else None
                    except Exception:
                        pass

    if wm_idx < len(sheets):
        ws3 = wb[sheets[wm_idx]]
        rows3 = list(ws3.iter_rows())  # cell objects so formula cells can be resolved
        data3 = [r for r in rows3[1:] if any(c is not None and c.value is not None for c in r)]
        data3 = [r for r in data3 if not _is_total_row(r)]
        check("Weekly_Menu has at least 5 days", len(data3) >= 5, f"Found {len(data3)}")
        all_text3 = " ".join(str(c.value) for r in rows3 for c in r if c is not None and c.value is not None).lower()
        check("Menu includes Monday", _contains_any(all_text3, MONDAY_TOKENS))
        check("Menu includes Friday", _contains_any(all_text3, FRIDAY_TOKENS))
        # Check cost column has values
        if rows3:
            headers3 = [str(c.value).lower() if c is not None and c.value is not None else "" for c in rows3[0]]
            cost_idx = next((i for i, h in enumerate(headers3) if "cost" in h), -1)
            recipe_idx = next((i for i, h in enumerate(headers3) if "recipe" in h or "lunch" in h), 1)
            if cost_idx >= 0 and data3:
                costs = [r[cost_idx] for r in data3 if r[cost_idx] is not None and r[cost_idx].value is not None]
                check("Cost values present for all 5 days", len(costs) >= 5, f"Found {len(costs)} costs")
                # Validate cost formula: 8 meat/seafood, 5 veg, 6 staple, 4 soup
                # (matches both the Chinese categories returned by the HowToCook db
                # and their English translations)
                cost_errs = 0
                cat_sequence = []
                diff_errs = 0
                for r in data3:
                    rc = r[recipe_idx]
                    rname = str(rc.value).strip().lower() if rc is not None and rc.value is not None else ""
                    cat = recipe_to_cat.get(rname, "")
                    cat_sequence.append(cat)
                    expected_cost = _category_cost(cat)
                    if expected_cost is not None:
                        val = _cell_value(r[cost_idx], wb_cached)
                        if val is not None and abs(val - expected_cost) > 0.01:
                            cost_errs += 1
                    diff_val = recipe_to_diff.get(rname)
                    if diff_val is not None and diff_val > 4:
                        diff_errs += 1
                check("All 5 recipe costs follow 8/5/6/4 formula", cost_errs == 0,
                      f"{cost_errs} cost mismatches")
                check("All menu recipes have difficulty <= 4", diff_errs == 0,
                      f"{diff_errs} recipes exceed difficulty 4")
                # No two consecutive same-category
                consec_err = 0
                for i in range(1, len(cat_sequence)):
                    if cat_sequence[i] and cat_sequence[i] == cat_sequence[i-1]:
                        consec_err += 1
                check("No two consecutive same-category days", consec_err == 0,
                      f"{consec_err} consecutive same-category pairs")

    # Program_Summary
    ps_idx = next((i for i, s in enumerate(sheets_lower) if "program" in s or "summary" in s), 3)
    if ps_idx < len(sheets):
        ws4 = wb[sheets[ps_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Program_Summary has 4+ metrics", len(data4) >= 4, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has total weekly cost", "total" in all_text4 and ("weekly" in all_text4 or "cost" in all_text4))


def check_gform():
    global CURRENT_CATEGORY
    CURRENT_CATEGORY = "runtime"
    print("\n=== Check 2: Google Form Survey ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms ORDER BY id")
        forms = cur.fetchall()
        # Only match forms whose title signals the employee lunch survey (the task fixes the
        # title as "Employee Lunch Program Preferences"), then pick the candidate with the
        # most questions so a noise form (e.g. an archived survey) can never displace it.
        lunch_form = None
        best_q = -1
        for form_id, title in forms:
            t = str(title).lower() if title else ""
            if "lunch" not in t and "preference" not in t:
                continue
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
            q = cur.fetchone()[0]
            if q > best_q:
                best_q = q
                lunch_form = (form_id, title)
        check("Lunch preference survey form exists", lunch_form is not None,
              f"Forms: {[f[1] for f in forms]}")

        if lunch_form:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (lunch_form[0],))
            q_count = cur.fetchone()[0]
            check("Survey has 5 questions", q_count >= 5, f"Found {q_count}")

            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position", (lunch_form[0],))
            questions = cur.fetchall()
            q_text = " ".join(str(q[0]) for q in questions).lower()
            check("Has cuisine question", _contains_any(q_text, CUISINE_TOKENS),
                  f"Questions: {q_text[:100]}")
    except Exception as e:
        check("Gform check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_notion():
    global CURRENT_CATEGORY
    CURRENT_CATEGORY = "runtime"
    print("\n=== Check 3: Notion Recipe Knowledge Base ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM notion.databases")
        dbs = cur.fetchall()
        recipe_db = None
        for db_id, title in dbs:
            title_str = ""
            if isinstance(title, list):
                title_str = " ".join(item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
            elif isinstance(title, str):
                try:
                    parsed = json.loads(title)
                    if isinstance(parsed, list):
                        title_str = " ".join(item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                    else:
                        title_str = str(title)
                except Exception:
                    title_str = str(title)
            else:
                title_str = str(title) if title else ""
            if "recipe" in title_str.lower() and ("knowledge" in title_str.lower() or "base" in title_str.lower()):
                recipe_db = (db_id, title_str)
                break
        check("Recipe Knowledge Base database exists", recipe_db is not None,
              f"Databases: {[d[1] for d in dbs]}")

        if recipe_db:
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (recipe_db[0],))
            count = cur.fetchone()[0]
            check("Knowledge base has 5+ recipe entries", count >= 5, f"Found {count}")
    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    global CURRENT_CATEGORY
    CURRENT_CATEGORY = "runtime"
    print("\n=== Check 4: Email to All Staff ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%all_staff%%'
               OR subject ILIKE '%%lunch program%%'
               OR subject ILIKE '%%lunch%%survey%%'
               OR subject ILIKE '%%employee%%lunch%%'
            ORDER BY id DESC
        """)
        emails = cur.fetchall()
        if not emails:
            cur.execute("""
                SELECT id, subject, to_addr, body_text
                FROM email.drafts
                WHERE to_addr::text ILIKE '%%all_staff%%'
                   OR subject ILIKE '%%lunch%%'
                ORDER BY id DESC
            """)
            emails = cur.fetchall()
        check("Email about lunch program sent", len(emails) >= 1, "No matching email found")
        if emails:
            # Prefer the email whose subject actually mentions the lunch program; a
            # pre-injected noise email (e.g. "Parking Lot Update") must never be picked.
            target = emails[0]
            for e in emails:
                s = str(e[1]).lower() if e[1] else ""
                if "lunch" in s or "survey" in s or "meal" in s:
                    target = e
                    break
            subject = str(target[1]).lower() if target[1] else ""
            check("Email subject mentions lunch or survey",
                  "lunch" in subject or "survey" in subject or "meal" in subject,
                  f"Subject: {target[1]}")
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    global CURRENT_CATEGORY
    CURRENT_CATEGORY = "runtime"
    print("\n=== Check 5: menu_planner.py ===")
    path = os.path.join(workspace, "menu_planner.py")
    check("menu_planner.py exists", os.path.exists(path))


def check_reverse_validation_wrapper(workspace):
    global CURRENT_CATEGORY
    CURRENT_CATEGORY = "local"
    check_reverse_validation(workspace)


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no negative cost values
    path = os.path.join(workspace, "Meal_Program_Plan.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    val = _to_float(cell)
                    if val is not None and val < 0:
                        check("No negative values in Excel", False,
                              f"Found {cell} in sheet {sheet_name}")
                        return
        check("No negative values in Excel", True)

    # Email: no emails to unrelated recipients
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE subject ILIKE '%%lunch%%' AND to_addr::text ILIKE '%%competitor%%'
        """)
        bad_emails = cur.fetchone()[0]
        check("No lunch emails to competitor addresses", bad_emails == 0,
              f"Found {bad_emails}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_notion()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation_wrapper(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Require no local-file failures; runtime (gform/notion/email/script) may fail in GT-only mode
    print(f"Local FAIL_COUNT: {LOCAL_FAIL_COUNT}, Total FAIL_COUNT: {FAIL_COUNT}")
    if LOCAL_FAIL_COUNT == 0 and accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
