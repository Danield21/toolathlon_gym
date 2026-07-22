"""Evaluation for canvas-announcement-summary."""
import argparse
import os
import re
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def date_match(a, b):
    """Compare two date-like values by their YYYY-MM-DD prefix."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    sa = str(a).strip()
    sb = str(b).strip()
    # Extract YYYY-MM-DD if present.
    ma = re.search(r"\d{4}-\d{2}-\d{2}", sa)
    mb = re.search(r"\d{4}-\d{2}-\d{2}", sb)
    if ma and mb:
        return ma.group(0) == mb.group(0)
    return sa == sb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Course_Announcements.xlsx")
    gt_file = os.path.join(gt_dir, "Course_Announcements.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Announcement Stats
    print(f"  Checking Announcement Stats...")
    a_rows = load_sheet_rows(agent_wb, "Announcement Stats")
    g_rows = load_sheet_rows(gt_wb, "Announcement Stats")
    if a_rows is None:
        all_errors.append("Sheet 'Announcement Stats' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Announcement Stats' not found in groundtruth")
    else:
        sheet_name = "Announcement Stats"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Counts must match exactly (integers).
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    errors.append(f"{key}.Announcements: {a_row[1]} vs {g_row[1]} (tol=0)")
            # Earliest_Date (col 2) and Latest_Date (col 3) compared by date.
            if len(a_row) > 2 and len(g_row) > 2:
                if not date_match(a_row[2], g_row[2]):
                    errors.append(f"{key}.Earliest_Date: {a_row[2]} vs {g_row[2]}")
            if len(a_row) > 3 and len(g_row) > 3:
                if not date_match(a_row[3], g_row[3]):
                    errors.append(f"{key}.Latest_Date: {a_row[3]} vs {g_row[3]}")

        # Validate sort order: Announcements column (index 1) must be
        # descending across data rows. (Task: 'Sort by Announcements descending'.)
        a_counts = []
        for row in a_data:
            if row and row[0] is not None and len(row) > 1 and row[1] is not None:
                try:
                    a_counts.append(int(float(row[1])))
                except (TypeError, ValueError):
                    pass
        if len(a_counts) >= 2:
            sort_ok = all(a_counts[i] >= a_counts[i+1] for i in range(len(a_counts)-1))
            if not sort_ok:
                errors.append(f"Announcement Stats not sorted by Announcements desc: counts={a_counts}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                a_val = a_row[1]
                g_val = g_row[1]
                # If GT is numeric (int/float-like), require exact equality.
                # Otherwise (string like 'FFF-2014J'), require exact string match.
                try:
                    g_num = float(g_val)
                    is_num = True
                except (TypeError, ValueError):
                    is_num = False
                if is_num:
                    if not num_close(a_val, g_val, 0):
                        errors.append(f"{key}.Value: {a_val} vs {g_val} (tol=0)")
                else:
                    if not str_match(a_val, g_val):
                        errors.append(f"{key}.Value: {a_val} vs {g_val}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check email deliverable.
    print(f"  Checking email...")
    email_ok = False
    email_detail = ""
    try:
        import psycopg2
        DB_CONFIG = {
            "host": os.environ.get("PGHOST", "localhost"),
            "port": 5432,
            "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
            "user": "eigent", "password": "camel",
        }
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Tighten subject match: require exact subject "Course Announcement Activity Report"
        # (case-insensitive trim, no extra prefix/suffix). To avoid penalising
        # email clients that prepend tags, we accept the exact subject after
        # trimming a leading "Re: " or "Fwd: " (common reply markers).
        target_subj = "Course Announcement Activity Report"
        cur.execute(
            """
            SELECT subject FROM email.messages
            WHERE to_addr::text ILIKE %s
              AND from_addr NOT ILIKE %s
            """,
            (
                "%academic-affairs@openuniversity.ac.uk%",
                "%academic-affairs@openuniversity.ac.uk%",
            ),
        )
        msg_subjects = [row[0] or "" for row in cur.fetchall()]
        sent_subjects = []
        try:
            cur.execute(
                """
                SELECT subject FROM email.sent_log
                WHERE to_addr::text ILIKE %s
                """,
                ("%academic-affairs@openuniversity.ac.uk%",),
            )
            sent_subjects = [row[0] or "" for row in cur.fetchall()]
        except Exception:
            pass
        cur.close()
        conn.close()

        def subject_matches(s):
            ss = (s or "").strip()
            for prefix in ("Re:", "Fwd:", "Fw:"):
                if ss.lower().startswith(prefix.lower()):
                    ss = ss[len(prefix):].strip()
            return ss.lower() == target_subj.lower()

        cnt = sum(1 for s in msg_subjects if subject_matches(s))
        sent = sum(1 for s in sent_subjects if subject_matches(s))
        email_ok = (cnt + sent) >= 1
        email_detail = f"messages_match={cnt}, sent_log_match={sent}, all_subjects={(msg_subjects+sent_subjects)[:5]}"
    except Exception as e:
        email_detail = f"DB error: {e}"
    if email_ok:
        print(f"    PASS ({email_detail})")
    else:
        all_errors.append(f"Email to academic-affairs@openuniversity.ac.uk with exact subject 'Course Announcement Activity Report' missing ({email_detail})")
        print(f"    FAIL: {email_detail}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
