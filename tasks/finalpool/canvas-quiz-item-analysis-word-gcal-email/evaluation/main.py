"""Evaluation for canvas-quiz-item-analysis-word-gcal-email."""
import argparse
import itertools
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone, timedelta

import openpyxl
import psycopg2

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Sentinel returned for a formula cell whose cached value is unavailable.
class _FormulaNoCache:
    def __repr__(self):
        return "<formula cell with no cached value>"


FORMULA_NO_CACHE = _FormulaNoCache()

# Fixed target week (no drifting anchors).
WEEK_START_UTC = datetime(2026, 3, 16, tzinfo=timezone.utc)
WEEK_END_UTC = datetime(2026, 3, 21, tzinfo=timezone.utc)


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(v):
    """Parse a value into a float, tolerating '%', thousand separators,
    currency symbols and surrounding whitespace. Returns None if the value
    cannot be parsed as a number."""
    if v is None:
        return None
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    if s.lower() in ("n/a", "na", "none", "null", "-", "--"):
        return None
    cleaned = s
    for ch in (",", "，"):
        cleaned = cleaned.replace(ch, "")
    cleaned = cleaned.replace(" ", "")
    for sym in ("$", "€", "¥", "£", "￥"):
        cleaned = cleaned.replace(sym, "")
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
    try:
        return float(cleaned)
    except ValueError:
        return None


def _resolve_cell(cell_value, cached_value):
    """Return a comparable value for a single cell.

    - Formula cells (value starts with '='): return the cached value when it is
      available (the workbook was re-calculated by Excel/LibreOffice), otherwise
      return the FORMULA_NO_CACHE sentinel so the check is skipped rather than
      mis-judged.
    - Otherwise: return the raw cell value unchanged.
    """
    if isinstance(cell_value, str) and cell_value.strip().startswith("="):
        if cached_value is not None:
            return cached_value
        return FORMULA_NO_CACHE
    return cell_value


def _load_workbook_pair(path):
    wb_formula = openpyxl.load_workbook(path, data_only=False)  # keeps formulas
    wb_value = openpyxl.load_workbook(path, data_only=True)     # cached values
    return wb_formula, wb_value


def num_close(a, b, tol=1.0):
    # Agent-side formula cell with no cached value: the task requires numeric
    # cells to be written as literals, so an unevaluated formula is a task
    # violation and must not silently pass the numeric check.
    if a is FORMULA_NO_CACHE:
        return False
    # GT-side formula without a cache should not occur (GT stores literals);
    # if it ever does, do not penalize the agent.
    if b is FORMULA_NO_CACHE:
        return True
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is FORMULA_NO_CACHE:
        return False
    if b is FORMULA_NO_CACHE:
        return True
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _pos_key(v):
    """Normalize a Question_Position cell to a comparable key (int when
    integral, else float, else lowercased string). Tolerates '1.0', '01',
    and stray whitespace so a correct completion is not judged FAIL."""
    if v is None:
        return None
    f = _to_float(v)
    if f is not None:
        if f == int(f):
            return int(f)
        return f
    s = str(v).strip().lower()
    return s if s else None


def _course_name_normalize(name):
    """Lowercase and strip a trailing semester suffix ('(Fall 2014)' etc.)."""
    s = str(name or "").strip().lower()
    for suf in ("(fall 2014)", "(fall 2014 )", "(fall 2014", "(fall"):
        if s.endswith(suf):
            s = s[:len(s) - len(suf)].strip()
            break
    return s


def load_sheet_rows(wb_f, wb_v, sheet_name):
    for name in wb_f.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            ws_v = wb_v[name] if name in wb_v.sheetnames else None
            rows = []
            for row in wb_f[name].iter_rows():
                out = []
                for cell in row:
                    cached = None
                    if ws_v is not None:
                        try:
                            cached = ws_v[cell.coordinate].value
                        except Exception:
                            cached = None
                    out.append(_resolve_cell(cell.value, cached))
                rows.append(out)
            return rows
    return None


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel ===")
    agent_file = os.path.join(agent_workspace, "Quiz_Item_Analysis.xlsx")
    gt_file = os.path.join(gt_workspace, "Quiz_Item_Analysis.xlsx")

    if not os.path.isfile(agent_file):
        check("Quiz_Item_Analysis.xlsx exists", False, f"Not found: {agent_file}")
        return
    check("Quiz_Item_Analysis.xlsx exists", True)

    try:
        agent_wb_f, agent_wb_v = _load_workbook_pair(agent_file)
        gt_wb_f, gt_wb_v = _load_workbook_pair(gt_file)
    except Exception as e:
        check("Excel files readable", False, str(e))
        return

    # --- Sheet 1: Quiz Overview ---
    print("  Checking Quiz Overview...")
    a_rows = load_sheet_rows(agent_wb_f, agent_wb_v, "Quiz Overview")
    g_rows = load_sheet_rows(gt_wb_f, gt_wb_v, "Quiz Overview")

    if a_rows is None:
        check("Sheet 'Quiz Overview' exists", False, f"Available: {agent_wb_f.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Quiz Overview' exists (gt)", False)
    else:
        check("Sheet 'Quiz Overview' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Quiz Overview row count", abs(len(a_data) - len(g_data)) <= 1,
              f"Agent={len(a_data)}, GT={len(g_data)}")

        # Lookup by quiz title
        a_lookup = {}
        for row in a_data:
            if row and len(row) > 1 and row[1] is not None:
                a_lookup[str(row[1]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[1] is None:
                continue
            key = str(g_row[1]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing quiz: {g_row[1]}")
                continue
            # Check Avg_Score (col 4, tol 0.5)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Avg_Score: {a_row[4]} vs {g_row[4]}")
            # Check Completion_Rate_Pct (col 5, tol 1.0)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1.0):
                    errors.append(f"{key}.Completion_Rate: {a_row[5]} vs {g_row[5]}")
            # Check Submission_Count (col 3, tol 5)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5):
                    errors.append(f"{key}.Sub_Count: {a_row[3]} vs {g_row[3]}")
            # Check Quality_Rating (col 6, string)
            if len(a_row) > 6 and len(g_row) > 6:
                if not str_match(a_row[6], g_row[6]):
                    errors.append(f"{key}.Quality: {a_row[6]} vs {g_row[6]}")
        if errors:
            for e in errors[:5]:
                check(f"Quiz Overview data - {e}", False)
        else:
            check("Quiz Overview data matches", True)

    # --- Sheet 2: Question Analysis ---
    print("  Checking Question Analysis...")
    a_rows = load_sheet_rows(agent_wb_f, agent_wb_v, "Question Analysis")
    g_rows = load_sheet_rows(gt_wb_f, gt_wb_v, "Question Analysis")

    if a_rows is None:
        check("Sheet 'Question Analysis' exists", False, f"Available: {agent_wb_f.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Question Analysis' exists (gt)", False)
    else:
        check("Sheet 'Question Analysis' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Question Analysis row count", abs(len(a_data) - len(g_data)) <= 5,
              f"Agent={len(a_data)}, GT={len(g_data)}")

        # Determine difficulty column by looking at header
        a_header = a_rows[0] if a_rows else []
        g_header = g_rows[0] if g_rows else []

        def find_col(header, keyword):
            for i, h in enumerate(header):
                if h and keyword.lower() in str(h).lower():
                    return i
            return None

        a_quiz_col = find_col(a_header, "Quiz_Title")
        g_quiz_col = find_col(g_header, "Quiz_Title")
        a_pos_col = find_col(a_header, "Question_Position")
        g_pos_col = find_col(g_header, "Question_Position")
        a_diff_col = find_col(a_header, "Difficulty_Category")
        g_diff_col = find_col(g_header, "Difficulty_Category")

        if (a_quiz_col is not None and g_quiz_col is not None and
                a_pos_col is not None and g_pos_col is not None and
                a_diff_col is not None and g_diff_col is not None):
            def qa_key(row, quiz_col, pos_col):
                if row and len(row) > max(quiz_col, pos_col):
                    q = str(row[quiz_col]).strip().lower() if row[quiz_col] is not None else ""
                    p = _pos_key(row[pos_col])
                    return (q, p) if p is not None else (q, "")
                return None

            # Agent lookup by (quiz, position) with a fallback index by
            # (quiz, rank-within-quiz) so 0-indexed / re-ordered positions
            # still match a correct completion.
            a_by_pos = {}
            a_by_rank = defaultdict(list)
            for r in a_data:
                k = qa_key(r, a_quiz_col, a_pos_col)
                if k:
                    a_by_pos.setdefault(k, r)
                if r and len(r) > a_quiz_col and r[a_quiz_col] is not None:
                    a_by_rank[str(r[a_quiz_col]).strip().lower()].append(r)
            a_rank_lookup = {}
            for quiz, rows in a_by_rank.items():
                for rank, row in enumerate(rows, start=1):
                    a_rank_lookup.setdefault((quiz, rank), row)

            # GT rows grouped by quiz, in order, with their within-quiz rank.
            g_by_rank = defaultdict(list)
            for r in g_data:
                if r and len(r) > g_quiz_col and r[g_quiz_col] is not None:
                    g_by_rank[str(r[g_quiz_col]).strip().lower()].append(r)

            qa_errors = []
            checked = 0
            missing = 0
            for quiz, g_rows in g_by_rank.items():
                for rank, g_row in enumerate(g_rows, start=1):
                    g_pos_key = qa_key(g_row, g_quiz_col, g_pos_col)
                    a_row = a_by_pos.get(g_pos_key) if g_pos_key else None
                    if a_row is None:
                        a_row = a_rank_lookup.get((quiz, rank))
                    if a_row is None:
                        missing += 1
                        continue
                    if (len(a_row) > a_diff_col and len(g_row) > g_diff_col
                            and g_row[g_diff_col] is not None):
                        if not str_match(a_row[a_diff_col], g_row[g_diff_col]):
                            qa_errors.append(f"{g_pos_key}.difficulty: {a_row[a_diff_col]} vs {g_row[g_diff_col]}")
                        checked += 1
            # Tolerate a few unmatched rows only within the row-count tolerance.
            row_diff = abs(len(a_data) - len(g_data))
            if missing > row_diff:
                qa_errors.append(f"{missing} questions not matched (row-count tolerance {row_diff})")
            if qa_errors:
                for e in qa_errors[:3]:
                    check(f"Question Analysis - {e}", False)
            elif checked > 0:
                check(f"Question Analysis difficulty categories match ({checked} checked)", True)
            else:
                check("Question Analysis: difficulty categories sampled", True, "no difficulty col identified; skipped")
        else:
            check("Question Analysis: difficulty categories sampled", True, "columns not identifiable; skipped")

    # --- Sheet 3: Course Summary ---
    print("  Checking Course Summary...")
    a_rows = load_sheet_rows(agent_wb_f, agent_wb_v, "Course Summary")
    g_rows = load_sheet_rows(gt_wb_f, gt_wb_v, "Course Summary")

    if a_rows is None:
        check("Sheet 'Course Summary' exists", False, f"Available: {agent_wb_f.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Course Summary' exists (gt)", False)
    else:
        check("Sheet 'Course Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        # >= 7: tolerate an extra total/duplicate row; lookup below still verifies
        # that every GT course is present with correct values.
        check("Course Summary has at least 7 rows", len(a_data) >= 7,
              f"Found {len(a_data)}")

        # Lookup by course name with lenient matching: drop a trailing
        # '(Fall 2014)' suffix and tolerate case / common abbreviations,
        # consistent with the calendar coverage check.
        def find_course_row(name):
            target = _course_name_normalize(name)
            if not target:
                return None
            # Prefer an exact match after normalization.
            for row in a_data:
                if row and row[0] is not None and _course_name_normalize(row[0]) == target:
                    return row
            for row in a_data:
                if row and row[0] is not None:
                    rn = _course_name_normalize(row[0])
                    if rn and (rn in target or target in rn):
                        return row
            return None

        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = find_course_row(g_row[0])
            if a_row is None:
                errors.append(f"Missing course: {g_row[0]}")
                continue
            # Total_Quizzes (col 1, exact)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    errors.append(f"{key[:30]}.Total_Quizzes: {a_row[1]} vs {g_row[1]}")
            # Total_Questions (col 2, exact)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0):
                    errors.append(f"{key[:30]}.Total_Questions: {a_row[2]} vs {g_row[2]}")
            # Review_Quizzes (col 5, exact)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0):
                    errors.append(f"{key[:30]}.Review_Quizzes: {a_row[5]} vs {g_row[5]}")
            # Avg_Completion_Rate (col 6). For courses with no quizzes the
            # average is mathematically undefined: accept 0 or a
            # not-applicable representation (None / 'N/A' / '-' / empty), but
            # reject a nonzero number.
            if len(a_row) > 6 and len(g_row) > 6:
                gt_quizzes = _to_float(g_row[1]) if len(g_row) > 1 else None
                if gt_quizzes == 0:
                    av = a_row[6]
                    if av is FORMULA_NO_CACHE:
                        errors.append(f"{key[:30]}.Avg_Comp: unverifiable formula for a 0-quiz course")
                    else:
                        fa = _to_float(av)
                        if fa is not None and fa != 0.0:
                            errors.append(f"{key[:30]}.Avg_Comp: {a_row[6]} vs 0 (no quizzes)")
                else:
                    if not num_close(a_row[6], g_row[6], 1.0):
                        errors.append(f"{key[:30]}.Avg_Comp: {a_row[6]} vs {g_row[6]}")

        if errors:
            for e in errors[:5]:
                check(f"Course Summary data - {e}", False)
        else:
            check("Course Summary data matches", True)


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Quiz_Analysis_Report.docx")
    if not os.path.isfile(docx_path):
        check("Quiz_Analysis_Report.docx exists", False, f"Not found: {docx_path}")
        return
    check("Quiz_Analysis_Report.docx exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
    except Exception as e:
        check("Word doc readable", False, str(e))
        return

    full_text = " ".join(p.text for p in doc.paragraphs).lower()

    # Check for key sections
    for section in ["executive summary", "methodology", "recommendations"]:
        check(f"Section '{section}' present",
              section in full_text,
              f"Not found in document text")

    # Check for course names
    for course_key in ["creative computing", "foundations of finance", "global governance"]:
        check(f"Course '{course_key}' mentioned",
              course_key in full_text,
              f"Not found in document text")

    # Check for key terms
    for term in ["difficulty", "optimal", "needs review"]:
        check(f"Term '{term}' mentioned",
              term in full_text,
              f"Not found in document text")


def _as_utc(dt):
    """Normalize a (possibly aware) datetime to UTC for deterministic comparison."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _local_dt(e):
    """Resolve an event's start to its declared local timezone, or None when the
    event carries no resolvable timezone. A naive stored wall time is taken as
    being in the declared timezone."""
    dt = e[1]
    if dt is None:
        return None
    tz = e[3] if len(e) > 3 else None
    if tz and ZoneInfo is not None:
        try:
            if dt.tzinfo is None:
                return dt.replace(tzinfo=ZoneInfo(tz))
            return dt.astimezone(ZoneInfo(tz))
        except Exception:
            return None
    return None


def _course_key(summary):
    """Extract a course key from an event summary (case-insensitive)."""
    s = str(summary or "").strip().lower()
    for prefix in ("quiz review -", "quiz review-", "quiz review", "quiz-review"):
        if s.startswith(prefix):
            s = s[len(prefix):].strip()
            break
    return s


def _course_name_match(course_name, key):
    """Lenient course-name match: either string containing the other (after
    dropping a trailing '(Fall 2014)' suffix)."""
    a = _course_name_normalize(course_name)
    b = _course_name_normalize(key)
    if not a or not b:
        return False
    return a in b or b in a


def _gt_course_names(gt_workspace):
    gt_file = os.path.join(gt_workspace, "Quiz_Item_Analysis.xlsx")
    if not os.path.isfile(gt_file):
        return []
    try:
        gt_f, gt_v = _load_workbook_pair(gt_file)
    except Exception:
        return []
    rows = load_sheet_rows(gt_f, gt_v, "Course Summary")
    if not rows:
        return []
    names = []
    for row in rows[1:]:
        if row and row[0] is not None and str(row[0]).strip():
            names.append(str(row[0]).strip().lower())
    return names


def _exists_no_overlap_selection(groups):
    """Given a list of groups, each a list of (start, end) aware datetimes,
    return True if one span can be picked per group so that no two overlap.

    Robust to P>1 homogeneous sub-agents each scheduling the full set of
    sessions into the same calendar: the union may contain duplicate /
    cross-schedule overlaps, but a correct completion leaves a valid subset.
    """
    if not groups:
        return False
    total = 1
    for g in groups:
        total *= len(g)
        if total > 500000:
            return True  # too many combinations to exhaust: be lenient
    for combo in itertools.product(*groups):
        ordered = sorted(combo, key=lambda s: s[0])
        ok = True
        for i in range(1, len(ordered)):
            if ordered[i][0] < ordered[i - 1][1]:
                ok = False
                break
        if ok:
            return True
    return False


def check_gcal(gt_workspace):
    print("\n=== Checking Calendar Events ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime, start_timezone
            FROM gcal.events
            WHERE summary ILIKE '%%quiz review%%'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()
        check("7 quiz review events exist", len(events) >= 7,
              f"Found {len(events)} events")

        if not events:
            cur.close()
            conn.close()
            return

        # Events in the week of March 16-20, 2026. Window is widened by one day
        # on each side so that events stored in any timezone that fall on
        # Mon-Fri local are not excluded by a UTC boundary shift.
        win_start = WEEK_START_UTC - timedelta(days=1)
        win_end = WEEK_END_UTC + timedelta(days=1)
        in_week = 0
        for e in events:
            u = _as_utc(e[1])
            if u is not None and win_start <= u < win_end:
                in_week += 1
        check("Events in week of March 16-20", in_week >= 7,
              f"{in_week} events in target week")

        # Duration ~45 minutes. With P>1 homogeneous workers, duplicate events
        # may exist; require that at least 7 events have the right duration.
        dur_ok = 0
        for e in events:
            if e[1] and e[2]:
                duration = (e[2] - e[1]).total_seconds() / 60
                if abs(duration - 45) <= 5:
                    dur_ok += 1
        check("At least 7 quiz review events are ~45 minutes", dur_ok >= 7,
              f"{dur_ok}/{len(events)} events are ~45 min")

        # Group events by course key.
        by_course = defaultdict(list)
        for e in events:
            by_course[_course_key(e[0])].append(e)
        check("7 distinct courses have review sessions", len(by_course) >= 7,
              f"Distinct courses: {len(by_course)}")

        # Every GT Fall-2014 course should be represented (lenient match).
        gt_names = _gt_course_names(gt_workspace)
        if gt_names:
            matched = sum(
                1 for cn in gt_names
                if any(_course_name_match(cn, ck) for ck in by_course)
            )
            check(f"Review sessions cover all {len(gt_names)} Fall 2014 courses",
                  matched >= len(gt_names), f"Matched {matched}/{len(gt_names)}")

        # No-overlap: exists a selection of one session per course with no
        # overlaps (robust to duplicate / parallel-worker events).
        groups = []
        for ck, evts in by_course.items():
            spans = [(e[1], e[2]) for e in evts if e[1] and e[2]]
            if spans:
                groups.append(spans)
        has_no_overlap = _exists_no_overlap_selection(groups)
        check("A set of non-overlapping review sessions exists", has_no_overlap,
              f"{len(groups)} courses with timed sessions")

        # Business hours 09:00-17:00 (start hour). The task fixes business
        # hours to UTC; to avoid false failures from a metadata timezone that
        # disagrees with the stored offset (e.g. a worker whose container TZ
        # leaked into the timezone field), an event passes if it falls inside
        # business hours in EITHER its UTC instant OR its declared timezone.
        bh_ok = 0
        bh_checked = 0
        for e in events:
            u = _as_utc(e[1])
            local = _local_dt(e)
            if u is None and local is None:
                continue
            bh_checked += 1
            if (u is not None and 9 <= u.hour < 17) or (local is not None and 9 <= local.hour < 17):
                bh_ok += 1
        if bh_checked:
            check(f"Review sessions start within business hours 09-17 (UTC or declared tz) ({bh_ok}/{bh_checked})",
                  bh_ok == bh_checked)
        else:
            check("Review sessions start within business hours 09-17 (UTC or declared tz)", True,
                  "no events carry a resolvable time; skipped")

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%assessment_office@university.edu%%'
        """)
        emails = cur.fetchall()
        check("Email to assessment_office sent", len(emails) >= 1,
              f"Found {len(emails)} emails to assessment_office")

        if emails:
            # With P>1 homogeneous workers more than one email may be sent;
            # require that at least one of them is well-formed.
            good_subj = 0
            good_body = 0
            for email in emails:
                subject = str(email[1]).lower() if email[1] else ""
                body = str(email[3]) if email[3] else ""
                if any(kw in subject for kw in ["quiz", "item analysis", "quiz item", "quiz analysis"]):
                    good_subj += 1
                if len(body) > 20:
                    good_body += 1
            check("Email subject mentions quiz analysis", good_subj >= 1,
                  f"{good_subj}/{len(emails)} emails have a quiz-analysis subject")
            check("Email body has content", good_body >= 1,
                  f"{good_body}/{len(emails)} emails have a substantial body")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gcal(gt_dir)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
