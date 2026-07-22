"""Evaluation for terminal-wc-excel-gsheet-gcal-email.
Checks:
1. Inventory_Lifecycle_Report.xlsx with 4 sheets
2. Google Sheet "Inventory Dashboard"
3. Calendar event for restock review
4. Email sent to purchasing team
5. demand_forecast.py script exists
"""
import argparse
import json
import os
import sys
from datetime import datetime

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def norm(s):
    return str(s or "").strip().lower().replace("_", " ")


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _fetch_expected_product_count():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM wc.products")
        n = cur.fetchone()[0]
        conn.close()
        return int(n)
    except Exception as e:
        print(f"[WARN] product count fetch: {e}")
        return None


def check_excel(workspace, gt_workspace):
    print("\n=== Check 1: Inventory_Lifecycle_Report.xlsx ===")
    path = os.path.join(workspace, "Inventory_Lifecycle_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel parses", False, str(e))
        return
    check("Excel parses", True)

    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheet_norm = {norm(s): s for s in sheets}
    expected_pc = _fetch_expected_product_count() or 82

    # Product_Inventory
    pi_key = sheet_norm.get(norm("Product_Inventory"))
    check("Product_Inventory sheet exists (exact)", pi_key is not None, f"sheets={sheets}")
    if pi_key:
        ws1 = wb[pi_key]
        rows1 = list(ws1.iter_rows(values_only=True))
        data1 = [r for r in rows1[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        # Row count must equal product catalog (allow ±2 for borderline filters)
        check(f"Product_Inventory has {expected_pc} rows (±2)",
              abs(len(data1) - expected_pc) <= 2, f"Found {len(data1)}, expected ~{expected_pc}")
        if rows1:
            headers = [norm(c) for c in rows1[0]]
            for h in ["product name", "category", "price", "stock qty", "stock status", "total sales", "days of supply"]:
                check(f"PI has '{h}' column", h in headers, f"headers={headers}")
            # Check sorted by days_of_supply ascending
            if "days of supply" in headers:
                dos_idx = headers.index("days of supply")
                vals = [safe_float(r[dos_idx]) for r in data1 if r[dos_idx] is not None]
                # Filter Nones; check nondecreasing
                if vals:
                    check("PI sorted by days_of_supply asc",
                          vals == sorted(vals), f"first 5: {vals[:5]}")

    # Reorder_Alerts
    ra_key = sheet_norm.get(norm("Reorder_Alerts"))
    check("Reorder_Alerts sheet exists (exact)", ra_key is not None)
    if ra_key:
        ws2 = wb[ra_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        if rows2:
            headers = [norm(c) for c in rows2[0]]
            for h in ["product name", "current stock", "reorder point", "urgency"]:
                check(f"RA has '{h}' column", h in headers, f"headers={headers}")
        # Each row's urgency must be Critical or Out_of_Stock
        if data2 and rows2:
            headers = [norm(c) for c in rows2[0]]
            if "urgency" in headers:
                ui = headers.index("urgency")
                bad = [r[ui] for r in data2 if norm(r[ui]) not in ("critical", "out of stock")]
                check("All RA rows have Critical or Out_of_Stock urgency",
                      len(bad) == 0, f"bad urgencies: {bad[:5]}")
            # Out_of_Stock first
            if "urgency" in headers:
                ui = headers.index("urgency")
                urgencies = [norm(r[ui]) for r in data2]
                # All Out_of_Stock should appear before any Critical
                first_critical = next((i for i, u in enumerate(urgencies) if u == "critical"), None)
                last_oos = next((i for i in range(len(urgencies)-1, -1, -1) if urgencies[i] == "out of stock"), -1)
                check("RA sorted Out_of_Stock first then Critical",
                      first_critical is None or last_oos < first_critical,
                      f"first_critical={first_critical}, last_oos={last_oos}")
        # Has at least 1 entry
        check("RA has >= 1 entry", len(data2) >= 1, f"got {len(data2)}")

    # Category_Summary
    cs_key = sheet_norm.get(norm("Category_Summary"))
    check("Category_Summary sheet exists (exact)", cs_key is not None)
    if cs_key:
        ws3 = wb[cs_key]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        if rows3:
            headers = [norm(c) for c in rows3[0]]
            for h in ["category", "product count", "avg stock", "total value"]:
                check(f"CS has '{h}' column", h in headers, f"headers={headers}")
            # Validate sum of product_counts equals total products
            if "product count" in headers:
                pc_idx = headers.index("product count")
                total = sum(safe_float(r[pc_idx]) or 0 for r in data3)
                check(f"CS product_count sum == {expected_pc}",
                      abs(total - expected_pc) <= 2, f"got {total}")

    # Restock_Schedule
    rs_key = sheet_norm.get(norm("Restock_Schedule"))
    check("Restock_Schedule sheet exists (exact)", rs_key is not None)
    if rs_key:
        ws4 = wb[rs_key]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        if rows4:
            headers = [norm(c) for c in rows4[0]]
            for h in ["product name", "reorder date", "quantity", "supplier"]:
                check(f"RS has '{h}' column", h in headers, f"headers={headers}")
        if rows4 and data4:
            headers = [norm(c) for c in rows4[0]]
            # Validate all reorder_dates == 2026-03-10
            if "reorder date" in headers:
                rd_idx = headers.index("reorder date")
                bad_dates = []
                for r in data4:
                    v = r[rd_idx]
                    s = ""
                    if hasattr(v, "strftime"):
                        s = v.strftime("%Y-%m-%d")
                    else:
                        s = str(v).strip()
                    if not s.startswith("2026-03-10"):
                        bad_dates.append(s)
                check("All reorder_dates == 2026-03-10",
                      len(bad_dates) == 0, f"bad: {bad_dates[:3]}")
            # Validate suppliers
            if "supplier" in headers:
                sup_idx = headers.index("supplier")
                bad = [r[sup_idx] for r in data4 if norm(r[sup_idx]) != norm("Primary Supplier")]
                check("All suppliers == 'Primary Supplier'",
                      len(bad) == 0, f"bad: {bad[:3]}")
            # Quantity must be >= 1
            if "quantity" in headers:
                q_idx = headers.index("quantity")
                bad = [safe_float(r[q_idx]) for r in data4 if (safe_float(r[q_idx]) or 0) < 1]
                check("All quantities >= 1", len(bad) == 0, f"bad: {bad[:3]}")
        check("RS has >= 1 entry", len(data4) >= 1, f"got {len(data4)}")


def check_gsheet():
    print("\n=== Check 2: Google Sheet Inventory Dashboard ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()
        # Require exact match (case-insensitive)
        dashboard = None
        for ss_id, title in spreadsheets:
            if title and norm(title) == norm("Inventory Dashboard"):
                dashboard = (ss_id, title)
                break
        check("Inventory Dashboard spreadsheet exists (exact title)", dashboard is not None,
              f"Spreadsheets: {[s[1] for s in spreadsheets]}")

        if dashboard:
            cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (dashboard[0],))
            cell_count = cur.fetchone()[0]
            check("Dashboard has >= 8 data cells", cell_count >= 8, f"Found {cell_count} cells")
            # Check at least some category names appear in cells
            cur.execute("""SELECT value FROM gsheet.cells
                           WHERE spreadsheet_id = %s AND value IS NOT NULL""", (dashboard[0],))
            cell_vals = [str(r[0]).lower() for r in cur.fetchall()]
            joined = " ".join(cell_vals)
            check("Dashboard mentions 'category' header",
                  "category" in joined, f"sample: {joined[:200]}")
            # Validate that real WC product category names appear in the dashboard
            try:
                cur.execute(
                    "SELECT DISTINCT jsonb_array_elements(categories::jsonb)->>'name' "
                    "AS cat_name FROM wc.products WHERE categories IS NOT NULL"
                )
                db_cats = [r[0] for r in cur.fetchall() if r[0]]
            except Exception:
                db_cats = []
            if db_cats:
                matched = [c for c in db_cats if c.lower() in joined]
                # Require at least 4 distinct DB category names to appear in cells
                threshold = min(4, len(db_cats))
                check(
                    f"Dashboard contains >= {threshold} real category names",
                    len(matched) >= threshold,
                    f"matched={matched}, db_cats={db_cats}",
                )
        cur.close()
        conn.close()
    except Exception as e:
        check("Gsheet check", False, str(e))


def check_gcal():
    print("\n=== Check 3: Calendar Restock Meeting ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        target_event = None
        for summary, desc, start, end in events:
            if not summary:
                continue
            if norm(summary) == norm("Inventory Restock Review Meeting"):
                target_event = (summary, desc, start, end)
                break
        check("'Inventory Restock Review Meeting' event exists", target_event is not None,
              f"Events: {[e[0] for e in events]}")
        if target_event:
            summary, desc, start, end = target_event
            # Date == 2026-03-11, time 10:00 AM
            ok_date = False
            ok_time = False
            ok_dur = False
            try:
                if isinstance(start, datetime):
                    sd = start
                else:
                    sd = datetime.fromisoformat(str(start).replace("Z", "+00:00"))
                ok_date = (sd.year == 2026 and sd.month == 3 and sd.day == 11)
                ok_time = (sd.hour == 10 and sd.minute == 0)
                if isinstance(end, datetime):
                    ed = end
                elif end is not None:
                    ed = datetime.fromisoformat(str(end).replace("Z", "+00:00"))
                else:
                    ed = None
                if ed:
                    dur = (ed - sd).total_seconds()
                    ok_dur = abs(dur - 3600) < 300  # 1 hour ± 5 min
            except Exception as e:
                print(f"[WARN] event time parse: {e}")
            check("Event date == 2026-03-11", ok_date, f"start={start}")
            check("Event start time == 10:00", ok_time, f"start={start}")
            check("Event duration ~ 1 hour", ok_dur, f"start={start}, end={end}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Gcal check", False, str(e))


def check_email():
    print("\n=== Check 4: Email to Purchasing Team ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
        """)
        all_msgs = cur.fetchall()
        # Match by exact subject + to_addr containing purchasing@company.com
        matched = []
        for mid, subj, to_addr, body in all_msgs:
            if not subj:
                continue
            ok_subj = norm(subj) == norm("Critical Inventory Alert - Restock Required")
            to_s = str(to_addr or "").lower()
            ok_to = "purchasing@company.com" in to_s
            if ok_subj and ok_to:
                matched.append((mid, subj, to_addr, body))
        check("Email sent: exact subject + recipient",
              len(matched) >= 1,
              f"all subjects: {[r[1] for r in all_msgs]}")
        if matched:
            body = str(matched[0][3] or "").lower()
            # Body should mention out-of-stock and critical numbers
            ok_oos = "out" in body and "stock" in body
            ok_crit = "critical" in body
            check("Email body mentions out-of-stock", ok_oos, f"body[:200]: {body[:200]}")
            check("Email body mentions critical", ok_crit, f"body[:200]: {body[:200]}")
            # Require at least one numeric value in the body indicating count of items
            import re as _re
            has_number = bool(_re.search(r"\b\d+\b", body))
            check(
                "Email body contains numeric count (e.g., out-of-stock count)",
                has_number,
                f"body[:200]: {body[:200]}",
            )
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_script(workspace):
    print("\n=== Check 5: demand_forecast.py ===")
    path = os.path.join(workspace, "demand_forecast.py")
    exists = os.path.exists(path)
    check("demand_forecast.py exists", exists)
    if exists:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                src = f.read().lower()
        except Exception as e:
            check("demand_forecast.py readable", False, str(e))
            return
        # Script should reference forecast/demand-related logic
        check(
            "demand_forecast.py references forecast or demand",
            ("forecast" in src) or ("demand" in src),
            f"script length={len(src)}",
        )


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Inventory_Lifecycle_Report.xlsx")
    if os.path.exists(path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            return
        # Stock quantities should not be negative
        sheet_norm = {norm(s): s for s in wb.sheetnames}
        pi_key = sheet_norm.get(norm("Product_Inventory"))
        if pi_key:
            ws = wb[pi_key]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [norm(c) for c in rows[0]]
                if "stock qty" in headers:
                    sq_idx = headers.index("stock qty")
                    has_negative = False
                    for r in rows[1:]:
                        v = safe_float(r[sq_idx])
                        if v is not None and v < 0:
                            has_negative = True
                            break
                    check("No negative stock quantities", not has_negative,
                          "Found negative stock value")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gsheet()
    check_gcal()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
