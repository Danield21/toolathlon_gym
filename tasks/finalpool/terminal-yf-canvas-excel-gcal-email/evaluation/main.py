"""Evaluation for terminal-yf-canvas-excel-gcal-email.

Checks:
1. Financial_Literacy_Workshops.xlsx with 3 sheets (Student_Tiers, Market_Events, Workshop_Schedule)
2. Google Calendar events for 3 workshops
3. Emails sent to finance_students and department_head
4. workshop_materials.txt exists
5. market_events.json exists
6. categorize_students.py and find_market_events.py scripts exist
"""
import argparse
import json
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import psycopg2

# All DB settings come from environment variables with local defaults so the
# evaluator connects to the same database the agent used (the harness injects
# PGHOST/PGPORT/PGDATABASE per task container).
DB = dict(host=os.environ.get("PGHOST", "localhost"),
          port=int(os.environ.get("PGPORT", "5432")),
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user=os.environ.get("PGUSER", "eigent"),
          password=os.environ.get("PGPASSWORD", "camel"))

PASS_COUNT = 0
FAIL_COUNT = 0

# Sentinel marking a cell that is an uncalculated Excel formula (its cached
# value is None because the workbook was never opened/recalculated by a
# spreadsheet application). We cannot numerically assess such a cell, so the
# corresponding check is skipped instead of failing the agent.
UNEVAL = object()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def _to_float(v):
    """Robustly convert a value to float.

    Supports int/float, and strings with thousands separators, currency
    symbols, a trailing '%' or leading/trailing whitespace. Returns None when
    the value cannot be parsed (including None and uncalculated formulas).
    """
    if v is None or v is UNEVAL:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        for ch in (",", "$", "€", "¥", "£", "%", " ", "\t", "\n"):
            s = s.replace(ch, "")
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _to_int(v):
    f = _to_float(v)
    if f is None:
        return None
    return int(f)


def num_close(a, b, tol=2.0):
    """Compare two values numerically (with a tolerance), falling back to a
    case-insensitive string comparison only when a numeric comparison is not
    possible on one side (per R3: parse both, compare; otherwise string)."""
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # One side unparseable -> tolerant string comparison so e.g. a stray
    # label/unit in the agent cell does not produce a hard FAIL by itself.
    return str(a).strip().lower() == str(b).strip().lower()


def check_num(name, a, b, tol=2.0):
    """Emit a check for a numeric comparison.

    An uncalculated Excel formula cell (UNEVAL) FAILS rather than being
    skipped: docs/task.md requires every number to be written as a literal
    value ("do not use Excel formulas"), so a formula cell whose value was
    never calculated is a task violation, not a value we can grade. This keeps
    the numeric checks from being silently bypassed (a model writing formulas
    everywhere must not sail through). A formula with a cached value resolves
    to that value and is graded normally, so a recalculated workbook is still
    tolerated.
    """
    if a is UNEVAL or b is UNEVAL:
        check(name, False,
              f"expected literal {b}, got an uncalculated Excel formula cell "
              "(the task requires literal values, not formulas)")
        return
    check(name, num_close(a, b, tol), f"Expected {b}, got {a}")


def num_close_any(a, expected_values, tol=2.0):
    """True if `a` is numerically close to any of the expected values."""
    return any(num_close(a, b, tol) for b in expected_values)


def check_num_any(name, a, expected_values, tol=2.0):
    """Emit a check that passes when `a` matches any of several expected
    values. Used where the task wording admits more than one legitimate
    interpretation (see get_expected_tiers). Uncalculated formula cells still
    FAIL, exactly like check_num."""
    if a is UNEVAL:
        check(name, False,
              f"expected a literal value in {expected_values}, got an "
              "uncalculated Excel formula cell (the task requires literal "
              "values, not formulas)")
        return
    check(name, num_close_any(a, expected_values, tol),
          f"Expected one of {expected_values}, got {a}")


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_expected_tiers():
    """Query Canvas DB for the expected tier counts.

    The task's tiering rule ("average score as a percentage, score divided by
    points possible times 100, across all graded submissions") admits two
    defensible readings:
      (a) per-student average of the per-submission percentages
          AVG(score / points_possible * 100)         -- the reading clarified
          in docs/task.md;
      (b) per-student weighted average
          SUM(score) / SUM(points_possible) * 100.
    Both are computed and returned keyed by interpretation ('a' / 'b'). A
    graded cell is accepted if it matches EITHER interpretation within the
    check's tolerance, so a model that read the task the other way is never
    falsely failed, while genuinely wrong counts still fail both.
    """
    def _query(inner_sql):
        cur.execute("""
            SELECT sub.course_id,
              SUM(CASE WHEN avg_pct < 60 THEN 1 ELSE 0 END) as needs_support,
              SUM(CASE WHEN avg_pct >= 60 AND avg_pct < 75 THEN 1 ELSE 0 END) as developing,
              SUM(CASE WHEN avg_pct >= 75 THEN 1 ELSE 0 END) as proficient,
              COUNT(*) as total
            FROM (
              %s
            ) sub
            GROUP BY sub.course_id
            ORDER BY sub.course_id
        """ % inner_sql)
        out = {}
        for row in cur.fetchall():
            out[int(row[0])] = {
                'needs_support': int(row[1]),
                'developing': int(row[2]),
                'proficient': int(row[3]),
                'total': int(row[4]),
            }
        return out

    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        interp_a = _query("""
              SELECT a.course_id, s.user_id,
                AVG(CASE WHEN a.points_possible > 0 THEN s.score / a.points_possible * 100 ELSE NULL END) as avg_pct
              FROM canvas.submissions s
              JOIN canvas.assignments a ON s.assignment_id = a.id
              WHERE a.course_id IN (16, 17) AND s.score IS NOT NULL AND a.points_possible > 0
              GROUP BY a.course_id, s.user_id
        """)
        interp_b = _query("""
              SELECT a.course_id, s.user_id,
                SUM(s.score)::numeric
                  / NULLIF(SUM(CASE WHEN a.points_possible > 0 THEN a.points_possible ELSE NULL END), 0) * 100 as avg_pct
              FROM canvas.submissions s
              JOIN canvas.assignments a ON s.assignment_id = a.id
              WHERE a.course_id IN (16, 17) AND s.score IS NOT NULL AND a.points_possible > 0
              GROUP BY a.course_id, s.user_id
        """)
        cur.close()
        conn.close()
        return {cid: {'a': interp_a[cid], 'b': interp_b[cid]} for cid in interp_a}
    except Exception as e:
        print(f"  [WARN] Could not query Canvas: {e}")
        return {}


def get_expected_market_events():
    """Query YF DB for expected top 3 market events.

    The 30-day window is anchored to the most recent trading date present in
    the data (MAX(date)), which is deterministic and matches the task wording
    ("most recent 30 days of trading data").
    """
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT symbol, date, open, close,
              ROUND(((close - open) / open * 100)::numeric, 2) as change_pct
            FROM yf.stock_prices
            WHERE symbol IN ('GOOGL', 'AMZN', 'JPM')
              AND date >= (SELECT MAX(date) - INTERVAL '30 days' FROM yf.stock_prices WHERE symbol = 'GOOGL')
            ORDER BY ABS((close - open) / open) DESC
            LIMIT 3
        """)
        events = []
        for row in cur.fetchall():
            events.append({
                'symbol': row[0],
                'date': str(row[1]),
                'change_pct': float(row[4])
            })
        cur.close()
        conn.close()
        return events
    except Exception as e:
        print(f"  [WARN] Could not query YF: {e}")
        return []


def load_workbook_pair(path):
    """Load a workbook twice: once raw (data_only=False) and once for cached
    formula values (data_only=True). Reading with data_only=False means a cell
    holding a literal value is always read correctly; formula cells fall back
    to their cached value so a correctly-recalculated workbook still passes.
    """
    wb_raw = None
    wb_cached = None
    try:
        wb_raw = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        print(f"  [WARN] Could not open workbook {path}: {e}")
        return None, None
    try:
        wb_cached = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        wb_cached = None
    return wb_raw, wb_cached


def cell_value(wb_cached, sheet_title, cell):
    """Resolve a cell's value.

    - Literal values are returned as-is.
    - Formula cells return their cached value when available (the workbook was
      opened by Excel / recalculated).
    - Uncalculated formula cells (cache is None) return the UNEVAL sentinel so
      callers can skip the numeric check rather than fail the agent.
    """
    v = cell.value
    if isinstance(v, str) and v.startswith('='):
        if wb_cached is not None and sheet_title in wb_cached.sheetnames:
            cv = wb_cached[sheet_title].cell(row=cell.row, column=cell.column).value
            if cv is not None:
                return cv
        return UNEVAL
    return v


def read_data_rows(wb_raw, wb_cached, ws):
    """Return data rows (from row 2 down) as resolved value lists, skipping
    rows that are completely empty or made up entirely of uncalculated
    formulas."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell_value(wb_cached, ws.title, c) for c in row]
        if any(v is not None and v is not UNEVAL and str(v).strip() != "" for v in vals):
            rows.append(vals)
    return rows


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Financial_Literacy_Workshops.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Financial_Literacy_Workshops.xlsx")
    # groundtruth file is kept for archival/consistency; expected values below
    # are computed live from the DB (which is the source of truth).

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    wb_raw, wb_cached = load_workbook_pair(agent_file)
    if wb_raw is None:
        check("Excel readable", False, "Could not open workbook")
        return

    check("Has 3 sheets", len(wb_raw.sheetnames) >= 3, f"Got {wb_raw.sheetnames}")

    expected_tiers = get_expected_tiers()
    expected_events = get_expected_market_events()

    # Student_Tiers sheet
    print("  Checking Student_Tiers...")
    st_sheet = get_sheet(wb_raw, "Student_Tiers")
    check("Sheet 'Student_Tiers' exists", st_sheet is not None, f"Sheets: {wb_raw.sheetnames}")
    if st_sheet:
        try:
            rows = read_data_rows(wb_raw, wb_cached, st_sheet)
            check("Student_Tiers has 2 rows", len(rows) == 2, f"Got {len(rows)}")
            for row in rows:
                try:
                    cid = _to_int(row[0]) if row[0] is not UNEVAL else None
                    if cid is None:
                        continue
                    if cid in expected_tiers:
                        exp = expected_tiers[cid]
                        check_num_any(f"Course {cid} Needs_Support",
                                      row[2],
                                      [exp['a']['needs_support'], exp['b']['needs_support']], 1)
                        check_num_any(f"Course {cid} Developing",
                                      row[3],
                                      [exp['a']['developing'], exp['b']['developing']], 1)
                        check_num_any(f"Course {cid} Proficient",
                                      row[4],
                                      [exp['a']['proficient'], exp['b']['proficient']], 10)
                        check_num_any(f"Course {cid} Total",
                                      row[5],
                                      [exp['a']['total'], exp['b']['total']], 10)
                except Exception as e:
                    print(f"  [WARN] Student_Tiers row check failed: {e}")
        except Exception as e:
            print(f"  [WARN] Student_Tiers sheet check failed: {e}")

    # Market_Events sheet
    print("  Checking Market_Events...")
    me_sheet = get_sheet(wb_raw, "Market_Events")
    check("Sheet 'Market_Events' exists", me_sheet is not None, f"Sheets: {wb_raw.sheetnames}")
    if me_sheet:
        try:
            rows = read_data_rows(wb_raw, wb_cached, me_sheet)
            check("Market_Events has 3 rows", len(rows) == 3, f"Got {len(rows)}")
            if expected_events and rows:
                # Check events in order (the task requires sorting by abs change desc)
                for i, exp_event in enumerate(expected_events):
                    if i >= len(rows):
                        break
                    try:
                        row = rows[i]
                        sym = str(row[1]).strip().upper() if row[1] not in (None, UNEVAL) else ""
                        check(f"Event {i+1} symbol is {exp_event['symbol']}",
                              sym == exp_event['symbol'],
                              f"Got {sym}")
                        check_num(f"Event {i+1} change_pct",
                                  row[2], exp_event['change_pct'], 0.5)
                    except Exception as e:
                        print(f"  [WARN] Market_Events row {i+1} check failed: {e}")
        except Exception as e:
            print(f"  [WARN] Market_Events sheet check failed: {e}")

    # Workshop_Schedule sheet
    print("  Checking Workshop_Schedule...")
    ws_sheet = get_sheet(wb_raw, "Workshop_Schedule")
    check("Sheet 'Workshop_Schedule' exists", ws_sheet is not None, f"Sheets: {wb_raw.sheetnames}")
    if ws_sheet:
        try:
            rows = read_data_rows(wb_raw, wb_cached, ws_sheet)
            check("Workshop_Schedule has 3 rows", len(rows) == 3, f"Got {len(rows)}")

            topics_found = set()
            for row in rows:
                if row[1] not in (None, UNEVAL) and str(row[1]).strip():
                    topics_found.add(str(row[1]).strip().lower())

            check("Has 'Intro to Markets' workshop",
                  any("intro" in t and "market" in t for t in topics_found),
                  f"Topics: {topics_found}")
            check("Has 'Portfolio Basics' workshop",
                  any("portfolio" in t and "basic" in t for t in topics_found),
                  f"Topics: {topics_found}")
            check("Has 'Risk Management' workshop",
                  any("risk" in t and "manage" in t for t in topics_found),
                  f"Topics: {topics_found}")

            # Check expected attendance for each tier (accept either
            # interpretation of the tiering rule, see get_expected_tiers).
            if expected_tiers and rows:
                ns_vals = [sum(t[k].get('needs_support', 0) for t in expected_tiers.values())
                           for k in ('a', 'b')]
                dev_vals = [sum(t[k].get('developing', 0) for t in expected_tiers.values())
                            for k in ('a', 'b')]
                prof_vals = [sum(t[k].get('proficient', 0) for t in expected_tiers.values())
                             for k in ('a', 'b')]
                for row in rows:
                    try:
                        tier = str(row[2]).strip().lower() if row[2] not in (None, UNEVAL) else ""
                        if "needs" in tier or "support" in tier:
                            check_num_any("Needs Support attendance", row[3], ns_vals, 2)
                        elif "develop" in tier:
                            check_num_any("Developing attendance", row[3], dev_vals, 2)
                        elif "proficient" in tier:
                            check_num_any("Proficient attendance", row[3], prof_vals, 50)
                    except Exception as e:
                        print(f"  [WARN] Workshop_Schedule row check failed: {e}")
        except Exception as e:
            print(f"  [WARN] Workshop_Schedule sheet check failed: {e}")


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE lower(summary) LIKE '%%intro%%market%%'
               OR lower(summary) LIKE '%%portfolio%%basic%%'
               OR lower(summary) LIKE '%%risk%%manage%%'
            ORDER BY start_datetime
        """)
        workshops = cur.fetchall()
        check("3 workshop calendar events created", len(workshops) >= 3,
              f"Found {len(workshops)} workshop events")

        if workshops:
            topics = [w[0].lower() for w in workshops]
            check("Calendar has Intro to Markets",
                  any("intro" in t and "market" in t for t in topics))
            check("Calendar has Portfolio Basics",
                  any("portfolio" in t and "basic" in t for t in topics))
            check("Calendar has Risk Management",
                  any("risk" in t and "manage" in t for t in topics))

            # Check workshops are on weekdays
            for w in workshops:
                if w[2]:
                    dt = w[2]
                    check(f"'{w[0]}' on weekday", dt.weekday() < 5,
                          f"Day: {dt.strftime('%A')}")

            # Check no two workshops on same day
            dates = set()
            for w in workshops:
                if w[2]:
                    d = w[2].date()
                    check(f"'{w[0]}' unique date", d not in dates,
                          f"Duplicate: {d}")
                    dates.add(d)

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Check student announcement email
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE lower(subject) LIKE '%%workshop%%announcement%%'
               OR lower(subject) LIKE '%%financial literacy%%workshop%%'
        """)
        student_emails = cur.fetchall()
        check("Workshop announcement email sent", len(student_emails) > 0,
              f"Found {len(student_emails)}")
        if student_emails:
            to_str = str(student_emails[0][1]).lower()
            check("Announcement to finance_students",
                  "finance_students" in to_str,
                  f"To: {student_emails[0][1]}")
            body = (student_emails[0][2] or "").lower()
            check("Announcement mentions workshops",
                  "intro" in body or "portfolio" in body or "risk" in body or "workshop" in body,
                  f"Body length: {len(body)}")

        # Check department head summary email
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE lower(subject) LIKE '%%workshop%%planning%%summary%%'
               OR lower(subject) LIKE '%%workshop%%summary%%'
        """)
        head_emails = cur.fetchall()
        check("Department head summary email sent", len(head_emails) > 0,
              f"Found {len(head_emails)}")
        if head_emails:
            to_str = str(head_emails[0][1]).lower()
            check("Summary to department_head",
                  "department_head" in to_str,
                  f"To: {head_emails[0][1]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_scripts_and_outputs(agent_workspace):
    print("\n=== Checking Scripts and Output Files ===")

    # Check scripts exist
    check("categorize_students.py exists",
          os.path.isfile(os.path.join(agent_workspace, "categorize_students.py")),
          agent_workspace)
    check("find_market_events.py exists",
          os.path.isfile(os.path.join(agent_workspace, "find_market_events.py")),
          agent_workspace)
    check("generate_outline.py exists",
          os.path.isfile(os.path.join(agent_workspace, "generate_outline.py")),
          agent_workspace)

    # Check market_events.json
    mej = os.path.join(agent_workspace, "market_events.json")
    check("market_events.json exists", os.path.isfile(mej), agent_workspace)
    if os.path.isfile(mej):
        try:
            with open(mej) as f:
                events = json.load(f)
            if isinstance(events, list):
                check("market_events.json has 3 events", len(events) >= 3,
                      f"Got {len(events)}")
            elif isinstance(events, dict) and "events" in events:
                check("market_events.json has 3 events", len(events["events"]) >= 3,
                      f"Got {len(events['events'])}")
            else:
                check("market_events.json is valid list/dict", False, f"Type: {type(events)}")
        except Exception as e:
            check("market_events.json parseable", False, str(e))

    # Check workshop_materials.txt
    wmt = os.path.join(agent_workspace, "workshop_materials.txt")
    check("workshop_materials.txt exists", os.path.isfile(wmt), agent_workspace)
    if os.path.isfile(wmt):
        with open(wmt) as f:
            content = f.read().lower()
        check("workshop_materials.txt has content", len(content) > 200,
              f"Length: {len(content)}")
        check("Materials mentions Intro to Markets",
              "intro" in content and "market" in content)
        check("Materials mentions Portfolio",
              "portfolio" in content)
        check("Materials mentions Risk",
              "risk" in content)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    # Isolate each check section so an unexpected exception in one cannot abort
    # the whole evaluation and prevent the res_log_file from being written.
    for fn in (check_excel, check_calendar, check_emails, check_scripts_and_outputs):
        try:
            if fn is check_excel:
                fn(args.agent_workspace, gt_dir)
            elif fn is check_scripts_and_outputs:
                fn(args.agent_workspace)
            else:
                fn()
        except Exception as e:
            check(f"{getattr(fn, '__name__', 'check')} did not crash", False, str(e)[:300])

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)
    # 80% pass bar. The numeric core (tier counts, change %, attendance) is
    # ~14 of the ~53 checks. If a submission gets every structural artifact
    # right but every number wrong it lands around 74%; the 80% bar prevents
    # such a submission from passing while a correct solution (~100%) and one
    # with a handful of minor misses (~90%+) clear it comfortably.
    sys.exit(0 if accuracy >= 80 else 1)


if __name__ == "__main__":
    main()
