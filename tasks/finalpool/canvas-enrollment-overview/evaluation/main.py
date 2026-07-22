"""Evaluation for canvas-enrollment-overview."""
import argparse
import json
import os
import sys
import openpyxl

try:
    import psycopg2
    DB = {
        "host": os.environ.get("PGHOST", "localhost"),
        "port": int(os.environ.get("PGPORT", "5432")),
        "dbname": "toolathlon_gym",
        "user": "eigent",
        "password": "camel",
    }
except Exception:
    psycopg2 = None
    DB = None


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def num_close_rel(a, b, rel=0.02, abs_tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_email(errors_list):
    """Verify email to registrar@openuniversity.ac.uk with the right subject."""
    import re
    print("  Checking Email to registrar@openuniversity.ac.uk...")
    if psycopg2 is None or DB is None:
        errors_list.append("psycopg2 unavailable; cannot verify email")
        return
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        rows = cur.fetchall()
        cur.close(); conn.close()
        found = False
        for subject, to_addr, body in rows:
            subj_l = (subject or "").lower()
            if "canvas enrollment overview report" not in subj_l:
                continue
            to_str = str(to_addr or "").lower()
            if "registrar@openuniversity.ac.uk" not in to_str:
                continue
            found = True
            body_l = (body or "").lower()
            # Body should mention key totals - require labelled numeric values via regex.
            # Total_Students = 32593 (also accept 32,593 or 32 593)
            students_pat = re.compile(r"(?:total[_\s\-]*students?|students?\s*total)[^0-9]{0,40}(32[\s,]?593)", re.I)
            if not students_pat.search(body_l):
                errors_list.append("Email body missing labelled Total_Students 32593")
            # Total_Courses = 22 (must be after a Total_Courses label, isolated number)
            courses_pat = re.compile(r"(?:total[_\s\-]*courses?|courses?\s*total)[^0-9]{0,40}\b22\b", re.I)
            if not courses_pat.search(body_l):
                errors_list.append("Email body missing labelled Total_Courses 22")
            # Total_Teachers = 41
            teachers_pat = re.compile(r"(?:total[_\s\-]*teachers?|teachers?\s*total)[^0-9]{0,40}\b41\b", re.I)
            if not teachers_pat.search(body_l):
                errors_list.append("Email body missing labelled Total_Teachers 41")
            break
        if not found:
            errors_list.append(
                f"Required email to registrar@openuniversity.ac.uk with subject 'Canvas Enrollment Overview Report' not found ({len(rows)} emails)"
            )
        else:
            print("    Found email")
    except Exception as e:
        errors_list.append(f"Email check raised: {e}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Enrollment_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Enrollment_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Enrollment Data
    print(f"  Checking Enrollment Data...")
    a_rows = load_sheet_rows(agent_wb, "Enrollment Data")
    g_rows = load_sheet_rows(gt_wb, "Enrollment Data")
    if a_rows is None:
        all_errors.append("Sheet 'Enrollment Data' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Enrollment Data' not found in groundtruth")
    else:
        sheet_name = "Enrollment Data"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            # Course_Name (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not str_match(a_row[1], g_row[1]):
                    errors.append(f"{key}.Course_Name: {a_row[1]} vs {g_row[1]}")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close_rel(a_row[2], g_row[2], rel=0.01, abs_tol=2):
                    errors.append(f"{key}.Students: {a_row[2]} vs {g_row[2]} (rel 1%)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0):
                    errors.append(f"{key}.Teachers: {a_row[3]} vs {g_row[3]} (exact)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0):
                    errors.append(f"{key}.TAs: {a_row[4]} vs {g_row[4]} (exact)")

            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close_rel(a_row[5], g_row[5], rel=0.01, abs_tol=2):
                    errors.append(f"{key}.Total_Enrolled: {a_row[5]} vs {g_row[5]} (rel 1%)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # For string-valued metrics (Largest_Course, Smallest_Course), require exact match
                # For numeric metrics, use small tolerance
                gt_v = g_row[1]
                if isinstance(gt_v, (int, float)):
                    if not num_close(a_row[1], gt_v, 1.0):
                        errors.append(f"{key}.Value: {a_row[1]} vs {gt_v} (tol=1.0)")
                else:
                    if not str_match(a_row[1], gt_v):
                        errors.append(f"{key}.Value: '{a_row[1]}' vs '{gt_v}'")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")



    # Email check
    check_email(all_errors)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
