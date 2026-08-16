"""Evaluation for wc-shipping-analysis."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Shipping_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Shipping_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # === Shipping Zones sheet ===
    print(f"  Checking Shipping Zones...")
    a_rows = load_sheet_rows(agent_wb, "Shipping Zones")
    g_rows = load_sheet_rows(gt_wb, "Shipping Zones")
    if a_rows is None:
        all_errors.append("Sheet 'Shipping Zones' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Shipping Zones' not found in groundtruth")
    else:
        errors = []

        # Header check
        if a_rows and a_rows[0]:
            a_header = [str(h or "").strip().lower().replace(" ", "_") for h in a_rows[0]]
            if a_header[:2] != ["zone", "shipping_methods"]:
                errors.append(f"Shipping Zones header: {a_header[:2]} (expected ['zone','shipping_methods'])")

        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        g_data = [r for r in (g_rows[1:] if len(g_rows) > 1 else []) if r and r[0] is not None]

        if len(a_data) != len(g_data):
            errors.append(f"Shipping Zones row count: {len(a_data)} vs {len(g_data)}")

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                # Tightened tol=0 for deterministic count
                if not num_close(a_row[1], g_row[1], 0):
                    errors.append(f"{key}.Shipping_Methods: {a_row[1]} vs {g_row[1]}")

        # Sort: ascending by Zone name
        zone_names = [str(r[0]).strip() for r in a_data if r[0] is not None]
        if zone_names != sorted(zone_names, key=str.lower):
            errors.append(f"Shipping Zones not sorted by Zone name: {zone_names}")

        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # === Summary sheet ===
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")

        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # === Email check (NEW) ===
    print(f"  Checking email...")
    try:
        # Get expected zone names from GT
        zone_names_lower = []
        a_zones_rows = load_sheet_rows(gt_wb, "Shipping Zones")
        if a_zones_rows:
            for r in a_zones_rows[1:]:
                if r and r[0] is not None:
                    zone_names_lower.append(str(r[0]).strip().lower())

        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id FROM email.folders WHERE LOWER(name) IN ('sent', 'sent items')")
        sent_ids = [r[0] for r in cur.fetchall()] or [1]
        cur.execute("""
            SELECT subject, from_addr, to_addr, COALESCE(body_text, body_html, '')
            FROM email.messages
            WHERE folder_id = ANY(%s)
        """, (sent_ids,))
        rows = cur.fetchall()
        cur.close()
        conn.close()

        target_subj = "shipping zone configuration review"
        target_to = "logistics@shop.com"

        match = None
        for subj, from_addr, to_addr, body in rows:
            subj_l = (subj or "").strip().lower()
            tos = []
            if isinstance(to_addr, list):
                tos = [str(t).lower() for t in to_addr]
            elif isinstance(to_addr, str):
                try:
                    p = json.loads(to_addr)
                    tos = [str(t).lower() for t in p] if isinstance(p, list) else [to_addr.lower()]
                except Exception:
                    tos = [to_addr.lower()]
            if target_subj in subj_l and any(target_to in t for t in tos):
                match = (subj, from_addr, to_addr, body)
                break

        if not match:
            all_errors.append(f"No email with subject '{target_subj}' to {target_to}")
            print(f"    FAIL: No matching email")
        else:
            body_lower = (match[3] or "").lower()
            missing_zones = [z for z in zone_names_lower if z not in body_lower]
            if missing_zones:
                all_errors.append(f"Email body missing zones: {missing_zones}")
                print(f"    FAIL: Email body missing zones {missing_zones}")
            else:
                print(f"    PASS")
    except Exception as e:
        all_errors.append(f"Email check error: {e}")
        print(f"    ERROR: {e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
