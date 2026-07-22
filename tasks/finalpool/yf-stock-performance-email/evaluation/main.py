"""Evaluation for yf-stock-performance-email."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

EXPECTED_PERF = {
    "GOOGL": {"company": "Alphabet", "price_1y": 172.345, "latest": 300.88, "yoy": 74.58, "verdict": "Outperform"},
    "JNJ":   {"company": "Johnson", "price_1y": 160.6441, "latest": 239.63, "yoy": 49.17, "verdict": "Outperform"},
    "XOM":   {"company": "Exxon", "price_1y": 101.9375, "latest": 150.76, "yoy": 47.89, "verdict": "Outperform"},
    "JPM":   {"company": "Morgan", "price_1y": 246.4761, "latest": 293.55, "yoy": 19.10, "verdict": "Outperform"},
    "AMZN":  {"company": "Amazon", "price_1y": 208.36, "latest": 218.94, "yoy": 5.08, "verdict": "Neutral"},
}

EXPECTED_SUMMARY = {
    "Best_Performer": "GOOGL",
    "Worst_Performer": "AMZN",
    "Avg_YoY_Return": 39.16,
    "Outperform_Count": 4,
    "Underperform_Count": 0,
}


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.lower() in str(haystack).lower()


def str_match(a, b):
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_dir = args.agent_workspace or os.path.join(task_root, "initial_workspace")

    agent_file = os.path.join(agent_dir, "YoY_Stock_Performance.xlsx")
    gt_file = os.path.join(gt_dir, "YoY_Stock_Performance.xlsx")

    file_errors = []
    db_errors = []

    if not os.path.exists(agent_file):
        file_errors.append(f"Agent output not found: {agent_file}")
    if not os.path.exists(gt_file):
        file_errors.append(f"Groundtruth not found: {gt_file}")

    if not file_errors:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Performance sheet
        print("  Checking Performance...")
        a_rows = load_sheet_rows(agent_wb, "Performance")
        g_rows = load_sheet_rows(gt_wb, "Performance")
        if a_rows is None:
            file_errors.append("Sheet 'Performance' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Performance' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            if len(a_data) != len(g_data):
                file_errors.append(f"Performance row count: agent {len(a_data)} vs gt {len(g_data)}")

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().upper()] = row

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().upper()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing stock: {key}")
                    continue

                # Price_1Y_Ago (col 2) - price tol 0.5
                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 0.5):
                        file_errors.append(f"{key} Price_1Y_Ago: {a_row[2]} vs {g_row[2]}")

                # Latest_Price (col 3)
                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 0.5):
                        file_errors.append(f"{key} Latest_Price: {a_row[3]} vs {g_row[3]}")

                # YoY_Return_Pct (col 4) - percent tol 0.1
                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 0.1):
                        file_errors.append(f"{key} YoY_Return: {a_row[4]} vs {g_row[4]}")

                # Verdict (col 5)
                if len(a_row) > 5 and len(g_row) > 5:
                    if not str_match(a_row[5], g_row[5]):
                        file_errors.append(f"{key} Verdict: {a_row[5]} vs {g_row[5]}")

            # Sort order check: rows sorted by YoY_Return_Pct DESC
            try:
                yoy_values = [float(r[4]) for r in a_data if r and len(r) > 4 and r[4] is not None]
                if yoy_values != sorted(yoy_values, reverse=True):
                    file_errors.append(f"Performance sheet not sorted by YoY_Return_Pct DESC: {yoy_values}")
            except Exception as e:
                file_errors.append(f"Sort-order check error: {e}")

        # Check Summary sheet
        print("  Checking Summary...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            file_errors.append("Sheet 'Summary' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing summary metric: {g_row[0]}")
                    continue

                if len(a_row) > 1 and len(g_row) > 1:
                    g_val = g_row[1]
                    a_val = a_row[1]
                    # Counts must be exact; averages tolerate 0.1
                    try:
                        fa, fb = float(a_val), float(g_val)
                        if "count" in key:
                            tol = 0
                        elif "avg" in key or "average" in key:
                            tol = 0.1
                        else:
                            tol = 0.5
                        if abs(fa - fb) > tol:
                            file_errors.append(f"Summary {key}: {a_val} vs {g_val} (tol={tol})")
                    except (TypeError, ValueError):
                        if not str_match(a_val, g_val):
                            file_errors.append(f"Summary {key}: '{a_val}' vs '{g_val}'")

    # Check email sent (DB check) - BLOCKING
    print("  Checking email...")
    email_errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%investment-committee@fund.com%'
        """)
        email_rows = cur.fetchall()
        if not email_rows:
            email_errors.append("No email to investment-committee@fund.com found")
        else:
            # Validate subject AND body
            subject_ok = False
            body_ok = False
            for subj, to, body in email_rows:
                subj_l = (subj or "").lower()
                body_l = (body or "").lower()
                if "yoy stock performance report" in subj_l:
                    subject_ok = True
                # Body should mention best (GOOGL) and worst (AMZN) performers
                if "googl" in body_l and "amzn" in body_l:
                    body_ok = True
            if not subject_ok:
                email_errors.append("Email subject 'YoY Stock Performance Report' not found")
            if not body_ok:
                email_errors.append("Email body must reference best (GOOGL) and worst (AMZN) performers")
        cur.close()
        conn.close()
    except Exception as e:
        email_errors.append(f"Email check error: {e}")

    # Final result - email BLOCKING
    print(f"\n=== SUMMARY ===")
    print(f"  File errors:  {len(file_errors)}")
    print(f"  Email errors: {len(email_errors)}")
    if email_errors:
        for e in email_errors[:15]:
            print(f"    [EMAIL] {e}")
    if file_errors:
        for e in file_errors[:15]:
            print(f"    [FILE] {e}")
    if file_errors or email_errors:
        print(f"  Overall: FAIL")
        sys.exit(1)
    else:
        print(f"  Overall: PASS")
        sys.exit(0)


if __name__ == "__main__":
    main()
