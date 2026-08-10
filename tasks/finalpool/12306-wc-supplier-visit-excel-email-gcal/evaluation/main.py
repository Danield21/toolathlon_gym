"""
Evaluation for 12306-wc-supplier-visit-excel-email-gcal task.

Checks:
1. Supplier_Visit_Plan.xlsx exists with Products, Travel_Plan, Visit_Schedule sheets
2. Products sheet has >= 5 rows with Product_ID and Supplier_Name columns
3. Product rows are verified against the store DB (wc.products): real product IDs,
   matching supplier / stock / total sales / priority / supplier city
4. Travel_Plan has valid Beijing-Shanghai and Beijing-Guangzhou trains on 2026-03-10
   (verified against train.trains + train.train_seats first-class price)
5. Visit_Schedule has >= 5 rows consistent with Products + 30-minute buffer after arrival
6. GCal has >= 2 supplier visit events on 2026-03-10
7. Outgoing emails sent to required addresses with meeting/visit subject keywords

DB facts (do not "fix"):
- Product supplier names come from wc.products.meta_data key 'supplier_name'.
- Supplier cities are a fixed, derivable mapping (documented in supplier_visit_guide.md).
- There are two valid Beijing-South -> Shanghai-Hongqiao trains on 2026-03-10
  (G1 09:00->13:28 and G11 07:00->12:31); BOTH are accepted.
- First-class (一等座) price: VNP->SHH 553, VNP->GGQ 862.5.
"""
import json
import os
import re
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

TRAVEL_DATE = "2026-03-10"

# Supplier -> city mapping. Documented in initial_workspace/supplier_visit_guide.md;
# kept here as the single ground truth for the evaluator.
SUPPLIER_CITIES = {
    "Asia Tech Trading": "Shanghai",
    "TechWorld Distribution": "Shanghai",
    "Global Electronics Co.": "Shanghai",
    "Euro Electronics GmbH": "Guangzhou",
    "Premium Gadgets Ltd": "Guangzhou",
    "Digital Dreams Supply": "Guangzhou",
    "SmartHome Wholesale": "Guangzhou",
    "AudioVisual Partners": "Guangzhou",
}

# City -> destination station telecode for Beijing South (VNP) departures.
CITY_STATION = {"shanghai": "SHH", "guangzhou": "GGQ"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


# ---------------------------------------------------------------------------
# Value helpers
# ---------------------------------------------------------------------------

_CURRENCY_RE = re.compile(r"[\s$¥€£￥,%]")
_PERCENT_RE = re.compile(r"%")


def _to_float(v):
    """Robustly convert a value to float. Handles strings with commas, currency
    symbols, percent signs and whitespace. Returns None if not parseable."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    is_percent = "%" in s
    s = _PERCENT_RE.sub("", s)
    s = _CURRENCY_RE.sub("", s)
    try:
        f = float(s)
    except (TypeError, ValueError):
        return None
    # "90%" means 90.0 (not 0.9) unless it is clearly below 1 as a fraction.
    if is_percent and f <= 1 and not re.search(r"[0-9]\.[0-9]", s):
        f = f * 100.0
    return f


def _duration_to_minutes(v):
    """Convert a duration like '04:28', '4:28', '05:31', 268 or '268' to minutes."""
    f = _to_float(v)
    if f is not None:
        return f
    if v is None:
        return None
    s = str(v).strip()
    parts = re.split(r"[:：]", s)
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            return h * 60 + m
        except (TypeError, ValueError):
            return None
    return None


def _time_to_minutes(v):
    """Convert 'HH:MM' / 'H:MM' to minutes past midnight; None if unparseable."""
    if v is None:
        return None
    s = str(v).strip()
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", s)
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def num_close(a, b, tol=1.0):
    """Compare two values numerically when both parse; fall back to a
    case-insensitive string comparison only when a side cannot be parsed."""
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if fa is None and fb is None:
        return True
    # one side unparseable -> fall back to string equality
    return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def fuzzy_contains(a, b):
    """True if either normalized string contains the other (used for product
    names where the agent may copy a truncated/abbreviated form)."""
    if a is None or b is None:
        return a is None and b is None
    sa = str(a).strip().lower()
    sb = str(b).strip().lower()
    if not sa or not sb:
        return False
    return sa in sb or sb in sa


def _normalize(v):
    if v is None:
        return ""
    return str(v).strip().lower()


def _norm_city(v):
    """Normalize a city value to 'shanghai' / 'guangzhou' (English or Chinese)."""
    s = _normalize(v)
    if "shanghai" in s or "上海" in s:
        return "shanghai"
    if "guangzhou" in s or "广州" in s:
        return "guangzhou"
    return s


# ---------------------------------------------------------------------------
# Excel reading (formula-tolerant)
# ---------------------------------------------------------------------------


def _load_value_workbook(path):
    """Return (value_wb, formula_wb): value_wb reads cached values (data_only),
    formula_wb reads raw cells so we can detect formula cells."""
    value_wb = openpyxl.load_workbook(path, data_only=True)
    formula_wb = openpyxl.load_workbook(path, data_only=False)
    return value_wb, formula_wb


def _cell_is_formula(formula_wb, sheet_name, row_idx, col_idx):
    try:
        cell = formula_wb[sheet_name].cell(row=row_idx, column=col_idx)
        return cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
    except Exception:
        return False


def _cell_value(value_wb, formula_wb, sheet_name, row_idx, col_idx):
    """Return the cached value of a cell; return None if the cell is a formula
    without a cached value (so the caller can skip that comparison)."""
    try:
        cell = value_wb[sheet_name].cell(row=row_idx, column=col_idx)
        val = cell.value
    except Exception:
        return None
    if val is None and _cell_is_formula(formula_wb, sheet_name, row_idx, col_idx):
        return None
    return val


def _cell_num(value_wb, formula_wb, sheet_name, row_idx, col_idx):
    return _to_float(_cell_value(value_wb, formula_wb, sheet_name, row_idx, col_idx))


# ---------------------------------------------------------------------------
# DB access
# ---------------------------------------------------------------------------


def _query_all(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or [])
        return cur.fetchall()


def load_runtime_data():
    """Load the store products and the day's trains. Returns (products, trains)
    or (None, None) if the DB is unreachable."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except psycopg2.Error as e:
        record("DB connection available", False, f"cannot connect: {str(e)[:200]}")
        return None, None
    try:
        rows = _query_all(
            conn,
            "SELECT id, name, sku, stock_quantity, total_sales, meta_data::text "
            "FROM wc.products",
        )
        products = {}
        for rid, name, sku, stock, total_sales, meta_text in rows:
            supplier = None
            if meta_text:
                try:
                    md = json.loads(meta_text)
                    for kv in md:
                        if isinstance(kv, dict) and kv.get("key") == "supplier_name":
                            supplier = kv.get("value")
                            break
                except Exception:
                    pass
            products[int(rid)] = {
                "name": name,
                "sku": sku,
                "stock": stock,
                "total_sales": total_sales,
                "supplier": supplier,
            }

        train_rows = _query_all(
            conn,
            "SELECT t.station_train_code, t.to_station_telecode, t.start_time, "
            "       t.arrive_time, t.lishi, s.price "
            "FROM train.trains t "
            "LEFT JOIN train.train_seats s ON s.train_id = t.id AND s.seat_type_code = 'M' "
            "WHERE t.depart_date = %s AND t.from_station_telecode = 'VNP' "
            "  AND t.to_station_telecode IN ('SHH','GGQ')",
            [TRAVEL_DATE],
        )
        trains = []  # list of dicts; valid candidates for VNP->SHH / VNP->GGQ
        for code, to_code, start, arrive, lishi, price in train_rows:
            trains.append({
                "code": code,
                "to_code": to_code,
                "start": start,
                "arrive": arrive,
                "duration_min": _duration_to_minutes(lishi),
                "price": _to_float(price),
            })
    except psycopg2.Error as e:
        record("DB query available", False, f"query failed: {str(e)[:200]}")
        return None, None
    finally:
        conn.close()

    record("DB connection available", True)
    return products, trains


# ---------------------------------------------------------------------------
# Check 1: Excel
# ---------------------------------------------------------------------------


def _find_sheet(wb, *keywords):
    for sn in wb.sheetnames:
        low = sn.strip().lower()
        if any(kw in low for kw in keywords):
            return sn
    return None


def check_excel(agent_workspace, groundtruth_workspace, products, trains):
    print("\n=== Check 1: Excel Supplier_Visit_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Supplier_Visit_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Supplier_Visit_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Supplier_Visit_Plan.xlsx exists", True)

    try:
        value_wb, formula_wb = _load_value_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in value_wb.sheetnames]
    has_products = any("product" in s for s in sheet_names_lower)
    has_travel = any("travel" in s for s in sheet_names_lower)
    has_schedule = any("schedule" in s or "visit" in s for s in sheet_names_lower)

    record("Excel has Products sheet", has_products, f"Sheets: {value_wb.sheetnames}")
    record("Excel has Travel_Plan sheet", has_travel, f"Sheets: {value_wb.sheetnames}")
    record("Excel has Visit_Schedule sheet", has_schedule, f"Sheets: {value_wb.sheetnames}")

    # --- structural checks ---
    products_sheet = _find_sheet(value_wb, "product") if has_products else None
    travel_sheet = _find_sheet(value_wb, "travel") if has_travel else None
    schedule_sheet = _find_sheet(value_wb, "schedule", "visit") if has_schedule else None

    product_rows = []
    travel_rows = []
    schedule_rows = []

    if products_sheet:
        ws = value_wb[products_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        product_rows = data_rows
        record("Products sheet has >= 5 rows", len(data_rows) >= 5, f"Found {len(data_rows)} rows")
        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            has_id = any("id" in h or "product" in h for h in headers)
            has_supplier = any("supplier" in h for h in headers)
            has_sales = any("sales" in h for h in headers)
            record("Products has Product_ID column", has_id, f"Headers: {rows[0]}")
            record("Products has Supplier_Name column", has_supplier, f"Headers: {rows[0]}")
            record("Products has sales column", has_sales, f"Headers: {rows[0]}")

    if travel_sheet:
        ws = value_wb[travel_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        travel_rows = data_rows
        record("Travel_Plan has >= 2 rows", len(data_rows) >= 2, f"Found {len(data_rows)} rows")
        all_text = " ".join(str(c) for row in rows for c in row if c)
        g_trains = set(m.group(0).upper() for m in re.finditer(r"\b[Gg]\d{1,4}\b", all_text))
        record("Travel_Plan contains >=2 distinct G-class train codes",
               len(g_trains) >= 2, f"Trains found: {sorted(g_trains)}")
        text_lower = all_text.lower()
        record("Travel_Plan references Shanghai", "shanghai" in text_lower or "上海" in text_lower,
               f"Content: {all_text[:200]}")
        record("Travel_Plan references Guangzhou", "guangzhou" in text_lower or "广州" in text_lower,
               f"Content: {all_text[:200]}")

    if schedule_sheet:
        ws = value_wb[schedule_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        schedule_rows = data_rows
        record("Visit_Schedule has >= 5 rows", len(data_rows) >= 5, f"Found {len(data_rows)} rows")

    # --- groundtruth presence (light, structural only) ---
    gt_path = os.path.join(groundtruth_workspace, "Supplier_Visit_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        value_wb.close(); formula_wb.close()
        return
    record("Groundtruth xlsx exists", True)

    try:
        gt_value_wb, gt_formula_wb = _load_value_workbook(gt_path)
    except Exception as e:
        record("Groundtruth xlsx readable", False, str(e))
        value_wb.close(); formula_wb.close()
        return
    for gt_sheet_name in gt_value_wb.sheetnames:
        found = any(asn.strip().lower() == gt_sheet_name.strip().lower() for asn in value_wb.sheetnames)
        record(f"GT sheet '{gt_sheet_name}' exists in agent", found, f"Available: {value_wb.sheetnames}")
    gt_value_wb.close(); gt_formula_wb.close()

    # --- DB-backed semantic checks (only meaningful when the DB is reachable) ---
    if products is None or trains is None:
        value_wb.close(); formula_wb.close()
        print("  [SKIP] DB-backed product/travel/schedule semantic checks (DB unavailable)")
        return

    _check_products_semantics(value_wb, formula_wb, products_sheet, product_rows, products)
    _check_travel_semantics(value_wb, formula_wb, travel_sheet, travel_rows, trains)
    _check_schedule_semantics(value_wb, formula_wb, schedule_sheet, schedule_rows,
                              product_rows, travel_rows)

    value_wb.close(); formula_wb.close()


def _check_products_semantics(value_wb, formula_wb, sheet_name, rows, products):
    print("\n  -- Products vs store data (DB) --")
    if not rows:
        record("Products: at least 5 rows reference real store products", False, "no product rows")
        return

    valid_count = 0
    supplier_bad = []
    stock_bad = []
    sales_bad = []
    priority_bad = []
    city_bad = []
    city_details = []

    ncols = max(len(r) for r in rows)
    # locate the sales column by header
    sales_col = None
    try:
        header_row = next(value_wb[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(c).lower() if c else "" for c in header_row]
        for i, h in enumerate(headers):
            if "sales" in h:
                sales_col = i
                break
    except Exception:
        sales_col = None

    for idx, row in enumerate(rows):
        r0 = _to_float(row[0]) if len(row) > 0 else None
        pid = int(r0) if r0 is not None and r0 == int(r0) else None
        if pid is None or pid not in products:
            supplier_bad.append(f"row{idx+1}: product id {row[0] if len(row)>0 else None} not in store")
            continue
        dbp = products[pid]
        valid_count += 1

        def colv(i):
            return _cell_value(value_wb, formula_wb, sheet_name, idx + 2, i + 1)

        # Supplier
        agent_supplier = colv(5) if ncols > 5 else None
        if agent_supplier is None or not fuzzy_contains(str(agent_supplier).strip(), str(dbp["supplier"] or "")):
            supplier_bad.append(f"row{idx+1}: expected supplier {dbp['supplier']!r}, got {agent_supplier!r}")
        # Stock
        agent_stock = colv(3) if ncols > 3 else None
        if not num_close(agent_stock, dbp["stock"], max(abs(float(dbp["stock"] or 0)) * 0.2, 1.0)):
            stock_bad.append(f"row{idx+1}: expected stock {dbp['stock']}, got {agent_stock!r}")
        # Sales
        if sales_col is not None and len(row) > sales_col:
            agent_sales = colv(sales_col)
            if agent_sales is None:
                sales_bad.append(f"row{idx+1}: sales cell empty")
            else:
                tol = max(abs(float(dbp["total_sales"] or 0)) * 0.15, 1.0)
                if not num_close(agent_sales, dbp["total_sales"], tol):
                    sales_bad.append(f"row{idx+1}: expected total_sales {dbp['total_sales']}, got {agent_sales!r}")
        # Priority
        expected_priority = _priority_for_stock(dbp["stock"])
        agent_priority = colv(7) if ncols > 7 else None
        if not _priority_match(agent_priority, expected_priority):
            priority_bad.append(f"row{idx+1}: expected {expected_priority}, got {agent_priority!r}")
        # City
        agent_city = colv(6) if ncols > 6 else None
        expected_city = SUPPLIER_CITIES.get(dbp["supplier"])
        if expected_city is None or not str_match(agent_city, expected_city):
            city_bad.append(f"row{idx+1}: expected city {expected_city}, got {agent_city!r}")
        city_details.append(_normalize(agent_city))

    def attr_ok(bad_list):
        # Allow minor noise: at least 5 rows must be fully correct for the attribute.
        return valid_count - len(bad_list) >= 5

    record("Products: at least 5 rows reference real store products", valid_count >= 5,
           f"{valid_count}/{len(rows)} rows have a valid Product_ID")
    if valid_count >= 5:
        record("Products: Supplier_Name correct for each row", attr_ok(supplier_bad),
               "; ".join(supplier_bad[:2]) or "ok")
        record("Products: Current_Stock correct for each row", attr_ok(stock_bad),
               "; ".join(stock_bad[:2]) or "ok")
        record("Products: sales figure correct (where provided)", attr_ok(sales_bad),
               "; ".join(sales_bad[:2]) or "ok")
        record("Products: Priority follows guide thresholds", attr_ok(priority_bad),
               "; ".join(priority_bad[:2]) or "ok")
        record("Products: Supplier_City matches guide location", attr_ok(city_bad),
               "; ".join(city_bad[:2]) or "ok")
        has_sh = any("shanghai" in c for c in city_details)
        has_gz = any("guangzhou" in c for c in city_details)
        record("Products: both Shanghai and Guangzhou suppliers represented",
               has_sh and has_gz, f"cities seen: {sorted(set(city_details))}")


def _priority_for_stock(stock):
    stock = _to_float(stock)
    if stock is None:
        return None
    if stock < 5:
        return "high"
    if stock <= 15:
        return "medium"
    return "low"


def _priority_match(agent_val, expected):
    if expected is None:
        return False
    a = _normalize(agent_val)
    if not a:
        return False
    return expected in a


def _check_travel_semantics(value_wb, formula_wb, sheet_name, rows, trains):
    print("\n  -- Travel_Plan vs store data (DB) --")
    if not rows:
        record("Travel_Plan: valid Shanghai and Guangzhou train rows", False, "no travel rows")
        return

    by_to = {}
    for t in trains:
        by_to.setdefault(t["to_code"], []).append(t)

    valid_by_city = {}
    for city, to_code in CITY_STATION.items():
        candidates = by_to.get(to_code, [])
        codes = sorted({t["code"] for t in candidates})
        found = False
        for row in rows:
            if _norm_city(row[0] if len(row) > 0 else None) != city:
                continue
            code = str(row[1] or "").strip().upper() if len(row) > 1 else ""
            matched = [t for t in candidates if t["code"].upper() == code]
            if not matched:
                continue
            t = matched[0]
            dep = row[2] if len(row) > 2 else None
            arr = row[3] if len(row) > 3 else None
            dur = row[4] if len(row) > 4 else None
            seat = row[5] if len(row) > 5 else None
            price = row[6] if len(row) > 6 else None
            ok_dep = str_match(dep, t["start"])
            ok_arr = str_match(arr, t["arrive"])
            ok_dur = (dur is not None and t["duration_min"] is not None
                      and abs(_duration_to_minutes(dur) - t["duration_min"]) <= 3)
            ok_seat = any(k in _normalize(seat) for k in ("一等", "first"))
            ok_price = t["price"] is not None and num_close(price, t["price"], 10.0)
            if ok_dep and ok_arr and ok_dur and ok_seat and ok_price:
                found = True
                break
        valid_by_city[city] = found

    for city, to_code in CITY_STATION.items():
        codes = sorted({t["code"] for t in by_to.get(to_code, [])})
        valid = valid_by_city.get(city, False)
        if city == "shanghai":
            record(f"Travel_Plan: valid Shanghai train row (accept {'/'.join(codes)})", valid,
                   f"expected one of {codes}")
        else:
            record("Travel_Plan: valid Guangzhou train row (G105)", valid,
                   f"expected one of {codes}")


def _check_schedule_semantics(value_wb, formula_wb, sheet_name, rows, product_rows, travel_rows):
    print("\n  -- Visit_Schedule consistency --")
    if not rows:
        record("Visit_Schedule: dates/status/buffer consistent", False, "no schedule rows")
        return

    # build supplier -> city from Products sheet
    prod_city = {}
    prod_suppliers = set()
    for r in product_rows:
        if len(r) >= 7:
            name = _normalize(r[5])
            city = _norm_city(r[6])
            if name and city in ("shanghai", "guangzhou"):
                prod_city[name] = city
                prod_suppliers.add(name)

    # build city -> arrival minutes from the agent's Travel_Plan (first valid row per city)
    arrival_by_city = {}
    for r in travel_rows:
        if len(r) < 4:
            continue
        city = _norm_city(r[0])
        if city not in ("shanghai", "guangzhou") or city in arrival_by_city:
            continue
        arr = _time_to_minutes(r[3])
        if arr is not None:
            arrival_by_city[city] = arr

    date_ok = True
    supplier_ok = True
    city_ok = True
    status_ok = True
    buffer_ok = True
    date_bad = []
    supplier_bad = []
    city_bad = []
    status_bad = []
    buffer_bad = []

    for idx, row in enumerate(rows):
        rdate = _cell_value(value_wb, formula_wb, sheet_name, idx + 2, 1)
        if not _normalize(rdate).startswith("2026-03-10"):
            date_ok = False
            date_bad.append(f"row{idx+1}: date {rdate!r}")
        rsupplier = _normalize(row[1]) if len(row) > 1 else ""
        if rsupplier not in prod_suppliers:
            supplier_ok = False
            supplier_bad.append(f"row{idx+1}: supplier '{row[1] if len(row)>1 else ''}' not in Products")
        rcity = _norm_city(row[2]) if len(row) > 2 else ""
        if rsupplier in prod_city and prod_city[rsupplier] != rcity:
            city_ok = False
            city_bad.append(f"row{idx+1}: city {rcity} != Products {prod_city[rsupplier]}")
        rstatus = _normalize(row[5]) if len(row) > 5 else ""
        if "sched" not in rstatus:
            status_ok = False
            status_bad.append(f"row{idx+1}: status {row[5] if len(row)>5 else None!r}")
        # buffer: meeting time >= arrival + 30 min (2-min tolerance)
        rtime = _time_to_minutes(_cell_value(value_wb, formula_wb, sheet_name, idx + 2, 4))
        arrival = arrival_by_city.get(rcity) if rcity in ("shanghai", "guangzhou") else None
        if rtime is None:
            buffer_ok = False
            buffer_bad.append(f"row{idx+1}: meeting time missing/invalid {row[3] if len(row)>3 else None!r}")
        elif arrival is None:
            buffer_ok = False
            buffer_bad.append(f"row{idx+1}: no travel arrival found for city {rcity!r}")
        elif rtime < arrival + 28:
            buffer_ok = False
            buffer_bad.append(f"row{idx+1}: meeting {row[3] if len(row)>3 else None!r} before arrival({arrival//60}:{arrival%60:02d})+30min")

    record("Visit_Schedule: dates are 2026-03-10", date_ok, "; ".join(date_bad[:2]) or "ok")
    record("Visit_Schedule: suppliers appear in Products sheet", supplier_ok,
           "; ".join(supplier_bad[:2]) or "ok")
    record("Visit_Schedule: city consistent with Products", city_ok,
           "; ".join(city_bad[:2]) or "ok")
    record("Visit_Schedule: Status = Scheduled", status_ok, "; ".join(status_bad[:2]) or "ok")
    record("Visit_Schedule: meeting times respect 30-min buffer after arrival", buffer_ok,
           "; ".join(buffer_bad[:2]) or "ok")


# ---------------------------------------------------------------------------
# Check 2: GCal
# ---------------------------------------------------------------------------


def _event_on_date(sh_date, utc_date, target="2026-03-10"):
    """True if the event's start date is `target` under EITHER interpretation:
    (a) the proper timezone view (start_datetime AT TIME ZONE 'Asia/Shanghai'),
    or (b) the naive-as-UTC view (the agent typed a local wall-clock string
    without an offset and PG stored it as UTC, so the stored UTC wall-clock IS
    the intended local time).  Both are returned by the query below so the
    session timezone of the container (initdb default = UTC) cannot drift a
    correctly-typed naive event off the target day."""
    try:
        if str(sh_date).startswith(target) or str(utc_date).startswith(target):
            return True
    except Exception:
        pass
    return False


def check_gcal():
    print("\n=== Check 2: GCal supplier visit events on 2026-03-10 ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Fetch a wide window and decide the "event day" in Python, accepting
        # both the proper Asia/Shanghai view AND the naive-as-UTC view.  This
        # keeps a correctly-typed event (with or without a +08:00 offset) on
        # March 10 even when the agent writes a naive local time string and PG
        # stores it as UTC.
        cur.execute("""
            SELECT summary, start_datetime, location, description,
                   (start_datetime AT TIME ZONE 'Asia/Shanghai')::date AS sh_date,
                   (start_datetime AT TIME ZONE 'UTC')::date AS utc_date
            FROM gcal.events
            WHERE start_datetime >= '2026-03-09T00:00:00Z'
              AND start_datetime < '2026-03-12T00:00:00Z'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except psycopg2.Error as e:
        record("GCal has >= 2 events on 2026-03-10", False, f"DB unavailable: {str(e)[:200]}")
        return

    events = [r for r in rows if _event_on_date(r[4], r[5])]
    record("GCal has >= 2 events on 2026-03-10", len(events) >= 2, f"Found {len(events)} events")

    # A visit event may carry the city in its summary, location, or description.
    # Supplier names also disambiguate the city (Shanghai vs Guangzhou suppliers).
    def _event_texts(e):
        texts = [str(e[0] or ""), str(e[2] or "") if len(e) > 2 else "", str(e[3] or "") if len(e) > 3 else ""]
        return " ".join(t.lower() for t in texts)

    all_texts = [_event_texts(e) for e in events]
    has_any = any(("supplier" in t or "visit" in t or "meeting" in t) for t in all_texts)
    has_sh = any(("shanghai" in t or "上海" in t) for t in all_texts)
    has_gz = any(("guangzhou" in t or "广州" in t) for t in all_texts)
    # Supplier-name fallback: an event mentioning a Shanghai supplier counts as a
    # Shanghai visit, and vice versa for Guangzhou.
    sh_suppliers = ("asia tech trading", "techworld distribution", "global electronics")
    gz_suppliers = ("euro electronics", "premium gadgets", "digital dreams", "smarthome", "audiovisual")
    if not has_sh:
        has_sh = any(any(sp in t for sp in sh_suppliers) for t in all_texts)
    if not has_gz:
        has_gz = any(any(sp in t for sp in gz_suppliers) for t in all_texts)
    record("GCal has supplier/visit/meeting events", has_any,
           f"Event texts: {all_texts}")
    record("GCal covers Shanghai and Guangzhou visits", has_sh and has_gz,
           f"Event texts: {all_texts}")


# ---------------------------------------------------------------------------
# Check 3: Emails
# ---------------------------------------------------------------------------


def check_emails():
    print("\n=== Check 3: Emails sent ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        messages = cur.fetchall()
        cur.close()
        conn.close()
    except psycopg2.Error as e:
        record("At least 1 email sent to supplier or procurement", False,
               f"DB unavailable: {str(e)[:200]}")
        return

    def to_addresses(to_addr):
        if isinstance(to_addr, list):
            return " ".join(str(r).lower() for r in to_addr)
        elif to_addr:
            try:
                parsed = json.loads(str(to_addr))
                if isinstance(parsed, list):
                    return " ".join(str(r).lower() for r in parsed)
                return str(to_addr).lower()
            except Exception:
                return str(to_addr).lower()
        return ""

    outgoing_addrs = ["shanghai_supplier@techworld.com", "gz_supplier@supplier.com", "procurement@company.com"]
    outgoing = [m for m in messages if any(addr in to_addresses(m[2]) for addr in outgoing_addrs)]

    record("At least 1 email sent to supplier or procurement", len(outgoing) >= 1,
           f"Total messages: {len(messages)}, matching: {len(outgoing)}")
    record("At least 2 outgoing emails sent (visits + summary)", len(outgoing) >= 2,
           f"Found {len(outgoing)} outgoing (matching to_addr); total: {len(messages)}")

    to_shanghai = [m for m in messages if "shanghai_supplier@techworld.com" in to_addresses(m[2])]
    record("Email sent to shanghai_supplier@techworld.com", len(to_shanghai) >= 1,
           f"Total messages: {len(messages)}")
    to_gz = [m for m in messages if "gz_supplier@supplier.com" in to_addresses(m[2])]
    record("Email sent to gz_supplier@supplier.com", len(to_gz) >= 1,
           f"Total messages: {len(messages)}")
    to_proc = [m for m in messages if "procurement@company.com" in to_addresses(m[2])]
    record("Email sent to procurement@company.com (summary)", len(to_proc) >= 1,
           f"Total messages: {len(messages)}")

    meeting_kw = ("meeting", "visit", "supplier", "appointment", "trip", "march",
                  "schedule", "invit", "meet")
    _time_pat = re.compile(r"\b\d{1,2}[:：]\d{2}\b")

    def _has_meeting_evidence(m):
        # Evidence that an outgoing email is a real meeting request: the subject
        # OR body mentions a meeting/visit/schedule/invitation, the supplier, the
        # March date, or proposes a specific time.  Searching subject+body (not
        # subject alone) means a correctly-completed agent whose subject is terse
        # (e.g. "Schedule", "Invitation") is not wrongly failed as long as the
        # body actually requests the meeting.
        subj = str(m[0] or "").lower()
        body = str(m[3] or "").lower() if len(m) > 3 else ""
        combined = subj + " " + body
        if any(kw in combined for kw in meeting_kw):
            return True
        return bool(_time_pat.search(subj)) or bool(_time_pat.search(body))

    supplier_outgoing = to_shanghai + to_gz
    if supplier_outgoing:
        # A meeting request naturally mentions the date, meeting wording, or a
        # proposed time; accept any single supplier email that evidences a
        # meeting request rather than requiring every one to contain a keyword.
        ok_subj = sum(1 for m in supplier_outgoing if _has_meeting_evidence(m))
        record("Outgoing supplier emails reference meeting/visit in subject",
               ok_subj >= 1 and len(supplier_outgoing) > 0,
               f"{ok_subj}/{len(supplier_outgoing)} outgoing supplier emails have meeting-request evidence in subject/body")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    products, trains = load_runtime_data()
    check_excel(args.agent_workspace, args.groundtruth_workspace, products, trains)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    print(f"\nOverall: {PASS_COUNT}/{total} checks passed")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "failed": FAIL_COUNT,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL ({FAIL_COUNT} failures)")
        sys.exit(1)


if __name__ == "__main__":
    main()
