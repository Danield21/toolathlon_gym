"""Evaluation for wc-order-status-report."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*), ROUND(SUM(total::numeric), 2) FROM wc.orders")
    total_orders, total_rev = cur.fetchone()
    cur.execute("SELECT ROUND(SUM(total::numeric), 2), COUNT(*) FROM wc.orders WHERE status = 'completed'")
    completed_rev, completed_count = cur.fetchone()
    fulfillment_pct = round(completed_count / total_orders * 100, 1) if total_orders else 0
    cur.close()
    conn.close()
    return {
        "Total_Orders": int(total_orders),
        "Total_Revenue": float(total_rev),
        "Completed_Revenue": float(completed_rev),
        "Fulfillment_Rate_Pct": fulfillment_pct,
    }


def check_excel(args):
    print("\n=== Checking Excel Output ===")
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Order_Status_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Order_Status_Report.xlsx")

    check("Agent Excel exists", os.path.exists(agent_file), f"Expected {agent_file}")
    if not os.path.exists(agent_file):
        return
    if not os.path.exists(gt_file):
        check("Groundtruth Excel exists", False, f"Expected {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Order Status sheet
    a_rows = load_sheet_rows(agent_wb, "Order Status")
    g_rows = load_sheet_rows(gt_wb, "Order Status")
    check("Sheet 'Order Status' present", a_rows is not None)
    if a_rows is not None and g_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        # Sort order: Total_Revenue descending
        revenues = []
        for r in a_data:
            try:
                revenues.append(float(r[2]) if len(r) > 2 and r[2] is not None else 0.0)
            except (TypeError, ValueError):
                revenues.append(0.0)
        check("Order Status sorted by Total_Revenue descending",
              revenues == sorted(revenues, reverse=True), f"got {revenues}")

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            check(f"Status '{g_row[0]}' present", a_row is not None)
            if a_row is None:
                continue
            check(f"  {key}.Order_Count == {g_row[1]}",
                  num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], 0),
                  f"got {a_row[1] if len(a_row) > 1 else None}")
            check(f"  {key}.Total_Revenue ≈ {g_row[2]}",
                  num_close(a_row[2] if len(a_row) > 2 else None, g_row[2], 1.0),
                  f"got {a_row[2] if len(a_row) > 2 else None}")
            check(f"  {key}.Avg_Order_Value ≈ {g_row[3]}",
                  num_close(a_row[3] if len(a_row) > 3 else None, g_row[3], 0.01),
                  f"got {a_row[3] if len(a_row) > 3 else None}")
            check(f"  {key}.Revenue_Share_Pct ≈ {g_row[4]}",
                  num_close(a_row[4] if len(a_row) > 4 else None, g_row[4], 0.1),
                  f"got {a_row[4] if len(a_row) > 4 else None}")

    # Summary sheet
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    check("Sheet 'Summary' present", a_rows is not None)
    if a_rows is not None and g_rows is not None:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            check(f"Summary metric '{g_row[0]}' present", a_row is not None)
            if a_row is None:
                continue
            metric = key
            # Tighten tolerance: numerics ±1.0, percentages ±0.1
            if "pct" in metric or "rate" in metric:
                tol = 0.1
            else:
                tol = 1.0
            check(f"Summary.{g_row[0]} ≈ {g_row[1]}",
                  num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], tol),
                  f"got {a_row[1] if len(a_row) > 1 else None}")


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except psycopg2.Error as e:
        check("DB connect", False, str(e))
        return
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    expected = get_expected()

    def find_for(recipient):
        for subj, frm, to, body in all_emails:
            if to is None:
                continue
            recipients = []
            if isinstance(to, list):
                recipients = [str(r).strip().lower() for r in to]
            elif isinstance(to, str):
                try:
                    parsed = json.loads(to)
                    if isinstance(parsed, list):
                        recipients = [str(r).strip().lower() for r in parsed]
                    else:
                        recipients = [str(to).strip().lower()]
                except (json.JSONDecodeError, TypeError):
                    recipients = [str(to).strip().lower()]
            if recipient.lower() in recipients:
                return subj, frm, to, body
        return None

    result = find_for("operations@shop.com")
    check("Email sent to operations@shop.com", result is not None)
    if result is None:
        return
    subj, frm, to, body = result
    check("Email subject == 'Order Fulfillment Status Report'",
          (subj or "").strip().lower() == "order fulfillment status report",
          f"got: {subj}")
    body_text = body or ""
    # body must contain key metrics (Total_Orders count and Total_Revenue value)
    check(f"Email body mentions Total_Orders {expected['Total_Orders']}",
          str(expected["Total_Orders"]) in body_text,
          f"missing {expected['Total_Orders']}")
    rev_2 = f"{expected['Total_Revenue']:.2f}"
    check(f"Email body mentions Total_Revenue {expected['Total_Revenue']}",
          rev_2 in body_text or str(int(expected["Total_Revenue"])) in body_text,
          f"missing {expected['Total_Revenue']}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args)
    check_emails()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
