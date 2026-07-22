"""
Evaluation script for sf-sales-forecast-notion task.

Checks:
1. Notion page "Sales Performance Dashboard" exists
2. Notion databases "Monthly Revenue" and "Regional Performance" exist with
   correct entry counts and per-row values
3. Excel file Sales_Dashboard_Backup.xlsx with correct data per month/region
"""
import argparse
import json
import os
import re
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # All months present:
    cur.execute("""
        SELECT TO_CHAR("ORDER_DATE"::timestamp, 'YYYY-MM') as month,
               COUNT(*) as orders,
               ROUND(SUM("TOTAL_AMOUNT")::numeric, 2) as revenue
        FROM sf_data."SALES_DW__PUBLIC__ORDERS"
        GROUP BY TO_CHAR("ORDER_DATE"::timestamp, 'YYYY-MM')
        ORDER BY month
    """)
    rows_all = [(m, int(o), float(r)) for m, o, r in cur.fetchall()]
    # "Complete-data" rule: classify each month into 3 buckets to tolerate
    # different reasonable agent thresholds (e.g., 30%, 50%, 70% of median):
    # - must_include: order_count >= 50% of median (definitely complete)
    # - must_exclude: order_count < 25% of median (definitely partial)
    # - borderline:   25%-50% of median (agent may include or exclude)
    if rows_all:
        order_counts = sorted(r[1] for r in rows_all)
        median = order_counts[len(order_counts) // 2]
        complete = [r for r in rows_all if r[1] >= median * 0.5]
        partial = [r for r in rows_all if r[1] < median * 0.25]
        borderline = [r for r in rows_all if median * 0.25 <= r[1] < median * 0.5]
    else:
        complete = []
        partial = []
        borderline = []

    cur.execute("""
        SELECT c."REGION",
               COUNT(o."ORDER_ID"),
               ROUND(SUM(o."TOTAL_AMOUNT")::numeric, 2)
        FROM sf_data."SALES_DW__PUBLIC__ORDERS" o
        JOIN sf_data."SALES_DW__PUBLIC__CUSTOMERS" c ON o."CUSTOMER_ID" = c."CUSTOMER_ID"
        GROUP BY c."REGION"
        ORDER BY c."REGION"
    """)
    regions = [(r, int(o), float(rev)) for r, o, rev in cur.fetchall()]

    cur.close()
    conn.close()
    return complete, partial, regions, borderline


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def _extract_text(json_obj):
    if json_obj is None:
        return ""
    if isinstance(json_obj, str):
        return json_obj
    if isinstance(json_obj, list):
        out = []
        for item in json_obj:
            if isinstance(item, dict):
                t = item.get("plain_text") or (item.get("text") or {}).get("content") or ""
                out.append(t)
        return "".join(out)
    if isinstance(json_obj, dict):
        for k in ("title", "rich_text", "plain_text"):
            if k in json_obj:
                return _extract_text(json_obj[k])
    return ""


def _normalize(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _page_title_text(props):
    if not props:
        return ""
    for pname, pval in (props.items() if isinstance(props, dict) else []):
        if isinstance(pval, dict) and pval.get("type") == "title":
            return _extract_text(pval.get("title", []))
    return ""


def _page_number_field(props, candidates):
    if not props:
        return None
    for pname, pval in (props.items() if isinstance(props, dict) else []):
        if not isinstance(pval, dict):
            continue
        if pval.get("type") != "number":
            continue
        if _normalize(pname) in [c.lower() for c in candidates] or any(
            c.lower().replace(" ", "_") == _normalize(pname).replace(" ", "_") for c in candidates
        ):
            return pval.get("number")
    return None


def check_notion():
    print("\n=== Checking Notion ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Dashboard page (exact title)
    cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
    pages = cur.fetchall()
    dashboard_page_id = None
    for pid, props in pages:
        title = _normalize(_page_title_text(props))
        if title == "sales performance dashboard":
            dashboard_page_id = pid
            break
    check("Notion 'Sales Performance Dashboard' page exists",
          dashboard_page_id is not None,
          f"Pages count={len(pages)}")

    cur.execute("SELECT id, title, properties FROM notion.databases WHERE archived = false")
    dbs = cur.fetchall()

    monthly_db = None
    regional_db = None
    for did, title, props in dbs:
        title_text = _normalize(_extract_text(title))
        if title_text == "monthly revenue":
            monthly_db = did
        elif title_text == "regional performance":
            regional_db = did

    check("Notion 'Monthly Revenue' database exists with exact title",
          monthly_db is not None,
          f"Found: {[_extract_text(t) for _, t, _ in dbs]}")
    check("Notion 'Regional Performance' database exists with exact title",
          regional_db is not None,
          f"Found: {[_extract_text(t) for _, t, _ in dbs]}")

    expected_monthly, partial_months, expected_regions, borderline_months = load_expected()

    # Monthly entries
    if monthly_db:
        cur.execute("""SELECT properties FROM notion.pages
                       WHERE parent->>'database_id' = %s AND archived=false""",
                    (monthly_db,))
        monthly_pages = cur.fetchall()
        # Count is between len(expected_monthly) and len(expected_monthly)+len(borderline)
        # — agent free to include or exclude borderline months.
        min_count = len(expected_monthly)
        max_count = len(expected_monthly) + len(borderline_months)
        check(f"Monthly Revenue entry count in [{min_count}, {max_count}]",
              min_count <= len(monthly_pages) <= max_count,
              f"Found {len(monthly_pages)}, expected {min_count}-{max_count}")

        # Build lookup by month string
        actual_months = {}
        for (props,) in monthly_pages:
            tname = _normalize(_page_title_text(props))
            order_count = _page_number_field(props, ["Order Count", "Order_Count", "OrderCount", "Orders"])
            revenue = _page_number_field(props, ["Revenue", "Total Revenue", "Total_Revenue"])
            actual_months[tname] = (order_count, revenue)

        for m, oc, rev in expected_monthly:
            actual = actual_months.get(m.lower())
            check(f"Monthly Revenue '{m}' entry exists", actual is not None,
                  f"Months found: {list(actual_months.keys())}")
            if actual is not None:
                check(f"Monthly Revenue '{m}' Order Count correct",
                      actual[0] is not None and num_close(actual[0], oc, 1),
                      f"Got {actual[0]}, expected {oc}")
                check(f"Monthly Revenue '{m}' Revenue correct",
                      actual[1] is not None and num_close(actual[1], rev, 1.0),
                      f"Got {actual[1]}, expected {rev}")

        # Partial months (definitely partial: <25% of median) must be excluded
        for m, _, _ in partial_months:
            check(f"Partial month '{m}' excluded from Monthly Revenue",
                  m.lower() not in actual_months,
                  f"Should not appear; found in titles {list(actual_months.keys())}")

    # Regional entries
    if regional_db:
        cur.execute("""SELECT properties FROM notion.pages
                       WHERE parent->>'database_id' = %s AND archived=false""",
                    (regional_db,))
        regional_pages = cur.fetchall()
        check(f"Regional Performance has exactly {len(expected_regions)} entries",
              len(regional_pages) == len(expected_regions),
              f"Found {len(regional_pages)}")

        actual_regions = {}
        for (props,) in regional_pages:
            tname = _normalize(_page_title_text(props))
            order_count = _page_number_field(props, ["Order Count", "Order_Count", "Orders"])
            revenue = _page_number_field(props, ["Revenue", "Total Revenue"])
            actual_regions[tname] = (order_count, revenue)

        for region, oc, rev in expected_regions:
            actual = actual_regions.get(_normalize(region))
            check(f"Regional Performance '{region}' entry exists", actual is not None,
                  f"Regions found: {list(actual_regions.keys())}")
            if actual is not None:
                check(f"Regional '{region}' Order Count correct",
                      actual[0] is not None and num_close(actual[0], oc, 1),
                      f"Got {actual[0]}, expected {oc}")
                check(f"Regional '{region}' Revenue correct",
                      actual[1] is not None and num_close(actual[1], rev, 1.0),
                      f"Got {actual[1]}, expected {rev}")

    cur.close()
    conn.close()


def check_excel(agent_workspace):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Sales_Dashboard_Backup.xlsx")
    check("Excel file exists", os.path.isfile(xlsx_path), f"Expected {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        return

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    expected_monthly, partial_months, expected_regions, borderline_months = load_expected()

    def find_sheet(name_norm):
        for s in wb.sheetnames:
            if _normalize(s).replace("_", " ") == name_norm:
                return wb[s]
        return None

    # Monthly Revenue sheet
    ws_m = find_sheet("monthly revenue")
    check("'Monthly Revenue' sheet exists", ws_m is not None, f"Sheets: {wb.sheetnames}")
    if ws_m is not None:
        rows = list(ws_m.iter_rows(values_only=True))
        if not rows:
            check("Monthly Revenue sheet has rows", False, "Empty")
        else:
            header = rows[0]
            data = [r for r in rows[1:] if any(v is not None for v in r)]

            def find_col(names):
                for i, c in enumerate(header):
                    if _normalize(c).replace("_", " ") in [n.lower() for n in names]:
                        return i
                return None

            month_col = find_col(["month"])
            order_col = find_col(["order count"])
            rev_col = find_col(["revenue"])

            check("Monthly Revenue has Month/Order Count/Revenue columns",
                  None not in (month_col, order_col, rev_col),
                  f"header={header}")
            min_count = len(expected_monthly)
            max_count = len(expected_monthly) + len(borderline_months)
            check(f"Monthly Revenue row count in [{min_count}, {max_count}]",
                  min_count <= len(data) <= max_count,
                  f"Got {len(data)}")

            if month_col is not None:
                actual_lookup = {}
                for r in data:
                    if month_col < len(r) and r[month_col] is not None:
                        m = _normalize(r[month_col])
                        actual_lookup[m] = r
                for m, oc, rev in expected_monthly:
                    a = actual_lookup.get(m.lower())
                    check(f"Excel monthly '{m}' present", a is not None, f"Months: {list(actual_lookup.keys())}")
                    if a is not None:
                        if order_col is not None and order_col < len(a):
                            check(f"Excel monthly '{m}' Order Count",
                                  num_close(a[order_col], oc, 1),
                                  f"Got {a[order_col]}, expected {oc}")
                        if rev_col is not None and rev_col < len(a):
                            check(f"Excel monthly '{m}' Revenue",
                                  num_close(a[rev_col], rev, 1.0),
                                  f"Got {a[rev_col]}, expected {rev}")

                # Partial month must be excluded
                for m, _, _ in partial_months:
                    check(f"Excel excludes partial month '{m}'",
                          m.lower() not in actual_lookup,
                          f"Found in: {list(actual_lookup.keys())}")

    # Regional Performance sheet
    ws_r = find_sheet("regional performance")
    check("'Regional Performance' sheet exists", ws_r is not None, f"Sheets: {wb.sheetnames}")
    if ws_r is not None:
        rows = list(ws_r.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            data = [r for r in rows[1:] if any(v is not None for v in r)]

            def find_col(names):
                for i, c in enumerate(header):
                    if _normalize(c).replace("_", " ") in [n.lower() for n in names]:
                        return i
                return None

            region_col = find_col(["region"])
            order_col = find_col(["order count"])
            rev_col = find_col(["revenue"])
            check("Regional has Region/Order Count/Revenue columns",
                  None not in (region_col, order_col, rev_col),
                  f"header={header}")
            check(f"Regional row count == {len(expected_regions)}",
                  len(data) == len(expected_regions),
                  f"Got {len(data)}")

            if region_col is not None:
                actual_lookup = {}
                for r in data:
                    if region_col < len(r) and r[region_col] is not None:
                        actual_lookup[_normalize(r[region_col])] = r
                for region, oc, rev in expected_regions:
                    a = actual_lookup.get(_normalize(region))
                    check(f"Excel region '{region}' present", a is not None,
                          f"Found: {list(actual_lookup.keys())}")
                    if a is not None:
                        if order_col is not None and order_col < len(a):
                            check(f"Excel region '{region}' Order Count",
                                  num_close(a[order_col], oc, 1),
                                  f"Got {a[order_col]}, expected {oc}")
                        if rev_col is not None and rev_col < len(a):
                            check(f"Excel region '{region}' Revenue",
                                  num_close(a[rev_col], rev, 1.0),
                                  f"Got {a[rev_col]}, expected {rev}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_notion()
    check_excel(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_passed = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
