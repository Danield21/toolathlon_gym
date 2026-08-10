"""Evaluation for terminal-canvas-howtocook-gsheet-notion-email.

Checks:
1. Nutrition_Academic_Study.xlsx with 2 sheets (Student_Engagement, Meal_Plans)
2. student_engagement.json
3. meal_recommendations.json
4. nutrition_study_summary.txt
5. Google Sheet "Nutrition Study Data" with 2 sheets
6. Notion database "Wellness Pilot Program" with 3 entries
7. Two emails sent (student_affairs, dining_services)

Robustness notes (applied so a task-correct model reliably PASSes):
- DB connection reads env vars (PGHOST/PGPORT/PGDATABASE/PGUSER/PGPASSWORD).
- Excel values are read in a formula-aware way (data_only=True for cached
  values + data_only=False to detect formulas); formula cells whose cached
  value is None are skipped (structural checks only).
- Numeric comparison parses strings that carry thousands separators, currency
  symbols or a trailing '%' (a value like "25%" is treated as 0.25 when the
  expected value is a fraction <= 1).
- Row-count checks accept >= 3 and deduplicate by tier, so a swarm that
  re-creates an entry (or an agent that retries) cannot flip a correct result
  to FAIL.
- Ground truth tier statistics are recomputed from the immutable Canvas DB on
  every run (no CURRENT_DATE / NOW() time anchors).
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)

PASS_COUNT = 0
FAIL_COUNT = 0

TIER_KEYS = ("high", "medium", "low")

# Recipe categories the howtocook MCP serves are Chinese labels; accept the
# English words the task uses as well as the native labels.
BREAKFAST_HINTS = ("breakfast", "早餐")
LUNCH_HINTS = ("lunch", "午餐")
NONE_HINTS = ("none", "no meal", "no special", "no intervention", "n/a", "无")
HINT_BY_TIER = {"high": NONE_HINTS, "medium": BREAKFAST_HINTS, "low": LUNCH_HINTS}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def _to_float(v):
    """Robustly parse a numeric value.

    Accepts int/float/None/str; strips currency symbols, thousands separators,
    whitespace and a trailing '%'. A parenthesized number is negative
    (accounting style). Returns float or None if unparseable.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1].strip()
        s = s.replace("$", "").replace("¥", "").replace("€", "").replace("£", "")
        s = s.replace("元", "").replace("￥", "")
        s = s.replace(",", "").replace(" ", "").replace("%", "").replace("_", "")
        # Strip leading non-numeric decorations such as '≈', '~', '≥', '±'.
        s = s.lstrip("≈~≥≤±")
        try:
            f = float(s)
            return -f if neg else f
        except ValueError:
            pass
        # Last resort: pull the leading numeric token out of a decorated string
        # such as '1152 students', '≈82.9 per day' or '82.9分'. Only applied when
        # the string actually starts with a number, so words like 'Control' or
        # 'None' still fall through to the string-equality fallback.
        m = re.match(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)", s)
        if m:
            try:
                f = float(m.group(0))
                return -f if neg else f
            except ValueError:
                return None
    return None


def num_close(a, b, tol=2.0):
    """Compare two numeric-ish values.

    Prefer numeric comparison. A value written as a percentage (e.g. "25" or
    "25%") is interpreted as a fraction (0.25) when the expected value is a
    fraction <= 1. Only when one side cannot be parsed at all do we fall back
    to a case-insensitive string comparison.
    """
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        if abs(fa - fb) <= tol:
            return True
        # percent fallback: fa>1 while fb is a fraction <= 1 (and vice versa)
        if 0 < fb <= 1 and fa > 1:
            return abs(fa / 100.0 - fb) <= tol
        if 0 < fa <= 1 and fb > 1:
            return abs(fb / 100.0 - fa) <= tol
        return False
    try:
        return str(a).strip().lower() == str(b).strip().lower()
    except Exception:
        return False


def count_near(text, expected, tol=0):
    """True if an integer within `tol` of `expected` appears in text.

    Tolerates thousands separators ("1,152") and whitespace between digits.
    """
    for m in re.finditer(r"(?<!\d)\d{1,3}(?:[, ]\d{3})+|\d+", text):
        token = m.group(0)
        try:
            if abs(int(token.replace(",", "").replace(" ", "")) - expected) <= tol:
                return True
        except ValueError:
            continue
    return False


def match_tier(first_value):
    """Map a leading cell value to a canonical tier key, or None.

    Accepts 'High', 'High engagement', 'high_engagement', 'High - engagement',
    'High: 1152', etc. as well as the plain tier names. Note rows that merely
    *mention* a tier (e.g. "High engagement correlates with zero late
    submissions") also map here, but sheet_rows_by_tier keeps the most
    complete (data-like) row per tier, so such a note can never overwrite the
    real data row.
    """
    if first_value is None:
        return None
    s = str(first_value).strip().lower()
    if not s:
        return None
    for t in TIER_KEYS:
        if s == t:
            return t
        if re.match(rf"{t}[ _\-:\t]", s):
            return t
    return None


def load_workbook_dual(path):
    """Load a workbook twice.

    data_only=True gives cached values (formula results when the file has been
    recalculated); data_only=False gives raw strings so formulas ("=...") can
    be detected. Returns (values_wb, formulas_wb).
    """
    return (
        openpyxl.load_workbook(path, data_only=True),
        openpyxl.load_workbook(path, data_only=False),
    )


def _is_formula(wb_formulas, ws_name, row_idx, col_idx):
    try:
        v = wb_formulas[ws_name].cell(row=row_idx, column=col_idx).value
    except Exception:
        return False
    return isinstance(v, str) and v.startswith("=")


def _entry_score(entry):
    """Score a tier-row entry by how 'data-like' it is.

    A real data row carries parseable numbers in the numeric columns; a
    harmless note row that merely mentions a tier in column A (e.g. "High
    engagement correlates with zero late submissions") has no numbers. We
    prefer the entry with the most filled cells and, above that, the most
    numeric cells, so a note row can never overwrite the real data row.
    """
    filled = 0
    numeric = 0
    for _col, (val, _isf) in entry.items():
        if val is not None and str(val).strip() != "":
            filled += 1
            if _to_float(val) is not None:
                numeric += 1
    return (numeric, filled)


def sheet_rows_by_tier(ws_values, wb_formulas):
    """Scan a single worksheet for data rows, returning
    {tier_key: {col: (value, is_formula)}}.

    Header rows and empty rows are skipped, so the tier rows are found
    regardless of header position, extra title rows, or formatting-only rows.
    Duplicate tier rows collapse to a single entry; when several rows map to
    the same tier we keep the most complete (data-like) one. This lets a
    harmless note row such as "High engagement correlates with zero late
    submissions" (which a correct agent may add below the data) coexist with
    the real data row without overwriting it, while still collapsing equal
    duplicates from a swarm re-run (last wins on ties).
    """
    result = {}
    ws_formulas = wb_formulas[ws_values.title]
    for row_cells in ws_values.iter_rows():
        if not row_cells or row_cells[0].value is None:
            continue
        tier = match_tier(row_cells[0].value)
        if tier is None:
            continue
        entry = {}
        for cell in row_cells:
            entry[cell.column] = (
                cell.value,
                _is_formula(wb_formulas, ws_values.title, cell.row, cell.column),
            )
        prev = result.get(tier)
        if prev is None or _entry_score(entry) >= _entry_score(prev):
            result[tier] = entry
    return result


def _cell_num(entry, col):
    """Return (value, is_formula) for a column within a tier row entry."""
    return entry.get(col, (None, False))


def _compare_num(name, entry, col, expected, tol):
    """Compare a numeric column with formula-awareness.

    The task requires every numeric value to be written as a literal (no
    formulas). A formula cell whose cached value is None (e.g. a formula
    written by a script that never ran Excel) therefore carries no verifiable
    number and must FAIL rather than silently pass the arithmetic check - an
    empty / formula-only cell must never bypass the numeric comparison.
    """
    value, is_formula = _cell_num(entry, col)
    if value is None:
        check(name, False, f"Missing numeric value (formula? {is_formula})")
        return
    check(name, num_close(value, expected, tol),
          f"Expected {expected}, got {value}")


def get_groundtruth_tiers():
    """Compute expected tier values from the Canvas DB.

    The DB seed is immutable, so recomputing here is deterministic. As
    defense-in-depth the baked values below were verified against
    db/init.sql.gz (dated 2026-07-16) and match the live query; if the DB query
    fails a FAIL check is recorded so the run cannot silently pass on stale
    numbers.
    """
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT
                CASE
                    WHEN late_rate = 0 THEN 'High'
                    WHEN late_rate <= 0.25 THEN 'Medium'
                    ELSE 'Low'
                END as tier,
                COUNT(*) as student_count,
                ROUND(AVG(avg_score)::numeric, 1) as tier_avg_score,
                ROUND(AVG(late_rate)::numeric, 3) as avg_late_rate
            FROM (
                SELECT s.user_id,
                    SUM(CASE WHEN s.late THEN 1 ELSE 0 END)::float / COUNT(*) as late_rate,
                    AVG(s.score) as avg_score
                FROM canvas.submissions s
                JOIN canvas.assignments a ON s.assignment_id = a.id
                WHERE a.course_id IN (13, 14) AND s.score IS NOT NULL
                GROUP BY s.user_id
            ) sub
            GROUP BY tier
            ORDER BY tier
        """)
        tiers = {}
        for row in cur.fetchall():
            tiers[row[0]] = {
                "student_count": int(row[1]),
                "avg_score": float(row[2]),
                "late_rate": float(row[3]),
            }
        cur.close()
        conn.close()
        if not tiers:
            raise RuntimeError("DB returned 0 tier rows")
        return tiers
    except Exception as e:
        # Record failure so the strict all-pass gate cannot silently pass
        check("Canvas DB query for groundtruth tiers", False,
              f"DB query failed: {e}")
        # Baked values verified against db/init.sql.gz dated 2026-07-16.
        return {
            "High": {"student_count": 1152, "avg_score": 82.9, "late_rate": 0.0},
            "Medium": {"student_count": 252, "avg_score": 80.5, "late_rate": 0.25},
            "Low": {"student_count": 339, "avg_score": 72.4, "late_rate": 0.62},
        }


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Nutrition_Academic_Study.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Nutrition_Academic_Study.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        wb_values, wb_formulas = load_workbook_dual(agent_file)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    gt_tiers = get_groundtruth_tiers()

    # Sheet 1: Student_Engagement
    print("  -- Student_Engagement sheet --")
    eng = None
    for s in wb_values.sheetnames:
        if "engagement" in s.lower():
            eng = s
            break
    check("Student_Engagement sheet exists", eng is not None,
          f"Sheets: {wb_values.sheetnames}")

    if eng:
        rows_by_tier = sheet_rows_by_tier(wb_values[eng], wb_formulas)
        for tier_key in TIER_KEYS:
            check(f"Tier '{tier_key}' present in Excel",
                  tier_key in rows_by_tier, "Missing")
        expected_by_key = {
            "high": gt_tiers["High"],
            "medium": gt_tiers["Medium"],
            "low": gt_tiers["Low"],
        }
        for tier_key, expected in expected_by_key.items():
            entry = rows_by_tier.get(tier_key)
            if entry is None:
                continue
            _compare_num(
                f"'{tier_key}' Student_Count exact",
                entry, 2, expected["student_count"], 2,
            )
            _compare_num(
                f"'{tier_key}' Avg_Score",
                entry, 3, expected["avg_score"], 1.5,
            )
            # Late_Submission_Rate (column 4) - required; an empty cell must
            # not silently pass the numeric comparison.
            value, _ = _cell_num(entry, 4)
            if value is None:
                check(f"'{tier_key}' Late_Submission_Rate", False,
                      "Missing late submission rate")
            else:
                check(
                    f"'{tier_key}' Late_Submission_Rate",
                    num_close(value, expected["late_rate"], 0.05),
                    f"Expected {expected['late_rate']}, got {value}",
                )

    # Sheet 2: Meal_Plans
    print("  -- Meal_Plans sheet --")
    meal = None
    for s in wb_values.sheetnames:
        if "meal" in s.lower() or "plan" in s.lower():
            meal = s
            break
    check("Meal_Plans sheet exists", meal is not None,
          f"Sheets: {wb_values.sheetnames}")

    if meal:
        rows_by_tier = sheet_rows_by_tier(wb_values[meal], wb_formulas)
        for tier_key, expected_type in [
            ("high", "Control"),
            ("medium", "Partial"),
            ("low", "Full Meal Plan"),
        ]:
            entry = rows_by_tier.get(tier_key)
            if entry is None:
                check(f"Meal tier '{tier_key}' present", False, "Missing")
                continue
            # Intervention_Type (column 2) - contains match
            ival, _ = _cell_num(entry, 2)
            check(
                f"'{tier_key}' Intervention_Type",
                ival is not None and expected_type.lower() in str(ival).strip().lower(),
                f"Expected contains '{expected_type}', got {ival}",
            )
            # Recommended_Meals (column 3)
            mval, _ = _cell_num(entry, 3)
            if mval is None:
                check(f"'{tier_key}' Recommended_Meals", False, "Missing meals cell")
            else:
                meals_str = str(mval).strip().lower()
                hints = HINT_BY_TIER[tier_key]
                ok = any(h in meals_str for h in hints)
                if tier_key == "high":
                    ok = ok or meals_str == "" or meals_str in ("-", "na", "no", "none")
                check(
                    f"'{tier_key}' Recommended_Meals",
                    ok,
                    f"Got: {mval}",
                )
            # Estimated_Daily_Cost (column 4) - required; an empty cell must not
            # silently pass the budget comparison.
            cval, _ = _cell_num(entry, 4)
            if cval is None:
                check(f"'{tier_key}' Estimated_Daily_Cost", False,
                      "Missing cost cell")
            elif tier_key == "high":
                check(
                    f"'{tier_key}' Estimated_Daily_Cost is 0",
                    num_close(cval, 0, 0.01),
                    f"Got {cval}",
                )
            else:
                check(
                    f"'{tier_key}' Estimated_Daily_Cost is positive",
                    (_to_float(cval) or 0) > 0,
                    f"Got {cval}",
                )


def check_json_files(agent_workspace):
    print("\n=== Checking JSON files ===")
    gt_tiers = get_groundtruth_tiers()

    # student_engagement.json
    se_path = os.path.join(agent_workspace, "student_engagement.json")
    check("student_engagement.json exists", os.path.isfile(se_path))
    if os.path.isfile(se_path):
        try:
            with open(se_path) as f:
                data = json.load(f)
            check("student_engagement.json is valid JSON", True)
            if isinstance(data, list):
                check("Has at least 3 tier entries", len(data) >= 3, f"Got {len(data)}")
                for item in data:
                    tier = str(item.get("tier", ""))
                    key = tier.lower().strip()
                    if key in gt_tiers:
                        check(
                            f"JSON '{key}' student_count",
                            num_close(item.get("student_count", 0), gt_tiers[key]["student_count"], 5),
                            f"Expected {gt_tiers[key]['student_count']}, got {item.get('student_count')}",
                        )
            else:
                check("student_engagement.json is array", False, f"Got {type(data).__name__}")
        except Exception as e:
            check("student_engagement.json parseable", False, str(e))

    # meal_recommendations.json
    mr_path = os.path.join(agent_workspace, "meal_recommendations.json")
    check("meal_recommendations.json exists", os.path.isfile(mr_path))
    if os.path.isfile(mr_path):
        try:
            with open(mr_path) as f:
                data = json.load(f)
            check("meal_recommendations.json is valid JSON", True)
            if isinstance(data, list):
                check("Has at least 3 meal plan entries", len(data) >= 3, f"Got {len(data)}")
            else:
                check("meal_recommendations.json is array", False, f"Got {type(data).__name__}")
        except Exception as e:
            check("meal_recommendations.json parseable", False, str(e))

    # nutrition_study_summary.txt
    summary_path = os.path.join(agent_workspace, "nutrition_study_summary.txt")
    check("nutrition_study_summary.txt exists", os.path.isfile(summary_path))
    if os.path.isfile(summary_path):
        with open(summary_path) as f:
            text = f.read().lower()
        check("Summary has substantial content", len(text) > 200, f"Length: {len(text)}")
        check("Summary mentions all 3 engagement tiers",
              "high" in text and "low" in text and "medium" in text)
        check("Summary mentions meal AND plan", "meal" in text and "plan" in text)
        # The task explicitly requires the report to state each tier's student
        # count (see task.md), so these are required content. Tolerance matches
        # the student_engagement.json check (also +/-5), so a count that is
        # legitimate in the JSON can never be rejected here.
        for tier_name, expected in gt_tiers.items():
            check(
                f"Summary mentions {tier_name} count {expected['student_count']}",
                count_near(text, expected["student_count"], 5),
                f"Expected {expected['student_count']}",
            )


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        cur.execute(
            "SELECT id, title FROM gsheet.spreadsheets "
            "WHERE lower(title) LIKE '%nutrition%study%' ORDER BY title"
        )
        rows = cur.fetchall()
        check("Nutrition Study spreadsheet exists", len(rows) >= 1,
              f"Found {len(rows)} matching spreadsheets")
        if not rows:
            cur.close()
            conn.close()
            return

        gt_tiers = get_groundtruth_tiers()

        # A swarm may create several matching spreadsheets; use the most
        # complete one (the one with the most engagement-sheet data rows).
        best_ss = None
        best_score = -1
        for ss_id, _ in rows:
            cur.execute(
                "SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s ORDER BY index",
                (ss_id,),
            )
            sheets = cur.fetchall()
            eng_sheet = None
            for s in sheets:
                if "engagement" in s[1].lower():
                    eng_sheet = s
                    break
            score = 0
            if eng_sheet:
                cur.execute(
                    "SELECT COUNT(DISTINCT row_index) FROM gsheet.cells "
                    "WHERE spreadsheet_id = %s AND sheet_id = %s",
                    (ss_id, eng_sheet[0]),
                )
                score = cur.fetchone()[0] or 0
            if score > best_score:
                best_score = score
                best_ss = (ss_id, sheets, eng_sheet)
        if best_ss is None:
            cur.close()
            conn.close()
            return

        ss_id, sheets, engagement_sheet = best_ss
        sheet_names = [s[1].lower() for s in sheets]
        check(
            "Has Student_Engagement sheet",
            any("engagement" in n for n in sheet_names),
            f"Sheets: {sheet_names}",
        )
        check(
            "Has Meal_Plans sheet",
            any("meal" in n or "plan" in n for n in sheet_names),
            f"Sheets: {sheet_names}",
        )

        if engagement_sheet:
            cur.execute(
                "SELECT row_index, col_index, value FROM gsheet.cells "
                "WHERE spreadsheet_id = %s AND sheet_id = %s ORDER BY row_index, col_index",
                (ss_id, engagement_sheet[0]),
            )
            cells = cur.fetchall()
            data_rows = {}
            for r, c, v in cells:
                if r not in data_rows:
                    data_rows[r] = {}
                data_rows[r][c] = v
            check(
                "Engagement sheet has >= 4 rows (header + 3 tiers)",
                len(data_rows) >= 4,
                f"Got {len(data_rows)} rows",
            )
            blob = " ".join(str(v) for row in data_rows.values() for v in row.values()).lower()
            for tier_name, expected in gt_tiers.items():
                check(f"GSheet engagement has tier '{tier_name}'",
                      tier_name.lower() in blob,
                      f"Tier missing from gsheet")
                # Tolerance matches student_engagement.json (+/-5), so a count
                # that is legitimate in the JSON is never rejected here.
                check(
                    f"GSheet engagement has {tier_name} count {expected['student_count']}",
                    count_near(blob, expected["student_count"], 5),
                    f"Expected count {expected['student_count']}",
                )

        cur.close()
        conn.close()
    except Exception as e:
        check("GSheet check", False, str(e))


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        cur.execute(
            "SELECT id FROM notion.databases WHERE title::text LIKE '%Wellness Pilot Program%'"
        )
        dbs = cur.fetchall()
        check("Wellness Pilot Program database exists", len(dbs) >= 1, f"Found {len(dbs)}")
        if not dbs:
            cur.close()
            conn.close()
            return

        db_ids = [d[0] for d in dbs]
        # Gather pages from ANY matching database (a swarm/retry may create
        # several); duplicates do not fail the >= 3 count.
        like_clauses = " OR ".join(["p.parent::text LIKE %s"] * len(db_ids))
        params = [f"%{db_id}%" for db_id in db_ids]
        cur.execute(
            f"SELECT p.id, p.properties FROM notion.pages p "
            f"WHERE p.archived = false AND ({like_clauses})",
            params,
        )
        pages = cur.fetchall()
        check("Has at least 3 Notion pages", len(pages) >= 3, f"Found {len(pages)}")

        if pages:
            all_props_text = " ".join(str(p[1]) for p in pages).lower()
            check("Has Control group", "control" in all_props_text)
            check("Has Partial group", "partial" in all_props_text)
            check("Has Full Meal Plan group", "full meal plan" in all_props_text)
            # Start_Date 2026-04-01 - accept the literal date as well as a
            # one-day timezone shift (some clients serialize to UTC).
            check(
                "Has Start_Date 2026-04-01",
                any(d in all_props_text for d in ("2026-04-01", "2026-04-02", "2026-03-31")),
                f"Start_Date 2026-04-01 not found",
            )
            # Status 'Planning'
            check("Has Status 'Planning'",
                  "planning" in all_props_text,
                  f"'Planning' not found")
            tiers_for_notion = get_groundtruth_tiers()
            for tier_name, expected in tiers_for_notion.items():
                # Tolerance matches student_engagement.json (+/-5), so a count
                # that is legitimate in the JSON is never rejected here.
                check(
                    f"Notion mentions {tier_name} enrollment count {expected['student_count']}",
                    count_near(all_props_text, expected["student_count"], 5),
                    f"Expected {expected['student_count']}",
                )

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Check email to student_affairs - require correct subject.
        # In swarm mode several emails may be sent; concatenate all matching
        # bodies so any complete one satisfies the content checks.
        cur.execute(
            """SELECT subject, body_text, to_addr FROM email.messages
            WHERE to_addr::text LIKE '%student_affairs@university.edu%'
              AND LOWER(subject) LIKE '%nutrition-academic performance study%pilot proposal%'"""
        )
        sa_emails = cur.fetchall()
        check(
            "Email to student_affairs with subject 'Nutrition-Academic Performance Study: Pilot Proposal'",
            len(sa_emails) >= 1,
            f"Found {len(sa_emails)}",
        )
        if sa_emails:
            body = " ".join((e[1] or "") for e in sa_emails).lower()
            check("SA email mentions pilot AND study", "pilot" in body and "study" in body)
            check("SA email mentions all 3 tiers",
                  "high" in body and "medium" in body and "low" in body)
            check("SA email mentions april 2026",
                  "april" in body or "2026-04" in body or "april 2026" in body)

        # Check email to dining_services - require correct subject
        cur.execute(
            """SELECT subject, body_text, to_addr FROM email.messages
            WHERE to_addr::text LIKE '%dining_services@university.edu%'
              AND LOWER(subject) LIKE '%meal plan recommendations%wellness pilot%'"""
        )
        ds_emails = cur.fetchall()
        check(
            "Email to dining_services with subject 'Meal Plan Recommendations for Wellness Pilot'",
            len(ds_emails) >= 1,
            f"Found {len(ds_emails)}",
        )
        if ds_emails:
            body = " ".join((e[1] or "") for e in ds_emails).lower()
            check("DS email mentions meal AND plan", "meal" in body and "plan" in body)
            check("DS email mentions breakfast and lunch",
                  any(h in body for h in BREAKFAST_HINTS)
                  and any(h in body for h in LUNCH_HINTS))

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_no_noise(agent_workspace):
    """Verify noise data was not included in outputs.

    Scoped to the Tier column (first column) of every sheet, so compliant
    content in the cost/meals columns (e.g. a budget note that mentions
    'budget tracking' or 'dining services marketing') cannot be falsely
    flagged while a real leak (a marketing/parking row) still fails.
    """
    print("\n=== Reverse Validation (noise rejection) ===")
    xlsx_path = os.path.join(agent_workspace, "Nutrition_Academic_Study.xlsx")
    if os.path.isfile(xlsx_path):
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=False)
            all_text = ""
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    if row and row[0] is not None:
                        all_text += str(row[0]) + " "
            all_text = all_text.lower()
            noise_terms = ["budget tracking", "marketing", "parking", "faculty retreat"]
            found = [t for t in noise_terms if t in all_text]
            check("No noise data in Excel", len(found) == 0, f"Found: {found}")
        except Exception as e:
            check("Noise check readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_json_files(args.agent_workspace)
    check_gsheet()
    check_notion()
    check_emails()
    check_no_noise(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total,
              "total_failed": FAIL_COUNT, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Strict gate: ALL checks must pass (no FP via threshold)
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
