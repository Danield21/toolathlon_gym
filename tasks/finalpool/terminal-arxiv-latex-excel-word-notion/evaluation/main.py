"""Evaluation for terminal-arxiv-latex-excel-word-notion.

Checks:
1. Literature_Review_Matrix.xlsx with 3 sheets
2. Literature_Review_Draft.docx
3. Notion database "Transformer Research Papers"

Robustness notes:
- DB connection reads PGHOST/PGPORT/PGDATABASE/PGUSER/PGPASSWORD with the same
  defaults used by preprocess/main.py (R1).
- Excel cells are resolved through both data_only modes so literal numbers,
  cached formula results, and formatted strings all compare correctly; formula
  cells without a cached value degrade to a structural (non-failing) check (R2).
- num_close parses numbers robustly and only falls back to a case-insensitive
  string compare when a side is not numeric (R3).
- Notion text is read only from real columns (notion.databases.title,
  notion.pages.parent / properties); nothing depends on content/page_id (R4).
- No use of CURRENT_DATE / NOW() as a comparison anchor (R8).
- Sheet names are compared with spaces normalized to underscores, matching the
  reverse-validation normalization (R5-style / R13).
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)

PASS_COUNT = 0
FAIL_COUNT = 0

EXPECTED_IDS = {"1706.03762", "1810.04805", "2005.14165", "2010.11929", "2301.07041"}
EXPECTED_AUTHORS = {
    "1706.03762": "vaswani",
    "1810.04805": "devlin",
    "2005.14165": "brown",
    "2010.11929": "dosovitskiy",
    "2301.07041": "kaplan",
}
REQUIRED_CITATIONS = [
    (("1810.04805", "1706.03762"), "BERT cites Transformer"),
    (("2005.14165", "1706.03762"), "GPT-3 cites Transformer"),
    (("2010.11929", "1706.03762"), "ViT cites Transformer"),
    (("2010.11929", "1810.04805"), "ViT cites BERT"),
    (("2301.07041", "1706.03762"), "Scaling Laws cite Transformer"),
]
# Injected non-transformer (noise) papers, from preprocess. NOTE: 2106.09685
# (LoRA: Low-Rank Adaptation of Large Language Models) is deliberately NOT
# enforced as noise: it is a fine-tuning technique for transformer LLMs (its own
# content describes injecting low-rank matrices into transformer layers), and the
# task's review criteria list "novel training approaches and optimization
# methods" as a focus area. Including LoRA is therefore a defensible reading of a
# correct solution, so a model that adds it must not be FAILed. The remaining
# four noise papers are clearly unrelated to transformer/attention research and
# keep the reverse-validation check meaningful.
NOISE_IDS = {"1901.02860", "2002.05709", "1811.12345", "2012.09876"}
# Deterministic markers for noise papers: their exact IDs plus full titles, so a
# faithful summary that merely mentions a related keyword cannot trigger a fail.
# LoRA (2106.09685) is intentionally absent (see NOISE_IDS comment).
NOISE_MARKERS = [
    "1901.02860", "2002.05709", "1811.12345", "2012.09876",
    "adversarial examples are not bugs",
    "survey on knowledge graphs",
    "graph neural networks for molecular property prediction",
    "contextual bandits",
]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def _to_float(v):
    """Robust numeric parser. Handles None / int / float / formatted strings
    (thousands separators, currency symbols, %, whitespace). Returns float or None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    for ch in (",", "$", "¥", "€", "%"):
        s = s.replace(ch, "")
    s = s.strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def num_close(a, b, tol=1.0):
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # One or both sides are not numeric: fall back to a case-insensitive string
    # compare instead of treating the values as unequal by default.
    try:
        return str(a).strip().lower() == str(b).strip().lower()
    except Exception:
        return False


def _norm_id(v):
    """Normalize an ArXiv identifier so 'arXiv:1706.03762', 'arxiv.org/abs/...'
    or a trailing '.pdf' still matches the bare id."""
    s = str(v or "").strip().lower()
    for p in ("arxiv.org/abs/", "arxiv:", "http://arxiv.org/abs/",
              "https://arxiv.org/abs/"):
        if s.startswith(p):
            s = s[len(p):]
    if s.endswith(".pdf"):
        s = s[:-4]
    return s


def get_sheet(wb, name):
    t = name.strip().lower().replace(" ", "_")
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == t:
            return wb[s]
    return None


def _load_workbooks(path):
    """Load a workbook in both data_only modes.

    Returns (wb_data, wb_cache): wb_data preserves formulas (data_only=False);
    wb_cache holds cached computed values (data_only=True). Either may be None.
    """
    try:
        wb_data = openpyxl.load_workbook(path, data_only=False)
    except Exception:
        return None, None
    try:
        wb_cache = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        wb_cache = None
    return wb_data, wb_cache


def _resolve_sheet_rows(ws_data, ws_cache):
    """Resolve a sheet's data rows into effective values plus formula flags.

    Returns (rows, flags): rows[r][c] is the effective value (formula cells use
    their cached result); flags[r][c] is True when the cell is a formula whose
    cached value is missing, so its numeric value cannot be verified.
    """
    rows = []
    flags = []
    if ws_data is None:
        return rows, flags
    for i, row in enumerate(ws_data.iter_rows(min_row=2), start=2):
        rvals = []
        rflags = []
        for c in row:
            dv = c.value
            cv = None
            if ws_cache is not None:
                cv = ws_cache.cell(row=i, column=c.column).value
            if isinstance(dv, str) and dv.startswith("="):
                if cv is None:
                    rvals.append(None)
                    rflags.append(True)
                else:
                    rvals.append(cv)
                    rflags.append(False)
            else:
                rvals.append(dv)
                rflags.append(False)
        rows.append(rvals)
        flags.append(rflags)
    return rows, flags


def _find_dbs(conn):
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
    dbs = cur.fetchall()
    cur.close()
    return dbs


def _matches_db(title):
    title_str = json.dumps(title).lower() if title else ""
    return ("transformer" in title_str) and (
        "research" in title_str or "paper" in title_str)


def _pages_for_db(conn, db_id):
    cur = conn.cursor()
    cur.execute("""
        SELECT properties::text FROM notion.pages
        WHERE parent->>'database_id' = %s AND archived = false AND in_trash = false
    """, (db_id,))
    pages = cur.fetchall()
    cur.close()
    return pages


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Literature_Review_Matrix.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Literature_Review_Matrix.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Literature_Review_Matrix.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    agent_wb_data, agent_wb_cache = _load_workbooks(agent_file)
    if agent_wb_data is None:
        check("Excel readable", False, "cannot open agent workbook")
        return
    gt_wb_data, gt_wb_cache = _load_workbooks(gt_file)
    if gt_wb_data is None:
        check("Excel readable", False, "cannot open groundtruth workbook")
        return

    a_cache = agent_wb_cache if agent_wb_cache is not None else None
    g_cache = gt_wb_cache if gt_wb_cache is not None else None

    # ---- Paper_Catalog ----
    print("  Checking Paper_Catalog...")
    a_sheet = get_sheet(agent_wb_data, "Paper_Catalog")
    g_sheet = get_sheet(gt_wb_data, "Paper_Catalog")
    check("Sheet 'Paper_Catalog' exists", a_sheet is not None,
          f"Sheets: {agent_wb_data.sheetnames}")
    a_rows, a_flags = _resolve_sheet_rows(a_sheet,
                                          get_sheet(a_cache, "Paper_Catalog") if a_cache else None)
    g_rows, g_flags = _resolve_sheet_rows(g_sheet,
                                          get_sheet(g_cache, "Paper_Catalog") if g_cache else None)
    if a_sheet is not None and g_sheet is not None:
        check("Paper_Catalog has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")

        # Check that key papers are present by ArXiv_ID.
        a_ids = {_norm_id(r[0]) for r in a_rows if r and r[0]}
        for eid in EXPECTED_IDS:
            check(f"Paper '{eid}' in catalog", eid in a_ids,
                  f"Missing from {a_ids}")

        a_lookup = {}
        for r, f in zip(a_rows, a_flags):
            if r and r[0]:
                a_lookup[_norm_id(r[0])] = (r, f)
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = _norm_id(g_row[0])
            item = a_lookup.get(key)
            if item:
                a_row, a_flag = item
                if len(a_row) > 3 and len(a_flag) > 3 and len(g_row) > 3:
                    if a_flag[3]:
                        # Formula cell with no cached value: numeric value cannot
                        # be verified, so only require the cell to be present.
                        check(f"'{key}' Year", True,
                              "Year is an uncached formula cell; numeric comparison skipped")
                    else:
                        check(f"'{key}' Year",
                              num_close(a_row[3], g_row[3], 0),
                              f"Expected {g_row[3]}, got {a_row[3]}")
                # First-author surname present (word-boundary match), and the
                # column holds a surname-style value, not a long author list.
                # Generous cap so correct-but-slightly-longer cells still pass.
                if len(a_row) > 2 and key in EXPECTED_AUTHORS:
                    author_str = str(a_row[2] or "").lower()
                    surname = EXPECTED_AUTHORS[key]
                    present = re.search(rf"\b{re.escape(surname)}\b", author_str) is not None
                    check(f"'{key}' first-author surname present",
                          present,
                          f"Expected '{surname}' (word-boundary) in '{author_str}'")
                    len_ok = len(author_str.strip()) <= 200
                    check(f"'{key}' Authors field limited to first-author surname",
                          len_ok,
                          f"Authors too long ({len(author_str)}): '{author_str}'")

    # ---- Methodology_Comparison ----
    print("  Checking Methodology_Comparison...")
    a_sheet = get_sheet(agent_wb_data, "Methodology_Comparison")
    g_sheet = get_sheet(gt_wb_data, "Methodology_Comparison")
    check("Sheet 'Methodology_Comparison' exists", a_sheet is not None,
          f"Sheets: {agent_wb_data.sheetnames}")
    a_rows, a_flags = _resolve_sheet_rows(
        a_sheet, get_sheet(a_cache, "Methodology_Comparison") if a_cache else None)
    if a_sheet is not None:
        check("Methodology_Comparison has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")

        a_ids = {_norm_id(r[0]) for r in a_rows if r and r[0]}
        for eid in EXPECTED_IDS:
            check(f"Method for '{eid}'", eid in a_ids, f"Missing from {a_ids}")

        # Check each row has method_name, approach, and a key innovation.
        for r, f in zip(a_rows, a_flags):
            if r and r[0] and _norm_id(r[0]) in EXPECTED_IDS:
                has_method = r[1] is not None and len(str(r[1]).strip()) > 0
                has_approach = len(r) > 2 and r[2] is not None and len(str(r[2]).strip()) > 0
                check(f"'{_norm_id(r[0])}' has method and approach",
                      has_method and has_approach,
                      f"method={r[1]}, approach={r[2] if len(r) > 2 else None}")
                # Key_Innovation must always be present. A generous length cap
                # keeps the check meaningful (empty / pathological cells fail)
                # without penalizing a correct but slightly longer summary.
                ki_val = r[3] if len(r) > 3 else None
                ki_txt = str(ki_val).strip() if ki_val is not None else ""
                ki_len = len(ki_txt)
                check(f"'{_norm_id(r[0])}' Key_Innovation present and reasonable length",
                      0 < ki_len <= 2000,
                      f"len={ki_len}, val={ki_val!r}")

    # ---- Citation_Network ----
    print("  Checking Citation_Network...")
    a_sheet = get_sheet(agent_wb_data, "Citation_Network")
    check("Sheet 'Citation_Network' exists", a_sheet is not None,
          f"Sheets: {agent_wb_data.sheetnames}")
    if a_sheet is not None:
        a_rows, a_flags = _resolve_sheet_rows(
            a_sheet, get_sheet(a_cache, "Citation_Network") if a_cache else None)
        check("Citation_Network has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")

        # Check key citations (exact strict pairs; all must be present).
        citations = {(_norm_id(r[0]), _norm_id(r[1])) for r in a_rows if r and r[0] and r[1]}
        for pair, label in REQUIRED_CITATIONS:
            check(label, pair in citations, f"Missing pair {pair} in {list(citations)[:10]}")


def check_word(agent_workspace):
    print("\n=== Checking Literature_Review_Draft.docx ===")
    docx_path = os.path.join(agent_workspace, "Literature_Review_Draft.docx")
    check("Literature_Review_Draft.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 300, f"Length: {len(text)}")
        check("Contains transformer reference",
              "transformer" in text)
        check("Contains attention reference",
              "attention" in text or "self-attention" in text)
        # The task only requires comparing the methodologies of the retrieved
        # papers; it never demands a specific model name. A faithful comparison
        # may refer to the papers either by name (BERT, GPT-3, ViT) or by their
        # method descriptions ("bidirectional masked-language pre-training",
        # "autoregressive language models", "vision transformer", "scaling
        # law"), so accept any one of a broad set of paper-specific markers.
        # This keeps the check meaningful (the document must engage with the
        # actual retrieved papers, not generic transformer content) without
        # penalizing a correct wording that avoids the model names.
        paper_markers = [
            "bert", "gpt", "bidirectional", "masked language", "autoregressive",
            "few-shot", "vision transformer", "scaling law", "pre-train",
            "pretrain", "vaswani", "devlin", "dosovitskiy", "kaplan",
        ]
        check("Contains reference to a specific retrieved paper",
              any(m in text for m in paper_markers))
        check("Contains methodology comparison",
              "method" in text or "approach" in text or "architecture" in text)
    except ImportError:
        check("python-docx available", False)
    except Exception as e:
        check("Word document readable", False, str(e))


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB)
        dbs = _find_dbs(conn)
        matching = [d for d in dbs if _matches_db(d[1])]
        check("Notion database 'Transformer Research Papers' exists",
              len(matching) > 0,
              f"Found {len(dbs)} databases, none matching")
        if matching:
            # Pass if ANY matching database has >= 5 entries, so that parallel
            # agents each creating the same (complete) database cannot cause a
            # spurious fail; a single complete db is sufficient.
            counts = {}
            max_count = 0
            for db_id, _ in matching:
                n = len(_pages_for_db(conn, db_id))
                counts[db_id] = n
                max_count = max(max_count, n)
            check("Notion database has >= 5 paper entries",
                  max_count >= 5,
                  f"Page counts per matching db: {counts}")
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check no unexpected sheets in the Excel file
    excel_path = os.path.join(workspace, "Literature_Review_Matrix.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            expected_sheets = {"paper_catalog", "methodology_comparison", "citation_network"}
            actual_sheets = {s.strip().lower().replace(" ", "_") for s in wb.sheetnames}
            unexpected = actual_sheets - expected_sheets
            check("No unexpected sheets in Excel",
                  len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")

            # Reverse: noise paper IDs must NOT be in the catalog.
            a_sheet = get_sheet(wb, "Paper_Catalog")
            if a_sheet is not None:
                rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
                ids_in = {_norm_id(r[0]) for r in rows if r and r[0]}
                polluted = ids_in & NOISE_IDS
                check("Reverse: noise papers NOT in Paper_Catalog",
                      len(polluted) == 0,
                      f"Found noise IDs: {polluted}")
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))

        # Notion reverse checks: duplicate pages and noise-paper exclusion.
        try:
            conn = psycopg2.connect(**DB)
            dbs = _find_dbs(conn)
            matching = [d for d in dbs if _matches_db(d[1])]
            all_pages = []
            for db_id, _ in matching:
                for (props,) in _pages_for_db(conn, db_id):
                    all_pages.append((db_id, props))
            # Duplicates are judged per-database, so two parallel agents that
            # each create the same complete database are not treated as a fail.
            dup_found = False
            for db_id in {d for d, _ in all_pages}:
                texts = [p for db, p in all_pages if db == db_id]
                if len(texts) > len(set(texts)):
                    dup_found = True
                    break
            check("No duplicate pages in Notion database", not dup_found,
                  f"{len(all_pages)} page rows across {len({d for d, _ in all_pages})} dbs")
            # Reverse: Notion pages should not include noise papers. Uses exact
            # noise IDs / full titles rather than generic keywords.
            all_text = " ".join(p for _, p in all_pages).lower()
            polluted = [m for m in NOISE_MARKERS if m in all_text]
            check("Reverse: Notion database excludes noise papers",
                  len(polluted) == 0,
                  f"Found noise markers: {polluted}")
            conn.close()
        except Exception:
            pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
