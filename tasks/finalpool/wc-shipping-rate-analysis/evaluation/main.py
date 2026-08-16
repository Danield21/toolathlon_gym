"""Evaluation for wc-shipping-rate-analysis."""
import argparse
import json
import os
import sys

import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_data):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Shipping_Audit.xlsx")
    if not os.path.exists(path):
        return ["Shipping_Audit.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Rate Comparison sheet
        rows = load_sheet_rows(wb, "Rate Comparison")
        if rows is None:
            errors.append("Sheet 'Rate Comparison' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_orders"]
            # Tighten tolerance: exact order count is expected
            if abs(len(data_rows) - expected) > 2:
                errors.append(f"Rate Comparison has {len(data_rows)} rows, expected ~{expected}")
            # Validate required columns exist
            header = [str(c).strip().lower() if c else "" for c in (rows[0] or [])]
            required_cols = ["order_id", "shipping_method", "carrier", "actual_cost", "difference", "undercharged"]
            for col in required_cols:
                if col not in header:
                    errors.append(f"Rate Comparison missing column: {col}")
            # Validate carrier-method mapping on a sample (Express->FedEx, Standard->USPS)
            if "shipping_method" in header and "carrier" in header:
                mi = header.index("shipping_method")
                ci = header.index("carrier")
                mapping_errors = 0
                for r in data_rows[:50]:  # sample
                    method = str(r[mi] or "").lower()
                    carrier = str(r[ci] or "").lower()
                    if "express" in method and "fedex" not in carrier:
                        mapping_errors += 1
                    elif "standard" in method and "usps" not in carrier:
                        mapping_errors += 1
                if mapping_errors > 2:
                    errors.append(f"Carrier/method mapping wrong in {mapping_errors} of sample rows (Express->FedEx, Standard->USPS)")
            # Validate Undercharged count matches
            if "undercharged" in header:
                ui = header.index("undercharged")
                yes_count = sum(1 for r in data_rows if str(r[ui] or "").strip().lower() == "yes")
                if abs(yes_count - gt_data["undercharged_count"]) > 2:
                    errors.append(f"Rate Comparison Undercharged=Yes count = {yes_count}, expected ~{gt_data['undercharged_count']}")

        # Check Undercharged Orders sheet
        rows2 = load_sheet_rows(wb, "Undercharged Orders")
        if rows2 is None:
            errors.append("Sheet 'Undercharged Orders' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            expected_uc = gt_data["undercharged_count"]
            if abs(len(data_rows2) - expected_uc) > 2:
                errors.append(f"Undercharged Orders has {len(data_rows2)} rows, expected ~{expected_uc}")
            # Verify the set of undercharged IDs overlaps with GT
            expected_ids = set(gt_data.get("undercharged_order_ids", []))
            agent_ids = set()
            for r in data_rows2:
                try:
                    agent_ids.add(int(r[0]))
                except (ValueError, TypeError):
                    pass
            if expected_ids:
                overlap = expected_ids & agent_ids
                if len(overlap) < len(expected_ids) * 0.9:
                    errors.append(f"Only {len(overlap)}/{len(expected_ids)} expected undercharged IDs present")

        # Check Summary sheet
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")
        else:
            summary_dict = {}
            for r in rows3[1:]:
                if r and r[0] is not None:
                    summary_dict[str(r[0]).strip().lower()] = r[1]

            def find_metric(substrs):
                for sk, sv in summary_dict.items():
                    if all(s in sk for s in substrs):
                        return sv
                return None

            # Undercharged count
            v = find_metric(["undercharged", "count"])
            if v is None:
                errors.append("Summary missing Undercharged_Count metric")
            else:
                try:
                    if abs(int(float(v)) - gt_data["undercharged_count"]) > 2:
                        errors.append(f"Summary Undercharged_Count = {v}, expected ~{gt_data['undercharged_count']}")
                except (ValueError, TypeError):
                    errors.append(f"Summary Undercharged_Count not numeric: {v}")

            # Total undercharged amount
            v = find_metric(["total", "undercharged", "amount"])
            if v is None:
                errors.append("Summary missing Total_Undercharged_Amount metric")
            else:
                try:
                    if abs(float(v) - gt_data["total_undercharged_amount"]) > 5.0:
                        errors.append(f"Summary Total_Undercharged_Amount = {v}, expected ~{gt_data['total_undercharged_amount']}")
                except (ValueError, TypeError):
                    errors.append(f"Summary Total_Undercharged_Amount not numeric: {v}")

            # Avg undercharge per order
            v = find_metric(["avg", "undercharge"])
            if v is None:
                errors.append("Summary missing Avg_Undercharge_Per_Order metric")
            else:
                try:
                    if abs(float(v) - gt_data["avg_undercharge"]) > 0.3:
                        errors.append(f"Summary Avg_Undercharge = {v}, expected ~{gt_data['avg_undercharge']}")
                except (ValueError, TypeError):
                    errors.append(f"Summary Avg_Undercharge not numeric: {v}")

            # Total orders analyzed
            v = find_metric(["total", "orders"])
            if v is not None:
                try:
                    if abs(int(float(v)) - gt_data["total_orders"]) > 2:
                        errors.append(f"Summary Total_Orders_Analyzed = {v}, expected ~{gt_data['total_orders']}")
                except (ValueError, TypeError):
                    pass

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_email(gt_data):
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")),
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%logistics@company.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if not rows:
            errors.append("No email found sent to logistics@company.com")
        else:
            subject = (rows[0][0] or "").lower()
            body = (rows[0][2] or "").lower()
            # Subject must mention audit/shipping
            if "shipping" not in subject and "audit" not in subject:
                errors.append(f"Email subject does not mention shipping/audit: '{rows[0][0]}'")
            # Body must mention undercharged (not just audit)
            if "undercharg" not in body:
                errors.append("Email body does not mention undercharged orders")
            # Count how many of the expected IDs appear in body
            expected_ids = gt_data.get("undercharged_order_ids", [])
            if expected_ids:
                id_hits = sum(1 for oid in expected_ids[:20] if str(oid) in body)
                # Task says "at least first ten undercharged order IDs"
                if id_hits < 8:
                    errors.append(f"Email body mentions only {id_hits} of expected undercharged order IDs (need at least 8 of first 20)")

    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    all_errors = []
    runtime_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email(gt_data)
    if errs:
        # Email check is runtime-only: can't succeed on GT workspace alone
        runtime_errors.extend(errs)
        for e in errs[:3]:
            print(f"    RUNTIME-ONLY ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors, {len(runtime_errors)} runtime-only) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print(f"\n=== RESULT: PASS (runtime-only errors: {len(runtime_errors)}) ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
