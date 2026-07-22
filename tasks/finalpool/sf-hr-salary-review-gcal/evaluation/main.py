"""
Evaluation script for sf-hr-salary-review-gcal task.

Checks:
1. Excel file with salary band data (values verified against DB with tolerance)
2. At least 7 gcal events with "salary review" in summary
3. Email sent with correct subject
"""

import argparse
import json
import os
import sys

import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def get_expected_stats():
    """Query actual salary stats from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT "DEPARTMENT",
               MIN("SALARY"),
               PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY "SALARY"),
               PERCENTILE_CONT(0.50) WITHIN GROUP (ORDER BY "SALARY"),
               PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY "SALARY"),
               MAX("SALARY"),
               AVG("SALARY"),
               COUNT(*)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    dept_stats = cur.fetchall()

    cur.execute('SELECT COUNT(*) FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    total_emp = cur.fetchone()[0]

    cur.execute("""
        SELECT "DEPARTMENT", AVG("SALARY") as avg_sal
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY avg_sal DESC
    """)
    dept_avgs = cur.fetchall()
    highest_avg_dept = dept_avgs[0][0]
    lowest_avg_dept = dept_avgs[-1][0]

    cur.execute('SELECT AVG("SALARY") FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    overall_avg = float(cur.fetchone()[0])

    cur.close()
    conn.close()

    return dept_stats, total_emp, highest_avg_dept, lowest_avg_dept, overall_avg


def check_excel(workspace):
    """Check Excel file against DB values."""
    from openpyxl import load_workbook

    errors = []
    xlsx_path = os.path.join(workspace, "Salary_Band_Analysis.xlsx")
    if not os.path.exists(xlsx_path):
        return ["Salary_Band_Analysis.xlsx not found"]

    dept_stats, total_emp, highest_avg_dept, lowest_avg_dept, overall_avg = get_expected_stats()

    wb = load_workbook(xlsx_path)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Department Bands sheet
    if "department bands" not in sheet_names_lower:
        errors.append(f"Missing 'Department Bands' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("department bands")]]
        data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)
        # Strict == 7 (was < 7)
        if data_rows != 7:
            errors.append(f"Department Bands has {data_rows} rows, expected exactly 7")

        # Check headers
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else "" for cell in ws[1]]
        for rh in ["department", "min_salary", "p25_salary", "median_salary", "p75_salary",
                   "max_salary", "avg_salary", "employee_count"]:
            if not any(rh in h or rh.replace("_", "") in h.replace("_", "") for h in headers):
                errors.append(f"Department Bands missing header: {rh}")

        # Build column index map
        col_map = {}
        for idx, h in enumerate(headers):
            if h == "department" or h.startswith("department"):
                col_map["dept"] = idx
            elif "min" in h and "salary" in h:
                col_map["min"] = idx
            elif "p25" in h or "25" in h:
                col_map["p25"] = idx
            elif "median" in h or "p50" in h or "50" in h:
                col_map["median"] = idx
            elif "p75" in h or "75" in h:
                col_map["p75"] = idx
            elif "max" in h and "salary" in h:
                col_map["max"] = idx
            elif "avg" in h and "salary" in h:
                col_map["avg"] = idx
            elif "count" in h or "employee" in h:
                col_map["count"] = idx

        # Build dept lookup (DB)
        dept_db = {row[0]: row for row in dept_stats}

        if "dept" in col_map:
            seen = set()
            for row in ws.iter_rows(min_row=2):
                if row[col_map["dept"]].value is None:
                    continue
                dept_name = str(row[col_map["dept"]].value).strip()
                seen.add(dept_name.lower())
                # Look up case-insensitive
                exp = None
                for k, v in dept_db.items():
                    if k.lower() == dept_name.lower():
                        exp = v
                        break
                if exp is None:
                    errors.append(f"Department '{dept_name}' not found in DB")
                    continue
                # Validate each numeric column - tighten Min/Max to 1 (deterministic salaries)
                fields = [
                    ("min", float(exp[1]), 1),
                    ("p25", float(exp[2]), 200),
                    ("median", float(exp[3]), 200),
                    ("p75", float(exp[4]), 200),
                    ("max", float(exp[5]), 1),
                    ("avg", float(exp[6]), 200),
                    ("count", int(exp[7]), 5),
                ]
                for fname, expected, tol in fields:
                    if fname not in col_map:
                        continue
                    actual_raw = row[col_map[fname]].value
                    if actual_raw is None:
                        errors.append(f"{dept_name}.{fname}: missing value")
                        continue
                    try:
                        actual = float(actual_raw)
                        if abs(actual - expected) > tol:
                            errors.append(f"{dept_name}.{fname}: got {actual:.2f}, expected ~{expected:.2f} (tol {tol})")
                    except (TypeError, ValueError):
                        errors.append(f"{dept_name}.{fname}: cannot parse {actual_raw}")
            # Verify all 7 departments present
            for dept_db_name in dept_db.keys():
                if dept_db_name.lower() not in seen:
                    errors.append(f"Department '{dept_db_name}' missing in output")

    # Check Summary sheet
    if "summary" not in sheet_names_lower:
        errors.append(f"Missing 'Summary' sheet. Found: {wb.sheetnames}")
    else:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        summary_data = {}
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                key = str(row[0].value).lower().replace(" ", "_")
                summary_data[key] = row[1].value

        # Total_Employees
        total_key = next((k for k in summary_data if "total" in k and "emp" in k), None)
        if total_key:
            try:
                val = int(float(summary_data[total_key]))
                if val != total_emp:
                    errors.append(f"Total_Employees: got {val}, expected {total_emp}")
            except (TypeError, ValueError):
                errors.append(f"Cannot parse Total_Employees: {summary_data[total_key]}")
        else:
            errors.append("Summary missing Total_Employees row")

        # Highest_Avg_Dept
        h_key = next((k for k in summary_data if "highest" in k), None)
        if h_key:
            actual = str(summary_data[h_key] or "").strip().lower()
            if actual != highest_avg_dept.lower():
                errors.append(f"Highest_Avg_Dept: got '{summary_data[h_key]}', expected '{highest_avg_dept}'")
        else:
            errors.append("Summary missing Highest_Avg_Dept row")

        # Lowest_Avg_Dept
        l_key = next((k for k in summary_data if "lowest" in k), None)
        if l_key:
            actual = str(summary_data[l_key] or "").strip().lower()
            if actual != lowest_avg_dept.lower():
                errors.append(f"Lowest_Avg_Dept: got '{summary_data[l_key]}', expected '{lowest_avg_dept}'")
        else:
            errors.append("Summary missing Lowest_Avg_Dept row")

        # Overall_Avg_Salary
        o_key = next((k for k in summary_data if "overall" in k or ("avg" in k and "salary" in k and "dept" not in k)), None)
        if o_key:
            try:
                val = float(summary_data[o_key])
                if abs(val - overall_avg) > 100:
                    errors.append(f"Overall_Avg_Salary: got {val:.2f}, expected ~{overall_avg:.2f}")
            except (TypeError, ValueError):
                errors.append(f"Cannot parse Overall_Avg_Salary: {summary_data[o_key]}")
        else:
            errors.append("Summary missing Overall_Avg_Salary row")

    return errors


def check_gcal(cur):
    """Check for salary review calendar events."""
    errors = []
    cur.execute("""
        SELECT id, summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%salary review%'
        AND start_datetime >= '2026-03-10T00:00:00'
        AND start_datetime < '2026-03-17T00:00:00'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()

    # Strict count == 7
    if len(events) != 7:
        errors.append(f"Found {len(events)} salary review events, expected exactly 7")

    # Get department list from DB
    cur.execute("""
        SELECT DISTINCT "DEPARTMENT"
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        ORDER BY "DEPARTMENT"
    """)
    dept_names = [r[0] for r in cur.fetchall()]

    # Each department should appear in some event summary as a whole word
    # (avoid 'HR' matching 'CHRO').
    import re as _re_dept
    for dept in dept_names:
        pat = r"\b" + _re_dept.escape(dept.lower()) + r"\b"
        if not any(_re_dept.search(pat, (e[1] or "").lower()) for e in events):
            errors.append(f"GCal: no event for department '{dept}' (word-boundary)")

    # Each of the 7 days should have one event
    expected_dates = [f"2026-03-{d:02d}" for d in range(10, 17)]
    event_dates = []
    for e in events:
        if e[2]:
            try:
                event_dates.append(e[2].date().isoformat())
            except AttributeError:
                event_dates.append(str(e[2])[:10])
    for d in expected_dates:
        if d not in event_dates:
            errors.append(f"GCal: no event on {d}")

    # Each event 10:00-11:00 America/New_York. The DB may store as naive local time
    # (interpret as ET) or as UTC. Accept hour==10 or the UTC equivalent (14 EDT / 15 EST).
    # Tighten duration to abs(min-60) <= 1.
    for e in events:
        if e[2] and e[3]:
            try:
                start_hour = e[2].hour
                if start_hour not in (10, 14, 15):
                    errors.append(f"GCal: event '{e[1]}' not 10:00 ET (got {start_hour:02d}:00)")
                if e[2].minute != 0:
                    errors.append(f"GCal: event '{e[1]}' not on the hour (minute={e[2].minute})")
                duration_min = (e[3] - e[2]).total_seconds() / 60
                if abs(duration_min - 60) > 1:
                    errors.append(f"GCal: event '{e[1]}' duration must be 60 min ±1 (got {duration_min} min)")
            except AttributeError:
                pass

    return errors


def check_email(cur, highest_avg_dept, lowest_avg_dept):
    """Check for the summary email."""
    errors = []
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%quarterly salary band analysis%'
    """)
    emails = cur.fetchall()

    if not emails:
        errors.append("No email with subject 'Quarterly Salary Band Analysis' found")
    else:
        email = emails[0]
        from_str = str(email[1] or "").lower()
        if "hr-analytics@company.com" not in from_str:
            errors.append(f"Email from_addr not hr-analytics@company.com (got {email[1]})")
        to_str = str(email[2] or "").lower()
        if "hr-director@company.com" not in to_str:
            errors.append(f"Email not sent to hr-director@company.com, to_addr: {email[2]}")
        body = str(email[3] or "").lower()
        if highest_avg_dept and highest_avg_dept.lower() not in body:
            errors.append(f"Email body does not mention highest avg dept '{highest_avg_dept}'")
        if lowest_avg_dept and lowest_avg_dept.lower() not in body:
            errors.append(f"Email body does not mention lowest avg dept '{lowest_avg_dept}'")

    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    # Check Excel
    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    # Check GCal and Email
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Get expected info for email body check
        try:
            _, _, highest_avg_dept, lowest_avg_dept, _ = get_expected_stats()
        except Exception:
            highest_avg_dept = lowest_avg_dept = None

        print("\n=== Checking GCal Events ===")
        gcal_errors = check_gcal(cur)
        if gcal_errors:
            for e in gcal_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(gcal_errors)
        else:
            print("  [PASS] GCal check passed")

        print("\n=== Checking Email ===")
        email_errors = check_email(cur, highest_avg_dept, lowest_avg_dept)
        if email_errors:
            for e in email_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(email_errors)
        else:
            print("  [PASS] Email check passed")

        cur.close()
        conn.close()
    except Exception as e:
        err = f"DB check error: {e}"
        print(f"  [FAIL] {err}")
        all_errors.append(err)

    # Summary
    print(f"\n=== SUMMARY ===")
    if all_errors:
        for e in all_errors:
            print(f"  [ERROR] {e}")
        print("  Overall: FAIL")
    else:
        print("  Overall: PASS")

    if args.res_log_file:
        result = {"errors": all_errors, "success": len(all_errors) == 0}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(1 if all_errors else 0)


if __name__ == "__main__":
    main()
