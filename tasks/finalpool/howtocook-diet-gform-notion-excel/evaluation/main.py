"""
Evaluation for howtocook-diet-gform-notion-excel task.
Checks: GForm with questions, Notion page, Excel with 2 sheets, email.

Design notes (fixed vs original):
- All DB connections read PG* environment variables with safe defaults (no
  hardcoded dbname/port).
- Google-Forms questions are distinguished by title + options content, using
  only the question types the google-forms MCP can actually produce
  ('textQuestion' / 'choiceQuestion' with config.type 'RADIO'). No checkbox /
  scale / dropdown / paragraph requirements.
- Options are parsed from the JSONB config (array of {"value": ...} dicts or
  plain strings) and compared case-insensitively.
- Excel is read twice (data_only=True for cached values, data_only=False to
  detect formulas). Category-count cells that are formulas use their cached
  value, or a best-effort evaluation of simple =COUNTIF(...) formulas, so a
  natural Excel solution using COUNTIF is not unfairly failed.
- Category Summary columns are located by header names (fallback: position),
  and header / aggregation ("total") rows are skipped.
"""
import argparse
import os
import re
import sys

import psycopg2
import openpyxl
from openpyxl.utils import column_index_from_string

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0
RUNTIME_ONLY_FAIL = 0
# True when the eval runs against the groundtruth workspace itself; the gform /
# notion / email are solver side-effects absent from the GT DB snapshot, so
# their existence checks are skipped (not failed) in that mode.
IS_GT_SELF_TEST = False


def record(name, passed, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, RUNTIME_ONLY_FAIL
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if runtime_only:
            RUNTIME_ONLY_FAIL += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


# ---------------------------------------------------------------------------
# Robust value helpers
# ---------------------------------------------------------------------------

def _to_float(v):
    """Return float(v) for numbers and numeric strings; None if not parseable.

    Handles thousand separators, currency symbols, a trailing % and whitespace.
    Formula strings (starting with '=') return None here (caller handles them).
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("="):
        return None
    if s.endswith("%"):
        s = s[:-1]
    s = s.replace(",", "").replace("$", "").replace("¥", "").replace("€", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _cell_is_formula(v):
    return isinstance(v, str) and v.lstrip().startswith("=")


def _eval_countif(formula, wb, src_sheet_name=None):
    """Best-effort evaluation of a simple =COUNTIF(range, criteria) formula.

    Supported ranges:  Sheet!A:A, 'Sheet Name'!A:B, Sheet!$A$1:$B$10, bare A:A.
    Supported criteria: "literal", a cell reference (e.g. A2), or a plain token.
    src_sheet_name is the sheet the formula lives in (used to resolve criteria
    cell references such as A2, which are relative to that sheet).
    Returns an int count, or None if the formula cannot be understood.
    """
    s = formula.strip()
    m = re.fullmatch(r"=\s*COUNTIF\s*\(\s*(.+?)\s*,\s*(.+?)\s*\)", s, re.IGNORECASE)
    if not m:
        return None
    range_ref = m.group(1).strip()
    criteria = m.group(2).strip()

    mrange = re.match(
        r"^(?:'([^']+)'|([^!]+))!\s*\$?([A-Za-z]+)\$?(\d*)(?::\s*\$?([A-Za-z]+)\$?(\d*))?$",
        range_ref,
    )
    if mrange:
        sheet_name = mrange.group(1) or mrange.group(2)
        col1 = mrange.group(3)
        row1 = int(mrange.group(4)) if mrange.group(4) else 1
        col2 = mrange.group(5) if mrange.group(5) else col1
        row2 = int(mrange.group(6)) if mrange.group(6) else None
    else:
        mrange2 = re.match(r"^\$?([A-Za-z]+)\$?(\d*)(?::\s*\$?([A-Za-z]+)\$?(\d*))?$", range_ref)
        if not mrange2:
            return None
        sheet_name = None
        col1 = mrange2.group(1)
        row1 = int(mrange2.group(2)) if mrange2.group(2) else 1
        col2 = mrange2.group(3) if mrange2.group(3) else col1
        row2 = int(mrange2.group(4)) if mrange2.group(4) else None

    if criteria.startswith('"') and criteria.endswith('"') and len(criteria) >= 2:
        crit_val = criteria[1:-1]
    elif re.fullmatch(r"[A-Za-z]+\d+", criteria):
        mcell = re.match(r"([A-Za-z]+)(\d+)", criteria)
        col_letter, row_num = mcell.group(1).upper(), int(mcell.group(2))
        try:
            cc = column_index_from_string(col_letter)
        except Exception:
            return None
        # A cell reference in a formula is relative to the sheet the formula
        # lives in (e.g. =COUNTIF(Recipes!B:B, A2) on the Category Summary
        # sheet -> A2 is Category Summary!A2).
        if src_sheet_name and src_sheet_name in wb.sheetnames:
            ws = wb[src_sheet_name]
            crit_val = ws.cell(row=row_num, column=cc).value if 1 <= row_num <= ws.max_row else None
        else:
            crit_val = None
        if crit_val is None:
            return None
    else:
        crit_val = criteria

    target = None
    for sn in wb.sheetnames:
        if sheet_name is None or sn.lower().strip("'") == sheet_name.lower().strip("'"):
            target = wb[sn]
            break
    if target is None:
        return None

    def _cidx(c):
        try:
            return column_index_from_string(c)
        except Exception:
            return None

    c1, c2 = _cidx(col1), _cidx(col2 if col2 else col1)
    if c1 is None or c2 is None:
        return None
    if c1 > c2:
        c1, c2 = c2, c1

    crit = str(crit_val).strip().lower() if crit_val is not None else ""
    cnt = 0
    r_hi = row2 if row2 is not None else target.max_row
    for r in range(row1, min(r_hi, target.max_row) + 1):
        for c in range(c1, c2 + 1):
            v = target.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().lower() == crit:
                cnt += 1
    return cnt


def _extract_options(cfg):
    """Extract option value strings from a question config.

    The google-forms MCP stores options as a JSONB array of {"value": ...}
    dicts; direct DB inserts may store plain strings. Returns a list of str.
    """
    if not isinstance(cfg, dict):
        return []
    raw = cfg.get("options")
    if raw is None:
        raw = cfg.get("choices", [])
    if isinstance(raw, str):
        return [raw]
    if isinstance(raw, (list, tuple)):
        out = []
        for o in raw:
            if isinstance(o, dict):
                v = o.get("value")
                if v is None:
                    v = o.get("label", "")
                out.append(str(v))
            else:
                out.append(str(o))
        return out
    return []


def _question_kind(qtype, cfg):
    """Normalise a question to one of: 'text', 'choice', 'unknown'.

    The google-forms MCP writes question_type 'textQuestion'/'choiceQuestion'
    (choice config.type = 'RADIO'). Accept those plus common aliases. This
    deliberately never requires MCP-impossible types (checkbox / scale /
    dropdown / paragraph / linear-scale).
    """
    t = (qtype or "").strip().lower()
    cfgtype = ""
    if isinstance(cfg, dict) and cfg.get("type") is not None:
        cfgtype = str(cfg.get("type")).strip().upper()
    if t in ("textquestion", "text", "short_answer", "shortanswer", "textarea"):
        return "text"
    if t in ("choicequestion", "radio", "multiple_choice", "multiple-choice",
             "multiplechoice", "checkbox", "list", "dropdown", "drop_down"):
        return "choice"
    if cfgtype in ("RADIO", "CHECKBOX", "MULTIPLE_CHOICE", "DROPDOWN", "LIST"):
        return "choice"
    return "unknown"


# ---------------------------------------------------------------------------
# Google Forms check
# ---------------------------------------------------------------------------

def check_gform():
    print("\n=== Checking Google Form ===")
    if IS_GT_SELF_TEST:
        print("  [SKIP] Google Form check: GT self-test (side-effect not present in GT snapshot)")
        return
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()

        target_form = None
        for fid, title in forms:
            tlow = (title or "").lower().strip()
            if "dietary preference" in tlow and "survey" in tlow:
                target_form = fid
                break
        if target_form is None:
            for fid, title in forms:
                tlow = (title or "").lower().strip()
                if "dietary" in tlow and ("preference" in tlow or "survey" in tlow):
                    target_form = fid
                    break
        # Form existence is BLOCKING: the task requires creating the
        # "Dietary Preference Survey".
        record("GForm 'Dietary Preference Survey' exists",
               target_form is not None,
               f"Found forms: {[t for _, t in forms]}")

        if target_form is None:
            conn.close()
            return

        cur.execute(
            "SELECT title, question_type, required, config FROM gform.questions "
            "WHERE form_id = %s ORDER BY position",
            (target_form,),
        )
        questions = cur.fetchall()
        # Question count is BLOCKING: task.md requires at least four questions.
        record("GForm has at least 4 questions", len(questions) >= 4,
               f"Found {len(questions)} questions")

        q_type_labels = [q[1] for q in questions]
        kinds = [_question_kind(q[1], q[3]) for q in questions]
        has_choice = any(k == "choice" for k in kinds)
        has_text = any(k == "text" for k in kinds)
        record("GForm has a multiple-choice question", has_choice,
               f"Question types: {q_type_labels}",
               runtime_only=True)
        record("GForm has a text question", has_text,
               f"Question types: {q_type_labels}",
               runtime_only=True)

        # Question about dietary restrictions (by title/content, not type)
        q_titles_lower = [q[0].lower() if q[0] else "" for q in questions]
        has_dietary_q = any("dietary" in qt or "restriction" in qt or "vegetarian" in qt
                            for qt in q_titles_lower)
        record("GForm has dietary restrictions question", has_dietary_q,
               f"Question titles: {q_titles_lower}",
               runtime_only=True)

        # --- Option set verification via questions.config ---
        meal_q_ok = False
        meal_q_opts = None
        restr_q_ok = False
        restr_q_opts = None
        scale_ok = False
        scale_detail = None
        for title, qtype, required, cfg in questions:
            tlow = (title or "").lower()
            kind = _question_kind(qtype, cfg)
            opts_lower = [str(o).strip().lower() for o in _extract_options(cfg)]
            # Meal-types question (single choice with the four meal options)
            if ("meal" in tlow and "week" not in tlow and "how many" not in tlow
                    and "home-cooked" not in tlow) and kind == "choice":
                needed = {"breakfast", "lunch", "dinner", "snacks"}
                if needed.issubset(set(opts_lower)):
                    meal_q_ok = True
                meal_q_opts = opts_lower
            # Dietary-restrictions question (single choice with the five options)
            if ("dietary" in tlow or "restriction" in tlow) and kind == "choice":
                needed = {"vegetarian", "vegan", "gluten-free", "dairy-free", "none"}
                if needed.issubset(set(opts_lower)):
                    restr_q_ok = True
                restr_q_opts = opts_lower
            # Home-cooked-meals-per-week question: single choice with options 1..7
            if ("meal" in tlow and ("week" in tlow or "home-cooked" in tlow
                                    or "how many" in tlow)) and kind == "choice":
                nums = [_to_float(o) for o in _extract_options(cfg)]
                if nums and all(n is not None for n in nums):
                    ints = sorted(int(n) for n in nums)
                    if ints == list(range(1, 8)):
                        scale_ok = True
                scale_detail = f"options={opts_lower}"
        record("Meal-types question has Breakfast/Lunch/Dinner/Snacks options",
               meal_q_ok,
               f"Options: {meal_q_opts}",
               runtime_only=True)
        record("Dietary-restrictions question has Veg/Vegan/GF/DF/None options",
               restr_q_ok,
               f"Options: {restr_q_opts}",
               runtime_only=True)
        record("Home-cooked-meals question has options 1..7",
               scale_ok,
               f"{scale_detail}",
               runtime_only=True)

        conn.close()
    except Exception as e:
        record("GForm connection", False, str(e), runtime_only=True)


# ---------------------------------------------------------------------------
# Notion check
# ---------------------------------------------------------------------------

def check_notion():
    print("\n=== Checking Notion Page ===")
    if IS_GT_SELF_TEST:
        print("  [SKIP] Notion page check: GT self-test (side-effect not present in GT snapshot)")
        return
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false AND in_trash = false
        """)
        pages = cur.fetchall()

        # Strict title match: "Healthy Recipe Knowledge Base"
        target_page = None
        for pid, props in pages:
            props_text = str(props).lower() if props else ""
            if "healthy recipe knowledge base" in props_text:
                target_page = pid
                break
        # Fallback loose match (so agent-created pages with minor variations still found)
        if target_page is None:
            for pid, props in pages:
                props_text = str(props).lower() if props else ""
                if "healthy recipe" in props_text and ("knowledge" in props_text or "base" in props_text):
                    target_page = pid
                    break

        # Notion page existence is BLOCKING: the task requires creating the
        # "Healthy Recipe Knowledge Base" page.
        record("Notion page 'Healthy Recipe Knowledge Base' exists",
               target_page is not None,
               f"Searched {len(pages)} pages")

        if target_page is not None:
            cur.execute("""
                SELECT id, type, block_data FROM notion.blocks
                WHERE parent_id = %s AND archived = false AND in_trash = false
                ORDER BY position
            """, (target_page,))
            blocks = cur.fetchall()
            record("Notion page has at least 6 content blocks", len(blocks) >= 6,
                   f"Found {len(blocks)} blocks (task requires at least 6 recipes)",
                   runtime_only=True)

            recipe_keywords = [
                "tomato and egg", "steamed fish", "kung pao", "mapo tofu",
                "stir-fried broccoli", "congee",
                "vegetables", "seafood", "meat", "tofu", "porridge"
            ]
            recipe_mention_blocks = 0
            for _, btype, bdata in blocks:
                text = str(bdata or "").lower()
                # Match English recipe keywords, or any block with substantive
                # text content (agents may legitimately write Chinese recipes).
                content_len = len(re.sub(r"\s", "", text))
                if any(kw in text for kw in recipe_keywords) or content_len >= 8:
                    recipe_mention_blocks += 1
            record("Notion page has recipe-content blocks (>=6)",
                   recipe_mention_blocks >= 6,
                   f"Found {recipe_mention_blocks} recipe-related blocks",
                   runtime_only=True)

        conn.close()
    except Exception as e:
        record("Notion connection", False, str(e), runtime_only=True)


# ---------------------------------------------------------------------------
# Excel check
# ---------------------------------------------------------------------------

def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _col_index(headers, *candidates):
    for c in candidates:
        cnorm = str(c).strip().lower().replace("_", " ")
        for i, h in enumerate(headers):
            hnorm = str(h or "").strip().lower().replace("_", " ")
            if hnorm == cnorm or cnorm in hnorm or hnorm in cnorm:
                return i
    return None


def _find_header_row(ws, keywords, max_scan=5):
    """Return (header_row_1based_or_None, headers_list).

    Scans the first max_scan rows for the one containing the most keyword hits.
    Returns (None, row1_cells) when no row has any keyword hit (i.e. the sheet
    has no header row), so callers can decide whether data starts on row 1."""
    best_row, best_count, best_headers = None, 0, None
    for r in range(1, min(max_scan, ws.max_row or 1) + 1):
        cells = [c.value for c in ws[r]]
        joined = " ".join(str(v or "").strip().lower().replace("_", " ") for v in cells)
        count = sum(1 for k in keywords if k in joined)
        if count > best_count:
            best_count = count
            best_row, best_headers = r, cells
    if best_row is None:
        best_row, best_headers = None, [c.value for c in ws[1]]
    return best_row, best_headers


def _pairs_nonempty(pairs):
    for cval, rval in pairs:
        if cval is not None and str(cval).strip() != "":
            return True
        if rval is not None and str(rval).strip() != "":
            return True
    return False


def _resolve_cell(cval, rval, wb, src_sheet_name=None):
    """Resolve one cell's value to a float, or None if unresolvable.

    Handles literal numbers/strings and formula cells (cached value first,
    else best-effort COUNTIF evaluation against the raw workbook)."""
    num = _to_float(cval)
    if num is not None:
        return num
    if _cell_is_formula(rval):
        return _eval_countif(rval, wb, src_sheet_name)
    return None


def _is_agg_row(pairs):
    joined = " ".join(str((c if c is not None else r) or "").strip().lower()
                      for c, r in pairs)
    for w in ("total", "sum", "合计", "总计", "grand total"):
        if w in joined:
            return True
    return False


def _load_howtocook_recipe_names():
    """Try several known paths for the howtocook all_recipes.json file."""
    import json
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.environ.get("HOWTOCOOK_ALL_RECIPES", ""),
        # Relative to this evaluation file (repo layout NewBenchmark/T4/<task>/evaluation):
        os.path.join(here, "../../../../toolathon复现/local_servers/HowToCook-mcp/src/data/all_recipes.json"),
        os.path.join(here, "../../../../toolathon复现/local_servers/HowToCook-mcp/build/data/all_recipes.json"),
        os.path.join(here, "../../../toolathon复现/local_servers/HowToCook-mcp/src/data/all_recipes.json"),
        # Relative to cwd (harness may run from the repo root)
        "toolathon复现/local_servers/HowToCook-mcp/src/data/all_recipes.json",
        "local_servers/HowToCook-mcp/src/data/all_recipes.json",
        # Original dev-machine paths (kept for local runs)
        "/Users/puzhen/PycharmProjects/toolathon_new/Toolathlon_Pack/local_servers/HowToCook-mcp/build/data/all_recipes.json",
        "/Users/puzhen/PycharmProjects/toolathon_new/Toolathlon_Pack/local_servers/HowToCook-mcp/src/data/all_recipes.json",
        "/Users/puzhen/PycharmProjects/toolathon_new/local_mcp_servers/HowToCook-mcp/src/data/all_recipes.json",
    ]
    seen = set()
    for p in candidates:
        if not p or p in seen:
            continue
        seen.add(p)
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            names = set()
            for r in data:
                n = r.get("name") or ""
                if n:
                    names.add(str(n).strip())
                    if "的做法" in n:
                        names.add(n.replace("的做法", "").strip())
            if names:
                return names
        except Exception:
            continue
    return None


# A small known-good set of common English transliterations used historically
# in tasks. Allows agents using English names to pass even when the underlying
# DB has only Chinese names.
_ENGLISH_RECIPE_HINTS = [
    "tomato", "egg", "fish", "chicken", "tofu", "broccoli", "congee",
    "noodle", "rice", "soup", "pork", "beef", "shrimp", "vegetable",
    "salad", "stir-fr", "stir fr", "porridge", "dumpling", "bun",
    "pancake", "bread", "cake", "smoothie", "juice", "stew", "fry",
    "roast", "boil", "steam", "grill", "bake", "duck", "lamb", "mushroom",
    "potato", "cabbage", "spinach", "garlic", "ginger", "onion", "carrot",
    "bean", "pepper", "chilli", "chili", "soy", "sauce",
    "kung pao", "mapo", "ma po", "wonton", "dim sum", "scallion", "leek",
    "cucumber", "eggplant", "zucchini", "pumpkin", "squash",
]


def _recipe_in_db(name, db_names):
    """Return True if a recipe name is plausibly real.

    Accepts (a) exact match against db_names (Chinese), (b) English text
    that contains common food keywords from _ENGLISH_RECIPE_HINTS.
    """
    if not name:
        return False
    n = str(name).strip()
    nlow = n.lower()
    if n in db_names:
        return True
    if (n + "的做法") in db_names:
        return True
    for dn in db_names:
        if n in dn or dn in n:
            return True
    if any(kw in nlow for kw in _ENGLISH_RECIPE_HINTS):
        return True
    return False


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel File ===")
    xl_path = os.path.join(agent_workspace, "Recipe_Overview.xlsx")
    if not os.path.isfile(xl_path):
        record("Excel file Recipe_Overview.xlsx exists", False, f"Not found at: {xl_path}")
        return
    record("Excel file Recipe_Overview.xlsx exists", True)

    try:
        # data_only=True -> cached formula values; data_only=False -> raw
        # (formula strings). Using both lets us handle literal values AND
        # formula cells (cached value or best-effort COUNTIF evaluation).
        wb_cached = openpyxl.load_workbook(xl_path, data_only=True)
        wb_raw = openpyxl.load_workbook(xl_path, data_only=False)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names = [s.lower() for s in wb_cached.sheetnames]
    has_recipes = any("recipe" in s for s in sheet_names)
    has_summary = any("category" in s or "summary" in s for s in sheet_names)
    record("Excel has 'Recipes' sheet", has_recipes, f"Found sheets: {wb_cached.sheetnames}")
    record("Excel has 'Category Summary' sheet", has_summary, f"Found sheets: {wb_cached.sheetnames}")

    def _parallel(name):
        """Return (cached_sheet, raw_sheet) for a sheet present in both loads."""
        if name not in wb_cached.sheetnames or name not in wb_raw.sheetnames:
            return None, None
        return wb_cached[name], wb_raw[name]

    # ---- Constraint-based validation (NOT strict GT recipe-name matching) ----
    # task.md: "select at least 6 recipes representing at least 3 different categories"
    # Agents may legitimately pick any valid recipes from the howtocook MCP, in
    # English or Chinese. Strict per-recipe matching would FN compliant agents.

    # Find Recipes sheet
    recipes_pair = None
    recipes_sheet_name = None
    for sname in wb_cached.sheetnames:
        if "recipe" in sname.lower() and "summary" not in sname.lower() and "category" not in sname.lower():
            recipes_pair = _parallel(sname)
            recipes_sheet_name = sname
            break

    # Find Category Summary sheet
    cat_pair = None
    cat_sheet_name = None
    for sname in wb_cached.sheetnames:
        if "category" in sname.lower() or "summary" in sname.lower():
            cat_pair = _parallel(sname)
            cat_sheet_name = sname
            break

    if recipes_pair is None or recipes_pair[0] is None:
        record("Recipes sheet present", False, f"Sheets: {wb_cached.sheetnames}")
        return

    recipes_c, recipes_r = recipes_pair

    # Locate the header row (tolerate a leading title row) and read column ids
    header_row, headers = _find_header_row(recipes_c, ["recipe", "category", "description"])
    headers = [str(h or "").strip() for h in headers]
    headers_lower = [str(h or "").strip().lower().replace("_", " ") for h in headers]
    name_idx = _col_index(headers_lower, "recipe_name", "name", "recipe name")
    cat_idx = _col_index(headers_lower, "category")
    desc_idx = _col_index(headers_lower, "description")

    # Collect data rows (non-empty), reading (cached, raw) pairs per cell.
    # Without a recognizable header row the task's required columns are missing,
    # so the column-presence check below will fail (correctly).
    data_start = (header_row + 1) if header_row is not None else 1
    recipe_rows = []
    for r in range(data_start, max(recipes_c.max_row, recipes_r.max_row) + 1):
        cvals = [c.value for c in recipes_c[r]]
        rvals = [c.value for c in recipes_r[r]]
        pairs = list(zip(cvals, rvals))
        if _pairs_nonempty(pairs):
            recipe_rows.append(pairs)
    record("Excel Recipes sheet has at least 6 recipe rows", len(recipe_rows) >= 6,
           f"Found {len(recipe_rows)} data rows")

    # Validate each recipe row: name + category + description non-empty
    if name_idx is not None and cat_idx is not None and desc_idx is not None:
        empty_count = 0
        for pairs in recipe_rows:
            n = str(pairs[name_idx][0] if pairs[name_idx][0] is not None else pairs[name_idx][1] or "").strip() \
                if name_idx < len(pairs) else ""
            c = str(pairs[cat_idx][0] if pairs[cat_idx][0] is not None else pairs[cat_idx][1] or "").strip() \
                if cat_idx < len(pairs) else ""
            d = str(pairs[desc_idx][0] if pairs[desc_idx][0] is not None else pairs[desc_idx][1] or "").strip() \
                if desc_idx < len(pairs) else ""
            if not n or not c or not d:
                empty_count += 1
        record("All recipe rows have non-empty Name/Category/Description",
               empty_count == 0,
               f"{empty_count} rows have empty Name/Category/Description")
    else:
        record("Recipes sheet has Recipe_Name/Category/Description columns",
               False, f"Headers: {headers}")

    # Validate recipe names exist in the howtocook recipe DB (allow English or
    # Chinese). Best-effort: skipped gracefully if the JSON cannot be located.
    #
    # RUNTIME-ONLY, deliberately: the howtocook DB stores only Chinese names
    # (e.g. "咖喱炒蟹", "白灼菜心"). The task and its ground truth are in
    # English, so a fully compliant model legitimately translates recipe names
    # to English, and many such translations (e.g. "Curry Fried Crab",
    # "Blanched Choy Sum") contain none of the _ENGLISH_RECIPE_HINTS keywords.
    # A keyword/DB-exact matcher can therefore never decide "this English name
    # is fabricated" without translating Chinese -> English, which we cannot do
    # reliably. Keeping this check non-runtime would wrongly FAIL compliant
    # English deliveries, and would also make the verdict depend on whether the
    # deployment happens to include all_recipes.json (repo-relative path), i.e.
    # the same deliverable would get different verdicts per environment. So it
    # is reported as a metric only; the real "realism" enforcement stays with
    # the count-consistency and row-completeness checks below.
    real_recipe_names = _load_howtocook_recipe_names()
    if real_recipe_names and name_idx is not None:
        unknown = []
        for pairs in recipe_rows:
            if name_idx >= len(pairs):
                continue
            cval, rval = pairs[name_idx]
            n_raw = cval if cval is not None else rval
            if n_raw is None:
                continue
            n = str(n_raw).strip()
            if not _recipe_in_db(n, real_recipe_names):
                unknown.append(n)
        valid_count = len(recipe_rows) - len(unknown)
        record("At least 5 selected recipes match howtocook DB names",
               valid_count >= 5,
               f"{valid_count}/{len(recipe_rows)} recognised; unknown: {unknown[:5]}",
               runtime_only=True)

    # Distinct categories >= 3
    cats = set()
    if cat_idx is not None:
        for pairs in recipe_rows:
            if cat_idx < len(pairs):
                cval, rval = pairs[cat_idx]
                v = cval if cval is not None else rval
                if v is not None and str(v).strip():
                    cats.add(str(v).strip().lower())
        record("Recipes span at least 3 distinct categories",
               len(cats) >= 3,
               f"Found {len(cats)} distinct categories: {sorted(cats)}")

    # Category Summary sheet: counts must equal actual row counts in Recipes
    if cat_pair is None or cat_pair[0] is None:
        record("Category Summary sheet present", False, f"Sheets: {wb_cached.sheetnames}")
        return

    cat_c, cat_r = cat_pair
    cat_header_row, cat_headers = _find_header_row(cat_c, ["category", "count", "number"])
    cat_headers_lower = [str(h or "").strip().lower().replace("_", " ") for h in cat_headers]
    cat_name_idx = _col_index(cat_headers_lower, "category", "category name", "name", "recipe category")
    cat_count_idx = _col_index(cat_headers_lower, "count", "number", "quantity", "num")

    # Without a header row, fall back to positional layout (col0 = category,
    # col1 = count) and start reading from row 1.
    cat_data_start = (cat_header_row + 1) if cat_header_row is not None else 1
    if cat_header_row is None:
        cat_name_idx = 0
        cat_count_idx = 1

    cat_summary_rows = []
    for r in range(cat_data_start, max(cat_c.max_row, cat_r.max_row) + 1):
        cvals = [c.value for c in cat_c[r]]
        rvals = [c.value for c in cat_r[r]]
        pairs = list(zip(cvals, rvals))
        if _pairs_nonempty(pairs) and not _is_agg_row(pairs):
            cat_summary_rows.append(pairs)

    record("Category Summary has at least 3 categories",
           len(cat_summary_rows) >= 3,
           f"Found {len(cat_summary_rows)} category rows")

    # Build expected counts from Recipes sheet
    if cat_idx is not None:
        expected_counts = {}
        for pairs in recipe_rows:
            if cat_idx < len(pairs):
                cval, rval = pairs[cat_idx]
                v = cval if cval is not None else rval
                if v is not None and str(v).strip():
                    key = str(v).strip().lower()
                    expected_counts[key] = expected_counts.get(key, 0) + 1
        # Build agent_summary {category: count}
        agent_summary = {}
        unverifiable = []
        for pairs in cat_summary_rows:
            key_cell = (pairs[cat_name_idx] if cat_name_idx is not None and cat_name_idx < len(pairs)
                        else (pairs[0] if pairs else (None, None)))
            cnt_cell = (pairs[cat_count_idx] if cat_count_idx is not None and cat_count_idx < len(pairs)
                        else (pairs[1] if len(pairs) > 1 else (None, None)))
            key = str((key_cell[0] if key_cell[0] is not None else key_cell[1]) or "").strip()
            if not key:
                continue
            cval, rval = cnt_cell if cnt_cell is not None else (None, None)
            count_val = _resolve_cell(cval, rval, wb_raw, cat_sheet_name)
            if count_val is None:
                # count column unreadable (e.g. unresolvable formula)
                unverifiable.append(f"{key} (cell={rval if rval is not None else cval})")
                continue
            agent_summary[key.lower()] = int(round(count_val))
        detail_suffix = f"; unverifiable: {unverifiable[:3]}" if unverifiable else ""
        # Sum-of-counts must equal total recipe rows
        total_summary_count = sum(agent_summary.values())
        record("Category Summary total count equals Recipes row count",
               total_summary_count == len(recipe_rows),
               f"Summary total={total_summary_count}, Recipes rows={len(recipe_rows)}{detail_suffix}")
        # All categories in Recipes should appear in Summary with correct count
        mismatch = []
        for k, v in expected_counts.items():
            a = agent_summary.get(k)
            if a != v:
                mismatch.append(f"{k}: expected {v}, got {a}")
        record("Each category count in summary matches Recipes-sheet count",
               len(mismatch) == 0,
               f"Mismatches: {mismatch[:5]}{detail_suffix}")


# ---------------------------------------------------------------------------
# Email check
# ---------------------------------------------------------------------------

def check_email():
    print("\n=== Checking Email ===")
    if IS_GT_SELF_TEST:
        print("  [SKIP] Email check: GT self-test (side-effect not present in GT snapshot)")
        return
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE LOWER(subject) LIKE '%healthy eating%'
               OR LOWER(subject) LIKE '%recipe selection%'
               OR LOWER(subject) LIKE '%healthy recipe%'
        """)
        emails = cur.fetchall()
        # Email existence is BLOCKING: the task requires sending the email.
        record("Email about healthy eating/recipes sent", len(emails) > 0,
               f"Found {len(emails)} matching emails")

        if emails:
            target_found = False
            for subject, to_addr in emails:
                to_str = str(to_addr).lower() if to_addr else ""
                if "wellness.team@company.com" in to_str:
                    target_found = True
                    break
            record("Email sent to wellness.team@company.com", target_found,
                   f"Recipients: {[e[1] for e in emails]}",
                   runtime_only=True)

        conn.close()
    except Exception as e:
        record("Email connection", False, str(e), runtime_only=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    global IS_GT_SELF_TEST
    try:
        _gt = os.path.realpath(args.groundtruth_workspace) if args.groundtruth_workspace else ""
        _ag = os.path.realpath(args.agent_workspace) if args.agent_workspace else ""
        IS_GT_SELF_TEST = bool(_gt) and bool(_ag) and _gt == _ag
    except Exception:
        IS_GT_SELF_TEST = False

    check_gform()
    check_notion()
    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    non_runtime_fail = FAIL_COUNT - RUNTIME_ONLY_FAIL
    overall = non_runtime_fail == 0
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (runtime-only fails: {RUNTIME_ONLY_FAIL})")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
