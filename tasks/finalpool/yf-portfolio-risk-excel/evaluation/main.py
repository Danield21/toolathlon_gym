"""
Evaluation script for yf-portfolio-risk-excel task.

Checks:
1. Excel file Portfolio_Risk_Analysis.xlsx exists with 3 sheets
2. Price History sheet has correct structure
3. Risk Metrics sheet has correct statistics
4. Risk Assessment sheet has correct risk categories
"""
import argparse
import math
import os
import sys
from collections import defaultdict

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

TICKERS = ["GOOGL", "AMZN", "JPM", "JNJ", "XOM"]


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    prices = defaultdict(dict)
    for ticker in TICKERS:
        cur.execute("SELECT date, close FROM yf.stock_prices WHERE symbol = %s ORDER BY date",
                    (ticker,))
        for date, close in cur.fetchall():
            prices[date][ticker] = float(close)

    valid_dates = sorted([d for d, p in prices.items() if len(p) == len(TICKERS)])

    stats = {}
    for ticker in TICKERS:
        vals = [prices[d][ticker] for d in valid_dates]
        avg_val = sum(vals) / len(vals)
        std_val = math.sqrt(sum((v - avg_val)**2 for v in vals) / (len(vals) - 1))
        cv = std_val / avg_val
        if cv < 0.10:
            category = "Low Risk"
        elif cv <= 0.20:
            category = "Medium Risk"
        else:
            category = "High Risk"
        stats[ticker] = {
            "avg": round(avg_val, 2),
            "std": round(std_val, 2),
            "min": round(min(vals), 2),
            "max": round(max(vals), 2),
            "cv": round(cv, 4),
            "category": category,
        }

    cur.close()
    conn.close()
    return stats, len(valid_dates)


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def number_close_match(expected, actual, tolerance=0.05):
    """Check if actual is within tolerance of expected."""
    try:
        return abs(float(actual) - float(expected)) <= abs(float(expected) * tolerance) + 0.01
    except (ValueError, TypeError):
        return False


def check_excel(agent_workspace):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Portfolio_Risk_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(xlsx_path), f"Expected {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        return

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    expected_stats, expected_rows = load_expected()

    def find_sheet(keywords):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(k in sl for k in keywords):
                return wb[s]
        return None

    # Price History
    ws_ph = find_sheet(["price"])
    if not ws_ph:
        ws_ph = find_sheet(["history"])
    check("Price History sheet exists", ws_ph is not None, f"Sheets: {wb.sheetnames}")
    if ws_ph:
        # Check has all tickers in header
        header = [str(c.value).upper() if c.value else "" for c in ws_ph[1]]
        for t in TICKERS:
            check(f"Price History header contains {t}",
                  t in header,
                  f"Header: {header}")
        # Check row count (header + data)
        data_rows = ws_ph.max_row - 1
        check(f"Price History has ~{expected_rows} data rows",
              abs(data_rows - expected_rows) <= 5,
              f"Found {data_rows} rows, expected ~{expected_rows}")

    def _rows_to_dicts(ws):
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        hdr = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        result = []
        for r in rows[1:]:
            if r and any(v is not None for v in r):
                result.append({hdr[i]: r[i] for i in range(len(hdr))})
        return result

    def _find_col(d, candidates):
        keys_lower = list(d.keys())
        for cand in candidates:
            cl = cand.lower()
            for k in keys_lower:
                if cl == k:
                    return k
        for cand in candidates:
            cl = cand.lower()
            for k in keys_lower:
                if cl in k or k in cl:
                    return k
        return None

    STAT_ALIASES = {
        "avg": ["average close price", "average close", "average", "avg price", "avg_price", "avg"],
        "std": ["standard deviation", "std deviation", "std_dev", "stdev", "std"],
        "min": ["minimum close", "minimum", "min price", "min_price", "min"],
        "max": ["maximum close", "maximum", "max price", "max_price", "max"],
    }
    CATEGORY_ALIASES = ["risk category", "risk_category", "category", "risk"]

    # Risk Metrics — row-structured per-ticker comparison
    ws_rm = find_sheet(["risk", "metric"])
    if not ws_rm:
        ws_rm = find_sheet(["metric"])
    check("Risk Metrics sheet exists", ws_rm is not None, f"Sheets: {wb.sheetnames}")
    if ws_rm:
        rm_rows = _rows_to_dicts(ws_rm)
        rm_lookup = {}
        for r in rm_rows:
            tk_key = _find_col(r, ["ticker", "symbol"])
            if tk_key and r[tk_key]:
                rm_lookup[str(r[tk_key]).strip().upper()] = r
        for ticker in TICKERS:
            s = expected_stats[ticker]
            if ticker not in rm_lookup:
                check(f"Risk Metrics: {ticker} row present", False,
                      f"Available: {list(rm_lookup.keys())}")
                continue
            check(f"Risk Metrics: {ticker} row present", True)
            r = rm_lookup[ticker]
            for stat_name, tol in (("avg", 0.02), ("std", 0.05), ("min", 0.02), ("max", 0.02)):
                col = _find_col(r, STAT_ALIASES[stat_name])
                if col is None:
                    check(f"Risk Metrics {ticker}.{stat_name} column present", False,
                          f"Columns: {list(r.keys())}")
                    continue
                try:
                    actual = float(r[col])
                    expected = float(s[stat_name])
                    ok = abs(actual - expected) <= abs(expected) * tol + 0.01
                    check(f"Risk Metrics {ticker}.{stat_name}", ok,
                          f"Expected {expected}, got {actual}")
                except (TypeError, ValueError):
                    check(f"Risk Metrics {ticker}.{stat_name}", False,
                          f"Cannot parse: {r[col]}")

    # Risk Assessment — row-structured per-ticker category comparison
    ws_ra = find_sheet(["risk", "assess"])
    if not ws_ra:
        ws_ra = find_sheet(["assess"])
    check("Risk Assessment sheet exists", ws_ra is not None, f"Sheets: {wb.sheetnames}")
    if ws_ra:
        ra_rows = _rows_to_dicts(ws_ra)
        ra_lookup = {}
        for r in ra_rows:
            tk_key = _find_col(r, ["ticker", "symbol"])
            if tk_key and r[tk_key]:
                ra_lookup[str(r[tk_key]).strip().upper()] = r
        for ticker in TICKERS:
            s = expected_stats[ticker]
            if ticker not in ra_lookup:
                check(f"Risk Assessment: {ticker} row present", False,
                      f"Available: {list(ra_lookup.keys())}")
                continue
            check(f"Risk Assessment: {ticker} row present", True)
            r = ra_lookup[ticker]
            cat_col = _find_col(r, CATEGORY_ALIASES)
            if cat_col is None:
                check(f"Risk Assessment {ticker}.category col present", False,
                      f"Columns: {list(r.keys())}")
                continue
            actual_cat = str(r[cat_col]).strip().lower() if r[cat_col] else ""
            check(f"Risk Assessment {ticker}.category == '{s['category']}'",
                  s["category"].lower() == actual_cat,
                  f"Got '{actual_cat}'")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    all_passed = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
