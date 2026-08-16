"""
Evaluation for yt-top-videos-excel-gcal task.

Checks:
1. Top_Videos_Watchlist.xlsx exists with Watchlist and Summary sheets
2. Watchlist has 10 rows, correct columns, correct top video
3. Summary has correct combined view count
4. GCal has 10 learning sessions on Tuesdays April-June 2026
5. Email sent to team@company.com with correct subject
"""
import json
import os
import sys
from argparse import ArgumentParser
from datetime import date

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


EXPECTED_VIDEOS = [
    (1, "Nl7aCUsWykg", "deepseek", 3.6, 3878491, 152669, 3.94),
    (2, "-2k1rcRzsLA", "free chinese ai", 4.7, 3115193, 90213, 2.9),
    (3, "LKCVKw9CzFo", "100+ linux things", 12.4, 2891861, 107006, 3.7),
    (4, "Iq_r7IcNmUk", "25 crazy software bugs", 16.8, 2496715, 72881, 2.92),
    (5, "4yDm6xNeYas", "some bad code", 4, 2467547, 121950, 4.94),
    (6, "hpwoGjpYygI", "deepseek stole", 4.8, 2436569, 92019, 3.78),
    (7, "jwnez8HdN7E", "microsoft", 4.3, 2329914, 71310, 3.06),
    (8, "N3ZGNT5S5IU", "hackers", 4.3, 1927279, 107099, 5.56),
    (9, "x2WtHZciC74", "claude 3.7", 5.8, 1926147, 80601, 4.18),
    (10, "Tw18-4U7mts", "vibe coding", 4.8, 1846051, 65631, 3.56),
]
EXPECTED_COMBINED_VIEWS = 25315767
EXPECTED_AVG_DURATION = 6.5
EXPECTED_AVG_LIKE_RATE = 3.85


def check_excel(agent_workspace):
    print("\n=== Check 1: Top_Videos_Watchlist.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Top_Videos_Watchlist.xlsx")
    if not os.path.exists(xlsx_path):
        record("Top_Videos_Watchlist.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Top_Videos_Watchlist.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Watchlist sheet
    if "watchlist" not in sheet_names_lower:
        record("Watchlist sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Watchlist sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("watchlist")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Watchlist has exactly 10 rows", len(data_rows) == 10, f"Found {len(data_rows)}")

        # Check columns by header (presence of all required headers)
        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            for col in ["rank", "video_id", "title", "duration_min", "view_count", "like_count", "like_rate"]:
                ok = any(col in h or col.replace("_", " ") in h for h in headers)
                record(f"Has '{col}' column", ok, f"Headers: {rows[0]}")

        # Per-row exact checks: Rank, Video_ID, View_Count, Like_Count, Like_Rate
        if len(data_rows) >= 10:
            for i, row in enumerate(data_rows[:10]):
                expected = EXPECTED_VIDEOS[i]
                exp_rank, exp_vid, exp_title_kw, exp_dur, exp_views, exp_likes, exp_rate = expected
                # Rank
                try:
                    rank_val = int(row[0])
                    record(f"Row {i+1} Rank == {exp_rank}", rank_val == exp_rank,
                           f"Got {rank_val}")
                except (TypeError, ValueError):
                    record(f"Row {i+1} Rank == {exp_rank}", False, f"Got {row[0]}")
                # Video_ID
                if len(row) > 1:
                    record(f"Row {i+1} Video_ID = {exp_vid}",
                           str(row[1] or "").strip() == exp_vid,
                           f"Got {row[1]}")
                # View_Count
                if len(row) > 4:
                    try:
                        v = int(row[4])
                        record(f"Row {i+1} View_Count = {exp_views}",
                               v == exp_views, f"Got {v}")
                    except (TypeError, ValueError):
                        record(f"Row {i+1} View_Count is integer", False, f"Got {row[4]}")
                # Like_Count
                if len(row) > 5:
                    try:
                        v = int(row[5])
                        record(f"Row {i+1} Like_Count = {exp_likes}",
                               v == exp_likes, f"Got {v}")
                    except (TypeError, ValueError):
                        record(f"Row {i+1} Like_Count is integer", False, f"Got {row[5]}")
                # Like_Rate (tol=0.05)
                if len(row) > 6:
                    try:
                        v = float(row[6])
                        record(f"Row {i+1} Like_Rate ~= {exp_rate}",
                               abs(v - exp_rate) < 0.05, f"Got {v}")
                    except (TypeError, ValueError):
                        record(f"Row {i+1} Like_Rate is numeric", False, f"Got {row[6]}")

            # Verify Rank 1 is DeepSeek (most-viewed)
            first_row = data_rows[0]
            first_str = " ".join(str(c) for c in first_row if c).lower()
            record("Rank 1 row mentions DeepSeek/Nl7aCUsWykg",
                   "deepseek" in first_str or "nl7acuswykg" in first_str,
                   f"Row: {first_row[:5]}")

            # Verify sort by View_Count descending
            try:
                view_counts = [int(r[4]) for r in data_rows[:10] if r[4] is not None]
                record("Rows sorted by View_Count descending",
                       view_counts == sorted(view_counts, reverse=True),
                       f"Got: {view_counts}")
            except (TypeError, ValueError):
                pass

    # Summary sheet
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        # Build label-to-value lookup
        sm = {}
        for r in rows2:
            if r and r[0]:
                sm[str(r[0]).strip().lower()] = r[1]

        ts = sm.get("total_selected")
        try:
            record("Summary Total_Selected == 10",
                   int(ts) == 10, f"Got {ts}")
        except (TypeError, ValueError):
            record("Summary Total_Selected == 10", False, f"Got {ts}")

        cv = sm.get("combined_views")
        try:
            record(f"Summary Combined_Views == {EXPECTED_COMBINED_VIEWS}",
                   int(cv) == EXPECTED_COMBINED_VIEWS, f"Got {cv}")
        except (TypeError, ValueError):
            record(f"Summary Combined_Views == {EXPECTED_COMBINED_VIEWS}", False, f"Got {cv}")

        ad = sm.get("avg_duration_min")
        try:
            record(f"Summary Avg_Duration_Min ~= {EXPECTED_AVG_DURATION}",
                   abs(float(ad) - EXPECTED_AVG_DURATION) < 0.05, f"Got {ad}")
        except (TypeError, ValueError):
            record(f"Summary Avg_Duration_Min ~= {EXPECTED_AVG_DURATION}", False, f"Got {ad}")

        ar = sm.get("avg_like_rate")
        try:
            record(f"Summary Avg_Like_Rate ~= {EXPECTED_AVG_LIKE_RATE}",
                   abs(float(ar) - EXPECTED_AVG_LIKE_RATE) < 0.05, f"Got {ar}")
        except (TypeError, ValueError):
            record(f"Summary Avg_Like_Rate ~= {EXPECTED_AVG_LIKE_RATE}", False, f"Got {ar}")


def check_gcal():
    print("\n=== Check 2: GCal learning sessions ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE summary ILIKE '%team learning%'
        AND start_datetime >= TIMESTAMPTZ '2026-04-07T00:00:00Z' AND start_datetime < TIMESTAMPTZ '2026-07-01T00:00:00Z'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    # Tighten: exactly 10 events
    record("Exactly 10 'Team Learning' events scheduled", len(events) == 10,
           f"Found {len(events)} events")

    if not events:
        return

    # Each event title must be 'Team Learning: <title>' (start with this prefix)
    valid_titles = sum(1 for e in events
                       if e[0] and e[0].lower().startswith("team learning:"))
    record("All event titles start with 'Team Learning:'",
           valid_titles == len(events),
           f"Got {valid_titles}/{len(events)} valid titles. Examples: {[e[0] for e in events[:3]]}")

    # Verify dates are 10 consecutive Tuesdays starting April 7, 2026
    expected_dates = [
        "2026-04-07", "2026-04-14", "2026-04-21", "2026-04-28",
        "2026-05-05", "2026-05-12", "2026-05-19", "2026-05-26",
        "2026-06-02", "2026-06-09",
    ]
    actual_dates = sorted({str(e[1])[:10] for e in events if e[1]})
    seen_expected = sum(1 for d in expected_dates if d in actual_dates)
    record("All 10 sessions on consecutive Tuesdays (Apr 7 - Jun 9)",
           seen_expected == 10,
           f"Matched {seen_expected}/10 expected dates. Got: {actual_dates}")

    # 1-hour duration
    valid_durations = 0
    for s, sd, ed in events:
        if sd and ed:
            try:
                if abs((ed - sd).total_seconds() - 3600) <= 60:
                    valid_durations += 1
            except Exception:
                pass
    record("All 10 sessions are 1 hour long", valid_durations == len(events),
           f"Got {valid_durations}/{len(events)} valid durations")

    # 10:00 AM start — normalize to UTC before extracting hour (session TZ
    # may shift str() representation, causing false negatives).
    from datetime import timezone as _tz
    valid_starts = 0
    for _, sd, _ in events:
        if hasattr(sd, 'astimezone'):
            hr = sd.astimezone(_tz.utc).hour
        else:
            s = str(sd)
            hr = int(s[11:13]) if len(s) >= 13 else -1
        if hr in (10, 14, 15):
            valid_starts += 1
    record("All 10 sessions start at 10:00 AM (or UTC equivalent)",
           valid_starts == len(events),
           f"Got {valid_starts}/{len(events)} valid starts")


def check_email():
    print("\n=== Check 3: Email sent ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT to_addr, subject, body_text FROM email.messages
        WHERE to_addr::text ILIKE %s
        ORDER BY date DESC LIMIT 10
    """, ("%team@company.com%",))
    emails = cur.fetchall()
    cur.close()
    conn.close()

    record("Email sent to team@company.com", len(emails) >= 1,
           f"Found {len(emails)} emails")

    if emails:
        ok_subj = False
        ok_dates = False
        ok_titles = False
        for to_addr, subject, body in emails:
            sl = (subject or "").lower()
            bl = (body or "").lower()
            if "learning sessions scheduled" in sl or ("learning sessions" in sl and "scheduled" in sl):
                ok_subj = True
            # Body should list 10 dates
            date_count = sum(1 for d in [
                "2026-04-07", "2026-04-14", "2026-04-21", "2026-04-28",
                "2026-05-05", "2026-05-12", "2026-05-19", "2026-05-26",
                "2026-06-02", "2026-06-09",
            ] if d in (body or "") or d.replace("2026-", "").lstrip("0").replace("-0", "-") in bl)
            # Also accept 'April 7', 'April 14', etc.
            month_dates = ["april 7", "april 14", "april 21", "april 28",
                           "may 5", "may 12", "may 19", "may 26",
                           "june 2", "june 9"]
            md_count = sum(1 for d in month_dates if d in bl)
            # Tighten: require >=8 of 10 dates listed (task says 'list 10 dates')
            if date_count >= 8 or md_count >= 8:
                ok_dates = True
            # Body should mention some video titles (DeepSeek, Linux, etc.)
            keywords = ["deepseek", "linux", "claude", "vibe coding", "openai",
                        "windows", "microsoft"]
            kw_count = sum(1 for kw in keywords if kw in bl)
            if kw_count >= 3:
                ok_titles = True
        record("Email subject 'Learning Sessions Scheduled'",
               ok_subj, f"Subjects: {[e[1] for e in emails]}")
        record("Email body lists at least 8 of 10 session dates",
               ok_dates,
               f"Bodies (first 200): {[(e[2] or '')[:200] for e in emails][:1]}")
        record("Email body mentions at least 3 video titles/topics",
               ok_titles)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Use FAIL_COUNT == 0 (replaces 70% accuracy threshold)
    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
