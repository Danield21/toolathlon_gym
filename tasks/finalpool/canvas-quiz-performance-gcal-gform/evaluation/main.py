"""
Evaluation script for canvas-quiz-performance-gcal-gform task.

Checks:
1. Excel file Quiz_Performance_Report.xlsx - 2 sheets with correct structure and data
2. Google Form "Quiz Improvement Feedback" exists with at least 3 questions
   (1 multiple-choice "most challenging topic" question with the 4 prescribed options,
   1 multiple-choice 1-5 confidence question, 1 optional text question)
3. Google Calendar has 3 specific tutoring session events (date/time/title verified)
4. Email sent to ccc.instructor@university.edu with exact subject

Robustness notes:
- All DB access reads PGHOST/PGPORT/PGDATABASE/PGUSER/PGPASSWORD with sensible defaults.
- GForm question types match the google-forms MCP contract exactly:
  add_text_question -> 'textQuestion'; add_multiple_choice_question -> 'choiceQuestion'
  (config.type == 'RADIO'). Both spellings are accepted.
- Excel cells are read with data_only=False (formula detection) plus a data_only=True
  cache pass, so formula cells without a cached value are skipped rather than failing.
- Calendar times are normalized to UTC before comparing hour/minute.
"""

import argparse
import json
import os
import re
import sys
import unicodedata

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(v):
    """Robustly convert a value to float. Returns None when not parseable."""
    if v is None:
        return None
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Never try to parse formula strings as numbers.
    if s.startswith("="):
        return None
    cleaned = (
        s.replace(",", "")
        .replace("$", "")
        .replace("¥", "")
        .replace("€", "")
        .replace("%", "")
        .replace(" ", "")
        .strip()
    )
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def num_close(a, b, tol=1.0):
    """Numeric comparison with tolerance; falls back to string compare only when
    either side cannot be parsed as a number."""
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if a is None and b is None:
        return True
    return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def _effective_value(wb_val, wb_cache, sheet_name, row, col):
    """Return the effective cell value.

    Prefers the data_only cache; if the raw value is a formula and the cache is
    None (never recalculated), returns None so callers can skip the check.
    """
    raw = wb_val[sheet_name].cell(row=row, column=col).value
    cached = wb_cache[sheet_name].cell(row=row, column=col).value
    if isinstance(raw, str) and raw.strip().startswith("="):
        return cached
    return cached if cached is not None else raw


def load_sheet_by_name(wb_val, wb_cache, name):
    for sname in wb_val.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            ws_val = wb_val[sname]
            rows = []
            for r, row in enumerate(ws_val.iter_rows(), start=1):
                rows.append(
                    [
                        _effective_value(wb_val, wb_cache, sname, r, c)
                        for c, _cell in enumerate(row, start=1)
                    ]
                )
            return rows
    return None


def _canon_header(h):
    """Normalize a header cell to a canonical column key.

    Tolerates case, spaces, underscores, % signs and other punctuation so that
    minor phrasing variants (e.g. "Pass Rate %", "Avg Score", "Total Submissions")
    still resolve to the canonical column names used by the task.
    """
    s = str(h or "").strip().lower()
    s = re.sub(r"[^a-z0-9]", "", s)
    return HEADER_ALIASES.get(s, s)


HEADER_ALIASES = {
    "quiztitle": "quiz_title",
    "totalsubmissions": "total_submissions",
    "total_submissions": "total_submissions",
    "avgscore": "avg_score",
    "avg_score": "avg_score",
    "average": "avg_score",
    "minscore": "min_score",
    "min_score": "min_score",
    "maxscore": "max_score",
    "max_score": "max_score",
    "passratepct": "pass_rate_pct",
    "pass_rate_pct": "pass_rate_pct",
    "passrate": "pass_rate_pct",
    "passratepercent": "pass_rate_pct",
    "metric": "metric",
    "value": "value",
}


def _build_column_map(header_row):
    mapping = {}
    for idx, h in enumerate(header_row):
        key = _canon_header(h)
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


# ============================================================================
# Check 1: Excel file
# ============================================================================

QUIZ_STATS_COLS = [
    ("quiz_title", 0),
    ("total_submissions", 1),
    ("avg_score", 2),
    ("min_score", 3),
    ("max_score", 4),
    ("pass_rate_pct", 5),
]
QUIZ_TOLERANCES = {
    "total_submissions": 1,
    "avg_score": 0.5,
    "min_score": 1,
    "max_score": 1,
    "pass_rate_pct": 0.5,
}


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Quiz_Performance_Report.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        return

    agent_file = os.path.join(agent_workspace, "Quiz_Performance_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Quiz_Performance_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return
    record("Excel file exists", True)

    try:
        agent_wb_val = openpyxl.load_workbook(agent_file, data_only=False)
        agent_wb_cache = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb_val = openpyxl.load_workbook(gt_file, data_only=False)
        gt_wb_cache = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel workbooks readable", False, f"{e}")
        return
    record("Excel workbooks readable", True)

    # ---------- Sheet 1: Quiz Stats ----------
    a_quiz = load_sheet_by_name(agent_wb_val, agent_wb_cache, "Quiz Stats")
    g_quiz = load_sheet_by_name(gt_wb_val, gt_wb_cache, "Quiz Stats")
    record("Sheet 'Quiz Stats' exists", a_quiz is not None)

    if a_quiz is not None and g_quiz is not None:
        a_header = a_quiz[0] if a_quiz else []
        g_header = g_quiz[0] if g_quiz else []
        a_cols = _build_column_map(a_header)
        g_cols = _build_column_map(g_header)
        g_data = [r for r in g_quiz[1:] if any(v is not None for v in r)]
        a_data = [r for r in a_quiz[1:] if any(v is not None for v in r)]

        # Header validation (informational but recorded): expected columns present.
        expected_headers = {name for name, _pos in QUIZ_STATS_COLS}
        present_headers = set(a_cols.keys()) & expected_headers
        record(
            "Quiz Stats headers contain expected columns",
            expected_headers.issubset(present_headers),
            f"Found columns: {sorted(a_cols.keys())}",
        )

        # Build lookup by quiz title. The title column is located through the
        # header map, so an agent that reorders columns or adds a leading
        # column (e.g. a 'Rank' / '#' column) is still matched correctly.
        def _col_val(row, name, fallback_pos, colmap):
            idx = colmap.get(name, fallback_pos)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        a_key_col = a_cols.get("quiz_title", 0)
        g_key_col = g_cols.get("quiz_title", 0)
        a_lookup = {}
        for row in a_data:
            if row and len(row) > a_key_col and row[a_key_col] is not None:
                a_lookup[str(row[a_key_col]).strip().lower()] = row

        # Verify ascending sort by Avg_Score (agent rows must be in ascending order).
        # Requiring at least two parseable values prevents the empty-set
        # `all([]) == True` loophole from silently passing the sort check.
        try:
            avg_scores = [
                _to_float(_col_val(r, "avg_score", 2, a_cols))
                for r in a_data
                if _col_val(r, "avg_score", 2, a_cols) is not None
            ]
            ascending_ok = (
                len(avg_scores) >= 2
                and all(
                    avg_scores[i] <= avg_scores[i + 1]
                    for i in range(len(avg_scores) - 1)
                )
            )
        except Exception:
            ascending_ok = False
        record("Quiz Stats sorted ascending by Avg_Score", ascending_ok,
               f"Avg_Scores: {[_col_val(r,'avg_score',2,a_cols) for r in a_data]}")

        for g_row in g_data:
            if not g_row or len(g_row) <= g_key_col or g_row[g_key_col] is None:
                continue
            title = str(g_row[g_key_col]).strip()
            key = title.lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Quiz row exists: {title}", False, "Row not found")
                continue
            record(f"Quiz row exists: {title}", True)

            for colname, fallback_pos in QUIZ_STATS_COLS:
                if colname == "quiz_title":
                    continue
                a_val = _col_val(a_row, colname, fallback_pos, a_cols)
                g_val = _col_val(g_row, colname, fallback_pos, g_cols)
                # An empty cell, or a formula with no cached value, cannot be
                # verified and must FAIL rather than silently pass: the task
                # instructs agents to write literal numbers.
                if a_val is None and g_val is not None:
                    record(f"{title}: {colname}", False,
                           "agent cell empty or formula with no cached value")
                    continue
                record(f"{title}: {colname}",
                       num_close(a_val, g_val, QUIZ_TOLERANCES.get(colname, 1)),
                       f"got {a_val}, expected {g_val}")

    # ---------- Sheet 2: Course Summary ----------
    a_summ = load_sheet_by_name(agent_wb_val, agent_wb_cache, "Course Summary")
    g_summ = load_sheet_by_name(gt_wb_val, gt_wb_cache, "Course Summary")
    record("Sheet 'Course Summary' exists", a_summ is not None)

    if a_summ is not None and g_summ is not None:
        a_header = a_summ[0] if a_summ else []
        g_header = g_summ[0] if g_summ else []
        a_cols = _build_column_map(a_header)
        g_cols = _build_column_map(g_header)
        val_pos = a_cols.get("value", 1)
        g_val_pos = g_cols.get("value", 1)
        g_key_col = g_cols.get("metric", 0)
        a_key_col = a_cols.get("metric", 0)
        g_data = [r for r in g_summ[1:] if any(v is not None for v in r)]
        a_data = [r for r in a_summ[1:] if any(v is not None for v in r)]

        expected_headers = {"metric", "value"}
        present_headers = set(a_cols.keys()) & expected_headers
        record("Course Summary headers contain expected columns",
               expected_headers.issubset(present_headers),
               f"Found columns: {sorted(a_cols.keys())}")

        a_lookup = {}
        for row in a_data:
            if row and len(row) > a_key_col and row[a_key_col] is not None:
                a_lookup[str(row[a_key_col]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or len(g_row) <= g_key_col or g_row[g_key_col] is None:
                continue
            key = str(g_row[g_key_col]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary row: {g_row[g_key_col]}", False, "Row not found")
                continue
            record(f"Summary row: {g_row[g_key_col]}", True)
            a_val = a_row[val_pos] if val_pos < len(a_row) else None
            g_val = g_row[g_val_pos] if g_val_pos < len(g_row) else None

            if key == "total_quizzes":
                record("Total_Quizzes value", num_close(a_val, g_val, 0),
                       f"got {a_val}, expected {g_val}")
            elif key == "total_quiz_submissions":
                record("Total_Quiz_Submissions value", num_close(a_val, g_val, 1),
                       f"got {a_val}, expected {g_val}")
            elif key == "overall_avg_score":
                record("Overall_Avg_Score value", num_close(a_val, g_val, 0.5),
                       f"got {a_val}, expected {g_val}")
            elif key == "lowest_avg_quiz":
                record("Lowest_Avg_Quiz value", str_match(a_val, g_val),
                       f"got {a_val}, expected {g_val}")
            elif key == "highest_avg_quiz":
                record("Highest_Avg_Quiz value", str_match(a_val, g_val),
                       f"got {a_val}, expected {g_val}")


# ============================================================================
# Check 2: Google Form
# ============================================================================

MC_TYPES = {"choicequestion", "choice", "multiple_choice", "radio", "multi_choice", "mc"}
TEXT_TYPES = {"textquestion", "text", "paragraph", "short_answer", "long_answer"}


def _parse_config(config):
    if isinstance(config, str):
        try:
            return json.loads(config)
        except Exception:
            return {}
    if isinstance(config, dict):
        return config
    return {}


def _extract_option_texts(config):
    cfg = _parse_config(config)
    opts = cfg.get("options") or []
    out = []
    for o in opts:
        if isinstance(o, dict):
            v = o.get("value")
        else:
            v = o
        if v is None:
            continue
        s = str(v).strip()
        # A single string may embed a comma-separated option list
        # (e.g. "1 - Not confident, 2, 3, 4, 5 - Very confident").
        if "," in s:
            for part in s.split(","):
                p = part.strip()
                if p:
                    out.append(p)
        else:
            out.append(s)
    return out


def _first_number(t):
    """Extract the first numeric value embedded in a string.

    Handles option texts such as '1 - Not confident' or '3 (somewhat)',
    which cannot be parsed as a bare number. Returns None when no digit is
    present.
    """
    m = re.search(r"[-+]?\d+(?:\.\d+)?", str(t))
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _numeric_options(option_texts):
    result = set()
    for t in option_texts:
        n = _first_number(t)
        if n is not None:
            result.add(int(n))
    return result


def _norm(s):
    """Normalize a text for fuzzy comparison.

    Lowercases, expands the 'UI' abbreviation ('UI Design' -> 'User Interface
    Design'), and strips all non-alphanumeric characters.
    """
    s = unicodedata.normalize("NFKC", str(s or "")).lower()
    s = re.sub(r"\bui\b", "user interface", s)
    return re.sub(r"[^a-z0-9]+", "", s)


def _match_topic(option, topic):
    """True when an option string expresses the given expected topic.

    Accepts exact matches, common abbreviations ('UI Design' -> 'User
    Interface Design') and one-sided containment, provided the containing side
    is a substantial token (>= 4 chars) so single-letter noise cannot match.
    """
    o = _norm(option)
    t = _norm(topic)
    if not o or not t:
        return False
    if o == t:
        return True
    if len(o) >= 4 and t in o:
        return True
    if len(t) >= 4 and o in t:
        return True
    return False


def _covers_topics(opts, expected):
    """True when the option strings cover every expected topic, each topic
    matched by a distinct option."""
    unmatched = list(expected)
    for o in opts:
        for t in list(unmatched):
            if _match_topic(o, t):
                unmatched.remove(t)
                break
    return len(unmatched) == 0


def _is_mc(q_type, config):
    qt = (q_type or "").lower()
    if qt in MC_TYPES:
        return True
    cfg = _parse_config(config)
    if cfg.get("type") in ("RADIO", "radio", "choiceQuestion", "choicequestion"):
        return True
    return False


def _is_text(q_type):
    return ((q_type or "").lower() in TEXT_TYPES)


def check_gform():
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("GForm: DB connect", False, f"cannot connect to DB: {e}")
        return
    try:
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        print(f"[check_gform] Found {len(forms)} forms.")
        record("At least 1 form created", len(forms) >= 1, f"Found {len(forms)}")

        expected_title = "quiz improvement feedback"
        target_form_id = None
        for form_id, title in forms:
            if title and title.strip().lower() == expected_title:
                target_form_id = form_id
                break
        record("Form titled exactly 'Quiz Improvement Feedback' found",
               target_form_id is not None,
               f"Forms titles: {[f[1] for f in forms]}")

        if target_form_id is None:
            return

        cur.execute(
            "SELECT title, question_type, required, config FROM gform.questions "
            "WHERE form_id=%s ORDER BY position",
            (target_form_id,),
        )
        questions = cur.fetchall()
        record("Form has at least 3 questions", len(questions) >= 3,
               f"Found {len(questions)} questions")

        expected_opts = {
            "programming fundamentals",
            "data structures",
            "algorithms",
            "user interface design",
        }

        has_mc_topic = False
        has_scale_confidence = False
        has_text_resources = False

        mc_questions = []
        for q_title, q_type, q_required, q_config in questions:
            ql = (q_title or "").lower()
            opts = _extract_option_texts(q_config)
            numeric = _numeric_options(opts)
            if _is_mc(q_type, q_config):
                mc_questions.append((ql, opts, numeric))
                if ("challeng" in ql or "topic" in ql) and _covers_topics(opts, expected_opts):
                    has_mc_topic = True
                if ("confiden" in ql or "prepar" in ql) and {1, 2, 3, 4, 5}.issubset(numeric):
                    has_scale_confidence = True
            if _is_text(q_type) and ("resource" in ql or "additional" in ql or "help" in ql):
                has_text_resources = True

        # Fallbacks: any multiple-choice question carrying the four topics (or
        # the numeric options 1..5) counts, even if its title is phrased
        # differently. Topic/option wording is matched fuzzily so reasonable
        # rephrasing (e.g. 'UI Design' for 'User Interface Design') still
        # passes, while a form that omits topics/options still fails.
        if not has_mc_topic:
            for ql, opts, numeric in mc_questions:
                if _covers_topics(opts, expected_opts):
                    has_mc_topic = True
                    break

        if not has_scale_confidence:
            for ql, opts, numeric in mc_questions:
                if {1, 2, 3, 4, 5}.issubset(numeric):
                    has_scale_confidence = True
                    break

        record("Multiple-choice 'most challenging topic' question present", has_mc_topic,
               f"Questions: {[(q[0], q[1]) for q in questions]}")
        record("Scale (1-5 multiple-choice) confidence question present", has_scale_confidence,
               f"Questions: {[(q[0], q[1]) for q in questions]}")
        record("Text question on additional resources present", has_text_resources,
               f"Questions: {[(q[0], q[1]) for q in questions]}")
    finally:
        try:
            conn.close()
        except Exception:
            pass


# ============================================================================
# Check 3: Google Calendar
# ============================================================================

def _as_utc(v):
    from datetime import datetime, timezone

    if isinstance(v, datetime):
        dt = v
    else:
        s = str(v).replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _as_local(v):
    """Return the datetime in its own wall-clock view (timezone preserved).

    A naive datetime is returned as-is, which represents the DB session's
    interpretation of the stored value. This lets the 3pm-5pm check accept
    events written in the session/agent local timezone as well as in UTC.
    """
    from datetime import datetime

    if isinstance(v, datetime):
        return v
    s = str(v).replace("Z", "+00:00")
    return datetime.fromisoformat(s)


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("GCal: DB connect", False, f"cannot connect to DB: {e}")
        return
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT summary, start_datetime, end_datetime FROM gcal.events ORDER BY start_datetime"
        )
        events = cur.fetchall()
    except Exception as e:
        record("GCal: query events", False, f"query failed: {e}")
        return
    finally:
        try:
            conn.close()
        except Exception:
            pass

    print(f"[check_gcal] Found {len(events)} calendar events.")
    record("At least 3 calendar events created", len(events) >= 3, f"Found {len(events)}")

    expected_sessions = [
        ("CCC Spring 2014 Tutoring Session 1", "2026-03-12"),
        ("CCC Spring 2014 Tutoring Session 2", "2026-03-19"),
        ("CCC Spring 2014 Tutoring Session 3", "2026-03-26"),
    ]
    for exp_title, exp_date in expected_sessions:
        match = None
        for summary, start_dt, end_dt in events:
            if summary and summary.strip().lower() == exp_title.lower():
                if start_dt is not None:
                    s_utc = _as_utc(start_dt)
                    if s_utc.strftime("%Y-%m-%d") == exp_date:
                        match = (summary, start_dt, end_dt)
                        break
        record(f"Event '{exp_title}' on {exp_date} found", match is not None,
               f"Events: {[(e[0], str(e[1])[:10]) for e in events]}")
        if match:
            try:
                s_utc = _as_utc(match[1])
                e_utc = _as_utc(match[2])
                s_local = _as_local(match[1])
                e_local = _as_local(match[2])

                def _is_3_to_5(st, et):
                    return (
                        st.hour == 15 and st.minute == 0
                        and et.hour == 17 and et.minute == 0
                    )

                # The task says '3pm to 5pm' without a timezone. Accept the
                # event if it is 15:00-17:00 either in UTC or in the event's
                # own local wall-clock view (which covers agents that wrote the
                # local session time, e.g. 15:00+08:00, or that converted to
                # UTC, e.g. 23:00+08:00 == 15:00Z). The date is already matched
                # in UTC above.
                time_ok = _is_3_to_5(s_utc, e_utc) or _is_3_to_5(s_local, e_local)
            except Exception:
                time_ok = False
            record(f"Event '{exp_title}' time 15:00-17:00",
                   time_ok, f"start={match[1]}, end={match[2]}")


# ============================================================================
# Check 4: Email
# ============================================================================

def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Email: DB connect", False, f"cannot connect to DB: {e}")
        return
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, from_addr, to_addr, body_text FROM email.messages"
        )
        all_emails = cur.fetchall()
    except Exception as e:
        record("Email: query messages", False, f"query failed: {e}")
        return
    finally:
        try:
            conn.close()
        except Exception:
            pass

    print(f"[check_emails] Found {len(all_emails)} total emails.")
    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    target_email = None
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = json.dumps(to_addr).lower() if isinstance(to_addr, (list, dict)) else str(to_addr or "").lower()
        if "ccc.instructor@university.edu" in to_str:
            target_email = (subject, from_addr, to_addr, body_text)
            break
    record("Email to ccc.instructor@university.edu found", target_email is not None,
           f"Emails: {[(e[0], e[2]) for e in all_emails[:5]]}")

    if target_email is None:
        return

    subject, _, _, body_text = target_email
    expected_subject = "Creative Computing Quiz Performance Report"
    record("Email subject exactly 'Creative Computing Quiz Performance Report'",
           (subject or "").strip().lower() == expected_subject.lower(),
           f"Subject: {subject}")

    body_lower = (body_text or "").lower()
    record("Body summarizes quiz performance findings",
           ("quiz" in body_lower) and ("performance" in body_lower or "score" in body_lower or "average" in body_lower or "avg" in body_lower),
           "Missing performance/score/average mention")
    record("Body mentions tutoring sessions",
           "tutoring" in body_lower or "session" in body_lower,
           "Missing tutoring/session mention")


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_gcal()
    check_emails()

    all_passed = (FAIL_COUNT == 0)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
