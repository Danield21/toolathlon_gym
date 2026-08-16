#!/usr/bin/env python3
"""Evaluation script for financial-audit-cross-system-reconciliation."""

import argparse
import os
import sys
from pathlib import Path

PASS_COUNT = 0
FAIL_COUNT = 0
WARN_COUNT = 0
IS_GT_SELF_TEST = False


def check(name, condition, detail="", db_side=False):
    global PASS_COUNT, FAIL_COUNT, WARN_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        if IS_GT_SELF_TEST and db_side:
            WARN_COUNT += 1
            detail_str = str(detail)[:200] if detail else ""
            print(f"  [WARN] {name} (GT self-test mode, DB-side): {detail_str}")
        else:
            FAIL_COUNT += 1
            detail_str = str(detail)[:200] if detail else ""
            print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")),
        dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
        user="eigent", password="camel",
    )


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, WARN_COUNT, IS_GT_SELF_TEST
    PASS_COUNT = 0
    FAIL_COUNT = 0
    WARN_COUNT = 0

    if not agent_workspace:
        return False, "No agent workspace provided"

    agent_ws = Path(agent_workspace)
    if not agent_ws.exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    # Detect GT self-test mode
    try:
        if groundtruth_workspace and Path(groundtruth_workspace).exists():
            IS_GT_SELF_TEST = (
                os.path.realpath(str(agent_ws)) ==
                os.path.realpath(str(Path(groundtruth_workspace)))
            )
    except Exception:
        IS_GT_SELF_TEST = False

    # Phase 4: Excel audit report (any *.xlsx with reconciliation/audit content)
    xlsx_files = list(agent_ws.glob("*.xlsx"))
    check("At least one xlsx audit report exists",
          len(xlsx_files) >= 1, f"found {len(xlsx_files)} xlsx files")
    if xlsx_files:
        try:
            import openpyxl
            # Pick the largest xlsx (skip leftover q1_manual_tracking.xlsx fixture)
            audit_files = [p for p in xlsx_files
                           if 'q1_manual' not in p.name.lower()
                           and 'tracking' not in p.name.lower()]
            if not audit_files:
                audit_files = xlsx_files
            audit_path = max(audit_files, key=lambda p: p.stat().st_size)
            wb = openpyxl.load_workbook(str(audit_path), data_only=True)
            sheets = [s.lower() for s in wb.sheetnames]
            # Reconciliation-related sheet keywords
            has_recon_sheet = any(
                kw in s
                for s in sheets
                for kw in ['recon', 'audit', 'variance', 'exception', 'summary', 'transaction']
            )
            check(f"Audit xlsx ({audit_path.name}) has a reconciliation-related sheet",
                  has_recon_sheet, f"sheets={wb.sheetnames}")
            # Sheets must have non-trivial content (>=3 rows in at least one sheet)
            max_rows = max((ws.max_row for ws in wb.worksheets), default=0)
            check("Audit xlsx has >=3 rows of data in some sheet",
                  max_rows >= 3, f"max rows={max_rows}")
            # At least 2 distinct sheets (reconciliation + exceptions)
            check("Audit xlsx has >= 2 sheets (multi-section audit)",
                  len(wb.sheetnames) >= 2,
                  f"sheets={wb.sheetnames}")
            # Has a sheet specifically for exceptions / variances
            has_excep_sheet = any(
                kw in s
                for s in sheets
                for kw in ['except', 'variance', 'discrep', 'mismatch']
            )
            check("Audit xlsx has exceptions/variance sheet",
                  has_excep_sheet, f"sheets={wb.sheetnames}")
            # Numeric content: at least one numeric reconciliation/variance figure
            # (count, amount, rate, percentage) must appear somewhere in the xlsx.
            numeric_found = False
            keyword_with_number = False
            for ws in wb.worksheets:
                last_label = ""
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    for cell in row:
                        if isinstance(cell, (int, float)) and cell != 0:
                            numeric_found = True
                        if cell is not None:
                            text = str(cell).lower()
                            if any(kw in text for kw in
                                   ['reconcil', 'discrep', 'variance',
                                    'exception', 'mismatch', 'rate',
                                    'count', 'total', 'amount']):
                                last_label = text
                        if last_label and isinstance(cell, (int, float)):
                            keyword_with_number = True
                            break
                    if keyword_with_number:
                        break
                if keyword_with_number:
                    break
            check("Audit xlsx contains numeric values (counts/amounts)",
                  numeric_found, "no numeric data found")
            check("Audit xlsx pairs reconciliation labels with numeric values",
                  keyword_with_number,
                  "expected label like 'reconciled/discrepancy/exception/total' alongside a number")
        except Exception as e:
            check("Audit xlsx parse", False, str(e))

    # Phase 5: Word executive summary doc
    docx_files = list(agent_ws.glob("*.docx"))
    check("At least one docx executive summary exists",
          len(docx_files) >= 1, f"found {len(docx_files)} docx files")
    if docx_files:
        try:
            from docx import Document
            audit_docx = max(docx_files, key=lambda p: p.stat().st_size)
            doc = Document(str(audit_docx))
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word doc has substantive content (>=200 chars)",
                  len(text) >= 200, f"len={len(text)}")
            has_summary = any(kw in text for kw in
                              ['executive summary', 'summary', 'overview', 'finding'])
            has_recon = any(kw in text for kw in
                            ['reconcil', 'variance', 'discrep', 'audit', 'exception'])
            check("Word doc mentions executive summary/finding",
                  has_summary, f"text head: {text[:200]}")
            check("Word doc mentions reconciliation/variance",
                  has_recon, f"text head: {text[:200]}")
        except Exception as e:
            check("Word doc parse", False, str(e))

    # Phase 6: Email + Calendar event
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, subject FROM email.messages
               WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s
                  OR body_text ILIKE %s""",
            ('%audit%', '%reconcil%', '%financial%', '%audit%'))
        rows = cur.fetchall()
        check("Audit email sent to senior finance management",
              len(rows) >= 1, f"found {len(rows)} matching emails", db_side=True)
        conn.close()
    except Exception as e:
        check("Email check", False, str(e), db_side=True)

    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT summary FROM gcal.events
               WHERE summary ILIKE %s OR summary ILIKE %s OR summary ILIKE %s""",
            ('%audit%', '%reconcil%', '%review%'))
        rows = cur.fetchall()
        check("Calendar review meeting scheduled",
              len(rows) >= 1, f"found {len(rows)} matching events", db_side=True)
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e), db_side=True)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
