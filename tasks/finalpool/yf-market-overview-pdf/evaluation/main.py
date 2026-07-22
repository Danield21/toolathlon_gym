"""
Evaluation for yf-market-overview-pdf task.

Dynamically queries PostgreSQL to compute expected stock data values,
then checks agent output files for correctness with full column validation.
"""
from argparse import ArgumentParser
import sys
import os
from pathlib import Path


SYMBOLS = ['AMZN', 'GOOGL', 'JNJ', 'JPM', 'XOM']

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def get_expected_data():
    """Query PostgreSQL to get expected stock data."""
    import psycopg2

    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym",
        user="eigent", password="camel"
    )
    cur = conn.cursor()

    syms = tuple(SYMBOLS)
    cur.execute("""
        SELECT sp.symbol, sp.date, sp.open, sp.high, sp.low, sp.close, sp.volume
        FROM yf.stock_prices sp
        WHERE sp.symbol IN %s
          AND sp.date = (SELECT MAX(date) FROM yf.stock_prices WHERE symbol = sp.symbol)
        ORDER BY sp.symbol
    """, (syms,))
    prices = cur.fetchall()

    cur.execute("""
        SELECT symbol,
               data->>'shortName' as name,
               (data->>'marketCap')::float as market_cap,
               (data->>'trailingPE')::float as pe_ratio,
               (data->>'fiftyTwoWeekHigh')::float as high_52w,
               (data->>'fiftyTwoWeekLow')::float as low_52w,
               data->>'sector' as sector
        FROM yf.stock_info
        WHERE symbol IN %s
        ORDER BY symbol
    """, (syms,))
    info = {r[0]: r for r in cur.fetchall()}

    conn.close()
    return prices, info


def check_excel(workspace, prices, info):
    """Check Market_Overview_Report.xlsx for correctness."""
    import openpyxl

    xlsx_path = Path(workspace) / "Market_Overview_Report.xlsx"
    if not xlsx_path.exists():
        record("Excel file exists", False, str(xlsx_path))
        return
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    record("Has 'Stock Data' sheet", "Stock Data" in wb.sheetnames, f"sheets: {wb.sheetnames}")
    record("Has 'Summary' sheet", "Summary" in wb.sheetnames, f"sheets: {wb.sheetnames}")

    # ---------- Stock Data ----------
    if "Stock Data" not in wb.sheetnames:
        return
    ws1 = wb["Stock Data"]
    rows1 = list(ws1.iter_rows(values_only=True))
    if len(rows1) < 2:
        record("Stock Data has data", False, "no data rows")
        return

    header = [str(h).strip() if h else "" for h in rows1[0]]
    REQUIRED_COLS = [
        "Symbol", "Company_Name", "Date", "Open", "High", "Low", "Close",
        "Volume", "Market_Cap", "PE_Ratio", "52W_High", "52W_Low", "Sector",
    ]
    for col in REQUIRED_COLS:
        record(f"Stock Data column '{col}'", col in header, f"got {header}")

    # Normalised lookup
    h_idx = {h: i for i, h in enumerate(header)}
    data_rows = [r for r in rows1[1:] if r and r[0]]
    record("Stock Data has exactly 5 rows", len(data_rows) == 5, f"got {len(data_rows)}")

    # Sort order: alphabetical by Symbol
    sym_idx = h_idx.get("Symbol", 0)
    syms_in_order = [str(r[sym_idx]).strip() for r in data_rows]
    record("Stock Data sorted alphabetically by Symbol",
           syms_in_order == sorted(syms_in_order),
           f"got {syms_in_order}")

    price_map = {r[0]: r for r in prices}
    info_map = info

    for row in data_rows:
        sym = str(row[sym_idx]).strip() if row[sym_idx] else ""
        if sym not in price_map:
            record(f"Symbol '{sym}' is expected", False, f"symbol not in {SYMBOLS}")
            continue

        exp = price_map[sym]
        # Open/High/Low/Close (tol 0.05 - 2 decimals)
        for col_name, exp_idx in [("Open", 2), ("High", 3), ("Low", 4), ("Close", 5)]:
            if col_name in h_idx:
                actual = row[h_idx[col_name]]
                expected = float(exp[exp_idx]) if exp[exp_idx] is not None else None
                record(
                    f"{sym} {col_name}",
                    num_close(actual, expected, 0.05) if expected is not None else actual is None,
                    f"actual={actual} expected={expected}",
                )
        # Volume (exact)
        if "Volume" in h_idx:
            actual_vol = row[h_idx["Volume"]]
            try:
                ok = actual_vol is not None and int(actual_vol) == int(exp[6])
            except Exception:
                ok = False
            record(f"{sym} Volume exact", ok, f"actual={actual_vol} expected={exp[6]}")

        # Date - allow string or datetime
        if "Date" in h_idx:
            actual_date = row[h_idx["Date"]]
            exp_date_str = exp[1].isoformat() if hasattr(exp[1], "isoformat") else str(exp[1])
            ok = str(actual_date).strip().startswith(exp_date_str[:10])
            record(f"{sym} Date", ok, f"actual={actual_date} expected={exp_date_str}")

        # Info columns
        if sym in info_map:
            inf = info_map[sym]
            # Company_Name (string contain)
            if "Company_Name" in h_idx:
                actual = str(row[h_idx["Company_Name"]] or "").lower()
                expected = str(inf[1] or "").lower()
                # Allow partial match (e.g., 'Amazon' OK if GT has 'Amazon.com, Inc.')
                ok = bool(actual) and (expected.split(",")[0].split()[0] in actual or actual in expected)
                record(f"{sym} Company_Name", ok, f"actual={actual} expected={expected}")
            # Market_Cap (tol 0.5%)
            if "Market_Cap" in h_idx:
                actual = row[h_idx["Market_Cap"]]
                expected = inf[2]
                if expected is None:
                    record(f"{sym} Market_Cap", actual is None or actual == "", f"actual={actual}")
                else:
                    record(f"{sym} Market_Cap", num_close(actual, expected, expected * 0.005), f"actual={actual} expected={expected}")
            # PE_Ratio (tol 0.1 — precise from DB)
            if "PE_Ratio" in h_idx:
                actual = row[h_idx["PE_Ratio"]]
                expected = inf[3]
                if expected is None:
                    record(f"{sym} PE_Ratio", actual is None or actual == "", f"actual={actual}")
                else:
                    record(f"{sym} PE_Ratio", num_close(actual, expected, 0.1), f"actual={actual} expected={expected}")
            # 52W_High / 52W_Low (tol 1.0)
            for col_name, idx_pos in [("52W_High", 4), ("52W_Low", 5)]:
                if col_name in h_idx:
                    actual = row[h_idx[col_name]]
                    expected = inf[idx_pos]
                    if expected is None:
                        record(f"{sym} {col_name}", actual is None or actual == "", f"actual={actual}")
                    else:
                        record(f"{sym} {col_name}", num_close(actual, expected, 1.0), f"actual={actual} expected={expected}")
            # Sector (exact substring)
            if "Sector" in h_idx:
                actual = str(row[h_idx["Sector"]] or "").strip().lower()
                expected = str(inf[6] or "").strip().lower()
                record(f"{sym} Sector", actual == expected if expected else True, f"actual={actual} expected={expected}")

    # ---------- Summary ----------
    if "Summary" not in wb.sheetnames:
        return
    ws2 = wb["Summary"]
    rows2 = list(ws2.iter_rows(values_only=True))
    summary_map = {}
    for row in rows2[1:]:
        if row and row[0]:
            summary_map[str(row[0]).strip()] = row[1]

    record("Summary has 'Stocks_Covered'", "Stocks_Covered" in summary_map)
    if "Stocks_Covered" in summary_map:
        record("Stocks_Covered == 5", int(summary_map["Stocks_Covered"]) == 5)

    # Highest_Close_Symbol & Price
    best_close = max(prices, key=lambda x: float(x[5]))
    record("Highest_Close_Symbol", summary_map.get("Highest_Close_Symbol") == best_close[0],
           f"got {summary_map.get('Highest_Close_Symbol')}, expected {best_close[0]}")
    record("Highest_Close_Price", num_close(summary_map.get("Highest_Close_Price"), float(best_close[5]), 0.05),
           f"got {summary_map.get('Highest_Close_Price')}, expected {best_close[5]}")
    # Highest Volume
    best_vol = max(prices, key=lambda x: int(x[6]))
    record("Highest_Volume_Symbol", summary_map.get("Highest_Volume_Symbol") == best_vol[0],
           f"got {summary_map.get('Highest_Volume_Symbol')}, expected {best_vol[0]}")
    record(
        "Highest_Volume",
        summary_map.get("Highest_Volume") is not None
        and int(summary_map["Highest_Volume"]) == int(best_vol[6]),
        f"got {summary_map.get('Highest_Volume')}, expected {best_vol[6]}",
    )
    # Report_Date
    if best_close[1] and "Report_Date" in summary_map:
        date_str = best_close[1].isoformat() if hasattr(best_close[1], "isoformat") else str(best_close[1])
        record("Report_Date matches", str(summary_map["Report_Date"]).startswith(date_str[:10]),
               f"got {summary_map.get('Report_Date')}, expected {date_str}")

    wb.close()


def check_pdf(workspace, prices, info):
    """Check Market_Report.pdf has correct content."""
    pdf_path = Path(workspace) / "Market_Report.pdf"
    if not pdf_path.exists():
        record("PDF file exists", False, str(pdf_path))
        return
    record("PDF file exists", True)

    size = pdf_path.stat().st_size
    record("PDF size > 500 bytes", size > 500, f"size={size}")

    # Extract PDF text
    text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        for page in reader.pages:
            text += page.extract_text() + "\n"
    except Exception:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(pdf_path))
            for page in reader.pages:
                text += page.extract_text() + "\n"
        except Exception as e:
            try:
                import pdfplumber
                with pdfplumber.open(str(pdf_path)) as pdf:
                    for page in pdf.pages:
                        text += (page.extract_text() or "") + "\n"
            except Exception as e2:
                record("PDF text extractable", False, f"errors: {e}, {e2}")
                return
    record("PDF text extractable", True)
    text_l = text.lower()

    # Title
    record(
        "PDF title 'Market Overview Report'",
        "market overview report" in text_l,
        f"text snippet: {text[:300]}",
    )
    # Each ticker symbol must appear
    for sym in SYMBOLS:
        record(f"PDF mentions {sym}", sym in text or sym.lower() in text_l, f"snippet: {text[:300]}")
    # Each company sector must appear (substring, accept first word as truncation tolerance)
    for sym, inf in info.items():
        sector = str(inf[6] or "").lower()
        if sector:
            sector_first = sector.split()[0]
            record(
                f"PDF mentions sector '{sector}' for {sym} (or first word)",
                sector in text_l or sector_first in text_l,
                f"snippet: {text_l[:300]}",
            )
    # Closing prices: each closing price in numerical form
    for p in prices:
        sym, _date, _o, _h, _l, close, _v = p
        close_str = f"{float(close):.2f}"
        record(
            f"PDF mentions {sym} closing price {close_str}",
            close_str in text or str(int(float(close))) in text,
            f"snippet: {text[:300]}",
        )


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, help="Launch time")
    args = parser.parse_args()

    workspace = args.agent_workspace
    if not workspace:
        print("Error: --agent_workspace is required")
        sys.exit(1)

    print("Fetching expected data from database...")
    try:
        prices, info = get_expected_data()
        print(f"  Stocks: {len(prices)}")
    except Exception as e:
        print(f"Error querying database: {e}")
        sys.exit(1)

    print("\n--- Check 1: Excel File ---")
    try:
        check_excel(workspace, prices, info)
    except Exception as e:
        record("Excel check exception", False, str(e))

    print("\n--- Check 2: PDF File ---")
    try:
        check_pdf(workspace, prices, info)
    except Exception as e:
        record("PDF check exception", False, str(e))

    print(f"\nOverall: {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks passed")
    if FAIL_COUNT == 0:
        print("Pass all tests!")
        sys.exit(0)
    else:
        print(f"FAIL ({FAIL_COUNT} failures)")
        sys.exit(1)
