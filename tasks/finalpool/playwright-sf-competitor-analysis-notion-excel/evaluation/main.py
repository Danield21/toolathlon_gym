"""Evaluation for playwright-sf-competitor-analysis-notion-excel."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "Competitive_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Competitive_Analysis.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("Competitive_Analysis.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Competitor Profiles sheet
        print("  Checking Competitor Profiles sheet...")
        a_rows = load_sheet_rows(agent_wb, "Competitor Profiles")
        g_rows = load_sheet_rows(gt_wb, "Competitor Profiles")
        if a_rows is None:
            all_errors.append("Sheet 'Competitor Profiles' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            if len(a_data) < 5:
                all_errors.append(f"Competitor Profiles row count: {len(a_data)}, expected 5")
            else:
                a_lookup = {}
                for row in a_data:
                    if row and row[0]:
                        a_lookup[str(row[0]).strip().lower()] = row
                for g_row in g_rows[1:]:
                    if not g_row or not g_row[0]:
                        continue
                    key = str(g_row[0]).strip().lower()
                    a_row = a_lookup.get(key)
                    if a_row is None:
                        all_errors.append(f"Missing competitor: {g_row[0]}")
                        continue
                    # Revenue_M - tighten tol 50 -> 5 ($5M)
                    if not num_close(a_row[1], g_row[1], 5):
                        all_errors.append(f"{g_row[0]} Revenue: {a_row[1]} vs {g_row[1]}")
                    # Market_Share_Pct
                    if not num_close(a_row[2], g_row[2], 0.5):
                        all_errors.append(f"{g_row[0]} Market_Share: {a_row[2]} vs {g_row[2]}")
                    # Product_Count (col 3) - integer count
                    if len(a_row) > 3 and len(g_row) > 3 and g_row[3] is not None:
                        if not num_close(a_row[3], g_row[3], 2):
                            all_errors.append(f"{g_row[0]} Product_Count: {a_row[3]} vs {g_row[3]}")
                    # Satisfaction (col 4) - rating with 1 decimal
                    if len(a_row) > 4 and len(g_row) > 4 and g_row[4] is not None:
                        if not num_close(a_row[4], g_row[4], 0.1):
                            all_errors.append(f"{g_row[0]} Satisfaction: {a_row[4]} vs {g_row[4]}")
                    # Position
                    if a_row[5] and g_row[5]:
                        if str(a_row[5]).strip().lower() != str(g_row[5]).strip().lower():
                            all_errors.append(f"{g_row[0]} Position: {a_row[5]} vs {g_row[5]}")
            print("    Done.")

        # Check Internal Performance sheet
        print("  Checking Internal Performance sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Internal Performance")
        g_rows2 = load_sheet_rows(gt_wb, "Internal Performance")
        if a_rows2 is None:
            all_errors.append("Sheet 'Internal Performance' not found in agent output")
        else:
            # Build region / segment lookups from agent rows
            a_region_map = {}
            a_segment_map = {}
            region_names = {"europe", "asia pacific", "north america", "middle east", "latin america"}
            segment_names = {"consumer", "enterprise", "government", "smb"}
            for row in a_rows2:
                if not row or not row[0]:
                    continue
                key = str(row[0]).strip().lower()
                if key in region_names:
                    a_region_map[key] = row
                elif key in segment_names:
                    a_segment_map[key] = row
            if len(a_region_map) < 5:
                all_errors.append(f"Internal Performance: found {len(a_region_map)}/5 regions")
            if len(a_segment_map) < 4:
                all_errors.append(f"Internal Performance: found {len(a_segment_map)}/4 segments")

            # Compare numeric values against GT row-by-row (tol 5% of value or 5000 abs)
            def _cmp_num_group(label, gt_map):
                for key, g_row in gt_map.items():
                    a_row = (a_region_map if key in region_names else a_segment_map).get(key)
                    if a_row is None:
                        all_errors.append(f"{label} '{key}' missing in agent output")
                        continue
                    # Orders (col 1) - tighter tol since DB-derived counts are exact
                    if len(g_row) > 1 and g_row[1] is not None:
                        if not num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], 10):
                            all_errors.append(
                                f"{label} '{key}' Orders: {a_row[1] if len(a_row) > 1 else None} vs {g_row[1]}")
                    # Revenue (col 2) - 5% tolerance or 5000
                    if len(g_row) > 2 and g_row[2] is not None:
                        tol = max(abs(float(g_row[2])) * 0.02, 5000.0)
                        if not num_close(a_row[2] if len(a_row) > 2 else None, g_row[2], tol):
                            all_errors.append(
                                f"{label} '{key}' Revenue: {a_row[2] if len(a_row) > 2 else None} vs {g_row[2]}")

            # Build GT maps
            gt_region_map = {}
            gt_segment_map = {}
            for row in (g_rows2 or []):
                if not row or not row[0]:
                    continue
                key = str(row[0]).strip().lower()
                if key in region_names:
                    gt_region_map[key] = row
                elif key in segment_names:
                    gt_segment_map[key] = row
            _cmp_num_group("Region", gt_region_map)
            _cmp_num_group("Segment", gt_segment_map)
            print("    Done.")

        # Check SWOT Summary sheet - must have 4 specific factors with values
        print("  Checking SWOT Summary sheet...")
        a_rows3 = load_sheet_rows(agent_wb, "SWOT Summary")
        g_rows3 = load_sheet_rows(gt_wb, "SWOT Summary")
        if a_rows3 is None:
            all_errors.append("Sheet 'SWOT Summary' not found in agent output")
        else:
            a_data3 = a_rows3[1:] if len(a_rows3) > 1 else []
            if len(a_data3) < 4:
                all_errors.append(f"SWOT Summary: {len(a_data3)} rows, expected at least 4")

            # Build lookup by factor keyword substring (task names factors as
            # Total Revenue / Market Breadth / Customer Base / Satisfaction)
            FACTOR_KEYS = {
                "total revenue": 0,    # match by prefix
                "market breadth": 1,
                "customer base": 2,
                "satisfaction": 3,
            }
            a_by_factor = {}
            for row in a_data3:
                if not row or not row[0]:
                    continue
                factor_low = str(row[0]).strip().lower()
                for kw in FACTOR_KEYS:
                    if kw in factor_low:
                        a_by_factor[kw] = row
                        break
            for kw in FACTOR_KEYS:
                if kw not in a_by_factor:
                    all_errors.append(f"SWOT Summary missing factor: '{kw}'")

            # Compare Internal_Value and Competitor_Avg numeric cols vs GT
            g_by_factor = {}
            for row in (g_rows3 or [])[1:]:
                if not row or not row[0]:
                    continue
                factor_low = str(row[0]).strip().lower()
                for kw in FACTOR_KEYS:
                    if kw in factor_low:
                        g_by_factor[kw] = row
                        break
            for kw, g_row in g_by_factor.items():
                a_row = a_by_factor.get(kw)
                if a_row is None:
                    continue
                # Internal_Value (col 1), Competitor_Avg (col 2)
                for col, label in [(1, "Internal_Value"), (2, "Competitor_Avg")]:
                    gv = g_row[col] if len(g_row) > col else None
                    av = a_row[col] if len(a_row) > col else None
                    if gv is None:
                        continue
                    # Use tolerance proportional to value
                    try:
                        gv_f = float(gv)
                        tol = max(abs(gv_f) * 0.05, 1.0)
                    except (TypeError, ValueError):
                        continue
                    if not num_close(av, gv, tol):
                        all_errors.append(f"SWOT '{kw}' {label}: {av} vs {gv}")
            print("    Done.")

    # --- Check 2: Notion database ---
    print("Checking Notion database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM notion.databases
            WHERE title::text ILIKE '%competitive%intelligence%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            all_errors.append("Notion database 'Competitive Intelligence Tracker' not found")
        else:
            # Check for competitor pages
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages p
                JOIN notion.databases d ON p.parent::jsonb->>'database_id' = d.id
                WHERE d.title::text ILIKE '%competitive%intelligence%'
                AND p.archived = false
            """)
            page_count = cur.fetchone()[0]
            if page_count < 5:
                all_errors.append(f"Notion: found {page_count}/5 competitor pages")
            else:
                print(f"    Found {page_count} competitor pages")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking Notion: {e}")

    # --- Check 3: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT from_addr, to_addr, subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%board@company.com%'
            AND subject ILIKE '%competitive positioning%'
        """)
        rows = cur.fetchall()
        if not rows:
            all_errors.append("No email to board@company.com with 'Competitive Positioning' subject")
        else:
            ok_any = False
            for from_addr, to_addr, subject, body in rows:
                subject = (subject or "").lower()
                from_s = (str(from_addr) or "").lower()
                body_s = (body or "").lower()
                if ("competitive positioning analysis" in subject and
                    "q1 2026" in subject and
                    "strategy@company.com" in from_s and
                    "alphatech" in body_s):
                    ok_any = True
                    break
            if not ok_any:
                all_errors.append(
                    "Email found but subject/from/body don't match "
                    "(need subject 'Competitive Positioning Analysis - Q1 2026', "
                    "from strategy@company.com, body mentions leading competitor AlphaTech)"
                )
            else:
                print(f"    Email with correct subject/from/body found")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
