#!/usr/bin/env python3
"""Evaluation script for product-quality-assurance-multi-batch."""

import argparse
import os
import sys
from pathlib import Path

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432,
        dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
        user="eigent", password="camel",
    )


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    if not agent_workspace:
        return False, "No agent workspace provided"

    agent_ws = Path(agent_workspace)
    if not agent_ws.exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    # Detect GT self-test mode (V1 parity test).
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    # Phase 5: Excel quality report
    xlsx_files = [p for p in agent_ws.glob("*.xlsx")
                  if 'data.csv' not in p.name.lower()]
    check("At least one xlsx quality report exists",
          len(xlsx_files) >= 1, f"found {len(xlsx_files)} xlsx files")
    if xlsx_files:
        try:
            import openpyxl
            audit_path = max(xlsx_files, key=lambda p: p.stat().st_size)
            wb = openpyxl.load_workbook(str(audit_path), data_only=True)
            sheets = [s.lower() for s in wb.sheetnames]
            has_qa_sheet = any(
                kw in s
                for s in sheets
                for kw in ['inspection', 'quality', 'defect', 'control',
                          'chart', 'corrective', 'analysis', 'summary',
                          'batch', 'spc']
            )
            check(f"Quality xlsx ({audit_path.name}) has quality-related sheet",
                  has_qa_sheet, f"sheets={wb.sheetnames}")
            max_rows = max((ws.max_row for ws in wb.worksheets), default=0)
            check("Quality xlsx has >=6 rows of data in some sheet (>=5 batches + header)",
                  max_rows >= 6, f"max rows={max_rows}")
            # Value-level: validate at least one defect_rate / defects column has plausible numeric values
            defect_validated = False
            distinct_batch_ids = set()
            for ws in wb.worksheets:
                headers = [str(c.value).lower() if c.value else '' for c in ws[1]]
                # Count distinct batch IDs - accept any non-empty identifier (UUID,
                # B-NNNN, Q1-001, BATCH-001, etc.) so agent has freedom of format.
                for col_idx, h in enumerate(headers):
                    if 'batch' in h and ('id' in h or h == 'batch' or 'no' in h or '#' in h):
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if col_idx < len(row) and row[col_idx] is not None:
                                v = str(row[col_idx]).strip()
                                # Accept any non-empty, non-trivial identifier
                                if v and len(v) >= 2 and v.lower() not in ('nan', 'none', 'null'):
                                    distinct_batch_ids.add(v)
                # Validate defect rate columns
                for col_idx, h in enumerate(headers):
                    if 'defect' in h and ('rate' in h or 'pct' in h or '%' in h):
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if col_idx < len(row) and row[col_idx] is not None:
                                try:
                                    fv = float(row[col_idx])
                                    if 0 <= fv <= 100:
                                        defect_validated = True
                                        break
                                except (ValueError, TypeError):
                                    pass
                        if defect_validated:
                            break
                if defect_validated:
                    break
            check("Quality xlsx has at least 5 distinct batch IDs",
                  len(distinct_batch_ids) >= 5,
                  f"found {len(distinct_batch_ids)} distinct batch IDs")
            check("Quality xlsx has at least one defect-rate column with values in 0-100 range",
                  defect_validated, "no Defect_Rate / Defects column with valid percentage found")
        except Exception as e:
            check("Quality xlsx parse", False, str(e))

    # Phase 5: Word recommendations doc
    docx_files = list(agent_ws.glob("*.docx"))
    check("At least one docx QA recommendations doc exists",
          len(docx_files) >= 1, f"found {len(docx_files)} docx files")
    if docx_files:
        try:
            from docx import Document
            audit_docx = max(docx_files, key=lambda p: p.stat().st_size)
            doc = Document(str(audit_docx))
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word doc has substantive content (>=200 chars)",
                  len(text) >= 200, f"len={len(text)}")
            has_topic = any(kw in text for kw in
                            ['quality', 'defect', 'inspection', 'control',
                             'batch', 'process', 'manufacturing'])
            has_recs = any(kw in text for kw in
                           ['recommendation', 'corrective', 'action',
                            'finding', 'improvement'])
            check("Word doc mentions quality/defect/inspection topic",
                  has_topic, f"text head: {text[:200]}")
            check("Word doc mentions recommendations/corrective/improvement",
                  has_recs, f"text head: {text[:200]}")
        except Exception as e:
            check("Word doc parse", False, str(e))

    # Phase 6: Email to production management
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, subject FROM email.messages
               WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s
                  OR body_text ILIKE %s""",
            ('%quality%', '%defect%', '%batch%', '%quality%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Quality-management email sent: 0 found (GT self-test, non-blocking)")
        else:
            check("Quality-management email sent",
                  len(rows) >= 1, f"found {len(rows)} matching emails")
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))

    # Phase 6: Quality review meeting
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT summary FROM gcal.events
               WHERE summary ILIKE %s OR summary ILIKE %s OR summary ILIKE %s""",
            ('%quality%', '%review%', '%production%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Quality review meeting scheduled: 0 found (GT self-test, non-blocking)")
        else:
            check("Quality review meeting scheduled",
                  len(rows) >= 1, f"found {len(rows)} matching events")
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))

    # Phase 6: Google Sheet for quality tracking
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, title FROM gsheet.spreadsheets
               WHERE title ILIKE %s OR title ILIKE %s OR title ILIKE %s""",
            ('%quality%', '%qa%', '%tracking%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Quality tracking GSheet created: 0 found (GT self-test, non-blocking)")
        else:
            check("Quality tracking GSheet created",
                  len(rows) >= 1, f"found {len(rows)} matching spreadsheets")
        conn.close()
    except Exception as e:
        check("GSheet check", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
