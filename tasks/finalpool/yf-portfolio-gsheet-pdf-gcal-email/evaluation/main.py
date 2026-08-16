"""Evaluation script for yf-portfolio-gsheet-pdf-gcal-email."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Check agent-produced Python script and JSON artifacts (per task.md)
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists (portfolio_monitor.py)",
          any("portfolio" in f.lower() and "monitor" in f.lower() for f in py_files) or len(py_files) >= 1,
          f"found: {py_files}")
    for req_json in ("stock_data.json", "policy_data.json", "portfolio_status.json"):
        jpath = os.path.join(agent_workspace, req_json)
        check(f"{req_json} exists", os.path.exists(jpath), f"Not found: {jpath}")

    # Load GT reference xlsx for expected values (GT-side only, not required in agent workspace)
    gt_path = os.path.join(groundtruth_workspace, "Portfolio_Dashboard_Reference.xlsx")
    gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None
    gt_sheets = {}
    if gt_wb:
        for sname in gt_wb.sheetnames:
            ws = gt_wb[sname]
            gt_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            gt_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_sheets[sname.lower()] = {'headers': gt_headers, 'rows': gt_rows}

    # Database checks
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr FROM email.messages WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) AND subject ILIKE '%portfolio%' AND subject ILIKE '%drift%'")
        email_row = cur.fetchone()
        check("Email with correct subject 'Portfolio Drift Alert' sent", email_row is not None, "no matching email found")
        if email_row:
            check("Email has recipient", email_row[1] is not None, f"to_addr: {email_row[1]}")
        cur.execute("SELECT summary, start_datetime FROM gcal.events WHERE summary ILIKE '%portfolio%' AND (summary ILIKE '%rebalancing%' OR summary ILIKE '%review%')")
        event_row = cur.fetchone()
        check("Calendar event 'Portfolio Rebalancing Review'", event_row is not None, "no matching event found")
        if event_row and event_row[1]:
            dt_str = str(event_row[1])
            check("Calendar event on 2026-03-17", "2026-03-17" in dt_str, f"got {dt_str}")
        # GSheet: verify "Portfolio Monitor Dashboard" with 3 sheets (Holdings/Performance/Rebalancing)
        cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%portfolio%' AND LOWER(title) LIKE '%dashboard%'")
        ss_rows = cur.fetchall()
        check("Google Sheet 'Portfolio Monitor Dashboard' exists", len(ss_rows) >= 1, f"got spreadsheets: {ss_rows}")
        if ss_rows:
            ss_id = ss_rows[0][0]
            cur.execute("SELECT title FROM gsheet.sheets WHERE spreadsheet_id = %s", (ss_id,))
            tabs = [str(r[0]).strip().lower() for r in cur.fetchall()]
            for required_sheet in ("holdings", "performance", "rebalancing"):
                check(f"GSheet has '{required_sheet}' tab", any(required_sheet in t for t in tabs), f"tabs: {tabs}")
            # Fetch all cells for structure + value check
            cur.execute("""
                SELECT s.title, c.row_index, c.col_index, c.value
                FROM gsheet.cells c
                JOIN gsheet.sheets s ON c.spreadsheet_id = s.spreadsheet_id AND c.sheet_id = s.id
                WHERE c.spreadsheet_id = %s
                ORDER BY s.title, c.row_index, c.col_index
            """, (ss_id,))
            sheet_cells = {}
            for title, ri, ci, val in cur.fetchall():
                key = str(title).strip().lower()
                sheet_cells.setdefault(key, {})[(ri, ci)] = val
            # Cross-check against GT reference xlsx (if available)
            sheet_mapping = {'holdings': 'holdings', 'performance': 'performance', 'rebalancing': 'rebalancing'}
            for logical_name, gt_name in sheet_mapping.items():
                matched_tab = next((t for t in sheet_cells if logical_name in t), None)
                if matched_tab is None:
                    continue
                cells = sheet_cells[matched_tab]
                gt_match = next((k for k in gt_sheets if logical_name in k), None)
                gt_data = gt_sheets.get(gt_match, {}) if gt_match else {}
                gt_headers = gt_data.get('headers', [])
                gt_rows = gt_data.get('rows', [])
                if not cells:
                    check(f"GSheet '{matched_tab}' has data", False, "empty")
                    continue
                # Locate header row (row 0)
                header_row = {ci: str(v or "").strip().lower() for (ri, ci), v in cells.items() if ri == 0}
                for h in gt_headers:
                    if h:
                        check(f"GSheet '{matched_tab}' has '{h}' column",
                              any(h in hv or hv in h for hv in header_row.values()),
                              f"headers: {list(header_row.values())}")
                # Data row count
                data_rows = sorted({ri for (ri, ci) in cells.keys() if ri > 0})
                min_rows = max(1, len(gt_rows) - 2)
                check(f"GSheet '{matched_tab}' has >= {min_rows} data rows",
                      len(data_rows) >= min_rows, f"got {len(data_rows)}")
                # Per-cell value comparison (first column assumed symbol-like, match by first-col value)
                if gt_headers and gt_rows:
                    # Build agent row dict: (symbol->dict of header_lower->val)
                    header_col_map = {v: ci for ci, v in header_row.items()}
                    first_header = gt_headers[0] if gt_headers else ''
                    agent_by_key = {}
                    for ri in data_rows:
                        key_ci = header_col_map.get(first_header)
                        if key_ci is None:
                            continue
                        row_key = cells.get((ri, key_ci))
                        if row_key is None: continue
                        agent_row_d = {}
                        for h, ci in header_col_map.items():
                            agent_row_d[h] = cells.get((ri, ci))
                        agent_by_key[str(row_key).strip().upper()] = agent_row_d
                    for gt_row in gt_rows:
                        if not gt_row or gt_row[0] is None: continue
                        gt_key = str(gt_row[0]).strip().upper()
                        if gt_key not in agent_by_key:
                            check(f"GSheet '{matched_tab}' row {gt_key} exists", False, f"keys: {list(agent_by_key.keys())[:5]}")
                            continue
                        agent_row_d = agent_by_key[gt_key]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row): continue
                            gv = gt_row[ci]
                            av = agent_row_d.get(gt_h)
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                tol = max(0.5, abs(gf) * 0.15)
                                check(f"GSheet '{matched_tab}' {gt_key}.{gt_h} ~{gf:.2f}",
                                      abs(gf - af) <= tol, f"got {af}")
                            elif gv is not None and av is not None:
                                import re as _re
                                gs = str(gv).strip().lower()
                                avs = str(av).strip().lower()
                                # Normalize punctuation for company-name comparison
                                # (e.g. "amazon.com, inc." vs "amazon.com inc.")
                                gs_norm = _re.sub(r'[.,;:!?]', '', gs)
                                avs_norm = _re.sub(r'[.,;:!?]', '', avs)
                                if gs:
                                    check(f"GSheet '{matched_tab}' {gt_key}.{gt_h} text",
                                          gs_norm == avs_norm or gs_norm in avs_norm or avs_norm in gs_norm,
                                          f"expected {gs[:50]}, got {avs[:50]}")
        # Reverse verification: noise emails should not be in Sent folder
        cur.execute("SELECT COUNT(*) FROM email.messages WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) AND subject ILIKE '%newsletter%'")
        noise_sent = cur.fetchone()[0]
        check("No noise emails in Sent folder", noise_sent == 0, f"found {noise_sent} noise emails in Sent")
        # Reverse verification: noise events should not match task keyword
        cur.execute("SELECT COUNT(*) FROM gcal.events WHERE summary ILIKE '%standup%' OR summary ILIKE '%lunch%'")
        noise_events = cur.fetchone()[0]
        check("Noise events exist (not deleted by agent)", noise_events >= 1, f"noise events: {noise_events}")
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()