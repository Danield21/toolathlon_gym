#!/usr/bin/env python3
"""Regenerate Portfolio_Dashboard_Reference.xlsx (GT) for
yf-portfolio-gsheet-pdf-gcal-email from the yf.stock_prices seed data.

Spec (task.md + Investment_Policy.pdf, P3 of 2026-08-14-clean-and-rerun-v2.md):
  - tracked: GOOGL AMZN JPM JNJ XOM, equal $10,000 initial per position
  - Purchase_Price = close on 2025-03-07 (yf.stock_prices)
  - Current_Price  = latest close in the DB (2026-03-05)
  - Shares_Held    = 10000 / Purchase_Price (2dp)
  - Market_Value   = Shares_Held * Current_Price (2dp)
  - Allocation_Pct = Market_Value / total (1dp)
  - Return_Pct     = (Current/Purchase - 1) * 100 (2dp)
  - Status         = Gain / Loss
  - Target_Alloc   = 20% each; Drift = Current - Target (1dp)
  - Action         = Buy / Sell / Hold (drift direction & magnitude, 3pp threshold)
Keeps the original workbook layout (3 sheets, same headers, sorted by Symbol).
"""
import gzip
import re
import sys
from pathlib import Path

import openpyxl

DB_INIT = Path("/lintaoLab2/bowending/project_agent_swarm_benchmark/toolathlon_gym_eval_dockers/toolathlon-pack-rootfs/workspace/db/init.sql.gz")
GT_XLSX = Path("/lintaoLab2/bowending/project_agent_swarm_benchmark/toolathlon_gym/tasks/finalpool/yf-portfolio-gsheet-pdf-gcal-email/groundtruth_workspace/Portfolio_Dashboard_Reference.xlsx")

SYMBOLS = {"GOOGL", "AMZN", "JPM", "JNJ", "XOM"}
COMPANY = {
    "GOOGL": "Alphabet Inc.", "AMZN": "Amazon.com Inc.", "JPM": "JPMorgan Chase & Co.",
    "JNJ": "Johnson & Johnson", "XOM": "Exxon Mobil Corp.",
}
SECTOR = {
    "GOOGL": "Communication Services", "AMZN": "Consumer Cyclical",
    "JPM": "Financial Services", "JNJ": "Healthcare", "XOM": "Energy",
}
PURCHASE_DATE = "2025-03-07"
INITIAL_PER_STOCK = 10_000.0
TARGET = 20.0
DRIFT_THRESHOLD = 3.0


def load_prices():
    with gzip.open(DB_INIT, "rt", encoding="utf-8", errors="replace") as f:
        inside = False
        prices = {}  # symbol -> list[(date, close)]
        for line in f:
            if line.startswith("COPY yf.stock_prices"):
                inside = True
                continue
            if inside:
                if line.startswith("\\."):
                    break
                parts = line.rstrip("\n").split("\t")
                sym, date, close = parts[0], parts[1], float(parts[5])
                if sym in SYMBOLS:
                    prices.setdefault(sym, []).append((date, close))
        for sym in prices:
            prices[sym].sort()
        return prices


def main():
    prices = load_prices()
    rows = []
    for sym in sorted(SYMBOLS):
        hist = prices.get(sym, [])
        purchase = next((c for d, c in hist if d == PURCHASE_DATE), None)
        current = hist[-1][1] if hist else None
        current_date = hist[-1][0] if hist else None
        if purchase is None or current is None:
            sys.exit(f"FATAL: no data for {sym} (purchase={purchase}, current={current})")
        rows.append({
            "sym": sym, "purchase": purchase, "current": current,
            "current_date": current_date,
        })

    total = 0.0
    for r in rows:
        r["shares"] = round(INITIAL_PER_STOCK / r["purchase"], 2)
        r["mv"] = round(r["shares"] * r["current"], 2)
        total += r["mv"]

    for r in rows:
        r["alloc"] = round(r["mv"] / total * 100, 1)
        r["ret"] = round((r["current"] / r["purchase"] - 1) * 100, 2)
        r["status"] = "Gain" if r["ret"] >= 0 else "Loss"
        r["drift"] = round(r["alloc"] - TARGET, 1)
        if abs(r["drift"]) < DRIFT_THRESHOLD:
            r["action"] = "Hold"
        elif r["drift"] > 0:
            r["action"] = "Sell"
        else:
            r["action"] = "Buy"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holdings"
    ws.append(["Symbol", "Company", "Sector", "Current_Price", "Shares_Held",
               "Market_Value", "Allocation_Pc"])
    for r in rows:
        ws.append([r["sym"], COMPANY[r["sym"]], SECTOR[r["sym"]],
                   round(r["current"], 2), r["shares"], r["mv"], r["alloc"]])

    ws2 = wb.create_sheet("Performance")
    ws2.append(["Symbol", "Purchase_Price", "Current_Price", "Return_Pct", "Status"])
    for r in rows:
        ws2.append([r["sym"], round(r["purchase"], 2), round(r["current"], 2),
                    r["ret"], r["status"]])

    ws3 = wb.create_sheet("Rebalancing")
    ws3.append(["Symbol", "Current_Allocation", "Target_Allocation", "Drift_Pct", "Action"])
    for r in rows:
        ws3.append([r["sym"], r["alloc"], TARGET, r["drift"], r["action"]])

    wb.save(GT_XLSX)

    print(f"as-of dates: purchase={PURCHASE_DATE} current={rows[0]['current_date']}")
    print(f"{'sym':6} {'purchase':>9} {'current':>9} {'shares':>8} {'mv':>10} {'alloc%':>7} {'ret%':>7} {'action':>6}")
    for r in rows:
        print(f"{r['sym']:6} {r['purchase']:9.2f} {r['current']:9.2f} {r['shares']:8.2f} "
              f"{r['mv']:10.2f} {r['alloc']:7.1f} {r['ret']:7.2f} {r['action']:>6}")
    print(f"total market value: {total:.2f}")
    print(f"saved: {GT_XLSX}")


if __name__ == "__main__":
    main()
