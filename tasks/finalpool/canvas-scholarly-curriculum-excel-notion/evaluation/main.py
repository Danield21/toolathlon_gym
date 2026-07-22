"""Evaluation script for canvas-scholarly-curriculum-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Check Curriculum_Review_Report.xlsx
    excel_path = os.path.join(agent_workspace, "Curriculum_Review_Report.xlsx")
    check("Curriculum_Review_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Curriculum_Review_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        if gt_wb:
            for sheet_name in gt_wb.sheetnames:
                check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    gt_ws = gt_wb[sheet_name]
                    gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                    for h in gt_headers:
                        if h:
                            check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                    gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                               if r and any(c is not None for c in r)]
                    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                                 if r and any(c is not None for c in r)]
                    check(f"{sheet_name} has >= {len(gt_rows)} data rows",
                          len(data_rows) >= len(gt_rows), f"got {len(data_rows)}")

                    # Structural validation per sheet (no exact-cell match
                    # because positional row alignment is unreliable when
                    # agent sorts differently, and GT values are illustrative).
                    header_map = {h: i for i, h in enumerate(headers)}
                    sn_lower = sheet_name.strip().lower()

                    if "current_courses" in sn_lower or "current courses" in sn_lower:
                        # Validate sort by Course_Name (ascending) and non-empty key columns.
                        cn_idx = header_map.get("course_name", -1)
                        cc_idx = header_map.get("course_code", -1)
                        ec_idx = header_map.get("enrollment_count", -1)
                        non_empty_ok = True
                        sort_ok = True
                        prev_name = None
                        for r in data_rows:
                            for idx in [cn_idx, cc_idx]:
                                if idx >= 0 and (idx >= len(r) or r[idx] is None or str(r[idx]).strip() == ""):
                                    non_empty_ok = False
                            if cn_idx >= 0 and cn_idx < len(r) and r[cn_idx] is not None:
                                v = str(r[cn_idx]).strip().lower()
                                if prev_name is not None and v < prev_name:
                                    sort_ok = False
                                prev_name = v
                            if ec_idx >= 0 and ec_idx < len(r) and r[ec_idx] is not None:
                                try:
                                    val = float(r[ec_idx])
                                    if val < 0:
                                        non_empty_ok = False
                                except (TypeError, ValueError):
                                    pass
                        check(f"{sheet_name} all rows non-empty in Course_Name/Course_Code",
                              non_empty_ok, "")
                        check(f"{sheet_name} sorted by Course_Name ascending",
                              sort_ok, "")

                    elif "research_trends" in sn_lower or "research trends" in sn_lower:
                        # Validate Relevance_to_Curriculum enum and sort by Citations descending.
                        rel_idx = header_map.get("relevance_to_curriculum", -1)
                        cit_idx = header_map.get("citations", -1)
                        valid_rel = {"high", "medium", "low"}
                        rel_ok = True
                        sort_ok = True
                        prev_cit = None
                        non_empty_ok = True
                        title_idx = header_map.get("paper_title", -1)
                        for r in data_rows:
                            if rel_idx >= 0 and rel_idx < len(r) and r[rel_idx] is not None:
                                v = str(r[rel_idx]).strip().lower()
                                if v and v not in valid_rel:
                                    rel_ok = False
                            if cit_idx >= 0 and cit_idx < len(r) and r[cit_idx] is not None:
                                try:
                                    cv = float(r[cit_idx])
                                    if prev_cit is not None and cv > prev_cit + 0.01:
                                        sort_ok = False
                                    prev_cit = cv
                                except (TypeError, ValueError):
                                    pass
                            if title_idx >= 0 and (title_idx >= len(r) or r[title_idx] is None or str(r[title_idx]).strip() == ""):
                                non_empty_ok = False
                        check(f"{sheet_name} Relevance_to_Curriculum in {{High,Medium,Low}}",
                              rel_ok, "")
                        check(f"{sheet_name} sorted by Citations descending",
                              sort_ok, "")
                        check(f"{sheet_name} all rows have Paper_Title",
                              non_empty_ok, "")
                        # Anti-fabrication: at least 1 row's Paper_Title must
                        # reference one of the 5 injected scholarly papers
                        # (by topic-related keywords).
                        if title_idx >= 0:
                            titles = [
                                str(r[title_idx] or "").lower()
                                for r in data_rows
                                if title_idx < len(r) and r[title_idx] is not None
                            ]
                            relevant_keywords = [
                                "machine learning",
                                "data analytics",
                                "computational thinking",
                                "deep learning",
                                "big data",
                                "pattern recognition",
                            ]
                            grounded_count = sum(
                                1
                                for t in titles
                                if any(kw in t for kw in relevant_keywords)
                            )
                            check(
                                f"{sheet_name} at least 2 rows reference real injected research topics",
                                grounded_count >= 2,
                                f"grounded titles: {grounded_count}, total titles: {len(titles)}",
                            )

                    elif "gap_analysis" in sn_lower or "gap analysis" in sn_lower:
                        # Validate required metrics present.
                        metric_idx = header_map.get("metric", 0)
                        val_idx = header_map.get("value", 1)
                        metrics = {}
                        for r in data_rows:
                            if metric_idx < len(r) and r[metric_idx] is not None:
                                metrics[str(r[metric_idx]).strip().lower()] = (
                                    r[val_idx] if val_idx < len(r) else None
                                )
                        required = ["total_courses", "papers_reviewed", "high_relevance_papers",
                                    "curriculum_coverage_pct", "top_gap_area"]
                        for m in required:
                            present = any(m in k or k in m for k in metrics.keys())
                            check(f"{sheet_name} contains metric '{m}'",
                                  present, f"metrics found: {list(metrics.keys())[:8]}")

                        # Validate Total_Courses against canvas DB count.
                        try:
                            tc_conn = get_conn()
                            tc_cur = tc_conn.cursor()
                            tc_cur.execute("SELECT COUNT(DISTINCT id) FROM canvas.courses")
                            db_total = int(tc_cur.fetchone()[0] or 0)
                            tc_cur.close()
                            tc_conn.close()
                        except Exception:
                            db_total = None
                        if db_total is not None:
                            tc_val = None
                            for k, v in metrics.items():
                                if "total_courses" in k or "total courses" in k:
                                    tc_val = safe_float(v)
                                    break
                            check(
                                f"{sheet_name} Total_Courses matches canvas.courses count ({db_total})",
                                tc_val is not None and abs(tc_val - db_total) <= 0.5,
                                f"Total_Courses metric value: {tc_val}, expected: {db_total}",
                            )

                        # Validate Top_Gap_Area is a recognised research topic.
                        top_gap = None
                        for k, v in metrics.items():
                            if "top_gap_area" in k or "top gap area" in k:
                                top_gap = str(v or "").strip().lower()
                                break
                        valid_topics = [
                            "machine learning", "data analytics", "data science",
                            "computational thinking", "deep learning", "big data",
                            "cs education", "computer science education",
                            "pattern recognition", "ai", "artificial intelligence",
                        ]
                        if top_gap is not None:
                            top_gap_ok = any(t in top_gap for t in valid_topics)
                            check(
                                f"{sheet_name} Top_Gap_Area names a recognised topic",
                                top_gap_ok and len(top_gap) > 0,
                                f"Top_Gap_Area value: {top_gap!r}",
                            )

                        # Validate Curriculum_Coverage_Pct is in [0, 100] range.
                        cov_val = None
                        for k, v in metrics.items():
                            if "curriculum_coverage_pct" in k or "coverage_pct" in k:
                                cov_val = safe_float(v)
                                break
                        if cov_val is not None:
                            check(
                                f"{sheet_name} Curriculum_Coverage_Pct in [0, 100]",
                                0 <= cov_val <= 100,
                                f"Curriculum_Coverage_Pct value: {cov_val}",
                            )

                        # Validate High_Relevance_Papers <= Papers_Reviewed.
                        hrp_val = None
                        pr_val = None
                        for k, v in metrics.items():
                            if "high_relevance_papers" in k:
                                hrp_val = safe_float(v)
                            elif "papers_reviewed" in k:
                                pr_val = safe_float(v)
                        if hrp_val is not None and pr_val is not None:
                            check(
                                f"{sheet_name} High_Relevance_Papers <= Papers_Reviewed",
                                hrp_val <= pr_val + 0.001,
                                f"High_Relevance: {hrp_val}, Papers_Reviewed: {pr_val}",
                            )

    # Check required Python script filename and JSON outputs (per task spec).
    script_path = os.path.join(agent_workspace, "curriculum_reviewer.py")
    check("curriculum_reviewer.py exists", os.path.exists(script_path))
    workspace_files = set(os.listdir(agent_workspace))
    check("course_data.json exists",
          "course_data.json" in workspace_files,
          f"workspace: {sorted(workspace_files)[:25]}")
    check("research_papers.json exists",
          "research_papers.json" in workspace_files,
          f"workspace: {sorted(workspace_files)[:25]}")
    check("curriculum_review.json exists",
          "curriculum_review.json" in workspace_files,
          f"workspace: {sorted(workspace_files)[:25]}")

    # Database checks
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Notion page with title 'Curriculum Innovation Tracker'.
        # Tighten: read the title field specifically, not the entire
        # properties JSON text.
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
        page_rows = cur.fetchall()
        match_count = 0
        for pid, props in page_rows:
            title_text = ""
            if isinstance(props, dict):
                title_obj = props.get("title", {})
                if isinstance(title_obj, dict):
                    title_list = title_obj.get("title", [])
                    if isinstance(title_list, list):
                        for t in title_list:
                            if isinstance(t, dict):
                                title_text += t.get("text", {}).get("content", "")
            if "curriculum innovation tracker" in title_text.lower():
                match_count += 1
        check("Notion page 'Curriculum Innovation Tracker' created",
              match_count >= 1, f"matching title pages: {match_count}, total pages: {len(page_rows)}")

        # Notion content has heading 'Research-Curriculum Alignment Review'
        cur.execute("""
            SELECT block_data FROM notion.blocks
            WHERE archived = false AND type LIKE 'heading%'
        """)
        heading_blocks = cur.fetchall()
        joined_headings = " ".join(json.dumps(b[0]) if b[0] else "" for b in heading_blocks).lower()
        check("Notion has heading 'Research-Curriculum Alignment Review'",
              "research-curriculum alignment review" in joined_headings or "research curriculum alignment review" in joined_headings,
              f"sample: {joined_headings[:200]}")
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()