"""Evaluation for canvas-curriculum-audit."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_compliance():
    """Get expected compliance data from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name,
               (SELECT COUNT(*) FROM canvas.assignments a WHERE a.course_id = c.id) as asgn,
               (SELECT COUNT(*) FROM canvas.quizzes q WHERE q.course_id = c.id) as quiz,
               (SELECT COUNT(*) FROM canvas.modules m WHERE m.course_id = c.id) as mods,
               CASE WHEN c.syllabus_body IS NOT NULL AND c.syllabus_body != '' THEN true ELSE false END as has_syl
        FROM canvas.courses c ORDER BY c.name
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    result = {}
    for name, asgn, quiz, mods, syl in rows:
        compliant = asgn >= 8 and quiz >= 3 and mods >= 4 and syl
        result[name.lower()] = {
            "assignments": asgn, "quizzes": quiz, "modules": mods,
            "syllabus": syl, "compliant": compliant,
        }
    return result


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Curriculum_Audit.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Curriculum_Audit.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Curriculum_Audit.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_compliance()

    # Check Compliance Matrix sheet
    cm_rows = load_sheet_rows(wb, "Compliance Matrix")
    if cm_rows is None:
        check("Sheet 'Compliance Matrix' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Compliance Matrix' exists", True)
        data_rows = cm_rows[1:] if len(cm_rows) > 1 else []
        check("Compliance Matrix has 22 rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        # Check header
        header = cm_rows[0] if cm_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "assignments_count", "quizzes_count", "modules_count", "has_syllabus", "compliant_yn"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

        # Validate ALL 22 courses counts and compliance against DB-derived expected
        courses_validated = 0
        course_names_in_order = []
        for row in data_rows:
            if not row or not row[0]:
                continue
            course_key = str(row[0]).strip().lower()
            course_names_in_order.append(str(row[0]).strip())
            if course_key in expected:
                exp = expected[course_key]
                short = course_key[:40]
                check(f"'{short}' assignments={exp['assignments']}",
                      num_close(row[1], exp["assignments"]), f"Got {row[1]}")
                check(f"'{short}' quizzes={exp['quizzes']}",
                      num_close(row[2], exp["quizzes"]), f"Got {row[2]}")
                check(f"'{short}' modules={exp['modules']}",
                      num_close(row[3], exp["modules"]), f"Got {row[3]}")
                # Has_Syllabus (col 4) must match DB
                syl_cell = str(row[4]).strip().lower() if row[4] is not None else ""
                exp_syl = "yes" if exp["syllabus"] else "no"
                syl_ok = syl_cell in ("yes", "y", "true", "1") if exp["syllabus"] else syl_cell in ("no", "n", "false", "0")
                check(f"'{short}' has_syllabus={exp_syl}", syl_ok, f"Got {row[4]}")
                # Compliant_YN
                comp_cell = str(row[5]).strip().lower() if row[5] is not None else ""
                if exp["compliant"]:
                    comp_ok = comp_cell in ("yes", "y", "true", "1")
                else:
                    comp_ok = comp_cell in ("no", "n", "false", "0")
                check(f"'{short}' compliant_yn", comp_ok, f"Got {row[5]}")
                # Issues column (index 6): if non-compliant, must list at least one deficiency
                issues_cell = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
                if exp["compliant"]:
                    # Compliant courses: Issues should be 'None' or empty
                    issues_ok = issues_cell.lower() in ("none", "", "n/a", "-") or issues_cell.lower().startswith("no ")
                    check(f"'{short}' issues=None (compliant)", issues_ok,
                          f"Got {issues_cell[:100]}")
                else:
                    # Non-compliant: must list a deficiency keyword
                    issues_lower = issues_cell.lower()
                    deficiency_kws = ["assignment", "quiz", "module", "syllabus"]
                    issues_ok = any(kw in issues_lower for kw in deficiency_kws)
                    check(f"'{short}' issues lists deficiencies (non-compliant)", issues_ok,
                          f"Got {issues_cell[:100]}")
                courses_validated += 1
        check(f"All 22 courses validated", courses_validated == 22,
              f"Validated {courses_validated}")

        # Alphabetical sort check
        sorted_names = sorted(course_names_in_order)
        check("Courses sorted alphabetically",
              course_names_in_order == sorted_names,
              f"Order mismatch at first diff")

    # Check Summary sheet
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        check("Total_Courses = 22", num_close(lookup.get("total_courses"), 22),
              f"Got {lookup.get('total_courses')}")

        compliant_expected = sum(1 for v in expected.values() if v["compliant"])
        non_compliant_expected = len(expected) - compliant_expected
        check(f"Compliant_Courses = {compliant_expected}",
              num_close(lookup.get("compliant_courses"), compliant_expected),
              f"Got {lookup.get('compliant_courses')}")
        check(f"Non_Compliant_Courses = {non_compliant_expected}",
              num_close(lookup.get("non_compliant_courses"), non_compliant_expected),
              f"Got {lookup.get('non_compliant_courses')}")
        # Compliance rate: strip any % sign and compare as float
        rate_cell = lookup.get("compliance_rate")
        rate_expected = round(100.0 * compliant_expected / max(len(expected), 1), 1)
        def _parse_pct(v):
            if v is None:
                return None
            try:
                return float(str(v).strip().rstrip("%"))
            except (TypeError, ValueError):
                return None
        rate_actual = _parse_pct(rate_cell)
        check(f"Compliance_Rate ~= {rate_expected}%",
              rate_actual is not None and abs(rate_actual - rate_expected) <= 1.0,
              f"Got {rate_cell}")


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Audit_Report.docx")
    if not os.path.isfile(docx_path):
        check("Audit_Report.docx exists", False, f"Not found: {docx_path}")
        return
    check("Audit_Report.docx exists", True)
    check("Word doc has content (> 1KB)", os.path.getsize(docx_path) > 1000,
          f"Size: {os.path.getsize(docx_path)}")

    try:
        from docx import Document
        doc = Document(docx_path)
        paragraphs = [p.text for p in doc.paragraphs]
        all_text = " ".join(paragraphs).lower()
        check("Report mentions compliance", "compli" in all_text, f"Sample: {all_text[:200]}")
        check("Report mentions recommendations", "recommend" in all_text, f"Sample: {all_text[:200]}")
        # Executive summary section
        check("Report has executive summary phrase",
              "executive summary" in all_text,
              f"Sample: {all_text[:200]}")
        # Non-compliance / areas of non-compliance
        has_noncompliance = any(p in all_text for p in [
            "non-compliance", "non compliance", "non-compliant", "non compliant", "areas of"
        ])
        check("Report discusses non-compliance areas", has_noncompliance,
              f"Sample: {all_text[:200]}")
        # At least one recommendation paragraph + 'recommendations' heading
        recommend_paragraphs = [p for p in paragraphs if p and "recommend" in p.lower()]
        check("Report has at least 1 recommendations section paragraph",
              len(recommend_paragraphs) >= 1,
              f"Found {len(recommend_paragraphs)} paragraphs with 'recommend'")
        # Key findings — structural/semantic check (not a literal phrase gate).
        # task.md asks the executive summary to "describe the overall compliance
        # rate and key findings"; that is satisfied by any narrative that states
        # a compliance rate AND draws at least one data-grounded conclusion.
        # Requiring the literal substring "key finding(s)" is a brittle口径 that
        # fails semantically-correct reports (case-study 2026-08-12, case #9).
        import re
        has_rate = bool(re.search(r"\d+\s*%", all_text))
        finding_synonyms = [
            "finding", "findings", "observation", "observations",
            "insight", "insights", "highlight", "highlights",
            "notable", "takeaway", "takeaways", "conclusion", "conclusions",
        ]
        has_finding = any(s in all_text for s in finding_synonyms)
        check("Report states a compliance rate (%) and at least one finding/observation",
              has_rate and has_finding,
              f"rate={has_rate}, finding={has_finding}; sample: {all_text[:200]}")
    except ImportError:
        check("python-docx available", False, "pip install python-docx")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ===")
    if FAIL_COUNT > 0:
        print(f"{FAIL_COUNT} checks failed")
        sys.exit(1)
    else:
        print("All checks passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
