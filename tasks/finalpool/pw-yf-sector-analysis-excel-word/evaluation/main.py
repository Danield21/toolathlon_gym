"""Evaluation script for pw-yf-sector-analysis-excel-word."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-yf-sector-analysis-excel-word"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Sector_Analysis_Report.xlsx")
    check("Sector_Analysis_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Sector_Analysis_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        # Synonyms / alternative phrasings each required column may appear as.
        # Used as a deterministic pre-LLM fallback for smart_column_exists so
        # agents that follow the generic task.md ("primary dimension, internal
        # metric, external benchmark, gap") still pass even if the LLM mapper
        # is unavailable in CI.
        COLUMN_SYNONYMS = {
            'Symbol': ['symbol', 'ticker', 'stock', 'code'],
            'Name': ['name', 'company', 'company_name', 'company name', 'stock_name'],
            'Sector': ['sector', 'industry', 'category', 'group', 'dimension', 'primary_dimension', 'segment'],
            'Current_Price': ['current_price', 'current price', 'price', 'market_price',
                              'last_price', 'close', 'internal', 'internal_metric',
                              'internal_value', 'actual'],
            'Target_Price': ['target_price', 'target price', 'analyst_target', 'fair_value',
                             'benchmark', 'external', 'external_benchmark', 'expected',
                             'forecast', 'projected_price'],
            'Upside': ['upside', 'upside_pct', 'upside_percent', 'gap', 'difference',
                       'delta', 'variance', 'gap_pct', 'percent_change', 'pct_diff'],
            'Metric': ['metric', 'name', 'key', 'kpi', 'measure'],
            'Value': ['value', 'amount', 'number', 'val'],
            'Priority': ['priority', 'rank', 'order', 'level'],
            'Action': ['action', 'recommendation', 'recommended_action', 'next_step', 'item'],
        }

        def _header_match_synonym(expected, headers_lower):
            for syn in COLUMN_SYNONYMS.get(expected, []):
                if syn.lower() in headers_lower:
                    return syn
            # Substring fallback (e.g., "stock_symbol" contains "symbol")
            for syn in COLUMN_SYNONYMS.get(expected, []):
                for h in headers_lower:
                    if syn.lower() in h or h in syn.lower():
                        return h
            return None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2 -> deterministic synonym
            fallback -> strict header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            _raw_headers = [str(c.value).strip() if c.value else "" for c in _ws[1]]
            _headers_lower = [h.lower() for h in _raw_headers]
            for _exp in expected_cols:
                # 1) exact header match
                if _exp.lower() in _headers_lower:
                    check(f"{sheet_name} has {_exp} column", True, "exact match")
                    continue
                # 2) deterministic synonym fallback
                _syn = _header_match_synonym(_exp, _headers_lower)
                if _syn:
                    check(f"{sheet_name} has {_exp} column", True,
                          f"matched via synonym {_syn!r}")
                    continue
                # 3) verify_v2 LLM mapping (last resort, if available)
                if _HAS_VERIFY_V2 and gt_wb is not None:
                    _raw_h_v2, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_h_v2,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
                else:
                    check(f"{sheet_name} has {_exp} column",
                          False, f"headers: {_raw_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Core required: Symbol/ticker, Name, Sector identification.
            # Secondary: pricing/benchmark/gap (any naming).
            check_columns('Data_Analysis',
                          ['Symbol', 'Name', 'Sector',
                           'Current_Price', 'Target_Price', 'Upside'], 5)
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 3)
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action', 'Symbol'], 2)

    # Word/processor checks moved out of Excel-existence block to ensure they
    # are always enumerated even if Excel is missing.
    import glob as globmod
    word_path = os.path.join(agent_workspace, "Sector_Analysis_Analysis.docx")
    word_files = []
    if os.path.exists(word_path):
        check("Sector_Analysis_Analysis.docx exists", True)
        word_files = [word_path]
    else:
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Sector_Analysis_Analysis.docx exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
    if word_files:
        from docx import Document
        doc = Document(word_files[0])
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Word has content (>=200 chars)", len(text) > 200, f"text length: {len(text)}")
        # Required sections per task spec: executive summary, key findings, recommendations
        section_keywords = {
            "executive summary / overview": ["executive summary", "overview", "summary"],
            "key findings": ["key findings", "findings"],
            "recommendations": ["recommendation"],
        }
        for sect_name, kws in section_keywords.items():
            hit = any(kw in text for kw in kws)
            check(f"Word has {sect_name} section", hit, f"none of {kws} found")
        # Topic relevance: must mention sector or stocks
        topic_terms = ["sector", "stock", "market", "industry"]
        hits = [t for t in topic_terms if t in text]
        check("Word covers sector/market topic", len(hits) >= 1, f"missing: {topic_terms}")

    check("yf_sector_processor.py exists", os.path.exists(os.path.join(agent_workspace, "yf_sector_processor.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
