"""
Evaluation for arxiv-method-benchmark-tracker task.
Checks Excel and Notion output.

Robustness notes:
- DB connection reads PG* env vars with defaults (R1).
- Excel is read with data_only=False plus a cached-value fallback so that
  formula cells (which openpyxl can't compute) do not silently become None.
  Summary/metric values are compared numerically after tolerant parsing (R2/R3).
- check_excel is crash-safe: a corrupt/partial xlsx degrades to FAIL entries
  instead of killing the whole evaluator process (multi-agent shared-file risk).
"""
import argparse
import os
import sys
import time

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

LEADERBOARD = [
    {"task": "Image Classification", "method": "ViT-Large", "score": 91.2, "paper_id": "2402.10001"},
    {"task": "Image Classification", "method": "ConvNeXt-XL", "score": 89.5, "paper_id": "2402.10002"},
    {"task": "Text Generation", "method": "GPT-4", "score": 95.0, "paper_id": ""},
    {"task": "Text Generation", "method": "LLaMA-3", "score": 92.3, "paper_id": "2402.10003"},
    {"task": "Image Generation", "method": "DiffusionXL", "score": 2.1, "paper_id": "2402.10004"},
]

PAPERS_WITH_ID = {"2402.10001", "2402.10002", "2402.10003", "2402.10004"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(v):
    """Tolerant numeric parsing: str/int/float/None, strips currency symbols,
    thousands separators, percent signs and surrounding whitespace.
    Returns None when the value cannot be interpreted as a number."""
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        sign = 1.0
        if s.startswith("-"):
            sign = -1.0
            s = s[1:]
        elif s.startswith("+"):
            s = s[1:]
        s = s.replace(",", "").replace("$", "").replace("¥", "").replace("€", "")
        s = s.replace("%", "").replace(" ", "")
        if not s:
            return None
        try:
            return sign * float(s)
        except ValueError:
            return None
    return None


def num_close(a, b, tol=1.0):
    """Numeric closeness with tolerant parsing. If both sides parse as numbers,
    compare |a-b| <= tol. Otherwise fall back to a case-insensitive string
    comparison (e.g. missing/blank values)."""
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if a is None and b is None:
        return True
    return str(a).strip().lower() == str(b).strip().lower()


def load_workbook_safe(path):
    """Load a workbook crash-safely. Returns ((formula_wb, cached_wb), None) on
    success, or (None, error_message) when the file is unreadable/corrupt."""
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:  # corrupt zip / unsupported format
        return None, str(e)
    wb_cached = None
    try:
        wb_cached = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        wb_cached = None
    return (wb, wb_cached), None


def _cell_value(wb_formula, wb_cached, ws_name, row_idx, col_idx):
    """Read a cell, resolving formula cells to their cached computed value when
    available (a formula without a cached value stays a string and will fail
    numeric checks -- the task requires literal values)."""
    v = wb_formula[ws_name].cell(row=row_idx, column=col_idx).value
    if isinstance(v, str) and v.startswith("="):
        if wb_cached is not None:
            cached = wb_cached[ws_name].cell(row=row_idx, column=col_idx).value
            if cached is not None:
                return cached
        return v
    return v


def load_sheet_rows(wb, wb_cached, sheet_name):
    """Return the rows of the sheet whose name matches (case/space/underscore
    insensitive). None when no such sheet exists."""
    target = None
    for name in wb.sheetnames:
        n1 = name.strip().lower().replace(" ", "_")
        n2 = name.strip().lower().replace("_", " ")
        s1 = sheet_name.strip().lower().replace(" ", "_")
        s2 = sheet_name.strip().lower().replace("_", " ")
        if n1 == s1 or n2 == s2:
            target = name
            break
    if target is None:
        return None
    rows = []
    for row in wb[target].iter_rows():
        vals = []
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and wb_cached is not None:
                cached = wb_cached[target].cell(row=cell.row, column=cell.column).value
                if cached is not None:
                    v = cached
            vals.append(v)
        rows.append(vals)
    return rows


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Method_Benchmark.xlsx")
    if not os.path.isfile(path):
        record("Excel file exists", False, f"Not found: {path}")
        return False
    record("Excel file exists", True)

    try:
        loaded, err = load_workbook_safe(path)
        if loaded is None:
            # The workbook may be momentarily unreadable if a peer agent is
            # writing the shared file at the instant of evaluation; retry once
            # before concluding the deliverable is corrupt.
            time.sleep(0.5)
            loaded, err = load_workbook_safe(path)
            if loaded is None:
                record("Excel file readable", False, f"openpyxl error: {err[:200]}")
                return False
        wb, wb_cached = loaded

        # Leaderboard sheet
        lb_rows = load_sheet_rows(wb, wb_cached, "Leaderboard")
        if lb_rows is None:
            record("Sheet 'Leaderboard' exists", False, f"Sheets: {wb.sheetnames}")
            return False
        record("Sheet 'Leaderboard' exists", True)

        header = lb_rows[0] if lb_rows else []
        data = lb_rows[1:]
        # Content coverage (below) is the real gate; the count is a lower bound
        # so that a correct file with extra/duplicate rows is not penalised.
        record("Leaderboard has >= 5 rows", len(data) >= 5, f"Found {len(data)}")

        method_col = find_col(header, ["Method", "method"])

        if method_col is not None:
            found_methods = {
                str(r[method_col]).strip().lower()
                for r in data
                if method_col < len(r) and r[method_col] is not None and str(r[method_col]).strip()
            }
            for entry in LEADERBOARD:
                present = entry["method"].lower() in found_methods
                record(f"Method '{entry['method']}' present", present, f"Found: {found_methods}")

        # Method Details sheet
        md_rows = load_sheet_rows(wb, wb_cached, "Method Details")
        if md_rows is None:
            record("Sheet 'Method Details' exists", False, f"Sheets: {wb.sheetnames}")
        else:
            record("Sheet 'Method Details' exists", True)
            data2 = md_rows[1:]
            record("Method Details has >= 4 rows", len(data2) >= 4, f"Found {len(data2)}")

            id_col = find_col(md_rows[0], ["Paper_ID", "Paper ID", "paper_id"])
            if id_col is not None:
                found_ids = {
                    str(r[id_col]).strip()
                    for r in data2
                    if id_col < len(r) and r[id_col] is not None and str(r[id_col]).strip()
                }
                for pid in PAPERS_WITH_ID:
                    record(f"Paper {pid} in Method Details", pid in found_ids, f"Found: {found_ids}")

            # Key_Contribution and Dataset_Used should be non-empty for the
            # papers the leaderboard actually references (the 4 required papers).
            # Extra rows a model may have added (e.g. noise papers picked up from
            # the arxiv DB) must NOT trigger a FAIL -- only the required papers
            # gate the outcome. If no paper-id column exists, fall back to
            # requiring every row to be non-empty.
            kc_col = find_col(md_rows[0], ["Key_Contribution", "Key Contribution", "contribution"])
            ds_col = find_col(md_rows[0], ["Dataset_Used", "Dataset Used", "dataset"])

            def _required_rows(rows, id_col_idx):
                if id_col_idx is None:
                    return rows
                return [
                    r for r in rows
                    if id_col_idx < len(r) and str(r[id_col_idx]).strip() in PAPERS_WITH_ID
                ]

            def _nonempty_count(rows, col_idx):
                return sum(
                    1 for r in rows
                    if col_idx >= len(r) or not r[col_idx] or not str(r[col_idx]).strip()
                )

            if kc_col is not None:
                empty_kc = _nonempty_count(_required_rows(data2, id_col), kc_col)
                record("All Key_Contribution values non-empty", empty_kc == 0, f"{empty_kc} empty")
            if ds_col is not None:
                empty_ds = _nonempty_count(_required_rows(data2, id_col), ds_col)
                record("All Dataset_Used values non-empty", empty_ds == 0, f"{empty_ds} empty")

        # Summary sheet
        sum_rows = load_sheet_rows(wb, wb_cached, "Summary")
        if sum_rows is None:
            record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        else:
            record("Sheet 'Summary' exists", True)
            metrics = {}
            for row in sum_rows[1:]:
                if row and row[0] is not None and str(row[0]).strip():
                    metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

            tm_key = next((k for k in metrics if "total" in k and "method" in k), None)
            if tm_key:
                record("Total_Methods = 5", num_close(metrics[tm_key], 5, tol=0), f"Got {metrics[tm_key]}")

            mwp_key = next((k for k in metrics if "with" in k and "paper" in k), None)
            if mwp_key:
                record("Methods_With_Papers = 4", num_close(metrics[mwp_key], 4, tol=0), f"Got {metrics[mwp_key]}")

            tt_key = next((k for k in metrics if "total" in k and "task" in k), None)
            if tt_key:
                record("Total_Tasks = 3", num_close(metrics[tt_key], 3, tol=0), f"Got {metrics[tt_key]}")

            ts_key = next((k for k in metrics if "top" in k and "score" in k), None)
            if ts_key:
                # Exact 95.0 expected
                record("Top_Score = 95.0", num_close(metrics[ts_key], 95.0, tol=0), f"Got {metrics[ts_key]}")

        return True
    except Exception as e:
        # Crash-safety: never let a malformed workbook kill the whole evaluator.
        record("Excel checks (unexpected error)", False, str(e)[:300])
        return False


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # A page counts as the tracker if its title/properties mention the
        # benchmark tracker, or if a child block under it does (robust against
        # the title being stored in either place).
        cur.execute("""
            SELECT DISTINCT p.id, p.properties FROM notion.pages p
            WHERE p.properties::text ILIKE '%%benchmark%%'
               OR p.properties::text ILIKE '%%method%%tracker%%'
               OR p.properties::text ILIKE '%%leaderboard%%'
               OR EXISTS (
                   SELECT 1 FROM notion.blocks b
                   WHERE b.parent_id = p.id
                     AND b.block_data::text ILIKE '%%benchmark%%'
               )
        """)
        pages = cur.fetchall()

        if not pages:
            cur.execute("SELECT id, properties FROM notion.pages")
            all_pages = cur.fetchall()
            record("Notion page with benchmark/method content", False,
                   f"Found {len(all_pages)} pages but none matching")
            return False

        record("Notion page exists", True)

        page_ids = [p[0] for p in pages]
        cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = ANY(%s)", (page_ids,))
        count = cur.fetchone()[0]
        # Do NOT fall back to total block count -- must be under this page
        record("Notion page has content blocks", count >= 3, f"Found {count} blocks under target page")

        cur.execute("SELECT block_data FROM notion.blocks")
        blocks = cur.fetchall()
        text = " ".join(str(b[0]).lower() for b in blocks if b[0])
        props = " ".join(str(p[1]).lower() for p in pages if p[1])
        combined = text + " " + props

        has_content = any(kw in combined for kw in ["vit", "convnext", "llama", "diffusion", "benchmark", "leaderboard"])
        record("Notion mentions methods/benchmarks", has_content)

        cur.close()
        conn.close()
        return True
    except Exception as e:
        record("Notion accessible", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
