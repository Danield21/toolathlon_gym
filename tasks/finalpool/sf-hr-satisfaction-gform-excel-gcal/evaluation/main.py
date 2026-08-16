"""Evaluation script for sf-hr-satisfaction-gform-excel-gcal."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Check Satisfaction_Program_Report.xlsx
    excel_path = os.path.join(agent_workspace, "Satisfaction_Program_Report.xlsx")
    check("Satisfaction_Program_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Satisfaction_Program_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        if gt_wb:
            for sheet_name in gt_wb.sheetnames:
                check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    gt_ws = gt_wb[sheet_name]
                    # Check headers
                    gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                    for h in gt_headers:
                        if h:
                            check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                    # Check row count
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                    min_rows = max(1, len(gt_rows) - 2)
                    check(f"{sheet_name} has >= {min_rows} data rows", len(data_rows) >= min_rows, f"got {len(data_rows)}")

                    # Cell value comparison against groundtruth
                    header_map = {h: i for i, h in enumerate(headers)}
                    gt_header_map = {h: i for i, h in enumerate(gt_headers)}
                    for ri in range(min(3, len(gt_rows), len(data_rows))):
                        gt_row = gt_rows[ri]
                        agent_row = data_rows[ri]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row):
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                tol = max(0.5, abs(gf) * 0.15)
                                check(f"{sheet_name} R{ri+2} {gt_h} ~{gf:.1f}",
                                      abs(gf - af) <= tol, f"got {af}")
                            elif gv is not None and av is not None:
                                gs = str(gv).strip().lower()
                                avs = str(av).strip().lower()
                                if gs:
                                    check(f"{sheet_name} R{ri+2} {gt_h} text",
                                          gs == avs or gs in avs or avs in gs,
                                          f"expected {gs[:50]}, got {avs[:50]}")

    # Check Python script exists - exact name per task: satisfaction_analyzer.py
    analyzer_path = os.path.join(agent_workspace, "satisfaction_analyzer.py")
    check("satisfaction_analyzer.py exists",
          os.path.exists(analyzer_path),
          f"missing satisfaction_analyzer.py")

    # Check satisfaction_analysis.json (output of script)
    analysis_json = os.path.join(agent_workspace, "satisfaction_analysis.json")
    if os.path.exists(analysis_json):
        check("satisfaction_analysis.json exists", True)
    else:
        check("satisfaction_analysis.json exists", False, "missing analysis output")

    # Database checks
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Calendar event title must contain 'wellness' AND 'kickoff' or be 'Wellness Program Kickoff'
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE LOWER(summary) LIKE '%wellness%' AND LOWER(summary) LIKE '%kickoff%'
        """)
        event_row = cur.fetchone()
        check("Calendar event 'Wellness Program Kickoff' exists",
              event_row is not None, "no matching event found")
        if event_row is not None:
            summary, desc, start_dt, end_dt = event_row
            # Required: 2026-03-15, 2:00 PM to 3:30 PM UTC = 14:00-15:30 UTC
            from datetime import datetime, timezone
            try:
                if hasattr(start_dt, 'date'):
                    start_utc = start_dt.astimezone(timezone.utc) if start_dt.tzinfo else start_dt.replace(tzinfo=timezone.utc)
                    end_utc = end_dt.astimezone(timezone.utc) if end_dt.tzinfo else end_dt.replace(tzinfo=timezone.utc)
                    expected_start = datetime(2026, 3, 15, 14, 0, tzinfo=timezone.utc)
                    expected_end = datetime(2026, 3, 15, 15, 30, tzinfo=timezone.utc)
                    check("Event starts 2026-03-15 14:00 UTC",
                          start_utc == expected_start,
                          f"got {start_utc}")
                    check("Event ends 2026-03-15 15:30 UTC",
                          end_utc == expected_end,
                          f"got {end_utc}")
                else:
                    check("Event datetime parseable", False, f"start={start_dt}")
            except Exception as e:
                check("Event datetime check", False, str(e))
            # Description should mention departments and initiatives
            desc_l = (desc or "").lower()
            has_dept = any(k in desc_l for k in ["department", "engineering", "finance", "hr", "operations", "sales"])
            has_init = any(k in desc_l for k in ["initiative", "program", "wellness"])
            check("Event description mentions target departments + initiatives",
                  has_dept and has_init, "missing dept/initiative terms")

        cur.execute("SELECT COUNT(*) FROM gform.forms WHERE LOWER(title) LIKE '%wellness check-in%' OR LOWER(title) LIKE '%employee wellness%'")
        form_count = cur.fetchone()[0]
        check("Google Form 'Employee Wellness Check-In' exists",
              form_count >= 1, f"form count: {form_count}")
        cur.execute("SELECT COUNT(*) FROM gform.questions")
        q_count = cur.fetchone()[0]
        # Task asks 5 questions
        check("Form has at least 5 questions",
              q_count >= 5, f"question count: {q_count}")
        # Reverse verification: noise events should not match task keyword
        cur.execute("SELECT COUNT(*) FROM gcal.events WHERE summary ILIKE '%standup%' OR summary ILIKE '%lunch%'")
        noise_events = cur.fetchone()[0]
        check("Noise events exist (not deleted by agent)",
              noise_events >= 1, f"noise events: {noise_events}")
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()