"""Evaluation for terminal-yf-fetch-gsheet-excel-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def fetch_yf_expectations():
    """Fetch latest close, ~30-day-prior close, name, sector for each ticker."""
    out = {}
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        for sym in ["AMZN", "GOOGL", "JPM"]:
            cur.execute("SELECT close FROM yf.stock_prices WHERE symbol=%s ORDER BY date DESC LIMIT 1", (sym,))
            row = cur.fetchone()
            latest = float(row[0]) if row and row[0] is not None else None
            cur.execute("""
                SELECT close FROM yf.stock_prices WHERE symbol=%s
                ORDER BY date DESC OFFSET 20 LIMIT 1
            """, (sym,))
            row = cur.fetchone()
            past = float(row[0]) if row and row[0] is not None else None
            cur.execute("SELECT data FROM yf.stock_info WHERE symbol=%s", (sym,))
            row = cur.fetchone()
            name = ""
            sector = ""
            if row and row[0]:
                d = row[0] if isinstance(row[0], dict) else json.loads(row[0])
                name = d.get("shortName") or d.get("longName") or d.get("name") or ""
                sector = d.get("sector") or ""
            out[sym] = {"latest": latest, "past": past, "name": name, "sector": sector}
            if latest is not None and past is not None and past != 0:
                out[sym]["return_30d_pct"] = round((latest - past) / past * 100, 2)
            else:
                out[sym]["return_30d_pct"] = None
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[fallback] YF fetch failed: {e}")
    return out


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def check_excel(agent_ws, gt_dir):
    print("\n=== Checking Market_Analysis_Report.xlsx ===")
    agent_file = os.path.join(agent_ws, "Market_Analysis_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Market_Analysis_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Stock_Overview
    print("  Checking Stock_Overview...")
    ws1 = get_sheet(awb, "Stock_Overview")
    check("Sheet Stock_Overview exists", ws1 is not None, f"Sheets: {awb.sheetnames}")
    yf_exp = fetch_yf_expectations()
    if ws1:
        rows = list(ws1.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in rows if r and r[0]]
        check("Stock_Overview has 3 rows", len(data_rows) == 3, f"Got {len(data_rows)}")

        # Sorted alphabetically by Symbol
        try:
            syms = [str(r[0]).strip().upper() for r in data_rows]
            sorted_ok = syms == sorted(syms)
        except Exception:
            sorted_ok = False
        check("Stock_Overview sorted alphabetically by Symbol",
              sorted_ok, f"got {syms if sorted_ok else 'unparseable'}")

        symbols = {str(r[0]).strip().upper() for r in data_rows if r and r[0]}
        for sym in ["AMZN", "GOOGL", "JPM"]:
            check(f"Contains {sym}", sym in symbols, f"Symbols: {symbols}")

        for r in data_rows:
            if r and len(r) >= 5 and r[0]:
                sym = str(r[0]).strip().upper()
                exp = yf_exp.get(sym, {})
                # Name
                actual_name = (str(r[1]) if r[1] else "").strip()
                exp_name = exp.get("name", "")
                if exp_name:
                    matches = exp_name.lower() in actual_name.lower() or actual_name.lower() in exp_name.lower()
                    check(f"{sym} Name matches expected '{exp_name}'", matches,
                          f"got '{actual_name}'")
                # Sector
                actual_sector = (str(r[2]) if r[2] else "").strip()
                exp_sector = exp.get("sector", "")
                if exp_sector:
                    sec_match = exp_sector.lower() in actual_sector.lower() or actual_sector.lower() in exp_sector.lower()
                    check(f"{sym} Sector matches '{exp_sector}'", sec_match,
                          f"got '{actual_sector}'")
                # Latest_Price
                if exp.get("latest") is not None:
                    check(f"{sym} Latest_Price matches DB", num_close(r[3], exp["latest"], 1.0),
                          f"got {r[3]}, expected {exp['latest']}")
                # Return_30d_Pct (tol=1.0 — DB-derived expected, deterministic OFFSET 20 trading day approx)
                if exp.get("return_30d_pct") is not None:
                    check(f"{sym} Return_30d_Pct close to {exp['return_30d_pct']}",
                          num_close(r[4], exp["return_30d_pct"], 1.0),
                          f"got {r[4]}, expected ~{exp['return_30d_pct']}")

    # Sheet 2: Economic_Indicators
    print("  Checking Economic_Indicators...")
    ws2 = get_sheet(awb, "Economic_Indicators")
    check("Sheet Economic_Indicators exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    if ws2:
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        data_rows2 = [r for r in rows2 if r and r[0]]
        check("Economic_Indicators has 5 rows", len(data_rows2) == 5, f"Got {len(data_rows2)}")

        # Check specific indicator values from mock data
        ind_lookup = {}
        for r in data_rows2:
            if r and r[0]:
                ind_lookup[str(r[0]).strip().lower()] = r

        gdp = ind_lookup.get("gdp growth rate") or ind_lookup.get("gdp growth")
        if gdp:
            check("GDP Growth = 2.1", num_close(gdp[1], 2.1, 0.2), f"Got {gdp[1]}")
            check("GDP Trend is Favorable",
                  gdp[2] and "favorable" in str(gdp[2]).lower(),
                  f"Got {gdp[2]}")

        infl = ind_lookup.get("inflation rate") or ind_lookup.get("inflation")
        if infl:
            check("Inflation = 3.4", num_close(infl[1], 3.4, 0.2), f"Got {infl[1]}")
            check("Inflation Trend is Unfavorable",
                  infl[2] and "unfavorable" in str(infl[2]).lower(),
                  f"Got {infl[2]}")

    # Sheet 3: Correlation_Matrix
    print("  Checking Correlation_Matrix...")
    ws3 = get_sheet(awb, "Correlation_Matrix")
    check("Sheet Correlation_Matrix exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        data_rows3 = [r for r in rows3 if r and r[0]]
        check("Correlation_Matrix has 3 rows", len(data_rows3) == 3, f"Got {len(data_rows3)}")

        valid_values = {"positive", "negative", "neutral"}
        for r in data_rows3:
            if r and len(r) >= 4:
                for col_idx in [1, 2, 3]:
                    val = str(r[col_idx]).strip().lower() if r[col_idx] else ""
                    if val not in valid_values:
                        check(f"{r[0]} correlation values valid", False, f"Invalid: {r[col_idx]}")
                        break
                else:
                    continue
                break
        else:
            check("All correlation values are valid", True)

    # Sheet 4: Portfolio_Signals
    print("  Checking Portfolio_Signals...")
    ws4 = get_sheet(awb, "Portfolio_Signals")
    check("Sheet Portfolio_Signals exists", ws4 is not None, f"Sheets: {awb.sheetnames}")
    if ws4:
        rows4 = list(ws4.iter_rows(min_row=2, values_only=True))
        data_rows4 = [r for r in rows4 if r and r[0]]
        check("Portfolio_Signals has 3 rows", len(data_rows4) == 3, f"Got {len(data_rows4)}")

        # Compute expected signals using the rules
        # mock data: consumer confidence and inflation come from API; using expected values
        # GDP=2.1, inflation=3.4 (mock dashboard); consumer confidence ~ 95+ assumed default
        # We accept signals consistent with rules:
        # - 30d return > 0 AND consumer_conf > 95 -> BUY
        # - 30d return < 0 AND inflation > 3 -> SELL
        # - else -> HOLD
        # Inflation IS > 3 (=3.4), so any negative return -> SELL
        # If consumer confidence > 95 then any positive return -> BUY else HOLD
        valid_signals = {"buy", "sell", "hold"}
        sig_lookup = {}
        for r in data_rows4:
            if r and r[0]:
                sym = str(r[0]).strip().upper()
                sig_lookup[sym] = (str(r[1] or "").strip().lower(),
                                    str(r[2] or "").strip())
        for sym in ["AMZN", "GOOGL", "JPM"]:
            v = sig_lookup.get(sym)
            check(f"{sym} portfolio signal row present", v is not None)
            if v is None:
                continue
            sig, rat = v
            check(f"{sym} signal value valid (BUY/SELL/HOLD)", sig in valid_signals,
                  f"Invalid: {sig}")
            check(f"{sym} rationale present", len(rat) > 0, "Missing rationale")
            # Validate signal logic against the 30-day return rule
            exp_ret = yf_exp.get(sym, {}).get("return_30d_pct")
            if exp_ret is not None:
                # Inflation 3.4 > 3, so negative return -> SELL
                if exp_ret < 0:
                    check(f"{sym} signal is SELL (negative 30d return + infl>3)",
                          sig == "sell", f"got '{sig}', expected 'sell'")
                # If positive return, signal must be BUY or HOLD (not SELL)
                if exp_ret > 0:
                    check(f"{sym} signal is BUY or HOLD (positive 30d return)",
                          sig in ("buy", "hold"),
                          f"got '{sig}', expected 'buy' or 'hold'")


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        rows = cur.fetchall()
        target = None
        for sid, title in rows:
            if (title or "").strip().lower() == "market analysis live data":
                target = sid
                break
        check("Google Sheet titled exactly 'Market Analysis Live Data' exists",
              target is not None, f"Found titles: {[r[1] for r in rows]}")
        if target is not None:
            cur.execute("""
                SELECT c.value
                FROM gsheet.cells c
                JOIN gsheet.sheets s ON s.id = c.sheet_id
                WHERE s.spreadsheet_id = %s
            """, (target,))
            cells = [str(c[0]) for c in cur.fetchall() if c and c[0] is not None]
            cells_blob = " ".join(cells).upper()
            for sym in ["AMZN", "GOOGL", "JPM"]:
                check(f"Sheet contains {sym}", sym in cells_blob,
                      f"Cells: {cells_blob[:200]}")
        cur.close()
        conn.close()
    except Exception as e:
        check("GSheet check", False, str(e))


def check_notion():
    print("\n=== Checking Notion Database ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
        dbs = cur.fetchall()
        found_db = None
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "investment research log" in title_str:
                found_db = db_id
                break
        check("Notion 'Investment Research Log' database exists",
              found_db is not None, f"Found {len(dbs)} dbs")

        if found_db:
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE parent->>'database_id' = %s AND archived = false
            """, (found_db,))
            pages = cur.fetchall()
            check("Notion DB has exactly 3 stock pages", len(pages) == 3,
                  f"Got {len(pages)}")

            # Each page should reference a stock symbol; aggregate text
            page_blob = json.dumps([p[1] for p in pages]).upper()
            for sym in ["AMZN", "GOOGL", "JPM"]:
                check(f"Notion DB has page for {sym}",
                      sym in page_blob,
                      f"Pages do not contain symbol {sym}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Notion check", False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Market_Analysis_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets
        expected_keywords = {"stock", "overview", "economic", "indicator", "correlation",
                             "matrix", "portfolio", "signal"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        # Stock prices should not be negative
        ws1 = get_sheet(wb, "Stock_Overview")
        if ws1:
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 4 and row[3] is not None:
                    try:
                        price = float(row[3])
                        if price < 0:
                            check("No negative stock prices", False,
                                  f"Found {price} for {row[0]}")
                            break
                    except (ValueError, TypeError):
                        pass
            else:
                check("No negative stock prices", True)

    # Notion: no duplicate Investment Research Log databases
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM notion.databases WHERE archived = false")
        dbs = cur.fetchall()
        invest_dbs = []
        for db_id, title in dbs:
            title_str = json.dumps(title).lower() if title else ""
            if "investment" in title_str and "research" in title_str:
                invest_dbs.append(db_id)
        check("No duplicate Investment Research databases", len(invest_dbs) <= 1,
              f"Found {len(invest_dbs)} matching databases")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gsheet()
    check_notion()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
