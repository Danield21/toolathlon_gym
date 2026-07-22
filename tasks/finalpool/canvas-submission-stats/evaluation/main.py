"""Evaluation for canvas-submission-stats."""
import argparse
import json
import os
import sys
import openpyxl

try:
    import psycopg2
    DB = {
        "host": os.environ.get("PGHOST", "localhost"),
        "port": int(os.environ.get("PGPORT", "5432")),
        "dbname": "toolathlon_gym",
        "user": "eigent",
        "password": "camel",
    }
except Exception:
    psycopg2 = None
    DB = None


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def num_close_rel(a, b, rel=0.02, abs_tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_page_title(properties):
    """Extract title text from notion.pages.properties."""
    if not properties:
        return ""
    if isinstance(properties, str):
        try:
            properties = json.loads(properties)
        except Exception:
            return ""
    if not isinstance(properties, dict):
        return ""
    for key, val in properties.items():
        if not isinstance(val, dict):
            continue
        if val.get("type") == "title":
            title_arr = val.get("title", [])
            return "".join((t.get("plain_text") or "") for t in title_arr if isinstance(t, dict))
    return ""


def _fetch_expected_summary_from_gt(gt_dir):
    """Read GT Excel summary values; falls back to None if missing."""
    try:
        gt_file = os.path.join(gt_dir, "Canvas_Submissions.xlsx")
        if not os.path.exists(gt_file):
            return None
        wb = openpyxl.load_workbook(gt_file, data_only=True)
        rows = load_sheet_rows(wb, "Summary")
        wb.close()
        if not rows:
            return None
        out = {}
        for r in rows[1:]:
            if not r or r[0] is None:
                continue
            metric = str(r[0]).strip().lower()
            try:
                out[metric] = float(r[1]) if r[1] is not None else None
            except Exception:
                out[metric] = r[1]
        return out
    except Exception:
        return None


def _format_variants(label_keys, value):
    """Generate plausible string variants of a numeric value (with/without commas, 1-2 decimals)."""
    variants = list(label_keys)
    if value is None:
        return variants
    try:
        f = float(value)
    except Exception:
        return variants
    is_int = abs(f - round(f)) < 1e-6
    if is_int:
        i = int(round(f))
        variants.append(str(i))
        variants.append(f"{i:,}")  # 173,739
    else:
        variants.append(f"{f:.1f}")
        variants.append(f"{f:.2f}")
        variants.append(str(f))
    return variants


def check_notion(errors_list, gt_dir):
    print("  Checking Notion 'Canvas Submission Analysis'...")
    if psycopg2 is None or DB is None:
        errors_list.append("psycopg2 unavailable; cannot verify Notion")
        return
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false AND in_trash = false")
        rows = cur.fetchall()
        target_id = None
        for pid, props in rows:
            title = get_page_title(props)
            if title and "canvas submission analysis" in title.strip().lower():
                target_id = pid
                break
        if target_id is None:
            errors_list.append(
                f"Notion page 'Canvas Submission Analysis' not found (saw {len(rows)} pages)"
            )
            cur.close(); conn.close()
            return

        # Verify the page has at least some block content with summary metrics
        cur.execute(
            "SELECT block_data FROM notion.blocks WHERE parent_id = %s AND archived = false",
            (target_id,)
        )
        blocks = cur.fetchall()
        text_content = ""
        for (bd,) in blocks:
            if bd is None:
                continue
            if isinstance(bd, str):
                try:
                    bd = json.loads(bd)
                except Exception:
                    continue
            text_content += json.dumps(bd) + " "

        text_lower = text_content.lower()

        # Compute expected values live from GT (avoid hardcoded brittle constants)
        gt_summary = _fetch_expected_summary_from_gt(gt_dir) or {}
        total_subs = gt_summary.get("total_submissions")
        overall_avg = gt_summary.get("overall_avg_score")
        total_late = gt_summary.get("total_late")

        # Each metric: pass if either label or any plausible numeric format appears.
        ts_keys = _format_variants(("total_submissions", "total submissions"), total_subs)
        if not any(k.lower() in text_lower for k in ts_keys):
            errors_list.append(
                f"Notion page missing total submissions metric reference (looked for: {ts_keys[:5]})"
            )
        oa_keys = _format_variants(("overall_avg_score", "overall avg", "average score"), overall_avg)
        if not any(k.lower() in text_lower for k in oa_keys):
            errors_list.append(
                f"Notion page missing overall avg score reference (looked for: {oa_keys[:5]})"
            )
        tl_keys = _format_variants(("total_late", "total late"), total_late)
        if not any(k.lower() in text_lower for k in tl_keys):
            errors_list.append(
                f"Notion page missing total late reference (looked for: {tl_keys[:5]})"
            )
        cur.close(); conn.close()
        print(f"    Found Notion page with {len(blocks)} blocks")
    except Exception as e:
        errors_list.append(f"Notion check raised: {e}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Submissions.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Submissions.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Submission Stats
    print(f"  Checking Submission Stats...")
    a_rows = load_sheet_rows(agent_wb, "Submission Stats")
    g_rows = load_sheet_rows(gt_wb, "Submission Stats")
    if a_rows is None:
        all_errors.append("Sheet 'Submission Stats' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Submission Stats' not found in groundtruth")
    else:
        sheet_name = "Submission Stats"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close_rel(a_row[1], g_row[1], rel=0.02, abs_tol=2):
                    errors.append(f"{key}.Submissions: {a_row[1]} vs {g_row[1]} (rel 2%)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0.5):
                    errors.append(f"{key}.Avg_Score: {a_row[2]} vs {g_row[2]} (tol=0.5)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close_rel(a_row[3], g_row[3], rel=0.02, abs_tol=2):
                    errors.append(f"{key}.Late_Count: {a_row[3]} vs {g_row[3]} (rel 2%)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Late_Pct: {a_row[4]} vs {g_row[4]} (tol=0.5)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close_rel(a_row[1], g_row[1], rel=0.02, abs_tol=2.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (rel 2%)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")



    # Notion check
    check_notion(all_errors, gt_dir)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
