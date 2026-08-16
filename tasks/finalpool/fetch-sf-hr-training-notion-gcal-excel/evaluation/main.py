"""Evaluation for fetch-sf-hr-training-notion-gcal-excel."""
import argparse
import os
import sys

import psycopg2

# ──────────────────────────────────────────────────────────────────────────
# EVALUATION GROUND TRUTH SPEC (gcal tz root-fix v3, case-study 2026-08-13)
# gcal.events.start_datetime is TIMESTAMPTZ; bare start_dt.hour silently
# compares wrong in non-UTC PG sessions. Use gcal_helpers.
# ──────────────────────────────────────────────────────────────────────────
# task.md: SF company L&D program → America/Los_Angeles (PT)
EXPECTED_TIMEZONE = "America/Los_Angeles"

from utils.evaluation.gcal_helpers import get_zone_components  # noqa: E402

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

ALL_DEPARTMENTS = ["R&D", "Finance", "Operations", "Sales", "HR", "Support", "Engineering"]
PRIORITY_DEPTS = ["Operations", "Sales", "R&D"]  # 3 lowest perf in GT (avg 3.18, 3.19, 3.20)
TRAINING_DATES = ["2026-05-04", "2026-05-11", "2026-05-18"]

# Acceptable courses per dept (catalog skill_area-aligned matches; agent free to
# pick any of these as a "relevant course" without exact GT match enforcement).
ACCEPTABLE_COURSES = {
    "operations": {"advanced operations management", "project management professional"},
    "sales": {"sales excellence program", "communication skills workshop"},
    "r&d": {"r&d innovation workshop", "data analytics for business"},
    "finance": {"financial analysis fundamentals"},
    "hr": {"hr business partner certification"},
    "support": {"technical support mastery", "communication skills workshop"},
    "engineering": {"software engineering best practices"},
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    agent_file = os.path.join(agent_ws, "Training_Budget_Plan.xlsx")
    gt_file = os.path.join(gt_dir, "Training_Budget_Plan.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        all_errors.append("Training_Budget_Plan.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # ---------- Department Analysis sheet (full column validation) ----------
        print("  Checking Department Analysis sheet...")
        a_rows = load_sheet_rows(agent_wb, "Department Analysis")
        g_rows = load_sheet_rows(gt_wb, "Department Analysis")
        if a_rows is None:
            all_errors.append("Sheet 'Department Analysis' not found")
        else:
            # Validate header column count
            if not a_rows or len(a_rows[0]) < 9:
                all_errors.append(
                    f"Department Analysis header has only {len(a_rows[0]) if a_rows else 0} columns, expected >=9"
                )
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            non_empty = [r for r in a_data if r and r[0]]
            if len(non_empty) != 7:
                all_errors.append(
                    f"Department Analysis: {len(non_empty)} dept rows, expected 7"
                )
            else:
                a_lookup = {str(r[0]).strip().lower(): r for r in non_empty}
                for g_row in g_rows[1:]:
                    if not g_row or not g_row[0]:
                        continue
                    key = str(g_row[0]).strip().lower()
                    a_row = a_lookup.get(key)
                    if a_row is None:
                        all_errors.append(f"Missing department: {g_row[0]}")
                        continue
                    # Pad agent row if shorter
                    a_row_padded = list(a_row) + [None] * (9 - len(a_row))
                    # Total_Employees (tol=5)
                    if not num_close(a_row_padded[1], g_row[1], 5):
                        all_errors.append(
                            f"{g_row[0]} Total_Employees: {a_row_padded[1]} vs {g_row[1]}"
                        )
                    # Avg_Performance (tol=0.05)
                    if not num_close(a_row_padded[2], g_row[2], 0.05):
                        all_errors.append(
                            f"{g_row[0]} Avg_Performance: {a_row_padded[2]} vs {g_row[2]}"
                        )
                    # High_School / Diploma / Bachelors / Masters / PhD counts (tol=2 — deterministic SF data)
                    for idx, label in enumerate(
                        ["High_School_Count", "Diploma_Count", "Bachelors_Count", "Masters_Count", "PhD_Count"],
                        start=3,
                    ):
                        if not num_close(a_row_padded[idx], g_row[idx], 2):
                            all_errors.append(
                                f"{g_row[0]} {label}: {a_row_padded[idx]} vs {g_row[idx]}"
                            )
                    # Low_Education_Pct (tol=0.5)
                    if not num_close(a_row_padded[8], g_row[8], 0.5):
                        all_errors.append(
                            f"{g_row[0]} Low_Education_Pct: {a_row_padded[8]} vs {g_row[8]}"
                        )
            print("    Done.")

        # ---------- Training Catalog sheet ----------
        print("  Checking Training Catalog sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Training Catalog")
        g_rows2 = load_sheet_rows(gt_wb, "Training Catalog")
        if a_rows2 is None:
            all_errors.append("Sheet 'Training Catalog' not found")
        else:
            a_data2 = [r for r in a_rows2[1:] if r and r[0]]
            if len(a_data2) != 12:
                all_errors.append(f"Training Catalog: {len(a_data2)} courses, expected 12")
            # Required column count >=5
            if a_rows2 and len(a_rows2[0]) < 5:
                all_errors.append(
                    f"Training Catalog header has only {len(a_rows2[0])} columns, expected >=5"
                )
            # Verify each course name in GT appears in agent's catalog
            gt_courses = {str(r[0]).strip().lower() for r in g_rows2[1:] if r and r[0]}
            agent_courses = {str(r[0]).strip().lower() for r in a_data2 if r and r[0]}
            missing = gt_courses - agent_courses
            if missing:
                all_errors.append(
                    f"Training Catalog missing courses: {sorted(missing)[:5]}"
                )
            print("    Done.")

        # ---------- Budget Allocation sheet ----------
        print("  Checking Budget Allocation sheet...")
        a_rows3 = load_sheet_rows(agent_wb, "Budget Allocation")
        g_rows3 = load_sheet_rows(gt_wb, "Budget Allocation")
        if a_rows3 is None:
            all_errors.append("Sheet 'Budget Allocation' not found")
        else:
            # Pull non-empty department rows (skip Grand Total row)
            a_dept_rows = [
                r for r in a_rows3[1:]
                if r and r[0] and str(r[0]).strip()
                and not str(r[0]).strip().lower().startswith("grand")
            ]
            if len(a_dept_rows) != 3:
                all_errors.append(
                    f"Budget Allocation: {len(a_dept_rows)} dept rows, expected 3"
                )
            else:
                dept_names_lower = [str(r[0]).strip().lower() for r in a_dept_rows]
                for d in PRIORITY_DEPTS:
                    if d.lower() not in dept_names_lower:
                        all_errors.append(
                            f"Budget Allocation missing priority dept: {d}"
                        )
                # Build GT lookup
                g_data3 = [r for r in g_rows3[1:] if r and r[0] and not str(r[0]).strip().lower().startswith("grand")]
                g_lookup3 = {str(r[0]).strip().lower(): r for r in g_data3}
                a_lookup3 = {str(r[0]).strip().lower(): r for r in a_dept_rows}
                # Validate Trainees / Sessions_Needed / Cost_Per_Session / Total_Cost / Priority
                for d in PRIORITY_DEPTS:
                    g_row = g_lookup3.get(d.lower())
                    a_row = a_lookup3.get(d.lower())
                    if g_row is None or a_row is None:
                        continue
                    a_row_padded = list(a_row) + [None] * (7 - len(a_row))
                    # Assigned_Course: accept GT exact match OR any catalog course
                    # whose skill_area aligns with the dept (per ACCEPTABLE_COURSES).
                    # Task says "most relevant course" — multiple courses can be valid.
                    g_course = str(g_row[1] or "").strip().lower()
                    a_course = str(a_row_padded[1] or "").strip().lower()
                    acceptable = ACCEPTABLE_COURSES.get(d.strip().lower(), set())
                    course_ok = (a_course == g_course) or (a_course in acceptable)
                    if g_course and not course_ok:
                        all_errors.append(
                            f"{d} Assigned_Course: '{a_row_padded[1]}' (GT '{g_row[1]}'; "
                            f"or any of {sorted(acceptable)})"
                        )
                    # Trainees (tol=5 — fully deterministic SF count of HS+Diploma per dept)
                    if not num_close(a_row_padded[2], g_row[2], 5):
                        all_errors.append(
                            f"{d} Trainees: {a_row_padded[2]} vs {g_row[2]}"
                        )
                    # Sessions_Needed (tol=2)
                    if not num_close(a_row_padded[3], g_row[3], 2):
                        all_errors.append(
                            f"{d} Sessions_Needed: {a_row_padded[3]} vs {g_row[3]}"
                        )
                    # Cost_Per_Session (tol=10)
                    if not num_close(a_row_padded[4], g_row[4], 10):
                        all_errors.append(
                            f"{d} Cost_Per_Session: {a_row_padded[4]} vs {g_row[4]}"
                        )
                    # Total_Cost (tol=1000 — should be Sessions x Cost_Per_Session, deterministic)
                    if not num_close(a_row_padded[5], g_row[5], 1000):
                        all_errors.append(
                            f"{d} Total_Cost: {a_row_padded[5]} vs {g_row[5]}"
                        )
                    # Training_Priority (string equality, case-insensitive)
                    g_pri = str(g_row[6] or "").strip().lower()
                    a_pri = str(a_row_padded[6] or "").strip().lower()
                    if g_pri and a_pri != g_pri:
                        all_errors.append(
                            f"{d} Training_Priority: '{a_row_padded[6]}' vs '{g_row[6]}'"
                        )
            print("    Done.")

    # --- Check 2: Notion database ---
    print("Checking Notion database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM notion.databases
            WHERE title::text ILIKE '%training%program%2026%'
               OR title::text ILIKE '%training program 2026%'
        """)
        dbs = cur.fetchall()
        if not dbs:
            all_errors.append("Notion database 'Training Program 2026' not found")
        else:
            db_id = dbs[0][0]
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE parent::jsonb->>'database_id' = %s
                  AND archived = false
            """, (db_id,))
            pages = cur.fetchall()
            if len(pages) != 3:
                all_errors.append(
                    f"Notion: found {len(pages)} pages in 'Training Program 2026', expected exactly 3"
                )
            # Validate each priority dept appears in title with required priority
            # Parse properties as proper JSON: extract title plain_text and select.name values
            def _extract_property_values(props):
                """Return lowercased list of strings: title plain_texts and select option names."""
                out = []
                if not isinstance(props, dict):
                    return out
                for prop_name, prop_val in props.items():
                    if not isinstance(prop_val, dict):
                        continue
                    ptype = prop_val.get("type")
                    if ptype == "title":
                        for t in prop_val.get("title", []) or []:
                            txt = t.get("plain_text") if isinstance(t, dict) else None
                            if txt:
                                out.append(str(txt).lower())
                    elif ptype == "rich_text":
                        for t in prop_val.get("rich_text", []) or []:
                            txt = t.get("plain_text") if isinstance(t, dict) else None
                            if txt:
                                out.append(str(txt).lower())
                    elif ptype == "select":
                        sel = prop_val.get("select")
                        if isinstance(sel, dict) and sel.get("name"):
                            out.append(str(sel["name"]).lower())
                    elif ptype == "multi_select":
                        for s in prop_val.get("multi_select", []) or []:
                            if isinstance(s, dict) and s.get("name"):
                                out.append(str(s["name"]).lower())
                return out

            seen_depts = set()
            seen_priorities = set()
            for page_id, props in pages:
                values = _extract_property_values(props)
                # Combine for substring search of department names (which can be 'R&D' etc.)
                joined = " ".join(values)
                for dept in PRIORITY_DEPTS:
                    if dept.lower() in joined:
                        seen_depts.add(dept.lower())
                # Priorities must come from select / multi_select option names exactly
                for v in values:
                    v_strip = v.strip()
                    if v_strip in ("high", "medium", "low"):
                        seen_priorities.add(v_strip)
            for d in PRIORITY_DEPTS:
                if d.lower() not in seen_depts:
                    all_errors.append(f"Notion: priority dept '{d}' not found in any page title/props")
            for pri in ["high", "medium", "low"]:
                if pri not in seen_priorities:
                    all_errors.append(f"Notion: priority value '{pri}' not present across pages")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking Notion: {e}")

    # --- Check 3: GCal events ---
    print("Checking GCal events...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE summary ILIKE '%training%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        # Filter for training events with priority dept name in summary
        priority_events = []
        for summary, description, start_dt, end_dt in rows:
            summary_l = (summary or "").lower()
            if any(d.lower() in summary_l for d in PRIORITY_DEPTS) and "training" in summary_l:
                priority_events.append((summary, description, start_dt, end_dt))
        if len(priority_events) != 3:
            all_errors.append(
                f"GCal: found {len(priority_events)} priority-department training events, expected 3"
            )
        # Date coverage (any of priority events on each expected date)
        # gcal.events.start_datetime is TIMESTAMPTZ; convert to EXPECTED_TIMEZONE
        # via session-tz-independent helper. Bare start_dt.date() can shift a day.
        dates_seen = {
            get_zone_components(start_dt, EXPECTED_TIMEZONE)[0].isoformat()
            for _, _, start_dt, _ in priority_events
            if start_dt is not None
            and get_zone_components(start_dt, EXPECTED_TIMEZONE)[0] is not None
        }
        for d in TRAINING_DATES:
            if d not in dates_seen:
                all_errors.append(f"GCal: no priority training event on {d}")
        # Each priority event 9:00-17:00 (8-hour duration)
        from datetime import timedelta
        for summary, description, start_dt, end_dt in priority_events:
            if start_dt is None or end_dt is None:
                all_errors.append(f"GCal event '{summary}' missing start/end timestamps")
                continue
            _sd, s_h, s_m = get_zone_components(start_dt, EXPECTED_TIMEZONE)
            _ed, e_h, e_m = get_zone_components(end_dt, EXPECTED_TIMEZONE)
            if s_h != 9 or s_m != 0:
                all_errors.append(f"GCal event '{summary}' start hour={s_h}:{s_m}, expected 09:00")
            if e_h != 17 or e_m != 0:
                all_errors.append(f"GCal event '{summary}' end hour={e_h}:{e_m}, expected 17:00")
            duration = end_dt - start_dt
            if duration != timedelta(hours=8):
                all_errors.append(f"GCal event '{summary}' duration={duration}, expected 8h")
        # Event summary should follow '[Department] Training: [Course Name]' pattern
        # check that 'training:' appears
        for summary, description, _, _ in priority_events:
            if "training:" not in (summary or "").lower():
                all_errors.append(f"GCal event summary missing 'Training:' separator: {summary}")
    except Exception as e:
        all_errors.append(f"Error checking GCal: {e}")

    # --- Check 4: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%department-managers@company.com%'
              AND subject ILIKE '%training%'
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            all_errors.append("No email sent to department-managers@company.com about training")
        else:
            # Inspect the first matching email
            subject, from_addr, to_addr, body_text = rows[0]
            # From address must be ld-director@company.com
            if "ld-director@company.com" not in (from_addr or "").lower():
                all_errors.append(f"Email from_addr is '{from_addr}', expected ld-director@company.com")
            # Subject must include '2026' and 'Department Assignments'
            subj_l = (subject or "").lower()
            if "2026" not in subj_l or "department assignments" not in subj_l:
                all_errors.append(
                    f"Email subject '{subject}' missing '2026' or 'Department Assignments'"
                )
            # Body must mention each priority dept name
            body_l = (body_text or "").lower()
            for d in PRIORITY_DEPTS:
                if d.lower() not in body_l:
                    all_errors.append(f"Email body missing priority dept '{d}'")
            # Body must mention training budget number (some 7-digit total)
            if not any(c.isdigit() for c in body_l):
                all_errors.append("Email body has no numeric budget figure")
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:20]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
