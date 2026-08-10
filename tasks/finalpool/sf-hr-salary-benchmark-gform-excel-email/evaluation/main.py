"""Evaluation for sf-hr-salary-benchmark-gform-excel-email.

Checks:
1. Salary_Analysis.xlsx has correct sheets and data
2. Google Forms survey "Compensation Satisfaction Survey" with 5 questions
3. Email to hr-leadership@company.example.com with subject matching pattern
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _to_float(v):
    """Robustly convert a value to float.

    Handles int/float, and numeric strings with thousand separators, currency
    symbols, percent signs, spaces and a 'USD' suffix. Returns None for None,
    unparseable strings, and formula cells (values starting with '=').
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return None
    s = (s.replace(",", "").replace("$", "").replace("¥", "")
          .replace("€", "").replace("%", "").replace(" ", "")
          .replace("USD", "").replace("usd", ""))
    try:
        return float(s)
    except ValueError:
        return None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=1.0):
    """Numeric closeness with tolerant parsing.

    If both sides parse to numbers -> |a - b| <= tol. If one side is
    unparseable (e.g. a formula cell or a label), fall back to a
    case-insensitive string comparison rather than failing blindly.
    """
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    return str_match(a, b)


def _norm_sheet(name):
    """Normalise a sheet/column name: lowercase, drop spaces/underscores/etc."""
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def get_sheet(wb, name):
    target = _norm_sheet(name)
    for s in wb.sheetnames:
        if _norm_sheet(s) == target:
            return wb[s]
    return None


def _norm_metric(s):
    """Normalise a Summary metric label so label variants ('Total_Employees',
    'Total Employees', 'Company Average Salary', 'Highest Paid Department',
    'highest_paid_dept') all map to one key."""
    s = str(s).strip().lower()
    s = s.replace("average", "avg").replace("department", "dept")
    return re.sub(r"[^a-z0-9]", "", s)


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Salary_Analysis.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Salary_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Salary_Analysis.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return False

    # data_only=False: we never rely on formula-cached values; numeric cells are
    # compared via _to_float / num_close which handle literals and formulas
    # gracefully. (See docs/task.md: cells must contain literal numeric values.)
    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=False)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=False)
    except Exception as e:
        check("Excel readable", False, str(e))
        return False

    all_ok = True

    # ---- Department_Stats sheet ----
    agent_dept = get_sheet(agent_wb, "Department_Stats")
    gt_dept = get_sheet(gt_wb, "Department_Stats")
    check("Sheet 'Department_Stats' exists", agent_dept is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_dept is None:
        all_ok = False
    else:
        a_rows = list(agent_dept.iter_rows(min_row=2, values_only=True))
        g_rows = list(gt_dept.iter_rows(min_row=2, values_only=True))

        # Build lookup by department name (dedupe, skip blank/header rows) so
        # row order and any stray duplicates cannot cause false negatives.
        a_lookup = {}
        for r in a_rows:
            if not r or r[0] is None:
                continue
            key = str(r[0]).strip().lower()
            if not key or key == "department":
                continue
            a_lookup[key] = r
        check("Department_Stats has 7 departments (unique)",
              len(a_lookup) == 7, f"Got {len(a_lookup)} unique departments")
        if len(a_lookup) != 7:
            all_ok = False

        for g_row in g_rows:
            if not g_row or g_row[0] is None:
                continue
            dept = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(dept)
            if a_row is None:
                check(f"Dept '{g_row[0]}' present", False, "Missing")
                all_ok = False
                continue
            # Headcount (col 1)
            ok = num_close(a_row[1], g_row[1], 5)
            check(f"'{g_row[0]}' Headcount", ok, f"Expected {g_row[1]}, got {a_row[1]}")
            if not ok:
                all_ok = False
            # Min_Salary (col 2)
            ok = num_close(a_row[2], g_row[2], 1)
            check(f"'{g_row[0]}' Min_Salary", ok, f"Expected {g_row[2]}, got {a_row[2]}")
            if not ok:
                all_ok = False
            # Max_Salary (col 3)
            ok = num_close(a_row[3], g_row[3], 1)
            check(f"'{g_row[0]}' Max_Salary", ok, f"Expected {g_row[3]}, got {a_row[3]}")
            if not ok:
                all_ok = False
            # Avg_Salary (col 4)
            ok = num_close(a_row[4], g_row[4], 100)
            check(f"'{g_row[0]}' Avg_Salary", ok, f"Expected {g_row[4]}, got {a_row[4]}")
            if not ok:
                all_ok = False
            # Median_Salary (col 5)
            ok = num_close(a_row[5], g_row[5], 50)
            check(f"'{g_row[0]}' Median_Salary", ok, f"Expected {g_row[5]}, got {a_row[5]}")
            if not ok:
                all_ok = False

        # Sort by Avg_Salary descending verification (over the deduped rows)
        lookup_avgs = []
        for r in a_lookup.values():
            v = _to_float(r[4])
            if v is not None:
                lookup_avgs.append(v)
        sorted_ok = len(lookup_avgs) == 7 and all(
            lookup_avgs[i] >= lookup_avgs[i + 1] for i in range(len(lookup_avgs) - 1))
        check("Department_Stats sorted by Avg_Salary descending", sorted_ok,
              f"Avg_Salaries: {[r[4] for r in a_lookup.values()]}")
        if not sorted_ok:
            all_ok = False

    # ---- Summary sheet ----
    agent_summary = get_sheet(agent_wb, "Summary")
    gt_summary = get_sheet(gt_wb, "Summary")
    check("Sheet 'Summary' exists", agent_summary is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_summary is None:
        all_ok = False
    else:
        a_summary = {}
        for row in agent_summary.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None and str(row[0]).strip():
                a_summary[_norm_metric(row[0])] = row[1]

        gt_sum = {}
        if gt_summary is not None:
            for row in gt_summary.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None and str(row[0]).strip():
                    gt_sum[_norm_metric(row[0])] = row[1]

        cas_expected = _to_float(gt_sum.get("companyavgsalary")) or 58396.14

        # Total_Employees (accepts 'Total_Employees' / 'Total Employees' / ...)
        te = a_summary.get("totalemployees")
        check("Total_Employees = 50000", num_close(te, 50000, 10), f"Got {te}")
        if not num_close(te, 50000, 10):
            all_ok = False

        # Company_Avg_Salary
        cas = a_summary.get("companyavgsalary")
        check("Company_Avg_Salary close to 58396", num_close(cas, cas_expected, 200), f"Got {cas}")
        if not num_close(cas, cas_expected, 200):
            all_ok = False

        # Highest_Paid_Dept (substring: 'Engineering', 'Engineering Dept', ...)
        hpd = str(a_summary.get("highestpaiddept", "")).strip().lower()
        check("Highest_Paid_Dept identifies 'Engineering'", "engineering" in hpd, f"Got '{hpd}'")
        if "engineering" not in hpd:
            all_ok = False

        # Lowest_Paid_Dept
        lpd = str(a_summary.get("lowestpaiddept", "")).strip().lower()
        check("Lowest_Paid_Dept identifies 'Operations'", "operations" in lpd, f"Got '{lpd}'")
        if "operations" not in lpd:
            all_ok = False

    return all_ok


def _gform_options(config):
    """Extract the list of option values from a question's config jsonb.

    The MCP stores options as a JSONB array of {"value": "..."} objects. The
    column may come back from psycopg2 as either a string or an object, so we
    normalise both. Returns lowercased, stripped option strings.
    """
    if config is None:
        return []
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except Exception:
            return []
    if not isinstance(config, dict):
        return []
    opts = config.get("options")
    if not isinstance(opts, list):
        return []
    out = []
    for o in opts:
        if isinstance(o, dict):
            v = o.get("value")
        else:
            v = o
        if v is not None:
            out.append(str(v).strip().lower())
    return out


def _opt_like(opts, word):
    """True if some option is exactly `word` (case-insensitive) or starts with
    it followed by a non-letter, e.g. 'Yes, absolutely' / 'No, not really'.
    This avoids failing faithful verbose phrasings while 'not sure' / 'maybe'
    style options never false-match the short 'no' token."""
    w = word.lower()
    for o in opts:
        o = str(o).strip().lower()
        if o == w:
            return True
        if o.startswith(w) and (len(o) == len(w) or not o[len(w)].isalpha()):
            return True
    return False


def check_gform():
    print("\n=== Checking Google Forms ===")
    try:
        conn = psycopg2.connect(**DB)
    except psycopg2.Error as e:
        check("Form titled exactly 'Compensation Satisfaction Survey'", False,
              f"DB unavailable: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        target_id = None
        for fid, title in forms:
            if (title or "").strip().lower() == "compensation satisfaction survey":
                target_id = fid
                break
        check("Form titled exactly 'Compensation Satisfaction Survey'", target_id is not None,
              f"Found titles: {[f[1] for f in forms]}")

        if target_id is not None:
            cur.execute("SELECT title, question_type, config FROM gform.questions "
                        "WHERE form_id=%s ORDER BY position", (target_id,))
            rows = cur.fetchall()
            questions = [{"title": (r[0] or ""), "type": (r[1] or "").lower(), "config": r[2]}
                         for r in rows]
            check("Form has exactly 5 questions", len(questions) == 5, f"Got {len(questions)}")

            q_titles = [q["title"].lower() for q in questions]
            q_types = [q["type"] for q in questions]

            # Content checks. Each task.md question has a distinct topic, but a
            # faithful model may paraphrase with synonyms, so each check accepts
            # a small family of equivalent wordings (case-insensitive substring;
            # ANY hit in a title counts) instead of one exact keyword.
            _KW = (
                ("satisf", "happy"),                                 # Q1 satisfaction
                ("competitiv", "industry", "market", "fair"),        # Q2 competitiveness
                ("benefit", "perk", "advantage"),                    # Q3 benefits
                ("leav", "switch", "another company", "other company",
                 "new job", "better pay", "higher pay", "better salary",
                 "higher salary", "more money", "resign", "quit",
                 "job offer", "elsewhere"),                          # Q4 leaving for better pay
                ("comment", "additional", "feedback", "anything else",
                 "anything to add", "suggest"),                      # Q5 additional comments
            )

            def _kw_hit(idx, text):
                return any(k in text for k in _KW[idx])

            has_q1 = any(_kw_hit(0, t) for t in q_titles)
            has_q2 = any(_kw_hit(1, t) for t in q_titles)
            has_q3 = any(_kw_hit(2, t) for t in q_titles)
            has_q4 = any(_kw_hit(3, t) for t in q_titles)
            has_q5 = any(_kw_hit(4, t) for t in q_titles)

            check("Q1 satisfaction question present", has_q1, f"titles={q_titles}")
            check("Q2 pay competitiveness question present", has_q2, f"titles={q_titles}")
            check("Q3 benefits question present", has_q3, f"titles={q_titles}")
            check("Q4 leaving-for-better-pay question present", has_q4, f"titles={q_titles}")
            check("Q5 additional comments question present", has_q5, f"titles={q_titles}")

            # Question types: the google_forms MCP can only produce
            # 'textQuestion' (add_text_question) or 'choiceQuestion'
            # (add_multiple_choice_question), so the form must be 3 multiple
            # choice + 2 open text. We never require types the MCP cannot make.
            n_choice = sum(1 for t in q_types if t == "choicequestion")
            n_short = sum(1 for t in q_types if t == "textquestion")
            check("3 multiple-choice questions (satisfaction, competitiveness, leaving)",
                  n_choice >= 3, f"types={q_types}")
            check("2 open-ended text questions (benefits, comments)",
                  n_short >= 2, f"types={q_types}")

            def _find(idx):
                for q in questions:
                    if _kw_hit(idx, q["title"].lower()):
                        return q
                return None

            q1 = _find(0)
            q2 = _find(1)
            q3 = _find(2)
            q4 = _find(3)
            q5 = _find(4)

            def _check_choice(q, label, opt_pred):
                if q is None:
                    return
                if q["type"] != "choicequestion":
                    check(f"{label} should be a multiple-choice question", False,
                          f"type={q['type']}")
                    return
                opts = _gform_options(q["config"])
                check(f"{label} has expected answer options", opt_pred(opts), f"options={opts}")

            def _check_text(q, label):
                if q is None:
                    return
                check(f"{label} should be an open-ended text question",
                      q["type"] == "textquestion", f"type={q['type']}")

            _check_choice(q1, "Q1",
                          lambda opts: len(opts) > 0 and
                          any("satisfied" in o for o in opts) and
                          any("dissatisfied" in o for o in opts))
            _check_choice(q2, "Q2",
                          lambda opts: _opt_like(opts, "yes") and
                          _opt_like(opts, "no") and
                          any("not sure" in o or o == "unsure" for o in opts))
            _check_text(q3, "Q3")
            _check_choice(q4, "Q4",
                          lambda opts: _opt_like(opts, "yes") and
                          _opt_like(opts, "no") and
                          any("maybe" in o for o in opts))
            _check_text(q5, "Q5")
    finally:
        cur.close()
        conn.close()


def _expected_sender(task_root):
    """Derive the expected From sender from the task-dir email_config.json.

    The emails MCP has no `from` parameter: the From header is always taken
    from email_config.json (with a `name` field the header is
    "Name <email>", otherwise a bare email). We compare case-insensitively.
    """
    cfg_path = os.path.join(task_root, "email_config.json")
    try:
        with open(cfg_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None
    if isinstance(data, list):
        data = data[0] if data else {}
    if not isinstance(data, dict):
        return None
    email = (data.get("email") or "").strip().lower()
    name = (data.get("name") or "").strip().lower()
    return {"email": email, "name": name}


def _mentions_number(text, expected, tol=1000):
    """True if the text contains a number within tol of the expected value.

    Tol 1000 covers faithful coarse rounding such as '59k' for 58,396.14
    (diff 604) while still rejecting numbers that are materially wrong
    (e.g. '50k' or '60k', which are >1500 off)."""
    if not text:
        return False
    s = str(text).lower().replace(",", "")
    for m in re.finditer(r"\$?\s*(\d+(?:\.\d+)?)\s*([kK])?", s):
        try:
            val = float(m.group(1))
        except ValueError:
            continue
        if m.group(2):
            val *= 1000
        if abs(val - expected) <= tol:
            return True
    return False


def _to_addr_list(to_addr):
    """Normalize a to_addr cell (jsonb list or string) to lowercased addresses."""
    if to_addr is None:
        return []
    if isinstance(to_addr, list):
        return [str(x).strip().lower() for x in to_addr]
    s = str(to_addr).strip()
    if s.startswith("["):
        try:
            items = json.loads(s)
            if isinstance(items, list):
                return [str(x).strip().lower() for x in items]
        except Exception:
            pass
    return [s.lower()] if s else []


def check_email(task_root, company_avg):
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB)
    except psycopg2.Error as e:
        check("Email with exact subject 'Compensation Analysis Report - Action Required'",
              False, f"DB unavailable: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        all_emails = cur.fetchall()

        expected_subject = "Compensation Analysis Report - Action Required"
        target = None
        for em in all_emails:
            if (em[0] or "").strip().lower() == expected_subject.lower():
                target = em
                break
        check("Email with exact subject 'Compensation Analysis Report - Action Required'",
              target is not None, f"Subjects seen: {[e[0] for e in all_emails]}")
        if target:
            subj, from_addr, to_addr, body = target
            to_addrs = _to_addr_list(to_addr)
            check("Email to hr-leadership@company.example.com",
                  "hr-leadership@company.example.com" in to_addrs, f"to: {to_addr}")

            # Sender is determined by the task-dir email_config.json.
            sender = _expected_sender(task_root)
            from_l = (from_addr or "").lower()
            if sender and sender["email"]:
                from_ok = sender["email"] in from_l
                check("Email from configured sender address",
                      from_ok, f"from: {from_addr}")
            else:
                check("Email from configured sender address", True,
                      "no email_config.json found; sender not verified")

            body_l = (body or "").lower()
            has_avg = _mentions_number(body_l, company_avg)
            check("Email body mentions company avg salary number",
                  has_avg, f"expected number near {company_avg}")
            check("Email body identifies highest dept (Engineering)",
                  "engineering" in body_l, "")
            check("Email body identifies lowest dept (Operations)",
                  "operations" in body_l, "")
    finally:
        cur.close()
        conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    company_avg = 58396.14  # derived from GT when available (see check_excel)
    try:
        gt_wb = openpyxl.load_workbook(os.path.join(gt_dir, "Salary_Analysis.xlsx"),
                                       data_only=False)
        gt_sum = get_sheet(gt_wb, "Summary")
        if gt_sum is not None:
            for row in gt_sum.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None and _norm_metric(row[0]) == "companyavgsalary":
                    v = _to_float(row[1])
                    if v is not None:
                        company_avg = v
                    break
    except Exception:
        pass

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_email(task_root, company_avg)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
