"""Evaluation for wc-review-competitor-study."""
import argparse
import json
import os
import sys

import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_data):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Review_Benchmark.xlsx")
    if not os.path.exists(path):
        return ["Review_Benchmark.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Category Comparison sheet
        rows = load_sheet_rows(wb, "Category Comparison")
        if rows is None:
            errors.append("Sheet 'Category Comparison' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_categories"]
            if len(data_rows) != expected:
                errors.append(f"Category Comparison has {len(data_rows)} rows, expected {expected}")

            # Check specific category - all columns - with EXACT category match
            for gc in gt_data["comparisons"]:
                cat_rows = [r for r in data_rows if r[0] and str(r[0]).strip().lower() == gc["category"].lower()]
                if not cat_rows:
                    errors.append(f"Category '{gc['category']}' not found in Category Comparison")
                    continue
                row = cat_rows[0]
                # Check our avg rating (col 1) with tolerance
                try:
                    if row[1] is None or abs(float(row[1]) - gc["our_avg_rating"]) > 0.15:
                        errors.append(f"{gc['category']} our_avg_rating={row[1]}, expected ~{gc['our_avg_rating']}")
                except (ValueError, TypeError):
                    errors.append(f"{gc['category']} our_avg_rating not numeric: {row[1]}")
                # Our_Review_Count (col 2) - exact integer
                try:
                    if row[2] is None or int(row[2]) != gc["our_review_count"]:
                        errors.append(f"{gc['category']} our_review_count={row[2]}, expected {gc['our_review_count']}")
                except (ValueError, TypeError):
                    errors.append(f"{gc['category']} our_review_count not int: {row[2]}")
                # Competitor_Avg_Rating (col 3)
                try:
                    if row[3] is None or abs(float(row[3]) - gc["competitor_avg_rating"]) > 0.15:
                        errors.append(f"{gc['category']} competitor_avg_rating={row[3]}, expected ~{gc['competitor_avg_rating']}")
                except (ValueError, TypeError):
                    errors.append(f"{gc['category']} competitor_avg_rating not numeric: {row[3]}")
                # Competitor_Review_Count (col 4) - allow some leniency
                try:
                    if row[4] is None or abs(int(row[4]) - gc["competitor_review_count"]) > 2:
                        errors.append(f"{gc['category']} competitor_review_count={row[4]}, expected ~{gc['competitor_review_count']}")
                except (ValueError, TypeError):
                    errors.append(f"{gc['category']} competitor_review_count not int: {row[4]}")
                # Rating_Difference (col 5)
                try:
                    if row[5] is None or abs(float(row[5]) - gc["rating_difference"]) > 0.15:
                        errors.append(f"{gc['category']} rating_difference={row[5]}, expected ~{gc['rating_difference']}")
                except (ValueError, TypeError):
                    errors.append(f"{gc['category']} rating_difference not numeric: {row[5]}")
                # Status (col 6) - exact match (case-insensitive)
                if row[6] is None or str(row[6]).strip().lower() != gc["status"].lower():
                    errors.append(f"{gc['category']} status={row[6]}, expected {gc['status']}")

        # Check Products Below Benchmark sheet
        rows2 = load_sheet_rows(wb, "Products Below Benchmark")
        if rows2 is None:
            errors.append("Sheet 'Products Below Benchmark' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            expected_pb = gt_data["products_below_count"]
            if abs(len(data_rows2) - expected_pb) > 2:
                errors.append(f"Products Below Benchmark has {len(data_rows2)} rows, expected ~{expected_pb}")

        # Check Summary sheet
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            kv = {str(r[0]).strip().lower(): r[1] for r in data_rows3 if r[0] is not None}
            expected_summary = {
                "total_categories": gt_data["total_categories"],
                "categories_above_benchmark": gt_data["categories_above"],
                "categories_below_benchmark": gt_data["categories_below"],
                "products_below_benchmark": gt_data["products_below_count"],
            }
            for key, exp_val in expected_summary.items():
                v = kv.get(key)
                if v is None:
                    errors.append(f"Summary metric '{key}' missing")
                    continue
                try:
                    if key == "products_below_benchmark":
                        # Allow products tolerance
                        if abs(int(v) - exp_val) > 2:
                            errors.append(f"Summary {key}={v}, expected ~{exp_val}")
                    else:
                        if int(v) != exp_val:
                            errors.append(f"Summary {key}={v}, expected {exp_val}")
                except (ValueError, TypeError):
                    errors.append(f"Summary {key} not int: {v}")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def _extract_title_text(properties):
    """Pull plain text from a notion title property, handling list/dict/json shapes."""
    if not properties:
        return ""
    if isinstance(properties, str):
        try:
            properties = json.loads(properties)
        except Exception:
            return properties.lower()
    if not isinstance(properties, dict):
        return str(properties).lower()
    for key, prop in properties.items():
        if not isinstance(prop, dict):
            continue
        ptype = prop.get("type")
        if ptype == "title":
            title_arr = prop.get("title") or []
            if isinstance(title_arr, list):
                texts = []
                for el in title_arr:
                    if isinstance(el, dict):
                        if el.get("plain_text"):
                            texts.append(el["plain_text"])
                        elif el.get("text") and isinstance(el["text"], dict):
                            texts.append(el["text"].get("content", ""))
                if texts:
                    return " ".join(texts).lower()
        # fallback: 'Name' / 'title' key with simple string value
        if key.lower() in ("name", "title") and isinstance(prop, dict):
            for nested_key in ("plain_text", "content", "value"):
                if nested_key in prop:
                    return str(prop[nested_key]).lower()
    return ""


def check_notion(gt_data):
    errors = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties::text ILIKE '%review performance analysis%'
               OR properties::text ILIKE '%performance analysis q1%'
            ORDER BY created_time DESC NULLS LAST
            LIMIT 50
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # Stricter: the title field of the page (not random properties JSON)
        # must contain all of {review, performance, analysis, q1, 2026}.
        found = False
        for row in rows:
            props = row[1]
            if isinstance(props, str):
                try:
                    props = json.loads(props)
                except Exception:
                    pass
            title_text = _extract_title_text(props)
            if not title_text:
                # fallback to JSON substring if title parse fails
                title_text = (json.dumps(props) if isinstance(props, dict) else str(props)).lower()
            if all(kw in title_text for kw in ("review", "performance", "analysis", "q1", "2026")):
                found = True
                break

        if not found:
            errors.append("No Notion page with title 'Review Performance Analysis Q1 2026'")

    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion page...")
    errs = check_notion(gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
