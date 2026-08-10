"""
Evaluation for yt-fireship-gform-survey-excel-gcal task.

Checks:
1. Community_Report.xlsx exists with "Top_Videos" sheet having 8 data rows
2. Engagement_Rate column exists in Top_Videos
3. Engagement_Analysis sheet exists with >= 4 rows
4. GForm created with title "Fireship Community Preference Survey"
5. GForm has 5 questions with correct radio/text types and wording
6. GCal has new "Community Standup" event on 2026-04-01 18:00-19:00 UTC
   (not the noise Q&A at 16:00)
7. Email sent to community@devclub.io with engagement-report subject/body
   (sender is not checked; derived from the task's email_config.json)

Groundtruth workbook comparison:
- Top_Videos rows matched by Video_ID; numeric columns compared with tolerance;
  Topic_Tags is subjective (task says "derived from the video's content theme")
  so it is skipped; Title compared with punctuation-normalized fuzzy matching.
- Engagement_Analysis is validated via grouping-independent aggregate invariants
  (row count >= GT, sum(Video_Count) == number of top videos,
  sum(Total_Views) == sum of top views,
  sum(Avg_Engagement_Rate * Video_Count) == sum of per-video engagement rates)
  because topic labels are subjective and a correct model may legitimately use
  different topic names/groupings.
- Agent cells that are Excel formulas are evaluated in-process (lightweight
  evaluator for arithmetic + SUM/AVERAGE/COUNTIF/refs). Cells that cannot be
  evaluated are skipped rather than failed, so a correct model that computes
  engagement with formulas is never wrongly penalized (R2).

CLI contract (unchanged):
  python evaluation/main.py --agent_workspace <ws> --groundtruth_workspace <gt>
                            [--res_log_file <f>] [--launch_time <iso>]
"""
import os
import re
import sys
import json
from argparse import ArgumentParser
from datetime import timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import psycopg2
import openpyxl
from openpyxl.utils import column_index_from_string

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _parse_duration_str(s):
    """Parse a duration shown as 'm:ss', 'mm:ss' or 'h:mm:ss' into seconds.

    Excel/agents commonly display the task's Duration_Sec (a plain seconds
    number in GT) as a mm:ss wall-clock string. Accept those forms so a correct
    model that renders durations readably is not wrongly failed.
    """
    if not isinstance(s, str):
        return None
    m = re.fullmatch(r"\s*(\d+):([0-5]?\d)(?::([0-5]?\d))?\s*", s)
    if not m:
        return None
    if m.group(3) is not None:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3))
    return int(m.group(1)) * 60 + int(m.group(2))


def _to_float(v):
    """Robust numeric conversion: str/int/float/None -> float or None.

    Strips thousands separators, currency symbols, percent signs and spaces.
    Percent strings like '90%' parse to 90.0. Formula strings ('=...') return
    None here and are handled by the in-process formula evaluator instead.
    Duration display strings ('m:ss' / 'h:mm:ss') parse to their seconds value.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        if s.startswith("="):
            return None
        s2 = s.replace("$", "").replace("€", "").replace("¥", "").replace("£", "")
        s2 = s2.replace(",", "").replace("%", "").replace(" ", "")
        try:
            return float(s2)
        except ValueError:
            pass
        d = _parse_duration_str(s)
        if d is not None:
            return float(d)
        return None
    return None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=1.0):
    """Compare two values numerically when both are numeric; otherwise fall back
    to a case-insensitive string comparison. Never raises on bad types."""
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    return str_match(a, b)


# ---------------------------------------------------------------------------
# Lightweight in-process Excel formula evaluator (R2).
# Supports arithmetic, cell refs (incl. cross-sheet and absolute), and
# SUM/AVERAGE/MAX/MIN/COUNT/COUNTIF/ROUND/IF. Any unsupported construct makes
# the evaluation return None so the caller can skip rather than fail.
# ---------------------------------------------------------------------------

class FormulaError(Exception):
    pass


_CELL_RE = re.compile(r"\$?[A-Za-z]{1,3}\$?\d+")
_FUNCS = {"SUM", "AVERAGE", "MAX", "MIN", "COUNT", "COUNTIF", "ROUND", "IF"}


def _ref_to_rc(ref):
    m = re.match(r"([A-Za-z]{1,3})(\d+)", ref)
    if not m:
        raise FormulaError(f"bad cell ref {ref!r}")
    return column_index_from_string(m.group(1)), int(m.group(2))


def _tokenize(expr):
    tokens = []
    i, n = 0, len(expr)
    while i < n:
        c = expr[i]
        if c in " \t":
            i += 1
            continue
        if c == "'":  # quoted sheet name, e.g. 'Top Videos'!A1
            j = expr.find("'", i + 1)
            if j == -1:
                raise FormulaError("unterminated quoted sheet name")
            sheet = expr[i + 1:j]
            i = j + 1
            if i < n and expr[i] == "!":
                i += 1
            tokens.append(("SHEET", sheet))
            continue
        if c == '"':  # string literal
            j = expr.find('"', i + 1)
            if j == -1:
                raise FormulaError("unterminated string literal")
            tokens.append(("STR", expr[i + 1:j]))
            i = j + 1
            continue
        if c.isdigit() or (c == "." and i + 1 < n and expr[i + 1].isdigit()):
            j = i
            while j < n and (expr[j].isdigit() or expr[j] == "."):
                j += 1
            tokens.append(("NUM", float(expr[i:j])))
            i = j
            continue
        if c.isalpha() or c == "_" or c == "$":
            m = _CELL_RE.match(expr, i)
            if m:
                tokens.append(("CELL", m.group(0).replace("$", "")))
                i = m.end()
                continue
            j = i
            while j < n and (expr[j].isalnum() or expr[j] == "_"):
                j += 1
            word = expr[i:j]
            i = j
            if i < n and expr[i] == "!":
                i += 1
                tokens.append(("SHEET", word))
            else:
                tokens.append(("ID", word))
            continue
        if c in "(),:+-*/^%":
            tokens.append(("OP", c))
            i += 1
            continue
        raise FormulaError(f"unexpected char {c!r}")
    return tokens


def _matches_criterion(v, crit):
    if v is None or crit is None:
        return False
    if isinstance(v, str) and isinstance(crit, str):
        return v.strip().lower() == crit.strip().lower()
    fv = _to_float(v)
    fc = _to_float(crit)
    if fv is not None and fc is not None:
        return abs(fv - fc) < 1e-9
    return False


class _FormulaParser:
    def __init__(self, tokens, sheet_name, numeric_cell, raw_cell):
        self.tokens = tokens
        self.pos = 0
        self.sheet_name = sheet_name
        self.numeric_cell = numeric_cell
        self.raw_cell = raw_cell

    def _peek(self):
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def _next(self):
        t = self._peek()
        if t is None:
            raise FormulaError("unexpected end of formula")
        self.pos += 1
        return t

    def _expect(self, op):
        t = self._next()
        if t != ("OP", op):
            raise FormulaError(f"expected {op!r}, got {t}")

    def parse(self):
        val = self._expr()
        if self.pos != len(self.tokens):
            raise FormulaError("trailing tokens")
        return val

    def _expr(self):
        v = self._term()
        while True:
            t = self._peek()
            if t and t[0] == "OP" and t[1] in "+-":
                self._next()
                r = self._term()
                v = v + r if t[1] == "+" else v - r
            else:
                break
        return v

    def _term(self):
        v = self._factor()
        while True:
            t = self._peek()
            if t and t[0] == "OP" and t[1] in "*/":
                self._next()
                r = self._factor()
                if t[1] == "*":
                    v = v * r
                else:
                    if r == 0:
                        raise FormulaError("division by zero")
                    v = v / r
            else:
                break
        return v

    def _factor(self):
        t = self._next()
        if t[0] == "OP":
            if t[1] == "-":
                return -self._factor()
            if t[1] == "+":
                return self._factor()
            if t[1] == "(":
                v = self._expr()
                self._expect(")")
                return self._postfix_pct(v)
            raise FormulaError(f"unexpected operator {t[1]!r}")
        if t[0] == "NUM":
            return self._postfix_pct(t[1])
        if t[0] == "STR":
            return self._postfix_pct(t[1])
        if t[0] == "CELL":
            v = self.numeric_cell(self.sheet_name, *_ref_to_rc(t[1]))
            if v is None:
                raise FormulaError("unresolvable cell")
            return self._postfix_pct(v)
        if t[0] == "SHEET":
            sheet = t[1]
            t2 = self._peek()
            if t2 and t2[0] == "CELL":
                self._next()
                v = self.numeric_cell(sheet, *_ref_to_rc(t2[1]))
                if v is None:
                    raise FormulaError("unresolvable cell")
                return self._postfix_pct(v)
            raise FormulaError("sheet ref must precede a cell")
        if t[0] == "ID":
            fn = t[1].upper()
            if fn in _FUNCS:
                return self._call_func(fn)
            raise FormulaError(f"unknown function {t[1]!r}")
        raise FormulaError(f"unexpected token {t}")

    def _postfix_pct(self, v):
        t = self._peek()
        if t and t[0] == "OP" and t[1] == "%":
            self._next()
            if isinstance(v, (int, float)):
                return v * 0.01
            raise FormulaError("percent applied to non-number")
        return v

    def _parse_cell_or_range(self):
        t = self._peek()
        if t is None:
            raise FormulaError("missing argument")
        if t[0] == "CELL":
            self._next()
            t2 = self._peek()
            if t2 and t2[0] == "OP" and t2[1] == ":":
                self._next()
                t3 = self._next()
                if t3[0] != "CELL":
                    raise FormulaError("bad range end")
                return ("range", self.sheet_name, t[1], t3[1])
            return ("cell", self.sheet_name, t[1])
        if t[0] == "SHEET":
            self._next()
            t2 = self._next()
            if t2[0] != "CELL":
                raise FormulaError("sheet ref must precede a cell")
            t3 = self._peek()
            if t3 and t3[0] == "OP" and t3[1] == ":":
                self._next()
                t4 = self._next()
                if t4[0] != "CELL":
                    raise FormulaError("bad range end")
                return ("range", t[1], t2[1], t4[1])
            return ("cell", t[1], t2[1])
        return ("expr", self._expr())

    def _range_cells(self, sname, ref1, ref2):
        col1, row1 = _ref_to_rc(ref1)
        col2, row2 = _ref_to_rc(ref2)
        if col1 > col2:
            col1, col2 = col2, col1
        if row1 > row2:
            row1, row2 = row2, row1
        cells = []
        for row in range(row1, row2 + 1):
            for col in range(col1, col2 + 1):
                cells.append((sname, col, row))
        return cells

    def _resolve_range(self, sname, ref1, ref2):
        out = []
        for s, col, row in self._range_cells(sname, ref1, ref2):
            v = self.numeric_cell(s, col, row)
            if v is not None:
                out.append(v)
        return out

    def _parse_value_arg(self):
        res = self._parse_cell_or_range()
        if res[0] == "cell":
            v = self.numeric_cell(res[1], *_ref_to_rc(res[2]))
            if v is None:
                raise FormulaError("unresolvable cell")
            return v
        if res[0] == "range":
            return self._resolve_range(res[1], res[2], res[3])
        return res[1]

    def _parse_raw_value_arg(self):
        res = self._parse_cell_or_range()
        if res[0] == "cell":
            return self.raw_cell(res[1], *_ref_to_rc(res[2]))
        if res[0] == "range":
            raise FormulaError("range not allowed here")
        return res[1]

    def _call_func(self, fn):
        self._expect("(")
        if fn == "IF":
            cond = self._expr()
            self._expect(",")
            tv = self._expr()
            self._expect(",")
            fv = self._expr()
            self._expect(")")
            return tv if cond else fv
        if fn == "ROUND":
            a = self._expr()
            self._expect(",")
            d = self._expr()
            self._expect(")")
            return float(round(a, int(d)))
        if fn == "COUNTIF":
            rng = self._parse_countif_range()
            self._expect(",")
            crit = self._parse_raw_value_arg()
            self._expect(")")
            cnt = 0
            for s, col, row in rng:
                if _matches_criterion(self.raw_cell(s, col, row), crit):
                    cnt += 1
            return float(cnt)
        args = [self._parse_value_arg()]
        while self._peek() and self._peek()[0] == "OP" and self._peek()[1] == ",":
            self._next()
            args.append(self._parse_value_arg())
        self._expect(")")
        vals = []
        for a in args:
            if isinstance(a, list):
                vals.extend(a)
            elif a is not None:
                vals.append(a)
        if not vals:
            raise FormulaError("empty range for " + fn)
        if fn == "SUM":
            return sum(vals)
        if fn == "AVERAGE":
            return sum(vals) / len(vals)
        if fn == "MAX":
            return max(vals)
        if fn == "MIN":
            return min(vals)
        if fn == "COUNT":
            return float(len(vals))
        raise FormulaError(fn)

    def _parse_countif_range(self):
        res = self._parse_cell_or_range()
        if res[0] == "range":
            return self._range_cells(res[1], res[2], res[3])
        raise FormulaError("COUNTIF requires a range")


def evaluate_formula(formula, sheet_name, numeric_cell, raw_cell):
    body = formula[1:] if formula.startswith("=") else formula
    try:
        tokens = _tokenize(body)
    except FormulaError:
        return None
    parser = _FormulaParser(tokens, sheet_name, numeric_cell, raw_cell)
    try:
        val = parser.parse()
    except (FormulaError, ZeroDivisionError, TypeError, ValueError, OverflowError):
        return None
    if isinstance(val, str):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _build_resolver(wb, sheets_map):
    cache = {}
    evaluating = set()

    def raw_cell(sheet, col, row):
        ws = sheets_map.get(str(sheet).strip().lower())
        if ws is None:
            return None
        try:
            return ws.cell(row=row, column=col).value
        except Exception:
            return None

    def numeric_cell(sheet, col, row):
        key = (str(sheet).strip().lower(), col, row)
        if key in cache:
            return cache[key]
        if key in evaluating:
            return None
        evaluating.add(key)
        v = raw_cell(sheet, col, row)
        if isinstance(v, str) and v.startswith("="):
            val = evaluate_formula(v, sheet, numeric_cell, raw_cell)
        else:
            val = _to_float(v)
        evaluating.discard(key)
        cache[key] = val
        return val

    return raw_cell, numeric_cell


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

_HEADER_ABBR = {
    "avg": "average",
    "sec": "seconds",
}


def _norm_header(h):
    """Normalize a header for fuzzy matching: lowercase, keep only letters and
    digits, and expand common abbreviations (avg->average, sec->seconds).
    Tolerant of spaces, underscores, '%', '()', '/', '-' etc., so that e.g.
    'Avg Engagement Rate (%)' and 'Average Engagement Rate' both match GT's
    'Avg_Engagement_Rate', and 'Duration (sec)' matches 'Duration_Sec'."""
    tokens = re.findall(r"[a-z0-9]+", str(h).lower())
    return "".join(_HEADER_ABBR.get(t, t) for t in tokens)


def _gt_header_present(gt_key, agent_keys):
    """True if a GT header (normalized) is covered by any agent header
    (normalized). Uses exact match or one-way containment so natural column
    spellings ('Avg_Engagement_Rate' vs 'Average Engagement Rate',
    'Total_Views' vs 'Views', 'Topic_Tags' vs 'Topic') are accepted."""
    return any(gt_key == ak or gt_key in ak
               or (len(gt_key) >= 4 and ak in gt_key)
               for ak in agent_keys)


def _fuzzy_title_match(gv, av):
    """Title comparison tolerant of punctuation/typography differences
    (apostrophes, ellipses, dashes) and trailing truncation."""
    if gv is None or av is None:
        return gv is None and av is None
    g = "".join(ch for ch in str(gv).lower() if ch.isalnum())
    a = "".join(ch for ch in str(av).lower() if ch.isalnum())
    if not g or not a:
        return str_match(gv, av)
    if g == a:
        return True
    return len(g) >= 15 and (g in a or a in g)


def _resolve_numeric(av, sheet_name, numeric_cell, raw_cell):
    """Resolve an agent workbook cell to (status, numeric_value).

    status is one of:
      'ok'            -> value parsed to a float (av_num)
      'missing'       -> empty cell / blank string
      'unparsable'    -> a value that is not numeric and not a formula
      'unverifiable'  -> a formula the evaluator cannot resolve (lenient skip)
    Empty / unparsable cells are treated as missing data (fail); unverifiable
    formulas are skipped rather than failed so a correct model that computes
    values with formulas is never wrongly penalized.
    """
    if av is None:
        return ("missing", None)
    if isinstance(av, str):
        if av.startswith("="):
            val = evaluate_formula(av, sheet_name, numeric_cell, raw_cell)
            if val is None:
                return ("unverifiable", None)
            return ("ok", val)
        if not av.strip():
            return ("missing", None)
        val = _to_float(av)
        if val is None:
            return ("unparsable", None)
        return ("ok", val)
    val = _to_float(av)
    if val is None:
        return ("unparsable", None)
    return ("ok", val)


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-3: Community_Report.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Community_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Community_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Top_Videos sheet has 8 data rows", False, "File missing")
        record("Engagement_Rate column exists in Top_Videos", False, "File missing")
        record("Engagement_Analysis sheet has >= 4 rows", False, "File missing")
        return
    record("Community_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    sheets = {s.strip().lower(): s for s in wb.sheetnames}
    top_key = next((v for k, v in sheets.items() if "video" in k or "top" in k), None)
    eng_key = next((v for k, v in sheets.items() if "engagement" in k or "analysis" in k), None)

    if not top_key:
        record("Top_Videos sheet has 8 data rows", False,
               f"No Top_Videos sheet. Sheets: {wb.sheetnames}")
        record("Engagement_Rate column exists in Top_Videos", False, "Sheet missing")
    else:
        ws = wb[top_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)] if rows else []
        record("Top_Videos sheet has 8 data rows", len(data_rows) == 8,
               f"Found {len(data_rows)} data rows, expected 8")
        headers = [str(c).strip().lower() if c else "" for c in rows[0]] if rows else []
        record("Engagement_Rate column exists in Top_Videos",
               any("engagement" in h or "rate" in h for h in headers),
               f"Headers: {rows[0] if rows else None}")

    if not eng_key:
        record("Engagement_Analysis sheet has >= 4 rows", False,
               f"No Engagement_Analysis sheet. Sheets: {wb.sheetnames}")
    else:
        ws2 = wb[eng_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c is not None for c in r)] if rows2 else []
        record("Engagement_Analysis sheet has >= 4 rows", len(data_rows2) >= 4,
               f"Found {len(data_rows2)} data rows, expected at least 4")

    # --- Groundtruth workbook comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Community_Report.xlsx")
    if not os.path.isfile(gt_path):
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    try:
        a_wb = openpyxl.load_workbook(os.path.join(agent_workspace, "Community_Report.xlsx"),
                                      data_only=False)
    except Exception:
        a_wb = None

    if a_wb is not None:
        a_sheets_map = {s.strip().lower(): a_wb[s] for s in a_wb.sheetnames}
        raw_cell, numeric_cell = _build_resolver(a_wb, a_sheets_map)

        # Expected aggregates derived from the (trusted) GT Top_Videos sheet.
        gt_top_headers = []
        gt_top_rows = []
        for gt_sname in gt_wb.sheetnames:
            if "video" in gt_sname.strip().lower() or "top" in gt_sname.strip().lower():
                gtw = gt_wb[gt_sname]
                gt_top_headers = [str(c.value).strip().lower() if c.value else ""
                                  for c in next(gtw.iter_rows(max_row=1))]
                gt_top_rows = [r for r in gtw.iter_rows(min_row=2, values_only=True)
                               if any(c is not None for c in r)]
                break
        expected_count = float(len(gt_top_rows)) if gt_top_rows else 0.0
        expected_views = 0.0
        expected_rate_sum = 0.0
        if gt_top_rows:
            v_idx = next((i for i, h in enumerate(gt_top_headers) if h == "views"), None)
            r_idx = next((i for i, h in enumerate(gt_top_headers)
                          if "rate" in h or "engagement" in h), None)
            for r in gt_top_rows:
                if v_idx is not None and v_idx < len(r):
                    expected_views += _to_float(r[v_idx]) or 0.0
                if r_idx is not None and r_idx < len(r):
                    expected_rate_sum += _to_float(r[r_idx]) or 0.0

        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            gt_key = gt_sname.strip().lower()
            if "video" in gt_key or "top" in gt_key:
                agent_sheet = top_key if top_key else None
            else:
                agent_sheet = eng_key if eng_key else None
            if agent_sheet is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                       f"Available: {a_wb.sheetnames}")
                continue
            a_ws = a_wb[agent_sheet]
            gt_headers = [str(c.value).strip().lower() if c.value else ""
                          for c in next(gt_ws.iter_rows(max_row=1))]
            a_headers = [str(c.value).strip().lower() if c.value else ""
                         for c in next(a_ws.iter_rows(max_row=1))]
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                       if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None for c in r)]

            gt_norm = [_norm_header(h) for h in gt_headers]
            a_norm = [_norm_header(h) for h in a_headers]
            missing = [gh for gh, gn in zip(gt_headers, gt_norm)
                       if not _gt_header_present(gn, a_norm)]
            record(f"GT '{gt_sname}' all GT columns present in agent",
                   not missing, f"Missing columns: {missing}")

            if "video" in gt_key or "top" in gt_key:
                _compare_top_videos(a_ws, gt_headers, a_headers, gt_rows, a_rows,
                                    raw_cell, numeric_cell, agent_sheet)
            else:
                _compare_engagement_analysis(a_ws, gt_rows, a_rows, a_headers,
                                             raw_cell, numeric_cell, agent_sheet,
                                             expected_count, expected_views,
                                             expected_rate_sum)
        a_wb.close()
    gt_wb.close()


def _compare_top_videos(a_ws, gt_headers, a_headers, gt_rows, a_rows,
                        raw_cell, numeric_cell, a_sheet_name):
    gt_norm = [_norm_header(h) for h in gt_headers]
    a_norm_to_idx = {}
    for i, h in enumerate(a_headers):
        a_norm_to_idx.setdefault(_norm_header(h), i)
    gt_norm_to_idx = {}
    for i, hn in enumerate(gt_norm):
        gt_norm_to_idx.setdefault(hn, i)

    pk_col = "videoid"
    if pk_col not in gt_norm_to_idx or pk_col not in a_norm_to_idx:
        record("GT 'Top_Videos' has stable PK column video_id", False,
               f"No video_id header found: {gt_headers}")
        return
    gt_pk_idx = gt_norm_to_idx[pk_col]
    a_pk_idx = a_norm_to_idx[pk_col]
    record("GT 'Top_Videos' row count", len(a_rows) == len(gt_rows),
           f"Expected {len(gt_rows)}, got {len(a_rows)}")

    a_lookup = {}
    for r in a_rows:
        if r and len(r) > a_pk_idx and r[a_pk_idx] is not None:
            a_lookup[str(r[a_pk_idx]).strip().lower()] = r

    for gt_row in gt_rows:
        if not gt_row or len(gt_row) <= gt_pk_idx or gt_row[gt_pk_idx] is None:
            continue
        pk_val = str(gt_row[gt_pk_idx]).strip().lower()
        a_row = a_lookup.get(pk_val)
        record(f"GT 'Top_Videos' pk={pk_val} present", a_row is not None)
        if a_row is None:
            continue
        ok = True
        fail_detail = ""
        skipped = []
        for ci, gh in enumerate(gt_headers):
            if ci >= len(gt_row):
                continue
            gv = gt_row[ci]
            if gv is None:
                continue
            ghn = _norm_header(gh)
            if "topic" in ghn and "tag" in ghn:
                continue  # subjective
            a_ci = a_norm_to_idx.get(ghn)
            if a_ci is None:
                continue
            if a_ci >= len(a_row):
                continue
            av = a_row[a_ci]
            if isinstance(gv, (int, float)) and not isinstance(gv, bool):
                st, av_num = _resolve_numeric(av, a_sheet_name, numeric_cell, raw_cell)
                if st == "unverifiable":
                    skipped.append(gh)  # formula we can't evaluate -> don't penalize
                    continue
                if st != "ok":
                    ok = False
                    fail_detail = f"col '{gh}': gt={gv}, agent={av!r} ({st})"
                    break
                if "rate" in ghn or "engagement" in ghn:
                    tol = max(abs(float(gv)) * 0.05, 0.1)
                else:
                    tol = max(abs(float(gv)) * 0.01, 0.5)
                if abs(av_num - float(gv)) > tol:
                    ok = False
                    fail_detail = f"col '{gh}': gt={gv}, agent={av}"
                    break
            else:
                if "title" in ghn:
                    if not _fuzzy_title_match(gv, av):
                        ok = False
                        fail_detail = f"col '{gh}': gt={gv!r}, agent={av!r}"
                        break
                elif not str_match(av, gv):
                    ok = False
                    fail_detail = f"col '{gh}': gt={gv!r}, agent={av!r}"
                    break
        note = f" (skipped unverifiable formula cols: {skipped})" if skipped else ""
        record(f"GT 'Top_Videos' pk={pk_val} values", ok, fail_detail + note)


_TOTAL_ROW_HINTS = ("total", "summary", "overall", "grand", "合计", "总计", "汇总")


def _is_total_row(topic):
    """True if an analysis row is a summary/total row (a common spreadsheet
    practice the task neither requires nor forbids). Such rows must be excluded
    from the aggregate invariants, otherwise a correct model that appends a
    'Total' row is wrongly failed."""
    if topic is None:
        return False
    t = str(topic).strip().lower()
    return bool(t) and any(hint in t for hint in _TOTAL_ROW_HINTS)


def _compare_engagement_analysis(a_ws, gt_rows, a_rows, a_headers,
                                 raw_cell, numeric_cell, a_sheet_name,
                                 expected_count, expected_views, expected_rate_sum):
    record("GT 'Engagement_Analysis' row count", len(a_rows) >= len(gt_rows),
           f"Expected >= {len(gt_rows)}, got {len(a_rows)}")

    def _find_analysis_col(fallbacks):
        """Resolve one required analysis column by normalized-header keyword
        matching, trying each (must_contain, any_contain) fallback in order and
        never reusing a column already assigned to another slot. Accepts natural
        header spellings: 'Avg_Engagement_Rate' / 'Average Engagement Rate' /
        'Avg Engagement Rate (%)' / 'Avg Engagement (%)', 'Total_Views' /
        'Views', 'Video_Count' / 'Count', 'Topic' / 'Theme'."""
        for must, anyof in fallbacks:
            for i, h in enumerate(a_headers):
                if i in used:
                    continue
                k = _norm_header(h)
                if ((must and all(m in k for m in must))
                        or (anyof and any(m in k for m in anyof))):
                    return i
        return None

    used = set()
    topic_i = _find_analysis_col([(("topic",), None), (None, ("theme", "category", "subject"))])
    if topic_i is not None:
        used.add(topic_i)
    avg_i = _find_analysis_col([(("engagement", "rate"), None),
                                (None, ("rate",)),
                                (("engagement", "average"), None)])
    if avg_i is not None:
        used.add(avg_i)
    tot_i = _find_analysis_col([(("total", "view"), None), (None, ("view",))])
    if tot_i is not None:
        used.add(tot_i)
    cnt_i = _find_analysis_col([(("video", "count"), None), (None, ("count",))])
    if topic_i is None or avg_i is None or tot_i is None or cnt_i is None:
        record("GT 'Engagement_Analysis' has required columns", False,
               f"Headers: {a_headers}")
        return

    def cell_status(row, idx):
        if idx is None or row is None or idx >= len(row):
            return ("missing", None)
        return _resolve_numeric(row[idx], a_sheet_name, numeric_cell, raw_cell)

    resolved = 0
    unverifiable = 0
    missing_bad = ""
    sum_cnt = 0.0
    sum_views = 0.0
    sum_rate = 0.0
    row_ok = True
    detail = ""
    for r in a_rows:
        if not r:
            continue
        topic_v = r[topic_i] if topic_i < len(r) else None
        if topic_v is None or (isinstance(topic_v, str) and not topic_v.strip()):
            continue
        if _is_total_row(topic_v):
            continue  # summary/total row -> exclude from aggregate invariants
        c = cell_status(r, cnt_i)
        t = cell_status(r, tot_i)
        a = cell_status(r, avg_i)
        if any(st == "unverifiable" for st, _ in (c, t, a)):
            unverifiable += 1
            continue
        if any(st in ("missing", "unparsable") for st, _ in (c, t, a)):
            missing_bad = (f"missing/unparsable analysis values for "
                           f"topic={topic_v}")
            continue
        resolved += 1
        sum_cnt += c[1]
        sum_views += t[1]
        sum_rate += a[1] * c[1]
        if c[1] < 1 or t[1] <= 0 or a[1] < 0 or a[1] > 100:
            row_ok = False
            detail = (f"implausible analysis row topic={topic_v}: "
                      f"count={c[1]}, views={t[1]}, rate={a[1]}")

    record("GT 'Engagement_Analysis' per-row values plausible",
           row_ok and not missing_bad,
           missing_bad or detail or "")
    if resolved == 0:
        if unverifiable > 0 and not missing_bad:
            record("GT 'Engagement_Analysis' aggregate values", True,
                   f"All {unverifiable} analysis rows use formulas the "
                   "evaluator cannot resolve; aggregate checks skipped")
        else:
            record("GT 'Engagement_Analysis' aggregate values", False,
                   "No analysis row values could be resolved")
        return

    record("GT 'Engagement_Analysis' sum(Video_Count) == top videos",
           abs(sum_cnt - expected_count) <= max(0.5, 0.001 * expected_count),
           f"Expected {expected_count:.0f}, got {sum_cnt:.1f} "
           f"(resolved {resolved} of {len(a_rows)} rows)")
    record("GT 'Engagement_Analysis' sum(Total_Views) == top views",
           abs(sum_views - expected_views) <= max(1.0, 0.001 * expected_views),
           f"Expected {expected_views:.0f}, got {sum_views:.0f}")
    record("GT 'Engagement_Analysis' weighted avg engagement sum",
           abs(sum_rate - expected_rate_sum) <= max(0.2, 0.02 * abs(expected_rate_sum)),
           f"Expected {expected_rate_sum:.2f}, got {sum_rate:.2f}")


# ---------------------------------------------------------------------------
# GForm / GCal / Email checks
# ---------------------------------------------------------------------------

def check_gform():
    print("\n=== Check 4-5: Google Form ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT f.id, f.title, COUNT(q.id) AS qcount
                FROM gform.forms f
                LEFT JOIN gform.questions q ON q.form_id = f.id
                WHERE f.title ILIKE '%Fireship Community Preference Survey%'
                   OR (f.title ILIKE '%Fireship%' AND f.title ILIKE '%Community%' AND f.title ILIKE '%Survey%')
                GROUP BY f.id, f.title
                ORDER BY qcount DESC, MAX(f.created_at) DESC, f.id
            """)
            forms = cur.fetchall()
            if not forms:
                record("GForm 'Fireship Community Preference Survey' exists", False,
                       "No matching form found")
                record("GForm has 5 questions", False, "Form missing")
                record("GForm radio/text question types match task", False, "Form missing")
                conn.close()
                return
            record("GForm 'Fireship Community Preference Survey' exists", True,
                   f"Found: {[f[1] for f in forms]}")

            form_id = forms[0][0]
            cur.execute("""
                SELECT title, question_type FROM gform.questions
                WHERE form_id = %s ORDER BY "position", id
            """, (form_id,))
            rows = cur.fetchall()
            q_titles = [r[0] for r in rows]
            q_types = [r[1] for r in rows]
            record("GForm has 5 questions", len(q_titles) >= 5,
                   f"Found {len(q_titles)} questions, expected at least 5 "
                   "(the five specified questions are verified below by wording)")

            # Validate wording by distinctive keyword pairs (tolerant to phrasing).
            required_keywords = [
                [("topic", "interest"), ("interest", "most")],          # Q1
                [("often", "watch"), ("frequently", "watch"), ("how", "often")],  # Q2
                [("format", "prefer"), ("prefer",), ("video", "length")],  # Q3
                [("role",), ("job",), ("position",)],                   # Q4
                [("topic", "next"), ("suggest", "topic")],              # Q5
            ]
            for kw_sets in required_keywords:
                ok = any(all(kw in (q or "").lower() for kw in kwset)
                         for kwset in kw_sets for q in q_titles)
                record(
                    f"GForm has question matching keywords {kw_sets}",
                    ok,
                    f"Question titles: {q_titles}",
                )

            # The MCP only produces two real question types: textQuestion and
            # choiceQuestion (radio). Q1-Q4 must be radio, Q5 open text.
            # Q5 keyword pattern must mirror required_keywords[4] above: accept
            # either 'topic ... next' or 'suggest ... topic' phrasing, then
            # require the matching question be an open textQuestion. (Failing
            # to mirror the keyword set would reject a correct 'What topic do
            # you suggest Fireship covers?' Q5.)
            def _q5_matches(t):
                tl = (t or "").lower()
                return (("topic" in tl and "next" in tl)
                        or ("suggest" in tl and "topic" in tl))

            next_text_ok = any(
                _q5_matches(q) and qt == "textQuestion"
                for q, qt in rows
            )
            choice_count = sum(1 for qt in q_types if qt == "choiceQuestion")
            record("GForm radio/text question types match task",
                   next_text_ok and choice_count >= 3,
                   f"choiceQuestion={choice_count}, "
                   f"textQuestion-on-topic-next={next_text_ok}")
        conn.close()
    except Exception as e:
        record("GForm check", False, str(e))


def _as_utc(dt):
    """Coerce a datetime to an absolute instant expressed in UTC."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _tz_candidates(session_tz):
    """Timezone views in which the Community Standup event is accepted.

    The gcal MCP can write the event as an aware UTC timestamp ('...Z'/'+00:00')
    or as a naive local time. A naive timestamp is interpreted by Postgres in
    the *session* timezone, which is not pinned by the container entrypoint and
    depends on the base image (the container default is a real risk). So an
    event is accepted if it reads 18:00-19:00 on 2026-04-01 in any of: UTC
    (covers aware input), the DB session timezone (covers naive input under
    whatever default the container has), and Asia/Shanghai (safety net for this
    lab's most common default).
    """
    views = [timezone.utc]
    if session_tz:
        try:
            views.append(ZoneInfo(session_tz))
        except (ZoneInfoNotFoundError, ValueError, KeyError, TypeError):
            m = re.fullmatch(r"([+-])(\d{1,2})(?::?(\d{2}))?", str(session_tz).strip())
            if m:
                off = int(m.group(2)) * 3600 + int(m.group(3) or 0) * 60
                if m.group(1) == "-":
                    off = -off
                views.append(timezone(timedelta(seconds=off)))
    try:
        views.append(ZoneInfo("Asia/Shanghai"))
    except Exception:
        pass
    return views


def _is_april1(dt, views):
    """True if the instant dt falls on 2026-04-01 in any timezone view."""
    for view in views:
        d = dt.astimezone(view)
        if d.year == 2026 and d.month == 4 and d.day == 1:
            return True
    return False


def _is_standup_window(sd, ed, views):
    """True if the event reads as 18:00-19:00 on 2026-04-01 in any view."""
    for view in views:
        sdt = sd.astimezone(view)
        if not (sdt.year == 2026 and sdt.month == 4 and sdt.day == 1):
            continue
        if sdt.hour != 18 or sdt.minute != 0:
            continue
        if ed is not None:
            edt = ed.astimezone(view)
            if not (edt.year == 2026 and edt.month == 4 and edt.day == 1):
                continue
            if edt.hour != 19 or edt.minute != 0:
                continue
        return True
    return False


def check_gcal():
    print("\n=== Check 6: GCal Community Standup in April 2026 ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("SELECT current_setting('TimeZone')")
            session_tz = cur.fetchone()[0]
            views = _tz_candidates(session_tz)
            cur.execute("""
                SELECT summary, start_datetime, end_datetime FROM gcal.events
                WHERE (summary ILIKE '%Community%' OR summary ILIKE '%Standup%')
                  AND start_datetime >= '2026-04-01 00:00:00+00'
                  AND start_datetime < '2026-05-01 00:00:00+00'
                  AND summary NOT ILIKE '%Q&A%'
                ORDER BY start_datetime
            """)
            events = cur.fetchall()
        conn.close()
        record("GCal has new Community/Standup event in April 2026 (not Q&A noise)",
               len(events) > 0, f"Found {len(events)} events")

        target_evt = None
        for s, sd, ed in events:
            sl = (s or "").lower()
            if "community" in sl and "standup" in sl:
                sd_utc = _as_utc(sd)
                if sd_utc and _is_april1(sd_utc, views):
                    target_evt = (s, sd, ed)
                    break
        record(
            "GCal 'Community Standup' on 2026-04-01 exists",
            target_evt is not None,
            f"Candidates: {[(e[0], e[1]) for e in events]}",
        )
        if target_evt:
            s, sd, ed = target_evt
            sd_utc = _as_utc(sd)
            ed_utc = _as_utc(ed)
            time_ok = sd_utc is not None and _is_standup_window(sd_utc, ed_utc, views)
            record(
                "Community Standup time is 18:00-19:00",
                time_ok,
                f"start={sd}, end={ed}, session_tz={session_tz}",
            )
    except Exception as e:
        record("GCal check", False, str(e))


def check_email():
    print("\n=== Check 7: Email sent to community@devclub.io ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT subject, body_text FROM email.messages
                WHERE to_addr::text ILIKE '%community@devclub.io%'
                  AND from_addr != 'community@devclub.io'
                ORDER BY id DESC
            """)
            rows = cur.fetchall()
        conn.close()
        record("Email sent to community@devclub.io", len(rows) > 0, f"Found {len(rows)}")
        if rows:
            subj_ok = False
            for s, _ in rows:
                sl = (s or "").lower()
                if "engagement" in sl or ("monthly" in sl and "report" in sl):
                    subj_ok = True
                    break
            record("Email subject references engagement/monthly report", subj_ok,
                   f"Subjects: {[r[0] for r in rows]}")
            # Body must reference (a) the survey and (b) the community call on
            # April 1. The task uses the word 'standup', but a correct model may
            # legitimately write 'community call'/'meeting'/'call' instead, so
            # accept those synonyms. Word-boundary matching avoids a bare 'form'
            # matching e.g. 'information'.
            survey_ok = lambda bl: (("survey" in bl) or ("questionnaire" in bl)
                                    or re.search(r"\bform\b", bl) is not None)
            standup_ok = lambda bl: (("standup" in bl) or ("stand up" in bl)
                                     or ("community call" in bl)
                                     or ("community meeting" in bl)
                                     or ("community meetup" in bl)
                                     or re.search(r"\bmeeting\b", bl) is not None
                                     or re.search(r"\bcall\b", bl) is not None)
            body_ok = False
            for _, b in rows:
                bl = (b or "").lower()
                if survey_ok(bl) and standup_ok(bl):
                    body_ok = True
                    break
            record("Email body mentions survey/form + standup/call", body_ok,
                   f"First body excerpt: {(rows[0][1] or '')[:200]}")
    except Exception as e:
        record("Email check", False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print(f"Running evaluation for yt-fireship-gform-survey-excel-gcal")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_gform()
    check_gcal()
    check_email()

    all_passed = FAIL_COUNT == 0
    summary = f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"
    print(f"\n{'='*40}")
    print(f"Result: {'PASS' if all_passed else 'FAIL'} - {summary}")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "all_passed": all_passed}, f)

    return all_passed, summary


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
