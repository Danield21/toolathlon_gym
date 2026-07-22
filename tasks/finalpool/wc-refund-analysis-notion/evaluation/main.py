"""
Evaluation script for wc-refund-analysis-notion task.

Checks:
1. Excel file (Refund_Analysis.xlsx) - refund count matches DB
2. Notion page with 'refund' in title
3. Email with 'refund' in subject sent to cfo@company.com
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail="", db=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        if db:
            PASS_COUNT += 1
        else:
            PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            FAIL_COUNT += 1
        else:
            FAIL_COUNT += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def get_expected_refunds():
    """Get expected refund data from wc.refunds."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, order_id, amount, reason, date_created FROM wc.refunds ORDER BY id")
    refunds = cur.fetchall()
    cur.execute("SELECT COUNT(*) FROM wc.orders")
    order_count = cur.fetchone()[0]
    cur.close()
    conn.close()
    return refunds, order_count


def check_excel(agent_workspace, expected_refunds, order_count):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Refund_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return False

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # Check Refund Details sheet
    ws = None
    for s in wb.sheetnames:
        if "refund" in s.lower() and "detail" in s.lower():
            ws = wb[s]
            break
    if ws is None:
        for s in wb.sheetnames:
            if "refund" in s.lower():
                ws = wb[s]
                break
        if ws is None and len(wb.sheetnames) > 0:
            ws = wb[wb.sheetnames[0]]

    check("Sheet with refund details exists", ws is not None, f"Sheets: {wb.sheetnames}")
    if ws is None:
        return False

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    expected_count = len(expected_refunds)
    check(f"Refund count matches DB ({expected_count})",
          len(rows) == expected_count,
          f"Got {len(rows)} rows, expected {expected_count}")

    # Check Summary sheet
    ws2 = None
    for s in wb.sheetnames:
        if "summary" in s.lower():
            ws2 = wb[s]
            break
    check("Summary sheet exists", ws2 is not None, f"Sheets: {wb.sheetnames}")

    if ws2:
        summary_data = {}
        for row in ws2.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                key = str(row[0]).strip().lower().replace(" ", "_")
                summary_data[key] = row[1]

        # Verify all 4 required Summary metrics present and numeric
        for required_key in ["total_refunds", "total_refund_amount", "avg_refund_amount", "refund_rate"]:
            present = required_key in summary_data
            check(f"Summary has '{required_key}' entry", present,
                  f"Keys: {list(summary_data.keys())}")
            if present:
                val = summary_data[required_key]
                try:
                    float(str(val).replace(",", "").replace("%", "").replace("$", ""))
                    is_num = True
                except (TypeError, ValueError):
                    is_num = False
                check(f"Summary '{required_key}' is numeric", is_num, f"Got: {val}")

        # Verify expected counts
        if "total_refunds" in summary_data:
            try:
                actual = int(float(str(summary_data["total_refunds"]).replace(",", "")))
                check(f"Summary Total_Refunds == {expected_count}",
                      actual == expected_count,
                      f"Got {actual}")
            except (TypeError, ValueError):
                pass

    return True


def check_notion():
    print("\n=== Checking Notion ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Tightened: page title must contain 'Refund Analysis Dashboard'
    cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false AND in_trash = false")
    pages = cur.fetchall()

    dashboard_pages = []
    for pid, props in pages:
        if props:
            props_str = json.dumps(props).lower() if isinstance(props, dict) else str(props).lower()
            if "refund analysis dashboard" in props_str:
                dashboard_pages.append(pid)

    check("Notion page 'Refund Analysis Dashboard' exists",
          len(dashboard_pages) > 0,
          f"Found {len(pages)} pages, {len(dashboard_pages)} with title 'Refund Analysis Dashboard'")

    # Verify 'Refund Trends' database exists
    # Notion databases live in notion.databases or as page-children with type=database
    try:
        cur.execute("""
            SELECT id, title FROM notion.databases
            WHERE archived = false
        """)
        databases = cur.fetchall()
        rt_dbs = []
        for did, title in databases:
            title_str = json.dumps(title).lower() if isinstance(title, (dict, list)) else (title or "").lower()
            if "refund trends" in title_str:
                rt_dbs.append(did)
        check("Notion database 'Refund Trends' exists", len(rt_dbs) > 0,
              f"Found {len(databases)} databases, {len(rt_dbs)} with 'Refund Trends'")
        if rt_dbs:
            # Each database should have at least one entry (page child of the database)
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE archived = false AND in_trash = false AND parent::text ILIKE %s
            """, (f"%{rt_dbs[0]}%",))
            entry_count = cur.fetchone()[0]
            check("'Refund Trends' database has >=1 entry", entry_count >= 1,
                  f"entries={entry_count}")
    except Exception as e:
        # Schema may differ; treat missing databases-table as soft fail
        check("Notion database query ran", False, str(e))

    cur.close()
    conn.close()
    return len(dashboard_pages) > 0


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_email] Found {len(all_emails)} emails.")

    matched = None
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                if isinstance(parsed, list):
                    to_str = " ".join(str(r).lower() for r in parsed)
                else:
                    to_str = str(to_addr).lower()
            except (json.JSONDecodeError, TypeError):
                to_str = str(to_addr).lower()
        from_str = str(from_addr or "").lower()

        if "cfo@company.com" in to_str:
            matched = (subject, from_str, body_text)
            break

    check("Email sent to cfo@company.com", matched is not None)
    if matched:
        subject, from_str, body_text = matched
        # Tightened: from finance@store.com
        check("Email from 'finance@store.com'",
              "finance@store.com" in from_str,
              f"From: {from_str}")
        # Tightened: exact subject 'Monthly Refund Analysis Report'
        sl = (subject or "").lower()
        check("Email subject equals 'Monthly Refund Analysis Report'",
              "monthly refund analysis report" in sl or
              ("monthly" in sl and "refund" in sl and "analysis" in sl and "report" in sl),
              f"Subject: {subject}")
        # Body must mention 4 key items: total count, total amount, avg amount, most common reason
        body = (body_text or "").lower()
        body_terms = {"total": "total" in body, "amount": "amount" in body,
                      "average": ("average" in body) or ("avg" in body) or ("mean" in body),
                      "reason": "reason" in body}
        for term, present in body_terms.items():
            check(f"Email body mentions '{term}'", present, f"body sample: {body[:200]}")
    return matched is not None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_refunds, order_count = get_expected_refunds()
    print(f"Expected {len(expected_refunds)} refunds, {order_count} orders")

    check_excel(args.agent_workspace, expected_refunds, order_count)
    check_notion()
    check_email()

    total_pass = PASS_COUNT
    total_fail = FAIL_COUNT
    all_ok = FAIL_COUNT == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if args.res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": all_ok}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
