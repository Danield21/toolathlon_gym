"""
Evaluation for scholarly-grant-proposal-prep task.
Checks Excel, emails sent, and calendar events.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

AGENCIES = ["NSF", "DARPA", "NIH"]
COLLABORATOR_EMAILS = ["jpark@mit.edu", "sthompson@stanford.edu", "mchen@mayo.edu", "lwang@berkeley.edu"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Grant_Prep.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Funding Opportunities
    fo_rows = load_sheet_rows(wb, "Funding Opportunities") or load_sheet_rows(wb, "Funding_Opportunities")
    if fo_rows is None:
        record("Sheet 'Funding Opportunities' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Funding Opportunities' exists", True)
        data = fo_rows[1:]
        record("Funding Opportunities has 3 rows", len(data) == 3, f"Found {len(data)}")

        agency_col = find_col(fo_rows[0], ["Agency", "agency"])
        if agency_col is not None:
            agencies = {str(r[agency_col]).strip() for r in data if agency_col < len(r) and r[agency_col]}
            for a in AGENCIES:
                record(f"Agency '{a}' present", a in agencies, f"Found: {agencies}")

        score_col = find_col(fo_rows[0], ["Relevance_Score", "Relevance Score", "Score"])
        if score_col is not None:
            for r in data:
                if score_col < len(r) and r[score_col] is not None:
                    try:
                        s = float(r[score_col])
                        record(f"Score {s} in range 1-10", 1 <= s <= 10, f"Got {s}")
                    except (TypeError, ValueError):
                        record("Score is numeric", False, f"Got {r[score_col]}")

    # Supporting Literature
    sl_rows = load_sheet_rows(wb, "Supporting Literature") or load_sheet_rows(wb, "Supporting_Literature")
    if sl_rows is None:
        record("Sheet 'Supporting Literature' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Supporting Literature' exists", True)
        data = [r for r in sl_rows[1:] if any(c is not None for c in r)]
        record("Supporting Literature has exactly 5 rows", len(data) == 5, f"Found {len(data)}")

    # Collaborator Matrix
    cm_rows = load_sheet_rows(wb, "Collaborator Matrix") or load_sheet_rows(wb, "Collaborator_Matrix")
    if cm_rows is None:
        record("Sheet 'Collaborator Matrix' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Collaborator Matrix' exists", True)
        data = cm_rows[1:]
        record("Collaborator Matrix has 4 rows", len(data) == 4, f"Found {len(data)}")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Grant_Prep.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]; break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            # Iterate ALL gt rows but match by first-column key (row order may differ for paper lists)
            gt_keys = {}
            for r in gt_rows:
                k = str(r[0]).strip().lower() if r and r[0] else None
                if k:
                    gt_keys[k] = r
            a_keys = {}
            for r in a_rows:
                k = str(r[0]).strip().lower() if r and r[0] else None
                if k:
                    a_keys[k] = r
            # Validate each GT row exists in agent file
            # Determine column-name mapping for tighter type-aware checks.
            gt_header = list(gt_ws[1])
            gt_header_names = [str(c.value).strip().lower() if c.value else ""
                               for c in gt_header]
            for k, gt_row in gt_keys.items():
                a_row = None
                if k in a_keys:
                    a_row = a_keys[k]
                else:
                    # Fallback: substring/prefix match for paper title keys.
                    # This handles cases where the agent's title is a prefix
                    # of GT title (or vice versa).
                    for ak, av_row in a_keys.items():
                        if ak == k:
                            a_row = av_row
                            break
                        # 80%+ overlap by shared prefix length / shorter length
                        shorter = min(len(ak), len(k))
                        if shorter >= 20:
                            if ak.startswith(k[:shorter]) or k.startswith(ak[:shorter]):
                                a_row = av_row
                                break
                if a_row is None:
                    record(f"GT '{gt_sname}' row '{k[:50]}' present", False, f"missing")
                    continue
                ok = True
                fail_reason = ""
                for ci in range(min(len(gt_row), len(a_row))):
                    gv, av = gt_row[ci], a_row[ci]
                    col_name = gt_header_names[ci] if ci < len(gt_header_names) else ""
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        # Year: strict integer equality.
                        if "year" in col_name:
                            try:
                                if int(round(float(av))) != int(round(float(gv))):
                                    ok = False
                                    fail_reason = f"col={col_name} got={av} expected={gv}"
                                    break
                            except (TypeError, ValueError):
                                ok = False
                                fail_reason = f"col={col_name} non-numeric av={av}"
                                break
                        elif "citation" in col_name:
                            # Citations: tighter 5% tolerance, min abs 5.
                            if not num_close(av, gv, max(abs(gv) * 0.05, 5.0)):
                                ok = False
                                fail_reason = f"col={col_name} got={av} expected={gv}"
                                break
                        elif "score" in col_name or "relevance" in col_name:
                            # Relevance_Score 1-10 — accept +/-1 (subjective grading).
                            if not num_close(av, gv, 1.0):
                                ok = False
                                fail_reason = f"col={col_name} got={av} expected={gv}"
                                break
                        else:
                            if not num_close(av, gv, max(abs(gv) * 0.05, 1.0)):
                                ok = False
                                fail_reason = f"col={col_name} got={av} expected={gv}"
                                break
                    else:
                        # String columns: strict equality (case/whitespace
                        # insensitive). NO bidirectional substring fallback —
                        # EXCEPT for "authors" which is free-form free text and
                        # may be formatted as "First Last, First Last",
                        # "Last, F.", "F. Last and F. Last", etc. We accept any
                        # format that mentions every last name from GT.
                        if "author" in col_name:
                            gv_str = str(gv).lower()
                            av_str = str(av).lower() if av is not None else ""
                            # Extract last-name tokens from GT (split by ',' or 'and')
                            import re as _re
                            tokens = [t.strip() for t in _re.split(r"[,;]| and ", gv_str) if t.strip()]
                            # last word of each token = last name
                            last_names = [t.split()[-1] if t.split() else t for t in tokens]
                            missing = [ln for ln in last_names if ln and ln not in av_str]
                            if missing:
                                ok = False
                                fail_reason = (
                                    f"col={col_name} missing surnames {missing}; "
                                    f"got={av!r} expected={gv!r}"
                                )
                                break
                        elif not str_match(av, gv):
                            ok = False
                            fail_reason = f"col={col_name} got={av!r} expected={gv!r}"
                            break
                record(f"GT '{gt_sname}' row '{k[:50]}' values", ok,
                       fail_reason or f"gt={gt_row[:4]}, agent={a_row[:4]}")
        gt_wb.close()

    return True


def check_emails():
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%%grant%%' OR subject ILIKE '%%collaboration%%'
           OR subject ILIKE '%%nsf%%' OR subject ILIKE '%%darpa%%' OR subject ILIKE '%%nih%%'
           OR subject ILIKE '%%proposal%%' OR subject ILIKE '%%funding%%'
    """)
    emails = cur.fetchall()

    record("Grant-related emails sent", len(emails) >= 4, f"Found {len(emails)} emails (need >=4)")

    if emails:
        all_to = []
        all_from = []
        all_body = []
        per_email_to = []
        for e in emails:
            to = e[3]
            if isinstance(to, str):
                try:
                    to = json.loads(to)
                except Exception:
                    pass
            all_to.append(str(to).lower())
            per_email_to.append(str(to).lower())
            all_from.append(str(e[2]).lower() if e[2] else "")
            all_body.append(str(e[4]).lower() if e[4] else "")

        # Per-collaborator: each collaborator address must appear in at least
        # one to_addr, and ideally each in a distinct email.
        addr_in_email_count = {}
        for addr in COLLABORATOR_EMAILS:
            count = sum(1 for to_str in per_email_to if addr in to_str)
            addr_in_email_count[addr] = count
            record(f"Email to collaborator {addr}", count >= 1,
                   f"appears in {count} emails; expected >=1")

        # Validate FROM address
        has_pi_from = any("pi@university.edu" in f for f in all_from)
        record("Emails sent from pi@university.edu", has_pi_from,
               f"From addresses: {all_from[:5]}")

        # Validate at least one body mentions a grant program name
        program_names = ["ai research institutes", "explainable ai", "ai for health"]
        body_concat = " ".join(all_body)
        mentioned = sum(1 for p in program_names if p in body_concat)
        record(f"Email bodies mention grant program names ({mentioned}/3)", mentioned >= 2,
               f"Need >=2 program names in bodies")

    cur.close()
    conn.close()
    return True


def check_calendar():
    print("\n=== Checking Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, start_datetime, end_datetime FROM gcal.events")
    events = cur.fetchall()

    # Task says one deadline event per grant + one writing event per grant.
    # GT injects 2 noise events ('Team Standup', 'Lunch Break') so we expect
    # exactly 6 deadline+writing events related to grants. We require >=6 total
    # (allowing noise) and >=3 deadlines + >=3 writing sessions.
    record("Calendar events created (total >=6 with noise allowed)",
           len(events) >= 6, f"Found {len(events)} events")

    if events:
        summaries = [str(e[1]).lower() for e in events if e[1]]
        # Deadline events: per-agency, with the agency name in the summary.
        per_agency_deadline = {}
        per_agency_writing = {}
        # Map agency to expected deadline (from funding.json / GT)
        expected_deadlines = {
            "nsf": "2026-06-15",
            "darpa": "2026-05-01",
            "nih": "2026-07-01",
        }
        # Writing sessions are 14 days before deadline.
        from datetime import datetime as _dt, timedelta as _td
        expected_writing = {
            agc: (_dt.strptime(d, "%Y-%m-%d") - _td(days=14)).strftime("%Y-%m-%d")
            for agc, d in expected_deadlines.items()
        }
        for e in events:
            if not e[1]:
                continue
            s = str(e[1]).lower()
            sd = str(e[2]) if e[2] else ""
            for agc in ("nsf", "darpa", "nih"):
                if agc in s:
                    if "writing" in s or "proposal writing" in s:
                        per_agency_writing.setdefault(agc, []).append((s, sd))
                    else:
                        per_agency_deadline.setdefault(agc, []).append((s, sd))

        for agc, exp_date in expected_deadlines.items():
            evs = per_agency_deadline.get(agc, [])
            record(f"Deadline event for {agc.upper()}",
                   any(exp_date in sd for _, sd in evs),
                   f"events for {agc}: {[(s, sd) for s, sd in evs][:3]}; expected date {exp_date}")
        for agc, exp_date in expected_writing.items():
            evs = per_agency_writing.get(agc, [])
            record(f"Writing session for {agc.upper()} on {exp_date}",
                   any(exp_date in sd for _, sd in evs),
                   f"events for {agc}: {[(s, sd) for s, sd in evs][:3]}; expected {exp_date}")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_emails()
    check_calendar()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
