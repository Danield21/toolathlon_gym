"""Evaluation for sf-hr-salary-overview."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_header(h):
    return str(h or "").strip().lower().replace(" ", "_")


def _header_map(rows):
    """Map normalized header text -> 0-based column index from row 1 of a sheet.
    Lets checks locate columns by name regardless of column order."""
    if not rows or not rows[0]:
        return {}
    return {_norm_header(h): i for i, h in enumerate(rows[0]) if _norm_header(h)}


def _cell(row, idx):
    """Safe cell access (None when idx out of range)."""
    if row is None or idx is None:
        return None
    return row[idx] if idx < len(row) else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Salary_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Salary_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Department Analysis
    print(f"  Checking Department Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Department Analysis")
    g_rows = load_sheet_rows(gt_wb, "Department Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Department Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Department Analysis' not found in groundtruth")
    else:
        sheet_name = "Department Analysis"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Locate columns by header name (order-independent).
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        if len(a_data) != len(g_data):
            errors.append(f"{sheet_name} row count: agent={len(a_data)}, expected={len(g_data)}")
        a_lookup = {}
        for row in a_data:
            # Row key is the Department column (column 0 by convention).
            dept = _cell(row, 0)
            if dept is not None:
                a_lookup[str(dept).strip().lower()] = row
        for g_row in g_data:
            g_dept = _cell(g_row, 0)
            if g_dept is None:
                continue
            key = str(g_dept).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_dept}")
                continue

            # Employee_Count is integer — exact equality (tol=0 preserved)
            av = _cell(a_row, a_col.get("employee_count"))
            gv = _cell(g_row, g_col.get("employee_count"))
            if av is not None and gv is not None and not num_close(av, gv, 0):
                errors.append(f"{key}.Employee_Count: {av} vs {gv} (exact)")

            av = _cell(a_row, a_col.get("avg_salary"))
            gv = _cell(g_row, g_col.get("avg_salary"))
            if av is not None and gv is not None and not num_close(av, gv, 0.5):
                errors.append(f"{key}.Avg_Salary: {av} vs {gv} (tol=0.5)")

            av = _cell(a_row, a_col.get("min_salary"))
            gv = _cell(g_row, g_col.get("min_salary"))
            if av is not None and gv is not None and not num_close(av, gv, 1.0):
                errors.append(f"{key}.Min_Salary: {av} vs {gv} (tol=1.0)")

            av = _cell(a_row, a_col.get("max_salary"))
            gv = _cell(g_row, g_col.get("max_salary"))
            if av is not None and gv is not None and not num_close(av, gv, 1.0):
                errors.append(f"{key}.Max_Salary: {av} vs {gv} (tol=1.0)")

            av = _cell(a_row, a_col.get("benchmark"))
            gv = _cell(g_row, g_col.get("benchmark"))
            if av is not None and gv is not None and not num_close(av, gv, 1.0):
                errors.append(f"{key}.Benchmark: {av} vs {gv} (tol=1.0)")

            av = _cell(a_row, a_col.get("variance"))
            gv = _cell(g_row, g_col.get("variance"))
            if av is not None and gv is not None and not num_close(av, gv, 1.0):
                errors.append(f"{key}.Variance: {av} vs {gv} (tol=1.0)")

            av = _cell(a_row, a_col.get("variance_pct"))
            gv = _cell(g_row, g_col.get("variance_pct"))
            if av is not None and gv is not None and not num_close(av, gv, 0.1):
                errors.append(f"{key}.Variance_Pct: {av} vs {gv} (tol=0.1)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Locate columns by header name (order-independent).
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        a_lookup = {}
        for row in a_data:
            # Row key is the Metric column (column 0 by convention).
            metric = _cell(row, 0)
            if metric is not None:
                a_lookup[str(metric).strip().lower()] = row
        for g_row in g_data:
            g_metric = _cell(g_row, 0)
            if g_metric is None:
                continue
            key = str(g_metric).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_metric}")
                continue

            av = _cell(a_row, a_col.get("value"))
            gv = _cell(g_row, g_col.get("value"))
            if av is not None and gv is not None:
                # Per-metric tolerance: counts exact, others 0.5 abs
                key_l = key.lower()
                is_count = (
                    "count" in key_l
                    or "total_employees" in key_l
                    or "departments_above" in key_l
                    or "departments_below" in key_l
                )
                tol = 0 if is_count else 0.5
                if not num_close(av, gv, tol):
                    errors.append(f"{key}.Value: {av} vs {gv} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
