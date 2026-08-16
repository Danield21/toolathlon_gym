"""
Evaluation for yt-fireship-canvas-quiz-excel-gcal task.

Checks:
1. TypeScript_Course_Tracker.xlsx exists with "Videos" sheet having >= 4 rows
2. "Quiz_Questions" sheet exists with >= 8 rows
3. "Grade_Template" sheet exists
4. Canvas quiz exists (canvas.quizzes WHERE title ILIKE '%TypeScript%')
5. Canvas quiz has >= 5 questions (canvas.quiz_questions)
6. GCal has TypeScript/Quiz event in March/April 2026
7. Email sent to students@webdev.edu
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
from zoneinfo import ZoneInfo

# ──────────────────────────────────────────────────────────────────────────
# EVALUATION GROUND TRUTH SPEC (gcal tz root-fix v3, case-study 2026-08-13)
# gcal.events.start_datetime / end_datetime are TIMESTAMPTZ storing the UTC
# instant. psycopg2 returns a tz-aware datetime whose display tz follows the
# PG session TimeZone (case-study: compute node default Asia/Shanghai), so
# bare sd.hour / sd.date() silently compares against the wrong wall-clock.
# Use utils.evaluation.gcal_helpers (session-tz-independent, DST-aware).
# ──────────────────────────────────────────────────────────────────────────
# task.md line 9: TypeScript Quiz Deadline on March 25, 2026 (closes 23:59)
# and TypeScript Quiz Review Session on March 26, 2026 from 14:00 to 15:00
# (university / course context → America/New_York / ET).
from utils.evaluation.gcal_helpers import get_zone_components  # noqa: E402

EXPECTED_TIMEZONE = ZoneInfo("America/New_York")
EXPECTED_DEADLINE_DATE = "2026-03-25"
EXPECTED_REVIEW_DATE = "2026-03-26"
EXPECTED_REVIEW_HOUR = 14
EXPECTED_REVIEW_MINUTE = 0
EXPECTED_REVIEW_END_HOUR = 15
EXPECTED_REVIEW_END_MINUTE = 0

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-3: TypeScript_Course_Tracker.xlsx ===")
    # Require exact filename TypeScript_Course_Tracker.xlsx (case-insensitive); do NOT
    # fall back to ANY .xlsx — that would let an agent name any file and pass.
    xlsx_path = None
    for fname in os.listdir(agent_workspace):
        if fname.strip().lower() == "typescript_course_tracker.xlsx":
            xlsx_path = os.path.join(agent_workspace, fname)
            break
    if xlsx_path is None:
        # As a final permissive fallback, accept any .xlsx whose name STARTS with
        # 'typescript_course_tracker' (e.g. _v2, _final).
        for fname in os.listdir(agent_workspace):
            if fname.lower().startswith("typescript_course_tracker") and fname.lower().endswith(".xlsx"):
                xlsx_path = os.path.join(agent_workspace, fname)
                break

    record("TypeScript_Course_Tracker.xlsx exists", xlsx_path is not None,
           f"No file named TypeScript_Course_Tracker.xlsx in {agent_workspace}")

    if not xlsx_path:
        for chk in ["Videos sheet has >= 4 rows", "Quiz_Questions sheet has >= 8 rows", "Grade_Template sheet exists"]:
            record(chk, False, "xlsx not found")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)

        # Videos sheet
        videos_sheet = None
        for name in wb.sheetnames:
            if "video" in name.lower():
                videos_sheet = wb[name]
                break
        if not videos_sheet and wb.sheetnames:
            videos_sheet = wb[wb.sheetnames[0]]

        if videos_sheet:
            data_rows = [r for r in videos_sheet.iter_rows(min_row=2, values_only=True)
                         if any(c is not None for c in r)]
            # Real DB has 4 distinct TypeScript-related Fireship videos. Accept 4-5 rows
            # since the agent might include borderline matches to reach the requested 5.
            record("Videos sheet has 4-5 data rows", 4 <= len(data_rows) <= 5,
                   f"Found {len(data_rows)} rows, expected 4-5")
        else:
            record("Videos sheet has 4-5 data rows", False, "No Videos sheet")

        # Quiz_Questions sheet
        qsheet = None
        for name in wb.sheetnames:
            if "quiz" in name.lower() or "question" in name.lower():
                qsheet = wb[name]
                break

        record("Quiz_Questions sheet exists", qsheet is not None, f"Sheets: {wb.sheetnames}")

        if qsheet:
            data_rows = [r for r in qsheet.iter_rows(min_row=2, values_only=True)
                         if any(c is not None for c in r)]
            # 2 questions per video; if 4-5 videos selected, expect 8-10 rows
            record("Quiz_Questions sheet has 8-10 data rows",
                   8 <= len(data_rows) <= 10,
                   f"Found {len(data_rows)} rows, expected 8-10")
        else:
            record("Quiz_Questions sheet has 8-10 data rows", False, "Sheet not found")

        # Grade_Template sheet
        grade_sheet = None
        for name in wb.sheetnames:
            if "grade" in name.lower() or "template" in name.lower():
                grade_sheet = wb[name]
                break
        record("Grade_Template sheet exists", grade_sheet is not None,
               f"Sheets: {wb.sheetnames}")

        # --- Structural / topic-based comparison instead of exact Video_ID match ---
        # The synthetic Video_IDs in GT (ts100s, tsbegin, ...) do NOT exist in the real
        # video database. An agent that correctly fetches videos will get real IDs like
        # 'PQ2WjtaPfXU'. So we cannot verify exact Video_IDs — instead verify schema
        # and content patterns that any correct agent must produce.
        gt_path = os.path.join(groundtruth_workspace, "TypeScript_Course_Tracker.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = wb[asn]; break
                if a_ws is None:
                    record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                           f"Available: {wb.sheetnames}")
                    continue

                gt_headers = [str(c.value).strip().lower() if c.value else ""
                              for c in gt_ws[1]]
                a_headers = [str(c.value).strip().lower() if c.value else ""
                             for c in a_ws[1]]
                # Headers must match (case-insensitive, .strip())
                missing_headers = [h for h in gt_headers if h and h not in a_headers]
                record(f"GT '{gt_sname}' headers match",
                       len(missing_headers) == 0,
                       f"Missing: {missing_headers}; agent headers: {a_headers}")

                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                           if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                          if any(c is not None for c in r)]
                record(f"GT '{gt_sname}' row count matches",
                       len(a_rows) == len(gt_rows),
                       f"Expected {len(gt_rows)}, got {len(a_rows)}")

                # For Quiz_Questions: validate Question_No is 1..N and Points are 5
                if gt_sname.lower() == "quiz_questions" and a_rows:
                    qno_idx = a_headers.index("question_no") if "question_no" in a_headers else 0
                    points_idx = a_headers.index("points") if "points" in a_headers else None
                    qnos = sorted([r[qno_idx] for r in a_rows if qno_idx < len(r) and r[qno_idx] is not None])
                    n = len(a_rows)
                    expected = list(range(1, n + 1))
                    qno_ok = (qnos == expected) or (qnos == [str(i) for i in expected])
                    record(f"Quiz_Questions Question_No values are 1..{n}", qno_ok,
                           f"Got Question_Nos: {qnos}")
                    if points_idx is not None:
                        pts_ok = all(num_close(r[points_idx], 5, 0.5) for r in a_rows
                                     if points_idx < len(r) and r[points_idx] is not None)
                        record("Quiz_Questions all rows have Points=5", pts_ok)

                # For Videos: validate Concepts_Count==3, Quiz_Question_Count==2 per row
                if gt_sname.lower() == "videos" and a_rows:
                    cc_idx = a_headers.index("concepts_count") if "concepts_count" in a_headers else None
                    qqc_idx = a_headers.index("quiz_question_count") if "quiz_question_count" in a_headers else None
                    title_idx = a_headers.index("title") if "title" in a_headers else None
                    if cc_idx is not None:
                        cc_ok = all(num_close(r[cc_idx], 3, 0.5) for r in a_rows
                                    if cc_idx < len(r) and r[cc_idx] is not None)
                        record("Videos Concepts_Count==3 for all rows", cc_ok)
                    if qqc_idx is not None:
                        qqc_ok = all(num_close(r[qqc_idx], 2, 0.5) for r in a_rows
                                     if qqc_idx < len(r) and r[qqc_idx] is not None)
                        record("Videos Quiz_Question_Count==2 for all rows", qqc_ok)
                    if title_idx is not None:
                        # At least one row must have a TypeScript-related title
                        ts_titles = sum(1 for r in a_rows
                                        if title_idx < len(r) and r[title_idx]
                                        and "typescript" in str(r[title_idx]).lower())
                        record("Videos sheet contains TypeScript-related titles",
                               ts_titles >= 1,
                               f"TypeScript title count: {ts_titles}")

                # For Quiz_Questions: cross-reference Video_Source values back to Videos sheet
                if gt_sname.lower() == "quiz_questions" and a_rows:
                    vs_idx = a_headers.index("video_source") if "video_source" in a_headers else None
                    if vs_idx is not None:
                        # Build the set of video IDs from the Videos sheet, if present
                        video_ids = set()
                        for asn in wb.sheetnames:
                            if asn.strip().lower() == "videos":
                                v_ws = wb[asn]
                                v_headers = [str(c.value).strip().lower() if c.value else ""
                                             for c in v_ws[1]]
                                v_id_idx = v_headers.index("video_id") if "video_id" in v_headers else 0
                                for vr in v_ws.iter_rows(min_row=2, values_only=True):
                                    if v_id_idx < len(vr) and vr[v_id_idx] is not None:
                                        video_ids.add(str(vr[v_id_idx]).strip())
                                break
                        if video_ids:
                            sources = [str(r[vs_idx]).strip() for r in a_rows
                                       if vs_idx < len(r) and r[vs_idx] is not None]
                            ok = all(s in video_ids for s in sources)
                            record("Quiz_Questions.Video_Source all reference Videos.Video_ID",
                                   ok,
                                   f"Sources: {sources[:5]}, video_ids: {list(video_ids)[:5]}")
            gt_wb.close()

    except Exception as e:
        for chk in ["Videos sheet has >= 4 rows", "Quiz_Questions sheet has >= 8 rows", "Grade_Template sheet exists"]:
            record(chk, False, str(e))


def check_canvas():
    print("\n=== Check 4-5: Canvas Quiz ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, title FROM canvas.quizzes
            WHERE title ILIKE '%typescript%' OR title ILIKE '%type script%'
            ORDER BY id DESC LIMIT 1
        """)
        quiz = cur.fetchone()
        record("Canvas quiz with TypeScript in title exists", quiz is not None,
               "No TypeScript quiz found in canvas.quizzes")

        if quiz:
            quiz_id = quiz[0]
            cur.execute("""
                SELECT COUNT(*) FROM canvas.quiz_questions WHERE quiz_id = %s
            """, (quiz_id,))
            q_count = cur.fetchone()[0]
            record("Canvas quiz has 8-10 questions", 8 <= q_count <= 10,
                   f"Found {q_count} questions for quiz {quiz_id}, expected 8-10")
        else:
            record("Canvas quiz has 8-10 questions", False, "No quiz found")
    except Exception as e:
        record("Canvas quiz with TypeScript in title exists", False, str(e))
        record("Canvas quiz has >= 5 questions", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 6: GCal TypeScript Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-01' AND start_datetime < '2026-05-01'
        AND (summary ILIKE '%typescript%' OR summary ILIKE '%quiz%')
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    # Find deadline event on March 25 — must contain TypeScript AND Deadline
    deadline_evt = None
    review_evt = None
    for s, sd, ed in events:
        sl = (s or "").lower()
        if sd is None: continue
        # gcal.events.start_datetime is TIMESTAMPTZ (UTC instant); bare
        # sd.date() / sd.hour / sd.minute follow the PG session TimeZone
        # (case-study: Asia/Shanghai), so use the session-tz-independent helper
        # anchored to task.md's ET (university course hours).
        sd_date, _, _ = get_zone_components(sd, EXPECTED_TIMEZONE)
        if sd_date is None:
            continue
        if sd_date.isoformat() == EXPECTED_DEADLINE_DATE and "typescript" in sl and "deadline" in sl:
            deadline_evt = (s, sd, ed)
        # Review session — must contain TypeScript AND ('review' or 'review session')
        if sd_date.isoformat() == EXPECTED_REVIEW_DATE and "typescript" in sl and "review" in sl:
            review_evt = (s, sd, ed)

    record("GCal has TypeScript Quiz Deadline event on 2026-03-25",
           deadline_evt is not None,
           f"Found {len(events)} matching events, summaries: {[e[0] for e in events]}")
    record("GCal has TypeScript Quiz Review Session event on 2026-03-26",
           review_evt is not None,
           f"Found {len(events)} matching events, summaries: {[e[0] for e in events]}")
    if review_evt is not None:
        sd, ed = review_evt[1], review_evt[2]
        _, sd_hour, sd_minute = get_zone_components(sd, EXPECTED_TIMEZONE)
        _, ed_hour, ed_minute = get_zone_components(ed, EXPECTED_TIMEZONE)
        time_ok = (sd_hour == EXPECTED_REVIEW_HOUR and sd_minute == EXPECTED_REVIEW_MINUTE
                   and ed_hour == EXPECTED_REVIEW_END_HOUR and ed_minute == EXPECTED_REVIEW_END_MINUTE)
        record("Review session start=14:00 end=15:00", time_ok,
               f"start={sd}, end={ed}")


def check_email():
    print("\n=== Check 7: Email to students@webdev.edu ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else to_addr.lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "students@webdev.edu" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    record("Email sent to students@webdev.edu", matching is not None,
           f"Total messages: {len(messages)}")
    if matching:
        subj = (matching[0] or "").lower()
        body = matching[3] or ""
        # Subject must contain "TypeScript Essentials Quiz" and indicate availability
        subj_ok = ("typescript essentials quiz" in subj) and ("available" in subj or "now" in subj)
        record("Email subject 'TypeScript Essentials Quiz Now Available'", subj_ok, f"Subject: {matching[0]}")
        # Body must mention BOTH the deadline date AND the closing time;
        # AND BOTH the review date AND the review time.
        body_lower = body.lower()
        deadline_date = ("march 25" in body_lower or "03-25" in body_lower
                         or "3/25" in body_lower or "march 25th" in body_lower
                         or "2026-03-25" in body_lower)
        deadline_time = "23:59" in body_lower
        review_date = ("march 26" in body_lower or "03-26" in body_lower
                       or "3/26" in body_lower or "march 26th" in body_lower
                       or "2026-03-26" in body_lower)
        review_time = "14:00" in body_lower or "2:00 pm" in body_lower or "2 pm" in body_lower
        review_end = "15:00" in body_lower or "3:00 pm" in body_lower or "3 pm" in body_lower
        record("Email body mentions deadline date (Mar 25)", deadline_date,
               f"Body excerpt: {body[:200]}")
        record("Email body mentions deadline time (23:59)", deadline_time,
               f"Body excerpt: {body[:200]}")
        record("Email body mentions review date (Mar 26)", review_date,
               f"Body excerpt: {body[:200]}")
        record("Email body mentions review start time (14:00)", review_time,
               f"Body excerpt: {body[:200]}")
        record("Email body mentions review end time (15:00)", review_end,
               f"Body excerpt: {body[:200]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_canvas()
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL ({FAIL_COUNT} failed checks)")
        sys.exit(1)


if __name__ == "__main__":
    main()
