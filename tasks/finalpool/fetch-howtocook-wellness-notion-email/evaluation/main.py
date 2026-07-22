"""Evaluation script for fetch-howtocook-wellness-notion-email."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-howtocook-wellness-notion-email"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Wellness_Report.xlsx")
    check("Wellness_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Wellness_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames and gt_wb is not None:
            ws = wb["Data_Analysis"]
            gt_ws = gt_wb["Data_Analysis"] if "Data_Analysis" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Data_Analysis has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}, gt={len(gt_rows)}")
            check_columns('Data_Analysis', ['Recipe', 'Calories', 'Meets_Guidelines'],
                          len(gt_rows) if gt_ws is not None else 5)
            # Validate Data_Analysis structure (recipes are agent's choice; values are agent-derived)
            if gt_ws is not None:
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                header_map = {h: i for i, h in enumerate(headers)}
                agent_recipe_idx = header_map.get("recipe", 0)
                agent_calories_idx = header_map.get("calories")
                agent_meets_idx = header_map.get("meets_guidelines")
                rows_with_recipe = 0
                rows_with_calories = 0
                rows_with_meets = 0
                for r in data_rows:
                    if agent_recipe_idx is not None and agent_recipe_idx < len(r) and r[agent_recipe_idx] and str(r[agent_recipe_idx]).strip():
                        rows_with_recipe += 1
                    if agent_calories_idx is not None and agent_calories_idx < len(r):
                        cf = safe_float(r[agent_calories_idx])
                        if cf is not None and cf > 0:
                            rows_with_calories += 1
                    if agent_meets_idx is not None and agent_meets_idx < len(r) and r[agent_meets_idx]:
                        mv = str(r[agent_meets_idx]).strip().lower()
                        if mv in ("yes", "no", "true", "false", "1", "0"):
                            rows_with_meets += 1
                check("Data_Analysis has non-empty Recipe values",
                      rows_with_recipe >= len(gt_rows), f"{rows_with_recipe}/{len(gt_rows)}")
                check("Data_Analysis has positive Calories values",
                      rows_with_calories >= len(gt_rows), f"{rows_with_calories}/{len(gt_rows)}")
                check("Data_Analysis has Meets_Guidelines yes/no values",
                      rows_with_meets >= len(gt_rows), f"{rows_with_meets}/{len(gt_rows)}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames and gt_wb is not None:
            ws = wb["Metrics"]
            gt_ws = gt_wb["Metrics"] if "Metrics" in gt_wb.sheetnames else None
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if gt_ws is not None:
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                check(f"Metrics has == {len(gt_rows)} rows",
                      len(data_rows) == len(gt_rows), f"got {len(data_rows)}")
            check_columns('Metrics', ['Metric', 'Value'], len(gt_rows) if gt_ws is not None else 4)
            if gt_ws is not None:
                # Check metric NAMES present (values depend on agent's chosen recipes)
                gt_metric_names = [str(r[0]).strip().lower() for r in gt_rows if r and r[0]]
                agent_metric_map = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for gt_m in gt_metric_names:
                    norm_m = gt_m.replace("_", "").replace(" ", "")
                    found = any(norm_m in k.replace("_", "").replace(" ", "") or
                                k.replace("_", "").replace(" ", "") in norm_m
                                for k in agent_metric_map.keys())
                    check(f"Metrics has '{gt_m}' (or equivalent)", found,
                          f"agent metrics: {list(agent_metric_map.keys())[:6]}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check_columns('Recommendations', ['Priority', 'Action'], 2)

        # Email check: require recipient = team-lead@company.com AND exact subject 'Analysis Report Complete'
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                """SELECT subject, to_addr FROM email.messages
                   WHERE to_addr::text ILIKE %s
                     AND LOWER(TRIM(subject)) = LOWER(%s)""",
                ('%team-lead@company.com%', 'Analysis Report Complete'),
            )
            emails = cur.fetchall()
            check("Email to team-lead@company.com with subject 'Analysis Report Complete' sent",
                  len(emails) >= 1, f"found {len(emails)} matching emails")
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        # Notion check: require page titled 'Cook Wellness Dashboard' (extract title from properties)
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Try multiple notion title structures: properties->title->title[0]->plain_text
            cur.execute("""
                SELECT id, properties FROM notion.pages
                WHERE archived = false
            """)
            rows = cur.fetchall()
            target_lower = "cook wellness dashboard"
            found_count = 0
            for pid, props in rows:
                if not props:
                    continue
                # Extract title text from JSON structure
                try:
                    if isinstance(props, str):
                        props_obj = json.loads(props)
                    else:
                        props_obj = props
                    # Navigate through possible title structures
                    titles = []
                    for k, v in (props_obj.items() if isinstance(props_obj, dict) else []):
                        if isinstance(v, dict) and v.get("type") == "title":
                            tlist = v.get("title", [])
                            for t in tlist:
                                if isinstance(t, dict):
                                    pt = t.get("plain_text", "") or (t.get("text", {}).get("content", "") if isinstance(t.get("text"), dict) else "")
                                    if pt:
                                        titles.append(pt)
                    title_text = " ".join(titles).lower().strip()
                    if title_text == target_lower or target_lower in title_text:
                        found_count += 1
                except Exception:
                    # Fallback: substring search
                    if target_lower in str(props).lower():
                        found_count += 1
            check("Notion 'Cook Wellness Dashboard' page exists",
                  found_count >= 1, f"found {found_count} matching pages")
            conn.close()
        except Exception as e:
            check("Notion check", False, str(e))

    # Move script existence check OUTSIDE excel-exists guard (still per task requirement)
    check("cook_wellness_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "cook_wellness_processor.py")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
