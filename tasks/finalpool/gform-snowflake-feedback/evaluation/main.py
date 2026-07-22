"""
Evaluation script for gform-snowflake-feedback task.

ALIGNED with task.md spec:
- Excel: Support_Feedback_Analysis.xlsx with sheets [Ticket Analysis, Priority Breakdown, Improvement Plan]
- Word: Support_Improvement_Report.docx with title 'Support Service Improvement Report' + 4 sections
- Form: 'Customer Feedback Survey' with >=5 questions including issue type radio, satisfaction rating, free text
- Email: to support-team@company.com with subject 'Support Improvement Report - Action Required'

Expected values are computed at evaluation time from the PostgreSQL/Snowflake mirror.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=5.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Top issues by ticket count (full + top 5)
    cur.execute("""
        SELECT
            "ISSUE_TYPE",
            COUNT(*) AS ticket_count,
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) AS avg_response_hours,
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) AS avg_satisfaction
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE"
        ORDER BY COUNT(*) DESC
    """)
    issue_rows = cur.fetchall()
    issue_summary = []
    for row in issue_rows:
        issue_summary.append({
            "Issue_Type": row[0],
            "Ticket_Count": int(row[1]),
            "Avg_Response_Hours": float(row[2]),
            "Avg_Satisfaction": float(row[3]),
        })
    top5_issue_types = [r["Issue_Type"] for r in issue_summary[:5]]

    # Priority breakdown
    cur.execute("""
        SELECT
            "PRIORITY",
            COUNT(*) AS cnt,
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) AS avg_response_hours
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY"
        ORDER BY "PRIORITY"
    """)
    priority_rows = cur.fetchall()
    priority_breakdown = []
    for row in priority_rows:
        priority_breakdown.append({
            "Priority": row[0],
            "Count": int(row[1]),
            "Avg_Response_Hours": float(row[2]),
        })

    cur.execute('SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
    total_tickets = cur.fetchone()[0]

    cur.close()
    conn.close()

    return {
        "issue_summary": issue_summary,
        "priority_breakdown": priority_breakdown,
        "top5_issue_types": top5_issue_types,
        "total_tickets": total_tickets,
    }


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")

    # Aligned with task.md: Support_Feedback_Analysis.xlsx
    agent_file = os.path.join(agent_workspace, "Support_Feedback_Analysis.xlsx")
    check("Excel file 'Support_Feedback_Analysis.xlsx' exists",
          os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Excel file readable", True)

    # Sheet names per task.md: 'Ticket Analysis', 'Priority Breakdown', 'Improvement Plan'
    check("Sheet 'Ticket Analysis' exists",
          any(str_match(s, "Ticket Analysis") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")
    check("Sheet 'Priority Breakdown' exists",
          any(str_match(s, "Priority Breakdown") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")
    check("Sheet 'Improvement Plan' exists",
          any(str_match(s, "Improvement Plan") for s in wb.sheetnames),
          f"Found sheets: {wb.sheetnames}")

    # --- Sheet 1: Ticket Analysis ---
    print("\n--- Ticket Analysis ---")
    ws = get_sheet(wb, "Ticket Analysis")
    if ws:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        agent_rows = [r for r in agent_rows if r and r[0] is not None]
        check("Ticket Analysis has >=5 rows (top 5 issue types)",
              len(agent_rows) >= 5,
              f"Got {len(agent_rows)} rows")

        # Sorted by Ticket_Count desc
        if len(agent_rows) >= 2:
            try:
                first_count = float(agent_rows[0][1]) if agent_rows[0][1] else 0
                second_count = float(agent_rows[1][1]) if agent_rows[1][1] else 0
                check("Ticket Analysis sorted by Ticket_Count desc",
                      first_count >= second_count,
                      f"First: {first_count}, Second: {second_count}")
            except (TypeError, ValueError):
                check("Ticket Analysis sorted by Ticket_Count desc", False,
                      "Could not parse count values")

        # Match top 5 issue types from DB
        top5 = expected["top5_issue_types"]
        for issue_type in top5:
            matched = None
            for ar in agent_rows:
                if ar and ar[0] and issue_type.lower() in str(ar[0]).lower():
                    matched = ar
                    break
            if matched:
                exp = next((r for r in expected["issue_summary"]
                            if r["Issue_Type"] == issue_type), None)
                if exp:
                    check(f"Ticket Analysis '{issue_type}' Ticket_Count",
                          num_close(matched[1], exp["Ticket_Count"], 5),
                          f"Expected {exp['Ticket_Count']}, got {matched[1]}")
                    check(f"Ticket Analysis '{issue_type}' Avg_Response_Time",
                          num_close(matched[2], exp["Avg_Response_Hours"], 0.5),
                          f"Expected {exp['Avg_Response_Hours']}, got {matched[2]}")
                    check(f"Ticket Analysis '{issue_type}' Avg_Satisfaction",
                          num_close(matched[3], exp["Avg_Satisfaction"], 0.5),
                          f"Expected {exp['Avg_Satisfaction']}, got {matched[3]}")
            else:
                check(f"Ticket Analysis '{issue_type}' present",
                      False, "Top 5 issue type not found in agent output")

    # --- Sheet 2: Priority Breakdown ---
    print("\n--- Priority Breakdown ---")
    ws = get_sheet(wb, "Priority Breakdown")
    if ws:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        agent_rows = [r for r in agent_rows if r and r[0] is not None]
        exp_rows = expected["priority_breakdown"]
        check("Priority Breakdown row count >= len(priority types)",
              len(agent_rows) >= max(1, len(exp_rows) - 1),
              f"Expected ~{len(exp_rows)}, got {len(agent_rows)}")

        for exp_row in exp_rows:
            priority = exp_row["Priority"]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], priority):
                    matched = ar
                    break
            if matched:
                check(f"Priority '{priority}' Count",
                      num_close(matched[1], exp_row["Count"], 5),
                      f"Expected {exp_row['Count']}, got {matched[1]}")
                # Expected schema: Priority, Count, Pct_of_Total, Avg_Response_Time
                # so Avg_Response_Time at column index 3
                if len(matched) >= 4:
                    check(f"Priority '{priority}' Avg_Response_Time",
                          num_close(matched[3], exp_row["Avg_Response_Hours"], 0.5),
                          f"Expected {exp_row['Avg_Response_Hours']}, got {matched[3]}")
            else:
                check(f"Priority '{priority}' found", False,
                      "Priority not in agent output")

    # --- Sheet 3: Improvement Plan ---
    print("\n--- Improvement Plan ---")
    ws = get_sheet(wb, "Improvement Plan")
    if ws:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        agent_rows = [r for r in agent_rows if r and r[0] is not None]
        check("Improvement Plan has >=5 rows",
              len(agent_rows) >= 5, f"Got {len(agent_rows)}")
        # Verify Target_Satisfaction = Current + 0.5 logic in at least one row
        good_target_count = 0
        for ar in agent_rows:
            try:
                if len(ar) >= 3 and ar[1] is not None and ar[2] is not None:
                    cur_v = float(ar[1])
                    tgt_v = float(ar[2])
                    if abs((cur_v + 0.5) - tgt_v) < 0.05:
                        good_target_count += 1
            except (TypeError, ValueError):
                pass
        check("Improvement Plan Target_Satisfaction = Current + 0.5 (>=3 rows)",
              good_target_count >= 3,
              f"Matching rows: {good_target_count}")


def check_form(expected):
    print("\n=== Checking Google Form ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title, description
        FROM gform.forms
        WHERE LOWER(title) LIKE '%customer%feedback%survey%'
           OR LOWER(title) LIKE '%customer feedback survey%'
        ORDER BY created_at DESC
        LIMIT 1
    """)
    form_row = cur.fetchone()

    check("Form 'Customer Feedback Survey' exists",
          form_row is not None,
          "No form found with title 'Customer Feedback Survey'")
    if not form_row:
        cur.close()
        conn.close()
        return

    form_id = form_row[0]
    form_title = form_row[1]
    check("Form title matches 'Customer Feedback Survey'",
          "customer feedback survey" in form_title.lower(),
          f"Actual title: '{form_title}'")

    cur.execute("""
        SELECT title, question_type, required, config
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()

    check("Form has at least 5 questions", len(questions) >= 5,
          f"Found {len(questions)} questions")

    # Issue type radio with top 5 issues
    top5 = expected["top5_issue_types"]
    has_issue_radio = False
    for q in questions:
        title_lower = q[0].lower()
        if ("issue" in title_lower or "experience" in title_lower or "type" in title_lower) \
                and q[1] in ("choiceQuestion", "RADIO"):
            config = q[3] if isinstance(q[3], dict) else {}
            options = []
            if "options" in config:
                for opt in config["options"]:
                    if isinstance(opt, dict) and "value" in opt:
                        options.append(opt["value"].lower())
                    elif isinstance(opt, str):
                        options.append(opt.lower())
            matched_count = sum(
                1 for t in top5
                if any(t.lower() in o for o in options)
            )
            if matched_count >= 3:
                has_issue_radio = True
                break
    check("Has issue-type radio question with top issue types", has_issue_radio)

    # Satisfaction rating question (1-5)
    has_rating = any(
        ("rate" in q[0].lower() or "satisfaction" in q[0].lower()
         or "overall" in q[0].lower())
        for q in questions
    )
    check("Has satisfaction rating question (1-5)", has_rating)

    # Free text improvement / suggestions
    has_free_text = any(
        ("improvement" in q[0].lower() or "improv" in q[0].lower()
         or "suggestion" in q[0].lower() or "comment" in q[0].lower()
         or "feedback" in q[0].lower())
        and q[1] in ("textQuestion", "TEXT", "paragraphTextQuestion", "PARAGRAPH_TEXT")
        for q in questions
    )
    check("Has free-text improvement/suggestion question", has_free_text)

    # Recommendation likelihood (1-10)
    has_recommend = any(
        ("recommend" in q[0].lower() or "likely" in q[0].lower())
        for q in questions
    )
    check("Has recommendation likelihood question", has_recommend)

    cur.close()
    conn.close()


def check_word(agent_workspace, expected):
    print("\n=== Checking Word Document ===")

    # Aligned with task.md: Support_Improvement_Report.docx
    doc_path = os.path.join(agent_workspace, "Support_Improvement_Report.docx")
    check("Word file 'Support_Improvement_Report.docx' exists",
          os.path.isfile(doc_path), f"Expected {doc_path}")
    if not os.path.isfile(doc_path):
        return

    try:
        from docx import Document
        doc = Document(doc_path)
    except Exception as e:
        check("Word file readable", False, str(e))
        return

    check("Word file readable", True)

    headings = []
    full_text = []
    for para in doc.paragraphs:
        if para.style and para.style.name and "Heading" in para.style.name:
            headings.append(para.text.strip())
        full_text.append(para.text)
    all_text = " ".join(full_text).lower()

    check("Doc title contains 'Support Service Improvement Report'",
          "support service improvement report" in all_text,
          f"text head: {all_text[:300]}")
    check("Doc has 'Executive Summary' section",
          "executive summary" in all_text,
          f"text head: {all_text[:300]}")
    check("Doc has 'Key Findings' section",
          "key findings" in all_text or "findings" in all_text,
          f"text head: {all_text[:300]}")
    check("Doc has 'Recommendations' section",
          "recommendation" in all_text,
          f"text head: {all_text[:300]}")
    check("Doc mentions total tickets count",
          str(expected["total_tickets"]) in all_text,
          f"Expected '{expected['total_tickets']}' in document")

    # Top 5 issue types must appear (>=3 of 5)
    top5 = expected["top5_issue_types"]
    mentioned = sum(1 for t in top5 if t.lower() in all_text)
    check("Doc mentions >=3 of top 5 issue types",
          mentioned >= 3,
          f"Found {mentioned} of {len(top5)}: top5={top5}")


def check_email(expected):
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Search across all folders for to=support-team@company.com
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        ORDER BY date DESC
    """)
    all_emails = cur.fetchall()

    target_email = None
    for email_row in all_emails:
        subject, from_addr, to_addr, body = email_row
        to_str = json.dumps(to_addr).lower() if to_addr else ""
        if "support-team@company.com" in to_str:
            target_email = email_row
            break

    check("Email to support-team@company.com exists",
          target_email is not None,
          f"Found {len(all_emails)} total emails but none to support-team@company.com")

    if target_email:
        subject, from_addr, to_addr, body = target_email
        subject_lower = (subject or "").lower()
        check("Email subject is 'Support Improvement Report - Action Required'",
              "support improvement report" in subject_lower
              and "action required" in subject_lower,
              f"Subject: '{subject}'")

        body_lower = (body or "").lower()
        check("Email body mentions total tickets",
              str(expected["total_tickets"]) in body_lower,
              f"Expected '{expected['total_tickets']}' in body")

        # Top issue type
        if expected["top5_issue_types"]:
            top1 = expected["top5_issue_types"][0]
            check("Email body mentions top issue type",
                  top1.lower() in body_lower,
                  f"Top: '{top1}'")

    cur.close()
    conn.close()


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0
    print("=== Computing Expected Values from Database ===")
    try:
        expected = compute_expected_values()
        print(f"  Computed issue summary for {len(expected['issue_summary'])} issue types")
        print(f"  Top 5 issue types: {expected['top5_issue_types']}")
        print(f"  Total tickets: {expected['total_tickets']}")
    except Exception as e:
        print(f"  ERROR computing expected values: {e}")
        return False, f"Failed to compute expected values: {e}"

    check_excel(agent_workspace, expected)
    check_form(expected)
    check_word(agent_workspace, expected)
    check_email(expected)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if FAIL_COUNT == 0 else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "success": FAIL_COUNT == 0,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return FAIL_COUNT == 0, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
