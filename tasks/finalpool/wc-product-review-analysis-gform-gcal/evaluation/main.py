"""Evaluation for wc-product-review-analysis-gform-gcal."""
import argparse
import os
import sys
import psycopg2
import openpyxl


# R1: every DB connection parameter is read from the environment with defaults
# that match the harness / preprocess conventions (PGHOST/PGPORT/PGDATABASE/
# PGUSER/PGPASSWORD). Never hardcode dbname or port.
DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}


def _to_float(v):
    """Robust float extraction for cell values (R2/R3).

    Handles int/float/str/None; strips currency symbols, thousands separators,
    percent signs and whitespace. Returns None when the value cannot be parsed
    (including unresolved formula strings).
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        if s.startswith("="):
            return None  # unresolved formula; treat as unparseable
        for ch in ("$", "€", "¥", ",", "%"):
            s = s.replace(ch, "")
        s = s.strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def num_close(a, b, tol=1.0):
    """Numeric comparison when both sides parse; else string compare (R3)."""
    if a is None and b is None:
        return True
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if a is None or b is None:
        return False
    try:
        return str(a).strip().lower() == str(b).strip().lower()
    except Exception:
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _norm_label(s):
    """Normalize a Stats metric label for comparison: lowercase and drop
    punctuation/whitespace, so 'Total_Products_Reviewed', 'Total Products
    Reviewed' and 'totalproductsreviewed' all collapse to one key. A correct
    agent that writes the label in a readable form must not be marked FAIL."""
    if s is None:
        return ""
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


# Common human-readable variants of the three canonical Stats metric labels.
# Keys/values are already normalized (no punctuation, lowercase).
_METRIC_ALIASES = {
    "totalproductsreviewed": ["totalproducts", "totalreviewedproducts", "productscounted"],
    "lowestratedproduct": ["lowestrated", "worstproduct", "lowestproduct"],
    "avgratingoverall": ["averageratingoverall", "overallaverage", "overallavg", "averageoverall"],
}


def _lookup_metric(a_data, canonical):
    """Find a Stats metric value in a label->value map, tolerating spacing,
    case and a few common readable aliases (e.g. 'Avg Rating Overall')."""
    nl = _norm_label(canonical)
    if nl in a_data:
        return a_data[nl]
    for alias in _METRIC_ALIASES.get(nl, ()):
        if alias in a_data:
            return a_data[alias]
    return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _resolve_value_matrix(wb_f, wb_v, sheet_name):
    """Return a cell-value matrix, using cached results for formula cells (R2).

    wb_f is loaded with data_only=False (so formulas are visible as strings),
    wb_v with data_only=True (cached values). When a cell holds a formula whose
    cached value is available we use the cached value; otherwise we keep the
    formula string so downstream checks can decide how to handle it.
    """
    rows_f = load_sheet_rows(wb_f, sheet_name)
    if rows_f is None:
        return None
    rows_v = load_sheet_rows(wb_v, sheet_name) or []
    out = []
    for r, rf in enumerate(rows_f):
        rv = rows_v[r] if r < len(rows_v) else []
        row = []
        for c, vf in enumerate(rf):
            if isinstance(vf, str) and vf.strip().startswith("="):
                vv = rv[c] if c < len(rv) else None
                row.append(vv if vv is not None else vf)
            else:
                row.append(vf)
        out.append(row)
    return out


def get_review_data():
    """Compute expected review metrics from the seed database.

    Returns None when the DB is unreachable so the caller can skip the
    DB-derived checks instead of silently using hardcoded values (R11).
    """
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT p.id, p.name, p.categories,
                count(pr.id) as review_count,
                avg(pr.rating) as avg_rating
            FROM wc.products p
            LEFT JOIN wc.product_reviews pr ON p.id = pr.product_id
            WHERE pr.id IS NOT NULL
            GROUP BY p.id, p.name, p.categories
            HAVING count(pr.id) >= 1
            ORDER BY avg(pr.rating) ASC, count(pr.id) DESC
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        total = len(rows)
        lowest = rows[0] if rows else None
        overall_avg = round(sum(round(float(r[4]), 2) for r in rows) / len(rows), 2) if rows else 0
        return {
            "all_rows": rows,
            "total": total,
            "lowest_name": lowest[1][:40] if lowest else "",
            "lowest_avg": round(float(lowest[4]), 2) if lowest else 0,
            "top5": rows[:5],
            "overall_avg": overall_avg,
        }
    except Exception as e:
        print(f"WARNING: Could not query DB; DB-derived checks will be skipped: {e}")
        return None


def check_gform():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Prefer the form with the most questions. In a parallel-agent setup an
        # orchestrator might accidentally create a duplicate (shell) form; the
        # fully-populated one is the one we want to evaluate.
        cur.execute(
            "SELECT f.id, f.title FROM gform.forms f "
            "WHERE LOWER(f.title) LIKE '%quality improvement survey%' "
            "   OR LOWER(f.title) LIKE '%product quality%' "
            "ORDER BY (SELECT count(*) FROM gform.questions q WHERE q.form_id = f.id) DESC "
            "LIMIT 1"
        )
        form = cur.fetchone()
        if not form:
            cur.close()
            conn.close()
            return False, "Form 'Product Quality Improvement Survey' not found"
        form_id = form[0]
        cur.execute("SELECT count(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        q_count = cur.fetchone()[0]
        cur.close()
        conn.close()
        if q_count < 6:
            return False, f"Form has {q_count} questions, expected >= 6 (5 product rating + 1 text)"
        return True, ""
    except Exception as e:
        return False, str(e)


def check_gcal_event():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Timezone-tolerant date check. The meeting must be "on 2026-03-20 from
        # 2pm to 3pm" in the agent's local timezone; a correct agent may send
        # that wall time with any UTC offset in [-12, +14]. Anchoring on UTC,
        # every such representation maps to the interval
        # [2026-03-20T00:00Z, 2026-03-21T03:00Z] (14:00-15:00 on 03-20 at the
        # most extreme plausible offsets). Comparing the stored timestamptz
        # against that window removes the dependence on the DB session
        # timezone, which previously could shift an extreme-offset event onto
        # 03-21 (or 03-19) and wrongly FAIL a correct event.
        cur.execute(
            "SELECT count(*) FROM gcal.events "
            "WHERE (LOWER(summary) LIKE '%quality%' OR LOWER(summary) LIKE '%product%') "
            "AND start_datetime >= TIMESTAMPTZ '2026-03-20T00:00:00+00' "
            "AND start_datetime <= TIMESTAMPTZ '2026-03-21T03:00:00+00'"
        )
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
        return cnt >= 1
    except Exception:
        return False


def check_email_sent():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT count(*) FROM email.messages WHERE LOWER(to_addr::text) LIKE '%product.team%' AND (LOWER(subject) LIKE '%quality%' OR LOWER(subject) LIKE '%review%')")
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
        return cnt >= 1
    except Exception:
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Product_Review_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Product_Review_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    all_errors = []

    review_data = get_review_data()

    # Read with data_only=False first, then resolve formula cells against the
    # cached (data_only=True) values so formulas are handled gracefully (R2).
    agent_wb_f = openpyxl.load_workbook(agent_file, data_only=False)
    agent_wb_v = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=False)

    # Check Review Summary sheet
    print("  Checking Review Summary sheet...")
    a_rows = _resolve_value_matrix(agent_wb_f, agent_wb_v, "Review Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Review Summary' not found in agent output")
    else:
        # Header-row detection: the first row is a header when its Avg_Rating
        # cell (column 2) is non-numeric. This tolerates agents that omit the
        # header row entirely, so the first (lowest-rated) data row is never
        # wrongly dropped.
        data_rows = [r for r in a_rows if r and any(c is not None for c in r)]
        if data_rows and _to_float(data_rows[0][1]) is None:
            data_rows = data_rows[1:]
        if len(data_rows) < 5:
            all_errors.append(f"Review Summary has {len(data_rows)} rows, expected >= 5 (at least top 5 lowest rated)")
        else:
            print(f"    PASS ({len(data_rows)} data rows)")

        # Check that data is sorted by avg_rating ascending (first row should be lowest rated)
        if len(data_rows) >= 2:
            first_rating = _to_float(data_rows[0][1])
            last_rating = _to_float(data_rows[-1][1])
            if first_rating is not None and last_rating is not None:
                if first_rating > last_rating:
                    all_errors.append(f"Data not sorted by Avg_Rating ASC: first={data_rows[0][1]}, last={data_rows[-1][1]}")
                else:
                    print("    Sort order PASS")

    # Check Stats sheet
    print("  Checking Stats sheet...")
    a_rows = _resolve_value_matrix(agent_wb_f, agent_wb_v, "Stats")
    if a_rows is None:
        all_errors.append("Sheet 'Stats' not found in agent output")
    else:
        # Build the metric map from every row, skipping only a 'Metric' header
        # row if one is present. Normalized keys accept spacing/case variants of
        # the labels, so a header-less Stats sheet cannot drop the first metric.
        a_data = {}
        for r in a_rows:
            if not r or r[0] is None:
                continue
            nl = _norm_label(r[0])
            if not nl or nl == "metric":
                continue
            a_data[nl] = r[1]
        if review_data is None:
            print("    SKIPPED (no DB connection; expected values unavailable)")
        else:
            errors = []

            total_val = _lookup_metric(a_data, "total_products_reviewed")
            if total_val is None:
                errors.append("Missing metric: Total_Products_Reviewed")
            elif not num_close(total_val, review_data["total"], 0):
                errors.append(f"Total_Products_Reviewed: {total_val} vs {review_data['total']}")

            lowest_val = _lookup_metric(a_data, "lowest_rated_product")
            if lowest_val is None:
                errors.append("Missing metric: Lowest_Rated_Product")
            else:
                expected_lower = review_data["lowest_name"].lower().strip()
                agent_lower = str(lowest_val).lower().strip()
                if expected_lower not in agent_lower and agent_lower not in expected_lower:
                    errors.append(f"Lowest_Rated_Product: '{lowest_val}' vs '{review_data['lowest_name']}'")

            avg_overall_val = _lookup_metric(a_data, "avg_rating_overall")
            if avg_overall_val is None:
                errors.append("Missing metric: Avg_Rating_Overall")
            elif not num_close(avg_overall_val, review_data["overall_avg"], 0.2):
                errors.append(f"Avg_Rating_Overall: {avg_overall_val} vs {review_data['overall_avg']} (tol=0.2)")

            if errors:
                all_errors.extend(errors)
                for e in errors[:5]:
                    print(f"    ERROR: {e}")
            else:
                print("    PASS")

    # Check GForm
    print("  Checking Google Form...")
    ok, detail = check_gform()
    if ok:
        print("    PASS")
    else:
        all_errors.append(detail)

    # Check GCal event
    print("  Checking GCal event...")
    if check_gcal_event():
        print("    PASS")
    else:
        all_errors.append("Calendar event 'Product Quality Review Meeting' on March 20 2026 not found")

    # Check email sent
    print("  Checking email to product.team...")
    if check_email_sent():
        print("    PASS")
    else:
        all_errors.append("Email to product.team@company.com with 'quality' or 'review' subject not found")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
