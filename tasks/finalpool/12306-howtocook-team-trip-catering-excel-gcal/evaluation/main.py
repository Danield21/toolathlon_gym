"""
Evaluation for 12306-howtocook-team-trip-catering-excel-gcal task.

Checks:
1. Team_Trip_Plan.xlsx exists with Travel, Menu, Timeline sheets
2. Travel sheet has G11 and 07:00 departure
3. Menu sheet has >= 5 rows with Course_Type and Dish_Name columns
4. Timeline sheet has >= 5 rows
5. Notion page exists with team/trip/march in title
6. GCal has the required events on 2026-03-10

Robustness notes (see audit):
- DB config reads PGHOST/PGPORT/PGDATABASE/PGUSER/PGPASSWORD env vars with defaults.
- Agent workbook is read with data_only=False so formula cells do not collapse to
  None; formulas resolve to cached value when available, else the formula string is
  kept for structural checks.
- Numeric parsing tolerates thousands separators / currency symbols / percent signs.
- Travel GT comparison maps columns by header name (not position) and uses flexible
  matching for Direction / Seat_Class / times / durations.
- GCal events are matched by wall-clock interpretation in UTC OR Asia/Shanghai so
  agents that write +08:00 offsets are not penalised.
"""
import json
import os
import re
import sys
from argparse import ArgumentParser
from datetime import datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

# The task is a Beijing -> Shanghai trip on this fixed date; wall-clock times in the
# task (06:30 / 07:00 / 12:30 / 18:00 ...) are understood in China local time.
EVENT_DATE = "2026-03-10"


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


# ---------------- value-normalization helpers ----------------

def _to_float(v):
    """Parse a numeric value from int/float/str. Returns None when unparseable."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        v = str(v)
    s = v.strip()
    if s.startswith("="):
        return None
    # strip thousands separators, currency symbols, percent, chinese yuan mark
    s = (s.replace(",", "").replace("$", "").replace("¥", "")
         .replace("€", "").replace("%", "").replace("元", "")
         .replace("CNY", "").replace("cny", "").replace("RMB", "").replace("rmb", "").strip())
    try:
        return float(s)
    except ValueError:
        return None


def _extract_number(v):
    """First number appearing in a value (fallback for '15 servings', '20 min')."""
    f = _to_float(v)
    if f is not None:
        return f
    m = re.search(r"\d+\.?\d*", str(v))
    if m:
        try:
            return float(m.group())
        except ValueError:
            return None
    return None


def _to_duration_min(v):
    """Duration in minutes from int/float or 'HH:MM' / 'H:MM' strings."""
    f = _to_float(v)
    if f is not None:
        return f
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", str(v).strip())
    if m:
        try:
            return int(m.group(1)) * 60 + int(m.group(2))
        except ValueError:
            return None
    return None


def num_close(a, b, tol=1.0):
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # fall back to case-insensitive string comparison when either side is not numeric
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _norm_time(v):
    """Normalise a time value to 'HH:MM'."""
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    s = str(v).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?", s)
    if m:
        return "%02d:%s" % (int(m.group(1)), m.group(2))
    return s.lower()


_TIME_TOKEN_RE = re.compile(r"(\d{1,2})[:：.．。](\d{1,2})(?::\d{2})?")


def _times_in_text(text):
    """Minutes-since-midnight (24h) for each time-like token in text.

    Handles 'HH:MM' / 'HH.MM' / 'H:MM' with optional am/pm markers, so a
    12-hour-clock agent that writes '6:00 PM' for 18:00 is parsed correctly.
    """
    out = []
    if not text:
        return out
    lower = str(text).lower()
    for m in _TIME_TOKEN_RE.finditer(lower):
        h, mi = int(m.group(1)), int(m.group(2))
        if h > 23 or mi > 59:
            continue
        tail = lower[m.end():m.end() + 4]
        pm = "pm" in tail
        am = ("am" in tail) and not pm
        if pm and h < 12:
            h += 12
        if am and h == 12:
            h = 0
        out.append(h * 60 + mi)
    return out


def _train_no_match(a, gt):
    if a is None:
        return False
    an = re.sub(r"[^a-z0-9]", "", str(a).lower())
    gn = re.sub(r"[^a-z0-9]", "", str(gt).lower())
    return gn in an or an in gn


def _direction_ok(a):
    """Value expresses a route from Beijing to Shanghai (flexible notation)."""
    if a is None:
        return False
    s = str(a).lower()
    # tokenise on arrows / dashes / 'to' / 'from'
    s = s.replace("→", " ").replace("->", " ").replace("-", " ")
    s = s.replace("_", " ").replace("到", " ").replace("至", " ")
    s = s.replace(" from ", " ")
    parts = [p for p in s.split() if p]
    if not parts:
        return False
    bj = [i for i, p in enumerate(parts) if ("beijing" in p or "北京" in p)]
    sh = [i for i, p in enumerate(parts) if ("shanghai" in p or "上海" in p)]
    if not bj or not sh:
        return False
    return min(bj) < min(sh)


def _seat_class_norm(v):
    if v is None:
        return ""
    s = str(v).strip().lower()
    if ("二等" in s) or ("second" in s) or ("2nd" in s) or ("2等" in s) or (s == "2"):
        return "second"
    if ("一等" in s) or ("first" in s) or ("1st" in s) or ("1等" in s) or (s == "1"):
        return "first"
    if ("商务" in s) or ("business" in s):
        return "business"
    return s


def _formula_pp_x_15(formula, pp_idx):
    """True if an unresolved Excel formula multiplies the per-person value by 15."""
    f = re.sub(r"\s+", "", str(formula)).upper()
    if not f.startswith("="):
        return False
    if not (re.search(r"\*15(?!\d)", f) or re.search(r"(?<!\d)15\*", f)):
        return False
    return True


def _course_class(course):
    c = str(course).strip().lower()
    if any(k in c for k in ("appetizer", "starter", "前菜", "凉菜", "冷菜", "冷盘",
                            "凉", "salad", "沙拉", "cold")):
        return "appetizer"
    if any(k in c for k in ("main", "entr", "主菜", "热菜", "荤", "主")):
        return "main"
    if any(k in c for k in ("dessert", "soup", "甜品", "汤", "甜", "羹")):
        return "dessert/soup"
    return "other"


def _norm_key(c):
    return re.sub(r"[^a-z0-9]", "", str(c).strip().lower())


def _header_map(headers):
    return {_norm_key(c): i for i, c in enumerate(headers) if c is not None and str(c).strip() != ""}


_COL_ALIASES = {
    "direction": ["route", "from_to", "fromto", "path", "itinerary"],
    "train_no": ["train", "traincode", "traintype"],
    "seat_class": ["class", "seat", "seattype", "seatclass"],
    "departure": ["depart", "departuretime", "starttime"],
    "arrival": ["arrive", "arrivaltime", "endtime"],
    "duration_min": ["duration", "traveltime", "journeytime"],
    "per_person_cny": ["perperson", "ticketprice", "priceperperson"],
    "total_budget_cny": ["totalbudget", "totalcost", "budget"],
}


def _find_col(headers, key):
    """Find column index for a keyword: exact-normalized, alias, then fuzzy substring."""
    nk = _norm_key(key)
    norm_headers = [_norm_key(c) for c in headers]
    if nk in norm_headers:
        return norm_headers.index(nk)
    for alias in _COL_ALIASES.get(key, ()):
        na = _norm_key(alias)
        if na in norm_headers:
            return norm_headers.index(na)
    for i, k in enumerate(norm_headers):
        if nk in k or k in nk:
            return i
    for alias in _COL_ALIASES.get(key, ()):
        na = _norm_key(alias)
        for i, k in enumerate(norm_headers):
            if na in k or k in na:
                return i
    return -1


def _load_ws(ws_raw, ws_val):
    """Resolve a worksheet to a list of value rows.

    Formula cells resolve to the cached value (data_only=True) when available,
    otherwise the raw formula string is kept so the evaluator can decide.
    """
    out = []
    for row in ws_raw.iter_rows():
        r = []
        for c in row:
            raw = c.value
            if isinstance(raw, str) and raw.startswith("="):
                cached = ws_val.cell(row=c.row, column=c.column).value
                r.append(cached if cached is not None else raw)
            else:
                r.append(raw)
        out.append(r)
    return out


def _cell_check(key, a_val, gt_val):
    """Compare a single Travel cell, dispatching on the column keyword."""
    if key == "train_no":
        return _train_no_match(a_val, gt_val), ""
    if key == "direction":
        return _direction_ok(a_val), ""
    if key in ("departure", "arrival", "depart", "arrive", "depart_time", "arrive_time",
               "start_time", "end_time"):
        return _norm_time(a_val) == _norm_time(gt_val), ""
    if key in ("duration_min", "duration_mins", "duration"):
        da, dg = _to_duration_min(a_val), _to_duration_min(gt_val)
        if da is not None and dg is not None:
            return abs(da - dg) <= 2, ""
        return str_match(a_val, gt_val), ""
    if key in ("seat_class", "seat", "class", "seat_type", "seatclass"):
        return _seat_class_norm(a_val) == _seat_class_norm(gt_val), ""
    fa, fg = _to_float(a_val), _to_float(gt_val)
    if fa is not None and fg is not None:
        if isinstance(gt_val, float):
            tol = max(abs(fg) * 0.02, 1.0)
        elif isinstance(gt_val, int):
            tol = 0.0
        else:
            tol = max(abs(fg) * 0.02, 1.0)
        return abs(fa - fg) <= tol, ""
    if a_val is not None and isinstance(a_val, str) and a_val.strip().startswith("="):
        # unresolved formula vs literal GT: cannot verify numerically; skip (no false fail)
        return True, ""
    return str_match(a_val, gt_val), ""


# ---------------- GT sheet comparisons ----------------

def _compare_travel(gt_sheet_name, gt_rows, agent_rows, agent_data):
    if not gt_rows:
        record(f"GT '{gt_sheet_name}' has data rows", False, "GT sheet empty")
        return
    gt_header = [str(c).strip().lower() if c else "" for c in gt_rows[0]]
    agent_header = [str(c).strip().lower() if c else "" for c in (agent_rows[0] if agent_rows else [])]
    gmap = _header_map(gt_header)
    gt_data = [r for r in gt_rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
    record(f"GT '{gt_sheet_name}' agent has >= 1 data row", len(agent_data) >= 1,
           f"got {len(agent_data)}")
    if not gt_data:
        record(f"GT '{gt_sheet_name}' has data rows", True)
        return
    for gi, gt_row in enumerate(gt_data):
        a_row = None
        gt_train = gt_row[gmap["train_no"]] if ("train_no" in gmap and gmap["train_no"] < len(gt_row)) else None
        if gt_train is not None:
            for r in agent_data:
                a_ti = _find_col(agent_header, "train_no")
                if a_ti >= 0 and a_ti < len(r) and _train_no_match(r[a_ti], gt_train):
                    a_row = r
                    break
        if a_row is None and gi < len(agent_data):
            a_row = agent_data[gi]
        if a_row is None:
            record(f"GT '{gt_sheet_name}' row {gi + 1} exists", False, "No matching agent row")
            continue
        all_ok = True
        first_bad = None
        for ci, key in enumerate(gt_header):
            if not key or key in ("none",):
                continue
            gt_val = gt_row[ci] if ci < len(gt_row) else None
            if gt_val is None:
                continue
            a_ti = _find_col(agent_header, key)
            a_val = a_row[a_ti] if (a_ti >= 0 and a_ti < len(a_row)) else None
            ok, extra = _cell_check(key, a_val, gt_val)
            if not ok:
                all_ok = False
                first_bad = (key, gt_val, a_val, extra)
                break
        if all_ok:
            record(f"GT '{gt_sheet_name}' row {gi + 1} values match", True)
        else:
            key, gt_val, a_val, extra = first_bad
            record(f"GT '{gt_sheet_name}' row {gi + 1} col {key}",
                   False, f"Expected {gt_val}, got {a_val} {extra}".strip())


def _compare_menu(gt_sheet_name, gt_rows, agent_rows, agent_data):
    record(f"GT '{gt_sheet_name}' agent has >= 6 data rows", len(agent_data) >= 6,
           f"got {len(agent_data)}")
    agent_header = [str(c).lower() if c else "" for c in (agent_rows[0] if agent_rows else [])]
    course_idx = next((i for i, h in enumerate(agent_header) if "course" in h or "category" in h or "type" in h or "kind" in h), -1)
    dish_idx = next((i for i, h in enumerate(agent_header) if "dish" in h or "name" in h or "item" in h or "meal" in h or "menu" in h), -1)
    cook_idx = next((i for i, h in enumerate(agent_header) if "cook" in h or "time" in h), -1)
    serv_idx = next((i for i, h in enumerate(agent_header) if "serv" in h or "yield" in h), -1)

    record("Menu has cook time column", cook_idx >= 0, f"Headers: {agent_header}")
    record("Menu has servings yield column", serv_idx >= 0, f"Headers: {agent_header}")

    appetizer_count = main_count = dessert_count = 0
    dish_names_valid = True
    for r in agent_data:
        course = (str(r[course_idx]) if course_idx >= 0 and course_idx < len(r) else "").strip().lower()
        dish = (str(r[dish_idx]) if dish_idx >= 0 and dish_idx < len(r) else "").strip()
        if not dish or dish.lower() == "none":
            dish_names_valid = False
        cls = _course_class(course)
        if cls == "appetizer":
            appetizer_count += 1
        elif cls == "main":
            main_count += 1
        elif cls == "dessert/soup":
            dessert_count += 1

    record("Menu has >= 2 appetizers, >= 3 mains, >= 1 dessert/soup (from Course_Type)",
           appetizer_count >= 2 and main_count >= 3 and dessert_count >= 1,
           f"appetizer={appetizer_count}, main={main_count}, dessert/soup={dessert_count}")
    record("Menu dish names are all non-empty", dish_names_valid,
           "All Dish_Name cells must be non-empty")

    cook_time_ok = True
    servings_ok = True
    for r in agent_data:
        if cook_idx >= 0 and cook_idx < len(r):
            ct = _extract_number(r[cook_idx])
            if ct is None or ct <= 0 or ct > 600:
                cook_time_ok = False
        if serv_idx >= 0 and serv_idx < len(r):
            sv = _extract_number(r[serv_idx])
            if sv is not None:
                # accept either scaled-to-team-size (15) or the base recipe yield (2-4)
                if abs(sv - 15) > 1 and not (2 <= sv <= 4):
                    servings_ok = False
    record("Menu cook times are numeric and reasonable (1-600 min)", cook_time_ok, "")
    record("Menu servings_yield is scaled-to-15 (or base recipe 2-4)", servings_ok, "")


def _compare_timeline(gt_sheet_name, gt_rows, agent_rows, agent_data):
    record(f"GT '{gt_sheet_name}' agent has >= 6 data rows", len(agent_data) >= 6,
           f"got {len(agent_data)}")
    agent_header = [str(c).lower() if c else "" for c in (agent_rows[0] if agent_rows else [])]
    time_idx = next((i for i, h in enumerate(agent_header) if "time" in h), -1)
    act_idx = next((i for i, h in enumerate(agent_header) if "activ" in h or "task" in h), -1)

    timeline_ok = True
    for r in agent_data:
        t = (str(r[time_idx]) if time_idx >= 0 and time_idx < len(r) else "").strip()
        a = (str(r[act_idx]) if act_idx >= 0 and act_idx < len(r) else "").strip()
        if not t or not a or t.lower() == "none" or a.lower() == "none":
            timeline_ok = False
            break
    record("Timeline rows have non-empty Time and Activity", timeline_ok, "")

    all_times_text = " ".join(
        str(r[time_idx]) if time_idx >= 0 and time_idx < len(r) else "" for r in agent_data).lower()
    # Parse times into 24h minutes so 12-hour ('6:00 PM' / '6:30 AM') and 24-hour
    # ('18:00' / '06:30') notations are both recognised for every window.
    t24 = _times_in_text(all_times_text)
    has_morning = any(300 <= t <= 540 for t in t24)    # 05:00 - 09:00
    has_arrival = any(660 <= t <= 840 for t in t24)    # 11:00 - 14:00
    has_dinner = any(1020 <= t <= 1260 for t in t24)   # 17:00 - 21:00
    record("Timeline covers morning departure, midday arrival, evening dinner times",
           has_morning and has_arrival and has_dinner,
           f"morning:{has_morning} arrival:{has_arrival} dinner:{has_dinner}")


# ---------------- Excel check ----------------

def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Team_Trip_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Team_Trip_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Team_Trip_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Team_Trip_Plan.xlsx exists", True)

    try:
        wb_raw = openpyxl.load_workbook(xlsx_path, data_only=False)
        wb_val = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb_raw.sheetnames]
    has_travel = any("travel" in s for s in sheet_names_lower)
    has_menu = any("menu" in s for s in sheet_names_lower)
    has_timeline = any("timeline" in s for s in sheet_names_lower)

    record("Excel has Travel sheet", has_travel, f"Sheets: {wb_raw.sheetnames}")
    record("Excel has Menu sheet", has_menu, f"Sheets: {wb_raw.sheetnames}")
    record("Excel has Timeline sheet", has_timeline, f"Sheets: {wb_raw.sheetnames}")

    sheets_resolved = {}
    for sn in wb_raw.sheetnames:
        sheets_resolved[sn] = _load_ws(wb_raw[sn], wb_val[sn])

    if has_travel:
        ws_name = wb_raw.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "travel" in s)]
        rows = sheets_resolved[ws_name]
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        record("Travel sheet has >= 1 data row", len(data_rows) >= 1, f"Found {len(data_rows)} rows")

        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        # match the GT comparison's _train_no_match tolerance: strip non-alphanumeric
        # and look for the code as a substring (accepts 'G11次' / 'G 11' / 'G11').
        g11_norm = re.sub(r"[^a-z0-9]", "", all_text)
        has_g11 = "g11" in g11_norm
        record("Travel sheet contains G11", has_g11, f"Content sample: {all_text[:200]}")
        has_time = bool(re.search(r"\b07:00(:00)?\b", all_text)) or bool(re.search(r"\b7:00\b", all_text))
        record("Travel sheet shows 07:00 departure", has_time, f"Content: {all_text[:200]}")

        if rows:
            pp_idx = _find_col(rows[0], "per_person_cny")
            tb_idx = _find_col(rows[0], "total_budget_cny")
            if pp_idx >= 0 and tb_idx >= 0 and data_rows:
                ok = True
                bad = None
                for r in data_rows:
                    pp = _to_float(r[pp_idx])
                    tb_val = r[tb_idx]
                    tb = _to_float(tb_val)
                    if pp is None:
                        ok = False
                        bad = ("Per_Person_CNY", r[pp_idx])
                        break
                    if tb is not None:
                        if abs(tb - pp * 15) > 1.0:
                            ok = False
                            bad = ("Total_Budget_CNY", r[tb_idx])
                            break
                    elif isinstance(tb_val, str) and tb_val.strip().startswith("="):
                        # unresolved formula: verify it structurally multiplies by 15
                        if not _formula_pp_x_15(tb_val, pp_idx):
                            ok = False
                            bad = ("Total_Budget_CNY formula", r[tb_idx])
                            break
                    else:
                        ok = False
                        bad = ("Total_Budget_CNY", r[tb_idx])
                        break
                record("Total_Budget_CNY = Per_Person_CNY * 15", ok,
                       f"Travel rows: {data_rows}")

    if has_menu:
        ws_name = wb_raw.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "menu" in s)]
        rows = sheets_resolved[ws_name]
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        record("Menu sheet has >= 6 rows", len(data_rows) >= 6, f"Found {len(data_rows)} rows")

        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            has_course = any("course" in h or "type" in h or "category" in h or "kind" in h for h in headers)
            has_dish = any("dish" in h or "name" in h or "item" in h or "meal" in h or "menu" in h for h in headers)
            record("Menu has course type column", has_course, f"Headers: {rows[0]}")
            record("Menu has dish name column", has_dish, f"Headers: {rows[0]}")

    if has_timeline:
        ws_name = wb_raw.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "timeline" in s)]
        rows = sheets_resolved[ws_name]
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        record("Timeline sheet has >= 6 rows", len(data_rows) >= 6, f"Found {len(data_rows)} rows")

    # ---- Groundtruth comparison ----
    gt_path = os.path.join(groundtruth_workspace, "Team_Trip_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    try:
        gt_wb_raw = openpyxl.load_workbook(gt_path, data_only=False)
        gt_wb_val = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        record("Groundtruth xlsx readable", False, str(e))
        return
    record("Groundtruth xlsx exists", True)

    for gt_sheet_name in gt_wb_raw.sheetnames:
        agent_ws_name = None
        for asn in wb_raw.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws_name = asn
                break
        if agent_ws_name is None:
            record(f"GT sheet '{gt_sheet_name}' exists in agent", False,
                   f"Available: {wb_raw.sheetnames}")
            continue

        gt_rows = _load_ws(gt_wb_raw[gt_sheet_name], gt_wb_val[gt_sheet_name])
        agent_rows = sheets_resolved[agent_ws_name]
        agent_data = [r for r in agent_rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        sn_lower = gt_sheet_name.strip().lower()

        if "travel" in sn_lower:
            _compare_travel(gt_sheet_name, gt_rows, agent_rows, agent_data)
        elif "menu" in sn_lower:
            _compare_menu(gt_sheet_name, gt_rows, agent_rows, agent_data)
        elif "timeline" in sn_lower:
            _compare_timeline(gt_sheet_name, gt_rows, agent_rows, agent_data)

    try:
        wb_raw.close()
        wb_val.close()
        gt_wb_raw.close()
        gt_wb_val.close()
    except Exception:
        pass


# ---------------- Notion check ----------------

def check_notion():
    print("\n=== Check 2: Notion page for team trip ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Notion pages query", False, f"DB connect/query error: {e}")
        return

    def _rich_text_to_str(title_block):
        if not isinstance(title_block, dict):
            return ""
        title_list = title_block.get("title")
        if isinstance(title_list, list):
            out = ""
            for t in title_list:
                if not isinstance(t, dict):
                    continue
                tc = t.get("text")
                if isinstance(tc, dict):
                    out += tc.get("content") or tc.get("plain_text") or ""
                else:
                    out += t.get("plain_text") or ""
            return out
        return ""

    def _page_title(props):
        if not isinstance(props, dict):
            return ""
        # title may live under any property key; try common ones then any dict
        # whose value carries a 'title' rich-text list
        for key in ("title", "Name", "name", "Page", "page", "Title"):
            if key in props:
                s = _rich_text_to_str(props[key])
                if s:
                    return s
        for k, v in props.items():
            if isinstance(v, dict) and "title" in v:
                s = _rich_text_to_str(v)
                if s:
                    return s
        return ""

    trip_page = None
    for page_id, props in pages:
        title = _page_title(props)
        title_lower = title.lower()
        has_trip_keyword = any(kw in title_lower for kw in
                               ["team", "trip", "建设", "出行", "团建", "团队", "旅行", "旅游"])
        has_date_or_dest = any(kw in title_lower for kw in
                               ["march", "shanghai", "上海", "3月", "03-10", "03/10",
                                "2026-03-10", "team building"])
        if has_trip_keyword and has_date_or_dest:
            trip_page = (page_id, title)
            break

    record("Notion page exists with trip+date/destination context (team/trip AND march/shanghai/team-building)",
           trip_page is not None,
           f"Pages found: {len(pages)}, titles: {[str(p[1])[:80] for p in pages[:3]]}")


# ---------------- GCal check ----------------

def check_gcal():
    print("\n=== Check 3: GCal events on 2026-03-10 ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, start_timezone, end_datetime, end_timezone
            FROM gcal.events
            ORDER BY start_datetime
        """)
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("GCal events query", False, f"DB connect/query error: {e}")
        return

    # Required: 3 specific events with the correct time slots.
    expected = [
        (["depart", "departure", "station", "train"], EVENT_DATE, "06:30", "07:30",
         "GCal departure event on 2026-03-10 06:30-07:30"),
        (["arriv", "check", "hotel"], EVENT_DATE, "12:30", "14:00",
         "GCal arrival/check-in event on 2026-03-10 12:30-14:00"),
        (["dinner", "team dinner", "banquet"], EVENT_DATE, "18:00", "20:00",
         "GCal team dinner event on 2026-03-10 18:00-20:00"),
    ]

    def _wallclock_candidates(sdt, stz, edt, etz):
        """Candidate (start, end) wall-clock interpretations of stored instants.

        psycopg2 returns naive datetimes in the session timezone (UTC). The agent may
        have written naive/UTC times or +08:00 (Asia/Shanghai) times; we accept both,
        plus any explicit timezone annotation stored on the event.
        """
        try:
            if sdt is None or edt is None:
                return []
            if getattr(sdt, "tzinfo", None) is not None:
                su = sdt.astimezone(timezone.utc).replace(tzinfo=None)
            else:
                su = sdt
            if getattr(edt, "tzinfo", None) is not None:
                eu = edt.astimezone(timezone.utc).replace(tzinfo=None)
            else:
                eu = edt
        except Exception:
            return []
        cands = [(su, eu)]  # interpretation A: values are the wall-clock as written
        cands.append((su + timedelta(hours=8), eu + timedelta(hours=8)))  # B: +08:00 agent
        for tzname in (stz, etz):
            if not tzname:
                continue
            try:
                tz = ZoneInfo(tzname)
                s_local = su.replace(tzinfo=timezone.utc).astimezone(tz).replace(tzinfo=None)
                e_local = eu.replace(tzinfo=timezone.utc).astimezone(tz).replace(tzinfo=None)
                cands.append((s_local, e_local))
                break
            except Exception:
                continue
        return cands

    def _has_event(keywords, exp_date, exp_start, exp_end):
        exp_d = datetime.strptime(exp_date, "%Y-%m-%d").date()
        exp_s_min = int(exp_start[:2]) * 60 + int(exp_start[3:5])
        exp_e_min = int(exp_end[:2]) * 60 + int(exp_end[3:5])
        for summary, sdt, stz, edt, etz in events:
            if not any(kw in (summary or "").lower() for kw in keywords):
                continue
            for (s_local, e_local) in _wallclock_candidates(sdt, stz, edt, etz):
                if s_local.date() != exp_d or e_local.date() != exp_d:
                    continue
                s_min = s_local.hour * 60 + s_local.minute
                e_min = e_local.hour * 60 + e_local.minute
                if abs(s_min - exp_s_min) <= 30 and abs(e_min - exp_e_min) <= 30:
                    return True
        return False

    summaries = [f"({summary or ''}, {sdt})" for summary, sdt, _, _, _ in events]

    def _check_expected(exp):
        return _has_event(exp[0], exp[1], exp[2], exp[3])

    has_depart = _check_expected(expected[0])
    has_arrival = _check_expected(expected[1])
    has_dinner = _check_expected(expected[2])
    record("GCal departure event on 2026-03-10 06:30-07:30", has_depart, f"Mar10 events: {summaries}")
    record("GCal arrival/check-in event on 2026-03-10 12:30-14:00", has_arrival, f"Mar10 events: {summaries}")
    record("GCal team dinner event on 2026-03-10 18:00-20:00", has_dinner, f"Mar10 events: {summaries}")


# ---------------- Email check ----------------

def check_email():
    print("\n=== Check 4: Email sent to events@company.com ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        messages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email messages query", False, f"DB connect/query error: {e}")
        return

    outgoing = []
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif to_addr:
            try:
                parsed = json.loads(str(to_addr))
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "events@company.com" in to_str:
            outgoing.append((subject, from_addr, to_addr, body_text))

    record("Email sent to events@company.com", len(outgoing) >= 1,
           f"Total messages: {len(messages)}, matching: {len(outgoing)}")

    if outgoing:
        subject, _, _, body = outgoing[0]
        body_lower = ((subject or "") + " " + (body or "")).lower()
        has_trip = any(kw in body_lower for kw in
                       ["trip", "travel", "train", "shanghai", "g11", "excel", "plan"])
        record("Email mentions trip/travel/plan content", has_trip,
               f"Subject: {subject}")


# ---------------- main ----------------

def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0 and PASS_COUNT > 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
