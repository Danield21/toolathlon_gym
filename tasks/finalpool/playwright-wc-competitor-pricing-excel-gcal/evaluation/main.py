"""
Evaluation script for playwright-wc-competitor-pricing-excel-gcal task.

Checks:
1. Competitive_Pricing_Analysis.xlsx with 4 sheets and correct data
2. Calendar events for 3 category price review meetings
3. Email sent with pricing analysis results
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=5.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


def _compute_competitor_avgs():
    """Parse the static playwright mock catalog HTML to compute competitor avg per category.
    Falls back to hardcoded values if the file is missing."""
    try:
        import re as _re
        # Try to locate the mock catalog file relative to the task directory.
        here = os.path.dirname(os.path.abspath(__file__))
        candidates = [
            os.path.join(here, "..", "tmp", "mock_pages", "catalog.html"),
        ]
        path = None
        for p in candidates:
            if os.path.isfile(p):
                path = p
                break
        if not path:
            raise FileNotFoundError("mock catalog.html not found")
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            html = f.read()
        # Each row: <td>Name</td><td>Category</td><td class="price">$X.YZ</td>
        rows = _re.findall(
            r"<tr>\s*<td>[^<]+</td>\s*<td>([^<]+)</td>\s*<td class=\"price\">\$([0-9.]+)</td>",
            html,
        )
        buckets = {}
        for cat, price in rows:
            cat_l = cat.strip().lower()
            try:
                p = float(price)
            except ValueError:
                continue
            buckets.setdefault(cat_l, []).append(p)
        return {k: round(sum(v) / len(v), 2) for k, v in buckets.items() if v}
    except Exception as e:
        print(f"[WARN] competitor_avgs parse failed: {e}")
        return {"electronics": 59.06, "headphones": 74.98, "speakers": 138.73}


COMPETITOR_AVGS = _compute_competitor_avgs()
OUR_AVG_REGULAR = {"electronics": 97.21, "headphones": 66.00, "speakers": 150.59}
OUR_AVG_SALE = {"electronics": 65.13, "headphones": 13.11, "speakers": 135.16}
# Gaps are derived: gap = our - competitor (so they auto-update with COMPETITOR_AVGS).
PRICE_GAP_REGULAR = {
    cat: round(OUR_AVG_REGULAR[cat] - COMPETITOR_AVGS.get(cat, 0), 2)
    for cat in OUR_AVG_REGULAR
}
PRICE_GAP_SALE = {
    cat: round(OUR_AVG_SALE[cat] - COMPETITOR_AVGS.get(cat, 0), 2)
    for cat in OUR_AVG_SALE
}


def check_excel(agent_workspace):
    """Check Competitive_Pricing_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    def _find_sheet(needle, exact_first=True):
        # Try exact first, then case-insensitive equality, then substring
        for name in wb.sheetnames:
            if name == needle:
                return name
        for name in wb.sheetnames:
            if name.strip().lower() == needle.strip().lower():
                return name
        for name in wb.sheetnames:
            if needle.strip().lower() in name.strip().lower():
                return name
        return None

    # --- Sheet 1: Competitor Products ---
    comp_sheet = _find_sheet("Competitor Products")
    if not comp_sheet:
        record("Competitor Products sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Competitor Products sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 15
        record("Competitor Products has 15 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 2: Our Products ---
    our_sheet = _find_sheet("Our Products")
    if not our_sheet:
        record("Our Products sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Our Products sheet exists", True)
        ws = wb[our_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        # GT has exactly 45 (Electronics + Headphones + Speakers from WC catalog).
        # Allow tight ±1 to absorb any minor catalog variation.
        ok = abs(len(data_rows) - 45) <= 1
        record("Our Products has 45 rows (±1)", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 3: Category Comparison ---
    cat_sheet = _find_sheet("Category Comparison")
    if not cat_sheet:
        record("Category Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Category Comparison sheet exists", True)
        ws = wb[cat_sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 3
        record("Category Comparison has 3 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

        # Header positions (validate by header name; positional fallback)
        def _col_idx(target_keys):
            for i, h in enumerate(header):
                for tk in target_keys:
                    if tk in h:
                        return i
            return None

        idx_cat = _col_idx(["category"]) or 0
        idx_comp_avg = _col_idx(["competitor_avg", "competitor avg"]) or 1
        idx_our_reg = _col_idx(["our_avg_regular", "regular"]) or 2
        idx_our_sale = _col_idx(["our_avg_sale", "sale"]) or 3
        idx_gap_reg = _col_idx(["price_gap_regular", "gap_regular", "gap regular"]) or 4
        idx_gap_sale = _col_idx(["price_gap_sale", "gap_sale", "gap sale"]) or 5

        for row in data_rows:
            if not row or not row[idx_cat]:
                continue
            cat = str(row[idx_cat]).strip().lower()
            if cat in COMPETITOR_AVGS:
                # Competitor avg
                v = row[idx_comp_avg] if idx_comp_avg < len(row) else None
                ok_comp = num_close(v, COMPETITOR_AVGS[cat], tol=2.0)
                record(f"Competitor avg for {cat}", ok_comp,
                       f"Expected ~{COMPETITOR_AVGS[cat]}, got {v}")
                if not ok_comp:
                    all_ok = False
                # Our_Avg_Regular
                v = row[idx_our_reg] if idx_our_reg < len(row) else None
                ok_reg = num_close(v, OUR_AVG_REGULAR[cat], tol=3.0)
                record(f"Our_Avg_Regular_Price for {cat}", ok_reg,
                       f"Expected ~{OUR_AVG_REGULAR[cat]}, got {v}")
                if not ok_reg:
                    all_ok = False
                # Our_Avg_Sale
                v = row[idx_our_sale] if idx_our_sale < len(row) else None
                ok_sale = num_close(v, OUR_AVG_SALE[cat], tol=3.0)
                record(f"Our_Avg_Sale_Price for {cat}", ok_sale,
                       f"Expected ~{OUR_AVG_SALE[cat]}, got {v}")
                if not ok_sale:
                    all_ok = False
                # Price_Gap_Regular
                v = row[idx_gap_reg] if idx_gap_reg < len(row) else None
                ok_gap_r = num_close(v, PRICE_GAP_REGULAR[cat], tol=3.0)
                record(f"Price_Gap_Regular for {cat}", ok_gap_r,
                       f"Expected ~{PRICE_GAP_REGULAR[cat]}, got {v}")
                if not ok_gap_r:
                    all_ok = False
                # Price_Gap_Sale
                v = row[idx_gap_sale] if idx_gap_sale < len(row) else None
                ok_gap_s = num_close(v, PRICE_GAP_SALE[cat], tol=3.0)
                record(f"Price_Gap_Sale for {cat}", ok_gap_s,
                       f"Expected ~{PRICE_GAP_SALE[cat]}, got {v}")
                if not ok_gap_s:
                    all_ok = False

    # --- Sheet 4: Summary ---
    sum_sheet = _find_sheet("Summary")
    if not sum_sheet:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Summary sheet exists", True)
        ws = wb[sum_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        seen_metrics = set()
        for row in data_rows:
            if row and row[0]:
                metric = str(row[0]).strip().lower()
                val = row[1]
                if "total_competitor" in metric:
                    seen_metrics.add("total_competitor")
                    ok = num_close(val, 15, tol=0)
                    record("Total_Competitor_Products = 15", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "total_our" in metric:
                    seen_metrics.add("total_our")
                    # GT has 45; allow >= 30 (mock catalog page may shift slightly)
                    ok = isinstance(val, (int, float)) and float(val) >= 30
                    record("Total_Our_Products >= 30", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "cheapest" in metric:
                    seen_metrics.add("cheapest")
                    ok = str_contains(val, "headphone")
                    record("Cheapest category is Headphones", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "expensive" in metric:
                    seen_metrics.add("expensive")
                    ok = str_contains(val, "electronics")
                    record("Most expensive category is Electronics", ok, f"Got {val}")
                    if not ok:
                        all_ok = False

        for required in ["total_competitor", "total_our", "cheapest", "expensive"]:
            if required not in seen_metrics:
                record(f"Summary metric '{required}' present", False,
                       f"Found metrics: {seen_metrics}")
                all_ok = False

    wb.close()
    return all_ok


def check_calendar():
    """Check calendar events for 3 category price review meetings."""
    print("\n=== Checking Google Calendar ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        return False

    all_ok = True
    # Required (category, expected_date) pairs
    expected = [
        ("electronics", "2026-03-10"),
        ("headphones", "2026-03-11"),
        ("speakers", "2026-03-12"),
    ]
    found_per_cat = {cat: [] for cat, _ in expected}

    for summary, description, start_dt, end_dt in events:
        summary_lower = (summary or "").lower()
        if "price review" not in summary_lower:
            continue
        for cat, _ in expected:
            if cat in summary_lower:
                found_per_cat[cat].append((summary, start_dt, end_dt))

    for cat, ymd in expected:
        evs = found_per_cat[cat]
        ok_exists = len(evs) >= 1
        record(f"Price Review event for {cat}", ok_exists,
               f"Found {len(evs)}: {[(s, str(sd)) for s, sd, _ in evs]}")
        if not ok_exists:
            all_ok = False
            continue
        ok_date = False
        ok_time = False
        from datetime import datetime as _dt
        for summary, sd, ed in evs:
            sd_str = str(sd)
            if ymd in sd_str:
                ok_date = True
            # Check that hour is 14 (NY local) OR 19 (UTC equivalent for EST -05).
            # Parse start/end as datetime to require explicit hour values.
            try:
                if isinstance(sd, _dt):
                    sd_obj = sd
                else:
                    sd_obj = _dt.fromisoformat(str(sd).replace("Z", "+00:00"))
                if isinstance(ed, _dt):
                    ed_obj = ed
                elif ed:
                    ed_obj = _dt.fromisoformat(str(ed).replace("Z", "+00:00"))
                else:
                    ed_obj = None
                start_hour = sd_obj.hour
                end_hour = ed_obj.hour if ed_obj else None
                # NY local 14:00-15:00 (hour=14, end_hour=15) OR UTC 19:00-20:00.
                hour_pair = (start_hour, end_hour)
                if hour_pair in {(14, 15), (19, 20), (18, 19), (13, 14)}:
                    ok_time = True
                # Also check minute is 00
                if sd_obj.minute != 0:
                    ok_time = False
            except Exception:
                continue
        record(f"  Event for {cat} dated {ymd}", ok_date,
               f"start_datetimes: {[str(sd) for _, sd, _ in evs]}")
        if not ok_date:
            all_ok = False
        record(f"  Event for {cat} 14:00-15:00 (NY local or 19:00-20:00 UTC)", ok_time,
               f"start/end: {[(str(sd), str(ed)) for _, sd, ed in evs]}")
        if not ok_time:
            all_ok = False

    return all_ok


def check_email():
    """Check email was sent with pricing analysis."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    all_ok = True
    candidates = []

    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        # Filter: subject must contain BOTH 'competitive' and 'pricing' (loose still matches the task subject)
        if ("competitive" in subj_lower and "pricing" in subj_lower) or \
           ("competitive pricing analysis" in subj_lower):
            candidates.append((subject, from_addr, to_addr, body_text))

    record("Email with 'Competitive Pricing Analysis' subject exists",
           len(candidates) >= 1,
           f"Found {len(candidates)} candidate emails out of {len(emails)} total. "
           f"Subjects: {[(e[0] or '') for e in emails][:5]}")
    if not candidates:
        return False

    subject, from_addr, to_addr, body_text = candidates[0]

    # Subject contains 'Competitive Pricing Analysis Results'
    subj_lower = (subject or "").lower()
    ok_subj = "competitive pricing analysis" in subj_lower and "result" in subj_lower
    record("Subject contains 'Competitive Pricing Analysis Results'",
           ok_subj, f"Got: '{subject}'")
    if not ok_subj:
        all_ok = False

    # From address
    from_lower = (from_addr or "").lower()
    from_ok = "pricing@ecommerce.com" in from_lower or (
        "pricing" in from_lower and "ecommerce" in from_lower
    )
    record("Email from pricing@ecommerce.com", from_ok, f"From: {from_addr}")
    if not from_ok:
        all_ok = False

    # To address: must include manager@ecommerce.com
    to_str = ""
    if isinstance(to_addr, str):
        try:
            ta_parsed = json.loads(to_addr)
            if isinstance(ta_parsed, list):
                to_str = ",".join(str(x) for x in ta_parsed)
            else:
                to_str = str(ta_parsed)
        except (TypeError, ValueError, json.JSONDecodeError):
            to_str = to_addr
    elif isinstance(to_addr, list):
        to_str = ",".join(str(x) for x in to_addr)
    else:
        to_str = str(to_addr or "")
    to_lower = to_str.lower()
    to_ok = "manager@ecommerce.com" in to_lower
    record("Email to manager@ecommerce.com", to_ok, f"To: {to_addr}")
    if not to_ok:
        all_ok = False

    # Body should mention all 3 categories AND pricing position (above/below)
    body_lower = (body_text or "").lower()
    cats_in_body = sum(1 for c in ["electronics", "headphones", "speakers"] if c in body_lower)
    body_cat_ok = cats_in_body == 3
    record("Email body mentions all 3 categories", body_cat_ok,
           f"Found {cats_in_body}/3 categories. Body preview: {(body_text or '')[:300]}")
    if not body_cat_ok:
        all_ok = False

    body_pos_ok = any(p in body_lower for p in ["above", "below", "higher", "lower",
                                                "cheaper", "more expensive", "less expensive",
                                                "premium"])
    record("Email body mentions pricing position", body_pos_ok,
           f"Body preview: {(body_text or '')[:300]}")
    if not body_pos_ok:
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    cal_ok = check_calendar()
    email_ok = check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Calendar: {'PASS' if cal_ok else 'FAIL'}")
    print(f"  Email:    {'PASS' if email_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    overall = excel_ok and cal_ok and email_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
