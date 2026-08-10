"""Evaluation for wc-coupon-analysis."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_header(h):
    return str(h or "").strip().lower().replace(" ", "_")


def _header_map(rows):
    """Map normalized header text -> 0-based column index from row 1 of a sheet.
    Lets checks locate columns by name regardless of column order."""
    if not rows or not rows[0]:
        return {}
    return {_norm_header(h): i for i, h in enumerate(rows[0]) if _norm_header(h)}


def _cell(row, idx):
    """Safe cell access (None when idx out of range)."""
    if row is None or idx is None:
        return None
    return row[idx] if idx < len(row) else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Coupon_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Coupon_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Coupon Analysis
    print(f"  Checking Coupon Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Coupon Analysis")
    g_rows = load_sheet_rows(gt_wb, "Coupon Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Coupon Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Coupon Analysis' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Required columns located by header name (order-independent).
        REQ_COLS = ["code", "discount_type", "amount", "usage_count",
                    "usage_limit", "utilization_pct"]
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)
        for cname in REQ_COLS:
            if cname not in a_col:
                errors.append(f"Coupon Analysis missing column '{cname}'")

        # Row count check
        if len(a_data) != len(g_data):
            errors.append(f"Coupon Analysis row count: {len(a_data)} vs {len(g_data)}")

        a_lookup = {}
        for row in a_data:
            code_v = _cell(row, a_col.get("code", 0))
            if code_v is not None:
                a_lookup[str(code_v).strip().lower()] = row
        for g_row in g_data:
            g_code = _cell(g_row, g_col.get("code", 0))
            if g_code is None:
                continue
            key = str(g_code).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_code}")
                continue
            # Discount_Type
            av = _cell(a_row, a_col.get("discount_type"))
            gv = _cell(g_row, g_col.get("discount_type"))
            if av is not None and gv is not None and not str_match(av, gv):
                errors.append(f"{key}.Discount_Type: '{av}' vs '{gv}'")
            # Amount (tol 0.1)
            av = _cell(a_row, a_col.get("amount"))
            gv = _cell(g_row, g_col.get("amount"))
            if av is not None and gv is not None and not num_close(av, gv, 0.1):
                errors.append(f"{key}.Amount: {av} vs {gv}")
            # Usage_Count (tol 0)
            av = _cell(a_row, a_col.get("usage_count"))
            gv = _cell(g_row, g_col.get("usage_count"))
            if av is not None and gv is not None and not num_close(av, gv, 0):
                errors.append(f"{key}.Usage_Count: {av} vs {gv}")
            # Usage_Limit — None (no limit) or specific number
            av = _cell(a_row, a_col.get("usage_limit"))
            gv = _cell(g_row, g_col.get("usage_limit"))
            if gv is None:
                if av is not None and str(av).strip() != "":
                    # Allow "None"/"-"/"" for "no limit"
                    try:
                        if float(av) != 0:
                            errors.append(f"{key}.Usage_Limit: {av} vs None (no limit)")
                    except (TypeError, ValueError):
                        pass  # accept string variants like "-", "None"
            else:
                if av is not None and not num_close(av, gv, 0):
                    errors.append(f"{key}.Usage_Limit: {av} vs {gv}")
            # Utilization_Pct (tol 0.2)
            av = _cell(a_row, a_col.get("utilization_pct"))
            gv = _cell(g_row, g_col.get("utilization_pct"))
            if av is not None and gv is not None and not num_close(av, gv, 0.2):
                errors.append(f"{key}.Utilization_Pct: {av} vs {gv}")

        # Sort: descending by Usage_Count (resolve column by header name, not a
        # fixed positional index, so a column reorder doesn't silently misread).
        try:
            uc_idx = a_col.get("usage_count")
            counts = []
            for r in a_data:
                v = _cell(r, uc_idx) if uc_idx is not None else None
                counts.append(int(float(v)) if v not in (None, "") else 0)
            if counts != sorted(counts, reverse=True):
                errors.append(f"Coupon Analysis not sorted descending by Usage_Count: {counts}")
        except (TypeError, ValueError):
            pass

        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # Tighter tolerances per metric type
                if key in ("total_coupons", "total_usage"):
                    if not num_close(a_row[1], g_row[1], 0):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
                elif key == "avg_utilization":
                    if not num_close(a_row[1], g_row[1], 0.2):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
                elif key == "most_used_code":
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}.Value: '{a_row[1]}' vs '{g_row[1]}'")
                else:
                    if not num_close(a_row[1], g_row[1], 0.5):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    
    docx_path = os.path.join(args.agent_workspace, "Coupon_Strategy.docx")
    if not os.path.exists(docx_path):
        all_errors.append("Coupon_Strategy.docx not found")
    else:
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs)
            _text_lower = _text.lower()
            if len(_text.strip()) < 200:
                all_errors.append(f"Coupon_Strategy.docx has too little text ({len(_text)} chars, need >=200)")
            # Both 'coupon' AND 'strategy'/'recommendation' required
            for kw in ["coupon"]:
                if kw not in _text_lower:
                    all_errors.append(f"Coupon_Strategy.docx missing keyword '{kw}'")
            if not any(k in _text_lower for k in ["strategy", "recommendation", "recommend"]):
                all_errors.append("Coupon_Strategy.docx missing strategy/recommendation content")
            # At least one specific coupon code mentioned (e.g. HOLIDAY30, VIP20)
            specific_codes_found = sum(1 for code in ["HOLIDAY30", "VIP20", "SAVE20", "WELCOME10", "BULK10", "ELECTRONICS15", "FREESHIP", "SUMMER25", "FLASH50", "NEWUSER5"] if code in _text)
            if specific_codes_found < 1:
                all_errors.append("Coupon_Strategy.docx mentions no specific coupon codes")
        except ImportError:
            if os.path.getsize(docx_path) < 200:
                all_errors.append("Coupon_Strategy.docx too small")
        except Exception as _e:
            all_errors.append(f"Error reading Coupon_Strategy.docx: {_e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
