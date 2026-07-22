"""
Evaluation for yt-fireship-monthly-stats-excel-notion task.

Checks:
1. Fireship_Monthly_Stats.xlsx exists with Monthly_Stats and Summary sheets
2. Monthly_Stats has all expected months with correct values
3. Summary sheet has correct Label/Value pairs (all 6 rows)
4. Notion page "Fireship Channel Analysis 2024-2025" exists with content
5. Email sent to analytics@company.com with exact subject
"""
import json
import os
import sys
from collections import defaultdict
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _fetch_expected_stats():
    """Compute expected monthly stats from DB; fallback to known values."""
    fallback = {
        "2024-06": (9, 812167, 34638, 7309503),
        "2024-07": (10, 1356776, 56758, 13567757),
        "2024-08": (10, 1180603, 52977, 11806026),
        "2024-09": (6, 990342, 44147, 5942049),
        "2024-10": (9, 1006886, 45001, 9061971),
        "2024-11": (10, 999280, 38817, 9992799),
        "2024-12": (4, 1230416, 40701, 4921663),
        "2025-01": (9, 1873748, 70688, 16863736),
        "2025-02": (9, 1419322, 52849, 12773898),
        "2025-03": (9, 1349811, 53162, 12148299),
        "2025-04": (6, 975134, 36639, 5850805),
        "2025-05": (6, 975866, 35282, 5855196),
        "2025-06": (7, 827588, 26592, 5793115),
    }
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT published_at, view_count, like_count
            FROM youtube.videos
            WHERE channel_id = 'UCsBjURrPoezykLs9EqgamOA'
              AND published_at IS NOT NULL
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            return fallback
        months = defaultdict(list)
        for ts, vc, lc in rows:
            m = ts.strftime("%Y-%m")
            months[m].append((vc or 0, lc or 0))
        result = {}
        for m, vl in months.items():
            n = len(vl)
            avg_v = round(sum(v for v, _ in vl) / n)
            avg_l = round(sum(l for _, l in vl) / n)
            total_v = sum(v for v, _ in vl)
            result[m] = (n, avg_v, avg_l, total_v)
        return result
    except Exception as e:
        print(f"[fallback] DB fetch failed: {e}")
        return fallback


EXPECTED_STATS = _fetch_expected_stats()


def _norm(s):
    return str(s or "").strip().lower().replace("_", " ")


def check_excel(agent_workspace):
    print("\n=== Check 1: Fireship_Monthly_Stats.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Fireship_Monthly_Stats.xlsx")
    if not os.path.exists(xlsx_path):
        record("Fireship_Monthly_Stats.xlsx exists", False, f"Not found at {xlsx_path}")
        # Record one fail per remaining expected check so the count reflects severity
        record("Excel file readable", False, "missing file")
        record("Monthly_Stats sheet exists", False, "missing file")
        record("Summary sheet exists", False, "missing file")
        return
    record("Fireship_Monthly_Stats.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # ---- Monthly_Stats sheet ----
    if "monthly_stats" not in sheet_names_lower:
        record("Monthly_Stats sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Monthly_Stats sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("monthly_stats")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None and str(c).strip() != "" for c in r)]

        # row count
        record(f"Monthly_Stats has {len(EXPECTED_STATS)} data rows",
               len(data_rows) == len(EXPECTED_STATS),
               f"Found {len(data_rows)} rows, expected {len(EXPECTED_STATS)}")

        # header check
        if rows:
            headers = [_norm(c) for c in rows[0]]
            expected_headers = ["month", "video count", "avg views", "avg likes", "total views"]
            header_ok = all(any(eh in h for h in headers) for eh in expected_headers)
            record("Monthly_Stats headers (Month, Video_Count, Avg_Views, Avg_Likes, Total_Views)",
                   header_ok, f"Headers: {rows[0]}")

        # build map of actual rows by Month
        actual_by_month = {}
        for r in data_rows:
            if r and r[0]:
                key = str(r[0]).strip()
                actual_by_month[key] = r

        # check each expected month exists with correct values
        missing = []
        wrong_value_count = 0
        for m, (vc, av, al, tv) in EXPECTED_STATS.items():
            r = actual_by_month.get(m)
            if r is None:
                missing.append(m)
                continue
            try:
                a_vc = int(r[1]) if r[1] is not None else None
                a_av = int(r[2]) if r[2] is not None else None
                a_al = int(r[3]) if r[3] is not None else None
                a_tv = int(r[4]) if r[4] is not None else None
            except (TypeError, ValueError):
                wrong_value_count += 1
                continue
            # tolerance: video_count exact, avg/total within tiny rounding diff
            ok_vc = a_vc == vc
            ok_av = a_av is not None and abs(a_av - av) <= 2
            ok_al = a_al is not None and abs(a_al - al) <= 2
            ok_tv = a_tv == tv
            if not (ok_vc and ok_av and ok_al and ok_tv):
                wrong_value_count += 1
                if wrong_value_count <= 3:
                    print(f"     [debug] {m} expected={(vc, av, al, tv)} actual={(a_vc, a_av, a_al, a_tv)}")
        record("Monthly_Stats has all expected months",
               not missing, f"Missing months: {missing[:5]}")
        record("Monthly_Stats values match expected (vc/avg_v/avg_l/total)",
               wrong_value_count == 0, f"{wrong_value_count} mismatching rows")

        # check chronological sort
        months_actual = [str(r[0]).strip() for r in data_rows if r and r[0]]
        record("Monthly_Stats sorted ascending by Month",
               months_actual == sorted(months_actual),
               f"first 3: {months_actual[:3]}")

    # ---- Summary sheet ----
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        # Build map Label -> Value (case-insensitive label)
        label_value = {}
        for r in rows2[1:]:
            if not r or len(r) < 2:
                continue
            if r[0] is None:
                continue
            label_value[_norm(r[0])] = r[1]

        # Compute expected from EXPECTED_STATS
        total_videos = sum(v[0] for v in EXPECTED_STATS.values())
        total_months = len(EXPECTED_STATS)
        peak_count_max = max(v[0] for v in EXPECTED_STATS.values())
        peak_candidates = [m for m, v in EXPECTED_STATS.items() if v[0] == peak_count_max]
        best_avg_max = max(v[1] for v in EXPECTED_STATS.values())
        best_avg_candidates = [m for m, v in EXPECTED_STATS.items() if v[1] == best_avg_max]

        def _check_summary_int(label, expected, name=None, tol=0):
            v = label_value.get(_norm(label))
            display = name or label
            if v is None:
                record(f"Summary {display} present", False, f"Label not found in: {list(label_value.keys())}")
                return
            try:
                ok = abs(int(float(v)) - int(expected)) <= tol
            except (TypeError, ValueError):
                ok = False
            record(f"Summary {display} = {expected}", ok, f"Got: {v}")

        def _check_summary_in(label, allowed, name=None):
            v = label_value.get(_norm(label))
            display = name or label
            if v is None:
                record(f"Summary {display} present", False, f"Label not found in: {list(label_value.keys())}")
                return
            ok = str(v).strip() in allowed
            record(f"Summary {display} in {allowed}", ok, f"Got: {v}")

        _check_summary_int("Total_Videos", total_videos, tol=0)
        _check_summary_int("Total_Months_Active", total_months, tol=0)
        _check_summary_in("Peak_Month", peak_candidates)
        _check_summary_int("Peak_Month_Videos", peak_count_max, tol=0)
        _check_summary_in("Best_Avg_Views_Month", best_avg_candidates)
        _check_summary_int("Best_Avg_Views", best_avg_max, tol=2)


def _extract_notion_title(properties):
    """Extract plain-text title from Notion page properties dict."""
    if not isinstance(properties, dict):
        return ""
    for pname, pval in properties.items():
        if not isinstance(pval, dict):
            continue
        ptype = pval.get("type")
        if ptype == "title":
            parts = []
            for item in pval.get("title", []) or []:
                if isinstance(item, dict):
                    t = item.get("plain_text") or (item.get("text") or {}).get("content") or ""
                    parts.append(t)
                elif isinstance(item, str):
                    parts.append(item)
            return "".join(parts).strip()
    return ""


def check_notion():
    print("\n=== Check 2: Notion page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Strict: parse the title property of every page and require exact match.
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties IS NOT NULL
              AND archived = false
              AND COALESCE(in_trash, false) = false
        """)
        all_pages = cur.fetchall()
        target_phrase = "fireship channel analysis 2024-2025"
        target_phrase2 = "fireship channel analysis 2024/2025"
        pages = []
        for pid, props in all_pages:
            title = _extract_notion_title(props).strip().lower()
            if not title:
                continue
            # Accept either dash or slash variant of the year range
            if title == target_phrase or title == target_phrase2 or (
                "fireship channel analysis" in title and ("2024-2025" in title or "2024/2025" in title)
            ):
                pages.append((pid, props))

        record("Notion page titled 'Fireship Channel Analysis 2024-2025' exists (title prop)",
               len(pages) >= 1,
               f"All page titles: {[_extract_notion_title(p[1]) for p in all_pages][:10]}")

        if not pages:
            cur.close()
            conn.close()
            return

        page_id = pages[0][0]
        # Pull page content (blocks)
        cur.execute("""
            SELECT type, content FROM notion.blocks
            WHERE parent_id = %s
        """, (page_id,))
        blocks = cur.fetchall()
        body_text = " ".join(str(b[1]) for b in blocks if b[1]).lower()

        # Compute expected
        total_videos = sum(v[0] for v in EXPECTED_STATS.values())
        total_months = len(EXPECTED_STATS)
        peak_month = max(EXPECTED_STATS.keys(), key=lambda m: EXPECTED_STATS[m][0])
        best_avg_month = max(EXPECTED_STATS.keys(), key=lambda m: EXPECTED_STATS[m][1])

        record("Notion page mentions total video count",
               str(total_videos) in body_text, f"Looking for {total_videos}")
        record("Notion page mentions number of active months",
               str(total_months) in body_text, f"Looking for {total_months}")
        record("Notion page mentions peak publishing month",
               peak_month in body_text or peak_month.replace("-", "/") in body_text,
               f"Looking for {peak_month}")
        record("Notion page mentions best avg views month",
               best_avg_month in body_text or best_avg_month.replace("-", "/") in body_text,
               f"Looking for {best_avg_month}")

        cur.close()
        conn.close()
    except Exception as e:
        record("Notion page check completed", False, f"DB error: {e}")


def check_email():
    print("\n=== Check 3: Email sent ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr, subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s
            ORDER BY date DESC NULLS LAST LIMIT 10
        """, ("%analytics@company.com%",))
        emails = cur.fetchall()

        cur.close()
        conn.close()

        record("Email sent to analytics@company.com", len(emails) >= 1,
               f"Found {len(emails)} emails")

        if not emails:
            record("Email subject is exactly 'Fireship Channel Analysis Complete'", False, "no email")
            record("Email body summarizes findings", False, "no email")
            return

        # Find an email matching exact subject (per task)
        target_subject = "fireship channel analysis complete"
        match_email = None
        for to_addr, subj, body in emails:
            if subj and subj.strip().lower() == target_subject:
                match_email = (to_addr, subj, body)
                break
        if match_email is None:
            # accept the most recent and report subject mismatch
            match_email = emails[0]
        to_addr, subj, body = match_email

        subj_lower = (subj or "").strip().lower()
        record("Email subject is exactly 'Fireship Channel Analysis Complete'",
               subj_lower == target_subject, f"Got: {subj!r}")

        body_lower = (body or "").lower()
        total_videos = sum(v[0] for v in EXPECTED_STATS.values())
        peak_count_max = max(v[0] for v in EXPECTED_STATS.values())
        peak_candidates = [m for m, v in EXPECTED_STATS.items() if v[0] == peak_count_max]
        best_avg_max = max(v[1] for v in EXPECTED_STATS.values())
        best_avg_candidates = [m for m, v in EXPECTED_STATS.items() if v[1] == best_avg_max]

        def _mentions_any(months, text):
            for m in months:
                if m in text or m.replace("-", "/") in text:
                    return True
            return False

        # Require the total_videos number to appear NEAR a video-related keyword
        # to avoid coincidental integer matches.
        import re as _re
        ok_total = False
        for m in _re.finditer(r"\b" + _re.escape(str(total_videos)) + r"\b", body_lower):
            window = body_lower[max(0, m.start() - 80): m.end() + 80]
            if "video" in window or "total" in window:
                ok_total = True
                break
        ok_peak = _mentions_any(peak_candidates, body_lower)
        ok_best = _mentions_any(best_avg_candidates, body_lower)

        record(f"Email body mentions total videos ({total_videos}) near 'video'/'total'",
               ok_total, f"body: {body_lower[:200]}")
        record(f"Email body mentions peak month (any of {peak_candidates})", ok_peak)
        record(f"Email body mentions best avg month (any of {best_avg_candidates})", ok_best)
    except Exception as e:
        record("Email check completed", False, f"DB error: {e}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_notion()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
