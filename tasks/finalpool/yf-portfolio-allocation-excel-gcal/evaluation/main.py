"""Evaluation for yf-portfolio-allocation-excel-gcal."""
import os
import argparse, os, sys

import psycopg2
from zoneinfo import ZoneInfo

# ──────────────────────────────────────────────────────────────────────────
# EVALUATION GROUND TRUTH SPEC (gcal tz root-fix v3, case-study 2026-08-13)
# gcal.events.start_datetime / end_datetime are TIMESTAMPTZ storing the UTC
# instant. psycopg2 returns a tz-aware datetime whose display tz follows the
# PG session TimeZone (case-study: compute node default Asia/Shanghai), so
# bare sd.hour / sd.date() silently compares against the wrong wall-clock.
# Use utils.evaluation.gcal_helpers (session-tz-independent, DST-aware).
# ──────────────────────────────────────────────────────────────────────────
# task.md line 9: schedule "Portfolio Rebalancing" on April 30, 2026 from
# 10:00 to 11:30 (portfolio / financial context → America/New_York / ET).
from utils.evaluation.gcal_helpers import get_zone_components  # noqa: E402

EXPECTED_TIMEZONE = ZoneInfo("America/New_York")
EXPECTED_START_HOUR = 10
EXPECTED_START_MINUTE = 0
EXPECTED_END_HOUR = 11
EXPECTED_END_MINUTE = 30


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_pdf(agent_workspace):
    errors = []
    pdf_path = os.path.join(agent_workspace, "Portfolio_Guidelines.pdf")
    if not os.path.exists(pdf_path):
        errors.append("Portfolio_Guidelines.pdf not found in agent workspace")
    return errors


# Expected weights per task.md (these are explicitly listed in the prompt)
EXPECTED_WEIGHTS = {"GOOGL": 25.0, "AMZN": 25.0, "JPM": 20.0, "JNJ": 15.0, "XOM": 15.0}


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Portfolio_Allocation.xlsx")
    if not os.path.exists(path):
        return ["Portfolio_Allocation.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        # Check Stock Analysis sheet
        rows = load_sheet_rows(wb, "Stock Analysis")
        if rows is None:
            errors.append("Sheet 'Stock Analysis' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) != 5:
                errors.append(f"Stock Analysis has {len(data_rows)} rows, expected exactly 5")
            # Check all expected symbols and their weights
            sym_to_row = {}
            for r in data_rows:
                if r and r[0]:
                    sym_to_row[str(r[0]).strip().upper()] = r
            for sym, expected_w in EXPECTED_WEIGHTS.items():
                if sym not in sym_to_row:
                    errors.append(f"Symbol {sym} missing from Stock Analysis")
                    continue
                row = sym_to_row[sym]
                weight = row[4] if len(row) > 4 else None
                if not num_close(weight, expected_w, 0.5):
                    errors.append(f"{sym} Allocated_Weight_Pct={weight}, expected {expected_w}")
            # Check sort order (alphabetical by Symbol)
            actual_syms = [str(r[0]).strip().upper() for r in data_rows if r and r[0]]
            if actual_syms != sorted(actual_syms):
                errors.append(f"Stock Analysis not sorted alphabetically by Symbol: {actual_syms}")

        # Check Allocation Summary sheet
        rows2 = load_sheet_rows(wb, "Allocation Summary")
        if rows2 is None:
            errors.append("Sheet 'Allocation Summary' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows2 if r[0]}
            # Total_Invested == 100000 (tighten tol)
            if "total_invested" in lookup:
                if not num_close(lookup["total_invested"], 100000, 1):
                    errors.append(f"Total_Invested={lookup['total_invested']}, expected 100000")
            else:
                errors.append("Total_Invested not found in Allocation Summary")
            # Strong_Buy_Count == 2 (per the GT yf data: AMZN+GOOGL = strong_buy)
            if "strong_buy_count" in lookup:
                if not num_close(lookup["strong_buy_count"], 2, 0):
                    errors.append(f"Strong_Buy_Count={lookup['strong_buy_count']}, expected 2")
            else:
                errors.append("Strong_Buy_Count not found in Allocation Summary")
            # Buy_Count == 3
            if "buy_count" in lookup:
                if not num_close(lookup["buy_count"], 3, 0):
                    errors.append(f"Buy_Count={lookup['buy_count']}, expected 3")
            else:
                errors.append("Buy_Count not found in Allocation Summary")
            # Weighted_Avg_Price ~ 247.22 (allow 0.5 tolerance)
            if "weighted_avg_price" in lookup:
                if not num_close(lookup["weighted_avg_price"], 247.22, 0.5):
                    errors.append(f"Weighted_Avg_Price={lookup['weighted_avg_price']}, expected ~247.22")
            else:
                errors.append("Weighted_Avg_Price not found in Allocation Summary")
    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime, description FROM gcal.events
            WHERE start_datetime >= '2026-04-29'
              AND start_datetime <  '2026-05-01'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No GCal event found near 2026-04-30")
        else:
            # Find the rebalancing event
            target = None
            for r in rows:
                s = (r[0] or "").lower()
                if "portfolio" in s and ("rebalanc" in s):
                    target = r
                    break
            if target is None:
                # Fallback: any portfolio/rebalance event
                for r in rows:
                    s = (r[0] or "").lower()
                    if "portfolio" in s or "rebalanc" in s:
                        target = r
                        break
            if target is None:
                errors.append(f"No portfolio rebalancing event near 2026-04-30 (found: {[r[0] for r in rows]})")
            else:
                # gcal.events.start_datetime is TIMESTAMPTZ (UTC instant). The PG
                # session TimeZone (case-study: Asia/Shanghai) changes the display
                # wall-clock, so bare start_dt.hour/.minute would miscompare.
                start_dt, end_dt, desc = target[1], target[2], target[3]
                if start_dt is not None:
                    sd_date, sd_hour, sd_minute = get_zone_components(start_dt, EXPECTED_TIMEZONE)
                    if sd_date is None \
                       or sd_date.isoformat() != "2026-04-30" \
                       or sd_hour != EXPECTED_START_HOUR \
                       or sd_minute != EXPECTED_START_MINUTE:
                        errors.append(f"GCal event start time={start_dt}, expected 2026-04-30 10:00 ET")
                if end_dt is not None:
                    ed_date, ed_hour, ed_minute = get_zone_components(end_dt, EXPECTED_TIMEZONE)
                    if ed_date is None \
                       or ed_hour != EXPECTED_END_HOUR \
                       or ed_minute != EXPECTED_END_MINUTE:
                        errors.append(f"GCal event end time={end_dt}, expected 2026-04-30 11:30 ET")
                # Verify description mentions the 5 stock symbols and the $100,000 total
                desc_low = (desc or "").lower()
                missing_syms = [sym for sym in ("googl", "amzn", "jpm", "jnj", "xom") if sym not in desc_low]
                if missing_syms:
                    errors.append(f"GCal description missing stock symbol(s): {missing_syms}")
                if "100,000" not in desc_low and "100000" not in desc_low and "$100" not in desc_low:
                    errors.append("GCal description should mention the $100,000 total portfolio value")
    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")), dbname="toolathlon_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%portfolio_manager@wealth.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No email found to portfolio_manager@wealth.com")
        else:
            # Verify subject contains literal "Portfolio Allocation Plan" and "April 2026"
            subj_match = False
            target_body = None
            for r in rows:
                s = (r[0] or "").lower()
                if "portfolio allocation plan" in s and "april" in s and "2026" in s:
                    subj_match = True
                    target_body = (r[2] or "").lower()
                    break
            if not subj_match:
                errors.append(f"No email with subject containing 'Portfolio Allocation Plan - April 2026' (found: {[r[0] for r in rows]})")
            else:
                # Validate body summarizes the allocation strategy
                if target_body is not None:
                    has_strategy = any(kw in target_body for kw in ("alloc", "weight", "portfolio"))
                    has_metric = any(kw in target_body for kw in ("100,000", "100000", "weighted", "buy"))
                    if not (has_strategy and has_metric):
                        errors.append(
                            f"Email body should summarize allocation strategy and key metrics; got: {target_body[:200]}"
                        )
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking PDF in workspace...")
    errs = check_pdf(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal event...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
