"""Evaluation script for canvas-grades-gsheet-pdf-email."""
import os
import argparse, json, os, sys, re
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


def _env(*names, default):
    """Read the first defined env var among ``names`` (supports both the
    PGHOST/PGPORT/... dash form used by the harness runtime and the
    PG_HOST/PG_PORT/... underscore form used by the MCP servers)."""
    for n in names:
        v = os.environ.get(n)
        if v:
            return v
    return default


DB_CONFIG = {
    "host": _env("PGHOST", "PG_HOST", default="localhost"),
    "port": int(_env("PGPORT", "PG_PORT", default="5432")),
    "dbname": _env("PGDATABASE", "PG_DATABASE", default="toolathlon_gym"),
    "user": _env("PGUSER", "PG_USER", default="eigent"),
    "password": _env("PGPASSWORD", "PG_PASSWORD", default="camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


# ---------------------------------------------------------------------------
# Course-name matching helpers.
# LMS course names carry a semester suffix, e.g.
#   "Environmental Economics & Ethics (Fall 2014)".
# A model that completes the task correctly may write only the course family
# ("Environmental Economics & Ethics") or use "&"/"and" interchangeably, both
# in the email body and in the Department_Summary highest/lowest rows.  These
# helpers accept the full name, the family name, and &/and spelling variants so
# a correct model is never penalised for phrasing.
# ---------------------------------------------------------------------------
def _course_family(name):
    """Return the course name with a trailing '(Fall|Spring|Summer|Winter YYYY)'
    parenthetical removed (identity if there is none)."""
    s = str(name or "").strip()
    s = re.sub(r"\s*\((?:fall|spring|summer|winter)\s*20\d\d\)\s*$", "", s, flags=re.IGNORECASE)
    return s


def mentions_course(body_l, course_name):
    """True if the (lowercased) email body mentions the course, allowing the
    family name without semester suffix and '&'/'and' spelling variants."""
    if not course_name:
        return False
    full = str(course_name).lower().strip()
    fam = _course_family(course_name).lower().strip()
    candidates = set(f for f in (full, fam) if f)
    if any(c in body_l for c in candidates):
        return True
    body_and = body_l.replace("&", "and")
    return any(c.replace("&", "and") in body_and for c in candidates)


def course_names_match(a, b):
    """Tolerant equality for course-name cells (highest/lowest rows): exact
    match, family-name match, &/and variants, or one a distinctive substring of
    the other (truncated family name)."""
    if a is None or b is None:
        return False
    a, b = str(a).strip().lower(), str(b).strip().lower()
    if a == b:
        return True

    def norm(s):
        s = _course_family(s).lower().replace("&", "and")
        return re.sub(r"\s+", " ", s).strip()

    na, nb = norm(a), norm(b)
    if na and nb and na == nb:
        return True
    if len(na) >= 5 and len(nb) >= 5 and (na in nb or nb in na):
        return True
    return False

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def get_gt_data(groundtruth_workspace):
    """Load reference data from groundtruth Grade_Dashboard_Reference.xlsx."""
    gt_path = os.path.join(groundtruth_workspace, "Grade_Dashboard_Reference.xlsx")
    if not os.path.exists(gt_path):
        return None
    wb = openpyxl.load_workbook(gt_path, data_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            out[sheet_name] = {"headers": headers, "rows": [tuple(r) for r in rows[1:]]}
    return out


# ---------------------------------------------------------------------------
# Expected-value computation from the shared DB (immutable seed).
# Rule (mirrors docs/task.md): for each course, take every student enrollment's
# grades.final_score as that student's overall score. Students with no recorded
# final score are excluded. Bucket with A>=90, B 80-89, C 70-79, D 60-69, F<60.
# ---------------------------------------------------------------------------
def _extract_final_score(grades):
    if grades is None:
        return None
    if isinstance(grades, dict):
        fs = grades.get("final_score")
    else:
        try:
            fs = json.loads(grades).get("final_score")
        except (TypeError, ValueError):
            return None
    if fs is None:
        return None
    try:
        return float(fs)
    except (TypeError, ValueError):
        return None


def compute_expected_from_db(cur):
    """Compute expected Grade_Distribution + Department_Summary from canvas seed.
    Returns (expected_dict, summary_dict). expected_dict has the same shape as
    get_gt_data() output so check_gsheet_against_gt() can consume it."""
    cur.execute("SELECT id, name FROM canvas.courses ORDER BY name")
    courses = cur.fetchall()
    grade_rows = []
    total_students = 0
    weighted_sum = 0.0
    pass_count_total = 0
    course_avgs = {}
    for cid, name in courses:
        cur.execute("SELECT grades FROM canvas.enrollments WHERE course_id = %s", (cid,))
        scores = []
        for (g,) in cur.fetchall():
            fs = _extract_final_score(g)
            if fs is not None:
                scores.append(fs)
        n = len(scores)
        a = sum(1 for s in scores if s >= 90)
        b = sum(1 for s in scores if 80 <= s < 90)
        c = sum(1 for s in scores if 70 <= s < 80)
        d = sum(1 for s in scores if 60 <= s < 70)
        f = sum(1 for s in scores if s < 60)
        avg = sum(scores) / n if n else 0.0
        pass_rate = round(100.0 * (a + b + c) / n, 1) if n else 0.0
        grade_rows.append((name, a, b, c, d, f, n, pass_rate, round(avg, 1)))
        total_students += n
        weighted_sum += avg * n
        pass_count_total += (a + b + c)
        course_avgs[name] = avg

    summary = {
        "Total_Courses": len(grade_rows),
        "Total_Students": total_students,
        "Overall_Pass_Rate": round(100.0 * pass_count_total / total_students, 1) if total_students else 0.0,
        "Overall_Avg_Grade": round(weighted_sum / total_students, 1) if total_students else 0.0,
        "Highest_Avg_Course": max(course_avgs, key=course_avgs.get) if course_avgs else None,
        "Lowest_Avg_Course": min(course_avgs, key=course_avgs.get) if course_avgs else None,
    }
    expected = {
        "Grade_Distribution": {
            "headers": ["course_name", "a_count", "b_count", "c_count", "d_count",
                        "f_count", "total_students", "pass_rate_pct", "course_avg"],
            "rows": grade_rows,
        },
        "Department_Summary": {
            "headers": ["metric", "value"],
            "rows": [
                ("Total_Courses", summary["Total_Courses"]),
                ("Total_Students", summary["Total_Students"]),
                ("Overall_Pass_Rate", summary["Overall_Pass_Rate"]),
                ("Overall_Avg_Grade", summary["Overall_Avg_Grade"]),
                ("Highest_Avg_Course", summary["Highest_Avg_Course"]),
                ("Lowest_Avg_Course", summary["Lowest_Avg_Course"]),
            ],
        },
    }
    return expected, summary


def check_gsheet_against_gt(cur, gt_data):
    """Verify the 'Department Grade Dashboard' Google Sheet matches the reference data."""
    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    sheets = cur.fetchall()
    target_id = None
    target_title = None
    for sid, title in sheets:
        if title and "department grade dashboard" in title.strip().lower():
            target_id = sid
            target_title = title
            break
    check("Google Sheet 'Department Grade Dashboard' exists", target_id is not None,
          f"got: {[t for _, t in sheets]}")
    if target_id is None:
        return

    # List sheets / tabs
    cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s", (target_id,))
    tabs = cur.fetchall()
    tab_titles_lower = {t[1].strip().lower(): t[0] for t in tabs if t[1]}

    for required in ("grade_distribution", "department_summary"):
        present = any(required.replace("_", "") == k.replace("_", "").replace(" ", "")
                      or required in k for k in tab_titles_lower)
        check(f"Tab '{required}' present in Google Sheet", present,
              f"tabs: {list(tab_titles_lower)}")

    # Fetch all cells for the spreadsheet
    cur.execute(
        "SELECT sheet_id, row_index, col_index, value FROM gsheet.cells WHERE spreadsheet_id = %s",
        (target_id,)
    )
    cells = cur.fetchall()
    by_sheet = {}
    for sheet_id, ri, ci, val in cells:
        by_sheet.setdefault(sheet_id, {})[(ri, ci)] = val

    # For each GT sheet, locate the matching tab and compare
    for gt_sheet_name, info in (gt_data or {}).items():
        sheet_key = gt_sheet_name.strip().lower().replace("_", "")
        match_id = None
        for tab_title, tid in tab_titles_lower.items():
            if tab_title.replace("_", "").replace(" ", "") == sheet_key:
                match_id = tid; break
        if match_id is None:
            continue  # already flagged above
        cell_map = by_sheet.get(match_id, {})
        if not cell_map:
            check(f"Tab '{gt_sheet_name}' has cells", False, "no cells found")
            continue
        # Build header -> col_index from cell row 0
        header_cols = {}
        for (r, c), v in cell_map.items():
            if r == 0 and v:
                header_cols[str(v).strip().lower()] = c
        gt_headers = info["headers"]
        for h in gt_headers:
            if h and h not in header_cols:
                check(f"GSheet '{gt_sheet_name}' has column '{h}'", False,
                      f"got: {list(header_cols)}")
        # Compare every GT row by Course_Name (or Metric)
        # Find key column index in cell_map
        key_label = gt_headers[0]
        key_col = header_cols.get(key_label)
        if key_col is None:
            check(f"GSheet '{gt_sheet_name}' has key column '{key_label}'", False)
            continue
        # Build agent_lookup: {key_lower: {col_label: val}}
        agent_lookup = {}
        max_row = max((r for (r, c) in cell_map.keys()), default=0)
        for r in range(1, max_row + 1):
            kval = cell_map.get((r, key_col))
            if kval is None:
                continue
            key_lower = str(kval).strip().lower()
            row_data = {}
            for label, c in header_cols.items():
                row_data[label] = cell_map.get((r, c))
            agent_lookup[key_lower] = row_data

        for gt_row in info["rows"]:
            if not gt_row or gt_row[0] is None:
                continue
            key_lower = str(gt_row[0]).strip().lower()
            agent_row = agent_lookup.get(key_lower)
            if agent_row is None:
                check(f"GSheet '{gt_sheet_name}' row '{gt_row[0]}' present", False)
                continue
            for ci, h in enumerate(gt_headers):
                if ci == 0 or not h or ci >= len(gt_row):
                    continue
                gv = gt_row[ci]
                av = agent_row.get(h)
                gf = safe_float(gv)
                af = safe_float(av)
                if gf is not None and af is not None:
                    tol = max(0.5, abs(gf) * 0.05)
                    if abs(gf - af) > tol:
                        check(f"GSheet '{gt_sheet_name}' '{gt_row[0]}' {h}={gf:.1f}",
                              False, f"got {af}")
                elif gv is not None:
                    gs = str(gv).strip().lower()
                    avs = str(av or "").strip().lower()
                    if gs and gs != avs:
                        # Highest_Avg_Course / Lowest_Avg_Course values are
                        # course names; tolerate family names without the
                        # semester suffix and &/and spelling variants.
                        if key_lower in ("highest_avg_course", "lowest_avg_course"):
                            match_ok = course_names_match(gs, avs)
                        else:
                            match_ok = False
                        if not match_ok:
                            check(f"GSheet '{gt_sheet_name}' '{gt_row[0]}' {h}",
                                  False, f"expected '{gs[:50]}', got '{avs[:50]}'")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    gt_data = get_gt_data(groundtruth_workspace)

    # Check Python script exists (terminal usage required by task)
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")
    # Specifically grade_reporter.py mentioned in task
    check("grade_reporter.py exists",
          any("grade_reporter" in f for f in py_files),
          f"py files: {py_files}")
    # Output JSON files mentioned in task
    has_grade_data = os.path.exists(os.path.join(agent_workspace, "grade_data.json"))
    has_grade_report = os.path.exists(os.path.join(agent_workspace, "grade_report.json"))
    check("grade_data.json exists", has_grade_data)
    check("grade_report.json exists", has_grade_report)

    # Database checks
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Expected values: prefer self-computed from the immutable canvas seed
        # (deterministic); fall back to the static GT xlsx if the canvas tables
        # are unavailable in this database.
        expected = None
        db_summary = None
        try:
            expected, db_summary = compute_expected_from_db(cur)
        except Exception as e:
            print(f"  [warn] DB self-compute of expected values failed, using GT xlsx: {e}")
        if expected is None:
            expected = gt_data

        # Email check - require exact subject and recipient.
        # Match any Sent-like folder (seed has 'Sent', 'SENT', 'Sent Messages',
        # 'INBOX.Sent', 'Sent Items') so the check works whichever name the
        # emails MCP writes sent messages under.
        cur.execute(
            "SELECT subject, to_addr, body_text FROM email.messages "
            "WHERE folder_id IN (SELECT id FROM email.folders WHERE LOWER(name) LIKE '%sent%')"
        )
        sent = cur.fetchall()
        target_subj = "q1 2026 grade distribution report"
        match = None
        for subj, to_addr, body in sent:
            subj_l = (subj or "").lower()
            to_l = str(to_addr or "").lower()
            if target_subj in subj_l and "dept-heads@university.edu" in to_l:
                match = (subj, to_addr, body)
                break
        check("Email to dept-heads@university.edu with correct subject", match is not None,
              f"{len(sent)} sent emails")
        if match is not None:
            body_l = (match[2] or "").lower()
            check("Email body mentions overall pass rate",
                  "pass" in body_l and ("rate" in body_l or "%" in body_l))
            # Highest / lowest performing courses, taken from the expected data.
            highest_name = None
            lowest_name = None
            if db_summary is not None:
                highest_name = db_summary.get("Highest_Avg_Course")
                lowest_name = db_summary.get("Lowest_Avg_Course")
            elif gt_data and "Department_Summary" in gt_data:
                for row in gt_data["Department_Summary"]["rows"]:
                    if row and len(row) >= 2:
                        label = str(row[0] or "").strip().lower()
                        val = str(row[1] or "").strip()
                        if label == "highest_avg_course":
                            highest_name = val
                        elif label == "lowest_avg_course":
                            lowest_name = val
            if highest_name:
                check(f"Email body mentions highest course '{highest_name}'",
                      mentions_course(body_l, highest_name))
            else:
                check("Email body mentions highest course (label-only fallback)",
                      "highest" in body_l)
            if lowest_name:
                check(f"Email body mentions lowest course '{lowest_name}'",
                      mentions_course(body_l, lowest_name))
            else:
                check("Email body mentions lowest course (label-only fallback)",
                      "lowest" in body_l)
            # Courses with pass rate below 70% require review (task requirement).
            # Accept the common phrasings a correct model may use for "below 70%".
            review_signal = any(k in body_l for k in
                                ("review", "below", "under", "less", "threshold",
                                 "attention", "flag", "investigate"))
            check("Email flags courses below 70% pass rate for review",
                  "70" in body_l and review_signal)

        # Google Sheet check using expected values
        if expected is not None:
            check_gsheet_against_gt(cur, expected)
        else:
            cur.execute("SELECT COUNT(*) FROM gsheet.spreadsheets")
            check("Google Sheet created", cur.fetchone()[0] >= 1)

        # Reverse verification: noise emails should not be forwarded
        cur.execute(
            "SELECT COUNT(*) FROM email.messages "
            "WHERE folder_id IN (SELECT id FROM email.folders WHERE LOWER(name) LIKE '%sent%') "
            "AND subject ILIKE '%newsletter%'"
        )
        noise_sent = cur.fetchone()[0]
        check("No noise emails forwarded", noise_sent == 0,
              f"found {noise_sent} noise emails in Sent")

        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
