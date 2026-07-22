"""Evaluation script for pw-sf-support-sla-benchmark-excel-gcal."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "SLA_Benchmark_Report.xlsx")
    check("SLA_Benchmark_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "SLA_Benchmark_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        # Helper: per-row validation by Priority
        def _validate_rows_by_priority(sheet_name, gt_check_cols, tol_pct=0.05):
            if sheet_name not in wb.sheetnames or not gt_wb or sheet_name not in gt_wb.sheetnames:
                return
            ws = wb[sheet_name]
            gt_ws = gt_wb[sheet_name]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
            agent_by = {str(r[0]).strip().lower(): r for r in data_rows if r and r[0]}
            for gt_row in gt_rows:
                if not gt_row or not gt_row[0]:
                    continue
                pri = str(gt_row[0]).strip().lower()
                arow = agent_by.get(pri)
                check(f"{sheet_name} has priority '{pri}'", arow is not None)
                if arow is None:
                    continue
                for col in gt_check_cols:
                    if col not in gt_headers or col not in headers:
                        continue
                    gv = gt_row[gt_headers.index(col)]
                    av = arow[headers.index(col)]
                    gf = safe_float(gv)
                    af = safe_float(av)
                    if gf is not None and af is not None:
                        tol = max(0.5, abs(gf) * tol_pct)
                        check(f"{sheet_name} '{pri}' {col} ~{gf}",
                              abs(gf - af) <= tol, f"got {av}")
                    elif gv is not None and av is not None:
                        check(f"{sheet_name} '{pri}' {col} text",
                              str(gv).strip().lower() == str(av).strip().lower(),
                              f"expected {gv}, got {av}")

        def _gt_rows_count(name):
            if gt_wb and name in gt_wb.sheetnames:
                return len(list(gt_wb[name].iter_rows(min_row=2, values_only=True)))
            return 0

        # Sort order helper for SLA priorities
        _PRIO_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}

        check("SLA_Comparison sheet exists", "SLA_Comparison" in wb.sheetnames)
        if "SLA_Comparison" in wb.sheetnames:
            ws = wb["SLA_Comparison"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_sc = _gt_rows_count("SLA_Comparison")
            min_sc = max(3, gt_sc)
            if gt_sc:
                check(f"SLA_Comparison has == {min_sc} rows",
                      len(data_rows) == min_sc, f"got {len(data_rows)}")
            else:
                check(f"SLA_Comparison has >= {min_sc} rows",
                      len(data_rows) >= min_sc, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Ticket_Count', 'Our_Avg_Response_Hrs',
                                 'Industry_Avg_Response_Hrs', 'Response_Gap',
                                 'Avg_CSAT', 'Compliance_Status']:
                check(f"SLA_Comparison has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            _validate_rows_by_priority("SLA_Comparison",
                ['ticket_count', 'our_avg_response_hrs',
                 'industry_avg_response_hrs', 'response_gap',
                 'avg_csat', 'compliance_status'])

            # Sort: priorities in Critical, High, Medium, Low order
            prio_seq = [str(r[0]).strip().lower() for r in data_rows if r and r[0]]
            prio_keys = [_PRIO_ORDER.get(p, 99) for p in prio_seq]
            check("SLA_Comparison sorted by priority Critical>High>Medium>Low",
                  prio_keys == sorted(prio_keys), f"order={prio_seq}")

        check("Action_Items sheet exists", "Action_Items" in wb.sheetnames)
        if "Action_Items" in wb.sheetnames:
            ws = wb["Action_Items"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_ai = _gt_rows_count("Action_Items")
            min_ai = max(3, gt_ai)
            if gt_ai:
                check(f"Action_Items has == {min_ai} rows",
                      len(data_rows) == min_ai, f"got {len(data_rows)}")
            else:
                check(f"Action_Items has >= {min_ai} rows",
                      len(data_rows) >= min_ai, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Response_Gap', 'Improvement_Needed_Pct',
                                 'Recommended_Action']:
                check(f"Action_Items has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            _validate_rows_by_priority("Action_Items",
                ['response_gap', 'improvement_needed_pct', 'recommended_action'])

        check("Summary sheet exists", "Summary" in wb.sheetnames)
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            gt_su = _gt_rows_count("Summary")
            min_su = max(5, gt_su)
            if gt_su:
                check(f"Summary has == {min_su} rows",
                      len(data_rows) == min_su, f"got {len(data_rows)}")
            else:
                check(f"Summary has >= {min_su} rows",
                      len(data_rows) >= min_su, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Validate Summary metric values
            if gt_wb and "Summary" in gt_wb.sheetnames:
                gt_ws = gt_wb["Summary"]
                gt_metrics = {str(r[0]).strip().lower(): r[1] for r in gt_ws.iter_rows(min_row=2, values_only=True) if r and r[0]}
                agent_metrics = {str(r[0]).strip().lower(): r[1] for r in data_rows if r and r[0]}
                for gt_m, gt_v in gt_metrics.items():
                    found = gt_m in agent_metrics
                    check(f"Summary has '{gt_m}'", found)
                    if found:
                        av = agent_metrics[gt_m]
                        gf = safe_float(gt_v)
                        af = safe_float(av)
                        if gf is not None and af is not None:
                            tol = max(1.0, abs(gf) * 0.05)
                            check(f"Summary '{gt_m}' ~{gf}",
                                  abs(gf - af) <= tol, f"got {av}")
                        elif gt_v is not None and av is not None:
                            check(f"Summary '{gt_m}' text",
                                  str(gt_v).strip().lower() == str(av).strip().lower(),
                                  f"expected {gt_v}, got {av}")

    # --- Calendar events: validate timing too (March 14 2-3:30 PM UTC, March 21 10 AM-12 PM UTC) ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Q1 SLA Review Meeting: anchor title with the leading 'Q1' qualifier so
        # 'My SLA Review notes' won't match.
        cur.execute(
            "SELECT summary, start_datetime, end_datetime, description FROM gcal.events "
            "WHERE summary ILIKE %s OR summary ILIKE %s",
            ('Q1 SLA Review%', '%Q1 SLA Review Meeting%'))
        review_events = cur.fetchall()
        check("Q1 SLA Review Meeting event created",
              len(review_events) >= 1, f"found {len(review_events)}")
        if review_events:
            ev = review_events[0]
            sd = str(ev[1]) if ev[1] else ""
            ed = str(ev[2]) if ev[2] else ""
            desc = str(ev[3] or "")
            check("SLA Review event date is 2026-03-14",
                  "2026-03-14" in sd, f"start_datetime: {sd}")
            # Hour 14 (2 PM) and end at 15:30
            check("SLA Review starts at 14:00",
                  "14:00" in sd or "T14:" in sd, f"start_datetime: {sd}")
            check("SLA Review ends at 15:30",
                  "15:30" in ed or "T15:30" in ed, f"end_datetime: {ed}")
            check("SLA Review description has substantive content (>30 chars)",
                  len(desc.strip()) > 30, f"desc len {len(desc)}")

        cur.execute(
            "SELECT summary, start_datetime, end_datetime, description FROM gcal.events "
            "WHERE summary ILIKE %s",
            ('SLA Improvement Workshop%',))
        wk_events = cur.fetchall()
        check("SLA Improvement Workshop event created",
              len(wk_events) >= 1, f"found {len(wk_events)}")
        if wk_events:
            ev = wk_events[0]
            sd = str(ev[1]) if ev[1] else ""
            ed = str(ev[2]) if ev[2] else ""
            wdesc = str(ev[3] or "")
            check("SLA Workshop event date is 2026-03-21",
                  "2026-03-21" in sd, f"start_datetime: {sd}")
            check("SLA Workshop starts at 10:00",
                  "10:00" in sd or "T10:" in sd, f"start_datetime: {sd}")
            check("SLA Workshop ends at 12:00",
                  "12:00" in ed or "T12:" in ed, f"end_datetime: {ed}")
            check("SLA Workshop description has substantive content (>20 chars)",
                  len(wdesc.strip()) > 20, f"desc len {len(wdesc)}")

        conn.close()
    except Exception as e:
        check("Calendar verification", False, str(e))

    check("sla_analyzer.py exists",
          os.path.exists(os.path.join(agent_workspace, "sla_analyzer.py")))
    for jname in ("sla_raw_data.json", "sla_results.json"):
        jpath = os.path.join(agent_workspace, jname)
        check(f"{jname} exists", os.path.exists(jpath))
        if os.path.exists(jpath):
            try:
                with open(jpath, "r") as _f:
                    _d = json.load(_f)
                check(f"{jname} is valid JSON and non-empty", bool(_d), "empty/null")
            except Exception as _e:
                check(f"{jname} is valid JSON and non-empty", False, str(_e))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
