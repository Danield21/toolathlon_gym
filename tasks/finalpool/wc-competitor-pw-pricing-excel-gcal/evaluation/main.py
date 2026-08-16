"""Evaluation script for wc-competitor-pw-pricing-excel-gcal."""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Check Competitive_Pricing_Analysis.xlsx
    excel_path = os.path.join(agent_workspace, "Competitive_Pricing_Analysis.xlsx")
    check("Competitive_Pricing_Analysis.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Competitive_Pricing_Analysis.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        if gt_wb:
            for sheet_name in gt_wb.sheetnames:
                check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    gt_ws = gt_wb[sheet_name]
                    # Check headers
                    gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                    for h in gt_headers:
                        if h:
                            check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                    # Check row count - exact match
                    gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                               if any(c is not None for c in r)]
                    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                                 if any(c is not None for c in r)]
                    check(f"{sheet_name} row count == {len(gt_rows)}",
                          len(data_rows) == len(gt_rows), f"got {len(data_rows)}")

                    # Cell value comparison against groundtruth - iterate ALL rows
                    header_map = {h: i for i, h in enumerate(headers)}

                    # Build agent lookup by Product_Name / Metric / Product (col 0) for stable
                    # alignment regardless of row order
                    a_lookup = {}
                    for row in data_rows:
                        if row and row[0] is not None:
                            a_lookup[str(row[0]).strip().lower()] = row

                    for ri, gt_row in enumerate(gt_rows):
                        if gt_row[0] is None:
                            continue
                        key = str(gt_row[0]).strip().lower()
                        agent_row = a_lookup.get(key)
                        if agent_row is None:
                            check(f"{sheet_name} row '{key}' present", False,
                                  f"agent missing key '{key}'")
                            continue
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row):
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                # Market_Position values (Products_Above/Below/At, Avg_Gap_Pct, Revenue_At_Risk)
                                # depend on agent's full-catalog classification; widen tolerance.
                                # Action_Items Suggested_Price is also derivative.
                                if sheet_name.lower() == "market_position":
                                    # Tightened: 15% relative or 2.0 absolute floor (was 50% / 5.0)
                                    tol = max(2.0, abs(gf) * 0.15)
                                elif sheet_name.lower() == "action_items" and gt_h.lower() == "suggested_price":
                                    # Tightened: 15% relative or 10.0 absolute floor (was 25% / 20.0)
                                    tol = max(10.0, abs(gf) * 0.15)
                                else:
                                    # Tighter tol: 5% relative or 0.5 absolute floor
                                    tol = max(0.5, abs(gf) * 0.05)
                                check(f"{sheet_name} '{key}' {gt_h} ~{gf}",
                                      abs(gf - af) <= tol, f"got {af}")
                            elif gv is not None and av is not None:
                                gs = str(gv).strip().lower()
                                avs = str(av).strip().lower()
                                if gs:
                                    # For subjective Action_Items text columns (Expected_Impact),
                                    # only require non-empty string. Priority requires {High,Medium,Low}.
                                    # Recommendation in Price_Comparison stays strict (defined values).
                                    if sheet_name.lower() == "action_items" and gt_h.lower() == "expected_impact":
                                        check(f"{sheet_name} '{key}' {gt_h} non-empty",
                                              len(avs) > 0,
                                              f"got empty/missing")
                                    elif sheet_name.lower() == "action_items" and gt_h.lower() == "priority":
                                        check(f"{sheet_name} '{key}' {gt_h} valid",
                                              avs in ("high", "medium", "low"),
                                              f"expected high|medium|low, got '{avs[:50]}'")
                                    else:
                                        # Strict: exact match (case-insensitive trim)
                                        check(f"{sheet_name} '{key}' {gt_h} text",
                                              gs == avs,
                                              f"expected '{gs[:50]}', got '{avs[:50]}'")

    # Check pricing_analyzer.py exists by name (task.md requirement)
    pricing_analyzer = os.path.join(agent_workspace, "pricing_analyzer.py")
    check("pricing_analyzer.py exists", os.path.exists(pricing_analyzer),
          f"agent_workspace files: {os.listdir(agent_workspace) if os.path.isdir(agent_workspace) else 'N/A'}")

    # Database checks
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Strict: exact summary 'Pricing Strategy Review'
        cur.execute("""
            SELECT summary, start_datetime, end_datetime, description
            FROM gcal.events
            WHERE summary = 'Pricing Strategy Review'
        """)
        rows = cur.fetchall()
        check("Calendar event 'Pricing Strategy Review' exists", len(rows) >= 1,
              "no matching event found")

        if rows:
            summary, start_dt, end_dt, desc = rows[0]

            def fmt_dt(dt):
                try:
                    # psycopg2 renders timestamptz in the session TZ
                    # (Asia/Shanghai): a 10:00Z event prints as 18:00+08.
                    # Normalize to UTC before comparing wall-clock.
                    import datetime as _dt
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=_dt.timezone.utc)
                    dt = dt.astimezone(_dt.timezone.utc)
                    return dt.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    return str(dt) if dt else ""

            # Task says March 14, 2026 from 10:00 AM to 11:30 AM UTC
            check("Event start = 2026-03-14 10:00:00",
                  fmt_dt(start_dt) == "2026-03-14 10:00:00",
                  f"got {fmt_dt(start_dt)}")
            check("Event end = 2026-03-14 11:30:00",
                  fmt_dt(end_dt) == "2026-03-14 11:30:00",
                  f"got {fmt_dt(end_dt)}")

            # Description must mention the top-3 largest-gap products.
            # Derive the expected keywords from the GT's own Action_Items
            # (sorted by |Price_Diff| desc) instead of hardcoding old-mock
            # generic names ("tv"/"headphones"/"smartphone") — the mock is
            # now product-level and those words no longer appear in any
            # product name (case-study 2026-08-16 round-3).
            desc_low = (desc or "").lower()
            gt_ai = []
            try:
                gt_wb_ai = openpyxl.load_workbook(
                    os.path.join(groundtruth_workspace, "Competitive_Pricing_Analysis.xlsx"),
                    read_only=True)
                ai_ws = gt_wb_ai["Action_Items"] if "Action_Items" in gt_wb_ai.sheetnames else None
                pc_ws = gt_wb_ai["Price_Comparison"] if "Price_Comparison" in gt_wb_ai.sheetnames else None
                diffs = {}
                if pc_ws:
                    hdr = [str(c.value).strip().lower() if c.value else "" for c in next(pc_ws.iter_rows(max_row=1))]
                    if "product_name" in hdr and "price_diff" in hdr:
                        ni, di = hdr.index("product_name"), hdr.index("price_diff")
                        for r in pc_ws.iter_rows(min_row=2, values_only=True):
                            if r and r[ni]:
                                try:
                                    diffs[str(r[ni]).strip().lower()] = abs(float(r[di]))
                                except (TypeError, ValueError):
                                    pass
                names = []
                if ai_ws:
                    for r in ai_ws.iter_rows(min_row=2, values_only=True):
                        if r and r[0]:
                            names.append(str(r[0]).strip().lower())
                names.sort(key=lambda n: -diffs.get(n, 0))
                gt_ai = names[:3]
                gt_wb_ai.close()
            except Exception:
                gt_ai = []
            for prod_key in (gt_ai or ["tv", "headphones", "smartphone"]):
                check(f"Event description mentions '{prod_key[:40]}'",
                      prod_key in desc_low, f"desc: {desc_low[:200]}")

        # Reverse verification: noise events should not match task keyword
        cur.execute("SELECT COUNT(*) FROM gcal.events WHERE summary ILIKE '%standup%' OR summary ILIKE '%lunch%'")
        noise_events = cur.fetchone()[0]
        check("Noise events exist (not deleted by agent)", noise_events >= 1, f"noise events: {noise_events}")
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()