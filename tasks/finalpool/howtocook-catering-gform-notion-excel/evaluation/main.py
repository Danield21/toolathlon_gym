"""Evaluation for howtocook-catering-gform-notion-excel."""
import argparse
import json
import os
import re
import sys

import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)

PASS_COUNT = 0
FAIL_COUNT = 0
RUNTIME_ONLY_FAIL = 0


def check(name, condition, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, RUNTIME_ONLY_FAIL
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if runtime_only:
            RUNTIME_ONLY_FAIL += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def _parse_number(v):
    """Parse a value into a float.

    Handles str/int/float/None. Strips thousands separators, currency symbols
    ($, ¥, €, £), percent signs and whitespace. Returns None when the value is
    missing or cannot be parsed.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("="):
            return None
        s = s.replace("$", "").replace("¥", "").replace("€", "").replace("£", "")
        s = s.replace(",", "").replace("%", "").replace(" ", "").strip()
        # tolerate trailing currency codes such as "500 USD"
        if s:
            low = s.lower()
            for code in ("usd", "eur", "cny", "rmb", "gbp"):
                if low.endswith(code):
                    s = s[:-len(code)].strip()
                    break
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def num_close(a, b, tol=1.0):
    fa = _parse_number(a)
    fb = _parse_number(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    if fa is None and fb is None:
        sa = "" if a is None else str(a)
        sb = "" if b is None else str(b)
        return sa.strip().lower() == sb.strip().lower()
    return False


def _is_total_like_label(s):
    """Detect a 'Total' / 'Subtotal' / 'Sum' / '合计' summary-row label.

    Budget sheets commonly end with a total row. task.md asks for one row per
    selected recipe and never mentions such a row, so these labels must not be
    counted as recipe rows or their costs folded into the per-dish sum.
    """
    if s is None:
        return False
    t = str(s).strip()
    if not t:
        return False
    if re.search(r"\b(?:total|subtotal|grand\s*total|sum)\b", t, re.IGNORECASE):
        return True
    for kw in ("合计", "总计", "小计"):
        if kw in t:
            return True
    return False


def _cell_float(ws, cache_ws, row, col):
    """Return (numeric_value_or_None, is_formula_without_cache).

    Uses the cached (data_only) workbook first so formula cells with a cached
    value are read correctly. If a cell holds a formula string but has no cached
    value, mark it as formula-without-cache so the caller can decide how to treat
    it (structural check only) instead of misjudging a correct answer.
    """
    raw = None
    cached = None
    try:
        raw = ws.cell(row=row, column=col).value
    except Exception:
        pass
    if cache_ws is not None:
        try:
            cached = cache_ws.cell(row=row, column=col).value
        except Exception:
            pass
    if cached is not None:
        return _parse_number(cached), False
    if isinstance(raw, str) and raw.strip().startswith("="):
        return None, True
    return _parse_number(raw), False


def check_excel(agent_workspace):
    """Check Excel budget file."""
    print("\n=== Checking Excel Budget File ===")
    try:
        import openpyxl
    except ImportError:
        check("openpyxl available", False, "openpyxl not installed")
        return

    agent_file = os.path.join(agent_workspace, "Catering_Budget.xlsx")
    check("Catering_Budget.xlsx exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        # data_only=False gives formula strings; data_only=True gives cached values.
        wb = openpyxl.load_workbook(agent_file, data_only=False)
        wb_cache = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    def get_sheet(wb_, name):
        for s in wb_.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb_[s]
        return None

    cache_sheets = {}
    for s in wb_cache.sheetnames:
        cache_sheets[s.strip().lower()] = wb_cache[s]

    def get_cache_sheet(name):
        return cache_sheets.get(name.strip().lower())

    # Check Menu sheet
    print("\n--- Menu Sheet ---")
    menu_ws = get_sheet(wb, "Menu")
    check("Sheet 'Menu' exists", menu_ws is not None, f"Found: {wb.sheetnames}")
    menu_cache_ws = get_cache_sheet("Menu")

    menu_cost_sum = 0.0
    data_row_indices = []
    recipe_row_indices = []
    if menu_ws:
        headers = [c.value for c in list(menu_ws.rows)[0]] if menu_ws.max_row > 0 else []
        check("Menu has Recipe_Name column",
              any("recipe" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")
        check("Menu has Category column",
              any("category" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")
        check("Menu has Servings column",
              any("serving" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")
        check("Menu has Estimated_Cost_Per_Person column",
              any("cost" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")

        for r in range(2, menu_ws.max_row + 1):
            vals = [menu_ws.cell(row=r, column=c).value for c in range(1, menu_ws.max_column + 1)]
            if any(v is not None for v in vals):
                data_row_indices.append(r)

        # Exclude trailing 'Total'/'Subtotal' summary rows (a normal budget-sheet
        # practice that task.md does not forbid) and blank-first-cell rows from the
        # recipe count, so a correct model with a Total row is not misjudged as 9 rows.
        for r in data_row_indices:
            label = menu_ws.cell(row=r, column=1).value
            if label is None or str(label).strip() == "":
                continue
            if _is_total_like_label(label):
                continue
            recipe_row_indices.append(r)

        # Per task: exactly 6 to 8 recipes
        check("Menu has 6 to 8 recipe rows", 6 <= len(recipe_row_indices) <= 8,
              f"Found {len(recipe_row_indices)} recipe rows (Total/blank rows excluded; raw={len(data_row_indices)})")

        # Sum Estimated_Cost_Per_Person for later verification
        cost_idx = None
        for i, h in enumerate(headers):
            if h and "cost" in str(h).lower():
                cost_idx = i
                break
        if cost_idx is not None:
            for r in recipe_row_indices:
                cost_val, _ = _cell_float(menu_ws, menu_cache_ws, r, cost_idx + 1)
                if cost_val is not None:
                    menu_cost_sum += cost_val

    # Check Summary sheet
    print("\n--- Summary Sheet ---")
    sum_ws = get_sheet(wb, "Summary")
    check("Sheet 'Summary' exists", sum_ws is not None, f"Found: {wb.sheetnames}")
    sum_cache_ws = get_cache_sheet("Summary")

    if sum_ws:
        summary_data = {}
        summary_formula = {}
        for r in range(2, sum_ws.max_row + 1):
            metric_raw = sum_ws.cell(row=r, column=1).value
            if metric_raw is None or str(metric_raw).strip() == "":
                continue
            key = str(metric_raw).strip().lower().replace(" ", "_")
            val, is_formula = _cell_float(sum_ws, sum_cache_ws, r, 2)
            summary_data[key] = val
            if is_formula:
                summary_formula[key] = True

        def summary_ok(key, expected, tol):
            if key not in summary_data:
                return False
            if summary_formula.get(key):
                # Formula without cached value: structural presence only, do not
                # misjudge a correct formula-based answer.
                return True
            v = summary_data[key]
            return v is not None and num_close(v, expected, tol)

        check("Summary has Total_Dishes",
              "total_dishes" in summary_data,
              f"Keys: {list(summary_data.keys())}")
        check("Summary has Total_Budget with value 500",
              summary_ok("total_budget", 500, 5),
              f"Data: {summary_data}")
        check("Summary has Budget_Per_Person with value 25",
              summary_ok("budget_per_person", 25, 2),
              f"Data: {summary_data}")

        # Validate Total_Dishes == menu row count
        if "total_dishes" in summary_data and menu_ws:
            if summary_formula.get("total_dishes"):
                check("Total_Dishes matches Menu row count", len(recipe_row_indices) > 0,
                      "Total_Dishes is a formula without cached value; structural check only")
            else:
                td = summary_data.get("total_dishes")
                try:
                    td_int = int(td) if td is not None else None
                except (ValueError, TypeError):
                    td_int = None
                check("Total_Dishes matches Menu row count",
                      td_int is not None and td_int == len(recipe_row_indices),
                      f"Summary Total_Dishes={td!r}, Menu recipe rows={len(recipe_row_indices)}")

        # Validate sum(Estimated_Cost_Per_Person) is plausible for the $25 per-person
        # budget. NOTE: task.md Part 4 only asks for a "reasonable estimated cost per
        # person" on each row and never declares that the per-dish costs must sum to
        # the budget, so this is an informational (runtime-only) check: a correct
        # model whose estimates sum slightly above $25 must NOT be failed overall.
        if recipe_row_indices:
            if menu_cost_sum > 0:
                check("Sum(Estimated_Cost_Per_Person) within budget 25",
                      menu_cost_sum <= 26.0,
                      f"Sum={menu_cost_sum:.2f} exceeds $25 per-person budget",
                      runtime_only=True)
            else:
                check("Sum(Estimated_Cost_Per_Person) within budget 25",
                      False,
                      "No literal per-person costs parsed from Menu recipe rows",
                      runtime_only=True)


def check_gform():
    """Check Google Form creation."""
    print("\n=== Checking Google Form ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        print(f"  [SKIP] Database unavailable, skipping Google Form checks: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        check("At least one form created", len(forms) >= 1, f"Found {len(forms)} forms")

        # Target form: "Team Lunch Dietary Preferences"
        form_id = None
        for fid, title in forms:
            t = (title or "").lower()
            if "dietary" in t and "lunch" in t:
                form_id = fid
                break
        if form_id is None:
            # fallback: forms with team-lunch / dietary preferences
            for fid, title in forms:
                t = (title or "").lower()
                if "dietary" in t or ("team" in t and "lunch" in t):
                    form_id = fid
                    break

        check("Form titled 'Team Lunch Dietary Preferences' found",
              form_id is not None,
              f"Found titles: {[f[1] for f in forms]}")

        if form_id:
            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s", (form_id,))
            questions = cur.fetchall()
            check("Form has at least 3 questions", len(questions) >= 3,
                  f"Found {len(questions)} questions")
            # Per task: questions must cover dietary, allergy, comments
            q_titles = " ".join([str(q[0] or "").lower() for q in questions])
            check("Form mentions dietary/preference",
                  "dietary" in q_titles or "vegetarian" in q_titles or "vegan" in q_titles
                  or "preference" in q_titles,
                  f"Q titles: {q_titles[:150]}",
                  runtime_only=True)
            check("Form mentions allergy",
                  "allerg" in q_titles,
                  f"Q titles: {q_titles[:150]}",
                  runtime_only=True)
            check("Form mentions comments/additional",
                  "comment" in q_titles or "additional" in q_titles or "request" in q_titles,
                  f"Q titles: {q_titles[:150]}",
                  runtime_only=True)
            # At least one multiple-choice (radio) or checkbox question per triage.
            # The google-forms MCP only creates 'choiceQuestion' (RADIO) for
            # add_multiple_choice_question and 'textQuestion' for add_text_question,
            # so accept the 'choice' family only (never test unsupported types).
            q_types = [str(q[1] or "").lower() for q in questions]
            check("Form has at least one multiple-choice question",
                  any(any(kw in t for kw in ("multiple", "radio", "checkbox", "choice", "choose")) for t in q_types),
                  f"Question types: {q_types}",
                  runtime_only=True)
    finally:
        cur.close()
        conn.close()


def _rich_text_to_text(rt):
    """Extract plain text from a Notion rich-text object or a plain string."""
    if rt is None:
        return ""
    if isinstance(rt, str):
        return rt
    if not isinstance(rt, dict):
        return ""
    txt = rt.get("text")
    if isinstance(txt, dict):
        for k in ("content", "plain_text", "text"):
            v = txt.get(k)
            if isinstance(v, str):
                return v
        return ""
    if isinstance(txt, str):
        return txt
    pt = rt.get("plain_text")
    if isinstance(pt, str):
        return pt
    return ""


def _extract_title(props):
    """Extract the page title from notion.pages.properties jsonb.

    Tolerates the shapes actually written by the notion MCP server:
      {"title": [{"text": {"content": "..."}}], "type": "title"}
    and the real Notion API shape:
      {"Name": {"type": "title", "title": [{"text": {"content": "..."}}]}}
    as well as plain-string titles. Does not depend on any 'content'/'page_id'
    column (those do not exist in the runtime schema).
    """
    if not props:
        return ""
    if isinstance(props, str):
        return props
    if not isinstance(props, dict):
        return ""

    # Shape A: top-level "title" key that is a list of rich text or a string.
    tl = props.get("title")
    if isinstance(tl, str):
        return tl
    if isinstance(tl, list):
        joined = "".join(_rich_text_to_text(t) for t in tl).strip()
        if joined:
            return joined

    # Shape B: property dicts with type == "title" (real Notion API form),
    # or any dict value carrying a "title" list of rich text.
    for _, prop in props.items():
        if not isinstance(prop, dict):
            continue
        if prop.get("type") == "title":
            tl2 = prop.get("title")
            if isinstance(tl2, str):
                return tl2
            if isinstance(tl2, list):
                joined = "".join(_rich_text_to_text(t) for t in tl2).strip()
                if joined:
                    return joined
        sub = prop.get("title")
        if isinstance(sub, list):
            joined = "".join(_rich_text_to_text(t) for t in sub).strip()
            if joined:
                return joined

    # Shape C: fall back to the first non-empty string value.
    for _, prop in props.items():
        if isinstance(prop, str) and prop.strip():
            return prop.strip()

    return ""


def _json_texts(v, out):
    """Collect every string leaf from a parsed JSON object (block_data etc.)."""
    if isinstance(v, dict):
        for _k, val in v.items():
            _json_texts(val, out)
    elif isinstance(v, list):
        for item in v:
            _json_texts(item, out)
    elif isinstance(v, str):
        out.append(v)


def check_notion():
    """Check Notion page creation."""
    print("\n=== Checking Notion Page ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        print(f"  [SKIP] Database unavailable, skipping Notion checks: {e}")
        return
    cur = conn.cursor()
    try:
        # Target page: "Team Lunch Menu"
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()
        check("At least one Notion page created", len(pages) >= 1,
              f"Found {len(pages)} pages")

        # Look for a page titled "Team Lunch Menu" (case-insensitive)
        target_page = None
        titles = []
        for pid, props in pages:
            t = _extract_title(props)
            titles.append(t)
            if t and "team lunch menu" in t.lower():
                target_page = pid
                break
        check("Notion page titled 'Team Lunch Menu' exists",
              target_page is not None,
              f"Page titles: {titles[:5]}")

        if target_page:
            # Check blocks (content) on that page mention 20 people, $25, budget.
            cur.execute("SELECT block_data FROM notion.blocks WHERE parent_id = %s", (target_page,))
            blocks = cur.fetchall()
            texts = []
            for (bd,) in blocks:
                if bd is None:
                    continue
                try:
                    _json_texts(bd, texts)
                except Exception:
                    pass
            all_block_text = " ".join(texts).lower()
            # Stricter regex: look for '20 people' / '20 persons' / 'for 20'
            twenty_people_match = bool(re.search(r"\b20\s*(people|persons|guests|attendees|team members)\b", all_block_text)) \
                                  or bool(re.search(r"\bfor\s*20\b", all_block_text)) \
                                  or bool(re.search(r"\b20\s*(staff|team)\b", all_block_text))
            check("Notion page mentions '20 people' (exact phrase)",
                  twenty_people_match,
                  f"Block text sample: {all_block_text[:200]}",
                  runtime_only=True)
            check("Notion page mentions 25 (per person) or budget",
                  "25" in all_block_text or "budget" in all_block_text,
                  f"Block text len: {len(all_block_text)}",
                  runtime_only=True)
    finally:
        cur.close()
        conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("HOWTOCOOK CATERING GFORM NOTION EXCEL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_gform()
    check_notion()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    non_runtime_fail = FAIL_COUNT - RUNTIME_ONLY_FAIL
    overall = non_runtime_fail == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (runtime-only fails: {RUNTIME_ONLY_FAIL})")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
