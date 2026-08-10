"""Evaluation for sf-hr-experience-distribution."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def is_numeric(val):
    """Check if a value can be parsed as float."""
    if val is None:
        return False
    try:
        # strip commas / currency
        s = str(val).replace(",", "").replace("$", "").strip()
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def check_notion_page(errors_list):
    """Verify Notion page titled 'HR Experience Distribution' exists with content."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE archived = false
              AND (properties::text ILIKE '%hr experience distribution%'
                   OR properties::text ILIKE '%experience distribution%')
        """)
        rows = cur.fetchall()
        if not rows:
            errors_list.append("Notion page 'HR Experience Distribution' not found")
            cur.close(); conn.close()
            return
        page_id = rows[0][0]
        cur.execute("SELECT COUNT(*) FROM notion.blocks WHERE parent_id = %s", (page_id,))
        block_count = cur.fetchone()[0]
        if block_count < 1:
            errors_list.append(f"Notion page has no body blocks (got {block_count})")
        cur.close(); conn.close()
    except Exception as e:
        errors_list.append(f"Notion check error: {e}")


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_header(h):
    return " ".join(str(h or "").strip().lower().replace("-", " ").replace("_", " ").split())


def _header_map(rows):
    """Map normalized header text -> 0-based column index from row 1 of a sheet."""
    if not rows or not rows[0]:
        return {}
    return {_norm_header(h): i for i, h in enumerate(rows[0]) if _norm_header(h)}


def _cell(row, idx):
    """Safe cell access (None when idx out of range)."""
    if row is None or idx is None:
        return None
    try:
        return row[idx] if idx < len(row) else None
    except (TypeError, IndexError):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Experience_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Experience_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Experience Analysis
    print(f"  Checking Experience Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Experience Analysis")
    g_rows = load_sheet_rows(gt_wb, "Experience Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Experience Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Experience Analysis' not found in groundtruth")
    else:
        sheet_name = "Experience Analysis"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Employee_Count (exact)
            a_v = _cell(a_row, a_col.get("employee count"))
            g_v = _cell(g_row, g_col.get("employee count"))
            if a_v is not None and g_v is not None:
                if not num_close(a_v, g_v, 0):
                    errors.append(f"{key}.Employee_Count: {a_v} vs {g_v} (exact)")

            # Avg_Salary (±$1 due to rounding)
            a_v = _cell(a_row, a_col.get("avg salary"))
            g_v = _cell(g_row, g_col.get("avg salary"))
            if a_v is not None and g_v is not None:
                if not num_close(a_v, g_v, 1.0):
                    errors.append(f"{key}.Avg_Salary: {a_v} vs {g_v} (tol=1.0)")

            # Avg_Rating
            a_v = _cell(a_row, a_col.get("avg rating"))
            g_v = _cell(g_row, g_col.get("avg rating"))
            if a_v is not None and g_v is not None:
                if not num_close(a_v, g_v, 0.1):
                    errors.append(f"{key}.Avg_Rating: {a_v} vs {g_v} (tol=0.1)")

            # Avg_Satisfaction
            a_v = _cell(a_row, a_col.get("avg satisfaction"))
            g_v = _cell(g_row, g_col.get("avg satisfaction"))
            if a_v is not None and g_v is not None:
                if not num_close(a_v, g_v, 0.1):
                    errors.append(f"{key}.Avg_Satisfaction: {a_v} vs {g_v} (tol=0.1)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Value: numeric compare when GT value is numeric, else string match
            a_v = _cell(a_row, a_col.get("value"))
            g_v = _cell(g_row, g_col.get("value"))
            if a_v is not None and g_v is not None:
                if is_numeric(g_v):
                    if not num_close(a_v, g_v, 0):  # exact integer count
                        errors.append(f"{key}.Value: '{a_v}' vs '{g_v}' (numeric)")
                else:
                    if not str_match(a_v, g_v):
                        errors.append(f"{key}.Value: '{a_v}' vs '{g_v}' (string)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    # Notion check
    print(f"  Checking Notion page...")
    notion_errors = []
    check_notion_page(notion_errors)
    all_errors.extend(notion_errors)
    if notion_errors:
        for e in notion_errors:
            print(f"    {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
