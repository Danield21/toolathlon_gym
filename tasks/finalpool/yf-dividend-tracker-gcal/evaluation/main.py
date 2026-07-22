"""
Evaluation script for yf-dividend-tracker-gcal task.

Checks:
1. Excel file (Dividend_Tracker.xlsx) - Dividend Stocks and Summary sheets
2. Google Calendar events for ex-dividend dates (>=3 events with dividend in summary)
3. Google Sheet with "dividend" in title
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail="", db=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        if db:
            PASS_COUNT += 1
        else:
            PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        if db:
            FAIL_COUNT += 1
        else:
            FAIL_COUNT += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_expected_dividends():
    """Get expected dividend data from yf.stock_info."""
    import datetime
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT symbol,
               data->>'shortName' as name,
               (data->>'dividendRate')::float as div_rate,
               (data->>'dividendYield')::float as div_yield,
               (data->>'exDividendDate')::float as ex_div_ts,
               (data->>'payoutRatio')::float as payout_ratio
        FROM yf.stock_info
        WHERE symbol IN ('GOOGL','AMZN','JPM','JNJ','XOM')
          AND (data->>'dividendRate') IS NOT NULL
          AND (data->>'dividendRate')::float > 0
        ORDER BY symbol
    """)
    rows = []
    for r in cur.fetchall():
        symbol, name, div_rate, div_yield, ex_div_ts, payout_ratio = r
        ex_div_date = ""
        if ex_div_ts:
            try:
                ex_div_date = datetime.datetime.utcfromtimestamp(float(ex_div_ts)).strftime("%Y-%m-%d")
            except Exception:
                ex_div_date = ""
        rows.append((symbol, name, div_rate, div_yield, ex_div_date, payout_ratio))
    cur.close()
    conn.close()
    return rows


def check_excel(agent_workspace, expected_rows):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Dividend_Tracker.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return False

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # Check Dividend Stocks sheet
    ws = get_sheet(wb, "Dividend Stocks")
    check("Sheet 'Dividend Stocks' exists", ws is not None, f"Sheets: {wb.sheetnames}")
    if ws is None:
        return False

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    num_expected = len(expected_rows)
    check(f"Dividend Stocks has {num_expected} rows",
          len(rows) == num_expected,
          f"Got {len(rows)} rows, expected {num_expected}")

    # Per-ticker value validation
    ticker_lookup = {}
    for r in rows:
        if r and r[0]:
            ticker_lookup[str(r[0]).strip().upper()] = r

    for exp_r in expected_rows:
        symbol, name, div_rate, div_yield, ex_div_date, payout_ratio = exp_r
        a_row = ticker_lookup.get(symbol)
        if a_row is None:
            check(f"Ticker {symbol} present", False, f"Found: {list(ticker_lookup.keys())}")
            continue
        check(f"Ticker {symbol} present", True)
        # Company name - allow inclusion (e.g. 'Alphabet Inc.' substring of 'Alphabet, Inc. (GOOGL)')
        if a_row[1] is not None and name:
            ok = name.lower().split()[0] in str(a_row[1]).lower() if name.split() else True
            check(f"{symbol} Company name", ok, f"Expected '{name}', got '{a_row[1]}'")
        # Dividend_Rate - tol 0.05
        if a_row[2] is not None and div_rate is not None:
            try:
                ok = abs(float(a_row[2]) - float(div_rate)) <= 0.05
                check(f"{symbol} Dividend_Rate", ok, f"Expected {div_rate}, got {a_row[2]}")
            except (ValueError, TypeError):
                check(f"{symbol} Dividend_Rate numeric", False, f"got {a_row[2]}")
        # Dividend_Yield - task says "as a percentage". Accept any of:
        #   - raw decimal (e.g. 0.28)
        #   - percent value (e.g. 28)
        #   - percent of percent (e.g. 28%)
        if a_row[3] is not None and div_yield is not None:
            try:
                af = float(a_row[3])
                gf = float(div_yield)
                # gf is the raw decimal (e.g. 0.28). Agent might report as 0.28 OR 28
                ok = (abs(af - gf) <= max(0.01, abs(gf) * 0.1)
                      or abs(af - gf*100) <= max(0.5, abs(gf*100) * 0.05))
                check(f"{symbol} Dividend_Yield", ok, f"Expected ~{div_yield} or {div_yield*100}, got {a_row[3]}")
            except (ValueError, TypeError):
                check(f"{symbol} Dividend_Yield numeric", False, f"got {a_row[3]}")
        # Ex_Dividend_Date - exact match (date string or convertible)
        if a_row[4] is not None and ex_div_date:
            a_date = str(a_row[4])[:10]
            ok = a_date == ex_div_date
            check(f"{symbol} Ex_Dividend_Date", ok, f"Expected {ex_div_date}, got {a_date}")
        # Payout_Ratio - tol 0.02
        if a_row[5] is not None and payout_ratio is not None:
            try:
                ok = abs(float(a_row[5]) - float(payout_ratio)) <= 0.02
                check(f"{symbol} Payout_Ratio", ok, f"Expected {payout_ratio}, got {a_row[5]}")
            except (ValueError, TypeError):
                check(f"{symbol} Payout_Ratio numeric", False, f"got {a_row[5]}")

    # Check Summary sheet - validate values
    ws2 = get_sheet(wb, "Summary")
    check("Sheet 'Summary' exists", ws2 is not None, f"Sheets: {wb.sheetnames}")
    if ws2:
        summary_data = {}
        for row in ws2.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                key = str(row[0]).strip().lower().replace(" ", "_")
                summary_data[key] = row[1]

        # Total_Dividend_Stocks - exact integer
        total_k = next((k for k in summary_data if "total" in k and "dividend" in k), None)
        check("Summary has Total_Dividend_Stocks", total_k is not None, f"Keys: {list(summary_data.keys())}")
        if total_k:
            try:
                ok = int(summary_data[total_k]) == num_expected
                check(f"Total_Dividend_Stocks={summary_data[total_k]}", ok, f"Expected {num_expected}")
            except (ValueError, TypeError):
                check("Total_Dividend_Stocks integer", False, f"got {summary_data[total_k]}")
        # Highest_Yield_Ticker - find expected
        if expected_rows:
            highest_yield_row = max(expected_rows, key=lambda r: r[3] if r[3] is not None else 0)
            highest_yield_ticker = highest_yield_row[0]
            high_k = next((k for k in summary_data if "highest" in k and "yield" in k), None)
            check("Summary has Highest_Yield_Ticker", high_k is not None, f"Keys: {list(summary_data.keys())}")
            if high_k:
                ok = str(summary_data[high_k] or "").strip().upper() == highest_yield_ticker
                check(f"Highest_Yield_Ticker={summary_data[high_k]}", ok, f"Expected {highest_yield_ticker}")
        # Avg_Yield - existence + value check (avg yields across dividend stocks)
        avg_k = next((k for k in summary_data if "avg" in k and "yield" in k), None)
        check("Summary has Avg_Yield", avg_k is not None, f"Keys: {list(summary_data.keys())}")
        if avg_k and expected_rows:
            try:
                # Compute expected avg yield (raw decimal), and also as percent
                yields = [r[3] for r in expected_rows if r[3] is not None]
                if yields:
                    expected_avg_dec = sum(yields) / len(yields)
                    expected_avg_pct = expected_avg_dec * 100
                    af = float(summary_data[avg_k])
                    # Accept either decimal or percent representation
                    ok = (
                        abs(af - expected_avg_dec) <= max(0.005, abs(expected_avg_dec) * 0.10)
                        or abs(af - expected_avg_pct) <= max(0.5, abs(expected_avg_pct) * 0.05)
                    )
                    check(
                        f"Avg_Yield value (~{expected_avg_dec:.4f} or ~{expected_avg_pct:.2f}%)",
                        ok,
                        f"got {summary_data[avg_k]}"
                    )
            except (ValueError, TypeError) as _e:
                check("Avg_Yield numeric", False, f"got {summary_data[avg_k]}: {_e}")

    return True


def check_calendar(expected_rows):
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime FROM gcal.events")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_calendar] Found {len(events)} events.")

    # For each expected dividend ticker, require an event with summary 'Ex-Dividend: TICKER'
    all_ok = True
    for exp_r in expected_rows:
        symbol = exp_r[0]
        ex_date = exp_r[4]
        # Find matching event
        matching = [e for e in events
                    if e[0] and "ex-dividend" in e[0].lower() and symbol.upper() in e[0].upper()]
        ok = len(matching) >= 1
        check(f"Calendar event 'Ex-Dividend: {symbol}'", ok, f"summaries: {[e[0] for e in events][:5]}")
        if not ok:
            all_ok = False
            continue
        # Optional: verify date matches if start_datetime present
        if matching[0][2] is not None and ex_date:
            try:
                event_date = str(matching[0][2])[:10]
                date_ok = event_date == ex_date
                check(f"Event {symbol} date matches ex-dividend {ex_date}", date_ok, f"got {event_date}")
                if not date_ok:
                    all_ok = False
            except Exception:
                pass

    return all_ok


def check_gsheet(expected_rows):
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    sheets = cur.fetchall()

    print(f"[check_gsheet] Found {len(sheets)} spreadsheets.")

    # Require exact title 'Dividend Watch List'
    target_sheets = [s for s in sheets if s[1] and "dividend watch list" in s[1].lower()]
    title_ok = len(target_sheets) > 0
    check("Google Sheet titled 'Dividend Watch List' exists", title_ok,
          f"Sheet titles: {[s[1] for s in sheets]}")
    if not title_ok:
        cur.close()
        conn.close()
        return False

    # Check 'Overview' tab
    spreadsheet_id = target_sheets[0][0]
    cur.execute("""
        SELECT id, title FROM gsheet.sheets
        WHERE spreadsheet_id = %s AND LOWER(title) = 'overview'
    """, (spreadsheet_id,))
    tabs = cur.fetchall()
    tab_ok = len(tabs) > 0
    check("GSheet has 'Overview' tab", tab_ok)

    # Validate Overview tab content: tickers must be present in cell values
    if tab_ok and expected_rows:
        sheet_id = tabs[0][0]
        try:
            cur.execute("""
                SELECT row_index, col_index, value
                FROM gsheet.cells
                WHERE sheet_id = %s
                ORDER BY row_index, col_index
            """, (sheet_id,))
            cells = cur.fetchall()
            all_text = " ".join(str(c[2] or "") for c in cells).upper()
            for exp_r in expected_rows:
                symbol = exp_r[0]
                check(f"GSheet Overview contains ticker {symbol}", symbol in all_text,
                      f"sheet text excerpt: {all_text[:200]}")
        except Exception as _e:
            check("GSheet Overview cell content readable", False, str(_e))

    cur.close()
    conn.close()
    return title_ok and tab_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_rows = get_expected_dividends()
    print(f"Expected {len(expected_rows)} dividend-paying stocks from DB")

    check_excel(args.agent_workspace, expected_rows)
    check_calendar(expected_rows)
    check_gsheet(expected_rows)

    total_pass = PASS_COUNT
    total_fail = FAIL_COUNT
    all_ok = FAIL_COUNT == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if args.res_log_file:
        result = {"passed": total_pass, "failed": total_fail, "success": all_ok}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
