"""
Evaluation script for canvas-grade-distribution-gform task.

Checks:
1. Excel with Spring 2014 course grade data (verified against canvas DB)
2. Google Form with >=4 questions
3. Email with correct subject
"""

import argparse
import json
import os
import re
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}


def _to_float(x):
    """Robustly convert a cell value to float. Returns None if unparseable."""
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    # Strip thousands separators, currency symbols, percent sign, whitespace.
    s = s.replace(",", "").replace("$", "").replace("€", "").replace("￥", "")
    s = s.replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _cell_status(raw_cell, val_cell):
    """Classify an xlsx cell for value comparison.

    Returns (status, float|None):
      - ("empty", None): cell is blank
      - ("formula", None): cell holds a formula with no cached value (structure-only)
      - ("value", float): cell holds a parseable numeric literal or a formula
        whose cached value parses (float may still be None if cached value is
        unparseable text).
    """
    raw = raw_cell.value
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return "empty", None
    if isinstance(raw, str) and raw.lstrip().startswith("="):
        if val_cell is None or val_cell.value is None:
            return "formula", None
        return "value", _to_float(val_cell.value)
    return "value", _to_float(raw)


def _course_name_matches(v, gt_name, gt_code=None):
    """Case-insensitive, tolerant match of an agent-supplied course identifier.

    Accepts the full course name, the name without its parenthetical term suffix
    (e.g. "(Spring 2014)"), or the course code.
    """
    if v is None:
        return False
    v = str(v).strip()
    if not v:
        return False
    v_l = v.lower()
    g_l = str(gt_name).strip().lower()
    if g_l == v_l or g_l in v_l or v_l in g_l:
        return True
    # Compare without parenthetical suffixes, e.g. "(spring 2014)".
    core = lambda s: re.sub(r"\s*\(.*\)\s*$", "", s).strip().lower()
    gc, vc = core(gt_name), core(v)
    if gc and vc and (gc == vc or gc in vc or vc in gc):
        return True
    if gt_code:
        code_l = str(gt_code).strip().upper()
        if code_l == v.upper() or code_l in v.upper():
            return True
    return False


def get_expected_course_data():
    """Query actual grade stats from canvas DB."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT c.course_code, c.name,
                   AVG(s.score) as avg_grade,
                   PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY s.score) as median_grade,
                   COUNT(DISTINCT s.user_id) as total_students,
                   AVG(CASE WHEN s.score >= 60 THEN 1.0 ELSE 0.0 END) * 100.0 as pass_rate
            FROM canvas.courses c
            JOIN canvas.assignments a ON a.course_id = c.id
            JOIN canvas.submissions s ON s.assignment_id = a.id
            WHERE c.name LIKE '%%Spring 2014%%'
            AND s.score IS NOT NULL
            GROUP BY c.course_code, c.name
            ORDER BY c.name
        """)
        course_stats = cur.fetchall()
        cur.close()
        conn.close()
        return course_stats
    except psycopg2.OperationalError:
        return []


def check_excel(workspace):
    """Check Excel file."""
    from openpyxl import load_workbook

    errors = []
    xlsx_path = os.path.join(workspace, "Grade_Distribution_Report.xlsx")
    if not os.path.exists(xlsx_path):
        return ["Grade_Distribution_Report.xlsx not found"]

    course_stats = get_expected_course_data()

    try:
        # Read formulas in raw form plus cached values so that either a literal
        # number or a formula with a cached result is accepted.
        wb_raw = load_workbook(xlsx_path, data_only=False)
        wb_val = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return [f"Could not open Grade_Distribution_Report.xlsx: {e}"]

    raw_sheet_names = [s.lower() for s in wb_raw.sheetnames]
    val_sheet_names = [s.lower() for s in wb_val.sheetnames]

    def _sheet_pair(name):
        if name not in raw_sheet_names:
            return None, None
        raw = wb_raw[wb_raw.sheetnames[raw_sheet_names.index(name)]]
        val = wb_val[wb_val.sheetnames[val_sheet_names.index(name)]]
        return raw, val

    # Check Course Grades sheet
    raw_ws, val_ws = _sheet_pair("course grades")
    if raw_ws is None:
        errors.append(f"Missing 'Course Grades' sheet. Found: {wb_raw.sheetnames}")
    else:
        headers = [str(c.value).lower().replace(" ", "_") if c.value else "" for c in raw_ws[1]]

        # Pass_Rate and Median_Grade are required by task
        for rh in ["course_code", "course_name", "avg_grade", "total_students",
                   "pass_rate", "median_grade"]:
            if not any(rh in h or rh.replace("_", "") in h.replace("_", "") for h in headers):
                errors.append(f"Course Grades missing header: {rh}")

        data_rows = sum(1 for row in raw_ws.iter_rows(min_row=2) if row[0].value is not None)
        if data_rows < len(course_stats):
            errors.append(f"Course Grades has {data_rows} rows, expected at least {len(course_stats)}")

        # Locate columns by name
        def _col_index(*aliases):
            for idx, h in enumerate(headers):
                hk = h.replace("_", "")
                for a in aliases:
                    if a in h or a.replace("_", "") in hk:
                        return idx
            return None

        code_col = _col_index("course_code", "code")
        avg_col = _col_index("avg_grade")
        med_col = _col_index("median_grade", "median")
        pass_col = _col_index("pass_rate")
        ts_col = _col_index("total_students", "students")

        # Verify course codes are present
        if code_col is not None:
            found_codes = set()
            row_data = {}
            raw_rows = list(raw_ws.iter_rows(min_row=2))
            val_rows = list(val_ws.iter_rows(min_row=2))
            for raw_row, val_row in zip(raw_rows, val_rows):
                code_cell = raw_row[code_col] if code_col < len(raw_row) else None
                if code_cell is not None and code_cell.value:
                    code = str(code_cell.value).strip().upper()
                    found_codes.add(code)
                    row_data[code] = (raw_row, val_row)
            expected_codes = set(cs[0].upper() for cs in course_stats)
            missing_codes = expected_codes - found_codes
            if missing_codes:
                errors.append(f"Missing course codes: {missing_codes}")

            # Per-row value validation when DB groundtruth is available
            if course_stats:
                for cs in course_stats:
                    code, _name, gt_avg, gt_med, gt_ts, gt_pr = cs
                    code_u = code.upper()
                    if code_u not in row_data:
                        continue
                    raw_row, val_row = row_data[code_u]

                    if avg_col is not None and gt_avg is not None:
                        raw_c = raw_row[avg_col] if avg_col < len(raw_row) else None
                        val_c = val_row[avg_col] if avg_col < len(val_row) else None
                        status, v = _cell_status(raw_c, val_c)
                        if status == "empty":
                            errors.append(f"{code_u}: avg_grade missing")
                        elif status == "value":
                            if v is None:
                                errors.append(f"{code_u}: avg_grade not numeric ('{raw_c.value}')")
                            elif abs(v - float(gt_avg)) > 1.0:
                                errors.append(
                                    f"{code_u}: avg_grade {v} != GT {float(gt_avg):.2f} (tol 1.0)")

                    if med_col is not None and gt_med is not None:
                        raw_c = raw_row[med_col] if med_col < len(raw_row) else None
                        val_c = val_row[med_col] if med_col < len(val_row) else None
                        status, v = _cell_status(raw_c, val_c)
                        if status == "empty":
                            errors.append(f"{code_u}: median_grade missing")
                        elif status == "value":
                            if v is None:
                                errors.append(f"{code_u}: median_grade not numeric ('{raw_c.value}')")
                            elif abs(v - float(gt_med)) > 2.0:
                                errors.append(
                                    f"{code_u}: median_grade {v} != GT {float(gt_med):.2f} (tol 2.0)")

                    if ts_col is not None and gt_ts is not None:
                        raw_c = raw_row[ts_col] if ts_col < len(raw_row) else None
                        val_c = val_row[ts_col] if ts_col < len(val_row) else None
                        status, v = _cell_status(raw_c, val_c)
                        if status == "empty":
                            errors.append(f"{code_u}: total_students missing")
                        elif status == "value":
                            if v is None:
                                errors.append(f"{code_u}: total_students not numeric ('{raw_c.value}')")
                            elif abs(v - float(gt_ts)) > 1.0:
                                # tol 1: seed has courses where one student has only a
                                # NULL-score submission, so "students who submitted work"
                                # differs by 1 between the all-submissions and
                                # graded-submissions interpretations. Accept either.
                                errors.append(
                                    f"{code_u}: total_students {v} != GT {float(gt_ts)} (tol 1)")

                    if pass_col is not None and gt_pr is not None:
                        raw_c = raw_row[pass_col] if pass_col < len(raw_row) else None
                        val_c = val_row[pass_col] if pass_col < len(val_row) else None
                        status, v = _cell_status(raw_c, val_c)
                        if status == "empty":
                            errors.append(f"{code_u}: pass_rate missing")
                        elif status == "value":
                            if v is None:
                                errors.append(f"{code_u}: pass_rate not numeric ('{raw_c.value}')")
                            else:
                                # Accept either fraction (0-1) or percent (0-100)
                                if min(abs(v - float(gt_pr)),
                                       abs(v * 100.0 - float(gt_pr))) > 5.0:
                                    errors.append(
                                        f"{code_u}: pass_rate {v} != GT {float(gt_pr):.1f}% (tol 5)")

    # Check Summary sheet
    raw_ws, val_ws = _sheet_pair("summary")
    if raw_ws is None:
        errors.append(f"Missing 'Summary' sheet. Found: {wb_raw.sheetnames}")
    else:
        summary_data = {}
        raw_rows = list(raw_ws.iter_rows(min_row=2))
        val_rows = list(val_ws.iter_rows(min_row=2))
        for raw_row, val_row in zip(raw_rows, val_rows):
            if raw_row[0].value:
                key = str(raw_row[0].value).lower().replace(" ", "_")
                summary_data[key] = (raw_row[1], val_row[1])

        # Check total courses
        total_key = None
        for k in summary_data:
            if "total" in k and "course" in k:
                total_key = k
                break
        if total_key:
            raw_c, val_c = summary_data[total_key]
            status, v = _cell_status(raw_c, val_c)
            if status == "empty":
                errors.append("Total_Courses: empty value")
            elif status == "value":
                if v is None:
                    errors.append(f"Cannot parse Total_Courses: {raw_c.value}")
                elif int(v) != len(course_stats):
                    errors.append(f"Total_Courses: got {int(v)}, expected {len(course_stats)}")
        else:
            errors.append("Summary missing Total_Courses row")

        # Validate Overall_Avg_Grade
        if course_stats:
            gt_overall = sum(float(cs[2]) for cs in course_stats) / len(course_stats)
            overall_key = next(
                (k for k in summary_data if "overall" in k and "avg" in k), None)
            if overall_key:
                raw_c, val_c = summary_data[overall_key]
                status, v = _cell_status(raw_c, val_c)
                if status == "empty":
                    errors.append("Overall_Avg_Grade: empty value")
                elif status == "value":
                    if v is None:
                        errors.append(f"Cannot parse Overall_Avg_Grade: {raw_c.value}")
                    elif abs(v - gt_overall) > 1.0:
                        errors.append(
                            f"Overall_Avg_Grade: got {v}, expected ~{gt_overall:.2f}")
            else:
                errors.append("Summary missing Overall_Avg_Grade row")

            # Validate Highest_Avg_Course / Lowest_Avg_Course
            sorted_by_avg = sorted(course_stats, key=lambda x: float(x[2]))
            gt_lowest_name = sorted_by_avg[0][1]
            gt_lowest_code = sorted_by_avg[0][0]
            gt_highest_name = sorted_by_avg[-1][1]
            gt_highest_code = sorted_by_avg[-1][0]
            highest_key = next(
                (k for k in summary_data if "highest" in k), None)
            lowest_key = next(
                (k for k in summary_data if "lowest" in k), None)
            if highest_key:
                v = str(summary_data[highest_key][0].value or "")
                if not _course_name_matches(v, gt_highest_name, gt_highest_code):
                    errors.append(
                        f"Highest_Avg_Course: got '{v}', expected '{gt_highest_name}'")
            else:
                errors.append("Summary missing Highest_Avg_Course row")
            if lowest_key:
                v = str(summary_data[lowest_key][0].value or "")
                if not _course_name_matches(v, gt_lowest_name, gt_lowest_code):
                    errors.append(
                        f"Lowest_Avg_Course: got '{v}', expected '{gt_lowest_name}'")
            else:
                errors.append("Summary missing Lowest_Avg_Course row")

    return errors


def check_gform(cur):
    """Check Google Form."""
    errors = []

    cur.execute("""
        SELECT id, title
        FROM gform.forms
        WHERE LOWER(title) LIKE '%%spring 2014%%'
        AND (LOWER(title) LIKE '%%survey%%' OR LOWER(title) LIKE '%%feedback%%')
        ORDER BY created_at DESC
        LIMIT 1
    """)
    form_row = cur.fetchone()

    if not form_row:
        # Try broader search
        cur.execute("""
            SELECT id, title
            FROM gform.forms
            WHERE LOWER(title) LIKE '%%spring 2014%%'
            ORDER BY created_at DESC
            LIMIT 1
        """)
        form_row = cur.fetchone()

    if not form_row:
        errors.append("No Google Form with 'Spring 2014' in title found")
        return errors

    form_id = form_row[0]
    form_title = form_row[1]

    cur.execute("""
        SELECT id, title, question_type, required, config
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()

    if len(questions) < 4:
        errors.append(f"Form has {len(questions)} questions, expected at least 4")
        return errors

    # Collect option labels per question from config JSON
    q_options = {}
    for qid, qtitle, qtype, qreq, qconfig in questions:
        try:
            cfg = qconfig if isinstance(qconfig, dict) else (
                json.loads(qconfig) if qconfig else {})
        except Exception:
            cfg = {}
        opts = cfg.get("options", []) or []
        labels = []
        for o in opts:
            if isinstance(o, dict):
                labels.append(str(o.get("value") or o.get("text") or "").strip().lower())
            else:
                labels.append(str(o or "").strip().lower())
        q_options[qid] = labels

    # Required choice option labels
    needed_satisfaction = {"excellent", "good", "fair", "poor"}
    needed_workload = {"too light", "manageable", "heavy", "overwhelming"}
    needed_rating = {"1", "2", "3", "4", "5"}

    found_sat = any(needed_satisfaction.issubset(set(opts)) for opts in q_options.values())
    found_wl = any(needed_workload.issubset(set(opts)) for opts in q_options.values())
    found_rating = any(needed_rating.issubset(set(opts)) for opts in q_options.values())

    if not found_sat:
        errors.append("Form missing satisfaction question with Excellent/Good/Fair/Poor options")
    if not found_wl:
        errors.append("Form missing workload question with Too Light/Manageable/Heavy/Overwhelming options")
    if not found_rating:
        errors.append("Form missing teaching quality 1-5 rating options")

    # Open-ended (text) question that is NOT required
    optional_text_q = any(
        ("text" in (qtype or "").lower() or "paragraph" in (qtype or "").lower()
         or "short" in (qtype or "").lower())
        and not qreq
        for _qid, _qt, qtype, qreq, _cfg in questions
    )
    if not optional_text_q:
        errors.append("Form missing optional (not required) open-ended text question")

    return errors


def _expected_sender():
    """Derive expected From address from the task-dir email_config.json.

    Returns (email, name). Falls back to the sender named in the task story.
    """
    candidates = []
    here = os.path.dirname(os.path.abspath(__file__))
    candidates.append(os.path.join(here, os.pardir, "email_config.json"))
    candidates.append(os.path.join(os.getcwd(), "email_config.json"))
    for path in candidates:
        p = os.path.abspath(path)
        if not os.path.exists(p):
            continue
        try:
            with open(p, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            if isinstance(cfg, list):
                cfg = cfg[0] if cfg else {}
            cfg = cfg or {}
            email = cfg.get("email", "")
            if email:
                return email, cfg.get("name", "")
        except Exception:
            continue
    return "registrar@university.edu", ""


def check_email(cur):
    """Check email."""
    errors = []

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%%spring 2014%%'
        AND (LOWER(subject) LIKE '%%survey%%' OR LOWER(subject) LIKE '%%feedback%%')
    """)
    emails = cur.fetchall()

    if not emails:
        errors.append("No email with 'Spring 2014' and 'survey/feedback' in subject found")
        return errors

    # Must be sent to students@university.edu (AND not OR with subject)
    matched = [e for e in emails if "students@university.edu" in str(e[2] or "").lower()]
    if not matched:
        errors.append(
            f"No email with subject match AND to_addr=students@university.edu; "
            f"to_addrs observed: {[e[2] for e in emails]}")
        return errors

    # Check from_addr matches the sender derived from the task's email_config.json
    exp_email, exp_name = _expected_sender()
    exp_email_l = exp_email.lower()
    from_ok = any(exp_email_l in str(e[1] or "").lower() for e in matched)
    if not from_ok:
        errors.append(
            f"Email from_addr should be {exp_email}; "
            f"got: {[e[1] for e in matched]}")

    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    all_errors = []

    # Check Excel
    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    # Check GForm and Email
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        print("\n=== Checking Google Form ===")
        gform_errors = check_gform(cur)
        if gform_errors:
            for e in gform_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(gform_errors)
        else:
            print("  [PASS] Google Form check passed")

        print("\n=== Checking Email ===")
        email_errors = check_email(cur)
        if email_errors:
            for e in email_errors:
                print(f"  [FAIL] {e}")
            all_errors.extend(email_errors)
        else:
            print("  [PASS] Email check passed")

        cur.close()
        conn.close()
    except Exception as e:
        err = f"DB check error: {e}"
        print(f"  [FAIL] {err}")
        all_errors.append(err)

    # Summary
    print(f"\n=== SUMMARY ===")
    if all_errors:
        for e in all_errors:
            print(f"  [ERROR] {e}")
        print("  Overall: FAIL")
    else:
        print("  Overall: PASS")

    if args.res_log_file:
        result = {"errors": all_errors, "success": len(all_errors) == 0}
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(1 if all_errors else 0)


if __name__ == "__main__":
    main()
