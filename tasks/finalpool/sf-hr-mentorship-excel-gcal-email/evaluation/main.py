"""Evaluation for sf-hr-mentorship-excel-gcal-email.

Checks:
1. Mentorship_Pairs.xlsx has correct sheets and data
2. Google Calendar event "Mentorship Kickoff Meeting" scheduled 7 days from launch
3. Email sent to program@hr.example.com with subject containing "Mentorship"
"""
import argparse
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="toolathlon_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Mentorship_Pairs.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Mentorship_Pairs.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Mentorship_Pairs.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return False

    all_ok = True

    # Check Pairs sheet
    agent_pairs = get_sheet(agent_wb, "Pairs")
    gt_pairs = get_sheet(gt_wb, "Pairs")
    check("Sheet 'Pairs' exists", agent_pairs is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_pairs is None:
        all_ok = False
    else:
        a_rows = list(agent_pairs.iter_rows(min_row=2, values_only=True))
        g_rows = list(gt_pairs.iter_rows(min_row=2, values_only=True))
        check("Pairs has 10 rows", len(a_rows) == 10, f"Got {len(a_rows)}")
        if len(a_rows) != 10:
            all_ok = False

        # Build lookup by mentor name
        a_lookup = {}
        for r in a_rows:
            if r and r[0]:
                a_lookup[str(r[0]).strip().lower()] = r

        mentor_names_present = 0
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().lower()
            if key in a_lookup:
                mentor_names_present += 1
        check("At least 9 mentor names match", mentor_names_present >= 9,
              f"Matched {mentor_names_present}/10")
        if mentor_names_present < 9:
            all_ok = False

        # Check column count - require all 6 (was >= 5)
        if a_rows:
            ncols = len([v for v in a_rows[0] if v is not None])
            check("Pairs has 6 columns", ncols == 6, f"Got {ncols}")
            if ncols != 6:
                all_ok = False

        # Verify per-row that mentor department/rating, mentee department/experience match GT
        # for matched mentor names
        match_field_errors = []
        a_by_mentor = a_lookup
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_by_mentor.get(key)
            if a_row is None:
                continue
            if not str_match(a_row[1], g_row[1]):
                match_field_errors.append(f"{key} Mentor_Department: '{a_row[1]}' vs '{g_row[1]}'")
            if not num_close(a_row[2], g_row[2], 0.05):
                match_field_errors.append(f"{key} Mentor_Rating: {a_row[2]} vs {g_row[2]}")
            # Mentee_Experience - tighten per-row tol to 0.1 (matches summary tol)
            if len(a_row) > 5 and len(g_row) > 5 and not num_close(a_row[5], g_row[5], 0.1):
                match_field_errors.append(f"{key} Mentee_Experience: {a_row[5]} vs {g_row[5]}")
        # Allow at most 1 mismatch (was 2) — single tie-break swap.
        check("Per-row mentor/mentee fields match GT (at most 1 mismatch)",
              len(match_field_errors) <= 1,
              f"Errors: {match_field_errors[:5]}")
        if len(match_field_errors) > 1:
            all_ok = False

    # Check Program_Summary sheet
    agent_summary = get_sheet(agent_wb, "Program_Summary")
    gt_summary = get_sheet(gt_wb, "Program_Summary")
    check("Sheet 'Program_Summary' exists", agent_summary is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if agent_summary is None:
        all_ok = False
    else:
        a_summary = {}
        for row in agent_summary.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        # Total_Pairs
        tp = a_summary.get("total_pairs")
        check("Total_Pairs = 10", num_close(tp, 10, 0), f"Got {tp}")
        if not num_close(tp, 10, 0):
            all_ok = False

        # Avg_Mentor_Rating
        amr = a_summary.get("avg_mentor_rating")
        check("Avg_Mentor_Rating close to 5.0", num_close(amr, 5.0, 0.1), f"Got {amr}")

        # Avg_Mentee_Experience - tighten tol from 0.5 to 0.1
        ame = a_summary.get("avg_mentee_experience")
        check("Avg_Mentee_Experience close to 0.1", num_close(ame, 0.1, 0.1), f"Got {ame}")
        if not num_close(ame, 0.1, 0.1):
            all_ok = False

    return all_ok


def check_gcal(launch_time_str):
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")
    check("At least 1 calendar event created", len(events) >= 1, f"Found {len(events)}")

    # Check for kickoff meeting
    kickoff_events = [e for e in events if "mentorship" in (e[0] or "").lower() and "kickoff" in (e[0] or "").lower()]
    check("Mentorship Kickoff Meeting event exists", len(kickoff_events) >= 1,
          f"Events: {[e[0] for e in events]}")

    if kickoff_events:
        ev = kickoff_events[0]
        # Description must mention 'mentorship'
        desc = (ev[1] or "").lower()
        check("Kickoff event description mentions mentorship",
              "mentorship" in desc or "mentor" in desc,
              f"Description: {desc[:200]}")
        # Duration ~ 1 hour
        if ev[2] and ev[3]:
            try:
                dur_min = (ev[3] - ev[2]).total_seconds() / 60.0
                check("Kickoff event duration is 1 hour (60 min)",
                      abs(dur_min - 60) <= 5, f"Got {dur_min:.1f} min")
            except Exception:
                pass

    if launch_time_str and kickoff_events:
        try:
            launch_dt = datetime.fromisoformat(launch_time_str)
            expected_dt = launch_dt + timedelta(days=7)
            ev = kickoff_events[0]
            if ev[2]:
                ev_dt = ev[2]
                if hasattr(ev_dt, 'date'):
                    diff = abs((ev_dt.replace(tzinfo=None) - expected_dt).total_seconds())
                    # Tighten from 2 days to 1 hour
                    check("Kickoff meeting exactly 7 days from launch (within 1 hour)",
                          diff <= 3600,
                          f"Expected around {expected_dt}, got {ev_dt} (diff {diff:.0f} sec)")
        except Exception as e:
            print(f"  [INFO] Could not verify date: {e}")

    return len(kickoff_events) >= 1


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%mentorship program launch%'
        ORDER BY date DESC
    """)
    emails = cur.fetchall()
    if not emails:
        # Fallback: any 'mentorship' subject (still validated below by exact match check)
        cur.execute("""
            SELECT subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE subject ILIKE '%mentorship%'
            ORDER BY date DESC
        """)
        emails = cur.fetchall()
    cur.close()
    conn.close()

    check("Email with 'mentorship' in subject exists", len(emails) >= 1,
          f"Found {len(emails)} matching emails")

    if emails:
        e = emails[0]
        subj = (e[0] or "").lower()
        # Subject must contain "Mentorship Program Launch" (case-insensitive)
        check("Subject contains 'Mentorship Program Launch'",
              "mentorship program launch" in subj or
              ("mentorship" in subj and "program" in subj and "launch" in subj),
              f"Subject: '{e[0]}'")
        # Check recipient
        to_str = str(e[2])
        check("Email sent to program@hr.example.com",
              "program@hr.example.com" in to_str.lower(),
              f"to_addr: {to_str}")
        # Check sender
        check("Email from hr@company.example.com",
              "hr@company.example.com" in (e[1] or "").lower(),
              f"from_addr: {e[1]}")
        # Body should mention total pairs - require '10' to appear adjacent to
        # 'pair'/'pairs'/'mentor' (avoids matching dates/years).
        body = (e[3] or "").lower()
        import re as _re
        scoped_ok = bool(_re.search(r"\b10\s*(pair|pairs|mentor)", body))
        check("Email body mentions total pairs scoped to '10 pair(s)/mentor'",
              scoped_ok,
              f"Body sample: {body[:200]}")

    return len(emails) >= 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    check_gcal(args.launch_time)
    check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
