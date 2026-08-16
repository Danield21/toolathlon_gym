"""
Evaluation for howtocook-grocery-budget-planner task.
Checks Excel, Google Sheet, and Calendar events.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Meal_Budget.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Weekly Plan
    wp_rows = load_sheet_rows(wb, "Weekly Plan") or load_sheet_rows(wb, "Weekly_Plan")
    if wp_rows is None:
        record("Sheet 'Weekly Plan' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Weekly Plan' exists", True)
        data = [r for r in wp_rows[1:] if r and r[0] is not None]
        record("Weekly Plan has >= 21 rows (7 days x 3 meals)", len(data) >= 21,
               f"Found {len(data)}")

        cost_col = find_col(wp_rows[0], ["Cost", "cost", "Estimated_Cost"])
        recipe_col = find_col(wp_rows[0], ["Recipe_Name", "Recipe", "recipe_name"])
        ingredients_col = find_col(wp_rows[0], ["Ingredients", "ingredients"])
        if cost_col is not None:
            total = 0
            costs_positive = 0
            for r in data:
                if cost_col < len(r) and r[cost_col] is not None:
                    try:
                        v = float(r[cost_col])
                        total += v
                        if v > 0:
                            costs_positive += 1
                    except (TypeError, ValueError):
                        pass
            record("Total cost > 0 and <= $150",
                   0 < total <= 150, f"Total: ${total:.2f}")
            record("All Weekly Plan rows have positive cost",
                   costs_positive >= 21, f"{costs_positive}/{len(data)} rows positive")

        # Shellfish exclusion: agent must avoid shrimp/crab/lobster/clam/oyster/mussel/scallop
        shellfish_kw = ("shrimp", "crab", "lobster", "clam", "oyster", "mussel", "scallop", "prawn")
        shellfish_hits = 0
        offending = []
        for r in data:
            check_text = ""
            if recipe_col is not None and recipe_col < len(r) and r[recipe_col]:
                check_text += " " + str(r[recipe_col])
            if ingredients_col is not None and ingredients_col < len(r) and r[ingredients_col]:
                check_text += " " + str(r[ingredients_col])
            cl = check_text.lower()
            for kw in shellfish_kw:
                if kw in cl:
                    shellfish_hits += 1
                    offending.append(kw)
                    break
        record("Weekly Plan has no shellfish (allergy constraint)",
               shellfish_hits == 0, f"Found shellfish in {shellfish_hits} rows: {offending[:5]}")

    # Shopping List
    sl_rows = load_sheet_rows(wb, "Shopping List") or load_sheet_rows(wb, "Shopping_List")
    if sl_rows is None:
        record("Sheet 'Shopping List' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Shopping List' exists", True)
        data = [r for r in sl_rows[1:] if r and r[0] is not None]
        # GT has 8 ingredient rows; be lenient at >=7 to allow minor consolidation/split
        record("Shopping List has >= 7 ingredients", len(data) >= 7, f"Found {len(data)}")

    # Budget Summary
    bs_rows = load_sheet_rows(wb, "Budget Summary") or load_sheet_rows(wb, "Budget_Summary")
    if bs_rows is None:
        record("Sheet 'Budget Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Budget Summary' exists", True)
        metrics = {}
        for row in bs_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        twc_key = next((k for k in metrics if "total" in k and "cost" in k), None)
        if twc_key:
            try:
                cost = float(metrics[twc_key])
                record("Total_Weekly_Cost <= 150", cost <= 150, f"Got ${cost}")
            except (TypeError, ValueError):
                record("Total_Weekly_Cost is numeric", False)

        br_key = next((k for k in metrics if "budget" in k and "remain" in k), None)
        if br_key:
            try:
                rem = float(metrics[br_key])
                record("Budget_Remaining >= 0", rem >= 0, f"Got ${rem}")
            except (TypeError, ValueError):
                record("Budget_Remaining is numeric", False)

    # --- Groundtruth XLSX comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Meal_Budget.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)

        # Weekly Plan: validate all 21 rows have Day, Meal, Recipe_Name, Cost
        if "Weekly Plan" in gt_wb.sheetnames or "Weekly_Plan" in gt_wb.sheetnames:
            gt_sname = "Weekly Plan" if "Weekly Plan" in gt_wb.sheetnames else "Weekly_Plan"
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower().replace(" ", "_") == gt_sname.lower().replace(" ", "_"):
                    a_ws = wb[asn]; break
            if a_ws is not None:
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                record(f"GT 'Weekly Plan' row count == {len(gt_rows)}",
                       len(a_rows) == len(gt_rows),
                       f"Expected {len(gt_rows)}, got {len(a_rows)}")
                # Validate ALL Day/Meal pairs are present (7 days x 3 meals = 21 entries)
                gt_pairs = {(str(r[0]).strip().lower(), str(r[1]).strip().lower()) for r in gt_rows if r[0] and r[1]}
                agent_pairs = {(str(r[0]).strip().lower(), str(r[1]).strip().lower()) for r in a_rows if r[0] and r[1]}
                missing_pairs = gt_pairs - agent_pairs
                record(f"All {len(gt_pairs)} Day/Meal pairs present",
                       not missing_pairs,
                       f"Missing: {list(missing_pairs)[:5]}")
                # Each row has a non-empty Recipe_Name and positive Cost
                bad_rows = 0
                for r in a_rows:
                    if len(r) < 5: bad_rows += 1; continue
                    rn = r[2]
                    cost = r[4]
                    if rn is None or str(rn).strip() == "":
                        bad_rows += 1; continue
                    try:
                        c = float(cost)
                        if c <= 0: bad_rows += 1
                    except (TypeError, ValueError):
                        bad_rows += 1
                record("All Weekly Plan rows have non-empty recipe + positive cost",
                       bad_rows == 0,
                       f"{bad_rows} bad rows out of {len(a_rows)}")

        # Budget Summary: validate Total_Weekly_Cost reasonably matches GT
        if "Budget Summary" in gt_wb.sheetnames or "Budget_Summary" in gt_wb.sheetnames:
            gt_sname = "Budget Summary" if "Budget Summary" in gt_wb.sheetnames else "Budget_Summary"
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower().replace(" ", "_") == gt_sname.lower().replace(" ", "_"):
                    a_ws = wb[asn]; break
            if a_ws is not None:
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                record(f"GT 'Budget Summary' row count == {len(gt_rows)}",
                       len(a_rows) == len(gt_rows),
                       f"Expected {len(gt_rows)}, got {len(a_rows)}")
                # Verify required metrics keys are present
                a_metrics = {str(r[0]).strip().lower(): r[1] for r in a_rows if r and r[0]}
                for k in ["total_weekly_cost", "daily_average_cost", "most_expensive_day",
                          "cheapest_day", "budget_remaining"]:
                    record(f"Budget Summary has '{k}'", k in a_metrics,
                           f"keys: {list(a_metrics.keys())}")
                # Total_Weekly_Cost must be <= 150 (already checked) AND reasonable (> 0)
                t = a_metrics.get("total_weekly_cost")
                try:
                    tv = float(t)
                    record("Total_Weekly_Cost is plausible (>0 and <=150)",
                           0 < tv <= 150, f"got {tv}")
                except (TypeError, ValueError):
                    record("Total_Weekly_Cost numeric", False, f"got {t}")
                # Budget_Remaining = 150 - Total_Weekly_Cost
                br = a_metrics.get("budget_remaining")
                try:
                    brv = float(br)
                    tv = float(a_metrics.get("total_weekly_cost"))
                    record("Budget_Remaining = 150 - Total_Weekly_Cost",
                           num_close(brv, 150 - tv, 0.5),
                           f"got {brv}, expected {150-tv}")
                except (TypeError, ValueError):
                    record("Budget_Remaining numeric", False, f"got {br}")

        gt_wb.close()

    return True


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    sheets = cur.fetchall()
    record("Cloud spreadsheet created", len(sheets) >= 1, f"Found {len(sheets)}")

    if sheets:
        # Require title equals 'Family Weekly Meal Plan' (case-insensitive trim)
        target = "family weekly meal plan"
        target_id = None
        for sid, st in sheets:
            if str(st).strip().lower() == target:
                target_id = sid
                break
        record("Sheet titled 'Family Weekly Meal Plan'",
               target_id is not None, f"Titles: {[s[1] for s in sheets]}")

        # 21 weekly plan rows + 8 shopping list rows + 5 budget metrics = ~70+ cells minimum
        if target_id is not None:
            cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (target_id,))
            count = cur.fetchone()[0]
            record("Cloud spreadsheet has substantive data (>= 60 cells)",
                   count >= 60, f"Found {count} cells in target sheet")
        else:
            cur.execute("SELECT COUNT(*) FROM gsheet.cells")
            count = cur.fetchone()[0]
            record("Cloud spreadsheet has substantive data (>= 60 cells)",
                   count >= 60, f"Found {count} cells (no target sheet)")

    cur.close()
    conn.close()
    return True


def check_calendar():
    print("\n=== Checking Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, description, start_datetime, end_datetime FROM gcal.events")
    events = cur.fetchall()

    # Trip 1: 'Weekly Grocery Shopping - Trip 1' on 2026-03-14 10:00-11:00 (1h)
    # Trip 2: 'Midweek Grocery Shopping - Trip 2' on 2026-03-18 17:00-18:00 (1h)
    from datetime import datetime, timezone
    expected_trips = [
        ("Trip 1", "Weekly Grocery Shopping", "2026-03-14",
         datetime(2026, 3, 14, 10, 0, 0, tzinfo=timezone.utc),
         datetime(2026, 3, 14, 11, 0, 0, tzinfo=timezone.utc)),
        ("Trip 2", "Midweek Grocery Shopping", "2026-03-18",
         datetime(2026, 3, 18, 17, 0, 0, tzinfo=timezone.utc),
         datetime(2026, 3, 18, 18, 0, 0, tzinfo=timezone.utc)),
    ]
    def to_utc(dt):
        if dt is None:
            return None
        if hasattr(dt, "tzinfo") and dt.tzinfo is not None:
            return dt.astimezone(timezone.utc)
        return dt.replace(tzinfo=timezone.utc) if hasattr(dt, "replace") else None
    import re
    for label, sub, exp_date, exp_start_utc, exp_end_utc in expected_trips:
        title_match = False
        date_match = False
        start_match = False
        end_match = False
        budget_in_desc = False
        for _, summary, desc, sd, ed in events:
            sm = str(summary or "").lower()
            if sub.lower() in sm and label.lower() in sm:
                title_match = True
                sd_utc = to_utc(sd)
                ed_utc = to_utc(ed)
                if sd_utc and sd_utc.date() == exp_start_utc.date():
                    date_match = True
                if sd_utc and abs((sd_utc - exp_start_utc).total_seconds()) < 60:
                    start_match = True
                if ed_utc and abs((ed_utc - exp_end_utc).total_seconds()) < 60:
                    end_match = True
                d = str(desc or "")
                # Use word-boundary check for '150' (avoid matching 1500, 21500 etc.)
                # Accept '150', '$150', '150.00', '150 USD', '150 dollars', '150 budget'
                if re.search(r'(?<!\d)\$?150(?:\.0{1,2})?(?!\d)', d):
                    budget_in_desc = True
        record(f"'{label}' calendar event title contains '{sub}'", title_match)
        record(f"'{label}' on {exp_date}", date_match)
        record(f"'{label}' starts at {exp_start_utc.strftime('%H:%M')} UTC", start_match)
        record(f"'{label}' ends at {exp_end_utc.strftime('%H:%M')} UTC", end_match)
        record(f"'{label}' description includes total budget (150)", budget_in_desc)

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gsheet()
    check_calendar()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
