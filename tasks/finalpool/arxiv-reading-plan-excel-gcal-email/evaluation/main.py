"""Evaluation for arxiv-reading-plan-excel-gcal-email.

Checks:
1. Reading_Plan.xlsx with Papers sheet (8 rows) and Schedule sheet (8 rows)
2. 8 Google Calendar events for reading sessions
3. Email to reading-group@lab.example.com with "LLM Agent Research Reading Plan" in subject
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0

ARXIV_IDS = ["2301.13379", "2302.01560", "2303.12528", "2305.10403",
             "2308.12950", "2309.17453", "2201.11903", "2310.06825"]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


# --- Fuzzy / normalized comparison helpers -----------------------------------
# The task leaves some fields intentionally flexible (authors separator, the
# "abbreviated title or topic area", and exact 100-char truncation of the
# abstract). These helpers make the groundtruth comparison robust to the
# legitimate variations a correct agent may produce, while still catching
# genuinely wrong output. Published_Date and Category are also tolerant:
# the date accepts the same year-month (published vs updated ambiguity) and
# the category accepts any valid category string for the paper (primary vs
# secondary), because the repository does not expose a direct-by-ID lookup.

def _date_norm(v):
    """Normalize a value to its date portion if it looks like a date."""
    s = str(v).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s


def _date_loose(agent_val, gt_val):
    """Tolerant date check for retrieval-dependent dates.

    Accepts an exact YYYY-MM-DD match OR the same year-month. This tolerates
    legitimate published-vs-updated or off-by-days confusion that a correct
    agent may produce when the repository only exposes one of the two dates,
    while still rejecting a genuinely wrong date (wrong year or wrong month).
    """
    if agent_val is None:
        return False
    a = _date_norm(agent_val)
    g = _date_norm(gt_val)
    if a == g:
        return True
    return len(a) >= 7 and len(g) >= 7 and a[:7] == g[:7]


def _category_ok(agent_val, gt_val):
    """Tolerant arXiv-category check.

    The task asks for the paper's primary category, but the repository returns
    the full category list, so a correct agent may report a secondary category
    or append a human-readable label. Accept an exact match, or one string
    containing the other (e.g. "cs.CL", "cs.CL (Computation and Language)",
    or "cs.CL, cs.AI"). A genuinely wrong category shares no substring and
    still fails.
    """
    if not agent_val or not gt_val:
        return False
    a = str(agent_val).strip().lower()
    g = str(gt_val).strip().lower()
    return a == g or g in a or a in g


def _norm(v):
    return str(v).strip().lower() if v is not None else None


def _authors_ok(agent_val, gt_val):
    """Every groundtruth author must appear in the agent value.

    Order- and separator-insensitive, so ", " / "; " / newline joins all pass.
    """
    if not agent_val or not gt_val:
        return False
    a = str(agent_val).lower()
    parts = [p.strip().lower() for p in str(gt_val).replace(";", ",").split(",") if p.strip()]
    return all(p in a for p in parts)


def _abstract_ok(agent_val, gt_val):
    """Abstract_Summary must be substantive and agree on the first 60 chars."""
    if not agent_val or not gt_val:
        return False
    a = str(agent_val).strip().lower()
    g = str(gt_val).strip().lower()
    if len(a) < 40:
        return False
    return a[:60] == g[:60] or a in g or g in a


_STOPWORDS = {"the", "of", "and", "a", "an", "for", "to", "with", "via", "in",
              "on", "by", "is", "are", "from", "its", "their", "that", "this",
              "models", "model", "using", "use"}


def _topics_ok(agent_val, paper_title):
    """Topics_Covered must share >=1 significant word token with the paper title.

    A single distinctive token is enough because an "abbreviated title" may be
    as short as the paper's short name (e.g. "Toolformer"), while a genuinely
    wrong topic label (e.g. "quantum computing" for a Toolformer session) shares
    zero significant tokens and fails.
    """
    import re
    if not agent_val or not paper_title:
        return False
    sig = lambda s: {w for w in re.findall(r"[a-zA-Z0-9]+", s.lower())
                     if len(w) >= 4 and w not in _STOPWORDS}
    return len(sig(str(agent_val)) & sig(str(paper_title))) >= 1


def check_excel(agent_ws, groundtruth_ws="."):
    print("\n=== Check 1: Reading_Plan.xlsx ===")
    path = os.path.join(agent_ws, "Reading_Plan.xlsx")
    check("File Reading_Plan.xlsx exists", os.path.isfile(path))
    if not os.path.isfile(path):
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel is readable", False, str(e))
        return

    # Check Papers sheet
    papers_ws = None
    for sname in wb.sheetnames:
        if "paper" in sname.lower():
            papers_ws = wb[sname]
            break
    check("Sheet 'Papers' exists", papers_ws is not None, f"Sheets: {wb.sheetnames}")

    if papers_ws is not None:
        rows = list(papers_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Papers sheet has 8 rows", len(non_empty) == 8, f"Got {len(non_empty)}")

        # Check arxiv IDs appear
        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        found_ids = sum(1 for arxiv_id in ARXIV_IDS if arxiv_id in all_text)
        check(f"Papers sheet contains all 8 arXiv IDs",
              found_ids == 8, f"Found {found_ids}/8 IDs in: {all_text[:200]}")

        # Check required columns exist in header row
        header_row = list(papers_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        header_text = " ".join(str(c).lower() for c in header_row if c is not None)
        check("Papers header has ArXiv_ID or arxiv column", "arxiv" in header_text, f"Header: {header_row}")
        check("Papers header has Title column", "title" in header_text, f"Header: {header_row}")
        check("Papers header has Session column", "session" in header_text or "assigned" in header_text, f"Header: {header_row}")

    # Check Schedule sheet
    schedule_ws = None
    for sname in wb.sheetnames:
        if "schedule" in sname.lower():
            schedule_ws = wb[sname]
            break
    check("Sheet 'Schedule' exists", schedule_ws is not None, f"Sheets: {wb.sheetnames}")

    if schedule_ws is not None:
        rows = list(schedule_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Schedule sheet has 8 rows", len(non_empty) == 8, f"Got {len(non_empty)}")

        # Check dates are in March-April 2026
        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        check("Schedule has March 2026 dates",
              "2026" in all_text and ("march" in all_text.lower() or "2026-03" in all_text or "03/09" in all_text),
              f"Date content: {all_text[:200]}")

    # --- Groundtruth value comparison ---
    gt_path = os.path.join(groundtruth_ws, "Reading_Plan.xlsx")
    if not os.path.isfile(gt_path):
        check("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)

    # Paper titles in order (from the GT Papers sheet); used for the fuzzy
    # Topics_Covered check on the Schedule sheet.
    gt_paper_titles = []
    papers_gt = None
    for sn in gt_wb.sheetnames:
        if "paper" in sn.lower():
            papers_gt = gt_wb[sn]
            break
    if papers_gt is not None:
        gt_paper_titles = [str(r[1]) for r in papers_gt.iter_rows(min_row=2, values_only=True)
                           if r and len(r) > 1 and r[1] is not None]

    def cell_ok(is_papers, col_idx, gt_val, a_val, row_idx):
        """Column-aware comparison for a single cell."""
        if gt_val is None:
            return True, "skip"
        if is_papers:
            if col_idx == 2:      # Authors (flexible separator)
                return _authors_ok(a_val, gt_val), "authors"
            if col_idx == 3:      # Published_Date (year-month tolerant)
                return _date_loose(a_val, gt_val), "date"
            if col_idx == 4:      # Category (primary vs secondary category, tolerant)
                return _category_ok(a_val, gt_val), "category"
            if col_idx == 5:      # Abstract_Summary (first 100 chars, fuzzy)
                return _abstract_ok(a_val, gt_val), "abstract"
        else:                     # Schedule
            if col_idx == 1:      # Session_Date (normalized YYYY-MM-DD)
                return (a_val is not None and _date_norm(a_val) == _date_norm(gt_val)), "date"
            if col_idx == 3:      # Topics_Covered (fuzzy vs paper title)
                title = gt_paper_titles[row_idx] if row_idx < len(gt_paper_titles) else ""
                return _topics_ok(a_val, title), "topics"
        # Deterministic columns: exact (numeric or string) match.
        if isinstance(gt_val, (int, float)):
            return num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0)), "value"
        return str_match(a_val, gt_val), "value"

    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws_sheet = gt_wb[gt_sheet_name]
        agent_ws_sheet = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws_sheet = wb[asn]
                break
        if agent_ws_sheet is None:
            check(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws_sheet.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws_sheet.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        check(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        is_papers = "paper" in gt_sheet_name.lower()
        check_indices_list = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices_list.append(len(gt_rows) - 1)
        for idx in check_indices_list:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    ok, rule = cell_ok(is_papers, col_idx, gt_val, a_val, idx)
                    if not ok:
                        check(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx+1} ({rule})",
                              False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    check(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                check(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_gcal():
    print("\n=== Check 2: Google Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, summary, start_datetime FROM gcal.events
        WHERE summary ILIKE '%reading session%'
           OR summary ILIKE '%reading%session%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("At least 8 'Reading Session' calendar events created",
          len(events) >= 8, f"Found {len(events)} events")

    if events:
        summaries = [e[1] for e in events]
        start_dates = [str(e[2]) for e in events]
        check("Events start in March 2026",
              any("2026-03" in d for d in start_dates),
              f"Dates: {start_dates[:4]}")
        check("Events cover 8 weeks (April 2026 included)",
              any("2026-04" in d for d in start_dates),
              f"Dates: {start_dates}")
        check("Events have 'Reading Session' in title",
              all("reading" in s.lower() for s in summaries[:8]),
              f"Titles: {summaries[:4]}")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%reading plan%'
           OR subject ILIKE '%LLM%reading%'
           OR subject ILIKE '%LLM Agent%'
           OR to_addr::text ILIKE '%reading-group%'
        LIMIT 10
    """)
    rows = cur.fetchall()
    check("Email with reading plan subject found",
          len(rows) > 0, "No matching email found")

    if rows:
        to_addrs = [str(r[1]) for r in rows]
        check("Email sent to reading-group@lab.example.com",
              any("reading-group" in addr for addr in to_addrs),
              f"To addresses: {to_addrs}")
        subjects = [r[0] or "" for r in rows]
        check("Email subject contains 'Reading Plan' or 'LLM'",
              any("reading plan" in s.lower() or "llm" in s.lower() for s in subjects),
              f"Subjects: {subjects}")
        bodies = [str(r[2] or "").lower() for r in rows]
        check("Email body mentions total papers (8) or session dates",
              any("8" in b or "march" in b or "reading" in b for b in bodies),
              f"Body: {bodies[0][:200] if bodies else ''}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Evaluation: arxiv-reading-plan-excel-gcal-email ===")

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    excel_failures = FAIL_COUNT  # capture failures so far (Excel only)
    check_gcal()
    check_email()

    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"pass": PASS_COUNT, "fail": FAIL_COUNT}, f)

    # Require >=85% accuracy (tolerates GCal/Email runtime-only failures but catches
    # real errors). Excel (local file) portion must have zero failures.
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0
    excel_ok = excel_failures == 0
    overall_ok = excel_ok and accuracy >= 85
    print(f"Excel failures: {excel_failures}; accuracy: {accuracy:.1f}%")
    sys.exit(0 if overall_ok else 1)


if __name__ == "__main__":
    main()
