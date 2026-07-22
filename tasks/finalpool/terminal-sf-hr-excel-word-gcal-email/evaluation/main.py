"""Evaluation for terminal-sf-hr-excel-word-gcal-email.
Checks:
1. Performance_Review_Report.xlsx with 4 sheets and per-dept correct values
2. Review_Policy_Memo.docx with required sections
3. Calendar events for 7 department reviews on March 9-13, 2026
4. Email to hr_leadership@company.com (exact subject)
5. rating_analysis.py script exists
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

DEPARTMENTS = ["Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"]

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=0.5):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def fetch_dept_stats():
    out = {}
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT "DEPARTMENT", COUNT(*),
                   ROUND(AVG("PERFORMANCE_RATING")::numeric, 2),
                   ROUND(SUM(CASE WHEN "PERFORMANCE_RATING" >= 4 THEN 1 ELSE 0 END)::numeric * 100.0 / COUNT(*), 1),
                   ROUND(SUM(CASE WHEN "PERFORMANCE_RATING" < 2 THEN 1 ELSE 0 END)::numeric * 100.0 / COUNT(*), 1)
            FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
            GROUP BY "DEPARTMENT"
            ORDER BY "DEPARTMENT"
        """)
        for dept, cnt, avg_r, p4, p2 in cur.fetchall():
            out[dept] = {
                "headcount": int(cnt),
                "avg_rating": float(avg_r),
                "pct_above_4": float(p4),
                "pct_below_2": float(p2),
            }
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[fallback] {e}")
        out = {
            "Engineering": {"headcount": 7096, "avg_rating": 3.21, "pct_above_4": 35.7, "pct_below_2": 5.3},
            "Finance":     {"headcount": 7148, "avg_rating": 3.21, "pct_above_4": 35.1, "pct_below_2": 4.7},
            "HR":          {"headcount": 7077, "avg_rating": 3.20, "pct_above_4": 35.1, "pct_below_2": 5.1},
            "Operations":  {"headcount": 7120, "avg_rating": 3.18, "pct_above_4": 34.1, "pct_below_2": 5.3},
            "R&D":         {"headcount": 7083, "avg_rating": 3.20, "pct_above_4": 35.0, "pct_below_2": 5.1},
            "Sales":       {"headcount": 7232, "avg_rating": 3.19, "pct_above_4": 34.9, "pct_below_2": 5.2},
            "Support":     {"headcount": 7244, "avg_rating": 3.20, "pct_above_4": 34.7, "pct_below_2": 5.1},
        }
    return out


def fetch_rating_distribution():
    out = {}
    total = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT "PERFORMANCE_RATING", COUNT(*)
            FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
            GROUP BY 1 ORDER BY 1
        """)
        for r, c in cur.fetchall():
            out[int(r)] = int(c)
            total += int(c)
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[fallback] {e}")
        out = {1: 2553, 2: 7459, 3: 22516, 4: 12464, 5: 5008}
        total = sum(out.values())
    return out, total


def check_excel(workspace):
    print("\n=== Check 1: Performance_Review_Report.xlsx ===")
    path = os.path.join(workspace, "Performance_Review_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]
    check("Has 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    dept_stats = fetch_dept_stats()
    rating_dist, total_emps = fetch_rating_distribution()

    def find_sheet(*keywords):
        for sn in sheets:
            sl = sn.lower().replace(" ", "_")
            if all(kw in sl for kw in keywords):
                return wb[sn]
        return None

    # Department_Ratings
    ws_dr = find_sheet("department", "rating")
    check("Sheet 'Department_Ratings' exists", ws_dr is not None)
    if ws_dr is not None:
        rows = list(ws_dr.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
            data = [r for r in rows[1:] if any(c is not None for c in r)]
            check("Department_Ratings has 7 rows", len(data) == 7, f"Found {len(data)}")

            def col(name):
                if name in header: return header.index(name)
                return None

            d_col = col("department")
            hc_col = col("headcount")
            avg_col = col("avg_rating")
            med_col = col("median_rating")
            p4_col = col("pct_above_4")
            p2_col = col("pct_below_2")

            check("Has all 6 columns (department/headcount/avg_rating/median_rating/pct_above_4/pct_below_2)",
                  None not in (d_col, hc_col, avg_col, med_col, p4_col, p2_col),
                  f"header={header}")

            if d_col is not None:
                lookup = {}
                for r in data:
                    if d_col < len(r) and r[d_col]:
                        lookup[str(r[d_col]).strip()] = r
                for dept, exp in dept_stats.items():
                    a = lookup.get(dept)
                    check(f"Dept '{dept}' present", a is not None,
                          f"departments: {list(lookup.keys())}")
                    if a is None:
                        continue
                    if hc_col is not None and hc_col < len(a):
                        check(f"{dept} headcount", num_close(a[hc_col], exp["headcount"], 1),
                              f"got {a[hc_col]}, expected {exp['headcount']}")
                    if avg_col is not None and avg_col < len(a):
                        check(f"{dept} avg_rating", num_close(a[avg_col], exp["avg_rating"], 0.05),
                              f"got {a[avg_col]}, expected {exp['avg_rating']}")
                    if p4_col is not None and p4_col < len(a):
                        check(f"{dept} pct_above_4", num_close(a[p4_col], exp["pct_above_4"], 0.5),
                              f"got {a[p4_col]}, expected {exp['pct_above_4']}")
                    if p2_col is not None and p2_col < len(a):
                        check(f"{dept} pct_below_2", num_close(a[p2_col], exp["pct_below_2"], 0.5),
                              f"got {a[p2_col]}, expected {exp['pct_below_2']}")

    # Rating_Distribution
    ws_rd = find_sheet("rating", "dist")
    check("Sheet 'Rating_Distribution' exists", ws_rd is not None)
    if ws_rd is not None:
        rows = list(ws_rd.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
            data = [r for r in rows[1:] if any(c is not None for c in r)]
            check("Rating_Distribution has exactly 5 rows", len(data) == 5, f"Found {len(data)}")

            def col(name):
                if name in header: return header.index(name)
                return None

            rv_col = col("rating_value") or col("rating")
            cnt_col = col("count")
            pct_col = col("pct")

            if rv_col is not None and cnt_col is not None:
                lookup = {}
                for r in data:
                    if rv_col < len(r) and r[rv_col] is not None:
                        try:
                            lookup[int(float(r[rv_col]))] = r
                        except (ValueError, TypeError):
                            pass
                for rv, exp_cnt in rating_dist.items():
                    a = lookup.get(rv)
                    check(f"Rating {rv} present in distribution", a is not None,
                          f"keys={list(lookup.keys())}")
                    if a is None:
                        continue
                    if cnt_col is not None and cnt_col < len(a):
                        check(f"Rating {rv} count", num_close(a[cnt_col], exp_cnt, 1),
                              f"got {a[cnt_col]}, expected {exp_cnt}")
                    if pct_col is not None and pct_col < len(a):
                        exp_pct = round(exp_cnt * 100.0 / total_emps, 1)
                        check(f"Rating {rv} pct", num_close(a[pct_col], exp_pct, 0.5),
                              f"got {a[pct_col]}, expected ~{exp_pct}")

    # Review_Calendar
    ws_rc = find_sheet("review", "calendar")
    check("Sheet 'Review_Calendar' exists", ws_rc is not None)
    if ws_rc is not None:
        rows = list(ws_rc.iter_rows(values_only=True))
        data = [r for r in rows[1:] if any(c is not None for c in r)]
        check("Review_Calendar has 7 rows", len(data) == 7, f"Found {len(data)}")

    # Policy_Summary
    ws_ps = find_sheet("policy", "summary")
    if ws_ps is None:
        ws_ps = find_sheet("policy")
    check("Sheet 'Policy_Summary' exists", ws_ps is not None)
    if ws_ps is not None:
        rows = list(ws_ps.iter_rows(values_only=True))
        data = [r for r in rows[1:] if any(c is not None for c in r)]
        check("Policy_Summary has at least 6 rows", len(data) >= 6, f"Found {len(data)}")

        # Build metric -> value lookup
        ps_lookup = {}
        for r in data:
            if r and r[0]:
                ps_lookup[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        # Required metrics
        required = [
            "total_employees",
            "overall_avg_rating",
            "highest_rated_department",
            "lowest_rated_department",
            "total_high_performers",
            "total_underperformers",
        ]
        for m in required:
            check(f"Policy_Summary metric '{m}' present", m in ps_lookup,
                  f"metrics found: {list(ps_lookup.keys())}")

        # Total employees
        if "total_employees" in ps_lookup:
            check("total_employees correct",
                  num_close(ps_lookup["total_employees"], total_emps, 5),
                  f"got {ps_lookup['total_employees']}, expected {total_emps}")
        # Overall avg rating
        try:
            overall_avg = sum(rv * cnt for rv, cnt in rating_dist.items()) / total_emps
            if "overall_avg_rating" in ps_lookup:
                check("overall_avg_rating correct",
                      num_close(ps_lookup["overall_avg_rating"], overall_avg, 0.05),
                      f"got {ps_lookup['overall_avg_rating']}, expected ~{overall_avg:.2f}")
        except Exception:
            pass
        # Highest/lowest rated department
        ranked = sorted(dept_stats.items(), key=lambda x: x[1]["avg_rating"], reverse=True)
        top_avg = ranked[0][1]["avg_rating"]
        bot_avg = ranked[-1][1]["avg_rating"]
        top_candidates = {d for d, s in dept_stats.items() if s["avg_rating"] == top_avg}
        bot_candidates = {d for d, s in dept_stats.items() if s["avg_rating"] == bot_avg}
        if "highest_rated_department" in ps_lookup:
            v = str(ps_lookup["highest_rated_department"]).strip()
            check("highest_rated_department is in top tied set",
                  v in top_candidates,
                  f"got '{v}', candidates {top_candidates}")
        if "lowest_rated_department" in ps_lookup:
            v = str(ps_lookup["lowest_rated_department"]).strip()
            check("lowest_rated_department is in bottom tied set",
                  v in bot_candidates,
                  f"got '{v}', candidates {bot_candidates}")
        # Counts
        exp_high = rating_dist.get(4, 0) + rating_dist.get(5, 0)
        exp_under = rating_dist.get(1, 0)  # below 2 = rating 1 only (integer ratings)
        if "total_high_performers" in ps_lookup:
            check("total_high_performers correct",
                  num_close(ps_lookup["total_high_performers"], exp_high, 5),
                  f"got {ps_lookup['total_high_performers']}, expected {exp_high}")
        if "total_underperformers" in ps_lookup:
            check("total_underperformers correct",
                  num_close(ps_lookup["total_underperformers"], exp_under, 5),
                  f"got {ps_lookup['total_underperformers']}, expected {exp_under}")


def check_word(workspace):
    print("\n=== Check 2: Review_Policy_Memo.docx ===")
    path = os.path.join(workspace, "Review_Policy_Memo.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs)
    full_lower = full_text.lower()

    # Title
    check("Title 'Annual Performance Review Policy Memo' present",
          "annual performance review policy memo" in full_lower,
          "")
    # Sections — task spec says 'overview section' so require literal 'overview'
    check("Has overview section",
          "overview" in full_lower,
          "")
    # Rating distribution section (FIXED parens)
    check("Has rating distribution section",
          ("rating" in full_lower) and ("distribution" in full_lower),
          "")
    # Calibration recommendations
    check("Has calibration/recommendation section",
          ("calibration" in full_lower) or ("recommendation" in full_lower),
          "")
    # All 7 departments mentioned
    for d in DEPARTMENTS:
        check(f"Mentions department {d}", d.lower() in full_lower, "")
    # Substantial content
    check("Has substantial content (>=400 chars)",
          len(full_text) >= 400, f"Length: {len(full_text)}")


def check_gcal():
    print("\n=== Check 3: Calendar Department Reviews ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
        review_events = [e for e in events
                         if e[0] and "performance review" in str(e[0]).lower()]
        check("At least 7 'Performance Review - X' events", len(review_events) >= 7,
              f"Found {len(review_events)}")

        # All 7 departments covered exactly
        summaries = " ".join(str(e[0]) for e in review_events).lower()
        for d in DEPARTMENTS:
            check(f"Event covers '{d}'",
                  d.lower() in summaries,
                  f"Looking for {d} in {summaries[:300]}")

        # Dates fall within March 9-13, 2026
        in_range = [e for e in review_events
                    if e[2] is not None and e[2].strftime("%Y-%m-%d") in
                       ("2026-03-09", "2026-03-10", "2026-03-11", "2026-03-12", "2026-03-13")]
        check("All 7 events within 2026-03-09..13", len(in_range) >= 7,
              f"Got {len(in_range)} in date range out of {len(review_events)}")

        # Friday (2026-03-13) has 2 meetings
        friday_events = [e for e in review_events
                         if e[2] is not None and e[2].strftime("%Y-%m-%d") == "2026-03-13"]
        check("Two meetings on Friday 2026-03-13", len(friday_events) == 2,
              f"Got {len(friday_events)} on Friday")
        if friday_events:
            hours = sorted(e[2].hour for e in friday_events)
            check("Friday meetings at 14:00 and 16:00",
                  hours == [14, 16],
                  f"Got {hours}")
    except Exception as e:
        check("Gcal check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Check 4: Email to HR Leadership ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, subject, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        expected_subject = "Annual Performance Review Cycle - Department Summary"
        target = None
        for em in emails:
            to_str = str(em[2] or "").lower()
            if (em[1] or "").strip().lower() == expected_subject.lower() and "hr_leadership@company.com" in to_str:
                target = em
                break
        check("Email with exact subject + recipient hr_leadership@company.com sent",
              target is not None,
              f"Subjects+to: {[(e[1], e[2]) for e in emails[:5]]}")
        if target:
            body = (target[3] or "").lower()
            check("Body mentions overall avg rating", "average" in body or "avg" in body,
                  "")
            check("Body mentions a department by name",
                  any(d.lower() in body for d in DEPARTMENTS), "")
            # Compute expected underperformer count from DB
            try:
                rd, _ = fetch_rating_distribution()
                exp_under = rd.get(1, 0)
            except Exception:
                exp_under = None
            # Require body to mention 'underperform'/'below 2' AND contain the actual count
            has_term = ("underperform" in body) or ("below 2" in body) or ("below_2" in body)
            count_ok = True
            if exp_under is not None:
                # Match the count as a standalone integer (with optional thousands separator)
                count_str_plain = str(exp_under)
                # Build comma-formatted version (e.g., 2553 -> 2,553)
                count_str_comma = format(exp_under, ',')
                count_ok = bool(re.search(r"(?<!\d)" + re.escape(count_str_plain) + r"(?!\d)", body)) or \
                           bool(re.search(r"(?<!\d)" + re.escape(count_str_comma) + r"(?!\d)", body))
            check("Body mentions underperformer count term + numeric count",
                  has_term and count_ok,
                  f"underperformer term: {has_term}; expected count {exp_under} found: {count_ok}")
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: rating_analysis.py ===")
    path = os.path.join(workspace, "rating_analysis.py")
    check("rating_analysis.py exists", os.path.exists(path))
    if os.path.exists(path):
        with open(path) as f:
            content = f.read().lower()
        # script should at least mention median and percentile somewhere
        check("Script computes median rating",
              "median" in content, "")
        check("Script computes percentile",
              "percentile" in content or "quantile" in content or "25" in content, "")


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Performance_Review_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        has_negative = False
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative values in Excel", not has_negative,
              "Found negative rating/count value")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_gcal()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed")
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "success": FAIL_COUNT == 0}, f)

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
