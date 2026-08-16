"""
Regenerate groundtruth_workspace/Grade_Equity_Analysis.xlsx directly from the
Canvas DB that preprocess leaves in place.

Why this exists
---------------
The previous hand-built GT xlsx drifted from the deterministic seed data: for
AAA-2014J it had Fall_2014_Pass_Rate = 90.9, while the real per-student means in
canvas.submissions give 310/340 = 91.176... -> 91.2. The evaluator then flagged
the model's correct 91.2 as wrong. This script makes GT self-consistent with the
DB so such drift cannot happen again.

Usage (from the project root, with the task's PG instance running):

    python -m tasks.finalpool.canvas-grade-equity-excel-word-gcal.files.generate_groundtruth \
        --output tasks/finalpool/canvas-grade-equity-excel-word-gcal/groundtruth_workspace/Grade_Equity_Analysis.xlsx

    # or, if you only want to print the computed values without writing xlsx:
    python .../generate_groundtruth.py --dry-run

Methodology (must match docs/task.md + initial_workspace/guide.md exactly)
-------------------------------------------------------------------------
- Fall courses: course_code ends with "2013J" or "2014J"; match across years by
  stripping the year suffix (e.g. AAA-2013J <-> AAA-2014J).
- A student's score in a course = mean of their non-null `score` values across
  all submissions for that course.
- Only students with >=1 scored submission are counted (no-submission students
  excluded from every statistic for that course-year).
- Mean / pass_rate rounded to 1 decimal.
- Pass = mean score >= 50.
- Grade bands (guide.md): A 90-100, B 70-89, C 50-69, D 30-49, F 0-29.
- Equity status by |Score_Difference|: <5 Acceptable, 5-10 Concerning, >10 Action Required.
- Overall_Avg_YYYY = sum of every included student's mean score / total number
  of included students across the six compared courses in that year.
- All numeric values written as literal numbers (no Excel formulas).
"""

import argparse
import os
import re
import sys
from collections import defaultdict

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

YEAR_SUFFIXES = {2013: "2013J", 2014: "2014J"}


def fetch_per_student_means(conn):
    """Return {course_base_name: {year: {user_id: mean_score}}}.

    A student's mean is over all their non-null submission scores in that
    course-year. Only users with >=1 scored submission appear.
    """
    cur = conn.cursor()
    cur.execute(
        """
        SELECT c.course_code, s.user_id, AVG(s.score) AS mean_score
        FROM canvas.submissions s
        JOIN canvas.assignments a ON s.assignment_id = a.id
        JOIN canvas.courses c ON a.course_id = c.id
        WHERE s.score IS NOT NULL
          AND (c.course_code LIKE '%%2013J' OR c.course_code LIKE '%%2014J')
        GROUP BY c.course_code, s.user_id
        """
    )
    out = defaultdict(lambda: defaultdict(dict))
    for course_code, user_id, mean_score in cur.fetchall():
        base, year = _split_course_code(course_code)
        if base is None:
            continue
        out[base][year][user_id] = float(mean_score)
    cur.close()
    return out


def _split_course_code(course_code):
    """AAA-2013J -> ('AAA', 2013); AAA-2014J -> ('AAA', 2014); else (None, None)."""
    m = re.match(r"^(.+?)-(2013|2014)J$", (course_code or "").strip())
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


def fetch_course_full_names(conn):
    """Map base course code (AAA) -> short course name (Applied Analytics & Algorithms).

    Derives the short name by stripping the trailing "(Fall YYYY)" / semester
    suffix from canvas.courses.name.
    """
    cur = conn.cursor()
    cur.execute(
        """
        SELECT DISTINCT
            regexp_replace(course_code, '-(2013|2014)J$', '') AS base,
            regexp_replace(name, '\\s*\\(Fall\\s*\\d{4}\\)\\s*$', '') AS short_name
        FROM canvas.courses
        WHERE course_code LIKE '%%2013J' OR course_code LIKE '%%2014J'
        """
    )
    mapping = {}
    for base, short_name in cur.fetchall():
        # Prefer the first non-null short_name seen for each base.
        if short_name and base not in mapping:
            mapping[base] = short_name
    cur.close()
    return mapping


def classify_band(mean_score):
    """guide.md bands. Returns one of A/B/C/D/F."""
    if mean_score >= 90:
        return "A"
    if mean_score >= 70:
        return "B"
    if mean_score >= 50:
        return "C"
    if mean_score >= 30:
        return "D"
    return "F"


def equity_status(score_diff):
    ad = abs(score_diff)
    if ad < 5:
        return "Acceptable"
    if ad <= 10:
        return "Concerning"
    return "Action Required"


def compute_all(conn):
    per_student = fetch_per_student_means(conn)
    name_by_base = fetch_course_full_names(conn)

    # Only bases that exist in BOTH years qualify for comparison.
    bases = sorted(b for b in per_student if 2013 in per_student[b] and 2014 in per_student[b])

    course_compare = []
    grade_distribution = []
    per_year_student_means = {2013: [], 2014: []}  # for Overall_Avg

    for base in bases:
        name = name_by_base.get(base, base)
        row = {"Course_Name": name}
        dist_by_year = {}
        for year in (2013, 2014):
            means = per_student[base][year]
            total = len(means)
            if total == 0:
                row[f"FALL_{year}_MEAN"] = None
                row[f"FALL_{year}_PASS_RATE"] = None
                dist_by_year[year] = None
                continue
            mean_of_means = sum(means.values()) / total
            passed = sum(1 for v in means.values() if v >= 50)
            pass_rate = passed / total * 100.0
            row[f"FALL_{year}_MEAN"] = round(mean_of_means, 1)
            row[f"FALL_{year}_PASS_RATE"] = round(pass_rate, 1)

            bands = defaultdict(int)
            for v in means.values():
                bands[classify_band(v)] += 1
            dist_by_year[year] = {
                "A": bands["A"], "B": bands["B"], "C": bands["C"],
                "D": bands["D"], "F": bands["F"], "Total": total,
                "Passed": passed,
            }
            grade_distribution.append({
                "Course_Name": name, "Year": year,
                "A_Count": bands["A"], "B_Count": bands["B"], "C_Count": bands["C"],
                "D_Count": bands["D"], "F_Count": bands["F"],
                "Total_Students": total,
            })
            per_year_student_means[year].extend(means.values())

        sd = round(row["FALL_2014_MEAN"] - row["FALL_2013_MEAN"], 1)
        row["Score_Difference"] = sd
        row["Pass_Rate_Change"] = round(row["FALL_2014_PASS_RATE"] - row["FALL_2013_PASS_RATE"], 1)
        row["Equity_Status"] = equity_status(sd)
        # Rename internal keys to the GT column names.
        course_compare.append({
            "Course_Name": row["Course_Name"],
            "Fall_2013_Mean": row["FALL_2013_MEAN"],
            "Fall_2014_Mean": row["FALL_2014_MEAN"],
            "Score_Difference": row["Score_Difference"],
            "Fall_2013_Pass_Rate": row["FALL_2013_PASS_RATE"],
            "Fall_2014_Pass_Rate": row["FALL_2014_PASS_RATE"],
            "Pass_Rate_Change": row["Pass_Rate_Change"],
            "Equity_Status": row["Equity_Status"],
            "_dist_by_year": dist_by_year,  # kept for assertion/dry-run, stripped before xlsx write
        })

    summary = []
    n = len(course_compare)
    summary.append(("Total_Courses_Compared", n))
    summary.append(("Courses_Acceptable", sum(1 for r in course_compare if r["Equity_Status"] == "Acceptable")))
    summary.append(("Courses_Concerning", sum(1 for r in course_compare if r["Equity_Status"] == "Concerning")))
    summary.append(("Courses_Action_Required", sum(1 for r in course_compare if r["Equity_Status"] == "Action Required")))
    for year in (2013, 2014):
        vals = per_year_student_means[year]
        summary.append((f"Overall_Avg_{year}", round(sum(vals) / len(vals), 1) if vals else None))
    summary.append(("Overall_Change", round(
        dict(summary)["Overall_Avg_2014"] - dict(summary)["Overall_Avg_2013"], 1)))

    return course_compare, grade_distribution, summary


def write_xlsx(course_compare, grade_distribution, summary, output_path):
    from openpyxl import Workbook

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Course Comparison"
    ws1.append(["Course_Name", "Fall_2013_Mean", "Fall_2014_Mean", "Score_Difference",
                "Fall_2013_Pass_Rate", "Fall_2014_Pass_Rate", "Pass_Rate_Change", "Equity_Status"])
    for r in course_compare:
        ws1.append([r["Course_Name"], r["Fall_2013_Mean"], r["Fall_2014_Mean"],
                    r["Score_Difference"], r["Fall_2013_Pass_Rate"], r["Fall_2014_Pass_Rate"],
                    r["Pass_Rate_Change"], r["Equity_Status"]])

    ws2 = wb.create_sheet("Grade Distribution")
    ws2.append(["Course_Name", "Year", "A_Count", "B_Count", "C_Count", "D_Count", "F_Count", "Total_Students"])
    for d in grade_distribution:
        ws2.append([d["Course_Name"], d["Year"], d["A_Count"], d["B_Count"], d["C_Count"],
                    d["D_Count"], d["F_Count"], d["Total_Students"]])

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    for k, v in summary:
        ws3.append([k, v])

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"[gt] wrote {output_path}")


def assert_self_consistent(course_compare, grade_distribution):
    """Fail loud if GT is internally inconsistent.

    These are the invariants the evaluator now also checks on the GT itself, so
    a drifted GT is caught at generation time rather than silently mis-scoring a
    run.
    """
    dist_lookup = {(d["Course_Name"], d["Year"]): d for d in grade_distribution}
    errors = []
    for r in course_compare:
        name = r["Course_Name"]
        for year, pr_key in [(2013, "Fall_2013_Pass_Rate"), (2014, "Fall_2014_Pass_Rate")]:
            d = dist_lookup.get((name, year))
            if not d:
                continue
            total = d["Total_Students"]
            band_sum = d["A"] + d["B"] + d["C"] + d["D"] + d["F"]
            if total != band_sum:
                errors.append(f"{name} {year}: Total={total} but A+B+C+D+F={band_sum}")
            passed = d["A"] + d["B"] + d["C"]
            expected_pr = round(passed / total * 100, 1) if total else None
            if expected_pr is not None and abs(expected_pr - r[pr_key]) > 0.05:
                errors.append(
                    f"{name} {year}: {pr_key}={r[pr_key]} but ({passed}/{total})*100={expected_pr}"
                )
    if errors:
        print("[gt] SELF-CONSISTENCY CHECK FAILED:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        sys.exit(2)


def main():
    ap = argparse.ArgumentParser()
    default_out = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "groundtruth_workspace", "Grade_Equity_Analysis.xlsx",
    )
    ap.add_argument("--output", default=default_out)
    ap.add_argument("--dry-run", action="store_true", help="print computed GT, write nothing")
    args = ap.parse_args()

    conn = psycopg2.connect(**DB_CONFIG)
    try:
        cc, gd, sm = compute_all(conn)
    finally:
        conn.close()

    assert_self_consistent(cc, gd)

    import json
    print("[gt] Course Comparison:")
    print(json.dumps([{k: v for k, v in r.items() if not k.startswith("_")} for r in cc], indent=2))
    print("[gt] Grade Distribution:")
    print(json.dumps(gd, indent=2))
    print("[gt] Summary:")
    print(json.dumps(dict(sm), indent=2))

    if not args.dry_run:
        write_xlsx(cc, gd, sm, args.output)


if __name__ == "__main__":
    main()
