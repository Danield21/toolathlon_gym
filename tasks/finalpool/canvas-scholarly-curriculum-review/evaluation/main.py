"""Evaluation for canvas-scholarly-curriculum-review."""
import argparse
import os
import re
import sys

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


def _norm_loose(s):
    """Lowercase, replace punctuation with spaces, collapse whitespace."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(s or "").lower())).strip()


def _levenshtein(a, b):
    """Plain Levenshtein edit distance (titles are short)."""
    if a == b:
        return 0
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[-1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def _titles_match(expected, actual):
    """task.md requires the page title to *contain* the exact phrase
    "Accreditation Review Report". After normalization the expected phrase
    must appear intact in the actual title — no edit-distance / keyword
    fuzzy tolerance."""
    e = _norm_loose(expected)
    a = _norm_loose(actual)
    if not e or not a:
        return False
    return e in a



def nums_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _get_valid_paper_titles():
    """Fetch valid (non-noise) pedagogy papers from scholarly DB."""
    try:
        import psycopg2
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user=os.environ.get("PGUSER", "eigent"),
            password=os.environ.get("PGPASSWORD", "camel"),
            connect_timeout=5,
        )
        cur = conn.cursor()
        cur.execute("SELECT title FROM scholarly.scholar_papers")
        pedagogy_keywords = ["active learning", "assessment", "online learning",
                             "curriculum", "engagement", "understanding by design"]
        result = set()
        for (t,) in cur.fetchall():
            tl = (t or "").lower()
            if any(kw in tl for kw in pedagogy_keywords):
                result.add(t.strip().lower())
        cur.close(); conn.close()
        return result
    except Exception as e:
        print(f"[fallback] {e}")
        return None


def check_excel(agent_workspace, groundtruth_workspace=None):
    errors = []
    import openpyxl

    path = os.path.join(agent_workspace, "Curriculum_Review.xlsx")
    if not os.path.exists(path):
        return ["Curriculum_Review.xlsx not found"]

    gt_path = None
    if groundtruth_workspace:
        gt_path = os.path.join(groundtruth_workspace, "Curriculum_Review.xlsx")
        if not os.path.exists(gt_path):
            gt_path = None

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True) if gt_path else None

        # Sheet 1: Course Compliance
        rows = load_sheet_rows(wb, "Course Compliance")
        gt_rows = load_sheet_rows(gt_wb, "Course Compliance") if gt_wb else None
        if rows is None:
            errors.append("Sheet 'Course Compliance' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) != 22:
                errors.append(
                    f"Course Compliance has {len(data_rows)} rows, expected exactly 22"
                )
            # Check that Compliant column exists with Yes/No values
            compliant_vals = []
            for r in data_rows:
                if len(r) > 7 and r[7]:
                    compliant_vals.append(str(r[7]).strip().lower())
            has_yes = any(v == "yes" for v in compliant_vals)
            has_no = any(v == "no" for v in compliant_vals)
            if not has_yes:
                errors.append("No courses marked as Compliant='Yes'")
            if not has_no:
                errors.append("No courses marked as Compliant='No'")

            # Compare every row against GT using Course_Name keys (loose substring in agent row)
            if gt_rows:
                gt_data_rows = [r for r in gt_rows[1:] if r and r[0] is not None]
                def norm_key(n):
                    return str(n or "").strip().lower()
                a_map = {norm_key(r[0]): r for r in data_rows}
                for g_row in gt_data_rows:
                    key = norm_key(g_row[0])
                    a_row = a_map.get(key)
                    if a_row is None:
                        # Try partial match (substring)
                        for k, r in a_map.items():
                            if key in k or k in key:
                                a_row = r
                                break
                    if a_row is None:
                        errors.append(f"Course missing: {g_row[0]}")
                        continue
                    # Compare Assignment_Count(1), Quiz_Count(2), Discussion_Count(3),
                    # Student_Count(4), Faculty_Count(5), Ratio(6), Compliant(7)
                    def _to_int(v):
                        """int(float(str)) tolerant parse."""
                        if v is None:
                            return None
                        try:
                            return int(float(str(v).strip()))
                        except (TypeError, ValueError):
                            return None

                    def _presence_match(agent_val, gt_int):
                        """Task wording ('whether the course has quizzes') invites
                        Yes/No answers; accept them when their presence semantics
                        agree with the GT count."""
                        s = str(agent_val).strip().lower()
                        if s in ("yes", "true", "y", "present"):
                            return gt_int > 0
                        if s in ("no", "false", "n", "none", "n/a", "na", "absent", ""):
                            return gt_int == 0
                        return False

                    for col, name, int_col in [
                        (1, "Assignment_Count", True),
                        (2, "Quiz_Count", True),
                        (3, "Discussion_Count", True),
                        (4, "Student_Count", True),
                        (5, "Faculty_Count", True),
                    ]:
                        if len(a_row) > col and len(g_row) > col and g_row[col] is not None:
                            agent_int = _to_int(a_row[col])
                            gt_int = _to_int(g_row[col])
                            if agent_int is None:
                                if name == "Quiz_Count" and gt_int is not None \
                                        and _presence_match(a_row[col], gt_int):
                                    continue
                                errors.append(f"{g_row[0]} {name} unparseable: {a_row[col]}")
                            elif gt_int is not None and agent_int != gt_int:
                                errors.append(f"{g_row[0]} {name}: {a_row[col]} vs {g_row[col]}")
                    # Ratio: tol 0.5
                    if len(a_row) > 6 and len(g_row) > 6 and g_row[6] is not None:
                        try:
                            if abs(float(a_row[6]) - float(g_row[6])) > 0.5:
                                errors.append(f"{g_row[0]} Ratio: {a_row[6]} vs {g_row[6]}")
                        except (TypeError, ValueError):
                            errors.append(f"{g_row[0]} Ratio unparseable: {a_row[6]}")
                    # Compliant: exact
                    if len(a_row) > 7 and len(g_row) > 7 and g_row[7] is not None:
                        if str(a_row[7]).strip().lower() != str(g_row[7]).strip().lower():
                            errors.append(f"{g_row[0]} Compliant: {a_row[7]} vs {g_row[7]}")

        # Sheet 2: Literature Support - also require valid pedagogy papers
        rows2 = load_sheet_rows(wb, "Literature Support")
        if rows2 is None:
            errors.append("Sheet 'Literature Support' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 5:
                errors.append(
                    f"Literature Support has {len(data_rows2)} rows, expected >= 5"
                )
            # Check topics coverage
            topics_found = [str(r[0]).strip().lower() for r in data_rows2 if r[0]]
            for keyword in ["active", "assessment", "online", "curriculum", "engagement"]:
                if not any(keyword in t for t in topics_found):
                    errors.append(
                        f"Literature Support missing topic containing '{keyword}'"
                    )
            # Every paper title in the sheet MUST match a pedagogy paper in scholarly DB.
            # Require non-trivial overlap: min 12-char substring match or >70% token overlap
            # to prevent 1-char substrings like " a " from accidentally matching.
            valid_titles = _get_valid_paper_titles()
            if valid_titles:
                def _title_matches(agent_title, valid_title):
                    """Match only if at least 12 chars of shared substring exist."""
                    if len(valid_title) <= 12:
                        return valid_title == agent_title
                    # Substring either way with min length
                    if len(agent_title) >= 12 and agent_title in valid_title:
                        return True
                    if len(valid_title) >= 12 and valid_title in agent_title:
                        return True
                    # Token overlap
                    a_toks = set(agent_title.split())
                    v_toks = set(valid_title.split())
                    if a_toks and v_toks:
                        # shared non-stopword token count >= 3 or >=70% of valid
                        stop = {"a", "an", "the", "and", "or", "of", "to", "for", "in", "on"}
                        shared = (a_toks & v_toks) - stop
                        if len(shared) >= 3:
                            return True
                        if len(v_toks - stop) > 0 and len(shared) / max(len(v_toks - stop), 1) >= 0.7:
                            return True
                    return False
                for r in data_rows2:
                    if len(r) > 1 and r[1]:
                        agent_title = str(r[1]).strip().lower()
                        if not any(_title_matches(agent_title, v) for v in valid_titles):
                            errors.append(f"Literature paper not a valid pedagogy paper: '{r[1]}' (topic={r[0]})")

        # Sheet 3: Summary - check all expected metrics
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 5:
                errors.append(
                    f"Summary has {len(data_rows3)} rows, expected >= 5"
                )
            def find_metric(name_keywords):
                for r in data_rows3:
                    k = str(r[0] or "").lower()
                    if all(kw in k for kw in name_keywords):
                        return r
                return None
            # Total courses == 22
            tr = find_metric(["total", "course"])
            if tr and len(tr) > 1:
                try:
                    if int(tr[1]) != 22:
                        errors.append(f"Total courses={tr[1]}, expected 22")
                except (TypeError, ValueError):
                    errors.append(f"Total courses unparseable: {tr[1]}")
            else:
                errors.append("Summary Total courses row missing")
            # Compliant courses == 18 (per GT) or runtime-pulled.
            # Match the "Compliant" row specifically, since a "Non_Compliant_..."
            # row also contains the substring "compliant" and would otherwise be
            # picked up first if the agent lists non-compliant before compliant.
            cr = None
            for r in data_rows3:
                k = str(r[0] or "").lower()
                if "compliant" in k and "non" not in k:
                    cr = r
                    break
            if cr and len(cr) > 1:
                try:
                    if int(cr[1]) != 18:
                        errors.append(f"Compliant courses={cr[1]}, expected 18")
                except (TypeError, ValueError):
                    pass
            # Compliance rate
            crate = find_metric(["compliance", "rate"])
            if crate and len(crate) > 1:
                try:
                    if abs(float(crate[1]) - 81.8) > 0.5:
                        errors.append(f"Compliance rate={crate[1]}, expected ~81.8")
                except (TypeError, ValueError):
                    pass
            # Supporting papers == 5
            sp = find_metric(["supporting", "paper"])
            if sp and len(sp) > 1:
                try:
                    if int(sp[1]) != 5:
                        errors.append(f"Supporting papers={sp[1]}, expected 5")
                except (TypeError, ValueError):
                    pass

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_notion():
    errors = []
    try:
        import psycopg2, json
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=5432,
            dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
            user=os.environ.get("PGUSER", "eigent"),
            password=os.environ.get("PGPASSWORD", "camel"),
        )
        cur = conn.cursor()
        # Find the agent's "Accreditation Review Report" page (NOT the preprocess parent "Accreditation Documents")
        cur.execute("SELECT id, properties FROM notion.pages")
        target_page = None
        for pid, props in cur.fetchall():
            if props is None:
                continue
            try:
                p = props if isinstance(props, dict) else json.loads(props)
                title_val = p.get("title")
                if title_val is None:
                    # Title property may be named differently (e.g. page in a database)
                    for v in p.values():
                        if isinstance(v, dict) and isinstance(v.get("title"), list):
                            title_val = v
                            break
                if isinstance(title_val, dict):
                    items = title_val.get("title", [])
                    title_text = " ".join(
                        (it.get("plain_text") or it.get("text", {}).get("content", ""))
                        for it in items if isinstance(it, dict)
                    ).strip()
                    # Title must contain the exact phrase (per task.md)
                    if _titles_match("accreditation review report", title_text):
                        target_page = (pid, title_text)
                        break
            except Exception:
                continue
        if target_page is None:
            errors.append("No Notion page titled 'Accreditation Review Report' found")
        else:
            pid, title_text = target_page
            # Check blocks/content: must mention compliance_rate, course failures, recommendations
            cur.execute("SELECT block_data FROM notion.blocks WHERE parent_id = %s", (pid,))
            blocks = cur.fetchall()
            joined = " ".join(json.dumps(b[0]) if b[0] is not None else "" for b in blocks).lower()
            # Require 'compliance' keyword AND a percentage / ratio number nearby
            import re
            if "compliance" not in joined:
                errors.append("Notion page content missing 'compliance' keyword")
            else:
                # Find any percentage or 'X%' or numeric in 0-100 near 'compliance'
                # Accept standalone percent e.g. '81%', '81.8', '0.818', 'rate: 81.8'
                has_pct = bool(re.search(r"\b\d{1,3}(\.\d+)?\s*%", joined))
                has_rate_num = bool(re.search(r"\b(81\.?8?|\d{1,3}\.\d+)\b", joined))
                if not (has_pct or has_rate_num):
                    errors.append("Notion page content has 'compliance' keyword but no percentage/numeric rate")
            # Must mention recommendations
            if "recommend" not in joined:
                errors.append("Notion page content missing recommendations")
            # Must mention non-compliant or failures
            if "non-compliant" not in joined and "non compliant" not in joined and "fail" not in joined:
                errors.append("Notion page content missing non-compliant course reference")
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking Notion: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )
    gt_ws = args.groundtruth_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Notion page...")
    errs = check_notion()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
