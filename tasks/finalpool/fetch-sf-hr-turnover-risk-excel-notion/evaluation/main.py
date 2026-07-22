"""Evaluation script for fetch-sf-hr-turnover-risk-excel-notion."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Turnover_Risk_Assessment.xlsx")
    check("Turnover_Risk_Assessment.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Turnover_Risk_Assessment.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        # Helper to load a sheet to dict-rows by header
        def _sheet_to_dicts(ws):
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return [], []
            hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
            out = []
            for r in rows[1:]:
                if r is None:
                    continue
                d = {}
                for i, h in enumerate(hdr):
                    d[h.lower()] = r[i] if i < len(r) else None
                out.append(d)
            return hdr, out

        check("Risk_Overview sheet exists", "Risk_Overview" in wb.sheetnames)
        if "Risk_Overview" in wb.sheetnames:
            ws = wb["Risk_Overview"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Risk_Overview has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Employee_Count', 'Avg_Salary', 'Risk_Level']:
                check(f"Risk_Overview has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Value-level: department names must overlap with GT departments
            if gt_wb is not None and "Risk_Overview" in gt_wb.sheetnames:
                try:
                    _, agent_dicts = _sheet_to_dicts(ws)
                    _, gt_dicts = _sheet_to_dicts(gt_wb["Risk_Overview"])
                    gt_depts = {str(d.get('department') or '').strip().lower() for d in gt_dicts}
                    gt_depts.discard('')
                    agent_depts = {str(d.get('department') or '').strip().lower() for d in agent_dicts}
                    overlap = gt_depts & agent_depts
                    check("Risk_Overview Department values match GT (>=6 of 7)",
                          len(overlap) >= 6, f"overlap={overlap}, gt={gt_depts}")
                    # Risk_Level must be one of {High, Medium, Low} for at least 6 rows
                    valid_levels = {'high', 'medium', 'low'}
                    valid_count = sum(1 for d in agent_dicts
                                      if str(d.get('risk_level') or '').strip().lower() in valid_levels)
                    check("Risk_Overview has >=6 valid Risk_Level values",
                          valid_count >= 6, f"valid count={valid_count}")

                    # Per-row Risk_Level correctness check (vs computed from
                    # Avg_Satisfaction & Risk_Threshold per task definition)
                    correct = 0
                    total_check = 0
                    for d in agent_dicts:
                        sat = safe_float(d.get('avg_satisfaction'))
                        thr = safe_float(d.get('risk_threshold'))
                        rl = str(d.get('risk_level') or '').strip().lower()
                        if sat is None or thr is None or not rl:
                            continue
                        total_check += 1
                        if sat < thr:
                            expected = 'high'
                        elif sat <= thr + 0.5:
                            expected = 'medium'
                        else:
                            expected = 'low'
                        if rl == expected:
                            correct += 1
                    check("Risk_Overview per-row Risk_Level matches task formula (>=5 correct)",
                          correct >= 5,
                          f"{correct}/{total_check} correct per-row classifications")
                except Exception as _e:
                    check("Risk_Overview value-check", False, str(_e))

        check("Risk_Summary sheet exists", "Risk_Summary" in wb.sheetnames)
        if "Risk_Summary" in wb.sheetnames:
            ws = wb["Risk_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Risk_Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Risk_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Value-level: required metric keys present + value sanity
            try:
                metric_keys = {str(r[0] or '').strip().lower() for r in data_rows
                               if r and len(r) > 0 and r[0]}
                required = {'total_departments', 'highest_risk_department',
                            'high_risk_count', 'medium_risk_count', 'low_risk_count',
                            'total_at_risk_employees'}
                missing = required - metric_keys
                check("Risk_Summary contains all 6 required metric keys",
                      not missing, f"missing={missing}, present={metric_keys}")

                # Value-level: Total_Departments must be 6-8 (GT has 7)
                metric_dict = {}
                for r in data_rows:
                    if r and len(r) >= 2 and r[0]:
                        metric_dict[str(r[0]).strip().lower()] = r[1]
                td = safe_float(metric_dict.get('total_departments'))
                check("Risk_Summary Total_Departments in [6,8]",
                      td is not None and 6 <= td <= 8,
                      f"got Total_Departments={td}")

                # Sum of High+Medium+Low must equal Total_Departments (sanity)
                hi = safe_float(metric_dict.get('high_risk_count'), 0) or 0
                me = safe_float(metric_dict.get('medium_risk_count'), 0) or 0
                lo = safe_float(metric_dict.get('low_risk_count'), 0) or 0
                check("Risk_Summary High+Medium+Low == Total_Departments",
                      td is not None and abs((hi + me + lo) - td) < 0.5,
                      f"sum={hi+me+lo}, total={td}")
            except Exception as _e:
                check("Risk_Summary metric-values check", False, str(_e))

        check("Detailed_Metrics sheet exists", "Detailed_Metrics" in wb.sheetnames)
        if "Detailed_Metrics" in wb.sheetnames:
            ws = wb["Detailed_Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Detailed_Metrics has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Satisfaction_Gap', 'Estimated_Turnover_Cost']:
                check(f"Detailed_Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Value-level: Estimated_Turnover_Cost must be numeric and positive in at least 5 rows
            try:
                _, agent_dicts = _sheet_to_dicts(ws)
                pos_cost = sum(1 for d in agent_dicts
                               if safe_float(d.get('estimated_turnover_cost')) is not None
                               and safe_float(d.get('estimated_turnover_cost'), -1) > 0)
                check("Detailed_Metrics has >=5 positive Estimated_Turnover_Cost values",
                      pos_cost >= 5, f"positive count={pos_cost}")
            except Exception as _e:
                check("Detailed_Metrics cost-check", False, str(_e))

        # Check Notion page (tightened: must include 'dashboard' too per task)
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                """SELECT properties FROM notion.pages
                   WHERE archived = false
                     AND properties::text ILIKE %s
                     AND properties::text ILIKE %s
                     AND properties::text ILIKE %s""",
                ('%turnover%', '%risk%', '%dashboard%'))
            pages = cur.fetchall()
            check("Notion 'Turnover Risk Dashboard' page created",
                  len(pages) >= 1, f"found {len(pages)} pages")
            conn.close()
        except Exception as e:
            check("Notion verification", False, str(e))

        check("risk_scorer.py exists", os.path.exists(os.path.join(agent_workspace, "risk_scorer.py")))
        check("risk_assessment.json exists", os.path.exists(os.path.join(agent_workspace, "risk_assessment.json")))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
