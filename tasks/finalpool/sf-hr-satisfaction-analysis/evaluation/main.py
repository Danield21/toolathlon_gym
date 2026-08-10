"""Evaluation for sf-hr-satisfaction-analysis."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_header(h):
    return " ".join(str(h or "").strip().lower().replace("-", " ").replace("_", " ").split())


def _header_map(rows):
    """Map normalized header text -> 0-based column index from row 1 of a sheet.
    Lets checks locate columns by name regardless of column order."""
    if not rows or not rows[0]:
        return {}
    return {_norm_header(h): i for i, h in enumerate(rows[0]) if _norm_header(h)}


def _cell(row, idx):
    """Safe cell access (None when idx out of range)."""
    if row is None or idx is None:
        return None
    try:
        return row[idx] if idx < len(row) else None
    except (TypeError, IndexError):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Satisfaction_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Satisfaction_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Satisfaction Analysis
    print(f"  Checking Satisfaction Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Satisfaction Analysis")
    g_rows = load_sheet_rows(gt_wb, "Satisfaction Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Satisfaction Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Satisfaction Analysis' not found in groundtruth")
    else:
        sheet_name = "Satisfaction Analysis"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Columns located by header name (order-independent).
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            # Avg_Satisfaction (tol 0.1)
            av = _cell(a_row, a_col.get("avg satisfaction"))
            gv = _cell(g_row, g_col.get("avg satisfaction"))
            if av is not None and gv is not None and not num_close(av, gv, 0.1):
                errors.append(f"{key}.Avg_Satisfaction: {av} vs {gv} (tol=0.1)")

            # Avg_Work_Life_Balance (tol 0.1)
            av = _cell(a_row, a_col.get("avg work life balance"))
            gv = _cell(g_row, g_col.get("avg work life balance"))
            if av is not None and gv is not None and not num_close(av, gv, 0.1):
                errors.append(f"{key}.Avg_Work_Life_Balance: {av} vs {gv} (tol=0.1)")

            # Avg_Rating (tol 0.1)
            av = _cell(a_row, a_col.get("avg rating"))
            gv = _cell(g_row, g_col.get("avg rating"))
            if av is not None and gv is not None and not num_close(av, gv, 0.1):
                errors.append(f"{key}.Avg_Rating: {av} vs {gv} (tol=0.1)")

            # Employees (tol 1)
            av = _cell(a_row, a_col.get("employees"))
            gv = _cell(g_row, g_col.get("employees"))
            if av is not None and gv is not None and not num_close(av, gv, 1):
                errors.append(f"{key}.Employees: {av} vs {gv} (tol=1)")

        # Validate sort order: descending by Avg_Satisfaction (resolve column by
        # header name; still compares the rounded cell values, semantics unchanged).
        sat_idx = a_col.get("avg satisfaction")
        a_satisfaction = []
        for r in a_data:
            v = _cell(r, sat_idx) if sat_idx is not None else None
            if r and r[0] is not None and v is not None:
                try:
                    a_satisfaction.append(float(v))
                except (TypeError, ValueError):
                    pass
        if a_satisfaction:
            sorted_desc = sorted(a_satisfaction, reverse=True)
            if a_satisfaction != sorted_desc:
                errors.append(f"Satisfaction Analysis not sorted by Avg_Satisfaction descending: {a_satisfaction}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    # Accept either of the top two raw-satisfaction departments for Happiest_Dept,
    # since rounding to 2 decimals creates a near-tie between Finance and Sales.
    HAPPIEST_ACCEPTED = {"finance", "sales"}
    LEAST_HAPPY_ACCEPTED = {"r&d", "rd", "r and d", "r & d"}
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Columns located by header name (order-independent).
        a_col = _header_map(a_rows)
        g_col = _header_map(g_rows)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            av = _cell(a_row, a_col.get("value"))
            gv = _cell(g_row, g_col.get("value"))
            if av is not None and gv is not None:
                # Tightened: 0.1 for numeric averages; string fields use exact-set match
                if key == "happiest_dept":
                    avs = str(av or "").strip().lower()
                    if avs not in HAPPIEST_ACCEPTED:
                        errors.append(f"{key}.Value: {av} not in accepted set {HAPPIEST_ACCEPTED}")
                elif key == "least_happy_dept":
                    avs = str(av or "").strip().lower()
                    if avs not in LEAST_HAPPY_ACCEPTED:
                        errors.append(f"{key}.Value: {av} not in accepted set {LEAST_HAPPY_ACCEPTED}")
                else:
                    if not num_close(av, gv, 0.1):
                        errors.append(f"{key}.Value: {av} vs {gv} (tol=0.1)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    
    docx_path = os.path.join(args.agent_workspace, "Satisfaction_Summary.docx")
    if not os.path.exists(docx_path):
        all_errors.append("Satisfaction_Summary.docx not found")
    else:
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs).lower()
            _headings = " ".join(p.text for p in _doc.paragraphs if p.style.name.startswith("Heading")).lower()
            # Tighter requirement: at least 200 chars and multiple topic keywords
            if len(_text.strip()) < 200:
                all_errors.append(f"Satisfaction_Summary.docx has too little text content (< 200 chars, got {len(_text.strip())})")
            _required_kws = ["satisfaction"]
            # Need at least 2 of these supporting topic words to demonstrate narrative depth
            _supporting_kws = ["work-life", "work life", "balance", "department", "rating", "trend"]
            _missing = [k for k in _required_kws if k not in _text and k not in _headings]
            if _missing:
                all_errors.append(f"Satisfaction_Summary.docx missing required keywords: {_missing}")
            _support_hits = sum(1 for k in _supporting_kws if k in _text or k in _headings)
            if _support_hits < 2:
                all_errors.append(
                    f"Satisfaction_Summary.docx narrative too shallow: only {_support_hits} of "
                    f"{_supporting_kws} found (need at least 2)"
                )
        except ImportError:
            if os.path.getsize(docx_path) < 100:
                all_errors.append("Satisfaction_Summary.docx too small")
        except Exception as _e:
            all_errors.append(f"Error reading Satisfaction_Summary.docx: {_e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
