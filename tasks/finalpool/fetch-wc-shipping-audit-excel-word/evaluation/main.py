"""Evaluation script for fetch-wc-shipping-audit-excel-word."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "fetch-wc-shipping-audit-excel-word"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    
    excel_path = os.path.join(agent_workspace, "Shipping_Audit_Report.xlsx")
    check("Shipping_Audit_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Shipping_Audit_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Category', 'Product_Count', 'Our_Avg_Price', 'Total_Sales', 'Market_Avg_Price', 'Price_Gap_Pct'], 6)

            # Value-level: category names overlap with GT (>=3 of 6) AND row-level value match
            if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames:
                try:
                    gt_ws = gt_wb["Data_Analysis"]
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    gt_cats = {str(r[0]).strip().lower() for r in gt_rows if r and r[0]}
                    agent_cats = {str(r[0]).strip().lower() for r in data_rows if r and r[0]}
                    overlap = gt_cats & agent_cats
                    check("Data_Analysis Category values match GT (>=3)",
                          len(overlap) >= 3, f"overlap={overlap}, gt={gt_cats}")
                    # Price_Gap_Pct must be numeric in >=3 rows (column index 5)
                    pg_count = 0
                    for r in data_rows:
                        if r and len(r) > 5 and safe_float(r[5]) is not None:
                            pg_count += 1
                    check("Data_Analysis Price_Gap_Pct numeric in >=3 rows",
                          pg_count >= 3, f"numeric count={pg_count}")
                    # Row-level value check: build category->row maps, compare numeric cols.
                    gt_map = {}
                    for r in gt_rows:
                        if r and r[0]:
                            gt_map[str(r[0]).strip().lower()] = r
                    matched_value_rows = 0
                    for r in data_rows:
                        if not r or not r[0]:
                            continue
                        key = str(r[0]).strip().lower()
                        gt_r = gt_map.get(key)
                        if gt_r is None or len(gt_r) < 6 or len(r) < 6:
                            continue
                        # Compare Our_Avg_Price (col 2): 5% relative
                        gtp = safe_float(gt_r[2])
                        ap = safe_float(r[2])
                        # Compare Price_Gap_Pct (col 5): ±1.0 absolute
                        gtg = safe_float(gt_r[5])
                        ag = safe_float(r[5])
                        ok_p = (gtp is not None and ap is not None and
                                abs(gtp - ap) <= max(0.5, abs(gtp) * 0.05))
                        ok_g = (gtg is not None and ag is not None and
                                abs(gtg - ag) <= 1.0)
                        if ok_p and ok_g:
                            matched_value_rows += 1
                    check("Data_Analysis at least 3 rows with matching Our_Avg_Price (5% rel) AND Price_Gap_Pct (+/-1.0)",
                          matched_value_rows >= 3,
                          f"matched_value_rows={matched_value_rows}")
                except Exception as _e:
                    check("Data_Analysis value-check", False, str(_e))
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 3)
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action', 'Category'], 2)
        # Check Word document — task explicitly names Shipping_Audit_Analysis.docx
        import glob as globmod
        target_word_path = os.path.join(agent_workspace, "Shipping_Audit_Analysis.docx")
        all_word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        # Accept the explicit name OR any *shipping*audit*analysis*.docx variant
        if os.path.exists(target_word_path):
            word_files = [target_word_path]
        else:
            word_files = [
                p for p in all_word_files
                if "shipping" in os.path.basename(p).lower()
                and "audit" in os.path.basename(p).lower()
            ]
        check(
            "Shipping_Audit_Analysis.docx exists (or shipping-audit-named .docx)",
            len(word_files) >= 1,
            f"target: Shipping_Audit_Analysis.docx; all docx found: {[os.path.basename(p) for p in all_word_files]}",
        )
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content (>=200 chars)", len(text) >= 200, f"text length: {len(text)}")
            has_findings = any(kw in text for kw in ['finding', 'summary', 'analysis', 'overview'])
            has_recs = any(kw in text for kw in ['recommendation', 'action', 'priority', 'next step'])
            check("Word has summary/findings section", has_findings, f"text head: {text[:200]}")
            check("Word has recommendations/action section", has_recs, f"text head: {text[:200]}")

        proc_path = os.path.join(agent_workspace, "wc_shipping_processor.py")
        check("wc_shipping_processor.py exists", os.path.exists(proc_path))
        if os.path.exists(proc_path):
            try:
                with open(proc_path, "r", encoding="utf-8", errors="ignore") as _pf:
                    _proc_content = _pf.read()
                # Tightened: must be non-trivial - >200 bytes AND contains 'def' and 'import'
                check("wc_shipping_processor.py is non-trivial (>200 bytes + def + import)",
                      len(_proc_content) >= 200 and "def" in _proc_content and "import" in _proc_content,
                      f"size={len(_proc_content)}, has def={'def' in _proc_content}, has import={'import' in _proc_content}")
            except Exception as _e:
                check("wc_shipping_processor.py readable", False, str(_e))


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
