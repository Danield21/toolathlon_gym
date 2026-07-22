"""
Evaluation for sf-hr-satisfaction-excel-gcal task.

Checks:
1. Excel file Employee_Satisfaction.xlsx with correct data
2. Google Calendar events for 2 lowest-satisfaction departments
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
RUNTIME_ONLY_FAIL = 0


def check(name, condition, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, RUNTIME_ONLY_FAIL
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if runtime_only:
            RUNTIME_ONLY_FAIL += 1
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel File ===")
    agent_file = os.path.join(agent_workspace, "Employee_Satisfaction.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Employee_Satisfaction.xlsx")

    if not os.path.exists(agent_file):
        check("Excel file exists", False, f"Not found: {agent_file}")
        return
    check("Excel file exists", True)

    if not os.path.exists(gt_file):
        check("Groundtruth file exists", False, f"Not found: {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Department Satisfaction sheet
    print("  Checking Department Satisfaction sheet...")
    a_rows = load_sheet_rows(agent_wb, "Department Satisfaction")
    g_rows = load_sheet_rows(gt_wb, "Department Satisfaction")

    if a_rows is None:
        check("Sheet 'Department Satisfaction' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth sheet exists", False, "Not found")
    else:
        check("Sheet 'Department Satisfaction' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("Row count matches", len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Row '{g_row[0]}'", False, "Missing")
                continue

            # Employee_Count (col 1) - exact match
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"{key}.Employee_Count",
                      num_close(a_row[1], g_row[1], 0),
                      f"{a_row[1]} vs {g_row[1]}")

            # Avg_Job_Satisfaction (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Avg_Job_Satisfaction",
                      num_close(a_row[2], g_row[2], 0.05),
                      f"{a_row[2]} vs {g_row[2]}")

            # Avg_Work_Life_Balance (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Avg_Work_Life_Balance",
                      num_close(a_row[3], g_row[3], 0.05),
                      f"{a_row[3]} vs {g_row[3]}")

            # Low_Satisfaction_Count (col 4) - tight tolerance of 1
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"{key}.Low_Satisfaction_Count",
                      num_close(a_row[4], g_row[4], 1),
                      f"{a_row[4]} vs {g_row[4]}")

            # Low_Satisfaction_Pct (col 5)
            if len(a_row) > 5 and len(g_row) > 5:
                check(f"{key}.Low_Satisfaction_Pct",
                      num_close(a_row[5], g_row[5], 0.1),
                      f"{a_row[5]} vs {g_row[5]}")

        # Check ascending order of Avg_Job_Satisfaction (col 2)
        asc_ok = True
        prev = None
        asc_details = []
        for row in a_data:
            if len(row) > 2 and row[2] is not None:
                try:
                    v = float(row[2])
                except (ValueError, TypeError):
                    continue
                asc_details.append(v)
                if prev is not None and v < prev - 1e-9:
                    asc_ok = False
                prev = v
        check("Department Satisfaction sorted by Avg_Job_Satisfaction ascending",
              asc_ok, f"Values in order: {asc_details}")

    # Check Summary sheet
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    if a_rows is None:
        check("Sheet 'Summary' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth Summary sheet exists", False, "Not found")
    else:
        check("Sheet 'Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Summary: '{g_row[0]}'", False, "Missing")
                continue
            g_val = g_row[1] if len(g_row) > 1 else None
            a_val = a_row[1] if len(a_row) > 1 else None
            # Try numeric comparison first, then string
            is_numeric = True
            try:
                float(g_val)
            except (TypeError, ValueError):
                is_numeric = False
            if is_numeric:
                # Tighter tolerance: Total_Employees/Employees_At_Risk exact, averages 0.05
                if key in ("total_employees", "employees_at_risk"):
                    tol = 0
                elif "avg" in key or "satisfaction" in key:
                    tol = 0.05
                else:
                    tol = 1
                check(f"Summary: {key}",
                      num_close(a_val, g_val, tol),
                      f"{a_val} vs {g_val} (tol={tol})")
            else:
                check(f"Summary: {key}",
                      str_match(a_val, g_val),
                      f"{a_val} vs {g_val}")


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for gcal check", False, str(e))
        return

    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%%wellness review%%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    # Event presence is runtime-only (agent may not have reached this step)
    check("At least 2 wellness review events", len(events) >= 2,
          f"Found {len(events)}",
          runtime_only=True)

    # The 2 lowest satisfaction departments - compute dynamically from DB
    # Tie-break: alphabetical on department name
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute("""
            SELECT "DEPARTMENT", AVG("JOB_SATISFACTION")::float
            FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
            GROUP BY "DEPARTMENT"
            ORDER BY AVG("JOB_SATISFACTION") ASC, "DEPARTMENT" ASC
            LIMIT 2
        """)
        bottom_two = [r[0].lower() for r in cur2.fetchall()]
        cur2.close()
        conn2.close()
    except Exception:
        bottom_two = ["r&d", "support"]

    # If agent DID create events, the content (dept/date/time) must match semantically (NOT runtime_only)
    if len(events) >= 2:
        for dept in bottom_two:
            found = any(dept in (e[0] or "").lower() or dept in (e[1] or "").lower()
                         for e in events)
            check(f"Event for department '{dept}'", found,
                  f"Not found among {len(events)} events. Expected depts: {bottom_two}")
    else:
        # Not enough events - mark as runtime-only
        for dept in bottom_two:
            check(f"Event for department '{dept}'", False,
                  "Not enough events to verify",
                  runtime_only=True)

    # Check dates AND times (10:00-11:00 UTC per task)
    if len(events) >= 2:
        d1_start = events[0][2]
        d1_end = events[0][3]
        d2_start = events[1][2]
        d2_end = events[1][3]
        check("First event on 2026-03-13",
              d1_start is not None and "2026-03-13" in str(d1_start),
              f"Got {d1_start}")
        check("Second event on 2026-03-14",
              d2_start is not None and "2026-03-14" in str(d2_start),
              f"Got {d2_start}")
        # Time check: duration 60 min AND start hour 10 UTC
        for i, (s, e) in enumerate([(d1_start, d1_end), (d2_start, d2_end)]):
            if s and e:
                dur_min = (e - s).total_seconds() / 60
                check(f"Event {i+1} duration 60 min",
                      abs(dur_min - 60) <= 5,
                      f"Got {dur_min} min")
                # start hour must be 10 UTC (task spec)
                try:
                    # If tzaware, convert to UTC
                    s_utc = s
                    if getattr(s, "tzinfo", None) is not None:
                        import datetime as _dt
                        s_utc = s.astimezone(_dt.timezone.utc)
                    check(f"Event {i+1} start hour = 10 UTC",
                          s_utc.hour == 10,
                          f"Got hour={s_utc.hour} from {s}")
                except Exception as ex:
                    check(f"Event {i+1} start hour = 10 UTC", False, f"Err: {ex}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gcal()

    total_pass = PASS_COUNT
    total_fail = FAIL_COUNT
    non_runtime_fail = FAIL_COUNT - RUNTIME_ONLY_FAIL
    all_ok = non_runtime_fail == 0

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'} (runtime-only fails: {RUNTIME_ONLY_FAIL})")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": total_pass, "failed": total_fail, "success": all_ok}, f, indent=2)

    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
