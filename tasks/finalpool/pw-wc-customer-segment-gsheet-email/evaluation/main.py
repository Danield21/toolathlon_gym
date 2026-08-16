"""Evaluation script for pw-wc-customer-segment-gsheet-email."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-wc-customer-segment-gsheet-email"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Detect GT self-test
    is_gt_self_test = False
    try:
        if groundtruth_workspace and os.path.realpath(groundtruth_workspace) == os.path.realpath(agent_workspace):
            is_gt_self_test = True
    except Exception:
        pass

    excel_path = os.path.join(agent_workspace, "Customer_Segment_Report.xlsx")
    check("Customer_Segment_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Customer_Segment_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Category', 'Product_Count', 'Our_Avg_Price', 'Total_Sales', 'Market_Avg_Price', 'Price_Gap_Pct'], 6)
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 3)
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action', 'Category'], 2)
        try:
            conn = get_conn()
            cur = conn.cursor()
            # Per task.md: subject is "Analysis Report Complete" and recipient
            # is team-lead@company.com. Tightened so the noise newsletter
            # ("Weekly Newsletter" from newsletter@company.com) cannot satisfy
            # this check by accident.
            cur.execute(
                "SELECT subject, to_addr FROM email.messages "
                "WHERE subject ILIKE %s AND to_addr::text ILIKE %s",
                ('%Analysis Report Complete%', '%team-lead@company.com%'))
            emails = cur.fetchall()
            if is_gt_self_test and len(emails) == 0:
                check(
                    "Analysis email sent to team-lead with required subject (GT self-test toleration)",
                    True,
                    "GT self-test: email is agent runtime artifact, skipped",
                )
            else:
                check(
                    "Analysis email sent to team-lead with required subject",
                    len(emails) >= 1,
                    f"found {len(emails)} emails with subject 'Analysis Report Complete' to team-lead",
                )
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        try:
            conn = get_conn()
            cur = conn.cursor()
            # Per task.md: Google Sheet must be named exactly "Wc Segment Tracker".
            cur.execute(
                "SELECT title FROM gsheet.spreadsheets WHERE title ILIKE %s",
                ('%Wc Segment Tracker%',))
            sheets = cur.fetchall()
            if is_gt_self_test and len(sheets) == 0:
                check(
                    "Google Sheet 'Wc Segment Tracker' created (GT self-test toleration)",
                    True,
                    "GT self-test: gsheet is agent runtime artifact, skipped",
                )
            else:
                check(
                    "Google Sheet 'Wc Segment Tracker' created",
                    len(sheets) >= 1,
                    f"found {len(sheets)} matching sheets",
                )
            conn.close()
        except Exception as e:
            check("GSheet check", False, str(e))

        proc_path = os.path.join(agent_workspace, "wc_segment_processor.py")
        if is_gt_self_test and not os.path.exists(proc_path):
            check("wc_segment_processor.py exists (GT self-test toleration)",
                  True, "GT self-test: agent-generated script, skipped")
        else:
            check("wc_segment_processor.py exists", os.path.exists(proc_path))
            if os.path.exists(proc_path):
                try:
                    with open(proc_path, "r", encoding="utf-8", errors="ignore") as _pf:
                        _src = _pf.read()
                    check(
                        "wc_segment_processor.py has Python import statements",
                        ('import ' in _src) or ('from ' in _src and 'import' in _src),
                        "no 'import' statement found in processor script",
                    )
                except OSError as e:
                    check("wc_segment_processor.py read", False, str(e))

        # Value-level: at least one numeric Price_Gap_Pct value in plausible
        # range (-100..100) so the agent cannot pad with garbage.
        if "Data_Analysis" in wb.sheetnames:
            try:
                ws = wb["Data_Analysis"]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                gap_col = None
                for i, h in enumerate(headers):
                    if 'gap' in h and ('pct' in h or 'percent' in h or '%' in h):
                        gap_col = i
                        break
                gap_ok = False
                if gap_col is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if gap_col < len(row) and row[gap_col] is not None:
                            fv = safe_float(row[gap_col])
                            if fv is not None and -100 <= fv <= 100:
                                gap_ok = True
                                break
                check("Data_Analysis Price_Gap_Pct has plausible numeric value",
                      gap_ok,
                      "no numeric Price_Gap_Pct value in [-100,100] range")
            except Exception:
                pass


    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
