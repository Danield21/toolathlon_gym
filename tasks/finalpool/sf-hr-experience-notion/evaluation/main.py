"""Evaluation for sf-hr-experience-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")), "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected tenure data from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               ROUND(AVG("YEARS_EXPERIENCE")::numeric, 2) as avg_tenure,
               MIN("YEARS_EXPERIENCE") as min_tenure,
               MAX("YEARS_EXPERIENCE") as max_tenure,
               COUNT(*) as emp_count
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    rows = cur.fetchall()

    cur.execute("""
        SELECT ROUND(AVG("YEARS_EXPERIENCE")::numeric, 2),
               COUNT(*)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
    """)
    overall = cur.fetchone()
    cur.close()
    conn.close()
    return rows, overall


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Tenure_Analysis.xlsx."""
    print("\n=== Checking Tenure_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Tenure_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True
    dept_data, overall = get_expected_data()

    # Check Department Tenure sheet
    dept_sheet = None
    for name in wb.sheetnames:
        if "department" in name.lower() and "tenure" in name.lower():
            dept_sheet = wb[name]
            break
    if dept_sheet is None:
        record("Sheet 'Department Tenure' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Department Tenure' exists", True)
        rows = list(dept_sheet.iter_rows(min_row=2, values_only=True))
        record("Department Tenure has 7 rows", len(rows) == 7, f"Got {len(rows)}")

        agent_lookup = {}
        for r in rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for exp in dept_data:
            dept = exp[0]
            a_row = agent_lookup.get(dept.lower())
            if a_row is None:
                record(f"Department '{dept}' present", False, "Missing")
                all_ok = False
                continue

            # Avg_Tenure: tighter ±0.05 (rounded to 2 decimals)
            ok_avg = num_close(a_row[1], exp[1], 0.05)
            record(f"'{dept}' Avg_Tenure", ok_avg,
                   f"Expected {exp[1]}, got {a_row[1]}")
            if not ok_avg:
                all_ok = False

            # Min_Tenure (col 2)
            ok_min = num_close(a_row[2], exp[2], 0)
            record(f"'{dept}' Min_Tenure", ok_min,
                   f"Expected {exp[2]}, got {a_row[2]}")
            if not ok_min:
                all_ok = False

            # Max_Tenure (col 3)
            ok_max = num_close(a_row[3], exp[3], 0)
            record(f"'{dept}' Max_Tenure", ok_max,
                   f"Expected {exp[3]}, got {a_row[3]}")
            if not ok_max:
                all_ok = False

            # Employee_Count: exact integer
            ok_count = num_close(a_row[4], exp[4], 0)
            record(f"'{dept}' Employee_Count", ok_count,
                   f"Expected {exp[4]}, got {a_row[4]}")
            if not ok_count:
                all_ok = False

    # Check Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        summary = {}
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

        # Determine expected highest_avg_tenure_dept
        sorted_by_avg = sorted(dept_data, key=lambda x: (-float(x[1]), x[0]))
        expected_highest_dept = sorted_by_avg[0][0]
        expected_overall_avg = float(overall[0])

        # Total_Employees: exact
        total_emp_key = next((k for k in summary if "total" in k and "employee" in k), None)
        if total_emp_key:
            ok = num_close(summary[total_emp_key], overall[1], 0)
            record("Summary Total_Employees", ok,
                   f"Expected {overall[1]}, got {summary[total_emp_key]}")
            if not ok: all_ok = False
        else:
            record("Summary Total_Employees row present", False, "missing key")
            all_ok = False

        # Department_Count: exact 7
        dc_key = next((k for k in summary if "department" in k and "count" in k), None)
        if dc_key:
            ok = num_close(summary[dc_key], 7, 0)
            record("Summary Department_Count", ok,
                   f"Expected 7, got {summary[dc_key]}")
            if not ok: all_ok = False
        else:
            record("Summary Department_Count row present", False, "missing key")
            all_ok = False

        # Highest_Avg_Tenure_Dept: string match against expected
        h_key = next((k for k in summary if "highest" in k and "tenure" in k and "dept" in k), None)
        if h_key:
            actual = str(summary[h_key]).strip().lower()
            ok = actual == expected_highest_dept.strip().lower()
            record("Summary Highest_Avg_Tenure_Dept", ok,
                   f"Expected '{expected_highest_dept}', got '{summary[h_key]}'")
            if not ok: all_ok = False
        else:
            record("Summary Highest_Avg_Tenure_Dept row present", False, "missing key")
            all_ok = False

        # Overall_Avg_Tenure: numeric ±0.05
        oa_key = next((k for k in summary if "overall" in k and ("tenure" in k or "avg" in k)), None)
        if oa_key:
            ok = num_close(summary[oa_key], expected_overall_avg, 0.05)
            record("Summary Overall_Avg_Tenure", ok,
                   f"Expected {expected_overall_avg}, got {summary[oa_key]}")
            if not ok: all_ok = False
        else:
            record("Summary Overall_Avg_Tenure row present", False, "missing key")
            all_ok = False

    return all_ok


def check_notion():
    """Check Notion page 'Employee Tenure Dashboard' has all 7 department names in body."""
    print("\n=== Checking Notion Page ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, properties FROM notion.pages WHERE archived = false")
    pages = cur.fetchall()

    target_page_id = None
    for page in pages:
        props = page[1] if isinstance(page[1], dict) else (json.loads(page[1]) if page[1] else {})
        title_text = ""
        if "title" in props:
            t = props["title"]
            if isinstance(t, dict) and "title" in t:
                for item in t["title"]:
                    if isinstance(item, dict):
                        title_text += item.get("plain_text", item.get("text", {}).get("content", ""))
        # Match required title (exact or close)
        if title_text.strip().lower() in ("employee tenure dashboard",) or \
           ("employee tenure" in title_text.lower() and "dashboard" in title_text.lower()):
            target_page_id = page[0]
            break

    found_page = target_page_id is not None
    record("Notion page 'Employee Tenure Dashboard' exists", found_page,
           "No matching page title")
    if not found_page:
        cur.close(); conn.close()
        return False

    # Verify body content includes all 7 department names
    try:
        dept_data, _ = get_expected_data()
    except Exception as e:
        record("Notion content check (DB access)", False, str(e))
        cur.close(); conn.close()
        return False
    expected_depts = [d[0] for d in dept_data]

    cur.execute("SELECT block_data FROM notion.blocks WHERE parent_id = %s", (target_page_id,))
    blocks = cur.fetchall()
    body_text = ""
    for (bd,) in blocks:
        try:
            body_text += " " + json.dumps(bd if isinstance(bd, (dict, list)) else bd, default=str).lower()
        except Exception:
            body_text += " " + str(bd).lower()

    missing = [d for d in expected_depts if d.lower() not in body_text]
    ok = len(missing) == 0
    record("Notion page body mentions all 7 department names",
           ok, f"missing: {missing}")
    cur.close(); conn.close()
    return found_page and ok


def check_email():
    """Check exact email subject and recipient + body content."""
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, COALESCE(body_text, body_html, '')
        FROM email.messages
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()

    target_subj = "employee tenure analysis report"
    target_to = "chro@company.com"
    target_from = "hr-analytics@company.com"
    matched = None
    for subj, frm, to_addr, body in rows:
        subj_l = (subj or "").strip().lower()
        to_str = str(to_addr or "").lower()
        if target_subj in subj_l and target_to in to_str:
            matched = (subj, frm, to_addr, body)
            break
    record(f"Email '{target_subj}' to {target_to} exists",
           matched is not None, f"checked {len(rows)} emails")
    if not matched:
        return False
    body = (matched[3] or "").lower()
    # Body should mention total employees, department count, highest tenure dept
    must = ["employee", "department"]
    missing = [m for m in must if m not in body]
    record("Email body mentions employee + department",
           len(missing) == 0, f"missing: {missing}")
    # Numeric content (digits) for total
    has_digits = any(ch.isdigit() for ch in body)
    record("Email body contains numeric figure", has_digits, "no digits found")
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    notion_ok = check_notion()
    email_ok = check_email()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    overall = FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
