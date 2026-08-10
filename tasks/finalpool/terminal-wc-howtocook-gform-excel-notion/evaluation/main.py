"""Evaluation for terminal-wc-howtocook-gform-excel-notion.
Checks:
1. Meal_Kit_Analysis.xlsx with 4 sheets (Appliance_Catalog, Recipe_Matches, Survey_Results, Product_Roadmap)
2. Google Form "Meal Kit Interest Survey" with 6 questions
3. Notion "Meal Kit Development Tracker" with 5 kit pages
4. appliance_recipe_matcher.py script exists
5. appliance_recipe_matches.json exists
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


import re as _re

_AMOUNT_RE = _re.compile(r"[-+]?\d*\.?\d+")


def _to_float(value):
    """Robustly parse a numeric cell into float.

    Handles int/float, strings with thousands separators, currency symbols
    ($, ¥, ￥, €), percent signs, and surrounding whitespace. Returns None when
    the value cannot be parsed as a number (never raises).
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    m = _AMOUNT_RE.search(s.replace(",", "").replace("%", ""))
    if not m:
        return None
    try:
        return float(m.group())
    except ValueError:
        return None


def num_close(a, b, tol=2.0):
    """Numeric close comparison.

    Both sides parseable as numbers -> |a-b| <= tol. If either side is not
    parseable, fall back to case-insensitive string equality (never a hard
    False that would wrongly fail an otherwise-correct value).
    """
    fa, fb = _to_float(a), _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    return str(a).strip().lower() == str(b).strip().lower()


def _get_sheet_by_name(wb, *candidates):
    """Return the first worksheet whose lowercase/normalized name matches
    one of the given candidates (case-insensitive, treating space=underscore)."""
    norms = []
    for c in candidates:
        norms.append(c.lower().replace(" ", "_"))
    for s in wb.sheetnames:
        n = s.lower().replace(" ", "_")
        if n in norms:
            return wb[s]
    # fallback: substring match
    for s in wb.sheetnames:
        n = s.lower().replace(" ", "_")
        for c in norms:
            if c in n:
                return wb[s]
    return None


def check_excel(workspace):
    print("\n=== Check 1: Meal_Kit_Analysis.xlsx ===")
    path = os.path.join(workspace, "Meal_Kit_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    # data_only=False: read formulas as formulas; numeric extraction is done
    # via _to_float so cells with currency symbols/separators/percent still
    # parse, and cached formula values (None) never crash the numeric checks.
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Appliance_Catalog (exact name lookup)
    ws1 = _get_sheet_by_name(wb, "Appliance_Catalog", "Appliance Catalog")
    if ws1 is None:
        check("Sheet 'Appliance_Catalog' exists", False, f"sheets: {sheets}")
        ws1 = wb[sheets[0]]
    else:
        check("Sheet 'Appliance_Catalog' exists", True)
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    # Per task.md: "There should be 8 products in this category" — enforce exactly 8.
    check("Appliance_Catalog has exactly 8 products",
          len(data1) == 8, f"Found {len(data1)}")

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has price column", any("price" in h for h in headers), f"Headers: {rows1[0]}")
        check("Has avg_rating column", any("rating" in h for h in headers), f"Headers: {rows1[0]}")
        check("Has compatible_recipes_count column",
              any("recipe" in h and "count" in h for h in headers) or any("compatible" in h for h in headers),
              f"Headers: {rows1[0]}")

    # Check known appliance prices
    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Contains Blender product", "blender" in all_text1, f"Text snippet: {all_text1[:150]}")
    check("Contains Vacuum Sealer product", "vacuum" in all_text1 or "sealing" in all_text1,
          f"Text snippet: {all_text1[:150]}")
    check("Contains Cooking Pot product", "cooking pot" in all_text1 or "cooker" in all_text1 or "steamer" in all_text1,
          f"Text snippet: {all_text1[:150]}")

    # Verify a known price
    price_col = next((i for i, h in enumerate(headers) if "price" in h and "regular" not in h), -1) if rows1 else -1
    if price_col >= 0:
        prices = [p for r in data1 if (p := _to_float(r[price_col])) is not None]
        # Blender should be 214.00
        has_blender_price = any(num_close(p, 214.0, 1.0) for p in prices)
        check("Blender price ~214.00", has_blender_price, f"Prices: {prices}")

    # Recipe_Matches (exact name lookup)
    ws2 = _get_sheet_by_name(wb, "Recipe_Matches", "Recipe Matches")
    if ws2 is None:
        check("Sheet 'Recipe_Matches' exists", False, f"sheets: {sheets}")
    else:
        check("Sheet 'Recipe_Matches' exists", True)
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Recipe_Matches has 10+ pairings", len(data2) >= 10, f"Found {len(data2)}")
        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has matched_appliance column",
                  any("appliance" in h or "matched" in h for h in headers2),
                  f"Headers: {rows2[0]}")
            check("Has difficulty column", any("difficult" in h for h in headers2),
                  f"Headers: {rows2[0]}")

    # Survey_Results (exact name lookup + per-question top-answer validation)
    ws3 = _get_sheet_by_name(wb, "Survey_Results", "Survey Results")
    if ws3 is None:
        check("Sheet 'Survey_Results' exists", False, f"sheets: {sheets}")
    else:
        check("Sheet 'Survey_Results' exists", True)
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Survey_Results has exactly 6 rows", len(data3) == 6, f"Found {len(data3)}")
        if rows3:
            headers3 = [str(c).lower() if c else "" for c in rows3[0]]
            # Tighter: require both 'top' AND 'answer' tokens (or exact 'top_answer'),
            # not just 'answer' which would match 'answer_count'.
            check("Has top_answer column",
                  any(("top" in h and "answer" in h) for h in headers3),
                  f"Headers: {rows3[0]}")
            check("Has response_count column",
                  any(("count" in h or h == "response_count" or h == "responses") for h in headers3),
                  f"Headers: {rows3[0]}")
        # Per-question top-answer validation. The preprocess injects 25
        # responses with these top answers (see task.md / preprocess fixture).
        # Counts match the preprocess fixture as of 2026-08:
        # Q1 "How often do you cook" -> "Several times a week" (14)
        # Q2 "What cuisine types"    -> "Chinese" (16)
        # Q3 "Which kitchen appliances" -> "Blender" (12)
        # Q4 "What is your monthly budget" -> "30 to 60 dollars" (15)
        # Q5 likelihood 1-5          -> "4" (12)
        # Q6 features (text)         -> any free text answer
        # q_kw: substring required in the question cell; ans: expected top
        # answer; must: (optional) keyword that further pins down the question;
        # exclude: (optional) keywords that rule out a *different* question that
        # also happens to contain q_kw. The evaluator scans ALL rows for a
        # keyword (order-independent), because task.md only requires "one row
        # per question" without prescribing an order — a correct agent may emit
        # the rows in any order (e.g. alphabetical, or by response-key order).
        expected_top = [
            ("how often", "several times a week", None, []),
            ("cuisine", "chinese", None, []),
            # 'appliance' occurs in BOTH Q3 ("Which kitchen appliances do you
            # currently own?") and Q5 ("How likely are you to purchase a meal
            # kit bundled with an appliance?"). Pin Q3 via 'own' and rule out
            # Q5's distinctive wording so the blender check can never read Q5's
            # numeric top-answer.
            ("appliance", "blender", "own", ["likely", "purchase", "buy", "bundled", "bundle"]),
            ("budget", "30 to 60 dollars", None, []),
            ("likely", "4", None, []),
        ]
        rows3_data = [
            [str(c).strip().lower() if c is not None else "" for c in r]
            for r in data3
        ]
        for q_kw, expected_ans, must_kw, excl_kws in expected_top:
            candidates = [r for r in rows3_data
                          if r and len(r) > 0 and q_kw in r[0]]
            pinned = [r for r in candidates
                      if must_kw is None or must_kw in r[0]]
            non_excluded = [r for r in pinned
                            if not any(x in r[0] for x in excl_kws)]
            selected = (non_excluded or pinned or candidates or [None])[0]
            matched = False
            if selected is not None:
                top_answer_cell = selected[1] if len(selected) > 1 else ""
                # Word-boundary match for expected_ans to avoid sub-word
                # collisions (e.g., 'blender' in 'meal blender starter').
                pat = r"\b" + _re.escape(expected_ans) + r"\b"
                matched = bool(_re.search(pat, top_answer_cell))
            check(f"Survey top answer for '{q_kw}' is '{expected_ans}'",
                  matched,
                  f"row data: {rows3_data}")

    # Product_Roadmap (exact name + formula validation)
    ws4 = _get_sheet_by_name(wb, "Product_Roadmap", "Product Roadmap")
    if ws4 is None:
        check("Sheet 'Product_Roadmap' exists", False, f"sheets: {sheets}")
    else:
        check("Sheet 'Product_Roadmap' exists", True)
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        # Per task.md: 5 meal kit concepts, one per major appliance type
        check("Product_Roadmap has exactly 5 kits", len(data4) == 5, f"Found {len(data4)}")
        if rows4:
            headers4 = [str(c).lower() if c else "" for c in rows4[0]]
            check("Has priority column", any("priority" in h for h in headers4),
                  f"Headers: {rows4[0]}")
            check("Has estimated_price column", any("price" in h or "estimated" in h for h in headers4),
                  f"Headers: {rows4[0]}")
            # Formula validation: estimated_price = appliance_price + 15
            # priority = "High" if recipe_count > 3 else "Medium"
            # Locate column indexes
            def _find_col(*kws):
                for i, h in enumerate(headers4):
                    if all(k in h for k in kws):
                        return i
                return None
            recipe_col = _find_col("recipe", "count")
            price_col = _find_col("estimated", "price")
            if price_col is None:
                price_col = _find_col("price")
            priority_col = _find_col("priority")
            appliance_col = _find_col("appliance")
            # Look up appliance prices from Appliance_Catalog by partial name
            appliance_prices = {}
            if data1:
                ap_headers = [str(c).lower() if c else "" for c in rows1[0]]
                ap_name_col = next((i for i, h in enumerate(ap_headers) if "name" in h), 0)
                ap_price_col = next(
                    (i for i, h in enumerate(ap_headers)
                     if "price" in h and "regular" not in h),
                    1,
                )
                for r in data1:
                    if r and r[ap_name_col] is not None:
                        ap_price_val = _to_float(r[ap_price_col])
                        if ap_price_val is not None:
                            appliance_prices[str(r[ap_name_col]).strip().lower()] = ap_price_val
            for r in data4:
                if not r or len(r) <= max(filter(None, [recipe_col, price_col, priority_col, appliance_col]) or [0]):
                    continue
                appliance_name = (str(r[appliance_col]).strip().lower()
                                  if appliance_col is not None and r[appliance_col]
                                  else "")
                # Find best matching appliance price by token overlap or
                # substring. Use longest common token sequence as a heuristic.
                ap_price = None
                if appliance_name:
                    appliance_tokens = set(appliance_name.split())
                    best_overlap = 0
                    for ap_name, p in appliance_prices.items():
                        ap_tokens = set(ap_name.split())
                        overlap = len(appliance_tokens & ap_tokens)
                        if overlap > best_overlap:
                            best_overlap = overlap
                            ap_price = p
                    # Fall back to substring if token approach failed.
                    if ap_price is None:
                        for ap_name, p in appliance_prices.items():
                            if appliance_name in ap_name or ap_name in appliance_name:
                                ap_price = p
                                break
                # Validate estimated_price = ap_price + 15
                if ap_price is not None and price_col is not None:
                    ep = _to_float(r[price_col])
                    if ep is not None:
                        check(f"Roadmap '{appliance_name[:30]}' estimated_price = price+15",
                              num_close(ep, ap_price + 15.0, 1.0),
                              f"Expected {ap_price + 15.0}, got {ep}")
                # Validate priority logic
                if recipe_col is not None and priority_col is not None:
                    rc = _to_float(r[recipe_col])
                    pri = (str(r[priority_col]).strip().lower()
                           if r[priority_col] is not None else "")
                    if rc is not None:
                        expected_priority = "high" if rc > 3 else "medium"
                        check(f"Roadmap '{appliance_name[:30]}' priority='{expected_priority}' (rc={rc})",
                              pri == expected_priority,
                              f"Expected '{expected_priority}', got '{pri}'")


def check_gform():
    print("\n=== Check 2: Google Form Survey ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("Meal Kit Interest Survey form exists", False, f"DB connect error: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        meal_form = None
        for form_id, title in forms:
            if title and ("meal kit" in title.lower() or "meal_kit" in title.lower()):
                meal_form = (form_id, title)
                break
        check("Meal Kit Interest Survey form exists", meal_form is not None,
              f"Forms: {[f[1] for f in forms]}")

        if meal_form:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (meal_form[0],))
            q_count = cur.fetchone()[0]
            check("Survey has exactly 6 questions", q_count == 6, f"Found {q_count}")

            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position",
                        (meal_form[0],))
            questions = cur.fetchall()
            q_text = " ".join(str(q[0]) for q in questions).lower()
            q_types = [str(t[1] or "").upper() for t in questions]
            check("Has cooking frequency question", "cook" in q_text or "often" in q_text,
                  f"Questions: {q_text[:150]}")
            check("Has cuisine question", "cuisine" in q_text, f"Questions: {q_text[:150]}")
            check("Has appliance ownership question", "appliance" in q_text or "own" in q_text,
                  f"Questions: {q_text[:150]}")
            check("Has budget question", "budget" in q_text, f"Questions: {q_text[:150]}")
            # Per task.md: also need a likelihood question and a feature/text question
            check("Has likelihood/purchase scale question",
                  "likely" in q_text or "purchase" in q_text or "buy" in q_text,
                  f"Questions: {q_text[:200]}")
            check("Has features/appealing question",
                  "feature" in q_text or "appealing" in q_text,
                  f"Questions: {q_text[:200]}")
            # Question types: the google-forms MCP can only create
            # 'textQuestion' and 'choiceQuestion' (RADIO single-select). The
            # preprocess fixture uses the same vocabulary. Accept both the MCP
            # vocabulary and the legacy type names, and never require a type
            # the MCP cannot produce (dropdown/linear scale/paragraph/checkbox).
            def _is_choice_type(t):
                t = str(t or "").upper()
                return "CHOICE" in t or t in ("RADIO", "CHECKBOX", "MULTIPLE_CHOICE")

            def _is_text_type(t):
                t = str(t or "").upper()
                return "TEXT" in t or t in ("PARAGRAPH", "SHORT_ANSWER")

            check("Has multiple-choice question type",
                  any(_is_choice_type(t) for t in q_types),
                  f"Types: {q_types}")
            check("Has text/short-answer question type",
                  any(_is_text_type(t) for t in q_types),
                  f"Types: {q_types}")
    except Exception as e:
        check("Gform check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_notion():
    print("\n=== Check 3: Notion Meal Kit Development Tracker ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("Meal Kit Development Tracker database exists", False, f"DB connect error: {e}")
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title, properties FROM notion.databases")
        dbs = cur.fetchall()
        tracker_db = None
        for db_id, title, props in dbs:
            title_str = ""
            if isinstance(title, list):
                title_str = " ".join(
                    item.get("text", {}).get("content", "") for item in title if isinstance(item, dict))
            elif isinstance(title, str):
                try:
                    parsed = json.loads(title)
                    if isinstance(parsed, list):
                        title_str = " ".join(
                            item.get("text", {}).get("content", "") for item in parsed if isinstance(item, dict))
                    else:
                        title_str = str(title)
                except Exception:
                    title_str = str(title)
            else:
                title_str = str(title) if title else ""
            if "meal kit" in title_str.lower() and ("tracker" in title_str.lower() or "development" in title_str.lower()):
                tracker_db = (db_id, title_str, props)
                break
        check("Meal Kit Development Tracker database exists", tracker_db is not None,
              f"Databases: {[d[1] for d in dbs]}")

        if tracker_db:
            # Check properties
            props = tracker_db[2]
            if isinstance(props, str):
                try:
                    props = json.loads(props)
                except:
                    props = {}
            props_lower = {k.lower(): v for k, v in (props or {}).items()}
            check("Has Status property",
                  any("status" in k for k in props_lower),
                  f"Properties: {list(props_lower.keys())}")
            check("Has Priority property",
                  any("priority" in k for k in props_lower),
                  f"Properties: {list(props_lower.keys())}")
            check("Has Estimated_Revenue property",
                  any("revenue" in k or "estimated" in k for k in props_lower),
                  f"Properties: {list(props_lower.keys())}")

            # Check pages
            cur.execute("""
                SELECT COUNT(*) FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (tracker_db[0],))
            count = cur.fetchone()[0]
            check("Tracker has exactly 5 kit pages", count == 5, f"Found {count}")

            # Check page content — every page must have status=Planning, a priority value,
            # and a numeric Estimated_Revenue.
            cur.execute("""
                SELECT properties FROM notion.pages
                WHERE parent->>'database_id' = %s
            """, (tracker_db[0],))
            pages = cur.fetchall()
            all_page_text = ""
            pages_with_planning = 0
            pages_with_priority = 0
            pages_with_revenue = 0
            for (page_props,) in pages:
                if isinstance(page_props, str):
                    try:
                        page_props = json.loads(page_props)
                    except:
                        page_props = {}
                if not isinstance(page_props, dict):
                    page_props = {}
                page_str = json.dumps(page_props).lower()
                all_page_text += page_str + " "

                # Walk properties for status/priority/revenue
                page_status = ""
                page_priority = ""
                page_revenue = None
                for k, v in page_props.items():
                    k_lower = k.lower()
                    v_str = json.dumps(v).lower() if v is not None else ""
                    # status: select.name=="Planning"
                    if "status" in k_lower:
                        if isinstance(v, dict):
                            sel = v.get("select") or {}
                            if isinstance(sel, dict):
                                name = (sel.get("name") or "").strip().lower()
                                if name:
                                    page_status = name
                        if not page_status and "planning" in v_str:
                            page_status = "planning"
                    if "priority" in k_lower:
                        if isinstance(v, dict):
                            sel = v.get("select") or {}
                            if isinstance(sel, dict):
                                name = (sel.get("name") or "").strip().lower()
                                if name:
                                    page_priority = name
                        if not page_priority:
                            for keyword in ("high", "medium", "low"):
                                if f'"{keyword}"' in v_str:
                                    page_priority = keyword
                                    break
                    if "revenue" in k_lower or "estimated" in k_lower:
                        if isinstance(v, dict):
                            num = v.get("number")
                            if isinstance(num, (int, float)):
                                page_revenue = num
                if page_status == "planning":
                    pages_with_planning += 1
                if page_priority in ("high", "medium", "low"):
                    pages_with_priority += 1
                if isinstance(page_revenue, (int, float)) and page_revenue > 0:
                    pages_with_revenue += 1

            check("All 5 pages have Planning status",
                  pages_with_planning == 5,
                  f"Got {pages_with_planning}/5")
            check("All 5 pages have a Priority value (High/Medium/Low)",
                  pages_with_priority == 5,
                  f"Got {pages_with_priority}/5")
            check("All 5 pages have a positive Estimated_Revenue number",
                  pages_with_revenue == 5,
                  f"Got {pages_with_revenue}/5")
    except Exception as e:
        check("Notion check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Scripts ===")
    matcher = os.path.join(workspace, "appliance_recipe_matcher.py")
    check("appliance_recipe_matcher.py exists", os.path.exists(matcher))

    matches_json = os.path.join(workspace, "appliance_recipe_matches.json")
    check("appliance_recipe_matches.json exists", os.path.exists(matches_json))
    if os.path.exists(matches_json):
        with open(matches_json) as f:
            data = json.load(f)
        check("JSON is a dict mapping appliances to recipe lists",
              isinstance(data, dict),
              f"Got {type(data).__name__}")
        if isinstance(data, dict):
            keys_lower = " ".join(str(k).lower() for k in data.keys())
            # Per task.md: blender, vacuum sealer, cooking pot/cooker/steamer, kitchen scale, exhaust fan
            # Require at least 4 of the 5 required appliance categories
            required_categories = {
                "blender": ["blender"],
                "vacuum_sealer": ["vacuum", "sealer"],
                "cooking_pot": ["cooking pot", "cooker", "steamer", "pot"],
                "kitchen_scale": ["kitchen scale", "scale"],
                "exhaust_fan": ["exhaust", "fan"],
            }
            categories_found = 0
            for cat, kws in required_categories.items():
                if any(kw in keys_lower for kw in kws):
                    categories_found += 1
            check("JSON has all 5 appliance categories",
                  categories_found == 5,
                  f"Found {categories_found}/5 in keys: {list(data.keys())}")
            # Values must be lists with at least one non-empty string element
            non_list = [k for k, v in data.items() if not isinstance(v, list)]
            check("All JSON values are lists",
                  len(non_list) == 0,
                  f"Non-list keys: {non_list}")
            non_empty_strings = sum(
                sum(1 for r in v if isinstance(r, str) and r.strip())
                for v in data.values()
                if isinstance(v, list)
            )
            # task.md requires retrieving >=12 recipes and matching each by the
            # appliance rules. The JSON only contains recipes that landed in an
            # appliance bucket; a faithful matcher may legitimately leave a few
            # retrieved recipes unmatched (e.g. a dish that is neither
            # stir-fry/deep-fry, soup/stew/steam, nor a measured/baking item).
            # Requiring every one of the 12+ retrieved recipes to appear as a
            # pair would wrongly FAIL a correct agent, so require a solid >=10
            # pairings: still proves real matching happened, with room for a
            # few legitimately-unmatched recipes.
            check("JSON has 10+ recipe-name entries (non-empty strings)",
                  non_empty_strings >= 10,
                  f"Total non-empty: {non_empty_strings}")
            # All recipe names must be strings (not e.g. ints) — prevents fabricated stub data
            non_string_values = []
            for k, v in data.items():
                if isinstance(v, list):
                    non_string_values.extend([r for r in v if not isinstance(r, str)])
            check("All recipe entries are strings",
                  len(non_string_values) == 0,
                  f"Non-string entries: {non_string_values[:5]}")


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Meal_Kit_Analysis.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        check("No negative values in Excel", False,
                              f"Found {cell} in sheet {sheet_name}")
                        return
        check("No negative values in Excel", True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_notion()
    check_scripts(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL: {FAIL_COUNT} check(s) failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
