"""
Evaluation script for memory-multi-source-research task.

Checks:
1. Excel file has Paper Summary and Research Progress sheets with correct data
2. Word document has report with required sections
3. Memory file has been updated with entities

Usage:
    python -m evaluation.main --agent_workspace <path> --groundtruth_workspace <path>
"""
import argparse
import json
import os
import re
import sys
from datetime import datetime

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def normalize(text):
    return re.sub(r'\s+', ' ', text.lower().strip())


EXPECTED_PAPER_FRAGMENTS = [
    "concrete problems in ai safety",
    "ai safety via debate",
    "risks from learned optimization",
    "red teaming language models",
    "alignment of language agents",
    "scalable oversight of ai systems via recursive reward modeling",
]

# Fallback relevance signal for when the exact expected titles are not all
# present (e.g. scholarly search is unavailable and the model falls back on
# its own knowledge to find different, equally valid AI-safety papers).
# A data row counts as "AI-safety relevant" if any keyword appears in it
# (matched against the title / authors / key finding text).
AI_SAFETY_RELEVANCE_KEYWORDS = [
    "ai safety", "alignment", "align", "learned optimization",
    "mesa-optimization", "reward", "red team", "oversight", "language model",
    "machine learning", "reinforcement learning", "debate", "human feedback",
    "human preference", "human values", "preference", "objective", "scalable",
    "safety", "harm", "shutdown", "switch", "deep learning", "rlhf", "agent",
]

# AI-safety-focused relevance signals, used for the memory paper-entity check
# and the Word prose fallback. Broader than the expected-title keywords so a
# model that correctly found different-but-valid AI-safety papers (e.g. when
# scholarly search is unavailable) still counts, but focused on
# AI-safety/alignment content rather than generic machine learning, so
# off-topic deliverables (generic ML / retail / grocery) do not slip through.
AI_SAFETY_FOCUSED_KEYWORDS = [
    "ai safety", "safety", "alignment", "align", "learned optimization",
    "mesa-optimization", "red team", "oversight", "human feedback",
    "human preference", "human values", "shutdown", "off-switch", "harm",
    "debate", "reward model", "reward hacking", "rlhf", "value alignment",
    "scalable",
]


def check_excel(agent_workspace):
    """Check Research_Analysis.xlsx has correct data."""
    print("\n=== Checking Excel Output ===")

    import openpyxl

    excel_path = os.path.join(agent_workspace, "Research_Analysis.xlsx")
    if not os.path.isfile(excel_path):
        check("Research_Analysis.xlsx exists", False, f"Not found: {excel_path}")
        return

    check("Research_Analysis.xlsx exists", True)

    try:
        # Read formulas too (data_only=False) so formula cells never come back
        # as None when the workbook was never recalculated.
        wb = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    # Check Paper Summary sheet
    paper_sheet = None
    for name in wb.sheetnames:
        if "paper" in name.lower() and "summary" in name.lower():
            paper_sheet = wb[name]
            break

    if paper_sheet is None:
        check("Paper Summary sheet exists", False,
              f"Sheets found: {wb.sheetnames}")
    else:
        check("Paper Summary sheet exists", True)

        rows = list(paper_sheet.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        check("Paper Summary has at least 5 data rows",
              len(data_rows) >= 5,
              f"Found {len(data_rows)} data rows")

        # Check that target papers appear. Lenient by design: pass if either
        # at least 3 of the expected paper titles are present, OR at least 5
        # rows are AI-safety-relevant content. This lets a model that correctly
        # found >=5 different-but-valid AI-safety papers (e.g. when scholarly
        # search is unavailable) still pass.
        all_text = " ".join(
            str(cell).lower() for row in data_rows for cell in row if cell is not None
        )
        found_count = sum(
            1 for f in EXPECTED_PAPER_FRAGMENTS if f in all_text
        )
        relevant_rows = 0
        for row in data_rows:
            row_text = " ".join(str(c).lower() for c in row if c is not None)
            if any(k in row_text for k in AI_SAFETY_RELEVANCE_KEYWORDS):
                relevant_rows += 1
        papers_ok = (found_count >= 3) or (relevant_rows >= 5)
        check("Paper Summary contains expected AI-safety papers",
              papers_ok,
              f"Found {found_count} of {len(EXPECTED_PAPER_FRAGMENTS)} expected papers; "
              f"{relevant_rows}/{len(data_rows)} rows AI-safety relevant")

        # Check header has expected columns
        if rows:
            header_cells = [str(h).lower() if h is not None else "" for h in rows[0]]
            header = " ".join(c for c in header_cells if c)
            has_title_col = "title" in header
            has_year_col = "year" in header
            has_citations_col = "citation" in header
            has_authors_col = "author" in header
            check("Paper Summary has Title column", has_title_col,
                  f"Header: {header}")
            check("Paper Summary has Authors column", has_authors_col,
                  f"Header: {header}")
            check("Paper Summary has Year column", has_year_col,
                  f"Header: {header}")
            check("Paper Summary has Citations column", has_citations_col,
                  f"Header: {header}")

            # Authors column non-empty per row
            if has_authors_col:
                try:
                    author_col_idx = next(i for i, c in enumerate(header_cells) if "author" in c)
                    empty_authors = 0
                    for r in data_rows:
                        if author_col_idx < len(r):
                            val = r[author_col_idx]
                            if val is None or str(val).strip() == "":
                                empty_authors += 1
                    check("Paper Summary Authors column non-empty for all rows",
                          empty_authors == 0,
                          f"{empty_authors} rows with empty Authors")
                except StopIteration:
                    pass

    # Check Research Progress sheet. Prefer 'progress' first (unambiguous),
    # then fall back to 'research' but only for sheets that aren't the paper
    # summary sheet. This avoids "research" accidentally matching an unrelated
    # sheet before finding the actual progress sheet.
    progress_sheet = None
    for name in wb.sheetnames:
        if "progress" in name.lower():
            progress_sheet = wb[name]
            break
    if progress_sheet is None:
        for name in wb.sheetnames:
            if "research" in name.lower():
                if paper_sheet is None or wb[name] != paper_sheet:
                    progress_sheet = wb[name]
                    break

    if progress_sheet is None:
        check("Research Progress sheet exists", False,
              f"Sheets found: {wb.sheetnames}")
    else:
        check("Research Progress sheet exists", True)

        rows = list(progress_sheet.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        # Task: 2 search rounds + analysis + writing => at least 4 rows
        check("Research Progress has at least 4 data rows",
              len(data_rows) >= 4,
              f"Found {len(data_rows)} data rows")

        if rows:
            header = " ".join(str(h).lower() for h in rows[0] if h is not None)
            has_phase = "phase" in header
            has_status = "status" in header
            check("Research Progress has Phase column", has_phase,
                  f"Header: {header}")
            check("Research Progress has Status column", has_status,
                  f"Header: {header}")


def check_word(agent_workspace):
    """Check Research_Report.docx has required content."""
    print("\n=== Checking Word Report ===")

    from docx import Document

    docx_path = os.path.join(agent_workspace, "Research_Report.docx")
    if not os.path.isfile(docx_path):
        check("Research_Report.docx exists", False, f"Not found: {docx_path}")
        return

    check("Research_Report.docx exists", True)

    try:
        doc = Document(docx_path)
        full_text = "\n".join([para.text for para in doc.paragraphs])
    except Exception as e:
        check("Word document readable", False, str(e))
        return

    normalized = normalize(full_text)

    check("Report has at least 500 characters",
          len(full_text.strip()) >= 500,
          f"Document has {len(full_text.strip())} characters")

    # Check sections
    check("Report has Introduction section",
          "introduction" in normalized,
          "No Introduction section found")

    # Require phrase "literature review" (not just "literature")
    has_lit_review = "literature review" in normalized
    check("Report has Literature Review section (phrase)",
          has_lit_review,
          "Expected phrase 'literature review'")

    has_key_findings = "key findings" in normalized
    check("Report has Key Findings section (phrase)",
          has_key_findings,
          "Expected phrase 'key findings'")

    has_gaps = "research gaps" in normalized
    check("Report has Research Gaps section (phrase)",
          has_gaps,
          "Expected phrase 'research gaps'")

    check("Report has Conclusion section",
          "conclusion" in normalized,
          "No Conclusion section found")

    # Check paper mentions - lenient, mirroring the Excel/memory fallback.
    # A model that correctly found different-but-valid AI-safety papers (when
    # scholarly search is unavailable) may discuss them by title fragment, by
    # year mention in any common form, or by thematic AI-safety prose. Pass if
    # any signal is strong enough:
    #   - >=3 expected-title mentions, OR
    #   - >=4 publication-year mentions in ANY common citation form: "(2017)",
    #     "[2017]", "in 2017", "published 2018", "2017,", or a bare year in
    #     prose (restricted to a plausible publication-year range so citation
    #     counts / ids don't inflate the count), OR
    #   - the prose is clearly AI-safety-relevant (>=4 focused keywords).
    paper_mention_count = 0
    for title_fragment in EXPECTED_PAPER_FRAGMENTS:
        if title_fragment in normalized:
            paper_mention_count += 1
    current_year = datetime.now().year
    year_citations = 0
    for m in re.finditer(r'\b(\d{4})\b', full_text):
        y = int(m.group(1))
        if 1900 <= y <= current_year:
            year_citations += 1
    keyword_hits = sum(1 for k in AI_SAFETY_FOCUSED_KEYWORDS if k in normalized)
    mentions_ok = ((paper_mention_count >= 3) or (year_citations >= 4)
                   or (keyword_hits >= 4))
    check("Report discusses expected papers",
          mentions_ok,
          f"Found {paper_mention_count} title mentions, {year_citations} year mentions, "
          f"{keyword_hits} AI-safety keywords")


def check_memory(agent_workspace):
    """Check that memory.json has research tracking entities."""
    print("\n=== Checking Memory ===")

    memory_path = os.path.join(agent_workspace, "memory", "memory.json")
    if not os.path.isfile(memory_path):
        check("memory.json exists", False, f"Not found: {memory_path}")
        return

    check("memory.json exists", True)

    with open(memory_path, "r") as f:
        content = f.read().strip()

    if not content or content in ("{}", '{"entities": [], "relations": []}'):
        check("Memory has content", False, "memory.json is empty or unchanged")
        return

    check("Memory has content", True)

    # The memory MCP server can emit either a single JSON object/array or
    # newline-delimited JSON (JSONL, one entity/record per line). Parse both
    # (audit §B.1.5): first try a single-document parse; on failure, fall back
    # to line-by-line parsing and merge into one entity list.
    memory_data = None
    try:
        memory_data = json.loads(content)
    except json.JSONDecodeError:
        parsed_lines = []
        ok = True
        for ln in content.splitlines():
            ln = ln.strip()
            if not ln:
                continue
            try:
                parsed_lines.append(json.loads(ln))
            except json.JSONDecodeError:
                ok = False
                break
        if ok and parsed_lines:
            memory_data = parsed_lines
        else:
            check("Memory is valid JSON", False, "Cannot parse memory.json (tried JSON and JSONL)")
            return

    check("Memory is valid JSON", True)

    # Check for entities
    entities = memory_data.get("entities", []) if isinstance(memory_data, dict) else []
    if isinstance(memory_data, list):
        entities = memory_data

    # 5 paper entities + 1 research tracking entity = at least 6
    check("Memory has at least 6 entities",
          len(entities) >= 6,
          f"Found {len(entities)} entities (expect 5 papers + research tracker)")

    entity_text = ""
    for ent in entities:
        if isinstance(ent, dict):
            entity_text += json.dumps(ent).lower() + " "

    # Check for research tracking entity
    has_research = ("research" in entity_text or "ai_safety" in entity_text
                    or "ai safety" in entity_text)
    check("Memory has research tracking entity",
          has_research,
          "No research tracking entity found")

    # Check for paper entities - lenient, mirroring the Excel/Word fallback.
    # The expected-title keywords are all bound to the six preprocess-injected
    # papers; a model that correctly found >=5 different-but-valid AI-safety
    # papers (when scholarly search is unavailable) may not hit any of them.
    # Use the AI-safety-focused keyword set so such a model still passes while
    # off-topic content (generic ML / retail / grocery) does not.
    kw_count = sum(1 for kw in AI_SAFETY_FOCUSED_KEYWORDS if kw in entity_text)
    check("Memory has paper-related entities (at least 2 keywords)",
          kw_count >= 2,
          f"Found {kw_count} AI-safety keywords in memory")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_memory(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== SUMMARY ===")
    print(f"Results: {PASS_COUNT}/{total} passed, {FAIL_COUNT} failed")

    sys.exit(0 if FAIL_COUNT == 0 else 1)


if __name__ == "__main__":
    main()
