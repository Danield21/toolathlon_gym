"""Evaluation for wc-tax-compliance-excel-gcal-gform."""
import os
import json
import re
import argparse, os, sys
from datetime import date
import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=int(os.environ.get("PGPORT", "5432")),
    dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
    user=os.environ.get("PGUSER", "eigent"),
    password=os.environ.get("PGPASSWORD", "camel"),
)


def _to_float(v):
    """Robustly parse a value to float. Supports numbers and strings with
    thousands separators, currency symbols ($, EUR, JPY), % and surrounding
    whitespace. Returns None for None/non-numeric/formula cells."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan"):
        return None
    if s.startswith("="):
        return None
    s = s.replace(",", "").replace("$", "").replace("€", "").replace("¥", "").replace("£", "").replace("%", "")
    # Strip common currency words that might prefix a value ("USD 1506.33").
    s = re.sub(r"\b(usd|eur|jpy|cny|gbp|rmb|yuan)\b", "", s, flags=re.IGNORECASE)
    s = s.strip()
    if s.startswith("+"):
        s = s[1:].strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _norm(s):
    """Lowercase and strip all non-alphanumerics for tolerant key/status
    comparison ('Over-Collection' == 'overcollection', 'Total_Orders_Audited'
    == 'totalordersaudited')."""
    if s is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _to_int(v):
    """Robustly parse an integer (order id / counts). Returns None when the
    value is missing or not integral (never raises)."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if float(v).is_integer() else None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(s)
    except (ValueError, TypeError):
        try:
            f = float(s)
            return int(f) if f.is_integer() else None
        except (ValueError, TypeError):
            return None


def _num_mismatch(a, g, tol):
    """True when the GT cell holds a number that the agent's cell does not
    match within tolerance. A GT value that parses requires an agent value
    that parses and is close; an empty/unparseable agent cell is a mismatch
    (it must not silently pass a core numeric check). Both non-numeric -> not
    a mismatch."""
    if g is None:
        return False
    if a is None:
        return True
    return abs(a - g) > tol


def num_close(a, b, tol=None, rel_tol=0.15, abs_tol=0.5):
    """Robust numeric closeness. Parses both sides to float; only falls back to
    case-insensitive string equality when either side is not numeric."""
    fa, fb = _to_float(a), _to_float(b)
    if fa is None or fb is None:
        return str(a or "").strip().lower() == str(b or "").strip().lower()
    return abs(fa - fb) <= max(abs_tol, abs(fb) * rel_tol)


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _num_for(cached_matrix, raw_matrix, r, c):
    """Resolve a cell to a numeric float for comparison.
    - If the raw cell holds a formula (starts with '='), use its cached result
      (which is None if the formula was never recalculated by Excel).
    - Otherwise parse the raw literal value.
    Returns None when the cell cannot be resolved to a number.
    """
    if raw_matrix is None or r >= len(raw_matrix) or c >= len(raw_matrix[r]):
        return None
    rawv = raw_matrix[r][c]
    if isinstance(rawv, str) and rawv.strip().startswith("="):
        cachev = None
        if cached_matrix is not None and r < len(cached_matrix) and c < len(cached_matrix[r]):
            cachev = cached_matrix[r][c]
        return _to_float(cachev)
    return _to_float(rawv)


def _sheet_pair(wb_raw, wb_cached, sheet_name):
    return load_sheet_rows(wb_raw, sheet_name), load_sheet_rows(wb_cached, sheet_name)


def check_excel(agent_workspace, groundtruth_workspace):
    errors = []
    import openpyxl

    agent_path = os.path.join(agent_workspace, "Tax_Compliance_Report.xlsx")
    if not os.path.exists(agent_path):
        return ["Tax_Compliance_Report.xlsx not found"]

    gt_path = os.path.join(groundtruth_workspace, "Tax_Compliance_Report.xlsx")
    if not os.path.exists(gt_path):
        return ["Groundtruth Tax_Compliance_Report.xlsx not found"]

    try:
        # Read raw (data_only=False) so literal values are compared directly,
        # plus a cached (data_only=True) view to read formula results when an
        # agent writes formulas. GT is authored with literal values only.
        wb_agent = openpyxl.load_workbook(agent_path, data_only=False)
        wb_agent_c = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=False)
        wb_gt_c = openpyxl.load_workbook(gt_path, data_only=True)

        # --- Sheet 1: Order Tax Audit ---
        agent_rows, agent_rows_c = _sheet_pair(wb_agent, wb_agent_c, "Order Tax Audit")
        gt_rows, gt_rows_c = _sheet_pair(wb_gt, wb_gt_c, "Order Tax Audit")
        if agent_rows is None:
            errors.append("Sheet 'Order Tax Audit' not found")
        elif gt_rows is None:
            errors.append("Groundtruth sheet 'Order Tax Audit' not found")
        else:
            # Carry the actual matrix row index along with each data row so a
            # blank separator row in the agent workbook (r[0] is None) does not
            # shift the per-column lookups below.
            agent_data = [(m, r) for m, r in enumerate(agent_rows[1:], start=1) if r and r[0] is not None]
            gt_data = [(m, r) for m, r in enumerate(gt_rows[1:], start=1) if r and r[0] is not None]
            if len(agent_data) != len(gt_data):
                errors.append(f"Order Tax Audit: {len(agent_data)} rows, expected {len(gt_data)}")
            else:
                # Build lookup by Order_ID -> (matrix row index, row)
                gt_lookup = {}
                for gm, r in gt_data:
                    oid = _to_int(r[0])
                    if oid is not None:
                        gt_lookup[oid] = (gm, r)

                mismatches = 0
                for am, r in agent_data:
                    oid = _to_int(r[0])
                    if oid is None or oid not in gt_lookup:
                        mismatches += 1
                        continue
                    gm, gt_r = gt_lookup[oid]
                    # Check Order_Total (col 1), tolerance 0.5
                    a_total = _num_for(agent_rows_c, agent_rows, am, 1)
                    g_total = _num_for(gt_rows_c, gt_rows, gm, 1)
                    if _num_mismatch(a_total, g_total, 0.5):
                        mismatches += 1
                        continue
                    # Check Applicable_Rate (col 3), tolerance 0.001
                    a_rate = _num_for(agent_rows_c, agent_rows, am, 3)
                    g_rate = _num_for(gt_rows_c, gt_rows, gm, 3)
                    if _num_mismatch(a_rate, g_rate, 0.001):
                        mismatches += 1
                        continue
                    # Check Expected_Tax (col 4), tolerance 0.5
                    a_exp = _num_for(agent_rows_c, agent_rows, am, 4)
                    g_exp = _num_for(gt_rows_c, gt_rows, gm, 4)
                    if _num_mismatch(a_exp, g_exp, 0.5):
                        mismatches += 1
                        continue
                    # Check Status (col 7) - tolerant of casing / punctuation
                    a_status = _norm(r[7])
                    g_status = _norm(gt_r[7])
                    if a_status != g_status:
                        mismatches += 1

                if mismatches > 5:
                    errors.append(f"Order Tax Audit: {mismatches} row mismatches (>5 threshold)")

        # --- Sheet 2: State Summary ---
        agent_ss, agent_ss_c = _sheet_pair(wb_agent, wb_agent_c, "State Summary")
        gt_ss, gt_ss_c = _sheet_pair(wb_gt, wb_gt_c, "State Summary")
        if agent_ss is None:
            errors.append("Sheet 'State Summary' not found")
        elif gt_ss is None:
            errors.append("Groundtruth sheet 'State Summary' not found")
        else:
            agent_ss_data = [(m, r) for m, r in enumerate(agent_ss[1:], start=1) if r and r[0] is not None]
            gt_ss_data = [(m, r) for m, r in enumerate(gt_ss[1:], start=1) if r and r[0] is not None]
            if abs(len(agent_ss_data) - len(gt_ss_data)) > 2:
                errors.append(f"State Summary: {len(agent_ss_data)} rows, expected ~{len(gt_ss_data)}")
            else:
                gt_state_lookup = {_norm(r[0]): (m, r) for m, r in gt_ss_data if r[0] is not None}
                ss_mismatches = 0
                for am, r in agent_ss_data:
                    state = _norm(r[0])
                    if not state or state not in gt_state_lookup:
                        ss_mismatches += 1
                        continue
                    gm, gt_r = gt_state_lookup[state]
                    # Check Order_Count (col 1)
                    a_count = _num_for(agent_ss_c, agent_ss, am, 1)
                    g_count = _num_for(gt_ss_c, gt_ss, gm, 1)
                    if _num_mismatch(a_count, g_count, 0):
                        ss_mismatches += 1
                        continue
                    # Check compliance rate (col 6), tolerance 5.0
                    a_comp = _num_for(agent_ss_c, agent_ss, am, 6)
                    g_comp = _num_for(gt_ss_c, gt_ss, gm, 6)
                    if _num_mismatch(a_comp, g_comp, 5.0):
                        ss_mismatches += 1

                if ss_mismatches > 1:
                    errors.append(f"State Summary: {ss_mismatches} state mismatches (>1 threshold)")

        # --- Sheet 3: Compliance Overview ---
        agent_co = load_sheet_rows(wb_agent, "Compliance Overview")
        gt_co = load_sheet_rows(wb_gt, "Compliance Overview")
        if agent_co is None:
            errors.append("Sheet 'Compliance Overview' not found")
        elif gt_co is None:
            errors.append("Groundtruth sheet 'Compliance Overview' not found")
        else:
            # Label keys are normalized so "Total Orders Audited", "Total_Orders_Audited"
            # and "totalordersaudited" are equivalent.
            agent_co_data = {_norm(r[0]): r[1] for r in agent_co[1:] if r and r[0] is not None}
            gt_co_data = {_norm(r[0]): r[1] for r in gt_co[1:] if r and r[0] is not None}

            # Check total orders
            a_total = _to_float(agent_co_data.get(_norm("total_orders_audited")))
            g_total = _to_float(gt_co_data.get(_norm("total_orders_audited")))
            if _num_mismatch(a_total, g_total, 0):
                errors.append(f"Total_Orders_Audited: {a_total}, expected {g_total}")

            # Check compliant orders (tolerance 5)
            a_comp = _to_float(agent_co_data.get(_norm("compliant_orders")))
            g_comp = _to_float(gt_co_data.get(_norm("compliant_orders")))
            if _num_mismatch(a_comp, g_comp, 5):
                errors.append(f"Compliant_Orders: {a_comp}, expected {g_comp}")

            # Check overall compliance rate (tolerance 5)
            a_rate = _to_float(agent_co_data.get(_norm("overall_compliance_rate")))
            g_rate = _to_float(gt_co_data.get(_norm("overall_compliance_rate")))
            if _num_mismatch(a_rate, g_rate, 5.0):
                errors.append(f"Overall_Compliance_Rate: {a_rate}, expected {g_rate}")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_gcal():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date FROM gcal.events
            WHERE summary ILIKE '%tax filing%' OR summary ILIKE '%tax deadline%'
               OR summary ILIKE '%filing deadline%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) < 4:
            errors.append(f"Expected 4 tax filing deadline events in GCal, found {len(rows)}")
            return errors

        # Enforce specific quarter label -> date mapping (dates come from the
        # regulations PDF in the task materials, so they are fixed, not drift).
        expected = {
            "q1": date(2026, 4, 15),
            "q2": date(2026, 7, 15),
            "q3": date(2026, 10, 15),
            "q4": date(2026, 1, 15),
        }
        for q, ed in expected.items():
            matched = False
            for summary, dt in rows:
                s_lower = (summary or "").lower()
                # Require whole-word 'qN' in summary
                if not re.search(r"\b" + q + r"\b", s_lower):
                    continue
                if dt is None:
                    continue
                if abs((dt - ed).days) <= 1:
                    matched = True
                    break
            if not matched:
                errors.append(f"No '{q.upper()}' event on {ed} (+/- 1 day) found in GCal")

    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
    return errors


def _is_radio(qtype, qconfig):
    """True when the question is a single-choice (radio) question. google-forms
    MCP stores question_type='choiceQuestion' with config.type='RADIO'; accept
    either signal (and other choice-ish labels) so a correct agent is never
    penalized for how the MCP happened to serialize the type."""
    t = str(qtype or "").lower()
    if any(k in t for k in ("choice", "radio", "multiple", "dropdown")):
        return True
    if qconfig is None:
        return False
    cfg = qconfig
    if isinstance(cfg, str):
        try:
            cfg = json.loads(cfg)
        except (ValueError, TypeError):
            return False
    if isinstance(cfg, dict):
        ctype = str(cfg.get("type", "")).upper()
        if any(k in ctype for k in ("RADIO", "CHOICE", "MULTIPLE", "DROP", "CHECKBOX")):
            return True
    return False


def check_gform():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%vendor%' OR title ILIKE '%tax information%'
               OR title ILIKE '%tax info%'
            ORDER BY created_at DESC LIMIT 5
        """)
        forms = cur.fetchall()

        if not forms:
            cur.close()
            conn.close()
            return ["No GForm found matching 'Vendor Tax Information'"]

        form_id = forms[0][0]

        cur.execute("""
            SELECT title, question_type, config FROM gform.questions
            WHERE form_id = %s ORDER BY position
        """, (form_id,))
        questions = cur.fetchall()
        cur.close()
        conn.close()

        if len(questions) < 5:
            errors.append(f"Vendor Tax Information form has {len(questions)} questions, expected 5")

        # Structural completeness. google-forms MCP only produces text questions
        # and single-choice radios, so a question that is not a radio/choice is
        # a free-form text question. The task needs 4 text fields (vendor name,
        # tax ID, state, certificate) plus 1 radio (tax-exempt status).
        radio_qs = []
        text_count = 0
        for qtitle, qtype, qconfig in questions:
            if _is_radio(qtype, qconfig):
                radio_qs.append((qtitle, qconfig))
            else:
                text_count += 1

        if not radio_qs:
            errors.append("No radio/choice question found (tax-exempt status must be a single-choice radio)")
        if text_count < 4:
            errors.append(
                f"Form must include at least 4 short-text questions "
                f"(vendor name, tax ID, state, certificate); found {text_count}"
            )

        # The tax-exempt status question must be a radio offering exactly the
        # two options "Yes" and "No". Checked structurally so a correctly-made
        # radio whose title rephrases the topic (e.g. "Tax status (choose one)")
        # is not penalized, while a radio with wrong/missing options still fails.
        yn_ok = False
        for qtitle, qconfig in radio_qs:
            vals = [str(o).strip().lower() for o in _extract_options(qconfig)]
            if len(vals) == 2 and sorted(vals) == ["no", "yes"]:
                yn_ok = True
                break
        if not yn_ok:
            desc = "; ".join(
                f"{qtitle or '(untitled)'}: {[str(o).strip().lower() for o in _extract_options(qconfig)] or ['(none)']}"
                for qtitle, qconfig in radio_qs
            )
            errors.append(
                "Tax-exempt status question must be a radio offering exactly Yes and No options; "
                f"found: {desc or 'no radio question'}"
            )

        # Topic coverage check with broad synonym groups, so a correct agent
        # that rephrases a title is not penalized (e.g. 'EIN' instead of 'tax
        # ID number', 'Company legal name' instead of 'Vendor name', 'US region'
        # / 'Jurisdiction' instead of 'State'). Matching is case-insensitive and
        # ignores whitespace. Tax-exempt is handled structurally above.
        q_titles = " ".join(str(q[0]) for q in questions).lower()
        q_titles_nospace = q_titles.replace(" ", "")
        required_topics = [
            (["vendor", "company", "business", "supplier", "merchant",
              "legal name", "organization", "firm", "entity", "name"], "vendor/company name"),
            (["tax id", "taxid", "tin", "ein", "employer id",
              "employer identification", "identification", "taxpayer",
              "tax number", "irs", "id number"], "tax ID number"),
            (["state", "jurisdiction", "region", "province", "registered",
              "registration", "location"], "state of registration"),
            (["certificate", "certification", "cert", "document", "proof", "upload"], "tax certificate"),
        ]
        for aliases, label in required_topics:
            if not any((a in q_titles) or (a in q_titles_nospace) for a in aliases):
                errors.append(f"Missing question about '{label}' in vendor form")

    except Exception as e:
        errors.append(f"Error checking GForm: {e}")
    return errors


def _extract_options(config):
    """Extract option value strings from a gform.questions.config JSONB value.

    The value may arrive as a dict (psycopg2 auto-parses jsonb), a JSON string,
    or a list. Supported shapes:
      {"type": "RADIO", "options": [{"value": "A"}, {"value": "B"}]}
      {"type": "RADIO", "options": ["A", "B"]}
      [{"value": "A"}, {"value": "B"}]
      ["A", "B"]
    """
    if config is None:
        return []
    cfg = config
    if isinstance(cfg, str):
        try:
            cfg = json.loads(cfg)
        except (ValueError, TypeError):
            return []
    if isinstance(cfg, dict):
        opts = cfg.get("options")
        if opts is None:
            opts = cfg.get("choices")
        if opts is None:
            opts = cfg.get("values")
        if opts is None and ("value" in cfg or "label" in cfg):
            opts = [cfg]
    elif isinstance(cfg, list):
        opts = cfg
    else:
        return []
    if isinstance(opts, dict):
        # Some serializers nest the option list, e.g. {"values": ["Yes", "No"]}
        opts = opts.get("values") or opts.get("options") or opts.get("value") or opts.get("choices")
    if not isinstance(opts, list):
        return []
    vals = []
    for o in opts:
        if isinstance(o, dict):
            v = o.get("value")
            if v is None:
                v = o.get("label")
            if v is not None:
                vals.append(str(v))
        elif isinstance(o, str):
            vals.append(o)
    return vals


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GForm...")
    errs = check_gform()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
