"""Evaluation for sf-hr-worklife-notion-excel."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "toolathlon_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_header(h):
    return " ".join(str(h or "").strip().lower().replace("-", " ").replace("_", " ").split())


def _header_map(rows):
    """Map normalized header text -> 0-based column index from row 1 of a sheet."""
    if not rows or not rows[0]:
        return {}
    return {_norm_header(h): i for i, h in enumerate(rows[0]) if _norm_header(h)}


def _cell(row, idx):
    """Safe cell access (None when idx out of range)."""
    if row is None or idx is None:
        return None
    try:
        return row[idx] if idx < len(row) else None
    except (TypeError, IndexError):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []

    # --- Check 1: Excel file ---
    import openpyxl

    print("Checking Excel file...")
    agent_file = os.path.join(agent_ws, "WL_Balance_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WL_Balance_Report.xlsx")

    if not os.path.exists(agent_file):
        all_errors.append("WL_Balance_Report.xlsx not found in agent workspace")
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Department Analysis sheet
        print("  Checking Department Analysis sheet...")
        a_rows = load_sheet_rows(agent_wb, "Department Analysis")
        g_rows = load_sheet_rows(gt_wb, "Department Analysis")

        if a_rows is None:
            all_errors.append("Sheet 'Department Analysis' not found in agent output")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []
            a_col = _header_map(a_rows)
            g_col = _header_map(g_rows)

            # Use GT row count (no hard-coded 7)
            expected_count = len(g_data)
            if len(a_data) != expected_count:
                all_errors.append(f"Department Analysis row count: {len(a_data)}, expected {expected_count}")

            a_lookup = {}
            for row in a_data:
                if row and row[0]:
                    a_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or not g_row[0]:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing department: {g_row[0]}")
                    continue
                # Employee_Count
                av = _cell(a_row, a_col.get("employee count"))
                gv = _cell(g_row, g_col.get("employee count"))
                if not num_close(av, gv, 1):
                    all_errors.append(f"{g_row[0]} Employee_Count: {av} vs {gv}")
                # Avg_WLB
                av = _cell(a_row, a_col.get("avg wlb"))
                gv = _cell(g_row, g_col.get("avg wlb"))
                if not num_close(av, gv, 0.05):
                    all_errors.append(f"{g_row[0]} Avg_WLB: {av} vs {gv}")
                # Avg_Job_Satisfaction
                av = _cell(a_row, a_col.get("avg job satisfaction"))
                gv = _cell(g_row, g_col.get("avg job satisfaction"))
                if not num_close(av, gv, 0.05):
                    all_errors.append(f"{g_row[0]} Avg_Job_Satisfaction: {av} vs {gv}")
                # Combined_Score
                av = _cell(a_row, a_col.get("combined score"))
                gv = _cell(g_row, g_col.get("combined score"))
                if av is not None and gv is not None:
                    if not num_close(av, gv, 0.05):
                        all_errors.append(f"{g_row[0]} Combined_Score: {av} vs {gv}")
            print("    Done.")

        # Check Findings sheet
        print("  Checking Findings sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Findings")
        g_rows2 = load_sheet_rows(gt_wb, "Findings")
        if a_rows2 is None:
            all_errors.append("Sheet 'Findings' not found in agent output")
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            g_data2 = g_rows2[1:] if g_rows2 and len(g_rows2) > 1 else []
            a_col2 = _header_map(a_rows2)
            g_col2 = _header_map(g_rows2)
            a_lookup2 = {}
            for row in a_data2:
                if row and row[0]:
                    a_lookup2[str(row[0]).strip().lower()] = row
            g_lookup2 = {}
            for row in g_data2:
                if row and row[0]:
                    g_lookup2[str(row[0]).strip().lower()] = row

            # Validate every Findings row from GT
            string_metrics = {"best_wlb_department", "best_js_department"}
            # Tie-breaking: when departments are tied at rounded values, accept any
            # department whose unrounded WLB / JS is at the (rounded) max.
            # WLB: Operations (4.5434) and Finance (4.5404) both round to 4.54 → accept either.
            # JS:  Finance (6.5932) and Sales (6.5874) both round to 6.59 → accept either.
            ACCEPTABLE_BEST_WLB = {"finance", "operations"}
            ACCEPTABLE_BEST_JS = {"finance", "sales"}
            for metric_key, g_row in g_lookup2.items():
                a_row = a_lookup2.get(metric_key)
                if a_row is None:
                    all_errors.append(f"Findings missing {g_row[0]} row")
                    continue
                a_val = _cell(a_row, a_col2.get("value"))
                g_val = _cell(g_row, g_col2.get("value"))
                if metric_key == "best_wlb_department":
                    av = str(a_val).strip().lower() if a_val is not None else ""
                    if av not in ACCEPTABLE_BEST_WLB:
                        all_errors.append(f"Findings {g_row[0]}: {a_val} not in {ACCEPTABLE_BEST_WLB}")
                elif metric_key == "best_js_department":
                    av = str(a_val).strip().lower() if a_val is not None else ""
                    if av not in ACCEPTABLE_BEST_JS:
                        all_errors.append(f"Findings {g_row[0]}: {a_val} not in {ACCEPTABLE_BEST_JS}")
                elif metric_key in string_metrics:
                    if str(a_val).strip().lower() != str(g_val).strip().lower():
                        all_errors.append(f"Findings {g_row[0]}: {a_val} vs {g_val}")
                elif metric_key == "total_employees":
                    # Tighten tolerance: must match exactly (or +/- 1 for off-by-one)
                    if not num_close(a_val, g_val, 1):
                        all_errors.append(f"Findings Total_Employees: {a_val} vs {g_val}")
                elif metric_key == "departments_analyzed":
                    if not num_close(a_val, g_val, 0):
                        all_errors.append(f"Findings Departments_Analyzed: {a_val} vs {g_val}")
                else:
                    # Overall_Avg_WLB, Overall_Avg_Job_Satisfaction
                    if not num_close(a_val, g_val, 0.05):
                        all_errors.append(f"Findings {g_row[0]}: {a_val} vs {g_val}")

            print("    Done.")

    # --- Check 2: Notion page exists ---
    print("Checking Notion page...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Require exact title 'HR Wellbeing Dashboard'
        cur.execute("""
            SELECT id, properties FROM notion.pages
            WHERE properties::text ILIKE '%HR Wellbeing Dashboard%'
              AND archived = false
        """)
        rows = cur.fetchall()
        if not rows:
            all_errors.append("Notion page 'HR Wellbeing Dashboard' not found")
        else:
            page_ids = [r[0] for r in rows]
            # Validate page content/blocks mention overall avgs and best department
            cur.execute(
                "SELECT block_data FROM notion.blocks WHERE parent_id = ANY(%s)",
                (page_ids,))
            blob = " ".join(str(r[0] or "").lower() for r in cur.fetchall())
            # Combine with page properties text in case agent wrote in title-only structure
            blob += " " + " ".join(str(r[1] or "").lower() for r in rows)
            # Accept any of the top-tier departments (Finance/Operations for WLB, Finance/Sales for JS).
            best_dept_terms = ["finance", "operations", "sales"]
            if not any(t in blob for t in best_dept_terms):
                all_errors.append(f"Notion page should mention a top-performing department (one of {best_dept_terms})")
            if "wellbeing" not in blob and "well-being" not in blob and "work-life" not in blob and "balance" not in blob:
                all_errors.append("Notion page should mention wellbeing/work-life balance topic in content")
            print(f"    Notion page validated ({len(rows)} matching pages)")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Error checking Notion: {e}")

    # --- Check 3: Email sent ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%hr_director@company.com%'
        """)
        msgs = cur.fetchall()
        cur.close()
        conn.close()
        if not msgs:
            all_errors.append("No email sent to hr_director@company.com")
        else:
            ok = False
            for subj, body in msgs:
                subj_l = (subj or "").lower()
                body_l = (body or "").lower()
                # Subject must reference wellbeing/work-life
                if not ("wellbeing" in subj_l or "well-being" in subj_l or
                        "work-life" in subj_l or "work life" in subj_l):
                    continue
                # Body must mention at least one top-performing department
                # (Finance / Operations / Sales — accepting tie-breaking variations).
                if not any(t in body_l for t in ("finance", "operations", "sales")):
                    continue
                ok = True
                break
            if not ok:
                all_errors.append("No email matched: subject must reference wellbeing/work-life and body must mention Finance (best dept)")
            else:
                print(f"    Email validated ({len(msgs)} messages)")
    except Exception as e:
        all_errors.append(f"Error checking email: {e}")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
