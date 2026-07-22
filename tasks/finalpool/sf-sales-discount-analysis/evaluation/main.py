"""Evaluation for sf-sales-discount-analysis."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def check_email(errors):
    """Verify email to finance@company.com with subject 'Discount Impact Analysis'."""
    try:
        conn = psycopg2.connect(**DB); cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, COALESCE(body_text, body_html, '') FROM email.messages")
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        errors.append(f"Email check error: {e}")
        return
    target_subj = "discount impact analysis"
    target_to = "finance@company.com"
    matched = None
    for subj, to_addr, body in rows:
        if target_subj in (subj or "").lower() and target_to in str(to_addr or "").lower():
            matched = (subj, to_addr, body)
            break
    if not matched:
        errors.append(f"Email '{target_subj}' to {target_to} not found (checked {len(rows)})")
        return
    body_l = (matched[2] or "").lower()
    must = ["discount", "revenue"]
    missing = [m for m in must if m not in body_l]
    if missing:
        errors.append(f"Email body missing: {missing}")


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Discount_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Discount_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    
    # Check sheet: Discount Analysis
    print(f"  Checking Discount Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Discount Analysis")
    g_rows = load_sheet_rows(gt_wb, "Discount Analysis")
    if a_rows is None:
        all_errors.append("Sheet 'Discount Analysis' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Discount Analysis' not found in groundtruth")
    else:
        sheet_name = "Discount Analysis"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):  # exact order count
                    errors.append(f"{key}.Orders: {a_row[1]} vs {g_row[1]} (exact)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1.0):  # 2-decimal rounding
                    errors.append(f"{key}.Revenue: {a_row[2]} vs {g_row[2]} (tol=1.0)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0.05):
                    errors.append(f"{key}.Avg_Order_Value: {a_row[3]} vs {g_row[3]} (tol=0.05)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                # Total_Orders count exact; revenue values ±1.0
                tol = 0 if "order" in key and "value" not in key else 1.0
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    # Email check
    print(f"  Checking email...")
    email_errors = []
    check_email(email_errors)
    all_errors.extend(email_errors)
    if email_errors:
        for e in email_errors:
            print(f"    {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
