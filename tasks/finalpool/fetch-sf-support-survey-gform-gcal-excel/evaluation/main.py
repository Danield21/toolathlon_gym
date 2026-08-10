"""
Evaluation script for fetch-sf-support-survey-gform-gcal-excel task.

Checks:
1. Support_Satisfaction_Analysis.xlsx with 4 sheets and correct data
2. Google Form for ongoing feedback
3. Calendar events for 4 quarterly reviews (runtime_only by default;
   becomes blocking if agent populated relevant events)

Robustness notes (applied so that a correct solution always PASSes):
- The workbook is read BOTH with data_only=True (cached values) and
  data_only=False (formula strings). The excel MCP's apply_formula writes
  only the formula string with no computed cache, so a cell written via a
  formula yields None under data_only=True.
- Formula cells are best-effort EVALUATED (arithmetic, AVERAGE/SUM/COUNT/
  COUNTA/... over ranges and cell refs). A formula that evaluates is compared
  numerically for real, so a weak solution that writes =1 or =AVERAGE(0)
  fails; a correct solution that legitimately wrote =AVERAGE(B2:B21) passes.
  Only a formula the evaluator cannot parse keeps a soft pass, so a plausible
  correct formula is never misjudged. A truly EMPTY cell fails the numeric
  check (task.md explicitly asks for literal values).
- Numeric comparisons parse int/float/str and strip thousands separators,
  currency symbols, percent signs and whitespace.
- All DB connections read env vars with defaults aligned with preprocess.
"""

import argparse
import math
import os
import re
import sys

import openpyxl
import psycopg2
from openpyxl.utils import get_column_letter

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0
RUNTIME_ONLY_FAIL = 0

# Sentinel for a cell that holds an Excel formula with no cached value.
_FORMULA_ONLY = "FORMULA_ONLY"


class _FormulaValue(str):
    """A cell holding an Excel formula string whose cached value is missing."""


# Global evaluation context set while a sheet is being resolved:
# (data_only workbook, formula workbook, current sheet name).
_EVAL_CTX = None

_CELL_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+)$")


def _col_to_idx(col):
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n  # 1-based


def _norm_text(s):
    """Normalize an area/metric name for fuzzy matching."""
    s = str(s).strip().lower()
    s = s.replace("_", " ").replace("-", " ").replace("(", " ").replace(")", " ")
    return re.sub(r"\s+", " ", s).strip()


def _search_key(s):
    """Dense lowercase key used for substring area matching."""
    return re.sub(r"[^a-z0-9]", "", _norm_text(s))


def _num(v):
    """Best-effort float of a resolved value; None when not numeric."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return _to_float(v)


def _resolve_cell_ref(sheet_name, sheet_ref, coord, depth):
    """Resolve one cell reference (possibly cross-sheet) to a value."""
    if _EVAL_CTX is None or depth > 8:
        return _FORMULA_ONLY
    wb_donly, wb_formula, _ = _EVAL_CTX
    sh = sheet_name
    if sheet_ref is not None:
        want = sheet_ref.lower()
        sh = None
        for n in wb_formula.sheetnames:
            if n.lower() == want:
                sh = n
                break
        if sh is None:
            return _FORMULA_ONLY
    try:
        ws_d = wb_donly[sh]
        ws_f = wb_formula[sh]
        dval = ws_d[coord].value
        fval = ws_f[coord].value
    except Exception:
        return _FORMULA_ONLY
    if dval is not None:
        return dval
    if isinstance(fval, str) and fval.startswith("="):
        v = _evaluate_formula(fval, sh, depth + 1)
        return v if v is not None else _FORMULA_ONLY
    return fval


def _apply_func(fname, args):
    """Apply a supported Excel aggregate/scalar function to already-parsed args."""
    fname = fname.upper()
    if fname in ("CONCATENATE", "CONCAT"):
        parts = []
        for a in args:
            items = a if isinstance(a, list) else [a]
            for x in items:
                if x is None or x == _FORMULA_ONLY:
                    continue
                parts.append(str(x))
        return "".join(parts)
    nums = []
    for a in args:
        items = a if isinstance(a, list) else [a]
        for x in items:
            v = _num(x)
            if v is not None:
                nums.append(v)
    if fname == "SUM":
        return sum(nums)
    if fname == "AVERAGE":
        return sum(nums) / len(nums) if nums else None
    if fname == "COUNT":
        return float(len(nums))
    if fname == "COUNTA":
        cnt = 0
        for a in args:
            items = a if isinstance(a, list) else [a]
            for x in items:
                if x is not None and x != _FORMULA_ONLY:
                    cnt += 1
        return float(cnt)
    if fname == "MIN":
        return min(nums) if nums else None
    if fname == "MAX":
        return max(nums) if nums else None
    if fname == "MEDIAN":
        if not nums:
            return None
        s = sorted(nums)
        m = len(s) // 2
        return s[m] if len(s) % 2 else (s[m - 1] + s[m]) / 2.0
    if fname == "PRODUCT":
        p = 1.0
        for x in nums:
            p *= x
        return p if nums else None
    if fname == "ABS":
        return abs(nums[0]) if len(nums) == 1 else None
    if fname == "INT":
        return float(int(nums[0])) if len(nums) == 1 else None
    if fname in ("ROUND", "ROUNDUP", "ROUNDDOWN"):
        if len(nums) not in (1, 2):
            return None
        digits = int(nums[1]) if len(nums) == 2 else 0
        factor = 10 ** digits
        x = nums[0]
        if fname == "ROUND":
            return round(x, digits)
        if fname == "ROUNDUP":
            return math.ceil(x * factor) / factor if x >= 0 else math.floor(x * factor) / factor
        return math.floor(x * factor) / factor if x >= 0 else math.ceil(x * factor) / factor
    if fname == "MOD":
        return nums[0] % nums[1] if len(nums) == 2 else None
    if fname == "SQRT":
        return math.sqrt(nums[0]) if len(nums) == 1 else None
    if fname == "POWER":
        return nums[0] ** nums[1] if len(nums) == 2 else None
    raise ValueError("unsupported function " + fname)


class _ExcelEval:
    """Minimal recursive-descent evaluator for simple Excel formulas."""

    def __init__(self):
        self.tokens = []
        self.pos = 0
        self.sheet_name = None
        self.depth = 0

    def parse(self, expr, sheet_name, depth):
        self.sheet_name = sheet_name
        self.depth = depth
        self._tokenize(expr)
        val = self._expr()
        if self._peek()[0] != "EOF":
            raise ValueError("trailing tokens")
        return val

    def _tokenize(self, s):
        toks = []
        i, n = 0, len(s)
        while i < n:
            c = s[i]
            if c.isspace():
                i += 1
                continue
            if c == '"' or c == "'":
                # Double quotes delimit text values; single quotes delimit
                # sheet names ('Survey Results'!B2). Treat both as strings.
                close = s.find(c, i + 1)
                if close < 0:
                    raise ValueError("unterminated string")
                toks.append(("STR", s[i + 1:close]))
                i = close + 1
                continue
            if c.isdigit() or (c == "." and i + 1 < n and s[i + 1].isdigit()):
                j = i
                while j < n and (s[j].isdigit() or s[j] == "."):
                    j += 1
                toks.append(("NUM", s[i:j]))
                i = j
                continue
            if c.isalpha() or c in "_$":
                j = i
                while j < n and (s[j].isalnum() or s[j] in "_.$ "):
                    j += 1
                toks.append(("IDENT", s[i:j]))
                i = j
                continue
            if c in "+-*/^(),:%!<>=":
                toks.append((c, c))
                i += 1
                continue
            raise ValueError("unexpected char " + c)
        toks.append(("EOF", ""))
        self.tokens = toks
        self.pos = 0

    def _peek(self):
        return self.tokens[self.pos]

    def _peek2(self):
        return self.tokens[min(self.pos + 1, len(self.tokens) - 1)]

    def _next(self):
        tok = self.tokens[self.pos]
        self.pos += 1
        return tok

    def _expr(self):
        v = self._term()
        while self._peek()[0] in ("+", "-"):
            op = self._next()[0]
            v = self._arith(v, op, self._term())
        return v

    def _term(self):
        v = self._factor()
        while self._peek()[0] in ("*", "/"):
            op = self._next()[0]
            v = self._arith(v, op, self._factor())
        return v

    def _factor(self):
        tok = self._peek()
        if tok[0] in ("+", "-"):
            self._next()
            v = self._factor()
            return v if tok[0] == "+" else -v
        v = self._primary()
        while self._peek()[0] == "%":
            self._next()
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                v = v / 100.0
            else:
                raise ValueError("bad percent")
        return v

    def _arith(self, a, op, b):
        if (isinstance(a, bool) or isinstance(b, bool)
                or not isinstance(a, (int, float)) or not isinstance(b, (int, float))):
            raise ValueError("non-numeric operand")
        if op == "+":
            return a + b
        if op == "-":
            return a - b
        if op == "*":
            return a * b
        if op == "/":
            if b == 0:
                raise ValueError("div by zero")
            return a / b
        raise ValueError("bad op")

    def _primary(self):
        tok = self._peek()
        if tok[0] == "NUM":
            self._next()
            return float(tok[1])
        if tok[0] == "STR":
            self._next()
            if self._peek()[0] == "!":
                self._next()
                return self._ref(self.sheet_name, tok[1])
            return tok[1]
        if tok[0] == "IDENT":
            if self._peek2()[0] == "(":
                fname = self._next()[1]
                self._next()  # '('
                return self._call(fname)
            if self._peek2()[0] == "!":
                sheet_token = self._next()[1]
                self._next()  # '!'
                return self._ref(self.sheet_name, sheet_token)
            raw = self._next()[1]
            return self._ref(self.sheet_name, None, raw)
        if tok[0] == "(":
            self._next()
            v = self._expr()
            if self._peek()[0] != ")":
                raise ValueError("missing )")
            self._next()
            return v
        raise ValueError("unexpected token " + tok[0])

    def _call(self, fname):
        args = []
        if self._peek()[0] == ")":
            self._next()
        else:
            while True:
                args.append(self._expr())
                if self._peek()[0] == ")":
                    self._next()
                    break
                if self._peek()[0] == ",":
                    self._next()
                    continue
                raise ValueError("expected , or )")
        return _apply_func(fname, args)

    def _ref(self, sheet_name, sheet_ref, raw=None):
        if raw is None:
            raw = self._next()[1]
        m = _CELL_RE.match(raw)
        if not m:
            raise ValueError("bad ref " + raw)
        r1 = int(m.group(2))
        c1 = _col_to_idx(m.group(1))
        if self._peek()[0] == ":":
            self._next()
            tok2 = self._next()
            m2 = _CELL_RE.match(tok2[1])
            if not m2:
                raise ValueError("bad range end")
            r2 = int(m2.group(2))
            c2 = _col_to_idx(m2.group(1))
            vals = []
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    vals.append(_resolve_cell_ref(
                        sheet_name, sheet_ref, f"{get_column_letter(cc)}{rr}", self.depth))
            return vals
        return _resolve_cell_ref(sheet_name, sheet_ref, raw, self.depth)


def _evaluate_formula(formula, sheet_name, depth=0):
    """Best-effort evaluate a simple Excel formula string.

    Returns the computed value (number or string) or None when the formula is
    outside the supported subset (the caller then keeps the soft pass). Never
    raises.
    """
    if depth > 8 or _EVAL_CTX is None:
        return None
    if isinstance(formula, _FormulaValue):
        f = str(formula)
    elif isinstance(formula, str) and formula.startswith("="):
        f = formula
    else:
        return formula
    f = f[1:].strip() if f.startswith("=") else f.strip()
    if not f:
        return None
    if f.startswith('"') and f.endswith('"') and f.count('"') == 2:
        return f[1:-1]
    # Pure numeric arithmetic (no identifiers/functions) -> safe literal eval.
    if not re.search(r"[A-Za-z_]", f):
        if not re.fullmatch(r"[0-9+\-*/().^% ]+", f):
            return None
        try:
            return float(eval(f.replace("^", "**"), {"__builtins__": {}}, {}))
        except Exception:
            return None
    try:
        return _ExcelEval().parse(f, sheet_name, depth)
    except Exception:
        return None


def _materialize(v):
    """Resolve a formula value to a concrete value (or the soft-pass sentinel)."""
    if isinstance(v, _FormulaValue):
        if _EVAL_CTX is not None:
            val = _evaluate_formula(str(v), _EVAL_CTX[2])
            return val if val is not None else _FORMULA_ONLY
        return _FORMULA_ONLY
    return v


def record(name, passed, detail="", runtime_only=False):
    global PASS_COUNT, FAIL_COUNT, RUNTIME_ONLY_FAIL
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if runtime_only:
            RUNTIME_ONLY_FAIL += 1
        msg = f": {detail[:300]}" if detail else ""
        suffix = " (runtime-only)" if runtime_only else ""
        print(f"  [FAIL] {name}{suffix}{msg}")


def _to_float(v):
    """Robustly convert a cell value to float.

    Supports int/float/str (and None). Strips thousands separators, currency
    symbols, percent signs and whitespace before parsing. Returns None when
    the value is missing or unparseable.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    for ch in (",", "$", "¥", "€", "%", " ", "_"):
        s = s.replace(ch, "")
    try:
        return float(s)
    except ValueError:
        return None


def _cell_value(donly, formula):
    """Resolve one cell into a comparable value.

    - donly: value from the data_only=True workbook (None for a formula cell
      whose cache is missing).
    - formula: raw cell value from the data_only=False workbook (formula
      string starting with '=' when the cell holds a formula).

    Returns the resolved value, a _FormulaValue (formula string to be
    evaluated on demand), or None when the cell is truly empty.
    """
    if donly is not None:
        return donly
    if isinstance(formula, str) and formula.startswith("="):
        return _FormulaValue(formula)
    return formula


def _resolve_sheet_rows(wb_donly, wb_formula, sheet_name):
    """Return rows (list of row-lists) with each cell resolved via _cell_value.

    Sets the module-level evaluation context so on-demand formula evaluation
    (num_close / str_contains) resolves references in this sheet's workbook.
    """
    global _EVAL_CTX
    _EVAL_CTX = (wb_donly, wb_formula, sheet_name)
    ws_donly = wb_donly[sheet_name]
    ws_formula = wb_formula[sheet_name]
    rows = []
    for row in ws_formula.iter_rows():
        row_vals = []
        for cell in row:
            dval = ws_donly[cell.coordinate].value
            row_vals.append(_cell_value(dval, cell.value))
        rows.append(row_vals)
    return rows


def num_close(a, b, tol=0.5):
    a = _materialize(a)
    b = _materialize(b)
    if a == _FORMULA_ONLY or b == _FORMULA_ONLY:
        # A formula whose value we could not verify: soft pass so a plausible
        # correct formula is never misjudged (task.md asks for literals anyway).
        return True
    if a is None or b is None:
        # A truly empty cell must not silently pass a numeric check.
        return a is None and b is None
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol
    # One or both sides unparseable -> fall back to case-insensitive
    # string comparison.
    return str(a).strip().lower() == str(b).strip().lower()


def str_contains(haystack, needle):
    haystack = _materialize(haystack)
    if haystack == _FORMULA_ONLY:
        # Unverifiable formula -> soft pass (mirrors num_close).
        return True
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


def check_excel(agent_workspace):
    """Check Support_Satisfaction_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Support_Satisfaction_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)

    try:
        wb_donly = openpyxl.load_workbook(fpath, data_only=True)
        wb_formula = openpyxl.load_workbook(fpath, data_only=False)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet 1: Survey Results ---
    survey_sheet = None
    for name in wb_formula.sheetnames:
        if "survey" in name.lower() and "summary" not in name.lower():
            survey_sheet = name
            break
    if not survey_sheet:
        record("Survey Results sheet exists", False, f"Sheets: {wb_formula.sheetnames}")
        all_ok = False
    else:
        record("Survey Results sheet exists", True)
        rows = _resolve_sheet_rows(wb_donly, wb_formula, survey_sheet)
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 20
        record("Survey Results has 20 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 2: Survey Summary ---
    summary_sheet = None
    for name in wb_formula.sheetnames:
        if "summary" in name.lower():
            summary_sheet = name
            break
    if not summary_sheet:
        record("Survey Summary sheet exists", False, f"Sheets: {wb_formula.sheetnames}")
        all_ok = False
    else:
        record("Survey Summary sheet exists", True)
        rows = _resolve_sheet_rows(wb_donly, wb_formula, summary_sheet)
        data_rows = rows[1:] if len(rows) > 1 else []

        for row in data_rows:
            if row and row[0]:
                metric = str(row[0]).strip().lower()
                val = row[1]
                if "total_respondents" in metric or ("total" in metric and "respondent" in metric):
                    ok = num_close(val, 20, tol=0)
                    record("Total respondents = 20", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "avg_overall" in metric or ("overall" in metric and "satisfaction" in metric):
                    ok = num_close(val, 3.25, tol=0.3)
                    record("Avg overall satisfaction ~3.25", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "lowest" in metric:
                    ok = str_contains(val, "low")
                    record("Lowest rated priority is Low", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "highest" in metric:
                    ok = str_contains(val, "high")
                    record("Highest rated priority is High", ok, f"Got {val}")
                    if not ok:
                        all_ok = False

    # --- Sheet 3: Ticket System Comparison ---
    comp_sheet = None
    for name in wb_formula.sheetnames:
        if "ticket" in name.lower() or "comparison" in name.lower():
            comp_sheet = name
            break
    if not comp_sheet:
        record("Ticket System Comparison sheet exists", False, f"Sheets: {wb_formula.sheetnames}")
        all_ok = False
    else:
        record("Ticket System Comparison sheet exists", True)
        rows = _resolve_sheet_rows(wb_donly, wb_formula, comp_sheet)
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 3
        record("Ticket Comparison has 3 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

        # Check ticket avg response hours for High priority
        for row in data_rows:
            if row and str_contains(row[0], "high"):
                # Ticket avg response hours should be ~6.23
                found = False
                for cell in row[1:]:
                    if num_close(cell, 6.23, tol=1.0):
                        found = True
                        break
                record("High priority ticket response ~6.23 hrs", found,
                       f"Row: {str(row)[:200]}")
                if not found:
                    all_ok = False

    # --- Sheet 4: Improvement Areas ---
    # task.md says: "Include rows for each metric (Response Time, Resolution
    # Quality, Agent Professionalism) where the survey average is below 4.0.
    # The target score should be 4.5 for all metrics. The gap is the target
    # minus the current score."
    # GT data: Response Time=3.10, Resolution Quality=3.55 (both below 4.0).
    # Agent Professionalism avg = 4.0 (NOT below) so it should NOT be present.
    # → Expect exactly 2 rows: Response Time and Resolution Quality.
    imp_sheet = None
    for name in wb_formula.sheetnames:
        if "improvement" in name.lower():
            imp_sheet = name
            break
    if not imp_sheet:
        record("Improvement Areas sheet exists", False, f"Sheets: {wb_formula.sheetnames}")
        all_ok = False
    else:
        record("Improvement Areas sheet exists", True)
        rows = _resolve_sheet_rows(wb_donly, wb_formula, imp_sheet)
        data_rows = [r for r in rows[1:] if r and r[0]]

        # Build lookup by normalized area name ('Response_Time',
        # 'Response Time', 'Response-Time' all map to 'response time').
        by_area = {}
        for r in data_rows:
            if r and r[0]:
                by_area[_search_key(r[0])] = r

        # Expected areas + scores per GT. Matching uses the dense search key
        # so separator/case variants (Response_Time, RESPONSE TIME, ...) pass.
        expected_imp = {
            "response time": {"current": 3.10, "target": 4.5, "gap": 1.40},
            "resolution quality": {"current": 3.55, "target": 4.5, "gap": 0.95},
        }

        for area_key, exp in expected_imp.items():
            ak = _search_key(area_key)
            row = None
            for k, v in by_area.items():
                if ak in k:
                    row = v
                    break
            if row is None:
                record(f"Improvement Area '{area_key}' present", False,
                       f"Got areas: {list(_norm_text(k) for k in by_area)}")
                all_ok = False
                continue
            record(f"Improvement Area '{area_key}' present", True)
            # row format: (Area, Current_Score, Target_Score, Gap)
            if len(row) >= 4:
                cur_score = row[1]
                tgt_score = row[2]
                gap = row[3]
                ok = num_close(cur_score, exp["current"], 0.05)
                record(f"'{area_key}' Current_Score ~{exp['current']}",
                       ok, f"Got {cur_score}")
                if not ok:
                    all_ok = False
                # Target_Score must be 4.5 — task.md is explicit
                ok_tgt = num_close(tgt_score, 4.5, 0.01)
                record(f"'{area_key}' Target_Score = 4.5",
                       ok_tgt, f"Got {tgt_score}")
                if not ok_tgt:
                    all_ok = False
                # Gap = Target - Current
                ok_gap = num_close(gap, exp["gap"], 0.05)
                record(f"'{area_key}' Gap ~{exp['gap']}",
                       ok_gap, f"Got {gap}")
                if not ok_gap:
                    all_ok = False
            else:
                record(f"'{area_key}' has 4 columns", False,
                       f"Row has {len(row)} columns: {row}")
                all_ok = False

        # Agent Professionalism (avg = 4.0). The strict reading of task.md
        # ('below 4.0') EXCLUDES it (2 rows total). A lenient reading
        # ('not reaching the 4.5 target', i.e. <= 4.0) INCLUDES it. Both are
        # accepted: if the row is present it must carry the correct values
        # (current 4.0, target 4.5, gap 0.5); if absent that is the strict
        # reading. Rows must be exactly 2 (strict) or 3 (with professionalism).
        prof_keys = [k for k in by_area if "professionalism" in k or "agent" in k]
        if prof_keys:
            prof_row = by_area[prof_keys[0]]
            ok_prof = (len(prof_row) >= 4
                       and num_close(prof_row[1], 4.0, 0.05)
                       and num_close(prof_row[2], 4.5, 0.01)
                       and num_close(prof_row[3], 0.5, 0.05))
            record("Agent Professionalism row (if present) values correct",
                   ok_prof, f"Got row: {str(prof_row)[:120]}")
            if not ok_prof:
                all_ok = False
        else:
            record("Agent Professionalism row absent (strict 'below 4.0' reading)", True)

        # Row-count sanity: 2 core rows, optionally +1 professionalism row.
        expected_rows = 2 + (1 if prof_keys else 0)
        ok_count = len(data_rows) == expected_rows
        record("Improvement Areas row count consistent (2, or 3 incl. professionalism)",
               ok_count, f"Found {len(data_rows)}, expected {expected_rows}")
        if not ok_count:
            all_ok = False

    wb_donly.close()
    wb_formula.close()
    return all_ok


def _gform_title_score(title):
    """Score a form title for the customer-feedback task.

    The task.md title is "Customer Support Feedback Form" but a correct agent
    may reasonably name it "Satisfaction Feedback Survey", "Customer Feedback
    Survey", etc. Accept any title carrying a feedback vocabulary; exclude the
    injected noise form ("Employee Satisfaction Survey") and any other
    employee-targeted form. 'support'/'customer'/'satisfaction' are stronger
    signals than 'feedback'/'survey'/'form', so a real customer form wins over
    an unrelated seed form (e.g. "Coupon Campaign Feedback") when both exist.
    """
    tl = (title or "").lower()
    if "employee" in tl:
        return -1
    score = 0
    for kw in ("support", "customer", "satisfaction"):
        if kw in tl:
            score += 2
    for kw in ("feedback", "survey", "form"):
        if kw in tl:
            score += 1
    return score


def check_gform():
    """Check Google Form for ongoing feedback.

    The form is a hard task requirement (task.md), so the structural checks
    (form exists, enough questions) are BLOCKING. The content checks
    (satisfaction / comments question) stay runtime_only so that a correct
    agent whose question wording differs is not misjudged. The form-title
    detection accepts any customer-feedback vocabulary title (see
    _gform_title_score), so reasonable variants of "Customer Support Feedback
    Form" pass while the injected noise form ("Employee Satisfaction Survey")
    and unrelated seed forms do not.
    """
    print("\n=== Checking Google Form ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title, description FROM gform.forms")
        forms = cur.fetchall()

        found_form = False
        form_id = None
        best_score = 0
        for fid, title, desc in forms:
            score = _gform_title_score(title)
            if score > best_score:
                best_score = score
                found_form = True
                form_id = fid

        record("Customer feedback form exists", found_form,
               f"Found forms: {[(t, d[:50] if d else '') for _, t, d in forms]}")

        all_ok = found_form

        if form_id:
            cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s", (form_id,))
            questions = cur.fetchall()
            q_count = len(questions)
            ok = q_count >= 4
            record("Form has >= 4 questions", ok, f"Found {q_count}")
            if not ok:
                all_ok = False

            # Content checks are informational (runtime_only): a correct
            # agent may word a question title differently than the evaluator
            # guesses, so these must not block.
            q_titles = " ".join((t or "").lower() for t, _ in questions)
            has_satisfaction = any(k in q_titles for k in
                                   ("satisfaction", "satisfied", "overall", "rating", "rate"))
            record("Has satisfaction question", has_satisfaction,
                   f"Q titles: {q_titles[:200]}", runtime_only=True)

            # Open text / comments question: the google-forms MCP's
            # add_text_question stores question_type='textQuestion'; also
            # accept a title containing comment/feedback keywords.
            has_comment = any("text" in (qt or "").lower() for _, qt in questions)
            has_comment = has_comment or "comment" in q_titles or "feedback" in q_titles
            record("Has comments/text question", has_comment, runtime_only=True)

        cur.close()
        conn.close()
        return all_ok

    except Exception as e:
        record("Google Form DB accessible", False, str(e), runtime_only=True)
        return False


def _quarter_from_summary(sl):
    """Extract quarter 'q1'..'q4' from an event summary.

    Tolerant of wording: 'Q1', 'Q1 2026', 'Quarter 1 2026', 'Quarter1',
    'QTR 1', '1st Quarter' are all recognized. Numbers embedded in a larger
    number ('Q10') are deliberately NOT matched; the date fallback then
    supplies the quarter.
    """
    m = re.search(r"([1-4])(?:st|nd|rd|th)\s*(?:quarter|qtr)\b", sl)
    if m:
        return "q" + m.group(1)
    m = re.search(r"\bq(?:uarter|tr)?\s*([1-4])(?:\D|$)", sl)
    if m:
        return "q" + m.group(1)
    return None


def _quarter_from_date(dt):
    """Quarter implied by the event date (Mar/Jun/Sep/Dec -> q1..q4)."""
    if dt is None:
        return None
    return {3: "q1", 6: "q2", 9: "q3", 12: "q4"}.get(dt.month)


def check_calendar():
    """Check calendar events for 4 quarterly review meetings.

    Pattern: runtime_only by default (won't block on V1 GT-only test). BUT
    when the agent has populated 'Support Satisfaction Review' events, all
    quality checks (count, dates, durations) are blocking.
    """
    print("\n=== Checking Google Calendar ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e), runtime_only=True)
        return False

    # Detect agent population: any event whose summary carries the review
    # series vocabulary ('support'/'satisfaction' + 'review'). Seed and noise
    # events never match, so calendar checks stay runtime_only until the agent
    # actually creates the review meetings (then they become blocking).
    relevant = []
    for s, d, st, et in events:
        sl = (s or "").lower()
        if ("support" in sl or "satisfaction" in sl) and "review" in sl:
            relevant.append((s, d, st, et))
    agent_populated = len(relevant) > 0
    is_runtime_only = not agent_populated

    # Identify the quarter of each relevant event. The quarter is read from
    # the summary when possible ('Q1 2026' / 'Quarter 1 2026' / 'QTR 1' / ...)
    # and falls back to the event's own date (Mar/Jun/Sep/Dec -> q1..q4), so a
    # correct agent that worded the summary differently is not misjudged. Only
    # events that clearly belong to the review series are counted, so seeded
    # noise events (Campaign launches, All-Hands, ...) never inflate the count.
    quarters_found = set()
    quarter_events = []
    for summary, description, start_dt, end_dt in events:
        sl = (summary or "").lower()
        if not (("support" in sl or "satisfaction" in sl) and "review" in sl):
            continue
        q = _quarter_from_summary(sl) or _quarter_from_date(start_dt)
        if q:
            quarters_found.add(q)
            quarter_events.append((q, summary, start_dt, end_dt))

    ok = len(quarters_found) >= 4
    record("All 4 quarterly review events found", ok,
           f"Found quarters: {quarters_found}",
           runtime_only=is_runtime_only)

    # Per-event verifications: when agent populated, BLOCKING
    if agent_populated:
        # Expected dates: 2026-03-15, 2026-06-15, 2026-09-15, 2026-12-15
        expected_dates_by_q = {
            "q1": "2026-03-15",
            "q2": "2026-06-15",
            "q3": "2026-09-15",
            "q4": "2026-12-15",
        }
        for q, summary, start_dt, end_dt in quarter_events:
            if start_dt and end_dt:
                duration_min = (end_dt - start_dt).total_seconds() / 60
                record(f"Event {q} duration 90 min (10:00-11:30)",
                       abs(duration_min - 90) <= 10,
                       f"Got {duration_min} min",
                       runtime_only=False)
                # Date check
                exp_date = expected_dates_by_q.get(q)
                if exp_date:
                    actual_date = str(start_dt)[:10]
                    record(f"Event {q} date = {exp_date}",
                           actual_date == exp_date,
                           f"Got {actual_date}",
                           runtime_only=False)
    else:
        # Not populated → keep all per-event checks as runtime_only
        for q, summary, start_dt, end_dt in quarter_events:
            if start_dt and end_dt:
                duration_min = (end_dt - start_dt).total_seconds() / 60
                record(f"Event {q} duration 90 min (10:00-11:30)",
                       abs(duration_min - 90) <= 10,
                       f"Got {duration_min} min",
                       runtime_only=True)

    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    gform_ok = check_gform()
    cal_ok = check_calendar()

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  GForm:    {'PASS' if gform_ok else 'FAIL'}")
    print(f"  Calendar: {'PASS' if cal_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} (runtime-only fails: {RUNTIME_ONLY_FAIL})")

    blocking_fail = FAIL_COUNT - RUNTIME_ONLY_FAIL
    overall = blocking_fail == 0
    print(f"  Blocking failures: {blocking_fail}")
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
