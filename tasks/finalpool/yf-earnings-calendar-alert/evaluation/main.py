"""Evaluation for yf-earnings-calendar-alert."""
import os
import argparse
import json
import re
import sys
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, abs_tol=0.05, rel_tol=0.0):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


# Calendar page expectations
CALENDAR_EXPECTED = {
    "GOOGL": {"name": "Alphabet Inc.", "date": "2026-04-15", "expected_eps": 1.95, "expected_rev": 95.2},
    "AMZN":  {"name": "Amazon.com, Inc.", "date": "2026-04-22", "expected_eps": 1.15, "expected_rev": 160.5},
    "JPM":   {"name": "JP Morgan Chase & Co.", "date": "2026-04-10", "expected_eps": 4.50, "expected_rev": 44.8},
}


def fetch_yf_quarters():
    """Most recent 4 quarters of diluted EPS + total revenue per symbol."""
    out = {}
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        for sym in ["AMZN", "GOOGL", "JPM"]:
            cur.execute("""
                SELECT period_end, data
                FROM yf.financial_statements
                WHERE symbol=%s AND stmt_type='income_stmt' AND freq='quarterly'
                ORDER BY period_end DESC LIMIT 4
            """, (sym,))
            quarters = []
            for period_end, data in cur.fetchall():
                d = data if isinstance(data, dict) else json.loads(data)
                eps = d.get("Diluted EPS")
                rev = d.get("Total Revenue")
                quarters.append({
                    "period": period_end,
                    "eps": float(eps) if eps is not None else None,
                    "revenue": float(rev) if rev is not None else None,
                })
            out[sym] = quarters
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[fallback] yf fetch failed: {e}")
    return out


def hist_avg_eps(quarters):
    eps = [q["eps"] for q in quarters if q["eps"] is not None]
    if not eps:
        return None
    return round(sum(eps) / len(eps), 2)


def expected_trend(expected_eps, hist_avg):
    if hist_avg is None or hist_avg == 0:
        return None
    pct_diff = abs(expected_eps - hist_avg) / abs(hist_avg)
    if pct_diff <= 0.05:
        return "in line"
    return "above" if expected_eps > hist_avg else "below"


def check_excel(agent_workspace):
    import openpyxl
    path = os.path.join(agent_workspace, "Earnings_Analysis.xlsx")
    record("Earnings_Analysis.xlsx exists", os.path.exists(path), path)
    if not os.path.exists(path):
        return
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    yf_quarters = fetch_yf_quarters()

    # === Earnings Calendar ===
    rows = load_sheet_rows(wb, "Earnings Calendar")
    record("Sheet 'Earnings Calendar' exists", rows is not None,
           f"sheets: {wb.sheetnames}")
    if rows is not None:
        header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
        data = [r for r in rows[1:] if r and r[0]]
        record("Earnings Calendar has 3 rows", len(data) == 3, f"got {len(data)}")

        # Required columns
        for col in ["symbol", "name", "earnings_date", "expected_eps", "historical_avg_eps", "surprise_trend"]:
            record(f"EC has column '{col}'", col in header,
                   f"headers: {header}")

        # Sort alphabetically by symbol
        syms = [str(r[0]).strip().upper() for r in data]
        record("EC sorted alphabetically by symbol",
               syms == sorted(syms), f"order: {syms}")

        idx = {h: i for i, h in enumerate(header)}
        for r in data:
            sym = str(r[0]).strip().upper()
            exp = CALENDAR_EXPECTED.get(sym)
            if exp is None:
                record(f"EC unexpected symbol {sym}", False, "")
                continue
            # Earnings_Date — must start with the exact YYYY-MM-DD prefix (allowing trailing time)
            ci = idx.get("earnings_date")
            if ci is not None and ci < len(r):
                date_str = str(r[ci]).strip()
                # Match: 'YYYY-MM-DD' optionally followed by space+time (no extra digits glued on)
                date_ok = bool(re.match(rf"^{re.escape(exp['date'])}(?:[\sT].*)?$", date_str))
                record(f"{sym} Earnings_Date == {exp['date']}",
                       date_ok,
                       f"got {r[ci]}")
            # Expected_EPS
            ci = idx.get("expected_eps")
            if ci is not None and ci < len(r):
                record(f"{sym} Expected_EPS",
                       num_close(r[ci], exp["expected_eps"], 0.01),
                       f"got {r[ci]}, expected {exp['expected_eps']}")
            # Historical_Avg_EPS
            ci = idx.get("historical_avg_eps")
            qd = yf_quarters.get(sym, [])
            exp_avg = hist_avg_eps(qd)
            if exp_avg is not None and ci is not None and ci < len(r):
                record(f"{sym} Historical_Avg_EPS",
                       num_close(r[ci], exp_avg, 0.01),
                       f"got {r[ci]}, expected {exp_avg}")
            # Surprise_Trend
            ci = idx.get("surprise_trend")
            if ci is not None and ci < len(r):
                exp_tr = expected_trend(exp["expected_eps"], exp_avg)
                actual = str(r[ci] or "").strip().lower()
                record(f"{sym} Surprise_Trend == {exp_tr}",
                       exp_tr is not None and actual == exp_tr,
                       f"got '{actual}', expected '{exp_tr}'")

    # === Financial Trends ===
    rows = load_sheet_rows(wb, "Financial Trends")
    record("Sheet 'Financial Trends' exists", rows is not None,
           f"sheets: {wb.sheetnames}")
    if rows is not None:
        header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
        data = [r for r in rows[1:] if r and r[0]]
        record("Financial Trends has exactly 12 rows", len(data) == 12,
               f"got {len(data)}")
        for col in ["symbol", "period", "diluted_eps", "total_revenue"]:
            record(f"FT has column '{col}'", col in header, f"headers: {header}")

        # Sort: by symbol asc then period desc
        try:
            sym_col = header.index("symbol")
            per_col = header.index("period")
            ordered = list(data)
            sym_seq = [str(r[sym_col]).strip().upper() for r in ordered]
            sym_sorted = sym_seq == sorted(sym_seq)
            record("FT sorted by symbol ascending",
                   sym_sorted, f"order: {sym_seq}")

            # Within each symbol, period must be in descending order
            from collections import OrderedDict
            sym_periods = OrderedDict()
            for r in ordered:
                s = str(r[sym_col]).strip().upper()
                p = str(r[per_col]).strip() if r[per_col] is not None else ""
                sym_periods.setdefault(s, []).append(p)
            within_desc_ok = True
            problems = []
            for s, ps in sym_periods.items():
                # Sort descending — string-sort works for ISO-like 'YYYY-MM-DD' or quarters 'YYYYQn'
                if ps != sorted(ps, reverse=True):
                    within_desc_ok = False
                    problems.append((s, ps))
            record("FT periods within each symbol sorted descending",
                   within_desc_ok,
                   f"problems: {problems[:3]}")
        except Exception:
            pass

        # Per-row EPS / Revenue match
        idx = {h: i for i, h in enumerate(header)}
        for r in data:
            sym = str(r[idx["symbol"]]).strip().upper()
            per = str(r[idx["period"]]).strip()
            qd = yf_quarters.get(sym, [])
            # Find matching quarter
            match = None
            for q in qd:
                if q["period"] is not None and str(q["period"]) in per:
                    match = q
                    break
            if match is None:
                continue
            if "diluted_eps" in idx and idx["diluted_eps"] < len(r):
                record(f"FT {sym}/{per} Diluted_EPS",
                       num_close(r[idx["diluted_eps"]], match["eps"], 0.01),
                       f"got {r[idx['diluted_eps']]}, expected {match['eps']}")
            if "total_revenue" in idx and idx["total_revenue"] < len(r) and match["revenue"] is not None:
                # Total_Revenue: relax abs_tol to better tolerate billion-scale rounding (rel 0.5% or abs 1e6)
                record(f"FT {sym}/{per} Total_Revenue",
                       num_close(r[idx["total_revenue"]], match["revenue"], abs_tol=1e6, rel_tol=0.005),
                       f"got {r[idx['total_revenue']]}, expected {match['revenue']}")

    # === Alert Summary ===
    rows = load_sheet_rows(wb, "Alert Summary")
    record("Sheet 'Alert Summary' exists", rows is not None, f"sheets: {wb.sheetnames}")
    if rows is not None:
        data = [r for r in rows[1:] if r and r[0]]
        lookup = {str(r[0]).strip().lower(): r[1] if len(r) > 1 else None for r in data}
        # Stocks_Reporting
        if "stocks_reporting" in lookup:
            record("Stocks_Reporting == 3",
                   num_close(lookup["stocks_reporting"], 3, 0),
                   f"got {lookup['stocks_reporting']}")
        else:
            record("Alert Summary has 'stocks_reporting'", False, "")

        # Earliest_Report = JPM 2026-04-10
        if "earliest_report" in lookup:
            v = str(lookup["earliest_report"]).strip()
            record("Earliest_Report contains '2026-04-10'",
                   "2026-04-10" in v, f"got {v}")
        else:
            record("Alert Summary has 'earliest_report'", False, "")

        # Latest_Report = AMZN 2026-04-22
        if "latest_report" in lookup:
            v = str(lookup["latest_report"]).strip()
            record("Latest_Report contains '2026-04-22'",
                   "2026-04-22" in v, f"got {v}")
        else:
            record("Alert Summary has 'latest_report'", False, "")

        # Avg_Expected_EPS = (1.95+1.15+4.50)/3 = 2.5333... ≈ 2.53
        exp_avg = round(sum(c["expected_eps"] for c in CALENDAR_EXPECTED.values()) / 3, 2)
        if "avg_expected_eps" in lookup:
            record(f"Avg_Expected_EPS == {exp_avg}",
                   num_close(lookup["avg_expected_eps"], exp_avg, 0.01),
                   f"got {lookup['avg_expected_eps']}")
        else:
            record("Alert Summary has 'avg_expected_eps'", False, "")

        # Stocks_Above_Historical
        above = 0
        for sym, exp in CALENDAR_EXPECTED.items():
            qd = yf_quarters.get(sym, [])
            ha = hist_avg_eps(qd)
            if ha is not None and exp["expected_eps"] > ha:
                above += 1
        if "stocks_above_historical" in lookup:
            record(f"Stocks_Above_Historical == {above}",
                   num_close(lookup["stocks_above_historical"], above, 0),
                   f"got {lookup['stocks_above_historical']}, expected {above}")
        else:
            record("Alert Summary has 'stocks_above_historical'", False, "")


def check_email():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        all_emails = cur.fetchall()
        cur.close()
        conn.close()
        # Strict to + exact subject
        target = None
        for subj, _, to_addr, body in all_emails:
            to_str = json.dumps(to_addr).lower() if isinstance(to_addr, (list, dict)) else str(to_addr or "").lower()
            if (subj or "").strip().lower() == "earnings season alert - q1 2026" and "investment-team@company.com" in to_str:
                target = (subj, to_addr, body)
                break
        record("Email with exact subject 'Earnings Season Alert - Q1 2026' to investment-team@company.com",
               target is not None,
               f"emails: {[(e[0], e[2]) for e in all_emails[:5]]}")
        if target:
            body = (target[2] or "").lower()
            for sym in ["GOOGL", "AMZN", "JPM"]:
                record(f"Body mentions {sym}", sym.lower() in body, "")
            for d in ["2026-04-10", "2026-04-15", "2026-04-22"]:
                record(f"Body mentions date {d}", d in body, "")
    except Exception as e:
        record("Email check error", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    print("  Checking Excel file...")
    check_excel(agent_ws)
    print("  Checking email...")
    check_email()

    all_passed = (FAIL_COUNT == 0)
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "success": all_passed}, f)
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
