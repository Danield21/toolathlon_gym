#!/usr/bin/env python3
"""Evaluation script for learning-outcomes-assessment-workflow."""

import argparse
import os
import sys
from pathlib import Path

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")),
        dbname=os.environ.get("PGDATABASE", "toolathlon_gym"),
        user="eigent", password="camel",
    )


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    if not agent_workspace:
        return False, "No agent workspace provided"

    agent_ws = Path(agent_workspace)
    if not agent_ws.exists():
        return False, f"Agent workspace not found: {agent_workspace}"

    # Detect GT self-test mode (V1 parity test).
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    # Phase 4: Excel learning-outcomes report
    xlsx_files = [p for p in agent_ws.glob("*.xlsx")
                  if 'data.csv' not in p.name.lower()]
    check("At least one xlsx learning-outcomes report exists",
          len(xlsx_files) >= 1, f"found {len(xlsx_files)} xlsx files")
    if xlsx_files:
        try:
            import openpyxl
            audit_path = max(xlsx_files, key=lambda p: p.stat().st_size)
            wb = openpyxl.load_workbook(str(audit_path), data_only=True)
            sheets = [s.lower() for s in wb.sheetnames]
            has_outcome_sheet = any(
                kw in s
                for s in sheets
                for kw in ['outcome', 'achievement', 'mastery', 'objective',
                          'item', 'assessment', 'analysis', 'summary',
                          'demographic', 'performance']
            )
            check(f"Learning-outcomes xlsx ({audit_path.name}) has outcome-related sheet",
                  has_outcome_sheet, f"sheets={wb.sheetnames}")
            # Tightened: at least one sheet should have >=4 rows
            max_rows = max((ws.max_row for ws in wb.worksheets), default=0)
            check("Learning-outcomes xlsx has >=4 rows of data in some sheet",
                  max_rows >= 4, f"max rows={max_rows}")

            # Value-level GT comparison
            gt_xlsx_path = Path(groundtruth_workspace) / "Learning_Outcomes_Report.xlsx"
            if gt_xlsx_path.exists():
                try:
                    gt_wb = openpyxl.load_workbook(str(gt_xlsx_path), data_only=True)
                    # Check column-coverage for learning outcome keywords
                    agent_headers = set()
                    for ws in wb.worksheets:
                        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                        for h in first:
                            if h:
                                agent_headers.add(str(h).strip().lower().replace(' ', '_'))
                    expected_kw = ['mastery', 'objective', 'student', 'assessment',
                                   'difficulty', 'discrimination', 'group',
                                   'learning_objective', 'rate', 'score', 'item']
                    matched_kw = sum(
                        1 for kw in expected_kw
                        if any(kw in h for h in agent_headers)
                    )
                    # Tightened: >=3 matched keywords (was >=2). Catches xlsx that
                    # incidentally has one matching column from a different domain.
                    check("Learning-outcomes xlsx column-coverage (>=3 of expected keywords)",
                          matched_kw >= 3,
                          f"matched {matched_kw}/{len(expected_kw)} headers={list(agent_headers)[:10]}")

                    # Has at least 3 numeric values
                    agent_numerics = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            for v in row:
                                if isinstance(v, (int, float)):
                                    agent_numerics.append(float(v))
                    check("Learning-outcomes xlsx has >=3 numeric metric values",
                          len(agent_numerics) >= 3,
                          f"got {len(agent_numerics)} numerics")
                    # Has at least 1 mastery-rate-style value (0-1 fraction or 0-100 percent)
                    rate_values = [v for v in agent_numerics if 0 <= v <= 100]
                    check("Learning-outcomes xlsx has >=2 plausible rate/percentage values (0-100)",
                          len(rate_values) >= 2,
                          f"got {len(rate_values)} values in [0, 100]")
                    gt_wb.close()
                except Exception as _e:
                    check("Learning-outcomes xlsx GT comparison", False, str(_e))
        except Exception as e:
            check("Learning-outcomes xlsx parse", False, str(e))

    # Phase 4-5: Word recommendations/findings doc
    docx_files = list(agent_ws.glob("*.docx"))
    check("At least one docx recommendations/findings doc exists",
          len(docx_files) >= 1, f"found {len(docx_files)} docx files")
    if docx_files:
        try:
            from docx import Document
            audit_docx = max(docx_files, key=lambda p: p.stat().st_size)
            doc = Document(str(audit_docx))
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word doc has substantive content (>=200 chars)",
                  len(text) >= 200, f"len={len(text)}")
            has_topic = any(kw in text for kw in
                            ['learning', 'outcome', 'objective', 'mastery',
                             'assessment', 'curriculum', 'student'])
            has_recs = any(kw in text for kw in
                           ['recommendation', 'improvement', 'finding',
                            'evidence', 'gap'])
            check("Word doc mentions learning/outcome/curriculum topic",
                  has_topic, f"text head: {text[:200]}")
            check("Word doc mentions recommendations/improvements/findings",
                  has_recs, f"text head: {text[:200]}")
        except Exception as e:
            check("Word doc parse", False, str(e))

    # Phase 2: Survey form created (Google Form)
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, title FROM gform.forms
               WHERE title ILIKE %s OR title ILIKE %s OR title ILIKE %s""",
            ('%learning%', '%course%', '%student%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Learning survey form created: 0 found (GT self-test, non-blocking)")
        else:
            check("Learning survey form created",
                  len(rows) >= 1, f"found {len(rows)} matching forms")
        conn.close()
    except Exception as e:
        check("GForm check", False, str(e))

    # Phase 6: Calendar curriculum review meeting
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT summary FROM gcal.events
               WHERE summary ILIKE %s OR summary ILIKE %s OR summary ILIKE %s""",
            ('%curriculum%', '%review%', '%learning%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Curriculum review meeting scheduled: 0 found (GT self-test, non-blocking)")
        else:
            check("Curriculum review meeting scheduled",
                  len(rows) >= 1, f"found {len(rows)} matching events")
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))

    # Phase 6: Email to faculty/leadership
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, subject FROM email.messages
               WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s
                  OR body_text ILIKE %s""",
            ('%learning%', '%curriculum%', '%assessment%', '%outcome%'))
        rows = cur.fetchall()
        if is_gt_self_test and len(rows) == 0:
            print("  [WARN] Faculty/leadership email sent: 0 found (GT self-test, non-blocking)")
        else:
            check("Faculty/leadership email sent",
                  len(rows) >= 1, f"found {len(rows)} matching emails")
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
