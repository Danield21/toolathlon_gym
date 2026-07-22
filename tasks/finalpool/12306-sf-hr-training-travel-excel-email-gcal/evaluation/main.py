"""
Evaluation for 12306-sf-hr-training-travel-excel-email-gcal.

Checks:
1. Training_Travel_Report.xlsx exists
2. "Employees" sheet has at least 4 data rows with Name/Department columns
3. "Travel_Plan" sheet has at least 8 rows (5 employees x 2 directions)
4. "Budget_Summary" sheet has at least 2 rows with price columns
5. "Budget_Summary" total row exists with value around 5530 (+-500)
6. GCal has at least 2 new events on or around 2026-03-10
7. Email sent to training@hr-dept.com
8. At least 1 additional email sent (to employee addresses)
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Training_Travel_Report.xlsx ===")
    import glob

    pattern = os.path.join(agent_workspace, "*.xlsx")
    all_xlsx = glob.glob(pattern)
    train_files = [f for f in all_xlsx if any(
        kw in os.path.basename(f).lower()
        for kw in ["training", "travel", "report", "hr"]
    )]

    if not train_files:
        record("Training travel xlsx exists", False,
               f"No matching xlsx in {agent_workspace}")
        return
    record("Training travel xlsx exists", True)

    xlsx_path = train_files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Employees sheet
    emp_match = [s for s in sheet_names_lower if "employ" in s or "staff" in s or "people" in s]
    if not emp_match:
        record("Employees sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Employees sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(emp_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Employees has exactly 5 data rows", len(data_rows) == 5,
               f"Found {len(data_rows)} data rows")
        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        # Task says employees from Sales OR Marketing depts (one or both acceptable).
        has_dept = "sales" in all_text or "marketing" in all_text
        record("Employees contains Sales or Marketing department", has_dept,
               f"Text: {all_text[:200]}")

    # Check Travel_Plan sheet
    plan_match = [s for s in sheet_names_lower if "travel" in s or "plan" in s or "train" in s]
    if not plan_match:
        record("Travel_Plan sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Travel_Plan sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(plan_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel_Plan has exactly 10 data rows", len(data_rows) == 10,
               f"Found {len(data_rows)} data rows")

    # Check Budget_Summary sheet
    budget_match = [s for s in sheet_names_lower if "budget" in s or "summar" in s or "cost" in s]
    if not budget_match:
        record("Budget_Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Budget_Summary sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(budget_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Budget_Summary has exactly 3 data rows (outbound, return, total)",
               len(data_rows) == 3, f"Found {len(data_rows)} data rows")

        # Check total appears in the budget. We compare to GT-derived total
        # (2 directions x 5 employees x ticket_price). GT is 553*10=5530 CNY.
        all_vals = []
        for r in rows:
            for c in r:
                try:
                    all_vals.append(float(c))
                except Exception:
                    pass
        # Tighter tolerance: ±50 CNY to allow round-off only.
        has_total = any(5480 <= v <= 5580 for v in all_vals)
        record("Budget total ~5530 CNY (5480-5580)", has_total,
               f"Numeric values found: {sorted(all_vals)}")

    # --- Groundtruth value comparison ---
    # Critical: Task is open-ended ("up to 5 employees" with criteria,
    # "one suitable return train"). The GT pins specific employees and a
    # specific return train (G2), but any valid selection should pass.
    # We validate criteria, not exact GT rows.
    gt_path = os.path.join(groundtruth_workspace, "Training_Travel_Report.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx exists", False, gt_path)
        return

    # Validate Employees sheet by criteria (Sales OR Marketing department,
    # >=3 years experience, Training_Eligible == Yes), NOT by exact names.
    emp_match2 = [s for s in sheet_names_lower if "employ" in s or "staff" in s or "people" in s]
    if emp_match2:
        ws_e = wb[wb.sheetnames[sheet_names_lower.index(emp_match2[0])]]
        rows_e = list(ws_e.iter_rows(values_only=True))
        if rows_e:
            headers_e = [str(c).lower().strip() if c else "" for c in rows_e[0]]
            try:
                dept_idx = next(i for i, h in enumerate(headers_e) if "department" in h or "dept" in h)
            except StopIteration:
                dept_idx = -1
            try:
                yrs_idx = next(i for i, h in enumerate(headers_e) if "year" in h or "experience" in h)
            except StopIteration:
                yrs_idx = -1
            try:
                elig_idx = next(i for i, h in enumerate(headers_e) if "elig" in h or "training" in h and "elig" in h)
            except StopIteration:
                # Try simpler
                elig_idx = next((i for i, h in enumerate(headers_e) if "elig" in h or "training_elig" in h), -1)

            data_rows_e = [r for r in rows_e[1:] if any(c is not None for c in r)]
            criteria_ok = (dept_idx >= 0 and yrs_idx >= 0)
            elig_ok = (elig_idx >= 0)
            # Iterate over all rows; do NOT break early so structural
            # mis-alignment (missing column header) still fails cleanly.
            for r in data_rows_e:
                if dept_idx >= 0 and dept_idx < len(r):
                    d = str(r[dept_idx] or "").lower()
                    if not ("sales" in d or "marketing" in d):
                        criteria_ok = False
                else:
                    criteria_ok = False
                if yrs_idx >= 0 and yrs_idx < len(r):
                    try:
                        yrs = float(r[yrs_idx])
                        if yrs < 3:
                            criteria_ok = False
                    except (TypeError, ValueError):
                        criteria_ok = False
                else:
                    criteria_ok = False
                if elig_idx >= 0 and elig_idx < len(r):
                    e = str(r[elig_idx] or "").strip().lower()
                    if e != "yes":
                        elig_ok = False
                else:
                    elig_ok = False
            record("All Employees in Sales OR Marketing with >=3 yrs experience",
                   criteria_ok,
                   f"Inspected {len(data_rows_e)} employees")
            record("All Employees have Training_Eligible == Yes",
                   elig_ok, "")

    # Validate Travel_Plan: structure - 10 rows, 5 unique employees, each with
    # one Outbound + one Return row, outbound train G11/07:00 (derivable from
    # task spec "earliest departure ... before early afternoon"), prices 553 CNY.
    plan_match2 = [s for s in sheet_names_lower if "travel" in s or "plan" in s]
    if plan_match2:
        ws_p = wb[wb.sheetnames[sheet_names_lower.index(plan_match2[0])]]
        rows_p = list(ws_p.iter_rows(values_only=True))
        if rows_p:
            headers_p = [str(c).lower().strip() if c else "" for c in rows_p[0]]
            try:
                name_idx = next(i for i, h in enumerate(headers_p) if "employee_name" in h or "name" in h)
                train_idx = next(i for i, h in enumerate(headers_p) if "train" in h)
                dir_idx = next(i for i, h in enumerate(headers_p) if "direction" in h or "type" in h)
                price_idx = next(i for i, h in enumerate(headers_p) if "price" in h or "cny" in h)
            except StopIteration:
                name_idx = train_idx = dir_idx = price_idx = -1

            data_rows_p = [r for r in rows_p[1:] if any(c is not None for c in r)]
            unique_names = set()
            outbound_count = 0
            return_count = 0
            outbound_train_g11 = True
            prices_ok = True
            return_trains = []
            for r in data_rows_p:
                if name_idx >= 0 and name_idx < len(r):
                    unique_names.add(str(r[name_idx] or "").strip())
                # Tighten direction matching: use word-boundary tokens, NOT
                # loose substring "in". Outbound: starts with "out", "depart",
                # contains "to shanghai". Return: starts with "return", "back",
                # contains "to beijing", or "inbound" (full token, not "in").
                if dir_idx >= 0 and dir_idx < len(r):
                    direction = str(r[dir_idx] or "").lower().strip()
                    is_outbound = (
                        direction.startswith("out")
                        or direction.startswith("depart")
                        or "to shanghai" in direction
                    )
                    is_return = (
                        direction.startswith("return")
                        or direction.startswith("back")
                        or direction.startswith("inbound")
                        or "to beijing" in direction
                    )
                    if is_outbound and not is_return:
                        outbound_count += 1
                        if train_idx >= 0 and train_idx < len(r):
                            tn = str(r[train_idx] or "").upper().strip()
                            if "G11" not in tn:
                                outbound_train_g11 = False
                    elif is_return and not is_outbound:
                        return_count += 1
                        if train_idx >= 0 and train_idx < len(r):
                            return_trains.append(str(r[train_idx] or "").upper().strip())
                if price_idx >= 0 and price_idx < len(r):
                    try:
                        p = float(r[price_idx])
                        # Second class price is 553 CNY (per GT, derivable from 12306)
                        if abs(p - 553) > 5:
                            prices_ok = False
                    except (TypeError, ValueError):
                        prices_ok = False

            record("Travel_Plan has 5 unique employees", len(unique_names) == 5,
                   f"unique names: {len(unique_names)}")
            record("Travel_Plan has 5 outbound rows", outbound_count == 5,
                   f"outbound rows: {outbound_count}")
            record("Travel_Plan has 5 return rows", return_count == 5,
                   f"return rows: {return_count}")
            record("All outbound rows use train G11", outbound_train_g11,
                   "Outbound is the earliest train (G11 07:00 per 12306 query)")
            # Return train: enforce all return rows use the SAME train number
            # (any valid Shanghai->Beijing train, but consistent across all 5 employees).
            return_consistent = (
                len(return_trains) == 5
                and len({t for t in return_trains if t}) == 1
                and all(t.startswith("G") for t in return_trains)
            )
            record("All return rows use the same return train (consistent G-series train)",
                   return_consistent,
                   f"return trains: {return_trains}")
            record("All ticket prices ~553 CNY (second class Beijing-Shanghai)",
                   prices_ok, "")

    # Validate Budget_Summary: total ~5530 already checked above; further verify
    # 3 rows: outbound, return, total; total_cny == count * unit_price.
    budget_match2 = [s for s in sheet_names_lower if "budget" in s or "summar" in s]
    if budget_match2:
        ws_b = wb[wb.sheetnames[sheet_names_lower.index(budget_match2[0])]]
        rows_b = list(ws_b.iter_rows(values_only=True))
        if rows_b:
            headers_b = [str(c).lower().strip() if c else "" for c in rows_b[0]]
            try:
                cnt_idx = next(i for i, h in enumerate(headers_b) if h == "count" or "count" in h)
                up_idx = next(i for i, h in enumerate(headers_b) if "unit" in h or "unit_price" in h)
                total_idx = next(i for i, h in enumerate(headers_b) if h == "total" or "total_cny" in h or "total" in h)
            except StopIteration:
                cnt_idx = up_idx = total_idx = -1

            data_rows_b = [r for r in rows_b[1:] if any(c is not None for c in r)]
            arith_ok = True
            for r in data_rows_b:
                try:
                    c = float(r[cnt_idx]) if cnt_idx >= 0 else None
                    u = float(r[up_idx]) if up_idx >= 0 else None
                    t = float(r[total_idx]) if total_idx >= 0 else None
                    if c is not None and u is not None and t is not None:
                        if abs(t - c * u) > 1.0:
                            arith_ok = False
                            break
                except (TypeError, ValueError):
                    pass
            record("Budget_Summary arithmetic Total = Count * Unit_Price",
                   arith_ok, "")

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    gt_wb.close()


def check_gcal():
    print("\n=== Check 2: Calendar events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-10'
          AND start_datetime < '2026-03-11'
          AND summary NOT ILIKE '%training program kickoff%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()
    summaries = [(e[0] or "").lower() for e in events]
    has_depart = any(("depart" in s or "outbound" in s) and "shanghai" in s for s in summaries)
    has_return = any(("return" in s or "inbound" in s) and "beijing" in s for s in summaries)
    record(
        "Calendar event 'Corporate Training - Depart for Shanghai' on Mar 10",
        has_depart,
        f"Mar10 events: {[e[0] for e in events]}",
    )
    record(
        "Calendar event 'Corporate Training - Return to Beijing' on Mar 10",
        has_return,
        f"Mar10 events: {[e[0] for e in events]}",
    )


def _count_email(to_pattern, exclude_from_pattern=None):
    cnt = 0
    sent = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        if exclude_from_pattern:
            cur.execute(
                "SELECT COUNT(*) FROM email.messages"
                " WHERE to_addr::text ILIKE %s AND from_addr NOT ILIKE %s",
                (f"%{to_pattern}%", f"%{exclude_from_pattern}%"),
            )
        else:
            cur.execute(
                "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s",
                (f"%{to_pattern}%",),
            )
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
    except Exception:
        pass
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COUNT(*) FROM email.sent_log WHERE to_addr::text ILIKE %s",
            (f"%{to_pattern}%",),
        )
        sent = cur2.fetchone()[0]
        cur2.close()
        conn2.close()
    except Exception:
        pass
    return cnt, sent


def check_emails():
    print("\n=== Check 3: Emails ===")

    hr_cnt, hr_sent = _count_email("training@hr-dept.com", "training@hr-dept.com")
    record("Email sent to training@hr-dept.com", hr_cnt >= 1 or hr_sent >= 1,
           f"messages: {hr_cnt}, sent_log: {hr_sent}")

    # At least 1 additional email to a corporate-style employee address.
    # Tighten: require employee email pattern (containing @company.com,
    # @corp.com, or similar) AND outgoing (from_addr NOT employee).
    emp_cnt = 0
    emp_sent = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM email.messages"
            " WHERE to_addr::text NOT ILIKE '%training@hr-dept.com%'"
            "   AND to_addr::text NOT ILIKE '%hr-dept.com%'"
            "   AND from_addr NOT ILIKE '%training@hr-dept.com%'"
            "   AND ("
            "     to_addr::text ILIKE '%@company.com%'"
            "     OR to_addr::text ILIKE '%@corp.%'"
            "     OR to_addr::text ILIKE '%@example.%'"
            "     OR to_addr::text ~* '@[a-z0-9.-]+\\.(com|org|co|edu|net)'"
            "   )"
        )
        emp_cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
    except Exception:
        pass
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COUNT(*) FROM email.sent_log"
            " WHERE to_addr::text NOT ILIKE '%training@hr-dept.com%'"
            "   AND to_addr::text NOT ILIKE '%hr-dept.com%'"
            "   AND ("
            "     to_addr::text ILIKE '%@company.com%'"
            "     OR to_addr::text ILIKE '%@corp.%'"
            "     OR to_addr::text ILIKE '%@example.%'"
            "     OR to_addr::text ~* '@[a-z0-9.-]+\\.(com|org|co|edu|net)'"
            "   )"
        )
        emp_sent = cur2.fetchone()[0]
        cur2.close()
        conn2.close()
    except Exception:
        pass
    record("At least 1 additional email to a non-HR corporate address (employee notification)",
           emp_cnt >= 1 or emp_sent >= 1,
           f"messages: {emp_cnt}, sent_log: {emp_sent}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0 and PASS_COUNT > 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
