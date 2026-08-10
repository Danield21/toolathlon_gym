"""
Evaluation for yt-12306-tech-conference-travel-notion-excel task.

Checks:
1. Tech_Conference_Plan.xlsx exists and is readable
2. Prep_Videos sheet: Title + View_Count columns, >= 5 data rows, GT entries
   matched order-insensitively on Rank / Title (fuzzy) / View_Count (numeric tolerance)
3. Travel_Details sheet: Train_No column, >= 2 data rows, valid outbound/return
   train codes (DB-derived), GT entries matched on Departure / Arrival / Price_CNY
   plus Train_No on its extracted short code (a correct agent may record either the
   short code "G235" or the MCP's full train_no "G235_260312_1"; both are accepted).
   Station names and seat class are intentionally NOT pinned, because the rail system
   returns them in Chinese and the task allows the agent to record them as returned)
4. Conference_Schedule sheet: >= 6 data rows, covers >= 3 of the four required
   dates (03-12 / 03-13 / 03-14 / 03-15) -- the schedule itself is agent-designed,
   so no exact per-row GT comparison is performed
5. Notion page titled 'Tech Conference Preparation - Qufu March 2026' exists
6. GCal has the 4 required events (title + start/end time, +/-8h timezone tolerance)
"""
import json
import os
import re
import sys
from argparse import ArgumentParser
from datetime import date, datetime, time, timedelta

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0

FORMULA_MARKER = object()


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


# ---------------------------------------------------------------------------
# Cell / column helpers for the (agent-vs-GT) Excel comparison
# ---------------------------------------------------------------------------
def _norm_cell(v):
    """Robust cell normalizer.

    Returns None, a float, a normalized lowercase string, or FORMULA_MARKER
    (when the agent wrote an Excel formula we cannot evaluate).
    - datetime / time -> 'HH:MM' (or 'HH:MM:SS' when seconds are non-zero)
    - date -> ISO date string
    - numbers -> float
    - strings: time-like -> 'HH:MM'; numeric-like (after stripping separators,
      currency symbols, % and spaces) -> float; otherwise lowercase string.
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        t = v.time()
    elif isinstance(v, time):
        t = v
    elif isinstance(v, date):
        return v.isoformat()
    else:
        t = None
    if t is not None:
        if t.second == 0 and t.microsecond == 0:
            return t.strftime("%H:%M")
        return t.strftime("%H:%M:%S")
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("="):
        return FORMULA_MARKER
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        sec = int(m.group(3)) if m.group(3) else 0
        return f"{h:02d}:{mi:02d}" if sec == 0 else f"{h:02d}:{mi:02d}:{sec:02d}"
    try:
        cleaned = s
        for ch in (",", "¥", "￥", "$", "€", "%", " "):
            cleaned = cleaned.replace(ch, "")
        return float(cleaned)
    except Exception:
        return s.lower()


def _sheet_headers(ws):
    row = next(ws.iter_rows(max_row=1), None)
    if row is None:
        return []
    return [str(c.value).strip().lower() if c.value else "" for c in row]


def _find_col(headers, gt_header):
    """Locate an agent column matching a GT header (exact, then contains, then token)."""
    g = gt_header.strip().lower()
    for i, h in enumerate(headers):
        if h == g:
            return i
    for i, h in enumerate(headers):
        if h and (g in h or h in g):
            return i
    gt_tokens = set(re.findall(r"[a-z0-9]+", g))
    best, best_score = None, 0
    for i, h in enumerate(headers):
        ht = set(re.findall(r"[a-z0-9]+", h))
        if gt_tokens and ht:
            score = len(gt_tokens & ht)
            if score > best_score:
                best, best_score = i, score
    return best if best_score >= 1 else None


def _col_category(h):
    """Classify a GT column for comparison.

    - 'skip': free-form / language-sensitive / agent-discretionary columns
    - 'time': departure / arrival / time columns (normalized to HH:MM)
    - 'num': rank / view_count / price (numeric with tolerance; rank exact)
    - 'title': video titles (fuzzy match)
    - 'exact': everything else (normalized string equality)
    """
    h = h.lower()
    if any(k in h for k in (
        "topic", "key_tech", "key tech", "technolog",
        "activity", "location", "note", "remark", "comment", "description",
        "reach", "leg", "station", "seat", "duration", "date",
    )):
        return "skip"
    # Train numbers: the rail MCP reports both a short code (station_train_code,
    # e.g. "G235") and a full train_no ("G235_260312_1"). A correct agent may
    # record either form, so compare on the extracted code rather than exact text.
    if "train" in h:
        return "train"
    if any(k in h for k in ("depart", "arriv", "time")):
        return "time"
    if any(k in h for k in ("view", "count", "price", "rank", "num")):
        return "num"
    if any(k in h for k in ("title", "video")):
        return "title"
    return "exact"


def _title_match(g, a):
    g, a = str(g).strip(), str(a).strip()
    if not g or not a:
        return False
    if g == a:
        return True
    if g in a or a in g:
        return True
    sig = lambda s: re.sub(r"[^a-z0-9]", "", s.lower())[:20]
    sg, sa = sig(g), sig(a)
    if sg and sa and sg == sa:
        return True
    toks_g = set(re.findall(r"[a-z0-9]+", g.lower()))
    toks_a = set(re.findall(r"[a-z0-9]+", a.lower()))
    if toks_g and toks_a:
        inter = toks_g & toks_a
        if len(inter) >= 3 and len(inter) >= 0.5 * max(len(toks_g), len(toks_a)):
            return True
    return False


def _coerce_number(v):
    """Best-effort numeric coercion tolerating unit suffixes / abbreviations.

    Returns a float when v can be read as a number -- possibly wrapped in a unit
    word ("3,878,491 views", "174.5元"), a K/M/B abbreviation ("3.8M"), or a
    short currency label ("CNY174.5") -- otherwise None. Used so a numeric cell
    written with a unit is compared by value, not by raw string equality.
    """
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    t = s
    for ch in (",", "¥", "￥", "$", "€", "%", " "):
        t = t.replace(ch, "")
    if not t:
        return None
    try:
        return float(t)
    except Exception:
        pass
    # Pure K/M/B suffix: "3.8M", "12k"
    m = re.fullmatch(r"([\d.]+)\s*([kKmMbB])", t)
    if m:
        mult = {"k": 1e3, "m": 1e6, "b": 1e9}[m.group(2).lower()]
        try:
            return float(m.group(1)) * mult
        except Exception:
            return None
    # Leading number (optionally with a K/M/B multiplier) before a unit word:
    # "3,878,491 views", "174.5元", "3.8M views". Requires a trailing unit so a
    # bare number is never mis-read here.
    m = re.match(r"^([\d.]+)\s*([kKmMbB])?([A-Za-z\u4e00-\u9fff]+)$", t)
    if m:
        base = float(m.group(1))
        if m.group(2):
            mult = {"k": 1e3, "m": 1e6, "b": 1e9}[m.group(2).lower()]
            base *= mult
        try:
            return base
        except Exception:
            return None
    # Short currency label prefix then a number: "CNY174.5", "RMB388"
    m = re.match(r"^(?:[A-Za-z\u4e00-\u9fff]{1,4})?([\d.]+)$", t)
    if m:
        try:
            return float(m.group(1))
        except Exception:
            return None
    return None


def _num_ok(h, g, a):
    if g is None or a is None:
        return g is None and a is None
    gn, an = _coerce_number(g), _coerce_number(a)
    if gn is not None and an is not None:
        if "rank" in h.lower():
            return gn == an
        return abs(gn - an) <= max(0.5, 0.01 * abs(gn))
    return str(g).strip().lower() == str(a).strip().lower()


def _train_code(v):
    """Extract the short train code (e.g. 'G235') from any of its forms:
    short code 'G235', full MCP train_no 'G235_260312_1', or annotated text."""
    if v is None:
        return ""
    s = str(v).strip()
    m = re.match(r"^\s*([A-Za-z0-9]+)", s)
    if m:
        return m.group(1).upper()
    m = re.search(r"\b([A-Za-z]+\d{2,5})\b", s)
    if m:
        return m.group(1).upper()
    return re.sub(r"[^A-Za-z0-9]", "", s).upper()


def _train_codes_equal(g, a):
    return bool(_train_code(g)) and _train_code(g) == _train_code(a)


def _col_match(cat, h, g, a):
    # If the agent wrote an Excel formula we cannot evaluate its cached value,
    # treat that column as a structural-only (lenient) pass.
    if a is FORMULA_MARKER:
        return True
    if g is None or a is None:
        return g is None and a is None
    if cat == "num":
        return _num_ok(h, g, a)
    if cat == "train":
        return _train_codes_equal(g, a)
    if cat == "time":
        return g == a
    if cat == "title":
        return _title_match(g, a)
    return g == a


def _find_sheet(wb, gt_name):
    g = gt_name.strip().lower()
    for sn in wb.sheetnames:
        if sn.strip().lower() == g:
            return wb[sn]
    if "prep" in g or "video" in g:
        for sn in wb.sheetnames:
            s = sn.lower()
            if "prep" in s or "video" in s:
                return wb[sn]
    elif "travel" in g:
        for sn in wb.sheetnames:
            if "travel" in sn.lower():
                return wb[sn]
    elif "sched" in g or "conference" in g:
        for sn in wb.sheetnames:
            s = sn.lower()
            if "sched" in s or "conference" in s:
                return wb[sn]
    return None


MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
          "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def _parse_month_day(v):
    """Parse (month, day) from a date-like cell value, or None.

    Handles datetime/date objects, "2026-03-12" / "03-12" / "3/13" /
    "March 13" strings. Digit boundaries are anchored so "2026" is not
    mistaken for a month.
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return (v.month, v.day)
    if isinstance(v, date):
        return (v.month, v.day)
    s = str(v).strip().lower()
    # "2026-03-12" / "2026/03/12" (year first)
    m = re.search(r"(?<!\d)(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?!\d)", s)
    if m:
        mo, da = int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= da <= 31:
            return (mo, da)
    # "03-12" / "3/13" (month first)
    m = re.search(r"(?<![\d/])(\d{1,2})[-/](\d{1,2})(?!\d)", s)
    if m:
        mo, da = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12 and 1 <= da <= 31:
            return (mo, da)
    # "march 13" / "mar 13"
    m = re.search(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s+(\d{1,2})\b", s)
    if m:
        mo = MONTHS[m.group(1)[:3]]
        da = int(m.group(2))
        if 1 <= da <= 31:
            return (mo, da)
    return None


def _check_schedule_dates(a_ws, a_headers):
    date_col = _find_col(a_headers, "date")
    if date_col is None:
        record("Conference_Schedule covers required dates", False,
               "No Date column found in agent sheet")
        return
    dates_found = set()
    for r in a_ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in r):
            continue
        v = r[date_col] if date_col < len(r) else None
        md = _parse_month_day(v)
        if md:
            dates_found.add(md)
    expected = {(3, 12), (3, 13), (3, 14), (3, 15)}
    covered = dates_found & expected
    record(
        "Conference_Schedule covers >=3 of required dates (03-12..03-15)",
        len(covered) >= 3,
        f"Expected dates {sorted(expected)}, dates found: {sorted(dates_found)}",
    )


# ---------------------------------------------------------------------------
# Check 1: Excel workbook
# ---------------------------------------------------------------------------
def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Tech_Conference_Plan.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Tech_Conference_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Conference_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Tech_Conference_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    # --- Prep_Videos sheet ---
    prep_sheet = None
    for name in wb.sheetnames:
        if "prep" in name.lower() or "video" in name.lower():
            prep_sheet = wb[name]
            break
    if prep_sheet is None:
        record("Prep_Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Prep_Videos sheet exists", True)
        rows = list(prep_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_title = any("title" in h for h in headers)
        has_viewcount = any("view" in h or "count" in h for h in headers)
        record("Prep_Videos has Title and View_Count columns", has_title and has_viewcount,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Prep_Videos has >= 5 data rows", len(data_rows) >= 5,
               f"Found {len(data_rows)} data rows")

    # --- Travel_Details sheet ---
    travel_sheet = None
    for name in wb.sheetnames:
        if "travel" in name.lower():
            travel_sheet = wb[name]
            break
    if travel_sheet is None:
        record("Travel_Details sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Travel_Details sheet exists", True)
        rows = list(travel_sheet.iter_rows(values_only=True))
        all_text = " ".join(str(c) for r in rows for c in r if c).upper()
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_train = any("train" in h or "no" in h for h in headers)
        record("Travel_Details has Train_No column", has_train, f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel_Details has >= 2 data rows", len(data_rows) >= 2,
               f"Found {len(data_rows)} rows")
        # Dynamic check: query the actual train DB for valid Beijing South<->Qufu
        # East trains on the task dates rather than hard-coding train codes.
        try:
            tconn = psycopg2.connect(**DB_CONFIG)
            tcur = tconn.cursor()
            tcur.execute("""
                SELECT station_train_code FROM train.trains
                WHERE from_station_telecode='VNP' AND to_station_telecode='QFB'
                  AND depart_date='2026-03-12'
            """)
            outbound_codes = {r[0] for r in tcur.fetchall()}
            tcur.execute("""
                SELECT station_train_code FROM train.trains
                WHERE from_station_telecode='QFB' AND to_station_telecode='VNP'
                  AND depart_date='2026-03-15'
            """)
            return_codes = {r[0] for r in tcur.fetchall()}
            tcur.close()
            tconn.close()
        except Exception as e:
            outbound_codes, return_codes = set(), set()
            print(f"  [WARN] Train DB lookup failed: {e}")
        if outbound_codes:
            ok_out = any(code in all_text for code in outbound_codes)
            record(
                f"Travel_Details outbound train code present",
                ok_out,
                f"Expected one of {sorted(outbound_codes)} in Travel_Details text",
            )
        if return_codes:
            ok_ret = any(code in all_text for code in return_codes)
            record(
                f"Travel_Details return train code present",
                ok_ret,
                f"Expected one of {sorted(return_codes)} in Travel_Details text",
            )

    # --- Conference_Schedule sheet ---
    sched_sheet = None
    for name in wb.sheetnames:
        if "schedule" in name.lower() or "conference" in name.lower():
            sched_sheet = wb[name]
            break
    if sched_sheet is None:
        record("Conference_Schedule sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Conference_Schedule sheet exists", True)
        rows = list(sched_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Conference_Schedule has >= 6 data rows", len(data_rows) >= 6,
               f"Found {len(data_rows)} rows")

    # --- Groundtruth XLSX value comparison (order-insensitive, column-aware) ---
    gt_path = os.path.join(groundtruth_workspace, "Tech_Conference_Plan.xlsx")
    if not os.path.isfile(gt_path):
        return
    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    try:
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = _find_sheet(wb, gt_sname)
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                       f"Available: {wb.sheetnames}")
                continue
            gt_headers = _sheet_headers(gt_ws)
            a_headers = _sheet_headers(a_ws)
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                       if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count (agent >= GT)",
                   len(a_rows) >= len(gt_rows),
                   f"Expected >= {len(gt_rows)}, got {len(a_rows)}")
            is_schedule = "sched" in gt_sname.lower() or "conference" in gt_sname.lower()
            if is_schedule:
                # The schedule is agent-designed; only verify required-date coverage.
                _check_schedule_dates(a_ws, a_headers)
                continue
            # Build the list of compared columns: (gt_idx, agent_idx, category, header)
            cols = []
            for gi, gh in enumerate(gt_headers):
                cat = _col_category(gh)
                if cat == "skip":
                    continue
                ai = _find_col(a_headers, gh)
                if ai is None:
                    continue
                cols.append((gi, ai, cat, gh))
            for gt_row in gt_rows:
                gt_cells = {}
                for gi, _, _, _ in cols:
                    if gi < len(gt_row):
                        gt_cells[gi] = _norm_cell(gt_row[gi])

                def _agent_row_matches(ar, _cols=cols, _gt_cells=gt_cells):
                    for gi, ai, cat, gh in _cols:
                        gv = _gt_cells.get(gi)
                        av = _norm_cell(ar[ai]) if ai < len(ar) else None
                        if not _col_match(cat, gh, gv, av):
                            return False
                    return True

                found = any(_agent_row_matches(ar) for ar in a_rows)
                first_val = next((v for v in gt_row if v is not None), None)
                gt_key = [
                    (gh, _norm_cell(gt_row[gi]))
                    for gi, _, _, gh in cols if gi < len(gt_row)
                ]
                record(f"GT '{gt_sname}' entry '{first_val}' found (order-insensitive)",
                       found, f"gt_key={gt_key}")
    finally:
        gt_wb.close()


# ---------------------------------------------------------------------------
# Check 2: Notion page
# ---------------------------------------------------------------------------
def check_notion():
    print("\n=== Check 2: Notion Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Notion page titled 'Tech Conference Preparation - Qufu March 2026' exists",
               False, f"DB connection failed: {e}")
        return
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()
        cur.close()
    except Exception as e:
        record("Notion page titled 'Tech Conference Preparation - Qufu March 2026' exists",
               False, f"DB query failed: {e}")
        return
    finally:
        conn.close()

    # Task title: "Tech Conference Preparation - Qufu March 2026".
    # Search for the discriminating keywords (conference + Qufu + March 2026)
    # so a correct agent that phrases the title slightly differently is not
    # falsely failed; preprocess clears notion.pages, so only agent-created
    # pages are considered.
    found_exact = False
    for page_id, props in pages:
        try:
            props_str = json.dumps(props).lower()
            if (
                "conference" in props_str
                and "qufu" in props_str
                and "march 2026" in props_str
            ):
                found_exact = True
                break
        except Exception:
            continue

    record(
        "Notion page titled 'Tech Conference Preparation - Qufu March 2026' exists",
        found_exact,
        f"Pages found: {len(pages)}",
    )


# ---------------------------------------------------------------------------
# Check 3: GCal events
# ---------------------------------------------------------------------------
def check_gcal():
    print("\n=== Check 3: GCal Conference Events ===")
    import datetime as _dt
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("GCal required conference events exist", False,
               f"DB connection failed: {e}")
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime FROM gcal.events
            WHERE start_datetime >= '2026-03-12' AND start_datetime < '2026-03-16'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()
        cur.close()
    except Exception as e:
        record("GCal required conference events exist", False,
               f"DB query failed: {e}")
        return
    finally:
        conn.close()

    def _to_naive(d):
        if d is None:
            return None
        if hasattr(d, "tzinfo") and d.tzinfo is not None:
            return d.astimezone(_dt.timezone.utc).replace(tzinfo=None)
        return d

    def _norm_title(s):
        # Keep only alphanumeric / CJK tokens so punctuation variants
        # ("Conference Travel: Beijing to Qufu" vs "...- Beijing to Qufu")
        # compare equal.
        return " ".join(re.findall(r"[a-z0-9一-鿿]+", str(s).lower()))

    def _gcal_title_matches(exp, summary):
        if not summary:
            return False
        e, s = _norm_title(exp), _norm_title(summary)
        if not e or not s:
            return False
        return e == s or e in s or s in e

    # Required 4 events with exact titles + times
    expected = [
        (
            "Conference Travel: Beijing to Qufu",
            _dt.datetime(2026, 3, 12, 17, 0),
            _dt.datetime(2026, 3, 12, 20, 0),
        ),
        (
            "Tech Conference Day 1",
            _dt.datetime(2026, 3, 13, 9, 0),
            _dt.datetime(2026, 3, 13, 18, 0),
        ),
        (
            "Tech Conference Day 2",
            _dt.datetime(2026, 3, 14, 9, 0),
            _dt.datetime(2026, 3, 14, 18, 0),
        ),
        (
            "Return: Qufu to Beijing",
            _dt.datetime(2026, 3, 15, 14, 30),
            _dt.datetime(2026, 3, 15, 17, 30),
        ),
    ]

    for exp_title, exp_start, exp_end in expected:
        match = None
        for summary, sdt, edt in events:
            if _gcal_title_matches(exp_title, summary):
                match = (summary, _to_naive(sdt), _to_naive(edt))
                break
        record(
            f"GCal event '{exp_title}' exists with correct title",
            match is not None,
            f"All summaries: {[e[0] for e in events]}",
        )
        if match is not None:
            _, sdt, edt = match
            # Allow timezone interpretations: accept expected, expected-8h, or
            # expected+8h (Beijing vs UTC storage), within 60s tolerance.
            ok_start = sdt is not None and (
                abs((sdt - exp_start).total_seconds()) <= 60
                or abs((sdt - (exp_start - _dt.timedelta(hours=8))).total_seconds()) <= 60
                or abs((sdt - (exp_start + _dt.timedelta(hours=8))).total_seconds()) <= 60
            )
            ok_end = edt is not None and (
                abs((edt - exp_end).total_seconds()) <= 60
                or abs((edt - (exp_end - _dt.timedelta(hours=8))).total_seconds()) <= 60
                or abs((edt - (exp_end + _dt.timedelta(hours=8))).total_seconds()) <= 60
            )
            record(
                f"GCal event '{exp_title}' start={exp_start.strftime('%H:%M')}",
                ok_start,
                f"got {sdt}",
            )
            record(
                f"GCal event '{exp_title}' end={exp_end.strftime('%H:%M')}",
                ok_end,
                f"got {edt}",
            )


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_notion()
    check_gcal()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if FAIL_COUNT == 0:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
