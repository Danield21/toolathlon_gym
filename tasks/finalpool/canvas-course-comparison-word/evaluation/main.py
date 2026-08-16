"""Evaluation for canvas-course-comparison-word."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")), "dbname": "toolathlon_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected course comparison data from DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        WITH course_data AS (
            SELECT
                SPLIT_PART(c.course_code, '-', 1) as prefix,
                c.course_code,
                c.name,
                c.total_students,
                COUNT(DISTINCT a.id) as assignment_count,
                ROUND(AVG(s.score)::numeric, 2) as avg_grade
            FROM canvas.courses c
            LEFT JOIN canvas.assignments a ON a.course_id = c.id
            LEFT JOIN canvas.submissions s ON s.assignment_id = a.id AND s.score IS NOT NULL
            WHERE c.course_code LIKE '%%2013J' OR c.course_code LIKE '%%2014J'
            GROUP BY c.id, c.course_code, c.name, c.total_students
        )
        SELECT
            f13.prefix,
            f13.name as name_2013,
            f13.total_students as enroll_2013,
            f14.total_students as enroll_2014,
            f13.assignment_count as assign_2013,
            f14.assignment_count as assign_2014,
            f13.avg_grade as grade_2013,
            f14.avg_grade as grade_2014
        FROM course_data f13
        JOIN course_data f14 ON f13.prefix = f14.prefix
        WHERE f13.course_code LIKE '%%2013J' AND f14.course_code LIKE '%%2014J'
        ORDER BY f13.prefix
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Year_Over_Year_Comparison.xlsx (compare to GT file directly)."""
    print("\n=== Checking Year_Over_Year_Comparison.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Year_Over_Year_Comparison.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Year_Over_Year_Comparison.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True) if os.path.isfile(gt_file) else None
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True

    # Build expected from GT directly (preferred) so all 9 columns are validated.
    expected_rows = []
    if gt_wb is not None:
        for s in gt_wb.sheetnames:
            if "comparison" in s.lower() or "course" in s.lower():
                gt_ws = gt_wb[s]
                gt_rows_full = list(gt_ws.iter_rows(min_row=2, values_only=True))
                expected_rows = [r for r in gt_rows_full if r and r[0]]
                break

    # Check Course Comparison sheet
    comp_sheet = None
    for name in wb.sheetnames:
        if "comparison" in name.lower() or "course" in name.lower():
            comp_sheet = wb[name]
            break
    if comp_sheet is None:
        record("Sheet 'Course Comparison' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Course Comparison' exists", True)
        rows = list(comp_sheet.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if r and r[0]]
        record("Course Comparison has correct row count",
               len(rows) == len(expected_rows),
               f"Expected {len(expected_rows)}, got {len(rows)}")

        agent_lookup = {}
        for r in rows:
            key = str(r[0]).strip().lower()
            agent_lookup[key] = r

        for exp_row in expected_rows:
            course_name = exp_row[0]
            key = str(course_name).strip().lower()
            matched = agent_lookup.get(key)
            if matched is None:
                # Try fuzzy: split on '(' to drop year suffix
                base = key.split("(")[0].strip()
                for k, r in agent_lookup.items():
                    if base in k or k in base:
                        matched = r
                        break

            if matched is None:
                record(f"Course '{course_name}' found", False, "Missing row")
                all_ok = False
                continue

            # Validate ALL 9 columns:
            #  0: Course_Name (string)
            #  1: Fall_2013_Enrollment (int, exact)
            #  2: Fall_2014_Enrollment (int, exact)
            #  3: Enrollment_Change (int, exact)
            #  4: Fall_2013_Assignments (int, exact)
            #  5: Fall_2014_Assignments (int, exact)
            #  6: Fall_2013_Avg_Grade (float, tol 0.05)
            #  7: Fall_2014_Avg_Grade (float, tol 0.05)
            #  8: Grade_Change (float, tol 0.05)
            int_cols = [(1, "Fall_2013_Enrollment"),
                        (2, "Fall_2014_Enrollment"),
                        (3, "Enrollment_Change"),
                        (4, "Fall_2013_Assignments"),
                        (5, "Fall_2014_Assignments")]
            for idx, label in int_cols:
                if idx < len(matched) and idx < len(exp_row):
                    ok = num_close(matched[idx], exp_row[idx], 0)
                    record(f"'{course_name}' {label}", ok,
                           f"Expected {exp_row[idx]}, got {matched[idx]}")
                    if not ok:
                        all_ok = False

            float_cols = [(6, "Fall_2013_Avg_Grade"),
                          (7, "Fall_2014_Avg_Grade"),
                          (8, "Grade_Change")]
            for idx, label in float_cols:
                if idx < len(matched) and idx < len(exp_row):
                    ok = num_close(matched[idx], exp_row[idx], 0.05)
                    record(f"'{course_name}' {label}", ok,
                           f"Expected {exp_row[idx]}, got {matched[idx]}")
                    if not ok:
                        all_ok = False

    # Check Summary sheet (now validates Avg_Enrollment_Change / Avg_Grade_Change)
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break

    expected_summary = {}
    if gt_wb is not None:
        for s in gt_wb.sheetnames:
            if "summary" in s.lower():
                for r in gt_wb[s].iter_rows(min_row=2, values_only=True):
                    if r and r[0] is not None:
                        expected_summary[str(r[0]).strip().lower()] = r[1]
                break

    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        agent_summary = {}
        for r in sum_sheet.iter_rows(min_row=2, values_only=True):
            if r and r[0] is not None:
                agent_summary[str(r[0]).strip().lower()] = r[1]
        for k, exp_v in expected_summary.items():
            agent_v = agent_summary.get(k)
            if agent_v is None:
                record(f"Summary '{k}'", False, "Missing")
                all_ok = False
                continue
            try:
                exp_f = float(exp_v)
                # Counts (e.g. Courses_Compared) get tol=0; otherwise 0.1
                tol = 0 if float(exp_f).is_integer() else 0.1
                ok = num_close(agent_v, exp_v, tol)
            except (TypeError, ValueError):
                ok = str(agent_v).strip().lower() == str(exp_v).strip().lower()
            record(f"Summary '{k}'", ok,
                   f"Expected {exp_v}, got {agent_v}")
            if not ok:
                all_ok = False

    return all_ok


def check_word(agent_workspace, groundtruth_workspace):
    """Check Academic_Year_Comparison.docx with per-course paragraph validation."""
    print("\n=== Checking Academic_Year_Comparison.docx ===")
    from docx import Document

    docx_file = os.path.join(agent_workspace, "Academic_Year_Comparison.docx")
    if not os.path.isfile(docx_file):
        record("Word file exists", False, f"Not found: {docx_file}")
        return False
    record("Word file exists", True)

    try:
        doc = Document(docx_file)
    except Exception as e:
        record("Word readable", False, str(e))
        return False

    all_text = " ".join(p.text.lower() for p in doc.paragraphs)

    # Title check
    record("Word has title 'Fall 2013 vs Fall 2014 Academic Performance Review'",
           "fall 2013" in all_text and "fall 2014" in all_text
           and ("academic performance review" in all_text
                or "performance review" in all_text),
           "Missing required title text")

    # Per-course paragraph: each GT course name should appear in body
    expected_names = []
    gt_file = os.path.join(groundtruth_workspace, "Year_Over_Year_Comparison.xlsx")
    if os.path.isfile(gt_file):
        try:
            gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
            for s in gt_wb.sheetnames:
                if "comparison" in s.lower() or "course" in s.lower():
                    for r in gt_wb[s].iter_rows(min_row=2, values_only=True):
                        if r and r[0]:
                            expected_names.append(str(r[0]).strip())
                    break
        except Exception:
            pass

    if expected_names:
        for nm in expected_names:
            # Use the leading 8 chars (cuts off year suffix and special chars)
            key = nm.split("(")[0].strip()
            ok = key.lower() in all_text
            record(f"Word mentions course '{nm}'", ok,
                   f"Course '{key}' not in document")

    record("Word mentions '2013'", "2013" in all_text, "No mention of '2013'")
    record("Word mentions '2014'", "2014" in all_text, "No mention of '2014'")
    record("Word mentions 'performance' or 'comparison' or 'review'",
           "performance" in all_text or "comparison" in all_text or "review" in all_text,
           "No relevant keywords found")

    return True


def check_gsheet():
    """Check Google Sheet 'Academic Year Comparison' with 'Comparison Data' sheet."""
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Find spreadsheet titled 'Academic Year Comparison'
    cur.execute(
        "SELECT id, title FROM gsheet.spreadsheets "
        "WHERE title ILIKE '%academic year comparison%'"
    )
    rows = cur.fetchall()
    record("GSheet titled 'Academic Year Comparison' exists",
           len(rows) >= 1, f"Found {len(rows)} matching")
    if len(rows) == 0:
        cur.close()
        conn.close()
        return False

    spreadsheet_id = rows[0][0]
    # Validate the 'Comparison Data' sheet exists with rows.
    # gsheet schema typically has 'sheets' with name and a 'cells' or 'values' table.
    try:
        cur.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_schema='gsheet' AND table_name='sheets'"
        )
        sheet_cols = [r[0] for r in cur.fetchall()]
    except Exception:
        sheet_cols = []

    found_sheet = False
    if "title" in sheet_cols and "spreadsheet_id" in sheet_cols:
        cur.execute(
            "SELECT id, title FROM gsheet.sheets "
            "WHERE spreadsheet_id = %s AND title ILIKE '%%comparison data%%'",
            (spreadsheet_id,),
        )
        srows = cur.fetchall()
        found_sheet = len(srows) >= 1
    record("GSheet has 'Comparison Data' sheet", found_sheet,
           "Sheet 'Comparison Data' not found within spreadsheet")

    cur.close()
    conn.close()
    return len(rows) > 0 and found_sheet


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)
    word_ok = check_word(args.agent_workspace, args.groundtruth_workspace)

    gsheet_ok = check_gsheet()

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    overall = excel_ok and word_ok and gsheet_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
