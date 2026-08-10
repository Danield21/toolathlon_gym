"""Evaluation script for terminal-arxiv-latex-fetch-excel-notion."""
import os
import re
import argparse, json, os, sys
import openpyxl

# All DB connection params come from env with defaults, so preprocess and evaluator
# always talk to the same database regardless of harness port assignment (R1).
DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel")
}


def _to_float(value):
    """Robustly convert a cell value to float, or None if unparseable.

    Handles ints, floats, and strings with thousand-separators, currency
    symbols, percent signs, and stray text around a number. Returns None for
    empty cells, booleans, or values with no extractable number, so callers
    can safely skip them instead of crashing (R2).
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", "").replace("$", "").replace("¥", "").replace("€", "").replace("%", "").strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            pass
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if m:
            try:
                return float(m.group())
            except ValueError:
                return None
    return None


def _effective_value(raw, ws_v, row_idx, col_idx):
    """If a cell holds a formula string, return the cached value instead."""
    if isinstance(raw, str) and raw.startswith("="):
        cached = ws_v.cell(row=row_idx, column=col_idx).value
        return cached
    return raw

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Conference_Prep_Tracker.xlsx")
    check("Conference_Prep_Tracker.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return

    wb = openpyxl.load_workbook(excel_path)  # data_only=False: read headers/strings as written
    wb_values = openpyxl.load_workbook(excel_path, data_only=True)  # cached values for any formulas

    # Paper_Sections sheet
    check("Paper_Sections sheet exists", "Paper_Sections" in wb.sheetnames)
    if "Paper_Sections" in wb.sheetnames:
        ws = wb["Paper_Sections"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Paper_Sections has >= 10 rows", len(data_rows) >= 10, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Paper_ID', 'Paper_Title', 'Section_Title', 'Section_Word_Count']:
            check(f"Paper_Sections has {col}", col.lower() in headers, f"headers: {headers[:5]}")
        # Check exact paper IDs present (stricter than substring; noise paper '2301.99901' must NOT match '2301')
        paper_ids = set(str(r[0]).strip() for r in data_rows if r[0])
        def has_paper(target_full):
            # Match the exact full ID, a prefix-extended form, or an 'arxiv:2301.07041'
            # prefixed variant. Adjacent-digit guard stops noise '2301.99901' matching '2301'.
            for pid in paper_ids:
                if pid == target_full or pid.startswith(target_full):
                    return True
                if target_full in pid:
                    idx = pid.find(target_full)
                    end = idx + len(target_full)
                    after_ok = end >= len(pid) or not pid[end].isdigit()
                    before_ok = idx == 0 or not pid[idx - 1].isdigit()
                    if after_ok and before_ok:
                        return True
            return False
        check("Has scaling laws paper (2301.07041)", has_paper("2301.07041"), f"IDs: {paper_ids}")
        check("Has RLHF paper (2203.11171)", has_paper("2203.11171"), f"IDs: {paper_ids}")
        check("Has OPT paper (2205.01068)", has_paper("2205.01068"), f"IDs: {paper_ids}")
        # Noise paper must NOT be present
        check("Noise paper 2301.99901 NOT present", not has_paper("2301.99901"),
              f"IDs: {paper_ids}")
        check("Noise paper 2302.99902 NOT present", not has_paper("2302.99902"),
              f"IDs: {paper_ids}")
        # Section_Word_Count: positive
        wc_idx_local = None
        headers_full = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for i, h in enumerate(headers_full):
            if "word_count" in h or "wordcount" in h:
                wc_idx_local = i
                break
        if wc_idx_local is not None:
            positive_count = 0
            counted = 0
            ws_v = wb_values[ws.title]
            for row_idx, r in enumerate(data_rows, start=2):
                if not r or len(r) <= wc_idx_local or r[wc_idx_local] is None:
                    continue
                raw = _effective_value(r[wc_idx_local], ws_v, row_idx, wc_idx_local + 1)
                val = _to_float(raw)
                if val is None:
                    # Unparseable/uncached (e.g. a formula with no cached value): skip it
                    # rather than counting it as a failing value (R2).
                    continue
                counted += 1
                if val > 0:
                    positive_count += 1
            check("All Section_Word_Count values are positive",
                  counted > 0 and positive_count == counted,
                  f"only {positive_count} positive of {counted} parseable values")

    # Conference_Schedule sheet
    check("Conference_Schedule sheet exists", "Conference_Schedule" in wb.sheetnames)
    if "Conference_Schedule" in wb.sheetnames:
        ws = wb["Conference_Schedule"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Conference_Schedule has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Session_ID', 'Session_Title', 'Date', 'Related_Papers']:
            check(f"Conference_Schedule has {col}", col.lower() in headers, f"headers: {headers[:7]}")
        # Related_Papers content check: at least some rows must have non-empty entries
        rel_idx = next((i for i, h in enumerate(headers) if "related" in h and "paper" in h), None)
        if rel_idx is not None:
            non_empty = sum(
                1 for r in data_rows
                if r and len(r) > rel_idx and r[rel_idx] is not None and str(r[rel_idx]).strip()
            )
            # At least 1 session must have related papers (cross-referencing requirement)
            check("At least 2 Conference_Schedule rows have non-empty Related_Papers",
                  non_empty >= 2, f"only {non_empty} non-empty")
            # Verify mention of at least one of the 3 expected paper titles or IDs
            keywords = ["scaling", "rlhf", "instructgpt", "instruct", "opt", "open pre-trained",
                        "2301.07041", "2203.11171", "2205.01068"]
            blob = " ".join(str(r[rel_idx]).lower() for r in data_rows
                            if r and len(r) > rel_idx and r[rel_idx] is not None)
            keyword_hits = sum(1 for k in keywords if k in blob)
            check("Related_Papers mentions expected papers (>=2 keyword hits)",
                  keyword_hits >= 2, f"hits={keyword_hits}, blob={blob[:200]}")

    # Presentation_Notes sheet
    check("Presentation_Notes sheet exists", "Presentation_Notes" in wb.sheetnames)
    if "Presentation_Notes" in wb.sheetnames:
        ws = wb["Presentation_Notes"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Presentation_Notes has >= 8 rows", len(data_rows) >= 8, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Slide_Number', 'Topic', 'Key_Points', 'Source_Paper']:
            check(f"Presentation_Notes has {col}", col.lower() in headers, f"headers: {headers[:5]}")


def check_notion():
    print("\n=== Checking Notion ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()

        found_page = False
        for page_id, props in pages:
            if isinstance(props, str):
                props = json.loads(props)
            if isinstance(props, dict):
                # Check if title contains "Conference Prep"
                title_prop = props.get("title", props.get("Title", {}))
                title_text = ""
                if isinstance(title_prop, dict):
                    title_items = title_prop.get("title", [])
                    if isinstance(title_items, list):
                        title_text = " ".join(t.get("plain_text", t.get("text", {}).get("content", "")) for t in title_items if isinstance(t, dict))
                    elif isinstance(title_items, str):
                        title_text = title_items
                elif isinstance(title_prop, str):
                    title_text = title_prop

                if "conference" in title_text.lower() and "prep" in title_text.lower():
                    found_page = True
                    check("Conference Prep page exists", True)

                    # Check for properties
                    has_conf_name = any("conference" in k.lower() and "name" in k.lower() for k in props.keys())
                    has_status_key = any("status" in k.lower() for k in props.keys())
                    has_paper_count_key = any("paper" in k.lower() and "count" in k.lower() for k in props.keys())
                    check("Page has Conference_Name property", has_conf_name, f"props: {list(props.keys())}")
                    check("Page has Status property", has_status_key, f"props: {list(props.keys())}")
                    check("Page has Paper_Count property", has_paper_count_key, f"props: {list(props.keys())}")

                    # Validate property VALUES
                    props_blob = json.dumps(props).lower()
                    # Status must be 'In Progress'
                    check("Status property value is 'In Progress'",
                          "in progress" in props_blob,
                          f"Status not 'In Progress' in props blob")
                    # Paper_Count must be 3 (3 ML papers, not 5 with noise)
                    # Look for "3" in any number-typed property
                    paper_count_correct = False
                    for k, v in props.items():
                        if "paper" in k.lower() and "count" in k.lower():
                            v_str = json.dumps(v).lower()
                            # Look for numeric value 3
                            # Check 'number': 3 form
                            if re.search(r'"number"\s*:\s*3\b', v_str):
                                paper_count_correct = True
                                break
                            # Or simply 3 in string form
                            if re.search(r'\b3\b', v_str):
                                paper_count_correct = True
                                break
                    check("Paper_Count property value is 3",
                          paper_count_correct,
                          f"Paper_Count != 3")
                    break

        if not found_page:
            check("Conference Prep page exists", False, f"Found {len(pages)} pages, none matching 'Conference Prep'")

        # Check for content blocks - must summarize each of the 3 papers' contributions.
        # Threshold is >= 3 (one descriptive block per paper) so a correct, concise page
        # that adds one summary block per paper is not penalized (R11).
        cur.execute("SELECT COUNT(*) FROM notion.blocks")
        block_count = cur.fetchone()[0]
        check("Notion has >= 3 content blocks", block_count >= 3, f"found {block_count} blocks")

        # Check that block content references the 3 expected papers
        cur.execute("SELECT block_data FROM notion.blocks")
        blocks = cur.fetchall()
        all_blocks_text = " ".join(str(b[0] or "").lower() for b in blocks)
        # At least 2 of the 3 paper IDs should be referenced in blocks (or paper topic keywords)
        paper_id_hits = sum(1 for pid in ("2301.07041", "2203.11171", "2205.01068")
                            if pid in all_blocks_text)
        topic_hits = sum(1 for kw in ("scaling", "instruct", "rlhf", "opt", "open pre-trained")
                         if kw in all_blocks_text)
        check("Notion content blocks reference >=2 expected papers (by id or topic)",
              paper_id_hits + topic_hits >= 2,
              f"id hits={paper_id_hits}, topic hits={topic_hits}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Notion accessible", False, str(e))


def check_script(agent_workspace):
    print("\n=== Checking Terminal Script ===")
    check("conference_prep_builder.py exists",
          os.path.exists(os.path.join(agent_workspace, "conference_prep_builder.py")))
    # Verify intermediate JSON outputs explicitly referenced in task.md
    for fname in ("conference_schedule.json", "paper_analysis.json", "conference_prep_results.json"):
        fp = os.path.join(agent_workspace, fname)
        check(f"{fname} exists", os.path.exists(fp))
        if os.path.exists(fp):
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    data = json.load(f)
                check(f"{fname} is valid non-empty JSON",
                      isinstance(data, (list, dict)) and len(data) > 0,
                      f"len={len(data) if hasattr(data, '__len__') else 0}")
            except Exception as e:
                check(f"{fname} is valid non-empty JSON", False, str(e)[:120])


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Conference_Prep_Tracker.xlsx")
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        wb_values = openpyxl.load_workbook(excel_path, data_only=True)
        # No "unexpected sheets" reverse check here: task.md requires the three
        # named sheets but never forbids extra descriptive sheets, so a correct
        # model may reasonably add e.g. a Summary / Key_References sheet. Flagging
        # such sheets would be an unstated constraint that mis-FAILs valid work
        # (fix2-R1). Sheet naming correctness is already covered by the positive
        # "sheet exists" + column checks in check_excel.

        # Word counts should not be negative
        if "Paper_Sections" in wb.sheetnames:
            ws = wb["Paper_Sections"]
            ws_v = wb_values[ws.title]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            wc_idx = next((i for i, h in enumerate(headers) if "word_count" in h), None)
            if wc_idx is not None:
                negative_found = False
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if row and len(row) > wc_idx and row[wc_idx] is not None:
                        raw = _effective_value(row[wc_idx], ws_v, row_idx, wc_idx + 1)
                        wc = _to_float(raw)
                        if wc is not None and wc < 0:
                            check("No negative word counts", False, f"Found {wc}")
                            negative_found = True
                            break
                if not negative_found:
                    check("No negative word counts", True)

    # Notion: no duplicate Conference Prep pages
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, properties FROM notion.pages")
        pages = cur.fetchall()
        conf_pages = 0
        for page_id, props in pages:
            if isinstance(props, str):
                props = json.loads(props)
            if isinstance(props, dict):
                title_prop = props.get("title", props.get("Title", {}))
                title_text = ""
                if isinstance(title_prop, dict):
                    title_items = title_prop.get("title", [])
                    if isinstance(title_items, list):
                        title_text = " ".join(t.get("plain_text", t.get("text", {}).get("content", ""))
                                              for t in title_items if isinstance(t, dict))
                if "conference" in title_text.lower() and "prep" in title_text.lower():
                    conf_pages += 1
        check("No duplicate Conference Prep pages", conf_pages <= 1,
              f"Found {conf_pages} matching pages")
        cur.close()
        conn.close()
    except Exception:
        pass


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    check_excel(agent_workspace)
    check_notion()
    check_script(agent_workspace)
    check_reverse_validation(agent_workspace)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
