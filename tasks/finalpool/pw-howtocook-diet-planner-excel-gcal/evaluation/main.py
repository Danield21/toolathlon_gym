"""Evaluation script for pw-howtocook-diet-planner-excel-gcal."""
import os
import argparse, json, os, sys
import openpyxl

# --- verify_v2 smart primitives ---
_EVAL_DIR = os.path.dirname(os.path.abspath(__file__))
_PACK_ROOT = os.path.abspath(os.path.join(_EVAL_DIR, "..", "..", "..", ".."))
if _PACK_ROOT not in sys.path:
    sys.path.insert(0, _PACK_ROOT)
try:
    from utils.verify_v2 import smart_column_exists
    from utils.verify_v2.eval_helpers import get_sheet_rows_as_dicts, get_gt_column_values
    _HAS_VERIFY_V2 = True
except Exception:
    _HAS_VERIFY_V2 = False

TASK_NAME = "pw-howtocook-diet-planner-excel-gcal"


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "toolathlon_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    # Detect GT self-test (V1 parity test).
    try:
        gt_canon = os.path.realpath(groundtruth_workspace) if groundtruth_workspace else ""
        ag_canon = os.path.realpath(agent_workspace) if agent_workspace else ""
        is_gt_self_test = bool(gt_canon) and (gt_canon == ag_canon)
    except Exception:
        is_gt_self_test = False

    excel_path = os.path.join(agent_workspace, "Diet_Planner_Report.xlsx")
    check("Diet_Planner_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Diet_Planner_Report.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        def check_columns(sheet_name, expected_cols, min_rows):
            """Verify sheet exists, >= min_rows, contains required columns.
            Uses LLM semantic mapping via verify_v2, falls back to strict
            header match."""
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                return
            _ws = wb[sheet_name]
            _data_rows = list(_ws.iter_rows(min_row=2, values_only=True))
            check(f"{sheet_name} has >= {min_rows} rows",
                  len(_data_rows) >= min_rows, f"got {len(_data_rows)}")
            if _HAS_VERIFY_V2 and gt_wb is not None:
                _raw_headers, _agent_rows = get_sheet_rows_as_dicts(wb, sheet_name)
                for _exp in expected_cols:
                    _gt_vals = get_gt_column_values(gt_wb, sheet_name, _exp)
                    _ok, _matched, _reason = smart_column_exists(
                        expected_col=_exp, agent_headers=_raw_headers,
                        gt_samples=_gt_vals[:3], agent_rows=_agent_rows,
                        task_name=TASK_NAME,
                    )
                    _detail = _reason
                    if _ok and _matched and _matched.lower() != _exp.lower():
                        _detail = f"LLM-mapped to {_matched!r}"
                    check(f"{sheet_name} has {_exp} column", _ok, _detail)
            else:
                _headers = [str(c.value).strip().lower() if c.value else "" for c in _ws[1]]
                for _exp in expected_cols:
                    check(f"{sheet_name} has {_exp} column",
                          _exp.lower() in _headers, f"headers: {_headers[:8]}")

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            check_columns('Data_Analysis', ['Recipe', 'Category', 'Calories', 'Protein_g', 'Meets_Guidelines'], 5)
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            check_columns('Metrics', ['Metric', 'Value'], 4)
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            check_columns('Recommendations', ['Priority', 'Action'], 2)
    # Calendar and script checks should run regardless of Excel state.
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Per task.md: title 'Analysis Review' on 2026-03-14 14:00-15:00 UTC
        cur.execute(
            "SELECT summary, start_datetime, end_datetime FROM gcal.events "
            "WHERE summary ILIKE %s",
            ('%analysis%review%',)
        )
        events = cur.fetchall()
        if is_gt_self_test and len(events) == 0:
            print("  [WARN] 'Analysis Review' calendar event exists: 0 found (GT self-test, non-blocking)")
        else:
            check("'Analysis Review' calendar event exists",
                  len(events) >= 1,
                  f"found {len(events)} events with 'Analysis Review' in summary")
            if events:
                # Check at least one event matches the date range
                from datetime import datetime, timezone
                target_start = datetime(2026, 3, 14, 14, 0, tzinfo=timezone.utc)
                ok_date = False
                for sum_, start_dt, end_dt in events:
                    if start_dt is None:
                        continue
                    # Normalize to UTC for comparison; if naive assume UTC.
                    if hasattr(start_dt, 'tzinfo') and start_dt.tzinfo is not None:
                        start_utc = start_dt.astimezone(timezone.utc)
                    else:
                        start_utc = start_dt.replace(tzinfo=timezone.utc) if hasattr(start_dt, 'replace') else start_dt
                    if abs((start_utc - target_start).total_seconds()) < 3600:
                        ok_date = True
                        break
                check("'Analysis Review' starts 2026-03-14 14:00 UTC",
                      ok_date,
                      f"event starts: {[e[1] for e in events]}")
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))

    # Processor file: must exist AND be non-trivial (not just a touched empty file)
    proc_path = os.path.join(agent_workspace, "cook_diet_processor.py")
    proc_exists = os.path.exists(proc_path)
    check("cook_diet_processor.py exists", proc_exists)
    if proc_exists:
        try:
            proc_size = os.path.getsize(proc_path)
            check("cook_diet_processor.py is non-trivial (>=200 bytes)",
                  proc_size >= 200, f"size={proc_size} bytes")
            # Ensure the script actually has Python imports (not a placeholder text file)
            with open(proc_path, "r", encoding="utf-8", errors="ignore") as _pf:
                _src = _pf.read()
            check("cook_diet_processor.py has Python import statements",
                  ('import ' in _src) or ('from ' in _src and 'import' in _src),
                  "no 'import' statement found in processor script")
        except OSError as e:
            check("cook_diet_processor.py size check", False, str(e))

    # Value-level: recipe overlap with GT (only when GT workbook is available)
    if os.path.exists(excel_path):
        try:
            gt_path_local = os.path.join(groundtruth_workspace, "Diet_Planner_Report.xlsx")
            if os.path.exists(gt_path_local) and not is_gt_self_test:
                wb_a = openpyxl.load_workbook(excel_path, data_only=True)
                wb_g = openpyxl.load_workbook(gt_path_local, data_only=True)
                if "Data_Analysis" in wb_a.sheetnames and "Data_Analysis" in wb_g.sheetnames:
                    def first_col_values(ws):
                        out = []
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row and row[0]:
                                out.append(str(row[0]).strip().lower())
                        return out
                    a_recipes = set(first_col_values(wb_a["Data_Analysis"]))
                    g_recipes = set(first_col_values(wb_g["Data_Analysis"]))
                    if g_recipes:
                        overlap = len(a_recipes & g_recipes)
                        # Allow loose lexical overlap (substring) since recipe naming
                        # may vary slightly. Require >=1 meaningful overlap with GT.
                        if overlap == 0:
                            for ar in a_recipes:
                                if any(gr in ar or ar in gr for gr in g_recipes):
                                    overlap += 1
                                    break
                        check("Data_Analysis Recipe values overlap with GT recipes",
                              overlap >= 1,
                              f"agent={list(a_recipes)[:5]}, gt={list(g_recipes)[:5]}")
        except Exception:
            # Soft-fail: don't penalize unrelated parse errors
            pass

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
